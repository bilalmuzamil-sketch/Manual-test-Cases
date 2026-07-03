#!/usr/bin/env python3
"""Build a non-technical, manual-QA-readable Excel workbook from SV5319 role result files."""
import json
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
RESULTS = os.path.join(BASE, "results")
OUT = os.path.join(BASE, "SV5319_QA_Report.xlsx")

ROLE_FILES = [
    "administrator", "service-manager", "senior-service-advisor", "service-advisor",
    "foreman", "technician", "parts-manager", "parts-technician", "office",
    "sales-representative", "time-clock",
]

# ---- styling constants ----
HEADER_FILL = PatternFill("solid", fgColor="1F3864")       # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(size=11)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")          # very light blue-grey
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")        # light green
RED_FILL = PatternFill("solid", fgColor="FFC7CE")          # light red
GRAND_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load_all():
    cases = []
    for f in ROLE_FILES:
        path = os.path.join(RESULTS, f + ".json")
        with open(path) as fh:
            cases.extend(json.load(fh))
    # sort by role then case number (numeric cases first, then any non-numeric)
    def num_key(c):
        n = c.get("num", 0)
        try:
            return (0, float(n), "")
        except (TypeError, ValueError):
            return (1, 0.0, str(n))

    cases.sort(key=lambda c: (c.get("role", ""), num_key(c)))
    return cases


def short(text, limit=90):
    text = (text or "").strip()
    if len(text) <= limit:
        return text
    cut = text[:limit].rsplit(" ", 1)[0]
    return cut + "..."


def what_we_tested(c):
    area = (c.get("feature_area") or "").strip()
    todo = short(c.get("what_to_do") or "", 90)
    if area and todo:
        return f"{area} - {todo}"
    return area or todo


def spec_or_expected(c):
    return (c.get("spec_reference_plain") or "").strip() or (c.get("expected_plain") or "").strip()


def actual_or_expected(c):
    return (c.get("actual_result_plain") or "").strip() or (c.get("expected_result") or "").strip()


def numbered_steps(text):
    """Return replication steps as a clean numbered list, one per line."""
    text = (text or "").strip()
    if not text:
        return ""
    # If already numbered like "1. ..." or "1) ...", split on the step markers.
    # Step numbers are 1-2 digits and the marker must be preceded by whitespace,
    # so values like "200." inside a step are not mistaken for a new step.
    if re.search(r"(?:^|\s)\d{1,2}[.)]\s", text):
        parts = re.split(r"(?<=\s)(?=\d{1,2}[.)]\s)", text)
        lines = [p.strip() for p in parts if p.strip()]
        return "\n".join(lines)
    # Otherwise split on sentence boundaries into a numbered list.
    sentences = re.split(r"(?<=[.!?])\s+", text)
    sentences = [s.strip() for s in sentences if s.strip()]
    return "\n".join(f"{i}. {s}" for i, s in enumerate(sentences, 1))


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_body(ws, first_data_row, last_row, ncols, alt=True):
    for r in range(first_data_row, last_row + 1):
        shade = alt and ((r - first_data_row) % 2 == 1)
        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if shade:
                cell.fill = ALT_FILL


def est_row_height(values, widths):
    """Estimate a row height so wrapped text isn't cut off."""
    max_lines = 1
    for val, w in zip(values, widths):
        s = str(val or "")
        chars_per_line = max(int(w * 1.05), 1)
        lines = 0
        for seg in s.split("\n"):
            lines += max(1, -(-len(seg) // chars_per_line))  # ceil div
        max_lines = max(max_lines, lines)
    return min(max(max_lines * 15, 20), 420)


def main():
    cases = load_all()
    passed = [c for c in cases if c.get("status") == "PASS"]
    failed = [c for c in cases if c.get("status") == "FAIL"]

    wb = Workbook()

    # ---------------- TAB 1: Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    headers = ["Role", "Total Tested", "Passed", "Failed"]
    ws.append(headers)
    style_header(ws, len(headers))

    # per-role counts, ordered by role name
    roles = sorted({c.get("role", "") for c in cases})
    g_total = g_pass = g_fail = 0
    for role in roles:
        rc = [c for c in cases if c.get("role") == role]
        p = sum(1 for c in rc if c.get("status") == "PASS")
        f = sum(1 for c in rc if c.get("status") == "FAIL")
        ws.append([role, len(rc), p, f])
        g_total += len(rc)
        g_pass += p
        g_fail += f
    last_role_row = ws.max_row
    apply_body(ws, 2, last_role_row, len(headers))

    grand_row = ws.max_row + 1
    ws.append(["GRAND TOTAL", g_total, g_pass, g_fail])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=grand_row, column=col)
        cell.font = GRAND_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.append([])  # blank row
    sentence_row = ws.max_row + 1
    sentence = (
        f"Overall: {g_total} tests were run. {g_pass} passed and {g_fail} failed. "
        f"Full step-by-step details for every failed case are on the 'Failed Tests' tab."
    )
    ws.cell(row=sentence_row, column=1, value=sentence)
    ws.cell(row=sentence_row, column=1).font = Font(size=11, italic=True)
    ws.cell(row=sentence_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=sentence_row, start_column=1, end_row=sentence_row, end_column=4)
    ws.row_dimensions[sentence_row].height = 45
    set_widths(ws, [30, 14, 12, 12])

    # ---------------- TAB 2: Passed Tests ----------------
    ws = wb.create_sheet("Passed Tests")
    headers = [
        "Role", "Case #", "What We Tested", "Where In The App",
        "What Should Happen (per spec)", "What Actually Happened", "Screenshot",
    ]
    widths = [22, 8, 40, 30, 60, 60, 28]
    ws.append(headers)
    style_header(ws, len(headers))
    set_widths(ws, widths)
    for c in passed:
        row = [
            c.get("role", ""),
            c.get("num", ""),
            what_we_tested(c),
            (c.get("where_to_go") or "").strip(),
            spec_or_expected(c),
            actual_or_expected(c),
            (c.get("screenshot_ref") or "").strip(),
        ]
        ws.append(row)
        ws.row_dimensions[ws.max_row].height = est_row_height(row, widths)
    apply_body(ws, 2, ws.max_row, len(headers))
    # green status marker on Case # cell
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        cell.fill = GREEN_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=11, bold=True)

    # ---------------- TAB 3: Failed Tests ----------------
    ws = wb.create_sheet("Failed Tests")
    headers = [
        "Role", "Case #", "What We Tested", "Where In The App",
        "Steps To Reproduce", "What Should Happen (per spec)",
        "What Actually Happened (the bug)", "Screenshot",
    ]
    widths = [22, 8, 38, 28, 68, 60, 60, 28]
    ws.append(headers)
    style_header(ws, len(headers))
    set_widths(ws, widths)
    for c in failed:
        row = [
            c.get("role", ""),
            c.get("num", ""),
            what_we_tested(c),
            (c.get("where_to_go") or "").strip(),
            numbered_steps(c.get("replication_steps_plain")),
            spec_or_expected(c),
            (c.get("actual_result_plain") or "").strip(),
            (c.get("screenshot_ref") or "").strip(),
        ]
        ws.append(row)
        ws.row_dimensions[ws.max_row].height = est_row_height(row, widths)
    apply_body(ws, 2, ws.max_row, len(headers))
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=2)
        cell.fill = RED_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=11, bold=True)

    wb.save(OUT)
    print("Saved:", OUT)
    print("Tabs:", wb.sheetnames)
    print("Passed rows:", len(passed), "Failed rows:", len(failed), "Total:", len(cases))


if __name__ == "__main__":
    main()
