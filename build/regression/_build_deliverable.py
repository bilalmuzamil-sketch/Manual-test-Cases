#!/usr/bin/env python3
"""Compile the final consolidated SV5319 regression deliverable from the 11 per-role result files."""
import json, os, csv
from collections import OrderedDict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
RES = os.path.join(BASE, "results")

# Order matters: preserve the task-specified file order
FILE_ORDER = ["administrator", "service-manager", "senior-service-advisor",
              "service-advisor", "foreman", "technician", "parts-manager",
              "parts-technician", "office", "sales-representative", "time-clock"]

cases = []
for f in FILE_ORDER:
    with open(os.path.join(RES, f + ".json")) as fh:
        cases.extend(json.load(fh))

def norm_status(s):
    return (s or "").strip().upper()

# ---- Column definitions (main sheets) ----
MAIN_COLS = [
    ("Role", "role"), ("Case #", "num"), ("Feature Area", "feature_area"),
    ("Where To Go", "where_to_go"), ("What To Do", "what_to_do"),
    ("Expected Result", "expected_result"), ("Depth", "depth"), ("Status", "status"),
    ("Spec Reference (plain)", "spec_reference_plain"), ("Expected (layman)", "expected_plain"),
    ("Actual Result (layman)", "actual_result_plain"),
    ("Steps To Replicate (layman)", "replication_steps_plain"),
    ("Screenshot", "screenshot_ref"), ("Notes", "notes"),
]
MAIN_WIDTHS = [22, 8, 22, 24, 26, 34, 8, 10, 46, 40, 60, 50, 16, 46]
STATUS_COL_IDX = 8  # 1-based column index of Status in MAIN_COLS

# ---- Styles ----
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
GREEN_FONT = Font(color="006100")
RED_FONT = Font(color="9C0006", bold=True)
AMBER_FONT = Font(color="9C6500")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def status_style(st):
    st = norm_status(st)
    if st == "PASS":
        return GREEN, GREEN_FONT
    if st == "FAIL":
        return RED, RED_FONT
    return AMBER, AMBER_FONT  # PARTIAL / other

def write_sheet(ws, cols, widths, rows, status_idx=None):
    # header
    for ci, (title, _) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=ci, value=title)
        c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = HEADER_ALIGN
        c.border = BORDER
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    # data
    for ri, row in enumerate(rows, start=2):
        for ci, (_, key) in enumerate(cols, start=1):
            val = row.get(key, "") if isinstance(row, dict) else row[ci-1]
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = WRAP; c.border = BORDER
        if status_idx:
            st = row.get("status", "") if isinstance(row, dict) else row[status_idx-1]
            fill, font = status_style(st)
            sc = ws.cell(row=ri, column=status_idx)
            sc.fill = fill; sc.font = font
            sc.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

# ---- Counts ----
role_order = []
per_role = OrderedDict()
for c in cases:
    r = c["role"]
    if r not in per_role:
        per_role[r] = Counter(); role_order.append(r)
    per_role[r][norm_status(c["status"])] += 1

grand = Counter()
for r in role_order:
    grand.update(per_role[r])

# ---- Backend-Enforcement Gaps (curated from FAILs/notes) ----
GAP_ROWS = [
    ["Technician", "35", "POST /api/work-orders/lines/change-status",
     "As the Technician, called change-status with status='authorized'; HTTP 200 and the line moved authorization_required -> authorized (verified on re-read).",
     "UI correctly hides the approve/authorize control, but the API applies the status change even though the Technician lacks Work Orders: Create and Edit (the 'authorized' status requires ROLE_WORK_ORDER::CREATE_AND_EDIT). No 403 returned."],
    ["Technician", "37", "POST /api/work-orders/change-service-advisor",
     "As the Technician, posted change-service-advisor (service_advisor_id=null); HTTP 201 and the field was cleared and persisted on re-read.",
     "The WO header is rendered read-only in the UI, but the change-* endpoint family accepts edits from a View-only Technician; header edits are not blocked at the API."],
    ["Technician", "41", "POST /api/work-orders/part/perform-request-status-action (action:'order')",
     "As the Technician, ordered part c40b2d97; HTTP 201 and the part moved authorized_to_order -> waiting_to_receive (Awaiting).",
     "The Technician has Order Parts OFF (woOrderParts=false) and the UI hides ordering, but the endpoint executes the order with no permission check. No 403."],
    ["Parts Manager", "44", "POST /api/work-orders/delete (for a part sale)",
     "Deleting part sale P9-994 fires work-orders/delete and returns HTTP 403 Access denied; the user is redirected to /access-denied.",
     "Wrong-permission gate: part-sale deletion is wired to Work Orders: Delete (which PM lacks) instead of Part Sales: Delete (which PM has). The UI shows a Delete control that always fails - a UX bug plus a permission-mapping inconsistency vs the spec."],
    ["Parts Technician", "49", "POST /api/parts-catalogue/remove-catalogue-part?id=<id>",
     "As the Parts Tech, created a throwaway part then deleted it via remove-catalogue-part; HTTP 200 and the part was removed from the catalogue.",
     "The Parts Technician has no catalogInventoryDelete permission and the UI hides delete, but the delete endpoint has no permission check and executes the delete."],
    ["Parts Technician", "N1", "POST /api/work-orders/part/remove-return-request",
     "As the Parts Tech, deleted a return (part 47-0680, WO S-13556); HTTP 200 and the return was removed.",
     "Per spec v33 (SV-7911) delete-return should require Invoicing & Payments: Delete, which the Parts Tech lacks. The endpoint still uses the old Vendor & Order Delete gate (which the role has), and the UI still exposes 'Delete Return'."],
    ["Office", "55", "POST /api/parts-catalogue/change-catalogue-part",
     "As the view-only Office user, posted a catalog part name change; HTTP 200 and the change persisted on read-back (restored afterward).",
     "Office has catalogInventoryView only (no Create and Edit) and the UI hides all edit controls, but the edit endpoint accepts and persists the change from a view-only user."],
    ["Time Clock", "N1", "POST /api/note/create",
     "As the Time Clock user (most restricted role), created a WO note; HTTP 201 and the note persisted in the work order's notes list.",
     "Notes should be blocked, but note/create (and other WO write endpoints: lines/create, lines/change-status, tasks/create, parts/create, invoices/create, note/update, note/delete) pass the permission check (reach 400 validation, not 403). Returns/DI/customers endpoints correctly return 403."],
]
GAP_COLS = [("Role","0"),("Case #","1"),("Endpoint","2"),("What Happened","3"),("Why It's a Gap","4")]
GAP_WIDTHS = [18, 8, 46, 60, 62]

# ---- One-line FAIL reasons (layman) ----
FAIL_REASON = {
    ("Service Manager","11"): "Service Manager was able to fully reverse a work-order invoice end-to-end (expected: blocked).",
    ("Service Manager","N2"): "Service Manager could reverse a work-order invoice end-to-end (expected: blocked); the staging role carries Work Orders: Delete per SV-8093.",
    ("Service Manager","14"): "Service Manager's Settings area is not limited to App Settings + Wages - Staff, Roles & Permissions, Locations, Departments and Feature Flags all open and are editable.",
    ("Service Manager","N3"): "Service Manager could delete another user's work-order note (expected: only own notes editable).",
    ("Service Advisor (a.k.a. Junior SA)","26"): "AP/AR customer tabs are visible and load real data for the Service Advisor even though AP/AR is OFF (they unlock via the invoicing-delete permission).",
    ("Technician","35"): "Technician could authorize/approve a work-order line via the API despite the control being hidden and the role lacking Create and Edit.",
    ("Technician","37"): "Technician could change a work-order header field (service advisor) via the API despite being View-only.",
    ("Technician","41"): "Technician could order a part via the API despite Order Parts being OFF.",
    ("Parts Manager","44"): "Parts Manager cannot delete a part sale - the Delete control is shown but the API returns Access Denied (gated by the wrong permission).",
    ("Parts Technician","49"): "Parts Technician could delete a catalog part via the API despite having no delete permission.",
    ("Parts Technician","N1"): "Parts Technician could delete a part return despite lacking Invoicing & Payments: Delete (the v33/SV-7911 gate is not enforced).",
    ("Office","55"): "Office (view-only) could edit a catalog part via the API despite all edit controls being hidden.",
    ("Time Clock","N1"): "Time Clock could create a work-order note via the API despite being the most restricted role.",
}
def reason_for(c):
    return FAIL_REASON.get((c["role"], str(c["num"])), c.get("feature_area",""))

fails = [c for c in cases if norm_status(c["status"]) == "FAIL"]

SYSTEMIC = [
    "Backend enforcement gaps: several granular permissions (line-authorize, WO header edit, order-parts, catalog delete, catalog edit, return delete, note-create) are enforced only in the front-end; the API accepts the action. Cases: Technician 35/37/41, Parts Technician 49/N1, Office 55, Time Clock N1.",
    "Wrong-permission gating: part-sale delete is gated by Work Orders: Delete instead of Part Sales: Delete (Parts Manager 44); AP/AR customer tabs unlock via an invoicing-delete OR-gate (Service Advisor 26).",
    "Role-config vs spec mismatch: the staging Service Manager role includes Work Orders: Delete (SV-8093), letting SM reverse invoices and delete others' notes (SM 11/N2/N3); the SM Settings area is over-exposed (SM 14); the Time Clock role grants workOrdersView/scheduleView.",
]

# ==================== BUILD XLSX ====================
wb = Workbook()

# --- Summary sheet ---
ws = wb.active; ws.title = "Summary"
ws["A1"] = "SV5319 Regression - Consolidated Results"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Compiled 2026-07-03 | 11 roles | flagged cases re-tested with real execution (see AUDIT-flagged-cases.json)"
ws["A2"].font = Font(italic=True, size=10, color="666666")

hdr_row = 4
headers = ["Role", "Total", "PASS", "FAIL", "PARTIAL"]
for ci, h in enumerate(headers, start=1):
    c = ws.cell(row=hdr_row, column=ci, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = HEADER_ALIGN; c.border = BORDER
r = hdr_row + 1
for role in role_order:
    ct = per_role[role]
    tot = sum(ct.values())
    vals = [role, tot, ct.get("PASS",0), ct.get("FAIL",0), ct.get("PARTIAL",0)]
    for ci, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=ci, value=v); c.border = BORDER
        c.alignment = Alignment(vertical="center")
    r += 1
# grand total
gt = ["GRAND TOTAL", sum(grand.values()), grand.get("PASS",0), grand.get("FAIL",0), grand.get("PARTIAL",0)]
for ci, v in enumerate(gt, start=1):
    c = ws.cell(row=r, column=ci, value=v); c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor="DDEBF7"); c.border = BORDER
grand_row = r
r += 2
ws.cell(row=r, column=1, value="Key systemic findings").font = Font(bold=True, size=12)
r += 1
for f in SYSTEMIC:
    c = ws.cell(row=r, column=1, value="- " + f)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 60
    r += 1
for ci, w in enumerate([22, 10, 10, 10, 12], start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.freeze_panes = "A5"

# --- All Results ---
ws = wb.create_sheet("All Results")
write_sheet(ws, MAIN_COLS, MAIN_WIDTHS, cases, status_idx=STATUS_COL_IDX)

# --- Passed (PASS + PARTIAL) ---
passed = [c for c in cases if norm_status(c["status"]) in ("PASS", "PARTIAL")]
ws = wb.create_sheet("Passed")
write_sheet(ws, MAIN_COLS, MAIN_WIDTHS, passed, status_idx=STATUS_COL_IDX)

# --- Failed ---
ws = wb.create_sheet("Failed")
write_sheet(ws, MAIN_COLS, MAIN_WIDTHS, fails, status_idx=STATUS_COL_IDX)

# --- Backend-Enforcement Gaps ---
ws = wb.create_sheet("Backend-Enforcement Gaps")
for ci, (title, _) in enumerate(GAP_COLS, start=1):
    c = ws.cell(row=1, column=ci, value=title)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = HEADER_ALIGN; c.border = BORDER
for ci, w in enumerate(GAP_WIDTHS, start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
for ri, row in enumerate(GAP_ROWS, start=2):
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=ri, column=ci, value=val); c.alignment = WRAP; c.border = BORDER
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 24

xlsx_path = os.path.join(BASE, "SV5319_Regression_Results.xlsx")
wb.save(xlsx_path)

# ==================== All Results CSV ====================
csv_path = os.path.join(BASE, "SV5319_Regression_Results.csv")
with open(csv_path, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow([t for t, _ in MAIN_COLS])
    for c in cases:
        w.writerow([c.get(k, "") for _, k in MAIN_COLS])

# ==================== Failures Layman (CSV + XLSX) ====================
LAY_COLS = ["Role", "Case #", "What Failed (plain)", "Steps To Replicate (numbered, layman)",
            "Expected Behavior (per spec, plain)", "Actual Result We Got (plain)", "Screenshot"]
lay_rows = []
for c in fails:
    lay_rows.append([
        c["role"], str(c["num"]), reason_for(c),
        c.get("replication_steps_plain", ""), c.get("expected_plain", ""),
        c.get("actual_result_plain", ""), c.get("screenshot_ref", ""),
    ])

lay_csv = os.path.join(BASE, "SV5319_Failures_Layman.csv")
with open(lay_csv, "w", newline="") as fh:
    w = csv.writer(fh)
    w.writerow(LAY_COLS)
    w.writerows(lay_rows)

wb2 = Workbook(); ws = wb2.active; ws.title = "Failures (Layman)"
for ci, h in enumerate(LAY_COLS, start=1):
    c = ws.cell(row=1, column=ci, value=h)
    c.fill = HEADER_FILL; c.font = HEADER_FONT; c.alignment = HEADER_ALIGN; c.border = BORDER
for ci, w in enumerate([18, 8, 50, 55, 45, 60, 16], start=1):
    ws.column_dimensions[get_column_letter(ci)].width = w
for ri, row in enumerate(lay_rows, start=2):
    for ci, val in enumerate(row, start=1):
        c = ws.cell(row=ri, column=ci, value=val); c.alignment = WRAP; c.border = BORDER
ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30
lay_xlsx = os.path.join(BASE, "SV5319_Failures_Layman.xlsx")
wb2.save(lay_xlsx)

print("OK")
print("total", len(cases), "fails", len(fails))
print("files:", xlsx_path, csv_path, lay_csv, lay_xlsx)
