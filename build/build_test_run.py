#!/usr/bin/env python3
"""Build the final test-run deliverable from the 5 committed batch result files.

Adds four result sheets to custom-roles-test-cases.xlsx (keeping all existing
tabs) and writes four CSV exports to the repo root. Pure data/file work.
"""
import csv
import json
import os
from collections import Counter, OrderedDict

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BUILD = os.path.join(REPO, "build")
RESULTS = os.path.join(BUILD, "test-run-results")
XLSX = os.path.join(REPO, "custom-roles-test-cases.xlsx")

BATCHES = OrderedDict([
    ("A", 52), ("B", 65), ("C", 41), ("D", 18), ("E", 121),
])

RUN_DATE = "2026-07-02"
ENVIRONMENT = "staging (Foothills Group Inc), per-role as Tech via dev-login"

SP_MAP = {
    "WO": "Work Orders", "WOL": "Work Order Lines", "SCH": "Schedule",
    "CUST": "Customer Management", "PS": "Part Sales",
    "CAT": "Catalog & Inventory", "VEND": "Vendor & Order Management",
    "INV": "Invoicing & Payments", "TS": "Timesheets",
    "WOSUB": "WO Sub-Settings", "RPT": "Reports", "CPORT": "Customer Portal",
    "BPORT": "Billing Portal", "PDEPT": "Parts Department", "SET": "Settings",
    "VM": "View Mode", "FIN": "See Financial Data", "APAR": "Manage AP/AR",
    "HIST": "View History Logs",
}
TE_MAP = {
    "ADMIN": "Administrator", "SM": "Service Manager",
    "SSA": "Senior Service Advisor", "SADV": "Service Advisor",
    "FORE": "Foreman", "TECH": "Technician", "PM": "Parts Manager",
    "PT": "Parts Technician", "OFFICE": "Office",
    "SALES": "Sales Representative", "TIMECLK": "Time Clock",
}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
TOP_LEFT = Alignment(vertical="top", horizontal="left", wrap_text=False)
TOP_WRAP = Alignment(vertical="top", horizontal="left", wrap_text=True)

DISCREPANCIES = [
    "VIU-32 [SP-INV-005, FAIL]: Invoicing Delete with See Financial Data ON "
    "but View/Manage AP/AR OFF does NOT trigger the AP/AR dependency gate/modal "
    "(financial gate works; AP/AR gate missing).",
    "VIU-33 [DI-111 / DI-117, FAIL]: The 'Time Clock' system role grants "
    "workOrdersView, so Time-Clock users reach Work Orders / Digital Inspections "
    "read-only — contradicts its clock-in/out-only intent.",
    "VIU-34: CRUD dependency cascade is UI-editor-only, NOT server-enforced — "
    "a direct API PUT of a Delete/Edit permission without its View parent is "
    "accepted by the backend.",
    "VIU-35: /administration/inspection-templates route lacks a settingsService "
    "guard — reachable read-only via direct URL by roles holding any other "
    "settings sub-permission (the Settings>Service menu path is correctly denied).",
    "VIU-36: Several granted surfaces return 404 in this build even when the "
    "permission is granted (Customer/Billing Portal, some Invoicing/AP-AR pages) "
    "— grant path reaches a 404 rather than the feature (possible build gap).",
    "VIU-37 [VERIFIED, informational]: Setting view_mode full/tech auto-adds a "
    "woFullViewMode / woTechViewMode marker permission to the stored role.",
]


def area_section(tid):
    parts = tid.split("-")
    if tid.startswith("SP-"):
        return "Single Permission > " + SP_MAP.get(parts[1], parts[1])
    if tid.startswith("TE-"):
        return "Template Edit > " + TE_MAP.get(parts[1], parts[1])
    if tid.startswith("CB-REP"):
        return "Combination (Representative)"
    if tid.startswith("CB-RND"):
        return "Combination (Random)"
    if tid.startswith("DI-"):
        return "Digital Inspections"
    return parts[0]


def load_all():
    per_batch = OrderedDict()
    allr = []
    for b, expected in BATCHES.items():
        data = json.load(open(os.path.join(RESULTS, f"batch{b}.json")))
        assert len(data) == expected, (b, len(data), expected)
        per_batch[b] = data
        allr.extend(data)
    return per_batch, allr


def style_data_sheet(ws, headers, widths, wrap_cols, nrows):
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = TOP_WRAP
        ws.column_dimensions[get_column_letter(ci)].width = widths[h]
    for r in range(2, nrows + 2):
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=r, column=ci)
            cell.alignment = TOP_WRAP if h in wrap_cols else TOP_LEFT
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(nrows + 1, 1)}"


def add_or_replace(wb, name):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def build_run_summary(wb, allr, per_batch):
    ws = add_or_replace(wb, "Run Summary")
    counts = Counter(x["status"] for x in allr)
    npass, nfail, nblock = counts.get("PASS", 0), counts.get("FAIL", 0), counts.get("BLOCKED", 0)

    ws["A1"] = "ShopView Custom Roles & Permissions — Live Test Run"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = f"Run date: {RUN_DATE}"
    ws["A4"] = f"Environment: {ENVIRONMENT}"

    ws["A6"] = "Overall Results"
    ws["A6"].font = Font(bold=True, size=12)
    over = [("Status", "Count"), ("PASS", npass), ("FAIL", nfail),
            ("BLOCKED", nblock), ("TOTAL", len(allr))]
    r = 7
    for label, val in over:
        ca, cb = ws.cell(row=r, column=1, value=label), ws.cell(row=r, column=2, value=val)
        if label in ("Status", "TOTAL"):
            ca.font, cb.font = HEADER_FONT, HEADER_FONT
        if label == "Status":
            ca.fill = cb.fill = HEADER_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Per-Batch Breakdown").font = Font(bold=True, size=12)
    r += 1
    hdr = ["Batch", "PASS", "FAIL", "BLOCKED", "Total"]
    for ci, h in enumerate(hdr, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font, c.fill = HEADER_FONT, HEADER_FILL
    r += 1
    for b, data in per_batch.items():
        cc = Counter(x["status"] for x in data)
        vals = [f"Batch {b}", cc.get("PASS", 0), cc.get("FAIL", 0),
                cc.get("BLOCKED", 0), len(data)]
        for ci, v in enumerate(vals, start=1):
            ws.cell(row=r, column=ci, value=v)
        r += 1
    tot = ["TOTAL", npass, nfail, nblock, len(allr)]
    for ci, v in enumerate(tot, start=1):
        ws.cell(row=r, column=ci, value=v).font = HEADER_FONT
    r += 2

    ws.cell(row=r, column=1, value="Run Discrepancies (see VIU Findings Log VIU-32..37)").font = Font(bold=True, size=12)
    r += 1
    for d in DISCREPANCIES:
        c = ws.cell(row=r, column=1, value="• " + d)
        c.alignment = TOP_WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 42
        r += 1

    ws.column_dimensions["A"].width = 22
    for col in ("B", "C", "D", "E"):
        ws.column_dimensions[col].width = 14
    # merged discrepancy cells need wide col A visually; keep A moderate, rows merged across B-E
    return ws


def build_passed(wb, allr):
    ws = add_or_replace(wb, "Passed")
    headers = ["Test ID", "Title", "Area/Section", "Status", "Actual (brief)", "Evidence"]
    widths = {"Test ID": 16, "Title": 50, "Area/Section": 34, "Status": 10,
              "Actual (brief)": 70, "Evidence": 24}
    wrap = {"Title", "Area/Section", "Actual (brief)", "Evidence"}
    ws.append(headers)
    rows = [x for x in allr if x["status"] == "PASS"]
    for x in rows:
        ws.append([x["test_id"], x["title"], area_section(x["test_id"]),
                   x["status"], x.get("notes") or x.get("actual", ""),
                   x.get("evidence", "")])
    style_data_sheet(ws, headers, widths, wrap, len(rows))
    return len(rows)


def build_failed(wb, allr):
    ws = add_or_replace(wb, "Failed")
    headers = ["Test ID", "Title", "Area/Section", "Expected Result",
               "Actual Result", "Difference/Notes", "Evidence"]
    widths = {"Test ID": 16, "Title": 44, "Area/Section": 28,
              "Expected Result": 55, "Actual Result": 60,
              "Difference/Notes": 45, "Evidence": 22}
    wrap = {"Title", "Area/Section", "Expected Result", "Actual Result",
            "Difference/Notes", "Evidence"}
    ws.append(headers)
    rows = [x for x in allr if x["status"] == "FAIL"]
    for x in rows:
        ws.append([x["test_id"], x["title"], area_section(x["test_id"]),
                   x.get("expected", ""), x.get("actual", ""),
                   x.get("notes", ""), x.get("evidence", "")])
    style_data_sheet(ws, headers, widths, wrap, len(rows))
    return len(rows)


def build_blocked(wb, allr):
    ws = add_or_replace(wb, "Blocked (Not Verified)")
    headers = ["Test ID", "Title", "Area/Section", "Reason Blocked", "Notes"]
    widths = {"Test ID": 16, "Title": 46, "Area/Section": 30,
              "Reason Blocked": 70, "Notes": 55}
    wrap = {"Title", "Area/Section", "Reason Blocked", "Notes"}
    ws.append(headers)
    rows = [x for x in allr if x["status"] == "BLOCKED"]
    for x in rows:
        ws.append([x["test_id"], x["title"], area_section(x["test_id"]),
                   x.get("actual", ""), x.get("notes", "")])
    style_data_sheet(ws, headers, widths, wrap, len(rows))
    return len(rows)


def write_csvs(allr):
    def w(path, headers, rowfn, pred):
        with open(os.path.join(REPO, path), "w", newline="", encoding="utf-8") as f:
            cw = csv.writer(f)
            cw.writerow(headers)
            n = 0
            for x in allr:
                if pred(x):
                    cw.writerow(rowfn(x))
                    n += 1
        return n

    n_pass = w("custom-roles-test-run-passed.csv",
               ["Test ID", "Title", "Area/Section", "Status", "Actual (brief)", "Evidence"],
               lambda x: [x["test_id"], x["title"], area_section(x["test_id"]),
                          x["status"], x.get("notes") or x.get("actual", ""),
                          x.get("evidence", "")],
               lambda x: x["status"] == "PASS")

    n_fail = w("custom-roles-test-run-failed.csv",
               ["Test ID", "Title", "Area/Section", "Expected Result",
                "Actual Result", "Difference/Notes", "Evidence"],
               lambda x: [x["test_id"], x["title"], area_section(x["test_id"]),
                          x.get("expected", ""), x.get("actual", ""),
                          x.get("notes", ""), x.get("evidence", "")],
               lambda x: x["status"] == "FAIL")

    n_block = w("custom-roles-test-run-blocked.csv",
                ["Test ID", "Title", "Area/Section", "Reason Blocked", "Notes"],
                lambda x: [x["test_id"], x["title"], area_section(x["test_id"]),
                           x.get("actual", ""), x.get("notes", "")],
                lambda x: x["status"] == "BLOCKED")

    n_all = w("custom-roles-test-run-all.csv",
              ["Test ID", "Title", "Area/Section", "Status", "Config",
               "Expected", "Actual", "Notes", "Evidence"],
              lambda x: [x["test_id"], x["title"], area_section(x["test_id"]),
                         x["status"], x.get("config", ""), x.get("expected", ""),
                         x.get("actual", ""), x.get("notes", ""),
                         x.get("evidence", "")],
              lambda x: True)
    return n_pass, n_fail, n_block, n_all


def main():
    per_batch, allr = load_all()
    assert len(allr) == 297, len(allr)
    counts = Counter(x["status"] for x in allr)
    print("=== RECONCILED COUNTS ===")
    print(f"PASS={counts['PASS']} FAIL={counts['FAIL']} "
          f"BLOCKED={counts['BLOCKED']} TOTAL={len(allr)}")
    assert counts["PASS"] == 232 and counts["FAIL"] == 3 and counts["BLOCKED"] == 62

    orig = ["Index", "VIU Findings Log", "Single Permission", "Template Edit",
            "Combination - Representative", "Combination - Random",
            "Digital Inspections"]
    wb = load_workbook(XLSX)
    assert all(s in wb.sheetnames for s in orig), wb.sheetnames

    build_run_summary(wb, allr, per_batch)
    np_ = build_passed(wb, allr)
    nf = build_failed(wb, allr)
    nb = build_blocked(wb, allr)
    wb.save(XLSX)

    n_pass, n_fail, n_block, n_all = write_csvs(allr)

    # ---- verify ----
    wb2 = load_workbook(XLSX)
    new = ["Run Summary", "Passed", "Failed", "Blocked (Not Verified)"]
    print("\n=== WORKBOOK SHEETS ===")
    print("  " + " | ".join(wb2.sheetnames))
    for s in orig + new:
        assert s in wb2.sheetnames, s
    print(f"  original {len(orig)} + new {len(new)} present: OK")

    print("\n=== SHEET DATA ROWS ===")
    print(f"  Passed={wb2['Passed'].max_row - 1} "
          f"Failed={wb2['Failed'].max_row - 1} "
          f"Blocked={wb2['Blocked (Not Verified)'].max_row - 1}")
    assert wb2["Passed"].max_row - 1 == 232
    assert wb2["Failed"].max_row - 1 == 3
    assert wb2["Blocked (Not Verified)"].max_row - 1 == 62

    print("\n=== CSV ROW COUNTS ===")
    print(f"  passed={n_pass} failed={n_fail} blocked={n_block} all={n_all}")
    assert (n_pass, n_fail, n_block, n_all) == (232, 3, 62, 297)
    print("\nAll assertions passed.")


if __name__ == "__main__":
    main()
