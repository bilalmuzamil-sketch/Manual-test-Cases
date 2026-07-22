#!/usr/bin/env python3
"""Traceability-status deliverable (post-backfill) for Simple Flow + Fees & Discounts.
Reads the per-project testrail-id-map.csv (now carrying the `refs` + `trace_note`
columns after the 2026-07-22 backfill) and emits, per project, a .md + .xlsx showing
each case now Traceable with its authentic per-story Jira ticket(s) + spec anchor.
Human-readable filenames (Rule 19); Case ID + TestRail link columns (Rule 8);
mirrors the Missing-Traceability layout (Rule 16)."""
import csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
DATE="2026-07-22"
LINK="https://shopview.testrail.io/index.php?/cases/view/{}"
BASE=os.path.dirname(os.path.abspath(__file__))
SPEC={'SF':'https://shopview.atlassian.net/wiki/spaces/PM/pages/646021121',
      'FD':'https://shopview.atlassian.net/wiki/spaces/~712020aa00b8d6a71f4259891982a304227c20/pages/622297094'}

PROJECTS=[
 ('Simple Flow','SF','build/simple-flow/testrail-id-map.csv','sf_id','build/simple-flow/SimpleFlow_TraceabilityStatus_2026-07-22'),
 ('Fees and Discounts','FD','build/fees-discounts/testrail-id-map.csv','fd_id','build/fees-discounts/FeesAndDiscounts_TraceabilityStatus_2026-07-22'),
]

def rows_for(path,idcol):
    rows=[]
    for r in csv.DictReader(open(path)):
        if not r['ID'].strip(): continue
        rows.append(r)
    return rows

hf=PatternFill('solid',fgColor='1F4E78'); hfont=Font(bold=True,color='FFFFFF')
wrap=Alignment(wrap_text=True,vertical='top')
warn=PatternFill('solid',fgColor='FCE4D6'); ok=PatternFill('solid',fgColor='E2EFDA')
thin=Border(*[Side(style='thin',color='D9D9D9')]*4)
HDR=['Case ID','TestRail link','Internal ID','Section (spec anchor)','refs — ticket + spec (set)','Status']

for name,code,path,idcol,outbase in PROJECTS:
    rows=rows_for(path,idcol)
    total=len(rows)
    flagged=[r for r in rows if r.get('trace_note','').strip()]
    traceable=[r for r in rows if not r.get('trace_note','').strip() and r.get('refs','').strip()]
    # XLSX
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='Traceability status'
    ws.append([f'{name} — traceability status after backfill ({DATE})'])
    ws['A1'].font=Font(bold=True,size=13)
    ws.append([f'{total} cases in scope · {len(traceable)} Traceable (per-story ticket + spec anchor together in TestRail refs) · {len(flagged)} ignored per user 2026-07-22 (id-map C-ID absent in live TestRail). Spec: {SPEC[code]}'])
    ws.append([])
    ws.append(HDR)
    for c in ws[ws.max_row]: c.fill=hf; c.font=hfont; c.alignment=wrap
    for r in rows:
        cid=r['ID'].strip()
        note=r.get('trace_note','').strip()
        status='Traceable' if not note else 'IGNORED per user (TestRail case absent; refs kept locally)'
        ws.append([f'C{cid}',LINK.format(cid),r[idcol],r['section'],r.get('refs',''),status])
        rr=ws[ws.max_row]
        for c in rr:
            c.alignment=wrap; c.border=thin
            c.fill = warn if note else ok
    for i,w in enumerate([9,42,14,40,22,40],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes='A5'
    wb.save(outbase+'.xlsx')
    # MD
    with open(outbase+'.md','w') as fh:
        fh.write(f"# {name} — traceability status after backfill ({DATE})\n\n")
        fh.write(f"**{total} cases in scope. {len(traceable)} Traceable** — each carries its authentic **per-story ticket + spec anchor together** in the TestRail References (`refs`) field, format `<TICKET(S)> (<spec-anchor>)` (Standing Rule 20). Spec URL: `{SPEC[code]}`. **{len(flagged)} ignored** (per user 2026-07-22) below.\n\n")
        fh.write("| Case | Internal ID | Section (spec anchor) | refs — ticket + spec (set) | Status |\n|---|---|---|---|---|\n")
        for r in rows:
            cid=r['ID'].strip(); note=r.get('trace_note','').strip()
            status='Traceable' if not note else 'IGNORED (user decision)'
            fh.write(f"| [C{cid}]({LINK.format(cid)}) | {r[idcol]} | {r['section']} | {r.get('refs','')} | {status} |\n")
        if flagged:
            fh.write(f"\n## Ignored — id-map C-ID absent in live TestRail ({len(flagged)})\n\n")
            fh.write("These case numbers return HTTP 400 in TestRail (created then deleted; all sit on the Create-Purchase-Orders toggle, a known build Deviation). **User decision 2026-07-22: ignore — not re-created, not treated as an open item.** Authentic ticket + spec refs kept in the id-map for the local record only.\n\n")
            fh.write("| Case | Internal ID | Title | refs — ticket + spec |\n|---|---|---|---|\n")
            for r in flagged:
                fh.write(f"| C{r['ID'].strip()} | {r[idcol]} | {r['title']} | {r.get('refs','')} |\n")
    print(f"{name}: total={total} traceable={len(traceable)} flagged={len(flagged)} -> {outbase}.xlsx/.md")
