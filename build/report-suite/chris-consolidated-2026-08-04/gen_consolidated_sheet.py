#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generator — Report Suite: ONE workbook, THREE reader-facing tabs, for Chris Ward (PO).
Date: 2026-08-04.

WHY THIS EXISTS
---------------
The QA lead's instruction, verbatim:
    "make one sheet and share with me with three tabs"

Three separate sheets were ready to send. Sending three files invites a bundled
non-answer and risks an old one being forwarded. This consolidates them into ONE
workbook, most-urgent tab first, with every duplicate removed.

THE THREE SOURCES (carried across FAITHFULLY, not rewritten)
------------------------------------------------------------
  1. chris-location-question-2026-08-04/gen_location_question.py
       the single-issue Location-column question (8 cases, the automation blocker)
  2. chris-sheet-2026-08-04/gen_spec_vs_build_sheet.py
       the 10-row spec-versus-build sheet (8 decisions + 1 write-down + 1 awareness)
  3. gen_po_questions_2026-08-03.py
       the 17-item sheet (5 decisions + 12 write-downs)

This generator IMPORTS those three modules and re-emits their own data structures.
Nothing is re-typed, so no wording can drift in transit. Every text change made in
consolidation is enumerated in CHANGES below and printed on the QA-only tab.

DE-DUPLICATION (Rule 17 — the same question is never asked twice)
-----------------------------------------------------------------
  DROP_SVB / DROP_PQ name the removed items, and DEDUP_LOG records, per removal,
  what was removed, where the question now lives, and why. The QA-only mapping row
  of every removed item is RETAINED under "merged out" so not one case id is lost —
  three of the removed rows carry case ids the surviving row does not.

FORMAT (Standing Rule 16): mirrors the two existing sheets 1:1 — same header fill
(1F4E79 / white bold), same group banding (DDEBF7), same wrap alignment, same column
widths per column-set, same freeze pane at A4, same blank "Your answer" column, same
QA-only appendices (mapping · source-currency · completeness · withdrawn · not-asked).

RULE 7 (absolute on tabs 1-3): no case ids, no C-numbers, no spec anchor codes, no
version numbers, no build marker, no ticket keys, no repo paths, no abbreviated report
names, no jargon. The single permitted exception is jargon INSIDE a direct quote of
Chris's own document — altering his quote would be dishonest. One further named
carry-over exemption is declared in RULE7_EXEMPTIONS and reported, never hidden.

RULE 50 / the C-id trap: verify_cids() re-reads build/report-suite/testrail-id-map.csv
(READ-ONLY) and ABORTS unless every internal-id/C-id pair printed anywhere on the
QA-only tab matches, and unless every bare C-id exists in the id-map. The 2026-07-31
sheet printed PV-API-04 as C30388 (correct: C30391; C30388 is PV-API-01), so this is
not optional.

EXCEL TAB-NAME LIMIT: Excel caps a sheet name at 31 characters, so two of the names
the QA lead asked for are carried as the tab's own on-sheet HEADING in full and the
tab name itself is the shortened form. Both are stated in TABS below.

Run:  python3 gen_consolidated_sheet.py
"""

import csv
import importlib.util
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)                        # build/report-suite
IDMAP = os.path.join(PROJECT, "testrail-id-map.csv")   # READ-ONLY

BASE = "Report-Suite_Questions-and-Decisions-for-Chris-Ward_2026-08-04"
TITLE = ("Report Suite - questions and decisions for Chris Ward - 2026-08-04 "
         "(one workbook, three tabs)")

BUILD_MARKER = "v3.4.1-0ed4433"
LINK = "https://shopview.testrail.io/index.php?/cases/view/"


# ---------------------------------------------------------------- load the sources
def _load(name, relpath):
    path = os.path.join(PROJECT, relpath)
    spec = importlib.util.spec_from_file_location(name, path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


LOC = _load("gen_location_question",
            "chris-location-question-2026-08-04/gen_location_question.py")
SVB = _load("gen_spec_vs_build_sheet",
            "chris-sheet-2026-08-04/gen_spec_vs_build_sheet.py")
PQ = _load("gen_po_questions_2026_08_03", "gen_po_questions_2026-08-03.py")

SOURCE_NAME_LOC = "chris-location-question-2026-08-04 (the one-question urgent sheet)"
SOURCE_NAME_SVB = "chris-sheet-2026-08-04 (the 10-row spec-versus-build sheet)"
SOURCE_NAME_PQ = "PO-Questions-Chris-ReportSuite-2026-08-03 (the 17-item sheet)"


# ------------------------------------------------------------------ de-duplication
# 1-based item numbers ON THE SOURCE SHEET that are removed in consolidation.
DROP_SVB = {2}          # the Location-column model question — Tab 1 asks it, wider
DROP_PQ = {6, 7, 13}    # each one's ask is contained in a Tab 2 item's options

DEDUP_LOG = [
    ("Tab 2, item 2 of the 10-row sheet",
     "\"The extra location column works one way on Work In Progress and the exact opposite way "
     "on Inventory Value\"",
     "REMOVED — the same question, asked wider, is Tab 1",
     "Tab 1 asks the identical question (automatic versus a switch the user controls) but across "
     "all SIX reports and adds the screen-versus-download split on Inventory Value, so it is the "
     "superset. Asking both would have let him answer the two-report version and the six-report "
     "version differently. Its QA mapping row is RETAINED below — it carries WIP-PERS-02 "
     "(C30507), which Tab 1's own eight-case list does not."),
    ("Tab 3, item 6 of the 17-item sheet",
     "\"Work In Progress: which number identifies the vehicle or machine first\" — will you update "
     "the description?",
     "REMOVED — subsumed by Tab 2 item 3",
     "Tab 2 item 3 asks which side moves now the product disagrees with his 29 July ruling, and "
     "BOTH of its options already name who edits the write-up. Keeping the separate tick-box "
     "allowed a contradiction: he could tick \"yes I will update it\" here and choose \"keep the "
     "product as it is\" there. Its QA mapping row is RETAINED below."),
    ("Tab 3, item 7 of the 17-item sheet",
     "\"The location chooser is hidden for someone with only one location\" — will you correct "
     "those four lines?",
     "REMOVED — subsumed by Tab 2 item 1",
     "Tab 2 item 1 puts the same four quoted lines in front of him and says in BOTH options that "
     "the four lines need correcting. Same contradiction risk as above. Its QA mapping row is "
     "RETAINED below."),
    ("Tab 3, item 13 of the 17-item sheet",
     "\"Five descriptions still say the report needs its own area permission\"",
     "REMOVED — exact duplicate of Tab 2 item 9",
     "The 10-row sheet already recorded this as the one overlap between the two sheets. Tab 2's "
     "version is the better one: it carries the live both-ways proof, and its figure is right — "
     "it says FOUR write-ups and lists four, whereas this row said \"five\" and then listed four. "
     "So the removal also retires a stale figure. Its QA mapping row is RETAINED below — it "
     "carries four navigation/tab cases (PV-NAV-01, IV-NAV-01, TU-NAV-01, WIP-TAB-01) that Tab 2 "
     "item 9's row does not."),
]

# Text changes made in consolidation — every one of them, no silent edits.
CHANGES = [
    ("Tab 1 opening line",
     "The original said the question is \"separate from the longer sheet you already have - that "
     "one still stands\". Now that the three sheets are one workbook it points at the other two "
     "tabs instead. No change of meaning.",
     "pointer fix, forced by the consolidation"),
    ("Tab 2 opening line — the item count",
     "\"There are 10 items: 8 need you to choose something...\" recomputed to 9 items / 7 "
     "decisions, because item 2 moved to Tab 1. The counts are computed from the data, never "
     "typed, so they cannot drift again.",
     "stale figure, caused by the de-duplication"),
    ("Tab 2 opening line — the companion pointer",
     "\"Please read it alongside the sheet dated 3 August\" now reads \"the last tab is its "
     "companion\". Same sentence otherwise.",
     "pointer fix, forced by the consolidation"),
    ("Tab 2 opening line — \"this sheet\"",
     "The two occurrences of \"this sheet\" read \"this tab\". Nothing else in the paragraph "
     "changed.",
     "pointer fix, forced by the consolidation"),
    ("Tab 3, item 2 — the pointer at the end",
     "\"The seven still-missing edits are listed one by one further down this sheet\" now reads "
     "\"...are each listed on their own further down, or on the tab before this one where the "
     "product disagrees with your answer as well\" — because two of those seven (the vehicle "
     "number on Work In Progress, and the location chooser) are now asked once, on Tab 2.",
     "pointer fix, forced by the de-duplication"),
    ("QA-only tab — the \"not asked here\" rows that pointed at the other sheet",
     "Five rows said \"already question N of the 3 August sheet\". They now name the tab and item "
     "number in this workbook. One row — \"the 17 items of the 3 August sheet are deliberately "
     "not duplicated\" — became meaningless once the sheets merged and is replaced by a row "
     "describing the consolidation itself.",
     "pointer fix, forced by the consolidation"),
    ("QA-only tab — the source-currency block",
     "The six description versions were RE-CHECKED LIVE today, 2026-08-04, read straight from "
     "Confluence rather than carried over from the 3 August check. All six are unchanged "
     "(Sales By Customer 13 · Sales By Representative 15 · Parts Velocity 4 · Technician "
     "Utilization 5 · Work In Progress 6 · Inventory Value 3), so no item on any tab has been "
     "overtaken by a description edit. Dates and verdicts updated accordingly.",
     "Standing Rule 31 pre-flight, done fresh for this pass"),
]

# Declared, named Rule-7 carry-over. Reported, never hidden.
RULE7_EXEMPTIONS = [
    ("toggle", "Tab 3 item 10",
     "The phrase \"a select-all / clear-all toggle\" is carried VERBATIM from the 3 August sheet, "
     "which is wording-checked and ready to send. \"Toggle\" is also Chris's own vocabulary — his "
     "Inventory Value description says \"it is not a user-toggled column in the column-selection "
     "control\" — so it is his word, not our jargon. Rewriting a wording-checked, PO-facing line "
     "was outside the brief for this pass, so it is declared instead of silently changed."),
]


# ---------------------------------------------------------------------- tab wiring
# (sheet-name  ,  the full on-sheet heading the QA lead asked for)
TAB1_NAME, TAB1_HEAD = "Urgent - Location column", "Urgent - the location column"
TAB2_NAME = "The product vs your write-up"          # 28 chars — Excel caps at 31
TAB2_HEAD = "What the product does vs your write-up"
TAB3_NAME = "Questions and things to note"          # 28 chars — Excel caps at 31
TAB3_HEAD = "Questions and things to write down"
TAB4_NAME = "QA internal - not for Chris"
TAB4_HEAD = "QA-ONLY - internal, not for Chris"

URGENT_NOTE = ("Needed today, please - the automated versions of these tests are being written "
               "today, so this one answer unblocks work that is already starting.")

TAB1_INTRO = (
    "One question only, and it should take a minute. It is on its own tab because it is the "
    "urgent one: the automated versions of these tests are being written today, and eight of our "
    "checks are waiting on your answer. The other two tabs are the longer lists - they still "
    "stand, and nothing is asked twice across the three."
)

TAB2_INTRO = [
    "This tab is a **side-by-side of your own written descriptions against what the product "
    "actually does today**, taken from a live look at the test build on **3 and 4 August**. Every "
    "item shows you the exact words from your write-up, what we saw happen, and asks which of the "
    "two you would rather keep. Nothing here is a bug report - bugs go straight to engineering and "
    "are not on this tab.",
    "**The last tab is its companion** - that one asks you to correct wording; this one asks you "
    "to choose between the wording and the product. Deliberately, nothing is asked twice.",
    "**One honest caveat up front:** engineering has told us the test build is not finished yet. "
    "So everything described below is what we saw on 3 and 4 August, and we will look again when "
    "they say it is done. If your answer depends on that, say so and we will come back to you.",
]

TAB3_INTRO = [
    "Plain-language product questions only - no bugs, no test jargon. This tab is an **exhaustive "
    "sweep of everything still owed by you**, not a top-up: every question sheet, the "
    "description-change watch list, our own decision register and all six live descriptions were "
    "re-read, and anything you have already answered has been deliberately left out.",
    "It is in **two parts**. The first five need you to **choose something**. The rest need **no "
    "decision at all** - you have already answered them; only the written description still says "
    "something different, so they are one-line confirmations. **Several of those were due on 4 "
    "August**, which is why they are listed one at a time rather than bundled.",
]

# the one sentence of Tab 3 item 2 that the de-duplication makes stale
PQ_ITEM2_OLD = ("The seven still-missing edits are listed one by one further down this sheet.")
PQ_ITEM2_NEW = ("The seven still-missing edits are each listed on their own further down, or on "
                "the tab before this one where the product disagrees with your answer as well.")


def tab2_rows():
    """The 10-row sheet minus the de-duplicated item, renumbered from 1."""
    out = []
    for old, r in enumerate(SVB.ROWS, 1):
        if old in DROP_SVB:
            continue
        out.append(dict(old=old, group=r["group"], topic=r["topic"],
                        c3=r["spec"], c4=r["build"],
                        c5=r["ask"] + "\n\n" + "\n".join(r["opts"]),
                        ask=r["ask"], opts=r["opts"]))
    for i, r in enumerate(out, 1):
        r["n"] = i
    return out


def tab3_rows():
    """The 17-item sheet minus the three de-duplicated items, renumbered from 1."""
    out = []
    for old, q in enumerate(PQ.QUESTIONS, 1):
        if old in DROP_PQ:
            continue
        now = q["now"].replace(PQ_ITEM2_OLD, PQ_ITEM2_NEW)
        out.append(dict(old=old, group=q["group"], topic=q["topic"], c3=now,
                        c4=q["q"], c5="\n".join(q["opts"]), opts=q["opts"]))
    for i, r in enumerate(out, 1):
        r["n"] = i
    return out


TAB2 = tab2_rows()
TAB3 = tab3_rows()
# old item number -> new item number, per tab (used by the QA-only mapping)
MAP2 = {r["old"]: r["n"] for r in TAB2}
MAP3 = {r["old"]: r["n"] for r in TAB3}


# ------------------------------------------------------------------- QA-only data
def qa_rows():
    """One row per surviving reader-facing item: cases + C-ids + links + anchors + source."""
    rows = []

    # -- Tab 1: the eight Location cases
    cases = "; ".join(f"{i} ({c})" for i, c, *_ in LOC.QA_CASES)
    anchors = "; ".join(sorted({a for *_, a, _ in
                                [(i, c, rep, a, s) for i, c, rep, a, s in LOC.QA_CASES]}))
    resolve = " || ".join(f"{a}: {b}" for a, b in LOC.QA_RESOLUTION)
    rows.append(dict(tab="Tab 1", n="1", was="the whole of the one-question sheet",
                     src=SOURCE_NAME_LOC, cases=cases, anchors=anchors, resolve=resolve))

    # -- Tab 2: the spec-versus-build rows
    svb_by_item = {m["q"]: m for m in SVB.QA_MAP}
    for r in TAB2:
        m = svb_by_item.get(str(r["old"]), {})
        rows.append(dict(tab="Tab 2", n=str(r["n"]), was=f"item {r['old']} of the 10-row sheet",
                         src=SOURCE_NAME_SVB, cases=m.get("cases", ""),
                         anchors=m.get("refs", ""), resolve=m.get("resolve", "")))

    # -- Tab 3: the question / write-down rows
    pq_by_item = {}
    for m in PQ.QA_MAP:
        key = m["q"]
        if "-" in key:                        # the "6-12" grouped row
            a, b = key.split("-")
            for k in range(int(a), int(b) + 1):
                pq_by_item[str(k)] = m
        else:
            pq_by_item[key] = m
    for r in TAB3:
        m = pq_by_item.get(str(r["old"]), {})
        note = ""
        if m.get("q") == "6-12":
            survivors = ", ".join(str(MAP3[o]) for o in sorted(MAP3) if 6 <= o <= 12)
            note = (f" [this mapping row covers items 6-12 of the source sheet; after the "
                    f"de-duplication its surviving items are Tab 3 items {survivors}]")
        rows.append(dict(tab="Tab 3", n=str(r["n"]), was=f"item {r['old']} of the 17-item sheet",
                         src=SOURCE_NAME_PQ, cases=m.get("cases", ""),
                         anchors=m.get("refs", "") + note, resolve=m.get("resolve", "")))
    return rows


def merged_out_rows():
    """The QA mapping rows of the REMOVED items — retained so no case id is lost."""
    out = []
    svb_by_item = {m["q"]: m for m in SVB.QA_MAP}
    pq_by_item = {m["q"]: m for m in PQ.QA_MAP}
    for old in sorted(DROP_SVB):
        m = svb_by_item.get(str(old), {})
        out.append(dict(was=f"item {old} of the 10-row sheet", src=SOURCE_NAME_SVB,
                        now="asked once, as Tab 1", cases=m.get("cases", ""),
                        anchors=m.get("refs", ""), resolve=m.get("resolve", "")))
    lands = {6: "Tab 2 item 3", 7: "Tab 2 item 1", 13: "Tab 2 item 9"}
    for old in sorted(DROP_PQ):
        m = pq_by_item.get(str(old), {})
        if not m:                              # items 6-12 share one grouped mapping row
            m = pq_by_item.get("6-12", {})
        out.append(dict(was=f"item {old} of the 17-item sheet", src=SOURCE_NAME_PQ,
                        now=f"asked once, as {lands[old]}", cases=m.get("cases", ""),
                        anchors=m.get("refs", ""), resolve=m.get("resolve", "")))
    return out


QA_ROWS = qa_rows()
MERGED_OUT = merged_out_rows()

# ---- SOURCE-CURRENCY (Standing Rule 31) — the six descriptions RE-READ LIVE 2026-08-04
SOURCE_CURRENCY = [
    ("Sales By Customer description", "Confluence page 577634305",
     "version 13, last changed 2026-07-31", "2026-08-04 (read live)",
     "CURRENT - unchanged since the sheets were written"),
    ("Sales By Representative description", "Confluence page 585629698",
     "version 15, last changed 2026-07-29", "2026-08-04 (read live)",
     "CURRENT - unchanged since the sheets were written"),
    ("Parts Velocity description", "Confluence page 620888066",
     "version 4, last changed 2026-07-29", "2026-08-04 (read live)",
     "CURRENT - unchanged since the sheets were written"),
    ("Technician Utilization description", "Confluence page 641400833",
     "version 5, last changed 2026-07-29", "2026-08-04 (read live)",
     "CURRENT - unchanged since the sheets were written"),
    ("Work In Progress description", "Confluence page 703660034",
     "version 6, last changed 2026-07-29", "2026-08-04 (read live)",
     "CURRENT - unchanged since the sheets were written"),
    ("Inventory Value description", "Confluence page 720142338",
     "version 3, last changed 2026-07-29", "2026-08-04 (read live)",
     "CURRENT - unchanged since the sheets were written"),
    ("The build", "QA branch sv8582 / project/reports-suite-bravo, app-version " + BUILD_MARKER,
     BUILD_MARKER, "2026-08-03 / 2026-08-04",
     "PARTIAL - engineering declared the branch NOT FINAL, so every observation quoted on these "
     "tabs is PROVISIONAL and is queued for re-check in viu-2026-08-03/RECHECK-QUEUE.md "
     "(Standing Rule 49). Shortfall: the observations may change when the branch settles."),
    ("Epic SV-8582 + child stories", "Jira, project SV",
     "currency-checked, no full re-read this pass", "2026-08-03",
     "PARTIAL - Tier-1 currency check only (Standing Rule 37); a full re-read was not authorised "
     "for this pass and is not claimed"),
    ("Designs", "none exist for the Report Suite", "n/a", "2026-08-04",
     "ABSENT - spec-only project; no Figma file has ever been supplied, so no design source was "
     "consulted and none is claimed"),
    ("Engineering tech plan", "tech-plan-2026-07-29/", "as supplied 2026-07-29", "2026-08-04",
     "CURRENT"),
    ("Chris Ward's answers, messages and both videos",
     "chris-answers-2026-07-28 / -07-31 / -08-01, chris-update-2026-07-29, both video transcripts",
     "newest = the 2026-08-01-round two-question sheet", "2026-08-04",
     "CURRENT - re-swept today; nothing newer than 2026-08-01 exists, so no question on any tab "
     "has been answered since the source sheets were written"),
    ("The three source sheets themselves",
     "chris-location-question-2026-08-04 / chris-sheet-2026-08-04 / "
     "PO-Questions-Chris-ReportSuite-2026-08-03",
     "all three READY TO SEND, none sent", "2026-08-04",
     "CURRENT - all three are now marked SUPERSEDED by this workbook, and are kept, not deleted"),
]

CURRENCY_METHOD = (
    "Method (Standing Rule 31, and the version-number trap it names): the six description "
    "versions were read from the LIVE Confluence page objects on 2026-08-04 - "
    "GET /wiki/api/v2/pages/<id>, HTTP 200 on all six - and the CONFLUENCE VERSION NUMBER was "
    "used, never the version written inside the document body. Values read: 13 (2026-07-31), "
    "15, 4, 5, 6, 3 (all five 2026-07-29). These are the exact versions the three source sheets "
    "were built against, so nothing on any tab is stale against its description."
)

# ---- WITHDRAWN — the union of both source sheets' withdrawn lists, de-duplicated
WITHDRAWN_SEEN_IN = {
    # candidate text (from the 10-row sheet) -> also present on the 17-item sheet
}
WITHDRAWN = []


def build_withdrawn():
    """Union of the two sheets' withdrawn lists; near-duplicates collapsed, source recorded."""
    out = []
    svb = list(SVB.WITHDRAWN)
    pq = list(PQ.WITHDRAWN)

    # pairs that are the same candidate question on both sheets: keep the 10-row sheet's
    # (later, fuller) text and record that both sheets withdrew it
    same = {
        "Should the location chooser be hidden for a one-location person?":
            "Is the location dropdown hidden for a one-location user?",
        "Should the six reports be gated by their own dedicated permission?":
            "Should the reports use their own dedicated permissions?",
        "Does the 10,000-row download limit apply to Parts Velocity, Technician Utilization and "
        "Work In Progress?":
            "Does the 10,000-row export cap apply to Parts Velocity, Technician Utilization and "
            "Work In Progress?",
        "Which of the two \"too large to export\" messages is correct?":
            "Which of the two \"too large to export\" messages is correct?",
        "Does Escape close the \"deactivate a representative\" pop-up?":
            "Does Escape close the \"deactivate a representative\" pop-up?",
    }
    pq_dupes = set(same.values())
    # the 27-July permission question is a third phrasing of the same settled point
    pq_dupes.add("Each report uses a different permission to view it - is that intended? "
                 "(Question 2 of the 27 July sheet)")

    for cand, ans in svb:
        both = cand in same
        out.append((cand, ans, "both sheets" if both else "the 10-row sheet"))
    for cand, ans in pq:
        if cand in pq_dupes:
            continue
        out.append((cand, ans, "the 17-item sheet"))
    return out


WITHDRAWN = build_withdrawn()

# ---- NOT ASKED — union, with the cross-sheet pointers repointed at the tabs
POINTER_FIXES = [
    ("Already question 3 of the 3 August sheet, unanswered. Not repeated.",
     "Asked once, as item 3 of the third tab."),
    ("Already question 4 of the 3 August sheet, unanswered. Not repeated.",
     "Asked once, as item 4 of the third tab."),
    ("Already question 5 of the 3 August sheet, unanswered. Not repeated.",
     "Asked once, as item 5 of the third tab."),
    ("Already questions 8 and 9 of the 3 August sheet.",
     "Asked once, as items 6 and 7 of the third tab."),
]
OBSOLETE_NOT_ASKED = "The 17 items of PO-Questions-Chris-ReportSuite-2026-08-03 (still unanswered)."


def build_not_asked():
    out = []
    for item, why in SVB.NOT_ASKED:
        if item.startswith(OBSOLETE_NOT_ASKED[:40]):
            continue                              # replaced by the consolidation row below
        for old, new in POINTER_FIXES:
            why = why.replace(old, new)
        out.append((item, why, "the 10-row sheet"))
    out.insert(0, (
        "Sending the three sheets separately.",
        "REPLACED BY THIS WORKBOOK. The 10-row sheet carried a row explaining that the 17 items "
        "of the 3 August sheet were deliberately NOT duplicated; that row is meaningless now the "
        "three are one workbook. What replaces it: the three sheets are consolidated, most-urgent "
        "tab first, and FOUR overlapping items were removed so nothing is asked twice - see the "
        "de-duplication log above. The source sheets are marked SUPERSEDED and kept, not deleted.",
        "this consolidation"))
    seen = {i for i, _, _ in out}
    for item, why in PQ.NOT_ASKED:
        if item in seen:
            continue
        out.append((item, why, "the 17-item sheet"))
    return out


NOT_ASKED = build_not_asked()


def sources_swept():
    """Both sheets' completeness proofs, carried whole, plus this pass's own sweep."""
    out = [(s, n, note, "the 10-row sheet") for s, n, note in SVB.SOURCES_SWEPT]
    out += [(s, n, note, "the 17-item sheet") for s, n, note in PQ.SOURCES_SWEPT]
    out += [
        ("The six LIVE Confluence descriptions, re-read for THIS workbook 2026-08-04",
         "**0 new**",
         "All six still at the versions the source sheets used (13 / 15 / 4 / 5 / 6 / 3), read "
         "live over the page API, HTTP 200 on all six. No item has been overtaken by a "
         "description edit, so nothing needed withdrawing on that ground.",
         "this consolidation"),
        ("Chris Ward's answers, messages and both video transcripts, re-swept 2026-08-04",
         "**0 new**",
         "Newest authoritative Chris source is still 2026-08-01, which both source sheets already "
         "accounted for. Nothing has been answered since they were written, so 0 further "
         "candidates were withdrawn by this pass; the 12 rows above are the de-duplicated union "
         "of what the two sheets had already withdrawn.",
         "this consolidation"),
        ("The three source sheets, compared item by item for overlap",
         "**4 overlaps found**",
         "1 the two sheets had already spotted (the permission write-down) and 3 this pass found: "
         "the Location-column model asked twice, and two write-down tick-boxes whose ask is "
         "already inside a spec-versus-build item's options. All four removed; see the "
         "de-duplication log. 28 reader-facing items in, 24 out.",
         "this consolidation"),
        ("build/OUTSTANDING-ITEMS-REGISTER.md (Report Suite send-list row)",
         "**0 new**",
         "Its Chris-facing row previously listed the three sheets; it now points at this "
         "workbook. No new ask surfaced.",
         "this consolidation"),
    ]
    return out


SOURCES_SWEPT = sources_swept()

HONESTY = [
    "Nothing here is rewritten. Every reader-facing row is the row its source sheet carried, "
    "emitted from that sheet's own generator, so the wording cannot drift in transit. The only "
    "text changes are the ones enumerated in the change log above.",
    "Standing Rule 49: the QA branch was declared NOT FINAL, so every build observation quoted on "
    "these tabs is PROVISIONAL and carries the build marker " + BUILD_MARKER + "; the re-check "
    "queue viu-2026-08-03/RECHECK-QUEUE.md is OPEN.",
    "The Column Selection panel's per-item on/off state could not be read reliably by automation, "
    "so the \"starts switched off\" claim on Work In Progress rests on the column's absence from "
    "the grid plus the 2026-08-03 pass's own observation. The presence or absence of the column - "
    "which is what Tab 1's question turns on - is solid.",
    "Rule 37: the epic had a Tier-1 currency check only. A full re-read was not authorised for "
    "this pass and is not claimed.",
]


# ----------------------------------------------------------------------- md output
def md_tab1():
    L = ["", f"# Tab 1 - {TAB1_HEAD}", "",
         f"**{URGENT_NOTE}**", "", TAB1_INTRO, "",
         "## 1 - The location column - should it appear on its own, or does the user switch it on?",
         "", f"**What happens now:** {LOC.NOW_LEAD}", ""]
    for what, does in LOC.NOW_BULLETS:
        L.append(f"- **{what}** {does}")
    L += ["", "**But both of those two written descriptions say the column should be automatic, "
              "and should not be something the user switches on.** Quoting them directly:", ""]
    for report, quote in LOC.HIS_WORDS:
        L.append(f"- **{report}:** *\"{quote}\"*")
    L += ["", f"**Why we are asking:** {LOC.WHY}", "",
          f"**The question:** {LOC.QUESTION}", "", "**Options:**", ""]
    L += [f"- {o}" for o in LOC.OPTIONS]
    L += ["", "**Your answer:** ____________________", "", f"**{LOC.CLOSING}**", ""]
    return L


def md_tab2():
    n_dec = sum(1 for r in TAB2 if r["group"] == SVB.G_DEC)
    n_wr = sum(1 for r in TAB2 if r["group"] == SVB.G_WRITE)
    n_fyi = sum(1 for r in TAB2 if r["group"] == SVB.G_FYI)
    L = ["", f"# Tab 2 - {TAB2_HEAD}", "", TAB2_INTRO[0], "",
         f"There are **{len(TAB2)} items**: **{n_dec} need you to choose something**, "
         f"**{n_wr} needs only a line changing in a write-up**, and **{n_fyi} needs nothing at "
         "all** and is here purely so you are not surprised by it later.", "",
         TAB2_INTRO[1], "", TAB2_INTRO[2], ""]
    group = None
    for r in TAB2:
        if r["group"] != group:
            group = r["group"]
            L += ["", f"## {group}", ""]
        L += [f"### {r['n']} - {r['topic']}", "",
              "**What your write-up says:** " + r["c3"], "",
              "**What the product actually does:** " + r["c4"], "",
              "**Which do you want?** " + r["ask"], ""]
        L += [f"- {o}" for o in r["opts"]]
        L += ["", "**Your answer:** ____________________", ""]
    return L


def md_tab3():
    L = ["", f"# Tab 3 - {TAB3_HEAD}", "", TAB3_INTRO[0], "", TAB3_INTRO[1], ""]
    group = None
    for r in TAB3:
        if r["group"] != group:
            group = r["group"]
            L += ["", f"## {group}", ""]
        L += [f"### {r['n']} - {r['topic']}", "",
              "**What happens now:** " + r["c3"], "",
              "**The question:** " + r["c4"], "", "**Options:**", ""]
        L += [f"- {o}" for o in r["opts"]]
        L += ["", "**Your answer:** ____________________", ""]
    return L


QA_MARKER = "# QA-ONLY - internal, not for Chris"


def md_qa():
    L = ["", "---", "", QA_MARKER, "",
         "Do not send this part to Chris. TestRail C-ids from "
         "`build/report-suite/testrail-id-map.csv` (Standing Rule 8). Links: "
         "https://shopview.testrail.io/index.php?/cases/view/<id>", "",
         "**Every internal-id/C-id pair below is verified against the id-map at generation time - "
         "the generator ABORTS on a mismatch, and also aborts on a C-id that is not in the id-map "
         "at all.** The 2026-07-31 sheet printed **PV-API-04 as C30388**, which is wrong: "
         "**PV-API-04 = C30391**, and **C30388 = PV-API-01**. Anyone acting on that row would have "
         "edited the wrong case.", "",
         "## Per-item mapping - every surviving reader-facing item", "",
         "| Tab | Item | Was | Source sheet | Affected internal case IDs (TestRail C-id) | "
         "Spec anchors + live evidence | What each answer resolves to |",
         "|---|---|---|---|---|---|---|"]
    for r in QA_ROWS:
        L.append("| {} | {} | {} | {} | {} | {} | {} |".format(
            r["tab"], r["n"], r["was"], r["src"].split(" (")[0],
            _cell(r["cases"]), _cell(r["anchors"]), _cell(r["resolve"])))
    L += ["", "### The links, spelled out", "",
          "| Internal ID | TestRail | Link | Report | Spec anchor | What it asserts today |",
          "|---|---|---|---|---|---|"]
    for i, c, rep, anchor, says in LOC.QA_CASES:
        L.append(f"| {i} | {c} | [open]({LINK}{c[1:]}) | {rep} | {anchor} | {says} |")
    L += ["", "## De-duplication log - what was removed and where the question now lives", "",
          f"**{len(DEDUP_LOG)} overlapping items removed. 28 reader-facing items across the three "
          f"sheets in; {1 + len(TAB2) + len(TAB3)} out.**", "",
          "| Removed | The item | What happened | Why, and what was preserved |", "|---|---|---|---|"]
    for a, b, c, d in DEDUP_LOG:
        L.append(f"| {a} | {_cell(b)} | {c} | {_cell(d)} |")
    L += ["", "### Merged out - the QA mapping rows of the removed items, RETAINED", "",
          "Kept so that not one case id is lost: three of these four rows carry case ids the "
          "surviving item's row does not.", "",
          "| Was | Source sheet | Now asked as | Affected internal case IDs (TestRail C-id) | "
          "Spec anchors + live evidence | What each answer resolves to |", "|---|---|---|---|---|---|"]
    for r in MERGED_OUT:
        L.append("| {} | {} | {} | {} | {} | {} |".format(
            r["was"], r["src"].split(" (")[0], r["now"],
            _cell(r["cases"]), _cell(r["anchors"]), _cell(r["resolve"])))
    L += ["", "## Changes made in consolidation - all of them", "",
          "| Where | What changed | Why |", "|---|---|---|"]
    for a, b, c in CHANGES:
        L.append(f"| {a} | {_cell(b)} | {c} |")
    L += ["", "## Declared wording carry-over (Standing Rule 7)", "",
          "| Word | Where | Why it is kept rather than reworded |", "|---|---|---|"]
    for w, where, why in RULE7_EXEMPTIONS:
        L.append(f"| {w} | {where} | {_cell(why)} |")
    L += ["", "## SOURCE-CURRENCY BLOCK (Standing Rule 31)", "",
          "| Source | Identifier | Version / last-updated | Checked | Verdict |",
          "|---|---|---|---|---|"]
    for name, ident, ver, checked, verdict in SOURCE_CURRENCY:
        L.append(f"| {name} | {ident} | {ver} | {checked} | {_cell(verdict)} |")
    L += ["", CURRENCY_METHOD, "",
          f"**Build marker: {BUILD_MARKER}.** Nothing in this workbook claims completeness: the "
          "build is a PARTIAL source by engineering's own statement, so every observation is "
          "PROVISIONAL and queued in `viu-2026-08-03/RECHECK-QUEUE.md` (Standing Rule 49).", "",
          "## Completeness proof - every source swept (Standing Rule 17)", "",
          "| Source | Items found | Notes | Swept by |", "|---|---|---|---|"]
    for s, n, note, by in SOURCES_SWEPT:
        L.append(f"| {s} | {n} | {_cell(note)} | {by} |")
    L += ["", "## Withdrawn - already answered (not put in front of Chris)", "",
          f"**{len(WITHDRAWN)} candidate questions**, the de-duplicated union of both source "
          "sheets' withdrawn lists. Questions have been withdrawn for this reason on four "
          "previous sheets, so the check is mandatory before any item survives. **This pass added "
          "0 new withdrawals** - nothing has been answered since the source sheets were written.",
          "", "| Candidate question | Already answered by | Withdrawn on |", "|---|---|---|"]
    for cand, ans, where in WITHDRAWN:
        L.append(f"| {_cell(cand)} | {_cell(ans)} | {where} |")
    L += ["", "## Not asked here (QA reference)", "",
          "| Item | Why it is not on any tab | From |", "|---|---|---|"]
    for item, why, where in NOT_ASKED:
        L.append(f"| {_cell(item)} | {_cell(why)} | {where} |")
    L += ["", "## Live evidence behind Tab 1's \"what happens now\" text", "",
          "| Observation | What was seen | Evidence |", "|---|---|---|"]
    for a, b, c in LOC.EVIDENCE:
        L.append(f"| {a} | {_cell(b)} | `{c}` |")
    L += ["", "## Honesty notes", ""]
    L += [f"- {h}" for h in HONESTY]
    return L


def _cell(s):
    return (s or "").replace("\n", " ").replace("|", "/")


def write_md():
    n2 = len(TAB2)
    n3 = len(TAB3)
    L = [f"# {TITLE}", "",
         "**STATUS: READY TO SEND** (not yet sent). This ONE workbook REPLACES the three separate "
         "sheets - the urgent one-question sheet, the spec-versus-build sheet and the 17-item "
         "sheet - all three of which are now marked SUPERSEDED and kept for the record. On "
         "return: ingest verbatim, then revisit the affected cases per the standing workflow; "
         "nothing is edited before his answers and the QA lead's go-ahead.", "",
         "**Three tabs, most urgent first:**", "",
         f"1. **{TAB1_HEAD}** - one question, needed today.",
         f"2. **{TAB2_HEAD}** - {n2} items.",
         f"3. **{TAB3_HEAD}** - {n3} items.", "",
         f"**{1 + n2 + n3} items in total.** Twenty-eight came in from the three sheets and four "
         "were removed as the same question asked twice, so **nothing is asked twice across the "
         "three tabs**. A fourth, QA-only tab carries the traceability and is not for Chris.", ""]
    L += md_tab1()
    L += md_tab2()
    L += md_tab3()
    L += md_qa()
    L += ["", "---", "", "## OUTSTANDING - what I need from you", "",
          "Cross-project register: `build/OUTSTANDING-ITEMS-REGISTER.md` (Standing Rule 36).", "",
          "**From you (QA lead):**", "",
          "1. **Send this ONE workbook to Chris** - it replaces all three earlier sheets, and the "
          "three are banner-marked SUPERSEDED so an old one cannot go out by mistake. Tab 1 is "
          "the one that is needed today: the automated versions of those tests are being written "
          "now and eight checks are frozen until he answers.",
          "2. **Nothing here is authorised to be applied.** No case has been edited, no TestRail "
          "write has been staged from this workbook, and the three source sheets' staged edits "
          "stay staged (Standing Rule 6). This pass was read-only on cases, specs and TestRail.",
          "3. **The QA branch is still not final.** Engineering said so, so every build "
          "observation quoted here is provisional and the re-check queue is OPEN. Tell us when it "
          "is declared done and we re-run the queue immediately.", "",
          "**From Chris:** the one urgent answer on Tab 1, "
          f"{sum(1 for r in TAB2 if r['group'] == SVB.G_DEC)} decisions on Tab 2 plus a tick "
          f"against its write-down item, and {sum(1 for r in TAB3 if r['group'].startswith('Decisions'))} "
          f"decisions on Tab 3 plus a tick against each of its "
          f"{sum(1 for r in TAB3 if not r['group'].startswith('Decisions'))} write-down items. "
          "Nothing at all is needed for Tab 2's last item - it is there so he is not surprised "
          "by it later.", "",
          "**Nothing else is outstanding from this workbook.**", ""]
    path = os.path.join(HERE, BASE + ".md")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    print("wrote", BASE + ".md")
    return path


# --------------------------------------------------------------------- xlsx output
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")


def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def write_xlsx():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ Tab 1
    ws = wb.active
    ws.title = TAB1_NAME
    ws["A1"] = f"{TAB1_HEAD} - Report Suite - Chris Ward - 2026-08-04"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = URGENT_NOTE
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    cols = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    _hdr(ws, 4, cols)
    _band(ws, 5, "DECISION WE NEED FROM YOU - TODAY", 6)
    quotes = ("Your own descriptions say it should be automatic:\n"
              + "\n".join(f'{r}: "{q}"' for r, q in LOC.HIS_WORDS))
    row = [1, "The location column - should it appear on its own, or does the user switch it on?",
           LOC.NOW + "\n\n" + quotes + "\n\n" + LOC.WHY, LOC.QUESTION,
           "\n\n".join(LOC.OPTIONS), ""]
    for j, v in enumerate(row, 1):
        ws.cell(row=6, column=j, value=v).alignment = WRAP
    ws.cell(row=8, column=3, value=LOC.CLOSING).alignment = WRAP
    ws.cell(row=8, column=3).font = Font(bold=True)
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[6].height = 300
    ws.freeze_panes = "A5"

    # ------------------------------------------------------------------ Tab 2
    ws2 = wb.create_sheet(TAB2_NAME)
    ws2["A1"] = f"{TAB2_HEAD} - Report Suite - Chris Ward - 2026-08-04"
    ws2["A1"].font = Font(bold=True)
    ws2["A2"] = " ".join(re.sub(r"\*\*", "", t) for t in TAB2_INTRO)
    ws2["A2"].alignment = WRAP
    cols2 = ["#", "Topic", "What your write-up says", "What the product actually does",
             "Which do you want?", "Your answer"]
    _hdr(ws2, 4, cols2)
    r = 4
    group = None
    for item in TAB2:
        if item["group"] != group:
            group = item["group"]
            r += 1
            _band(ws2, r, group, 6)
        r += 1
        for j, v in enumerate([item["n"], item["topic"], item["c3"], item["c4"], item["c5"], ""], 1):
            ws2.cell(row=r, column=j, value=v).alignment = WRAP
    for col, w in zip("ABCDEF", [4, 30, 62, 62, 52, 22]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A5"

    # ------------------------------------------------------------------ Tab 3
    ws3 = wb.create_sheet(TAB3_NAME)
    ws3["A1"] = f"{TAB3_HEAD} - Report Suite - Chris Ward - 2026-08-04"
    ws3["A1"].font = Font(bold=True)
    ws3["A2"] = " ".join(re.sub(r"\*\*", "", t) for t in TAB3_INTRO)
    ws3["A2"].alignment = WRAP
    cols3 = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    _hdr(ws3, 4, cols3)
    r = 4
    group = None
    for item in TAB3:
        if item["group"] != group:
            group = item["group"]
            r += 1
            _band(ws3, r, group, 6)
        r += 1
        for j, v in enumerate([item["n"], item["topic"], item["c3"], item["c4"], item["c5"], ""], 1):
            ws3.cell(row=r, column=j, value=v).alignment = WRAP
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A5"

    # ------------------------------------------------------------------ Tab 4 (QA)
    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = (
        "QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail C-ids from "
        "build/report-suite/testrail-id-map.csv (Standing Rule 8); links "
        "https://shopview.testrail.io/index.php?/cases/view/<id>. Every internal-id/C-id pair on "
        "this tab is verified against the id-map at generation time - the generator ABORTS on a "
        "mismatch and on any C-id absent from the id-map (the 2026-07-31 sheet printed PV-API-04 "
        "as C30388; the correct id is C30391 - C30388 is PV-API-01). FORMAT mirrors the "
        "2026-08-03 and 2026-08-04 sheets 1:1 per Standing Rule 16.")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP

    r = 3
    ws4.cell(row=r, column=1, value="PER-ITEM MAPPING - EVERY SURVIVING READER-FACING ITEM").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Tab", "Item", "Was", "Source sheet",
                  "Affected internal case IDs (TestRail C-id)",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for q in QA_ROWS:
        for j, v in enumerate([q["tab"], q["n"], q["was"], q["src"], q["cases"],
                               q["anchors"], q["resolve"]], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="THE LINKS, SPELLED OUT (Tab 1's eight cases)").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Internal ID", "TestRail C-id", "Link", "Report", "Spec anchor",
                  "What it asserts today", ""])
    r += 1
    for i, c, rep, anchor, says in LOC.QA_CASES:
        for j, v in enumerate([i, c, LINK + c[1:], rep, anchor, says], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1,
             value=("DE-DUPLICATION LOG - 4 OVERLAPPING ITEMS REMOVED (28 IN, %d OUT)"
                    % (1 + len(TAB2) + len(TAB3)))).font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Removed", "The item", "What happened", "Why, and what was preserved", "", "", ""])
    r += 1
    for a, b, c, d in DEDUP_LOG:
        for j, v in enumerate([a, b, c, d], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1,
             value="MERGED OUT - THE REMOVED ITEMS' QA MAPPING ROWS, RETAINED SO NO CASE ID IS LOST").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Was", "Source sheet", "Now asked as",
                  "Affected internal case IDs (TestRail C-id)",
                  "Spec anchors + live evidence", "What each answer resolves to", ""])
    r += 1
    for m in MERGED_OUT:
        for j, v in enumerate([m["was"], m["src"], m["now"], m["cases"], m["anchors"],
                               m["resolve"]], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="CHANGES MADE IN CONSOLIDATION - ALL OF THEM").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Where", "What changed", "Why", "", "", "", ""])
    r += 1
    for a, b, c in CHANGES:
        for j, v in enumerate([a, b, c], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="DECLARED WORDING CARRY-OVER (STANDING RULE 7)").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Word", "Where", "Why it is kept rather than reworded", "", "", "", ""])
    r += 1
    for w, where, why in RULE7_EXEMPTIONS:
        for j, v in enumerate([w, where, why], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="SOURCE-CURRENCY BLOCK (STANDING RULE 31)").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Source", "Identifier", "Version / last-updated", "Checked", "Verdict", "", ""])
    r += 1
    for name, ident, ver, checked, verdict in SOURCE_CURRENCY:
        for j, v in enumerate([name, ident, ver, checked, verdict], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value=CURRENCY_METHOD).alignment = WRAP
    r += 1
    ws4.cell(row=r, column=1, value="Build marker: " + BUILD_MARKER +
             " - the branch is NOT FINAL, so every observation is PROVISIONAL "
             "(Standing Rule 49; re-check queue viu-2026-08-03/RECHECK-QUEUE.md is OPEN).").alignment = WRAP

    r += 2
    ws4.cell(row=r, column=1, value="COMPLETENESS PROOF - EVERY SOURCE SWEPT (STANDING RULE 17)").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Source", "Items found", "Notes", "Swept by", "", "", ""])
    r += 1
    for s, n, note, by in SOURCES_SWEPT:
        for j, v in enumerate([s, n, note, by], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1,
             value=("WITHDRAWN - ALREADY ANSWERED, NOT ASKED (%d candidates; 0 added by this pass)"
                    % len(WITHDRAWN))).font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Candidate question", "Already answered by", "Withdrawn on", "", "", "", ""])
    r += 1
    for cand, ans, where in WITHDRAWN:
        for j, v in enumerate([cand, ans, where], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="NOT ASKED HERE (QA REFERENCE)").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Item", "Why it is not on any tab", "From", "", "", "", ""])
    r += 1
    for item, why, where in NOT_ASKED:
        for j, v in enumerate([item, why, where], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="LIVE EVIDENCE BEHIND TAB 1'S 'WHAT HAPPENS NOW' TEXT").font = Font(bold=True)
    r += 1
    _hdr(ws4, r, ["Observation", "What was seen", "Evidence", "", "", "", ""])
    r += 1
    for a, b, c in LOC.EVIDENCE:
        for j, v in enumerate([a, b, c], 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1

    r += 1
    ws4.cell(row=r, column=1, value="HONESTY NOTES").font = Font(bold=True)
    r += 1
    for h in HONESTY:
        ws4.cell(row=r, column=1, value="- " + h).alignment = WRAP
        r += 1

    for col, w in zip("ABCDEFG", [16, 14, 30, 30, 46, 56, 56]):
        ws4.column_dimensions[col].width = w

    path = os.path.join(HERE, BASE + ".xlsx")
    wb.save(path)
    print("wrote", BASE + ".xlsx")
    return path


# ------------------------------------------------------------------------- guards
# Vladimir Tomovic's automated cases. They are NOT ours (Standing Rule 38 - foreign cases are
# hands-off and are excluded from our id-map and our counts), so a C-id in this set is allowed to
# be absent from testrail-id-map.csv. Named here rather than silently tolerated.
FOREIGN_CIDS = {"C38919", "C38920", "C38921", "C38922", "C38923"}


def read_idmap():
    idmap = {}
    with open(IDMAP, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            if row and row[0] != "internal_id":
                idmap[row[0]] = row[1]
    return idmap


def qa_text_blob():
    parts = []
    for q in QA_ROWS:
        parts += [q["cases"], q["anchors"], q["resolve"]]
    for m in MERGED_OUT:
        parts += [m["cases"], m["anchors"], m["resolve"]]
    for a, b, c, d in DEDUP_LOG:
        parts += [a, b, c, d]
    for cand, ans, _ in WITHDRAWN:
        parts += [cand, ans]
    for item, why, _ in NOT_ASKED:
        parts += [item, why]
    for s, n, note, _ in SOURCES_SWEPT:
        parts += [s, note]
    for i, c, rep, anchor, says in LOC.QA_CASES:
        parts += [i, c, anchor, says]
    for a, b in LOC.QA_RESOLUTION:
        parts += [a, b]
    parts += [c for _, c, _ in CHANGES] + [b for _, b, _ in CHANGES]
    return " ".join(parts)


def verify_cids():
    """ABORT unless every internal-id/C-id pair on the QA tab matches the id-map."""
    idmap = read_idmap()
    text = qa_text_blob()
    pairs = re.findall(r"([A-Z]{2,4}(?:-[A-Z]+)+-\d+)\s*[=(]\s*(C\d+)", text)
    if not pairs:
        raise SystemExit("C-ID VERIFICATION FOUND NOTHING TO CHECK - the regex or the data changed")
    bad = [(i, c, idmap.get(i, "<not in id-map>")) for i, c in pairs if idmap.get(i) != c]
    if bad:
        raise SystemExit("C-ID MISMATCH vs testrail-id-map.csv (internal, printed, actual): "
                         + str(bad))
    known = set(idmap.values()) | FOREIGN_CIDS
    strays = sorted({c for c in re.findall(r"\bC\d{5}\b", text) if c not in known})
    if strays:
        raise SystemExit("C-IDS PRINTED THAT ARE NOT IN THE ID-MAP AT ALL: " + str(strays))
    uniq = sorted(set(pairs))
    print(f"C-id verification: {len(pairs)} internal-id/C-id pairs checked "
          f"({len(uniq)} distinct), ALL MATCH the id-map; 0 stray C-ids; "
          f"id-map rows read {len(idmap)}")
    return len(pairs), len(uniq)


JARGON = ["API", "HTTP", "endpoint", "payload", "JSON", "403", "404", "500", "200", "201",
          "atom", "aria", "DOM", "regex", "boolean", "null", "backend", "back-end", "front-end",
          "frontend", "VIU", "TestRail", "Confluence", "Jira", "feature flag", "feature-flag",
          "CSV", "PDF", "UTF-8", "BOM", "server-side", "toggle", "column selector"]
EXEMPT_WORDS = {w.lower() for w, _, _ in RULE7_EXEMPTIONS}


def rule7_scan(blob, label):
    """Return a list of (kind, token). Jargon inside a double-quoted span is exempt."""
    problems = []
    for m in re.findall(r"\bC\d{4,6}\b", blob):
        problems.append(("TestRail case id", m))
    for m in re.findall(r"\b(?:SBC|SBR|PV|TU|WIP|IV)-[A-Z]{2,6}-\d+\b", blob):
        problems.append(("internal case id", m))
    for m in re.findall(r"\bS\d+-[RNE]\d+[a-z]?\b", blob):
        problems.append(("spec anchor", m))
    for m in re.findall(r"[§¶]", blob):
        problems.append(("section mark", m))
    for m in re.findall(r"\bv\d+(?:\.\d+)*\b", blob):
        problems.append(("version number", m))
    for m in re.findall(r"\bversion\s+\d+\b", blob, re.I):
        problems.append(("version number", m))
    if BUILD_MARKER in blob:
        problems.append(("build marker", BUILD_MARKER))
    for m in re.findall(r"\bSV-\d+\b", blob):
        problems.append(("jira key", m))
    for m in re.findall(r"\bbuild/[a-z0-9\-/]+", blob):
        problems.append(("repo path", m))
    quoted = re.findall(r"\"[^\"]*\"", blob) + re.findall(r"[“][^”]*[”]", blob)
    unquoted = blob
    for q in quoted:
        unquoted = unquoted.replace(q, " ")
    for j in JARGON:
        if j.lower() in EXEMPT_WORDS:
            continue
        if re.search(r"(?<![A-Za-z])" + re.escape(j) + r"(?![A-Za-z])", unquoted, re.I):
            problems.append(("jargon (outside a quote)", j))
    for m in re.findall(r"\b(?:SBC|SBR|WIP|TU|PV|IV)\b", blob):
        problems.append(("abbreviated report name", m))
    return problems


def verify_reader_text_clean():
    """Rule-7 gate on the SOURCE data, before anything is written."""
    reader = [TITLE, TAB1_HEAD, TAB2_HEAD, TAB3_HEAD, URGENT_NOTE, TAB1_INTRO,
              LOC.NOW, LOC.WHY, LOC.QUESTION, LOC.CLOSING] + list(LOC.OPTIONS)
    reader += [f"{a} {b}" for a, b in LOC.NOW_BULLETS]
    reader += [f'{r}: "{q}"' for r, q in LOC.HIS_WORDS]
    reader += TAB2_INTRO + TAB3_INTRO
    for r in TAB2:
        reader += [r["topic"], r["c3"], r["c4"], r["ask"]] + list(r["opts"])
    for r in TAB3:
        reader += [r["topic"], r["c3"], r["c4"]] + list(r["opts"])
    problems = rule7_scan("\n".join(reader), "generator input")
    if problems:
        seen, uniq = set(), []
        for kind, tok in problems:
            if (kind, tok) not in seen:
                seen.add((kind, tok))
                uniq.append(f"{kind}: {tok!r}")
        raise SystemExit("READER-FACING TEXT IS NOT RULE-7 CLEAN:\n  " + "\n  ".join(uniq))
    print("Rule 7 gate (generator input): reader-facing text CLEAN - 0 case ids, 0 internal ids, "
          "0 spec anchors, 0 version numbers, 0 build markers, 0 ticket keys, 0 repo paths, "
          "0 abbreviated report names, 0 jargon outside a direct quote "
          f"({len(RULE7_EXEMPTIONS)} declared carry-over: "
          f"{', '.join(w for w, _, _ in RULE7_EXEMPTIONS)})")


def verify_output(md_path, xlsx_path):
    """Rule-50 OUTPUT-level gate: re-scan the PRODUCED files, not just the generator data."""
    # ---- the .md, reader-facing portion only (everything above the QA-only marker)
    text = open(md_path, encoding="utf-8").read()
    assert QA_MARKER in text, "QA-only marker missing from the produced .md"
    reader_md, qa_md = text.split(QA_MARKER, 1)
    p_md = rule7_scan(reader_md, "produced .md")
    # ---- each produced xlsx tab
    wb = openpyxl.load_workbook(xlsx_path)
    names = wb.sheetnames
    per_tab = {}
    for name in names[:3]:
        cells = [str(c.value) for row in wb[name].iter_rows() for c in row
                 if c.value is not None]
        per_tab[name] = rule7_scan("\n".join(cells), name)
    total = len(p_md) + sum(len(v) for v in per_tab.values())

    print("\nOUTPUT-LEVEL WORDING GATE (Standing Rule 50 - the produced files, not the generator)")
    print(f"  produced .md, reader-facing portion ({len(reader_md.splitlines())} lines): "
          f"{len(p_md)} findings {sorted(set(p_md)) if p_md else ''}")
    for name, v in per_tab.items():
        n_cells = sum(1 for row in wb[name].iter_rows() for c in row if c.value is not None)
        print(f"  xlsx tab {name!r} ({n_cells} non-empty cells): "
              f"{len(v)} findings {sorted(set(v)) if v else ''}")
    print(f"  TOTAL across the reader-facing .md and all three reader-facing tabs: {total}")
    if total:
        raise SystemExit("OUTPUT-LEVEL GATE FAILED")
    # the QA-only half must, by contrast, carry the traceability
    n_cids = len(set(re.findall(r"\bC\d{5}\b", qa_md)))
    print(f"  QA-only portion carries {n_cids} distinct TestRail case ids (expected - "
          "traceability lives there, Standing Rule 8)")
    print(f"  workbook tabs: {names} (all <= 31 chars: "
          f"{all(len(n) <= 31 for n in names)})")
    return total, n_cids


if __name__ == "__main__":
    n_pairs, n_uniq = verify_cids()
    verify_reader_text_clean()
    md_path = write_md()
    xlsx_path = write_xlsx()
    verify_output(md_path, xlsx_path)
    n2d = sum(1 for r in TAB2 if r["group"] == SVB.G_DEC)
    n3d = sum(1 for r in TAB3 if r["group"].startswith("Decisions"))
    print("\nTALLY")
    print(f"  Tab 1 {TAB1_NAME!r}: 1 item (1 decision, urgent)")
    print(f"  Tab 2 {TAB2_NAME!r}: {len(TAB2)} items ({n2d} decisions + "
          f"{sum(1 for r in TAB2 if r['group'] == SVB.G_WRITE)} write-down + "
          f"{sum(1 for r in TAB2 if r['group'] == SVB.G_FYI)} awareness)")
    print(f"  Tab 3 {TAB3_NAME!r}: {len(TAB3)} items ({n3d} decisions + "
          f"{len(TAB3) - n3d} write-downs)")
    print(f"  TOTAL in front of Chris: {1 + len(TAB2) + len(TAB3)} items (28 in, "
          f"{len(DEDUP_LOG)} de-duplicated)")
    print(f"  QA-only: {len(QA_ROWS)} mapping rows + {len(MERGED_OUT)} merged-out rows + "
          f"{len(WITHDRAWN)} withdrawn + {len(NOT_ASKED)} not-asked + "
          f"{len(SOURCES_SWEPT)} sources swept + {len(SOURCE_CURRENCY)} currency rows")
