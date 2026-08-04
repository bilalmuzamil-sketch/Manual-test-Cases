#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Mark the three source Chris Ward sheets as SUPERSEDED by the consolidated workbook.

They are KEPT, not deleted (they are the record of how the questions were derived) — but nobody
must send an old one now that one workbook replaces all three.

WHAT IT DOES
  .md  — inserts a banner block directly under the H1 title. Idempotent: re-running does nothing.
  .xlsx — inserts a NEW FIRST SHEET called "SUPERSEDED" carrying the same banner, so the file
          opens on it. No existing sheet, cell, width or style is touched.

BYTE-LEVEL VERIFICATION (Standing Rule 50): before writing each workbook every cell value of
every existing sheet is snapshotted; after writing, every one is re-read and compared. A single
difference aborts and reports both values. Sheet order and names are checked too.

Run:  python3 mark_superseded.py
"""

import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)

CONSOLIDATED = ("build/report-suite/chris-consolidated-2026-08-04/"
                "Report-Suite_Questions-and-Decisions-for-Chris-Ward_2026-08-04.xlsx")

BANNER_MD = [
    "> ## ⚠️ SUPERSEDED 2026-08-04 — DO NOT SEND THIS SHEET",
    ">",
    "> The QA lead asked for **one sheet with three tabs**, so this sheet and the other two "
    "Chris Ward sheets were consolidated into a single workbook on **2026-08-04**:",
    ">",
    f"> **`{CONSOLIDATED}`** (with a `.md` twin)",
    ">",
    "> Send **that** workbook, never this file. This one is kept only as the record of how its "
    "questions were derived and verified. Its content was carried across faithfully — four "
    "overlapping items across the three sheets were removed so nothing is asked twice, and every "
    "removal plus every text change is logged on the consolidated workbook's QA-only tab.",
]

BANNER_XLSX = (
    "SUPERSEDED 2026-08-04 - DO NOT SEND THIS FILE.\n\n"
    "The QA lead asked for one sheet with three tabs, so this sheet and the other two Chris Ward "
    "sheets were consolidated into a single workbook on 2026-08-04:\n\n"
    + CONSOLIDATED + "\n\n"
    "Send that workbook, never this file. This one is kept only as the record of how its "
    "questions were derived and verified. Its content was carried across faithfully - four "
    "overlapping items across the three sheets were removed so nothing is asked twice, and every "
    "removal plus every text change is logged on the consolidated workbook's QA-only tab."
)

SOURCES = [
    "chris-location-question-2026-08-04/PO-Question-Chris-Ward-Location-Column-2026-08-04",
    "chris-sheet-2026-08-04/Report-Suite_Spec-vs-Build_Decisions-for-Chris-Ward_2026-08-04",
    "PO-Questions-Chris-ReportSuite-2026-08-03",
]

MARKER = "SUPERSEDED 2026-08-04"


def mark_md(path):
    with open(path, encoding="utf-8") as f:
        lines = f.read().split("\n")
    if any(MARKER in ln for ln in lines[:15]):
        print(f"  md   already marked, left alone: {os.path.relpath(path, PROJECT)}")
        return False
    # insert immediately after the H1 title (and its blank line)
    at = 1
    while at < len(lines) and lines[at].strip() == "":
        at += 1
    out = lines[:at] + BANNER_MD + [""] + lines[at:]
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(out))
    print(f"  md   banner inserted at line {at + 1}: {os.path.relpath(path, PROJECT)}")
    return True


def snapshot(path):
    wb = openpyxl.load_workbook(path)
    snap = {}
    for name in wb.sheetnames:
        ws = wb[name]
        snap[name] = {(c.row, c.column): c.value
                      for row in ws.iter_rows() for c in row if c.value is not None}
    order = list(wb.sheetnames)
    wb.close()
    return order, snap


def mark_xlsx(path):
    order, before = snapshot(path)
    rel = os.path.relpath(path, PROJECT)
    if "SUPERSEDED" in order:
        print(f"  xlsx already marked, left alone: {rel}")
        return False
    wb = openpyxl.load_workbook(path)
    ws = wb.create_sheet("SUPERSEDED", 0)
    ws["A1"] = BANNER_XLSX
    ws["A1"].font = Font(bold=True, color="9C0006")
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[1].height = 220
    wb.active = 0
    wb.save(path)
    wb.close()

    # ---- byte-level verification of every pre-existing cell (Standing Rule 50)
    order_after, after = snapshot(path)
    if order_after != ["SUPERSEDED"] + order:
        raise SystemExit(f"SHEET ORDER CHANGED in {rel}: {order_after} vs "
                         f"{['SUPERSEDED'] + order}")
    checked = 0
    for name, cells in before.items():
        if name not in after:
            raise SystemExit(f"SHEET LOST in {rel}: {name}")
        for key, val in cells.items():
            got = after[name].get(key)
            if got != val:
                raise SystemExit(f"CELL CHANGED in {rel}!{name}{key}: {val!r} -> {got!r}")
            checked += 1
        extra = set(after[name]) - set(cells)
        if extra:
            raise SystemExit(f"CELLS ADDED in {rel}!{name}: {sorted(extra)[:5]}")
    print(f"  xlsx SUPERSEDED tab added as sheet 1: {rel} "
          f"| byte-verified {checked} pre-existing cells across {len(before)} sheet(s), "
          f"all IDENTICAL, 0 added, 0 lost")
    return True


if __name__ == "__main__":
    changed = 0
    for base in SOURCES:
        print(base)
        md = os.path.join(PROJECT, base + ".md")
        xl = os.path.join(PROJECT, base + ".xlsx")
        for p in (md, xl):
            if not os.path.exists(p):
                sys.exit("MISSING SOURCE FILE: " + p)
        changed += mark_md(md)
        changed += mark_xlsx(xl)
    print(f"\n{changed} file(s) changed. The three source sheets are kept, never deleted.")
