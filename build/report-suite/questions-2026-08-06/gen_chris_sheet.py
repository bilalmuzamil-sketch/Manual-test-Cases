#!/usr/bin/env python3
"""Generate the Chris Ward question sheet of 2026-08-06 (.xlsx + .md).

Authorised by the QA lead, verbatim: "If there are more questions for Chris make
sheet for him."

MIRRORS 1:1 (Standing Rule 16) the established peers:
  build/report-suite/chris-consolidated-2026-08-04/
      Report-Suite_Questions-and-Decisions-for-Chris-Ward_2026-08-04.xlsx
  build/report-suite/rulings-2026-08-05/
      Follow-up-Question-for-Chris-Ward_2026-08-05.xlsx
Same six reader columns in the same order, the same row layout (A1 title /
A2 note / A4 header / A5 band / items from row 6), the same fills and fonts,
the same freeze pane, and the same QA-only tab that is never sent.

Standing Rule 55: every reader-facing question row NAMES THE PROJECT AND THE
REPORT, because Chris Ward owns BOTH the Report Suite AND Fees & Discounts, so
"the export" or "the date filter" is genuinely ambiguous to him.

RESEARCH ONLY - this script writes two files into this folder. It makes no
TestRail call, no Jira call and no application call of any kind.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_2026-08-06.xlsx")
MD = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_2026-08-06.md")

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Needed first - Location column"
TAB2_NAME = "Decisions we need"
TAB3_NAME = "Small tidy-ups - one tick each"
TAB4_NAME = "QA internal - not for Chris"

THANKS = (
    "Thank you - your last round of answers landed well and most of the reporting work moved "
    "straight away. This sheet is the rest of what is genuinely still open. Every question says "
    "which project and which report it is about, because we know you look after more than one "
    "thing here."
)

# ------------------------------------------------------------------- the questions
# (topic, what-happens-now, the question, options)

TAB1 = [
    (
        "Report Suite - the Location column - all six reports "
        "(the \"show or hide columns\" story on each report, under epic SV-8582)",

        "This is a one-line confirmation rather than a fresh decision, and it is the only thing "
        "holding seven of our tests.\n\n"
        "You have already decided this. The note recording your decision appears in every one of "
        "the six written descriptions, and it says the same thing each time: anyone who can see "
        "more than one location gets the Location column, it is on by default, and they can switch "
        "it on or off themselves from the list of columns. Someone who can see only one location "
        "never sees it at all.\n\n"
        "What is missing is only the tidy-up. Older sentences are still sitting in all six "
        "descriptions saying the column appears and disappears on its own depending on how many "
        "locations are picked - and on three of the six one of those sentences goes further and "
        "says flatly that the person cannot switch it on or off. Those are listed one per report "
        "on the third tab, as ticks rather than questions.\n\n"
        "Why we are asking: we do not want to release seven tests on our own reading of a note. "
        "One word from you and they go.",

        "Is this right - anyone who can see more than one location gets the Location column, on by "
        "default, and can switch it on or off themselves?",

        "A) Yes, that is the rule - it is an access thing, not a picking thing, and the person can "
        "switch it on or off whenever they like. (Then the older sentences on the third tab are "
        "just leftovers to delete, and our seven tests come off hold the same day.)\n\n"
        "B) No - the column should appear and disappear on its own depending on how many locations "
        "are picked, and the person should not be able to switch it. (Then the newer note in all "
        "six descriptions is the one that needs removing, and we rewrite the seven tests.)\n\n"
        "C) Something else, or it should differ between reports - please describe it.",
    ),
]

TAB2 = [
    (
        "Report Suite - all six reports - which heading each report sits under in the Reports menu "
        "(the \"report access and navigation placement\" story on each report, for example SV-8600)",

        "The six new reports are currently spread across three headings in the Reports menu. Work "
        "In Progress, Technician Utilization and Sales By Representative sit under PERFORMANCE. "
        "Parts Velocity and Inventory Value sit under PARTS. Sales By Customer sits on its own "
        "under SALES.\n\n"
        "None of the six written descriptions says which heading a report belongs under. They only "
        "say the report \"appears in the Reports left-side navigation\", which it does.\n\n"
        "Why we are asking: one of our tests had been written expecting Sales By Customer under "
        "Performance. That expectation came from a walkthrough video last month, not from any "
        "current description, so we have taken it out rather than guess. Right now the test records "
        "the heading and does not judge it, which is honest but not much of a test.",

        "Is that arrangement the one you want - three of them under Performance, two under Parts, "
        "and Sales By Customer under Sales?",

        "A) Yes, leave them spread as they are. (Then please add the heading to each description so "
        "there is something to test against.)\n\n"
        "B) No - all six should sit together under one heading. Please say which.\n\n"
        "C) Something else - please describe where each one should go.",
    ),
    (
        "Report Suite - Sales By Customer - an invoice the person is not allowed to open "
        "(the invoice-link story, under epic SV-8582)",

        "Your Sales By Customer description says two different things about the same thing, and "
        "both are in the current version.\n\n"
        "One part says a person who cannot open an invoice is not given a link at all - the invoice "
        "number is shown as ordinary text.\n\n"
        "Another part says that person does get a link, clicks it, and lands on the standard "
        "\"you are not allowed in\" page, from which they can press back.\n\n"
        "Why we are asking: three of our tests are on hold on this one point, and the two answers "
        "need completely different tests. This is the same question that went to you on 5 August - "
        "it is repeated because it is still open, not because we forgot.",

        "Which one do you want?",

        "A) No link at all. For that person the invoice number is plain text and there is nothing "
        "to click.\n\n"
        "B) A link they can click, which takes them to the \"you are not allowed in\" page.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "Report Suite - Sales By Representative - the same link rule, which never got written into "
        "the numbered requirements (the invoice-link and customer-link story, under epic SV-8582)",

        "On 5 August you added the rule that a link is only a link when the person is allowed to "
        "open what it points at. Thank you - it was written into Sales By Customer properly.\n\n"
        "On Sales By Representative it only went into the opening paragraph. The numbered "
        "requirements underneath still say, flatly, that each invoice number on a detail row is a "
        "clickable link, and that each customer name on a detail row is a clickable link.\n\n"
        "Why we are asking: a tester reading only the numbered requirements would expect a link for "
        "everybody, and would raise a fault against a correct build.",

        "Should the numbered requirements on Sales By Representative be updated to match Sales By "
        "Customer?",

        "A) Yes, please update them so they say the same as Sales By Customer.\n\n"
        "B) No - Sales By Representative should keep links for everybody, whether or not they can "
        "open what the link points at.",
    ),
    (
        "Report Suite - Sales By Representative - what paper size and orientation the printable "
        "downloads should use (the download story, under epic SV-8582)",

        "Your two descriptions disagree with each other about the same thing.\n\n"
        "The Sales By Representative description says its printable downloads are A4 portrait. The "
        "Sales By Customer description says A4 landscape.\n\n"
        "The Sales By Representative table has sixteen columns, which would not fit across a "
        "portrait page at all, and when we last looked both reports came out landscape.\n\n"
        "Why we are asking: one of our tests currently states portrait, word for word from your "
        "description, and is marked ready for the automation team to pick up. If portrait is not "
        "what you want, that test is wrong and we would rather fix it before it is automated than "
        "after.",

        "For Sales By Representative, should the printable downloads be A4 landscape or A4 "
        "portrait?",

        "A) A4 landscape, the same as Sales By Customer - and the Sales By Representative "
        "description should be corrected to say landscape.\n\n"
        "B) A4 portrait, as its description currently says. (Then we raise the difference with the "
        "developers.)",
    ),
    (
        "Report Suite - Sales By Representative - the word \"Representative\" on the screen and on "
        "the customer's card (the representative-assignment story, under epic SV-8582)",

        "You confirmed that \"Representative\" on its own is fine in the downloaded files, and we "
        "have matched our tests to that.\n\n"
        "Two tests are still on hold because they are about the word on the SCREEN and on the "
        "customer's card, not in the files. You were only asked about the files, so we did not want "
        "to stretch your answer to cover screens you had not been shown.\n\n"
        "Why we are asking: two tests come off hold the moment you answer, whichever way you "
        "answer. This one was asked on 5 August and is still open.",

        "Should the full word \"Representative\" also be used on the screen and on the customer's "
        "card?",

        "A) Yes - use the full word everywhere it appears, not only in the downloaded files.\n\n"
        "B) No - only the downloaded files matter. The screen can stay as it is.",
    ),
]

TAB3 = [
    (
        "Report Suite - Parts Velocity - one leftover sentence about the Location column",

        "Your decision note in this description says the person can switch the Location column on "
        "and off. A sentence further down still says the column \"is not user-toggleable\" and is "
        "not one of the columns offered in the picker.\n\n"
        "That is the strongest of the six leftovers: it says the opposite of your decision in plain "
        "words. Parts Velocity is also the one report where the older wording was never revisited "
        "on this point.",

        "Nothing to decide - please delete or reword that sentence next time you have the "
        "description open.",

        "(No options - a tick is enough. Tick here if you would like us to keep reminding you until "
        "it is done.)",
    ),
    (
        "Report Suite - Work In Progress - one leftover sentence about the Location column",

        "Your decision note in this description says the person can switch the Location column on "
        "and off. A sentence further down still says the column is shown automatically and that "
        "\"the user does not toggle it in the column selector\", and two other places describe it as "
        "appearing automatically.\n\n"
        "So this description currently contains both answers, one of them in so many words.",

        "Nothing to decide - please delete or reword those sentences next time you have the "
        "description open.",

        "(No options - a tick is enough. Tick here if you would like us to keep reminding you until "
        "it is done.)",
    ),
    (
        "Report Suite - Inventory Value - one leftover sentence about the Location column",

        "Your decision note in this description says the person can switch the Location column on "
        "and off. A sentence further down still says its visibility \"follows the location scope "
        "automatically\" and that it \"is not one of the columns offered in the column-selection "
        "control\".\n\n"
        "So this description also contains both answers.",

        "Nothing to decide - please delete or reword that sentence next time you have the "
        "description open.",

        "(No options - a tick is enough. Tick here if you would like us to keep reminding you until "
        "it is done.)",
    ),
    (
        "Report Suite - Sales By Customer - one milder leftover sentence about the Location column",

        "Milder than the three above. This description does not say the person cannot switch the "
        "column - it simply still has a summary sentence saying the column is shown only when more "
        "than one location is in view and hidden when a single location is in view.\n\n"
        "Your newer wording in the same document says the opposite: that it is on by default and "
        "stays available whatever the person has picked.",

        "Nothing to decide - please tidy that summary sentence next time you have the description "
        "open.",

        "(No options - a tick is enough. Tick here if you would like us to keep reminding you until "
        "it is done.)",
    ),
    (
        "Report Suite - Sales By Representative - one milder leftover sentence about the Location "
        "column",

        "The same mild leftover. A numbered requirement still says the column is shown only when the "
        "view spans more than one location and is hidden when the view is narrowed to a single "
        "location, \"because that one location is already unambiguous\".\n\n"
        "Your newer wording earlier in the same document says it is on by default and can be "
        "switched on or off whatever is picked.",

        "Nothing to decide - please tidy that requirement next time you have the description open.",

        "(No options - a tick is enough. Tick here if you would like us to keep reminding you until "
        "it is done.)",
    ),
    (
        "Report Suite - Technician Utilization - one milder leftover sentence about the Location "
        "column",

        "The same mild leftover, in two places: the column is described as hidden whenever a single "
        "location is in view.\n\n"
        "Your newer wording in the same document says it is on by default and can be switched on or "
        "off regardless of how many locations are picked.",

        "Nothing to decide - please tidy those two sentences next time you have the description "
        "open.",

        "(No options - a tick is enough. Tick here if you would like us to keep reminding you until "
        "it is done.)",
    ),
    (
        "Report Suite - Parts Velocity, Technician Utilization and Work In Progress - the download "
        "size limit you have already written down elsewhere",

        "Every one of the six reports refuses to build a download once the result would run past "
        "about ten thousand rows, and shows \"This report is too large to export. Narrow the date "
        "range or filters, then try again.\" That is deliberate and it is correct.\n\n"
        "Three of your six descriptions already say so - Sales By Customer, Sales By Representative "
        "and Inventory Value. The other three do not mention it at all.\n\n"
        "So this is not a question about whether there is a limit. It is a request to copy the "
        "sentence you have already written into the three that are missing it.\n\n"
        "Why we are asking: on those three reports a tester who meets that message has nothing to "
        "check it against, and the honest thing for them to do is raise it as a fault - against "
        "behaviour that is entirely correct.\n\n"
        "One correction to our own record, so you are not misled: we previously told you that NONE "
        "of the six mentioned this limit. That was wrong - three of them do.",

        "Please add the limit to Parts Velocity, Technician Utilization and Work In Progress.",

        "A) Yes, I will add it to those three.\n\n"
        "B) No - leave it out of those three. (Then please say why they are different, because the "
        "product behaves the same on all six.)",
    ),
]


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws, r


# --------------------------------------------------------------- QA-only mapping
TR = "https://shopview.testrail.io/index.php?/cases/view/"

QA_ROWS = [
    ("Tab 1", "1",
     "The Location column - is the access-gate + user-toggleable model the ruling?",
     "Q5 of full-viu-2026-08-06/QUESTIONS-FOR-CHRIS.md. Every one of the six descriptions carries "
     "the 2026-08-04 QA-review decision note stating the access-gate + toggleable model, while "
     "older requirement text in the same document states the selection-scope / auto-visibility "
     "model. He has decided; the edit is unfinished.",
     "SEVEN cases on AUTOMATION: HOLD naming this contradiction - WIP-COL-02 (C30467); "
     "IV-COL-01 (C30551); IV-COL-04 (C30554); IV-EXP-02 (C30588); SBC-LOC-04 (C38912); "
     "IV-LOC-06 (C38917); WIP-PERS-05 (C43551)",
     " · ".join(f"C{i} {TR}{i}" for i in
                [30467, 30551, 30554, 30588, 38912, 38917, 43551]),
     "LIVE-VERIFIED 2026-08-06 over the six Confluence page bodies (HTTP 200 on all six): "
     "SBC v15 · SBR v17 · PV v5 · TU v6 · WIP v9 · IV v4. HARD contradiction (the text denies the "
     "user can toggle it) in PV v5 (near S4-R1, \"is not user-toggleable\"), WIP v9 S7-R13 (\"the "
     "user does not toggle it in the column selector\") and IV v4 S7-R6 (\"not one of the columns "
     "offered in the column-selection control\"). MILDER contradiction (visibility follows the "
     "SELECTION rather than ACCESS) in SBC v15 (overview line), SBR v17 S21-R7 and TU v6 (two "
     "lines). The newer model is stated at SBC S4-R12, SBR (near S21-R7's own story), PV (near "
     "S3-R3), TU (near S7-R10), WIP and IV in their access-gate paragraphs, all with the phrase "
     "\"regardless of how many locations are currently selected\" on SBC/PV/TU.",
     "A -> the 7 come off hold unchanged in substance, and the 7 tidy-up rows on tab 3 are just "
     "spec edits. B -> the 7 are rewritten to the auto-visibility model AND the decision note in "
     "all six descriptions has to come out. THE COUNT: 7, not 8, 12 or 16. Counted live from "
     "TestRail on 2026-08-06 - 476 of our cases under group 4281, 35 mention the Location column, "
     "43 are on HOLD for any reason, and exactly 7 of those name this contradiction. The brief "
     "said 8, build/OUTSTANDING-ITEMS-REGISTER.md C2 says 12, and an earlier draft said 16 - all "
     "three are stale; the 2026-08-06 pass released the others."),

    ("Tab 2", "1",
     "Which heading each of the six reports sits under in the Reports menu",
     "Q1 of QUESTIONS-FOR-CHRIS.md. The specifications are SILENT: SBC v15 Story 1 says only "
     "\"'Sales By Customer' appears in the Reports left-side navigation\". The 'Performance group' "
     "claim traced to a 30 July walkthrough video and the spec has been revised twice since "
     "without naming a group.",
     "SBC-NAV-01 (C30096) - repaired by the 2026-08-06 pass to assert only what the specification "
     "asserts, and now AUTOMATION: READY. NOTHING is on hold for this one.",
     f"C30096 {TR}30096",
     "Spec silence verified live 2026-08-06. Build-side heading observation (SALES / PARTS / "
     "PERFORMANCE) comes from the 2026-08-06 pass on build v3.5-16cf83f and was NOT re-verified "
     "today - the shared QA sign-in returns HTTP 401 sso_required, and quick-login / switch-user "
     "were deliberately not called.",
     "A -> he adds the heading to each description and C30096 regains a real assertion. B -> a "
     "developer job plus six description edits. Either way C30096 stops being a record-it-only "
     "case. Recorded also in full-viu-2026-08-06/NO-SOURCE-DEFECTS.md item 3 as deliberately not "
     "filed, because filing it would assert a requirement no current document contains (Rule 57)."),

    ("Tab 2", "2",
     "Sales By Customer - is a person who cannot open an invoice given a link at all?",
     "Q3 of QUESTIONS-FOR-CHRIS.md, and it was already asked on 2026-08-05. SBC v15 states BOTH: "
     "S9-R1a says the number is a link only when the user may open the target and otherwise plain "
     "text; S9-N2 says that user is shown the standard access-denied page and can press back.",
     "THREE cases on AUTOMATION: HOLD - SBC-PERM-04 (C30100); SBC-LINK-05 (C43558); "
     "SBR-LINK-06 (C43559)",
     " · ".join(f"C{i} {TR}{i}" for i in [30100, 43558, 43559]),
     "Verified live 2026-08-06 in SBC v15: \"Each invoice number on a detail row is a clickable "
     "link.\" sits alongside \"The invoice number is rendered as a link only when the user has "
     "permission to open the target it links to ... a user without that permission sees the "
     "invoice number as plain text.\" and \"If the user lacks permission to open the destination "
     "invoice, the destination page shows the application's standard access-denied state.\"",
     "A -> C43558/C43559 assert plain text and C30100 is retired or rescoped to the direct-URL "
     "case only. B -> C30100 stands and the plain-text half of C43558/C43559 is removed. NOTE: all "
     "three ALSO need the second non-administrator sign-in (register row C1), so his answer alone "
     "does not make them runnable - it makes them writable."),

    ("Tab 2", "3",
     "Sales By Representative - the link-permission rule is missing from the numbered requirements",
     "Q4 of QUESTIONS-FOR-CHRIS.md. His 2026-08-05 rule reached SBR's introductory §2 but not its "
     "numbered requirements.",
     "SBR-LINK-06 (C43559) - already counted above; no additional case is held by this row",
     f"C43559 {TR}43559",
     "Verified live 2026-08-06 in SBR v17: §2 carries \"rendered as a link only when the user has "
     "permission to open that target, otherwise plain text\", while the numbered requirements "
     "still read \"Each invoice number on a detail row is a clickable link.\" and \"Each customer "
     "name on a detail row is a clickable link that navigates the current tab to the customer's "
     "record.\" The 2026-08-05 change-log entry in the same page confirms the rule was applied.",
     "A -> a spec edit only; our cases already follow the conditional rule. B -> SBR diverges from "
     "SBC on the same rule and C43559 splits into two cases."),

    ("Tab 2", "4",
     "Sales By Representative - A4 landscape or A4 portrait for the printable downloads?",
     "Q7 of QUESTIONS-FOR-CHRIS.md (added by the third session). SBR v17 says A4 portrait; SBC v15 "
     "says A4 landscape. A sixteen-column table cannot fit portrait.",
     "SBR-EXP-01 (C30278) asserts \"The PDF is A4 portrait\" word for word and is currently "
     "AUTOMATION: READY. Its sibling SBR-EXP-02 (C30279) is READY - EXPECT FAIL (SV-8981) for a "
     "different, unrelated fault.",
     " · ".join(f"C{i} {TR}{i}" for i in [30278, 30279]),
     "Verified live 2026-08-06. SBR v17, verbatim: \"Both PDFs are server-rendered and delivered "
     "as a file attachment, in A4 portrait, edge-to-edge ...\". SBC v15, verbatim: \"The PDF is A4 "
     "landscape with 25px margins on all sides ...\". PV v5 separately documents \"A3 landscape\", "
     "so A3 is not undocumented across the suite.",
     "A (landscape) -> C30278's orientation line is rewritten AND its marker must move, because a "
     "case asserting portrait against a landscape build cannot honestly stay plain READY. "
     "B (portrait) -> a developer job, and C30278 becomes READY - EXPECT FAIL. HONEST CORRECTION: "
     "QUESTIONS-FOR-CHRIS.md says \"the test is on hold for this one point\" - it is NOT; C30278 "
     "reads AUTOMATION: READY live today. The A3 half is already reported and is not asked here."),

    ("Tab 2", "5",
     "Does \"Representative\" in full also apply on screen and on the customer's card?",
     "Carried forward from rulings-2026-08-05 Tab 3 item 2, still unanswered. His answer covered "
     "only the download column heading; extending it to the screen would be stretching it past "
     "what he wrote (Rule 12).",
     "TWO cases on AUTOMATION: HOLD - SBR-WO-01 (C30310); SBR-WO-06 (C30315)",
     " · ".join(f"C{i} {TR}{i}" for i in [30310, 30315]),
     "Both cases read AUTOMATION: HOLD - \"waiting on an answer from the product owner\" live on "
     "2026-08-06, so the question is demonstrably still open. Their provenance lines cite SBR v17 "
     "S19-R1/S19-N1 and S19-R7/S19-E1.",
     "A -> both keep the full-word expectation and a developer job is raised. B -> both are "
     "reworded to the short form. Either way 2 cases come off hold."),

    ("Tab 3", "1-6",
     "Six per-report tidy-ups of the leftover Location-column wording",
     "The unfinished half of his own 2026-08-04 decision. Deliberately one row PER REPORT, because "
     "each is a different sentence in a different document and a single combined row would give "
     "him nothing to act on.",
     "No case is held by these six rows on their own - they are the spec-side twin of tab 1's "
     "single question, whose 7 held cases are listed above.",
     "(see tab 1's row)",
     "Live 2026-08-06, per report. HARD (denies toggling): PV v5 near S4-R1 · WIP v9 S7-R13 · "
     "IV v4 S7-R6. MILD (selection-scope wording): SBC v15 overview line · SBR v17 S21-R7 · "
     "TU v6, two lines. WIP additionally carries the auto-visibility wording near S9-E1 and in its "
     "own change-log entry.",
     "ALL SIX ARE EDITS, NOT DECISIONS - which is why they are ticks. IMPORTANT COUNT CORRECTION: "
     "the brief and the register both say FOUR specs still state it both ways (naming SBR S21-R7, "
     "WIP S7-R13, IV S7-R6, SBC S13-R4, plus PV as never-touched). Live today it is SIX - all six "
     "descriptions still contain both models, three of them in explicit terms. SBC's live "
     "contradiction is NOT at S13-R4; it is an overview line. Nothing was taken on trust."),

    ("Tab 3", "7",
     "The ~10,000-row download limit is missing from three descriptions",
     "Q6 of QUESTIONS-FOR-CHRIS.md, narrowed. It is real and deliberate - epic story SV-8591 "
     "\"Export contract + 10k row-cap guard\".",
     "No case is on hold for this. It is a documentation gap that would make a tester raise a "
     "fault against correct behaviour.",
     "(no held cases)",
     "COUNTED LIVE 2026-08-06 over the six spec bodies. PRESENT in three: SBC v15 (\"Each CSV is "
     "capped at 10,000 data rows\" and the same for PDF), SBR v17 (\"the export's row cap of "
     "10,000 data rows\", S14-E2), IV v4 (\"10,000 rows\" plus the verbatim message). ABSENT from "
     "three: PV v5, TU v6, WIP v9 - zero matches for 10,000 / 10k / \"too large to export\" in any "
     "of them.",
     "A -> three spec edits and the message becomes testable everywhere. B -> he must explain why "
     "three reports differ, because the product does not. HONEST CORRECTION CARRIED ON THE READER "
     "TAB ITSELF: full-viu-2026-08-06/FINDINGS.md and FILED.md both say \"none of the six "
     "specifications mentions the cap\" - that is WRONG, and the sheet tells him so rather than "
     "quietly asking the narrow question."),
]

QA_NOTES = [
    "HOW THE POPULATION WAS SEARCHED (Standing Rule 50): all cases under TestRail group 4281 were "
    "pulled LIVE read-only on 2026-08-06 - 481 live, of which 476 are ours and 5 are Vladimir "
    "Tomovic's (C38919-C38923, hands off per Rule 38). Every one of the 476 was searched in full "
    "across title, preconditions, steps and expected results. No sampling. Counts reported to "
    "Chris are the live counts, not the counts in our own notes.",

    "WHAT WAS DELIBERATELY LEFT OFF, AND WHY - (1) Q2 of QUESTIONS-FOR-CHRIS.md, \"what should the "
    "totals line do when nothing matches?\" IS NOT ON THIS SHEET. Its premise is false. SBC v15 "
    "answers it TWICE, verbatim: \"If an export (CSV or PDF) is triggered while the active filters "
    "match no customers ... the export still downloads, containing the column headers and a totals "
    "row of zeros, with no data rows and no warning\" and \"When no customer is selected (every "
    "customer cleared), the report shows the empty state (Story 17) and the totals row shows "
    "zeros\" (both near S18-R10/R11, read live 2026-08-06). So the build showing NO totals line is "
    "a DEFECT against his own document, not a product decision - and Rule 7 forbids putting a bug "
    "in front of a PO. Asking it would also have re-asked a question his own spec answers, which "
    "we have embarrassed ourselves with once already. FOR THE QA LEAD: "
    "full-viu-2026-08-06/NO-SOURCE-DEFECTS.md item 2 says \"No requirement says what the totals "
    "line should do when nothing matches\" - that is wrong, and SBC-EXP-11 (C30173) still asserts "
    "the zeros row and reads AUTOMATION: READY. This wants a ticket, not a question. Those files "
    "were NOT edited - they belong to another pass.",

    "WHAT WAS DELIBERATELY LEFT OFF - (2) the en-dash vs em-dash in the Sales By Customer PDF date "
    "heading (NO-SOURCE-DEFECTS.md item 4). It is a defect, it is sourced to S15-R11, and it needs "
    "a one-word QA-lead ruling (fold into SV-8937, file separately, or drop) - not a PO question. "
    "(3) The Location column not sorting - already filed as SV-8963. (4) The Work In Progress "
    "export failing with a server error - already covered by SV-8907. (5) The ~10,000-row refusal "
    "as a suspected defect - it is DELIBERATE and in the epic (SV-8591), so only the documentation "
    "half is asked.",

    "WHAT WAS DELIBERATELY LEFT OFF - (6) the date-range chooser's Today and Yesterday presets. "
    "QUESTIONS-FOR-CHRIS.md records that we nearly asked this and then found he ANSWERED it on "
    "5 August, with our own case already following the answer. Not re-asked. That check - our own "
    "newer sources first - is what kept it off the sheet.",

    "WORDING RULES APPLIED (Standing Rules 7 and 55): every reader-facing question row names the "
    "PROJECT (Report Suite) and the REPORT, because Chris Ward also owns Fees & Discounts and a "
    "bare \"the export\" or \"the date filter\" would be genuinely ambiguous to him days later on "
    "a phone. Story or epic references appear in plain form only where they orient him. No case "
    "IDs, no requirement anchors, no HTTP terms, no endpoint names, no internal names and not the "
    "word VIU appear anywhere he reads. Each question carries a one-line \"why we are asking\" so "
    "the consequence is visible, and every question is a plain A/B/C with a blank for the answer.",

    "SOURCE-CURRENCY (Standing Rules 31 and 59), all fetched LIVE at the start of this pass AND "
    "re-read immediately before the files were written, on 2026-08-06: the six Confluence page "
    "bodies over GET /wiki/rest/api/content/<id>?expand=version,body.storage, HTTP 200 on all six "
    "- SBC 577634305 v15 (2026-08-05T17:53:06Z) · SBR 585629698 v17 (2026-08-05T17:53:08Z) · "
    "PV 620888066 v5 (2026-08-05T13:21:40Z) · TU 641400833 v6 (2026-08-05T13:33:10Z) · "
    "WIP 703660034 v9 (2026-08-05T17:54:07Z) · IV 720142338 v4 (2026-08-05T13:33:13Z). None moved "
    "during the pass. The CONFLUENCE version number was used, never the version written inside the "
    "document body (Rule 31's trap). Epic SV-8582 was not re-read this pass and no claim rests on "
    "it. Designs: ABSENT - spec-only project.",

    "THE BUILD IS A SOURCE AND IT IS STALE FOR US TODAY (Standing Rules 12, 49, 60). The shared "
    "QA sign-in is dead - GET /api/auth/me/fe-permissions on sv8582api returns HTTP 401 "
    "{\"error\":\"sso_required\"} - and quick-login and switch-user were deliberately NOT called, "
    "because both rotate the shared session and would sign concurrent workers out. So NO build-side "
    "claim on this sheet was re-verified today. Worse, the branch has REDEPLOYED since the pass we "
    "are citing: index.html on sv8582.qa.shopview.com now reads app-version v3.5-f77875c, "
    "last-modified Thu 06 Aug 2026 10:43:37 GMT, where the 2026-08-06 pass ran on v3.5-16cf83f. "
    "Every build observation quoted to Chris (the three menu headings, the landscape rendering, the "
    "missing totals line) therefore comes from a build that no longer exists, and is presented to "
    "him as \"when we last looked\" rather than as a current fact. All Report Suite verdicts remain "
    "PROVISIONAL; the Rule-49 queues are OPEN and the 2026-08-06 pass opened none (register RS2).",

    "NOTHING HAS BEEN WRITTEN ANYWHERE. This sheet is a draft for the QA lead to send. Read-only "
    "on TestRail (get_sections, get_cases) and on Confluence (GET content). No TestRail write, no "
    "Jira call at all in the Report Suite half, no case edit, no run write. CLAUDE.md, "
    "build/OUTSTANDING-ITEMS-REGISTER.md and build/APP-ACTIONS-PLAYBOOK.md were not touched.",

    "THREE NUMBERS IN OUR OWN RECORDS ARE STALE AND ARE CORRECTED HERE RATHER THAN REPEATED: the "
    "held Location count is 7 (the brief said 8, the register says 12, an earlier draft said 16); "
    "the number of specs still stating the Location model both ways is 6, not 4; and three of the "
    "six specs DO document the export cap, where two of our own 2026-08-06 files say none does. An "
    "inflated blocked-count or a wrong \"none of them\" in front of a PO costs credibility on every "
    "other row of the sheet, so each was counted from the live source before it went on.",
]


def write_xlsx():
    wb = openpyxl.Workbook()
    w_narrow = [4, 34, 52, 42, 46, 22]
    w_wide = [4, 34, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "Needed first - the Location column - Report Suite - Chris Ward - 2026-08-06",
           "Needed first, please - this is a one-line confirmation and it releases seven tests that "
           "are on hold today. Everything else on this sheet can wait until you have a spare ten "
           "minutes. " + THANKS,
           "One confirmation - it releases seven held tests", TAB1, w_wide, first=True)

    _sheet(wb, TAB2_NAME,
           "Decisions we need - Report Suite - Chris Ward - 2026-08-06",
           "Five questions, each a plain A or B. Two of them were asked on 5 August and are here "
           "again because they are still open, not because we forgot. " + THANKS,
           "Five decisions - five tests are on hold across them", TAB2, w_narrow)

    _sheet(wb, TAB3_NAME,
           "Small tidy-ups - Report Suite - Chris Ward - 2026-08-06",
           "NO DECISIONS ON THIS TAB. Seven small wording edits in your own descriptions - six of "
           "them are the unfinished half of a decision you have already made, one is a sentence you "
           "have already written for three reports and we are asking for it on the other three. A "
           "tick is enough on each.",
           "Seven wording edits - a tick is enough", TAB3, w_narrow)

    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail C-ids, "
                 "requirement anchors, live evidence and the honest count corrections live here so "
                 "the reader-facing tabs stay plain (Standing Rules 7, 8 and 55).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Tab", "Item", "What it asks", "Where the question comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 62, 62]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, 1):
            out.append(f"### Item {i}.0 — {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    md = f"""# Questions for Chris Ward — Report Suite — 2026-08-06

**Project: Report Suite (the six reports) · epic SV-8582 · Product Owner: Chris Ward**

**This is the plain-language twin of `Questions-for-Chris-Ward_Report-Suite_2026-08-06.xlsx`.**
The spreadsheet is the version to send; it mirrors the 2026-08-04 and 2026-08-05 Chris Ward sheets'
format exactly, and it carries a QA-only tab that must not be forwarded.

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

Authorised by the QA lead, verbatim: *"If there are more questions for Chris make sheet for him."*

{THANKS}

**Thirteen items in total: one confirmation that releases seven held tests, five decisions, and
seven small wording edits that need no decision at all.**

**Live source versions confirmed on 2026-08-06, immediately before writing** — Sales By Customer
**version 15** · Sales By Representative **version 17** · Parts Velocity **version 5** ·
Technician Utilization **version 6** · Work In Progress **version 9** · Inventory Value
**version 4**. Every sentence quoted below comes from that fetch.

---

## Tab 1 — {TAB1_NAME}

Needed first, please — this is a one-line confirmation and it releases seven tests that are on hold
today.

{block(TAB1)}

---

## Tab 2 — {TAB2_NAME}

Five questions, each a plain A or B. Two of them were asked on 5 August and are here again because
they are still open, not because we forgot.

{block(TAB2)}

---

## Tab 3 — {TAB3_NAME}

**No decisions on this tab.** Seven small wording edits in your own descriptions — six are the
unfinished half of a decision you have already made, and one is a sentence you have already written
for three reports which we are asking for on the other three. A tick is enough on each.

{block(TAB3)}

---

## QA-only — not for Chris

The internal question-to-case mapping lives on the spreadsheet's `{TAB4_NAME}` tab: every question's
affected TestRail case IDs with links, the requirement anchors quoted from the live pages, and what
each possible answer resolves to. It also records the method notes — how all 476 of our cases were
searched live, what was deliberately left OFF the sheet and why, the source-currency block, and the
three stale numbers in our own records that were corrected before anything went in front of Chris.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
