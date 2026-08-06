#!/usr/bin/env python3
"""Friendly, forward-as-is version of the 2026-08-06 Chris Ward question sheet.

SAME SUBSTANCE, EASIER READING. Every item of the 13-item sheet in this folder is
carried over - none dropped, none added. What changed is the reading experience,
which is what the QA lead asked for: "Give me the friendly and easy to read and
understandable files for Chris and Branko."

  * a short warm note at the top - who it is for, what it covers, how long it
    takes, and that a letter or one line is a perfectly good answer
  * ordered by WHAT TO DO FIRST: the one confirmation that releases held tests,
    then the real decisions, then the tick-box tidy-ups - said so in the headings
  * shorter sentences, no walls of text

UNCHANGED AND NON-NEGOTIABLE (Standing Rules 7 + 55): every row names the PROJECT
and the REPORT, because Chris owns the Report Suite AND Fees & Discounts, so "the
export" or "the date filter" is genuinely ambiguous to him. Nothing he reads
carries a case ID, a requirement anchor, an HTTP term, an endpoint or the word
VIU. The question-to-case mapping stays on the QA-only tab, imported verbatim
from gen_chris_sheet.py so it cannot drift.

RESEARCH ONLY - writes two files into this folder. No TestRail, Jira,
Confluence or application call of any kind.
"""

import importlib.util
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
PRIOR = os.path.join(HERE, "gen_chris_sheet.py")

XLSX = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_Friendly-Version_2026-08-06.xlsx")
MD = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_Friendly-Version_2026-08-06.md")

_spec = importlib.util.spec_from_file_location("prior_chris", PRIOR)
prior = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(prior)

HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Which project and report", "What happens now", "The question", "Options",
        "Your answer"]

TAB1_NAME = "1 Start here"
TAB2_NAME = "2 Decisions"
TAB3_NAME = "3 Just a tick"
TAB4_NAME = "QA internal - not for Chris"

HELLO = (
    "Hello Chris - thank you, your last round of answers cleared most of the reporting work "
    "straight away. This is the rest of what is genuinely still open on the REPORT SUITE (the six "
    "new reports). Thirteen items, and most are one word: about ten minutes if you go straight "
    "down the list. SHORT ANSWERS ARE PERFECT - a letter, or one line. Nothing here needs an "
    "essay.\n\n"
    "WHERE TO START. Section 1 is a single yes-or-no and it releases seven tests that are on hold "
    "today - that is the only thing we are really waiting on. Section 2 is five ordinary "
    "decisions. Section 3 is seven small tidy-ups in your own descriptions: a tick each, nothing "
    "to decide.\n\n"
    "Every question says which project and which report it is about, because we know you look "
    "after Fees & Discounts as well as the Report Suite. And to be clear - we have not edited any "
    "of your descriptions or anyone's tickets. Where two of your own documents disagree we simply "
    "say so and ask which one to keep."
)

# ------------------------------------------------------------------ Section 1
TAB1 = [
    (
        "REPORT SUITE - the Location column - all six reports",

        "This is a one-line confirmation rather than a fresh decision, and it is the only thing "
        "holding seven of our tests.\n\n"
        "You have already decided it. The same note now appears in all six report descriptions: "
        "anyone who can see more than one location gets the Location column, it is on by default, "
        "and they can switch it on or off themselves from the list of columns. Someone who can see "
        "only one location never sees it at all.\n\n"
        "What is left is only tidying. Older sentences still sit in all six descriptions saying "
        "the column appears and disappears on its own depending on how many locations are picked - "
        "and in three of them a sentence says flatly that the person cannot switch it. Those six "
        "are in Section 3 as ticks, not questions.\n\n"
        "Why we are asking: we would rather not release seven tests on our own reading of a note. "
        "One word from you and they go.",

        "Is this right - anyone who can see more than one location gets the Location column, on by "
        "default, and can switch it on or off themselves?",

        "A) Yes, that is the rule. It depends on what someone is allowed to see, not on what they "
        "have picked, and they can switch it whenever they like.\n\n"
        "B) No - the column should appear and disappear on its own depending on how many locations "
        "are picked, and the person should not be able to switch it.\n\n"
        "C) Something else, or it should differ between reports - please describe it.",
    ),
]

# ------------------------------------------------------------------ Section 2
TAB2 = [
    (
        "REPORT SUITE - all six reports - which heading each one sits under in the Reports menu",

        "The six reports are spread across three headings today. Work In Progress, Technician "
        "Utilization and Sales By Representative sit under PERFORMANCE. Parts Velocity and "
        "Inventory Value sit under PARTS. Sales By Customer sits on its own under SALES.\n\n"
        "None of the six descriptions says which heading a report belongs under - only that it "
        "appears in the Reports menu, which it does.\n\n"
        "Why we are asking: one of our tests used to expect Sales By Customer under Performance. "
        "That came from a walkthrough video last month rather than from any current description, "
        "so we have taken it out rather than guess. Today the test only records the heading, which "
        "is honest but not much of a test.",

        "Is that arrangement the one you want - three under Performance, two under Parts, and "
        "Sales By Customer under Sales?",

        "A) Yes, leave them as they are - and please add the heading to each description so there "
        "is something to test against.\n\n"
        "B) No - all six should sit together under one heading. Please say which.\n\n"
        "C) Something else - please describe where each one should go.",
    ),
    (
        "REPORT SUITE - Sales By Customer - an invoice the person is not allowed to open",

        "Your Sales By Customer description says two different things about the same thing, and "
        "both are in the current version.\n\n"
        "One part says a person who cannot open an invoice gets no link at all - the invoice "
        "number is shown as ordinary text.\n\n"
        "Another part says they do get a link, click it, and land on the standard \"you are not "
        "allowed in\" page, from which they can go back.\n\n"
        "Why we are asking: three of our tests are on hold on this one point, and the two answers "
        "need completely different tests. This went to you on 5 August - it is here again only "
        "because it is still open, not because we forgot.",

        "Which one do you want?",

        "A) No link at all. For that person the invoice number is plain text and there is nothing "
        "to click.\n\n"
        "B) A link they can click, which takes them to the \"you are not allowed in\" page.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "REPORT SUITE - Sales By Representative - the same link rule, which never reached the "
        "numbered requirements",

        "On 5 August you added the rule that a link is only a link when the person is allowed to "
        "open what it points at. Thank you - it went into Sales By Customer properly.\n\n"
        "On Sales By Representative it only reached the opening paragraph. The numbered "
        "requirements underneath still say, flatly, that every invoice number and every customer "
        "name on a detail row is a clickable link.\n\n"
        "Why we are asking: a tester reading only those numbered requirements would expect a link "
        "for everybody, and would raise a fault against a build that is correct.",

        "Should the numbered requirements on Sales By Representative be updated to match Sales By "
        "Customer?",

        "A) Yes, please update them so they say the same as Sales By Customer.\n\n"
        "B) No - Sales By Representative should keep links for everybody, whether or not they can "
        "open what the link points at.",
    ),
    (
        "REPORT SUITE - Sales By Representative - what paper size and orientation the printable "
        "downloads should use",

        "Your two descriptions disagree with each other. Sales By Representative says its "
        "printable downloads are A4 portrait. Sales By Customer says A4 landscape.\n\n"
        "The Sales By Representative table has sixteen columns, which would not fit across a "
        "portrait page at all, and when we last looked both reports came out landscape.\n\n"
        "Why we are asking: one of our tests says portrait, word for word from your description, "
        "and it is queued for the automation team to pick up. If portrait is not what you want, we "
        "would much rather fix that test now than after it is automated.",

        "For Sales By Representative, should the printable downloads be A4 landscape or A4 "
        "portrait?",

        "A) A4 landscape, the same as Sales By Customer - and please correct the Sales By "
        "Representative description to say landscape.\n\n"
        "B) A4 portrait, as its description says today. (Then we will raise the difference with "
        "the developers.)",
    ),
    (
        "REPORT SUITE - Sales By Representative - the word \"Representative\" on the screen and on "
        "the customer's card",

        "You confirmed that \"Representative\" on its own is fine in the downloaded files, and we "
        "have matched our tests to that.\n\n"
        "Two tests are still on hold because they are about the word on the SCREEN and on the "
        "customer's card, not in the files. You were only asked about the files, so we did not "
        "want to stretch your answer to cover screens you had not been shown.\n\n"
        "Why we are asking: two tests come off hold the moment you answer, whichever way you "
        "answer. Asked on 5 August, still open.",

        "Should the full word \"Representative\" also be used on the screen and on the customer's "
        "card?",

        "A) Yes - use the full word everywhere it appears, not only in the downloaded files.\n\n"
        "B) No - only the downloaded files matter. The screen can stay as it is.",
    ),
]

TICK = "(Nothing to choose - a tick is enough.)"

# ------------------------------------------------------------------ Section 3
TAB3 = [
    (
        "REPORT SUITE - Parts Velocity - one leftover sentence about the Location column",

        "Your decision note in this description says the person can switch the Location column on "
        "and off. A sentence further down still says the column \"is not user-toggleable\" and is "
        "not offered in the column list.\n\n"
        "That is the strongest of the six leftovers - it says the opposite of your decision, in "
        "plain words.",

        "Nothing to decide - please delete or reword that sentence next time you have the "
        "description open.",
        TICK,
    ),
    (
        "REPORT SUITE - Work In Progress - one leftover sentence about the Location column",

        "Your decision note here says the person can switch the column on and off. A sentence "
        "further down still says it is shown automatically and that the user does not toggle it, "
        "and two other places describe it as appearing on its own.\n\n"
        "So this description currently contains both answers, one of them in so many words.",

        "Nothing to decide - please delete or reword those sentences next time you have the "
        "description open.",
        TICK,
    ),
    (
        "REPORT SUITE - Inventory Value - one leftover sentence about the Location column",

        "Your decision note here says the person can switch the column on and off. A sentence "
        "further down still says its visibility follows the location scope automatically, and that "
        "it is not one of the columns offered in the column-selection control.\n\n"
        "So this description also contains both answers.",

        "Nothing to decide - please delete or reword that sentence next time you have the "
        "description open.",
        TICK,
    ),
    (
        "REPORT SUITE - Sales By Customer - one milder leftover about the Location column",

        "Milder than the three above. This description does not say the person cannot switch the "
        "column - it simply still has a summary sentence saying the column shows only when more "
        "than one location is in view and hides when a single location is in view.\n\n"
        "Your newer wording in the same document says the opposite: on by default, and available "
        "whatever the person has picked.",

        "Nothing to decide - please tidy that summary sentence next time you have the description "
        "open.",
        TICK,
    ),
    (
        "REPORT SUITE - Sales By Representative - one milder leftover about the Location column",

        "The same mild leftover. A numbered requirement still says the column shows only when the "
        "view spans more than one location, and hides when it is narrowed to a single one.\n\n"
        "Your newer wording earlier in the same document says it is on by default and can be "
        "switched on or off whatever is picked.",

        "Nothing to decide - please tidy that requirement next time you have the description open.",
        TICK,
    ),
    (
        "REPORT SUITE - Technician Utilization - one milder leftover about the Location column",

        "The same mild leftover, in two places: the column is described as hidden whenever a "
        "single location is in view.\n\n"
        "Your newer wording in the same document says it is on by default and can be switched on "
        "or off regardless of how many locations are picked.",

        "Nothing to decide - please tidy those two sentences next time you have the description "
        "open.",
        TICK,
    ),
    (
        "REPORT SUITE - Parts Velocity, Technician Utilization and Work In Progress - the download "
        "size limit you have already written down elsewhere",

        "All six reports refuse to build a download once the result would run past about ten "
        "thousand rows, and show \"This report is too large to export. Narrow the date range or "
        "filters, then try again.\" That is deliberate and it is correct.\n\n"
        "Three of your six descriptions already say so - Sales By Customer, Sales By "
        "Representative and Inventory Value. The other three do not mention it at all. So this is "
        "not a question about whether there is a limit; it is a request to copy a sentence you "
        "have already written into the three that are missing it.\n\n"
        "Why it matters: on those three a tester who meets that message has nothing to check it "
        "against, and the honest thing for them to do is raise it as a fault - against behaviour "
        "that is entirely correct.\n\n"
        "One correction to our own record, so you are not misled: we previously told you that NONE "
        "of the six mentioned this limit. That was wrong - three of them do.",

        "Please add the limit to Parts Velocity, Technician Utilization and Work In Progress.",

        "A) Yes, I will add it to those three.\n\n"
        "B) No - leave it out of those three. (Then please say why they are different, because the "
        "product behaves the same on all six.)",
    ),
]

EXTRA_NOTES = [
    "THIS IS THE FRIENDLY, FORWARD-AS-IS VERSION of "
    "Questions-for-Chris-Ward_Report-Suite_2026-08-06.xlsx, produced on the QA lead's instruction: "
    "\"Give me the friendly and easy to read and understandable files for Chris and Branko.\" "
    "ALL 13 ITEMS ARE CARRIED OVER - none dropped, none added, no substance changed. What changed: "
    "a short warm opening note, ordering by what to do first (the confirmation that releases held "
    "tests, then the decisions, then the tick-box tidy-ups) with the headings saying so, and "
    "shorter sentences. The three tabs are the same three groups the earlier sheet had, renamed "
    "\"1 Start here\", \"2 Decisions\" and \"3 Just a tick\" so the order to work in is visible "
    "from the tab strip on a phone.",

    "THE EARLIER WORKBOOK IS SUPERSEDED, NOT DELETED. Both files sit in this folder. Send THIS "
    "one; the earlier pair is kept as the record of what was verified and when. Its QA tab and "
    "this one carry the same mapping - the rows and notes below are IMPORTED from "
    "gen_chris_sheet.py at build time, not retyped, so the two cannot drift.",

    "NO SOURCE WAS RE-FETCHED FOR THIS PASS AND NONE NEEDED TO BE. The content was verified live "
    "against the six specifications earlier on 2026-08-06 (Sales By Customer v15, Sales By "
    "Representative v17, Parts Velocity v5, Technician Utilization v6, Work In Progress v9, "
    "Inventory Value v4) and this pass only rewrote the presentation. NOTHING WAS WRITTEN "
    "ANYWHERE: no TestRail call, no Jira call, no Confluence call, no application call.",
]


def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    gc.alignment = WRAP
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws


def write_xlsx():
    wb = openpyxl.Workbook()
    w = [4, 40, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "Report Suite - questions for Chris Ward - 2026-08-06",
           HELLO,
           "Section 1 of 3 - please start here. One yes-or-no, and seven of our tests come off "
           "hold the same day.",
           TAB1, w, first=True)

    _sheet(wb, TAB2_NAME,
           "Section 2 - five decisions we need from you",
           "Five questions, each a plain A or B. Two of them were asked on 5 August and are here "
           "again because they are still open, not because we forgot. Short answers are perfect.",
           "Section 2 of 3 - five decisions. Six of our tests are waiting across them.",
           TAB2, w)

    _sheet(wb, TAB3_NAME,
           "Section 3 - seven small tidy-ups in your own descriptions",
           "NOTHING TO DECIDE ON THIS TAB. Six of these are the unfinished half of a decision you "
           "have already made, and one is a sentence you have already written for three reports "
           "which we are asking for on the other three. A tick is enough on each.",
           "Section 3 of 3 - seven tidy-ups, a tick each. No test is blocked by these.",
           TAB3, w)

    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail case IDs, "
                 "requirement anchors, live evidence and the corrections to our own records live "
                 "here so the tabs Chris reads stay plain (Standing Rules 7, 8 and 55).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY QUESTION CHRIS SEES").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Tab", "Item", "What it asks", "Where the question comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in prior.QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in EXTRA_NOTES + list(prior.QA_NOTES):
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, wd in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 62, 62]):
        ws4.column_dimensions[col].width = wd

    wb.save(XLSX)
    return XLSX


def block(items):
    out = []
    for i, (topic, now, q, opts) in enumerate(items, 1):
        out.append(f"### {i}. {topic}\n")
        out.append("**What happens now**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
        out.append("**The question**\n")
        out.append(f"> {q}\n")
        out.append("**Options**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
        out.append("**Your answer:** _______________________________________________\n")
    return "\n".join(out)


def write_md():
    hello_md = HELLO.replace("\n\n", "\n\n")
    md = f"""# Questions for Chris Ward — Report Suite — 2026-08-06

**Project: Report Suite (the six reports) · Product Owner: Chris Ward**

*This is the friendly, forward-as-is version. It carries all 13 items of
`Questions-for-Chris-Ward_Report-Suite_2026-08-06.md` — nothing dropped, nothing added — reordered
by what to do first and rewritten to read easily on a phone. The spreadsheet twin is
`Questions-for-Chris-Ward_Report-Suite_Friendly-Version_2026-08-06.xlsx`; it carries a QA-only tab
that must not be forwarded.*

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

---

{hello_md}

---

## Section 1 — Start here: one line, and seven tests can run

**This is the only thing we are really waiting on.**

{block(TAB1)}

---

## Section 2 — Five decisions we need from you

Each one is a plain A or B. Two were asked on 5 August and are here again because they are still
open, not because we forgot.

{block(TAB2)}

---

## Section 3 — Seven small tidy-ups: a tick each, nothing to decide

Six are the unfinished half of a decision you have already made. One is a sentence you have already
written for three reports, which we are asking for on the other three. No test is blocked by any of
them.

{block(TAB3)}

---

## QA-only — not for Chris

The question-to-case mapping is on the spreadsheet's `{TAB4_NAME}` tab — every question's affected
TestRail case IDs with links, the requirement anchors quoted from the live pages, and what each
possible answer resolves to. It is imported verbatim from the earlier sheet's generator so the two
cannot drift, with three notes added recording that this friendly version supersedes the earlier
pair for sending, that all 13 items are carried over unchanged, and that no source was re-fetched
because only the presentation changed.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
