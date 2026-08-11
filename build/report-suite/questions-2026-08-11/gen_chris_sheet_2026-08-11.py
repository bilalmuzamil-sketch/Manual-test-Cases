#!/usr/bin/env python3
"""Chris Ward question sheet — Report Suite, 2026-08-11 spec-delta round.

Mirrors the 2026-08-06 friendly sheet 1:1 (Standing Rule 16): same four columns, same
"#, Which project and report, What happens now, The question, Options, Your answer",
same QA-only final tab.

Standing Rules 7 + 55: every row names the PROJECT and the REPORT, because Chris owns
the Report Suite AND Fees & Discounts. Nothing he reads carries a case id, a
requirement anchor, an HTTP term or the word VIU — that lives on the QA-only tab.

RESEARCH ONLY. No TestRail, Jira, Confluence or application call.
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_2026-08-11.xlsx")
MD = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_2026-08-11.md")

HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Which project and report", "What happens now", "The question", "Options",
        "Your answer"]

HELLO = (
    "Hello Chris - thank you for the specification updates on the 7th and the 10th. Your rewrite of "
    "the location-column wording answered a question we had been holding tests for, so four tests "
    "have come off hold already - no reply needed on that one.\n\n"
    "Three things are still open on the REPORT SUITE (the six new reports). Question 1 is the "
    "important one and it holds three tests today. The other two are small. Short answers are "
    "perfect - a letter, or one line."
)

ROWS = [
    ("GROUP", "QUESTION 1 OF 3 - THIS IS THE ONE THAT MATTERS. IT HOLDS THREE TESTS TODAY."),
    ("1",
     "REPORT SUITE - the Work In Progress report - the four tabs",
     "The Work In Progress report splits jobs across four tabs: Estimates, Approved - not started, "
     "Approved - partially completed, and Completed. Two parts of the written description now say "
     "different things about how a job gets into a tab, and we cannot tell which you meant.\n\n"
     "One part says a whole job goes into ONE tab, chosen by the job's own status.\n\n"
     "A newer part, added on the 10th, says each LINE on the job is placed separately - so a job "
     "with some lines approved and some not would show up in TWO tabs at once, each showing only "
     "that tab's share of the money.",
     "Which of these is right?",
     "A) One job, one tab, chosen by the job's status. (This is what our tests check today.)\n\n"
     "B) Lines are placed separately, so one job can appear in more than one tab, each showing "
     "only its own share of the money.\n\n"
     "C) Something else - please describe it in a line.",
     ""),
    ("GROUP", "QUESTION 2 OF 3 - SMALL, BUT IT IS THE LAST REPORT WHERE THIS IS UNCLEAR"),
    ("2",
     "REPORT SUITE - the Parts Velocity report - the Location column",
     "You settled this for five of the six reports on the 7th, and thank you - anyone who can reach "
     "more than one location sees the Location column, it is on by default, and they can switch it "
     "off from the column list.\n\n"
     "Parts Velocity still reads differently in two places. One part now matches the other five "
     "reports. But another part still says the column simply disappears when you narrow down to a "
     "single location, and the list of columns you can switch on and off still names twenty columns "
     "without Location among them.",
     "Should Parts Velocity work the same way as the other five reports?",
     "A) Yes - same as the others. On by default for anyone who can reach more than one location, "
     "switchable from the column list, and narrowing your selection does not hide it.\n\n"
     "B) No - Parts Velocity is different, and it should behave the way the older wording says.\n\n"
     "C) Something else - please describe it in a line.",
     ""),
    ("GROUP", "QUESTION 3 OF 3 - A NUMBER TO CONFIRM. YOUR OWN NOTE ASKS FOR THIS."),
    ("3",
     "REPORT SUITE - the download size limit - all six reports",
     "All six reports now refuse a download when there is too much in it, and show: \"This report "
     "is too large to export. Narrow the date range or filters, then try again.\"\n\n"
     "The limit is written as 10,000 rows. On the Inventory Value report your own note beside it "
     "says the figure is a proposed default and asks for the exact number to be confirmed before "
     "the work is built.",
     "Is 10,000 rows the right limit, and does it apply to all six reports?",
     "A) Yes - 10,000 rows, and the same on all six.\n\n"
     "B) A different number - please tell us what it is.\n\n"
     "C) It differs between reports - please tell us which is which.",
     ""),
]

QA = [
    ["Q", "Report", "Requirement", "Cases affected", "What it blocks", "Our position"],
    ["1", "Work In Progress",
     "WIP spec v11 §3 Key Decisions (line-state bucketing, per SV-9027) CONTRADICTS the unchanged "
     "S2-R4 / S3-R1 / S3-R2 / S3-R3 / S3-R4, which place a whole WO in one tab by status.",
     "C30458, C30462 (Automated), C30464 held. Consequentially affected and NOT edited: C30480, "
     "C30491, C30488, C30452, and C30528 (the nightly snapshot's one-row-per-job shape).",
     "3 cases on AUTOMATION: HOLD. Also blocks authoring for two surfaces with no coverage: the "
     "per-tab money slice, and line-level vs WO-level ageing (Days Open).",
     "No side picked (Rules 15/57/58). Assertions preserved verbatim; divergence disclosed in the "
     "case text; markers HOLD."],
    ["2", "Parts Velocity",
     "PV v6 rewrote S3-R10 to access-gated + picker-toggleable but left S2-R12 scope-driven, and "
     "S4-R1/S4-R2/S4-R3 still enumerate 20 picker columns excluding Location.",
     "C38914, C30352 (Automated).",
     "Nothing on hold — both cases assert only the uncontested half and tell the tester not to fail "
     "on the contested points.",
     "Uncontested: the column exists, values are per-row, merged special-order rows read Multiple, "
     "a single-location-access user never sees it."],
    ["3", "All six",
     "PV S6-R12 / TU S7-R14 / WIP S9-R11 / SBC S14-R16+S15-R25 / SBR S14-E2 / IV S10-R12. IV v5 "
     "carries the bracketed note: \"[Cap value 10,000 is a proposed default — confirm the exact "
     "suite-standard value with the owner before dev.]\"",
     "C30172, C30290, C38885, C38887, C38918, C30593 (one per report; all six covered).",
     "Nothing on hold. All six cases assert 10,000 and the verbatim message already.",
     "Coverage is complete; only the NUMBER is unconfirmed. If it changes, six cases need one edit "
     "each."],
    ["-", "ANSWERED ALREADY - do not re-ask", 
     "The 2026-08-06 sheet's question 1 (is the Location column access-gated and toggleable?) was "
     "ANSWERED by Chris's own spec edits of 2026-08-07: SBR v18, PV v6 and IV v5 reworded the "
     "governing requirement, message \"reworded the Location-column visibility to the access-gated, "
     "column-selector-toggleable rule\".",
     "C38917, C30551, C30554, C30588 — all four HOLDs lifted 2026-08-11.",
     "-",
     "Recorded so nobody re-asks a question a source has already answered."],
]


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"
    ws["A1"] = "Report Suite - questions for Chris Ward - 2026-08-11"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = HELLO
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 96
    ws.append([])
    ws.append(COLS)
    for i, c in enumerate(ws[4], 1):
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, WRAP
    for r in ROWS:
        if r[0] == "GROUP":
            ws.append([r[1]])
            ws.cell(ws.max_row, 1).fill = GRP_FILL
            ws.cell(ws.max_row, 1).font = Font(bold=True)
            continue
        ws.append(list(r))
        for c in ws[ws.max_row]:
            c.alignment = WRAP
        ws.row_dimensions[ws.max_row].height = 170
    for col, w in zip("ABCDEF", (5, 34, 60, 40, 52, 26)):
        ws.column_dimensions[col].width = w

    q = wb.create_sheet("QA internal - not for Chris")
    for row in QA:
        q.append(row)
        for c in q[q.max_row]:
            c.alignment = WRAP
    for i, c in enumerate(q[1], 1):
        c.fill, c.font = HDR_FILL, HDR_FONT
    for col, w in zip("ABCDEF", (5, 22, 58, 46, 46, 46)):
        q.column_dimensions[col].width = w
    wb.save(XLSX)

    md = ["# Report Suite - questions for Chris Ward - 2026-08-11", "", HELLO, ""]
    for r in ROWS:
        if r[0] == "GROUP":
            md += ["", f"## {r[1]}", ""]
            continue
        md += [f"### Question {r[0]} - {r[1]}", "", f"**What happens now.** {r[2]}", "",
               f"**The question.** {r[3]}", "", "**Options.**", "", r[4], "",
               "**Your answer:** ______", ""]
    open(MD, "w").write("\n".join(md) + "\n")
    print("wrote", XLSX)
    print("wrote", MD)


if __name__ == "__main__":
    build()
