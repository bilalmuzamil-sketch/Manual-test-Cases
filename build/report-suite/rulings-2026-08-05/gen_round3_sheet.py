#!/usr/bin/env python3
"""Generate the Chris Ward ROUND-3 clarification sheet (.xlsx + .md).

Mirrors gen_followup_sheet.py 1:1 (Standing Rule 16), which in turn mirrors
build/report-suite/chris-consolidated-2026-08-04/
Report-Suite_Questions-and-Decisions-for-Chris-Ward_2026-08-04.xlsx:
same six reader columns in the same order, same row layout (A1 title / A2 note /
A4 header / A5 band / items from row 6), same fills and fonts, same freeze pane,
same column widths, and the same QA-only mapping tab that is never sent.

Standing Rule 55: reader-facing text is extremely simple, names the PROJECT and
the REPORT on every question row (Chris owns the Report Suite AND Fees &
Discounts), and cites stories/the epic only where the reference orients him.

Standing Rule 45(e): where this sheet says two sentences disagree, BOTH are
quoted, in plain words, from the live Confluence page fetched immediately before
writing (versions recorded on the QA-only tab).

DOCUMENTATION ONLY - this script writes two files into this folder. It makes no
TestRail or Jira call of any kind.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_Round-3_2026-08-05.xlsx")
MD = os.path.join(HERE, "Questions-for-Chris-Ward_Report-Suite_Round-3_2026-08-05.md")

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Finish the location column"
TAB2_NAME = "Carried over from last sheet"
TAB3_NAME = "Still holding tests"
TAB4_NAME = "QA internal - not for Chris"

THANKS = (
    "Thank you - you have had a very heavy day on these: fifteen answers, and then you went and "
    "updated all six written descriptions on top of that. Almost everything we were waiting on is "
    "now settled, and this sheet is much shorter than the last one because of it. Nothing here is a "
    "complaint. Most of it is simply \"please finish the sentence you started\" - you have already "
    "made the decision, and a few paragraphs further down the same documents still say the old "
    "thing. Every question says which project and which report it is about, because we know you "
    "look after more than one thing here."
)

TAB1_NOTE = (
    "PLEASE START HERE - this is the one group that is holding real work. You have already DECIDED "
    "this: your updates today say the location column belongs to anyone who can reach more than one "
    "branch, that they see it straight away, and that they can switch it off in the column list. "
    "The trouble is that four of the six documents still contain an older paragraph saying the "
    "opposite, and one report was not changed at all. We are not asking you to decide again - only "
    "to tell us which sentence to keep. " + THANKS
)

# ------------------------------------------------------------------- the questions
# (topic, what-happens-now, the question, options)
TAB1 = [
    (
        "Report Suite - the Sales By Customer report - the column list "
        "(the \"show or hide columns\" story for this report, under epic SV-8582)",

        "You updated this document today and it now says the location column can be switched on and "
        "off in the column list.\n\n"
        "Further down, the same document still lists what is in that column list and says there are "
        "\"exactly nine\" switches - Date, Inv. Hrs, Labor Invoiced, Labor Margin, Parts Invoiced, "
        "Parts Margin, Shop Supplies, Margin, Margin % - and Location is not one of them. It also "
        "says the columns that are always on \"do not appear in the toggle list\", and it does not "
        "name Location as an exception either.\n\n"
        "So the two halves of the same document cannot both be true.\n\n"
        "Why we are asking: the tester has to count the switches in that list, and right now we "
        "cannot tell them whether to expect nine or ten.",

        "When someone who can reach several branches opens the column list on this report, how many "
        "switches should be in it?",

        "A) TEN - the nine you already list, plus Location. (This matches your update today; the "
        "\"exactly nine\" sentence is the leftover and can be corrected.)\n\n"
        "B) NINE - no Location switch. (Then the location column is not something a person can "
        "switch off, and your update today is the part that needs changing.)",
    ),
    (
        "Report Suite - the Sales By Representative report - the location column "
        "(the \"column selector\" story for this report, under epic SV-8582)",

        "The top of this document, and your own note on today's update, both say the location column "
        "is for anyone who can reach more than one branch, that they see it by default, and that "
        "they can switch it off in the column list.\n\n"
        "Further down, the requirement itself still says the opposite: the column is shown \"only "
        "when the current view spans more than one location\" and \"when the view is scoped to a "
        "single location the column is hidden\" - in other words it comes and goes on its own, with "
        "nobody switching anything.\n\n"
        "The same document also says the column list holds \"the seven toggleable metric columns\" "
        "and names them, and Location is not among the seven.\n\n"
        "Why we are asking: these are two different tests, and we can only write one of them.",

        "For this report, is the location column something a person switches on and off, or does it "
        "appear and disappear on its own?",

        "A) A PERSON SWITCHES IT - anyone who can reach more than one branch sees it by default and "
        "can switch it off. (This matches the top of your document and today's note; the "
        "\"only when the view spans more than one location\" paragraph is the leftover.)\n\n"
        "B) IT APPEARS ON ITS OWN - only while more than one branch is being looked at, and it is "
        "never in the column list. (Then the top of the document is the part that needs changing.)",
    ),
    (
        "Report Suite - the Work In Progress report - the location column "
        "(the \"show or hide columns\" story for this report, under epic SV-8582)",

        "This one is the clearest of the four, because both sentences are in the version you saved "
        "today, a few paragraphs apart.\n\n"
        "One says: the location column \"is offered in the column selector to any user with access "
        "to more than one location; for that user it is shown by default and can be toggled on or "
        "off.\"\n\n"
        "The other says: the location column \"is shown automatically whenever the current scope "
        "spans more than one location, and is hidden whenever a single location is in scope; the "
        "user does not toggle it in the column selector.\"\n\n"
        "Why we are asking: one of those two sentences is the leftover, and we would rather you told "
        "us which than have us guess and get it wrong in both directions.",

        "For this report, can a person switch the location column off in the column list?",

        "A) YES they can switch it off - the \"user does not toggle it\" sentence is the leftover.\n\n"
        "B) NO they cannot - it comes and goes on its own, and the \"offered in the column selector\" "
        "sentence is the leftover.",
    ),
    (
        "Report Suite - the Inventory Value report - the location column "
        "(the \"show or hide columns\" story for this report, under epic SV-8582)",

        "The top of this document, and your note on today's update, both say the location column is "
        "shown by default to anyone who can reach more than one branch and \"can be toggled off from "
        "the column selector\".\n\n"
        "Further down, the requirement still says it \"is shown only when the current scope spans "
        "more than one location\" and that \"it is not one of the columns offered in the "
        "column-selection control\".\n\n"
        "Why we are asking: this is the same leftover as on the other two reports - one sentence, "
        "and then a whole group of our tests can come off hold.",

        "For this report, is the location column offered in the column list?",

        "A) YES it is offered, switched on to start with - the \"not one of the columns offered\" "
        "sentence is the leftover.\n\n"
        "B) NO it is not offered - it comes and goes on its own, and the top of the document is the "
        "part that needs changing.",
    ),
    (
        "Report Suite - the Parts Velocity report - the location column "
        "(the \"choose which columns to show\" story for this report, under epic SV-8582)",

        "This is the one report you have not changed on this point at all - and we want to be "
        "careful, because you did save a new version of it today, so it is not that you missed it in "
        "an old document.\n\n"
        "It still says the location column \"is auto-managed by the location scope\", that it is "
        "\"not one of the 20 columns in the picker\", and that it \"is not user-toggleable\".\n\n"
        "That is now the odd one out: the other five reports have all moved to the version where a "
        "person can switch it off.\n\n"
        "Why we are asking: we do not want to change this report on the strength of what you decided "
        "for the others. If you meant it to be the same, say so and we will match it; if this report "
        "really is different, we will leave it exactly as it is and stop treating it as a leftover.",

        "Should this report work the same way as the other five, or is it deliberately different?",

        "A) THE SAME - a person who can reach more than one branch sees the location column and can "
        "switch it off, exactly as on the others. (Then this document has one paragraph left to "
        "update.)\n\n"
        "B) DELIBERATELY DIFFERENT - on this report the column comes and goes on its own and cannot "
        "be switched off. (Then nothing needs changing here and we will note it as intended.)",
    ),
]

TAB2 = [
    (
        "Report Suite - the Parts Velocity and Technician Utilization reports - downloads "
        "(the download stories on those two reports, under epic SV-8582)",

        "For Sales By Representative and for Inventory Value your descriptions say exactly where the "
        "line naming the branches sits: in the header area of a printable document, and as one of "
        "the short summary lines above the column headings in a spreadsheet.\n\n"
        "Parts Velocity and Technician Utilization do not say.\n\n"
        "Why we are asking: the tester has to be told where to look, and we would rather not invent "
        "a position for it.",

        "Should those two reports put that line in the same places as the others?",

        "A) YES - the same as the other reports: the header area of a printable document, and a "
        "summary line above the column headings in a spreadsheet.\n\n"
        "B) NO - please say where it should go.",
    ),
    (
        "Report Suite - all six reports - anything on screen naming the branches "
        "(the branch-filter stories on each report, under epic SV-8582)",

        "Our tests used to tell the tester to look for something on the page naming the branches "
        "currently being shown - separate from the branch chooser itself.\n\n"
        "We searched all six of your descriptions for it and found no mention of such a thing "
        "anywhere. So we have taken it out of the tests rather than leave a tester hunting for "
        "something that may never have been meant to exist. We would rather tell you that plainly "
        "than quietly leave it in.\n\n"
        "Why we are asking: if it should be there, we need to put it back and it is a developer job. "
        "If not, we have already done the right thing and you can simply confirm it.",

        "Should there be something on the page naming the branches you are looking at, beyond the "
        "branch chooser itself?",

        "A) NO - the branch chooser already shows it. Nothing else is needed, and removing it was "
        "correct.\n\n"
        "B) YES - please say what it should say and where it should sit.",
    ),
    (
        "Report Suite - the Sales By Customer report - downloaded file names "
        "(the download story for this report, under epic SV-8582)",

        "Your update today changed the date list to nine choices and removed Today and Yesterday.\n\n"
        "Another part of the same document still explains what a downloaded file should be called for "
        "a Today range and for a Yesterday range.\n\n"
        "Why we are asking: it is almost certainly just a leftover, but we do not want to delete "
        "something from a test on our own guess.",

        "Can that leftover be removed, or are Today and Yesterday still meant to exist somewhere?",

        "A) REMOVE IT - Today and Yesterday are gone for good.\n\n"
        "B) THEY STILL EXIST somewhere - please say where.",
    ),
    (
        "Report Suite - the Technician Utilization report - the column button "
        "(the \"show or hide columns\" story for this report, under epic SV-8582)",

        "Your description says that hovering over the column button shows the words "
        "\"Column Selection\".\n\n"
        "It does not say what someone using a screen reader should hear when they land on that "
        "button.\n\n"
        "Why we are asking: a screen reader has to read out something, and if we make the words up "
        "the test is only checking our own invention.",

        "Should a screen reader read out the same words, \"Column Selection\"?",

        "A) YES - the same words.\n\n"
        "B) Something else - please say what.",
    ),
    (
        "Report Suite - the Technician Utilization report - spreadsheet downloads "
        "(the download story for this report, under epic SV-8582)",

        "Following your answer about this report's download menu, we wrote a test for two different "
        "spreadsheet downloads - a short one and a full one.\n\n"
        "The product currently offers only one spreadsheet, and your description does not mention a "
        "second.\n\n"
        "Why we are asking: if there should only ever be one, we will delete the test rather than "
        "leave it sitting there unused.",

        "Should this report offer two spreadsheet downloads, or only one?",

        "A) ONE is correct - we will delete the extra test.\n\n"
        "B) TWO - please say what each one should contain.",
    ),
]

TAB3 = [
    (
        "Report Suite - the Sales By Customer report - where it sits in the menu "
        "(the \"where the report lives and who can open it\" story, under epic SV-8582)",

        "This is the third time we have put this one in front of you, and we are only repeating it "
        "because a test is genuinely stuck on it - it is not a nag.\n\n"
        "We need to know which menu group this report belongs in. Our notes from the product show it "
        "under a group named SALES; your description says Performance. We cannot tell which one is "
        "meant to be right, so we cannot tell whether the product is wrong or our test is.\n\n"
        "Why we are asking: one word from you and the test is either correct as written or reworded "
        "in a minute.",

        "Which menu group should this report appear in?",

        "A) Performance - below the links that were already there.\n\n"
        "B) Sales - below the links that were already there.\n\n"
        "C) Somewhere else - please say where.",
    ),
    (
        "Report Suite - the Sales By Representative report - the word on screen "
        "(the sales-representative stories for this report, under epic SV-8582)",

        "You confirmed that the full word \"Representative\" is right in the downloaded files, and we "
        "have matched our tests to it.\n\n"
        "Two tests are still parked because they are about the word on the SCREEN and on the "
        "customer's record, not in the files. You were only asked about the files at the time, and we "
        "did not want to stretch your answer to cover screens you had not been shown.\n\n"
        "Why we are asking: two tests come unparked the moment you answer, whichever way you "
        "answer.",

        "Should the full word \"Representative\" also be used on the screen and on the customer's "
        "record?",

        "A) YES - the full word everywhere it appears, not only in the downloaded files.\n\n"
        "B) NO - only the downloaded files matter; the screen can stay as it is.",
    ),
    (
        "Report Suite - the date chooser - all six reports share it "
        "(the date-range stories on each report, under epic SV-8582)",

        "Your update today set the date list to nine choices, with \"Last 12 Months\" first, and said "
        "there is no Today, no Yesterday and nothing called Custom.\n\n"
        "When we tried the reports, the two you removed still work, and the new first choice - Last "
        "12 Months - is refused. So the product looks like it is still on the old list.\n\n"
        "We have NOT raised this as a fault yet, for one honest reason: the new list is only hours "
        "old, and we would rather check that it is what you intend before asking anyone to build it.\n\n"
        "Why we are asking: this one chooser is shared by all six reports, so it decides the wording "
        "of six tests.",

        "Is the nine-choice list you wrote today - starting with Last 12 Months, and with no Today, "
        "Yesterday or Custom - what you want built?",

        "A) YES - that is the intended list. We will write the tests to it and raise the difference "
        "with the developers.\n\n"
        "B) NO - please say what the list should be.",
    ),
]

REMINDER = (
    "NO ANSWER NEEDED - two notes for when you are next in the descriptions. FIRST: the Technician "
    "Utilization description is now correct about the location column in its requirements, so it is "
    "not on the list above - but three sentences elsewhere in it still describe the old behaviour "
    "(\"hidden when a single location is in scope\"). Nothing is blocked by it; it is just untidy, "
    "and a reader of only those sentences would test the wrong thing. SECOND: the seven small wording "
    "tidy-ups from the last sheet are still open and still holding nothing up - where the Technician "
    "Utilization report sits in the menu; the machine chooser on Work In Progress; the line calling "
    "Parts Velocity the only report in its group; the line saying the Escape key closes the "
    "deactivate pop-up; the download size limit missing from three descriptions; a short note that "
    "the vehicle-number field also holds serial numbers for things that are not vehicles; and some "
    "garbled characters in two descriptions. Tell us if you would like us to keep listing those or "
    "to stop."
)

ALREADY_ANSWERED = (
    "THINGS WE DELIBERATELY DID NOT ASK YOU, because you have already answered them. Your updates "
    "today settled four things we had queued up, and we would rather show you that than have you "
    "wonder why they vanished. (1) The logo rule: the ShopView logo stands in only when a logo has "
    "been uploaded but will not load, and when no logo has been uploaded at all no logo is printed "
    "and the text fills the space - now written into the description, so we have stopped citing your "
    "message and cite the description instead. (2) One reports permission for all six reports, not a "
    "separate permission per report. (3) On the Sales By Customer report, a person who can only "
    "reach one branch never sees Location and it never appears in their column list - that answers a "
    "question we had open. (4) On the same report, whether a person who can reach several branches "
    "but is looking at only one still gets the column - your update says yes, \"regardless of how "
    "many locations are currently selected\". Thank you: that was the single hardest one and it is "
    "now closed."
)


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws, r


# --------------------------------------------------------------- QA-only mapping
TR = "https://shopview.testrail.io/index.php?/cases/view/"


def links(*ids):
    return " · ".join(f"C{i} {TR}{i}" for i in ids)


QA_ROWS = [
    ("Tab 1", "1",
     "SBC: does the column selector list nine columns or ten?",
     "SBC v14 contradicts ITSELF, same version: S4-R12 (new today) says the Location column "
     "\"is shown by default and can be toggled on or off from the column selector, regardless of "
     "how many locations are currently selected\"; S13-R4 says \"The nine toggleable columns are, "
     "in order: Date, Inv. Hrs, Labor Invoiced, Labor Margin, Parts Invoiced, Parts Margin, Shop "
     "Supplies, Margin, Margin %\" and S13-R6 says the always-present columns \"do not appear in "
     "the toggle list\" without naming Location as an exception.",
     "SBC-COL-01 (C30156); SBC-LOC-04 (C38912) - the two contradict each other and each has a live "
     "requirement on its side",
     links(30156, 38912),
     "Confluence page 577634305, LIVE version 14, fetched 2026-08-05T16:56Z. Both anchors extracted "
     "verbatim from that fetch. This is audit item 2 and round-2 Q2.",
     "A (ten) -> C30156 is reworded and S13-R4 is Chris's to correct. B (nine) -> C38912 loses the "
     "toggle assertion and S4-R12 is Chris's to correct. Either way both come off HOLD."),
    ("Tab 1", "2",
     "SBR: is the Location column user-toggleable or automatic?",
     "SBR v16 Feature Overview + the v16 changelog message say access-gated and toggleable; S21-R7 "
     "still says \"shown on the report only when the current view spans more than one location… "
     "When the view is scoped to a single location the column is hidden\"; S20-R1 closes the "
     "dropdown at \"the seven toggleable metric columns\" and S20-R3 says the five always-visible "
     "columns \"do not appear in the dropdown\".",
     "SBR-LOC-05 (C38913). NOTE: SBR-COL-01 (C30265) is CORRECT AS WRITTEN against S21-R7 + "
     "S20-R1/R3 and must NOT be changed until this is answered - the 2026-08-05 audit records that "
     "explicitly.",
     links(38913, 30265),
     "Confluence page 585629698, LIVE version 16, saved 2026-08-05T13:33:14Z, fetched "
     "2026-08-05T16:56Z. v16 changelog verbatim: \"Location column changed to an access gate and "
     "made toggleable in the column selector\".",
     "A (toggleable) -> C38913 stands; S21-R7 and S20-R1/R2 are Chris's to correct; C30265 needs "
     "rewording. B (automatic) -> C38913 loses the toggle assertion; C30265 stays exactly as it is; "
     "the Feature Overview is Chris's to correct."),
    ("Tab 1", "3",
     "WIP: can the user toggle the Location column off?",
     "WIP v7 contradicts ITSELF, same version, a few paragraphs apart. S4-R3: \"The Location column "
     "is offered in the column selector to any user with access to more than one location; for that "
     "user it is shown by default and can be toggled on or off.\" S7-R13: \"…is shown automatically "
     "whenever the current scope spans more than one location, and is hidden whenever a single "
     "location is in scope; the user does not toggle it in the column selector.\"",
     "WIP-COL-01 (C30466); WIP-COL-02 (C30467); WIP-EXP-02 (C30511); WIP-FLT-09 (C38916). ALSO "
     "UNBLOCKS WIP-PERS-05 (C43551), which is a separate HOLD and is NOT asked as its own question: "
     "WIP v7 S8-R7 already says the report remembers \"column selection\", so once S4-R3 wins, the "
     "memory question is answered by the spec.",
     links(30466, 30467, 30511, 38916, 43551),
     "Confluence page 703660034, LIVE version 7, saved 2026-08-05T13:33:12Z, fetched "
     "2026-08-05T16:56Z. Both anchors extracted verbatim. C30466's precondition currently instructs "
     "the tester to switch the column back on, which S7-R13 says is impossible.",
     "A (toggleable) -> the 4 WIP cases take the access-gate model, C43551 is answered by S8-R7, "
     "S7-R13 is Chris's to correct. B (automatic) -> the 4 keep the scope model, C43551 stays "
     "blocked, S4-R3 is Chris's to correct."),
    ("Tab 1", "4",
     "IV: is the Location column offered in the column-selection control?",
     "IV v4 Feature Overview + the v4 changelog say access-gated, \"shown by default and can be "
     "toggled off from the column selector\"; S7-R6 still says \"shown only when the current scope "
     "spans more than one location… Its visibility follows the location scope automatically and it "
     "is not one of the columns offered in the column-selection control\".",
     "IV-COL-01 (C30551); IV-COL-04 (C30554); IV-EXP-02 (C30588); IV-LOC-06 (C38917)",
     links(30551, 30554, 30588, 38917),
     "Confluence page 720142338, LIVE version 4, saved 2026-08-05T13:33:13Z, fetched "
     "2026-08-05T16:56Z. Both texts extracted verbatim from that fetch.",
     "A (offered) -> the 4 take the access-gate model and S7-R6 is Chris's to correct. "
     "B (automatic) -> the 4 keep the scope model and the Feature Overview is Chris's to correct."),
    ("Tab 1", "5",
     "PV: same as the other five, or deliberately different? (NO contradiction here - a genuine "
     "ruling is needed)",
     "PV v5 was SAVED TODAY at 13:21:40Z and was NOT changed on this point. S3-R10 still reads: "
     "\"The column is auto-managed by the location scope (it is not one of the 20 columns in the "
     "picker, S4-R1, and is not user-toggleable) and is hidden entirely when a single location is "
     "in scope.\" Because he republished the page AFTER giving the ambiguous answer, latest-wins "
     "(Rule 32) currently points AT the specification for this report.",
     "PV-COL-02 (C30352); PV-FILT-14 (C38914). C30352 matters historically: an earlier pass "
     "overwrote its spec-correct line and recorded it as \"wrong under both readings\" - it was PV "
     "S3-R10 almost verbatim.",
     links(30352, 38914),
     "Confluence page 620888066, LIVE version 5, fetched 2026-08-05T16:56Z. This is the only one of "
     "the five that is a RULING rather than a contradiction, and the sheet says so in its own text.",
     "A (same as others) -> the 2 take the access-gate model and S3-R10 is Chris's to correct. "
     "B (different) -> the 2 keep S3-R10 exactly and we record PV as a deliberate exception in the "
     "decisions register."),
    ("Tab 2", "1",
     "PV and TU: where does the \"Locations:\" line sit in an export?",
     "Round-2 Q3 / audit item 4. SBR S14-R20 and IV S10-R15 specify it; PV and TU are silent.",
     "PV-EXP-02 (C30376) - currently READY, carries a \"confirmed in the build\" hedge; "
     "TU-EXP-04 (C30437) - also one of the 16 Location HOLDs",
     links(30376, 30437),
     "Searched PV v5 and TU v6 live for an equivalent of SBR S14-R20: none. The hedge in both cases "
     "is an unsourced deferral to the build (audit class B).",
     "A -> both cite the SBR/IV position and the hedge is replaced by a citation. B -> both are "
     "reworded to whatever he specifies. Either way the class-B hedge goes."),
    ("Tab 2", "2",
     "Is there an on-screen scope indicator at all? (we removed it - telling him plainly)",
     "Round-2 Q4 / audit item 5. Searched ALL SIX live specs: ZERO mentions of any on-screen scope "
     "indicator. We invented it and then deferred its appearance to the build - audit class D, "
     "repaired by DELETION per Rule 25/57 (never by substituting build behaviour).",
     "The 7 already repaired by removal: SBC-LOC-03 (C30111); SBR-LOC-03 (C30215); PV-FILT-10 "
     "(C30337); TU-LOC-02 (C30443); WIP-FLT-06 (C30503); IV-LOC-01 (C30574); IV-LOC-02 (C30575)",
     links(30111, 30215, 30337, 30443, 30503, 30574, 30575),
     "All six specs fetched live 2026-08-05T16:56Z. What each spec DOES support - that the Location "
     "FILTER shows the current selection - is already asserted elsewhere in the same cases, so "
     "nothing was lost by the deletion.",
     "A -> the removal is confirmed, no case changes, no ticket. B -> the sentence goes back in and "
     "it is a developer job. No case is on HOLD for this - all 7 are READY - so this question "
     "confirms a decision rather than unblocking anything."),
    ("Tab 2", "3",
     "SBC: the export-filename map still lists Today and Yesterday",
     "Round-2 Q5 / audit item 6. SBC v14 S2-R2 deleted both presets today; S14-R14 still maps "
     "\"Today -> today; Yesterday -> yesterday\".",
     "SBC-EXP-02 (C30160) - currently READY",
     links(30160),
     "Both anchors read live from page 577634305 v14, 2026-08-05T16:56Z. Spec residue from his own "
     "edit of the same day.",
     "A -> C30160's filename map drops today/yesterday and cites S14-R14 as corrected. B -> the "
     "presets survive somewhere and S2-R2 needs correcting instead."),
    ("Tab 2", "4",
     "TU: is the accessible name of the column-selection button specified?",
     "Round-2 Q6 / audit item 7. TU S10-R1 gives the hover tooltip \"Column Selection\"; no "
     "requirement gives the accessible name.",
     "TU-COL-01 (C38859) - currently HOLD, reason \"this part of the report is not built yet\"",
     links(38859),
     "TU v6 read live 2026-08-05T16:56Z. The case's own text hedges with \"the exact wording is "
     "confirmed in the build\" - audit class B.",
     "A -> the case asserts \"Column Selection\" on his authority and the hedge goes. B -> it "
     "asserts whatever he names. The case stays HOLD on the not-built ground either way."),
    ("Tab 2", "5",
     "TU: two spreadsheet downloads, or one?",
     "Round-2 Q7. His answer chose the longer download-menu wording but never gave the count; we "
     "authored a second spreadsheet case on the strength of it.",
     "TU-EXP-10 (C43552) - HOLD, \"the two spreadsheet downloads do not exist yet and the product "
     "owner has not confirmed them\"",
     links(43552),
     "TU v6 S7-R7 says the spreadsheet always holds the technician rows and the Summary row and "
     "does not change with the Summary/Expanded choice - i.e. it describes ONE spreadsheet.",
     "A (one) -> C43552 is deleted, ready-to-automate +0 (it leaves the population). B (two) -> "
     "C43552 comes off HOLD, +1, and a developer ticket is written."),
    ("Tab 3", "1",
     "SBC: which menu group, and below which links?",
     "THIRD time of asking - left blank on the 2026-08-04 sheet and again on the 2026-08-05 sheet.",
     "SBC-NAV-01 (C30096) - HOLD, \"waiting on an answer from the product owner\"",
     links(30096),
     "SBC v14 Story 1 says Performance. Our own build notes of 2026-08-03 record a SALES group. NOT "
     "re-observed this pass - stated as such rather than implied (Rule 12).",
     "A (Performance) -> C30096 stands and a developer ticket is raised. B (Sales) -> C30096 is "
     "reworded. Either way +1 to ready-to-automate."),
    ("Tab 3", "2",
     "SBR: \"Representative\" on screen and on the customer record too?",
     "THIRD time of asking - left blank twice. His earlier answer covered only the download "
     "heading, and extending it to screens he was not shown would be stretching it (Rule 12).",
     "SBR-WO-01 (C30310); SBR-WO-06 (C30315) - both HOLD",
     links(30310, 30315),
     "SBR v16 changelog verbatim: \"download column heading set to 'Representative'\" - the download "
     "only. Read live 2026-08-05T16:56Z.",
     "A -> both keep the full-word expectation and a developer ticket is raised. B -> both are "
     "reworded to the short form. Either way +2 to ready-to-automate."),
    ("Tab 3", "3",
     "The nine-preset date list: is it what he intends? (asked WITHOUT technical detail, Rule 51)",
     "SBC v14 S2-R2 (six hours old when observed) sets nine presets with \"Last 12 Months\" first "
     "and states \"There is no Today, no Yesterday, and no option labeled 'Custom'\". Observed live "
     "on v3.5-16cf83f: the two deleted presets still work and the NEW FIRST ONE is refused.",
     "The six date cases, all READY: SBC-DATE-01 (C30102); SBR-DATE-01 (C30201); PV-FILT-03 "
     "(C30330); TU date case; WIP-FLT-04 (C30501); IV-DATE-01 (C30561). Plus SBC-EXP-02 (C30160) "
     "via Tab 2 question 3.",
     links(30102, 30201, 30330, 30501, 30561, 30160),
     "The evidence is an endpoint observation (API-ASK.md ASK 1) and the picker itself was NOT "
     "driven on screen. Per Rule 51 NOTHING has been filed and the reader-facing wording carries no "
     "technical detail at all - it says only that the two work and the new one is refused.",
     "A -> the six cases are written to the nine-preset list and a developer ticket becomes "
     "writable AFTER the picker is driven on screen. B -> he names the real list and the six are "
     "written to that instead."),
]

QA_NOTES = [
    "SOURCE CURRENCY (Standing Rule 31 + 45(e)) - ALL SIX SPECS FETCHED LIVE IMMEDIATELY BEFORE "
    "WRITING, at 2026-08-05T16:56:26Z, Confluence REST /wiki/rest/api/content/{id}"
    "?expand=version,body.storage, all HTTP 200: Sales By Customer page 577634305 = v14 "
    "(2026-08-05T13:07:07Z) · Sales By Representative 585629698 = v16 (13:33:14Z) · Parts Velocity "
    "620888066 = v5 (13:21:40Z) · Technician Utilization 641400833 = v6 (13:33:10Z) · Work In "
    "Progress 703660034 = v7 (13:33:12Z) · Inventory Value 720142338 = v4 (13:33:13Z). Every "
    "quoted sentence on this sheet comes from THAT fetch, not from a mirror. The Confluence version "
    "NUMBER is used throughout, never any in-body version field (Rule 31(a)).",

    "WHY THAT MATTERED TODAY: the 2026-08-05 final-VIU pass ran against SBR v15 / TU v5 / WIP v6 / "
    "IV v3 and all four moved UNDER IT between 13:33Z and 14:23Z. Four of the five Tab-1 questions "
    "were re-confirmed against the NEW versions before being written, and the wording changed as a "
    "result: they are no longer \"please decide\", they are \"please finish\". Nothing on this sheet "
    "asks him about a paragraph he has already fixed.",

    "TU IS DELIBERATELY NOT ON TAB 1, and this is the clearest example of the re-check paying for "
    "itself. TU v6 S9-R9 AND S10-R4 both now state the access-gate + toggleable model, so its "
    "REQUIREMENTS are internally consistent and there is no question to ask. Three sentences "
    "elsewhere in TU v6 still carry the old \"more than one location in scope\" phrasing (the "
    "Feature Overview paragraph, and the cross-references in S7-R13 and S8-R15), so TU appears in "
    "the no-answer-needed reminder as a tidy-up, NOT as a question. Asking it as a fifth "
    "contradiction would have been wrong.",

    "THREE CASES ARE ON HOLD 'WAITING ON THE PRODUCT OWNER' AND SHOULD NOT BE - THEY ARE OUR OWN "
    "DEFECT, NOT HIS. Deliberately NOT asked, because a source already answers each one, and asking "
    "would be noise (Rule 55's 'if he has answered it, do not ask again'). (1) SBC-VIS-02 = C30186 "
    "https://shopview.testrail.io/index.php?/cases/view/30186 - all five of its assertions are "
    "SBC v14 S20-R8, S20-R9, S20-R10, S20-R11 and S20-R14 almost verbatim, checked live this pass; "
    "there is no product question behind it. (2) WIP-FLT-05 = C30502 "
    "https://shopview.testrail.io/index.php?/cases/view/30502 - WIP v7 S7-R8 says \"A Custom range "
    "is capped at a 366-day maximum span\" and the build accepts 367; per Rule 57 that is a "
    "DEVELOPER DEFECT with the documented expectation kept, not a PO question. (3) SBC-COL-04 = "
    "C43550 https://shopview.testrail.io/index.php?/cases/view/43550 - SBC v14 S4-R12 answers it "
    "outright: \"a user with access to a single location is never shown it and it never appears in "
    "their column selector\". ALL THREE need an authorised repair pass, not an answer: +3 to "
    "ready-to-automate for free, and one defect ticket to write.",

    "CASE-COUNT ARITHMETIC, stated so it can be checked. Current ready-to-automate = 440 (423 READY "
    "+ 17 READY-EXPECT-FAIL, out of 473). 33 cases carry HOLD: 16 on the Location column · 8 not "
    "built · 7 'waiting on an answer from the product owner' · 1 needs a logo-load check · 1 the two "
    "spreadsheet downloads. ANSWERING TAB 1 releases exactly the 16 -> 456. Tab 3 questions 1 and 2 "
    "release C30096, C30310, C30315 -> 459. Tab 2 question 5 answered 'two' releases C43552 -> 460. "
    "The three mis-held cases in the note above add +3 with no answer needed at all -> 463. "
    "WIP-PERS-05 C43551 is released by Tab 1 question 3 rather than by a question of its own.",

    "WORDING RULES APPLIED (Standing Rules 7 + 55). Every reader-facing row names the PROJECT "
    "(Report Suite) and the REPORT, in the Topic column, because Chris also owns Fees & Discounts "
    "and 'the date filter' would be ambiguous to him. Story/epic references are given in plain form "
    "only where they orient him - epic SV-8582 and a plain description of the story ('the show or "
    "hide columns story'), never a bare key on its own. NO case IDs, NO requirement anchors, NO HTTP "
    "terms, NO endpoint names, NO enum names, NO bug codes and NOT ONE use of the word VIU appear in "
    "anything he reads. Each question carries a one-line 'why we are asking' so the consequence is "
    "visible. Where the sheet says two sentences disagree, BOTH are quoted in plain words (Rule "
    "45(e)).",

    "TONE. He answered 15 questions and edited all six specifications in one day. Tab 1's note says "
    "in his own reading order that he has ALREADY DECIDED and is only being asked which leftover "
    "sentence to drop. The 'already answered' block exists so he can see what his work today closed, "
    "rather than wondering why four queued questions disappeared.",

    "NOTHING HAS BEEN WRITTEN ANYWHERE. This sheet is a draft for the QA lead to send. No TestRail "
    "write, no Jira write, no case edit and no CLAUDE.md edit was made in producing it. The only "
    "calls made were read-only Confluence GETs and read-only Jira GETs.",

    "THE BUILD IS STILL A PARTIAL SOURCE. v3.5-16cf83f, branch not declared final, so the Rule-49 "
    "queue viu-2026-08-03/RECHECK-QUEUE.md stays OPEN and every pass/fail verdict on this project "
    "is PROVISIONAL. That does not affect any question on this sheet - all thirteen are decided by "
    "documents, not by the build.",
]


def write_xlsx():
    wb = openpyxl.Workbook()
    w_narrow = [4, 34, 52, 42, 46, 22]
    w_wide = [4, 34, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "Please finish the location column - Report Suite - Chris Ward - 2026-08-05 (round 3)",
           TAB1_NOTE,
           "You have already decided this - we only need to know which sentence to keep",
           TAB1, w_wide, first=True)

    _sheet(wb, TAB2_NAME,
           "Carried over from the last sheet - Report Suite - Chris Ward - 2026-08-05 (round 3)",
           "Five questions from the last sheet that your updates today did not touch. None of them "
           "is urgent. " + THANKS,
           "Still open from the previous sheet", TAB2, w_narrow)

    ws3, r = _sheet(wb, TAB3_NAME,
                    "Still holding tests - Report Suite - Chris Ward - 2026-08-05 (round 3)",
                    "Three questions, each of which frees a parked test the moment you answer it. "
                    "The first two have been asked twice before and we are only repeating them "
                    "because a test is genuinely stuck. " + THANKS,
                    "Three answers would free three parked tests", TAB3, w_narrow)
    ws3.cell(row=r + 1, column=2, value=REMINDER).alignment = WRAP
    ws3.cell(row=r + 1, column=2).font = Font(bold=True)
    ws3.row_dimensions[r + 1].height = 200
    ws3.cell(row=r + 3, column=2, value=ALREADY_ANSWERED).alignment = WRAP
    ws3.cell(row=r + 3, column=2).font = Font(bold=True)
    ws3.row_dimensions[r + 3].height = 220

    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail C-ids, "
                 "requirement anchors and live evidence live here so the reader-facing tabs stay "
                 "plain (Standing Rules 7, 8 and 55).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Tab", "Item", "What it asks", "Where the ambiguity comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 60, 60]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items, start=1):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, start):
            out.append(f"### Item {i}.0 — {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    tab1_note_md = TAB1_NOTE.replace(THANKS, "").strip()

    md = f"""# Questions for Chris Ward — Report Suite — round 3 — 2026-08-05

**Project: Report Suite (the six reports) · epic SV-8582 · Product Owner: Chris Ward**

**This is the plain-language twin of `Questions-for-Chris-Ward_Report-Suite_Round-3_2026-08-05.xlsx`.**
The spreadsheet is the version to send; it mirrors the 2026-08-04 and 2026-08-05 sheets' format
exactly, and it carries a QA-only tab that must not be forwarded.

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

{THANKS}

**Thirteen questions in total, each one a plain A or B, plus two notes that need no answer.**

**Live source versions confirmed at 2026-08-05T16:56:26Z, immediately before writing** — Sales By
Customer **v14** · Sales By Representative **v16** · Parts Velocity **v5** · Technician Utilization
**v6** · Work In Progress **v7** · Inventory Value **v4**. Every sentence quoted below comes from
that fetch.

---

## Tab 1 — {TAB1_NAME}

{tab1_note_md}

{block(TAB1)}

---

## Tab 2 — {TAB2_NAME}

Five questions from the last sheet that your updates today did not touch. None of them is urgent.

{block(TAB2)}

---

## Tab 3 — {TAB3_NAME}

Three questions, each of which frees a parked test the moment you answer it. The first two have been
asked twice before and we are only repeating them because a test is genuinely stuck.

{block(TAB3)}

### And two notes that need no answer

> {REMINDER}

> {ALREADY_ANSWERED}

---

## QA-only — not for Chris

The internal question-to-case mapping lives on the spreadsheet's `{TAB4_NAME}` tab: every question's
affected TestRail case IDs with links, the requirement anchors quoted verbatim from the live pages,
and what each possible answer resolves to. It also records the source-currency block, the
case-count arithmetic (**440 now → 456 on Tab 1 alone**), why Technician Utilization is deliberately
NOT on Tab 1, and the three cases we found are on hold **wrongly** and therefore did not ask about.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
