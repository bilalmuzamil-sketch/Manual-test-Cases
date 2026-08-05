#!/usr/bin/env python3
"""Generate the Chris Ward follow-up clarification sheet (.xlsx + .md).

Mirrors build/report-suite/chris-consolidated-2026-08-04/
Report-Suite_Questions-and-Decisions-for-Chris-Ward_2026-08-04.xlsx 1:1
(Standing Rule 16): same six reader columns in the same order, same row
layout (A1 title / A2 note / A4 header / A5 band / items from row 6),
same fills and fonts, same freeze pane, same column widths, and the same
QA-only mapping tab that is never sent.

Standing Rule 7 + the QA lead's ruling of 2026-08-05: reader-facing text is
extremely simple, names the PROJECT and the FEATURE on every question row
(Chris owns more than one thing), and cites stories/the epic only where the
reference orients him.

RESEARCH ONLY - this script writes two files into this folder. It makes no
TestRail or Jira call of any kind.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Follow-up-Question-for-Chris-Ward_2026-08-05.xlsx")
MD = os.path.join(HERE, "Follow-up-Question-for-Chris-Ward_2026-08-05.md")

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Needed first - location column"
TAB2_NAME = "Quick confirmations"
TAB3_NAME = "Still holding tests"
TAB4_NAME = "QA internal - not for Chris"

THANKS = (
    "Thank you for these - 15 answers in one go, and they unblocked most of the reporting work "
    "straight away. This is one short follow-up sheet, not a complaint: a handful of small "
    "clarifications, each one a plain A or B. Every question says which project and which report "
    "it is about, because we know you look after more than one thing here."
)

TAB1_NOTE = (
    "Needed first, please - this is the one that is holding a developer job. Everything else on "
    "this sheet can wait until you have a spare ten minutes. " + THANKS
)

# ------------------------------------------------------------------- the questions
# (topic, what-happens-now, the question, options)
TAB1 = [
    (
        "Report Suite - the location column - all six reports "
        "(this is the \"show or hide columns\" story on each report, under epic SV-8582)",

        "Your answer on the location column gave us a clear rule for two kinds of person, and we "
        "have used it. There is a third kind of person in between, and your answer points two ways "
        "for them, so we would rather ask than guess.\n\n"
        "You said the column shows by itself when someone (1) can see more than one branch AND "
        "(2) has chosen more than one branch. You also said the Location option should not be in "
        "the list of columns if someone cannot see more than one branch.\n\n"
        "The person in between is a manager who CAN see three branches but is looking at just one "
        "of them right now.\n\n"
        "This is not a rare case. Every one of the six reports opens on the single branch the "
        "person is working in, so this is exactly what a multi-branch manager sees the first time "
        "they open any report.\n\n"
        "Why we are asking: a developer has to build one of the two, and we cannot write the job "
        "down until we know which.",

        "For a manager who can see several branches but is looking at just one branch right now: "
        "is \"Location\" offered in the list of columns, or not?",

        "A) It is NOT in the list. While they are looking at one branch there is no Location "
        "option at all - they would have to add a second branch before the option appears. "
        "(Fewer things on screen; nobody can add a column that just repeats the same branch name "
        "on every row.)\n\n"
        "B) It IS in the list, switched off. They can switch it on whenever they like. "
        "(The manager is always in control; the column can end up repeating the same branch name "
        "on every row, which is harmless but a little pointless.)\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "Report Suite - the location column - all six reports "
        "(the \"column selection and persistence\" stories, for example SV-8664 and SV-8675)",

        "Your answer means a person can switch the location column off by hand. We do not know "
        "whether that choice should be remembered.\n\n"
        "The reports already remember which columns a person picked, so this may simply follow the "
        "same habit - but we did not want to assume it.\n\n"
        "Why we are asking: it is one line in a test either way, and we would rather have your "
        "word than our guess.",

        "If someone switches the location column off by hand, should it stay off the next time "
        "they open the report?",

        "A) Yes - remember it, exactly like every other column they turn on or off.\n\n"
        "B) No - it should come back on by itself each time they open the report with several "
        "branches chosen.",
    ),
]

TAB2 = [
    (
        "Report Suite - the Technician Utilization report - the download menu "
        "(story SV-8654, \"Export to PDF and CSV\")",

        "You chose the longer wording for this report's download menu and said consistency is key. "
        "We agree - we just need to know how many options there should be.\n\n"
        "The two reports you compared it with each offer four: a short version and a full version, "
        "each one as a document and as a spreadsheet. This report offers four today too, but its "
        "written description only describes three.\n\n"
        "Why we are asking: if it is four, there is a whole spreadsheet nobody has written a test "
        "for yet.",

        "Should this report's download menu offer four options, matching the other two reports?",

        "A) Yes - four: a short version and a full version, each as a document and as a "
        "spreadsheet.\n\n"
        "B) No - three is right, and one of the four should be taken away.",
    ),
    (
        "Report Suite - the date chooser - all six reports share it "
        "(the \"filter by date range\" story on each report, for example SV-8601)",

        "You chose \"keep what the product does today\", and added that the original date picker is "
        "the intentional one. We want to be certain we have read that the way you meant it.\n\n"
        "What is in the product today: nine ready-made ranges - Last 12 Months, This Year, Last "
        "Year, This Quarter, Last Quarter, This Month, Last Month, This Week, Last Week - plus a "
        "calendar you click dates on to build your own range. There is no Today, no Yesterday, and "
        "nothing called Custom.\n\n"
        "Why we are asking: this one chooser is shared by all six reports, so we are rewriting six "
        "tests on the strength of it.",

        "Is the date chooser that is in the product today the one you want kept?",

        "A) Yes - keep exactly what is there now. We correct our tests, and the written "
        "descriptions get tidied when you next have them open.\n\n"
        "B) No - Today, Yesterday and a Custom option should be put back in.",
    ),
    (
        "Report Suite - the Inventory Value report - the \"As of\" line in the downloads "
        "(stories SV-8672 and SV-8677)",

        "You confirmed the \"As of\" line belongs in the spreadsheet as well as in the printable "
        "document. Thank you - that settles it.\n\n"
        "The two files word it slightly differently. The spreadsheet says \"As of: 2026-08-03\" "
        "with a colon; the printable document says \"As of 2026-08-04\" with no colon.\n\n"
        "Right now our tests tell the tester not to report that difference. We would rather that "
        "instruction came from you than from us.\n\n"
        "Why we are asking: if it should match, it is a small developer job; if not, we leave it "
        "alone for good.",

        "Should the two files word that line in exactly the same way?",

        "A) No - the small punctuation difference is fine. Leave both as they are.\n\n"
        "B) Yes - both should read the same way. (We would raise it with the developers.)",
    ),
    (
        "Report Suite - the logo on printable downloads - all six reports "
        "(the download stories, for example SV-8613 and SV-8646)",

        "Your logo rule was clear about what happens when a logo is missing or will not load, and "
        "we have followed it. One word in it could mean two different things.\n\n"
        "You wrote \"if the customer has a logo selected, it appears\", and then in your corrected "
        "rule \"use the company's own uploaded logo\".\n\n"
        "A customer and a company are different people in ShopView: the company is the shop "
        "running it, and the customer is the person the shop is doing work for. Our tests today "
        "expect the SHOP's logo.\n\n"
        "Why we are asking: the two would look completely different on a printed report.",

        "Whose logo should appear at the top of a printable download?",

        "A) The shop's own logo - the business running ShopView.\n\n"
        "B) The customer's logo - the client the report is about.",
    ),
    (
        "Report Suite - all six reports - \"what is on screen should match the download\" "
        "(the download stories on each report, for example SV-8631)",

        "Alongside your answer about the missing columns you added a note: \"on-screen should match "
        "download\".\n\n"
        "That sentence turned out to be more useful than the question it came with - we have used "
        "it to settle a second thing, whether a download carries the location column whenever the "
        "screen shows it.\n\n"
        "Why we are asking: we would rather have your word that it is a general rule than lean on "
        "a remark you made about one report.",

        "Should \"whatever you see on screen is what comes out in the download\" be the rule for "
        "all six reports?",

        "A) Yes - treat it as a general rule for all six.\n\n"
        "B) No - it was only about the one report we were discussing at the time.",
    ),
]

TAB3 = [
    (
        "Report Suite - the Sales By Customer report - where it sits in the menu "
        "(story SV-8600, \"Report access and navigation placement\")",

        "This is one of the items you have not got to yet, and it is the only reason one of our "
        "tests is still parked.\n\n"
        "We need to know which menu group this report belongs in, and whether it sits below the "
        "links that were already there. Our notes from the product show it under a group named "
        "SALES; the written description says Performance. We cannot tell which one is right.\n\n"
        "Why we are asking: the test currently checks the wrong group, or the product does - and "
        "we do not know which to raise.",

        "Which menu group should the Sales By Customer report appear in?",

        "A) Performance - below the links that were already there.\n\n"
        "B) Sales - below the links that were already there.\n\n"
        "C) Somewhere else - please say where.",
    ),
    (
        "Report Suite - the Sales By Representative report - the word on screen "
        "(stories SV-8599 and SV-8632)",

        "You confirmed that \"Representative\" on its own is fine in the downloaded files, and we "
        "have matched our tests to it.\n\n"
        "Two tests are still parked because they are about the word on the SCREEN and on the "
        "customer's card, not in the files. You were only asked about the files, so we did not want "
        "to stretch your answer to cover screens you had not seen.\n\n"
        "Why we are asking: two tests come unparked the moment you answer, either way.",

        "Should the full word \"Representative\" also be used on the screen and on the customer's "
        "card?",

        "A) Yes - use the full word everywhere it appears, not only in the downloaded files.\n\n"
        "B) No - only the downloaded files matter. The screen can stay as it is.",
    ),
    (
        "Report Suite - how a machine is named - the other five reports "
        "(stories SV-8660 for Work In Progress and SV-8606 for Sales By Customer)",

        "You confirmed the Work In Progress report should keep showing the unit number on top with "
        "the vehicle number underneath. That is settled and we are not reopening it - it is already "
        "what your written description asks for.\n\n"
        "What we would like to record properly is what that means for the other reports. Back in "
        "July you told us a machine should be named by its vehicle number first, everywhere, and "
        "that has since been written into the Sales By Customer description.\n\n"
        "Why we are asking: nothing needs to change either way - we just want your answer on paper, "
        "so that when somebody asks why one report is different, we can show them your words "
        "instead of our reasoning.",

        "Does your earlier instruction - name a machine by its vehicle number first - still stand "
        "for the other reports?",

        "A) Yes - it still applies everywhere else. Work In Progress is the one exception, because "
        "its two-line layout is already right.\n\n"
        "B) No - drop it. Show the unit number first everywhere, and we will change the other "
        "report to match.",
    ),
]

REMINDER = (
    "NO ANSWER NEEDED - just a note for when you are next in the descriptions. Seven small "
    "wording tidy-ups are still open from the last sheet, and not one of them is holding any test "
    "up: where the Technician Utilization report sits in the menu; the machine chooser on Work In "
    "Progress; the line calling Parts Velocity the only report in its group; the line saying the "
    "Escape key closes the deactivate pop-up; the download size limit missing from three "
    "descriptions; a short note that the vehicle-number field also holds serial numbers for things "
    "that are not vehicles; and some garbled characters in two descriptions. Also, two lines still "
    "describe the Print feature you deliberately dropped, and there is an open job for it - tell us "
    "if you would like us to keep reminding you about those, or to stop."
)


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws, r


# --------------------------------------------------------------- QA-only mapping
QA_ROWS = [
    ("Tab 1", "1",
     "The location column - the one unresolved state (multi-ACCESS, single-SELECTED)",
     "T1-1 of the 2026-08-04 sheet; his answer C contradicts itself (gap U1)",
     "GENUINELY BLOCKED (11): WIP-COL-02 (C30467); IV-COL-04 (C30554); PV-COL-02 (C30352); "
     "IV-COL-01 (C30551); SBC-LOC-04 (C38912); SBR-LOC-05 (C38913); PV-FILT-14 (C38914); "
     "TU-LOC-06 (C38915); WIP-FLT-09 (C38916); IV-LOC-06 (C38917); TU-EXP-04 (C30437) 2nd assertion",
     "C30467 https://shopview.testrail.io/index.php?/cases/view/30467 · "
     "C30554 https://shopview.testrail.io/index.php?/cases/view/30554 · "
     "C30352 https://shopview.testrail.io/index.php?/cases/view/30352 · "
     "C30551 https://shopview.testrail.io/index.php?/cases/view/30551 · "
     "C38912 https://shopview.testrail.io/index.php?/cases/view/38912 · "
     "C38913 https://shopview.testrail.io/index.php?/cases/view/38913 · "
     "C38914 https://shopview.testrail.io/index.php?/cases/view/38914 · "
     "C38915 https://shopview.testrail.io/index.php?/cases/view/38915 · "
     "C38916 https://shopview.testrail.io/index.php?/cases/view/38916 · "
     "C38917 https://shopview.testrail.io/index.php?/cases/view/38917 · "
     "C30437 https://shopview.testrail.io/index.php?/cases/view/30437",
     "WIP spec v6 S4-R3 and IV spec v3 S7-R6 both still say the column is NOT in the selector - "
     "both STALE against his answer. Read live 2026-08-05. Default scope = the user's active "
     "location (one branch) is proven by SBC-PERS-05 C30178, SBR-PERS-04 C30274, PV-FILT-10 "
     "C30337, TU-NAV-03 C30394, WIP-FLT-06 C30503, IV-NAV-03 C30536.",
     "A (entry absent at single scope) -> the 11 are reworded to 'no Location option until a "
     "second branch is chosen'; new case N1 authorable; N2 becomes a negative case. "
     "B (entry present, off) -> WIP-COL-02 and IV-COL-04 are nearly correct as written; the other "
     "9 need the 'by default - you can still switch it on' qualifier. EITHER WAY: developer ticket "
     "B4 becomes writable, and the 5 cases in DELTAS Group 2 can be fixed now without him."),
    ("Tab 1", "2",
     "Does a hand-made location-column toggle persist?",
     "T1-1 gap U2 - his answer is silent on memory",
     "WIP-PERS-03 (C30508); IV-PERS-03 (C30581) - both currently TRUE either way; new case N3 "
     "cannot be authored until he answers",
     "C30508 https://shopview.testrail.io/index.php?/cases/view/30508 · "
     "C30581 https://shopview.testrail.io/index.php?/cases/view/30581",
     "Both cases restore 'column selection' generically, so neither is wrong today. No spec anchor "
     "covers the new switch's memory.",
     "A -> N3 authored as 'remembered like any other column'. B -> N3 authored as 'returns to "
     "automatic'. Neither answer changes C30508 or C30581."),
    ("Tab 2", "1",
     "Technician Utilization download menu - four options or three?",
     "T2-6; he answered B (longer wording) but not the count",
     "TU-EXP-01 (C30434) staged for rewrite; TU-EXP-02 (C30435) released unchanged; "
     "coverage gap N4 (a second spreadsheet download) not authored",
     "C30434 https://shopview.testrail.io/index.php?/cases/view/30434 · "
     "C30435 https://shopview.testrail.io/index.php?/cases/view/30435",
     "TU spec v5 S7-R2 describes three options. Live build showed four, short-form. Read "
     "2026-08-03.",
     "A (four) -> N4 is real coverage and should be authored; C30434's rewrite stands. "
     "B (three) -> N4 is not needed and the developer ticket changes shape."),
    ("Tab 2", "2",
     "The date chooser - confirm 'keep the product'",
     "T2-5; he answered A but his gloss 'the original datepicker is the intentional one' could be "
     "read either way",
     "SBC-DATE-01 (C30102); SBC-DATE-03 (C30104) title; SBR-DATE-01 (C30201); PV-FILT-03 (C30330); "
     "WIP-FLT-04 (C30501); IV-DATE-01 (C30561); SBC-EXP-02 (C30160) - NOT frozen",
     "C30102 https://shopview.testrail.io/index.php?/cases/view/30102 · "
     "C30104 https://shopview.testrail.io/index.php?/cases/view/30104 · "
     "C30201 https://shopview.testrail.io/index.php?/cases/view/30201 · "
     "C30330 https://shopview.testrail.io/index.php?/cases/view/30330 · "
     "C30501 https://shopview.testrail.io/index.php?/cases/view/30501 · "
     "C30561 https://shopview.testrail.io/index.php?/cases/view/30561 · "
     "C30160 https://shopview.testrail.io/index.php?/cases/view/30160",
     "SBC spec v13 S2-R2 closes an eleven-item list including Today/Yesterday/Custom. All six "
     "specs STALE against his answer.",
     "A confirmed -> the 6 rewrites plus the C30104 title fix stand as staged; C30160's file-name "
     "map loses today/yesterday/custom. B -> every one of those staged edits is wrong and a "
     "developer ticket is needed instead. This is the highest-volume confirmation on the sheet."),
    ("Tab 2", "3",
     "Inventory Value 'As of' - must both files word it identically?",
     "T2-7; he answered A (belongs in the spreadsheet) and left the wording half unanswered - "
     "option C was the 'make them identical' option and he did not take it",
     "IV-EXP-04 (C30590) - already tells the tester not to raise the difference",
     "C30590 https://shopview.testrail.io/index.php?/cases/view/30590",
     "IV spec v3 S10-R8 describes the PDF header only. The CSV line is undocumented.",
     "A -> C30590 stands exactly as written and the stance is his, not ours. B -> a developer "
     "ticket, and C30590 flips to expect matching wording."),
    ("Tab 2", "4",
     "The logo - whose logo, the shop's or the customer's?",
     "T3-4; his C says 'if the customer has a logo selected' then 'the company's own uploaded logo'",
     "SBC-EXP-10 (C30168); TU-EXP-06 (C30439) incl. title; PV-EXP-05 (C30379); "
     "SBR-EXP-06 (C30281) - NOT frozen; IV-EXP-04 (C30590) unchanged; coverage gap N5",
     "C30168 https://shopview.testrail.io/index.php?/cases/view/30168 · "
     "C30439 https://shopview.testrail.io/index.php?/cases/view/30439 · "
     "C30379 https://shopview.testrail.io/index.php?/cases/view/30379 · "
     "C30281 https://shopview.testrail.io/index.php?/cases/view/30281 · "
     "C30590 https://shopview.testrail.io/index.php?/cases/view/30590",
     "Every one of our cases says the SHOP's logo ('the shop logo shows at the top of the PDF when "
     "one is set'). No spec says whose logo it is.",
     "A (shop) -> the 4 staged logo rewrites are correct as staged. B (customer) -> all 4 change "
     "again AND the seeding steps change, because a customer logo has to be set per customer."),
    ("Tab 2", "5",
     "'On-screen should match download' - a general rule?",
     "T2-4's extra remark; we USED it as a derivation to settle T1-1 gap U3",
     "IV-EXP-02 (C30588); WIP-EXP-02 (C30511); TU-EXP-04 (C30437) - the download half of the "
     "location-column rule rests on this derivation",
     "C30588 https://shopview.testrail.io/index.php?/cases/view/30588 · "
     "C30511 https://shopview.testrail.io/index.php?/cases/view/30511 · "
     "C30437 https://shopview.testrail.io/index.php?/cases/view/30437",
     "DELTAS.md labels it a DERIVATION from T2-4 + T3-1, not something he said. Rule 12 says a "
     "derivation is not an answer.",
     "A -> the derivation becomes an answer and can be cited directly; recommend the QA lead "
     "records it as a durable ruling. B -> the download half of the location rule loses its basis "
     "and must be asked separately."),
    ("Tab 3", "1",
     "Sales By Customer - which menu group, and below which links?",
     "T3-7 - LEFT BLANK on the 2026-08-04 sheet, second time of asking",
     "SBC-NAV-01 (C30096) - stays frozen until he answers",
     "C30096 https://shopview.testrail.io/index.php?/cases/view/30096",
     "SBC spec v13 Story 1 says Performance. Our own build notes from 2026-08-03 record a SALES "
     "group. Not re-observed this pass.",
     "A (Performance) -> C30096 stands and a developer ticket is raised. B (Sales) -> C30096 is "
     "reworded and he tidies the description. Either way the case comes unfrozen."),
    ("Tab 3", "2",
     "'Representative' on screen and on the customer card too?",
     "T3-9 - LEFT BLANK; his T2-3 answer covered only the download heading",
     "SBR-WO-01 (C30310); SBR-WO-06 (C30315) - both stay frozen",
     "C30310 https://shopview.testrail.io/index.php?/cases/view/30310 · "
     "C30315 https://shopview.testrail.io/index.php?/cases/view/30315",
     "Applying T2-3 to the screen would be extending his answer past what he wrote - deliberately "
     "not done (Rule 12).",
     "A -> both cases keep the full-word expectation and a developer ticket is raised. B -> both "
     "are reworded to the short form. Either way, 2 cases come unfrozen."),
    ("Tab 3", "3",
     "Does the July 'vehicle number everywhere' instruction still stand for the other reports?",
     "T2-2; his B narrows it to Work In Progress, but his July words were expressly cross-project",
     "Work In Progress (4 staged, decided): WIP-COL-05 (C30470); WIP-SORT-03 (C30485); "
     "WIP-FLT-03 (C30500); WIP-EXP-07 (C30516). MUST NOT CHANGE: SBC-LBL-01 (C30134); "
     "WIP-VIS-07 (C30525) - live and already correct",
     "C30470 https://shopview.testrail.io/index.php?/cases/view/30470 · "
     "C30485 https://shopview.testrail.io/index.php?/cases/view/30485 · "
     "C30500 https://shopview.testrail.io/index.php?/cases/view/30500 · "
     "C30516 https://shopview.testrail.io/index.php?/cases/view/30516 · "
     "C30134 https://shopview.testrail.io/index.php?/cases/view/30134 · "
     "C30525 https://shopview.testrail.io/index.php?/cases/view/30525",
     "WIP spec v6 (page 703660034, read live 2026-08-05) says unit-number-first in five places and "
     "NEVER said otherwise. SBC spec v13 S8-R7..R10 carries the vehicle-number chain as ratified "
     "text. The other four specs are SILENT (0 hits). Epic SV-8582 is SILENT (all 102 children's "
     "descriptions + all comments searched). Tech plan line 532 independently says unit-first for "
     "WIP.",
     "THE WORK IN PROGRESS HALF IS ALREADY DECIDED by the QA lead 2026-08-05 (latest-wins) - this "
     "question does NOT reopen it. A -> the durable CLAUDE.md rule is narrowed with a Work In "
     "Progress exception and C30134 is protected. B -> C30134 changes too and SBC spec v13 "
     "S8-R7..R10 must be edited by him. See VIN-ORDER-RULING.md section 3."),
]

QA_NOTES = [
    "HOW THE POPULATION WAS SEARCHED (Standing Rule 50): all 469 of our cases were pulled LIVE "
    "read-only from TestRail on 2026-08-05 (474 under group 4281 minus 5 authored by Vladimir "
    "Tomovic, C38919-C38923, hands off per Rule 38) and searched in full - title, preconditions, "
    "steps and every expected result. No sampling.",
    "WHY THESE TEN AND NOT MORE: all 15 of his answers were swept for ambiguity. Ten produced a "
    "question. The five that are genuinely clear and need nothing are: T2-1 (hide the branch "
    "chooser for a one-branch person = B), T2-4 (add the four missing columns back = A), T2-8 (one "
    "reports permission = A), T3-1 (downloads carry the location column = A), T3-2 (keep testing "
    "to his answers = A), T3-3 (Summary-file column position = A) and T3-5 (the five dropped "
    "features = A). T2-9 (Print) produced only the reminder row, because his answer was praise "
    "rather than a decision.",
    "TWO ANSWERS ARE NOT RE-ASKED ON PURPOSE: T2-2's Work In Progress half is DECIDED by the QA "
    "lead's latest-wins ruling of 2026-08-05, so Tab 3 question 3 asks only about the OTHER "
    "reports and says so in its own text. And the 9 items he left blank are not re-asked one by "
    "one - only the two that hold live tests (T3-7, T3-9) are asked as questions; the other seven "
    "are one no-answer-needed reminder row, because he already told the QA lead he had not done "
    "them.",
    "WORDING RULES APPLIED (Standing Rule 7 + the QA lead's ruling 2026-08-05): every question row "
    "names the PROJECT (Report Suite) and the FEATURE (which report), because Chris also owns Fees "
    "& Discounts and a bare question would be ambiguous to him. Story or epic references are given "
    "in plain form only where they orient him. No case IDs, no requirement anchors, no HTTP terms "
    "and no internal names appear in anything he reads. Each question carries a one-line 'Why we "
    "are asking' so he can see the consequence.",
    "NOTHING HAS BEEN WRITTEN ANYWHERE. This sheet is a draft for the QA lead to send. No TestRail "
    "write, no Jira write, no case edit was made in producing it, and CLAUDE.md was not touched.",
    "SOURCE-CURRENCY (Standing Rule 31), all checked 2026-08-05: his answers = CURRENT (newest "
    "authoritative product source, sha256 6da732152589a31b842adf6e1a16549c3fce0dd0ca0c4da0e5792aac"
    "924993cd). All six Confluence descriptions fetched LIVE, HTTP 200 - SBC v13 / SBR v15 / PV v4 "
    "/ TU v5 / WIP v6 / IV v3, none moved, all STALE against his answers except the WIP identifier "
    "text which AGREES with them. Epic SV-8582 = CURRENT, 102 children verified two ways with the "
    "key sets equal in both directions. Tech plan = CURRENT as supplied 2026-07-29. Designs = "
    "ABSENT (spec-only project). The BUILD was NOT observed in this pass, and the Rule-49 re-check "
    "queue viu-2026-08-03/RECHECK-QUEUE.md stays OPEN, so every verdict on this project is "
    "provisional.",
]


def write_xlsx():
    wb = openpyxl.Workbook()
    w_narrow = [4, 34, 52, 42, 46, 22]
    w_wide = [4, 34, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "Follow-up - the location column - Report Suite - Chris Ward - 2026-08-05",
           TAB1_NOTE, "Needed first - one detail is holding a developer job", TAB1,
           w_wide, first=True)

    _sheet(wb, TAB2_NAME,
           "Quick confirmations - Report Suite - Chris Ward - 2026-08-05",
           "Five one-word confirmations. In each of these we have read something INTO your answer "
           "rather than out of it, and we would rather you corrected us now than after the tests "
           "are written. " + THANKS,
           "Things we have assumed - please confirm or correct", TAB2, w_narrow)

    ws3, r = _sheet(wb, TAB3_NAME,
                    "Still holding tests - Report Suite - Chris Ward - 2026-08-05",
                    "Three questions, each of which frees up a parked test the moment you answer "
                    "it. Two are from the block you told us you had not got to yet. " + THANKS,
                    "Three answers would unfreeze five tests", TAB3, w_narrow)
    ws3.cell(row=r + 1, column=2, value=REMINDER).alignment = WRAP
    ws3.cell(row=r + 1, column=2).font = Font(bold=True)
    ws3.row_dimensions[r + 1].height = 160

    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail C-ids, "
                 "requirement anchors and live evidence live here so the reader-facing tabs stay "
                 "plain (Standing Rules 7 and 8).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Tab", "Item", "What it asks", "Where the ambiguity comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 60, 60]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items, tab, start=1):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, start):
            out.append(f"### Item {i}.0 — {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    # In the spreadsheet each tab repeats the thank-you, because tabs are read
    # independently. In this one-page twin it would read as padding, so it is
    # printed once at the top and stripped from the per-tab notes.
    tab1_note_md = TAB1_NOTE.replace(THANKS, "").strip()

    md = f"""# Follow-up questions for Chris Ward — Report Suite — 2026-08-05

**Project: Report Suite (the six reports) · epic SV-8582 · Product Owner: Chris Ward**

**This is the plain-language twin of `Follow-up-Question-for-Chris-Ward_2026-08-05.xlsx`.**
The spreadsheet is the version to send; it mirrors the 2026-08-04 sheet's format exactly, and it
carries a QA-only tab that must not be forwarded.

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

{THANKS}

**Ten questions in total, each one a plain A or B, plus one note that needs no answer.**

---

## Tab 1 — {TAB1_NAME}

{tab1_note_md}

{block(TAB1, 1)}

---

## Tab 2 — {TAB2_NAME}

Five one-word confirmations. In each of these we have read something INTO your answer rather than
out of it, and we would rather you corrected us now than after the tests are written.

{block(TAB2, 1)}

---

## Tab 3 — {TAB3_NAME}

Three questions, each of which frees up a parked test the moment you answer it. Two are from the
block you told us you had not got to yet.

{block(TAB3, 1)}

### And one note that needs no answer

> {REMINDER}

---

## QA-only — not for Chris

The internal question-to-case mapping lives on the spreadsheet's `{TAB4_NAME}` tab: every question's
affected TestRail case IDs with links, the requirement anchors, the live evidence, and what each
possible answer resolves to. It also records the method notes — how all 469 cases were searched,
which of his 15 answers were judged clear and therefore not re-asked, and the source-currency block.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
