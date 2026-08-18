#!/usr/bin/env python3
"""Generate the Chris Ward question sheet - Report Suite - 2026-08-17 (.xlsx + .md).

The SIMPLEST POSSIBLE sheet: Chris is very busy, so this asks only the TWO
genuinely-open product decisions surfaced by the 2026-08-17 Fabian design-review
reconciliation, both on the Work In Progress report. Each is a plain A / B / C.

Mirrors the established format 1:1 (Standing Rule 16), following
build/report-suite/rulings-2026-08-05/gen_followup_sheet.py:
 - same six reader columns in the same order (#, Topic, What happens now,
   The question, Options, Your answer);
 - same row layout (A1 title / A2 note / A4 header / A5 band / items from row 6);
 - same fills, fonts, freeze pane, column widths;
 - a QA-only mapping tab (Standing Rule 8) that is NEVER sent to Chris.

Standing Rule 7 + Rule 55: reader-facing text is extremely simple, names the
PROJECT (Report Suite) and the FEATURE (Work In Progress report) on every
question row (Chris also owns Fees & Discounts), cites the epic in plain form,
and carries no case IDs / spec anchors / jargon / the word "VIU".

RESEARCH ONLY - this script writes two files into this folder. It makes NO
TestRail or Jira call of any kind.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Report-Suite_Questions-for-Chris-Ward_2026-08-17.xlsx")
MD = os.path.join(HERE, "Report-Suite_Questions-for-Chris-Ward_2026-08-17.md")

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Questions for Chris"
TAB2_NAME = "QA internal - not for Chris"

INTRO = (
    "Two quick questions, both about the Work In Progress report. Each one is a plain A / B / C, "
    "and both are small tidy-ups where your written description says a thing two different ways and "
    "we would rather have your word than guess. Every question names the project and the report, "
    "because we know you look after more than one thing here. There are no bugs on this sheet - "
    "just two wording decisions. Thank you."
)

# ------------------------------------------------------------------- the questions
# (topic, what-happens-now, the question, options)
TAB1 = [
    (
        "Report Suite - the Work In Progress report - the Estimates help text "
        "(the little information icon next to the Estimates figure; under epic SV-8582)",

        "On the Work In Progress report, the Estimates figure has a small information icon that "
        "shows a short plain explanation of what that number means.\n\n"
        "Your written description gives that explanation TWO different ways, in two different places "
        "of the same document - one short, and one longer that the recent design review locked in. "
        "Word for word:\n\n"
        "- Short version: \"Quotes the customer has not approved yet - not counted in the "
        "totals.\"\n\n"
        "- Longer version (locked in the design review): \"The total value of all estimate lines "
        "that have not yet been approved, including lines awaiting authorization on open work "
        "orders.\"\n\n"
        "We are using the longer one, because it is the most recent - it came from the design "
        "review. We just want to record which one you want, so the description says it only once.\n\n"
        "Why we are asking: it is a one-line tidy-up in your description either way, and we would "
        "rather have your word than leave the description saying two things.",

        "Which explanation should the Estimates information icon show - and may we drop the other so "
        "the description states it once?",

        "A) Keep the longer one from the design review (\"The total value of all estimate lines "
        "that have not yet been approved, including lines awaiting authorization on open work "
        "orders\") and drop the short one. This is the one we have already built our check to, so "
        "if you choose A we just need your confirmation and the short line tidied out.\n\n"
        "B) Keep the short one (\"Quotes the customer has not approved yet - not counted in the "
        "totals\") and drop the longer one. We change our check back to the short wording.\n\n"
        "C) Something else - please write the exact wording you want.",
    ),
    (
        "Report Suite - the Work In Progress report - which tab a job appears in "
        "(the tabs across the top of the report; under epic SV-8582)",

        "The Work In Progress report is split into tabs across the top - Estimate, In Progress, "
        "Review, Complete, and so on.\n\n"
        "Your written description says two different things about which tab a job belongs in, in the "
        "same document:\n\n"
        "- One part says a job appears in exactly ONE tab, chosen by the job's overall status.\n\n"
        "- Another part, added more recently, says the tabs go by the state of each individual LINE "
        "on the job - so a job with lines in more than one state would show up in more than one "
        "tab.\n\n"
        "We have not picked a side - our tests follow the wording each was written against. This is "
        "the one thing on the report we cannot settle ourselves.\n\n"
        "Why we are asking: the two behave completely differently for a job that has work in more "
        "than one state, and we do not want to guess which one is right.",

        "When a single job has work in more than one state, should it appear in just one tab, or in "
        "every tab that matches?",

        "A) In just ONE tab, chosen by the job's overall status - a job is only ever in one place "
        "on the report.\n\n"
        "B) In EVERY tab that matches - a job with some lines estimated and some in progress shows "
        "in both the Estimate tab and the In Progress tab.\n\n"
        "C) Something else - please describe it.",
    ),
]

# --------------------------------------------------------------- QA-only mapping
QA_ROWS = [
    ("1",
     "Work In Progress - Estimates info-icon help text: longer (design-review-locked) wording "
     "vs the short leftover - which governs, and drop the other",
     "WIP spec states the Estimates explanation two ways in the same live document: S5-R12 (short) "
     "vs S5a-R2 (locked verbatim, Fabian design review). We follow S5a-R2 (latest wins, Rule 32) "
     "and disclose the divergence (Rule 56). Spec-hygiene contradiction raised for Chris.",
     "WIP-SUM-07 (C30493)",
     "C30493 https://shopview.testrail.io/index.php?/cases/view/30493",
     "WIP spec v21 2026-08-14, S5-R12 short: \"Quotes the customer has not approved yet - not "
     "counted in the totals.\" vs S5a-R2 locked verbatim (Fabian design review 2026-08-17): "
     "\"The total value of all estimate lines that have not yet been approved, including lines "
     "awaiting authorization on open work orders.\" Read live 2026-08-17. The other six info-icon "
     "explanations re-verified against S5-R12 and match byte-for-byte (Rule 41).",
     "A (keep longer, drop short) -> C30493 stands exactly as written; Chris drops the S5-R12 "
     "short leftover so the spec states it once. Blocks nothing (case is correct). "
     "B (keep short) -> C30493 reworded back to the short wording. "
     "C -> C30493 reworded to his exact text."),
    ("2",
     "Work In Progress - tab placement: one work order in exactly one tab (by work-order status) "
     "vs a work order in every matching tab (by line state)",
     "WIP spec says two incompatible things in the same live document: v11 §3 Key Decision "
     "(SV-9027) keys buckets on LINE STATE so a work order can appear in several tabs, while S2-R4 "
     "and S3-R1..S3-R4 place a work order ONCE by its STATUS. No side picked (Rules 15/57/58); "
     "cases keep the requirement they cite, word for word.",
     "WIP-SCOPE-03 (C30458) - primary, on AUTOMATION: HOLD; also affected: WIP-PLACE-01 (C30462, "
     "Automated) and WIP-PLACE-03 (C30464). C30528 (nightly snapshot shape) may also be wrong "
     "under the line-state reading - flagged HIGH risk, not edited.",
     "C30458 https://shopview.testrail.io/index.php?/cases/view/30458 · "
     "C30462 https://shopview.testrail.io/index.php?/cases/view/30462 · "
     "C30464 https://shopview.testrail.io/index.php?/cases/view/30464 · "
     "C30528 https://shopview.testrail.io/index.php?/cases/view/30528",
     "WIP spec v11 2026-08-10, Story 2 S2-R4: \"exactly once, in exactly one tab\"; S3-R1..S3-R4 "
     "place a work order by its status. v11 §3 Key Decision (added 2026-08-10, SV-9027): "
     "\"Buckets are keyed on line state, not work-order status ... a work order carrying lines in "
     "more than one state appears in each matching tab.\" Read live 2026-08-17. Epic SV-8582 "
     "silent on the tiebreak.",
     "A (one tab, by status) -> C30458/C30462/C30464 stand as written; the line-state Key Decision "
     "is removed from the spec by Chris; C30528 confirmed correct. "
     "B (every matching tab, by line state) -> C30458 reworded (drops 'exactly once'), C30462/"
     "C30464 re-checked, and the per-tab money slice + line-level vs work-order-level ageing become "
     "authorable (no coverage today); C30528 likely needs rework (HIGH risk - the snapshot feeds "
     "trend history). C -> per his description."),
]

QA_NOTES = [
    "SCOPE - WHY ONLY TWO QUESTIONS: this sheet is deliberately the simplest possible (the QA lead "
    "flagged Chris is very busy). It carries ONLY the two genuinely-open PRODUCT decisions surfaced "
    "by the 2026-08-17 Fabian design-review reconciliation "
    "(build/report-suite/fabian-review-2026-08-17/REPORT-SUITE-COMPLETION-REPORT.md OUTSTANDING "
    "section + COVERAGE-REDERIVATION.md). Both are Work In Progress spec-hygiene contradictions. "
    "No bugs (Rule 7). Nothing already answered is re-asked.",
    "NOT ON THIS SHEET, ON PURPOSE: the older Location-column ruling, the four unfinished spec "
    "contradictions and the ~10,000-row export cap (register rows C2 / RS-Q2 / RS-Q3) are already "
    "captured in the HELD sheet build/report-suite/questions-2026-08-06/ and are not re-asked here "
    "(no drip - Rule 55). Run-359 sync, marker/spec-version passes and build verification are QA-"
    "lead items, not product decisions for Chris, so they are excluded.",
    "WORDING RULES APPLIED (Standing Rule 7 + Rule 55): every reader-facing question names the "
    "PROJECT (Report Suite) and the FEATURE (the Work In Progress report), because Chris also owns "
    "Fees & Discounts. The epic is named in plain form only to orient him. No case IDs, requirement "
    "anchors, HTTP terms or internal names appear in anything he reads. Each question carries a "
    "one-line 'Why we are asking' so he can see the consequence.",
    "NOTHING HAS BEEN WRITTEN ANYWHERE. This sheet is a draft for the QA lead to send. No TestRail "
    "write, no Jira write, no case edit was made in producing it, and CLAUDE.md was not touched.",
    "SOURCE-CURRENCY (Standing Rule 31), checked 2026-08-17: all six Confluence descriptions "
    "fetched live 2026-08-17 - SBC v20 / SBR v22 / PV v10 / TU v9 / WIP v21 / IV v10, none moved; "
    "epic SV-8582 = CURRENT (114 children, verified two ways). The design artifact (Fabian Claude "
    "share link) is undated/unfetchable = PARTIAL. Build NOT observed this pass (documents-only by "
    "instruction); the Rule-49 re-check queue stays OPEN, so every verdict on this project is "
    "provisional.",
]


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def write_xlsx():
    wb = openpyxl.Workbook()
    widths = [4, 36, 60, 44, 50, 22]

    ws = wb.active
    ws.title = TAB1_NAME
    ws["A1"] = "Questions for Chris Ward - Report Suite - 2026-08-17"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = INTRO
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, "Two quick wording decisions on the Work In Progress report", 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(TAB1, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 340
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # -------------------------------------------------- QA-only mapping tab
    ws4 = wb.create_sheet(TAB2_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail C-ids, "
                 "requirement anchors and live evidence live here so the reader-facing tab stays "
                 "plain (Standing Rules 7 and 8).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Item", "What it asks", "Where the ambiguity comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFG", [7, 42, 44, 44, 52, 62, 62]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, 1):
            out.append(f"### Question {i} - {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    md = f"""# Questions for Chris Ward - Report Suite - 2026-08-17

**Project: Report Suite (the six reports) - epic SV-8582 - Product Owner: Chris Ward**

**This is the plain-language twin of `Report-Suite_Questions-for-Chris-Ward_2026-08-17.xlsx`.**
The spreadsheet is the version to send; it mirrors the established sheet format exactly, and it
carries a QA-only tab that must not be forwarded.

**DRAFT - NOT SENT. Nothing has been written to TestRail or Jira.**

{INTRO}

**Two questions in total, both about the Work In Progress report, each a plain A / B / C.**

---

## {TAB1_NAME}

{block(TAB1)}

---

## QA-only - not for Chris

The internal question-to-case mapping lives on the spreadsheet's `{TAB2_NAME}` tab: each question's
affected TestRail case IDs with links, the requirement anchors, the live evidence, and what each
possible answer resolves to. It also records the scope, wording rules and source-currency notes.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
