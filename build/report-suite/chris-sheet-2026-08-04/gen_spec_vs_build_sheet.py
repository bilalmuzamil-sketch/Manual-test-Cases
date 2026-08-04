#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generator — Report Suite: SPEC-versus-BUILD decision sheet for Chris Ward (PO).
Date: 2026-08-04.

WHAT THIS IS (and why it is shaped this way)
--------------------------------------------
The QA lead's instruction, verbatim:
  "Anything which is blocked on Chris, make a SHeet for him to confirm which behavior he
   would prefer to use, give him reference from specs what specs says vs what is in the
   build."

So this is NOT a question sheet about intentions. Every reader-facing row carries FOUR things:
  1. What your write-up says      — the requirement quoted VERBATIM (Rule 25), anchor codes stripped
  2. What the product actually does — the live observation, with the evidence date
  3. Which do you want?           — plain options: keep the product / change the product / a third
  4. Your answer                  — a blank cell

FORMAT (Standing Rule 16): mirrors 1:1 the sheets Chris has already answered —
  build/report-suite/PO-Questions-Chris-ReportSuite-2026-08-03.{md,xlsx} (and the 07-27 / 07-31
  ones): same two tabs ("Questions for PO" + "QA Internal Mapping"), same header fills/fonts,
  same group banding, same numbering, same QA-only mapping table plus the completeness-proof,
  withdrawn and not-asked appendices. Only the three middle reader-facing COLUMN HEADINGS change,
  because the QA lead asked for the spec-vs-build shape rather than the what-happens-now shape.

RULE 7 (absolute): no case IDs, no C-numbers, no spec anchor codes, no version numbers, no
jargon and no HTTP/API terms in anything Chris reads. Everything traceable lives on the
QA-only tab. Where a quote is HIS OWN document's wording it is preserved verbatim (altering a
quote would be dishonest); our own prose says "spreadsheet download" / "printable download".

RULE 50 / the C-id trap: verify_cids() re-reads build/report-suite/testrail-id-map.csv and
aborts unless EVERY internal-id/C-id pair printed on the QA-only tab matches. The 2026-07-31
sheet printed PV-API-04 as C30388 (correct: C30391; C30388 is PV-API-01), so this is not
optional.

Run:  python3 gen_spec_vs_build_sheet.py
"""

import csv
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.dirname(HERE)                      # build/report-suite
IDMAP = os.path.join(PROJECT, "testrail-id-map.csv")  # READ-ONLY

BASE = "Report-Suite_Spec-vs-Build_Decisions-for-Chris-Ward_2026-08-04"
TITLE = ("Report Suite - what your write-ups say vs what the product does: "
         "decisions we need from Chris Ward - 2026-08-04")

BUILD_MARKER = "v3.4.1-0ed4433"

# --------------------------------------------------------------------------------------
# SOURCE-CURRENCY BLOCK (Standing Rule 31) - QA-facing, kept off the reader-facing tab
# --------------------------------------------------------------------------------------
SOURCE_CURRENCY = [
    ("Sales By Customer description", "Confluence page 577634305", "version 13, last changed 2026-07-31",
     "2026-08-03", "CURRENT"),
    ("Sales By Representative description", "Confluence page 585629698", "version 15, last changed 2026-07-29",
     "2026-08-03", "CURRENT"),
    ("Parts Velocity description", "Confluence page 620888066", "version 4, last changed 2026-07-29",
     "2026-08-03", "CURRENT"),
    ("Technician Utilization description", "Confluence page 641400833", "version 5, last changed 2026-07-29",
     "2026-08-03", "CURRENT"),
    ("Work In Progress description", "Confluence page 703660034", "version 6, last changed 2026-07-29",
     "2026-08-03", "CURRENT"),
    ("Inventory Value description", "Confluence page 720142338", "version 3, last changed 2026-07-29",
     "2026-08-03", "CURRENT"),
    ("The build", "QA branch sv8582 / project/reports-suite-bravo, app-version " + BUILD_MARKER,
     BUILD_MARKER, "2026-08-03 / 2026-08-04",
     "PARTIAL - engineering declared the branch NOT FINAL, so every observation on this sheet is "
     "PROVISIONAL and is queued for re-check in viu-2026-08-03/RECHECK-QUEUE.md (Standing Rule 49). "
     "Shortfall: the observations may change when the branch settles."),
    ("Epic SV-8582 + child stories", "Jira, project SV", "currency-checked, no full re-read this pass",
     "2026-08-03", "PARTIAL - Tier-1 currency check only (Standing Rule 37); a full re-read was not "
     "authorised and is not claimed"),
    ("Designs", "none exist for the Report Suite", "n/a", "2026-08-03",
     "ABSENT - spec-only project; no Figma file has ever been supplied, so no design source was "
     "consulted and none is claimed"),
    ("Engineering tech plan", "tech-plan-2026-07-29/", "as supplied 2026-07-29", "2026-08-03", "CURRENT"),
    ("Chris Ward's answers, messages and both videos",
     "chris-answers-2026-07-28 / -07-31 / -08-01, chris-update-2026-07-29, both video transcripts",
     "newest = the 2026-08-01-round two-question sheet", "2026-08-03", "CURRENT"),
]

# --------------------------------------------------------------------------------------
# THE READER-FACING ROWS
# --------------------------------------------------------------------------------------
G_DEC = "Decisions we need from you"
G_WRITE = "Only needs writing down (no decision needed)"
G_FYI = "For your awareness (nothing needed from you)"

ROWS = [
    # ---------------------------------------------------------------- 1
    dict(
        group=G_DEC,
        topic="The location chooser is still shown to someone who only has one location - on all six reports",
        spec=(
            "Four of the six write-ups say the opposite of the ruling you gave us on 31 July. Quoted "
            "word for word:\n\n"
            "Sales By Representative: \"A single-location user still sees the filter with one selectable "
            "location; behavior is unchanged from single-location use.\"\n\n"
            "Parts Velocity: \"A user with access to only one location still sees the Location filter with "
            "a single selectable location; behavior is unchanged from single-location use.\"\n\n"
            "Technician Utilization: \"A user with access to only one location still sees the filter with a "
            "single selectable location; behavior is unchanged from single-location use.\"\n\n"
            "Inventory Value: \"A user with access to only one location still sees the filter with a single "
            "selectable location.\"\n\n"
            "Work In Progress and Sales By Customer say nothing either way. Your ruling on 31 July was the "
            "opposite - you chose hidden, and called it \"classic spec drift\"."
        ),
        build=(
            "We gave a person access to exactly one location, signed in as them, and opened all six reports "
            "on 3 August. The location chooser was still on screen on every single one of the six.\n\n"
            "The location COLUMN is a separate thing and it behaves correctly: on Sales By Customer and "
            "Sales By Representative the column was properly absent for that person.\n\n"
            "So the product is following the four written lines above, not your ruling."
        ),
        ask="Should the location chooser stay visible for a person who only has one location, or be hidden?",
        opts=[
            "A) Keep what the product does - the chooser stays visible. We change our tests back, and you "
            "tidy the four lines so they stop contradicting the ruling.",
            "B) Change the product to match your ruling - hide it. We raise it with engineering, and the "
            "four lines still need correcting because they say it stays.",
            "C) Something else - please describe it.",
        ],
    ),
    # ---------------------------------------------------------------- 2  (items 2 + 3 merged)
    dict(
        group=G_DEC,
        topic="The extra location column works one way on Work In Progress and the exact opposite way on Inventory Value",
        spec=(
            "Both write-ups describe the SAME model - automatic, and deliberately NOT something the user "
            "switches on. Quoted word for word:\n\n"
            "Work In Progress: \"The Location column is not offered in the column selector; its visibility "
            "is automatic - shown only when more than one location is in scope.\"\n\n"
            "Inventory Value: \"the column is shown only when the current scope spans more than one location "
            "and is hidden for a single-location scope. It is governed by scope automatically - it is not a "
            "user-toggled column in the column-selection control.\""
        ),
        build=(
            "Neither report does that, and the two of them do opposite things. Both observed on 3 and 4 August.\n\n"
            "Work In Progress: Location IS one of the switches in the column-picker panel (sixth in the list), "
            "and it starts switched OFF. Switching it on adds the column; switching it off removes it. It never "
            "appeared on its own when two locations were in scope.\n\n"
            "Inventory Value: Location is also one of the switches in the column-picker panel (fifth in the "
            "list), but it starts switched ON - and when we narrowed the location chooser down to one location "
            "the column stayed put, with every row simply repeating that one location's name.\n\n"
            "So: two reports, two different models, and neither one is what either write-up describes."
        ),
        ask=("Which single model should both reports use - and, since these are the last two, the whole set: "
             "automatic, or a switch the user controls?"),
        opts=[
            "A) Automatic everywhere, exactly as both write-ups say - not in the column picker at all, appears "
            "by itself when more than one location is in scope, disappears when there is only one. We raise both "
            "reports with engineering; the write-ups need no change.",
            "B) A switch the user controls, everywhere - always in the column picker. If you pick this, please "
            "also say whether it should start on or off. We correct both write-ups and our tests.",
            "C) Leave each report as it is today and we will write down what each one does - but please confirm "
            "you are happy for the two to differ.",
        ],
    ),
    # ---------------------------------------------------------------- 3
    dict(
        group=G_DEC,
        topic="On Work In Progress the machine is still identified by its unit number first, not its vehicle number",
        spec=(
            "Your Work In Progress write-up still puts the unit number first. Quoted word for word:\n\n"
            "\"The Asset column is a two-line cell: the unit number on the first line in bold, and the vehicle "
            "identification number on the second line in a smaller, muted style.\"\n\n"
            "and\n\n"
            "\"The Asset column sorts by unit number.\"\n\n"
            "Your ruling on 29 July was the other way round - you answered \"A is the correct answer\" to the "
            "vehicle-number-first chain (vehicle number, then unit number, then plate), and added: \"Not just "
            "for these specs though -- really good to keep this in mind for all actions moving forward.\""
        ),
        build=(
            "On 3 and 4 August the machine's cell showed the unit number first, in bold, with the vehicle number "
            "underneath it in smaller grey text - for example 6548 on the top line and 1FDSE3EL1EDB20609 "
            "underneath. Sorting on that column also used the unit number.\n\n"
            "So the product is following the written line above, not your ruling.\n\n"
            "One thing worth knowing before you decide, because it is your own point back to you: you told us "
            "\"we just have to be careful with using the acronym VIN ... it stands for VEHICLE identification "
            "number. So for a generator for example, it gets confusing when we say VIN rather than serial #.\" "
            "That is already happening in the real data. The field labelled as the vehicle number is holding "
            "serial-number-style values for things that are not vehicles - live examples we read include "
            "BULK PARTS1, 12-06696 and P631627 - sitting alongside genuine 17-character vehicle numbers like "
            "1FDSE3EL1EDB20609."
        ),
        ask="Which number should lead on this report, and what should the heading call it?",
        opts=[
            "A) Change the product so the vehicle number leads (then the unit number, then the plate), matching "
            "your ruling and the other report. We raise it with engineering and you update the Work In Progress "
            "write-up.",
            "B) Keep the product as it is - the unit number leads on this report. We change our tests back, and "
            "we record that your ruling does not reach this one report.",
            "C) Lead with the vehicle number as in A, but change the wording so it also reads sensibly for a "
            "machine that is not a vehicle - please tell us the word you want on screen.",
        ],
    ),
    # ---------------------------------------------------------------- 4
    dict(
        group=G_DEC,
        topic="The Sales By Representative downloads say \"Representative\" - a third spelling",
        spec=(
            "Your Sales By Representative write-up still uses the short form in the download column list. "
            "Quoted word for word, from the summary spreadsheet requirement:\n\n"
            "\"Headers, in order: `Sales Rep`, `# Invoices`, `# Customers`, ...\"\n\n"
            "Your ruling on 31 July was: \"Rep is too much slang, let's do representative everywhere\" - so we "
            "changed our tests to the full \"Sales Representative\"."
        ),
        build=(
            "The file that actually downloads says neither of those. Read straight out of the downloaded summary "
            "file on 3 August, the first column heading is simply:\n\n"
            "Representative\n\n"
            "So three different words are in play: your document says \"Sales Rep\", your ruling says \"Sales "
            "Representative\", and the product says \"Representative\". The same single word is used in the "
            "detailed download too."
        ),
        ask="Which word should the download column heading use?",
        opts=[
            "A) \"Representative\" on its own is fine - it is not slang, so it satisfies your ruling. We match "
            "our tests to it and you tidy the write-up.",
            "B) It must read \"Sales Representative\" in full, as your ruling said. We raise it with engineering.",
            "C) Something else - please write the exact wording you want.",
        ],
    ),
    # ---------------------------------------------------------------- 5
    dict(
        group=G_DEC,
        topic="Four columns are missing from the Sales By Representative summary download",
        spec=(
            "Your write-up closes the list of columns for that download. Quoted word for word:\n\n"
            "\"Headers, in order: `Sales Rep`, `# Invoices`, `# Customers`, `Hrs Worked`, `Hrs Invoiced`, "
            "`Inv. Hrs`, `Labor Invoiced`, `Labor Margin`, `Parts Invoiced`, `Parts Margin`, `Margin`, "
            "`Margin %`, `Subtotal`.\"\n\n"
            "That is thirteen columns."
        ),
        build=(
            "We downloaded the file on 3 August for a single location. It has nine columns, and this is the "
            "heading line read straight out of the file:\n\n"
            "Representative, Inv. Hrs, Labor Invoiced, Labor Margin, Parts Invoiced, Parts Margin, Margin, "
            "Margin %, Subtotal\n\n"
            "Four are missing: # Invoices, # Customers, Hrs Worked and Hrs Invoiced.\n\n"
            "The figures themselves are not missing - the information the screen is built from does carry the "
            "invoice count and both hours figures. It is only the download that is short. That is why we read "
            "this as an unfinished download rather than missing data."
        ),
        ask="Should those four columns be added to the download, or should the download be the shorter nine?",
        opts=[
            "A) They are missing by mistake - add the four back. We raise it with engineering and your write-up "
            "stays exactly as it is.",
            "B) The shorter nine-column file is what you want - we correct our tests and you shorten the "
            "write-up's list.",
            "C) Something else - please say which of the four you want.",
        ],
    ),
    # ---------------------------------------------------------------- 6
    dict(
        group=G_DEC,
        topic="The date chooser offers nine choices and has no \"Custom\" option",
        spec=(
            "Three write-ups describe the same eleven-item list. Quoted word for word from Sales By Customer:\n\n"
            "\"The picker offers eleven options, in this order: Today, Yesterday, This Week, Last Week, This "
            "Month, Last Month, This Year, Last Year, This Quarter, Last Quarter, Custom.\"\n\n"
            "Work In Progress and Inventory Value describe the same eleven, each in their own words."
        ),
        build=(
            "The chooser we opened on 3 August offers nine, and they are not those nine. Read off the screen, "
            "in the order shown:\n\n"
            "Last 12 Months, This Year, Last Year, This Quarter, Last Quarter, This Month, Last Month, This "
            "Week, Last Week\n\n"
            "Beside them it shows a month calendar you click dates on, a live readout of how many days your range "
            "covers (it read \"Range: 3 days\" when we looked) and an Apply button. "
            "There is no Today, no Yesterday, and no item called Custom - you build your own range "
            "by clicking the calendar instead. \"All Time\" is correctly not offered.\n\n"
            "Two things you should know. First, this is ONE shared chooser used by all six reports, so whatever "
            "you decide here lands on every one of them. Second, one of our tests cannot be run at all today, "
            "because it tells the tester to click \"Custom\" and there is nothing to click."
        ),
        ask="Should the date chooser be changed to match the write-ups, or should the write-ups be changed to match it?",
        opts=[
            "A) Keep what the product does - we correct the write-ups and our tests to the nine choices plus the "
            "calendar, and the unrunnable test gets fixed the same day.",
            "B) Add Today, Yesterday and a Custom option so it matches the write-ups. We raise it with "
            "engineering; it affects all six reports.",
            "C) Somewhere in between - please tell us which of the three (Today, Yesterday, Custom) you want "
            "added.",
        ],
    ),
    # ---------------------------------------------------------------- 7
    dict(
        group=G_DEC,
        topic="The Technician Utilization download menu has four options, all worded differently",
        spec=(
            "Your Technician Utilization write-up describes three options, each quoted word for word:\n\n"
            "\"The menu has an option labeled 'Download Summary (PDF)'.\"\n\n"
            "\"The menu has an option labeled 'Download Expanded View (PDF)'.\"\n\n"
            "\"The menu has an option labeled 'Download (CSV)'.\""
        ),
        build=(
            "The menu we opened on 3 August has four options, and not one of them is worded that way. Read "
            "straight off the screen, exactly as they appear:\n\n"
            "\"Summary (PDF)\" · \"Summary (CSV)\" · \"Expanded (PDF)\" · \"Expanded (CSV)\"\n\n"
            "So there is one more spreadsheet option than the write-up describes, and the word \"Download\" is "
            "missing from the front of all four.\n\n"
            "For comparison, Sales By Customer and Sales By Representative both show the longer wording, and it "
            "matches their own write-ups exactly: \"Download Summary (PDF)\", \"Download Expanded View (PDF)\", "
            "\"Download Summary (CSV)\", \"Download Expanded View (CSV)\". So this report is the odd one out "
            "rather than the whole set being different."
        ),
        ask="Should Technician Utilization keep its four shorter options, or be brought into line with the other two reports?",
        opts=[
            "A) Keep the four options and their shorter wording - we match our tests to them and you tidy the "
            "write-up (which will also mean listing the fourth option).",
            "B) Bring it into line with Sales By Customer and Sales By Representative - the longer \"Download ...\" "
            "wording. We raise it with engineering.",
            "C) Something else - please write the wording you want.",
        ],
    ),
    # ---------------------------------------------------------------- 8
    dict(
        group=G_DEC,
        topic="The Inventory Value spreadsheet carries an \"As of\" line that no write-up asks for",
        spec=(
            "Your Inventory Value write-up asks for that line in the printable download ONLY. Quoted word "
            "for word:\n\n"
            "\"The PDF header shows the report name 'Inventory Value', the organization name, the selected "
            "period, and an 'as of' line naming the day the values represent (or a message that no snapshot is "
            "available for the period).\"\n\n"
            "Nothing in any of the six write-ups asks for it in the spreadsheet."
        ),
        build=(
            "The spreadsheet has it anyway. The very first line of the downloaded file, read on 3 August, is:\n\n"
            "\"As of: 2026-08-03\"\n\n"
            "with the locations line directly beneath it. The printable download carries the same information but "
            "words it slightly differently - it reads \"As of 2026-08-04\", with no colon. (The two dates differ "
            "only because we downloaded the two files on different days; the wording is the difference, not the "
            "date.) So the line is in both files, only one of them is written down, and the two are punctuated "
            "differently."
        ),
        ask="Should the spreadsheet carry the \"As of\" line, and should both files word it the same way?",
        opts=[
            "A) Yes, it belongs in the spreadsheet too - you add it to the write-up and we keep testing for it.",
            "B) No, it should not be in the spreadsheet - we raise it with engineering to take it out.",
            "C) Keep it in both, but make them word it identically - please say which wording you prefer.",
        ],
    ),
    # ---------------------------------------------------------------- 9  (write-down only)
    dict(
        group=G_WRITE,
        topic="Four write-ups still say each report needs its own separate permission",
        spec=(
            "You have already ruled on this - all report access collapses into one single reports permission. "
            "Four write-ups still say otherwise. Quoted word for word:\n\n"
            "Parts Velocity: \"Both loading the report and exporting it require the Inventory Reports -> View "
            "permission. A user without that permission is denied the report data and the export.\"\n\n"
            "Inventory Value: \"The user must have the permission that grants access to the inventory reports.\"\n\n"
            "Technician Utilization: \"The user must have the permission that grants access to the timesheet "
            "reports.\"\n\n"
            "Work In Progress: \"The user must have the permission that grants access to Work In Progress "
            "reports.\"\n\n"
            "Sales By Customer's write-up has already been corrected - thank you."
        ),
        build=(
            "The product already does exactly what you ruled. Checked on 3 August, and proven both ways round:\n\n"
            "There is exactly ONE reports permission in the whole product, and no per-report one exists anywhere "
            "in the list a manager picks from.\n\n"
            "A person whose entire set of permissions was eight - including that one reports permission, and no "
            "report-specific permission at all - could open and download all six reports.\n\n"
            "A person without that one permission was refused all six, both on screen and on download.\n\n"
            "So nothing needs deciding. It is only the four written lines that still disagree, which makes it "
            "look to an outside reader as though our tests are wrong."
        ),
        ask="Will you update those four write-ups so they all name the single reports permission?",
        opts=[
            "A) Yes, I will update them.",
            "B) It is already done (please point us at it).",
            "C) No - and here is why.",
        ],
    ),
    # ---------------------------------------------------------------- 10 (awareness)
    dict(
        group=G_FYI,
        topic="Print has gone from the product everywhere - two of your lines and one open job still describe it",
        spec=(
            "Two Sales By Customer requirements still list Print as one of the ways the report goes out. Quoted "
            "word for word:\n\n"
            "\"Exports (CSV, PDF, Print) are generated on the server and contain exactly the customers matching "
            "the active filters ...\"\n\n"
            "\"If an export (CSV, PDF, or Print) is triggered while the active filters match no customers - for "
            "example, no customer is selected - the export still downloads, containing the column headers and a "
            "totals row of zeros, with no data rows and no warning.\"\n\n"
            "There is also still an open job in the tracker for building Print."
        ),
        build=(
            "We searched every download menu, button and link on all six reports on 3 August. There is no Print "
            "anywhere in the product. That matches your decision to retire it, so the product is right and our "
            "tests already assume it is gone."
        ),
        ask=("Nothing is needed from you today - this is purely so you are not surprised by it. Next time you "
             "are in the document, dropping Print from those two lines (and closing that open job) would tidy "
             "up the last trace of it."),
        opts=[
            "No decision needed. Tick here if you would like us to keep it on the reminder list until the two "
            "lines and the job are closed.",
        ],
    ),
]

# --------------------------------------------------------------------------------------
# QA-ONLY MAPPING (never sent to the PO)
# --------------------------------------------------------------------------------------
QA_MAP = [
    dict(q="1",
         cases=("SBR-LOC-04 (C30216); PV-FILT-13 (C30340); TU-LOC-05 (C30446); WIP-FLT-06 (C30503); "
                "IV-LOC-04 (C30577). NOTE: there is NO Sales By Customer case asserting the hidden "
                "filter - SBC-LOC-01 (C30109) only asserts the control's position, so SBC is a "
                "coverage question in its own right if he picks B."),
         refs=("OUR SOURCE: Chris Ward 2026-07-31 Q1=A, verbatim \"A -- classic spec drift\" "
               "(chris-answers-2026-07-31/answers-ingested.md). THE BUILD'S SOURCE: SBR v15 S21-N1, "
               "PV v4 S2-E4, TU v5 S9-N1, IV v3 S7-N1 - all four still read \"still sees the filter\". "
               "LIVE: viu-2026-08-03/evidence/singleloc-matrix.json - hasLocationControl TRUE on all "
               "six for the single-workplace subject; build " + BUILD_MARKER + ". "
               "Verdicts: batch-sbc-sbr/VERDICTS.md SBR-LOC-04, batch-pv-tu/VERDICTS.md PV-FILT-13 + "
               "TU-LOC-05, batch-wip-iv/VERDICTS.md IV-LOC-04. RECHECK-QUEUE row B18 (flagged there as "
               "the single most important row to re-check)."),
         resolve=("A -> the 5 cases flip from 'filter hidden' to 'filter shown', and SBC needs no new "
                  "case. B -> no case change; a dev ticket is raised for all six, and SBC needs a NEW "
                  "case for parity. Either way the four spec notes need his edit. All 5 currently sit "
                  "DEVIATION and are HELD (batch-wip-iv/STAGED-CHANGES.md group C2) - deliberately not "
                  "edited, because editing them would assert behaviour no written source supports.")),
    dict(q="2",
         cases=("Work In Progress: WIP-COL-02 (C30467); WIP-COL-01 (C30466); WIP-PERS-02 (C30507); "
                "WIP-FLT-09 (C38916). Inventory Value: IV-LOC-06 (C38917); IV-COL-04 (C30554)."),
         refs=("SPEC: WIP v6 S4-R3 + S7-R13 (\"not offered in the column selector; its visibility is "
               "automatic\"); IV v3 Story-7 context note + S7-R7 (\"not a user-toggled column in the "
               "column-selection control\"). LIVE: viu-2026-08-03/batch-wip-iv/evidence/ui/"
               "colsel-work-in-progress.json - Location is item index 5 of 16, ariaChecked=false; "
               "colsel-inventory-value.json - Location is item index 4 of 11, ariaChecked=true; "
               "iv-singleloc.png shows the column still present with the chooser narrowed to Staging "
               "Lethbridge - 4310. Build " + BUILD_MARKER + ". "
               "ALSO OUR OWN CONTRADICTION (Rule 28 cross-case sweep): C30466 and C30507 both list "
               "Location inside the toggleable order while C30467 says it is not offered - that self-"
               "contradiction is resolved whichever way he rules. "
               "HONEST LIMIT: the separate one-location-USER read of the IV screen was CONFOUNDED by a "
               "persisted column selection (RECHECK-QUEUE B34, recorded NOT VERIFIED); the observation "
               "quoted to Chris is the ADMIN-NARROWING one, which is clean. Note also that the IV "
               "single-location CSV has NO Location column "
               "(evidence/location-matrix/inventory-value__SINGLE__plain.csv), so the download already "
               "follows the automatic rule while the screen follows the toggle."),
         resolve=("A -> dev ticket for both reports; the 6 cases stand as written and the two specs need "
                  "no change. B -> both specs need his edit and the 6 cases are reworded to the toggle "
                  "model (WIP off-by-default, IV on-by-default, or one agreed default). C -> the two "
                  "reports are documented as deliberately different and our cross-case contradiction is "
                  "closed by writing each report's own model down.")),
    dict(q="3",
         cases=("WIP-COL-05 (C30470); WIP-SORT-03 (C30485); WIP-FLT-03 (C30500); WIP-EXP-07 (C30516). "
                "Cross-report reference case, no change proposed: SBC-LBL-01 (C30134)."),
         refs=("OUR SOURCE: Chris Ward 2026-07-29, verbatim \"A is the correct answer\" plus the durable "
               "instruction \"Not just for these specs though -- really good to keep this in mind for all "
               "actions moving forward\" (chris-update-2026-07-29/wip-identifier-answer-2026-07-29.md). "
               "THE BUILD'S SOURCE: WIP v6 S4-R7 (\"the unit number on the first line in bold, and the "
               "vehicle identification number on the second line\") and S4-R9 (\"The Asset column sorts "
               "by unit number\"). LIVE DOM, verbatim: <div class=\"wip-asset\"><span class=\"wip-asset__"
               "unit text-weight-bold\">6548</span><span class=\"wip-asset__vin text-caption text-grey-7\">"
               "1FDSE3EL1EDB20609</span></div>. Serial-style values live in the VIN field: BULK PARTS1, "
               "12-06696, P631627, 86J8FAC1VALJ43SJY. Build " + BUILD_MARKER + ". "
               "HONEST LIMIT: we did not CREATE a non-vehicle asset - the report has no asset-creation "
               "surface; the terminology point is evidenced from existing records only "
               "(batch-wip-iv/STAGED-CHANGES.md group C1)."),
         resolve=("A -> dev ticket; the 4 cases stand as written and he edits S4-R7/S4-R9. B -> the 4 "
                  "cases revert to unit-number-leads and the durable CLAUDE.md ruling is narrowed to "
                  "exclude this report. C -> a label change is a new question for the whole suite, "
                  "including SBC-LBL-01. All 4 are HELD, not edited.")),
    dict(q="4",
         cases="SBR-EXP-10 (C30285); SBR-EXP-11 (C30286).",
         refs=("Three sources, three words. SPEC: SBR v15 S14-R15/S14-R16 both open with `Sales Rep`. "
               "OUR SOURCE: Chris Ward 2026-07-31 Q5=A, verbatim \"Rep is too much slang, let's do "
               "representative everywhere\" - so our cases correctly say \"Sales Representative\" "
               "(Rule 32). BUILD: header line read from "
               "evidence/location-matrix/sales-by-representative__SINGLE__summary.csv line 2 = "
               "`Representative,\"Inv. Hrs\",...`; the expanded file agrees. Build " + BUILD_MARKER + ". "
               "Analysis: viu-2026-08-03/LABEL-DIFF.md section A4, which explicitly says do NOT edit "
               "these two to \"Representative\" before he rules."),
         resolve=("A -> both cases reworded to \"Representative\" and he tidies S14-R15/R16. B -> dev "
                  "ticket; the cases stand. Either way these two cases also carry item 5 (the four "
                  "missing columns) and the separate LABEL-DIFF A6 findings (the expanded file puts "
                  "Invoice # before Date and heads the status column \"Invoice Status\"), so all of it "
                  "lands as ONE combined edit per case, each re-verified WHOLE against the current spec "
                  "per Rule 41.")),
    dict(q="5",
         cases="SBR-EXP-10 (C30285) - the same case as item 4, so both answers land in one edit.",
         refs=("SPEC (Rule 25, verbatim): SBR v15 S14-R15 \"Headers, in order: `Sales Rep`, `# Invoices`, "
               "`# Customers`, `Hrs Worked`, `Hrs Invoiced`, `Inv. Hrs`, `Labor Invoiced`, `Labor Margin`, "
               "`Parts Invoiced`, `Parts Margin`, `Margin`, `Margin %`, `Subtotal`.\" = 13. BUILD: 9 - "
               "`Representative, Inv. Hrs, Labor Invoiced, Labor Margin, Parts Invoiced, Parts Margin, "
               "Margin, Margin %, Subtotal` (evidence/location-matrix/"
               "sales-by-representative__SINGLE__summary.csv). Missing: # Invoices, # Customers, "
               "Hrs Worked, Hrs Invoiced. The payload DOES carry invoice_count, hours_worked and "
               "hours_invoiced, which is why LABEL-DIFF.md A5 reads it as an unfinished export, not a "
               "data defect. Build " + BUILD_MARKER + ". "
               "NOTE: S14-R16 itself carries a build note asking engineering to align the hours columns, "
               "so his own document already half-anticipates this."),
         resolve=("A -> dev ticket; C30285 stands as written. B -> C30285's enumeration shortens to nine "
                  "and he shortens S14-R15. Rule 42 applies either way: the rewritten enumeration must "
                  "carry a version-pinned anchor.")),
    dict(q="6",
         cases=("SBC-DATE-01 (C30102); SBC-DATE-03 (C30104) - NOT RUNNABLE TODAY; SBR-DATE-01 (C30201); "
                "PV-FILT-03 (C30330); WIP-FLT-04 (C30501); IV-DATE-01 (C30561)."),
         refs=("SPEC: SBC v13 S2-R2 closes the eleven-item list verbatim; WIP v6 S7-R6 and IV v3 S5-R1 "
               "close the same list in their own words. BUILD: nine presets + inline calendar + "
               "\"Range: N days\" + Apply, captured verbatim in viu-2026-08-03/evidence/"
               "date-range-picker.json. Build " + BUILD_MARKER + ". "
               "This is the application's SHARED date component, so it is a suite-wide product decision. "
               "Registered as DELIBERATE-DECISIONS.md entry 2.1 at risk MEDIUM precisely because "
               "C30104's steps cannot be executed today."),
         resolve=("A -> 6 cases reworded to the nine presets + calendar, C30104 becomes runnable, and 3 "
                  "specs need his edit. B -> dev ticket against the shared component; all 6 cases stand "
                  "and C30104 stays unrunnable until it ships. C -> partial dev ticket plus a partial "
                  "case rewrite.")),
    dict(q="7",
         cases="TU-EXP-01 (C30434); TU-EXP-02 (C30435).",
         refs=("SPEC (Rule 25, verbatim): TU v5 S7-R2 'an option labeled \"Download Summary (PDF)\"', "
               "S7-R3 '\"Download Expanded View (PDF)\"', S7-R4 '\"Download (CSV)\"' = three items with "
               "the Download prefix. BUILD: four items - `Summary (PDF)`, `Summary (CSV)`, "
               "`Expanded (PDF)`, `Expanded (CSV)`, no prefix (batch-pv-tu/VERDICTS.md TU-EXP-01; "
               "evidence/tu/ui/tu-ui-*.json). CONTRAST captured the same run: SBC and SBR both show the "
               "four long labels and MATCH their specs exactly (batch-sbc-sbr/VERDICTS.md F2, "
               "LABEL-DIFF.md B row SBC-EXP-01). Build " + BUILD_MARKER + ". "
               "Registered as DELIBERATE-DECISIONS.md entry 2.4, risk MEDIUM."),
         resolve=("A -> C30434 reworded to the four shipped strings (and C30435's Summary-PDF scope "
                  "re-checked), and he edits S7-R2/R3/R4 to four items. B -> dev ticket; both cases "
                  "stand.")),
    dict(q="8",
         cases="IV-EXP-04 (C30590).",
         refs=("SPEC (Rule 25, verbatim): IV v3 S10-R8 \"The PDF header shows the report name 'Inventory "
               "Value', the organization name, the selected period, and an 'as of' line naming the day "
               "the values represent...\" - the PDF only; no requirement mentions the CSV. S10-R15 "
               "governs only the \"Locations:\" line. BUILD: CSV line 1 = \"As of: 2026-08-03\" with the "
               "locations line on line 2 (evidence/location-matrix/inventory-value__SINGLE__plain.csv); "
               "extracted PDF header block reads `As of 2026-08-04` with NO colon "
               "(batch-wip-iv/STAGED-CHANGES.md B28). Build " + BUILD_MARKER + ". "
               "This surfaced from the SURFACE-MATRIX 1b sweep (Rule 40) - the IV \"Locations:\" line is "
               "the only one of six that is line 2 rather than line 1, because the as-of line sits above "
               "it."),
         resolve=("A -> he adds the CSV as-of line to S10-R8 (or a new requirement) and C30590 gains the "
                  "CSV half. B -> dev ticket to remove it; C30590 unchanged. C -> a wording ticket plus a "
                  "one-line spec edit; C30590 quotes whichever string he picks.")),
    dict(q="9",
         cases=("PV-PERM-01 (C30325); PV-PERM-03 (C30327); PV-API-04 (C30391); TU-NAV-07 (C30398); "
                "WIP-PERM-01 (C30526); WIP-PERM-02 (C30527); IV-PERM-01 (C30603); IV-PERM-02 (C30604)."),
         refs=("ALREADY RULED - no product decision is being re-asked. Chris Ward 2026-07-31 Q4=A, "
               "verbatim \"A - the intention is to not hide these from normal reports access. These were "
               "specced before CRP was built :)\" (and the same answer 2026-07-28), plus the QA LEAD's "
               "ruling 2026-08-03, verbatim: \"Yes all the reports will be gated by ONE permission FOR "
               "NOW.\" SPEC TEXT STILL STALE: PV v4 S1-R4 + S1-N2, IV v3 Story-1 prerequisite, TU v5 "
               "Story-1 prerequisite, WIP v6 Story-1 prerequisite; SBC v13 S1-R2 has been corrected. "
               "LIVE PROOF BOTH WAYS (viu-2026-08-03/SURFACE-MATRIX.md Matrix 2 + "
               "evidence/permissions/permission-matrix.json + minimal-role-proof.json): the FE permission "
               "catalogue holds exactly one report atom; an 8-atom Sales Representative holding only it "
               "got 200 on data AND export for all six; a Foreman without it got 403 on all six, data and "
               "export. Build " + BUILD_MARKER + ". "
               "This is why C30327 and C30391 are verified MORE strongly than written - the extra "
               "per-report permission does not merely fail to enforce, it does not exist."),
         resolve=("No case change on any answer - all 8 already follow the ruling. This row only closes a "
                  "documentation debt on four spec pages. The separate rescope-or-retire decision on "
                  "C30327 and C30391 is the QA LEAD's, not Chris's "
                  "(chris-answers-2026-08-01/staged-case-plan-CDE-2026-08-03.md).")),
    dict(q="10",
         cases="SBC-EXP-01 (C30159); SBC-EXP-14 (C30172).",
         refs=("SPEC: SBC v13 S18-R7 \"Exports (CSV, PDF, Print) are generated on the server...\" and "
               "S18-R10 \"If an export (CSV, PDF, or Print) is triggered while the active filters match "
               "no customers...\" - both still name Print, although Chris retired it in Story 16 "
               "(\"(removed - Print retired)\"). JIRA SV-8614 \"SBC - Story 16 - Print the report\" is "
               "still OPEN. BUILD: a sweep of every button, menu item and link for 'print' in text or "
               "aria-label returned an EMPTY list on all six reports "
               "(batch-sbc-sbr/VERDICTS.md F3; evidence/sales-by-customer/observe-full.json"
               "#toolbar.printControls). Build " + BUILD_MARKER + ". "
               "Registered as DELIBERATE-DECISIONS.md entry 1.4 at risk LOW, and as an OUTSIDE-IN.md "
               "external signal."),
         resolve=("No decision and no case change - C30159 explicitly asserts the ABSENCE of Print and "
                  "PASSED live. Included only so a documentation tidy-up and the closure of SV-8614 are "
                  "not forgotten; SV-8614 is a dev/ticket action, not Chris's.")),
]

# --------------------------------------------------------------------------------------
# COMPLETENESS PROOF (Standing Rule 17)
# --------------------------------------------------------------------------------------
SOURCES_SWEPT = [
    ("viu-2026-08-03/batch-sbc-sbr/VERDICTS.md + STAGED-CHANGES.md", "3 spec-vs-build items",
     "items 4, 5, 6 (SBC/SBR halves) and 10. The SBC nav-group question (spec says Performance, build "
     "says SALES) is NOT on this sheet - it is already question 9 of the 3 August sheet, unanswered"),
    ("viu-2026-08-03/batch-pv-tu/VERDICTS.md + STAGED-CHANGES.md", "3 spec-vs-build items",
     "items 1 (PV/TU halves), 6 (PV half) and 7. Its section-C product questions map 1:1 onto items 1, "
     "6 and 7"),
    ("viu-2026-08-03/batch-wip-iv/VERDICTS.md + STAGED-CHANGES.md", "4 spec-vs-build items",
     "items 1 (WIP/IV halves), 2, 3 and 8. Its group C (HELD pending a Chris ruling, 6 cases) is fully "
     "represented here: C1 -> item 3, C2 -> item 1, C3 is a build defect not a product question"),
    ("viu-2026-08-03/LABEL-DIFF.md", "4 items",
     "A2 -> item 7, A3 -> item 6, A4 -> item 4, A5 -> item 5. A6/A7/A8/A9 are build defects or wording "
     "fixes, not product decisions - excluded with reason"),
    ("viu-2026-08-03/SURFACE-MATRIX.md", "2 items",
     "Matrix 2 -> item 9; the 1b as-of/locations line sweep -> item 8. Matrix 1a (Location column "
     "placement) is read as an implementation slip with no product question"),
    ("viu-2026-08-03/DELIBERATE-DECISIONS.md", "4 of 35 entries name Chris Ward as the closer",
     "1.4 -> item 10, 2.1 -> item 6, 2.4 -> item 7; the fourth (logo treatment) is already question 4 of "
     "the 3 August sheet and is NOT repeated"),
    ("viu-2026-08-03/RECHECK-QUEUE.md", "0 new",
     "B18/B19/B34 are the same observations as items 1, 7 and 2. Every row on this sheet inherits the "
     "queue's PROVISIONAL status (Standing Rule 49)"),
    ("spec-watch-verification-2026-08-03/VERIFICATION.md + ADDENDUM", "0 new",
     "confirms the six live versions used in the source-currency block (SBC 13 / SBR 15 / PV 4 / TU 5 / "
     "WIP 6 / IV 3) and that only SBC moved since 07-31"),
    ("The six live descriptions, re-read for this sheet", "10 verbatim quotes extracted",
     "every reader-facing quote on this sheet was pulled from the live capture or the version-matched "
     "mirror, not from a summary (Standing Rule 15)"),
    ("PO-Questions-Chris-ReportSuite-2026-08-03.md (17 items, unanswered)", "0 duplicated",
     "checked item by item - nothing on this sheet repeats it. Its 17 items are description-text asks; "
     "this sheet is spec-versus-build behaviour. The overlap is only item 9 here vs its item 13, and "
     "ours adds the live both-ways proof it lacked"),
    ("chris-answers-2026-07-28 / -07-31 / -08-01, chris-update-2026-07-29, both videos",
     "8 candidates WITHDRAWN", "see the withdrawn appendix - every one quoted"),
    ("coverage-rederivation-2026-07-31/DELIBERATE-DECISIONS.md + COVERAGE-REDERIVATION.md", "0 new",
     "its open bucket is fully covered by the 3 August sheet; nothing there is a spec-vs-build item"),
    ("OUTSTANDING-ITEMS-REGISTER.md (Report Suite rows)", "0 new",
     "its Chris-facing rows are the 3 August sheet itself and the spec-watch deadline; the QA-branch row "
     "is the QA lead's, not Chris's"),
]

# --------------------------------------------------------------------------------------
# WITHDRAWN - already answered (QA-only appendix; NEVER put in front of Chris)
# --------------------------------------------------------------------------------------
WITHDRAWN = [
    ("Should the location chooser be hidden for a one-location person?",
     "ANSWERED 2026-07-31 Q1=A, verbatim \"A -- classic spec drift\". NOT re-asked. Item 1 asks a "
     "genuinely NEW question that could not exist before 3 August: the build disagrees with his ruling, "
     "so he must choose whether the product changes or the ruling does. The reader-facing text states "
     "his ruling rather than asking for it again."),
    ("Should the six reports be gated by their own dedicated permission?",
     "ANSWERED THREE TIMES: 2026-07-28 (\"these should be gated by normal reports access\"), 2026-07-31 "
     "Q4=A (\"the intention is to not hide these from normal reports access. These were specced before "
     "CRP was built :)\"), and the separate permissions sheet Q1=A. Plus the QA lead 2026-08-03: \"Yes "
     "all the reports will be gated by ONE permission FOR NOW.\" Item 9 asks ONLY for the four spec "
     "edits and says so in its own text."),
    ("Does the 10,000-row download limit apply to Parts Velocity, Technician Utilization and Work In "
     "Progress?",
     "ANSWERED 2026-07-31 Q3=A, verbatim \"A - this was not well thought out by me (the specs were "
     "written at different times)\". Suite-wide. The three cases exist and are pushed. Only his spec "
     "edit remains, and that is already item 15 of the 3 August sheet."),
    ("Which of the two \"too large to export\" messages is correct?",
     "ANSWERED 2026-07-31 Q2=A, verbatim \"A - great catch\". One suite-wide string, and the build "
     "returns it verbatim (confirmed live 2026-08-03 on the SBC guard). Nothing left to ask."),
    ("Should the asset identifier be the VIN chain rather than the serial number?",
     "ANSWERED 2026-07-29, verbatim \"A is the correct answer\", with the durable instruction to apply it "
     "everywhere. Item 3 does NOT re-ask it - it reports that the build did not implement it and asks "
     "which side now moves. His own VIN/serial caution is quoted inside item 3 rather than raised as its "
     "own question, because it is his point, not ours."),
    ("Should \"Rep\" be spelled out as \"Representative\"?",
     "ANSWERED 2026-07-31 Q5=A, verbatim \"Rep is too much slang, let's do representative everywhere\". "
     "Item 4 accepts that ruling and asks only about the THIRD spelling the build produced, which did "
     "not exist as a fact until 3 August."),
    ("Does Escape close the \"deactivate a representative\" pop-up?",
     "ANSWERED 2026-07-28, verbatim \"B.\" - Escape must NOT dismiss it. Our case follows it; only his "
     "spec edit remains and that is item 14 of the 3 August sheet."),
    ("Should \"All Time\" be offered in the date chooser?",
     "ANSWERED and already implemented - he removed it (SBC 2026-07-16 change log; WIP 2026-07-21 change "
     "log, recorded as \"a Chris product/UX decision\"). Live check 2026-08-03: not offered anywhere. "
     "Item 6 confirms its absence as a MATCH and asks only about the nine-versus-eleven list."),
]

# --------------------------------------------------------------------------------------
# NOT ASKED HERE (QA reference)
# --------------------------------------------------------------------------------------
NOT_ASKED = [
    ("The 17 items of PO-Questions-Chris-ReportSuite-2026-08-03 (still unanswered).",
     "Deliberately NOT duplicated. That sheet asks him to correct DESCRIPTION TEXT; this one asks him to "
     "choose between the description and the BUILD. Sending the same question twice on two sheets is how "
     "a bundled non-answer happens. Both sheets go together; item 9 here is the only overlap and it adds "
     "the live both-ways permission proof that sheet could not carry."),
    ("Where the location column sits in the shorter Summary downloads (spec-silent).",
     "Already question 3 of the 3 August sheet, unanswered. Not repeated."),
    ("\"The same logo treatment\" - three descriptions describe three different rules.",
     "Already question 4 of the 3 August sheet, unanswered. Not repeated."),
    ("Which Sales By Customer features were dropped.",
     "Already question 5 of the 3 August sheet, unanswered. Not repeated."),
    ("The Sales By Customer navigation group (description says Performance, build shows SALES).",
     "Already questions 8 and 9 of the 3 August sheet. Not repeated, although the build observation is "
     "new - it is recorded in batch-sbc-sbr/STAGED-CHANGES.md for whenever he answers."),
    ("Two printable downloads fail with a server error at full size (Parts Velocity, Inventory Value).",
     "A DEFECT, not a product decision - Standing Rule 7 forbids putting bugs in front of the PO. It is a "
     "dev ticket; the friendly over-size guard exists on the spreadsheet path and the printable path fails "
     "instead of using it. Request ids captured in batch-wip-iv/evidence/api/."),
    ("The location column's on-screen POSITION on Parts Velocity and Technician Utilization "
     "(sixth/second, not leftmost).",
     "Read as an implementation slip, not a product decision - no source asks for anything other than "
     "leftmost, so there is nothing for him to choose. Dev ticket."),
    ("The Work In Progress Estimates figure showing zero, and \"Inv. Hrs\" being shown but not "
     "downloadable.",
     "Defects. Dev tickets, not PO questions."),
    ("The QA branch being non-final, and fresh sign-in credentials.",
     "Not Chris's to give - the QA lead's / engineering's. Every observation on this sheet is PROVISIONAL "
     "until the branch is declared final (Standing Rule 49); the re-check queue is OPEN."),
]


# --------------------------------------------------------------------------------------
def md_rows():
    out = []
    group = None
    for i, r in enumerate(ROWS, 1):
        if r["group"] != group:
            group = r["group"]
            out.append("")
            out.append("# " + group)
        out.append("")
        out.append(f"## {i} — {r['topic']}")
        out.append("")
        out.append("**What your write-up says:** " + r["spec"])
        out.append("")
        out.append("**What the product actually does:** " + r["build"])
        out.append("")
        out.append("**Which do you want?** " + r["ask"])
        out.append("")
        for o in r["opts"]:
            out.append("- " + o)
        out.append("")
        out.append("**Your answer:** ____________________")
    return out


def write_md():
    n_dec = sum(1 for r in ROWS if r["group"] == G_DEC)
    n_wr = sum(1 for r in ROWS if r["group"] == G_WRITE)
    n_fyi = sum(1 for r in ROWS if r["group"] == G_FYI)
    L = [
        "# " + TITLE,
        "",
        "**STATUS: READY TO SEND** (not yet sent). On return: ingest verbatim, then revisit the "
        "affected cases per the standing workflow.",
        "",
        "This sheet is a **side-by-side of your own written descriptions against what the product "
        "actually does today**, taken from a live look at the test build on **3 and 4 August**. Every "
        "item shows you the exact words from your write-up, what we saw happen, and asks which of the "
        "two you would rather keep. Nothing here is a bug report - bugs go straight to engineering and "
        "are not on this sheet.",
        "",
        f"There are **{len(ROWS)} items**: **{n_dec} need you to choose something**, "
        f"**{n_wr} needs only a line changing in a write-up**, and **{n_fyi} needs nothing at all** and is "
        "here purely so you are not surprised by it later.",
        "",
        "**Please read it alongside the sheet dated 3 August** - that one asks you to correct wording; "
        "this one asks you to choose between the wording and the product. Deliberately, nothing is "
        "asked twice.",
        "",
        "**One honest caveat up front:** engineering has told us the test build is not finished yet. So "
        "everything described below is what we saw on 3 and 4 August, and we will look again when they "
        "say it is done. If your answer depends on that, say so and we will come back to you.",
    ]
    L += md_rows()
    L += [
        "",
        "---",
        "",
        "## QA Internal Mapping (QA-only — not for the PO)",
        "",
        "TestRail C-ids from `build/report-suite/testrail-id-map.csv` (Standing Rule 8). Links: "
        "https://shopview.testrail.io/index.php?/cases/view/<id>",
        "",
        "**Every C-id in this table is verified against the id-map at generation time — the generator "
        "aborts on a mismatch.** The 2026-07-31 sheet printed **PV-API-04 as C30388**, which is wrong: "
        "**PV-API-04 = C30391**, and **C30388 = PV-API-01**. Anyone acting on that row would have "
        "edited the wrong case.",
        "",
        "**Format** mirrors `PO-Questions-Chris-ReportSuite-2026-08-03.{md,xlsx}` 1:1 (Standing Rule "
        "16) — same tabs, headers, group banding and appendices. Only the three middle reader-facing "
        "column headings differ, because the QA lead asked for the spec-versus-build shape: *what your "
        "write-up says* · *what the product actually does* · *which do you want* · *blank answer*.",
        "",
        "| Item # | Affected internal case IDs (TestRail C-id) | Source refs (spec anchors + live evidence) | What each answer resolves to |",
        "|---|---|---|---|",
    ]
    for m in QA_MAP:
        L.append("| {} | {} | {} | {} |".format(
            m["q"],
            m["cases"].replace("\n", " "),
            m["refs"].replace("\n", " "),
            m["resolve"].replace("\n", " ")))
    L += [
        "",
        "### SOURCE-CURRENCY BLOCK (Standing Rule 31)",
        "",
        "| Source | Identifier | Version / last-updated | Checked | Verdict |",
        "|---|---|---|---|---|",
    ]
    for s in SOURCE_CURRENCY:
        L.append("| {} | {} | {} | {} | {} |".format(*s))
    L += [
        "",
        "**Nothing on this sheet claims completeness.** The build is a PARTIAL source by engineering's "
        "own statement, so every observation is PROVISIONAL and queued in "
        "`viu-2026-08-03/RECHECK-QUEUE.md` (Standing Rule 49). The designs are ABSENT and the epic had "
        "a Tier-1 currency check only (Standing Rule 37) — neither is claimed as more than that.",
        "",
        "### Completeness proof — every source swept (Standing Rule 17)",
        "",
        "| Source | Items found | Notes |",
        "|---|---|---|",
    ]
    for s, n, note in SOURCES_SWEPT:
        L.append(f"| {s} | **{n}** | {note} |")
    L += [
        "",
        f"**Totals: {len(SOURCES_SWEPT)} sources swept · deduplicated to {len(ROWS)} items on this "
        f"sheet ({n_dec} decisions + {n_wr} write-down + {n_fyi} awareness) · "
        f"{len(WITHDRAWN)} candidates WITHDRAWN as already answered · {len(NOT_ASKED)} not asked here, "
        "each with a reason.**",
        "",
        "### Withdrawn — already answered (QA-only appendix)",
        "",
        "Each of these was a candidate item; the source that answers it is quoted. **Not put in front "
        "of Chris.** Questions have been withdrawn for this reason on four previous sheets, so the "
        "check is mandatory before any item survives.",
        "",
        "| Candidate question | Already answered by |",
        "|---|---|",
    ]
    for c, a in WITHDRAWN:
        L.append(f"| {c} | {a} |")
    L += [
        "",
        "### Not asked here (QA reference)",
        "",
        "| Item | Why it is not on the sheet |",
        "|---|---|",
    ]
    for c, a in NOT_ASKED:
        L.append(f"| {c} | {a} |")
    L += [
        "",
        "---",
        "",
        "## OUTSTANDING — what I need from you",
        "",
        "Cross-project register: `build/OUTSTANDING-ITEMS-REGISTER.md` (Standing Rule 36).",
        "",
        "**From you (QA lead):**",
        "",
        "1. **Send this sheet to Chris, together with the 3 August one.** They are complementary and "
        "neither repeats the other. Eight of these ten items block a decision we cannot make for him: "
        "with each one open, the affected cases are frozen exactly as they are.",
        "2. **Nothing here is authorised to be applied.** No case has been edited and no TestRail write "
        "has been staged from this sheet (Standing Rule 6). The staged edits from the three VIU batches "
        "stay staged.",
        "3. **The QA branch is not final.** Engineering said so, so every observation on this sheet is "
        "provisional and the re-check queue is OPEN. Tell us when it is declared done and we re-run the "
        "queue immediately.",
        "",
        "**From Chris:** the eight decisions, a tick against item 9, and nothing at all for item 10.",
        "",
        "**Nothing else is outstanding from this sheet.**",
    ]
    open(os.path.join(HERE, BASE + ".md"), "w", encoding="utf-8").write("\n".join(L) + "\n")


def write_xlsx():
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    grp_fill = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True)
    cols = ["#", "Topic", "What your write-up says", "What the product actually does",
            "Which do you want?", "Your answer"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r = 3
    group = None
    for i, q in enumerate(ROWS, 1):
        if q["group"] != group:
            group = q["group"]
            r += 1
            gc = ws.cell(row=r, column=1, value=group.upper())
            gc.font = Font(bold=True)
            for j in range(1, 7):
                ws.cell(row=r, column=j).fill = grp_fill
        r += 1
        vals = [i, q["topic"], q["spec"], q["build"],
                q["ask"] + "\n\n" + "\n".join(q["opts"]), ""]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).alignment = wrap
    for col, w in zip("ABCDEF", [4, 30, 62, 62, 52, 22]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids from "
                 "build/report-suite/testrail-id-map.csv (Standing Rule 8); links "
                 "https://shopview.testrail.io/index.php?/cases/view/<id>. Every C-id verified against "
                 "the id-map at generation time (the 2026-07-31 sheet printed PV-API-04 as C30388; the "
                 "correct id is C30391 - C30388 is PV-API-01). FORMAT mirrors the 2026-08-03 sheet 1:1 "
                 "per Standing Rule 16.")
    ws2["A1"].font = Font(bold=True)
    cols2 = ["Item #", "Affected internal case IDs (TestRail C-id)",
             "Source refs (spec anchors + live evidence)", "What each answer resolves to"]
    for j, c in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r = 4
    for m in QA_MAP:
        for j, v in enumerate([m["q"], m["cases"], m["refs"], m["resolve"]], 1):
            ws2.cell(row=r, column=j, value=v).alignment = wrap
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="SOURCE-CURRENCY BLOCK (Standing Rule 31)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["Source", "Identifier", "Version / last-updated", "Checked / Verdict"], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for name, ident, ver, checked, verdict in SOURCE_CURRENCY:
        ws2.cell(row=r, column=1, value=name).alignment = wrap
        ws2.cell(row=r, column=2, value=ident).alignment = wrap
        ws2.cell(row=r, column=3, value=ver).alignment = wrap
        ws2.cell(row=r, column=4, value=checked + " — " + verdict).alignment = wrap
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="COMPLETENESS PROOF - EVERY SOURCE SWEPT").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Source", "Items found", "Notes"], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for s, n, note in SOURCES_SWEPT:
        ws2.cell(row=r, column=2, value=s).alignment = wrap
        ws2.cell(row=r, column=3, value=n).alignment = wrap
        ws2.cell(row=r, column=4, value=note).alignment = wrap
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="WITHDRAWN - already answered (not asked)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Candidate question", "Already answered by", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for c, a in WITHDRAWN:
        ws2.cell(row=r, column=2, value=c).alignment = wrap
        ws2.cell(row=r, column=3, value=a).alignment = wrap
        r += 1

    r += 1
    ws2.cell(row=r, column=1, value="NOT ASKED HERE (QA reference)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Item", "Why it is not on the sheet", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for c, a in NOT_ASKED:
        ws2.cell(row=r, column=2, value=c).alignment = wrap
        ws2.cell(row=r, column=3, value=a).alignment = wrap
        r += 1

    for col, w in zip("ABCD", [10, 36, 66, 60]):
        ws2.column_dimensions[col].width = w
    wb.save(os.path.join(HERE, BASE + ".xlsx"))


# --------------------------------------------------------------------------------------
def verify_cids():
    """Abort unless every internal-id/C-id pair printed on the QA-only sheet matches the id-map."""
    idmap = {}
    with open(IDMAP, newline="", encoding="utf-8") as f:
        for row in csv.reader(f):
            if row and row[0] != "internal_id":
                idmap[row[0]] = row[1]
    text = " ".join(m["cases"] + " " + m["refs"] + " " + m["resolve"] for m in QA_MAP)
    text += " " + " ".join(a for _, a in WITHDRAWN)
    text += " " + " ".join(a for _, a in NOT_ASKED)
    text += " " + " ".join(n for _, _, n in SOURCES_SWEPT)
    pairs = re.findall(r"([A-Z]{2,4}(?:-[A-Z]+)+-\d+)\s*[=(]\s*(C\d+)", text)
    if not pairs:
        raise SystemExit("C-ID VERIFICATION FOUND NOTHING TO CHECK - the regex or the data changed")
    bad = [(i, c, idmap.get(i, "<not in id-map>")) for i, c in pairs if idmap.get(i) != c]
    if bad:
        raise SystemExit("C-ID MISMATCH vs testrail-id-map.csv (internal, printed, actual): " + str(bad))
    # every bare C-id must also exist in the id-map
    known = set(idmap.values())
    strays = sorted({c for c in re.findall(r"\bC\d{5}\b", text) if c not in known})
    if strays:
        raise SystemExit("C-IDS PRINTED THAT ARE NOT IN THE ID-MAP AT ALL: " + str(strays))
    print(f"C-id verification: {len(pairs)} internal-id/C-id pairs checked, all MATCH the id-map; "
          f"0 stray C-ids")


def verify_reader_text_clean():
    """Rule 7 gate: reader-facing text must carry no case IDs, anchors, versions or jargon."""
    reader = []
    for r in ROWS:
        reader += [r["topic"], r["spec"], r["build"], r["ask"]] + r["opts"]
    blob = "\n".join(reader)

    problems = []

    # 1. TestRail C-ids
    for m in re.findall(r"\bC\d{4,6}\b", blob):
        problems.append(("TestRail case id", m))
    # 2. internal case IDs (SBC-EXP-10 etc.)
    for m in re.findall(r"\b(?:SBC|SBR|PV|TU|WIP|IV)-[A-Z]{2,6}-\d+\b", blob):
        problems.append(("internal case id", m))
    # 3. spec anchor codes (S4-R3, S14-R20, S1-N2, S2-E4) and section marks
    for m in re.findall(r"\bS\d+-[RNE]\d+[a-z]?\b", blob):
        problems.append(("spec anchor", m))
    for m in re.findall(r"[§¶]", blob):
        problems.append(("section mark", m))
    # 4. version numbers / build markers
    for m in re.findall(r"\bv\d+(?:\.\d+)*\b", blob):
        problems.append(("version number", m))
    for m in re.findall(r"\bversion\s+\d+\b", blob, re.I):
        problems.append(("version number", m))
    if BUILD_MARKER in blob:
        problems.append(("build marker", BUILD_MARKER))
    # 5. jira keys
    for m in re.findall(r"\bSV-\d+\b", blob):
        problems.append(("jira key", m))
    # 6. jargon / HTTP / API terms. NOTE: quoted spans of Chris's OWN document wording are
    #    exempt (altering a quote would be dishonest); everything else must be clean.
    quoted = re.findall(r"\"[^\"]*\"", blob)
    unquoted = blob
    for q in quoted:
        unquoted = unquoted.replace(q, " ")
    JARGON = [
        "API", "HTTP", "endpoint", "payload", "JSON", "403", "404", "500", "200", "201",
        "atom", "aria", "DOM", "regex", "boolean", "null", "backend", "back-end", "front-end",
        "frontend", "VIU", "TestRail", "Confluence", "Jira", "feature flag", "feature-flag",
        "CSV", "PDF", "UTF-8", "BOM", "server-side", "toggle", "column selector",
    ]
    for j in JARGON:
        if re.search(r"(?<![A-Za-z])" + re.escape(j) + r"(?![A-Za-z])", unquoted, re.I):
            problems.append(("jargon (outside a quote)", j))
    # 7. abbreviated report names
    for m in re.findall(r"\b(?:SBC|SBR|WIP|TU|PV|IV)\b", blob):
        problems.append(("abbreviated report name", m))

    if problems:
        seen, uniq = set(), []
        for kind, tok in problems:
            if (kind, tok) not in seen:
                seen.add((kind, tok))
                uniq.append(f"{kind}: {tok!r}")
        raise SystemExit("READER-FACING TEXT IS NOT RULE-7 CLEAN:\n  " + "\n  ".join(uniq))
    print("Rule 7 gate: reader-facing text clean — 0 case ids, 0 internal ids, 0 spec anchors, "
          "0 version numbers, 0 build markers, 0 ticket keys, 0 abbreviated report names, "
          "0 jargon outside a direct quote")


if __name__ == "__main__":
    verify_cids()
    verify_reader_text_clean()
    write_md()
    write_xlsx()
    n_dec = sum(1 for r in ROWS if r["group"] == G_DEC)
    n_wr = sum(1 for r in ROWS if r["group"] == G_WRITE)
    n_fyi = sum(1 for r in ROWS if r["group"] == G_FYI)
    print("wrote", BASE + ".md", "and", BASE + ".xlsx",
          f"| {len(ROWS)} items ({n_dec} decisions + {n_wr} write-down + {n_fyi} awareness)"
          f" | {len(WITHDRAWN)} withdrawn | {len(NOT_ASKED)} not-asked"
          f" | {len(SOURCES_SWEPT)} sources swept | {len(SOURCE_CURRENCY)} sources in the currency block")
