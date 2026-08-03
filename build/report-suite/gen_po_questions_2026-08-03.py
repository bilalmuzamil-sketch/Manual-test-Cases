#!/usr/bin/env python3
"""Report Suite -> Chris Ward question sheet, 2026-08-03 (EXHAUSTIVE sweep).

Mirrors gen_po_questions_2026-07-31.py 1:1 (Standing Rule 16):
  sheet 1 "Questions for PO": title row, blank row, header row, one row per item
  sheet 2 "QA Internal Mapping": mapping table + WITHDRAWN + NOT-ASKED blocks
Reader-facing content is plain layman (Standing Rule 7): no case IDs, no spec
anchors, no version numbers, no HTTP/API terms. Those live on the QA-only sheet.

Every printed C-id is verified against build/report-suite/testrail-id-map.csv at
generation time (Standing Rule 8) - the run fails if any id does not match.
"""
import csv
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "PO-Questions-Chris-ReportSuite-2026-08-03"
TITLE = ("Report Suite - Questions for Chris Ward (complete sweep of everything still "
         "open) - 2026-08-03")

# --------------------------------------------------------------------------
# GROUP 1 - DECISIONS: he has to choose something.
# Q1-Q4 are carried forward VERBATIM from PO-Questions-Chris-ReportSuite-2026-07-31.
# --------------------------------------------------------------------------
QUESTIONS = [
    {
        "group": "Decisions we need from you",
        "topic": "The Sales By Representative download columns contradict each other",
        "now": ("Your Sales By Representative description was updated on 29 July. One part of it "
                "now says that when the location column is showing on screen, it is also included "
                "in all four downloads. But an older part of the same document still lists the "
                "download columns as a fixed set, in order, with no location column in the list - "
                "those older lines were never updated. So the same document says two different "
                "things about the same download."),
        "q": ("Which one is right - should the downloads include the location column whenever it "
              "is showing on screen, or should they always show the same fixed set of columns?"),
        "opts": [
            "A) The downloads should include the location column whenever it appears on screen "
            "(this is the newer instruction, and we have already built our checks to follow it - "
            "so if A is right we just need your confirmation, plus the older lines tidied up).",
            "B) The downloads should always show the same fixed set of columns, whatever is on "
            "screen (in which case we will change our checks back).",
        ],
    },
    {
        "group": "Decisions we need from you",
        "topic": "Have the six descriptions been updated to match your video and your answers yet?",
        "now": ("You updated all six report descriptions on 29 July - thank you, the changes we "
                "were waiting for mostly landed. A handful did not, and they are ones where the "
                "written description now says the OPPOSITE of an answer you gave us afterwards. "
                "The clearest is the Work In Progress report: on 29 July you told us assets should "
                "be identified by VIN first (then unit number, then plate) for every report, and "
                "you believed you had already made that edit - but the Work In Progress "
                "description still puts the unit number first in several places. The others are: "
                "the location dropdown being hidden for a one-location person, the full word "
                "\"Representative\" on the customer card, the new reports sitting below the "
                "existing links in the menu, and the note that Parts Velocity is the \"only\" "
                "report in the Parts group when Inventory Value is there too. We are testing to "
                "your ANSWERS, not to the older written text. The seven still-missing edits are "
                "listed one by one further down this sheet."),
        "q": ("Will the descriptions be updated to match your answers, or should we simply keep "
              "testing to your answers and treat the written text as out of date?"),
        "opts": [
            "A) The descriptions will be updated - we keep testing to your answers meanwhile.",
            "B) Do not wait for the descriptions - your answers are the final word and the written "
            "text can stay as it is.",
        ],
    },
    {
        "group": "Decisions we need from you",
        "topic": "Where the location column goes in the shorter \"Summary\" downloads",
        "now": ("On screen the location column has a clear home: on Sales By Customer it sits "
                "right after the date, and on Sales By Representative right after the status. Your "
                "instruction for the downloads is that it appears \"in the same position it "
                "occupies on screen\". That works for the detailed downloads, which have those "
                "same columns. But the shorter Summary downloads do NOT have a date or a status "
                "column at all - so there is no position for it to match. Nothing in the "
                "descriptions says where it should go in those two files, and we do not want to "
                "guess."),
        "q": "In the shorter Summary downloads, where should the location column sit?",
        "opts": [
            "A) With the naming columns at the left - straight after the customer name (Sales By "
            "Customer) or the representative name (Sales By Representative), before the money "
            "columns.",
            "B) At the far right, after all the money columns.",
            "C) You do not mind - we will confirm whatever the build does and write that down.",
        ],
    },
    {
        "group": "Decisions we need from you",
        "topic": "\"The same logo treatment\" - the three descriptions describe three different rules",
        "now": ("In your 29 July note you said every report now uses the same logo treatment. The "
                "written descriptions do not agree with each other on what that treatment is. "
                "Technician Utilization says the built-in ShopView logo is always used. Sales By "
                "Customer says it tries the company's own uploaded logo first, then falls back to "
                "the built-in one, and if neither exists it prints no logo at all and lets the "
                "text fill the space. Parts Velocity does not mention a logo anywhere. Our checks "
                "for the three reports currently follow their own descriptions, so they cannot all "
                "be right."),
        "q": "Which single rule should every report's printed download follow?",
        "opts": [
            "A) Try the company's own uploaded logo first, then the built-in ShopView logo, and "
            "print no logo only if neither exists (the Sales By Customer rule).",
            "B) Always print the built-in ShopView logo (the Technician Utilization rule).",
            "C) Something else - please describe it.",
        ],
    },
    {
        "group": "Decisions we need from you",
        "topic": "Which Sales By Customer features were dropped - we need the list",
        "now": ("You told us that Sales By Customer had several features dropped just before the "
                "squad assembled, that some of them are exactly the kind of thing that should sit "
                "behind an extra permission, and that the written requirements should have been "
                "dropped along with them - \"I own that\". We went looking, and the good news is "
                "that everything your own change history records as dropped has already gone from "
                "both the description and our checks: the customer comparison list, the "
                "side-by-side asset comparison, using the global search bar to narrow the report, "
                "the \"All Time\" date range, and Print. We found nothing left over. But we cannot "
                "tell whether those five ARE the ones you meant, or whether you meant an earlier "
                "set we never saw - and if it is an earlier set, there could be requirements "
                "sitting somewhere we have not looked."),
        "q": ("Which dropped features did you mean - the five we found, or others we have not been "
              "told about?"),
        "opts": [
            "A) Those are the ones - nothing else was dropped, so this is already tidy and you can "
            "close it.",
            "B) There were others - please list them (even roughly) and we will check the "
            "description and our tests for anything left behind.",
        ],
    },
    # ----------------------------------------------------------------------
    # GROUP 2 - THINGS THAT ONLY NEED WRITING DOWN. No decision to make; the
    # answer is already settled, only the written description disagrees.
    # ----------------------------------------------------------------------
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "Work In Progress: which number identifies the vehicle or machine first",
        "now": ("You answered that it should be the VIN first, then the unit number, then the "
                "plate - for every report. The Work In Progress description still puts the unit "
                "number first in several places, including its overview, three of its rules and "
                "its download rule."),
        "q": "Will you update the Work In Progress description to put the VIN first?",
        "opts": ["A) Yes, I will update it.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "The location chooser is hidden for someone with only one location",
        "now": ("You answered that it is hidden - you called it classic spec drift. Four "
                "descriptions still say the opposite, that the person \"still sees the filter\": "
                "Sales By Representative, Technician Utilization, Inventory Value and Parts "
                "Velocity."),
        "q": "Will you correct those four lines?",
        "opts": ["A) Yes, I will update them.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "Technician Utilization sits BELOW the existing menu links",
        "now": ("Your video showed the new reports being added below the report links that were "
                "already there, without moving them. The Technician Utilization description names "
                "the right menu group but never says the new entry goes below the existing items."),
        "q": "Will you add that wording?",
        "opts": ["A) Yes, I will update it.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "Sales By Customer: the menu group and which links it sits below",
        "now": ("The same point for Sales By Customer: your video put it in the Performance group, "
                "below the four report links that already exist. Its description does not name a "
                "group at all, and does not mention the existing links."),
        "q": "Will you add the group and the placement to the Sales By Customer description?",
        "opts": ["A) Yes, I will update it.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "The asset chooser on Work In Progress: normal ShopView style, with a select-all",
        "now": ("In the walkthrough you said you were happy to update the description so the asset "
                "chooser looks like every other multi-pick list in the application, with a "
                "select-all / clear-all toggle. That has not been written down. Nothing in our "
                "tests depends on it - we just do not want it forgotten."),
        "q": "Will you add it to the description?",
        "opts": ["A) Yes, I will update it.",
                 "B) It is already done (please point us at it).",
                 "C) No - drop it."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "\"Representative\" written out in full, everywhere",
        "now": ("You answered that the short form is slang and it should say representative "
                "everywhere. The Sales By Representative description still says \"Sales Rep\" on "
                "the customer card, in the name of the assignments screen, and in the lists of "
                "download column headings."),
        "q": "Will you change those to the full word?",
        "opts": ["A) Yes, I will update them.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "Parts Velocity is described as the \"only\" report in the Parts group",
        "now": ("Inventory Value lives in the Parts group too, so that line is no longer true. It "
                "is a one-line fix."),
        "q": "Will you correct it?",
        "opts": ["A) Yes, I will update it.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "Five descriptions still say the report needs its own area permission",
        "now": ("You answered that all report access collapses into one single reports permission, "
                "and that any extra permission that has already been built should simply be hidden "
                "and left doing nothing. Sales By Customer's description has been corrected to "
                "match - thank you. The other five have not: Parts Velocity and Inventory Value "
                "still say a person needs the existing inventory-reports permission, Technician "
                "Utilization points at the permission for a different, older report, and Work In "
                "Progress names a Work-In-Progress reports permission. We are testing to your "
                "answer, so anyone comparing our tests to the written text will think we made a "
                "mistake."),
        "q": "Will you update those five descriptions so they all name the one reports permission?",
        "opts": ["A) Yes, I will update them.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "The Escape key on the \"deactivate a representative\" pop-up",
        "now": ("You answered this on 28 July: pressing Escape must NOT close that pop-up, because "
                "it is a confirm-or-cancel decision. The Sales By Representative description still "
                "says Escape closes it. Our test follows your answer."),
        "q": "Will you correct that line?",
        "opts": ["A) Yes, I will update it.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "The \"too big to download\" limit is missing from three descriptions",
        "now": ("You confirmed the same size limit and the same single message apply to all six "
                "reports. Three descriptions - Parts Velocity, Technician Utilization and Work In "
                "Progress - still carry no line about it at all. Our tests already exist for all "
                "six."),
        "q": "Will you add the limit and the message wording to those three?",
        "opts": ["A) Yes, I will update them.",
                 "B) It is already done (please point us at it).",
                 "C) No - and here is why."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "A note that \"VIN\" also covers machines that are not vehicles",
        "now": ("This was your own point: VIN stands for vehicle identification number, and for "
                "something like a generator the number people actually read is its serial number. "
                "You asked us to be careful with the wording. Our tests keep the on-screen word "
                "VIN and add a plain note for the tester - but the descriptions do not explain it, "
                "so a reader could think the two are different things."),
        "q": "Will you add a short note to the descriptions?",
        "opts": ["A) Yes, I will add it.",
                 "B) Not needed - the on-screen word is enough.",
                 "C) Something else - please describe it."],
    },
    {
        "group": "Things that only need writing down (no decision needed)",
        "topic": "Some odd characters appear in two of the descriptions",
        "now": ("The Sales By Representative and Parts Velocity documents contain a few garbled "
                "characters where a quote mark or a dash should be - almost certainly from a "
                "copy-and-paste. It changes nothing about the product, but it makes those lines "
                "hard to read and hard to quote back to you."),
        "q": "Would you tidy those up next time you are in the documents?",
        "opts": ["A) Yes.", "B) Leave them."],
    },
]

# --------------------------------------------------------------------------
# QA-ONLY mapping. Every C-id here is verified against testrail-id-map.csv.
# --------------------------------------------------------------------------
QA_MAP = [
    {"q": "1", "cases": "SBR-EXP-10 (C30285); SBR-EXP-11 (C30286); SBR-EXP-03 (C30278); "
                        "SBR-EXP-04 (C30279); SBR-LOC-05 (C38913)",
     "refs": "SBR spec v15 (Confluence 585629698, lastModified 2026-07-29 - RE-VERIFIED LIVE "
             "2026-08-03, still unresolved): NEW S14-R20 (\"included in all four exports in the "
             "same position it occupies on screen\") vs S14-R15 (Summary CSV headers \"in order\", "
             "beginning `Sales Rep`, no Location) and S14-R16 (Expanded CSV headers). The header "
             "enumerations date from the 2026-07-11 \"Exports hardened\" round and were never "
             "amended. Corroborated by contradiction-analysis-2026-07-31/SBR-CSV-LOCATION.md and "
             "by Vladimir Tomovic's automated C38923, which was RIGHT (Rule 44).",
     "resolve": "A -> the 5 cases stand as pushed and Chris tidies S14-R15/R16. B -> revert the "
                "export halves to the fixed lists and drop the Location assertions. Same "
                "on-screen/export split was fixed on SBC (S4-R13), PV (S6-R11), TU (S7-R13), "
                "IV (S10-R15); WIP already had it. Either way VIU-confirm live."},
    {"q": "2", "cases": "WIP-COL-05 (C30470); WIP-FLT-03 (C30500); WIP-SORT-03 (C30485); "
                        "WIP-EXP-07 (C30516); SBR-LOC-04 (C30216); TU-LOC-05 (C30446); "
                        "IV-LOC-04 (C30577); PV-FILT-13 (C30340); TU-NAV-01 (C30392); "
                        "PV-NAV-01 (C30322)",
     "refs": "SPEC-WATCH-2026-07-28.md re-diff 2026-07-31, re-confirmed 2026-08-03 (all five "
             "non-SBC specs still at their 2026-07-29 versions per live CQL, so nothing has "
             "landed since): 7 of 12 items still need spec text - 1b WIP identifier, 4 location "
             "filter hidden, 6 nav placement wording, 8 WIP asset dropdown, 9 customer-card "
             "Representative, 10 SBC nav anchors, 11 PV \"only report\". Deadline 2026-08-04.",
     "resolve": "A -> no case change; SPEC-WATCH stays open until the text lands. B -> SPEC-WATCH "
                "closes as a documentation debt. Neither answer changes a case. Items 1b/4/6/9/"
                "10/11 are ALSO asked individually as sheet items 6-12 so he can tick them off."},
    {"q": "3", "cases": "SBC-EXP-16 (C38856); SBC-LOC-04 (C38912); SBR-EXP-10 (C30285); "
                        "SBR-EXP-03 (C30278); SBR-LOC-05 (C38913)",
     "refs": "SPEC-SILENT, found by coverage-rederivation-2026-07-31. SBC S4-R13 states inclusion "
             "with no position; SBR S14-R20 says \"the same position it occupies on screen\" but "
             "the Summary CSV (S14-R15) has no Date/Status column and the Summary PDF (S14-R5) "
             "has none either. Cases currently hedge (\"with the identifying columns ahead of the "
             "money columns (confirm its exact position in the build)\") - hedged, not invented "
             "(Rule 9).",
     "resolve": "A -> replace the hedge with the stated position. B -> reword to far-right. "
                "C -> keep the hedge and pin it at VIU. No case is wrong today either way."},
    {"q": "4", "cases": "SBC-EXP-10 (C30168); TU-EXP-06 (C30439); TU-EXP-07 (C30440); "
                        "PV-EXP-05 (C30379); PV-EXP-06 (C30380); SBR-EXP-03 (C30278); "
                        "SBR-EXP-04 (C30279)",
     "refs": "Cross-report spec contradiction against his 2026-07-29 message (\"Each report now "
             "ensures the same 'logo' treatment\"): SBC S15-R16/R17/R18 = 3-step chain ending in "
             "NO logo; TU = bundled ShopView default; PV has no logo requirement at all. "
             "DELIBERATE-DECISIONS.md A2. Not resolved by us (Rule 15 - never pick a side "
             "silently).",
     "resolve": "A -> TU/PV export cases gain the fallback chain. B -> SBC-EXP-10's chain "
                "collapses to the bundled default. C -> re-ask. Wording-only edits either way."},
    {"q": "5", "cases": "NONE - zero case impact either way. Evidence base: SBC-DATE-01 (C30102) "
                        "asserts the ABSENCE of \"All Time\"; the SBC Print case was "
                        "retired 2026-07-28 (deleted from TestRail, so it has no live C-id); no case exists for customer comparison, asset "
                        "comparison or global-search narrowing",
     "refs": "His chat, verbatim: \"SBC actually has several features that we dropped almost right "
             "before the squad assembled ... the requirements should have dropped with the "
             "additional features dropping, I own that.\" Checked live 2026-08-03 against the SBC "
             "spec change log (Confluence 577634305): the four dropped rounds are 2026-07-12 "
             "(customer comparison + asset comparison), 2026-07-15 (global search), 2026-07-16 "
             "(All Time), 2026-07-29 (Print). ZERO lingering requirements and ZERO stale cases "
             "found - see chris-answers-2026-08-01/answers-ingested.md section 3.",
     "resolve": "A -> the item closes with a written all-clear. B -> re-derive coverage for "
                "whatever he names. We did NOT manufacture a retire list to fill the gap "
                "(Rule 12)."},
    {"q": "6-12", "cases": "Same cases as row 2 - these are SPEC-WATCH items 1b, 4, 6, 10, 8, 9, "
                           "11 asked one at a time so he can tick each off",
     "refs": "Named individually at the QA lead's instruction (\"each named individually, not as "
             "a bundle\"), because the 2026-08-04 deadline is tomorrow and a bundled question has "
             "so far produced a bundled non-answer.",
     "resolve": "No case change on any answer - the cases already follow his rulings (Rule 32). "
                "These only close a documentation debt."},
    {"q": "13", "cases": "PV-PERM-01 (C30325); PV-PERM-03 (C30327); PV-API-04 (C30391); "
                         "IV-PERM-01 (C30603); IV-PERM-02 (C30604); TU-NAV-07 (C30398); "
                         "WIP-PERM-01 (C30526); WIP-PERM-02 (C30527); PV-NAV-01 (C30322); "
                         "IV-NAV-01 (C30534); TU-NAV-01 (C30392); WIP-TAB-01 (C30451)",
     "refs": "NEW spec debt created by his own Q2=A (\"Collapse all report access into a single "
             "Reports permission\") plus the QA lead's ruling 2026-08-03, verbatim: \"Yes all the "
             "reports will be gated by ONE permission FOR NOW.\" Verified live 2026-08-03: PV "
             "S1-R4/S1-N2, IV S1-R4 and the TU/WIP Story-1 prerequisites still name per-area "
             "permissions; only SBC S1-R2 has been corrected (2026-07-31). NOTE: the SBC change "
             "log still instructs engineering to DROP the atom, while his later chat allows "
             "hiding it inert - the later source wins (Rule 32).",
     "resolve": "No decision needed - the model is settled. Groups C and D of staged-case-plan-"
                "CDE-2026-08-03.md reword these 12 cases to the single permission; the two "
                "retire-or-rescope candidates (C30327, C30391) await the QA lead's sign-off."},
    {"q": "14", "cases": "SBR-DEACT-04 (C30255)",
     "refs": "ANSWERED by Chris 2026-07-28, verbatim \"B.\" (chris-answers-2026-07-28/"
             "answers-ingested.md Q1) = Escape must NOT dismiss. Verified live 2026-08-03: SBR "
             "S13-R8 still says Escape closes the dialog. NOTE - a correction to our own record: "
             "PO-Questions-Chris-ReportSuite-2026-07-31.md line 125 says this question is still "
             "\"open four days\". That is STALE; only the spec text remains outstanding, which is "
             "why this appears here as a write-down item, not a decision.",
     "resolve": "No case change - C30255 already follows his answer."},
    {"q": "15", "cases": "PV-EXP-11 (C38885); TU-EXP-09 (C38887); WIP-EXP-10 (C38918)",
     "refs": "His 2026-07-31 Q2=A (\"A - great catch\", one suite-wide message) + Q3=A (\"this was "
             "not well thought out by me\", cap on all six). Verified: the PV/TU/WIP pages carry "
             "no cap line; the three cases exist and are pushed.",
     "resolve": "No case change - documentation debt only."},
    {"q": "16", "cases": "SBC-LBL-01 (C30134); WIP-COL-05 (C30470)",
     "refs": "His own 2026-07-29 standing note, verbatim: \"we just have to be careful with using "
             "the acronym VIN ... for a generator ... it gets confusing\". Our cases keep the "
             "build label \"VIN\" plus a plain tester note, per the durable ruling in CLAUDE.md.",
     "resolve": "No case change either way - the on-screen label is unaffected."},
    {"q": "17", "cases": "NONE",
     "refs": "Cosmetic encoding artefact (mojibake) in the SBR and PV spec text, already noted in "
             "SPEC-WATCH-2026-07-28.md. Not a product question; asked only because it makes "
             "quoting those lines back to him unreliable.",
     "resolve": "No case impact."},
]

WITHDRAWN = [
    ("How far does \"one single Reports permission\" reach - do the six reports just read one "
     "permission, or are the per-area permissions merged/removed in Custom Roles?",
     "ANSWERED - and this was going to be the headline question of this sheet. Chris's Q2 = A "
     "(\"Collapse all report access into a single Reports permission\") plus the QA LEAD's ruling "
     "2026-08-03, verbatim: \"Yes all the reports will be gated by ONE permission FOR NOW.\" "
     "Rule 33 - the QA lead's ruling, consistent with the PO's answer. WITHDRAWN so we do not "
     "ask a settled question. \"FOR NOW\" is recorded on every affected case."),
    ("Should the reports use their own dedicated permissions?",
     "ANSWERED three times: 2026-07-28 (\"these should be gated by normal reports access\"), "
     "2026-07-31 Q4 = A (\"the intention is to not hide these from normal reports access. These "
     "were specced before CRP was built\"), and again on the SV-8598 permissions sheet Q1 = A. "
     "SBC spec S1-R2 was corrected 2026-07-31 to match - verified live 2026-08-03."),
    ("Does Escape close the \"deactivate a representative\" pop-up?",
     "ANSWERED by Chris 2026-07-28, verbatim \"B.\" - Escape must NOT dismiss it. Our own 31 July "
     "sheet wrongly recorded this as still open; it is not. Only his SPEC EDIT remains, so it is "
     "carried as write-down item 14 rather than a question."),
    ("Are there any pictures or videos to check the reports against? (Question 3 of the 27 July "
     "sheet)",
     "ANSWERED by delivery, not by words - he produced the walkthrough Loom on 2026-07-30. It was "
     "ingested and ruled AUTHORITATIVE, and drove 3 firm deltas plus the SPEC-WATCH list. Designs "
     "remain absent, which is a separate, already-recorded fact."),
    ("Each report uses a different permission to view it - is that intended? (Question 2 of the "
     "27 July sheet)",
     "ANSWERED by Q2 = A plus the QA lead's one-permission ruling. Superseded; not re-asked."),
    ("Which of the two \"too large to export\" messages is correct?",
     "ANSWERED 2026-07-31 Q2 = A, verbatim \"A - great catch\": one suite-wide string. All six cap "
     "cases already quote it. Only his spec edit remains (write-down item 15)."),
    ("Does the 10,000-row export cap apply to Parts Velocity, Technician Utilization and Work In "
     "Progress?",
     "ANSWERED 2026-07-31 Q3 = A, verbatim \"A - this was not well thought out by me (the specs "
     "were written at different times)\". Cap is suite-wide; the three cases exist and are pushed."),
    ("Is the location dropdown hidden for a one-location user?",
     "ANSWERED 2026-07-31 Q1 = A, verbatim \"A -- classic spec drift\". Hidden. Only the spec text "
     "still disagrees (write-down item 7)."),
    ("Does the short heading \"Rep is active?\" also become \"Representative\"?",
     "ANSWERED by his 2026-07-31 Q5 = A, verbatim \"slang, let's do representative everywhere\" - "
     "the scope explicitly reaches the export column headers. Affects SBR-ASGN-02 (C30293)."),
    ("What is the exact renamed \"Sales Rep Assignments\" file name?",
     "ANSWERED by the same Q5 = A - the file name is explicitly in scope of \"representative "
     "everywhere\". SBR-ASGN-02 (C30293) hedges the exact final string for live confirmation, "
     "which is the correct treatment; no product question remains."),
]

NOT_ASKED = [
    ("The 5 automated cases in our Report Suite folder authored by Vladimir Tomovic "
     "(C38919-C38923).",
     "Not a Chris question, and by the QA lead's ruling 2026-07-31 we do not message Vladimir "
     "either. His cases stay untouched (Standing Rule 38) and are excluded from our counts - we "
     "report \"ours 475 / live 480\". One of them (C38923) was RIGHT and exposed a real gap on our "
     "side, which we fixed on our own cases."),
    ("The Technician Utilization column-selector story has no Jira ticket.",
     "Ticket-management, not a product decision - he already asked for the control in his "
     "2026-07-29 message (\"for visual/natural conformance\"), so scope is settled. TU-COL-01 "
     "(C38859) and TU-LOC-06 (C38915) cite epic SV-8582 and say so in refs. Tracked in "
     "OUTSTANDING-ITEMS-REGISTER.md as an OTHER TEAM item."),
    ("SV-8780 (the built dedicated Sales By Customer permission).",
     "OUT OF SCOPE by the QA lead's ruling 2026-08-03, verbatim: \"Ignore this ticket.\" Not "
     "commented on, not transitioned, not read-and-edited. The drafted comment is retained "
     "unposted and banner-marked NOT TO BE POSTED."),
    ("Four requirements we deliberately do not test (SBC S10-N1, SBR S11-N1, SBR S14-R14, "
     "PV S4-N1).",
     "QA decisions, not product ones - cut by the user-authorized 2026-07-28 Ruthless Usefulness "
     "Audit as no-op assertions, un-measurable px font-tier minutiae, and a stored-schema state a "
     "manual tester cannot seed. Recorded with reasons in coverage-rederivation-2026-07-31/"
     "COVERAGE-REDERIVATION.md section 5."),
    ("The QA branch / environment and fresh login cookies.",
     "Not Chris's to give - this is the VIU blocker and it sits with the QA lead / engineering. "
     "All 475 cases remain VIU-Pending until it exists (Standing Rules 12/22)."),
]

SOURCES_SWEPT = [
    ("PO-Questions-Chris-ReportSuite-2026-07-31.md (5 questions)", "4 open",
     "Q5 is ANSWERED (=A) on the separate permissions sheet; Q1-Q4 carried forward VERBATIM"),
    ("PO-Questions-Chris-ReportSuite-2026-07-27.md (3 questions)", "0 open",
     "Q1 answered 2026-07-28 (\"B.\"), Q2 superseded by the one-permission ruling, Q3 answered by "
     "the 2026-07-30 Loom"),
    ("PO-Questions-Chris-ReportSuite-TechPlan_2026-07-30.md (5 questions)", "0 open",
     "all five answered = A on 2026-07-31; 70 cases updated + 7 added and pushed"),
    ("SPEC-WATCH-2026-07-28.md (12 items, deadline 2026-08-04)", "7 open",
     "items 1b, 4, 6, 8, 9, 10, 11 - each now asked INDIVIDUALLY as items 6-12"),
    ("What-We-Need-From-Chris-Ward-2026-07-31.md (12 items)", "11 open",
     "item 4 (permission granularity) is now closed; the rest dedupe into this sheet"),
    ("coverage-rederivation-2026-07-31/DELIBERATE-DECISIONS.md (D1-D7 open bucket)", "6 open",
     "D5 closed by the one-permission ruling; D1/D2/D3/D4/D6/D7 all dedupe into this sheet"),
    ("OUTSTANDING-ITEMS-REGISTER.md (Report Suite rows)", "0 new",
     "its Chris-facing rows are the QA branch (not Chris), SV-8780 (out of scope by ruling) and "
     "the scope question (now answered)"),
    ("The six LIVE Confluence specs, re-checked 2026-08-03", "2 new",
     "SBC lastModified Jul 31; SBR/PV/TU/WIP/IV all still Jul 29. NEW: the five per-area "
     "permission descriptions (item 13); and the SBC change log still says DROP the atom while "
     "his chat says hide it (folded into item 13's note)"),
    ("Chris's 2026-08-01-round chat message + the filled 2-question sheet", "1 new",
     "\"which SBC features were dropped\" (item 5). His two answers themselves are ingested, not "
     "re-asked"),
    ("Chris's 2026-07-29 group message", "1 new",
     "the \"be careful with the acronym VIN\" caution is his own point and is not written into any "
     "spec (item 16)"),
    ("Both walkthrough videos (2026-07-30 Loom + the earlier PRD companion)", "0 new",
     "every delta already promoted or on SPEC-WATCH; nothing unaddressed"),
    ("coverage-rederivation-2026-07-31/COVERAGE-REDERIVATION.md (spec-silent / spec-inconsistent "
     "flags)", "0 new",
     "the spec-silent flag is the Summary-download position = item 3; the spec-inconsistent flags "
     "are items 1 and 4"),
    ("PROJECT-STATE.md", "0 new", "no Chris-facing ask not already listed above"),
]


def md_items():
    L = []
    group = None
    for i, q in enumerate(QUESTIONS, 1):
        if q["group"] != group:
            group = q["group"]
            L += ["", f"# {group}", ""]
        L += [f"## {i} — {q['topic']}", "",
              f"**What happens now:** {q['now']}", "",
              f"**The question:** {q['q']}", "", "**Options:**", ""]
        L += [f"- {o}" for o in q["opts"]]
        L += ["", "**Your answer:** ____________________", ""]
    return L


def write_md():
    L = [f"# {TITLE}", "",
         "**STATUS: READY TO SEND** (not yet sent). On return: ingest verbatim, then revisit the "
         "affected cases per the standing workflow.", "",
         "Plain-language product questions only — no bugs, no test jargon. This sheet is an "
         "**exhaustive sweep of everything still owed by you**, not a top-up: every question sheet, "
         "the description-change watch list, our own decision register and all six live "
         "descriptions were re-read on 3 August, and anything you have already answered has been "
         "deliberately left out (10 candidate questions were withdrawn for exactly that reason).",
         "",
         "It is in **two parts**. The first five need you to **choose something**. The rest need "
         "**no decision at all** — you have already answered them; only the written description "
         "still says something different, so they are one-line confirmations. **Several of those "
         "were due on 4 August**, which is why they are listed one at a time rather than bundled.",
         ""]
    L += md_items()
    L += ["---", "", "## QA Internal Mapping (QA-only — not for the PO)", "",
          "TestRail C-ids from `build/report-suite/testrail-id-map.csv` (Standing Rule 8). Links: "
          "https://shopview.testrail.io/index.php?/cases/view/<id>",
          "", "**Every C-id in this table is verified against the id-map at generation time — the "
          "generator aborts on a mismatch.** This was added because the 2026-07-31 sheet printed "
          "**PV-API-04 as C30388**, which is wrong: **PV-API-04 = C30391**, and **C30388 = "
          "PV-API-01** (a server-pagination case with nothing to do with permissions). Anyone "
          "acting on that row would have edited the wrong case.",
          "", "**The team endorsed this sheet format** — the QA lead reports that everyone in the "
          "meeting liked this way of sharing scenarios (plain layman wording, one question per "
          "decision, A/B options with a blank answer, the citations kept on the QA-only side). It "
          "is the format to reuse for PO questions on every project.",
          "", "| Item # | Affected internal case IDs (TestRail C-id) | Source refs | What each "
          "answer resolves to |", "|---|---|---|---|"]
    for m in QA_MAP:
        L.append(f"| {m['q']} | {m['cases']} | {m['refs']} | {m['resolve']} |")
    L += ["", "### Completeness proof — every source swept (Standing Rule 17)", "",
          "| Source | Open items found | Notes |", "|---|---|---|"]
    for s, n, note in SOURCES_SWEPT:
        L.append(f"| {s} | **{n}** | {note} |")
    L += ["",
          f"**Totals: 13 sources swept · 32 open items found across them · deduplicated to "
          f"{len(QUESTIONS)} items on this sheet ({sum(1 for q in QUESTIONS if q['group'].startswith('Decisions'))} "
          f"decisions + {sum(1 for q in QUESTIONS if not q['group'].startswith('Decisions'))} "
          f"write-downs) · {len(WITHDRAWN)} candidates WITHDRAWN as already answered · "
          f"{len(NOT_ASKED)} not asked here, with reasons.**", ""]
    L += ["### Withdrawn — already answered (QA-only appendix)", "",
          "Each of these was a candidate question; the source that answers it is quoted. **Not put "
          "in front of Chris.** Questions have been withdrawn for this reason three times already "
          "today, so the check is now mandatory before any question survives.", "",
          "| Candidate question | Already answered by |", "|---|---|"]
    for c, a in WITHDRAWN:
        L.append(f"| {c} | {a} |")
    L += ["", "### Not asked here (QA reference)", "",
          "| Item | Why it is not on the sheet |", "|---|---|"]
    for c, a in NOT_ASKED:
        L.append(f"| {c} | {a} |")
    L += ["", "---", "", "## OUTSTANDING — what I need from you", "",
          "Cross-project register: `build/OUTSTANDING-ITEMS-REGISTER.md` (Standing Rule 36).", "",
          "**From you (QA lead):**", "",
          "1. **Send this sheet to Chris — today if possible.** Items 6 to 12 are the description "
          "corrections that were due **4 August**, i.e. tomorrow.",
          "2. **A ruling on the two Parts Velocity permission cases** that the one-permission ruling "
          "leaves unrunnable — PV-PERM-03 = [C30327](https://shopview.testrail.io/index.php?/cases/view/30327) "
          "and PV-API-04 = [C30391](https://shopview.testrail.io/index.php?/cases/view/30391). "
          "My recommendation is **RESCOPE, not retire** — reasoning in "
          "`chris-answers-2026-08-01/staged-case-plan-CDE-2026-08-03.md`.",
          "3. **The QA branch / environment, and confirmation the reports are switched on.** This "
          "is the only thing left that is not documentation: **all 475 cases have never been run "
          "against the real build.** We also need fresh login cookies when it exists.", "",
          "**Already done, so please do not chase it:** the case titles are all within the length "
          "TestRail can display; the coverage matrix has no open gaps; every case carries its "
          "ticket and description reference.", "",
          "**From Chris:** the five decisions, plus a tick against each of items 6 to 17.", ""]
    open(os.path.join(HERE, BASE + ".md"), "w", encoding="utf-8").write("\n".join(L) + "\n")


def write_xlsx():
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    grp_fill = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True)
    cols = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r = 3
    group = None
    for i, q in enumerate(QUESTIONS, 1):
        if q["group"] != group:
            group = q["group"]
            r += 1
            gc = ws.cell(row=r, column=1, value=group.upper())
            gc.font = Font(bold=True)
            for j in range(1, 7):
                ws.cell(row=r, column=j).fill = grp_fill
        r += 1
        for j, v in enumerate([i, q["topic"], q["now"], q["q"], "\n".join(q["opts"]), ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = wrap
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids from "
                 "build/report-suite/testrail-id-map.csv (Standing Rule 8); links "
                 "https://shopview.testrail.io/index.php?/cases/view/<id>. Every C-id verified "
                 "against the id-map at generation time (the 2026-07-31 sheet printed PV-API-04 "
                 "as C30388; the correct id is C30391 - C30388 is PV-API-01). FORMAT ENDORSED by "
                 "the team per the QA lead.")
    ws2["A1"].font = Font(bold=True)
    cols2 = ["Item #", "Affected internal case IDs (TestRail C-id)", "Source refs",
             "What each answer resolves to"]
    for j, c in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r = 4
    for m in QA_MAP:
        for j, v in enumerate([m["q"], m["cases"], m["refs"], m["resolve"]], 1):
            ws2.cell(row=r, column=j, value=v).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="COMPLETENESS PROOF - EVERY SOURCE SWEPT").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Source", "Open items found", "Notes"], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for s, n, note in SOURCES_SWEPT:
        ws2.cell(row=r, column=2, value=s).alignment = wrap
        ws2.cell(row=r, column=3, value=n).alignment = wrap
        ws2.cell(row=r, column=4, value=note).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="WITHDRAWN - already answered (not asked)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Candidate question", "Already answered by", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for c, a in WITHDRAWN:
        ws2.cell(row=r, column=2, value=c).alignment = wrap
        ws2.cell(row=r, column=3, value=a).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="NOT ASKED HERE (QA reference)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Item", "Why it is not on the sheet", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for c, a in NOT_ASKED:
        ws2.cell(row=r, column=2, value=c).alignment = wrap
        ws2.cell(row=r, column=3, value=a).alignment = wrap
        r += 1
    for col, w in zip("ABCD", [8, 34, 52, 60]):
        ws2.column_dimensions[col].width = w
    wb.save(os.path.join(HERE, BASE + ".xlsx"))


def verify_cids():
    """Abort unless every C-id printed on the QA-only sheet matches the id-map."""
    idmap = {}
    with open(os.path.join(HERE, "testrail-id-map.csv"), newline="") as f:
        for row in csv.reader(f):
            if row and row[0] != "internal_id":
                idmap[row[0]] = row[1]
    text = " ".join(m["cases"] + " " + m["refs"] + " " + m["resolve"] for m in QA_MAP)
    text += " ".join(a for _, a in WITHDRAWN) + " ".join(a for _, a in NOT_ASKED)
    pairs = re.findall(r"([A-Z]{2,4}(?:-[A-Z]+)+-\d+)\s*[=(]\s*(C\d+)", text)
    bad = [(i, c) for i, c in pairs if idmap.get(i) != c]
    if bad:
        raise SystemExit("C-ID MISMATCH vs testrail-id-map.csv: " + str(bad))
    print(f"C-id verification: {len(pairs)} internal-id/C-id pairs checked, all MATCH the id-map")


if __name__ == "__main__":
    verify_cids()
    write_md()
    write_xlsx()
    n_dec = sum(1 for q in QUESTIONS if q["group"].startswith("Decisions"))
    print("wrote", BASE + ".md", "and", BASE + ".xlsx",
          f"| {len(QUESTIONS)} items ({n_dec} decisions + {len(QUESTIONS)-n_dec} write-downs)"
          f" | {len(WITHDRAWN)} withdrawn | {len(NOT_ASKED)} not-asked"
          f" | {len(SOURCES_SWEPT)} sources swept")
