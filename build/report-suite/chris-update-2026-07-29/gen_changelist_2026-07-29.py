#!/usr/bin/env python3
"""Report Suite — Chris-update change-list (2026-07-29). Emits ChangeList-2026-07-29.md + .xlsx.

Source: chris-message-2026-07-29.md (verbatim; Chris Ward group message 8:53 AM 2026-07-29 —
NEWEST source, last-update-wins). ALL case edits this pass are LOCAL ONLY (Rule 6) except the
three separately-authorized fixes already executed (marked TESTRAIL-PUSHED). This change-list is
the approval gate for the eventual authorized push.

Mirrors the established change-list format (reconciliation-2026-07-28/gen_changelist.py →
CustomRoles_SpecRecheck_ChangeList_2026-07-20 lineage): plain rows, TestRail Case ID +
/cases/view link (Rule 8), driving-source column, classification column, plain-English
"What needs to be done" column (Rule 7). Tab 2 = the push-authorization queue.
"""
import os
BASE = os.path.dirname(os.path.abspath(__file__))
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"
DATE = "2026-07-29"

AUDIT_TALLY = ("Rule-28 three-dimension audit (all 26 touched cases): USEFUL 26/26 KEEP (0 merge / "
 "0 weak-keep / 0 cut - no slop introduced; the SBC-EXP-01 vs SBC-EXP-16 menu overlap is deliberate: "
 "menu composition vs file contents). MAKES SENSE 26/26 SENSIBLE (cold-read, 6 fail conditions; the "
 "stale 'no asset layer by design' contradiction was repaired inside this pass). GENUINE + "
 "LAYMAN-RUNNABLE 26/26 (every touched case carries ticket + spec/message refs - 11 story tickets "
 "backfilled; plain wording; all titles now 80 characters or fewer).")

PUSH = "Awaiting push authorization (update_case)."
ROWS = [
 # ---- TESTRAIL-PUSHED (separate explicit authorization 2026-07-29; already live) ----
 ("TU-DAY-01","C30418","TU","Authorized fix (import artifact)","TESTRAIL-PUSHED",
  "DONE in TestRail: the swallowed angle-bracket placeholder ('Expand 's daily breakdown') repaired "
  "to plain words. HTTP 200 + re-GET MATCH. LATER THE SAME DAY the case also got LOCAL-only edits "
  "(title trimmed 87->61, story ticket SV-8651 added to refs) - those still need the next authorized push."),
 ("PV-API-02","C30389","PV","Authorized fix (title 100 chars)","TESTRAIL-PUSHED",
  "DONE in TestRail: title trimmed to 71 chars ('Each filter or search change re-queries the server "
  "and returns page one'). HTTP 200 + re-GET MATCH. Nothing further pending."),
 ("PV-FILT-09","C30336","PV","Authorized fix (title 96 chars) + Chris message (rename)","TESTRAIL-PUSHED",
  "DONE in TestRail: title trimmed to 77 chars. HTTP 200 + re-GET MATCH. LATER THE SAME DAY the body "
  "was edited LOCALLY to the confirmed 'Special Order' label - that edit still needs the next authorized push."),
 # ---- APPLIED-LOCALLY: D1 SBC VIN identifier ----
 ("SBC-LBL-01","C30134","SBC","Chris message 2026-07-29 (VIN chain; supersedes video serial)","APPLIED-LOCALLY",
  "Asset identifier re-ruled: VIN, falling back to Unit #, then plate (was serial per the kickoff "
  "video; the message is newer and wins). Whether year/make/model text disappears entirely and the "
  "all-three-missing fallback are still to be confirmed from Chris's spec changelog + the build. " + PUSH),
 ("SBC-LBL-04","C30137","SBC","Chris message 2026-07-29 (VIN chain)","APPLIED-LOCALLY",
  "Notes-only: duplicate-label seeding context moved from serial to the VIN chain; the (#1)/(#2) "
  "rule itself is unchanged. " + PUSH),
 # ---- APPLIED-LOCALLY: D2 SBC Summary/Expanded exports ----
 ("SBC-EXP-01","C30159","SBC","Chris message 2026-07-29 (four menu items; Print stays removed)","APPLIED-LOCALLY",
  "Menu now the four exact items: Download Summary (PDF), Download Expanded View (PDF), Download "
  "Summary (CSV), Download Expanded View (CSV) - and still NO Print. Labels are now spec-stated by "
  "the message (hedge removed); still to be sighted live at VIU. " + PUSH),
 ("SBC-EXP-16","C38856","SBC","Chris message 2026-07-29 (Summary/Expanded both formats)","APPLIED-LOCALLY",
  "Reshaped to the confirmed split: Summary = one row per customer; Expanded View = the full "
  "Customer then Asset then Invoice breakdown; both PDF and CSV; four exact menu items. " + PUSH),
 ("SBC-EXP-03","C30161","SBC","Chris message 2026-07-29 (Expanded breakdown + Locations line)","APPLIED-LOCALLY",
  "Scoped to the Expanded View CSV; the old 'no asset layer by design' note removed (the Expanded "
  "View now includes the asset level - exact representation pending the spec changelog); Locations: "
  "line expectation added. " + PUSH),
 ("SBC-EXP-11","C30169","SBC","Chris message 2026-07-29 (Expanded breakdown)","APPLIED-LOCALLY",
  "Scoped to the Expanded View PDF body with the Customer/Asset/Invoice breakdown; formatting rules "
  "kept. " + PUSH),
 ("SBC-EXP-09","C30167","SBC","Chris message 2026-07-29 (Locations line in every export)","APPLIED-LOCALLY",
  "REVERSED: the old 'location is NOT shown in the header' expectation replaced by a 'Locations:' "
  "line in the PDF header naming the scoped location(s). " + PUSH),
 # ---- APPLIED-LOCALLY: D3 Locations line + on-screen scope indicator (all 6 reports) ----
 ("SBR-EXP-02","C30277","SBR","Chris message 2026-07-29 (Locations line, all reports)","APPLIED-LOCALLY",
  "Added: every one of the four downloads carries a 'Locations:' line naming the scoped "
  "location(s). Long title also trimmed. " + PUSH),
 ("PV-EXP-02","C30376","PV","Chris message 2026-07-29 (Locations line, all reports)","APPLIED-LOCALLY",
  "Added: each export (PDF and CSV) carries the 'Locations:' line. " + PUSH),
 ("TU-EXP-04","C30437","TU","Chris message 2026-07-29 (Locations line, all reports)","APPLIED-LOCALLY",
  "Added: every download carries the 'Locations:' line. Long title also trimmed. " + PUSH),
 ("IV-EXP-02","C30588","IV","Chris message 2026-07-29 (Locations line, all reports)","APPLIED-LOCALLY",
  "Added: each download carries the 'Locations:' line. Long title also trimmed. " + PUSH),
 ("WIP-EXP-02","C30511","WIP","Chris message 2026-07-29 (Locations line, all reports)","APPLIED-LOCALLY",
  "Added: each download carries the 'Locations:' line. Long title also trimmed. " + PUSH),
 ("SBC-LOC-03","C30111","SBC","Chris message 2026-07-29 (on-screen scope indicator)","APPLIED-LOCALLY",
  "Added: the page shows which location(s) the report is scoped to (exact placement confirmed in "
  "the build). Distinct from the per-row location label, which stays. " + PUSH),
 ("SBR-LOC-03","C30215","SBR","Chris message 2026-07-29 (on-screen scope indicator)","APPLIED-LOCALLY",
  "Same on-screen scope-indicator expectation added. " + PUSH),
 ("PV-FILT-10","C30337","PV","Chris message 2026-07-29 (on-screen scope indicator)","APPLIED-LOCALLY",
  "Same on-screen scope-indicator expectation added. " + PUSH),
 ("TU-LOC-02","C30443","TU","Chris message 2026-07-29 (on-screen scope indicator)","APPLIED-LOCALLY",
  "Same on-screen scope-indicator expectation added. Long title also trimmed. " + PUSH),
 ("IV-LOC-02","C30575","IV","Chris message 2026-07-29 (on-screen scope indicator)","APPLIED-LOCALLY",
  "Same on-screen scope-indicator expectation added. " + PUSH),
 ("WIP-FLT-06","C30503","WIP","Chris message 2026-07-29 (on-screen scope indicator)","APPLIED-LOCALLY",
  "Same on-screen scope-indicator expectation added. Long title also trimmed. " + PUSH),
 # ---- APPLIED-LOCALLY: D4 Special Order rename ----
 ("PV-FILT-01","C30328","PV","Chris message 2026-07-29 ('Catalogue' -> 'Special Order' CONFIRMED)","APPLIED-LOCALLY",
  "Type filter options now read exactly: Both, Inventory, Special Order; the rename hedge removed "
  "(label/display-only, no data change). " + PUSH),
 ("PV-FILT-09","C30336","PV","Chris message 2026-07-29 ('Special Order' label)","APPLIED-LOCALLY",
  "Body wording moved to the exact 'Special Order' label (title already pushed in the authorized "
  "fix, label-neutral). " + PUSH),
 ("PV-ROW-05","C30345","PV","Chris message 2026-07-29 ('Special Order' label)","APPLIED-LOCALLY",
  "Type column values now read exactly 'Inventory' or 'Special Order'; hedge removed. " + PUSH),
 ("PV-EXP-08","C30382","PV","Chris message 2026-07-29 ('Special Order' in the export)","APPLIED-LOCALLY",
  "Notes-only: exported Type values read 'Special Order' (rename covers the export); alignment rule "
  "unaffected. " + PUSH),
 # ---- APPLIED-LOCALLY: D5 TU column selector (NEW) + D6 PV logo ----
 ("TU-COL-01","(new, no C-ID yet)","TU","Chris message 2026-07-29 ('Column selector added for visual/natural conformance')","NEW-CASE",
  "ONE new case authored (section TU - Visual & Accessibility; refs SV-8655 TU Story 8 + the "
  "message). Reverses the video-era no-selector state. Column list/defaults unpinned until the spec "
  "changelog lands. Needs an authorized add_case."),
 ("PV-EXP-05","C30379","PV","Chris message 2026-07-29 (same logo treatment, all reports)","APPLIED-LOCALLY",
  "PV was the ONLY report with no logo coverage (the PV spec never mentions a logo) - the same-logo-"
  "treatment expectation added here; the other five reports already carry logo cases (SBC-EXP-10 "
  "C30168, SBR-EXP-06 C30281, TU-EXP-06 C30439, WIP-EXP-08 C30517, IV-EXP-04 C30590 - no edits "
  "needed). Long title also trimmed. " + PUSH),
 # ---- QUESTION / FYI ----
 ("WIP-COL-05 (+WIP-FLT-03, WIP-SORT-03)","C30470 (+C30500, C30485)","WIP",
  "Chris message 2026-07-29 scopes VIN to SBC only","QUESTION-PENDING-CHRIS",
  "NOT edited - WIP stays on the kickoff video's serial-number identifier. QUESTION for Chris "
  "(plain): 'On the Work In Progress report, how should each asset be identified? The kickoff plan "
  "was the serial number. Your update says Sales By Customer assets are identified by VIN (then "
  "Unit #, then plate) - but it only names Sales By Customer. A) WIP also switches to VIN (then "
  "Unit #, then plate). B) WIP keeps the serial number.' Confirm from the spec changelog when it lands."),
 ("(no case - Parts Sales report)","-","-","Chris message 2026-07-29 (rename on Parts Sales dropdown)","FYI",
  "The same Catalogue -> Special Order rename applies to the matching dropdown on the Parts Sales "
  "report - that is a DIFFERENT feature, OUT of this suite's scope. No case authored; noting it so "
  "the owning suite can pick it up."),
 ("(no case - SBR)","-","SBR","Chris message 2026-07-29 (padding flag)","FYI",
  "Chris renamed things in his local and is adding a spec flag for the SBR padding issue - PURELY "
  "VISUAL for his companion video, his words. No case change."),
 ("(no case - all reports)","-","-","Chris message 2026-07-29 second part (filters cross-squad)","FYI",
  "Branko + Milos are building an app-wide Filters project that WILL cross over with the report "
  "filters. Chris's instruction: build to spec for now, but EXPECT the filter portion to change "
  "once something workable is on staging (Branko/Milos will sweep the report filters; response "
  "awaited). Our filter cases stay as authored; re-reconcile when that sweep lands."),
 ("(watch - SBC exports)","-","SBC","Summary/Expanded split side-effect","FYI",
  "The SBC export FILENAME map (SBC-EXP-02 C30160) and the 10,000-row cap wording may shift once "
  "the four-way menu is spec'd - not edited now; watch the spec changelog."),
 ("(watch - all specs)","-","-","Chris message caveat","FYI",
  "The message's change summary was written by Chris's ASSISTANT, 'pending a human-eye-pass'. When "
  "the real spec changelog lands (~2026-07-30), re-run the spec capture + diff and verify every "
  "item above against it before the push. SPEC-WATCH deadline 2026-08-04 stands."),
]

HDR = ["Internal ID","TestRail Case ID","TestRail link","Report","Driving source","Status","What needs to be done (plain)"]

def link(cid):
    return LINK.format(cid[1:]) if cid.startswith("C") and cid[1:].isdigit() else ""

def md():
    L=[]
    L.append(f"# Report Suite — Chris-update change-list ({DATE})\n")
    L.append("**Source:** `chris-message-2026-07-29.md` (verbatim; Chris Ward group message 8:53 AM "
             "2026-07-29 — NEWEST source, last-update-wins over the kickoff video and the current six "
             "specs). **All Part-2 edits are LOCAL ONLY** — backups in `backup/` (+ MANIFEST.md); this "
             "change-list is the approval gate for the eventual authorized TestRail push "
             "(**24 update_case + 1 add_case**). The 3 TESTRAIL-PUSHED rows were a separate explicit "
             "authorization (2026-07-29), already executed + re-GET verified.\n")
    L.append(f"**Suite count:** 459 active in TestRail + 1 new authored locally (TU-COL-01) = **460 active authored**.\n")
    L.append(f"**{AUDIT_TALLY}**\n")
    L.append("| " + " | ".join(HDR) + " |")
    L.append("|" + "---|"*len(HDR))
    for iid,cid,rep,src,cls,todo in ROWS:
        lk = link(cid)
        lkmd = f"[{cid}]({lk})" if lk else cid
        L.append(f"| {iid} | {lkmd} | {lk or '—'} | {rep} | {src} | {cls} | {todo} |")
    L.append("")
    L.append("## Push queue (awaiting authorization — NOTHING pushed from Part 2)\n")
    L.append("- **24 × update_case:** SBC-LBL-01 C30134, SBC-LBL-04 C30137, SBC-EXP-01 C30159, "
             "SBC-EXP-16 C38856, SBC-EXP-03 C30161, SBC-EXP-11 C30169, SBC-EXP-09 C30167, SBR-EXP-02 "
             "C30277, PV-EXP-02 C30376, TU-EXP-04 C30437, IV-EXP-02 C30588, WIP-EXP-02 C30511, "
             "SBC-LOC-03 C30111, SBR-LOC-03 C30215, PV-FILT-10 C30337, TU-LOC-02 C30443, IV-LOC-02 "
             "C30575, WIP-FLT-06 C30503, PV-FILT-01 C30328, PV-FILT-09 C30336, PV-ROW-05 C30345, "
             "PV-EXP-08 C30382, PV-EXP-05 C30379, TU-DAY-01 C30418 (local title trim + refs ticket).")
    L.append("- **1 × add_case:** TU-COL-01 (section TU — Visual & Accessibility; "
             "custom_atmstatus:3 + custom_automation_type:0).")
    L.append("- Refs-length caution at push time: some combined refs exceed the TestRail refs cap — "
             "condense at push (full text stays in local spec_ref), same convention as SBC-EXP-01/"
             "SBR-LOC-03 on 2026-07-28.")
    return "\n".join(L)+"\n"

with open(os.path.join(BASE, "ChangeList-2026-07-29.md"), "w") as fh:
    fh.write(md())
print("wrote ChangeList-2026-07-29.md")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Change list"
    ws.append([f"Report Suite - Chris-update change list ({DATE})"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append(["Source: Chris Ward group message 8:53 AM 2026-07-29 (verbatim in chris-message-2026-07-29.md) - "
               "NEWEST source, last-update-wins. Part-2 edits LOCAL ONLY; this sheet is the approval gate for the "
               "push (24 update_case + 1 add_case). The 3 TESTRAIL-PUSHED rows were separately authorized and are done."])
    ws.append([AUDIT_TALLY])
    ws.append(["Suite count: 459 active in TestRail + 1 new authored locally (TU-COL-01) = 460 active authored."])
    ws.append([])
    ws.append(HDR)
    for c in ws[6]: c.font = Font(bold=True)
    for iid,cid,rep,src,cls,todo in ROWS:
        ws.append([iid, cid, link(cid), rep, src, cls, todo])
    for i,w in enumerate([26,20,52,7,52,24,110],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=6):
        for c in row: c.alignment = Alignment(wrap_text=True, vertical="top")
    ws2 = wb.create_sheet("Push queue")
    ws2.append(["Awaiting authorization - 24 update_case + 1 add_case (TU-COL-01). Nothing from Part 2 pushed."])
    ws2["A1"].font = Font(bold=True)
    ws2.append([])
    ws2.append(["Op","Internal ID","TestRail Case ID","TestRail link"])
    for c in ws2[3]: c.font = Font(bold=True)
    for iid,cid,rep,src,cls,todo in ROWS:
        if cls in ("APPLIED-LOCALLY","NEW-CASE"):
            ws2.append(["add_case" if cls=="NEW-CASE" else "update_case", iid.split(" ")[0], cid, link(cid)])
    ws2.append(["update_case","TU-DAY-01","C30418",LINK.format("30418")])
    for i,w in enumerate([12,20,20,52],1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    wb.save(os.path.join(BASE, "ChangeList-2026-07-29.xlsx"))
    print("wrote ChangeList-2026-07-29.xlsx")
except ImportError:
    print("openpyxl missing - xlsx skipped")
