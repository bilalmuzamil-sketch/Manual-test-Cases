#!/usr/bin/env python3
"""Generate PO-Questions-Chris-ReportSuite-2026-07-31.md + .xlsx.

FORMAT: pure 1:1 with PO-Questions-Chris-ReportSuite-TechPlan_2026-07-30.xlsx and
PO-Questions-Chris-ReportSuite-2026-07-27.xlsx (Standing Rule 16) —
  sheet 1 "Questions for PO": title row, blank row, header row
    [#, Topic, What happens now, The question, Options, Your answer],
    widths 4/24/48/42/46/20, header fill 1F4E79 white bold, wrap on, freeze A4
  sheet 2 "QA Internal Mapping": title row, blank row, header row
    [Q#, Affected internal case IDs (TestRail C-id), Source refs,
     What each answer resolves to], widths 4/34/52/60
Reader-facing content is plain layman English (Rule 7): no case IDs, no C-numbers, no
spec anchors, no HTTP/API terms. Those live on the QA-only sheet.
"""
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
BASE = "PO-Questions-Chris-ReportSuite-2026-07-31"
TITLE = ("Report Suite - Questions for Chris Ward (from the coverage re-check of the "
         "current specs) - 2026-07-31")

QUESTIONS = [
    {
        "topic": "The Sales By Representative download columns contradict each other",
        "now": ("Your Sales By Representative description was updated on 29 July. One part of it "
                "now says that when the location column is showing on screen, it is also included "
                "in all four downloads. But an older part of the same document still lists the "
                "download columns as a fixed set, in order, with no location column in the list - "
                "those older lines were never updated. So the same document says two different "
                "things about the same download."),
        "q": ("Which one is right - should the downloads include the location column whenever it "
              "is showing on screen, or should they always show the same fixed set of columns?"),
        "opts": [
            "A) The downloads should include the location column whenever it appears on screen "
            "(this is the newer instruction, and we have already built our checks to follow it - "
            "so if A is right we just need your confirmation, plus the older lines tidied up).",
            "B) The downloads should always show the same fixed set of columns, whatever is on "
            "screen (in which case we will change our checks back).",
        ],
    },
    {
        "topic": "Have the six descriptions been updated to match your video and your answers yet?",
        "now": ("You updated all six report descriptions on 29 July - thank you, the changes we "
                "were waiting for mostly landed. A handful did not, and they are ones where the "
                "written description now says the OPPOSITE of an answer you gave us afterwards. "
                "The clearest is the Work In Progress report: on 29 July you told us assets should "
                "be identified by VIN first (then unit number, then plate) for every report, and "
                "you believed you had already made that edit - but the Work In Progress "
                "description still puts the unit number first in several places. The others are: "
                "the location dropdown being hidden for a one-location person, the full word "
                "\"Representative\" on the customer card, the new reports sitting below the "
                "existing links in the menu, and the note that Parts Velocity is the \"only\" "
                "report in the Parts group when Inventory Value is there too. We are testing to "
                "your ANSWERS, not to the older written text."),
        "q": ("Will the descriptions be updated to match your answers, or should we simply keep "
              "testing to your answers and treat the written text as out of date?"),
        "opts": [
            "A) The descriptions will be updated - we keep testing to your answers meanwhile.",
            "B) Do not wait for the descriptions - your answers are the final word and the "
            "written text can stay as it is.",
        ],
    },
    {
        "topic": "Where the location column goes in the shorter \"Summary\" downloads",
        "now": ("On screen the location column has a clear home: on Sales By Customer it sits "
                "right after the date, and on Sales By Representative right after the status. "
                "Your instruction for the downloads is that it appears \"in the same position it "
                "occupies on screen\". That works for the detailed downloads, which have those "
                "same columns. But the shorter Summary downloads do NOT have a date or a status "
                "column at all - so there is no position for it to match. Nothing in the "
                "descriptions says where it should go in those two files, and we do not want to "
                "guess."),
        "q": ("In the shorter Summary downloads, where should the location column sit?"),
        "opts": [
            "A) With the naming columns at the left - straight after the customer name (Sales By "
            "Customer) or the representative name (Sales By Representative), before the money "
            "columns.",
            "B) At the far right, after all the money columns.",
            "C) You do not mind - we will confirm whatever the build does and write that down.",
        ],
    },
    {
        "topic": "\"The same logo treatment\" - the three descriptions describe three different rules",
        "now": ("In your 29 July note you said every report now uses the same logo treatment. The "
                "written descriptions do not agree with each other on what that treatment is. "
                "Technician Utilization says the built-in ShopView logo is always used. Sales By "
                "Customer says it tries the company's own uploaded logo first, then falls back to "
                "the built-in one, and if neither exists it prints no logo at all and lets the "
                "text fill the space. Parts Velocity does not mention a logo anywhere. Our checks "
                "for the three reports currently follow their own descriptions, so they cannot all "
                "be right."),
        "q": ("Which single rule should every report's printed download follow?"),
        "opts": [
            "A) Try the company's own uploaded logo first, then the built-in ShopView logo, and "
            "print no logo only if neither exists (the Sales By Customer rule).",
            "B) Always print the built-in ShopView logo (the Technician Utilization rule).",
            "C) Something else - please describe it.",
        ],
    },
    {
        "topic": "\"Normal reports access\" - one single permission, or the existing per-area ones?",
        "now": ("You told us twice that these reports should not be hidden behind their own "
                "special permissions - they should use normal reports access. We have changed our "
                "checks to that. What is still not settled is how far that goes. Two of the six "
                "reports (Parts Velocity and Inventory Value) are described as needing the "
                "EXISTING inventory-reports permission - which is a normal, already-existing "
                "reports permission, not a new special one. So your instruction can be read two "
                "ways."),
        "q": ("Should all six reports open with one single reports permission, or should the "
              "existing per-area reports permissions (like the inventory one) still apply?"),
        "opts": [
            "A) One single reports permission for all six - if someone can see reports, they can "
            "see all six of these.",
            "B) Keep the existing per-area reports permissions - the two parts reports stay behind "
            "the inventory-reports permission, the others behind their own existing areas.",
        ],
    },
]

QA_MAP = [
    {
        "q": 1,
        "cases": ("SBR-EXP-10 (C30285); SBR-EXP-11 (C30286); SBR-EXP-03 (C30278); "
                  "SBR-EXP-04 (C30279); SBR-LOC-05 (C38913)"),
        "refs": ("SBR spec v15 2026-07-29 self-inconsistency: NEW S14-R20 (\"included in all four "
                 "exports in the same position it occupies on screen\") vs S14-R15 (Summary CSV "
                 "headers \"in order, are exactly\") and S14-R16 (Expanded CSV headers) — the "
                 "header enumerations date from the 2026-07-11 \"Exports hardened\" round and were "
                 "never amended. Independently corroborated in "
                 "contradiction-analysis-2026-07-31/SBR-CSV-LOCATION.md. Found by "
                 "coverage-rederivation-2026-07-31 (S14-R20 was an uncovered requirement) and by "
                 "the Rule-28 Stage-2b sweep (the header lists contradicted SBR-LOC-05)."),
        "resolve": ("A -> cases stand as pushed 2026-07-31 (the 5 cases above are already "
                    "scope-conditional per S14-R20, user-authorized) and Chris tidies S14-R15/R16. "
                    "B -> revert the export halves to the fixed lists and drop the Location "
                    "assertions from the 4 export cases; SBR-LOC-05's export expected would be "
                    "removed. Either way VIU-confirm live. Same pattern was applied to SBC "
                    "(S4-R13), PV (S6-R11), TU (S7-R13), IV (S10-R15) — WIP already had it."),
    },
    {
        "q": 2,
        "cases": ("WIP-COL-05 (C30470); WIP-FLT-03 (C30500); WIP-SORT-03 (C30485); "
                  "WIP-EXP-07 (C30516) [VIN chain]; plus SPEC-WATCH items 4/6/8/9/10/11 "
                  "touching SBR-LOC-04 (C30216), TU-LOC-05 (C30446), IV-LOC-04 (C30577), "
                  "PV-FILT-13 (C30340), SBR-WO-06 (C30315), TU-NAV-01 (C30392), "
                  "PV-NAV-01 (C30323)"),
        "refs": ("SPEC-WATCH-2026-07-28.md re-diff 2026-07-31: 7 of 12 items still need spec text "
                 "after the 2026-07-29 changelog (1b WIP identifier, 4 location filter hidden, 6 "
                 "nav placement wording, 8 WIP asset dropdown, 9 customer-card Representative, 10 "
                 "SBC nav anchors, 11 PV \"only report\"). Deadline 2026-08-04 partly met. Chris's "
                 "2026-07-29 answer (\"A is the correct answer\") and his 2026-07-31 Q1/Q5 answers "
                 "are the NEWER sources and win (Rule 32)."),
        "resolve": ("A -> no case change; SPEC-WATCH stays open until the text lands; re-diff on "
                    "the next touch. B -> SPEC-WATCH can be CLOSED as a documentation debt and the "
                    "cases are already correct. Neither answer changes a case."),
    },
    {
        "q": 3,
        "cases": ("SBC-EXP-16 (C38856); SBC-LOC-04 (C38912); SBR-EXP-10 (C30285); "
                  "SBR-EXP-03 (C30278); SBR-LOC-05 (C38913)"),
        "refs": ("SPEC-SILENT, found by coverage-rederivation-2026-07-31. SBC S4-R13 says every "
                 "export \"also includes that Location column\" with no position stated; SBR "
                 "S14-R20 says \"in the same position it occupies on screen\" but the Summary CSV "
                 "(S14-R15) has no Date/Status column and the Summary PDF (S14-R5) has none "
                 "either. The Expanded files DO have Status/Date so their position is determined. "
                 "Cases currently say \"with the identifying columns ahead of the money columns "
                 "(confirm its exact position in the build)\" — hedged, not invented (Rule 9)."),
        "resolve": ("A -> replace the hedge with the stated position in the 5 cases. B -> reword to "
                    "far-right. C -> leave the hedge and pin it at VIU. No case is wrong today "
                    "either way."),
    },
    {
        "q": 4,
        "cases": ("SBC-EXP-10 (C30168); TU-EXP-06 (C30439); TU-EXP-07 (C30440); "
                  "PV-EXP-05 (C30379); PV-EXP-06 (C30380); SBR-EXP-03 (C30278); "
                  "SBR-EXP-04 (C30279)"),
        "refs": ("Cross-report spec contradiction against Chris's own 2026-07-29 group message "
                 "(\"Each report now ensures the same 'logo' treatment\"): SBC S15-R16/R17/R18 = "
                 "3-step chain ending in NO logo; TU spec = bundled ShopView default (the "
                 "2026-07-29 changelog added \"bundled-default logo\"); PV has no logo requirement "
                 "at all. Raised by our own sweep 2026-07-31; already listed in "
                 "OUTSTANDING-ITEMS-REGISTER.md. NOT resolved by us (Rule 15 — never pick a side "
                 "silently)."),
        "resolve": ("A -> TU/PV export cases gain the fallback chain (SBC-EXP-10 already asserts "
                    "it). B -> SBC-EXP-10's chain collapses to the bundled default and its no-logo "
                    "step is dropped. C -> re-ask. All are wording-only edits to the export-header "
                    "cases."),
    },
    {
        "q": 5,
        "cases": ("PV-PERM-01 (C30325); PV-PERM-02 (C30326); PV-PERM-03 (C30327); "
                  "IV-PERM-01 (C30603); IV-PERM-02 (C30604); PV-API-04 (C30388); "
                  "SBC-PERM-01 (C30098); SBC-PERM-02 (C30099); SBC-NAV-01 (C30096)"),
        "refs": ("Follow-on granularity question from his 2026-07-31 Q4 = A (\"the intention is to "
                 "not hide these from normal reports access\"). His ruling removes the DEDICATED "
                 "per-report atom (ROLE_SALES_BY_CUSTOMER_REPORT::VIEW) but PV S1-R4/S1-N2 and IV "
                 "S1-R4 name the pre-existing \"Inventory Reports -> View\" permission, which IS a "
                 "normal reports permission. Both readings are consistent with his words, so the "
                 "question is genuinely open (Rule 32 — do not infer). Dev note already drafted: "
                 "chris-answers-2026-07-31/Q4-permission-dev-note-2026-07-31.md."),
        "resolve": ("A -> the PV/IV permission cases lose the Inventory-Reports qualifier and use "
                    "the single reports permission. B -> PV/IV cases stand as authored and only the "
                    "SBC dedicated-atom cases change (already done per his Q4). VIU-confirm the "
                    "exact permission names live at the QA branch."),
    },
]

WITHDRAWN = [
    ("Does the short header \"Rep is active?\" also become \"Representative\"?",
     "ANSWERED by his 2026-07-31 Q5 = A. Verbatim: \"slang, let's do representative everywhere\" "
     "— our ingest records the scope as reaching \"the 'Sales Rep Assignments' export (dialog "
     "entry, file name, and the CSV column header)\". The header is a CSV column header, so it is "
     "covered. Affects SBR-ASGN-02 (C30293)."),
    ("What is the exact renamed \"Sales Rep Assignments\" file name?",
     "ANSWERED by the same Q5 = A — the file name is explicitly in scope of \"representative "
     "everywhere\". SBR-ASGN-02 (C30293) already hedges the exact final string for live "
     "confirmation, which is the correct treatment; no product question remains."),
    ("Which of the two \"too large to export\" messages is correct?",
     "ANSWERED 2026-07-31 Q2 = A, verbatim \"A - great catch\": one suite-wide string \"This "
     "report is too large to export. Narrow the date range or filters, then try again.\" All six "
     "cap cases already quote it. What remains is only his SPEC EDIT (folded into Q2 of this "
     "sheet), not a question."),
    ("Does the 10,000-row export cap apply to Parts Velocity, Technician Utilization and Work In "
     "Progress?",
     "ANSWERED 2026-07-31 Q3 = A, verbatim \"A - this was not well thought out by me (the specs "
     "were written at different times)\". Cap is suite-wide; PV-EXP-11 (C38885), TU-EXP-09 "
     "(C38887), WIP-EXP-10 (C38918) exist and are pushed."),
    ("Is the location dropdown hidden for a one-location user?",
     "ANSWERED 2026-07-31 Q1 = A, verbatim \"A -- classic spec drift\". Hidden. Only the spec text "
     "still disagrees, which is Q2 of this sheet."),
    ("Should the reports use their own dedicated permissions?",
     "ANSWERED 2026-07-31 Q4 = A — no dedicated per-report permission. Only the GRANULARITY "
     "follow-on survives, and it is asked as Q5 of this sheet rather than re-asking the settled "
     "part."),
]

NOT_ASKED = [
    ("The Technician Utilization column-selector story has no Jira ticket.",
     "Ticket-management, not a product decision — he already asked for the control in his "
     "2026-07-29 message (\"for visual/natural conformance\"), so scope is settled. TU-COL-01 "
     "(C38859) and TU-LOC-06 (C38915) cite epic SV-8582 and say so in refs. Tracked in "
     "OUTSTANDING-ITEMS-REGISTER.md as an OTHER TEAM item."),
    ("The Sales By Representative Escape-key question (deactivate dialog vs Golden Rule #9).",
     "Deliberately NOT re-asked — it is Q1 of PO-Questions-Chris-ReportSuite-2026-07-27, still "
     "awaiting his answer (open 4 days). Same decision the 2026-07-30 sheet took. Affects "
     "SBR-DEACT-04 (C30255)."),
    ("Four requirements we deliberately do not test (SBC S10-N1, SBR S11-N1, SBR S14-R14, "
     "PV S4-N1).",
     "QA decisions, not product ones — cut by the user-authorized 2026-07-28 Ruthless Usefulness "
     "Audit as no-op assertions, un-measurable px font-tier minutiae, and a stored-schema state a "
     "manual tester cannot seed. Recorded with reasons in "
     "coverage-rederivation-2026-07-31/COVERAGE-REDERIVATION.md §5. Raise with him only if he asks "
     "for 100% requirement-to-case parity."),
    ("The \"this spec is the source of truth\" sentences (PV S7-R7, SBR S18-R7.6) and the mojibake "
     "in the SBR/PV spec text.",
     "Not product questions — a statement about the document, and a cosmetic encoding artifact. "
     "The mojibake is already noted for him in SPEC-WATCH-2026-07-28.md."),
]


def write_md():
    L = [f"# {TITLE.replace(' - 2026-07-31', '') } — 2026-07-31", "",
         "**STATUS: READY TO SEND** (not yet sent). On return: ingest verbatim, then revisit the "
         "affected cases per the standing workflow.", "",
         "Plain-language product questions only — no bugs, no test jargon. These came out of a "
         "**coverage re-check of the six current descriptions** (the 29 July versions), read "
         "against your answers of 31 July, your 29 July note and your videos. Everything you have "
         "already answered has been deliberately left out.",
         "Please pick an option (or write your own answer) for each.", ""]
    for i, q in enumerate(QUESTIONS, 1):
        L += [f"## Question {i} — {q['topic']}", "",
              f"**What happens now:** {q['now']}", "",
              f"**The question:** {q['q']}", "", "**Options:**", ""]
        L += [f"- {o}" for o in q["opts"]]
        L += ["", "**Your answer:** ____________________", ""]

    L += ["---", "", "## QA Internal Mapping (QA-only — not for the PO)", "",
          "TestRail C-ids from `build/report-suite/testrail-id-map.csv` (Standing Rule 8). "
          "Links: https://shopview.testrail.io/index.php?/cases/view/<id>", "",
          "| Q# | Affected internal case IDs (TestRail C-id) | Source refs | "
          "What each answer resolves to |", "|---|---|---|---|"]
    for m in QA_MAP:
        L.append(f"| {m['q']} | {m['cases']} | {m['refs']} | {m['resolve']} |")

    L += ["", "### Withdrawn — already answered (QA-only appendix)", "",
          "Each of these was a candidate question; the source that answers it is quoted. "
          "**Not put in front of Chris.**", "",
          "| Candidate question | Already answered by |", "|---|---|"]
    for c, a in WITHDRAWN:
        L.append(f"| {c} | {a} |")

    L += ["", "### Not asked here (QA reference)", "", "| Item | Why it is not on the sheet |",
          "|---|---|"]
    for c, a in NOT_ASKED:
        L.append(f"| {c} | {a} |")

    L += ["", "---", "", "## OUTSTANDING — what I need from you", "",
          "Cross-project register: `build/OUTSTANDING-ITEMS-REGISTER.md` (Standing Rule 36).", "",
          "**From you (QA lead):**", "",
          "1. **Send this sheet to Chris.** Five questions. Two of them (Q1 and Q4) are places "
          "where his own descriptions contradict each other or contradict his own note, so a reader "
          "cannot tell which is right — and one of them (Q3) is a detail nobody has written down "
          "anywhere.",
          "2. **The QA branch / environment, and confirmation the reports are switched on.** This "
          "is still the single biggest gap on the project: **all 474 cases have never been run "
          "against the real build.** Everything we assert about wording and layout is \"the "
          "description says so\", not \"the build shows it\". We also need fresh login cookies when "
          "the branch exists.",
          "3. **A nudge to Chris on the seven description corrections he still owes** — headed by "
          "the Work In Progress asset identifier, which he believed he had already fixed. Our cases "
          "follow his answer, so until he edits the text anyone comparing us to the written "
          "description will read a mismatch as our mistake. The deadline we agreed was **4 August**.",
          "4. **Go-ahead for the title-trim pass** — 288 of the 474 case titles are still too long "
          "and get cut off on the TestRail case page. Cosmetic, but it costs the tester every day.",
          "",
          "**From Chris:** answers to the five questions here, plus Q1 of the 27 July sheet (the "
          "Escape-key question, now open four days).", "",
          "**Nothing else is outstanding on this pass** — coverage against the current descriptions "
          "is complete (888 of 895 requirements covered, no open gaps) and the suite has no "
          "internal contradictions left.", ""]
    open(os.path.join(HERE, BASE + ".md"), "w", encoding="utf-8").write("\n".join(L) + "\n")


def write_xlsx():
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True)
    cols = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    for i, q in enumerate(QUESTIONS, 1):
        r = 3 + i
        for j, v in enumerate([i, q["topic"], q["now"], q["q"],
                               "\n".join(q["opts"]), ""], 1):
            cell = ws.cell(row=r, column=j, value=v)
            cell.alignment = wrap
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids from "
                 "build/report-suite/testrail-id-map.csv (Standing Rule 8); links "
                 "https://shopview.testrail.io/index.php?/cases/view/<id>")
    ws2["A1"].font = Font(bold=True)
    cols2 = ["Q#", "Affected internal case IDs (TestRail C-id)", "Source refs",
             "What each answer resolves to"]
    for j, c in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r = 4
    for m in QA_MAP:
        for j, v in enumerate([m["q"], m["cases"], m["refs"], m["resolve"]], 1):
            ws2.cell(row=r, column=j, value=v).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="WITHDRAWN - already answered (not asked)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Candidate question", "Already answered by", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for c, a in WITHDRAWN:
        ws2.cell(row=r, column=2, value=c).alignment = wrap
        ws2.cell(row=r, column=3, value=a).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="NOT ASKED HERE (QA reference)").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Item", "Why it is not on the sheet", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for c, a in NOT_ASKED:
        ws2.cell(row=r, column=2, value=c).alignment = wrap
        ws2.cell(row=r, column=3, value=a).alignment = wrap
        r += 1
    for col, w in zip("ABCD", [4, 34, 52, 60]):
        ws2.column_dimensions[col].width = w
    wb.save(os.path.join(HERE, BASE + ".xlsx"))


if __name__ == "__main__":
    write_md()
    write_xlsx()
    print("wrote", BASE + ".md", "and", BASE + ".xlsx",
          f"| {len(QUESTIONS)} questions | {len(WITHDRAWN)} withdrawn | {len(NOT_ASKED)} not-asked")
