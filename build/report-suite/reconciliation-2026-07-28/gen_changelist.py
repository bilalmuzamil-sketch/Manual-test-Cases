#!/usr/bin/env python3
"""Report Suite — spec-relevance reconciliation change-list (2026-07-28).

Phase 2 deliverable, UPDATED after the 2026-07-28 video-promotion user ruling:
Chris Ward's kickoff video is AUTHENTIC + AUTHORITATIVE product intent (made
for Chris Amani, company VP) and NEWER than the six Confluence specs (specs
last updated 2026-07-21) — by last-update-wins the VIDEO overrides the spec
where they conflict. The previously PENDING-CHRIS rows are therefore PROMOTED:
applied as LOCAL case edits (audit log
video-promotion-edit-log-2026-07-28.md, applier
apply_video_promotion_2026-07-28.py). NO TestRail writes (Rule 6) — this
change-list remains the approval gate for the eventual authorized push.

Mirrors the established change-list workbook format
(build/custom-roles-run/CustomRoles_SpecRecheck_ChangeList_2026-07-20.*):
plain rows, TestRail Case ID + /cases/view link (Rule 8), a driving-source
column, a classification column, and a plain-English "What needs to be done"
column (Rule 7 / deliverable convention). Tab 2 = the rows applied locally and
awaiting TestRail push authorization.
"""
import os
BASE = os.path.dirname(os.path.abspath(__file__))
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"
DATE = "2026-07-28"
TOTAL_CASES = 516   # 515 in TestRail + 1 new authored locally (SBC-EXP-16)

APPLIED = ("APPLIED-LOCALLY (video-authoritative, user ruling 2026-07-28) - "
           "awaiting TestRail push authorization")

# Each row: (internal_id, cid, report, driving_source, classification, what_to_do)
# classification in {APPLIED-NOW, APPLIED-LOCALLY, RETIRE-PROPOSED,
#                    OPEN-DECISION, LIVE-VIU-PENDING}
ROWS = [
 # ---- APPLIED-NOW (edited locally in the first pass; firmly confirmed) ----
 ("SBR-DEACT-04","C30255","SBR","PO answer Q1 = B (overrides spec S13-R8)","APPLIED-NOW",
  "DONE (first pass). Reworded so pressing the Esc key does NOT close the 'deactivate a sales rep' "
  "confirm pop-up - only the Cancel button and the X icon close it (matches Chris's answer and the "
  "app's house rule that pop-ups do not close on Esc). Long title shortened to fit. Still to be "
  "confirmed live on the QA branch."),
 ("SBR-DEACT-05","C30256","SBR","PO answer Q1 = B (consistency)","APPLIED-NOW",
  "DONE (first pass). Small consistency fix: made clear the Esc key never closes this dialog at any "
  "time (aligns with the case above). Overlong title shortened. Confirm live."),

 # ---- APPLIED-LOCALLY (video promoted to authoritative, user ruling 2026-07-28) ----
 ("SBC-LBL-01","C30134","SBC","Video P24 - serial-number identifier","APPLIED-LOCALLY",
  "DONE LOCALLY. The asset label now uses the SERIAL NUMBER as the identifier (video: 'the holy "
  "grail... is the serial number, or in some cases the bin number'), replacing the spec's Unit -> "
  "plate -> VIN chain (S8-R8 OVERRIDDEN). What stands in when the serial is missing is flagged "
  "'confirmed in the build' - not invented. SBC-LBL-02 C30135 / SBC-LBL-03 C30136 / SBC-LBL-04 "
  "C30137 got notes/refs updates only (their own rules are not overridden). Needs an authorized "
  "update_case push, then live confirmation."),
 ("WIP-COL-05","C30470","WIP","Video P24 - serial-number identifier","APPLIED-LOCALLY",
  "DONE LOCALLY. The Work In Progress Asset cell's bold first line is now the SERIAL NUMBER "
  "(spec S4-R7 unit number OVERRIDDEN); the missing-serial placeholder text is flagged 'confirmed "
  "in the build'. Same serial-number change applied to WIP-FLT-03 C30500 (asset filter options + "
  "type-ahead), WIP-SORT-03 C30485 (Asset sort key), and a tester caveat added to WIP-EXP-07 "
  "C30516 (export header 'Unit' - record what it shows, do not file). Needs an authorized push."),
 ("SBC-EXP-01","C30159","SBC","Video P25 - remove SBC Print","APPLIED-LOCALLY",
  "DONE LOCALLY. The overflow menu expectation is now 'Download (CSV)', 'Download (PDF)' with NO "
  "'Print' item (video: 'print here, this should not exist. I'm going to make sure that's cut out "
  "of the spec' - spec Story 16 + S14-R1 'Print' third item OVERRIDDEN). Needs an authorized push."),
 ("SBC-EXP-14","C30172","SBC","Video P25 - remove SBC Print","APPLIED-LOCALLY",
  "DONE LOCALLY. The Print leg was removed from the 10,000-row-cap export negative (CSV + PDF legs "
  "kept); overlong title shortened. CORRECTION to the earlier list: the third Print reference "
  "lives HERE (SBC-EXP-14 C30172), not in SBC-EXP-15 C30173 (which has no Print reference and is "
  "untouched). Needs an authorized push."),
 ("SBC-EXP-13","C30171","SBC","Video P25 - remove SBC Print","RETIRE-PROPOSED",
  "RETIRE PROPOSED - this case's ONLY purpose is the Print behavior and Print is removed from "
  "Sales By Customer. NOT deleted: it stays in TestRail (C30171) and in the import until you "
  "authorize delete_case (Rule 6). Marked Retire-Proposed locally; body left as authored for the "
  "record."),
 ("SBC-EXP-16","","SBC","Video P21 - add compressed download","APPLIED-LOCALLY",
  "DONE LOCALLY - NEW CASE AUTHORED (new - no C-ID yet). The download menu also offers a "
  "compressed (summary) version of the report alongside the expanded (nested) one (Parth's "
  "suggestion, Chris: 'That's actually a good callout. Let's, let's add that' + 'we're gonna have "
  "to add to SVC... the CompressedView'). Menu wording + file shape flagged 'confirmed in the "
  "build'. Needs an authorized add_case, then live confirmation."),
 ("SBC-LOC-03","C30111","SBC","Video P10 - per-row location label","APPLIED-LOCALLY",
  "DONE LOCALLY. Added the expectation that with 'All locations' active each row identifies WHICH "
  "location it belongs to (video: 'how do I know which is for shop A and which is for shop B?... "
  "we should probably add that in there'). The exact label/control is flagged 'confirmed in the "
  "build' - the video names none. Same add applied to SBR-LOC-03 C30215, PV-FILT-10 C30337, "
  "TU-LOC-01 C30442 (worded for TU's pooled rows), IV-LOC-01 C30574. Work In Progress already has "
  "a Location column - no WIP edit. Needs an authorized push."),
 ("SBR-LOC-04","C30216","SBR","Video P33 - hide location filter when <=1 location","APPLIED-LOCALLY",
  "DONE LOCALLY - expectation FLIPPED. A user with permission to only ONE location does NOT see "
  "the Location filter at all; with two or more permitted locations the filter shows (video: 'the "
  "filter's just gone... if you had QA testing and QB location, then of course you'd see the "
  "filter'). Overrides spec S21-N1 ('a single-location user STILL SEES the filter'). Ties to the "
  "Q2 permission question. Needs an authorized push."),
 ("TU-LOC-05","C30446","TU","Video P33 - hide location filter when <=1 location","APPLIED-LOCALLY",
  "DONE LOCALLY - same flip (overrides spec S9-N1). Needs an authorized push."),
 ("IV-LOC-04","C30577","IV","Video P33 - hide location filter when <=1 location","APPLIED-LOCALLY",
  "DONE LOCALLY - same flip (overrides spec S7-N1). Needs an authorized push."),
 ("PV-FILT-13","C30340","PV","Video P33 - hide location filter when <=1 location","APPLIED-LOCALLY",
  "DONE LOCALLY - same flip (overrides spec S2-E4). Needs an authorized push."),
 ("TU-NAV-01","C30392","TU","Video P3 - move TU down in nav (additive, not interruptive)","APPLIED-LOCALLY",
  "DONE LOCALLY. Added the expectation that the Technician Utilization entry sits BELOW the "
  "previously existing report links (video: 'technician utilization is actually in a really bad "
  "spot right now. So, we want to move these down below what's already there'). The Performance-"
  "group expectation is kept. PV-NAV-01 C30322 + IV-NAV-01 C30534 already expect the new 'Parts' "
  "nav section (video P2) - verified, no edit needed. Needs an authorized push + live placement "
  "check."),

 # ---- OPEN-DECISION items APPLIED per LATEST info (user addition 2026-07-28:
 #      update per latest info now, correct at VIU later) ----
 ("PV-FILT-01","C30328","PV","Video P31 - Catalogue rename (latest info: leaning to rename)","APPLIED-LOCALLY",
  "DONE LOCALLY per LATEST INFO (correct at VIU later). The Type filter's third option is now "
  "worded by MEANING - special-order catalog parts never put into stock - with the exact "
  "on-screen label flagged 'confirmed in the build' (video: 'maybe we do rename it... you're "
  "absolutely right' + 'we'll have to truncate that down'; both possible labels noted in the case "
  "note - 'Catalogue' per spec, or a short special-order-parts name, not invented). Same "
  "treatment on PV-FILT-09 C30336 and PV-ROW-05 C30345; PV-EXP-08 C30382 notes-only. Needs an "
  "authorized push."),
 ("WIP-FLT-03","C30500","WIP","Video P12 - asset dropdown style (latest info: match native + toggle)","APPLIED-LOCALLY",
  "DONE LOCALLY per LATEST INFO. Latest information = match the native ShopView multi-select "
  "style plus a possible toggle (Stefan proposed, Chris agreed: 'let's please do this, happy to "
  "update the spec'). Verified NO case asserts the stay-open behavior - nothing to flip; this "
  "case's note now records the native+toggle expectation with the exact interaction = confirm "
  "live. (Its serial-number wording was also updated under P24 above.)"),
 ("PV-API-01","C30388","PV","Video P30 - pagination (latest info: pagination on every page)","APPLIED-LOCALLY",
  "DONE LOCALLY per LATEST INFO. Stefan (definitive): 'we are definitely having pagination on "
  "every page'. Verified NO case asserts infinite-scroll or all-rows-at-once; the pagination "
  "cases already expect server pagination. PV-API-01 + PV-API-02 C30389 notes now flag the "
  "pagination-behavior details (page size, control style) as confirm-live; Chris's "
  "infinite-scroll revisit stays a later product question."),
 ("(no case - TU)","","TU","Video P18 - TU column selector (latest info: veto stands)","NO-CHANGE-CONFIRMED",
  "CONFIRMED NO-OP per LATEST INFO: Chris's veto stands and the spec agrees (no TU column-selector "
  "story). Verified no TU column-selector case exists - cases already match. If Chris later adds "
  "one, author cases then."),

 # ---- OPEN-DECISION (still not settled - NOT applied) ----
 ("IV-PERS-01","C30579","IV","Video P18 / P36 - column-selector scope","OPEN-DECISION",
  "Column-selector scope was left open across the suite (Chris vetoed one on Technician Utilization, "
  "'not married to it'). Inventory Value currently HAS a column selector per the spec (S8-R1). "
  "Confirm with Chris whether Inventory Value keeps its selector. Not changed."),
 ("(follow-up - SBC/WIP)","","SBC/WIP","Video P33 - hide location filter when <=1 location","OPEN-DECISION",
  "FOLLOW-UP flagged for completeness: SBC and Work In Progress have NO single-location filter "
  "case at all (their specs never had the 'still sees the filter' rule), so nothing on those two "
  "reports asserts the old behavior. Decide whether each needs a NEW hidden-filter case - not "
  "authored this pass (outside the promotion's scope)."),

 # ---- LIVE-VIU-PENDING (needs the live build; no QA branch yet) ----
 ("WIP-CALC-08","C30481","WIP","Video P14 vs spec S4-R23 - labor-delta basis","LIVE-VIU-PENDING",
  "Video P14 says the labor delta = clocked/tech hours vs INVOICED hours; the WIP spec (S4-R23) "
  "uses QUOTED (estimate) hours minus worked hours because WIP is a pre-invoice report. The case "
  "matches the SPEC (quoted-basis) and is NOT changed - the video discusses the delta generally "
  "and does not explicitly override WIP's pre-invoice basis, so this stays a live-check, not a "
  "promotion. Confirm live which basis actually ships. Do NOT conflate with Sales By Customer / "
  "Sales By Representative, which correctly use invoiced-minus-worked (SBC-CALC-03 C30151, "
  "SBR-CALC-02 C30230)."),
]

CLASS_FILL = {
 "APPLIED-NOW":"C6E0B4",        # green
 "APPLIED-LOCALLY":"A9D08E",    # darker green
 "RETIRE-PROPOSED":"F8CBAD",    # salmon
 "NO-CHANGE-CONFIRMED":"E2EFDA",# pale green
 "OPEN-DECISION":"FFF2CC",      # yellow
 "LIVE-VIU-PENDING":"DDEBF7",   # blue
}

def link(cid): return LINK.format(cid[1:]) if cid.startswith("C") else ""


def write_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
    wrap = Alignment(wrap_text=True, vertical='top')
    thin = Border(*[Side(style='thin', color='D9D9D9')]*4)
    HDR = ['Internal ID','TestRail Case ID','TestRail link','Report','Driving source',
           'Classification','What needs to be done (plain)']

    def sheet(ws, rows):
        ws.append(HDR)
        for c in ws[ws.max_row]:
            c.fill=hf; c.font=hfont; c.alignment=wrap
        for iid,cid,rep,src,cls,todo in rows:
            ws.append([iid, cid, link(cid), rep, src, cls, todo])
            r = ws[ws.max_row]
            fill = CLASS_FILL.get(cls)
            for c in r:
                c.alignment=wrap; c.border=thin
            if fill:
                r[5].fill = PatternFill('solid', fgColor=fill)
        for i,w in enumerate([16,15,44,8,42,18,80],1):
            ws.column_dimensions[get_column_letter(i)].width=w
        ws.freeze_panes='A2'

    n_appl = sum(1 for r in ROWS if r[4]=='APPLIED-LOCALLY')
    n_ret = sum(1 for r in ROWS if r[4]=='RETIRE-PROPOSED')

    # Tab 1 — full change list
    ws = wb.active; ws.title='Change list'
    ws.append([f'Report Suite - spec-relevance reconciliation change list ({DATE}, video-promotion update)'])
    ws['A1'].font=Font(bold=True,size=13)
    ws.append([f"USER RULING {DATE}: Chris Ward's kickoff video is authentic + authoritative product "
               f"intent and NEWER than the specs (last updated 2026-07-21) - the video overrides the "
               f"spec where they conflict (last-update-wins). The previously 'blocked on Chris' rows "
               f"are now APPLIED-LOCALLY: {n_appl} rows edited/authored in the local case source + "
               f"{n_ret} retire-proposed. {TOTAL_CASES} cases total (515 in TestRail + 1 new). NO "
               f"TestRail writes - this list is the approval gate for the push."])
    ws.append(["Legend: APPLIED-NOW = edited locally in the first pass (Chris's written answer). "
               "APPLIED-LOCALLY = video-authoritative / latest-info edit applied to the local case "
               "source (user ruling 2026-07-28) - awaiting TestRail push authorization. "
               "RETIRE-PROPOSED = proposed for deletion, NOT deleted - awaiting authorization. "
               "NO-CHANGE-CONFIRMED = latest info confirms the cases already match. OPEN-DECISION = "
               "not settled. LIVE-VIU-PENDING = needs the live QA build to confirm (no QA branch "
               "yet). Backups of every touched case: build/report-suite/video-promotion-backup-"
               "2026-07-28/ (recover if Chris never ratifies). Spec-watch deadline 2026-08-04: "
               "build/report-suite/SPEC-WATCH-2026-07-28.md."])
    ws.append([])
    sheet(ws, ROWS)

    # Tab 2 — applied locally, awaiting push authorization
    appl = [r for r in ROWS if r[4] in ('APPLIED-LOCALLY','RETIRE-PROPOSED')]
    ws2 = wb.create_sheet('Applied locally - to push')
    ws2.append(["Report Suite - edits applied LOCALLY (video-authoritative), awaiting TestRail push authorization"])
    ws2['A1'].font=Font(bold=True,size=13)
    ws2.append(["Every row below is already edited/authored in build/report-suite/cases/ (audit log "
                "video-promotion-edit-log-2026-07-28.md, with the driving video quote + the overridden "
                "spec wording per case). NOTHING is pushed: the push needs your explicit permission "
                "(update_case for the edits, add_case for SBC-EXP-16, delete_case for the "
                "retire-proposed row). After the push: live VIU on the QA branch."])
    ws2.append([])
    sheet(ws2, appl)

    out = os.path.join(BASE, f'Report-Suite_Spec-Reconciliation_ChangeList_{DATE}.xlsx')
    wb.save(out)
    return out


def write_md():
    out = os.path.join(BASE, f'Report-Suite_Spec-Reconciliation_ChangeList_{DATE}.md')
    n_applied = sum(1 for r in ROWS if r[4]=='APPLIED-NOW')
    n_local = sum(1 for r in ROWS if r[4]=='APPLIED-LOCALLY')
    n_ret = sum(1 for r in ROWS if r[4]=='RETIRE-PROPOSED')
    n_open = sum(1 for r in ROWS if r[4]=='OPEN-DECISION')
    n_viu = sum(1 for r in ROWS if r[4]=='LIVE-VIU-PENDING')
    n_noop = sum(1 for r in ROWS if r[4]=='NO-CHANGE-CONFIRMED')
    with open(out,'w') as fh:
        fh.write(f"# Report Suite - spec-relevance reconciliation change list ({DATE}, video-promotion update)\n\n")
        fh.write(f"**USER RULING {DATE}:** Chris Ward's kickoff video is AUTHENTIC and AUTHORITATIVE "
                 f"product intent (created for Chris Amani, company Vice President) and is NEWER than "
                 f"the six Confluence specs (specs last updated 2026-07-21) - by last-update-wins the "
                 f"**video overrides the spec where they conflict**. The previously PENDING-CHRIS rows "
                 f"are promoted and **applied as LOCAL case edits** (audit log "
                 f"`video-promotion-edit-log-2026-07-28.md`; applier "
                 f"`apply_video_promotion_2026-07-28.py`).\n\n")
        fh.write(f"**{TOTAL_CASES} cases** (515 in TestRail + 1 newly authored). **NO TestRail writes "
                 f"this pass** (Rule 6) - this change-list remains the approval gate: the push needs "
                 f"explicit permission (update_case for the edits, add_case for SBC-EXP-16, delete_case "
                 f"for the retire-proposed row). Run R359 untouched.\n\n")
        fh.write(f"**Counts:** {n_applied} edited in the first pass (APPLIED-NOW) | {n_local} "
                 f"video-authoritative / latest-info rows APPLIED-LOCALLY (covering 27 edited cases + "
                 f"1 new case - see the audit log for the per-case list) | {n_ret} RETIRE-PROPOSED | "
                 f"{n_noop} confirmed no-op (NO-CHANGE-CONFIRMED) | {n_open} open decisions "
                 f"(OPEN-DECISION) | {n_viu} awaiting a live-build check (LIVE-VIU-PENDING).\n\n")
        fh.write("**Backups (recovery requirement):** every touched case's verbatim PRE-EDIT body is "
                 "saved in `build/report-suite/video-promotion-backup-2026-07-28/` (27 files + "
                 "MANIFEST.md; the new case SBC-EXP-16 is delete-to-recover) - if Chris never "
                 "ratifies these video items into the specs, the originals are recoverable exactly. "
                 "**Spec-watch:** `build/report-suite/SPEC-WATCH-2026-07-28.md` tracks every item "
                 "awaiting spec ratification, deadline **2026-08-04**.\n\n")
        fh.write("**Legend:**\n"
                 "- **APPLIED-NOW** - edited locally in the first pass (firmly confirmed by Chris's written answer).\n"
                 "- **APPLIED-LOCALLY** - video-authoritative / latest-info edit applied to the local "
                 "case source (user ruling 2026-07-28) - **awaiting TestRail push authorization**.\n"
                 "- **RETIRE-PROPOSED** - proposed for deletion, NOT deleted - awaiting authorization.\n"
                 "- **NO-CHANGE-CONFIRMED** - latest info confirms the cases already match; nothing touched.\n"
                 "- **OPEN-DECISION** - still not settled; not applied.\n"
                 "- **LIVE-VIU-PENDING** - needs the live QA build to confirm; no QA branch yet, so "
                 "labelled 'not live-verified this run' (Rule 22).\n\n")
        fh.write("| Internal ID | Case | Report | Driving source | Classification | What needs to be done (plain) |\n")
        fh.write("|---|---|---|---|---|---|\n")
        for iid,cid,rep,src,cls,todo in ROWS:
            case = f"[{cid}]({link(cid)})" if cid else "_new - no C-ID yet_" if iid.startswith("SBC-EXP-16") else "_-_"
            fh.write(f"| {iid} | {case} | {rep} | {src} | **{cls}** | {todo} |\n")

        appl = [r for r in ROWS if r[4] in ('APPLIED-LOCALLY','RETIRE-PROPOSED')]
        fh.write(f"\n## Applied locally (video-authoritative) - awaiting TestRail push authorization ({len(appl)})\n\n")
        fh.write("Every row below is already edited/authored in `build/report-suite/cases/` - see the "
                 "per-case audit log `video-promotion-edit-log-2026-07-28.md` (driving video quote + "
                 "overridden spec wording per case, Rule 25). NOTHING is pushed (Rule 6). After the "
                 "authorized push: live VIU on the QA branch.\n\n")
        fh.write("| Internal ID | Case | Report | Driving source | What needs to be done (plain) |\n")
        fh.write("|---|---|---|---|---|\n")
        for iid,cid,rep,src,cls,todo in appl:
            case = f"[{cid}]({link(cid)})" if cid else "_new - no C-ID yet_"
            fh.write(f"| {iid} | {case} | {rep} | {src} | {todo} |\n")

        fh.write("\n## Confirmed already-matching (no change needed) - notable\n\n")
        fh.write("- **All-Time removal (video P9):** every report's date-range case correctly shows "
                 "NO 'All Time' option (matches spec). Caveat: Stefan noted the backend caps history "
                 "at ~365 days, so a Custom range beyond ~1 year returns limited data - a VIU/data "
                 "caveat, not a wording change.\n")
        fh.write("- **'Sales By Representative' naming (video P5):** SBR-NAV cases already use the "
                 "full 'Sales By Representative' name; 'Associate' appears nowhere.\n")
        fh.write("- **'Parts' nav section (video P2):** PV-NAV-01 (C30322) and IV-NAV-01 (C30534) "
                 "already place Parts Velocity + Inventory Value under a 'Parts' nav group - "
                 "re-verified during the promotion, no edit needed.\n")
        fh.write("- **No 'snapshot taken X days ago' label (video P32):** no case expects that "
                 "label; the Inventory Value 'As of <date>' indicator is a different, kept "
                 "indicator.\n")
        fh.write("- **Labor-delta colors (video P14):** SBC/SBR (invoiced-minus-worked) and WIP "
                 "(quoted-minus-worked) cases render green(+)/black(0.0)/red(-) per spec; the WIP "
                 "basis-vs-video difference is the LIVE-VIU-PENDING row above.\n")

        fh.write("\n## Q2 permission model (user ruling 2026-07-28)\n\n")
        fh.write("All permission cases KEEP the shipped MIXED model (Sales By Customer = dedicated "
                 "permission; Parts Velocity + Inventory Value = inventory-reports access; Sales By "
                 "Rep = performance group). They are **not edited**. The PO-vs-build discrepancy "
                 "(Chris wants 'normal reports access' for all) is captured for Chris/dev in "
                 "`build/report-suite/chris-answers-2026-07-28/Q2-permission-discrepancy-for-Chris-dev.md`. "
                 "NOTE: the video-P33 location-filter gating (applied above) is a separate, "
                 "location-permission question and does not change this ruling.\n")
    return out


if __name__ == "__main__":
    xlsx = write_xlsx()
    md = write_md()
    print("Wrote:", xlsx)
    print("Wrote:", md)
    from collections import Counter
    c = Counter(r[4] for r in ROWS)
    print("Rows by classification:", dict(c), "| total rows:", len(ROWS))
