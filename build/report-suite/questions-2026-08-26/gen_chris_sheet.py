#!/usr/bin/env python3
"""Generate the Chris Ward question sheet - Report Suite - 2026-08-26 (.xlsx + .md).

TWO questions, added on 2026-08-26 by the source-currency + write pass. Both are places
where Chris's own written description says a thing two different ways, and neither can be
resolved by looking at the build (Standing Rule 58) - so the affected cases are HELD until
he answers.

Mirrors the established format 1:1 (Standing Rule 16), following
build/report-suite/questions-2026-08-17/gen_chris_sheet.py:
 - same six reader columns in the same order (#, Topic, What happens now,
   The question, Options, Your answer);
 - same row layout (A1 title / A2 note / A4 header / A5 band / items from row 6);
 - same fills, fonts, freeze pane, column widths;
 - a QA-only mapping tab (Standing Rule 8) that is NEVER sent to Chris.

Standing Rule 7 + Rule 55: reader-facing text is extremely simple, names the
PROJECT (Report Suite) and the FEATURE (the report) on every question row (Chris also
owns Fees & Discounts), cites the epic in plain form, and carries no case IDs / spec
anchors / jargon / the word "VIU". Answerable without opening TestRail.

RESEARCH ONLY - this script writes two files into this folder. It makes NO
TestRail or Jira call of any kind. NOTHING IS SENT.
"""
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Report-Suite_Questions-for-Chris-Ward_2026-08-26.xlsx")
MD = os.path.join(HERE, "Report-Suite_Questions-for-Chris-Ward_2026-08-26.md")

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Questions for Chris"
TAB2_NAME = "QA internal - not for Chris"

INTRO = (
    "Two quick questions - one about the Sales By Representative report, one about the Work In "
    "Progress report. Each one is a plain A / B / C. Both are places where your written "
    "description says a thing two different ways, and we would rather have your word than pick a "
    "side ourselves. Every question names the project and the report, because we know you look "
    "after more than one thing here. There are no bugs on this sheet - just two wording "
    "decisions. You do not need to open anything to answer. Thank you."
)

# ------------------------------------------------------------------- the questions
# (topic, what-happens-now, the question, options)
TAB1 = [
    (
        "Report Suite - the Sales By Representative report - the list of columns you can switch "
        "on and off (the column chooser; under epic SV-8582)",

        "The Sales By Representative report has a column chooser - a list of columns the user can "
        "switch on and off.\n\n"
        "Your written description says two different things about that list, in the same "
        "document:\n\n"
        "- One part says the report has NINE columns that can be switched on and off, and names "
        "them in order, and Shop Supplies is one of the nine.\n\n"
        "- Another part also says there are NINE, but then lists only EIGHT - it names Labor "
        "Delta, Labor Invoiced, Labor Margin, Parts Invoiced, Parts Margin, Adjustments, Margin "
        "and Margin %, and Shop Supplies is missing from the list.\n\n"
        "So the count says nine and the list shows eight. Shop Supplies is the one that differs.\n\n"
        "For context: Shop Supplies was added as a new column on 21 August, and the note recording "
        "that change says it joins the column chooser as the ninth one - which points to the list "
        "simply having been missed when the column was added. We have not assumed that, though.\n\n"
        "Why we are asking: two of our tests check exactly this list, and they are on hold until "
        "you tell us. If Shop Supplies belongs in the chooser, the list needs one word added; if "
        "it does not, the count needs changing to eight.",

        "Should Shop Supplies be one of the columns a user can switch on and off - so the list is "
        "nine - or should it always be shown, so the list is really eight?",

        "A) Nine, with Shop Supplies among them - Shop Supplies can be switched on and off like "
        "the others, and the shorter list in your description is just missing it and should have "
        "it added.\n\n"
        "B) Eight - Shop Supplies is NOT something the user can switch off, and the places that "
        "say nine should be changed to eight.\n\n"
        "C) Something else - please tell us which columns can be switched on and off.",
    ),
    (
        "Report Suite - the Work In Progress report - what the Estimates figure counts (the "
        "Estimates total in the summary strip at the top; under epic SV-8582)",

        "On the Work In Progress report, the strip along the top shows an Estimates figure.\n\n"
        "Your written description describes what that figure counts in three places, and they do "
        "not agree about whether whole-job fees and discounts (the ones added to the job as a "
        "whole, rather than to a single line of work) are counted in it. Word for word:\n\n"
        "- First place: \"Estimates is the total quoted value of the jobs in the 'Estimates' tab, "
        "including their work-order-level adjustments so the figure matches the estimate document "
        "the customer sees.\" - that is, fees and discounts ARE counted.\n\n"
        "- Second place, the small information icon next to the figure: \"The total value of all "
        "estimate lines that have not yet been approved, including lines awaiting authorization on "
        "open work orders.\" - that counts LINES of work only, and says nothing about whole-job "
        "fees and discounts.\n\n"
        "- Third place: \"The Estimates figure is excluded from Total Earned and from Total "
        "Remaining.\" - which tells us the figure stands on its own, but not what is inside it.\n\n"
        "There is also a rule elsewhere in the same description saying that when one job shows up "
        "in two tabs, its whole-job fees and discounts are counted on the other tab's row and "
        "never on its Estimates row - so the same money is never shown twice. That reads as the "
        "opposite of the first place above, which is what we are stuck on.\n\n"
        "Why we are asking: the figure comes out at a different number depending on which one is "
        "right, and one of our tests is on hold until you tell us.",

        "Should the Estimates figure include whole-job fees and discounts, or should it count only "
        "the value of the lines of work that have not been approved yet?",

        "A) Include them - Estimates counts the lines PLUS the whole-job fees and discounts, so "
        "the figure matches the estimate document the customer sees.\n\n"
        "B) Do not include them - Estimates counts only the value of the not-yet-approved lines of "
        "work, and whole-job fees and discounts are shown elsewhere on the report instead.\n\n"
        "C) Something else - please describe what the figure should count.",
    ),
]

# --------------------------------------------------------------- QA-only mapping
QA_ROWS = [
    ("1",
     "Sales By Representative - column selector membership: S20-R2 says NINE toggleable columns "
     "then enumerates EIGHT, omitting Shop Supplies; contradicts S5-R2",
     "SBR spec contradicts itself inside the same live document. S5-R2 lists 14 columns = 4 "
     "identifier + NINE metric + Subtotal, INCLUDING Shop Supplies. S20-R2 says 'The nine "
     "toggleable columns are:' and then names only eight, Shop Supplies absent. The 2026-08-21 "
     "change-log entry (SV-9423) says Shop Supplies 'joins the column selector as the ninth "
     "toggleable metric column', which points to S20-R2's list simply not having been updated - "
     "but Rule 58 forbids resolving an ambiguous source by looking at the build, so no side is "
     "picked and the cases are HELD.",
     "C30265 and C43831 - both HELD pending this answer; neither was fetched or touched by the "
     "2026-08-26 write pass (excluded absolutely).",
     "C30265 https://shopview.testrail.io/index.php?/cases/view/30265 - "
     "C43831 https://shopview.testrail.io/index.php?/cases/view/43831",
     "SBR spec v24 (Confluence page 585629698, lastmod 2026-08-24), body read live 2026-08-26. "
     "S20-R2 verbatim: 'The nine toggleable columns are: Labor Delta, Labor Invoiced, Labor "
     "Margin, Parts Invoiced, Parts Margin, Adjustments, Margin, Margin %.' (count nine, list "
     "eight). S5-R2 verbatim: 'The columns appear left-to-right: Date, Invoice, Customer, Status, "
     "Labor Delta, Labor Invoiced, Labor Margin, Parts Invoiced, Parts Margin, Shop Supplies, "
     "Adjustments, Margin, Margin %, Subtotal. (14 columns: four leading identifier columns, nine "
     "metric columns, and Subtotal.)' Change log 2026-08-21 verbatim: Shop Supplies '...joins the "
     "column selector as the ninth toggleable metric column...'. S20-R3 lists the five "
     "always-visible columns as Date, Invoice, Customer, Status, Subtotal - Shop Supplies is not "
     "among them, which is consistent with A.",
     "A (nine incl. Shop Supplies) -> S20-R2's enumeration gains Shop Supplies; C30265 and C43831 "
     "are released from hold and authored/updated to a nine-item selector. "
     "B (eight) -> S5-R2's 'nine metric columns' and the 2026-08-21 change-log line both need "
     "correcting, and S20-R3's always-visible list gains Shop Supplies; the two cases are written "
     "to an eight-item selector. C -> per his enumeration."),
    ("2",
     "Work In Progress - does the Estimates summary figure include WORK-ORDER-LEVEL adjustments? "
     "S5-R8 says yes, the S5-R12 tooltip is line-only, and S4-R29 says the adjustments never ride "
     "the Estimates row",
     "WIP spec states the composition of the Estimates figure three ways in the same live "
     "document and they do not reconcile. Rule 58: not resolvable from the build; the case is "
     "HELD. NOTE FOR THE QA LEAD: the task named this as an Inventory Value question - it is not. "
     "C30491 is a WORK IN PROGRESS case (section 4356, 'The Estimates figure is the Estimates tab "
     "total, shown at full opacity'), and S5-R8/S5-R9/S5-R12 are WIP anchors. Inventory Value v10 "
     "S5-R8 is 'Changing the \"as of\" date reloads the report' and has nothing to do with "
     "Estimates. The row is therefore written against the Work In Progress report.",
     "C30491 (WIP summary strip) - HELD pending this answer. Not Automated (custom_atmstatus 1). "
     "Related and worth re-checking once answered: C30493 (the info-icon wording).",
     "C30491 https://shopview.testrail.io/index.php?/cases/view/30491 - "
     "C30493 https://shopview.testrail.io/index.php?/cases/view/30493",
     "WIP spec v28 (Confluence page 703660034, lastmod 2026-08-24), body read live 2026-08-26. "
     "S5-R8 verbatim: 'Estimates is the total quoted value of the jobs in the \"Estimates\" tab, "
     "including their work-order-level adjustments so the figure matches the estimate document the "
     "customer sees. It is shown at full opacity, like every other figure (un-muted in the "
     "2026-08-13 review).' S5-R12 Estimates tooltip verbatim: 'The total value of all estimate "
     "lines that have not yet been approved, including lines awaiting authorization on open work "
     "orders.' S5-R9 verbatim: 'The Estimates figure is excluded from Total Earned and from Total "
     "Remaining.' AND the sharpest conflict, S4-R29 verbatim: 'When a work order produces two "
     "rows, its work-order-level Adjustments amount rides only the status-tab row, never its "
     "Estimates row, so the same dollars are never shown twice.' S5-R13 adds that the summary "
     "strip shows no Adjustments figure at all.",
     "A (include adjustments) -> S5-R8 governs; the S5-R12 tooltip gains the adjustments clause "
     "and S4-R29 must be reconciled (a job in two tabs would double-count unless S4-R29 is "
     "restated); C30491 keeps expectation 1 and gains an adjustments clause. "
     "B (lines only) -> S5-R8's 'including their work-order-level adjustments' is struck; C30491 "
     "stands as written today and S4-R29 needs no change. C -> per his description."),
]

QA_NOTES = [
    "SCOPE - WHY ONLY TWO QUESTIONS: this sheet carries only the two genuinely-open PRODUCT "
    "decisions surfaced by the 2026-08-26 source-currency and write pass. Both are self-"
    "contradictions inside Chris's own live specification that Rule 58 forbids resolving against "
    "the build. Nothing already asked on an earlier sheet is re-asked (no drip - Rule 55), and no "
    "bug is raised here (Rule 7).",
    "HELD CASES: C30265 and C43831 (question 1) were excluded ABSOLUTELY from the 2026-08-26 "
    "write pass - never fetched, never touched - and remain held. C30491 (question 2) is held and "
    "was not written to. Rule 58: an ambiguous source is never resolved by looking at the build; "
    "the case waits for the answer.",
    "WORDING RULES APPLIED (Standing Rule 7 + Rule 55): every reader-facing question names the "
    "PROJECT (Report Suite) and the FEATURE (the Sales By Representative report / the Work In "
    "Progress report), because Chris also owns Fees & Discounts. The epic is named in plain form "
    "only to orient him. No case IDs, requirement anchors, HTTP terms or internal names appear in "
    "anything he reads, and both questions are answerable WITHOUT opening TestRail or any other "
    "tool - every quotation he needs is printed in the question. Each carries a one-line 'Why we "
    "are asking'.",
    "A CORRECTION THE QA LEAD SHOULD SEE: question 2 was handed to us as an INVENTORY VALUE "
    "question about C30491. It is not. C30491 is a Work In Progress case, and S5-R8 / S5-R9 / "
    "S5-R12 are Work In Progress anchors; Inventory Value v10's S5-R8 reads 'Changing the \"as "
    "of\" date reloads the report'. The row is written against the correct report. The three "
    "quoted anchors are the ones named in the instruction; S4-R29 is added because it is the "
    "sharpest contradiction of the three and Chris cannot decide well without it.",
    "NOTHING HAS BEEN SENT AND NOTHING WAS WRITTEN FOR THIS SHEET. This is a draft for the QA "
    "lead. No TestRail write, no Jira write and no case edit was made in producing it, and "
    "CLAUDE.md was not touched. The Jira creation hold (Rule 62 / register row H1) is untouched "
    "by this sheet.",
    "SOURCE-CURRENCY (Standing Rule 31), 2026-08-26: both quoted specifications were fetched live "
    "the same day - SBR v24 (page 585629698, lastmod 2026-08-24) and WIP v28 (page 703660034, "
    "lastmod 2026-08-24), bodies stored under build/report-suite/source-verify-2026-08-26/specs/. "
    "Epic SV-8582 was NOT re-read this pass (no Jira access in this session) - the epic is cited "
    "only to orient the reader, and no expectation on this sheet rests on it. The build was NOT "
    "observed this pass, so the Rule-49 re-check queue stays OPEN and every verdict on this "
    "project remains provisional.",
]

# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def write_xlsx():
    wb = openpyxl.Workbook()
    widths = [4, 36, 60, 44, 50, 22]

    ws = wb.active
    ws.title = TAB1_NAME
    ws["A1"] = "Questions for Chris Ward - Report Suite - 2026-08-26"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = INTRO
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, "Two wording decisions - one on Sales By Representative, one on Work In Progress", 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(TAB1, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 340
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"

    # -------------------------------------------------- QA-only mapping tab
    ws4 = wb.create_sheet(TAB2_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR CHRIS. Do not send this tab. TestRail C-ids, "
                 "requirement anchors and live evidence live here so the reader-facing tab stays "
                 "plain (Standing Rules 7 and 8).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Item", "What it asks", "Where the ambiguity comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFG", [7, 42, 44, 44, 52, 62, 62]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, 1):
            out.append(f"### Question {i} - {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    md = f"""# Questions for Chris Ward - Report Suite - 2026-08-26

**Project: Report Suite (the six reports) - epic SV-8582 - Product Owner: Chris Ward**

**This is the plain-language twin of `Report-Suite_Questions-for-Chris-Ward_2026-08-26.xlsx`.**
The spreadsheet is the version to send; it mirrors the established sheet format exactly, and it
carries a QA-only tab that must not be forwarded.

**DRAFT - NOT SENT. Nothing has been written to TestRail or Jira.**

{INTRO}

**Two questions in total - one on the Sales By Representative report, one on the Work In\nProgress report - each a plain A / B / C.**

---

## {TAB1_NAME}

{block(TAB1)}

---

## QA-only - not for Chris

The internal question-to-case mapping lives on the spreadsheet's `{TAB2_NAME}` tab: each question's
affected TestRail case IDs with links, the requirement anchors, the live evidence, and what each
possible answer resolves to. It also records the scope, wording rules and source-currency notes.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
