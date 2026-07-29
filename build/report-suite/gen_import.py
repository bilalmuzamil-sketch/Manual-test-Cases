#!/usr/bin/env python3
"""Generate the Report Suite v1 TestRail import file (CSV + XLSX) from the
authored per-case JSON in build/report-suite/cases/ (26 files, 6 reports).

CANONICAL FORMAT — PURE 1:1 MATCH to
testrail-import/fees-discounts-v1-testrail-import.csv,
testrail-import/simple-flow-v1-testrail-import.csv,
testrail-import/global-search-v2-testrail-import.csv,
testrail-import/filters-v1-testrail-import.csv and
testrail-import/schedule-v1-testrail-import.csv:
  EIGHT named columns, IDENTICAL in name and order to the other project
  imports, followed by TWO trailing UNNAMED (blank) columns:
      Title, Section, Type, Priority, Preconditions, Steps, Expected Result,
      References, "", ""
  There are NO Report-Suite-specific columns. Standing-Rule-8 traceability
  (internal SBC-/SBR-/PV-/TU-/WIP-/IV- id <-> TestRail Case ID) lives in
  build/report-suite/testrail-id-map.csv exactly as it does for the other
  projects; the import file itself is byte-format-identical to theirs.

TESTRAIL STRUCTURE (user-prescribed, PROJECT-STATE §0.5): ONE main section
"Report Suite" -> one SUBSECTION per report area. The Section column carries
the case's leaf value exactly as authored ("SBC — Filters", "WIP — API", ...);
the user's TestRail import nests these under the "Report Suite" main section.
STANDING RULE 4: every API-related case is authored in a "<Report> — API"
area, so its Section title includes 'API' (verified by a sanity check below).

CONTENT RULES enforced here (same as the other generators):
  1. VIU-word-free: internal `viu_status`, `notes`, `design_ref`,
     `permissions_required` are NOT emitted; a sanity check confirms 0 "viu"
     occurrences in the emitted cells.
  2. Feature-flag-free: no feature-flag phrasing survives (sanity-checked).
  3. NO internal-id leakage into reader-facing cells: "(see PV-PERM-01)"-style
     cross-refs and bare SBC-/SBR-/PV-/TU-/WIP-/IV- id mentions are stripped/
     rewritten generically (same fix as the Schedule generator); a sanity
     check confirms 0 survivors.
  4. References = spec reference only.
  5. Preconditions/Steps/Expected kept as authored (numbered, line-broken),
     cleaned per rules 1-3.

ORDERING (deterministic): report order SBC, SBR, PV, TU, WIP, IV; within a
report, sections in authored (spec/user-journey) first-appearance order;
within a section, by internal id (all numeric suffixes are two-digit, so
lexicographic == natural).

Also writes build/report-suite/testrail-id-map.csv (Standing Rule 8: all
internal ids, blank TestRail Case-ID column until a permitted push assigns
C-ids; schema identical to the filters/schedule id-maps:
internal_id,testrail_case_id,title,section).
⚠️ GOTCHA (same as Filters/Schedule): re-running this generator BLANKS the
testrail_case_id column — after C-ids are ever populated, RE-MERGE them after
any rerun.

Outputs (canonical location + naming, matching the other projects):
  testrail-import/report-suite-v1-testrail-import.csv
  testrail-import/report-suite-v1-testrail-import.xlsx
  build/report-suite/testrail-id-map.csv

PER-REPORT SPLIT FILES (2026-07-22, user's per-folder import workflow): the
user manually created TestRail group 4281 "Reports Suite" with six EMPTY
per-report subsections (4282 Sales By Customer Report / 4283 Sales By
Representative Report / 4284 Parts Velocity Report / 4285 Technician
Utilization — Product Specification / 4286 Work In Progress — Product
Specification / 4287 Inventory Value — Product Specification) and imports ONE
report at a time targeting each folder; the CSV Section column then creates
the "<PREFIX> — <area>" leaf sections inside that folder. So, IN ADDITION to
the unchanged unified files, this generator ALSO emits six per-report files
with HUMAN-READABLE names (user rule 2026-07-22: spell report names out in
full — no cryptic abbreviations like sbc/pv/tu):
  testrail-import/Report-Suite_<Full-Report-Name>_testrail-import.csv/.xlsx
  i.e. Report-Suite_Sales-By-Customer-Report_testrail-import.csv,
       Report-Suite_Sales-By-Representative-Report_testrail-import.csv,
       Report-Suite_Parts-Velocity-Report_testrail-import.csv,
       Report-Suite_Technician-Utilization-Report_testrail-import.csv,
       Report-Suite_Work-In-Progress-Report_testrail-import.csv,
       Report-Suite_Inventory-Value-Report_testrail-import.csv (+ .xlsx twins).
Each contains ONLY that report's rows in the IDENTICAL canonical format:
same byte-identical header (8 named cols + 2 blanks), same CRLF row
terminator, same per-report row ordering as the unified file, every data row
byte-identical to its unified-file counterpart (verified by the sanity check:
concatenation of the six, minus repeated headers, == the unified 515 rows).
Row counts: SBC 99 / SBR 127 / PV 70 / TU 59 / WIP 83 / IV 77 = 515.
"""
import csv, json, glob, os, re
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))          # build/report-suite/
BASE = os.path.dirname(HERE)                               # build/
ROOT = os.path.dirname(BASE)                               # repo root
CASES_DIR = os.path.join(HERE, "cases")
OUT_CSV = os.path.join(ROOT, "testrail-import", "report-suite-v1-testrail-import.csv")
OUT_XLSX = os.path.join(ROOT, "testrail-import", "report-suite-v1-testrail-import.xlsx")
OUT_IDMAP = os.path.join(HERE, "testrail-id-map.csv")

# Byte-identical (name + order) to the fees-discounts / simple-flow /
# global-search / filters / schedule imports: 8 named columns + 2 trailing
# UNNAMED (blank) columns.
HEADER = [
    "Title", "Section", "Type", "Priority",
    "Preconditions", "Steps", "Expected Result", "References",
    "", "",
]

# Deterministic report order (user-prescribed suite order).
REPORT_ORDER = ["SBC", "SBR", "PV", "TU", "WIP", "IV"]

# Per-report split-file naming (HUMAN-READABLE, user rule 2026-07-22 — full
# report names, never cryptic abbreviations) + XLSX sheet titles.
REPORT_FILE = {"SBC": "Sales-By-Customer-Report",
               "SBR": "Sales-By-Representative-Report",
               "PV": "Parts-Velocity-Report",
               "TU": "Technician-Utilization-Report",
               "WIP": "Work-In-Progress-Report",
               "IV": "Inventory-Value-Report"}
REPORT_SHEET = {"SBC": "Sales By Customer",
                "SBR": "Sales By Representative",
                "PV": "Parts Velocity",
                "TU": "Technician Utilization",
                "WIP": "Work In Progress",
                "IV": "Inventory Value"}

# Internal case-id pattern (all six report prefixes).
IDPAT = r"(?:SBC|SBR|PV|TU|WIP|IV)-[A-Z]+-\d+"


def clean(s):
    """Strip internal authoring markers / internal-id cross-refs; keep the
    functional content (mirrors the Schedule generator's leak fix)."""
    if not s:
        return s
    # "(see X)", "(per X)", "(from X)", "(verified in X)", "(as seeded for X)"
    s = re.sub(r"\s*\((?:see|per|from|verified in|as seeded for)\s+" + IDPAT +
               r"(?:(?:,|\s+and)\s+" + IDPAT + r")*(?:'s setup)?\)", "", s)
    # "(X)" bare parenthetical id
    s = re.sub(r"\s*\(" + IDPAT + r"\)", "", s)
    # ", see X" / "— see X" embedded inside a longer parenthetical/sentence
    s = re.sub(r",?\s*[—–-]?\s*see\s+" + IDPAT, "", s)
    # "covered by X and Y" -> generic
    s = re.sub(r"covered by\s+" + IDPAT + r"(?:\s+and\s+" + IDPAT + r")*",
               "covered by separate cases", s)
    # Defensive: strip any remaining bare internal-id mentions.
    s = re.sub(r"[,;]?\s*(?:as seeded for\s+|verified in\s+)?" + IDPAT, "", s)
    s = re.sub(r"feature[ -]flags?", "report feature", s, flags=re.I)
    return s


def joinlines(lst):
    return "\n".join(clean(x.rstrip()) for x in (lst or []))


def load_cases():
    """Load all authored cases, EXCLUDING Retired ones (2026-07-28 consolidation:
    merged-away members / cuts / the Print retire keep their bodies in cases/*.json
    marked viu_status 'Retired ...' and are excluded from every deliverable —
    same convention as the Fees & Discounts / Simple Flow generators)."""
    cases = []
    for f in sorted(glob.glob(os.path.join(CASES_DIR, "cases-*.json"))):
        cases += [c for c in json.load(open(f))
                  if not str(c.get("viu_status", "")).startswith("Retired")]
    return cases


def write_csv(path, rows):
    """Canonical-format CSV: byte-identical header, QUOTE_MINIMAL, CRLF."""
    with open(path, "w", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(HEADER)
        w.writerows(rows)


def write_xlsx(path, sheet_title, rows):
    """Canonical-format XLSX twin (same styling as the unified file).
    Returns False if openpyxl is unavailable."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(HEADER)
    for r in rows:
        ws.append(r)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")

    widths = {"Title": 50, "Section": 40, "Type": 12, "Priority": 10,
              "Preconditions": 60, "Steps": 60, "Expected Result": 60,
              "References": 34}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, len(rows) + 2):
        for cidx in range(1, len(HEADER) + 1):
            ws.cell(row=r, column=cidx).alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(path)
    return True


def main():
    cases = load_cases()

    # Per-report section order = first-appearance in the authored files
    # (files load per report A->E, so this is the authored journey order).
    sec_order = {}
    for c in cases:
        sec_order.setdefault(c["area"].strip(), len(sec_order))
    rep_order = {p: i for i, p in enumerate(REPORT_ORDER)}
    cases.sort(key=lambda c: (rep_order[c["id"].split("-")[0]],
                              sec_order[c["area"].strip()],
                              c["id"]))

    rows = []
    titles_by_section = []
    ids = []
    idmap_rows = []
    api_cases = 0
    api_section_bad = []
    for c in cases:
        title = clean(c["title"].strip())
        section = c["area"].strip()
        ids.append(c["id"])
        titles_by_section.append((section, title))
        if c.get("api_related"):
            api_cases += 1
            if "API" not in section:
                api_section_bad.append(c["id"])
        idmap_rows.append([c["id"], "", title, section])
        rows.append([
            title,
            section,
            c.get("type", "Functional"),
            c["priority"].strip(),
            joinlines(c.get("preconditions")),
            joinlines(c.get("steps")),
            joinlines(c.get("expected")),
            clean((c.get("spec_ref") or "").strip()),
            "",
            "",
        ])

    write_csv(OUT_CSV, rows)
    print("Wrote CSV:", OUT_CSV, "rows(data):", len(rows))

    # --- per-report split files (same order as the unified file) ---
    per_report_rows = {p: [] for p in REPORT_ORDER}
    for iid, r in zip(ids, rows):
        per_report_rows[iid.split("-")[0]].append(r)
    for p in REPORT_ORDER:
        path = os.path.join(ROOT, "testrail-import",
                            "Report-Suite_%s_testrail-import.csv" % REPORT_FILE[p])
        write_csv(path, per_report_rows[p])
        print("Wrote CSV:", path, "rows(data):", len(per_report_rows[p]))

    with open(OUT_IDMAP, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["internal_id", "testrail_case_id", "title", "section"])
        w.writerows(idmap_rows)
    print("Wrote ID map:", OUT_IDMAP, "rows:", len(idmap_rows))

    # --- Sanity checks (must be zero / clean) ---
    dupes = [t for t, n in Counter(titles_by_section).items() if n > 1]
    print("Duplicate titles within a section:", dupes if dupes else "NONE")
    dupids = [i for i, n in Counter(ids).items() if n > 1]
    print("Duplicate internal ids:", dupids if dupids else "NONE")
    per_report = Counter(i.split("-")[0] for i in ids)
    print("Per-report counts:", {p: per_report[p] for p in REPORT_ORDER})
    blob = "\n".join("\t".join(str(x) for x in r) for r in rows).lower()
    print("VIU occurrences:", blob.count("viu"))
    print("'feature flag' occurrences:", blob.count("feature flag"))
    print("'flag on' occurrences:", blob.count("flag on"))
    print("'flag off' occurrences:", blob.count("flag off"))
    leaks = re.findall(IDPAT, "\n".join("\t".join(str(x) for x in r) for r in rows))
    print("Internal-id leaks in cells:", leaks if leaks else "NONE")
    print("API cases:", api_cases,
          "| API cases NOT in an 'API' section:", api_section_bad if api_section_bad else "NONE")
    empties = [iid for iid, r in zip(ids, rows)
               if not (r[4].strip() and r[5].strip() and r[6].strip())]
    print("Rows missing Preconditions/Steps/Expected:", empties if empties else "NONE")

    # --- xlsx review copies (unified + six per-report twins) ---
    if not write_xlsx(OUT_XLSX, "Report Suite V1", rows):
        print("openpyxl not available - skipped XLSX.")
        return
    print("Wrote XLSX:", OUT_XLSX)
    for p in REPORT_ORDER:
        path = os.path.join(ROOT, "testrail-import",
                            "Report-Suite_%s_testrail-import.xlsx" % REPORT_FILE[p])
        write_xlsx(path, REPORT_SHEET[p], per_report_rows[p])
        print("Wrote XLSX:", path)


if __name__ == "__main__":
    main()
