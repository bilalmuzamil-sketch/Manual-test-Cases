#!/usr/bin/env python3
"""Generate the single-issue Location-column question sheet for Chris Ward.

Mirrors build/report-suite/gen_po_questions_2026-08-03.py 1:1 (Standing Rule 16):
same two-tab shape, same header fill/font, same column widths, same QA-only banner,
same abort-on-C-id-mismatch guard.  Human-readable filename (Standing Rule 19).

Reader-facing text carries NO case ids, NO spec anchors, NO version numbers and no
jargon (Standing Rule 7); full report names throughout.  Chris's own specification
wording is quoted verbatim and keeps its own phrasing (Standing Rule 25).
"""
import csv
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
REPO_RS = os.path.dirname(HERE)
BASE = "PO-Question-Chris-Ward-Location-Column-2026-08-04"
TITLE = ("Report Suite - One question for Chris Ward: the location column - "
         "2026-08-04 (needed today)")

# ---------------------------------------------------------------- reader-facing
INTRO = (
    "One question only, and it should take a minute. It is separate from the longer sheet you "
    "already have - that one still stands. This one is urgent because the automated versions of "
    "these tests are being written today, and eight of our checks are waiting on your answer."
)

NOW_LEAD = ("The six reports can show a location column, telling you which branch each row "
            "belongs to. Right now they do not agree on how it should behave:")

NOW_BULLETS = [
    ("Sales By Customer, Sales By Representative, Parts Velocity and Technician Utilization",
     "handle it on their own - the column appears when you are looking at more than one location, "
     "and disappears when you narrow to one."),
    ("Work In Progress",
     "never shows it on its own. The column is missing until you switch it on yourself from the "
     "list of columns - even when you have every location in view."),
    ("Inventory Value",
     "does the opposite. The column is on from the start, and it stays on even after you narrow "
     "to a single location - so you get a column repeating the same branch name on every row."),
    ("One more oddity on Inventory Value -",
     "the screen and the downloaded file disagree with each other. The download drops the column "
     "when you narrow to one location, but the screen keeps it."),
]

NOW = NOW_LEAD + " " + " ".join(f"{a}: {b}" for a, b in NOW_BULLETS)

HIS_WORDS = [
    ("Work In Progress",
     "The Location column is not offered in the column selector; its visibility is automatic - "
     "shown only when more than one location is in scope (Story 7)."),
    ("Work In Progress",
     "...and is hidden whenever a single location is in scope; the user does not toggle it in the "
     "column selector."),
    ("Inventory Value",
     "Its visibility follows the location scope automatically and it is not one of the columns "
     "offered in the column-selection control (Story 8)."),
]

WHY = (
    "We are asking rather than assuming because our eight checks for those two reports currently "
    "describe what the product does today, not what your description asks for - which means if "
    "the product is the thing that is wrong, our tests would quietly pass it instead of catching it."
)

QUESTION = ("Which behaviour should all six reports use for the location column?")

OPTIONS = [
    "A) The column appears on its own whenever more than one location is in view, and disappears "
    "when only one is - it is not something the user switches on. (This matches what both your "
    "written descriptions already say, and what the other four reports already do. If you choose "
    "A we will raise the two reports that behave differently, and correct our eight checks so "
    "they would catch it.)",
    "B) The column is a switch the user turns on and off from the list of columns, and it stays "
    "however they left it. (If you choose B, the two written descriptions need updating to say "
    "so, and we will keep our eight checks as they are.)",
    "C) Something else, or it should differ between reports - please describe it.",
]

CLOSING = (
    "Needed today, please: the automated versions of these tests are being written today, and "
    "these eight checks cannot be finalised until we know which behaviour is the correct one."
)

# ------------------------------------------------------------------ QA-only tab
# internal_id -> (C-id is verified against testrail-id-map.csv before anything is written)
QA_CASES = [
    ("IV-COL-01", "C30551", "Inventory Value", "Inventory Value spec v3 S7-R6",
     "Asserts Location is in the column-selection control and appears 'when it is turned on', "
     "between Vendor and Qty."),
    ("IV-COL-04", "C30554", "Inventory Value", "Inventory Value spec v3 S7-R6",
     "Asserts Location can be 'turned on from the column-selection control' and then appears in "
     "its fixed position."),
    ("IV-PERS-02", "C30580", "Inventory Value", "Inventory Value spec v3 S7-R6",
     "Fixed column order stated 'with Location, when it is turned on in the column-selection "
     "control, between Vendor and Qty'."),
    ("IV-EXP-02", "C30588", "Inventory Value", "Inventory Value spec v3 S7-R6",
     "Tester note says the files carry Location 'when Location is turned ON in the "
     "column-selection control'."),
    ("IV-LOC-06", "C38917", "Inventory Value", "Inventory Value spec v3 S7-R6",
     "Step 1 instructs the tester to 'Turn Location on in the column-selection control'; expected "
     "says visibility 'follows that toggle, not the location selection'."),
    ("WIP-COL-01", "C30466", "Work In Progress", "WIP spec v6 S4-R3; S7-R13",
     "Precondition 4: 'Location is turned ON in the column-selection control (it is off by "
     "default).'"),
    ("WIP-COL-02", "C30467", "Work In Progress", "WIP spec v6 S4-R3; S7-R13",
     "Asserts Location IS offered in the selector, off by default, and does 'NOT appear on its "
     "own' - and says out loud 'That is what the build does today.'"),
    ("WIP-FLT-09", "C38916", "Work In Progress", "WIP spec v6 S4-R3; S7-R13",
     "Asserts 'The column does not appear or disappear on its own ... it follows the "
     "column-selection toggle only.'"),
]

QA_RESOLUTION = [
    ("If Chris answers A (automatic - matches both specs)",
     "All 8 cases are re-worded to assert the automatic, scope-driven model that "
     "SBC/SBR/PV/TU already use, and the observed selector-controlled build is recorded as a "
     "DEVIATION in the case notes (the pattern WIP-FLT-05 = C30502 already uses). Two build "
     "defects get raised: Work In Progress never shows it automatically; Inventory Value never "
     "hides it at single scope on screen. This is the outcome Standing Rule 33 already points to "
     "- the specs outrank our build observation - so A costs 8 re-words and 2 tickets."),
    ("If Chris answers B (user-toggled)",
     "All 8 cases stand exactly as they are; no TestRail write is needed. Chris updates the two "
     "written descriptions (WIP S4-R3 + S7-R13, IV S7-R6). The 11 cases on the other four "
     "reports that assert the automatic model stay correct, because B would apply only to the "
     "two reports whose descriptions change - CONFIRM THIS WITH HIM if he picks B, since a "
     "suite-wide B would invalidate those 11."),
    ("If Chris answers C (something else / differs per report)",
     "Re-derive per report from his answer; expect a further reconciliation pass and treat all 8 "
     "as blocked until then."),
    ("Regardless of the answer - a separate surface split to settle",
     "On Inventory Value the SCREEN keeps the Location column at single scope while the CSV "
     "download drops it (screen observed 2026-08-04; CSV observed 2026-08-03, "
     "viu-2026-08-03/evidence/location-matrix/inventory-value__SINGLE__plain.csv has no Location "
     "header, __MULTI__ does). Two surfaces, two behaviours - IV-EXP-02 (C30588) is the export "
     "case affected. Standing Rule 40: every surface gets its own verdict."),
]

EVIDENCE = [
    ("Work In Progress, every location in view",
     "Headers: WO # | Status | Customer | Asset | Advisor | Days Open | Earned | Remaining | "
     "Total. NO Location column. Location IS listed in the Column Selection panel.",
     "evidence/location-behaviour.json; evidence/work-in-progress-selector.png"),
    ("Work In Progress, one location in view",
     "Identical headers - still no Location column.",
     "evidence/location-single-vs-multi.json; evidence/wip-ONE-location-screen.png"),
    ("Inventory Value, every location in view",
     "Headers: Part # | Description | Category | Vendor | LOCATION | Qty | Unit Cost | Unit Sell "
     "| Margin | Margin % | Total Sell | Total Cost. Location present, and ALSO offered in the "
     "Column Selection panel.",
     "evidence/location-behaviour.json; evidence/inventory-value-selector.png"),
    ("Inventory Value, ONE location in view (the deviation)",
     "Location filter reads 'Staging Lethbridge - 4310' (single) yet the Location column is STILL "
     "shown, every row repeating 'Staging Lethbridge - 4310'. Reproduced twice.",
     "evidence/location-single-vs-multi.json; evidence/iv-ONE-location-screen.png"),
    ("Build marker (Standing Rule 49 - branch declared NOT FINAL)",
     "v3.4.1-0ed4433 on sv8582.qa.shopview.com; index.html last-modified Mon, 03 Aug 2026 "
     "13:40:38 GMT, etag 02091e9dc11f187d7739b4efa166ea21 - byte-identical to the 2026-08-03 "
     "marker, so the build has not moved. All observations PROVISIONAL.",
     "../viu-2026-08-03/RECHECK-QUEUE.md"),
]

LINK = "https://shopview.testrail.io/index.php?/cases/view/"


def verify_cids():
    """Abort unless every internal-id/C-id pair matches testrail-id-map.csv."""
    idmap = {}
    with open(os.path.join(REPO_RS, "testrail-id-map.csv"), newline="") as f:
        for row in csv.reader(f):
            if row and row[0] != "internal_id":
                idmap[row[0]] = row[1]
    bad = [(i, c) for i, c, *_ in QA_CASES if idmap.get(i) != c]
    # also check any pair mentioned in the resolution prose
    text = " ".join(a + " " + b for a, b in QA_RESOLUTION)
    for i, c in re.findall(r"([A-Z]{2,4}(?:-[A-Z]+)+-\d+)\s*[=(]\s*(C\d+)", text):
        if idmap.get(i) != c:
            bad.append((i, c))
    if bad:
        raise SystemExit("C-ID MISMATCH vs testrail-id-map.csv: " + str(bad))
    print(f"C-id verification: {len(QA_CASES)} case rows + prose pairs checked, all MATCH the id-map")
    return idmap


def check_reader_text_is_clean():
    """Guard Standing Rule 7 on the reader-facing text only.

    Chris's own quoted specification wording is deliberately EXCLUDED from this scan -
    it is his text and keeps its phrasing.
    """
    reader = " ".join([TITLE, INTRO, NOW, WHY, QUESTION, CLOSING] + OPTIONS)
    problems = []
    # internal case ids  (SBR-EXP-10 etc.) and TestRail C-ids
    problems += [("internal case id", m) for m in re.findall(r"\b[A-Z]{2,4}(?:-[A-Z]+)+-\d+\b", reader)]
    problems += [("TestRail C-id", m) for m in re.findall(r"\bC\d{4,6}\b", reader)]
    # spec anchors  (S7-R6, S14-R20, S11-E1)
    problems += [("spec anchor", m) for m in re.findall(r"\bS\d+-[REN]\d+\b", reader)]
    # version numbers  (v3, v6, v15, "version 6")
    problems += [("version number", m) for m in re.findall(r"\bv\d+\b|\bversion \d+\b", reader)]
    # jargon
    for w in ["API", "HTTP", "endpoint", "payload", "refs", "VIU", "TestRail", "Jira", "epic",
              "anchor", "regression", "backend", "front-end", "frontend", "CSV", "JSON",
              "boolean", "toggle state", "Standing Rule", "deviation", "spec"]:
        if re.search(r"\b" + re.escape(w) + r"\b", reader, re.I):
            problems.append(("jargon", w))
    if problems:
        raise SystemExit("READER-FACING TEXT NOT CLEAN: " + str(problems))
    # full report names, never abbreviations
    for abbr in ["SBC", "SBR", "PV", "TU", "WIP", "IV"]:
        if re.search(r"\b" + abbr + r"\b", reader):
            raise SystemExit(f"READER TEXT uses the abbreviation {abbr}; use the full report name")
    print("Reader-facing text: clean - no case ids, no spec anchors, no version numbers, "
          "no jargon, no report abbreviations")


def write_md():
    L = [f"# {TITLE}", "",
         "**STATUS: READY TO SEND** (not yet sent). Single issue. On return: ingest verbatim, "
         "then apply to the eight affected cases per the standing workflow - nothing is edited "
         "before his answer and the QA lead's go-ahead.", "",
         INTRO, "", "---", "",
         "## The location column - should it appear on its own, or does the user switch it on?", "",
         f"**What happens now:** {NOW_LEAD}", ""]
    for what, does in NOW_BULLETS:
        L.append(f"- **{what}** {does}")
    L += ["",
          "**But both of those two written descriptions say the column should be automatic, and "
          "should not be something the user switches on.** Quoting them directly:", ""]
    for report, quote in HIS_WORDS:
        L.append(f"- **{report}:** *\"{quote}\"*")
    L += ["",
          f"**Why we are asking:** {WHY}", "",
          f"**The question:** {QUESTION}", "",
          "**Options:**", ""]
    for o in OPTIONS:
        L.append(f"- {o}")
    L += ["", "**Your answer:** ____________________", "",
          f"**{CLOSING}**", "", "---", ""]

    # ---- QA-only appendix
    L += ["# QA-ONLY APPENDIX - do not send this part to Chris", "",
          "TestRail C-ids verified against `build/report-suite/testrail-id-map.csv` at generation "
          "time; the generator aborts on any mismatch (Standing Rule 8).", "",
          "## The affected cases - the TRUE count is 8, not 7", "",
          "The audit's contradiction group is headed *\"3 cases + 4 more to align\"* (= 7) but its "
          "own table lists **8**, and all **8** were re-read live from TestRail on 2026-08-04 and "
          "**every one asserts the selector-controlled model**. The undercount appears to be "
          "`WIP-COL-01` (C30466), whose assertion sits in a *precondition* rather than an expected "
          "result - which is exactly why it still needs the edit. **True count: 8** (5 Inventory "
          "Value + 3 Work In Progress).", "",
          "| Internal ID | TestRail | Link | Report | Spec anchor | What it asserts today |",
          "|---|---|---|---|---|---|"]
    for i, c, rep, anchor, says in QA_CASES:
        L.append(f"| {i} | {c} | [open]({LINK}{c[1:]}) | {rep} | {anchor} | {says} |")
    L += ["", "## What each answer resolves to", "",
          "| Answer | Consequence |", "|---|---|"]
    for a, b in QA_RESOLUTION:
        L.append(f"| {a} | {b} |")
    L += ["", "## Live evidence behind the \"what happens now\" text", "",
          "| Observation | What was seen | Evidence |", "|---|---|---|"]
    for a, b, c in EVIDENCE:
        L.append(f"| {a} | {b} | `{c}` |")
    L += ["",
          "## Honesty notes", "",
          "- Both reports were driven live on 2026-08-04 by mouse clicks; the single-location "
          "state was reached by de-selecting one of the two locations, leaving the filter reading "
          "a single branch, and the result was reproduced twice.",
          "- The Column Selection panel's per-item on/off state could NOT be read reliably by "
          "automation (the detector reported every item as off, including columns plainly "
          "displayed). So the claim *\"off by default\"* on Work In Progress rests on the "
          "**column's absence from the grid** plus the 2026-08-03 pass's own observation, not on "
          "a machine reading of the toggle. The presence/absence of the column - which is what "
          "the question turns on - is solid.",
          "- Standing Rule 49: the branch was declared NOT FINAL, so every observation here is "
          "PROVISIONAL and carries the build marker above.", ""]
    with open(os.path.join(HERE, BASE + ".md"), "w") as f:
        f.write("\n".join(L) + "\n")
    print("wrote", BASE + ".md")


def write_xlsx():
    wb = openpyxl.Workbook()
    hdr_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF")
    grp_fill = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True)
    cols = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    gc = ws.cell(row=4, column=1, value="DECISION WE NEED FROM YOU - TODAY")
    gc.font = Font(bold=True)
    for j in range(1, 7):
        ws.cell(row=4, column=j).fill = grp_fill
    quotes = "Your own descriptions say it should be automatic:\n" + "\n".join(
        f'{r}: "{q}"' for r, q in HIS_WORDS)
    row = [1, "The location column - should it appear on its own, or does the user switch it on?",
           NOW + "\n\n" + quotes + "\n\n" + WHY, QUESTION, "\n\n".join(OPTIONS), ""]
    for j, v in enumerate(row, 1):
        ws.cell(row=5, column=j, value=v).alignment = wrap
    ws.cell(row=7, column=3, value=CLOSING).alignment = wrap
    ws.cell(row=7, column=3).font = Font(bold=True)
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[5].height = 300
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids from "
                 "build/report-suite/testrail-id-map.csv (Standing Rule 8); links "
                 "https://shopview.testrail.io/index.php?/cases/view/<id>. Every C-id verified "
                 "against the id-map at generation time - the generator aborts on a mismatch. "
                 "TRUE COUNT IS 8, NOT 7: the audit group is headed '3 cases + 4 more' (=7) but "
                 "lists 8, and all 8 were re-read live on 2026-08-04 and all 8 assert the "
                 "selector-controlled model. The undercount appears to be WIP-COL-01 (C30466), "
                 "whose assertion sits in a precondition rather than an expected result.")
    ws2["A1"].font = Font(bold=True)
    cols2 = ["Internal ID", "TestRail C-id", "Link", "Report", "Spec anchor",
             "What it asserts today"]
    for j, c in enumerate(cols2, 1):
        cell = ws2.cell(row=3, column=j, value=c)
        cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r = 4
    for i, c, rep, anchor, says in QA_CASES:
        for j, v in enumerate([i, c, LINK + c[1:], rep, anchor, says], 1):
            ws2.cell(row=r, column=j, value=v).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="WHAT EACH ANSWER RESOLVES TO").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Answer", "Consequence", "", "", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for a, b in QA_RESOLUTION:
        ws2.cell(row=r, column=2, value=a).alignment = wrap
        ws2.cell(row=r, column=3, value=b).alignment = wrap
        r += 1
    r += 1
    ws2.cell(row=r, column=1, value="LIVE EVIDENCE BEHIND THE 'WHAT HAPPENS NOW' TEXT").font = Font(bold=True)
    r += 1
    for j, c in enumerate(["", "Observation", "What was seen", "Evidence", "", ""], 1):
        cell = ws2.cell(row=r, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = hdr_fill, hdr_font, wrap
    r += 1
    for a, b, c in EVIDENCE:
        ws2.cell(row=r, column=2, value=a).alignment = wrap
        ws2.cell(row=r, column=3, value=b).alignment = wrap
        ws2.cell(row=r, column=4, value=c).alignment = wrap
        r += 1
    for col, w in zip("ABCDEF", [14, 14, 46, 18, 26, 60]):
        ws2.column_dimensions[col].width = w
    wb.save(os.path.join(HERE, BASE + ".xlsx"))
    print("wrote", BASE + ".xlsx")


if __name__ == "__main__":
    verify_cids()
    check_reader_text_is_clean()
    write_md()
    write_xlsx()
