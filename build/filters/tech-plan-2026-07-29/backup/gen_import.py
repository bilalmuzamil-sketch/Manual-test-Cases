#!/usr/bin/env python3
"""Generate the Filters v1 TestRail import file (CSV + XLSX) from the authored
per-case JSON in build/filters/cases/.

CANONICAL FORMAT — PURE 1:1 MATCH to
testrail-import/fees-discounts-v1-testrail-import.csv,
testrail-import/simple-flow-v1-testrail-import.csv and
testrail-import/global-search-v2-testrail-import.csv:
  EIGHT named columns, IDENTICAL in name and order to the other project
  imports, followed by TWO trailing UNNAMED (blank) columns:
      Title, Section, Type, Priority, Preconditions, Steps, Expected Result,
      References, "", ""
  There are NO Filters-specific columns. Standing-Rule-8 traceability
  (internal FLT- id <-> TestRail Case ID) lives in
  build/filters/testrail-id-map.csv exactly as it does for the other
  projects; the import file itself is byte-format-identical to theirs.

CONTENT RULES enforced here (same as the other generators):
  1. VIU-word-free: internal `viu_status`, `notes`, `design_ref` are NOT emitted;
     any "(see FLT-...)" internal cross-refs are stripped.
  2. Feature-flag-free: no feature-flag phrasing survives (Filters has none in
     the reader-facing case text; a sanity check confirms 0 occurrences).
  3. Sections = leaf area names; API-related cases route to an "API — <leaf>"
     section (STANDING RULE 4) using the same em-dash convention as the other
     imports.
  4. References = spec reference only (no internal FLT- ids, no VIU text).
  5/6. Preconditions/Steps/Expected kept as authored (numbered, line-broken),
     cleaned per rule 1.

Also writes build/filters/testrail-id-map.csv (Standing Rule 8: all internal
ids, blank TestRail Case-ID column until a permitted push assigns C-ids).

Outputs (canonical location + naming, matching the other projects):
  testrail-import/filters-v1-testrail-import.csv
  testrail-import/filters-v1-testrail-import.xlsx
  build/filters/testrail-id-map.csv
"""
import csv, json, glob, os, re
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))          # build/filters/
BASE = os.path.dirname(HERE)                               # build/
ROOT = os.path.dirname(BASE)                               # repo root
CASES_DIR = os.path.join(HERE, "cases")
OUT_CSV = os.path.join(ROOT, "testrail-import", "filters-v1-testrail-import.csv")
OUT_XLSX = os.path.join(ROOT, "testrail-import", "filters-v1-testrail-import.xlsx")
OUT_IDMAP = os.path.join(HERE, "testrail-id-map.csv")

# Byte-identical (name + order) to the fees-discounts / simple-flow /
# global-search imports: 8 named columns + 2 trailing UNNAMED (blank) columns.
HEADER = [
    "Title", "Section", "Type", "Priority",
    "Preconditions", "Steps", "Expected Result", "References",
    "", "",
]

# Deterministic, tidy section order (functional first, API last).
SECTION_ORDER = [
    "Filter Bar Layout and Visibility",
    "Status Filter",
    "Customer Filter",
    "Lead Technician Filter",
    "Service Advisor Filter",
    "Asset on Site Filter",
    "Active Filter Chips and Clear Filters",
    "Collapse and Expand",
    "Empty State",
    "Tab Behaviour",
    "Persistence",
    "URL State and Shareable Links",
    "Mobile Filters",
    "Parts Page Filters",
    "Reports Page Filters",
    "Page Search (Command-K)",
    "API — Work Orders List Filtering",
]


def clean(s):
    """Strip internal authoring markers; keep functional content (mirrors the
    other project generators)."""
    if not s:
        return s
    s = re.sub(r"\s*\(see (?:FLT|GS|SF|FD)-[A-Z0-9-]+\)", "", s)
    s = re.sub(r"feature[ -]flags?", "Filters feature", s, flags=re.I)
    return s


def joinlines(lst):
    return "\n".join(clean(x.rstrip()) for x in (lst or []))


def load_cases():
    cases = []
    for f in sorted(glob.glob(os.path.join(CASES_DIR, "cases-*.json"))):
        cases += json.load(open(f))
    return cases


def section_for(c):
    """Leaf area name; API-related cases route to an 'API — <leaf>' section
    (STANDING RULE 4), using the same em-dash convention as the other imports."""
    area = c["area"].strip()
    if c.get("api_related"):
        leaf = re.sub(r"^API\s*[—-]\s*", "", area).strip()
        return "API — " + leaf
    return area


def main():
    cases = load_cases()

    order = {s: i for i, s in enumerate(SECTION_ORDER)}
    cases.sort(key=lambda c: (order.get(section_for(c), 999), c["id"]))

    rows = []
    titles = []
    ids = []                 # parallel internal-id list (for dupe/empty checks only)
    idmap_rows = []
    api_sections = set()
    api_moved = 0
    for c in cases:
        title = clean(c["title"].strip())
        titles.append(title)
        ids.append(c["id"])
        section = section_for(c)
        if c.get("api_related"):
            api_sections.add(section)
            api_moved += 1
        idmap_rows.append([c["id"], "", title, section])
        rows.append([
            title,
            section,
            c.get("type", "Functional"),
            c["priority"].strip(),
            joinlines(c.get("preconditions")),
            joinlines(c.get("steps")),
            joinlines(c.get("expected")),
            clean((c.get("spec_ref") or "").strip()),
            "",
            "",
        ])

    with open(OUT_CSV, "w", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(HEADER)
        w.writerows(rows)
    print("Wrote CSV:", OUT_CSV, "rows(data):", len(rows))

    with open(OUT_IDMAP, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["internal_id", "testrail_case_id", "title", "section"])
        w.writerows(idmap_rows)
    print("Wrote ID map:", OUT_IDMAP, "rows:", len(idmap_rows))

    # --- Sanity checks (must be zero / clean) ---
    dupes = [t for t, n in Counter(titles).items() if n > 1]
    print("Duplicate titles:", dupes if dupes else "NONE")
    dupids = [i for i, n in Counter(ids).items() if n > 1]
    print("Duplicate internal ids:", dupids if dupids else "NONE")
    blob = "\n".join("\t".join(str(x) for x in r) for r in rows).lower()
    print("VIU occurrences:", blob.count("viu"))
    print("'feature flag' occurrences:", blob.count("feature flag"))
    print("'flag on' occurrences:", blob.count("flag on"))
    print("'flag off' occurrences:", blob.count("flag off"))
    print("API sections created:", sorted(api_sections), "| API cases:", api_moved)
    empties = [iid for iid, r in zip(ids, rows)
               if not (r[4].strip() and r[5].strip() and r[6].strip())]
    print("Rows missing Preconditions/Steps/Expected:", empties if empties else "NONE")

    # --- xlsx review copy ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not available - skipped XLSX.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Filters V1"
    ws.append(HEADER)
    for r in rows:
        ws.append(r)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")

    widths = {"Title": 50, "Section": 40, "Type": 12, "Priority": 10,
              "Preconditions": 60, "Steps": 60, "Expected Result": 60,
              "References": 34}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, len(rows) + 2):
        for cidx in range(1, len(HEADER) + 1):
            ws.cell(row=r, column=cidx).alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)
    print("Wrote XLSX:", OUT_XLSX)


if __name__ == "__main__":
    main()
