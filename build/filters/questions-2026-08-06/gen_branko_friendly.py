#!/usr/bin/env python3
"""Friendly, forward-as-is version of the 2026-08-06 Branko Cicovic question sheet.

SAME SUBSTANCE, EASIER READING, ONE ITEM REMOVED ON PURPOSE.

Produced on the QA lead's instruction: "Give me the friendly and easy to read and
understandable files for Chris and Branko."

WHAT CHANGED
  * a short warm opening note - who it is for, what it covers, how long it takes,
    that a letter or one line is a fine answer, and which section actually blocks us
  * REORDERED BY WHAT TO DO FIRST: five questions that release stuck tests, then
    eight ordinary decisions, then six engineering-plan-only items nothing is
    waiting on, then one heads-up with nothing to decide. The headings say so.
  * shorter sentences throughout

ONE ITEM WAS REMOVED FROM WHAT HE READS - the exact shape of the Reports page web
address (the date-range URL contract). HIS OWN SPECIFICATION ALREADY STATES IT,
verbatim, in the Reports Filters part of the scope section:
    "The selected range is reflected in the URL (e.g.,
     range=custom&from=2026-04-01&to=2026-04-25) so a filtered report is shareable"
Confirmed LIVE for this pass: Confluence page 572030978, VERSION 19, HTTP 200,
version published 2026-08-06T11:48:47Z - the sentence is present in the live body.
Asking him would have re-asked something his own document answers, which is the
exact embarrassment Standing Rule 55 exists to prevent and which this project has
already had once. Vlad was right on that row; our recorded verdict in
build/filters/vlad-gap-review-2026-08-06/ROW-BY-ROW.md needs correcting - FLAGGED
for the QA lead, NOT rewritten here (that file is out of scope for this pass).

UNCHANGED AND NON-NEGOTIABLE (Standing Rules 7 + 55): every row names the PROJECT
- Filters or Schedule - because Branko owns Filters, Schedule AND Global Search,
so a row read alone, days later, on a phone must still be unambiguous. Nothing he
reads carries a case ID, a requirement anchor, an HTTP term, an endpoint or the
word VIU. The question-to-case mapping stays on the QA-only tab, imported verbatim
from gen_branko_sheet.py so it cannot drift.

RESEARCH ONLY - writes two files into this folder. No TestRail write, no Jira
call, no application call.
"""

import importlib.util
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
PRIOR = os.path.join(HERE, "gen_branko_sheet.py")

XLSX = os.path.join(
    HERE, "Questions-for-Branko-Cicovic_Filters-and-Schedule_Friendly-Version_2026-08-06.xlsx")
MD = os.path.join(
    HERE, "Questions-for-Branko-Cicovic_Filters-and-Schedule_Friendly-Version_2026-08-06.md")

_spec = importlib.util.spec_from_file_location("prior_branko_f", PRIOR)
prior = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(prior)

HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Which project and screen", "What happens now", "The question", "Options",
        "Your answer"]

TAB1_NAME = "1 Start here"
TAB2_NAME = "2 Decisions"
TAB3_NAME = "3 When you have time"
TAB4_NAME = "4 Just a heads-up"
TAB5_NAME = "QA internal - not for Branko"

HELLO = (
    "Hello Branko - this is everything we have open across TWO of your projects, FILTERS and "
    "SCHEDULE, gathered into one place so you can go through it in a single sitting instead of "
    "getting a trickle of separate messages. Twenty-two items; about twenty minutes if you go "
    "straight "
    "down the list. SHORT ANSWERS ARE PERFECT - a letter, or one line. Nothing here needs an "
    "essay.\n\n"
    "WHERE TO START. Section 1 is five questions that release tests which are stuck today - that "
    "is the part we are genuinely waiting on. Section 2 is nine ordinary decisions. Section 3 is "
    "seven things only the engineering plan describes, and nothing of ours is waiting on those, so "
    "they can keep for a quiet moment. Section 4 is one typo-level heads-up with nothing to "
    "decide.\n\n"
    "ONE OF THESE IS OUR OWN FAULT and we are sorry: question 2 in Section 1 was written on "
    "22 July and we never actually sent it to you. Two tests have been parked ever since waiting "
    "for an answer you were never asked for.\n\n"
    "Every question says which project it belongs to, because you look after Filters, Schedule and "
    "Global Search. And to be clear - we have not edited any of your tickets or your descriptions. "
    "Where two of your own documents disagree we simply say so and ask which one to keep."
)

# ================================================================= SECTION 1
SEC1 = [
    (
        "FILTERS - the Work Orders list - the Status button on the Estimates and Completed tabs",

        "Four of our tests are on hold on this one point, and two answers are on record that "
        "disagree with each other.\n\n"
        "The Work Orders list has tabs across the top. Two of them - Estimates and Completed - "
        "already show only one kind of work order. There is also a row of filter buttons below, and "
        "one of them is Status.\n\n"
        "Your written description says the Status button is NOT SHOWN AT ALL on those two tabs. It "
        "has said that since 14 May and it still says it today.\n\n"
        "You told us on 17 July that the Status button IS SHOWN, greyed out, already filled in with "
        "that tab's own status, and cannot be changed. Our QA lead agreed with that on 30 July, and "
        "the design shows it that way too.\n\n"
        "Why we are asking rather than choosing: we have put the four tests back to your July "
        "answer, because that is what you and our QA lead actually decided - but the product "
        "currently behaves the way the written description says. So one of the three has to change, "
        "and it is your call which.",

        "Which is right - is the Status button hidden on the Estimates and Completed tabs, or shown "
        "greyed out and already filled in?",

        "A) NOT SHOWN AT ALL on those two tabs - the written description is right, and my July "
        "answer is out of date.\n\n"
        "B) SHOWN, GREYED OUT AND ALREADY FILLED IN - my July answer stands, and the description "
        "needs correcting. (Then we will also raise it so the product can be fixed.)\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the technician calendar - planning a job across several days, and days the shop "
        "is closed",

        "FIRST, AN APOLOGY: we wrote this question on 22 July and never sent it. Two tests have "
        "been parked ever since. That delay is ours, not yours.\n\n"
        "When a job is too big for one day, the schedule spreads it across several working days. "
        "Your description tells us two opposite things about days the shop is closed - a public "
        "holiday, or an inventory day - and both sentences are still there today.\n\n"
        "One says closures and public holidays are NOT skipped in this first version: the spread "
        "puts shifts on a closed day like any other day.\n\n"
        "The other says shop closures are set at shop level and BLOCK the spread from placing "
        "shifts on those days: the spread jumps over them.\n\n"
        "Why we are asking: those are two completely different tests and we cannot write both. We "
        "have not guessed - the two tests say plainly that the point is undecided and are waiting.",

        "When a job is spread across several days, should the schedule skip days the shop is "
        "closed, or place shifts on them anyway?",

        "A) SKIP THEM - closed days are jumped over, the way weekends already are, and the job runs "
        "on to the next open day.\n\n"
        "B) DO NOT SKIP THEM - closed days get shifts like any other day for this first version, "
        "and somebody moves them by hand if needed.",
    ),
    (
        "SCHEDULE - the search box in the toolbar above the calendar (not the one in the job list "
        "down the left)",

        "THIS ONE DECIDES WHETHER ONE OF OUR TESTS IS RIGHT OR WRONG, so it is worth a minute.\n\n"
        "Your description used to say that when someone searches, the jobs that do not match go "
        "FADED BUT STAY ON SCREEN, so you keep sight of the whole week. That sentence was taken out "
        "THIS MORNING, after your team decided the description was wrong and that the drawing shows "
        "only the matching jobs. The description now says only WHAT the search looks through - "
        "customer name, work order number, unit number, technician name and line name. It says "
        "nothing at all about the jobs that do not match.\n\n"
        "Our test still says the non-matching jobs go faded, because that is what was written down "
        "when we wrote it. We will not quietly change it to match what the software does today - "
        "that would tell us what was built rather than what you wanted.\n\n"
        "One more thing in the same place: the developer ticket for the calendar layout STILL says "
        "the non-matching jobs should fade, in two places. So your description and that ticket now "
        "disagree. If your answer is A, somebody should tidy that ticket up - we have not touched "
        "it, because it is not ours.",

        "When someone searches, what should happen to the jobs that do not match?",

        "A) They should disappear from the calendar - only the matching jobs are shown. (This is "
        "what your team said the drawing shows.)\n\n"
        "B) They should stay on screen but faded, and the matching ones stand out.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the pop-up window that opens when you click a scheduled job - the estimated "
        "hours",

        "THIS ONE DECIDES WHETHER A TEST PASSES OR FAILS, which is exactly why we are asking "
        "instead of choosing.\n\n"
        "Earlier today you told us the little ESTIMATE BADGE should not be clickable, and that the "
        "time is changed in the fields higher up the window instead. That makes sense to us.\n\n"
        "What we cannot tell is how far your answer reaches. Your description STILL says the window "
        "should let someone type a new estimate straight into it, and the developer ticket says the "
        "same. Both are live today. So your sentence might mean the estimate cannot be changed "
        "anywhere in that window, or only that the small badge on the job line should not be "
        "clickable.\n\n"
        "We have one test that says the estimate CAN be typed into. If you mean the first, that "
        "test is wrong and we will correct it. If you mean the second, the test is right and the "
        "software has something to fix. We are deliberately not settling it by looking at what the "
        "software does today.",

        "In that pop-up window, should someone be able to change the ESTIMATED HOURS by typing into "
        "them?",

        "A) NO - the estimate cannot be changed in that window at all; only the start and end times "
        "can be changed, in the fields above.\n\n"
        "B) YES - the estimate itself can still be typed into; only the little badge on the job "
        "line should not be clickable.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - which drawing of the Schedule we should be working from",

        "This one is about which picture is the real one, and it affects roughly fifty of our "
        "tests.\n\n"
        "Back in July you told us the Schedule prototype we were given was the one to work from, "
        "and we pinned about fifty on-screen names and labels from it - button wording, column "
        "headings, the words used in warnings.\n\n"
        "Three Schedule faults raised on 5 August all point at a DIFFERENT drawing - a shared link "
        "to a live, editable design page with no version and no date on it. Because it can change "
        "at any moment and has nothing to say when it was finished, we cannot compare our tests "
        "against it, and we cannot tell whether it is newer or older than the one we hold.\n\n"
        "You have noticed something similar yourself: on one of those three you replied that the "
        "button being reported is not in the design, and asked where it had been found.\n\n"
        "Why we are asking: if that newer drawing is the finished one, about fifty of our labels "
        "may be out of date and we should go through it properly. If it is not finished, we should "
        "carry on from the prototype and leave it alone. We are not going to guess between two "
        "pictures.",

        "Which drawing of the Schedule is the one to work from - the prototype you pointed us at in "
        "July, or the newer shared design page?",

        "A) THE PROTOTYPE from July is still the one. The newer page is a work in progress and we "
        "should ignore it.\n\n"
        "B) THE NEWER SHARED PAGE is the finished one - please confirm it is final, and we will go "
        "through it and update whatever has changed.\n\n"
        "C) Neither is final yet - please say when a finished drawing will be available.",
    ),
]

# ================================================================= SECTION 2
SEC2 = [
    (
        "FILTERS - the Work Orders list - where the filter bar sits",

        "Your description says the filter bar sits BELOW the row of tabs (All, Estimates, "
        "Completed, My Work Orders), and the design shows the same. In the product the five filter "
        "buttons sit ON THE SAME ROW as the tabs.\n\n"
        "One of our own tests used to wave this away with a note saying the product behaves this "
        "way on purpose for now - and nothing anywhere backed that up. That note was wrong and has "
        "been removed. Our test now expects what your description says.\n\n"
        "Since then a developer job has been raised to move the bar below the tabs, on the grounds "
        "that the product does not match your description. So somebody is about to change the "
        "product on the strength of that reading.\n\n"
        "Why we are asking now: if you actually wanted them on one row, that job should be "
        "cancelled and your description updated instead. Better to ask before the change is made "
        "than after.",

        "Should the filter buttons be moved below the tabs, or did you want them on the same row as "
        "the tabs?",

        "A) MOVE THEM BELOW - as your description and the design say. The developer job is correct "
        "and nothing needs changing in writing.\n\n"
        "B) SAME ROW IS WHAT I WANTED - then the developer job should be cancelled and the "
        "description updated to say so.",
    ),
    (
        "FILTERS - the Work Orders list on a phone - the Imported choice",

        "Imported sits in the Status list but behaves differently: while it is chosen, the other "
        "filters cannot be used. That much is written down, and we have added a test for it on a "
        "phone.\n\n"
        "There is a second behaviour that is not written down anywhere. We are told the product "
        "also does the reverse: if you pick an ordinary status last, Imported is quietly un-picked "
        "for you.\n\n"
        "Why we are asking rather than just testing it: that behaviour exists only in the "
        "developers' own code checks. We do not turn something the code happens to do into "
        "something the product must do - that has to be your decision, or it stops being a test of "
        "the product and becomes a description of it.",

        "Is that reverse behaviour intended - picking an ordinary status last automatically "
        "un-picks Imported?",

        "A) Yes - that is intended. We will test it, and it should be written down.\n\n"
        "B) No - that is not intended. (Then we will raise it.)\n\n"
        "C) Something else - please describe what should happen.",
    ),
    (
        "FILTERS - the Work Orders list on a phone - the wording on the apply button",

        "A tiny one, and it exists only because our tests have to quote the exact words a tester "
        "will see on screen.\n\n"
        "Your description calls the button \"Apply filters\", with a small f. On a phone the button "
        "actually reads \"Apply Filters\", with a capital F.\n\n"
        "Why we are asking: we would rather your description and the screen said the same thing "
        "than have our test quietly differ from one of them.",

        "Which spelling is the right one?",

        "A) \"Apply Filters\" with a capital F - the description can be tidied to match the "
        "screen.\n\n"
        "B) \"Apply filters\" with a small f - then the button on screen should be corrected.",
    ),
    (
        "FILTERS - the filter buttons on the Parts pages and the Reports pages - the write-up",

        "A GENTLE STATUS ASK, not a new question - we know this is already with you and we are not "
        "chasing.\n\n"
        "Eight of our tests cover filter buttons on the Parts pages and the Reports pages. They "
        "were written from the designs back in July, and they are parked because your write-up for "
        "them has not arrived.\n\n"
        "A CORRECTION WE OWE YOU: an earlier draft of this question also told you the feature was "
        "not built yet. That was our mistake. An engineering handover we were given today says the "
        "filter buttons on the eight Parts pages ARE built, and so are the ones on six reports - "
        "Shop Billing Efficiency, My Timesheets, Timesheet Activities, Notes, Reminders and Sales "
        "Tax - all waiting on one final code review.\n\n"
        "That same handover says several other reports are deliberately NOT being done in this "
        "piece of work. Our tests name a good many of those, which is a separate tidy-up on our "
        "side and not something you need to answer here.\n\n"
        "Why we are asking: so we can finish these tests against something you have written, rather "
        "than leaving them resting on a drawing alone - and so we can tell our own management "
        "honestly whether this is weeks away or months.",

        "Roughly when do you expect the Parts and Reports write-up, and is that part of the product "
        "still planned for this release?",

        "A) It is coming shortly, and it is still in this release.\n\n"
        "B) It has moved to a later release - please say roughly when.\n\n"
        "C) It has been dropped. (Then we will ask about deleting the eight tests.)",
    ),
    (
        "SCHEDULE - the menu that opens on an empty part of the calendar",

        "Nobody had spotted this one, and no test is wrong because of it - but your description "
        "contradicts itself, so a reader who happens to read the wrong half would test the wrong "
        "thing.\n\n"
        "In two places it says a LEFT-CLICK on empty calendar space opens a menu offering Create "
        "event and New work order. In two other places, where it lists what each access level "
        "unlocks, it twice calls the same thing a RIGHT-CLICK menu.\n\n"
        "Our tests follow left-click and the product agrees with them, so nothing is broken "
        "today.\n\n"
        "Why we are asking: it is a one-word correction, and until it is made anyone reading only "
        "the access-level section will look for the wrong mouse click.",

        "Which is correct - does that menu open on a left-click or a right-click?",

        "A) LEFT-CLICK - as the two earlier places say. The access-level section is the wording "
        "that needs correcting.\n\n"
        "B) RIGHT-CLICK - then the product and our tests are both wrong and it becomes a developer "
        "job.",
    ),
    (
        "SCHEDULE - weekends for a technician who has no working hours set up",

        "A gap rather than a contradiction, and no test is parked on it - we are asking so that we "
        "are not quietly relying on our own reading.\n\n"
        "Three parts of your description do not quite add up for a technician with no working hours "
        "of their own:\n"
        "- the default working day is 7:00 AM to 7:00 PM, with nothing said about which days of the "
        "week that covers;\n"
        "- spreading a job automatically skips weekends;\n"
        "- but a shift placed on a weekend counts as a clash to be warned about.\n\n"
        "So we cannot tell whether a weekend is simply an ordinary working day for such a "
        "technician, or a day the schedule should push back on.\n\n"
        "Why it matters: it decides whether a warning should appear, and a warning that should not "
        "be there is just as much a bug as a missing one.",

        "For a technician with no working hours of their own set up, is a weekend a normal working "
        "day or a day the schedule should warn about?",

        "A) A DAY TO WARN ABOUT - weekends are outside normal hours for everyone unless someone has "
        "set weekend hours for them.\n\n"
        "B) A NORMAL WORKING DAY - the default 7 to 7 applies to all seven days, and only the "
        "spread step avoids weekends.",
    ),
    (
        "SCHEDULE - the list of jobs INSIDE that same pop-up window",

        "This one is small, and it is only two of your own documents disagreeing.\n\n"
        "Your description was changed this morning so each job line in that pop-up shows an "
        "estimate figure and a status label. In your own words earlier today: the lines should show "
        "the estimate and the status badge and there should not be any totals. Our test already "
        "expects exactly that - the hours and a status label, no money anywhere in the window - so "
        "nothing of ours is stuck.\n\n"
        "But the developer ticket for that window still says each line shows a labour TOTAL, and it "
        "has said so since before your change. You edited that same ticket three days ago and the "
        "word is still in it. So your description and that ticket now say different things, and "
        "only you can say which to keep.\n\n"
        "While you are in that ticket: it also still says the estimate can be typed in, which is "
        "question 4 in Section 1. Both would be tidied in the same visit.",

        "On each job line in that pop-up, should there be a money total, or only the hours and a "
        "status label?",

        "A) ONLY the hours and a status label - no money anywhere. Your description is right and "
        "the developer ticket needs tidying up.\n\n"
        "B) A MONEY TOTAL should be shown there - the developer ticket is right and your "
        "description needs correcting. (Then our test is wrong and we will change it.)\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - how much of the day the timeline shows when the day view opens",

        "Nothing is stuck on this one and no test of ours is wrong today. We are asking because two "
        "of your own documents now point in opposite directions.\n\n"
        "Your description says the day view keeps the WHOLE 24 HOURS there and scrollable, and "
        "simply scrolls itself so the start of the working day is on the left. That is still what "
        "it says today.\n\n"
        "The design review of 5 August asks for something different: that the timeline show ONLY "
        "THE WORKING HOURS plus a little after them, with anything outside reached by scrolling. "
        "That review lists it as IN SCOPE for this release, alongside the change that makes the day "
        "view open at the start of the working day.\n\n"
        "A CORRECTION WE OWE YOU: an earlier draft of this question told you that the narrower "
        "version was only a later improvement. That was our mistake - the review puts it in this "
        "release. We would rather correct it than have you answer on the strength of it.\n\n"
        "So your description was last changed on 7 August, two days AFTER that review, and it still "
        "says the full 24 hours. Only you can say which of the two you meant to stand. If the "
        "narrower version is meant for this release, your description needs changing first and then "
        "we will change the test to match.",

        "For THIS release, which is right?",

        "A) Keep the full 24 hours, as your description says today - the narrower version is for "
        "later.\n\n"
        "B) Change it now to show only the working hours plus a little after them.\n\n"
        "C) Something else - please describe it.",
    ),
    # ADDED 2026-08-11 (Schedule follow-up push, item 1) - the Status filter multi-select
    # question. It belongs in SECTION 2 rather than Section 1 because nothing of ours is
    # BLOCKED: the test still runs, it simply no longer asserts what happens with two
    # statuses chosen. Framed as an ordinary decision, not as a failure of his (Rule 55).
    (
        "SCHEDULE - the technician calendar - the Status filter on the work order list, when you "
        "want more than one status at a time",

        "Nothing is stuck on this one and no test of ours is wrong today. We are asking because we "
        "found one of our own tests claiming something your description does not say, and we have "
        "taken the claim out rather than leave it in.\n\n"
        "On the Schedule page there is a Filter panel, and one of its groups is Status. Your "
        "description lists what goes in that group - \"all work order statuses currently supported "
        "in the app\" - and it says that applying a filter narrows the list of work order cards.\n\n"
        "What it does not say, anywhere, is whether you can pick MORE THAN ONE status at the same "
        "time - for example Approved and Review together - and what the list should then show.\n\n"
        "Being straight with you about what we did: one of our tests had been claiming that "
        "choosing several statuses shows the work orders of all of them together. We could not find "
        "that in your description, in the story, in the design or in anything you have told us, so "
        "we have removed the claim. We have NOT replaced it with the opposite - we are not saying "
        "you can only pick one - because we do not know, and guessing either way would put words in "
        "your mouth.\n\n"
        "So the test now checks one status at a time, which is safe and true either way. If you "
        "tell us more than one is intended, we will add that back as a proper test and note that it "
        "came from you.",

        "Can more than one status be chosen in the Status filter at the same time, and if so what "
        "should the list show?",

        "A) YES - you can pick several statuses, and the list shows the work orders of ALL the "
        "chosen statuses together.\n\n"
        "B) NO - only one status at a time; picking another replaces the first.\n\n"
        "C) Something else - please describe it.",
    ),
]

# ================================================================= SECTION 3 + 4
# ADDED 2026-08-11 - a FILTERS documentation gap, appended as item 7 of section 3.
# It belongs here because section 3 is precisely "behaviours only the engineering plan
# describes". It differs from the other six in one way, and the row says so plainly: the
# other six are candidate coverage awaiting his word, whereas this one is ALREADY covered
# by two of our tests (C29600 and C29632), correctly, because the engineering technical
# design is an authoritative source under Standing Rule 57(d3). So it is framed as a
# DOCUMENTATION GAP - his own technical design states the rule and his product description
# does not - and not as a decision he has failed to make (Standing Rule 55).
FILTERS_CROSS_FILTER_DOCGAP = (
    "FILTERS - the Work Orders list - using two different filter buttons at the same time",

    "NOTHING OF OURS IS STUCK ON THIS ONE and we are not asking you to make a decision. We are "
    "asking you to write down in your own description something your engineers have already "
    "written down in theirs.\n\n"
    "The Work Orders list has five filter buttons: Status, Customer, Lead Technician, Service "
    "Advisor and Asset on Site.\n\n"
    "Your description says what each button does ON ITS OWN. For Status it says the list shows "
    "work orders matching ANY of the statuses you tick. For Customer it says the list shows work "
    "orders belonging to ANY of the customers you pick.\n\n"
    "What it never says is what the list should show when someone uses TWO DIFFERENT BUTTONS AT "
    "THE SAME TIME - for example ticking the status 'Estimate' and also picking the customer "
    "'Smith'. We searched the whole of your current description and five earlier versions of it, "
    "and the rule is not there in any of them.\n\n"
    "Your engineers' own working notes DO state it: the buttons must narrow together, so the list "
    "shows only the work orders that match both. That is also what the product does today, and it "
    "is what two of our tests already check - so nothing is broken and nothing is waiting.\n\n"
    "Why we are raising it anyway: a rule that lives only in an engineering note is one edit away "
    "from being changed by accident, and nobody would notice. One sentence in your description "
    "settles it for good.",

    "When someone ticks a status AND also picks a customer, what should the list show?",

    "A) ONLY the work orders that match BOTH - status 'Estimate' AND customer 'Smith'. (This is "
    "what your engineers' notes say and what the product does today. If you pick A, please add a "
    "sentence saying so to your description - nothing of ours needs to change.)\n\n"
    "B) Something else - please describe it. (Then two of our tests are wrong and we will correct "
    "them.)",
)

SEC3 = list(prior.PRIOR_SCHED_ENG) + [FILTERS_CROSS_FILTER_DOCGAP]   # 6 carried + 1 new
SEC4 = [prior.PRIOR_FILTERS[2]]             # the pointer heads-up, carried over verbatim

EXTRA_NOTES = [
    "THIS IS THE FRIENDLY, FORWARD-AS-IS VERSION of "
    "Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-06.xlsx, produced on the QA lead's "
    "instruction: \"Give me the friendly and easy to read and understandable files for Chris and "
    "Branko.\" 20 of the 21 items are carried over with their substance unchanged. What changed: a "
    "short warm opening note, REORDERING BY WHAT TO DO FIRST (five items that release stuck tests, "
    "then nine ordinary decisions, then the six engineering-plan-only items nothing is waiting on, "
    "then one heads-up with nothing to decide) with the headings and tab names saying so, and "
    "shorter sentences. The six engineering-only items and the pointer heads-up are IMPORTED "
    "verbatim from gen_branko_sheet.py so their wording cannot drift.",

    "ONE ITEM WAS REMOVED FROM THE READER-FACING SHEET, AND THIS IS WHY. The removed item asked "
    "whether the EXACT SHAPE OF THE REPORTS PAGE WEB ADDRESS (the date-range URL contract) is "
    "something we should be testing. HIS OWN SPECIFICATION ALREADY STATES IT, verbatim, in the "
    "\"Reports Filters\" part of the scope section: \"The selected range is reflected in the URL "
    "(e.g., range=custom&from=2026-04-01&to=2026-04-25) so a filtered report is shareable\". "
    "VERIFIED LIVE FOR THIS PASS - Confluence page 572030978, VERSION 19, HTTP 200, published "
    "2026-08-06T11:48:47Z; the sentence is present in the live body (one occurrence). Sending it "
    "would have re-asked a question his own document answers - the exact embarrassment Rule 55 "
    "exists to prevent, and one this project has already had once. SO THE ROW IS OFF THE SHEET and "
    "the sheet is 20 items, not 21. [SUPERSEDED 2026-08-11 ON THE COUNT ONLY - the removal itself "
    "stands and the reasoning above is unchanged; the sheet is 21 items again because a NEW item "
    "was added on 2026-08-11. See the next note.]",

    "⚠️ ADDED 2026-08-11 - ONE NEW ITEM, TAKING THE SHEET FROM 20 BACK TO 21: the FILTERS "
    "CROSS-FILTER DOCUMENTATION GAP, appended as item 7 of Section 3. WHY IT IS IN SECTION 3: that "
    "section is exactly \"behaviours only the engineering plan describes\", and this is one - the "
    "rule that two DIFFERENT filter buttons narrow together is in the engineering technical design "
    "and in no version of his product description. WHY IT IS FRAMED DIFFERENTLY FROM THE OTHER "
    "SIX: the other six are candidate coverage awaiting his word, whereas this rule is already "
    "correctly covered by C29600 and C29632, because the technical design is an authoritative "
    "source under Standing Rule 57(d3). So it is put to him as a DOCUMENTATION GAP - his own "
    "engineers state it, his own description does not - and expressly NOT as a decision he has "
    "failed to make (Rule 55). NOTHING OF OURS IS BLOCKED ON IT and the row says so on its face, "
    "so he is not misled into thinking it is urgent. DUPLICATE CHECK RUN BEFORE ADDING IT (Rule "
    "55): both markdown files, the README and EVERY XML part of BOTH workbooks were searched for "
    "29600, 29632, 'multi-criteria', 'matching both' and 'two different filter' - ZERO hits in "
    "all of them, so it duplicates nothing. This also CORRECTS our own record: "
    "build/unsourced-cases-2026-08-11/CANDIDATES.md states C29600 is already on Branko's sheet; "
    "IT WAS NOT. Source: build/filters/c29600-sourcing-2026-08-11/FINDINGS.md; the accompanying "
    "TestRail recording fix is build/filters/c29600-fix-2026-08-11/.",

    "NOTHING WAS SENT TO BRANKO. The QA lead's standing instruction is that nothing goes to a PO "
    "until our own work is done, so this remains a DRAFT. No Jira issue was created or edited by "
    "the 2026-08-11 pass either - the creation hold at the tail of Standing Rule 62 stands.",

    "⚠️ CORRECTION OWED TO OUR OWN RECORD, FLAGGED NOT REWRITTEN. The earlier sheet's QA tab says "
    "of that item: \"Filters v19 section 4 and S11-R1 say nothing about the address format\". That "
    "is true of those two places and MISSES THE SCOPE-SECTION BULLET quoted above, which does state "
    "the shape. VLAD WAS RIGHT ON THAT ROW (row 8 of his eleven-row Filters coverage-gap table), "
    "and our recorded verdict in build/filters/vlad-gap-review-2026-08-06/ROW-BY-ROW.md NEEDS "
    "CORRECTING. It was deliberately NOT edited by this pass - that file is out of scope here, and "
    "silently rewriting a recorded verdict is not ours to do (Rules 33 and 44). THE QA LEAD OWES "
    "THAT ONE-ROW CORRECTION. There is also a real coverage consequence: the shape IS documented, "
    "so it is now authorable as a test rather than an open product question - which needs "
    "authorisation, and nothing was authored.",

    "THE EARLIER WORKBOOK IS SUPERSEDED, NOT DELETED. Both pairs sit in this folder. Send THIS one; "
    "the earlier pair is kept as the record of what was verified and when. NOTE THAT THE EARLIER "
    "PAIR STILL CARRIES THE URL QUESTION as its Filters item 3 - so if an old file goes out by "
    "mistake, that question goes with it. The rows and notes below are IMPORTED from "
    "gen_branko_sheet.py at build time so the mapping cannot drift; the URL item's own mapping row "
    "is carried below re-labelled as REMOVED, rather than deleted, so the decision stays auditable.",

    "STILL OWED FROM THE EARLIER PASS AND NOT DONE HERE (both out of scope for this task): the "
    "2026-08-05 Branko workbook in build/branko-questions-2026-08-05/ should be marked SUPERSEDED "
    "so an old one cannot go out by mistake; and CLAUDE.md plus "
    "build/OUTSTANDING-ITEMS-REGISTER.md still name the Schedule specification at version 23 when "
    "it is at 25.",

    "NO OTHER SOURCE WAS RE-FETCHED AND NONE NEEDED TO BE. The content was verified live earlier on "
    "2026-08-06 (Filters specification version 19, Schedule specification version 25) and this pass "
    "only rewrote the presentation and settled the one URL question above. The single live call made "
    "here was the read-only Confluence GET of page 572030978. NO TestRail write, no Jira call, no "
    "application call, and no case, ticket or specification was edited by anyone in this pass.",
]


def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    gc.alignment = WRAP
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws


TR = "https://shopview.testrail.io/index.php?/cases/view/"

# ADDED 2026-08-11 - QA-only mapping for the new section-3 item 7. Rule 8: internal ID +
# C-id + link. NOT FOR BRANKO.
QA_ROW_CROSS_FILTER = (
    "Tab 3", "7",
    "Filters - what the list should show when two DIFFERENT filter buttons are used at once "
    "(for example status Estimate + customer Smith)",

    "NEW 2026-08-11, from build/filters/c29600-sourcing-2026-08-11/FINDINGS.md. NOT a blocked "
    "decision and NOT an undecided product question - a DOCUMENTATION GAP. The rule IS decided "
    "and IS written down, in the engineering technical design; it is his own product description "
    "that is silent. Framed to him accordingly (Rule 55).",

    "NO case is held on this. TWO cases assert the rule and both are CORRECT as written - "
    "FLT-COMBO-01 (C29600, AUTOMATION: READY, TestRail custom_atmstatus = 3 AUTOMATED) and "
    "FLT-API-08 (C29632, AUTOMATION: READY, custom_atmstatus = 1). Both had their refs and "
    "provenance corrected on 2026-08-11 to cite the technical design; NEITHER had any assertion, "
    "step or expected outcome changed.",

    " · ".join(f"C{i} {TR}{i}" for i in [29600, 29632]),

    "VERIFIED 2026-08-11 against Filters Confluence page 572030978 at VERSION 19 (HTTP 200), and "
    "against versions 4, 12, 17 and 18. The rule for combining two DIFFERENT filters is absent "
    "from all five. Boolean 'AND' appears EXACTLY ONCE in the whole specification, at S13-R10 - "
    "'Search and filters are additive (AND)' - which is search-versus-filters, not "
    "filter-versus-filter. S2-R2 and S3-R6 give OR *within* one filter ('matching any of the "
    "selected statuses' / 'belonging to any of the selected customers'). S8-R3, which C29600 and "
    "C29632 used to cite for this, is the EMPTY-STATE requirement and does not define the "
    "combination. All 14 epic stories (SV-8786 to SV-8799) were read live including comments: not "
    "one acceptance criterion states the result of combining two different filters. THE RULE IS "
    "STATED in the engineering technical design, build/filters/tech-plan-2026-07-29/"
    "TechPlan-AppWide-Filter-Redesign.md - §1.8 verbatim '...return the right WOs and AND across "
    "fields' (line 323) and §0.3 verbatim 'Same-field filters are grouped and OR'd ... repeated "
    "eq on one field is a de-facto IN' (line 218). That document is UNDATED internally and "
    "records 'Spec baseline: v1.3' (line 124), so it is a PARTIAL currency source.",

    "A -> nothing of ours changes; he adds one sentence to his description and the two cases' "
    "provenance is re-stamped to name it instead of the technical design. B -> both cases are "
    "wrong on their headline assertion, both must be corrected, and VLAD MUST BE TOLD because "
    "C29600 is flagged AUTOMATED in TestRail. Standing Rule 57's OPEN question - whether a "
    "technical design carries PRD-level authority on product behaviour, or whether Rule 30's "
    "'informs but never overrules' still holds - governs these two cases as well as the nine in "
    "class C-3 of build/unsourced-cases-2026-08-11/CANDIDATES.md, so eleven in total. That "
    "question is the QA lead's, not Branko's, and is NOT on this sheet.",
)


# ADDED 2026-08-11 (Schedule follow-up push) - QA-only mapping for the new Section 2 item.
# Rule 8: internal ID + C-id + link. NOT FOR BRANKO.
QA_ROW_STATUS_MULTISELECT = (
    "Tab 2", "9",
    "Schedule - whether more than one status can be chosen in the Status filter, and what the "
    "work order list should then show",

    "NEW 2026-08-11, from build/schedule/followup-push-2026-08-11/. This is a COVERAGE GAP we "
    "created by removing an unsourced assertion, not a blocked decision. Framed to him as an "
    "ordinary decision, and it says on its face that nothing of ours is stuck (Rule 55).",

    "NO case is held on this. ONE case is affected - SCH-FILT-03 (C29944, AUTOMATION: READY, "
    "TestRail custom_atmstatus = 1 NOT AUTOMATED at write time) - and it is CORRECT as it now "
    "stands. Its expected item 3, 'Choosing more than one status shows the work orders of all the "
    "chosen statuses together.', was REMOVED on 2026-08-11 and item 4 renumbered to 3. Its steps, "
    "preconditions, title, refs and provenance were NOT changed. The opposite was deliberately NOT "
    "asserted either: nothing claims multi-select is absent, because that is equally unsourced "
    "(Rules 25/42/57/58).",

    " · ".join(f"C{i} {TR}{i}" for i in [29944]),

    "VERIFIED LIVE 2026-08-11 before the removal. Schedule Confluence page 713031682 at VERSION 27 "
    "(HTTP 200, published 2026-08-07T15:01:20.801Z, 43,064 chars, body sha256 4c51fb72...), read at "
    "both the start and the end of the write window and byte-identical (Rule 59). §5.1 in full on "
    "this point is only: 'Status | All work order statuses currently supported in the app'. The "
    "words multi, multi-select, multiple, more than one, several, checkbox, one or more, combine, "
    "at once and simultane- appear in §5.1 ZERO times, in NONE of the 27 versions. Whole-document "
    "counts: 'multi-select' 0, 'multiple statuses' 0, 'more than one status' 0. The only 'Select "
    "multiple' in the specification is §4.3, the SCOPE PICKER turning line rows into checkboxes - "
    "a different feature. Story SV-8687 ('Work Order Sidebar & Mini Calendar', QA Complete) was "
    "read live and independently: 'multi' 0, 'more than one' 0; its filter sentence is 'Filter "
    "groups: Assignment (Assigned/Unassigned), Status (all WO statuses), Priority (High/Medium/"
    "Low)'. The case's own authoring note has said since it was written: 'Single vs multi-select "
    "within a group is not pinned - confirm live.' The claim was introduced by our own 2026-08-05 "
    "repair pass, whose audit had scored the case clean minutes earlier - Rule 58's failure mode "
    "exactly.",

    "A -> the assertion goes BACK, with a step that selects two statuses, cited to his answer with "
    "its date and file link (Rule 54), and the multi-status behaviour is covered again. B -> "
    "nothing changes; C29944 is already correct, and we may add a small negative check that picking "
    "a second status replaces the first. EITHER WAY VLAD NEED NOT BE TOLD for this case - C29944 is "
    "custom_atmstatus = 1, not flagged AUTOMATED (Rule 65). UNTIL HE ANSWERS, multi-status "
    "filtering on the Schedule work order list is UNCOVERED, and that is the honest cost of the "
    "removal - recorded rather than hidden.",
)


def _qa_rows():
    """Imported mapping, with the removed URL item re-labelled rather than dropped,
    and the 2026-08-11 cross-filter + status-multiselect rows appended."""
    out = []
    for row in prior.QA_ROWS:
        row = list(row)
        if "exact shape of the report page's web address" in str(row[2]):
            row[2] = "REMOVED FROM THE READER SHEET - " + row[2]
            row[3] = ("REMOVED 2026-08-06. His own specification already states the shape verbatim "
                      "in the Reports Filters scope bullet: \"The selected range is reflected in "
                      "the URL (e.g., range=custom&from=2026-04-01&to=2026-04-25) so a filtered "
                      "report is shareable\" - confirmed live, page 572030978 version 19, HTTP 200. "
                      "Asking it would re-ask a question his document answers (Rule 55). "
                      "ORIGINALLY: " + str(row[3]))
        out.append(row)
    out.append(list(QA_ROW_CROSS_FILTER))
    out.append(list(QA_ROW_STATUS_MULTISELECT))
    return out


def write_xlsx():
    wb = openpyxl.Workbook()
    w = [4, 40, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "Filters and Schedule - questions for Branko Cicovic - 2026-08-06",
           HELLO,
           "Section 1 of 4 - please start here. These five release tests that are stuck today.",
           SEC1, w, first=True)

    _sheet(wb, TAB2_NAME,
           "Section 2 - nine ordinary decisions",
           "Eight decisions, each a plain A or B. Nothing here is urgent this week, but each one "
           "settles a difference between two of your own documents or between a document and the "
           "screen. Short answers are perfect.",
           "Section 2 of 4 - eight decisions. Four are Filters, four are Schedule.",
           SEC2, w)

    _sheet(wb, TAB3_NAME,
           "Section 3 - seven things only the engineering plan describes",
           "NOTHING OF OURS IS WAITING ON THIS TAB, so it can keep for a quiet moment. These "
           "seven behaviours appear only in the engineering plan and nowhere in your own document. "
           "For the first six we are not going to turn an engineering note into something the "
           "product must do without your word - so each is either a decision for you or an item we "
           "close as out of scope. The seventh is different and says so: it is already covered by "
           "two of our tests and nothing is waiting on it - we are simply asking you to write the "
           "rule into your own description as well.",
           "Section 3 of 4 - seven engineering-plan-only items. No test is blocked by these.",
           SEC3, w)

    _sheet(wb, TAB4_NAME,
           "Section 4 - one heads-up, nothing to decide",
           "NO DECISION NEEDED. One typo-level slip in your own document, which has already cost "
           "one round of confusion. Noted for your next edit.",
           "Section 4 of 4 - one heads-up. Nothing to answer.",
           SEC4, w)

    ws5 = wb.create_sheet(TAB5_NAME)
    ws5["A1"] = ("QA-ONLY - INTERNAL - NOT FOR BRANKO. Do not send this tab. TestRail case IDs, "
                 "requirement anchors, live evidence and the corrections to our own records live "
                 "here so the tabs Branko reads stay plain (Standing Rules 7, 8 and 55).")
    ws5["A1"].font = Font(bold=True)
    ws5["A1"].alignment = WRAP
    r = 3
    ws5.cell(row=r, column=1,
             value="PER-ITEM MAPPING - INCLUDING THE ONE ITEM REMOVED").font = Font(bold=True)
    r += 2
    _hdr(ws5, r, ["Tab", "Item", "What it asks", "Where the question comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in _qa_rows():
        for j, v in enumerate(row, 1):
            ws5.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws5.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in EXTRA_NOTES + list(prior.QA_NOTES):
        ws5.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, wd in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 62, 62]):
        ws5.column_dimensions[col].width = wd

    wb.save(XLSX)
    return XLSX


def block(items):
    out = []
    for i, (topic, now, q, opts) in enumerate(items, 1):
        out.append(f"### {i}. {topic}\n")
        out.append("**What happens now**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
        out.append("**The question**\n")
        out.append(f"> {q}\n")
        out.append("**Options**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
        out.append("**Your answer:** _______________________________________________\n")
    return "\n".join(out)


def write_md():
    md = f"""# Questions for Branko Cicovic — Filters and Schedule — 2026-08-06

**Projects: Filters (epic SV-8785) and Schedule (epic SV-8685) · Product Owner: Branko Cicovic**

*This is the friendly, forward-as-is version. It carries 20 of the 21 items of
`Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-06.md`, reordered by what to do first and
rewritten to read easily on a phone. **One item was removed on purpose** — the exact shape of the
Reports page web address — because his own specification already states it verbatim; the reason and
the live evidence are on the QA-only tab. **One item was ADDED on 2026-08-11** — the Filters
cross-filter documentation gap, as item 7 of Section 3 — so the sheet is now **21 items**. The
spreadsheet twin is
`Questions-for-Branko-Cicovic_Filters-and-Schedule_Friendly-Version_2026-08-06.xlsx`; it carries a
QA-only tab that must not be forwarded.*

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

---

{HELLO}

---

## Section 1 — Start here: these five release tests that are stuck today

{block(SEC1)}

---

## Section 2 — Eight ordinary decisions, when you have a moment

Each one is a plain A or B. Four are Filters, four are Schedule.

{block(SEC2)}

---

## Section 3 — Seven things only the engineering plan describes

**Nothing of ours is waiting on this section**, so it can keep for a quiet moment. These seven
behaviours appear only in the engineering plan and nowhere in your own document. For the first six
we are not going to turn an engineering note into something the product must do without your word.
**The seventh is different and says so** — it is already covered by two of our tests and nothing is
waiting on it; we are simply asking you to write the rule into your own description as well.

{block(SEC3)}

---

## Section 4 — One heads-up, nothing to decide

{block(SEC4)}

---

## QA-only — not for Branko

The question-to-case mapping is on the spreadsheet's `{TAB5_NAME}` tab — every question's affected
TestRail case IDs with links, the requirement anchors quoted from the live pages, and what each
possible answer resolves to. It is imported verbatim from the earlier sheet's generator so the two
cannot drift, and it adds:

- **why the web-address item was removed**, with the verbatim sentence from his own specification
  and the live confirmation (page version 19, HTTP 200) — its mapping row is kept, re-labelled
  REMOVED rather than deleted;
- **⚠️ the correction owed to our own record**: Vlad was right on that row, and
  `build/filters/vlad-gap-review-2026-08-06/ROW-BY-ROW.md` needs a one-row correction. It was
  deliberately **not** rewritten here;
- that the earlier pair of files is **superseded, not deleted** — and that it still carries the
  removed question, so an old file must not go out by mistake.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
