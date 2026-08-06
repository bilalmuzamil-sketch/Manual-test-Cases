#!/usr/bin/env python3
"""Generate the Branko Cicovic question sheet of 2026-08-06 (.xlsx + .md).

WHY THIS IS A CONSOLIDATION, NOT A SECOND SHEET (Standing Rule 55).
build/branko-questions-2026-08-05/ holds a 13-item sheet that was written and is
READY TO SEND but WAS NEVER SENT (register row C4). Rule 55 says to sweep every
open ambiguity onto ONE sheet so a PO answers in a single sitting "rather than a
drip of separate asks" - and two unsent sheets to the same person is exactly
that drip. So this file CARRIES FORWARD every still-open item from the
2026-08-05 sheet and adds the items raised since, and the 2026-08-05 workbook
should be sent NOT AT ALL (it is superseded by this one).

Every carried-forward reader-facing row is IMPORTED from the 2026-08-05
generator rather than retyped, so the wording cannot drift in transit - the same
technique the 2026-08-04 Chris consolidation used.

MIRRORS 1:1 (Standing Rule 16) the established peers - same six reader columns
in the same order, same row layout (A1 title / A2 note / A4 header / A5 band /
items from row 6), same fills, fonts, freeze pane and QA-only tab.

Standing Rule 55: Branko owns THREE things - Filters, Schedule and Global Search
- so EVERY question row names its project and its screen.

RESEARCH ONLY - this script writes two files into this folder. It makes no
TestRail call, no Jira call and no application call of any kind.
"""

import importlib.util
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
PRIOR = os.path.join(ROOT, "branko-questions-2026-08-05", "gen_branko_sheet.py")

XLSX = os.path.join(HERE, "Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-06.xlsx")
MD = os.path.join(HERE, "Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-06.md")

# --- import the 2026-08-05 sheet's own reader-facing rows, verbatim -------------
_spec = importlib.util.spec_from_file_location("prior_branko", PRIOR)
prior = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(prior)

PRIOR_SCHED_DOC = prior.TAB1      # 3 items - his own Schedule document
PRIOR_SCHED_ENG = prior.TAB2      # 6 items - only the engineering plan describes them
PRIOR_FILTERS = prior.TAB3        # 4 items - Filters

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Filters"
TAB2_NAME = "Schedule - your document"
TAB3_NAME = "Schedule - engineering only"
TAB4_NAME = "QA internal - not for Branko"

THANKS = (
    "Nothing here is a complaint, and one of these is honestly our own fault for not sending it "
    "sooner. It covers TWO of your projects, so every question says which one it belongs to - "
    "Filters or Schedule - and the tabs are split the same way. It is gathered into one place so "
    "you can go through it in a single sitting rather than getting a trickle of separate messages."
)

# ------------------------------------------------------------- NEW Filters items
NEW_FILTERS = [
    (
        "Filters (the filter buttons on the Work Orders list) - the Status button on the Estimates "
        "and Completed tabs (the story about how filters behave on each tab, SV-8794, under epic "
        "SV-8785)",

        "PLEASE START HERE ON THIS TAB - four of our tests are on hold on this one point, and two "
        "answers are on record that disagree with each other.\n\n"
        "The Work Orders list has tabs across the top. Two of them - Estimates and Completed - "
        "already show you only one kind of work order. There is also a row of filter buttons below, "
        "and one of them is Status.\n\n"
        "Your written description says the Status button is NOT SHOWN AT ALL on those two tabs. It "
        "has said that since 14 May and it still says it in the version you published this morning.\n\n"
        "You told us on 17 July that the Status button IS SHOWN, greyed out, and already filled in "
        "with that tab's own status, and cannot be changed. Our QA lead agreed with that on 30 July, "
        "and the design drawing shows it that way too.\n\n"
        "Why we are asking rather than choosing: we have put the four tests back to your July "
        "answer, because that is what you and our QA lead actually decided - but the product "
        "currently behaves the way the written description says. So one of the three has to change "
        "and it is your call which.",

        "Which is right - is the Status button hidden on the Estimates and Completed tabs, or shown "
        "greyed out and already filled in?",

        "A) NOT SHOWN AT ALL on those two tabs - the written description is right, and my July "
        "answer is out of date.\n\n"
        "B) SHOWN, GREYED OUT AND ALREADY FILLED IN - my July answer stands, and the written "
        "description needs correcting. (Then we will also raise it so the product can be fixed.)\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "Filters (the filter buttons on the report pages) - the date filter and the page's web "
        "address (the story about sharing a filtered view by link, SV-8796, under epic SV-8785)",

        "When you pick a date range on a report, the page's web address changes so the view can be "
        "shared or bookmarked. Our tests check that the address changes and that a shared link "
        "works. They do not check the exact shape of the address, because nothing written down says "
        "what that shape should be.\n\n"
        "An engineering note does suggest a shape, but the product appears to do something "
        "different, and part of that same note has already been overtaken by your update of "
        "4 August.\n\n"
        "Why we are asking: our automation engineer raised this, and we would rather leave the gap "
        "open honestly than invent a rule from an engineering note. No test has been written for it.",

        "Is the exact shape of the web address something we should be testing?",

        "A) No - it is enough that the link works when it is shared. Do not test the exact shape.\n\n"
        "B) Yes, it matters - and here is the shape it must be: ____________________\n\n"
        "C) Ask engineering to settle it and write it down; treat it as their documentation rather "
        "than as a test.",
    ),
    (
        "Filters (the filter buttons on the Work Orders list, on a phone) - the Imported choice "
        "(the mobile filter bar story, SV-8797, under epic SV-8785)",

        "Imported sits in the Status list but behaves differently from the others: while it is "
        "chosen, the other filters cannot be used. That much is written down, and we have now added "
        "a test for it on a phone.\n\n"
        "There is a second behaviour that is not written down anywhere. We are told the product also "
        "does the reverse: if you pick an ordinary status last, Imported is quietly un-picked for "
        "you.\n\n"
        "Why we are asking rather than just testing it: that behaviour exists only in the "
        "developers' own code checks. We do not turn something the code happens to do into something "
        "the product must do - that has to be your decision, or it stops being a test of the product "
        "and becomes a description of it.",

        "Is that reverse behaviour intended - picking an ordinary status last automatically un-picks "
        "Imported?",

        "A) Yes - that is intended. We will test it, and it should be written down.\n\n"
        "B) No - that is not intended. (Then we will raise it.)\n\n"
        "C) Something else - please describe what should happen.",
    ),
]

# ------------------------------------------------------------ NEW Schedule item
NEW_SCHEDULE = [
    (
        "Schedule (the technician scheduling calendar) - which drawing of the Schedule we should be "
        "working from (under epic SV-8685)",

        "This one is about which picture is the real one, and it affects roughly fifty of our tests.\n\n"
        "Back in July you told us the Schedule prototype we were given was the one to work from, and "
        "we pinned about fifty on-screen names and labels from it - button wording, column headings, "
        "the words used in warnings.\n\n"
        "Three Schedule faults raised on 5 August all point at a DIFFERENT drawing instead - a "
        "shared link to a live, editable design page that carries no version and no date. Because it "
        "can be edited at any moment and has nothing on it to say when it was finished, we cannot "
        "compare our tests against it, and we cannot tell whether it is newer or older than the one "
        "we hold.\n\n"
        "You have already noticed the same thing yourself. On one of those three you replied that "
        "the button being reported \"is not in the design\" and asked where it had been found.\n\n"
        "Why we are asking: if that newer drawing is the finished one, then about fifty of our "
        "labels may be out of date and we should go through it properly. If it is not finished, we "
        "should carry on from the prototype and leave it alone. We are not going to guess between "
        "two pictures.",

        "Which drawing of the Schedule is the one to work from - the prototype you pointed us at in "
        "July, or the newer shared design page?",

        "A) THE PROTOTYPE from July is still the one. The newer page is a work in progress and we "
        "should ignore it.\n\n"
        "B) THE NEWER SHARED PAGE is the finished one - please confirm it is final, and we will go "
        "through it and update whatever has changed.\n\n"
        "C) Neither is final yet - please say when a finished drawing will be available.",
    ),
]

# ------------------------------------------- NEW Schedule items from spec v24/v25
# Raised by build/schedule/spec-v25-2026-08-06/QUESTIONS-FOR-BRANKO.md. Every
# sentence below was re-verified live against Confluence page 713031682 at VERSION
# 25 and against Jira SV-8686 / SV-8695 / SV-8829 / SV-8874 / SV-8915 on
# 2026-08-06 before it was written. Appended AFTER the existing four items so the
# established tab order is not restructured (Rule 16); the tab note names which of
# them are decisive so importance is still visible.
NEW_SCHEDULE_V25 = [
    (
        "Schedule (the technician scheduling calendar) - the search box in the toolbar above the "
        "calendar grid (not the one in the job list down the left) (the story about the calendar "
        "layout, SV-8686, under epic SV-8685)",

        "THIS ONE DECIDES WHETHER ONE OF OUR TESTS IS RIGHT OR WRONG, so it is worth a minute.\n\n"
        "Your written description used to say that when someone searches, the jobs that do not match "
        "go FADED BUT STAY ON SCREEN, so you keep sight of the whole week. That sentence was taken "
        "out of the description THIS MORNING, after your team decided the description was wrong "
        "and that the drawing shows only the matching jobs. The description now says only WHAT the "
        "search looks through - customer name, work order number, unit number, technician name and "
        "line name. It no longer says anything at all about what happens to the jobs that do NOT "
        "match.\n\n"
        "Our test still says the non-matching jobs go faded, because that is what was written down "
        "when we wrote it. We are not going to quietly change it to match what the software does "
        "today, because that would tell us what was built rather than what you wanted.\n\n"
        "One more thing in the same place: the developer ticket that describes the calendar layout "
        "STILL says the non-matching jobs should fade - in two places, in its requirements and again "
        "in its acceptance criteria. So your description and that ticket now disagree with each "
        "other. If your answer is A, somebody should tidy that ticket up. We have not touched it, "
        "because it is not ours.",

        "When someone searches, what should happen to the jobs that do not match?",

        "A) They should disappear from the calendar - only the matching jobs are shown. (This is what "
        "your team said the drawing shows.)\n\n"
        "B) They should stay on screen but faded, and the matching ones stand out.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "Schedule (the technician scheduling calendar) - the pop-up window that opens when you click "
        "a scheduled job (the story about that window, SV-8695, under epic SV-8685)",

        "THIS ONE DECIDES WHETHER A TEST PASSES OR FAILS, which is exactly why we are asking instead "
        "of choosing.\n\n"
        "Earlier today, on the report about that window, you told us the little ESTIMATE BADGE should "
        "not be clickable, and that the time is changed in the fields higher up the window instead. "
        "That makes sense to us.\n\n"
        "What we cannot tell is how far your answer reaches. Your written description STILL says the "
        "window should let someone type a new estimate straight into it, and the developer ticket "
        "for that window still says the same. Both of those are live today. So your sentence might "
        "mean the estimate cannot be changed anywhere in that window, or it might mean only that the "
        "small badge on the job line should not be clickable.\n\n"
        "We have one test that says the estimate CAN be typed into. If you mean the first, that test "
        "is wrong and we will correct it. If you mean the second, the test is right and the software "
        "has something to fix. We are deliberately not settling it by looking at what the software "
        "does today.",

        "In that pop-up window, should someone be able to change the ESTIMATED HOURS by typing into "
        "them?",

        "A) NO - the estimate cannot be changed in that window at all; only the start and end times "
        "can be changed, in the fields above.\n\n"
        "B) YES - the estimate itself can still be typed into; only the little badge on the job line "
        "should not be clickable.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "Schedule (the technician scheduling calendar) - the list of jobs INSIDE that same pop-up "
        "window (the story about that window, SV-8695, under epic SV-8685)",

        "This one is small, and it is only two of your own documents disagreeing with each other.\n\n"
        "Your written description was changed this morning so that each job line in that pop-up shows "
        "an estimate figure and a status label. In your own words earlier today, the lines should show "
        "the estimate and the status badge and there should not be any totals. Our test already "
        "expects exactly that - the hours and a status label, and no money anywhere in the window - "
        "so nothing of ours is stuck.\n\n"
        "But the developer ticket for that window still says each line shows a labour TOTAL, and it "
        "has said so since before your change. You edited that same ticket three days ago and the "
        "word is still sitting in it. So your description and that ticket now say different things, "
        "and only you can say which one is the one to keep.\n\n"
        "While you are in that ticket: it also still says the estimate can be typed in, which is the "
        "question just above this one. Both would be tidied up in the same visit.",

        "On each job line in that pop-up, should there be a money total, or only the hours and a "
        "status label?",

        "A) ONLY the hours and a status label - no money anywhere. Your description is right and the "
        "developer ticket needs tidying up.\n\n"
        "B) A MONEY TOTAL should be shown there - the developer ticket is right and your description "
        "needs correcting. (Then our test is wrong and we will change it.)\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "Schedule (the technician scheduling calendar) - how much of the day the timeline shows when "
        "the day view opens (under epic SV-8685)",

        "Nothing is stuck on this one and no test of ours is wrong today. We are asking so that we "
        "are not quietly relying on a note that disagrees with your own description.\n\n"
        "Your description says the day view keeps the WHOLE 24 HOURS there and scrollable, and "
        "simply scrolls itself so the start of the working day is on the left.\n\n"
        "A note on one of the design-review reports asks for something different: that the timeline "
        "show ONLY THE WORKING HOURS plus a small amount after them, with anything outside that "
        "reached by scrolling. That note says it is being tracked as a later improvement rather than "
        "for this release, which is why we have left our test alone.\n\n"
        "If the narrower version is meant for this release, your description needs changing first "
        "and then we will change the test to match it.",

        "For THIS release, which is right?",

        "A) Keep the full 24 hours, as your description says today - the narrower version is for "
        "later.\n\n"
        "B) Change it now to show only the working hours plus a small amount after them.\n\n"
        "C) Something else - please describe it.",
    ),
]


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws, r


# ------------------------------------------------------------------ the tab order
# Filters first: it holds the most blocked tests (4 + 10) and the newest asks.
# Within the tab, the new Status-chip question leads, then the carried-forward
# Filters items in the order the 2026-08-05 sheet had them.
TAB1 = [NEW_FILTERS[0], PRIOR_FILTERS[3], NEW_FILTERS[1], NEW_FILTERS[2],
        PRIOR_FILTERS[0], PRIOR_FILTERS[1], PRIOR_FILTERS[2]]
TAB2 = list(PRIOR_SCHED_DOC) + NEW_SCHEDULE + NEW_SCHEDULE_V25
TAB3 = list(PRIOR_SCHED_ENG)

TR = "https://shopview.testrail.io/index.php?/cases/view/"

QA_ROWS = [
    ("Tab 1", "1",
     "Filters - is the Status chip hidden or shown greyed-out on Estimates and Completed?",
     "NEW. Row 1 of Vlad's eleven-row Filters coverage-gap table, worked as Q1 of "
     "build/filters/vlad-gap-review-2026-08-06/QUESTIONS-FOR-BRANKO.md. Three sources disagree.",
     "FOUR cases on AUTOMATION: HOLD - FLT-TAB-02 (C29609); FLT-TAB-03 (C29610); "
     "FLT-BAR-03 (C29559); FLT-TAB-05 (C29612)",
     " · ".join(f"C{i} {TR}{i}" for i in [29609, 29610, 29559, 29612]),
     "LIVE-VERIFIED 2026-08-06 against Filters Confluence page 572030978 at VERSION 19 "
     "(2026-08-06T11:48:47Z, HTTP 200). S9-R2 verbatim: \"On the Estimates tab, the Status filter "
     "chip is hidden; the remaining four filters are shown and apply on top of the Estimates "
     "pre-filter\"; S9-R3 says the same for Completed. UNCHANGED by v19 - so the contradiction with "
     "Branko's Round-1 Q4=B answer of 2026-07-17 and the QA lead's ruling of 2026-07-30 is still "
     "live. All four HOLD reasons read \"waiting on Branko to confirm whether the Status chip is "
     "hidden or shown greyed out ...\" live today.",
     "A -> the 4 cases revert to hidden and Branko's July answer is retired. B -> the 4 stand as "
     "written, a spec edit is owed, and a defect is raised against the build. This is the sharpest "
     "row in Vlad's table because it alleges a Rule-57-class defect (our cases asserting the "
     "REJECTED behaviour), not merely a gap - which is why it leads the tab."),

    ("Tab 1", "2",
     "Filters - the Parts and Reports product write-up: when, and is the session happening?",
     "CARRIED FORWARD VERBATIM from the 2026-08-05 sheet's Filters item 4, and it is ALSO row 7 of "
     "Vlad's table (Q2 of QUESTIONS-FOR-BRANKO.md). Deliberately framed as a STATUS/COMMITMENT ASK, "
     "not a re-ask of the list.",
     "TEN cases on AUTOMATION: HOLD naming this write-up - C38880; C38882; C38904; C38905; C38906; "
     "C38907; C38908; C38909; C38910; C38911. (An eleventh, C43562, is held because the bar has "
     "only reached some Parts views - absent product, not a missing document.)",
     " · ".join(f"C{i} {TR}{i}" for i in
                [38880, 38882, 38904, 38905, 38906, 38907, 38908, 38909, 38910, 38911]),
     "COUNTED LIVE 2026-08-06: 114 Filters cases under group 4110, all ours (created_by 3), 20 on "
     "AUTOMATION: HOLD, of which exactly TEN name Branko's Parts/Reports write-up. Filters spec v19 "
     "still defines no chip set for any Parts view or report.",
     "⚠️ THIS IS NOT A RE-ASK, AND MUST NOT BE SENT AS ONE. He has ALREADY said twice that no list "
     "exists - 2026-07-31 Q3 and 2026-08-04 Q8, the latter verbatim: \"Same as before, we do not "
     "have concrete list. If this is really necessary i suggest Engineering + PO together make a "
     "list for remaining 6 Parts pages i Reports, using same format as Work Orders do.\" "
     "build/OUTSTANDING-ITEMS-REGISTER.md row 623 records the standing decision: \"We are "
     "deliberately NOT re-asking ... a third ask would be the fifth withdrawn question on this "
     "project.\" So the row asks only ROUGHLY WHEN and WHETHER IT IS STILL IN THIS RELEASE - a "
     "commitment on the session HE proposed. NOTE FOR THE QA LEAD: arranging that Engineering + PO "
     "session is on YOU, not on Branko (register row 636). HONEST CORRECTION: "
     "QUESTIONS-FOR-BRANKO.md says these have been paused \"for four weeks\". The PRD request went "
     "to him on 2026-07-17, which is THREE weeks (20 days) today; the cases were authored "
     "2026-07-27. The reader-facing row says neither figure - it says \"since 17 July\"."),

    ("Tab 1", "3",
     "Filters - is the exact shape of the report page's web address a requirement?",
     "NEW. Row 8 of Vlad's table, Q3 of QUESTIONS-FOR-BRANKO.md.",
     "NO case authored - the gap was left open on purpose. FLT-RPTS-23 (C38882) is the nearest "
     "case and is already held on the Parts/Reports write-up above.",
     f"C38882 {TR}38882",
     "Filters v19 §4 and S11-R1 say nothing about the address format. The engineering plan's item "
     "D19 proposes one, and part of that plan was already overtaken by his 2026-08-04 update. A "
     "5 August capture showed the product sending a different shape. Verified live 2026-08-06 that "
     "v19 adds nothing on this point.",
     "A -> the gap is closed by product decision and stays deliberately untested; record it in the "
     "decisions register. B -> one new case becomes authorable against his stated shape. C -> it "
     "becomes engineering documentation and we assert only that a shared link works. Rule 30: an "
     "engineering plan informs but never overrules product truth, which is exactly why this is a "
     "question and not a case."),

    ("Tab 1", "4",
     "Filters - on a phone, does picking an ordinary status un-pick Imported?",
     "NEW. Row 11 of Vlad's table, Q4 of QUESTIONS-FOR-BRANKO.md.",
     "The documented half is covered by the new FLT-MOB-11 (C43563). The reverse behaviour is ONE "
     "added assertion if he says yes.",
     f"C43563 {TR}43563",
     "Filters v19 S2-R7 verbatim: \"Imported is an exception to S2-R2 and cannot be combined with "
     "anything else ... selecting Imported switches the list to the imported records and disables "
     "the other filter chips while it is active\". S2-N4 covers the prevented combination. NEITHER "
     "describes the reverse. Searched S2-R7, S2-N4, Story 12 and every recorded Branko answer: 0 "
     "occurrences. It exists only in the developers' own MobileAllFiltersSheet spec file.",
     "A -> one assertion is added to C43563 and a spec line is owed. B -> a defect is raised. "
     "Rule 58 is the reason this is asked rather than settled from the code: an ambiguous or absent "
     "source is never resolved by looking at what the build does."),

    ("Tab 1", "5-7",
     "Filters - the three items carried forward from the unsent 2026-08-05 sheet: where the filter "
     "bar sits · \"Apply Filters\" with a capital F · the wrong cross-reference in his own spec",
     "CARRIED FORWARD VERBATIM (imported from the 2026-08-05 generator, so the wording cannot "
     "drift). All three re-verified as still open on 2026-08-06.",
     "The filter-bar position row is the one that matters most: a developer job is about to move "
     "the bar on the strength of somebody else's reading of his document. The cross-reference row "
     "needs no decision at all.",
     "(no additional held cases beyond those already listed)",
     "THE CROSS-REFERENCE IS STILL WRONG IN v19, verified live 2026-08-06: S12-R2 reads \"The "
     "filter chips behave like desktop with one exception (see S12-R5)\", but S12-R5 is about the "
     "page search control - \"The page search control is shown on mobile and behaves as it does on "
     "desktop\". The real exception is S12-R6: \"Unlike desktop, mobile does not filter in real "
     "time ... the table updates only when the user taps an 'Apply filters' button within the "
     "sheet.\" His own v17 renumbering moved the apply-button requirement down one and the pointer "
     "above it was left behind.",
     "The filter-bar row: A -> the developer job proceeds. B -> it is cancelled and the spec is "
     "edited. The capital-F row is a one-word choice either way. The cross-reference row is a tick."),

    ("Tab 2", "1-3",
     "Schedule - the three questions about his own document: shop closures on a multi-day spread · "
     "left-click or right-click on empty calendar space · weekends for a technician with no hours "
     "set",
     "CARRIED FORWARD VERBATIM from the 2026-08-05 sheet's Schedule tab. Item 1 was drafted on "
     "22 JULY AND NEVER SENT - the delay is OURS and the sheet says so to his face.",
     "THREE cases on AUTOMATION: HOLD waiting on a product-owner answer that has never been sent - "
     "SCH-SPREAD-07 (C29983); SCH-EDGE-05 (C30089); SCH-DND-09 (C43555)",
     " · ".join(f"C{i} {TR}{i}" for i in [29983, 30089, 43555]),
     "RE-VERIFIED LIVE 2026-08-06 against Schedule Confluence page 713031682, which has moved to "
     "VERSION 25 (2026-08-06T09:13:51Z, HTTP 200) from the version 23 our records name. BOTH "
     "contradictions SURVIVE v25. Shop closures, both sentences still present verbatim: \"Shop "
     "closures and public holidays are not skipped in V1.\" and \"Shop closures (holidays, "
     "inventory days) are defined at the shop level and block the spread step from placing shifts "
     "on those days.\" Click menu, both still present: \"Create via left-click on empty grid space, "
     "which opens a menu with 'Create event' and 'New work order'.\" against the access-level "
     "section's \"shift and event creation (including via right-click context men[u]\".",
     "Shop closures: A (skip) -> both cases assert skipping and a developer job follows. B (do not "
     "skip) -> both assert placement and the second sentence is deleted. HONEST NOTE CARRIED ON THE "
     "READER TAB: for two of these his answer alone will not make the case runnable, because the "
     "shop-closure setting does not exist anywhere in the product yet. The click-menu row is a "
     "one-word spec fix and no test is wrong today. ⚠️ THE SPEC MOVED 23 -> 25 AND WE HAVE NOT "
     "DIFFED IT - that is a genuine outstanding item, recorded below, not glossed."),

    ("Tab 2", "4",
     "Schedule - which drawing of the Schedule is canonical, and is Sasha Grosman's design final?",
     "NEW, and it is the precondition on a ruling the QA lead has ALREADY given, verbatim: \"Yes if "
     "Sasha's design is final.\" We cannot establish that ourselves - it needs a person to confirm.",
     "No case is on hold for it, and that is the point: about 48 Schedule labels were pinned from "
     "the July prototype and would silently be stale if the newer drawing is the finished one. "
     "SV-8916 could not be verified AT ALL - its button is in his design and in no requirement of "
     "the specification.",
     "(no held cases - the exposure is ~48 pinned labels across the Schedule suite)",
     "READ LIVE FROM JIRA 2026-08-06, read-only: SV-8915 = Bug, OBSOLETE/Done, parent SV-8685, "
     "reporter Sasha Grosman · SV-8916 = Bug, BLOCKED, parent SV-8685, Sasha Grosman · SV-8917 = "
     "Bug, IN PROGRESS, parent SV-8685, Sasha Grosman. All three cite the same source: a "
     "claude.ai/design share URL with NO version and NO date, each \"Raised in the Schedule design "
     "review with Fabian on 5 Aug 2026\" - NOT the artefact at build/schedule/design-2026-07-27/ "
     "that Branko ruled authoritative at Q0. BRANKO HAS ALREADY PUSHED BACK HIMSELF, on SV-8916 at "
     "2026-08-06T03:30:54-0500, verbatim: \"Hey there is no 'Add Existing Work Order' in the "
     "design. Can you clarify where you found this?\" - which is why the row quotes him back to "
     "himself rather than presenting Sasha's claim as fact.",
     "A -> the design baseline stands, C3 in the register closes, and Sasha's three claims are "
     "answered from the artefact we hold. B -> the QA lead's conditional authorisation FIRES and we "
     "fetch and diff the newer design, re-checking ~48 labels. C -> the design source stays a "
     "PARTIAL source under Rule 31 and every design-pinned label carries that caveat. Note the "
     "register names Sasha Grosman / Fabian as the people who can say whether it is final; it is "
     "asked of Branko because HE made the July ruling and is already engaged on exactly this point."),

    ("Tab 2", "5",
     "Schedule - what should happen to the calendar blocks that do NOT match the toolbar search?",
     "NEW, from the Schedule spec v23 -> v25 diff: Q1 of "
     "build/schedule/spec-v25-2026-08-06/QUESTIONS-FOR-BRANKO.md. Confluence **v24** DELETED the "
     "fade/highlight requirement; the PRD is now SILENT on what happens instead.",
     "ONE case, and his answer decides what it should ASSERT - SCH-TOOL-03 (C30041). It is NOT on "
     "HOLD today; it carries AUTOMATION: READY - EXPECT FAIL (SV-8874), and a HOLD is PROPOSED "
     "pending this answer (spec-v25-2026-08-06/PROPOSED-CHANGES.md §2).",
     f"C30041 {TR}30041",
     "VERIFIED LIVE 2026-08-06, read-only. Schedule Confluence page 713031682 at VERSION 25 "
     "(2026-08-06T09:13:51Z, HTTP 200): §6 now reads only \"Search - Filters grid blocks by matching "
     "against customer name, WO number, unit number, technician name, and line name.\" Searched the "
     "whole v25 body: \"non-matching\" = 0 hits, \"de-emphas\" = 0, \"dim\" = 0; the single \"faded\" "
     "hit is the series \"continues\" label and is unrelated. SO THE PRD IS SILENT, NOT REVERSED. "
     "JIRA READ LIVE: SV-8686 (Story, TESTING QA, parent SV-8685) STILL CARRIES IT TWICE - "
     "Requirements verbatim \"Non-matching blocks fade; matching blocks highlight. - (PRD: §6)\" and "
     "Acceptance Criteria verbatim \"...then matching blocks highlight and non-matching blocks "
     "fade.\" SV-8874 (Story Defect, parent SV-8686) is OBSOLETE/Done, resolved "
     "2026-08-06T03:32:42-0500 by Milos Vasic, whose comment reads \"All good on this one updated the "
     "PRD , i will close this ticket as absolute\", 81 seconds before v24. TESTRAIL READ LIVE: "
     "C30041's title is \"Toolbar search highlights matching blocks and fades non-matching ones\" and "
     "expected item 1 is \"Blocks that match the search are highlighted; blocks that do not match "
     "fade.\" - so the deleted requirement is in the title AND the body.",
     "A -> the fade assertion is REMOVED and the case asserts non-matching blocks are gone; marker "
     "becomes READY. B -> the build has a real defect and SV-8874 was closed wrongly. EITHER WAY THE "
     "MARKER IS ALREADY WRONG, because it names SV-8874 as an open expect-fail and SV-8874 is closed. "
     "Rule 57 is why this is asked and not read off the build: the PRD's SILENCE licenses REMOVING "
     "our assertion, it does not license asserting the opposite. Stefan Vukovic's \"per design we "
     "show only shifts/events that are matching the search\" (2026-08-06T03:15:35-0500) is an "
     "ENGINEER's statement about a design we do not hold, so it is not a PO ruling (Rule 33). "
     "SV-8686 is the OWNER's to correct - Rule 38, not ours to edit."),

    ("Tab 2", "6",
     "Schedule - can the estimated hours still be typed into inside the shift window?",
     "NEW, Q2 of the same file, and THE SHARPEST ROW ON THE SHEET: it is the difference between a "
     "case that passes and a case that fails on the same build. Branko's own ruling created the "
     "ambiguity 17 minutes after he published v25.",
     "ONE case - SCH-MODAL-05 (C30012). Currently AUTOMATION: READY, i.e. we expect it to PASS, while "
     "the PO has closed a report saying the build does not allow the edit. A HOLD is PROPOSED, not "
     "applied (PROPOSED-CHANGES.md §5).",
     f"C30012 {TR}30012",
     "VERIFIED LIVE 2026-08-06, read-only. v25 §4.9 STILL READS \"Estimated hours with inline "
     "edit.\" - the spec line was NOT changed by v24 or v25. SV-8695 (Story, Ready for QA) Requirements "
     "verbatim: \"...scope summary with scheduled line(s) and labor/total, estimated hours with inline "
     "edit...\" - SO TWO WRITTEN SOURCES STILL REQUIRE IT. Against that, SV-8829 (Story Defect, "
     "OBSOLETE/Done, parent SV-8695) carries Branko Cicovic's ONLY comment, at "
     "2026-08-06T04:31:05-0500 = 09:31:05Z, SEVENTEEN MINUTES AFTER v25 was published at 09:13:51Z, "
     "verbatim: \"Estimated badge should not be clickable, you can change time only in the input "
     "fields above. I updated PRD, for work order lines we just show estimate and status badge, there "
     "shouldn't be totals. Please always check the design as it is single source of truth.\" THE "
     "AMBIGUITY IS SCOPE: SV-8829's own steps distinguish the LINE ROW's badge (\"only 1h and an "
     "Authorized badge\") from the MODAL-level \"estimated hours with inline edit\", and his sentence "
     "names \"the Estimated badge\". TestRail read live: C30012 = \"Estimated hours can be edited "
     "inline in the modal\", refs SV-8695 (§4.9 (Estimated hours with inline edit)), marker READY.",
     "A -> the case is WRONG, is rewritten, and both the spec line and SV-8695 are stale. B -> the "
     "case is RIGHT and SV-8829 was closed wrongly, so a defect is owed. Rule 58 is the whole reason "
     "this is a question: an ambiguous source is NEVER resolved by looking at what the build does - "
     "and flipping the case off two written sources onto one ambiguous sentence would be the Filters "
     "mistake in reverse."),

    ("Tab 2", "7",
     "Schedule - his own story still says the shift window's job lines show a labour TOTAL, while "
     "his own spec now says a labour/status figure",
     "NEW, and it is the row the QA lead asked for explicitly: a small inconsistency between two of "
     "HIS OWN documents that he owes. It also carries the second half of item 6's tidy-up.",
     "ONE case, and it does NOT change under answer A - SCH-MODAL-04 (C30011). Marker AUTOMATION: "
     "READY. PROPOSED-CHANGES.md §1 proposes a provenance/divergence edit only, no assertion change.",
     f"C30011 {TR}30011",
     "VERIFIED LIVE 2026-08-06, read-only. v25 §4.9 verbatim: \"Scope summary and the scheduled "
     "line(s) with labor/status figures.\" SV-8695 Requirements verbatim, STILL: \"...scope summary "
     "with scheduled line(s) and labor/total...\" - and per the v25 diff pass that phrase is present "
     "both BEFORE AND AFTER Branko's own description edit of 2026-08-03T08:54:30Z, so the PO edited "
     "this story three days ago and left `total` standing, then changed the PRD on 2026-08-06. THREE "
     "SOURCES AGREE AND ONE LAGS: his 22 July no-money-on-the-Schedule ruling, PRD v25 "
     "(`labor/status`), and his SV-8829 comment (\"we just show estimate and status badge, there "
     "shouldn't be totals\") all say no totals; SV-8695 alone still says `total`. TestRail read live: "
     "C30011 = \"The modal lists the scheduled line(s) with no money fields\", expected items 2-3 "
     "verbatim \"...each showing its line number, title, hours, and a status pill only. 3. No labor "
     "figures and no total dollar amount appear anywhere in the modal.\" - so under Rules 32/33 the "
     "case ALREADY follows the three, which is why nothing is blocked on this row.",
     "A (expected) -> nothing changes in the case; SV-8695 is corrected BY ITS OWNER (Rule 38). "
     "B -> C30011 is wrong and item 3 of its expected result must be rewritten, and the PRD needs "
     "correcting too. NOTE FOR THE QA LEAD ON WHAT WAS DELIBERATELY *NOT* ASKED: the source file's "
     "Q2b asked what the WORD \"labor\" means in \"labor/status figures\" - hours or money. That was "
     "LEFT OFF, because his SV-8829 comment already answers it in his own words and, on the source "
     "file's own admission, NEITHER answer changes the case's substance. Asking it would have re-asked "
     "a settled point for zero test consequence. Its one live consequence - that two of his documents "
     "now disagree - is THIS row. Our own text using \"labor\" for both hours and money in C30011 is "
     "an internal wording repair, not a product question."),

    ("Tab 2", "8",
     "Schedule - for this release, does the day view keep the full 24 hours or only the working "
     "hours plus a buffer?",
     "NEW, Q3 of the same file. Asked so we are not silently relying on a change-request note that "
     "contradicts the live spec - not because any case is wrong today.",
     "ONE case, and NO change is proposed to it - SCH-DAY-01 (C30001). It is correct against v25 as "
     "written and already covers the full start-time hierarchy.",
     f"C30001 {TR}30001",
     "VERIFIED LIVE 2026-08-06, read-only. v25 §4.8 verbatim: \"The full 24-hour timeline remains "
     "intact and scrollable.\", and §3.2: \"Day view. A 24-hour timeline per technician row with "
     "time-positioned blocks.\" Against that, SV-8915 (Bug, OBSOLETE/Done, parent SV-8685, reporter "
     "Sasha Grosman) carries under \"Related change requests\", verbatim: \"Schedule width should "
     "render only business hours plus a small trailing buffer rather than the full 24 hours; "
     "after-hours scheduling is an edge case reachable by scrolling. Tracked separately on the "
     "enhancements list for the Schedule design.\" TestRail read live: C30001 = \"Day view "
     "auto-scrolls to the working-day start; manual scrolling stands\", refs SV-8694 (§4.8 "
     "(Auto-scroll to business hours)), marker AUTOMATION: READY - EXPECT FAIL (SV-8837) - and "
     "SV-8837 is live-confirmed STILL OPEN (Story Defect, Ready for QA, parent SV-8694), so that "
     "marker is correct and untouched by this question.",
     "A (recommended, and what the spec says today) -> nothing changes; the item closes in the "
     "decisions register as out of V1. B -> §4.8 becomes wrong and needs a spec edit BEFORE any case "
     "change - the spec is corrected first, never the case. SV-8915 is ALSO the design-authority "
     "ticket, which is why its other half sits on Schedule item 4 of this sheet rather than here."),

    ("Tab 3", "1-6",
     "Schedule - six behaviours only the engineering plan describes: pre-existing shifts and "
     "events · a multi-day job on the Dashboard · an appointment set while creating a job · jobs "
     "from another branch · the priority on a job · a limit on how long a spread can be",
     "CARRIED FORWARD VERBATIM from the 2026-08-05 sheet's engineering-only tab. All six were "
     "re-checked against spec v25 today and all six are STILL absent from it.",
     "No case is held on these six. They are candidate coverage rather than blocked coverage - "
     "Rule 30 forbids turning an engineering plan into a product requirement without the PO.",
     "(no held cases)",
     "VERIFIED LIVE 2026-08-06 against Schedule spec v25, searching the whole body: pre-existing "
     "shifts/migration = 0 matches · dashboard = 0 · appointment = 0 · another branch / "
     "cross-location = 0 · spread length limit (8 weeks / 120 shifts / maximum span) = 0. Only "
     "\"priority\" appears, and only as a bare field name with no behaviour attached. So none of the "
     "six was answered by the v23 -> v25 edits.",
     "Each answer either creates authorable coverage or closes the item as out of scope. Nothing "
     "here is urgent, which is why it is the last tab - but leaving them off would make six "
     "deliberate omissions indistinguishable from six misses (Rule 46)."),
]

QA_NOTES = [
    "ADDENDUM 2026-08-06 - FOUR SCHEDULE ITEMS ADDED, TAKING THE SHEET FROM 17 TO 21. They come from "
    "build/schedule/spec-v25-2026-08-06/QUESTIONS-FOR-BRANKO.md, the Schedule spec v23 -> v25 diff. "
    "They are APPENDED as Schedule items 5-8 rather than re-ordered into the tab, so the established "
    "structure and the carried-forward wording are untouched (Rule 16); the tab note names items 5 "
    "and 6 as the decisive ones so their importance is still visible on the reader tab. TWO OF THEM "
    "EACH DECIDE A VERDICT: item 5 decides what C30041 should assert, item 6 decides whether C30012 "
    "passes or fails. Nothing was applied to TestRail - no case was edited, and the HOLDs those two "
    "items imply are PROPOSED in PROPOSED-CHANGES.md, not written.",

    "ONE OF THE FOUR SOURCE QUESTIONS WAS DELIBERATELY LEFT OFF, AND THIS IS THE REASON (Rule 55: "
    "ask what is unclear; it does not say invent an ask). The source file's Q2b asked what the word "
    "\"labor\" means in v25's new phrase \"labor/status figures\" - the number of HOURS or a MONEY "
    "amount. It is NOT asked, on two grounds, both checked rather than assumed. FIRST, HE HAS ALREADY "
    "ANSWERED IT IN HIS OWN WORDS: SV-8829, 2026-08-06T04:31:05-0500, verbatim \"for work order lines "
    "we just show estimate and status badge, there shouldn't be totals\" - which, with his 22 July "
    "no-money-on-the-Schedule ruling, says plainly that no money appears. SECOND, THE SOURCE FILE "
    "ITSELF RECORDS THAT NEITHER ANSWER CHANGES THE CASE: \"Neither answer changes the case's "
    "substance; A confirms it as written.\" So it would have cost a PO round trip for zero test "
    "consequence - and re-asking something a source has already settled is the exact embarrassment "
    "this project has had once already. ITS ONE GENUINELY LIVE CONSEQUENCE IS ON THE SHEET as "
    "Schedule item 7: that his STORY and his SPEC now disagree (`labor/total` vs `labor/status`), "
    "which is a fork with a real answer and a named owner. The residual ambiguity is in OUR OWN text "
    "- C30011 uses \"labor\" for both hours and money in one expected result - and that is an "
    "internal wording repair, not a question for a product owner.",

    "THE DESIGN-AUTHORITY QUESTION WAS DELIBERATELY *NOT* ADDED AS A NEW ROW, ON THE QA LEAD'S "
    "INSTRUCTION AND BECAUSE IT WOULD HAVE BEEN A DUPLICATE. The source file's Q4 is a two-part item. "
    "Its PRODUCT half - which drawing of the Schedule is the one to work from - is ALREADY Schedule "
    "item 4 on this very sheet, and asking it twice inside one workbook would be the drip Rule 55 "
    "exists to prevent, in miniature. Its PROCESS half - whether a design may outrank the written "
    "description in OUR OWN method - is a decision for the QA LEAD, not for Branko, and has been put "
    "to him separately. Branko is asked which drawing is canonical; he is not asked to rule on our "
    "sourcing rules.",

    "WHY ONE FILE AND NOT TWO (Standing Rule 55). build/branko-questions-2026-08-05/ holds a "
    "13-item Branko sheet that is READY TO SEND and WAS NEVER SENT (register row C4). Producing a "
    "second sheet without absorbing it would leave the QA lead holding two unsent sheets for the "
    "same person - the exact 'drip of separate asks' Rule 55 exists to prevent. So this workbook "
    "CARRIES FORWARD all 13 of its items and adds 8 new ones = 21 (4 added when it was first built, and 4 more Schedule items added later the same day from the spec's move to version 25). ⚠️ ACTION NEEDED FROM THE QA "
    "LEAD: the 2026-08-05 workbook should be marked SUPERSEDED so an old one cannot go out by "
    "mistake. It was NOT edited by this pass, because this task was scoped to write only inside "
    "build/filters/questions-2026-08-06/ and build/report-suite/questions-2026-08-06/.",

    "HOW DRIFT WAS PREVENTED: every carried-forward reader-facing row is IMPORTED from the "
    "2026-08-05 generator's own TAB1/TAB2/TAB3 lists at build time, not retyped. So the 13 carried "
    "items are byte-identical to the rows that sheet carried - the same technique the 2026-08-04 "
    "Chris consolidation used. Only the 8 new items and the tab notes are new text.",

    "SV-8876 IS DELIBERATELY NOT ON THIS SHEET, AND THE BRIEF FOR THIS TASK WAS WRONG ABOUT IT. "
    "The brief listed it as \"Ahtasham's clarification question, still his to answer\", and "
    "CLAUDE.md's Filters section says \"Branko owes SV-8876\". READ LIVE FROM JIRA 2026-08-06: "
    "SV-8876 is a Task, status DONE, resolution Done, resolved 2026-08-05T08:38:16-0500, parent "
    "SV-8785, reporter Ahtasham Amjad. He CLOSED IT HIMSELF, with one comment, verbatim: \"closing "
    "this as it was a gap with test case , I've updated the test case here "
    ">>https://shopview.testrail.io/index.php?/cases/view/29557 And created a story defect >> as "
    "the build is not behaving as per PRD\". Putting it in front of Branko would have re-asked a "
    "closed question - the exact embarrassment this project has already had once. THE HALF THAT IS "
    "GENUINELY STILL HIS is the narrower one and IS on the sheet, as Filters item 5: did he want "
    "the buttons on one row, in which case the developer job should be cancelled? "
    "SEPARATE POINT FOR THE QA LEAD, NOT FOR BRANKO: that comment says Ahtasham EDITED OUR CASE "
    "C29557. Under Rule 38 we do not touch his cases and he should not be editing ours; this is "
    "reported, not acted on.",

    "THE FILTERS SPEC MOVED TO VERSION 19 THIS MORNING AND NO QUESTION WAS MANUFACTURED FROM IT. "
    "Page 572030978, version 19, 2026-08-06T11:48:47Z. Its visible change on the filter bar is "
    "S1-R3, verbatim: \"Each chip displays a leading type-icon identifying the filter, the filter "
    "name, and a chevron icon indicating it opens a dropdown\". THAT IS CLEAR ENOUGH TO TEST - a "
    "tester can check that each chip carries a leading icon plus a chevron, and which specific icon "
    "belongs to which chip comes from the design node the spec links. So NO ROW WAS ADDED for it "
    "(Rule 55 says ask what is unclear; it does not say invent an ask). What v19 did NOT do is "
    "resolve anything on this sheet: S9-R2 and S9-R3 are unchanged, and the S12-R2 cross-reference "
    "is still wrong. ⚠️ OWED: a proper v18 -> v19 diff. This pass checked the points the sheet "
    "depends on, not the whole document, and does not claim to have diffed it.",

    "THE SCHEDULE v23 -> v25 DIFF HAS SINCE BEEN DONE, AND THIS ADDENDUM IS ITS OUTPUT - so the "
    "warning that stood here (\"we have not diffed it\") IS NOW OUT OF DATE and is corrected rather "
    "than deleted. Page 713031682 is at VERSION 25 (2026-08-06T09:13:51Z, re-confirmed live for this "
    "addendum); CLAUDE.md and build/OUTSTANDING-ITEMS-REGISTER.md still name version 23, and NEITHER "
    "WAS EDITED (out of scope for this pass - the record correction is owed). The diff lives at "
    "build/schedule/spec-v25-2026-08-06/SPEC-DIFF.md: 2 requirements changed, 1 governing-source "
    "change that is not a page edit, 3 cases affected. Its four reader-facing questions are Schedule "
    "items 5-8 of this sheet. Also re-confirmed by that diff and unchanged in v25: both shop-closure "
    "sentences survive, both click-menu wordings survive, and all six engineering-only topics are "
    "still absent - so items 1-3 and tab 3 stand exactly as written.",

    "WORDING RULES APPLIED (Standing Rules 7 and 55): every reader-facing question row names the "
    "PROJECT (Filters or Schedule) and the SCREEN, because Branko owns THREE projects - Filters, "
    "Schedule and Global Search - so a row read on its own, days later, on a phone must still be "
    "unambiguous. Story and epic keys appear in plain form only where they orient him. No case IDs, "
    "no requirement anchors, no technical terms and not the word VIU appear anywhere he reads. Each "
    "question is a plain A/B/C with a blank for the answer, and each carries a one-line reason so "
    "the consequence is visible.",

    "SOURCE-CURRENCY (Standing Rules 31 and 59), fetched LIVE and re-read immediately before "
    "writing, 2026-08-06: Filters spec page 572030978 = VERSION 19, HTTP 200 (in-body field still "
    "reads \"Version: 1.6\" - the Rule-31(a) trap; the CONFLUENCE number is the one used). Schedule "
    "spec page 713031682 = VERSION 25, HTTP 200, and STALE against our records at 23. Jira read "
    "live for SV-8876, SV-8825, SV-8915, SV-8916, SV-8917. TestRail read live for all 114 Filters "
    "cases. Designs = PARTIAL, and which artefact is canonical is Schedule item 4 on this very sheet. "
    "ADDENDUM, 2026-08-06: for the four new Schedule items the Schedule spec was RE-FETCHED live "
    "(page 713031682, version 25, HTTP 200, by Branko Cicovic at 09:13:51Z) and every sentence they "
    "rest on was searched in its body; JIRA was read live and read-only for SV-8686, SV-8695, "
    "SV-8829, SV-8874, SV-8837 and SV-8915; and FOUR SCHEDULE CASES WERE PULLED LIVE FROM TESTRAIL - "
    "C30041, C30012, C30011 and C30001, all HTTP 200, titles, refs and automation markers read from "
    "the live cases rather than from our records. So the earlier line saying \"Schedule cases were "
    "NOT re-pulled this pass\" now holds only for the THREE HELD C-ids on items 1-3 (C29983, C30089, "
    "C43555), which still come from committed records and are still cited as such.",

    "THE BUILD WAS NOT OBSERVED (Standing Rules 12, 49, 60). The shared QA sign-in is dead - a "
    "read-only probe returned HTTP 401 {\"error\":\"sso_required\"} - and quick-login and "
    "switch-user were deliberately NOT called, because both rotate the shared session and would "
    "sign concurrent workers out of the other branches. BOTH BRANCHES HAVE ALSO REDEPLOYED: "
    "sv8785.qa.shopview.com now reads app-version v3.4.2-280ca5a (last-modified Thu 06 Aug 2026 "
    "09:37:49 GMT) where the Filters passes ran on v3.4.2-d00239b. So every build-side sentence "
    "quoted to Branko - the Status chip behaving as the description says, the filter bar on one row, "
    "the capital F - comes from a build that no longer exists, and is worded to him as what the "
    "product does rather than as a fresh measurement. All Filters and Schedule verdicts remain "
    "PROVISIONAL and the Rule-49 queues are OPEN.",

    "NOTHING HAS BEEN WRITTEN ANYWHERE. This sheet is a draft for the QA lead to send. Read-only on "
    "TestRail (get_sections, get_cases), on Confluence (GET content) and on Jira (GET issue). No "
    "TestRail write, no Jira write, no case edit, no run write, no application call. CLAUDE.md, "
    "build/OUTSTANDING-ITEMS-REGISTER.md and build/APP-ACTIONS-PLAYBOOK.md were not touched, and "
    "neither was the 2026-08-05 Branko workbook.",
]


def write_xlsx():
    wb = openpyxl.Workbook()
    w_narrow = [4, 34, 52, 42, 46, 22]
    w_wide = [4, 34, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "Filters - questions for Branko Cicovic - 2026-08-06",
           "PLEASE START WITH QUESTION 1 - four tests are on hold on it, and it is the one where "
           "your own answer and your written description disagree. Question 2 is the biggest single "
           "blocker on this project: ten tests are waiting on it. " + THANKS,
           "Seven Filters items - fourteen tests are on hold across them", TAB1,
           w_wide, first=True)

    _sheet(wb, TAB2_NAME,
           "Schedule - your document - questions for Branko Cicovic - 2026-08-06",
           "EIGHT ITEMS, AND AN APOLOGY WITH THE FIRST. Question 1 was written on 22 July and we "
           "never actually sent it to you. Two tests have been sitting parked ever since waiting "
           "for an answer you were never asked for. That delay is ours, not yours. QUESTIONS 5 AND 6 "
           "ARE THE TWO THAT MATTER MOST: each one decides whether a test of ours is right or wrong, "
           "and both come out of the changes made to your own description this morning. " + THANKS,
           "Eight Schedule items - five of our tests are waiting on your answers", TAB2, w_narrow)

    _sheet(wb, TAB3_NAME,
           "Schedule - engineering only - questions for Branko Cicovic - 2026-08-06",
           "NOTHING ON THIS TAB IS URGENT AND NO TEST IS WAITING ON IT. These six behaviours are "
           "described only in the engineering plan and appear nowhere in your own document. We are "
           "not going to turn an engineering note into something the product must do without your "
           "word, so each one is either a decision for you or an item we close as out of scope.",
           "Six items the engineering plan describes and your document does not", TAB3, w_narrow)

    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR BRANKO. Do not send this tab. TestRail C-ids, "
                 "requirement anchors, live evidence and the corrections to our own records live "
                 "here so the reader-facing tabs stay plain (Standing Rules 7, 8 and 55).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Tab", "Item", "What it asks", "Where the question comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 62, 62]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, 1):
            out.append(f"### Item {i}.0 — {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    md = f"""# Questions for Branko Cicovic — Filters and Schedule — 2026-08-06

**Projects: Filters (epic SV-8785) and Schedule (epic SV-8685) · Product Owner: Branko Cicovic**

**This is the plain-language twin of
`Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-06.xlsx`.**
The spreadsheet is the version to send; it mirrors the established Chris Ward sheets' format
exactly, and it carries a QA-only tab that must not be forwarded.

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

**⚠️ THIS WORKBOOK REPLACES `build/branko-questions-2026-08-05/`**, which was written, is ready, and
was never sent. Standing Rule 55 says to sweep every open ambiguity onto ONE sheet so a product
owner answers in a single sitting rather than receiving a drip of separate asks — two unsent sheets
to the same person is that drip. All 13 of that sheet's items are carried forward here, imported
from its own generator so the wording cannot drift, and 8 new items are added. **The 2026-08-05
workbook should be marked superseded so an old one cannot go out by mistake.**

{THANKS}

**Twenty-one items in total: seven Filters, eight Schedule questions about his own document, and six
Schedule behaviours that only the engineering plan describes. Seventeen of our tests are on hold
across them, and two more are waiting on the two new Schedule questions that decide whether they
pass or fail.**

**Schedule items 5 to 8 were added on 2026-08-06** from the Schedule description's move to version 25
earlier the same day. **Items 5 and 6 are the decisive ones:** item 5 decides what one of our tests should
say at all, and item 6 is the difference between a test that passes and a test that fails on the same
software. One further question that the same review raised was **deliberately left off**, because he
had already answered it in his own words and it changed no test — the reason is recorded on the
QA-only tab.

**Live source versions confirmed on 2026-08-06, immediately before writing** — Filters
specification **version 19** (published 11:48 UTC this morning) · Schedule specification **version
25**. Every sentence quoted below comes from that fetch. *(Note for us, not for him: the Filters
page's in-body field still reads "Version: 1.6" — the Confluence version number is the one used.
And our own records still say the Schedule specification is at version 23, so two versions of change
are uningested — recorded on the QA-only tab.)*

---

## Tab 1 — {TAB1_NAME}

**Please start with question 1** — four tests are on hold on it, and it is the one where your own
answer and your written description disagree. **Question 2 is the biggest single blocker on this
project:** ten tests are waiting on it.

{block(TAB1)}

---

## Tab 2 — {TAB2_NAME}

**Eight items, and an apology with the first.** Question 1 was written on 22 July and we never
actually sent it to you. Two tests have been sitting parked ever since waiting for an answer you
were never asked for. That delay is ours, not yours.

**Questions 5 and 6 are the two that matter most** — each one decides whether a test of ours is right
or wrong, and both come out of the changes made to your own description this morning.

{block(TAB2)}

---

## Tab 3 — {TAB3_NAME}

**Nothing on this tab is urgent and no test is waiting on it.** These six behaviours are described
only in the engineering plan and appear nowhere in your own document. We are not going to turn an
engineering note into something the product must do without your word.

{block(TAB3)}

---

## QA-only — not for Branko

The internal question-to-case mapping lives on the spreadsheet's `{TAB4_NAME}` tab: every question's
affected TestRail case IDs with links, the requirement anchors quoted from the live pages, and what
each possible answer resolves to. It also records why this is one consolidated file rather than two,
**why SV-8876 is deliberately NOT on the sheet** (it is closed — Ahtasham Amjad closed it himself on
5 August), why the new Filters version 19 produced no question, and the two source-currency gaps
this pass found and did not paper over.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
