#!/usr/bin/env python3
"""Branko Cicovic PO question sheet — Filters + Schedule — 2026-08-17.

WHY THIS FILE EXISTS. The 2026-08-17 Fabian-redesign reconciliation of Filters (spec v21) and
Schedule (spec v30) left a small set of GENUINE PRODUCT DECISIONS that only Branko can settle —
each one is a difference between two of his own documents, or a decision his written description
does not yet contain. Per Standing Rule 55 they go onto ONE plain sheet so he answers in a single
sitting; per Standing Rule 7 the reader-facing tabs carry no case IDs, anchors, HTTP terms, or
jargon, and the question->case mapping lives only on a QA-only tab that must not be forwarded.

EIGHT items. Questions 1-3 were SENT 2026-08-17; questions 4-8 are NEW FOLLOW-UPS added 2026-08-18
(Aug-5 design review, Schedule spec v30 silent) and are HELD pending the QA lead sending them:
  1. FILTERS  — Status button hidden vs greyed on Estimates/Completed  (FAB-2; held, C29609/C29610)
  2. FILTERS  — the per-view filter list: which filters on which page   (FAB-1; owed by engineering)
  3. SCHEDULE — the shift pop-up: should it offer a "Reassign" action    (C30015; genuine PO decision)
  4. SCHEDULE — a user-level "always schedule the whole work order" preference   (E6; v30 silent)
  5. SCHEDULE — restore the one-click carryover button                            (E15; v30 silent)
  6. SCHEDULE — rename "Carryover" to "Add a Day" / "Extend a Day"                (E7; v30 silent)
  7. SCHEDULE — multi-day carryover extends ONE day, not the whole span           (E8; v30 silent)
  8. SCHEDULE — drag a shift onto the next day in week view                       (E9; v30 silent)
Source for 4-8: build/schedule/aug5-review-2026-08-17/RECONCILIATION.md (OUTSTANDING items 1-2).

DELIBERATELY NOT ON THIS SHEET (already answered — task: "leave out anything already answered by
spec v21/v30"):
  * Schedule shop-closures / multi-day spread — ANSWERED by spec v30 (§4.5/§12: skip weekends only,
    closures receive shifts). The cases already follow v30; it is not a question for Branko.
  * The greyed-vs-hidden question is the ONLY Filters tab-behaviour item still open; everything else
    the redesign changed is settled by spec v21.

Sources for the wording (no live read this pass; sheet is HELD, not sent — Rule 66):
  * build/filters/fabian-review-2026-08-17/FILTERS-COMPLETION-REPORT.md  (OUTSTANDING items 1, 2, 5)
  * build/filters/fabian-review-2026-08-17/COVERAGE-REDERIVATION.md      (greyed-vs-hidden; per-view)
  * build/schedule/fabian-review-2026-08-17/SCHEDULE-COMPLETION-REPORT.md (C30015 open PO question)
  * build/OUTSTANDING-ITEMS-REGISTER.md                                   (FAB-1, FAB-2, SCH-FR-6)

RESEARCH ONLY — writes two files (.xlsx + .md) into this folder. No TestRail write, no Jira call,
no application call, no Confluence call.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Filters-and-Schedule_Questions-for-Branko_2026-08-17.xlsx")
MD = os.path.join(HERE, "Filters-and-Schedule_Questions-for-Branko_2026-08-17.md")

HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

TAB1 = "Questions for Branko"
TABQA = "QA internal - not for Branko"

HELLO = (
    "Hello Branko - this is everything we have open across your projects FILTERS and SCHEDULE, "
    "gathered into one place so you can answer it in one sitting instead of a trickle of separate "
    "messages. SHORT ANSWERS ARE PERFECT - a letter, or one line. Nothing here needs an essay.\n\n"
    "There are EIGHT questions in total. The first THREE (1 to 3) are the ones we already sent "
    "you - they are repeated here only so everything is in one place. Questions 4 to 8 are NEW "
    "FOLLOW-UPS, all about the Schedule's 'carryover' feature and one scheduling preference from "
    "the design review - so if you have already answered 1 to 3, you only need to look at 4 to 8.\n\n"
    "Every question says which project and screen it is about, because you look after Filters, "
    "Schedule and Global Search. Each one is a point where two of your own documents disagree, or "
    "where your written description does not yet say the thing we need - so we are asking you which "
    "to keep, rather than guessing. To be clear: we have not edited any of your tickets or "
    "descriptions.\n\n"
    "The first question has two tests parked on it right now, so it is the one that unblocks work."
)

# ================================================================ THE THREE QUESTIONS
# (topic, what-happens-now, the-question, options)
QUESTIONS = [
    (
        "FILTERS - the Work Orders list - the Status button on the Estimates and Completed tabs",

        "Two of our tests are on hold on this one point, and two answers are on record that "
        "disagree with each other.\n\n"
        "The Work Orders list has tabs across the top. Two of them - Estimates and Completed - "
        "already show only one kind of work order. There is also a row of filter buttons, and one "
        "of them is Status.\n\n"
        "Your written description says the Status button is NOT SHOWN AT ALL on those two tabs.\n\n"
        "But you told us on 17 July that the Status button IS SHOWN, greyed out, already filled in "
        "with that tab's own status, and cannot be changed. Our QA lead agreed with that on 30 "
        "July, and the design shows it that way too.\n\n"
        "Why we are asking rather than choosing: we have set the two tests to your July answer, "
        "because that is what you and our QA lead actually decided - but the written description "
        "still says the opposite. So one of them has to change, and it is your call which.",

        "On the Estimates and Completed tabs, is the Status button hidden, or shown greyed out and "
        "already filled in?",

        "A) NOT SHOWN AT ALL on those two tabs - the written description is right, and my July "
        "answer is out of date.\n\n"
        "B) SHOWN, GREYED OUT AND ALREADY FILLED IN - my July answer stands, and the description "
        "needs correcting. (Then we will also raise it so the product can be fixed to match.)\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "FILTERS - the filter buttons on the Parts pages and the Report pages",

        "The redesign puts a row of filter buttons on the Parts pages and on the Report pages. "
        "Your written description says the filters those pages already had are moved into the new "
        "row - but it does not list exactly WHICH buttons should appear on WHICH page.\n\n"
        "Your engineering team was going to send us that page-by-page list, and it has not arrived "
        "yet. Without it we can check that the buttons on those pages work, but we cannot yet check "
        "that each page is showing exactly the right set of buttons - so those tests say 'confirm "
        "the exact buttons later' and are waiting on this.\n\n"
        "Why we are asking you: it is a product decision - which filter buttons belong on each "
        "page - and you are the person who can confirm it.",

        "For the Parts pages and the Report pages, can you confirm which filter buttons should "
        "appear on each page?",

        "A) Every page keeps exactly the same filters it had before the redesign - nothing added "
        "or removed - so the old set for each page is the answer.\n\n"
        "B) There is a specific page-by-page list - you (or engineering) will send it, and we will "
        "check each page against it.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the pop-up window that opens when you click a scheduled job",

        "When someone clicks a job on the schedule, a small pop-up window opens with actions on it. "
        "Right now that window offers DELETE only.\n\n"
        "An earlier version of your description mentioned a REASSIGN action in that window, and a "
        "later version took it out. Separately, a job can already be moved to a different "
        "technician by DRAGGING it on the calendar.\n\n"
        "So there are two sensible possibilities and we do not want to guess: either the window is "
        "correct with Delete only and reassigning is done by dragging, or the window should also "
        "offer a Reassign action. We have kept the test flagged and left it for your decision.",

        "In that pop-up window, should there be a REASSIGN action, or is Delete the only action and "
        "reassigning is done by dragging the job to another technician?",

        "A) DELETE ONLY - reassigning is done by dragging the job to another technician. The window "
        "is correct as it is.\n\n"
        "B) ADD A REASSIGN ACTION to the window as well.\n\n"
        "C) Something else - please describe it.",
    ),
    # ---- Q4-Q8: NEW FOLLOW-UPS (Aug-5 design review, Schedule spec v30 silent) ----
    (
        "SCHEDULE - the scheduling pop-up - a personal 'always schedule the whole work order' "
        "preference",

        "When you drag a work order onto the schedule, it asks whether to schedule the WHOLE work "
        "order or to pick individual lines. Right now it asks this every time.\n\n"
        "The 5 August design review suggested letting each person set a preference - 'always "
        "schedule the whole work order' - so the system remembers their choice and stops asking "
        "every time.\n\n"
        "Your newer written description does not mention this preference at all, and the review "
        "itself listed it as an open question to decide before launch. So we are asking you rather "
        "than guessing.",

        "Should a person be able to set a preference to always schedule the whole work order (so "
        "the system remembers it and stops asking each time), in the first version?",

        "A) YES - add the 'always schedule the whole work order' preference in the first version.\n\n"
        "B) NO - leave it out for now and keep asking each time; maybe a later version.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the schedule grid - the one-click carryover button",

        "The old scheduler had a one-click 'carryover' button on a scheduled job. Pressing it "
        "copied that job to the next day at the same time.\n\n"
        "The new scheduler does not have this button. The 5 August design review asked to bring it "
        "back for the first version.\n\n"
        "Your newer written description does not mention carryover at all, so we are asking you "
        "rather than guessing.",

        "Should the one-click carryover button - copy a job to the next day at the same time - be "
        "brought back in the first version?",

        "A) YES - bring the carryover button back in the first version.\n\n"
        "B) NO - leave it out for now.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the schedule grid - what the carryover button is called",

        "This one is only about what the button is CALLED, if it comes back (see the previous "
        "question).\n\n"
        "The old name was 'Carryover'. The design review suggested a clearer name - 'Add a Day' or "
        "'Extend a Day' - but the final wording is not decided.",

        "If the button comes back, what should it be called?",

        "A) 'Add a Day'\n\n"
        "B) 'Extend a Day'\n\n"
        "C) Keep 'Carryover', or something else - please write the exact words you want.",
    ),
    (
        "SCHEDULE - the schedule grid - carryover on a job that runs several days",

        "Some jobs are scheduled across several days in a row.\n\n"
        "When you press carryover on one of these, there are two ways it could behave: it could add "
        "just ONE more day at the end, or it could copy the whole run of days again.\n\n"
        "The design review said it should add just one more day. Your newer written description "
        "does not cover this, so we are checking with you.",

        "When you press carryover on a job that already runs across several days, should it add just "
        "one more day, or copy the whole run of days again?",

        "A) Add just ONE more day at the end.\n\n"
        "B) Copy the whole run of days again.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the schedule grid - week view - dragging a job to the next day",

        "In week view you can see several days at once.\n\n"
        "The design review suggested that - as well as, or instead of, a carryover button - a "
        "person should be able to drag a scheduled job straight onto the next day.\n\n"
        "Your newer written description does not mention this, so it is your call.",

        "In week view, should a person be able to drag a scheduled job straight onto the next day, "
        "as another way to carry it over?",

        "A) YES - allow dragging a job onto the next day in week view.\n\n"
        "B) NO - keep a carryover button as the only way.\n\n"
        "C) Something else - please describe it.",
    ),
]

# ================================================================ QA-ONLY MAPPING
# Tab, Item, What it asks, Where the question comes from,
# Affected internal case IDs (C-id), TestRail links, Spec anchors + evidence, resolution
LINK = "https://shopview.testrail.io/index.php?/cases/view/"
QA_ROWS = [
    (
        "Questions", "1",
        "Filters: Status chip on Estimates/Completed tabs - hidden (spec v21 S9-R5) vs greyed-out "
        "pre-filled (Branko 17 Jul + QA-lead 30 Jul + design).",
        "FAB-2 (FILTERS-COMPLETION-REPORT OUTSTANDING item 2; COVERAGE-REDERIVATION; register FAB-2). "
        "Verdict HELD per Rule 33 - a recorded QA-lead ruling is not silently reversed.",
        "FLT-TAB-02 = C29609 ; FLT-TAB-03 = C29610",
        f"{LINK}29609 ; {LINK}29610",
        "Filters spec Confluence v21 (14 Aug 2026) S9-R5 (Status chip All-tab only; hidden on "
        "Estimates/Completed) vs Branko answer 2026-07-17 Q4 + QA-lead ruling 2026-07-30 "
        "(greyed-out/pre-filled) + design. Cases currently assert the July/30-Jul ruling.",
        "A -> the two cases move to 'Status hidden' (spec wins, July answer retired). "
        "B -> cases stay as written (greyed/pre-filled) AND a defect is raised so the build+"
        "description are corrected. C -> re-scope on his description.",
    ),
    (
        "Questions", "2",
        "Filters: the per-view filter list - exactly which filter chips appear on each Parts page "
        "and each Report page.",
        "FAB-1 (FILTERS-COMPLETION-REPORT OUTSTANDING item 1; COVERAGE-REDERIVATION 'per-view "
        "filter list PENDING from engineering'; register FAB-1).",
        "FLT-PARTS-01 = C38904 ; FLT-PARTS-09 = C38905 ; FLT-PARTS-11 = C38906 ; "
        "FLT-PARTS-12 = C38907 ; FLT-PARTS-13 = C38908 ; FLT-PARTS-14 = C43562 ; "
        "FLT-RPTS-01 = C38909 ; FLT-RPTS-21 = C38910 ; FLT-RPTS-22 = C38911 ; FLT-RPTS-23 = C38882",
        f"{LINK}38904 ; {LINK}38905 ; {LINK}38906 ; {LINK}38907 ; {LINK}38908 ; {LINK}43562 ; "
        f"{LINK}38909 ; {LINK}38910 ; {LINK}38911 ; {LINK}38882",
        "Filters spec v21 S1-R7 / S1-R8 / S13-R23 (relocate existing filters, none added/removed; "
        "per-view list owed by engineering, not in the spec). Cases are behavioural + 'confirm the "
        "exact buttons live'.",
        "A -> the pre-redesign per-page filter set is the baseline; cases confirm that set. "
        "B -> engineering per-view list becomes the baseline; cases check each page against it. "
        "C -> re-scope on his answer.",
    ),
    (
        "Questions", "3",
        "Schedule: shift-detail modal - Delete only (spec v30 §4.9, Reassign removed at v23) vs a "
        "Reassign action. Genuine open PO product decision.",
        "SCHEDULE-COMPLETION-REPORT 'Still deferred / flagged' item 2 + OUTSTANDING item 6; "
        "register SCH-FR-6. Left untouched per QA-lead instruction ('a genuine PO product decision "
        "... DO NOT act on it; keep it flagged').",
        "SCH-MODAL-08 = C30015",
        f"{LINK}30015",
        "Schedule spec Confluence v30 §4.9 (Actions - Delete only; Reassign removed in v23) and §7 "
        "(stacking order). Drag-to-reassign is covered separately (SCH-REAS-01). The case asserts "
        "Delete-only; the open question is whether Reassign belongs in the modal.",
        "A -> case stays as written (Delete only; drag reassigns) - confirms spec v30. "
        "B -> a Reassign action is added; the case is updated and a build/spec change follows. "
        "C -> re-scope on his answer.",
    ),
    (
        "Questions", "4 (NEW)",
        "Schedule: a user-level 'always schedule the whole work order' preference so the scope "
        "picker remembers the choice and stops asking each time.",
        "aug5-review-2026-08-17/RECONCILIATION.md E6 (DIVERGENCE / PO-Q; OUTSTANDING item 2). The "
        "Aug-5 review lists it as an open question to decide before V1; Schedule spec v30 is SILENT "
        "(0 hits).",
        "None - no V1 case authored (spec v30 silent; Rule 58 - not authored from silence).",
        "—",
        "Aug-5 design review E6 ('Yes - open question') vs Schedule spec Confluence v30 (13 Aug "
        "2026) - silent on a whole-WO preference. Newer v30 wins (Rule 32); the review item is "
        "unratified.",
        "A -> author one case for the preference per his answer. B -> confirm dropped for V1 (no "
        "case; close the review item). C -> re-scope on his answer.",
    ),
    (
        "Questions", "5 (NEW)",
        "Schedule: restore the one-click carryover button (duplicates a scheduled block to the next "
        "day at the same time).",
        "RECONCILIATION.md E15 (was bug B3; DIVERGENCE / PO-Q; OUTSTANDING item 1). Spec v30 SILENT "
        "on carryover entirely (0 hits).",
        "None - no V1 case authored (spec v30 silent; Rule 58).",
        "—",
        "Aug-5 review E15 ('Yes') vs Schedule spec v30 - no mention of carryover / add-a-day / "
        "extend. Rule 32 - v30 wins; Rule 58 - not authored from silence.",
        "A -> author carryover-button case(s). B -> confirm dropped; close the item. C -> re-scope.",
    ),
    (
        "Questions", "6 (NEW)",
        "Schedule: rename 'Carryover' to 'Add a Day' / 'Extend a Day' (final wording TBC).",
        "RECONCILIATION.md E7 (DIVERGENCE / PO-Q; OUTSTANDING item 1). Spec v30 SILENT.",
        "None - depends on E15 (Q5); no case authored (Rule 58).",
        "—",
        "Aug-5 review E7 ('Yes - TBC') vs spec v30 silent. Label decision only; contingent on the "
        "carryover button (Q5) existing.",
        "A / B -> the chosen label is used when authoring the carryover case. C -> use his exact "
        "words.",
    ),
    (
        "Questions", "7 (NEW)",
        "Schedule: for a multi-day shift, carryover should add ONE more day, not duplicate the "
        "whole span.",
        "RECONCILIATION.md E8 (DIVERGENCE / PO-Q; OUTSTANDING item 1). Spec v30 SILENT.",
        "None - depends on E15 (Q5); no case authored (Rule 58).",
        "—",
        "Aug-5 review E8 ('Yes' - extend one day) vs spec v30 silent on carryover behaviour.",
        "A -> author the carryover case to add exactly one day. B -> whole-span duplicate. C -> "
        "re-scope.",
    ),
    (
        "Questions", "8 (NEW)",
        "Schedule: allow dragging a shift onto the next day directly in week view, as an "
        "alternative to the carryover button.",
        "RECONCILIATION.md E9 (DIVERGENCE / PO-Q; OUTSTANDING item 1). Spec v30 SILENT on week-view "
        "cross-day drag as carryover (reassignment across tech rows IS covered separately).",
        "None - no V1 case for week-view carryover-by-drag. Reassignment across tech rows is "
        "covered by SCH-REAS-01 = C30052 (adjacent, not the same behaviour). Rule 58.",
        f"{LINK}30052 (adjacent - drag reassign across tech rows, not cross-day carryover)",
        "Aug-5 review E9 ('Yes') vs spec v30 - week-view cross-day drag is not a v30 requirement.",
        "A -> author a week-view drag-to-next-day case. B -> carryover button only. C -> re-scope.",
    ),
]

QA_NOTES = [
    "SEND STATUS: questions 1 to 3 were SENT to Branko on 2026-08-17 (per the QA lead). Questions "
    "4 to 8 are NEW FOLLOW-UPS added 2026-08-18 and are HELD pending the QA lead sending them "
    "(Standing Rule 66 - the question sheet is the last thing sent). A follow-up-only copy of the "
    "five new questions is 'Filters-and-Schedule_Questions-for-Branko_FOLLOW-UP-5_2026-08-17.docx'. "
    "Nothing on this sheet has been written to TestRail or Jira.",
    "ONLY PRODUCT DECISIONS (Rule 7): only genuine PRODUCT DECISIONS for the PO are on this sheet "
    "- never bugs. Bugs/defects go to dev tickets, not to Branko.",
    "FOLLOW-UPS ADDED 2026-08-18 (questions 4-8): the five Schedule carryover / whole-WO-preference "
    "items (E6, E15, E7, E8, E9) that the 5 August design review marked V1 but Schedule spec v30 "
    "(13 Aug 2026) is SILENT on. Source: build/schedule/aug5-review-2026-08-17/RECONCILIATION.md "
    "(OUTSTANDING items 1-2) + build/OUTSTANDING-ITEMS-REGISTER.md (Schedule / Branko, rows "
    "SCH-A5-1..5). Not authored from silence (Rule 58); Branko's decision is what unblocks "
    "authoring. Also tracked: whichever way he answers, only the QA lead sends the sheet.",
    "SOURCE CURRENCY (Rule 31): the eight items were reconciled against Filters spec Confluence "
    "v21 (14 Aug 2026) and Schedule spec Confluence v30 (13 Aug 2026) during the 2026-08-17 "
    "Fabian-review and 2026-08-18 Aug-5-review passes. THIS GENERATOR MADE NO LIVE CONFLUENCE READ "
    "- it consolidates already-established open items and the follow-ups are HELD. OWED BEFORE "
    "SENDING (Rule 59): re-read both live spec versions and each question's anchor immediately "
    "before the send, and re-verify none has been answered by a spec edit in the meantime.",
    "DELIBERATELY NOT ON THIS SHEET: (a) the Schedule shop-closures / multi-day-spread contradiction "
    "- ANSWERED by spec v30 (skip weekends only, closures receive shifts); cases already follow v30, "
    "so it is not a Branko question. (b) The greyed-vs-hidden question is the ONLY Filters "
    "tab-behaviour item still open - everything else the redesign changed is settled by spec v21. "
    "(c) Aug-5 review items already covered by spec v30 (E1-E5, E10-E12) or confirmed out-of-scope "
    "(E13, E14, E16) are not questions - only the 5 divergences v30 is silent on are asked. "
    "(d) No open Global Search product decision was found in the current passes.",
]


def _sheet_questions(wb):
    ws = wb.active
    ws.title = TAB1
    ws["A1"] = "Filters and Schedule - questions for Branko Cicovic - 2026-08-17"
    ws["A1"].font = Font(bold=True)
    ws["A2"] = HELLO
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 150
    r = 4
    ws.cell(row=r, column=1,
            value="Eight questions - 1 to 3 already sent; 4 to 8 are new follow-ups. Please answer "
                  "A / B / C or one line.").font = Font(bold=True)
    r += 1
    cols = ["#", "Which project and screen", "What happens now", "The question", "Options",
            "Your answer"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
    r += 1
    for i, (topic, now, q, opts) in enumerate(QUESTIONS, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 320
        r += 1
    for col, wd in zip("ABCDEF", [4, 40, 60, 44, 50, 22]):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A6"


def _sheet_qa(wb):
    ws = wb.create_sheet(TABQA)
    ws["A1"] = ("QA-ONLY - INTERNAL - NOT FOR BRANKO. Do not send this tab. TestRail case IDs, "
                "requirement anchors and evidence live here so the tab Branko reads stays plain "
                "(Standing Rules 7, 8 and 55).")
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = WRAP
    r = 3
    hdr = ["Tab", "Item", "What it asks", "Where the question comes from",
           "Affected internal case IDs (TestRail C-id)", "TestRail links",
           "Spec anchors + live evidence", "What each answer resolves to"]
    for j, c in enumerate(hdr, 1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, wd in zip("ABCDEFGH", [10, 6, 42, 42, 46, 50, 60, 60]):
        ws.column_dimensions[col].width = wd


def write_xlsx():
    wb = openpyxl.Workbook()
    _sheet_questions(wb)
    _sheet_qa(wb)
    wb.save(XLSX)
    return XLSX


def block(items):
    out = []
    for i, (topic, now, q, opts) in enumerate(items, 1):
        out.append(f"### {i}. {topic}\n")
        out.append("**What happens now**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
        out.append("**The question**\n")
        out.append(f"> {q}\n")
        out.append("**Options**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
        out.append("**Your answer:** _______________________________________________\n")
    return "\n".join(out)


def write_md():
    md = f"""# Questions for Branko Cicovic — Filters and Schedule — 2026-08-17

**Projects: Filters (epic SV-8785) and Schedule (epic SV-8685) · Product Owner: Branko Cicovic**

*Eight genuine product decisions left open after the 2026-08-17 filter-redesign reconciliation
(Filters spec v21, Schedule spec v30) and the 2026-08-18 Aug-5-design-review reconciliation. One
sheet, per Standing Rule 55. The spreadsheet twin is
`Filters-and-Schedule_Questions-for-Branko_2026-08-17.xlsx`; it carries a QA-only tab that must not
be forwarded.*

**QUESTIONS 1 to 3 were SENT to Branko on 2026-08-17 (per the QA lead). QUESTIONS 4 to 8 are NEW
FOLLOW-UPS added 2026-08-18, HELD pending the QA lead sending them (Standing Rule 66: the question
sheet is the LAST thing sent). A follow-up-only copy of the five new questions, for easy sending, is
`Filters-and-Schedule_Questions-for-Branko_FOLLOW-UP-5_2026-08-17.docx`. Nothing has been written to
TestRail or Jira.**

---

{HELLO}

---

## The questions

{block(QUESTIONS)}

---

## QA-only — not for Branko

The question-to-case mapping (internal IDs, TestRail C-ids and links, spec anchors, and what each
answer resolves to) is on the spreadsheet's `{TABQA}` tab, together with the source-currency record
and the list of what was deliberately left off and why.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
