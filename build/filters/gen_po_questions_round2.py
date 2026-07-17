#!/usr/bin/env python3
"""Generate the Filters ROUND-2 PO question sheet for Branko (Filters PO).

Outputs (regenerable — run from the repo root):
  - build/filters/PO-Questions-Filters-Round2_2026-07-17.xlsx
  - build/filters/PO-Questions-Filters-Round2_2026-07-17.md

Sheets / sections (mirrors build/filters/gen_po_questions.py = the round-1
sheet, which itself mirrors build/simple-flow/gen_po_questions_round3.py):
  - "Questions for PO"    : reader-facing, VERY SIMPLE layman language ONLY.
                            NO case IDs, FLT- codes, API/HTTP terms, or jargon.
                            Columns: # | Topic | What happens now | The question |
                            Options | Your answer (blank).
  - "QA Internal Mapping" : QA-only. Per question, the affected FLT- case IDs
                            WITH their TestRail C-ids + links (id-map populated
                            79/79 — testrail-id-map.csv, Standing Rule 8),
                            spec/design refs, and what each answer option
                            resolves to.

Only genuine PRODUCT DECISIONS are included (Standing Rule 7) — no bugs/defects
(the feature is still pre-VIU), no env questions (those are QA-side OQ-2/3/6/7
asked of the user, not the PO).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/filters/PO-Questions-Filters-Round2_2026-07-17.xlsx"
MD_OUT = "build/filters/PO-Questions-Filters-Round2_2026-07-17.md"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

TR = "https://shopview.testrail.io/index.php?/cases/view/"

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no codes, no tech terms)
# ---------------------------------------------------------------------------
questions = [
    {
        "topic": "Two older sentences in the write-up to correct",
        "now": ("When you update the write-up to add the other pages (as you "
                "mentioned), two older sentences in it no longer match your "
                "decisions. First, the write-up still says the Status choice "
                "is hidden on the Estimates and Completed tabs - but your "
                "decision was that it shows greyed-out and pre-filled there. "
                "Second, the write-up still says your chosen filters are kept "
                "only until the browser is closed - but your decision was "
                "that they are remembered permanently."),
        "question": ("When you update the write-up, will you also correct "
                     "these two places so the write-up matches your "
                     "decisions?"),
        "opts": ("A) Yes - I'll fix both sentences in the new write-up.\n"
                 "B) Something else (please explain)."),
    },
    {
        "topic": "A status called \"Reported\" in the interactive demo",
        "now": ("In the interactive design demo you shared, the list of "
                "work-order statuses shows a status called \"Reported\". The "
                "Figma design shows \"Imported\" instead - and \"Reported\" "
                "doesn't exist there."),
        "question": ("Which is correct - \"Imported\" or \"Reported\"?"),
        "opts": ("A) \"Imported\" is correct - the demo has a typo.\n"
                 "B) \"Reported\" is a real status - please tell us more."),
    },
    {
        "topic": "Do the filter lists depend on the user's role?",
        "now": ("The write-up doesn't say whether what a person can pick in "
                "the filters depends on their role. For example, should a "
                "technician see every customer and advisor in the filter "
                "lists, or only what their role allows?"),
        "question": ("Do the filter lists depend on the user's role?"),
        "opts": ("A) No - everyone sees the same filter options.\n"
                 "B) Yes - some roles see fewer options; please describe "
                 "which."),
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (Standing Rule 8: internal id + TestRail C-id + link;
# id-map POPULATED 79/79 — build/filters/testrail-id-map.csv).
# Each entry: (q_no, [internal ids w/ C-ids + links], refs, resolves_to)
# ---------------------------------------------------------------------------
internal_map = [
    (1,
     [f"FLT-TAB-02 (C29609 - {TR}29609)",
      f"FLT-TAB-03 (C29610 - {TR}29610)",
      f"FLT-PERS-02 (C29614 - {TR}29614)"],
     "Stale spec passages: S2-N1/S2-N2 + S9-R2/S9-R3 ('Status chip hidden on "
     "Estimates/Completed' - superseded by Branko 2026-07-17 answer Q4=B: "
     "disabled pre-filled chip) and S10-R2 ('kept for the duration of the "
     "browser session' - superseded by answer Q2=B: permanent per-user "
     "persistence). Cases C29609/C29610/C29614 were ALREADY updated + pushed "
     "to TestRail 2026-07-17 (branko-answers-2026-07-17/testrail-update-log.md); "
     "this question is a spec-cleanup reminder riding on his announced PRD "
     "update (Q1=A - Parts/Reports sections coming).",
     "A -> no QA action beyond re-checking the new PRD text against the "
     "already-updated cases when it arrives. "
     "B -> reconcile whatever he says against the Q2=B/Q4=B rulings "
     "(last-update-wins) and re-open the 3 cases only if he reverses himself."),
    (2,
     ["FLT-STAT status-list cases, esp. "
      f"FLT-STAT-01 (C29560 - {TR}29560)"],
     "Design-system zip prototype status-list anomaly "
     "(new-inputs-inventory-2026-07-17.md; PROJECT-STATE WHAT'S-LEFT item 4): "
     "the coded prototype's 9-status list ends in 'Reported' where the Figma "
     "9-status list (design-notes.md) ends in 'Imported'; 'Reported' exists "
     "nowhere in the Figma frames. Zip = reference prototype only, not "
     "authoritative frames.",
     "A -> no case change (cases are authored to the Figma/spec list with "
     "'Imported'); note the demo typo and move on. "
     "B -> update the status option list in the FLT-STAT cases (and any case "
     "enumerating the 9 statuses) to include 'Reported' per Branko's detail; "
     "TestRail edits need fresh user authorization."),
    (3,
     [f"FLT-CUST-01 (C29566 - {TR}29566) + FLT-CUST dropdown-content cases",
      f"FLT-TECH-01 (C29575 - {TR}29575) + FLT-TECH dropdown-content cases",
      f"FLT-ADV-01 (C29582 - {TR}29582) + FLT-ADV dropdown-content cases"],
     "requirements.md OQ-4 (permissions): the spec has NO permissions/role "
     "section - S1 prerequisite is only 'The user has access to the Work "
     "Orders page'. Affects what the Customer / Lead Technician / Service "
     "Advisor dropdowns list per role.",
     "A -> OQ-4 closed; dropdown-content cases stay role-agnostic; at VIU "
     "verify the lists once as any role. "
     "B -> author new role-based filter cases at VIU (per-role dropdown "
     "scoping), extend the dropdown-content cases with role preconditions; "
     "add_case push needs fresh user authorization."),
]


def build_xlsx():
    wb = Workbook()

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = ("Filters (Work Orders page) - Questions for Branko - Round 2 "
                "- 2026-07-17")
    ws["A1"].font = TITLE_FONT
    ws.append([])
    hdr = ["#", "Topic", "What happens now", "The question", "Options",
           "Your answer"]
    ws.append(hdr)
    for c in ws[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = BORDER
    for i, q in enumerate(questions, 1):
        ws.append([i, q["topic"], q["now"], q["question"], q["opts"], ""])
    for row in ws.iter_rows(min_row=4, max_row=3 + len(questions)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    widths = [4, 22, 46, 42, 46, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids "
                 "from build/filters/testrail-id-map.csv (populated 79/79).")
    ws2["A1"].font = Font(bold=True, color="C00000")
    ws2.append([])
    hdr2 = ["Q#", "Affected internal case IDs (TestRail C-id + link)",
            "Spec / design refs", "What each answer resolves to"]
    ws2.append(hdr2)
    for c in ws2[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = BORDER
    for qno, ids, refs, resolves in internal_map:
        ws2.append([qno, "\n".join(ids), refs, resolves])
    for row in ws2.iter_rows(min_row=4, max_row=3 + len(internal_map)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    for i, w in enumerate([4, 34, 52, 60], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A4"

    wb.save(XLSX_OUT)
    print("Wrote", XLSX_OUT)


def build_md():
    lines = [
        "# Filters (Work Orders page) — Questions for Branko — Round 2 — "
        "2026-07-17",
        "",
        "Plain-language product questions only (no bugs, no test jargon).",
        "Please pick an option (or write your own answer) for each.",
        "",
    ]
    for i, q in enumerate(questions, 1):
        lines += [
            f"## Question {i} — {q['topic']}",
            "",
            f"**What happens now:** {q['now']}",
            "",
            f"**The question:** {q['question']}",
            "",
            "**Options:**",
            "",
        ]
        lines += [f"- {o}" for o in q["opts"].split("\n")]
        lines += ["", "**Your answer:** ____________________", ""]
    lines += [
        "---",
        "",
        "## QA Internal Mapping (QA-only — not for the PO)",
        "",
        "TestRail C-ids from `build/filters/testrail-id-map.csv` "
        "(populated 79/79, Standing Rule 8).",
        "",
        "| Q# | Affected internal case IDs | Spec / design refs | Resolves to |",
        "|---|---|---|---|",
    ]
    for qno, ids, refs, resolves in internal_map:
        lines.append(
            f"| {qno} | {', '.join(ids)} | {refs} | {resolves} |")
    lines.append("")
    open(MD_OUT, "w").write("\n".join(lines))
    print("Wrote", MD_OUT)


if __name__ == "__main__":
    build_xlsx()
    build_md()
