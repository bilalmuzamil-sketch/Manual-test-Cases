#!/usr/bin/env python3
"""CONSOLIDATED Branko Cicovic question sheet — Filters + Schedule — 2026-08-13.

WHY THIS FILE EXISTS. Standing Rule 55 says every open ambiguity goes onto ONE sheet so the
product owner answers in a single sitting. Since the 2026-08-06 friendly sheet was written and
HELD (Standing Rule 66 — sheets go out LAST, on the QA lead's word), TEN more questions have
accumulated in four separate pass folders, each of which says on its face "add this to the
6 August sheet rather than send it separately":

  * build/filters/questions-2026-08-06/ADDENDUM-2026-08-11_SV-9041-collapse-toggle.md  (2 items)
  * build/schedule/coverage-rederivation-2026-08-10/QUESTIONS-FOR-BRANKO.md  (S-1, S-2)
  * build/schedule/panel-collapse-2026-08-11/QUESTIONS-FOR-BRANKO.md  (S-2 restated, cited on C43587)
  * build/schedule/coverage-gaps-2026-08-11/QUESTIONS-FOR-BRANKO.md  (S-1/S-2 restated + honesty notes)
  * build/handover-ingest-2026-08-10/QUESTIONS.md  (B-1 .. B-6)

This generator IMPORTS the 22 items of the 2026-08-06 Friendly sheet from its own generator
(gen_branko_friendly.py, which itself imports gen_branko_sheet.py) so the wording CANNOT drift,
and appends the ten new items as Section 5, plus their QA-only mapping rows.

ONE DATED CORRECTION is applied to an imported QA-only row: the cross-filter row (Section 3
item 7) said the technical-design authority question was OPEN. It was ANSWERED by the QA lead on
2026-08-12; the correction is APPENDED to that row, dated, with the superseded wording kept.

PRODUCE ONLY — THIS SHEET IS WRITTEN AND HELD, NOT SENT (Standing Rule 66, QA lead 2026-08-12,
verbatim: "This should be the last thing once you give me the report that everything else has
been done only this part is left...").

RESEARCH ONLY — writes two files into this folder. No TestRail write, no Jira call, no
application call, no Confluence call.
"""

import importlib.util
import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
PRIOR_FRIENDLY = os.path.normpath(os.path.join(HERE, "..", "questions-2026-08-06",
                                               "gen_branko_friendly.py"))

XLSX = os.path.join(HERE, "Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-13.xlsx")
MD = os.path.join(HERE, "Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-13.md")

_spec = importlib.util.spec_from_file_location("friendly_branko", PRIOR_FRIENDLY)
friendly = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(friendly)

HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

TAB1 = "1 Start here"
TAB2 = "2 Decisions"
TAB3 = "3 When you have time"
TAB4 = "4 Just a heads-up"
TAB5 = "5 Added 13 August"
TABQA = "QA internal - not for Branko"

HELLO = (
    "Hello Branko - this is everything we have open across TWO of your projects, FILTERS and "
    "SCHEDULE, gathered into one place so you can go through it in a single sitting instead of "
    "getting a trickle of separate messages. Thirty-two items; about half an hour if you go "
    "straight down the list. SHORT ANSWERS ARE PERFECT - a letter, or one line. Nothing here "
    "needs an essay.\n\n"
    "WHERE TO START. Section 1 is five questions that release tests which are stuck today - that "
    "is the part we are genuinely waiting on. Section 2 is nine ordinary decisions. Section 3 is "
    "seven things only the engineering plan describes, and nothing of ours is waiting on those, "
    "so they can keep for a quiet moment. Section 4 is one typo-level heads-up with nothing to "
    "decide. Section 5 is ten questions added on 13 August - everything that has come up since "
    "the earlier sections were written, so you get one sheet rather than several.\n\n"
    "ONE OF THESE IS OUR OWN FAULT and we are sorry: question 2 in Section 1 was written on "
    "22 July and we never actually sent it to you. Two tests have been parked ever since waiting "
    "for an answer you were never asked for.\n\n"
    "Every question says which project it belongs to, because you look after Filters, Schedule "
    "and Global Search. And to be clear - we have not edited any of your tickets or your "
    "descriptions. Where two of your own documents disagree we simply say so and ask which one "
    "to keep."
)

# ================================================================= SECTION 5 — NEW 2026-08-13
# Items carried from the four post-06-August pass folders named in the module docstring.
# Wording carried from those files; only trimmed to the row shape.
SEC5 = [
    (
        "FILTERS - the filter buttons across Work Orders, Parts and Reports - the show/hide "
        "control for the filter row",

        "Most list pages have a row of filter buttons, and a small control in the toolbar that "
        "hides that row to give the table more space.\n\n"
        "On 7 August one of your team raised a ticket saying that control should only appear when "
        "a page has more than one filter. If a page has only one filter, the control should not "
        "be there at all and that page's filter row should simply always be on display. Your QA "
        "has since checked it and confirmed the product already behaves that way.\n\n"
        "That rule is not in your written description. The description says, and has said "
        "unchanged since 13 May, only that the toolbar contains a control that hides and shows "
        "the filter row - it does not mention any condition about how many filters the page has.\n\n"
        "Why we are asking rather than choosing: we have followed the ticket, because it is the "
        "newer statement, and we have updated two tests so a page with one filter and no control "
        "is treated as correct rather than as a fault. But the description is the document QA "
        "works from, and right now it does not contain this rule.",

        "Should the show/hide control for the filter row be hidden on pages that have only one "
        "filter, and should that rule go into your written description?",

        "A) YES - the rule is correct, and please add it to the description so it is written "
        "down.\n\n"
        "B) YES, the rule is correct, but leave the description as it is - the ticket is "
        "enough.\n\n"
        "C) NO - the control should always be there whatever the page has on it, and the ticket "
        "is wrong.\n\n"
        "D) Something else - please describe it.",
    ),
    (
        "FILTERS - the same question, for the Parts pages and the Reports pages",

        "The ticket does not say which pages it covers. It says \"the page\", which reads as all "
        "of them.\n\n"
        "The evidence your QA attached to that ticket is a screenshot of the Part Sales page - a "
        "Parts page with a single filter and no show/hide control - so in practice it has already "
        "been treated as covering Parts.\n\n"
        "This matters to us because you told us on 31 July that hiding the filter row on Parts "
        "and Reports works the same way as it does on the Work Orders list. The Work Orders list "
        "has five filters, so the control is always there. Some Parts pages and some reports have "
        "only one, so on those the control would now be absent.",

        "Does this rule apply to the Parts pages and the Reports pages too, and not only the "
        "Work Orders list?",

        "A) YES - it applies everywhere there is a filter row.\n\n"
        "B) NO - it applies only to the Work Orders list; Parts and Reports always show the "
        "control.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the technician calendar - the small pop-up on the workload bar above each day",

        "Above every day on the calendar there is a small bar showing how full the shop's day "
        "is. When you rest your mouse on that bar, a little pop-up lists technicians and how many "
        "hours each one has been given against the hours they are available for.\n\n"
        "On 7 August your description changed by one word: it used to say the pop-up lists the "
        "technicians, and now it says it lists the technicians who have work assigned.\n\n"
        "In a shop with fifteen technicians where only three have work that day, that is the "
        "difference between a fifteen-line pop-up and a three-line one.\n\n"
        "What we have done in the meantime, so it is not a surprise: our test now expects only "
        "the assigned technicians, because that is what your description says today. If your "
        "answer is B, we change one test back - it is a five-minute fix and no other test "
        "depends on it.",

        "Was that change what you meant - should the pop-up list only the technicians who have "
        "work assigned that day?",

        "A) YES - only the technicians who have work assigned that day.\n\n"
        "B) NO - it should list all technicians, including those with nothing booked, and the "
        "change was a slip.",
    ),
    (
        "SCHEDULE - the technician calendar - whether the calendar remembers that you hid the "
        "job list",

        "Your description of 7 August adds a new button to the calendar that hides and shows the "
        "job list down the left-hand side, giving its space to the calendar. About whether the "
        "calendar remembers that choice, it says the setting lasts only while you are signed in - "
        "so if you hide the list, sign out, and sign back in tomorrow, the list is showing "
        "again.\n\n"
        "Separately, the design review of 5 August asks for the calendar's view settings to be "
        "remembered for each person even after they sign out and come back.\n\n"
        "Those two are different promises, and we do not want to guess which you meant. This is "
        "not the same as the question about remembering the other view settings, which is asked "
        "separately below - we are asking about this one button because your description and the "
        "design review give different answers for it.",

        "When someone hides the job list and then signs out, what should they see the next time "
        "they sign in?",

        "A) The job list is SHOWING AGAIN. Hiding it only lasts for the sign-in you are in. "
        "(This is what your 7 August description says.)\n\n"
        "B) The job list is STILL HIDDEN. The calendar remembers it for that person from one "
        "sign-in to the next. (This is what the 5 August design review asks for.)",
    ),
    (
        "FILTERS - the Work Orders list - whether the filter dropdowns should close when you "
        "pick something",

        "The five filter buttons on the Work Orders list open a small panel when you click them. "
        "They do not all behave the same way when you make a choice.\n\n"
        "The ones where you can tick several things - Customer, Lead Technician, Service Advisor, "
        "Status - stay open, so you can tick a second and a third. That is what we would "
        "expect.\n\n"
        "But Asset on Site, where you can only pick one answer, closes the moment you pick. And "
        "the date panel on the report pages closes when you pick a ready-made period, but stays "
        "open while you are typing your own dates.\n\n"
        "Your written description gives one rule for all of them - the panel closes when you "
        "click outside it - and does not mention closing when you choose something.\n\n"
        "Why we are asking: none of our tests is wrong today, because none of them says either "
        "way. But a tester will notice that the buttons behave differently from each other, and "
        "we would rather have your answer than let them guess.",

        "Should a filter panel where you can only pick ONE answer close by itself as soon as you "
        "pick?",

        "A) Yes - a one-choice panel closes as soon as you pick. Please have the description say "
        "so.\n\n"
        "B) No - every panel should stay open until you click outside it, as the description "
        "says today.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the technician calendar - the wording of the warning when a shift falls "
        "outside someone's hours",

        "When a shift is put before or after someone's working day, the calendar shows a "
        "warning. Today that warning says \"working hours\".\n\n"
        "A fault raised after the design review on 5 August asks for it to say \"business "
        "hours\" instead, because that is the wording used elsewhere in the product.\n\n"
        "Here is our worry, and it is why we have not simply changed our tests. Your own "
        "description treats those two as DIFFERENT things. It says a technician's own hours come "
        "first, and the shop's business hours are only used when that technician has no hours of "
        "their own.\n\n"
        "So if the warning is changed to say \"business hours\" for everybody, it will be wrong "
        "for any technician who has their own hours set - it would blame the shop's hours while "
        "actually measuring against the technician's.",

        "When a shift falls outside someone's hours, what should the warning say?",

        "A) It should refer to THAT TECHNICIAN'S hours, because those are what it measures "
        "against.\n\n"
        "B) It should say \"business hours\" for everyone - then your description needs changing "
        "to match.\n\n"
        "C) It should avoid both and just say something like \"outside the working day\".\n\n"
        "D) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the technician calendar - the \"Add Existing Work Order\" button",

        "A button called \"Add Existing Work Order\" appears in the drawing of the Schedule, but "
        "it is not in the product. It was raised as a fault after the design review on 5 August, "
        "and that report says openly that nobody is sure whether it was dropped while building or "
        "never planned at all. The fault has been parked since.\n\n"
        "We searched your written description for it and it is not there - not in the current "
        "version and not in any earlier one we hold.\n\n"
        "Why we are asking rather than writing a test: a drawing on its own is not enough for us "
        "to say the product must do something. If we wrote the test now we would be inventing a "
        "requirement, and if the button was never planned, that test would fail forever for no "
        "reason.",

        "Should there be an \"Add Existing Work Order\" button on the Schedule in this release?",

        "A) Yes - it was meant to be there. (Then we will write the test, and it should go in "
        "your description.)\n\n"
        "B) No - it was never planned, or it has been dropped. (Then nothing more is needed from "
        "us.)\n\n"
        "C) Not in this release, but later - please say roughly when.",
    ),
    (
        "SCHEDULE - the technician calendar - how many hours get planned for a job that is half "
        "done",

        "When a big job is spread across several days, the calendar decides how many hours to "
        "plan.\n\n"
        "Your description is clear about this today: it plans the WHOLE original estimate every "
        "time, and it says in as many words that planned hours, the estimate and the hours "
        "actually worked are three separate numbers that are not made to add up.\n\n"
        "The design review of 5 August asks for the opposite: that when a job is partly "
        "finished, planning should use the hours REMAINING rather than the original estimate. "
        "That review lists it as in scope for this release.\n\n"
        "Our test follows your description. We have not changed it, because your description was "
        "updated on 7 August - two days after that review - and still says the original "
        "estimate.",

        "When a job is already partly finished, should the calendar plan the hours remaining, or "
        "the whole original estimate?",

        "A) THE WHOLE ORIGINAL ESTIMATE - as your description says today. Nothing changes.\n\n"
        "B) THE HOURS REMAINING - then your description needs changing, and we will change our "
        "test to match.\n\n"
        "C) Something else - please describe it.",
    ),
    (
        "SCHEDULE - the technician calendar - whether the view settings are remembered",

        "The Schedule has a small settings panel that turns things on and off - the capacity "
        "bars along the top, which departments are shown, whether events appear, and so on.\n\n"
        "Your description tells us which of those start switched on and which start switched "
        "off. It does not say whether the product remembers a person's choices for next time.\n\n"
        "The design review of 5 August asks for them to be remembered for each person, and lists "
        "it as in scope for this release, but marks the details as still to be worked out.\n\n"
        "Why we are asking: we have no test for this, and we do not want to write one that says "
        "\"remembered\" when the only thing asking for it is a meeting note.",

        "Should each person's Schedule view settings be remembered for next time?",

        "A) YES - remembered for that person, and they should still be set that way after "
        "signing out and back in.\n\n"
        "B) NO - they go back to their starting positions each time.\n\n"
        "C) Remembered only until they close the browser, not beyond that.",
    ),
    (
        "SCHEDULE - the technician calendar - dragging a shift onto the next day",

        "In the week view you can drag a shift from one technician to another, and your "
        "description covers that.\n\n"
        "The design review of 5 August also asks that you be able to drag a shift onto the NEXT "
        "DAY for the same technician, as a quicker alternative to a button. It lists it as in "
        "scope for this release but marks the details as still to be worked out.\n\n"
        "Your description does not mention moving a shift to a different day. It only mentions "
        "moving one between technicians. It does allow meetings and other non-job blocks to be "
        "moved between days - but that is a different kind of block.\n\n"
        "We have no test for it, and we would rather ask than invent one.",

        "Should someone be able to drag a shift onto a different day for the same technician?",

        "A) YES - dragging a shift onto another day moves it there. (Then we will write the "
        "test.)\n\n"
        "B) NO - shifts move only between technicians; changing the day is done another way.\n\n"
        "C) Not in this release - please say roughly when.",
    ),
]

SEC5_INTRO = (
    "Ten questions added on 13 August - everything that has come up since the earlier sections "
    "were written, gathered here so you still get ONE sheet rather than a trickle. The first two "
    "are about a ticket one of your team raised on 7 August; five are about differences between "
    "your description and the design review of 5 August; the rest are single decisions. Nothing "
    "in this section is blocking a test today, but several decide what a test should say."
)

TR = "https://shopview.testrail.io/index.php?/cases/view/"

# QA-only mapping rows for the ten new items (8 columns, same as the imported tab):
# Tab, Item, What it asks, Where the question comes from, Affected cases, Links,
# Spec anchors + live evidence, What each answer resolves to.
QA_NEW = [
    ("Tab 5", "1-2",
     "Filters - the SV-9041 show/hide control rule (hidden when a page has one filter), and "
     "whether it reaches Parts and Reports",
     "build/filters/questions-2026-08-06/ADDENDUM-2026-08-11_SV-9041-collapse-toggle.md (2026-08-11), "
     "carried onto this consolidated sheet unchanged. Source ticket SV-9041 (Task, parent SV-8785, "
     "TESTING QA), condition stated at creation 2026-08-07T13:28:17Z.",
     "FLT-COLL-01 (C29601) and FLT-PR-PAR-01 (C43562) - neither is blocked; both were repaired "
     "2026-08-11 to follow SV-9041, the newer authoritative source (Rule 32).",
     " · ".join(f"C{i} {TR}{i}" for i in [29601, 43562]),
     "S1-R4 unchanged since spec v1 (2026-05-13), present byte-identically in all 19 versions; the "
     "condition ('more then 1 filter') absent from all 19. SV-9041 (7 Aug) is ~3 months newer, so "
     "the ticket prevails (Rule 32). Branko's own 31 July ruling (Round 3, Q5=A) is cited on C43562 "
     "and the divergence disclosed (Rules 33/56). Full evidence in the addendum's QA section.",
     "A -> he adds one sentence to the description; the two cases' provenance is re-stamped to name "
     "it. B -> description stays; nothing changes. C -> both cases are wrong on their headline "
     "assertion, both must be corrected, and VLAD MUST BE TOLD (C29601 is flagged Automated in "
     "TestRail - Rule 65)."),
    ("Tab 5", "3",
     "Schedule - does the capacity-bar pop-up list ONLY assigned technicians (the v26 one-word "
     "change) or all technicians?",
     "build/schedule/coverage-rederivation-2026-08-10/QUESTIONS-FOR-BRANKO.md item S-1; restated "
     "with the honesty note in build/schedule/coverage-gaps-2026-08-11/QUESTIONS-FOR-BRANKO.md Q3.",
     "SCH-CAP-04 (C30033) - already edited to the v26 wording WITHOUT waiting for this answer, and "
     "the sheet says so to his face.",
     f"C30033 {TR}30033",
     "Spec anchor §4.12; the word 'assigned' entered on 2026-08-07 (v26) with no change note. The "
     "wording had been stable since July.",
     "A -> C30033 is already correct; the open-question note comes off. B -> one update_case "
     "reverts it; no other test depends on it."),
    ("Tab 5", "4",
     "Schedule - panel persistence: does hiding the job list survive sign-out (spec says "
     "session-scoped; design review E12 says persist across sessions)?",
     "build/schedule/coverage-rederivation-2026-08-10/QUESTIONS-FOR-BRANKO.md item S-2; restated "
     "2026-08-11 in build/schedule/panel-collapse-2026-08-11/QUESTIONS-FOR-BRANKO.md when it became "
     "cited on a live case.",
     "SCH-PANEL-06 (C43587), expected item 2 - follows the specification (answer A) and states the "
     "open question in its own tester-facing text (Rule 58; the control is not built, so the build "
     "could not have settled it anyway). Marker AUTOMATION: READY, not HOLD.",
     f"C43587 {TR}43587",
     "Assertion at stake: §5.3-L195.A2 'Session-scoped per user for build' vs design-review item "
     "E12 'persist view options per user ... so it survives across sessions' (Fabian/Sasha review, "
     "2026-08-05). A different control from the view-settings question (Tab 5 item 9) - the two "
     "answers may legitimately differ.",
     "A -> the case is already correct; the open-question note is replaced with a Rule-54 "
     "confirmation (one update_case). B -> expected item 2 is reversed with a Rule-56 divergence "
     "sentence (one update_case)."),
    ("Tab 5", "5",
     "Filters - should a single-choice filter panel close on pick?",
     "build/handover-ingest-2026-08-10/QUESTIONS.md item B-1 (from the Filters handover document, "
     "a Rule-57(f) source).",
     "No case asserts either way - nothing is wrong today and nothing is held.",
     "-",
     "The specification gives one rule (closes on outside click) and is silent on close-on-pick; "
     "the build's panels differ from each other (multi-select stay open, Asset on Site closes, the "
     "date panel closes on preset only).",
     "A -> the description gains the rule and we may add a small assertion. B -> the differing "
     "panels become a defect to raise. C -> as described."),
    ("Tab 5", "6",
     "Schedule - the outside-hours warning wording: technician's own hours vs 'business hours' "
     "(the SV-8917 question)",
     "build/handover-ingest-2026-08-10/QUESTIONS.md item B-2. Deliberately NOT resolved from the "
     "build or the ticket: SV-8917 is Sasha Grosman's (Rule 38) and its fix, applied literally, "
     "would contradict spec §4.2 (technician's own hours take precedence over shop business hours).",
     "No case is held; our cases follow §4.2. Recorded at risk HIGH in "
     "build/schedule/coverage-rederivation-2026-08-10/DELIBERATE-DECISIONS.md entry 6 - SV-8917 is "
     "TESTING QA, so the change may already be in the build, and two of our cases would then fail "
     "against it.",
     "-",
     "Spec §4.2 vs §4.11 vs SV-8917's requested wording.",
     "A -> our cases stand; the ticket's wording should be adjusted. B -> the description changes "
     "and our two warning cases change with it. C/D -> as described."),
    ("Tab 5", "7",
     "Schedule - should an 'Add Existing Work Order' button exist in this release?",
     "build/handover-ingest-2026-08-10/QUESTIONS.md item B-3.",
     "No case exists - deliberately not authored: the phrase appears 0 times in all 27 versions of "
     "the specification; the only source is the design review's parked fault, whose own text says "
     "nobody is sure it was ever planned.",
     "-",
     "Searched all 27 spec versions; absent from every one.",
     "A -> we author the case and the description gains the requirement. B -> the parked fault "
     "closes. C -> stays out of this release's suite."),
    ("Tab 5", "8",
     "Schedule - plan the hours REMAINING or the WHOLE original estimate for a part-done job?",
     "build/handover-ingest-2026-08-10/QUESTIONS.md item B-4.",
     "Our test follows the specification (whole original estimate); the spec was updated 7 August, "
     "two days AFTER the review asked for the opposite, and still says the original estimate - so "
     "latest-wins currently favours the spec and the test was not changed.",
     "-",
     "Spec: planned hours / estimate / actual are three separate numbers not made to reconcile; "
     "design review E5 asks for remaining-hours planning, in scope for this release.",
     "A -> nothing changes. B -> the description changes and our test changes with it."),
    ("Tab 5", "9",
     "Schedule - are the view settings remembered per person?",
     "build/handover-ingest-2026-08-10/QUESTIONS.md item B-5. A DIFFERENT control from the "
     "panel-persistence question (Tab 5 item 4), and the sheet says so.",
     "No case exists - deliberately not authored against a meeting note alone (Rule 57).",
     "-",
     "The spec pins the defaults and is silent on persistence; design review E12 asks for per-user "
     "persistence with details still to be worked out.",
     "A -> we author the persistence case. B/C -> we author the matching reset/session case."),
    ("Tab 5", "10",
     "Schedule - drag a shift onto the NEXT DAY for the same technician?",
     "build/handover-ingest-2026-08-10/QUESTIONS.md item B-6.",
     "No case exists - deliberately not authored: the spec covers technician-to-technician moves "
     "only; day moves exist only in the review (E9), details to be worked out.",
     "-",
     "Spec silent on day moves for shifts (non-job blocks CAN move between days - a different "
     "block type).",
     "A -> we author the case. B -> nothing changes. C -> queued for the later release."),
]

EXTRA_NOTES_2026_08_13 = [
    "CONSOLIDATION 2026-08-13: this sheet supersedes, as the ONE file to send, the combination of "
    "the 2026-08-06 Friendly sheet + the 2026-08-11 SV-9041 addendum + the S-1/S-2 questions "
    "(coverage-rederivation 2026-08-10, restated in panel-collapse and coverage-gaps 2026-08-11) + "
    "the six handover-ingest questions B-1..B-6 (2026-08-10). All 22 earlier items are imported "
    "from the 2026-08-06 generator so the wording cannot drift; the ten new items are Section 5. "
    "None of the earlier files was edited or deleted.",
    "HELD, NOT SENT (Standing Rule 66, QA lead 2026-08-12 verbatim: 'This should be the last thing "
    "once you give me the report that everything else has been done only this part is left and "
    "save it as a rule for now and for the future projects too.'). The sheet goes out attached to "
    "the completion report, on the QA lead's word, once everything we can do ourselves on Filters "
    "and Schedule is finished.",
    "CORRECTION TO AN IMPORTED ROW, DATED 2026-08-13: the Section 3 item 7 (cross-filter) mapping "
    "row below still says Standing Rule 57's technical-design authority question is OPEN and held. "
    "IT WAS ANSWERED by the QA lead on 2026-08-12 (verbatim: 'Technical design is the authority "
    "but if that contradicts with specs/tickets/answer sheet/claude design/figma ... consider the "
    "specs/tickets/answer sheet/claude design/figma ... as the authority for the test cases but "
    "let me know where it contradicts with the tech design.'). The eleven held cases were released "
    "(build/rulings-2026-08-12/TECH-DESIGN-CONTRADICTIONS.md). The question is NOT on this sheet "
    "and must never be put to a PO - it was the QA lead's and it is closed. The imported row's "
    "superseded wording is kept visible per the standing dated-correction pattern.",
    "SOURCE CURRENCY FOR THIS CONSOLIDATION (Rule 31): the imported items were verified against "
    "Filters spec Confluence VERSION 19 and Schedule spec VERSION 25 on 2026-08-06; the S-1/S-2 "
    "items against Schedule VERSION 26/27 (2026-08-10/11 reads); the SV-9041 items against all 19 "
    "Filters versions (2026-08-11). THIS PASS (2026-08-13) MADE NO LIVE CONFLUENCE READ - it is a "
    "consolidation of held questions, and the sheet is HELD. OWED BEFORE SENDING (Rule 59): "
    "re-read both live spec versions and each question's anchor immediately before the send, and "
    "re-verify no question has been answered by a spec edit in the meantime.",
    "DELIBERATELY NOT ON THIS SHEET: (a) the technical-design authority question - answered "
    "2026-08-12, see above; (b) the 'Create Work Order' label - Branko answered it himself in "
    "SV-9076 on 2026-08-10; (c) SV-8906 (empty-state inconsistency, Task, Board Backlog, register "
    "row F14) - not analysed yet, so no row can honestly be written for it (a row would be a guess); "
    "(d) 'what New Work Order actually does' (register row A5, 2026-07-31) - nothing is blocked "
    "(C38855 passes either way) and the 2026-08-06 consolidation already left it off; kept in the "
    "register, not re-raised; (e) the Filters collapse-state-per-page gap behind SV-8905/C29603 - "
    "that is a QA-lead ruling (register row F-C29603: result vs case vs new-case), not a product "
    "question, and the ticket already exists.",
]


def _sheet(wb, name, title, note, band, items, widths, first=False, start_num=1):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 200 if first else 90
    r = 4
    ws.cell(row=r, column=1, value=band).font = Font(bold=True)
    r += 1
    cols = ["#", "Which project and screen", "What happens now", "The question", "Options",
            "Your answer"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
    r += 1
    for i, (topic, now, q, opts) in enumerate(items, start_num):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    return ws


def _qa_rows():
    """The imported mapping (with the friendly generator's own corrections), plus the dated
    2026-08-13 correction to the cross-filter row, plus the ten new rows."""
    out = []
    for row in friendly._qa_rows():
        row = list(row)
        for j, v in enumerate(row):
            if isinstance(v, str) and "Standing Rule 57's OPEN question" in v:
                row[j] = v + (
                    " *** CORRECTED 2026-08-13: that question is NO LONGER OPEN - the QA lead "
                    "answered it on 2026-08-12 (technical design sources a case alone where "
                    "nothing contradicts it; on a contradiction the specs/tickets/answer sheets/"
                    "Claude design/Figma win, latest-wins among them; every contradiction is "
                    "reported to him). The eleven held cases were released - see "
                    "build/rulings-2026-08-12/TECH-DESIGN-CONTRADICTIONS.md. The superseded "
                    "wording before this marker is kept visible, dated, per the standing "
                    "pattern. ***"
                )
        out.append(row)
    out.extend(list(r) for r in QA_NEW)
    return out


def write_xlsx():
    wb = openpyxl.Workbook()
    w = [4, 40, 58, 44, 48, 22]

    _sheet(wb, TAB1,
           "Filters and Schedule - questions for Branko Cicovic - 2026-08-13 (consolidated)",
           HELLO,
           "Section 1 of 5 - please start here. These five release tests that are stuck today.",
           friendly.SEC1, w, first=True)

    _sheet(wb, TAB2,
           "Section 2 - nine ordinary decisions",
           "Nine decisions, each a plain A or B. Nothing here is urgent this week, but each one "
           "settles a difference between two of your own documents or between a document and the "
           "screen. Short answers are perfect.",
           "Section 2 of 5 - nine decisions.",
           friendly.SEC2, w)

    _sheet(wb, TAB3,
           "Section 3 - seven things only the engineering plan describes",
           "NOTHING OF OURS IS WAITING ON THIS TAB, so it can keep for a quiet moment. These "
           "seven behaviours appear only in the engineering plan and nowhere in your own "
           "document. For the first six we are not going to turn an engineering note into "
           "something the product must do without your word. The seventh is different and says "
           "so: it is already covered by two of our tests and nothing is waiting on it - we are "
           "simply asking you to write the rule into your own description as well.",
           "Section 3 of 5 - seven engineering-plan-only items. No test is blocked by these.",
           friendly.SEC3, w)

    _sheet(wb, TAB4,
           "Section 4 - one heads-up, nothing to decide",
           "NO DECISION NEEDED. One typo-level slip in your own document, which has already cost "
           "one round of confusion. Noted for your next edit.",
           "Section 4 of 5 - one heads-up. Nothing to answer.",
           friendly.SEC4, w)

    _sheet(wb, TAB5,
           "Section 5 - ten questions added on 13 August",
           SEC5_INTRO,
           "Section 5 of 5 - ten questions added on 13 August. Two Filters, then mostly Schedule.",
           SEC5, w)

    ws = wb.create_sheet(TABQA)
    ws["A1"] = ("QA-ONLY - INTERNAL - NOT FOR BRANKO. Do not send this tab. TestRail case IDs, "
                "requirement anchors, live evidence and the corrections to our own records live "
                "here so the tabs Branko reads stay plain (Standing Rules 7, 8 and 55).")
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = WRAP
    r = 3
    ws.cell(row=r, column=1, value="PER-ITEM MAPPING - IMPORTED 2026-08-06 ROWS FIRST (INCLUDING "
                                   "THE ONE REMOVED ITEM), THEN THE TEN 13 AUGUST ROWS").font = \
        Font(bold=True)
    r += 2
    hdr = ["Tab", "Item", "What it asks", "Where the question comes from",
           "Affected internal case IDs (TestRail C-id)", "TestRail links",
           "Spec anchors + live evidence", "What each answer resolves to"]
    for j, c in enumerate(hdr, 1):
        cell = ws.cell(row=r, column=j, value=c)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
    r += 1
    for row in _qa_rows():
        for j, v in enumerate(row, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in EXTRA_NOTES_2026_08_13 + list(friendly.EXTRA_NOTES) + list(friendly.prior.QA_NOTES):
        ws.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, wd in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 62, 62]):
        ws.column_dimensions[col].width = wd

    wb.save(XLSX)
    return XLSX


def block(items, start=1):
    out = []
    for i, (topic, now, q, opts) in enumerate(items, start):
        out.append(f"### {i}. {topic}\n")
        out.append("**What happens now**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
        out.append("**The question**\n")
        out.append(f"> {q}\n")
        out.append("**Options**\n")
        out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
        out.append("**Your answer:** _______________________________________________\n")
    return "\n".join(out)


def write_md():
    md = f"""# Questions for Branko Cicovic — Filters and Schedule — 2026-08-13 (consolidated)

**Projects: Filters (epic SV-8785) and Schedule (epic SV-8685) · Product Owner: Branko Cicovic**

*This is the CONSOLIDATED sheet: the 22 items of the held 2026-08-06 Friendly sheet, imported from
its own generator so the wording cannot drift, plus the TEN questions that accumulated after it was
written (the SV-9041 addendum of 2026-08-11, the two Schedule questions of 2026-08-10/11, and the
six handover-ingest questions of 2026-08-10) — 32 items in total, one sheet, per Standing Rule 55.
The spreadsheet twin is `Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-13.xlsx`; it
carries a QA-only tab that must not be forwarded.*

**DRAFT — WRITTEN AND HELD, NOT SENT (Standing Rule 66: the question sheet is the LAST thing sent,
on the QA lead's word, once everything we can do ourselves is finished). Nothing has been written
to TestRail or Jira.**

---

{HELLO}

---

## Section 1 — Start here: these five release tests that are stuck today

{block(friendly.SEC1)}

---

## Section 2 — Nine ordinary decisions, when you have a moment

{block(friendly.SEC2)}

---

## Section 3 — Seven things only the engineering plan describes

**Nothing of ours is waiting on this section**, so it can keep for a quiet moment.

{block(friendly.SEC3)}

---

## Section 4 — One heads-up, nothing to decide

{block(friendly.SEC4)}

---

## Section 5 — Ten questions added on 13 August

{SEC5_INTRO}

{block(SEC5)}

---

## QA-only — not for Branko

The question-to-case mapping is on the spreadsheet's `{TABQA}` tab — the imported 2026-08-06 rows
(including the one removed item, kept re-labelled), the ten new 13 August rows, the dated
2026-08-13 correction to the cross-filter row (the technical-design authority question was ANSWERED
by the QA lead on 2026-08-12 and is NOT on this sheet), the source-currency record, the pre-send
re-read duty (Rule 59), and the list of what was deliberately NOT asked and why.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
