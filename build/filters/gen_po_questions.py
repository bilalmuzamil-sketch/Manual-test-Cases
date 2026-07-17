#!/usr/bin/env python3
"""Generate the Filters PO question sheet for Branko (Filters PO).

Outputs (regenerable — run from the repo root):
  - build/filters/PO-Questions-Filters_2026-07-17.xlsx
  - build/filters/PO-Questions-Filters_2026-07-17.md

Sheets / sections (mirrors build/simple-flow/gen_po_questions_round3.py):
  - "Questions for PO"    : reader-facing, VERY SIMPLE layman language ONLY.
                            NO case IDs, FLT- codes, API/HTTP terms, or jargon.
                            Columns: # | Topic | What happens now | The question |
                            Options | Your answer (blank).
  - "QA Internal Mapping" : QA-only. Per question, the affected FLT- case IDs
                            (TestRail C-ids blank until a permitted push;
                            testrail-id-map.csv per Standing Rule 8), spec/design
                            refs, and what each answer option resolves to.

Only genuine PRODUCT DECISIONS are included (Standing Rule 7) — no bugs/defects
(there are none yet: the feature is pre-VIU), no env questions (those are QA-side
OQ-2/3/6/7 asked of the user, not the PO).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/filters/PO-Questions-Filters_2026-07-17.xlsx"
MD_OUT = "build/filters/PO-Questions-Filters_2026-07-17.md"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no codes, no tech terms)
# ---------------------------------------------------------------------------
questions = [
    {
        "topic": "Filters on the Parts and Reports pages",
        "now": ("The final design pictures show the same new filter buttons not "
                "only on the Work Orders page, but also on nine Parts pages and "
                "on all the Reports pages. However, the written description we "
                "received only talks about the Work Orders page - it says "
                "nothing about Parts or Reports."),
        "question": ("Are the filters on the Parts pages and the Reports pages "
                     "part of this release, and should we test them now? If yes, "
                     "is there a written description for them like the one we "
                     "have for Work Orders?"),
        "opts": ("A) Yes - they are part of this release; a write-up exists or "
                 "will be provided, and they should be tested now.\n"
                 "B) No - only the Work Orders page is in this release; Parts "
                 "and Reports come later."),
    },
    {
        "topic": "How long the app remembers your filters",
        "now": ("The write-up says two slightly different things. In one place "
                "it says your chosen filters are kept only until you close the "
                "browser. In another place it says they are saved for you and "
                "come back whenever you return to the page - which sounds like "
                "they would still be there even after closing the browser or "
                "logging out."),
        "question": ("When someone picks filters on the Work Orders page, how "
                     "long should the app remember them?"),
        "opts": ("A) Only until they close the browser - after that the page "
                 "starts fresh.\n"
                 "B) Remembered for that person permanently - the filters are "
                 "still there the next day, even after closing the browser or "
                 "logging out."),
    },
    {
        "topic": "Spelling of \"Lead Technician\"",
        "now": ("In several of the design pictures the word is misspelled as "
                "\"Lead Tehnician\" (the letter c is missing) - for example in "
                "the list column heading and in the mobile filter list. In other "
                "places in the same designs it is spelled correctly."),
        "question": ("Can you confirm the app will ship with the correct "
                     "spelling \"Lead Technician\" everywhere, and that the "
                     "misspelling in the designs will be fixed?"),
        "opts": ("A) Yes - it must read \"Lead Technician\" everywhere; the "
                 "design will be corrected.\n"
                 "B) Something else (please explain)."),
    },
    {
        "topic": "The Status filter on the Estimates and Completed tabs",
        "now": ("The write-up says that on the Estimates and Completed tabs the "
                "Status filter button should be hidden completely (those tabs "
                "already show only one status). But the design picture of the "
                "Estimates tab shows the Status button still there - greyed "
                "out, pre-filled with \"Status: Estimate\", and not clickable."),
        "question": ("On the Estimates and Completed tabs, what should the "
                     "Status filter button do?"),
        "opts": ("A) Hidden completely - the button is not there at all (as the "
                 "write-up says).\n"
                 "B) Shown but greyed out, pre-filled with the tab's status, and "
                 "not clickable (as the design picture shows)."),
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (Standing Rule 8: internal id + TestRail C-id; C-ids are
# BLANK until a permitted TestRail push — see build/filters/testrail-id-map.csv).
# Each entry: (q_no, [internal ids], refs, resolves_to)
# ---------------------------------------------------------------------------
internal_map = [
    (1,
     ["(no cases authored - deliberate)"],
     "design-notes.md §B.5 (9 Parts screens) + §B.6 (22 Reports screens) are in "
     "the final ZIP design set but NO spec story covers them (requirements.md "
     "Stories 1-12 are all Work Orders page). Scope ruling recorded in "
     "coverage-matrix.md §C.",
     "A -> request the Parts/Reports spec from Branko, then author dedicated "
     "case sections (est. +30-50 cases) as a scope extension. "
     "B -> keep the 79-case WO-only suite; Parts/Reports screens stay listed as "
     "excluded-with-reason in coverage-matrix.md §C."),
    (2,
     ["FLT-PERS-02", "FLT-PERS-01", "FLT-PERS-03"],
     "requirements.md OQ-5: S10-R2 ('for the duration of the browser session') "
     "vs §2/§4 ('saved per user and reloaded when they return'). FLT-PERS-02 is "
     "authored to the common ground (same-session persistence) with the tension "
     "flagged in its notes.",
     "A -> FLT-PERS-02 stays as-is; add an explicit negative expectation that a "
     "browser restart clears the filters. "
     "B -> extend FLT-PERS-02 (or add a case) to verify filters survive browser "
     "close + re-login, per user."),
    (3,
     ["FLT-BAR-02", "FLT-TECH-01", "FLT-MOB-02", "FLT-MOB-06"],
     "design-notes.md §C.1: 'Lead Tehnician' recurs in the final ZIP set (WO "
     "table column header + mobile sheet rows); the filter chip itself is "
     "spelled correctly. All cases are authored with the CORRECT 'Lead "
     "Technician' and carry typo-flag notes.",
     "A -> no case change; at VIU, if the build shows 'Tehnician' anywhere, "
     "file a bug (do NOT rewrite cases to the typo). "
     "B -> per Branko's explanation."),
    (4,
     ["FLT-TAB-02", "FLT-TAB-03", "FLT-BAR-03"],
     "Spec S2-N1/S2-N2/S9-R2/S9-R3 say the Status chip is HIDDEN on "
     "Estimates/Completed; final design frame 11972:32318 shows a pale/disabled "
     "'Status: Estimate' chip (design-notes §C.7). Cases authored to the spec's "
     "intent (no USABLE Status filter) so they hold under either answer.",
     "A -> tighten FLT-TAB-02/03 expected to 'chip not present at all'. "
     "B -> rewrite FLT-TAB-02/03 expected to 'chip shown disabled, pre-filled "
     "with the tab status, not clickable'."),
]


def build_xlsx():
    wb = Workbook()

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = "Filters (Work Orders page) - Questions for Branko - 2026-07-17"
    ws["A1"].font = TITLE_FONT
    ws.append([])
    hdr = ["#", "Topic", "What happens now", "The question", "Options",
           "Your answer"]
    ws.append(hdr)
    for c in ws[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = BORDER
    for i, q in enumerate(questions, 1):
        ws.append([i, q["topic"], q["now"], q["question"], q["opts"], ""])
    for row in ws.iter_rows(min_row=4, max_row=3 + len(questions)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    widths = [4, 22, 46, 42, 46, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids are "
                 "BLANK until a permitted push (build/filters/testrail-id-map.csv).")
    ws2["A1"].font = Font(bold=True, color="C00000")
    ws2.append([])
    hdr2 = ["Q#", "Affected internal case IDs (TestRail C-id: pending push)",
            "Spec / design refs", "What each answer resolves to"]
    ws2.append(hdr2)
    for c in ws2[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = BORDER
    for qno, ids, refs, resolves in internal_map:
        ws2.append([qno, "\n".join(ids), refs, resolves])
    for row in ws2.iter_rows(min_row=4, max_row=3 + len(internal_map)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    for i, w in enumerate([4, 34, 52, 60], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A4"

    wb.save(XLSX_OUT)
    print("Wrote", XLSX_OUT)


def build_md():
    lines = [
        "# Filters (Work Orders page) — Questions for Branko — 2026-07-17",
        "",
        "Plain-language product questions only (no bugs, no test jargon).",
        "Please pick an option (or write your own answer) for each.",
        "",
    ]
    for i, q in enumerate(questions, 1):
        lines += [
            f"## Question {i} — {q['topic']}",
            "",
            f"**What happens now:** {q['now']}",
            "",
            f"**The question:** {q['question']}",
            "",
            "**Options:**",
            "",
        ]
        lines += [f"- {o}" for o in q["opts"].split("\n")]
        lines += ["", "**Your answer:** ____________________", ""]
    lines += [
        "---",
        "",
        "## QA Internal Mapping (QA-only — not for the PO)",
        "",
        "TestRail C-ids are blank until a permitted push "
        "(`build/filters/testrail-id-map.csv`, Standing Rule 8).",
        "",
        "| Q# | Affected internal case IDs | Spec / design refs | Resolves to |",
        "|---|---|---|---|",
    ]
    for qno, ids, refs, resolves in internal_map:
        lines.append(
            f"| {qno} | {', '.join(ids)} | {refs} | {resolves} |")
    lines.append("")
    open(MD_OUT, "w").write("\n".join(lines))
    print("Wrote", MD_OUT)


if __name__ == "__main__":
    build_xlsx()
    build_md()
