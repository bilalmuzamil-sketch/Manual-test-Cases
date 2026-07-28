#!/usr/bin/env python3
"""Generate the Filters (Parts + Reports + page search) PO question sheet for
Branko (Filters PO).

Mirrors build/filters/gen_po_questions.py 1:1 (same two sheets, same layman
rules, Standing Rules 7 & 8):
  - "Questions for PO"    : reader-facing, VERY SIMPLE layman language ONLY.
                            NO case IDs, FLT- codes, Figma nodes, API/HTTP terms,
                            or jargon. Columns: # | Topic | What happens now |
                            The question | Options | Your answer (blank).
  - "QA Internal Mapping" : QA-only. Per question, the affected FLT- case IDs
                            (TestRail C-ids blank until a permitted push;
                            testrail-id-map.csv), design refs, resolves-to.

Outputs (human-readable filename, Standing Rule 19):
  - build/filters/PO-Questions-Branko-PartsReports-2026-07-27.xlsx
  - build/filters/PO-Questions-Branko-PartsReports-2026-07-27.md
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/filters/PO-Questions-Branko-PartsReports-2026-07-27.xlsx"
MD_OUT = "build/filters/PO-Questions-Branko-PartsReports-2026-07-27.md"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no codes, no tech terms)
# ---------------------------------------------------------------------------
questions = [
    {
        "topic": "A written description for the Parts and Reports filters",
        "now": ("We now have the design pictures for the new filter buttons on the "
                "Parts pages and on the Reports pages, so we know which buttons "
                "appear on each page. What we do NOT have is a written description "
                "that says how they should behave - what each button does, what "
                "choices are inside it, and what the page should look like after "
                "you pick something."),
        "question": ("Can you share a written description (the same kind we have "
                     "for the Work Orders page) for the Parts filters and the "
                     "Reports filters, so we can test them properly?"),
        "opts": ("A) Yes - a write-up exists or will be provided.\n"
                 "B) No write-up yet - test only what the pictures show for now, "
                 "and treat all behaviour as 'to be confirmed later'."),
    },
    {
        "topic": "Which filter buttons actually filter each page",
        "now": ("The pictures show a row of filter buttons at the top of each "
                "Parts page and each Reports page (for example on Inventory: Bin "
                "Location, Category, Supply, Vendor). The pictures do not tell us "
                "whether every button really narrows the list, or whether some "
                "are just shown for now."),
        "question": ("On each Parts page and each Reports page, should every "
                     "filter button shown in the design actually filter the list "
                     "or the report when used?"),
        "opts": ("A) Yes - every button shown filters that page.\n"
                 "B) Some are not active yet (please tell us which ones)."),
    },
    {
        "topic": "The full list of choices inside each filter",
        "now": ("When you click a filter button it should open a list of choices "
                "(for example the Status button, or the Vendor button). The "
                "pictures only show the buttons, not the full list of choices "
                "behind each one."),
        "question": ("Can you give us the full list of choices for each filter on "
                     "the Parts and Reports pages (for example all the statuses, "
                     "all the vendor options, the date options)?"),
        "opts": ("A) Yes - here is the list / it will be provided.\n"
                 "B) The choices come from the shop's own data (for example the "
                 "list of real vendors), so there is no fixed list."),
    },
    {
        "topic": "How the new kinds of filter work",
        "now": ("The Parts and Reports pages show some filter buttons we do not "
                "have on the Work Orders page - for example Location, Transaction "
                "Type, Invoice Status, Type, User, Mention, and a Core / Non Core "
                "filter. We do not know how each of these is meant to work."),
        "question": ("For each of these new filter buttons, what choices should it "
                     "offer and how should it narrow the page - can you pick more "
                     "than one choice, and does the page update right away?"),
        "opts": ("A) Yes - here is how each one works / it will be described in "
                 "the write-up.\n"
                 "B) Something else (please explain)."),
    },
    {
        "topic": "Do the Parts and Reports filters work the same way as Work Orders",
        "now": ("On the Work Orders page the filters let you pick more than one "
                "choice, clear a single filter or clear them all, collapse the "
                "filter bar, remember your choices when you come back, share a "
                "link that keeps your filters, and work on a phone. We do not "
                "know if the Parts and Reports filters behave the same way."),
        "question": ("Should the Parts and Reports filters behave exactly like the "
                     "Work Orders filters for these things (multiple choices, "
                     "clearing, collapsing, remembering, shareable link, phone)?"),
        "opts": ("A) Yes - they should behave the same as the Work Orders "
                 "filters.\n"
                 "B) No - there are differences (please tell us which)."),
    },
    {
        "topic": "The pop-up search box (\"Search or ask a question\")",
        "now": ("The designs also include a pop-up search box that opens from the "
                "top bar (or with a keyboard shortcut) and searches across work "
                "orders, customers, assets, parts, vendors and part sales. It "
                "shows the words \"Search or ask a question\". This same search box "
                "is also part of a separate piece of work called Global Search."),
        "question": ("Is this pop-up search box part of THIS filters release (so we "
                     "test it here), or is it owned by the separate Global Search "
                     "work? And does the \"ask a question\" part (an AI answer) go "
                     "live now, or later?"),
        "opts": ("A) Test it as part of Global Search (not here) - and 'ask a "
                 "question' is for later.\n"
                 "B) It is part of this filters release - please confirm what "
                 "'ask a question' should do now.\n"
                 "C) Something else (please explain)."),
    },
    {
        "topic": "Do the filter choices depend on the person's role",
        "now": ("For the Work Orders page you already confirmed the filter lists "
                "are the same for everyone (they do not change by a person's "
                "role). We do not know if the same is true on the Parts and "
                "Reports pages."),
        "question": ("On the Parts and Reports pages, should the filter buttons and "
                     "their choices be the same for every user, or should some be "
                     "hidden or limited depending on the person's role?"),
        "opts": ("A) Same for everyone - the person's role does not change the "
                 "filters.\n"
                 "B) Some filters or choices depend on the role (please tell us "
                 "which)."),
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (Standing Rule 8: internal id + TestRail C-id; C-ids are
# BLANK until a permitted TestRail push — see build/filters/testrail-id-map.csv).
# ---------------------------------------------------------------------------
internal_map = [
    (1,
     ["FLT-PARTS-01..12", "FLT-RPTS-01..22"],
     "design-notes.md §B.5 (9 Parts screens) + §B.6 (23 Reports screens) give "
     "the chips + columns, but NO behaviour spec exists (requirements.md Stories "
     "1-12 are Work Orders only). All 34 Parts/Reports cases authored design-only, "
     "viu_status VIU-Pending.",
     "A -> ingest the PRD, run SPEC-RELEVANCE-RECONCILIATION + build-accurate "
     "wording, then VIU. B -> keep the design-level cases as-is; behaviour stays "
     "flagged 'to be confirmed' until the write-up lands."),
    (2,
     ["FLT-PARTS-11", "FLT-RPTS-21"],
     "Behaviour cases assert the list/report narrows when a filter is chosen; the "
     "design does not pin which chips actually apply. Flagged pending PRD.",
     "A -> tighten the behaviour expected per page. B -> mark the inactive chips "
     "and adjust the affected per-page cases."),
    (3,
     ["FLT-PARTS-09", "FLT-PARTS-11", "FLT-PARTS-12", "FLT-RPTS-21", "FLT-RPTS-22"],
     "Option lists behind each chip are not in the design. Part Type is the only "
     "pinned list (Core / Non Core / Clear selection, §B.5 #9).",
     "A -> add option-list checks per filter. B -> data-driven lists: verify "
     "against seeded shop data at VIU, no fixed expected list."),
    (4,
     ["FLT-PARTS-09", "FLT-RPTS-12", "FLT-RPTS-13", "FLT-RPTS-15", "FLT-RPTS-16",
      "FLT-RPTS-17", "FLT-RPTS-20", "FLT-RPTS-22"],
     "New filter types vs the WO page: Location, Transaction Type (A/R + A/P Aging "
     "Detail/Collection/Unpaid), Invoice Status (Sales Tax), Type + User (QB "
     "Unexported), Mention (Notes), Core/Non Core (Returns).",
     "A -> author per-type option + apply checks once described. B -> per Branko's "
     "explanation."),
    (5,
     ["FLT-PARTS-12", "FLT-RPTS-21", "and parity with FLT-* WO cases C29557-C29635"],
     "Parity of multi-select / Clear filters / collapse / persistence / URL / "
     "mobile with the Work Orders filters is assumed but not pinned by the Parts/"
     "Reports designs.",
     "A -> reuse the WO behaviour cases per Parts/Reports page. B -> author "
     "difference-specific cases for the exceptions."),
    (6,
     ["FLT-SRCH-01..09"],
     "Page-search / Command-K component (design-2026-07-27 screenshots, Figma "
     "11829-8908). OVERLAPS the Global Search project (86 cases already authored "
     "there). FLT-SRCH-09 is a scope-decision case. OQ-3 (AI 'ask a question' "
     "scope) still open.",
     "A -> retire/keep FLT-SRCH-01..09 in favour of Global Search's suite (avoid "
     "duplicate testing). B -> keep here and de-scope from Global Search; confirm "
     "AI behaviour."),
    (7,
     ["permissions_required flag on all FLT-PARTS-* / FLT-RPTS-*"],
     "OQ-4 was resolved for the Work Orders page (filter lists role-independent). "
     "Not confirmed for Parts/Reports; every new case carries a role-difference "
     "'to confirm' flag in permissions_required.",
     "A -> no per-role cases needed. B -> add role-scoped cases per affected "
     "filter."),
]


def build_xlsx():
    wb = Workbook()

    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = ("Filters (Parts, Reports & page search) - Questions for Branko - "
                "2026-07-27")
    ws["A1"].font = TITLE_FONT
    ws.append([])
    hdr = ["#", "Topic", "What happens now", "The question", "Options",
           "Your answer"]
    ws.append(hdr)
    for c in ws[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = BORDER
    for i, q in enumerate(questions, 1):
        ws.append([i, q["topic"], q["now"], q["question"], q["opts"], ""])
    for row in ws.iter_rows(min_row=4, max_row=3 + len(questions)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    widths = [4, 24, 48, 42, 46, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A4"

    ws2 = wb.create_sheet("QA Internal Mapping")
    ws2["A1"] = ("QA-ONLY - do not send this sheet to the PO. TestRail C-ids are "
                 "BLANK until a permitted push (build/filters/testrail-id-map.csv).")
    ws2["A1"].font = Font(bold=True, color="C00000")
    ws2.append([])
    hdr2 = ["Q#", "Affected internal case IDs (TestRail C-id: pending push)",
            "Design refs", "What each answer resolves to"]
    ws2.append(hdr2)
    for c in ws2[3]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = BORDER
    for qno, ids, refs, resolves in internal_map:
        ws2.append([qno, "\n".join(ids), refs, resolves])
    for row in ws2.iter_rows(min_row=4, max_row=3 + len(internal_map)):
        for c in row:
            c.alignment = WRAP
            c.border = BORDER
    for i, w in enumerate([4, 34, 52, 60], 1):
        ws2.column_dimensions[chr(64 + i)].width = w
    ws2.freeze_panes = "A4"

    wb.save(XLSX_OUT)
    print("Wrote", XLSX_OUT)


def build_md():
    lines = [
        "# Filters (Parts, Reports & page search) — Questions for Branko — 2026-07-27",
        "",
        "Plain-language product questions only (no bugs, no test jargon).",
        "Please pick an option (or write your own answer) for each.",
        "",
    ]
    for i, q in enumerate(questions, 1):
        lines += [
            f"## Question {i} — {q['topic']}",
            "",
            f"**What happens now:** {q['now']}",
            "",
            f"**The question:** {q['question']}",
            "",
            "**Options:**",
            "",
        ]
        lines += [f"- {o}" for o in q["opts"].split("\n")]
        lines += ["", "**Your answer:** ____________________", ""]
    lines += [
        "---",
        "",
        "## QA Internal Mapping (QA-only — not for the PO)",
        "",
        "TestRail C-ids are blank until a permitted push "
        "(`build/filters/testrail-id-map.csv`, Standing Rule 8).",
        "",
        "| Q# | Affected internal case IDs | Design refs | Resolves to |",
        "|---|---|---|---|",
    ]
    for qno, ids, refs, resolves in internal_map:
        lines.append(f"| {qno} | {', '.join(ids)} | {refs} | {resolves} |")
    lines.append("")
    open(MD_OUT, "w").write("\n".join(lines))
    print("Wrote", MD_OUT)


if __name__ == "__main__":
    build_xlsx()
    build_md()
