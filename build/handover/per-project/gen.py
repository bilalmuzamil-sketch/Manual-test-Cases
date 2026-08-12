#!/usr/bin/env python3
"""Generate the THREE per-project manual-tester handover sheets (.xlsx + .md each).

MIRRORS 1:1 (Standing Rule 16) the model this replaces:
  build/handover/Manual-Tester-Handover_2026-08-12.xlsx  (+ its .md twin)
Same four tabs, same row layout (A1 title / A2 note / A4 header / A5 band / items
from row 6), the same fills and fonts, the same wrap and freeze pane, the same
one-band-per-group shape. What changes is that there are now three sheets, one per
project, so each tester gets only what concerns them.

THE READER IS A MANUAL QA TESTER SEEING THIS FOR THE FIRST TIME (the QA lead,
verbatim: "make this manual tester handover sheet simple in words and understandable
as the manual tester will be reading it for the first time and it will be and can be
confusing for them"). So: short sentences, no jargon, no spec anchors, no endpoint
names, no HTTP codes, no internal ids beyond the TestRail C-id and its link (Rule 8).

This script writes files into this folder only. It makes NO TestRail call, NO Jira
call and NO application call. Its input is data.json, derived live earlier in the
same pass.
"""
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, 'data.json')))

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WARN_FILL = PatternFill("solid", start_color="FCE4E4", end_color="FCE4E4")
WRAP = Alignment(wrap_text=True, vertical="top")

TR = "https://shopview.testrail.io/index.php?/cases/view/"
JIRA = "https://shopview.atlassian.net/browse/"

TAB1 = "Read me first"
TAB2 = "Problems found, not reported"
TAB3 = "Old tickets that mislead"
TAB4 = "Tests that cannot be run yet"

FILES = {
    'Schedule': 'Schedule_Tester-Handover_2026-08-12',
    'Filters': 'Filters_Tester-Handover_2026-08-12',
    'Report Suite': 'Report-Suite_Tester-Handover_2026-08-12',
}

GROUP_ORDER = [
    'A problem was found but no ticket exists for it yet',
    'The feature or control is not in the build yet',
    'Waiting on a second sign-in as a different user',
    'Waiting on an answer from the product owner',
    "Waiting on the QA lead's ruling",
    'Depends on a nightly/background job the product never shows you',
    'The set-up this test needs cannot be produced on this environment',
]


def link(cid):
    return f"{TR}{cid}"


def n_words(n):
    return {0: 'no', 1: 'one', 2: 'two', 3: 'three', 4: 'four', 5: 'five', 6: 'six',
            7: 'seven', 8: 'eight', 9: 'nine', 10: 'ten'}.get(n, str(n))


# ----------------------------------------------------------------- tab 1 content
def tab1_rows(D):
    p = D['project']
    b = D['build']
    nh, npass = D['n_hold'], D['n_passed_on_hold']
    runnable = D['ready'] + D['expect_fail']
    nd, nt = len(D['defects']), len(D['tickets'])

    build_note = (
        f"The {p} build was rebuilt at {b['now_modified'].split(', ')[1]}, so it is NEWER than the "
        f"one most of this was checked against. What a test EXPECTS does not change when the build "
        f"changes - that comes from the written product description. What CAN change is the exact "
        f"wording of a button, or where something sits on screen."
        if b['moved'] else
        "The build has not changed since this was checked, so what is written here should match what "
        "you see."
    )

    rows = [
        ("1", "The one rule that matters most",
         "If a test says it cannot be run yet, mark it Blocked. Do not mark it Passed.\n\n"
         "Every test that cannot be run says so in its own words, at the very bottom of its "
         "Expected Results. If you open a test and it tells you it is waiting on something, that is "
         "the test telling you it cannot be judged today. Mark it Blocked and move on.",
         "This is not a formality. It has already happened on this project. Read from TestRail as "
         f"this sheet was written: {n_words(npass)} {p} test{'' if npass == 1 else 's'} that cannot "
         f"currently be judged already {'has' if npass == 1 else 'have'} a Passed result recorded "
         "against them. A test nobody could run cannot have passed. They are listed on the last tab "
         "and shaded in pink, and they need changing to Blocked before anyone reads them as evidence "
         "that the feature works.\n\n"
         "Across all three projects together the figure is 15."),

        ("2", "What is on each tab",
         f'Tab 2, "Problems found, not reported". {n_words(nd).capitalize()} real '
         f"fault{'' if nd == 1 else 's'} we found and confirmed on this project. No ticket has been "
         f"raised for {'it' if nd == 1 else 'any of them'}. If you want to raise the ticket "
         "yourself, everything you need to paste in is on the row, and there is a blank column for "
         "you to write the ticket number in.\n\n"
         f'Tab 3, "Old tickets that mislead". {n_words(nt).capitalize()} '
         f"case{'' if nt == 1 else 's'} where what a ticket says and what the product actually does "
         "no longer match.\n\n"
         f'Tab 4, "Tests that cannot be run yet". All {nh} of them, with one plain sentence each, '
         "grouped by what they are waiting on.",
         "Tabs 2 and 3 are suggestions for a person to action. Nothing on them has been done for "
         "you, and nothing has been raised in Jira - we were asked to hold off on that. That is why "
         "you are getting this sheet: so you can decide for yourself whether each one is worth "
         "raising."),

        ("3", "How to mark a test, in four lines",
         "The test tells you it is waiting on something -> Blocked.\n\n"
         "The test tells you what you will see today, and that is what you see -> Failed, and raise "
         "nothing new. It is already known.\n\n"
         "The test tells you what you will see today, and you see something different -> that is a "
         "NEW problem. Please report it.\n\n"
         "The test says nothing special and it works -> Passed.",
         "Anything else, or the test simply does not make sense to you: mark it Blocked and tell the "
         "QA lead. Never guess a result."),

        ("4", f"How many {p} tests there are",
         f"{runnable} to run.\n\n"
         f"{nh} to skip - they are all on tab 4.\n\n"
         f"{D['ours']} in total.",
         (f"A few {p} tests were written by a colleague and are not counted above and not listed "
          f"here. They are not ours to change and not yours to compare against this sheet: "
          f"{D['foreign']} of them."
          if D['foreign'] else
          "Every test in this project is ours, so there is nothing here written by someone else "
          "that you should ignore.")),

        ("5", "When these numbers were taken",
         f"Every count and every list here was read from TestRail at {D['read_at']}.\n\n"
         "PLEASE NOTE THE TIME. Other people are working on these same tests today, and the numbers "
         "move. Treat this as a photograph taken at that minute, not a permanent count.\n\n"
         "If a test on the last tab no longer says it is waiting on something when you open it, "
         "believe the test, not this sheet.",
         "As an example of how fast this moves: an earlier count this morning had 28 Schedule tests "
         "on the skip list. By the time this sheet was written it was 35, because more tests had "
         "been checked and found to be waiting on something. Neither figure was wrong - they were "
         "different minutes."),

        ("6", "One thing to know about the build",
         f"The {p} test branch is at {b['now']} right now.\n\n"
         f"Most of what is on tabs 2 and 3 was seen on {b['observed_on']}.\n\n" + build_note,
         ("BECAUSE THE BUILD HAS MOVED SINCE, please spend two minutes confirming a problem on tab 2 "
          "still happens before you raise a ticket for it. A ticket for something that was fixed "
          "this morning is worse than no ticket."
          if b['moved'] else
          "So if a button is named slightly differently from what a test says, that is worth telling "
          "the QA lead, but it is not automatically a fault in the product.")),
    ]
    return rows


def tab1_title(D):
    return f"{D['project']} - handover for the manual tester, 12 August 2026"


def tab1_note(D):
    return (
        f"Release is tomorrow. This sheet covers {D['project']} only, so everything in it is yours "
        "to act on. It has three lists in it, one per tab, and this page explains what they are and "
        "the one rule that matters most.\n\n"
        "You do not need to know anything about how the tests were written to use it. Row 5 says "
        "exactly when these numbers were read, and why that matters today."
    )


# ----------------------------------------------------------------- tab 2 content
def tab2_note(D):
    nd = len(D['defects'])
    b = D['build']
    extra = ("\n\nTHE BUILD HAS MOVED SINCE THESE WERE SEEN. They were seen on "
             f"{b['observed_on']}; the branch is now on {b['now']}. Please confirm the problem still "
             "happens before you raise a ticket for it - it takes two minutes."
             if b['moved'] else "")
    return (
        f"{n_words(nd).capitalize()} problem{'' if nd == 1 else 's'} on {D['project']}. "
        + ("It was seen with our own eyes on the build named on its row, and no Jira ticket covers "
           "it." if nd == 1 else
           "Every one was seen with our own eyes on the build named on its row, and no Jira ticket "
           "covers any of them.")
        + extra +
        "\n\nWHAT TO DO WITH THIS TAB: nothing, unless you want to. We were asked to hold off "
        "raising tickets, so these are written out in full in case you would rather raise them "
        "yourself. Everything you need is on the row. Use the last column to write down the ticket "
        "number if you do raise one.\n\n"
        "BEFORE YOU RAISE ONE, PLEASE READ THE LAST TWO COLUMNS. One says what is still missing; the "
        "other says the strongest argument AGAINST it being a fault, so you can answer that first. A "
        "ticket that a developer can argue with costs more than no ticket at all."
    )


T2_COLS = ["#", "What you will see", "What should happen instead",
           "Where that comes from (quoted)", "Exactly how to see it", "Tests affected",
           "Where we saw it", "What is still needed before filing",
           "The strongest argument AGAINST it being a fault",
           "Filed? / Ticket number"]


# ----------------------------------------------------------------- tab 3 content
def tab3_note(D):
    nt = len(D['tickets'])
    return (
        f"{n_words(nt).capitalize()} entr{'y' if nt == 1 else 'ies'}. "
        + ("It is a place where a ticket and the product no longer agree, so anyone reading that "
           "ticket today would get the wrong idea.\n\n" if nt == 1 else
           "Each one is a place where a ticket and the product no longer agree, so anyone reading "
           "the ticket today would get the wrong idea.\n\n")
        + "WHAT TO DO WITH THIS TAB: nothing has been done. We have not reopened, closed or commented "
        "on any of these, and we made no Jira calls at all while writing this sheet. The last column "
        "is what we would suggest, and it is only a suggestion - a person has to decide.\n\n"
        "The tests themselves are already honest about these, so you can run them normally. They "
        "tell you what you will see and what to do about it."
    )


T3_COLS = ["#", "Ticket", "What the ticket says", "What actually happens now",
           "Tests affected", "What we suggest (a suggestion only)"]


# ----------------------------------------------------------------- tab 4 content
def tab4_note(D):
    nh, npass = D['n_hold'], D['n_passed_on_hold']
    groups = {}
    for r in D['holds']:
        groups[r['group']] = groups.get(r['group'], 0) + 1
    biggest = max(groups.items(), key=lambda kv: kv[1])
    mis = len(D['mismarked'])
    extra = (
        f"\n\n(3) At the very bottom there is a short extra group of {mis} tests that are NOT marked "
        "to skip but which our own notes say cannot be run. They were meant to be marked and the "
        "change never went through. Please treat them the same way - Blocked, not Passed - and tell "
        "the QA lead."
        if mis else ""
    )
    return (
        "Mark every test on this tab BLOCKED. Do not mark any of them Passed.\n\n"
        "Each row says in one plain sentence what the test is waiting on. The same sentence is on "
        "the test itself, at the bottom of its Expected Results. They are grouped by what they are "
        "waiting on, so one thing being sorted out releases a whole group at once.\n\n"
        "TWO THINGS TO LOOK AT FIRST. (1) The column 'Already has a result?' - "
        + (f"{n_words(npass)} of these tests already says Passed in the test run, and it is "
           if npass == 1 else
           f"{n_words(npass)} of these tests already say Passed in the test run, and they are ")
        + "shaded pink. Those need changing to Blocked; they are the most useful thing on this tab. "
        f"(2) The biggest single group is '{biggest[0]}' with {biggest[1]} tests in it."
        + extra +
        "\n\nThe first group, 'A problem was found but no ticket exists for it yet', is the one "
        "written up in full on tab 2. Those tests are waiting on a ticket that nobody has been "
        "allowed to raise yet - which is exactly what tab 2 is for." +
        "\n\nThe 'What it is waiting on' wording is quoted from the test itself, so it matches what "
        "you will read on the case. The only thing changed is that document reference codes have "
        "been spelled out in plain words."
    )


T4_COLS = ["#", "Test", "Link", "What the test covers", "What it is waiting on",
           "Already has a result?"]


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    gc.alignment = WRAP
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, cols, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 130
    _hdr(ws, 4, cols)
    ws.freeze_panes = "A5"
    return ws


def _widths(ws, widths):
    for col, w in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[col].width = w


# ------------------------------------------------------------------------- build
def build_xlsx(D, path):
    wb = openpyxl.Workbook()

    # ---- tab 1
    ws = _sheet(wb, TAB1, tab1_title(D), tab1_note(D),
                ["#", "Topic", "What it means", "Why it matters"], first=True)
    _band(ws, 5, "start here", 4)
    r = 6
    for n, topic, what, why in tab1_rows(D):
        for j, v in enumerate([n, topic, what, why], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 190
        r += 1
    _widths(ws, [5, 30, 78, 70])

    # ---- tab 2
    ws = _sheet(wb, TAB2, f"{D['project']} - problems we found and confirmed, which have NOT been "
                          "reported anywhere", tab2_note(D), T2_COLS)
    nd = len(D['defects'])
    _band(ws, 5, f"{n_words(nd)} problem{'' if nd == 1 else 's'}, "
                 f"{'none of them' if nd != 1 else 'not'} reported anywhere", len(T2_COLS))
    r = 6
    for i, d in enumerate(D['defects'], 1):
        cases = "\n".join(f"C{c} - {link(c)}" for c in d['cases'])
        vals = [i, d['seen'], d['expected'], d['source'], d['steps'], cases,
                D['env'], d['owed'], d['against'], ""]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    _widths(ws, [4, 55, 50, 60, 62, 28, 34, 52, 52, 22])

    # ---- tab 3
    ws = _sheet(wb, TAB3, f"{D['project']} - tickets where the ticket and the product no longer "
                          "agree", tab3_note(D), T3_COLS)
    _band(ws, 5, "suggestions for a person - nothing here has been done", len(T3_COLS))
    r = 6
    for i, t in enumerate(D['tickets'], 1):
        cases = "\n".join(f"C{c} - {link(c)}" for c in t['cases']) or "-"
        key = t['key']
        tick = key if t.get('link') is None and not key.startswith('SV-') \
            else f"{key}\n{JIRA}{key}"
        vals = [i, tick, t['says'], t['actual'], cases, t['suggest']]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 210
        r += 1
    _widths(ws, [4, 30, 58, 66, 30, 56])

    # ---- tab 4
    ws = _sheet(wb, TAB4, f"{D['project']} - tests that cannot be run yet, and why - all "
                          f"{D['n_hold']}", tab4_note(D), T4_COLS)
    r = 5
    n = 0
    for g in GROUP_ORDER:
        rows = [x for x in D['holds'] if x['group'] == g]
        if not rows:
            continue
        _band(ws, r, f"{g} - {len(rows)} test{'' if len(rows) == 1 else 's'}", len(T4_COLS))
        r += 1
        for x in rows:
            n += 1
            res = x['result'] or "no result recorded - leave it, or mark Blocked"
            vals = [n, f"C{x['id']}", link(x['id']), x['title'], x['reason'], res]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.alignment = WRAP
                if x['result'] == 'Passed':
                    cell.fill = WARN_FILL
            ws.row_dimensions[r].height = 46
            r += 1
    if D['mismarked']:
        _band(ws, r, f"NOT marked to skip, but our own notes say they cannot be run - "
                     f"{len(D['mismarked'])} tests", len(T4_COLS))
        r += 1
        for x in D['mismarked']:
            n += 1
            vals = [n, f"C{x['id']}", link(x['id']), x['title'], x['why'],
                    "marked runnable by mistake - please mark Blocked and tell the QA lead"]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.alignment = WRAP
                cell.fill = WARN_FILL
            ws.row_dimensions[r].height = 60
            r += 1
    _widths(ws, [5, 11, 62, 62, 88, 44])

    wb.save(path)
    return path


def build_md(D, path):
    L = []
    A = L.append
    xlsx = os.path.basename(path).replace('.md', '.xlsx')
    A(f"# {tab1_title(D)}\n")
    A(f"**For the manual tester on {D['project']}. Release is tomorrow.**\n")
    A(f"> This is the plain-text twin of `{xlsx}`. Same content, same four sections.\n")
    A("---\n")

    A(f"## 1. {TAB1}\n")
    A(tab1_note(D) + "\n")
    for n, topic, what, why in tab1_rows(D):
        A(f"### {n}. {topic}\n")
        A(what + "\n")
        A(f"*Why it matters:* {why}\n")
    A("---\n")

    A(f"## 2. {TAB2}\n")
    A(tab2_note(D) + "\n")
    for i, d in enumerate(D['defects'], 1):
        A(f"### Problem {i} - {d['name']}\n")
        A(f"**What you will see.** {d['seen']}\n")
        A(f"**What should happen instead.** {d['expected']}\n")
        A(f"**Where that comes from.** {d['source']}\n")
        A(f"**Exactly how to see it.**\n\n{d['steps']}\n")
        A(f"**Where we saw it.** {D['env']}\n")
        A("**Tests affected.** " + " · ".join(f"[C{c}]({link(c)})" for c in d['cases']) + "\n")
        A(f"**What is still needed before filing.** {d['owed']}\n")
        A(f"**The strongest argument AGAINST it being a fault.** {d['against']}\n")
        A(f"*(Our own record of this: `{d['record']}`.)*\n")
        A("**Filed? / Ticket number:** ______________\n")
    A("---\n")

    A(f"## 3. {TAB3}\n")
    A(tab3_note(D) + "\n")
    for i, t in enumerate(D['tickets'], 1):
        key = t['key']
        head = f"[{key}]({JIRA}{key})" if key.startswith('SV-') else key
        A(f"### {i}. {head}\n")
        A(f"**What the ticket says.** {t['says']}\n")
        A(f"**What actually happens now.** {t['actual']}\n")
        if t['cases']:
            A("**Tests affected.** " + " · ".join(f"[C{c}]({link(c)})" for c in t['cases']) + "\n")
        A(f"**What we suggest (a suggestion only).** {t['suggest']}\n")
    A("---\n")

    A(f"## 4. {TAB4}\n")
    A(tab4_note(D) + "\n")
    for g in GROUP_ORDER:
        rows = [x for x in D['holds'] if x['group'] == g]
        if not rows:
            continue
        A(f"### {g} - {len(rows)} test{'' if len(rows) == 1 else 's'}\n")
        A("| Test | What the test covers | What it is waiting on | Already has a result? |")
        A("|---|---|---|---|")
        for x in rows:
            res = x['result'] or "-"
            flag = " ⚠️" if x['result'] == 'Passed' else ""
            A(f"| [C{x['id']}]({link(x['id'])}) | {x['title']} | {x['reason']} | {res}{flag} |")
        A("")
    if D['mismarked']:
        A(f"### NOT marked to skip, but our own notes say they cannot be run - "
          f"{len(D['mismarked'])} tests\n")
        A("| Test | What the test covers | Why it cannot be run |")
        A("|---|---|---|")
        for x in D['mismarked']:
            A(f"| [C{x['id']}]({link(x['id'])}) | {x['title']} | {x['why']} |")
        A("")
    A("---\n")
    A(f"*Every count and every list here was read from TestRail at {D['read_at']}. "
      f"The {D['project']} branch was on `{D['build']['now']}` when this was written "
      f"(last rebuilt {D['build']['now_modified']}). No Jira call was made while writing this "
      f"sheet, and nothing in TestRail was changed.*")
    open(path, 'w').write("\n".join(L) + "\n")
    return path


if __name__ == '__main__':
    for proj, stem in FILES.items():
        D = DATA[proj]
        x = build_xlsx(D, os.path.join(HERE, stem + '.xlsx'))
        m = build_md(D, os.path.join(HERE, stem + '.md'))
        print(f"{proj:14} {os.path.basename(x)}  +  {os.path.basename(m)}   "
              f"(defects {len(D['defects'])}, tickets {len(D['tickets'])}, "
              f"held {D['n_hold']}, passed-on-held {D['n_passed_on_hold']})")
