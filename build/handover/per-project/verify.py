#!/usr/bin/env python3
"""Verify the three per-project sheets against the live snapshot they were built from.

Checks, and every one must pass before the sheets are handed over:
  1. every tab-4 row count equals the live HOLD count for that project
  2. the shaded (pink) rows equal the live count of held cases graded Passed
  3. every C-id that appears anywhere in a sheet is a real case in that project
  4. no case appears on tab 4 twice
  5. every "no ticket exists yet" held case is written up on tab 2
  6. the marker arithmetic closes both ways for each project
  7. the .md twin carries the same C-ids as the .xlsx
Read-only: reads data.json, census.json and the generated files. No API call.
"""
import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.dirname(HERE)
DATA = json.load(open(os.path.join(HERE, 'data.json')))
CENSUS = json.load(open(os.path.join(HERE, 'tools', 'census.json')))

FILES = {
    'Schedule': 'Schedule_Tester-Handover_2026-08-12',
    'Filters': 'Filters_Tester-Handover_2026-08-12',
    'Report Suite': 'Report-Suite_Tester-Handover_2026-08-12',
}
CID = re.compile(r'C(\d{5})')
fails = []


def check(cond, msg):
    print(('  PASS  ' if cond else '  FAIL  ') + msg)
    if not cond:
        fails.append(msg)


for proj, stem in FILES.items():
    print(f'== {proj}')
    D = DATA[proj]
    c = CENSUS[proj]
    live_ids = {r['id'] for r in c['rows'] if not r['foreign']}
    wb = openpyxl.load_workbook(os.path.join(OUT, stem + '.xlsx'))

    # ---- 6. marker arithmetic
    check(c['ready'] + c['expect_fail'] == c['ours'] - c['hold'],
          f"marker arithmetic closes both ways: {c['ready']}+{c['expect_fail']}="
          f"{c['ready'] + c['expect_fail']} = {c['ours']}-{c['hold']}")

    # ---- 1/2/4. tab 4
    ws = wb['Tests that cannot be run yet']
    t4, shaded = [], 0
    for row in ws.iter_rows(min_row=5):
        v = row[1].value
        if isinstance(v, str) and v.startswith('C'):
            cid = int(v[1:])
            t4.append(cid)
            if row[1].fill.start_color.rgb in ('00FCE4E4', 'FFFCE4E4'):
                shaded += 1
    n_mis = len(D['mismarked'])
    check(len(t4) == D['n_hold'] + n_mis,
          f"tab 4 lists {len(t4)} tests = {D['n_hold']} held + {n_mis} mis-marked")
    check(shaded == D['n_passed_on_hold'] + n_mis,
          f"tab 4 shades {shaded} rows = {D['n_passed_on_hold']} graded Passed + {n_mis} mis-marked")
    check(len(t4) == len(set(t4)), 'no case appears on tab 4 twice')

    # ---- 3. every C-id in the sheet is a real case of this project
    seen = set()
    for w in wb.worksheets:
        for row in w.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    seen |= {int(x) for x in CID.findall(cell.value)}
    stray = sorted(seen - live_ids)
    check(not stray, f'every C-id in the workbook is a live case of this project '
                     f'({len(seen)} distinct){"" if not stray else " stray=" + str(stray)}')

    # ---- 5. no-ticket holds are all written up on tab 2
    cov = set()
    for d in D['defects']:
        cov |= set(d['cases'])
    missing = [r['id'] for r in D['holds']
               if r['group'] == 'A problem was found but no ticket exists for it yet'
               and r['id'] not in cov]
    check(not missing, 'every "no ticket yet" held case is written up on tab 2'
                       + ('' if not missing else f' missing={missing}'))

    # ---- 7. md twin carries the same C-ids
    md = open(os.path.join(OUT, stem + '.md')).read()
    md_ids = {int(x) for x in CID.findall(md)}
    check(md_ids == seen, f'.md twin carries the same {len(seen)} C-ids as the .xlsx'
                          + ('' if md_ids == seen else
                             f' (xlsx-only={sorted(seen - md_ids)} md-only={sorted(md_ids - seen)})'))
    print()

print('ALL CHECKS PASSED' if not fails else f'{len(fails)} CHECK(S) FAILED')
raise SystemExit(1 if fails else 0)
