#!/usr/bin/env python3
"""Generate the test-case change list for Vlad, the automation engineer (.xlsx + .md).

Standing Rule 65 — the QA lead, verbatim: "make sure to share with me the list of the
test cases always that you create/update/delete specially those which has that AUTOMATED
marker set by Vlad. because Vlad is our automation engineer and he always needs to know
if any test which he has already automated is updated/deleted."

MIRRORS the established sheet skeleton (Standing Rule 16) — A1 title / A2 note /
A4 header / A5 band / rows from 6, the same fills, fonts, wrap and freeze pane as
build/report-suite/questions-2026-08-06/gen_chris_sheet.py.

SCOPE: every case we created, updated or deleted on 11 and 12 August 2026, across
Filters, Schedule and the Report Suite. Established from the committed per-operation
logs AND verified against live TestRail — the two sets are equal in both directions.

RESEARCH ONLY — writes two files into this folder. No TestRail call, no Jira call, no
application call. Its input is vlad.json, derived live and read-only earlier in the pass.
"""
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
D = json.load(open(os.path.join(HERE, 'vlad.json')))
READ_AT = json.load(open(os.path.join(HERE, 'data.json')))['read_at']
XLSX = os.path.join(HERE, 'Test-Case-Changes-for-Vlad_2026-08-12.xlsx')
MD = os.path.join(HERE, 'Test-Case-Changes-for-Vlad_2026-08-12.md')

HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
HOT_FILL = PatternFill("solid", start_color="FCE4E4", end_color="FCE4E4")
WRAP = Alignment(wrap_text=True, vertical="top")
TR = "https://shopview.testrail.io/index.php?/cases/view/"

S = D['summary']

TITLE = "Test cases we changed on 11 and 12 August 2026 — for Vlad"

NOTE = (
    f"{S['total']} cases in total: {S['A'] + S['B']} updated, {S['C']} created, "
    f"{S['deleted']} deleted.\n\n"
    "SECTION A IS THE ONE THAT MATTERS TO YOU. It is the "
    f"{S['A']} cases TestRail flags as Automated. For each one, the last column says whether "
    "the change alters what an automated check should conclude. "
    f"{S['A_affecting']} of the {S['A']} do; the rest are sourcing and reference edits that "
    "change nothing you evaluate.\n\n"
    "ONE THING THAT APPLIES TO ALL OF SECTION A, so it is said once here rather than "
    f"repeated on {S['A']} rows: every one of them had a line added to the bottom of its "
    "expected results recording the date we read each source. If any of your checks compares "
    "the WHOLE expected-results field as one string, it will see a difference on all "
    f"{S['A']} — even the ones where nothing about the behaviour changed. If your checks read "
    "individual assertions, they will not.\n\n"
    "Sections B and C are for completeness. Section C is new cases, which you can see in "
    "TestRail yourself.\n\n"
    f"READ FROM TESTRAIL AT {READ_AT}. Another worker was editing the same suites during this "
    "pass, so a handful of Schedule cases moved after this snapshot; the case ids and the "
    "Automated set are unaffected.\n\n"
    "Please overrule us on the last column wherever you disagree — we have never seen your "
    "scripts, and something we call cosmetic can still break a check that matches an exact "
    "string."
)

CAVEAT = (
    "TWO THINGS THAT KEEP THIS LIST HONEST.\n\n"
    "1. Every case in Section A was checked individually against TestRail's own history to "
    "confirm that YOU set the Automated flag on it. All 44 came back as set by Vladimir "
    "Tomovic. None is on this list because a script assumed it.\n\n"
    "2. No Schedule case appears in Section A, and that is correct. Our own case-creation "
    "tooling used to fill in the Automated flag automatically on everything it created, which "
    "was wrong — that is our mistake, not your marking. Those flags were corrected to Not "
    "Automated on 11 August, so Schedule now shows nothing flagged Automated. A Schedule case "
    "that once read Automated was never evidence that you had automated it."
)

COLS_A = ["#", "Case", "Link", "Project", "Title", "What changed",
          "Does this change what an automated check should conclude?"]
COLS_B = ["#", "Case", "Link", "Project", "Title", "What changed",
          "Does this change what an automated check should conclude?"]
COLS_C = ["#", "Case", "Link", "Project", "Title", "Note"]


def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    gc.alignment = WRAP
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _widths(ws, widths):
    for col, w in zip("ABCDEFG", widths):
        ws.column_dimensions[col].width = w


def build_xlsx():
    wb = openpyxl.Workbook()

    # ---------------- Section A
    ws = wb.active
    ws.title = "A - Automated cases changed"
    ws["A1"] = "Section A — cases TestRail flags as Automated that we changed"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = NOTE
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 210
    ws["A3"] = CAVEAT
    ws["A3"].alignment = WRAP
    ws.row_dimensions[3].height = 130
    _hdr(ws, 5, COLS_A)
    ws.freeze_panes = "A6"
    r = 6
    hot = [x for x in D['A'] if x['affects'].startswith(('YES', 'PROBABLY'))]
    cold = [x for x in D['A'] if not x['affects'].startswith(('YES', 'PROBABLY'))]
    n = 0
    for band, rows in (("look at these first — the change alters what a check should conclude",
                        hot),
                       ("the rest — sourcing and reference edits only", cold)):
        _band(ws, r, f"{band}  ({len(rows)})", len(COLS_A))
        r += 1
        for x in rows:
            n += 1
            vals = [n, f"C{x['id']}", TR + str(x['id']), x['project'], x['title'],
                    x['what'], x['affects']]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.alignment = WRAP
                if rows is hot:
                    cell.fill = HOT_FILL
            ws.row_dimensions[r].height = 110 if rows is hot else 60
            r += 1
    _widths(ws, [5, 10, 56, 14, 60, 90, 66])

    # ---------------- Section B
    ws = wb.create_sheet("B - other cases updated")
    ws["A1"] = "Section B — the other cases we updated"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"{S['B']} cases. None of these is flagged Automated in TestRail, so none of them should "
        "affect anything you have already automated. They are here for completeness.\n\n"
        "The great majority were touched by two bulk passes that changed no wording a tester "
        "reads: one added the date we read each source to the sourcing line, and one corrected "
        "the product-description version cited in the References field."
    )
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 100
    _hdr(ws, 4, COLS_B)
    ws.freeze_panes = "A5"
    r = 5
    n = 0
    for proj in ("Filters", "Schedule", "Report Suite"):
        rows = [x for x in D['B'] if x['project'] == proj]
        hot = [x for x in rows if x['affects'].startswith(('YES', 'PROBABLY'))]
        cold = [x for x in rows if not x['affects'].startswith(('YES', 'PROBABLY'))]
        _band(ws, r, f"{proj} — {len(rows)}, of which {len(hot)} had a real change",
              len(COLS_B))
        r += 1
        for group in (hot, cold):
            for x in group:
                n += 1
                vals = [n, f"C{x['id']}", TR + str(x['id']), proj, x['title'],
                        x['what'], x['affects']]
                for j, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=j, value=v)
                    cell.alignment = WRAP
                    if group is hot:
                        cell.fill = HOT_FILL
                ws.row_dimensions[r].height = 80 if group is hot else 46
                r += 1
    _widths(ws, [5, 10, 56, 14, 60, 88, 60])

    # ---------------- Section C
    ws = wb.create_sheet("C - cases created")
    ws["A1"] = "Section C — cases we created"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (
        f"{S['C']} new cases. You can see new cases in TestRail yourself, so this section is "
        "for completeness rather than action. None of them is flagged Automated.\n\n"
        f"DELETIONS: {S['deleted']}. Stated rather than left out, because a deletion is the "
        "change that would break your suite hardest — no case was deleted on either day, on any "
        "of the three projects."
    )
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 90
    _hdr(ws, 4, COLS_C)
    ws.freeze_panes = "A5"
    r = 5
    n = 0
    for proj in ("Filters", "Schedule", "Report Suite"):
        rows = [x for x in D['C'] if x['project'] == proj]
        if not rows:
            continue
        _band(ws, r, f"{proj}  ({len(rows)})", len(COLS_C))
        r += 1
        for x in rows:
            n += 1
            vals = [n, f"C{x['id']}", TR + str(x['id']), proj, x['title'], x['what']]
            for j, v in enumerate(vals, 1):
                ws.cell(row=r, column=j, value=v).alignment = WRAP
            ws.row_dimensions[r].height = 40
            r += 1
    _widths(ws, [5, 10, 56, 14, 70, 30])

    wb.save(XLSX)
    return XLSX


def build_md():
    L = []
    A = L.append
    A(f"# {TITLE}\n")
    A("> Plain-text twin of `Test-Case-Changes-for-Vlad_2026-08-12.xlsx`. Same content.\n")
    A(NOTE + "\n")
    A("---\n")
    A(CAVEAT + "\n")
    A("---\n")

    A(f"## Section A — cases TestRail flags as Automated ({S['A']})\n")
    hot = [x for x in D['A'] if x['affects'].startswith(('YES', 'PROBABLY'))]
    cold = [x for x in D['A'] if not x['affects'].startswith(('YES', 'PROBABLY'))]
    for band, rows in (("Look at these first — the change alters what a check should conclude",
                        hot),
                       ("The rest — sourcing and reference edits only", cold)):
        A(f"### {band} ({len(rows)})\n")
        A("| Case | Project | Title | What changed | "
          "Does this change what an automated check should conclude? |")
        A("|---|---|---|---|---|")
        for x in rows:
            A(f"| [C{x['id']}]({TR}{x['id']}) | {x['project']} | {x['title']} | "
              f"{x['what']} | {x['affects']} |")
        A("")

    A(f"## Section B — the other cases we updated ({S['B']})\n")
    A("None of these is flagged Automated in TestRail. Listed for completeness.\n")
    for proj in ("Filters", "Schedule", "Report Suite"):
        rows = [x for x in D['B'] if x['project'] == proj]
        hot = [x for x in rows if x['affects'].startswith(('YES', 'PROBABLY'))]
        cold = [x for x in rows if not x['affects'].startswith(('YES', 'PROBABLY'))]
        A(f"### {proj} — {len(rows)}, of which {len(hot)} had a real change\n")
        A("| Case | Title | What changed | Affects a check? |")
        A("|---|---|---|---|")
        for x in hot + cold:
            A(f"| [C{x['id']}]({TR}{x['id']}) | {x['title']} | {x['what']} | {x['affects']} |")
        A("")

    A(f"## Section C — cases we created ({S['C']})\n")
    A(f"**Deletions: {S['deleted']}.** Stated rather than left out, because a deletion is the "
      "change that would break your suite hardest. No case was deleted on either day, on any of "
      "the three projects.\n")
    for proj in ("Filters", "Schedule", "Report Suite"):
        rows = [x for x in D['C'] if x['project'] == proj]
        if not rows:
            continue
        A(f"### {proj} ({len(rows)})\n")
        A("| Case | Title |")
        A("|---|---|")
        for x in rows:
            A(f"| [C{x['id']}]({TR}{x['id']}) | {x['title']} |")
        A("")

    A("---\n")
    A("*Established from our committed per-operation logs and verified against live TestRail on "
      "12 August 2026. The two sets are equal in both directions: every case TestRail shows as "
      f"changed on 11 or 12 August is in one of these three sections ({S['total']}), and every "
      "case our logs claim to have written shows as changed in TestRail. Neither set has a "
      "member the other lacks.*")
    open(MD, 'w').write("\n".join(L) + "\n")
    return MD


if __name__ == '__main__':
    print(build_xlsx())
    print(build_md())
