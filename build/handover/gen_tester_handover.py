#!/usr/bin/env python3
"""Generate the manual tester's handover sheet of 2026-08-12 (.xlsx + .md).

MIRRORS 1:1 (Standing Rule 16) the established question-sheet peers:
  build/report-suite/questions-2026-08-06/Questions-for-Chris-Ward_Report-Suite_2026-08-06.xlsx
  build/filters/questions-2026-08-06/Questions-for-Branko-Cicovic_Filters-and-Schedule_2026-08-06.xlsx
Same row layout (A1 title / A2 note / A4 header / A5 band / items from row 6), the
same fills and fonts, the same wrap and freeze pane, the same one-band-per-group shape.

The COLUMNS differ per tab because the content differs — a defect row is not a
question row — but the skeleton, the styling and the conventions are the peers'.

THE READER IS A MANUAL QA TESTER SEEING THIS FOR THE FIRST TIME (the QA lead,
verbatim: "make this manual tester handover sheet simple in words and understandable
as the manual tester will be reading it for the first time and it will be and can be
confusing for them"). So: short sentences, no jargon, no spec anchors, no endpoint
names, no HTTP codes, no internal ids beyond the TestRail C-id and its link (Rule 8).

RESEARCH ONLY — this script writes two files into this folder. It makes no TestRail
call, no Jira call and no application call of any kind. Its input is data.json, which
was derived live and read-only earlier in the same pass.
"""
import json
import os

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, 'data.json')))
READ_AT = DATA['read_at']
STAMP = (
    f"Every count and every list here was read from TestRail at {READ_AT}.\n\n"
    "PLEASE NOTE THE TIME, because it matters today. Another member of the team was working "
    "on the same tests while this sheet was being written, and the numbers moved twice in the "
    "hour before it was finished: three Schedule tests came off the skip list because they "
    "were checked and passed. So treat this as a photograph taken at that minute, not a "
    "permanent count. If a test on the last tab no longer says it is waiting on something "
    "when you open it, believe the test, not this sheet."
)
XLSX = os.path.join(HERE, 'Manual-Tester-Handover_2026-08-12.xlsx')
MD = os.path.join(HERE, 'Manual-Tester-Handover_2026-08-12.md')

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WARN_FILL = PatternFill("solid", start_color="FCE4E4", end_color="FCE4E4")
WRAP = Alignment(wrap_text=True, vertical="top")

TR = "https://shopview.testrail.io/index.php?/cases/view/"

TAB1 = "Read me first"
TAB2 = "Problems found, not reported"
TAB3 = "Old tickets that mislead"
TAB4 = "Tests that cannot be run yet"


def link(cid):
    return f"{TR}{cid}"


# --- every count below is COMPUTED from the live snapshot, never typed in ---
_H = DATA['holds']
N_HOLD = len(_H)
N_PASSED = sum(1 for h in _H if h['result'] == 'Passed')
PER = {p: sum(1 for h in _H if h['project'] == p)
       for p in ('Filters', 'Schedule', 'Report Suite')}
TOTALS = DATA['totals']          # ours / foreign, per project
N_TOTAL = sum(TOTALS[p]['ours'] for p in TOTALS)
N_RUN = N_TOTAL - N_HOLD
N_SIGNIN = DATA['signin_blocked']


# ============================================================ TAB 1 — Read me first
T1_TITLE = "Handover for the manual test team — 12 August 2026"
T1_NOTE = (
    "Release is tomorrow. This sheet is for you. It has three lists in it, one per tab, and "
    "this page explains what they are and the one rule that matters most. You do not need to "
    "know anything about how the tests were written to use it.\n\n"
    "Row 5 says exactly when these numbers were read, and why that matters today."
)

T1_ROWS = [
    ("1", "The one rule that matters most",
     "If a test says it cannot be run yet, mark it Blocked. Do not mark it Passed.\n\n"
     "Every test that cannot be run says so in its own words, at the very bottom of its "
     "Expected Results. If you open a test and it tells you it is waiting on something, "
     "that is the test telling you it cannot be judged today. Mark it Blocked and move on.",
     "This is not a formality. It has already happened. Read from TestRail as this sheet was "
     f"written: {N_PASSED} tests that cannot currently be judged already have a Passed result "
     "recorded against them. A test nobody could run cannot have passed. They are listed on "
     "the last tab and shaded, and they need changing to Blocked before anyone reads them as "
     "evidence that the feature works."),

    ("2", "What is on each tab",
     "Tab 2, \"Problems found, not reported\". Three real faults we found and confirmed. "
     "No ticket has been raised for any of them. If you want to raise the ticket yourself, "
     "everything you need to paste in is on the row, and there is a blank column for you to "
     "write the ticket number in.\n\n"
     "Tab 3, \"Old tickets that mislead\". Four tickets where what the ticket says and what "
     "the product actually does no longer match. Three say they are closed and the fault is "
     "still there. One is still open and the fault is gone.\n\n"
     f"Tab 4, \"Tests that cannot be run yet\". All {N_HOLD} of them, with one plain "
     "sentence each.",
     "Tabs 2 and 3 are suggestions for a person to action. Nothing on them has been done for "
     "you, and nothing has been raised in Jira - we were asked to hold off on that."),

    ("3", "How to mark a test, in four lines",
     "The test tells you it is waiting on something -> Blocked.\n\n"
     "The test tells you what you will see today, and that is what you see -> Failed, and "
     "raise nothing new. It is already known.\n\n"
     "The test tells you what you will see today, and you see something different -> that is "
     "a NEW problem. Please report it.\n\n"
     "The test says nothing special and it works -> Passed.",
     "Anything else, or the test simply does not make sense to you: mark it Blocked and tell "
     "the QA lead. Never guess a result."),

    ("4", "How many tests there are",
     "\n".join(
         f"{p}: {TOTALS[p]['ours'] - PER[p]} to run, {PER[p]} to skip "
         f"({TOTALS[p]['ours']} in total)."
         for p in ('Filters', 'Schedule', 'Report Suite')
     ) + f"\n\nAll three together: {N_RUN} to run, {N_HOLD} to skip, {N_TOTAL} in total.",
     "A few tests in Filters and the Report Suite were written by a colleague and are not "
     "counted above and not listed here. They are not ours to change and not yours to "
     "compare against this sheet: 5 in Filters, 12 in the Report Suite."),

    ("5", "When these numbers were taken",
     STAMP,
     "A count taken an hour earlier said 91 tests to skip and 16 of them wrongly marked "
     f"Passed. It now says {N_HOLD} and {N_PASSED}, because three Schedule tests were checked "
     "and passed while this was being written. Neither figure is wrong - they are different "
     "minutes."),

    ("6", "One thing to know about dates on the tests",
     "Near the bottom of every test there is a line saying which build it was last checked "
     "against. On most tests that build is older than the one you will be testing on today.\n\n"
     "That does not make the test wrong. What a test expects comes from the written product "
     "description, not from the build - so a newer build never changes what a test should "
     "expect. What can change is the exact wording of a button or where something sits on "
     "screen.",
     "So if a button is named slightly differently from what the test says, that is worth "
     "telling the QA lead, but it is not automatically a fault in the product."),
]


# ================================================ TAB 2 — problems found, not reported
T2_TITLE = "Problems we found and confirmed, which have NOT been reported anywhere"
T2_NOTE = (
    f"{len(DATA['defects'])} problems, all of them in Schedule. Every one was seen with our own "
    "eyes on the build named on its row, and no Jira ticket covers any of them.\n\n"
    "WHAT TO DO WITH THIS TAB: nothing, unless you want to. We were asked to hold off raising "
    "tickets, so these are written out in full in case you would rather raise them yourself. "
    "Everything you need is on the row. Use the last column to write down the ticket number if "
    "you do raise one.\n\n"
    "BEFORE YOU RAISE ONE, PLEASE READ THE 'What is still needed' COLUMN. Each row says what is "
    "already proven and what is still missing. A ticket that a developer can argue with costs "
    "more than no ticket at all."
)

T2_COLS = ["#", "Where", "What you will see", "What should happen instead",
           "Where that comes from (quoted)", "Exactly how to see it",
           "Tests affected", "Where we saw it", "What is still needed before filing",
           "Filed? / Ticket number"]

T2 = DATA['defects']


# ============================================== TAB 3 — old tickets that mislead
T3_TITLE = "Old tickets where the ticket and the product no longer agree"
T3_NOTE = (
    "FOUR tickets. Three are closed and the fault is still there. One is still open and the "
    "fault is gone. Either way, anyone reading the ticket today would get the wrong idea.\n\n"
    "WHAT TO DO WITH THIS TAB: nothing has been done. We have not reopened, closed or commented "
    "on any of these. The last column is what we would suggest, and it is only a suggestion - "
    "a person has to decide.\n\n"
    "The tests themselves are already honest about all four, so you can run them normally. They "
    "tell you what you will see and what to do about it."
)

T3_COLS = ["#", "Ticket", "Where", "What the ticket says", "What actually happens now",
           "Tests affected", "What we suggest"]

T3 = DATA['tickets']


# =============================================== TAB 4 — tests that cannot be run
T4_TITLE = f"Tests that cannot be run yet, and why — all {N_HOLD}"
T4_NOTE = (
    "Mark every test on this tab BLOCKED. Do not mark any of them Passed.\n\n"
    "Each row says in one plain sentence what the test is waiting on. The same sentence is on "
    "the test itself, at the bottom of its Expected Results.\n\n"
    f"TWO THINGS TO LOOK AT FIRST. (1) The column 'Already has a result?' - {N_PASSED} of these "
    "tests already say Passed in the test run, and they are shaded. Those need changing to "
    "Blocked; they are the most useful thing on this tab. (2) "
    f"{N_SIGNIN['Schedule']} Schedule tests are all waiting on the same one thing - a second "
    "login that is not an administrator. One login would release them all at once.\n\n"
    "The 'What it is waiting on' wording is quoted from the test itself, so it matches what you "
    "will read on the case. The only thing changed is that document reference codes have been "
    "spelled out as 'the written description'."
)

T4_COLS = ["#", "Project", "Test", "Link", "What the test covers",
           "What it is waiting on", "Already has a result?"]

T4 = DATA['holds']


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    gc.alignment = WRAP
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, cols, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    ws.row_dimensions[2].height = 120
    _hdr(ws, 4, cols)
    ws.freeze_panes = "A5"
    return ws


def _widths(ws, widths):
    for col, w in zip("ABCDEFGHIJ", widths):
        ws.column_dimensions[col].width = w


def build_xlsx():
    wb = openpyxl.Workbook()

    # ---- tab 1
    ws = _sheet(wb, TAB1, T1_TITLE, T1_NOTE, ["#", "Topic", "What it means", "Why it matters"],
                first=True)
    _band(ws, 5, "start here", 4)
    r = 6
    for n, topic, what, why in T1_ROWS:
        for j, v in enumerate([n, topic, what, why], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 190
        r += 1
    _widths(ws, [5, 30, 78, 70])

    # ---- tab 2
    ws = _sheet(wb, TAB2, T2_TITLE, T2_NOTE, T2_COLS)
    _band(ws, 5, "three problems, none of them reported anywhere", len(T2_COLS))
    r = 6
    for i, d in enumerate(T2, 1):
        cases = "\n".join(f"C{c} - {link(c)}" for c in d['cases'])
        vals = [i, d['where'], d['seen'], d['expected'], d['source'], d['steps'],
                cases, d['env'], d['owed'], ""]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    _widths(ws, [4, 22, 55, 50, 60, 62, 30, 34, 56, 22])

    # ---- tab 3
    ws = _sheet(wb, TAB3, T3_TITLE, T3_NOTE, T3_COLS)
    _band(ws, 5, "three closed but still broken, one open but already fixed", len(T3_COLS))
    r = 6
    for i, t in enumerate(T3, 1):
        cases = "\n".join(f"C{c} - {link(c)}" for c in t['cases'])
        vals = [i, f"{t['key']}\nhttps://shopview.atlassian.net/browse/{t['key']}",
                t['where'], t['says'], t['actual'], cases, t['suggest']]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 200
        r += 1
    _widths(ws, [4, 30, 20, 58, 66, 30, 56])

    # ---- tab 4
    ws = _sheet(wb, TAB4, T4_TITLE, T4_NOTE, T4_COLS)
    r = 5
    n = 0
    for proj in ("Filters", "Schedule", "Report Suite"):
        rows = [x for x in T4 if x['project'] == proj]
        _band(ws, r, f"{proj} — {len(rows)} tests to skip", len(T4_COLS))
        r += 1
        for x in rows:
            n += 1
            res = x['result'] or "no result recorded — leave it, or mark Blocked"
            vals = [n, proj, f"C{x['id']}", link(x['id']), x['title'], x['reason'], res]
            for j, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=j, value=v)
                cell.alignment = WRAP
                if x['result'] == 'Passed':
                    cell.fill = WARN_FILL
            ws.row_dimensions[r].height = 46
            r += 1
    _widths(ws, [5, 14, 11, 62, 62, 76, 44])

    wb.save(XLSX)
    return XLSX


def build_md():
    L = []
    A = L.append
    A(f"# {T1_TITLE}\n")
    A("**For the manual test team. Release is tomorrow.**\n")
    A("> This is the plain-text twin of "
      "`Manual-Tester-Handover_2026-08-12.xlsx`. Same content, same four sections.\n")
    A("---\n")

    A(f"## 1. {TAB1}\n")
    A(T1_NOTE + "\n")
    for n, topic, what, why in T1_ROWS:
        A(f"### {n}. {topic}\n")
        A(what + "\n")
        A(f"*Why it matters:* {why}\n")
    A("---\n")

    A(f"## 2. {TAB2}\n")
    A(T2_NOTE + "\n")
    for i, d in enumerate(T2, 1):
        A(f"### Problem {i} — {d['where']}\n")
        A(f"**What you will see.** {d['seen']}\n")
        A(f"**What should happen instead.** {d['expected']}\n")
        A(f"**Where that comes from.** {d['source']}\n")
        A(f"**Exactly how to see it.** {d['steps']}\n")
        A(f"**Where we saw it.** {d['env']}\n")
        A("**Tests affected.** " + " · ".join(f"[C{c}]({link(c)})" for c in d['cases']) + "\n")
        A(f"**What is still needed before filing.** {d['owed']}\n")
        A("**Filed? / Ticket number:** ______________\n")
    A("---\n")

    A(f"## 3. {TAB3}\n")
    A(T3_NOTE + "\n")
    A("| # | Ticket | Where | What the ticket says | What actually happens now | "
      "Tests affected | What we suggest |")
    A("|---|---|---|---|---|---|---|")
    for i, t in enumerate(T3, 1):
        cases = " · ".join(f"[C{c}]({link(c)})" for c in t['cases'])
        A(f"| {i} | [{t['key']}](https://shopview.atlassian.net/browse/{t['key']}) | "
          f"{t['where']} | {t['says']} | {t['actual']} | {cases} | {t['suggest']} |")
    A("")
    A("---\n")

    A(f"## 4. {TAB4}\n")
    A(T4_NOTE + "\n")
    for proj in ("Filters", "Schedule", "Report Suite"):
        rows = [x for x in T4 if x['project'] == proj]
        A(f"### {proj} — {len(rows)} tests to skip\n")
        A("| Test | What the test covers | What it is waiting on | Already has a result? |")
        A("|---|---|---|---|")
        for x in rows:
            res = x['result'] or "—"
            flag = " ⚠️" if x['result'] == 'Passed' else ""
            A(f"| [C{x['id']}]({link(x['id'])}) | {x['title']} | {x['reason']} | {res}{flag} |")
        A("")
    A("---\n")
    A("*" + STAMP.replace("\n\n", " ") + "*")
    open(MD, 'w').write("\n".join(L) + "\n")
    return MD


if __name__ == '__main__':
    print(build_xlsx())
    print(build_md())
