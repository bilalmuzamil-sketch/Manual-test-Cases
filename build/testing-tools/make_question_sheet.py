#!/usr/bin/env python3
"""Build a PO/dev question sheet as a REAL SPREADSHEET, in the established column shape.

QA lead, 2026-09-01, verbatim: "the questions should always be in Excel or google soreadsheet, in a
lay man language for a nontechnical person to understand."

So a question sheet is never a markdown table any more. This writes .xlsx in the schema the
2026-08-05 Branko sheet established (Rule 16 - mirror the format that already exists):

    # | Topic | What happens now | The question | Options | Your answer

with one sheet per feature and a final "QA internal" sheet that the PO is not meant to read.

LAYMAN RULES ENFORCED HERE (Rules 7 and 9), because a sheet that fails them is worse than no sheet:
  * no case ids, no spec anchors (S7-N1), no API paths, no HTTP terms, no field names
  * the build's own on-screen labels, in quotes
  * every question answerable by a non-technical person in one sitting, with OPTIONS offered
"""
import re, sys, json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

BANNED = [
    # 🛑 A CASE ID HAS NO DIGIT WIDTH. Until 2026-09-03 this read `\bC\d{5}\b`, so a case id of
    # any other length would have LEAKED INTO A PO-FACING SHEET UNDETECTED - Rules 7/9, the one
    # thing this gate exists to stop. It is not hypothetical in either direction: C281 is a live
    # case id in this workspace today, and the estate reaches C100000 by counting.
    # ANCHORING - `\bC\d+\b`, not a bare `C\d+` and not a wider digit count:
    #   the literal capital C plus a LEADING word boundary is the anchor, and the match is
    #   case-SENSITIVE, so nothing that is not C-prefixed can leak in - '8/5/2026' -> no match,
    #   'v3.10-49b5fe3' -> no match, 'SV-8582' -> no match, 'ABC123' -> no match (no boundary
    #   before that C). No digit floor: C281 proves short ids are real.
    #   This gate errs toward FLAGGING: another C-prefixed code (a part number like C1608054)
    #   is reported here too. That costs the author one re-word, whereas the miss it replaces
    #   costs a case id in front of the PO. The message says id-shaped, not "definitely a case".
    (re.compile(r'\bC\d+\b'), 'a case-id-shaped token (C followed by digits)'),
    (re.compile(r'\bS\d+-[RNE]\d+\b'), 'a specification anchor'),
    (re.compile(r'/api/|\bHTTP\b|\bJSON\b|\bAPI\b'), 'an API or HTTP term'),
    (re.compile(r'\b[a-z]+[A-Z][A-Za-z]*\b'), 'a camelCase identifier'),
    (re.compile(r'\bcustom_[a-z]+\b'), 'a database field name'),
]
HEAD = ['#', 'Topic', 'What happens now', 'The question', 'Options', 'Your answer']

def check(text):
    out = []
    for rx, why in BANNED:
        m = rx.search(text or '')
        if m:
            out.append(f'{why}: {m.group(0)!r}')
    return out

def build(path, title, intro, sheets, internal=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    problems = []
    for sheet_name, rows in sheets:
        ws = wb.create_sheet(sheet_name[:31])
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=13)
        ws['A2'] = intro
        ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.append([])
        ws.append(HEAD)
        for c in range(1, len(HEAD) + 1):
            cell = ws.cell(row=4, column=c)
            cell.font = Font(bold=True)
            cell.fill = PatternFill('solid', fgColor='DDDDDD')
        for i, r in enumerate(rows, 1):
            ws.append([str(i), r['topic'], r['now'], r['question'], r['options'], ''])
            for col in ('B', 'C', 'D', 'E'):
                ws[f'{col}{ws.max_row}'].alignment = Alignment(wrap_text=True, vertical='top')
            for field in ('topic', 'now', 'question', 'options'):
                for p in check(r[field]):
                    problems.append(f'{sheet_name} row {i} {field}: {p}')
        for col, w in zip('ABCDEF', (5, 34, 60, 62, 52, 30)):
            ws.column_dimensions[col].width = w
        ws.freeze_panes = 'A5'
    if internal:
        ws = wb.create_sheet('QA internal - not for the PO')
        ws.append(['This sheet is for us. It is the traceability the PO does not need to read.'])
        ws.append([])
        ws.append(['Question #', 'Cases it decides', 'Requirement', 'Why it is a question and not a defect'])
        for c in range(1, 5):
            ws.cell(row=3, column=c).font = Font(bold=True)
        for r in internal:
            ws.append(r)
        for col, w in zip('ABCD', (12, 34, 24, 76)):
            ws.column_dimensions[col].width = w
        for row in ws.iter_rows(min_row=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    wb.save(path)
    return problems

if __name__ == '__main__':
    spec = json.load(open(sys.argv[1]))
    probs = build(spec['path'], spec['title'], spec['intro'],
                  [(s['name'], s['rows']) for s in spec['sheets']], spec.get('internal'))
    print('written', spec['path'])
    if probs:
        print('LAYMAN CHECK FAILED — jargon a non-technical reader will not understand:')
        for p in probs: print('   ', p)
        sys.exit(1)
    print('layman check: clean — no case ids, spec anchors, API terms or field names')
