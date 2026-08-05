#!/usr/bin/env python3
"""Generate the Branko Cicovic question sheet (.xlsx + .md) - Schedule AND Filters.

ONE COMBINED FILE, deliberately. Reasoning is on the QA-only tab and in the .md:
Standing Rule 55 says "sweep every open one onto ONE sheet so he answers in a
single sitting rather than a drip of separate asks", and two files to the same
person on the same day IS that drip. Rule 55's actual worry - a PO who owns
several things answering the wrong one - is solved by naming the project on every
ROW, which this sheet does, and by giving each project its own TAB.

Mirrors build/report-suite/rulings-2026-08-05/gen_followup_sheet.py 1:1
(Standing Rule 16): same six reader columns in the same order, same row layout
(A1 title / A2 note / A4 header / A5 band / items from row 6), same fills and
fonts, same freeze pane, same column widths, and the same QA-only mapping tab
that is never sent.

Standing Rule 45(e): where this sheet says two sentences disagree, BOTH are quoted
in plain words from the live Confluence page fetched immediately before writing.

DOCUMENTATION ONLY - this script writes two files into this folder. It makes no
TestRail or Jira call of any kind.
"""

import os
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Questions-for-Branko-Cicovic_Schedule-and-Filters_2026-08-05.xlsx")
MD = os.path.join(HERE, "Questions-for-Branko-Cicovic_Schedule-and-Filters_2026-08-05.md")

# ---------------------------------------------------------------- mirrored style
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
GRP_FILL = PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7")
WRAP = Alignment(wrap_text=True, vertical="top")
COLS = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]

TAB1_NAME = "Schedule - your document"
TAB2_NAME = "Schedule - engineering only"
TAB3_NAME = "Filters"
TAB4_NAME = "QA internal - not for Branko"

THANKS = (
    "Thank you - you answered and closed the mobile question this morning, and the updated "
    "description settled it cleanly, so a whole group of tests came unstuck. This sheet is the rest "
    "of what we are waiting on, gathered into one place so you can go through it in a single sitting "
    "rather than getting a trickle of separate messages. It covers TWO of your projects, so every "
    "question says which one it belongs to - Schedule or Filters - and the tabs are split the same "
    "way. Nothing here is a complaint, and one of them is honestly our own fault for not sending it "
    "sooner."
)

TAB1_NOTE = (
    "PLEASE START WITH QUESTION 1 - and an apology with it. That question was written on 22 July and "
    "we never actually sent it to you. Two tests have been sitting parked ever since waiting for an "
    "answer you were never asked for. That delay is ours, not yours. " + THANKS
)

# ------------------------------------------------------------------- the questions
# (topic, what-happens-now, the question, options)
TAB1 = [
    (
        "Schedule (the technician scheduling calendar) - planning a job across several days "
        "(the multi-day spread story, SV-8691, under epic SV-8685)",

        "FIRST, AN APOLOGY: we wrote this question on 22 July and never sent it. Two tests have been "
        "parked ever since. That is our delay, not yours.\n\n"
        "When a job is too big for one day, the schedule spreads it across several working days. Your "
        "description tells us two opposite things about days the shop is closed - a public holiday, "
        "or an inventory day - and both sentences are still in the current version:\n\n"
        "One says: \"Shop closures and public holidays are NOT SKIPPED in V1.\" In other words the "
        "spread puts shifts on a closed day like any other day.\n\n"
        "The other says: \"Shop closures (holidays, inventory days) are defined at the shop level and "
        "BLOCK the spread step from placing shifts on those days.\" In other words the spread jumps "
        "over closed days.\n\n"
        "Why we are asking: these are two completely different tests and we cannot write both. We "
        "have not guessed - the two tests say plainly that the point is undecided and are waiting.",

        "When a job is spread across several days, should the schedule skip days the shop is closed, "
        "or place shifts on them anyway?",

        "A) SKIP THEM - closed days are jumped over, the same way weekends already are, and the job "
        "runs on to the next open day.\n\n"
        "B) DO NOT SKIP THEM - closed days get shifts like any other day for this first version, and "
        "somebody moves them by hand if needed.",
    ),
    (
        "Schedule (the technician scheduling calendar) - the menu on an empty part of the calendar "
        "(the events story, SV-8696, and the access-level section of your description)",

        "Nobody had spotted this one before, and no test is wrong because of it - but your "
        "description contradicts itself, so a reader who happens to read the wrong half would test "
        "the wrong thing.\n\n"
        "In two places it says: \"LEFT-CLICK on empty grid space opens a menu with: Create event, New "
        "work order.\"\n\n"
        "In two other places, where it lists what each access level unlocks, it twice calls the same "
        "thing a \"RIGHT-CLICK context menu\".\n\n"
        "Our tests follow left-click, and the product agrees with them, so nothing is broken today.\n\n"
        "Why we are asking: it is a one-word correction in your description, and until it is made "
        "anyone reading only the access-level section will look for the wrong mouse click.",

        "Which is correct - does that menu open on a left-click or a right-click?",

        "A) LEFT-CLICK - as the two earlier places say. The access-level section is the wording that "
        "needs correcting.\n\n"
        "B) RIGHT-CLICK - then the product and our tests are both wrong and it becomes a developer "
        "job.",
    ),
    (
        "Schedule (the technician scheduling calendar) - weekends for a technician with no hours set "
        "(the working hours story, SV-8699, and the conflict story, SV-8697)",

        "This one is a gap rather than a contradiction, and no test is parked on it - we are asking "
        "so that we are not quietly relying on our own reading.\n\n"
        "Three separate parts of your description do not quite add up for a technician who has no "
        "working hours of their own set up:\n\n"
        "- the default working day is given as 7:00 AM to 7:00 PM, with nothing said about which days "
        "of the week that covers;\n"
        "- spreading a job \"automatically skips weekends\";\n"
        "- but a shift placed on a weekend counts as a clash to be warned about.\n\n"
        "So we cannot tell whether a weekend is simply an ordinary working day for such a technician, "
        "or a day the schedule should push back on.\n\n"
        "Why we are asking: it decides whether a warning should appear, and a warning that should not "
        "be there is just as much a bug as a missing one.",

        "For a technician with no working hours of their own set up, is a weekend a normal working "
        "day or a day the schedule should warn about?",

        "A) A DAY TO WARN ABOUT - weekends are outside normal hours for everyone unless someone has "
        "set weekend hours for them.\n\n"
        "B) A NORMAL WORKING DAY - the default 7 to 7 applies to all seven days, and only the "
        "spread step avoids weekends.",
    ),
]

TAB2 = [
    (
        "Schedule (the technician scheduling calendar) - shifts and events that already existed "
        "(under epic SV-8685)",

        "Six of our tests describe things that only the ENGINEERING plan describes - your own "
        "description does not mention them at all. We have kept the tests, because they cover real "
        "risks and throwing them away would lose that cover, but each one says openly that it rests "
        "on an engineering note rather than on a product decision.\n\n"
        "We will not present something as a requirement when no product document says it. So until "
        "you tell us these are right, those tests stay parked - that is the honest consequence and we "
        "would rather you saw it than have us quietly promote an engineering note into a rule.\n\n"
        "This first one: after this feature goes out, everything that was already on the calendar "
        "should still be there - same technician, same day, same time, same job.\n\n"
        "Why we are asking: it is the difference between a safe release and losing somebody's "
        "existing plan.",

        "Should everything already on the calendar survive the release completely unchanged?",

        "A) YES - nothing already scheduled may move, change or disappear. (Please confirm and we "
        "will treat it as a requirement.)\n\n"
        "B) Not quite - please say what is allowed to change.",
    ),
    (
        "Schedule (the technician scheduling calendar) - how a multi-day job looks on the Dashboard "
        "(under epic SV-8685)",

        "Another one that only the engineering plan describes.\n\n"
        "When a job is spread over, say, five days, it becomes five separate day-by-day entries on "
        "the calendar. The engineering note says the Dashboard should still show that job as ONE "
        "line covering the whole stretch, not as five separate lines.\n\n"
        "Why we are asking: if it is five lines, the Dashboard suddenly looks five times busier than "
        "it is, and nobody has written down which is intended.",

        "On the Dashboard, should a job scheduled across several days appear as one line or as one "
        "line per day?",

        "A) ONE LINE covering the whole stretch.\n\n"
        "B) ONE LINE PER DAY.",
    ),
    (
        "Schedule (the technician scheduling calendar) - an appointment set while creating a job "
        "(under epic SV-8685)",

        "Another one that only the engineering plan describes.\n\n"
        "When somebody sets an appointment date and time while creating a job, the engineering note "
        "says that appointment should turn up on the schedule calendar at that date and time, and "
        "behave like anything else on the calendar.\n\n"
        "Why we are asking: your description does not mention appointments at all, so we cannot tell "
        "whether the two things are meant to be connected.",

        "Should an appointment set while creating a job appear on the schedule calendar?",

        "A) YES - it appears at that date and time and behaves like anything else on the calendar.\n\n"
        "B) NO - appointments and the schedule calendar are separate things.",
    ),
    (
        "Schedule (the technician scheduling calendar) - jobs from another branch "
        "(under epic SV-8685)",

        "Another one that only the engineering plan describes, and this is the one with the most at "
        "stake.\n\n"
        "The engineering note says a shift should only ever appear on the calendar of the branch the "
        "job belongs to - even when the technician doing it also works at another branch - and that "
        "somebody looking at one branch should not be able to reach or change another branch's shift "
        "at all.\n\n"
        "It also gives a reason we think you should see: if a shift's branch were worked out from the "
        "technician instead of from the job, then moving a technician to another branch would quietly "
        "move all of their PAST shifts too - history would rewrite itself.\n\n"
        "Why we are asking: this is about one branch seeing another branch's work, so it is worth "
        "your explicit yes rather than our assumption.",

        "Should a shift only ever appear on the calendar of the branch its job belongs to?",

        "A) YES - the job's branch decides, always, and one branch can never see or change "
        "another's shifts.\n\n"
        "B) Something else - please describe it.",
    ),
    (
        "Schedule (the technician scheduling calendar) - the priority on a job "
        "(under epic SV-8685)",

        "Another one that only the engineering plan describes.\n\n"
        "The engineering note says a job offers a priority choice of High, Medium and Low, that a "
        "brand-new job has NONE of them picked to begin with, and that the chosen priority then shows "
        "on the job.\n\n"
        "Why we are asking: \"nothing picked to start with\" is the kind of detail that is easy to "
        "get wrong in either direction, and no product document states it.",

        "Is that right - High, Medium, Low, with nothing pre-selected on a new job?",

        "A) YES - three choices, nothing pre-selected.\n\n"
        "B) Not quite - please say what the choices are and what a new job should start with.",
    ),
    (
        "Schedule (the technician scheduling calendar) - a limit on how long a spread can be "
        "(the multi-day spread story, SV-8691, under epic SV-8685)",

        "Two of our tests say that spreading a job further than about eight weeks, or into more than "
        "about a hundred and twenty separate days, should stop and ask the person to confirm before "
        "going ahead.\n\n"
        "Those two numbers appear ONLY in the engineering plan. We searched your description and "
        "neither number is in it anywhere. The product currently does not warn at all - a very long "
        "spread just goes through.\n\n"
        "We have NOT raised that as a fault, precisely because we cannot show a product document that "
        "asks for it.\n\n"
        "Why we are asking: there are three possible answers here and we are not going to guess "
        "between them.",

        "Should a very long spread warn the person first, and if so at what point?",

        "A) YES, at those numbers - about eight weeks, or about a hundred and twenty days. Please "
        "confirm and we will treat it as a requirement.\n\n"
        "B) YES, but at different numbers - please say what they should be.\n\n"
        "C) NO limit at all - a spread of any length just goes ahead. (Then we will delete the two "
        "tests.)",
    ),
]

TAB3 = [
    (
        "Filters (the filter buttons on the Work Orders list) - where the filter bar sits "
        "(the filter bar layout story, SV-8786, under epic SV-8785)",

        "Your description says the filter bar sits BELOW the row of tabs (All, Estimates, Completed, "
        "My Work Orders). The design shows the same thing. In the product the five filter buttons sit "
        "ON THE SAME ROW as the tabs.\n\n"
        "One of our own tests used to wave this away with a note saying the product behaves this way "
        "\"on purpose for now\" - and nothing anywhere backed that up. That note was wrong and it has "
        "been removed. Our test now expects what your description says.\n\n"
        "Since then a developer job has been raised to move the bar below the tabs, on the grounds "
        "that the product does not match your description. So somebody is about to change the product "
        "on the strength of that reading.\n\n"
        "Why we are asking: if you actually wanted them on one row, that developer job should be "
        "cancelled and your description updated instead. Better to ask you now than after the change "
        "has been made.",

        "Should the filter buttons be moved below the tabs, or did you want them on the same row as "
        "the tabs?",

        "A) MOVE THEM BELOW - as your description and the design say. The developer job is correct "
        "and nothing needs changing in writing.\n\n"
        "B) SAME ROW IS WHAT I WANTED - then the developer job should be cancelled and the "
        "description updated to say so.",
    ),
    (
        "Filters (the filter buttons on the Work Orders list) - the wording on the phone button "
        "(the mobile filter bar story, SV-8797, under epic SV-8785)",

        "This is a tiny one, and it exists only because our tests have to quote the exact words a "
        "tester will see on screen.\n\n"
        "Your description calls the button \"Apply filters\", with a small f. On a phone the button "
        "actually reads \"Apply Filters\", with a capital F.\n\n"
        "Why we are asking: we would rather your description and the screen said the same thing than "
        "have our test quietly differ from one of them.",

        "Which spelling is the right one?",

        "A) \"Apply Filters\" with a capital F - the description can be tidied to match the screen.\n\n"
        "B) \"Apply filters\" with a small f - then the button on screen should be corrected.",
    ),
    (
        "Filters (the filter buttons on the Work Orders list) - a pointer in your own description "
        "that leads to the wrong place (the mobile filter bar story, SV-8797, under epic SV-8785)",

        "NO DECISION NEEDED - this is just a helpful heads-up about a typo-level slip, and it has "
        "already cost one round of confusion.\n\n"
        "In the phone section, the paragraph about the filter buttons says they work like the desktop "
        "\"with one exception\", and then points the reader at a numbered paragraph further down. But "
        "the paragraph it points at is the one about the page search box - not an exception at all. "
        "The real exception, that a phone only filters when the person taps the apply button, is the "
        "very next paragraph after the one it points to.\n\n"
        "We think we know how it happened: when you tidied the numbering yesterday you moved the "
        "apply-button paragraph down one place, and the pointer above it kept pointing at the old "
        "number.\n\n"
        "Why we are telling you: a reader who follows that pointer lands on the wrong paragraph and "
        "concludes there is no exception.",

        "Nothing to decide - please just repoint that reference to the apply-button paragraph next "
        "time you have the document open.",

        "(No options - noted for your next edit.)",
    ),
    (
        "Filters (the filter buttons on the Work Orders list) - the Parts and Reports write-up "
        "(under epic SV-8785)",

        "GENTLE STATUS ASK, not a new question - we know this is already with you and we are not "
        "chasing.\n\n"
        "Eight of our tests cover filter buttons on the Parts pages and the Reports pages. They were "
        "written from the designs back in July, and they are parked because that part of the product "
        "is not built yet and because your write-up for it has not arrived.\n\n"
        "To be straight with you: the write-up on its own will not unpark them - the feature still "
        "has to be built before anybody can run them. What the write-up does is let us finish the "
        "tests properly instead of leaving them resting on a design alone.\n\n"
        "Why we are asking: only so we can tell our own management honestly whether this is weeks "
        "away or months.",

        "Roughly when do you expect the Parts and Reports write-up, and is that part of the product "
        "still planned for this release?",

        "A) It is coming shortly, and it is still in this release.\n\n"
        "B) It has moved to a later release - please say roughly when.\n\n"
        "C) It has been dropped. (Then we will ask about deleting the eight tests.)",
    ),
]

CLOSED_NOTE = (
    "ONE THING WE ARE NOT ASKING YOU, AND WHY. There was an open question in Jira about the filter "
    "bar's position - SV-8876, raised by Ahtasham. He CLOSED IT HIMSELF this morning, about two hours "
    "after raising it, and raised a developer job instead on the grounds that the product does not "
    "match your description. So that question is no longer waiting for you. We have not left it out "
    "though: the half of it that still genuinely needs YOUR answer - whether you actually wanted the "
    "buttons on one row, in which case that developer job should be cancelled - is question 1 on the "
    "Filters tab above. We would rather ask you the narrower question than let a change be built on "
    "somebody else's reading of your document."
)


# ------------------------------------------------------------------------ helpers
def _hdr(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        if c:
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP


def _band(ws, row, text, ncols):
    gc = ws.cell(row=row, column=1, value=text.upper())
    gc.font = Font(bold=True)
    for j in range(1, ncols + 1):
        ws.cell(row=row, column=j).fill = GRP_FILL


def _sheet(wb, name, title, note, band, items, widths, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws["A2"] = note
    ws["A2"].font = Font(bold=True)
    ws["A2"].alignment = WRAP
    _hdr(ws, 4, COLS)
    _band(ws, 5, band, 6)
    r = 6
    for i, (topic, now, q, opts) in enumerate(items, 1):
        for j, v in enumerate([i, topic, now, q, opts, ""], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        ws.row_dimensions[r].height = 300
        r += 1
    for col, w in zip("ABCDEF", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A5"
    return ws, r


# --------------------------------------------------------------- QA-only mapping
TR = "https://shopview.testrail.io/index.php?/cases/view/"


def links(*ids):
    return " · ".join(f"C{i} {TR}{i}" for i in ids)


QA_ROWS = [
    ("Tab 1", "1",
     "SCHEDULE: does the multi-day spread skip shop closures? (NEVER SENT - our delay)",
     "Schedule spec CONTRADICTS ITSELF and always has. §4.5: \"Uses the technician's own working "
     "hours. Automatically skips weekends when business hours are not set for them. Shop closures "
     "and public holidays are not skipped in V1.\" §12: \"Shop closures (holidays, inventory days) "
     "are defined at the shop level and block the spread step from placing shifts on those days.\" "
     "BOTH still present at live v23.",
     "SCH-EDGE-05 (C30089); SCH-SPREAD-07 (C29983) - the ONLY two Schedule cases whose marker reads "
     "\"HOLD - waiting on the product owner's answer, and the question has not been sent yet\"",
     links(30089, 29983),
     "Confluence page 713031682, LIVE version 23 (2026-07-30T10:40:32Z), fetched "
     "2026-08-05T17:00:12Z; both sentences extracted verbatim from that fetch. RULE 48 / RULE 36 "
     "HONESTY: the blocker is US. The question was drafted 2026-07-22 and never sent. The two cases "
     "have carried a do-not-automate note for two weeks waiting on an answer nobody asked for.",
     "A (skip) -> both cases take the §12 reading and §4.5 is Branko's to correct. B (do not skip) "
     "-> both keep the §4.5 reading, which is what they currently assert, and §12 is Branko's to "
     "correct. EITHER WAY +2 to ready-to-automate: 158 -> 160."),
    ("Tab 1", "2",
     "SCHEDULE: left-click or right-click for the empty-cell menu?",
     "NEW IN THE 2026-08-05 AUDIT - nobody had noticed it. §4.10 and §7 both say \"Left-click on "
     "empty grid space opens a menu with: Create event, New work order.\" §14.1 lists among the "
     "hidden affordances a \"right-click context menu\"; §14.2 lists among what Schedule: Edit "
     "unlocks \"shift and event creation (including via right-click context menu and day-view "
     "click-to-create)\".",
     "SCH-REAS-03 (C30054); SCH-PERM-02 (C30075); SCH-PERM-04 (C30077) - ALL THREE ARE READY, NOT "
     "HELD. Our cases follow §7 (left-click) and the build agrees, so no case is wrong.",
     links(30054, 30075, 30077),
     "All four passages extracted verbatim from the live v23 fetch, 2026-08-05T17:00:12Z. NO CASE "
     "COUNT MOVES on this answer - it is a specification defect, reported not worked around.",
     "A (left-click) -> no case changes; §14.1/§14.2 are Branko's to correct. B (right-click) -> "
     "3 cases change AND it becomes a developer defect, because the build does left-click."),
    ("Tab 1", "3",
     "SCHEDULE: is a weekend a working day for a technician with no custom hours?",
     "Audit gap 3 - asserted nowhere as a rule. §4.2's default \"7:00 AM to 7:00 PM\" carries no "
     "weekday restriction; §4.5 says the spread \"automatically skips weekends\"; §4.11 counts a "
     "weekend shift as a conflict. The three cannot be reconciled from the text.",
     "Touches SCH-CONF-02 (C30024); SCH-CAP-01 (C30030) - both READY, neither held",
     links(30024, 30030),
     "Live v23, 2026-08-05T17:00:12Z. NO CASE COUNT MOVES. Included because Rule 55 says sweep every "
     "open ambiguity onto one sheet rather than drip-feed them, and because we are otherwise relying "
     "on our own reading of three sentences that do not agree.",
     "A (warn) -> the conflict expectation is confirmed on his authority. B (normal day) -> the "
     "weekend-conflict expectation is narrowed to technicians who DO have hours set."),

    ("Tab 2", "1",
     "SCHEDULE: do pre-existing shifts and events survive the release? (engineering-sourced)",
     "Expected behaviour rests on the ENGINEERING TECH PLAN only. Standing Rule 30: engineering "
     "intent never overrules product truth; Rule 57: the sources of expected behaviour are the PRD, "
     "the epic's stories and the PO's answers - a tech plan is none of the three. SPEC IS ENTIRELY "
     "SILENT.",
     "SCH-REG-01 (C38867)",
     links(38867),
     "Searched live Schedule v23 for any data-migration requirement: none. The case says on its face "
     "that it has no numbered requirement behind it.",
     "A -> the case's basis is upgraded from an engineering note to a product decision and its "
     "provenance line is re-stamped. B -> the case is narrowed to what he allows to change."),
    ("Tab 2", "2",
     "SCHEDULE: Dashboard row model for a multi-day job (engineering-sourced)",
     "Engineering plan only; SPEC ENTIRELY SILENT. The tech plan says the Dashboard collapses a "
     "spread series into one row.",
     "SCH-REG-02 (C38868)",
     links(38868),
     "Live v23 searched: no requirement covers the Dashboard's row model at all.",
     "A (one row) -> the assertion becomes product-sourced. B (one per day) -> the case is rewritten "
     "and it is probably also a developer defect."),
    ("Tab 2", "3",
     "SCHEDULE: does a work-order appointment appear on the board? (engineering-sourced)",
     "Engineering plan only; SPEC ENTIRELY SILENT.",
     "SCH-REG-03 (C38869)",
     links(38869),
     "Live v23 searched: appointments are not mentioned.",
     "A -> product-sourced. B -> the case is deleted, because it would be testing a connection that "
     "is not meant to exist."),
    ("Tab 2", "4",
     "SCHEDULE: is a shift scoped to the JOB's branch, never the technician's? (engineering-sourced) "
     "- the highest-stakes of the six",
     "Engineering plan only; SPEC ENTIRELY SILENT on both halves. TWO cases, one screen-level and "
     "one at the interface level, so this row covers both.",
     "SCH-REG-04 (C38870); SCH-API-04 (C38875)",
     links(38870, 38875),
     "Live v23 searched: no requirement on location-scoping of a shift read. The tech plan's stated "
     "reason - that deriving the branch from the technician would relocate their historical shifts "
     "on transfer - is quoted TO HIM in plain words, because it is the strongest argument for A and "
     "he should see it rather than be asked cold.",
     "A -> both become product-sourced and cross-branch leakage becomes a real, citable requirement. "
     "B -> both are rewritten to whatever he describes."),
    ("Tab 2", "5",
     "SCHEDULE: work-order priority choices and the default (engineering-sourced)",
     "Engineering plan only; SPEC ENTIRELY SILENT.",
     "SCH-REG-05 (C38871)",
     links(38871),
     "Live v23 searched: priority is not mentioned.",
     "A -> product-sourced as written. B -> the choices and the default are corrected."),
    ("Tab 2", "6",
     "SCHEDULE: is there an 8-week / 120-shift spread cap at all? (engineering-sourced, and the "
     "product does NOT do it)",
     "The two numbers appear ONLY in the engineering plan - searched live Schedule v23 and NEITHER "
     "number is present anywhere. Observed live on v3.5-be42149: a 76h 36m spread produced 7 shifts "
     "with no warning and nothing asking for confirmation; no 409 and no 422 exists on the build.",
     "SCH-SPREAD-11 (C38863) - HOLD \"the feature is not built yet\"; SCH-API-02 (C38873) - same "
     "HOLD",
     links(38863, 38873),
     "NOTHING HAS BEEN FILED for this, deliberately, and Rule 51 is the reason for the second case: "
     "it is interface-only, so it may not be raised as a ticket without the QA lead's explicit "
     "separate go-ahead. Written up in build/schedule/final-viu-2026-08-05/API-ASK.md.",
     "A (those numbers) -> both become product-sourced and a developer defect becomes writable, "
     "though the interface half still needs the QA lead's Rule-51 go-ahead. B (different numbers) -> "
     "both are rewritten. C (no limit) -> BOTH CASES ARE DELETED, which is the only answer that "
     "reduces the suite."),

    ("Tab 3", "1",
     "FILTERS: filter bar below the tabs, or on the same row? (NARROWED - SV-8876 is CLOSED)",
     "Filters spec S1-R1, live v18, verbatim: \"The filter bar is displayed below the tab navigation "
     "row (All, Estimates, Completed, My Work Orders) by default.\" The build renders the chips on "
     "the tabs' own row. ⚠️ SV-8876 (Ahtasham, raised 2026-08-05T06:17:01-0500) is NOT OPEN - he "
     "CLOSED IT HIMSELF at 08:38:16-0500 with the comment \"closing this as it was a gap with test "
     "case, I've updated the test case here >> C29557 And created a story defect >> as the build is "
     "not behaving as per PRD\", and raised SV-8883 (Story Defect, Open, parent SV-8786) instead. "
     "BRANKO NEVER ANSWERED IT. His second question - please update S1-R1 if the same-row layout is "
     "approved - died with the ticket, and that is the half this row asks.",
     "FLT-BAR-01 (C29557) - one of the five cases whose \"known and accepted\" waiver was removed by "
     "the 2026-08-05 audit; now READY - EXPECT FAIL",
     links(29557),
     "Confluence page 572030978, LIVE version 18, fetched 2026-08-05T17:00:12Z. SV-8876 and SV-8883 "
     "both read live from Jira the same pass, read-only. Neither was touched (Rule 38). NOTE the "
     "Rule-31(a) trap on this page: its in-body field reads \"Version: 1.6\" while the Confluence "
     "version is 18.",
     "A (move it below) -> C29557 stands exactly as repaired, SV-8883 is correct, nothing changes in "
     "writing. B (same row was intended) -> SV-8883 should be cancelled AND S1-R1 updated - and "
     "note that would NOT restore our old waiver note, because a waiver in a test case was never a "
     "valid substitute for the document saying so (Rule 57)."),
    ("Tab 3", "2",
     "FILTERS: \"Apply Filters\" (capital F) on screen vs \"Apply filters\" in the spec",
     "Spec v18 writes \"Apply filters\" in BOTH places it appears (S12-R6 and the Key Decision). "
     "Observed live: on-screen text exactly \"Apply Filters\", data-test-id apply_filters. Standing "
     "Rule 9 requires our cases to quote the label the tester actually sees.",
     "The 8 mobile cases that name the button: FLT-MOB-01..07 (C29621-C29627) and FLT-MOB-10 "
     "(C29630). All READY - none held on this.",
     links(29621, 29622, 29623, 29624, 29625, 29626, 29627, 29630),
     "Live Filters v18 fetched 2026-08-05T17:00:12Z, both occurrences lowercase. The build label was "
     "observed on v3.4.2-d00239b in the 2026-08-05 final-check pass. NO CASE COUNT MOVES.",
     "A (capital F) -> the spec is tidied and our cases already quote the screen correctly. "
     "B (lowercase) -> a cosmetic developer defect, and our cases must keep quoting the screen "
     "(Rule 9) while expecting the spec's casing."),
    ("Tab 3", "3",
     "FILTERS: S12-R2's cross-reference points at the wrong paragraph (NO DECISION - a heads-up)",
     "PROVEN LIVE, INCLUDING THE CAUSE. S12-R2 (v18): \"The filter chips behave like desktop with "
     "one exception (see S12-R5)…\" But S12-R5 (v18) is \"The page search control is shown on mobile "
     "and behaves as it does on desktop…\" - not an exception at all. The real exception is S12-R6: "
     "\"Unlike desktop, mobile does not filter in real time…\" CAUSE CONFIRMED from Branko's OWN v17 "
     "version message, read live: \"Fix Story 12 numbering: deferred-apply requirement renumbered to "
     "S12-R6, placed after the page-search S12-R5.\" He moved the target and the pointer stayed.",
     "No case is blocked. It cost one round of confusion in the 2026-08-05 cleanup pass, which had "
     "to prove S12-R6 covers a SINGLE filter's sheet and not only the combined one.",
     "n/a",
     "Confluence page 572030978 v18 body + /wiki/rest/api/content/572030978/version history, both "
     "fetched live 2026-08-05T17:00Z. This row deliberately carries NO options - it is the one "
     "no-decision item on the sheet.",
     "Nothing to resolve. He repoints the reference at his next edit."),
    ("Tab 3", "4",
     "FILTERS: status of the Parts and Reports write-up (GENTLE STATUS ASK, per the QA lead's own "
     "ruling)",
     "RULE 48 - THE RULING THAT FREEZES THIS IS THE QA LEAD'S OWN, and it was the right call. He "
     "ruled \"lets wait for Brankos PRD\" when asked what it would take to apply the staged "
     "Parts/Reports groups. It was correct, because applying them means asserting behaviour that no "
     "written product source supports - exactly what Rule 57 forbids. Nothing has changed since to "
     "reopen it; this row only asks WHEN, so management can be told honestly.",
     "The 8 cases marked \"HOLD - the feature is not in the product yet\": FLT-PARTS-01 (C38904); "
     "FLT-PARTS-09 (C38905); FLT-PARTS-11 (C38906); FLT-PARTS-12 (C38907); FLT-PARTS-13 (C38908); "
     "FLT-RPTS-01 (C38909); FLT-RPTS-21 (C38910); FLT-RPTS-22 (C38911). Also related: FLT-RPTS-23 "
     "(C38882), \"HOLD - the report filter bars are not in the product yet beyond the first report "
     "tab\".",
     links(38904, 38905, 38906, 38907, 38908, 38909, 38910, 38911, 38882),
     "The cases were authored 2026-07-27 from the captured Figma designs and Branko's 2026-07-31 "
     "answers. The spec's §2 Feature Overview and §4 Key Decisions mention Parts and Reports filters "
     "in prose only - no numbered requirements exist for them at live v18.",
     "HONESTY, AND IT IS SAID TO HIM IN THE SHEET: the write-up ALONE does not move the "
     "ready-to-automate figure, because the feature still has to ship before anyone can run these. "
     "The write-up lets the cases stop resting on a design alone. C (dropped) is the only answer "
     "that changes the count, by removing 8 cases from the population."),
]

QA_NOTES = [
    "WHY ONE COMBINED FILE AND NOT TWO. The brief left the choice open. ONE file, for three reasons. "
    "(1) Standing Rule 55 says in terms: \"Do NOT let ambiguities stack up across days either: sweep "
    "every open one onto ONE sheet so he answers in a single sitting rather than a drip of separate "
    "asks.\" Two files to the same person on the same day IS that drip. (2) Rule 55's actual worry - "
    "a PO who owns several things answering the wrong one - is addressed by naming the project on "
    "every ROW, which every one of the 13 rows does, and by one TAB per project. It is not addressed "
    "by splitting the file; two files make it MORE likely that one is answered and the other is lost. "
    "(3) He owns three projects, so a per-project file convention would eventually mean three sheets "
    "in one week. Pointer stubs are placed in build/filters/branko-questions-2026-08-05/ and "
    "build/schedule/branko-questions-2026-08-05/ so the file is discoverable from either project.",

    "SOURCE CURRENCY (Standing Rule 31), all fetched LIVE IMMEDIATELY BEFORE WRITING. Schedule spec: "
    "Confluence page 713031682, version 23, last edited 2026-07-30T10:40:32Z, fetched "
    "2026-08-05T17:00:12Z, HTTP 200. Filters spec: page 572030978, version 18, last edited "
    "2026-08-04T18:19:21Z, fetched 2026-08-05T17:00:12Z, HTTP 200. Filters version HISTORY also "
    "fetched live to confirm the v17 renumbering that caused the stale cross-reference. Jira read "
    "live the same pass, read-only: epic SV-8685 = 15 stories (SV-8686..SV-8700, all Ready for QA); "
    "epic SV-8785 = 14 stories (SV-8786..SV-8799); SV-8876 full issue + comments + changelog; "
    "SV-8883. THE CONFLUENCE VERSION NUMBER IS USED THROUGHOUT - the Filters page's in-body field "
    "still reads \"Version: 1.6\" against a real version of 18, which is the Rule-31(a) trap exactly.",

    "WHAT THE LIVE RE-CHECK CHANGED, AND IT CHANGED A WHOLE QUESTION. The brief asked for SV-8876 to "
    "be put to Branko as Ahtasham's OPEN clarification question. It is NOT open: Ahtasham closed it "
    "himself at 2026-08-05T08:38:16-0500, two hours and twenty-one minutes after raising it, and "
    "converted it into SV-8883. Branko never answered it. Asking him to answer a closed ticket would "
    "have been wrong, so the question was NARROWED to the half that is still genuinely his and still "
    "genuinely undecided - whether he wanted the same-row layout, in which case SV-8883 should be "
    "cancelled. The sheet's closing note tells him plainly that the ticket was closed and why we are "
    "still asking. Ahtasham's resolution was, for the record, CORRECT under Rule 57: a closed ticket "
    "is not a specification amendment, so the build failing S1-R1 is a defect and not a spec question.",

    "SCHEDULE CASE ARITHMETIC. Current ready-to-automate = 158 (137 READY + 21 READY-EXPECT-FAIL, out "
    "of 165). 7 cases carry HOLD: 3 not built (C38873, C30017, C38863) · 2 waiting on the product "
    "owner with the question NEVER SENT (C30089, C29983) · 1 needs shop business hours switched on, a "
    "shared setting on this estate (C29970) · 1 needs a real daylight-saving clock change, next one "
    "1 November 2026 (C38865). TAB 1 QUESTION 1 RELEASES EXACTLY 2: 158 -> 160. Tab 2 question 6 "
    "answered A or B keeps its 2 cases held on the not-built ground; answered C DELETES them. Tab 1 "
    "questions 2 and 3, and Tab 2 questions 1 to 5, move NO count at all - what they change is "
    "whether 8 cases rest on a product decision or on an engineering note, which is a Rule-20 "
    "authenticity matter, not a readiness one. That distinction is stated rather than blurred.",

    "FILTERS CASE ARITHMETIC. Current ready-to-automate = 100 (82 READY + 18 READY-EXPECT-FAIL, out "
    "of 110). 10 cases carry HOLD: 8 Parts/Reports not in the product (C38904-C38911) · 1 report "
    "filter bars not built beyond the first tab (C38882) · 1 needs a second test login to prove one "
    "person's saved filters do not reach another (C38895, FLT-API-06). NONE of the four Filters "
    "questions moves that figure, and the sheet says so to his face on the Parts/Reports row rather "
    "than implying his answer unblocks work it does not.",

    "WORDING RULES APPLIED (Standing Rules 7 + 55). Every reader-facing row names the PROJECT in its "
    "Topic column - \"Schedule (the technician scheduling calendar)\" or \"Filters (the filter "
    "buttons on the Work Orders list)\" - because Branko also owns Global Search and a bare question "
    "would be ambiguous to him. Story and epic keys are given in plain form ONLY where they orient "
    "him, always alongside a plain description of the story, never as a bare key. NO case IDs, NO "
    "requirement anchors, NO HTTP terms, NO endpoint names, NO enum names, NO bug codes and NOT ONE "
    "use of the word VIU appear in anything he reads. Each question carries a one-line reason so the "
    "consequence is visible. Where the sheet says two sentences disagree, BOTH are quoted in plain "
    "words (Rule 45(e)).",

    "TONE AND THE APOLOGY. He answered and closed the mobile question this morning and his v18 update "
    "settled it cleanly, so the sheet opens by saying so. Tab 1's note carries an explicit apology: "
    "the shop-closures question was drafted 2026-07-22 and NEVER SENT, and two cases have been parked "
    "for two weeks waiting on an answer nobody asked him for. Rule 36 requires the blocker to be "
    "named honestly, and here the blocker is us. It is stated in the reader-facing text, not hidden "
    "on this tab.",

    "THE ENGINEERING-SOURCED SIX, AND WHY THE CONSEQUENCE IS SPELLED OUT TO HIM. Tab 2's first row "
    "tells him plainly that we will not present something as a requirement when no product document "
    "says it, and that those tests therefore stay parked until he answers. The brief asked for that "
    "to be explicit and it is - but note the precise position: the 8 cases are NOT counted as a "
    "readiness shortfall (6 are READY, 2 are HOLD-not-built), because a case can be automatable while "
    "its basis is un-ratified. What is at risk is AUTHENTICITY (Rule 20), not automatability, and "
    "conflating the two would have overstated the ask.",

    "NOTHING HAS BEEN WRITTEN ANYWHERE. This sheet is a draft for the QA lead to send. No TestRail "
    "write, no Jira write, no case edit and no CLAUDE.md edit was made in producing it. The only "
    "calls made were read-only Confluence GETs and read-only Jira GETs. SV-8876 and SV-8883 were read "
    "and NOT touched (Rule 38 - and SV-8883 is Ahtasham's).",

    "BOTH BUILDS ARE STILL PARTIAL SOURCES. Neither QA branch has been declared final, so the Rule-49 "
    "queues build/schedule/final-viu-2026-08-05/RECHECK-QUEUE.md and "
    "build/filters/final-viu-2026-08-05/RECHECK-QUEUE.md are both OPEN and every pass/fail verdict on "
    "both projects is PROVISIONAL. That does not weaken any question here: all 13 are decided by "
    "documents, not by a build.",
]


def write_xlsx():
    wb = openpyxl.Workbook()
    w_narrow = [4, 34, 52, 42, 46, 22]
    w_wide = [4, 34, 58, 44, 48, 22]

    _sheet(wb, TAB1_NAME,
           "SCHEDULE - questions about your written description - Branko Cicovic - 2026-08-05",
           TAB1_NOTE,
           "Schedule - three places where your description needs one decision", TAB1,
           w_wide, first=True)

    _sheet(wb, TAB2_NAME,
           "SCHEDULE - six things only the engineering plan describes - Branko Cicovic - 2026-08-05",
           "Six tests that describe behaviour your own document does not mention - only the "
           "engineering plan does. We have kept them because they cover real risk, but we will not "
           "call something a requirement when no product document says it, so they stay parked until "
           "you tell us they are right. " + THANKS,
           "Schedule - please ratify or correct each of these six", TAB2, w_narrow)

    ws3, r = _sheet(wb, TAB3_NAME,
                    "FILTERS - four questions - Branko Cicovic - 2026-08-05",
                    "Four Filters questions. One of them is about to change the product, so it is "
                    "worth a minute of your time; one needs no decision at all. " + THANKS,
                    "Filters - the filter buttons on the Work Orders list", TAB3, w_narrow)
    ws3.cell(row=r + 1, column=2, value=CLOSED_NOTE).alignment = WRAP
    ws3.cell(row=r + 1, column=2).font = Font(bold=True)
    ws3.row_dimensions[r + 1].height = 220

    ws4 = wb.create_sheet(TAB4_NAME)
    ws4["A1"] = ("QA-ONLY - INTERNAL - NOT FOR BRANKO. Do not send this tab. TestRail C-ids, "
                 "requirement anchors and live evidence live here so the reader-facing tabs stay "
                 "plain (Standing Rules 7, 8 and 55).")
    ws4["A1"].font = Font(bold=True)
    ws4["A1"].alignment = WRAP
    r = 3
    ws4.cell(row=r, column=1,
             value="PER-ITEM MAPPING - EVERY READER-FACING QUESTION").font = Font(bold=True)
    r += 2
    _hdr(ws4, r, ["Tab", "Item", "What it asks", "Where the ambiguity comes from",
                  "Affected internal case IDs (TestRail C-id)", "TestRail links",
                  "Spec anchors + live evidence", "What each answer resolves to"])
    r += 1
    for row in QA_ROWS:
        for j, v in enumerate(row, 1):
            ws4.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="HONESTY AND METHOD NOTES").font = Font(bold=True)
    r += 1
    for n in QA_NOTES:
        ws4.cell(row=r, column=1, value=n).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFGH", [10, 7, 40, 40, 50, 50, 60, 60]):
        ws4.column_dimensions[col].width = w

    wb.save(XLSX)
    return XLSX


def write_md():
    def block(items, start=1):
        out = []
        for i, (topic, now, q, opts) in enumerate(items, start):
            out.append(f"### Item {i}.0 — {topic}\n")
            out.append("**What happens now**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in now.split("\n")) + "\n")
            out.append("**The question**\n")
            out.append(f"> {q}\n")
            out.append("**Options**\n")
            out.append("\n".join("> " + l if l.strip() else ">" for l in opts.split("\n")) + "\n")
            out.append("**Your answer:** _______________________________________________\n")
        return "\n".join(out)

    tab1_note_md = TAB1_NOTE.replace(THANKS, "").strip()

    md = f"""# Questions for Branko Cicovic — Schedule and Filters — 2026-08-05

**Projects: Schedule (epic SV-8685) and Filters (epic SV-8785) · Product Owner: Branko Cicovic**

**This is the plain-language twin of
`Questions-for-Branko-Cicovic_Schedule-and-Filters_2026-08-05.xlsx`.**
The spreadsheet is the version to send; it mirrors the established Chris Ward sheets' format exactly,
and it carries a QA-only tab that must not be forwarded.

**DRAFT — NOT SENT. Nothing has been written to TestRail or Jira.**

**ONE COMBINED FILE, deliberately.** Standing Rule 55 says to sweep every open ambiguity onto one
sheet so a PO answers in a single sitting "rather than a drip of separate asks" — and two files to
the same person on the same day is that drip. Rule 55's real concern, a PO who owns several things
answering the wrong one, is handled by **naming the project on every single row** and by giving each
project its **own tab**. He owns three projects; a file-per-project habit would mean three sheets a
week.

{THANKS}

**Thirteen items in total: three Schedule questions about his own document, six Schedule items that
only the engineering plan describes, and four Filters questions — one of which needs no decision at
all.**

**Live source versions confirmed at 2026-08-05T17:00:12Z, immediately before writing** — Schedule
specification **version 23** · Filters specification **version 18**. Every sentence quoted below
comes from that fetch. (Note for us, not for him: the Filters page's in-body field still reads
"Version: 1.6" — the Confluence version number is the one used here.)

---

## Tab 1 — {TAB1_NAME}

{tab1_note_md}

{block(TAB1)}

---

## Tab 2 — {TAB2_NAME}

Six tests that describe behaviour his own document does not mention — only the engineering plan does.
We have kept them because they cover real risk, but we will not call something a requirement when no
product document says it, so they stay parked until he tells us they are right.

{block(TAB2)}

---

## Tab 3 — {TAB3_NAME}

Four Filters questions. One of them is about to change the product, so it is worth a minute of his
time; one needs no decision at all.

{block(TAB3)}

### And one note explaining something we are NOT asking

> {CLOSED_NOTE}

---

## QA-only — not for Branko

The internal question-to-case mapping lives on the spreadsheet's `{TAB4_NAME}` tab: every question's
affected TestRail case IDs with links, the requirement anchors quoted verbatim from the live pages,
and what each possible answer resolves to. It also records the source-currency block, the case-count
arithmetic for **both** projects (Schedule **158 → 160** on Tab 1 question 1; Filters **100**, which
none of these questions moves), the reason one combined file was chosen, and the live re-check that
changed a whole question — **SV-8876 is closed, not open**.

**Do not forward that tab.**
"""
    open(MD, "w").write(md)
    return MD


if __name__ == "__main__":
    print("wrote", write_xlsx())
    print("wrote", write_md())
