#!/usr/bin/env python3
"""Compile the four JSON test-case files into one multi-tab Excel workbook."""
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BUILD = os.path.join(ROOT, "build")
OUT = os.path.join(ROOT, "custom-roles-test-cases.xlsx")


def load(name):
    with open(os.path.join(BUILD, name), "r") as f:
        return json.load(f)


COLUMNS = [
    "Test ID", "Title", "Related Jira", "Permission(s) Under Test",
    "Dependency Mode", "Priority", "Type", "Preconditions", "Role Setup",
    "Test Data", "Steps", "Expected Results", "Expected Final Result",
    "Source / VIU Status",
]

# Column widths keyed by header name.
WIDTHS = {
    "Test ID": 14, "Title": 35, "Related Jira": 14,
    "Permission(s) Under Test": 26, "Dependency Mode": 24, "Priority": 10,
    "Type": 12, "Preconditions": 40, "Role Setup": 40, "Test Data": 30,
    "Steps": 60, "Expected Results": 60, "Expected Final Result": 45,
    "Source / VIU Status": 22,
}
WRAP_COLS = {"Preconditions", "Role Setup", "Steps", "Expected Results",
             "Expected Final Result", "Title", "Test Data"}

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
TOP = Alignment(vertical="top")
TOP_WRAP = Alignment(vertical="top", wrap_text=True)


def steps_cell(rec):
    return "\n".join("{}. {}".format(s["n"], s["action"]) for s in rec["steps"])


def expected_cell(rec):
    return "\n".join("{}. {}".format(s["n"], s["expected"]) for s in rec["steps"])


def row_for(rec):
    return [
        rec["test_id"], rec["title"], rec["jira"], rec["permission"],
        rec["dependency_mode"], rec["priority"], rec["type"],
        rec["preconditions"], rec["role_setup"], rec["test_data"],
        steps_cell(rec), expected_cell(rec), rec["expected_final"],
        rec["source_viu"],
    ]


def style_data_sheet(ws):
    # Header styling.
    for c, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[name]
    # Data cells alignment.
    for r in range(2, ws.max_row + 1):
        for c, name in enumerate(COLUMNS, start=1):
            ws.cell(row=r, column=c).alignment = TOP_WRAP if name in WRAP_COLS else TOP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(get_column_letter(len(COLUMNS)), ws.max_row)


def add_data_sheet(wb, title, records):
    ws = wb.create_sheet(title)
    ws.append(COLUMNS)
    for rec in records:
        ws.append(row_for(rec))
    style_data_sheet(ws)
    return len(records)


def main():
    sp = load("sp-crud.json") + load("sp-noncrud.json")
    te = load("te.json")
    combo = load("combo.json")
    combo_rep = [x for x in combo if x["test_id"].startswith("CB-REP")]
    combo_rnd = [x for x in combo if x["test_id"].startswith("CB-RND")]

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; Index created below

    # --- Index tab ---
    idx = wb.create_sheet("Index")
    idx["A1"] = "Custom Roles & Permissions — Manual Test Case Suite"
    idx["A1"].font = Font(bold=True, size=14)
    idx["A3"] = ("This suite covers Epic SV-7388 (Custom Roles and Permissions).")
    idx["A4"] = ("Status: ALL cases are UNVERIFIED - VIU pending until verified "
                 "against staging.")
    idx["A4"].font = Font(italic=True)

    counts = {
        "Single Permission": len(sp),
        "Template Edit": len(te),
        "Combination - Representative": len(combo_rep),
        "Combination - Random": len(combo_rnd),
    }
    covers = {
        "Single Permission": "One permission toggled at a time (CRUD + non-CRUD).",
        "Template Edit": "Editing/cloning permission templates and propagation.",
        "Combination - Representative": "Hand-picked representative permission combinations.",
        "Combination - Random": "Randomly sampled permission combinations.",
    }

    r = 6
    idx.cell(row=r, column=1, value="Tab").font = HEADER_FONT
    idx.cell(row=r, column=2, value="Covers").font = HEADER_FONT
    idx.cell(row=r, column=3, value="Row Count").font = HEADER_FONT
    for c in range(1, 4):
        idx.cell(row=r, column=c).fill = HEADER_FILL
    r += 1
    for tab in ["Single Permission", "Template Edit",
                "Combination - Representative", "Combination - Random"]:
        idx.cell(row=r, column=1, value=tab)
        idx.cell(row=r, column=2, value=covers[tab])
        idx.cell(row=r, column=3, value=counts[tab])
        r += 1
    idx.cell(row=r, column=1, value="TOTAL")
    idx.cell(row=r, column=1).font = HEADER_FONT
    idx.cell(row=r, column=3, value=sum(counts.values()))
    idx.cell(row=r, column=3).font = HEADER_FONT
    r += 2

    idx.cell(row=r, column=1, value="Dependency Mode legend").font = Font(bold=True)
    r += 1
    dep_legend = [
        "None - permission acts independently; no dependency behaviour expected.",
        "Cascade: auto-enable lower - enabling this permission auto-enables lower/child permissions.",
        "Cascade: auto-disable higher (reverse) - disabling this permission auto-disables higher/dependent permissions.",
        "Parent gate: hide children - child permissions are hidden/unavailable until the parent is enabled.",
        "Financial gate: confirm modal - a confirmation modal guards financially sensitive actions.",
        "AP/AR gate: confirm modal - a confirmation modal guards AP/AR-sensitive actions.",
        "Multiple - the case exercises more than one dependency mode at once.",
    ]
    for line in dep_legend:
        idx.cell(row=r, column=1, value=line)
        r += 1
    r += 1

    idx.cell(row=r, column=1, value="Priority legend").font = Font(bold=True)
    r += 1
    for line in ["Critical - blocks core security/permission enforcement.",
                 "High - important behaviour, likely user-facing impact.",
                 "Medium - meaningful but lower-risk behaviour.",
                 "Low - edge/cosmetic behaviour."]:
        idx.cell(row=r, column=1, value=line)
        r += 1
    r += 1

    idx.cell(row=r, column=1, value="Type legend").font = Font(bold=True)
    r += 1
    for line in ["Positive - verifies allowed behaviour works.",
                 "Negative - verifies disallowed behaviour is blocked.",
                 "Security - verifies access control / privilege enforcement.",
                 "Dependency - verifies permission dependency/cascade behaviour.",
                 "Combination - verifies interaction of multiple permissions."]:
        idx.cell(row=r, column=1, value=line)
        r += 1

    idx.column_dimensions["A"].width = 48
    idx.column_dimensions["B"].width = 55
    idx.column_dimensions["C"].width = 12
    for row in idx.iter_rows():
        for cell in row:
            if cell.alignment.vertical != "top":
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    # --- Data tabs ---
    n_sp = add_data_sheet(wb, "Single Permission", sp)
    n_te = add_data_sheet(wb, "Template Edit", te)
    n_rep = add_data_sheet(wb, "Combination - Representative", combo_rep)
    n_rnd = add_data_sheet(wb, "Combination - Random", combo_rnd)

    wb.save(OUT)
    print("Wrote", OUT)
    return {"Single Permission": n_sp, "Template Edit": n_te,
            "Combination - Representative": n_rep,
            "Combination - Random": n_rnd}


if __name__ == "__main__":
    main()
