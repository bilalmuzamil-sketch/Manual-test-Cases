#!/usr/bin/env python3
"""TASK B — Custom Roles developer bug drafts (2026-07-13).

Builds CustomRoles_Bug-Drafts_2026-07-13.md (+ .xlsx) — the genuine build
Deviations from the behavioural VIU pass, written for DEVELOPERS in very simple
layman language (Standing Rule 7): reader-facing Title + 'What happens now' +
'What should happen' + numbered Steps to see it + Expected vs Actual, with NO
case IDs / permission-enum / HTTP jargon in the reader-facing part. All that
technical detail lives in a separate 'Technical notes (QA internal)' block per
bug + a QA-internal mapping tab (bug -> case C-IDs + links + permission detail).

The QuickBooks/Integrations items (C26529/C26530/C26531) are EXCLUDED — re-check
of the state doc shows they are corrected stale CASE WORDING (build keeps
QuickBooks under Integrations, matching the current spec), NOT build defects.
The two RUN331 fixes now verified (C26475, C26482) are listed in a short 'Now
Fixed' awareness section (not bugs). New file only.
"""
import os, csv
BASE = os.path.dirname(os.path.abspath(__file__))
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

# Each bug: reader-facing (layman, no jargon) + tech notes (QA internal) + case map.
BUGS = [
 {
  "id": "CR-BUG-1",
  "title": "New Work Order screen shows 'Add' buttons for creating a customer and an asset to users who are not allowed to add customers",
  "now": ("A user is given a role that lets them create work orders and view customers, but NOT "
          "create or edit customers. When they open the New Work Order screen they still see and "
          "can use the 'Add' button to create a brand-new customer, and — after picking a customer "
          "— the 'Add' button to create a brand-new asset."),
  "should": ("Because this user is not allowed to add or edit customers, both 'Add' buttons (new "
             "customer and new asset) should be hidden or disabled on the New Work Order screen."),
  "steps": [
    "Create a role with 'Work Orders: Create & Edit' ON and 'Customers: View' ON, but 'Customers: Create & Edit' OFF.",
    "Assign this role to a test user and log in as that user.",
    "Start a New Work Order.",
    "Look at the 'Add' button next to the Customer field.",
    "Pick a customer, then look at the 'Add' button next to the Asset field.",
  ],
  "expected": "The 'Add' button next to Customer and the 'Add' button next to Asset are hidden or disabled.",
  "actual": "Both 'Add' buttons are shown and clickable — the user can create a new customer and a new asset even though they lack the permission.",
  "tech": ("RUN331 FAIL persists. Test role = workOrdersCreateAndEdit + customersView, "
           "customersCreateAndEdit OFF. The Customers Create&Edit gate is not applied to the "
           "'Add' affordances in the New Work Order modal. Asset 'Add' is only disabled before a "
           "customer is selected (customer-required dependency), then becomes enabled. "
           "Screenshots: newwo-modal-addbuttons / newwo-custselected-addasset."),
  "cases": [26387, 26388],
  "sev": "Medium (permission bypass in UI)",
 },
 {
  "id": "CR-BUG-2",
  "title": "Labor rate and pricing stay visible in the simplified technician view when the money permission is on",
  "now": ("The technician (simplified) view is meant to keep money off the screen. But when a "
          "user's role has the money permission ('See Financial Data') turned on, the work order "
          "lines screen still shows the labor Rate, Margin and Total columns and the labor dollar "
          "amount — even in the technician view."),
  "should": ("Please confirm the intended behaviour: the test case expected the labor rate to stay "
             "hidden in the technician view even when the money permission is on. If that is "
             "correct, the labor rate columns and amount should be hidden in the technician view."),
  "steps": [
    "Create a role set to the technician (simplified) view that ALSO has the money permission ('See Financial Data') ON.",
    "Assign it to a test user and log in.",
    "Open a work order and view its lines.",
    "Look for the Rate, Margin and Total columns and the labor dollar amount.",
  ],
  "expected": "In the technician view the labor rate columns and the labor dollar amount are hidden.",
  "actual": "The Rate, Margin and Total columns and the labor amount ($150) are all shown.",
  "tech": ("Labor-rate visibility follows See Financial Data (SFD), not the view mode. The overall "
           "money-by-SFD principle otherwise holds (plain Technician with SFD off shows no $ or "
           "prices anywhere). CAVEAT for dev: the test role also had workOrdersCreateAndEdit — dev "
           "to confirm that is not what surfaced the columns. Screenshot: techview-sfd-lines."),
  "cases": [26459, 26464],
  "sev": "Low/Needs-decision (view-mode vs SFD precedence)",
 },
 {
  "id": "CR-BUG-3",
  "title": "Turning on the invoice delete permission gives no confirmation prompt about the permission it depends on",
  "now": ("When editing a role, ticking the 'Invoicing & payments: Delete / Reverse' permission "
          "while 'View and Manage AP/AR Data' is OFF simply turns Delete on. No prompt appears."),
  "should": ("The test case expected a prompt to appear asking to also turn on the permission that "
             "Delete depends on, so the tester is not left with an inconsistent set of permissions. "
             "Please confirm which dependency should be enforced here."),
  "steps": [
    "Edit a role.",
    "Make sure 'View and Manage AP/AR Data' is OFF.",
    "Tick 'Invoicing & payments: Delete / Reverse'.",
    "Watch for a confirmation prompt.",
  ],
  "expected": "A confirmation prompt appears offering to also turn on the dependent permission.",
  "actual": "No prompt appears — Invoicing Delete just turns on and the other permission stays off.",
  "tech": ("The build actually links Invoicing to 'See Financial Data' (there IS an SFD-direction "
           "dialog), not to View and Manage AP/AR Data. The case's expected AP/AR prompt is "
           "stale / not implemented. Dev/PO to confirm the intended dependency (SFD vs AP/AR)."),
  "cases": [26424],
  "sev": "Low/Needs-decision (dependency-prompt design)",
 },
 {
  "id": "CR-BUG-4",
  "title": "Two roles can be created with the same name",
  "now": ("You can create a new role using a name that already exists. The system only warns you "
          "if the new role has the exact same set of permissions as an existing one, and even then "
          "it lets you continue with a 'Create Anyway' button. It does not stop you from reusing a "
          "name."),
  "should": ("Role names should be unique — the system should prevent creating a second role with a "
             "name that is already in use."),
  "steps": [
    "Note the name of an existing role.",
    "Start creating a new custom role and type that same name.",
    "Give it a different set of permissions.",
    "Save.",
  ],
  "expected": "The system blocks the save and tells you the name is already in use.",
  "actual": "The role saves with a duplicate name. (A warning only appears if the permissions are identical, and it can be overridden with 'Create Anyway'.)",
  "tech": ("The SimilarRoleWarningModal keys on IDENTICAL PERMISSIONS ('identical permissions "
           "already exists') + 'Create Anyway', not on the name. Name uniqueness is not enforced."),
  "cases": [26339],
  "sev": "Low (data hygiene / duplicate role names)",
 },
 {
  "id": "CR-BUG-5",
  "title": "The role template picker shows the same names and long descriptions as the roles list",
  "now": ("When creating a role you first pick a starting template. The template picker shows the "
          "same role names (Admin, Foreman, Office User, ...) and the same descriptions as the main "
          "Roles list."),
  "should": ("The test cases expected the template picker to use shorter names and/or different "
             "(shorter) descriptions than the Roles list. Please confirm whether distinct template "
             "labels were ever intended."),
  "steps": [
    "Start creating a new custom role.",
    "Look at the names and descriptions in the template picker.",
    "Compare them to the names and descriptions in the main Roles list.",
  ],
  "expected": "The template picker uses shorter names and/or different descriptions than the Roles list.",
  "actual": "The template picker names and descriptions are identical to the Roles list (e.g. Admin 'Full system access', Foreman 'Oversees technicians and work orders').",
  "tech": ("Premise likely stale — may be a spec/case expectation mismatch rather than a code "
           "defect. Dev/PO to confirm whether distinct/shorter template labels were ever intended."),
  "cases": [26340, 26341],
  "sev": "Low/Needs-decision (spec vs build wording)",
 },
]

# Not bugs — corrected stale case wording (excluded from the bug list, noted for the record).
NOT_BUGS = [
 {"cases": [26529, 26530, 26531],
  "why": ("QuickBooks / Integrations. Re-check of the state doc: the build KEEPS QuickBooks under "
          "the Integrations settings section (Integrations gates IBS / Open API / QuickBooks; "
          "Finance gates only Payment Methods / Taxes). The old case premise (QuickBooks moving to "
          "Finance and Integrations being removed) was STALE and has been corrected in the case "
          "wording to match the current build and the 09-Jul spec. This is corrected case wording, "
          "NOT a build defect — no dev ticket needed.")},
]

# Now-fixed (RUN331 fails re-verified this pass) — awareness only, not bugs.
NOW_FIXED = [
 {"case": 26475,
  "note": ("Turning 'See Financial Data' OFF now shows the 'Disable See Financial Data?' "
           "confirmation prompt ('Disabling See Financial Data will also disable Invoicing & "
           "Payments. Continue?' [Cancel | Disable]). RUN331 fail now fixed.")},
 {"case": 26482,
  "note": ("AP/AR aging reports now follow the Reports permission: with Reports ON but 'View and "
           "Manage AP/AR Data' OFF, all 6 AP/AR aging reports are still listed on the Reports "
           "page (A/R Aging Summary / Detail / Collection, A/P Aging Summary / Detail, A/P Unpaid "
           "Invoices). RUN331 fail now fixed.")},
]

# ---- Markdown (reader-facing + QA-internal per bug) ----
md = []
md.append('# Custom Roles — Developer Bug Drafts — 2026-07-13\n')
md.append('The genuine build **Deviations** found during the 2026-07-13 behavioural VIU pass, '
          'written for developers. The reader-facing part of each bug uses very simple, '
          'non-technical language (no case IDs, no permission codes, no HTTP terms). The '
          'per-bug **Technical notes (QA internal)** block and the mapping table hold the case '
          'IDs, links and permission detail.\n')
md.append(f'## {len(BUGS)} bug drafts (covering {sum(len(b["cases"]) for b in BUGS)} test cases)\n')
md.append('| # | Bug | Severity |')
md.append('|---|---|---|')
for b in BUGS:
    md.append(f'| {b["id"]} | {b["title"]} | {b["sev"]} |')
md.append('')
md.append('---\n')
for b in BUGS:
    md.append(f'## {b["id"]} — {b["title"]}\n')
    md.append(f'**What happens now:** {b["now"]}\n')
    md.append(f'**What should happen:** {b["should"]}\n')
    md.append('**Steps to see it:**')
    for i, s in enumerate(b['steps'], 1):
        md.append(f'{i}. {s}')
    md.append('')
    md.append(f'**Expected:** {b["expected"]}\n')
    md.append(f'**Actual:** {b["actual"]}\n')
    md.append(f'> **Technical notes (QA internal):** {b["tech"]} '
              + 'Affected cases: '
              + ', '.join(f'[C{c}]({LINK.format(c)})' for c in b['cases']) + '.\n')
    md.append('---\n')

md.append('## Not bugs — corrected stale case wording (excluded)\n')
for nb in NOT_BUGS:
    md.append('- ' + ', '.join(f'[C{c}]({LINK.format(c)})' for c in nb['cases']) + ': ' + nb['why'] + '\n')

md.append('## Now Fixed (RUN331 fails re-verified this pass — awareness only, not bugs)\n')
for nf in NOW_FIXED:
    md.append(f'- [C{nf["case"]}]({LINK.format(nf["case"])}): {nf["note"]}\n')

md.append('## Mapping (QA internal): bug -> affected cases + permission / enforcement detail\n')
md.append('| Bug | Case | Link | Permission / enforcement detail |')
md.append('|---|---|---|---|')
for b in BUGS:
    for c in b['cases']:
        md.append(f'| {b["id"]} | C{c} | [link]({LINK.format(c)}) | {b["tech"]} |')
with open(os.path.join(BASE, 'CustomRoles_Bug-Drafts_2026-07-13.md'), 'w') as fh:
    fh.write('\n'.join(md))

# ---- CSV (reader-facing) ----
with open(os.path.join(BASE, 'CustomRoles_Bug-Drafts_2026-07-13.csv'), 'w', newline='') as fh:
    w = csv.writer(fh)
    w.writerow(['Bug', 'Title', 'Severity', 'What happens now', 'What should happen',
                'Steps to see it', 'Expected', 'Actual', 'Affected cases'])
    for b in BUGS:
        w.writerow([b['id'], b['title'], b['sev'], b['now'], b['should'],
                    ' | '.join(f'{i}. {s}' for i, s in enumerate(b['steps'], 1)),
                    b['expected'], b['actual'],
                    ', '.join(f'C{c}' for c in b['cases'])])

# ---- XLSX ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
wb = openpyxl.Workbook()
hdr_fill = PatternFill('solid', fgColor='C00000')
hdr_font = Font(bold=True, color='FFFFFF')
wrap = Alignment(vertical='top', wrap_text=True)

# Bug Drafts tab (reader-facing)
ws = wb.active; ws.title = 'Bug Drafts'
cols = ['Bug', 'Title', 'Severity', 'What happens now', 'What should happen',
        'Steps to see it', 'Expected', 'Actual']
ws.append(cols)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
for b in BUGS:
    ws.append([b['id'], b['title'], b['sev'], b['now'], b['should'],
               '\n'.join(f'{i}. {s}' for i, s in enumerate(b['steps'], 1)),
               b['expected'], b['actual']])
    for col in range(1, len(cols) + 1):
        ws.cell(ws.max_row, col).alignment = wrap
for col, wd in zip('ABCDEFGH', [11, 44, 26, 60, 55, 55, 45, 55]):
    ws.column_dimensions[col].width = wd
ws.freeze_panes = 'A2'

# Mapping tab (QA internal)
ws = wb.create_sheet('Mapping (QA internal)')
mcols = ['Bug', 'Case ID', 'TestRail Link', 'Permission / enforcement detail']
ws.append(mcols)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
for b in BUGS:
    for c in b['cases']:
        ws.append([b['id'], f'C{c}', LINK.format(c), b['tech']])
        lc = ws.cell(ws.max_row, 3)
        lc.hyperlink = LINK.format(c); lc.font = Font(color='0563C1', underline='single')
        for col in range(1, len(mcols) + 1):
            ws.cell(ws.max_row, col).alignment = wrap
for col, wd in zip('ABCD', [11, 10, 46, 90]):
    ws.column_dimensions[col].width = wd
ws.freeze_panes = 'A2'

# Not bugs tab
ws = wb.create_sheet('Not bugs (excluded)')
ws.append(['Cases', 'Why excluded (corrected stale case wording, not a build defect)'])
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
for nb in NOT_BUGS:
    ws.append([', '.join(f'C{c}' for c in nb['cases']), nb['why']])
    for col in (1, 2):
        ws.cell(ws.max_row, col).alignment = wrap
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 110
ws.freeze_panes = 'A2'

# Now fixed tab
ws = wb.create_sheet('Now Fixed (awareness)')
fcols = ['Case ID', 'TestRail Link', 'Now-fixed note (not a bug)']
ws.append(fcols)
for c in ws[1]:
    c.fill = hdr_fill; c.font = hdr_font
for nf in NOW_FIXED:
    ws.append([f'C{nf["case"]}', LINK.format(nf['case']), nf['note']])
    lc = ws.cell(ws.max_row, 2)
    lc.hyperlink = LINK.format(nf['case']); lc.font = Font(color='0563C1', underline='single')
    for col in range(1, len(fcols) + 1):
        ws.cell(ws.max_row, col).alignment = wrap
for col, wd in zip('ABC', [11, 46, 100]):
    ws.column_dimensions[col].width = wd
ws.freeze_panes = 'A2'

wb.save(os.path.join(BASE, 'CustomRoles_Bug-Drafts_2026-07-13.xlsx'))

print('bug drafts:', len(BUGS), 'covering cases:', sum(len(b['cases']) for b in BUGS))
print('excluded (not bugs):', sum(len(nb['cases']) for nb in NOT_BUGS))
print('now fixed:', len(NOW_FIXED))
