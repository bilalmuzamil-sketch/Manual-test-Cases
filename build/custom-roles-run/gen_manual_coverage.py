#!/usr/bin/env python3
"""TASK A — Custom Roles manual-coverage hand-off sheet (2026-07-13).

Builds CustomRoles_Manual-Coverage_2026-07-13.xlsx (+ .md) listing every
Blocked-UI case (the genuine harness / environment residue from the
behavioural VIU pass) so a manual tester / QA lead can finish them off.

Each row: Category, Case ID (C#####), clickable TestRail Link (Rule 8),
Title, Section/Folder, the precise reason it can't be automated, and a plain
layman 'How to cover' for a manual tester. Summary tab has category counts.

Reads titles/sections/reasons from cases-2026-07-13/*.json. New file only —
does NOT edit any case JSON or the state doc.
"""
import json, os, csv
BASE = os.path.dirname(os.path.abspath(__file__))
CASEDIR = os.path.join(BASE, 'cases-2026-07-13')
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

# Section id -> tester-facing folder name (from the wording-VIU workbook AREAS map)
SECTIONS = {
    3532: 'Permission Summary', 3534: 'Work Orders Permissions',
    3535: 'Work Order Lines Permissions', 3536: 'Schedule Permissions',
    3537: 'Customer Management Permissions', 3538: 'Parts Department Permissions',
    3539: 'Invoicing and Payments Permissions', 3540: 'Timesheets Permissions',
    3541: 'Page Access Toggles', 3542: 'Settings Access', 3543: 'View Mode',
    3545: 'View and Manage AP/AR Data', 3547: 'Staff Page Role Assignment',
    3549: 'Migration', 3550: 'Staff Record Settings',
    3552: 'User Feedback Strings', 3553: 'Cross-Permission Combinations',
}

# Category -> ordered list of case ids. Root-cause grouping from the Blockers Tracker.
CATEGORIES = [
    ("Staff editor / staff record (needs a real browser or a 2nd real user account)",
     [26356, 26450, 26490, 26491, 26493, 26526, 26527, 26539, 27873]),
    ("Calendar drag / slot interaction (needs a real browser)",
     [26395, 26396, 27867]),
    ("In-page payment / terminal / return / financial / timesheet-entry editors (needs a real browser)",
     [26401, 26422, 26423, 26427, 27871, 29434, 29438, 26479, 26431]),
    ("Portal / Send-to-Portal surfaces (not exposed in this test environment)",
     [26437, 26438, 26439, 26440, 26466]),
    ("Parts delete / restock detail-page affordance (not reachable in the harness)",
     [26412, 26415, 26418, 26419]),
    ("Seeded work-order line-state operations (review / pick / core / set-line-status / line-delete / qty)",
     [26379, 26380, 26391, 27866, 27870, 29435]),
    ("Tech-view parts-request form field count (needs a real browser)",
     [26460]),
    ("Last-Administrator guard (shared org has 89 admins — cannot create the last-admin state)",
     [26550]),
    ("Migration (needs a pre-migration org with a legacy Owner user)",
     [27731]),
]

# Plain, layman 'How to cover' steps for a manual tester (one per case).
HOW_TO_COVER = {
 26356: "Open the Staff page, click a real staff member to open the Edit Staff Member window, click the eye icon next to the Role dropdown, and confirm a read-only Permission Summary opens.",
 26450: "Give a test user a role with View/Manage Wages turned ON, open that user's Edit Staff Member profile, and confirm the wage/pay fields are shown and editable. Repeat with the toggle OFF and confirm they are hidden.",
 26490: "Open Edit Staff Member for any staff member, open the Role dropdown, and confirm the roles are grouped under two headings: 'System Roles' and 'Custom Roles'.",
 26491: "On the Edit Staff Member window, click the 'View Permissions' (eye) button next to the role selector and confirm the Permission Summary window opens.",
 26493: "Needs a real browser plus a way to make the save fail (for example, disconnect the network mid-save). Change a staff member's role, force the save to fail, and confirm the user still shows their previous role.",
 26526: "Needs two test technicians set to different departments. Confirm whether each one can be scheduled based on their department setting, not their role.",
 26527: "Needs two test users with different per-staff Time Clock settings. On a work order line, confirm the clock-in option follows the Time Clock setting, not the role.",
 26539: "On the Staff page, reassign a staff member's role and confirm the row updates immediately with NO success-message popup.",
 27873: "Needs two real user accounts. As user A, add a note to a customer. Log in as user B (a role that lacks the note permission) and confirm the edit/delete options on user A's note are hidden.",
 26395: "Needs a real browser. Give a test user Schedule Create & Edit, open the Schedule calendar, create a new appointment, drag it to a different slot, and edit it — confirm all three work.",
 26396: "Needs a real browser. With Schedule Delete, select an existing appointment on the calendar and delete it — confirm it is removed.",
 27867: "Needs a real browser. With a Schedule Create & Edit only role, open 'Assign existing work order' on the calendar and confirm the unscheduled work order list loads.",
 26401: "Needs a customer with a payment on file. Give a test user Customers Delete but NOT Invoicing & payments Delete, and confirm they cannot delete the customer's payment.",
 26422: "Give a test user Invoicing & payments Delete. On a work order with a payment, confirm they can delete the payment and void the transaction (but cannot reverse the invoice).",
 26423: "Confirm that deleting a customer payment works only with Invoicing & payments Delete and NOT with Customers Delete (compare a role that has one against a role that has the other).",
 26427: "Give a test user Invoicing & payments Create & Edit. Open an invoice's take-payment flow and confirm the 'Send to Terminal' action is available.",
 27871: "First create a part return. Then confirm that deleting or cancelling the return is only allowed with Invoicing & payments Delete.",
 29434: "Confirm 'Send to Terminal' appears only when the user has BOTH Invoicing & payments Create & Edit AND Customer portal turned ON (test with each one missing).",
 29438: "On an open work order, confirm the edit control on the Financial Info card appears only for a user with Invoicing & payments Create & Edit, and not for a work-order-edit-only role.",
 26479: "Give a test user View and Manage AP/AR Data ON. On the Unpaid Invoices tab, confirm they can pay several invoices at once.",
 26431: "Give a test user Timesheets Create & Edit. Open a timesheet entry and confirm they can edit it. Compare with a view-only role that cannot edit.",
 26437: "Needs an environment with the Customer Portal feature enabled. With Customer portal ON, confirm the user can manage the customer portal.",
 26438: "Needs the Customer Portal surface enabled. With Customer portal OFF, confirm the customer portal item is hidden from navigation.",
 26439: "Needs the Billing Portal feature enabled. With Billing Portal ON, confirm the user can manage the billing portal.",
 26440: "Needs the Billing Portal surface enabled. With Billing Portal OFF, confirm the Billing Portal item is hidden in the Settings area.",
 26466: "In Full View, with a user who can approve lines, confirm the 'Send to Portal' button is shown (check the work order header action menu if it is not on the page body).",
 26412: "Give a test user Part sales Delete. On a part sale, confirm they can delete it and reverse the part-sales invoice. Compare with a role that lacks the permission.",
 26415: "Give a test user Catalog and Inventory Delete. Open a catalog part detail and confirm they can delete it. Compare with a role that lacks the permission.",
 26418: "Give a test user Vendor and order management Delete. Confirm they can delete a vendor and a purchase order and reverse a vendor transaction.",
 26419: "On a work order or purchase order that has a picked part, confirm the return-to-inventory (restock) action is controlled by Vendor and order management.",
 26379: "First put a work order line into the submitted/awaiting-review state. Then, with the 'Review work orders' sub-toggle ON, confirm the Review action appears; with it OFF, confirm it is hidden.",
 26380: "On a work order that has a pending part request to pick, give a test user Work orders View + Pick parts (no Create & Edit) and confirm they can pick the parts.",
 26391: "Give a test user Work order lines Delete. On a work order, confirm they can remove a line.",
 27866: "Seed a work order with lines assigned to a DIFFERENT technician. With a default Technician role (Work order lines Create & Edit), confirm they can bulk-complete those lines via 'Set Line Status'.",
 27870: "Seed a work order line with a cored part (use New Part Request with a cored catalog part number). With Work order lines Create & Edit, confirm the user can mark the core OK / Not-OK and see the line history.",
 29435: "On the Part Requests tab, with a Pick/Order Parts role, confirm the Quantity value can be edited inline.",
 26460: "Open the New Part Request form once in tech view and once in full view, and confirm the tech view shows fewer fields.",
 26550: "Needs an isolated org (not shared staging). Try to remove or reassign the last remaining Administrator and confirm the system blocks leaving zero Administrators. Route to the product team to confirm the rule exists.",
 27731: "Needs a pre-migration org that has a legacy 'Owner' user. After migration, confirm the Owner user becomes an 'Administrator' (Owner merged into Admin).",
}


def load(cid):
    with open(os.path.join(CASEDIR, f'C{cid}.json')) as fh:
        return json.load(fh)


def clean_reason(v):
    v = (v or '').strip()
    # strip a leading "Blocked-UI:" / "Blocked-UI (...)" label for readability
    for pre in ("Blocked-UI:", "Blocked-UI"):
        if v.startswith(pre):
            v = v[len(pre):].lstrip(': ').strip()
            break
    return v


rows = []
for cat, ids in CATEGORIES:
    for cid in ids:
        d = load(cid)
        sec = int(d['section_id'])
        rows.append({
            'category': cat,
            'case_id': cid,
            'link': LINK.format(cid),
            'title': d['title'],
            'section': f"{SECTIONS.get(sec, str(sec))} ({sec})",
            'reason': clean_reason(d.get('viu_status', '')),
            'how': HOW_TO_COVER.get(cid, ''),
        })

# ---- CSV ----
with open(os.path.join(BASE, 'CustomRoles_Manual-Coverage_2026-07-13.csv'), 'w', newline='') as fh:
    w = csv.DictWriter(fh, fieldnames=['category', 'case_id', 'link', 'title', 'section', 'reason', 'how'])
    w.writeheader()
    for r in rows:
        w.writerow(r)

# ---- XLSX ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
wb = openpyxl.Workbook()
hdr_fill = PatternFill('solid', fgColor='305496')
hdr_font = Font(bold=True, color='FFFFFF')
wrap = Alignment(vertical='top', wrap_text=True)
cols = ['Category', 'Case ID', 'TestRail Link', 'Title', 'Section / Folder',
        'Why it could not be automated', 'How to cover (manual tester)']


def write_rows(ws, data):
    ws.append(cols)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
    for r in data:
        ws.append([r['category'], f"C{r['case_id']}", r['link'], r['title'],
                   r['section'], r['reason'], r['how']])
        lc = ws.cell(ws.max_row, 3)
        lc.hyperlink = r['link']; lc.font = Font(color='0563C1', underline='single')
        for col in range(1, len(cols) + 1):
            ws.cell(ws.max_row, col).alignment = wrap
    widths = [40, 9, 46, 50, 34, 60, 62]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = wd
    ws.freeze_panes = 'A2'


# Summary tab
ws = wb.active; ws.title = 'Summary'
ws.append(['Custom Roles — Manual-Coverage Hand-off — 2026-07-13']); ws['A1'].font = Font(bold=True, size=14)
ws.append([])
ws.append(['Definitive hand-off list: the Blocked-UI cases from the 2026-07-13 behavioural VIU pass that could NOT be driven headless.'])
ws.append(['Each is genuine harness / environment residue — a manual tester or a second real user account is needed to finish it.'])
ws.append(['Wording of every case is already build-accurate and pushed to TestRail; only the BEHAVIOUR still needs a live pass.'])
ws.append([])
ws.append(['Category', 'Cases'])
for c in ws[ws.max_row]:
    c.fill = hdr_fill; c.font = hdr_font
for cat, ids in CATEGORIES:
    ws.append([cat, len(ids)])
ws.append(['TOTAL', len(rows)])
ws[ws.max_row][0].font = Font(bold=True); ws[ws.max_row][1].font = Font(bold=True)
ws.column_dimensions['A'].width = 78
ws.column_dimensions['B'].width = 10
for row in ws.iter_rows(min_row=7):
    row[0].alignment = wrap

# All cases tab
write_rows(wb.create_sheet('Manual Coverage'), rows)

# One tab per category (short handles)
HANDLES = {
    CATEGORIES[0][0]: 'Staff editor',
    CATEGORIES[1][0]: 'Calendar',
    CATEGORIES[2][0]: 'Payment editors',
    CATEGORIES[3][0]: 'Portals',
    CATEGORIES[4][0]: 'Parts delete',
    CATEGORIES[5][0]: 'Line-state ops',
    CATEGORIES[6][0]: 'Tech-view form',
    CATEGORIES[7][0]: 'Last-admin guard',
    CATEGORIES[8][0]: 'Migration',
}
for cat, ids in CATEGORIES:
    data = [r for r in rows if r['category'] == cat]
    write_rows(wb.create_sheet(HANDLES[cat][:31]), data)

wb.save(os.path.join(BASE, 'CustomRoles_Manual-Coverage_2026-07-13.xlsx'))

# ---- Markdown ----
md = []
md.append('# Custom Roles — Manual-Coverage Hand-off — 2026-07-13\n')
md.append('Definitive hand-off list: the **Blocked-UI** cases from the 2026-07-13 behavioural '
          'VIU pass that could NOT be driven headless. Each is genuine harness / environment '
          'residue — a manual tester (or a second real user account) is needed to finish it. '
          'Every case\'s **wording** is already build-accurate and pushed to TestRail; only the '
          '**behaviour** still needs a live pass.\n')
md.append('**Rule 8:** every row carries the TestRail Case ID (C#####) + a clickable link.\n')
md.append(f'## Total: {len(rows)} cases across {len(CATEGORIES)} categories\n')
md.append('| Category | Cases |')
md.append('|---|---:|')
for cat, ids in CATEGORIES:
    md.append(f'| {cat} | {len(ids)} |')
md.append(f'| **TOTAL** | **{len(rows)}** |')
md.append('')
for cat, ids in CATEGORIES:
    md.append(f'## {cat} — {len(ids)}\n')
    for cid in ids:
        r = next(x for x in rows if x['case_id'] == cid)
        md.append(f'### C{cid} — {r["title"]}')
        md.append(f'- **TestRail:** [C{cid}]({r["link"]}) · **Section:** {r["section"]}')
        md.append(f'- **Why it could not be automated:** {r["reason"]}')
        md.append(f'- **How to cover:** {r["how"]}')
        md.append('')
with open(os.path.join(BASE, 'CustomRoles_Manual-Coverage_2026-07-13.md'), 'w') as fh:
    fh.write('\n'.join(md))

print('manual-coverage rows:', len(rows))
print('categories:', [(HANDLES[c], len(i)) for c, i in CATEGORIES])
