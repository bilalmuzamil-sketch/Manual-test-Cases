#!/usr/bin/env python3
"""
Append a 'New-WO Create Dual LIVE' sheet to the LIVE-VERIFIED workbook.
Capabilities: Create Work Order ("New" button) + Create Customer from New-WO ("Add" next to
Customer) + Create Asset control present ("Add" next to Asset; enabled only after a customer is
picked = design gate, so presence is reported). Observed live per role on both envs (2026-07-15).
Observed-only (Rules 10 & 12) — every row DUAL LIVE-OBSERVED.
"""
import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = '/home/user/Manual-test-Cases/build/custom-roles-run'
STG = f'{BASE}/live-ui-2026-07-15/staging'
PROD = f'{BASE}/live-ui-2026-07-15/production'

def load_dir(base, fname):
    out = {}
    for d in os.listdir(base):
        f = f'{base}/{d}/{fname}'
        if os.path.exists(f):
            o = json.load(open(f)); out[o['role']] = o
    return out

stg = load_dir(STG, 'newwo-obs.json')
prod = load_dir(PROD, 'prod-newwo-obs.json')

MAP = {
 'Admin': ('Administrator', []),
 'Service_Manager': ('Service_Manager', []),
 'Senior_Service_Advisor': ('Service_Advisor', ['SA_Technician', 'SA_No_Reports']),
 'Parts_Manager': ('Parts_Manager', []),
 'Service_Advisor': ('SA_Limited_View', []),
 'Foreman': ('Foreman', []),
 'Office_User': ('Office_User', []),
 'Parts_Technician': ('Parts_Technician', []),
 'Sales_Representative': ('Sales_Representative', ['Reporting']),
 'Technician': ('Technician', []),
 'Time_Clock_User': ('Time_Clock_User', []),
}
ORDER = list(MAP.keys())
NICE = lambda s: s.replace('_', ' ')

CAPS = [
 ('Create Work Order ("New" button on Work Orders list)', 'newButton',
  'workOrdersCreateAndEdit — opens the New Work Order dialog.'),
 ('Create Customer from New-WO ("Add" next to Customer)', 'addCustomerPresent',
  'customersCreateAndEdit — the inline Add button in the New Work Order dialog.'),
 ('Create Asset control from New-WO ("Add" next to Asset)', 'addAssetPresent',
  'assetsCreateAndEdit — Add button present; enabled only after a customer is chosen (design gate), so presence is reported.'),
]

wb = load_workbook(f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx')
if 'New-WO Create Dual LIVE' in wb.sheetnames:
    del wb['New-WO Create Dual LIVE']
ws = wb.create_sheet('New-WO Create Dual LIVE')

H = Font(bold=True, color='FFFFFF'); HF = PatternFill('solid', fgColor='2F5496')
OK = PatternFill('solid', fgColor='C6EFCE'); YEL = PatternFill('solid', fgColor='FFF2CC')
ORG = PatternFill('solid', fgColor='FCE4D6'); NV = PatternFill('solid', fgColor='FFC7CE')
thin = Side(style='thin', color='BFBFBF'); BD = Border(thin, thin, thin, thin)
wrap = Alignment(wrap_text=True, vertical='top')

note = ws.cell(1, 1, 'NEW-WO CREATE flow — LIVE dual observation 2026-07-15. Both envs driven live: open Work Orders list, '
       'click "New", observe the New Work Order dialog (Add Customer / Add Asset). Staging via switch-user (7) + throwaway '
       'role-swap+switch-user (4, restored). Production via test-staff role-swap+self-login (14, restored to Office User). '
       'Every row DUAL LIVE-OBSERVED (Rules 10 & 12) — no inference.')
note.font = Font(bold=True, color='9C0006'); note.alignment = wrap
ws.merge_cells('A1:K1'); ws.row_dimensions[1].height = 55

hdr = ['Staging Role', 'Prod role compared', 'Capability', 'PROD observed', 'Staging observed',
       'Direction / verdict', 'Per-spec?', 'Confidence', 'Method', 'Prod screenshot', 'Staging screenshot / Notes']
ws.append(hdr)
for i, h in enumerate(hdr, 1):
    c = ws.cell(2, i); c.font = H; c.fill = HF; c.alignment = wrap; c.border = BD
for i, w in enumerate([20, 30, 42, 14, 15, 30, 10, 22, 30, 46, 50], 1):
    ws.column_dimensions[ws.cell(2, i).column_letter].width = w

def yn(v): return 'SHOWN' if v else 'hidden'

for role in ORDER:
    prole, merges = MAP[role]
    s = stg[role]; p = prod[prole]
    pshot = f'{PROD}/{prole}/new_wo_dialog.png'
    if not os.path.exists(pshot): pshot = f'{PROD}/{prole}/new_wo_nobutton.png'
    sshot = f'{STG}/{role}/new_wo_dialog.png'
    if not os.path.exists(sshot): sshot = f'{STG}/{role}/new_wo_nobutton.png'
    for label, key, capnote in CAPS:
        sv = bool(s.get(key)); pv = bool(p.get(key))
        merge_note = ''
        if merges:
            vals = {m: bool(prod[m].get(key)) for m in merges if m in prod}
            if all(v == pv for v in vals.values()):
                merge_note = f' (merge {", ".join(NICE(m) for m in vals)} also {yn(pv)} — consistent)'
            else:
                merge_note = ' (merge differ: ' + ', '.join(f'{NICE(m)}={yn(v)}' for m, v in vals.items()) + ')'
        verdict = 'MATCH' if pv == sv else ('STAGING-LESS (prod grants more)' if pv and not sv else 'STAGING-MORE (staging grants more)')
        ws.append([NICE(role), NICE(prole), label, yn(pv), yn(sv), verdict,
                   'yes' if verdict == 'MATCH' else 'see note', 'DUAL LIVE-OBSERVED',
                   'stg: switch-user/role-swap; prod: test-staff role-swap',
                   pshot if os.path.exists(pshot) else '', (sshot if os.path.exists(sshot) else '') + '  ' + capnote + merge_note])
        r = ws.max_row
        for i in range(1, len(hdr) + 1):
            ws.cell(r, i).alignment = wrap; ws.cell(r, i).border = BD
        ws.cell(r, 4).fill = OK if pv else YEL
        ws.cell(r, 5).fill = OK if sv else YEL
        ws.cell(r, 6).fill = OK if verdict == 'MATCH' else (NV if verdict.startswith('STAGING-LESS') else ORG)

wb.save(f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx')
print('WROTE New-WO Create Dual LIVE sheet.')
diffs = []
for role in ORDER:
    prole = MAP[role][0]
    for label, key, _ in CAPS:
        if bool(stg[role].get(key)) != bool(prod[prole].get(key)):
            diffs.append(f'{NICE(role)} / {label}: prod={yn(bool(prod[prole].get(key)))} stg={yn(bool(stg[role].get(key)))}')
print('NON-MATCH cells:', len(diffs))
for d in diffs: print('  -', d)
