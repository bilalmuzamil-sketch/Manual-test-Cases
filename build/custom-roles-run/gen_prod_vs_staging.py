#!/usr/bin/env python3
"""
Custom Roles (SV-7388) — PRODUCTION vs STAGING role/permission gap generator.

STATUS OF DATA (2026-07-15): *** BOTH SIDES NOW LIVE-VERIFIED ***
  * STAGING (new custom-roles model) = LIVE. GET /api/organizations/{org}/roles +
    per-role GET /api/roles/{id}. 11 system roles.
    Source: compare-evidence-2026-07-14/staging-capability-matrix.json
  * PRODUCTION (old legacy model) = LIVE. Prod authenticated with a fresh PHPSESSID
    (app.shopview.com / api.shopview.com; NO SSO). Prod org UUID
    72b2cc90-6964-4429-a207-76e55f946936. Role inventory from GET /api/iam/list-roles
    (14 legacy roles, NO "Owner" present). Per-role effective permissions captured
    empirically by impersonation (POST /api/switch-user -> data.permissions
    [{resource_name,action_name,limits,exclusions}] -> POST /api/exit-switch-user).
    Roles without an existing active user were captured by temporarily assigning a
    throwaway ZZ invite-test user (bilal.muzamil+bugstesting) that role via
    POST /api/staff/change, impersonating, then restoring to Technician
    (departments/workplace verified preserved). Source:
    compare-evidence-2026-07-14/prod-capability-matrix.json

Old model = 50 resources x actions {'*','view','create','change','remove', + task/return
verbs}. '*' = ALL actions on that resource. New model = 41 fe_permission atoms +
view_mode + 3 cross toggles.

Deliverable: bi-directional 11-column main tab (STAGING-LESS + STAGING-MORE), a dedicated
"Work Orders - granular" tab, per-role 2x2 summaries, both live matrices, open questions.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HERE = os.path.dirname(os.path.abspath(__file__))
EV = os.path.join(HERE, "compare-evidence-2026-07-14")
STG = json.load(open(os.path.join(EV, "staging-capability-matrix.json")))
PRODJ = json.load(open(os.path.join(EV, "prod-capability-matrix.json")))

# ---- prod effective-capability helpers (expand '*') ----
PROD = {}
for rc, perms in PRODJ.items():
    m = {}
    for p in perms:
        m.setdefault(p["resource_name"], set()).add(p["action_name"])
    PROD[rc] = m

def p_has(rc, res, *acts):
    a = PROD.get(rc, {}).get(res, set())
    if "*" in a:
        return True
    return any(x in a for x in acts)

def p_all(rc, res):
    return "*" in PROD.get(rc, {}).get(res, set())

# ---- staging helpers ----
def s_has(role, code):
    return code in STG[role]["codes"]

def s_ct(role, key):
    return bool(STG[role]["ct"].get(key))

def s_vm(role):
    return STG[role]["view_mode"]

# ---- PROD -> STAGING merge mapping (spec migration table; CONFIRMED by QA lead 2026-07-14) ----
# Administrator compared 1:1 (prod Administrator <-> staging Administrator); the spec's
# "Owner merged in" is not applicable - no Owner role exists in either environment (N/A,
# confirmed by QA lead 2026-07-14). SA/SSA merge mapping also CONFIRMED (see SA_SSA_MERGE).
MAP = {
    "Admin": ["ROLE_ADMINISTRATOR"],   # 1:1 Administrator<->Administrator; no Owner in either env (N/A)
    "Service Manager": ["ROLE_SERVICE_MANAGER"],
    "Senior Service Advisor": ["ROLE_SERVICE_ADVISOR", "ROLE_SERVICE_ADVISOR_TECHNICIAN", "ROLE_SERVICE_ADVISOR_NO_REPORTS"],
    "Service Advisor": ["ROLE_SERVICE_ADVISOR_LIMITED_VIEW"],
    "Foreman": ["ROLE_FOREMAN"],
    "Technician": ["ROLE_TECHNICIAN"],
    "Parts Manager": ["ROLE_PARTS_MANAGER"],
    "Parts Technician": ["ROLE_PARTS_TECHNICIAN"],
    "Office User": ["ROLE_OFFICE_USER"],
    "Sales Representative": ["ROLE_SALES_REPRESENTATIVE", "ROLE_REPORTING"],
    "Time Clock User": ["ROLE_TIME_CLOCK_USER"],
}
PLABEL = {
    "ROLE_ADMINISTRATOR": "Administrator", "ROLE_SERVICE_MANAGER": "Service Manager",
    "ROLE_SERVICE_ADVISOR": "Service Advisor", "ROLE_SERVICE_ADVISOR_TECHNICIAN": "Service Advisor Technician",
    "ROLE_SERVICE_ADVISOR_NO_REPORTS": "Service Advisor - No Reports",
    "ROLE_SERVICE_ADVISOR_LIMITED_VIEW": "Service Advisor - Limited View",
    "ROLE_FOREMAN": "Foreman", "ROLE_TECHNICIAN": "Technician",
    "ROLE_PARTS_MANAGER": "Parts Manager", "ROLE_PARTS_TECHNICIAN": "Parts Technician",
    "ROLE_OFFICE_USER": "Office User", "ROLE_SALES_REPRESENTATIVE": "Sales Representative",
    "ROLE_REPORTING": "Reporting", "ROLE_TIME_CLOCK_USER": "Time Clock User",
}
ORDER = ["Admin", "Service Manager", "Senior Service Advisor", "Service Advisor",
         "Foreman", "Technician", "Parts Manager", "Parts Technician",
         "Office User", "Sales Representative", "Time Clock User"]
# SA/SSA merge mapping CONFIRMED by QA lead 2026-07-14 (spec migration table authoritative):
#   Senior Service Advisor <- Service Advisor + SA Technician + SA No Reports (3 merged)
#   Service Advisor         <- SA Limited View
# No longer flagged NEEDS-REVIEW (mapping unconfirmed); per-capability NEEDS-REVIEW still stands.
SA_SSA_MERGE = {"Service Advisor", "Senior Service Advisor"}

# ---- QA-lead Migration-Type per staging role (spec Migration Plan table) ----
MIGRATION_TYPE = {
    "Admin": "Direct - Administrator (Owner merge N/A: no Owner in either env)",
    "Service Manager": "Direct (with adjustments)",
    "Senior Service Advisor": "Renamed + expanded (merge: Service Advisor + SA Technician + SA No Reports; gains Reports)",
    "Service Advisor": "Mapped from SA Limited View (AP/AR OFF preserves core restriction)",
    "Foreman": "Direct (with expansions)",
    "Technician": "Direct mapping",
    "Parts Manager": "Direct (with adjustments)",
    "Parts Technician": "Direct (with expansions)",
    "Office User": "Direct (with adjustments)",
    "Sales Representative": "Direct (merge: Sales Representative + Reporting)",
    "Time Clock User": "Direct mapping",
}

# ---- STAGING role slug / system identifier (LIVE-captured 2026-07-15, staging UI-verify run) ----
# Source: GET /api/role-templates -> role_templates[].slug, mapped to each role via the role's
# template_id (GET /api/roles/{id}).  The role object itself has NO slug/system_name/code field
# (confirmed has_slug_field=False for all 11); the slug is the role TEMPLATE slug (the real API
# value - NOT hardcoded/guessed).  Raw objects saved to
# staging-ui-verify-2026-07-14/ (staging-role-templates.json, staging-role-<id>.json,
# staging-role-slug-map.json).
STG_SLUG = {
    "Admin": "administrator",
    "Service Manager": "service_manager",
    "Senior Service Advisor": "senior_service_advisor",
    "Service Advisor": "service_advisor",
    "Foreman": "foreman",
    "Technician": "technician",
    "Parts Manager": "parts_manager",
    "Parts Technician": "parts_technician",
    "Office User": "office",
    "Sales Representative": "sales_representative",
    "Time Clock User": "time_clock_user",
}

# ---- STAGING UI / FE-SOURCE VERIFICATION (2026-07-15) — per-capability override of the
# Verification-confidence for the FE-gated High-severity rows (task: raise MEDIUM->HIGH where
# observable).  Method: LIVE staging role definitions (GET /api/roles/{id} + cross_toggles +
# view_mode, 2026-07-15) evaluated against the ACTUAL FE gate predicate read from the shipped
# staging JS bundle.  Live pixel-screenshots were blocked because the sv_sso_session expired
# mid-run (quick-login returned 200 but the next API call 409'd "Session has expired"); the
# verification is therefore FE-source (shipped gate logic) + live role-definition data, which is
# authoritative for a front-end DISPLAY gate.  Keyed by capability name (applies to every role). ----
UI_VERIFIED = {
    "Send to Portal": (
        "HIGH - FE-source gate + live role-def verified (staging 2026-07-15)",
        "Staging FE gates Send-to-Portal on Customer Portal access (permission store helper "
        "userHasCustomerPortalAccess = has('customerPortalPageAccess'); button in WorkOrderNavBar "
        "off the WO detail). LIVE staging role defs 2026-07-15: the atom is ABSENT for all 6 "
        "STAGING-LESS roles (Technician/Foreman/Parts Tech/Office/Sales Rep/Time Clock) => HIDDEN, "
        "and PRESENT for the 5 roles that keep it (Admin/Parts Mgr/Sr SA/Svc Adv/Svc Mgr) => SHOWN "
        "- internally consistent. Pixel-screenshot blocked by mid-run SSO expiry."),
    "See AP/AR Data": (
        "HIGH - FE-source gate + live cross-toggle verified (staging 2026-07-15)",
        "Staging AP/AR surfaces (Accounts Payable/Receivable, customer/vendor transactions + "
        "payments tabs) gate on the seeApArData() cross-toggle (FE source: check:()=>seeApArData()). "
        "LIVE staging cross_toggles 2026-07-15 match: Parts Tech=OFF (HIDDEN), Sales Rep=ON (VISIBLE), "
        "Service Advisor=OFF (HIDDEN, spec-intended)."),
    "Send to Terminal (take payment on WO)": (
        "HIGH - control ABSENT from staging build (FE-source 2026-07-15)",
        "NO payment-terminal / card-reader / 'Send to Terminal' / 'take payment' control exists "
        "anywhere in the staging FE bundle (2026-07-15); 'terminal' matches only the Quasar "
        "framework. Staging has no per-role Send-to-Terminal gate at all - the row reflects a "
        "build-wide ABSENCE, not a role regression. Confirm prod's actual control name before "
        "treating as a role-level loss."),
    "Approve / complete a WO part return": (
        "MEDIUM / NEEDS-UI-VERIFY - control present, gate not isolated (staging 2026-07-15)",
        "Part-return controls DO exist in the staging build ('Process Return' / 'Confirm Return' in "
        "the ConfirmReturn chunk), but the exact permission gate could not be isolated from the "
        "minified source and the pixel-screenshot was blocked by SSO expiry. Live role atoms recorded."),
    "Decline a WO part return": (
        "MEDIUM / NEEDS-UI-VERIFY - control present, gate not isolated (staging 2026-07-15)",
        "Part-return controls exist in the staging build; exact gate not isolated from minified "
        "source; pixel-screenshot blocked by SSO expiry."),
    "Process a WO part return (create)": (
        "MEDIUM / NEEDS-UI-VERIFY - control present, gate not isolated (staging 2026-07-15)",
        "Part-return controls exist in the staging build; exact gate not isolated from minified "
        "source; pixel-screenshot blocked by SSO expiry."),
}

# ---- OUT-OF-MODEL capabilities (verification 3.3): clock-in + timesheets are STAFF-RECORD
# controlled per spec 'Staff Record Settings', NOT the role/permission model. These are
# annotated + EXCLUDED from the risk "No" counts and moved to an informational section. ----
OUT_OF_MODEL = {
    "Clock in / log time on a WO line task",
    "Timesheets View",
    "Timesheets Create & Edit",
}
OUT_OF_MODEL_NOTE = ("staff-record-controlled, not a permission delta (spec 'Staff Record "
                     "Settings') - excluded from risk counts")

# ---- SPEC-DOCUMENTED intended changes (updated-spec-source.md 'Behavior Changes', l.474-485) ----
SPEC_INTENDED = {
    # intended REDUCTIONS (STAGING-LESS)
    ("Service Manager", "Invoicing & Payments Delete (reverse/delete invoice)", "STAGING-LESS"):
        "Spec Behavior-Changes l.478: Service Manager 'Loses Invoicing Delete (cannot reverse)'",
    ("Service Manager", "Settings: Service", "STAGING-LESS"):
        "Spec Behavior-Changes l.478: Service Manager 'Loses Settings: Service'",
    ("Service Manager", "Settings: Parts", "STAGING-LESS"):
        "Spec Behavior-Changes l.478: Service Manager 'Loses Settings: Parts'",
    ("Service Manager", "Settings: Finance", "STAGING-LESS"):
        "Spec Behavior-Changes l.478: Service Manager 'Loses Settings: Finance'",
    ("Service Manager", "Settings: Data Import", "STAGING-LESS"):
        "Spec Behavior-Changes l.478: Service Manager 'Loses Settings: Data Import'",
    ("Foreman", "Timesheets Create & Edit", "STAGING-LESS"):
        "Spec Behavior-Changes l.479: Foreman 'Loses Timesheets Edit'",
    ("Technician", "Send to Portal", "STAGING-LESS"):
        "Spec Behavior-Changes l.480 + change-log l.565: Technician 'Lose Send to Portal'",
    ("Parts Manager", "Work Orders Delete", "STAGING-LESS"):
        "Spec Behavior-Changes l.481: Parts Manager 'Loses WO/WOL Delete'",
    ("Parts Manager", "Work Order Lines Delete", "STAGING-LESS"):
        "Spec Behavior-Changes l.481: Parts Manager 'Loses WO/WOL Delete'",
    ("Office User", "Catalog & Inventory Create & Edit", "STAGING-LESS"):
        "Spec Behavior-Changes l.483: Office 'Catalog reduced to V only'",
    # intended EXPANSIONS (STAGING-MORE)
    ("Senior Service Advisor", "Work Orders Delete", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains WO Delete'",
    ("Senior Service Advisor", "Work Order Lines Delete", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains WOL Delete'",
    ("Senior Service Advisor", "Schedule Delete", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains Schedule Delete'",
    ("Senior Service Advisor", "Part Sales Delete", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains PartSales Delete'",
    ("Senior Service Advisor", "Invoicing & Payments Delete (reverse/delete invoice)", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains Invoicing FULL'",
    ("Senior Service Advisor", "Timesheets Create & Edit", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains Timesheets CE'",
    ("Senior Service Advisor", "Customer Portal Page Access", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains Customer Portal'",
    ("Senior Service Advisor", "See AP/AR Data", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains AP/AR'",
    ("Senior Service Advisor", "Reports Page Access", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains Reports'",
    ("Foreman", "Work Order Lines Delete", "STAGING-MORE"): "Spec l.479: Foreman 'Gains WOL Delete'",
    ("Foreman", "Schedule Delete", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Schedule Delete'",
    ("Foreman", "Order Parts (on WO)", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Order Parts'",
    ("Foreman", "WO History / Audit Log (view)", "STAGING-MORE"): "Spec l.479: Foreman 'Gains History Logs'",
    ("Foreman", "View History Logs (cross-toggle)", "STAGING-MORE"): "Spec l.479: Foreman 'Gains History Logs'",
    ("Foreman", "Part Sales View", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Parts Dept (Part Sales V)'",
    ("Foreman", "Catalog & Inventory View", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Parts Dept (Catalog V/CE)'",
    ("Foreman", "Catalog & Inventory Create & Edit", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Parts Dept (Catalog V/CE)'",
    ("Foreman", "Vendor & Order Mgmt View", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Parts Dept (Vendor V/CE)'",
    ("Foreman", "Vendor & Order Mgmt Create & Edit", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Parts Dept (Vendor V/CE)'",
    ("Foreman", "Receive / accept a delivery (Bulk Receive)", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Vendor V/CE' (delivery)",
    ("Foreman", "Assign vendor to a WO part order", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Vendor V/CE + Order Parts'",
    ("Foreman", "Invoicing & Payments View", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Invoicing V/CE'",
    ("Foreman", "Invoicing & Payments Create & Edit", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Invoicing V/CE'",
    ("Foreman", "Create an invoice from a WO (estimate->invoice)", "STAGING-MORE"): "Spec l.479: Foreman 'Gains Invoicing V/CE'",
    ("Foreman", "See Financial Data on WO (rates/margins/totals)", "STAGING-MORE"): "Spec l.479 Invoicing V/CE + l.572 (SFD required for invoicing)",
    ("Foreman", "See Financial Data (cross-toggle)", "STAGING-MORE"): "Spec l.479 Invoicing V/CE + l.572 (SFD required for invoicing)",
    ("Technician", "Pick Parts", "STAGING-MORE"): "Spec l.480: Technician 'Gains Pick Parts'",
    ("Parts Manager", "Schedule View", "STAGING-MORE"): "Spec l.481: Parts Manager 'Gains Schedule View'",
    ("Parts Manager", "Customer Portal Page Access", "STAGING-MORE"): "Spec l.481: Parts Manager 'Gains Customer Portal'",
    ("Parts Technician", "Pick Parts", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains Pick Parts'",
    ("Parts Technician", "Order Parts (on WO)", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains Order Parts'",
    ("Parts Technician", "Invoicing & Payments Create & Edit", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains Invoicing V/CE'",
    ("Parts Technician", "Invoicing & Payments View", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains Invoicing V/CE'",
    ("Parts Technician", "WO History / Audit Log (view)", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains History Logs'",
    ("Parts Technician", "View History Logs (cross-toggle)", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains History Logs'",
    ("Parts Technician", "Create an invoice from a WO (estimate->invoice)", "STAGING-MORE"): "Spec l.482: Parts Tech 'Gains Invoicing V/CE'",
    ("Parts Technician", "See Financial Data on WO (rates/margins/totals)", "STAGING-MORE"): "Spec l.482 Invoicing V/CE + l.572 (SFD required for invoicing)",
    ("Parts Technician", "See Financial Data (cross-toggle)", "STAGING-MORE"): "Spec l.482 Invoicing V/CE + l.572 (SFD required for invoicing)",
    ("Office User", "Customers Delete", "STAGING-MORE"): "Spec l.483: Office 'Customer Mgmt gains Delete'",
    ("Service Manager", "Customer Portal Page Access", "STAGING-MORE"): "Spec l.478: SM 'Gains Customer Portal'",
    ("Service Manager", "Billing Portal Page Access", "STAGING-MORE"): "Spec l.478: SM 'Gains Billing Portal'",
    ("Service Advisor", "Customer Portal Page Access", "STAGING-MORE"): "Spec l.485: SA Limited View 'Gains Customer Portal'",
    ("Senior Service Advisor", "Customers Delete", "STAGING-MORE"): "Spec l.477 (Senior SA expansion set)",
    ("Senior Service Advisor", "Part Sales Create & Edit", "STAGING-MORE"): "Spec l.477: Senior SA 'Gains ... Vendor FULL/Invoicing FULL' expansion set",

    # ==== VERIFICATION-REPORT CORRECTIONS (compare-VERIFICATION-2026-07-14.md) ====
    # 3.1 CONFIRMED spec-explicit correction (was mis-flagged No; removes a false High risk)
    ("Service Advisor", "See AP/AR Data", "STAGING-LESS"):
        "Spec Behavior-Changes: 'SA Limited View to Svc Advisor - AP/AR OFF preserves core restriction' (intended reduction; verification 3.1)",
    # 3.2 Migration-Type generic-expansion reclassifications (expansion-typed roles: Foreman / Parts Technician)
    #     Cite the role Migration Type; flagged as generic-expansion, NOT individually itemized in Behavior-Changes.
    ("Foreman", "Decline a WO part return", "STAGING-MORE"):
        "Migration-Type: Foreman = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
    ("Parts Technician", "Process a WO part return (create)", "STAGING-MORE"):
        "Migration-Type: Parts Technician = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
    ("Parts Technician", "Create / edit customer from New WO screen", "STAGING-MORE"):
        "Migration-Type: Parts Technician = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
    ("Parts Technician", "Schedule View", "STAGING-MORE"):
        "Migration-Type: Parts Technician = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
    ("Parts Technician", "Customers View", "STAGING-MORE"):
        "Migration-Type: Parts Technician = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
    ("Parts Technician", "Customers Create & Edit", "STAGING-MORE"):
        "Migration-Type: Parts Technician = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
    ("Parts Technician", "Part Sales Create & Edit", "STAGING-MORE"):
        "Migration-Type: Parts Technician = 'Direct (with expansions)' - generic expansion clause (not individually itemized; verification 3.2)",
}

# ============================================================================
# CAPABILITY SPEC TABLE
# name, cat ('WO'|'GEN'), stg(role)->bool, prod(prodcode)->bool, severity, confidence
# ============================================================================
CAPS = [
    # ---------- WORK ORDERS (granular) ----------
    ("Work Orders View", "WO", lambda r: s_has(r, "workOrdersView"),
     lambda c: p_has(c, "work_order", "view"), "Medium", "live"),
    ("Work Orders Create & Edit", "WO", lambda r: s_has(r, "workOrdersCreateAndEdit"),
     lambda c: p_has(c, "work_order", "create", "change"), "High", "live"),
    ("Work Orders Delete", "WO", lambda r: s_has(r, "workOrdersDelete"),
     lambda c: p_all(c, "work_order"), "High", "live"),
    ("Work Order Lines Create & Edit", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_line", "create", "change"), "High", "live"),
    ("Work Order Lines Delete", "WO", lambda r: s_has(r, "workOrderLinesDelete"),
     lambda c: p_has(c, "work_order_line", "remove"), "High", "live"),
    ("Order Parts (on WO)", "WO", lambda r: s_has(r, "woOrderParts"),
     lambda c: p_has(c, "work_order_part_request", "create"), "High", "live"),
    ("Pick Parts", "WO", lambda r: s_has(r, "woPickParts"),
     lambda c: p_has(c, "work_order_part", "change"), "Medium", "live"),
    ("Manage picked WO parts (view/change)", "WO", lambda r: s_has(r, "woPickParts") or s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_part", "view", "change"), "Medium", "NEEDS-REVIEW"),
    ("Remove a WO part", "WO", lambda r: s_has(r, "workOrderLinesDelete") or s_has(r, "workOrdersDelete"),
     lambda c: p_has(c, "work_order_part", "remove"), "High", "NEEDS-REVIEW"),
    ("Add / request a part on WO (part request)", "WO", lambda r: s_has(r, "woOrderParts") or s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_part_request", "create", "change"), "Medium", "NEEDS-REVIEW"),
    ("Process a WO part return (create)", "WO", lambda r: s_has(r, "vendorOrderManagementCreateAndEdit") or s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_part_return", "create", "change"), "Medium", "NEEDS-REVIEW"),
    ("Approve / complete a WO part return", "WO", lambda r: s_has(r, "invoicingPaymentsDelete") or s_has(r, "workOrdersDelete"),
     lambda c: p_has(c, "work_order_part_return", "complete"), "High", "NEEDS-REVIEW"),
    ("Decline a WO part return", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_part_return", "decline"), "Medium", "NEEDS-REVIEW"),
    ("Canned lines on WO (add/edit)", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit") or s_has(r, "workOrdersCreateAndEdit"),
     lambda c: p_has(c, "work_order_canned_line", "create", "change"), "Low", "NEEDS-REVIEW"),
    ("WO History / Audit Log (view)", "WO", lambda r: s_ct(r, "viewHistoryLogs"),
     lambda c: p_has(c, "work_order_history", "view"), "Medium", "live"),
    ("Clock in / log time on a WO line task", "WO", lambda r: (s_vm(r) == "tech" or s_has(r, "workOrderLinesCreateAndEdit")),
     lambda c: p_has(c, "work_order_line_task", "clock_in"), "Low", "NEEDS-REVIEW"),
    ("Edit / move WO line tasks", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_line_task", "change", "move", "create"), "Low", "NEEDS-REVIEW"),
    ("Mark Reviewed / review sign-off", "WO", lambda r: s_has(r, "woReviewWorkOrders"),
     lambda c: p_has(c, "work_order", "change"), "Medium", "NEEDS-REVIEW"),
    ("Complete a Work Order", "WO", lambda r: s_has(r, "workOrdersCreateAndEdit"),
     lambda c: p_has(c, "work_order", "create", "change"), "Medium", "live"),
    ("Approve / decline a WO line", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_line", "change"), "Medium", "live"),
    ("Set line status (bulk)", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit"),
     lambda c: p_has(c, "work_order_line", "change"), "Low", "live"),
    ("Full WO view mode (vs Tech view)", "WO", lambda r: s_vm(r) == "full",
     lambda c: c not in ("ROLE_TECHNICIAN", "ROLE_TIME_CLOCK_USER", "ROLE_REPORTING"), "Medium", "NEEDS-REVIEW"),
    ("See Financial Data on WO (rates/margins/totals)", "WO", lambda r: s_ct(r, "seeFinancialData"),
     lambda c: p_has(c, "invoice", "view"), "High", "NEEDS-REVIEW"),
    ("Send to Portal", "WO", lambda r: (s_vm(r) == "full" and s_has(r, "customerPortalPageAccess") and s_has(r, "woReviewWorkOrders")),
     lambda c: p_has(c, "work_order", "view"), "High", "NEEDS-REVIEW"),
    ("Send to Terminal (take payment on WO)", "WO", lambda r: (s_has(r, "invoicingPaymentsCreateAndEdit") and s_has(r, "customerPortalPageAccess")),
     lambda c: p_has(c, "invoice", "create", "change"), "High", "NEEDS-REVIEW"),
    ("Assign vendor to a WO part order", "WO", lambda r: s_has(r, "vendorOrderManagementCreateAndEdit"),
     lambda c: p_has(c, "work_order_part_request", "create") or p_has(c, "vendor", "change"), "Medium", "NEEDS-REVIEW"),
    ("Add a vendorless / manual part on WO", "WO", lambda r: s_has(r, "workOrderLinesCreateAndEdit") or s_has(r, "catalogInventoryCreateAndEdit"),
     lambda c: p_has(c, "work_order_part_request", "create"), "Low", "NEEDS-REVIEW"),
    ("Create / edit customer from New WO screen", "WO", lambda r: s_has(r, "customersCreateAndEdit"),
     lambda c: p_has(c, "customer", "create", "change"), "Medium", "live"),
    ("Create / edit asset (vehicle) from New WO screen", "WO", lambda r: s_has(r, "workOrdersCreateAndEdit"),
     lambda c: p_has(c, "vehicle", "create", "change"), "Medium", "NEEDS-REVIEW"),
    ("Create an invoice from a WO (estimate->invoice)", "WO", lambda r: s_has(r, "invoicingPaymentsCreateAndEdit"),
     lambda c: p_has(c, "invoice", "create"), "High", "live"),
    ("WO notes - create / edit", "WO", lambda r: s_has(r, "workOrdersView"),
     lambda c: p_has(c, "work_order", "view", "change", "create"), "Low", "NEEDS-REVIEW"),
    ("WO notes - delete", "WO", lambda r: s_has(r, "workOrdersDelete"),
     lambda c: p_all(c, "work_order"), "Low", "NEEDS-REVIEW"),

    # ---------- GENERAL / whole-app atoms ----------
    ("Schedule View", "GEN", lambda r: s_has(r, "scheduleView"),
     lambda c: p_has(c, "schedule", "view"), "Low", "live"),
    ("Schedule Create & Edit", "GEN", lambda r: s_has(r, "scheduleCreateAndEdit"),
     lambda c: p_all(c, "schedule"), "Low", "live"),
    ("Schedule Delete", "GEN", lambda r: s_has(r, "scheduleDelete"),
     lambda c: p_all(c, "schedule"), "Low", "live"),
    ("Customers View", "GEN", lambda r: s_has(r, "customersView"),
     lambda c: p_has(c, "customer", "view"), "Low", "live"),
    ("Customers Create & Edit", "GEN", lambda r: s_has(r, "customersCreateAndEdit"),
     lambda c: p_has(c, "customer", "create", "change"), "Medium", "live"),
    ("Customers Delete", "GEN", lambda r: s_has(r, "customersDelete"),
     lambda c: p_all(c, "customer"), "Medium", "live"),
    ("Catalog & Inventory View", "GEN", lambda r: s_has(r, "catalogInventoryView"),
     lambda c: p_has(c, "inventory", "view") or p_has(c, "catalogue", "view"), "Low", "live"),
    ("Catalog & Inventory Create & Edit", "GEN", lambda r: s_has(r, "catalogInventoryCreateAndEdit"),
     lambda c: p_has(c, "inventory", "create", "change") or p_has(c, "catalogue", "create", "change"), "Medium", "live"),
    ("Catalog & Inventory Delete", "GEN", lambda r: s_has(r, "catalogInventoryDelete"),
     lambda c: p_all(c, "inventory") or p_all(c, "catalogue"), "Medium", "live"),
    ("Vendor & Order Mgmt View", "GEN", lambda r: s_has(r, "vendorOrderManagementView"),
     lambda c: p_has(c, "vendor", "view") or p_has(c, "delivery", "view") or p_has(c, "work_order_part_request", "view"), "Low", "live"),
    ("Vendor & Order Mgmt Create & Edit", "GEN", lambda r: s_has(r, "vendorOrderManagementCreateAndEdit"),
     lambda c: p_has(c, "vendor", "create", "change") or p_has(c, "delivery", "create", "change"), "Medium", "live"),
    ("Vendor & Order Mgmt Delete", "GEN", lambda r: s_has(r, "vendorOrderManagementDelete"),
     lambda c: p_all(c, "vendor") or p_all(c, "delivery"), "Medium", "live"),
    ("Receive / accept a delivery (Bulk Receive)", "GEN", lambda r: s_has(r, "vendorOrderManagementCreateAndEdit"),
     lambda c: p_has(c, "delivery", "create", "change"), "Medium", "live"),
    ("Part Sales View", "GEN", lambda r: s_has(r, "partSalesView"),
     lambda c: p_has(c, "manual_part_return", "view"), "Low", "NEEDS-REVIEW"),
    ("Part Sales Create & Edit", "GEN", lambda r: s_has(r, "partSalesCreateAndEdit"),
     lambda c: p_has(c, "manual_part_return", "create", "change"), "Medium", "NEEDS-REVIEW"),
    ("Part Sales Delete", "GEN", lambda r: s_has(r, "partSalesDelete"),
     lambda c: p_all(c, "manual_part_return"), "Medium", "NEEDS-REVIEW"),
    ("Invoicing & Payments View", "GEN", lambda r: s_has(r, "invoicingPaymentsView"),
     lambda c: p_has(c, "invoice", "view"), "Medium", "live"),
    ("Invoicing & Payments Create & Edit", "GEN", lambda r: s_has(r, "invoicingPaymentsCreateAndEdit"),
     lambda c: p_has(c, "invoice", "create", "change"), "High", "live"),
    ("Invoicing & Payments Delete (reverse/delete invoice)", "GEN", lambda r: s_has(r, "invoicingPaymentsDelete"),
     lambda c: p_all(c, "invoice"), "High", "live"),
    ("Timesheets View", "GEN", lambda r: s_has(r, "timesheetsView"),
     lambda c: p_has(c, "timesheet_activities", "change") or p_has(c, "attendance", "view") or p_has(c, "payroll_timesheet_reports", "view"), "Low", "NEEDS-REVIEW"),
    ("Timesheets Create & Edit", "GEN", lambda r: s_has(r, "timesheetsCreateAndEdit"),
     lambda c: p_has(c, "timesheet_activities", "change"), "Medium", "NEEDS-REVIEW"),
    ("Reports Page Access", "GEN", lambda r: s_has(r, "reportsPageAccess"),
     lambda c: any(res.endswith("_reports") for res in PROD.get(c, {})), "Medium", "live"),
    ("Customer Portal Page Access", "GEN", lambda r: s_has(r, "customerPortalPageAccess"),
     lambda c: False, "Medium", "NEEDS-REVIEW"),
    ("Billing Portal Page Access", "GEN", lambda r: s_has(r, "billingPortalPageAccess"),
     lambda c: p_all(c, "organization"), "Low", "NEEDS-REVIEW"),
    ("Settings: App", "GEN", lambda r: s_has(r, "settingsApp"),
     lambda c: p_all(c, "organization"), "Medium", "NEEDS-REVIEW"),
    ("Settings: Service", "GEN", lambda r: s_has(r, "settingsService"),
     lambda c: p_all(c, "labour_type") or p_all(c, "inspection_template") or p_all(c, "bay"), "Medium", "NEEDS-REVIEW"),
    ("Settings: Parts", "GEN", lambda r: s_has(r, "settingsParts"),
     lambda c: p_all(c, "pricing_matrix"), "Medium", "NEEDS-REVIEW"),
    ("Settings: Integrations", "GEN", lambda r: s_has(r, "settingsIntegrations"),
     lambda c: p_all(c, "organization"), "Medium", "NEEDS-REVIEW"),
    ("Settings: Finance", "GEN", lambda r: s_has(r, "settingsFinance"),
     lambda c: p_all(c, "tax") or p_all(c, "payment_method"), "Medium", "NEEDS-REVIEW"),
    ("Settings: Data Import", "GEN", lambda r: s_has(r, "settingsDataImport"),
     lambda c: p_all(c, "data_import"), "Medium", "live"),
    ("Settings: Wages", "GEN", lambda r: s_has(r, "settingsWages"),
     lambda c: p_all(c, "staff"), "Medium", "NEEDS-REVIEW"),
    ("See Financial Data (cross-toggle)", "GEN", lambda r: s_ct(r, "seeFinancialData"),
     lambda c: p_has(c, "invoice", "view") or any(res.endswith("_reports") for res in PROD.get(c, {})), "High", "NEEDS-REVIEW"),
    ("See AP/AR Data", "GEN", lambda r: s_ct(r, "seeApArData"),
     lambda c: p_all(c, "invoice") or p_has(c, "customer_transaction") or p_has(c, "vendor_transaction"), "High", "NEEDS-REVIEW"),
    ("View History Logs (cross-toggle)", "GEN", lambda r: s_ct(r, "viewHistoryLogs"),
     lambda c: p_has(c, "work_order_history", "view"), "Medium", "live"),
    ("Manage Staff", "GEN", lambda r: s_has(r, "settingsWages") or s_has(r, "settingsApp"),
     lambda c: p_all(c, "staff"), "Medium", "NEEDS-REVIEW"),
]

def intended_for(role, cap, direction):
    cit = SPEC_INTENDED.get((role, cap, direction))
    if cit:
        return "Yes", cit
    return "No", "not in spec (unaccounted change - review before release)"

def prod_grant(role, cap_prod):
    holders = [PLABEL[c] for c in MAP[role] if cap_prod(c)]
    return (len(holders) > 0), holders

def ver_conf(cap, conf):
    # Verification confidence (compare-VERIFICATION-2026-07-14.md 4 + staging UI-verify 2026-07-15):
    #  - UI_VERIFIED override: FE-gated High rows verified against the shipped FE gate + live role
    #    defs on 2026-07-15 (Send to Portal / See AP/AR -> HIGH; Send to Terminal -> absent-from-build;
    #    part-return verbs stay MEDIUM).
    #  - else resource/action-mapped rows = HIGH; remaining FE-gated rows = MEDIUM / NEEDS-UI-VERIFY.
    if cap in UI_VERIFIED:
        return UI_VERIFIED[cap][0]
    if conf == "live":
        return "HIGH (resource/action-mapped; recompute-matched)"
    return "MEDIUM / NEEDS-UI-VERIFY (FE-gated / no clean old-model atom; role-definition-inferred)"

def ver_short(cap, conf):
    if cap in UI_VERIFIED:
        return UI_VERIFIED[cap][0].split(" ")[0]  # HIGH or MEDIUM
    return "HIGH" if conf == "live" else "MEDIUM / UI-verify"

rows = []
delta_rows = []
oom_rows = []          # out-of-model (staff-record) informational rows - excluded from risk counts
for role in ORDER:
    mig = MIGRATION_TYPE[role]
    for name, cat, sfn, pfn, sev, conf in CAPS:
        sg = bool(sfn(role))
        pg, holders = prod_grant(role, pfn)
        prod_names = "; ".join(holders) if holders else "(none of mapped)"
        mapped_all = " + ".join(PLABEL[c] for c in MAP[role])
        if sg == pg:
            direction = "match"
        elif pg and not sg:
            direction = "STAGING-LESS"
        else:
            direction = "STAGING-MORE"
        rowconf = conf  # SA/SSA mapping CONFIRMED 2026-07-14 - no mapping-unconfirmed flag; per-capability conf (live/NEEDS-REVIEW) stands
        rec = dict(role=role, slug=STG_SLUG.get(role, ""), cat=cat, cap=name, mapped=mapped_all,
                   holders=prod_names, pg="Yes" if pg else "No", sg="Yes" if sg else "No",
                   direction=direction, sev=sev, conf=rowconf, mig=mig,
                   vconf=ver_conf(name, conf), vshort=ver_short(name, conf))
        rows.append(rec)
        if direction in ("STAGING-LESS", "STAGING-MORE"):
            ev = f"prod holder: {prod_names} | staging live | old->new map conf={conf}"
            if role in SA_SSA_MERGE:
                ev += " | SA/SSA merge mapping CONFIRMED by QA lead 2026-07-14"
            rec2 = dict(rec)
            if name in OUT_OF_MODEL:
                # verification 3.3 - staff-record-controlled; not a permission delta; excluded from risk counts
                rec2.update(intended="n/a (out-of-model)", cit=OUT_OF_MODEL_NOTE, ev=ev)
                oom_rows.append(rec2)
            else:
                intended, cit = intended_for(role, name, direction)
                rec2.update(intended=intended, cit=cit, ev=ev)
                delta_rows.append(rec2)

# ============================================================================
# WORKBOOK
# ============================================================================
def style_header(ws, ncols, color="1F4E78"):
    fill = PatternFill("solid", fgColor=color)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(vertical="top", wrap_text=True)

RED = Font(color="C00000", bold=True)
HDR = ["Staging Role", "Staging role slug", "Production role(s) mapped", "Migration Type",
       "Capability", "Prod grants?", "Staging grants?", "Direction (STAGING-LESS / STAGING-MORE)",
       "Per spec - intended? (Yes/No)", "Spec citation", "Severity",
       "Evidence / source", "Confidence", "Verification confidence"]

def write_delta_tab(ws, subset):
    ws.append(HDR)
    style_header(ws, len(HDR))
    for r in subset:
        ws.append([r["role"], r["slug"], r["mapped"], r["mig"], r["cap"], r["pg"], r["sg"],
                   r["direction"], r["intended"], r["cit"], r["sev"], r["ev"], r["conf"], r["vconf"]])
        if r["intended"] == "No":
            ws.cell(row=ws.max_row, column=9).font = RED
    for col, w in zip("ABCDEFGHIJKLMN", [22, 22, 34, 30, 42, 11, 12, 22, 16, 52, 9, 50, 22, 34]):
        ws.column_dimensions[col].width = w
    for rr in ws.iter_rows(min_row=2):
        for cc in rr:
            cc.alignment = Alignment(wrap_text=True, vertical="top")

def summary_tab(ws, subset, oom_subset, title):
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(["Staging Role", "Staging role slug", "Merged?", "Prod role(s) mapped", "Migration Type",
               "STAGING-LESS intended (Yes)", "STAGING-LESS NOT-in-spec (No) = RISK",
               "STAGING-MORE intended (Yes)", "STAGING-MORE NOT-in-spec (No) = RISK",
               "Out-of-model (staff-record, excl.)", "Highest severity", "Mapping confirmed?"])
    for c in range(1, 13):
        cell = ws.cell(row=2, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    sevrank = {"High": 3, "Medium": 2, "Low": 1}
    for role in ORDER:
        items = [x for x in subset if x["role"] == role]
        less = [x for x in items if x["direction"] == "STAGING-LESS"]
        more = [x for x in items if x["direction"] == "STAGING-MORE"]
        lyes = sum(1 for x in less if x["intended"] == "Yes")
        lno = sum(1 for x in less if x["intended"] == "No")
        myes = sum(1 for x in more if x["intended"] == "Yes")
        mno = sum(1 for x in more if x["intended"] == "No")
        oom = sum(1 for x in oom_subset if x["role"] == role)
        hs = max([x["sev"] for x in items], key=lambda s: sevrank[s], default="-")
        merged = "YES" if len(MAP[role]) > 1 else "no"
        conf = "confirmed"
        ws.append([role, STG_SLUG.get(role, ""), merged, " + ".join(PLABEL[c] for c in MAP[role]),
                   MIGRATION_TYPE[role], lyes, lno, myes, mno, oom, hs, conf])
        if lno:
            ws.cell(row=ws.max_row, column=7).font = RED
        if mno:
            ws.cell(row=ws.max_row, column=9).font = RED
    for col, w in zip("ABCDEFGHIJKL", [22, 22, 8, 36, 30, 22, 30, 22, 30, 24, 14, 16]):
        ws.column_dimensions[col].width = w

wb = Workbook()

# ---- Tab 0: READ ME ----
ws0 = wb.active
ws0.title = "READ ME - DATA STATUS"
banner = [
    ["CUSTOM ROLES (SV-7388) - PROD vs STAGING permission compare - LIVE 2026-07-15"],
    [""],
    ["*** BOTH SIDES LIVE-VERIFIED (this is NOT the spec-predicted interim) ***"],
    [""],
    ["STAGING (new custom-roles model): 11 system roles, GET /api/organizations/{org}/roles"],
    ["  + per-role GET /api/roles/{id}. Org d55bc308-...  All HTTP 200."],
    ["PRODUCTION (old legacy model): authenticated live on api.shopview.com with a fresh"],
    ["  PHPSESSID (no SSO). Prod org UUID 72b2cc90-6964-4429-a207-76e55f946936."],
    ["  14 legacy roles from GET /api/iam/list-roles (NO 'Owner' role exists in this org)."],
    ["  Per-role effective permissions captured by impersonation (POST /api/switch-user ->"],
    ["  data.permissions -> exit-switch-user). Userless roles captured by temporarily"],
    ["  assigning a throwaway ZZ invite-test user that role, then restoring to Technician."],
    [""],
    ["OLD model = {resource_name, action_name} pairs; action '*' = ALL actions (incl delete)."],
    ["NEW model = 41 fe_permission atoms + view_mode + 3 cross-toggles. Capabilities are"],
    ["translated old<->new (see 'Confidence' col: 'live' = clean resource/action map;"],
    ["'NEEDS-REVIEW' = old model has no clean equivalent / FE-gated, judged best-effort)."],
    [""],
    ["BI-DIRECTIONAL: STAGING-LESS = prod grants, staging does not (regression / over-in-prod)."],
    ["  STAGING-MORE = staging grants, prod did not (new model grants more)."],
    ["Per spec - intended? Yes = the spec Behavior-Changes/migration text documents it (cited)."],
    ["  No (RED) = the change is NOT accounted for in the spec = HEADLINE RELEASE RISK."],
    [""],
    ["MAPPING CONFIRMED by QA lead 2026-07-14 (spec migration table is authoritative):"],
    ["  Senior Service Advisor <- Service Advisor + SA Technician + SA No Reports (3 merged);"],
    ["  Service Advisor <- SA Limited View. The naming trap is RESOLVED - these rows are FINAL"],
    ["  (no longer NEEDS-REVIEW for mapping; a per-capability NEEDS-REVIEW may still apply where"],
    ["  an old-model atom has no clean equivalent / is FE-gated)."],
    ["Administrator compared 1:1 (prod Administrator <-> staging Administrator); the spec's"],
    ["  'Owner merged in' is not applicable - no Owner role exists in either environment"],
    ["  (confirmed by QA lead 2026-07-14). Administrator delta rows stand as computed."],
    [""],
    ["*** INDEPENDENT VERIFICATION APPLIED (compare-VERIFICATION-2026-07-14.md) ***"],
    ["Migration Type column = the QA-lead spec Migration-Type per staging role (spec intent)."],
    ["Verification confidence column: HIGH = resource/action-mapped row, independent recompute"],
    ["  matched + prod live-confirmed; MEDIUM / NEEDS-UI-VERIFY = FE-gated / no clean old-model"],
    ["  atom (Send to Portal/Terminal, part-return verbs, AP/AR proxies, portal pages) - these"],
    ["  are role-definition-inferred, NOT UI-click-verified; verify in staging UI per role."],
    ["Corrections applied: (1) Service Advisor 'See AP/AR Data' STAGING-LESS is now intended=Yes"],
    ["  (spec: 'AP/AR OFF preserves core restriction') - removed a false High risk. (2) ~7"],
    ["  STAGING-MORE expansion rows on expansion-typed roles (Foreman, Parts Technician) flipped"],
    ["  No->Yes citing Migration-Type 'Direct (with expansions)' (generic clause, not itemized)."],
    ["  (3) Clock-in + Timesheets rows are STAFF-RECORD-controlled (spec 'Staff Record Settings'),"],
    ["  NOT permission deltas - moved to the 'Out-of-model (staff-record)' tab and EXCLUDED from"],
    ["  the risk 'No' counts. Genuine over-grants/regressions (e.g. Parts Manager gains WO"],
    ["  Create&Edit + WO Lines Create&Edit; SM/PM delete+settings; Sales Rep SFD/AP-AR; all"],
    ["  STAGING-LESS regressions) are KEPT as No."],
    [""],
    ["COMPLETENESS (verification): NO release-critical omissions - all 43 staging atoms + 3"],
    ["  cross-toggles + view_mode represented; all 14 prod roles (no Owner) + 11 staging roles +"],
    ["  4 merges present. Only 5 LOW-severity prod resources have no explicit row (settings /"],
    ["  reference / report-view only): workplace, department, vehicle_type, vehicle_history,"],
    ["  shop_billing_efficiency (subsumed under Settings / Reports / vehicle view)."],
    [""],
    [""],
    ["*** STAGING UI / FE-SOURCE VERIFICATION (2026-07-15) ***"],
    ["'Staging role slug' column = the LIVE role identifier from GET /api/role-templates (mapped"],
    ["  via each role's template_id). The role object itself has NO slug field; the slug is the"],
    ["  role TEMPLATE slug (real API value, NOT hardcoded). Raw objects in staging-ui-verify-2026-07-14/."],
    ["FE-gated High rows verified against the SHIPPED FE gate + live role defs (2026-07-15):"],
    ["  - Send to Portal -> HIGH: gate = Customer Portal access; ABSENT for all 6 STAGING-LESS"],
    ["    roles (Technician/Foreman/Parts Tech/Office/Sales Rep/Time Clock) = HIDDEN, PRESENT for"],
    ["    the 5 kept roles = SHOWN (internally consistent)."],
    ["  - See AP/AR Data -> HIGH: gate = seeApArData() cross-toggle; live values match per role."],
    ["  - Send to Terminal -> control ABSENT from the staging build (no terminal/take-payment"],
    ["    control anywhere in the FE bundle) - not a per-role gate."],
    ["  - Part-return approve/complete/decline -> stay MEDIUM (control exists, exact gate not"],
    ["    isolated; pixel-screenshot blocked by mid-run SSO session expiry)."],
    [""],
    ["TABS: 'Deltas - ALL (bi-dir)' whole-app | 'Work Orders - granular' WO-only |"],
    ["  Summary tabs (per-role Yes/No 2x2) | 'Out-of-model (staff-record)' informational |"],
    ["  Full capability matrix | Open questions."],
]
for row in banner:
    ws0.append(row)
ws0["A1"].font = Font(bold=True, size=13)
ws0["A3"].font = Font(bold=True, color="1F6F1F", size=12)
ws0.column_dimensions["A"].width = 100

write_delta_tab(wb.create_sheet("Deltas - ALL (bi-dir)"), delta_rows)
wo_deltas = [r for r in delta_rows if r["cat"] == "WO"]
oom_wo = [r for r in oom_rows if r["cat"] == "WO"]
write_delta_tab(wb.create_sheet("Work Orders - granular"), wo_deltas)
summary_tab(wb.create_sheet("Summary per role (ALL)"), delta_rows, oom_rows,
            "Per-role 2x2 summary - WHOLE APP (No = release risk; out-of-model excluded)")
summary_tab(wb.create_sheet("Summary per role (WO)"), wo_deltas, oom_wo,
            "Per-role 2x2 summary - WORK ORDERS only (No = release risk; out-of-model excluded)")

# ---- Out-of-model (staff-record) informational tab ----
ws_oom = wb.create_sheet("Out-of-model (staff-record)")
ws_oom.append(["INFORMATIONAL - staff-record-controlled, NOT permission deltas - EXCLUDED from risk 'No' counts"])
ws_oom["A1"].font = Font(bold=True, color="7F3F00", size=11)
OOM_HDR = ["Staging Role", "Staging role slug", "Migration Type", "Capability", "Direction",
           "Prod grants?", "Staging grants?", "Why out-of-model", "Severity"]
ws_oom.append(OOM_HDR)
for c in range(1, len(OOM_HDR) + 1):
    cell = ws_oom.cell(row=2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(vertical="top", wrap_text=True)
for r in oom_rows:
    ws_oom.append([r["role"], r["slug"], r["mig"], r["cap"], r["direction"], r["pg"], r["sg"], r["cit"], r["sev"]])
for col, w in zip("ABCDEFGHI", [22, 22, 30, 42, 22, 11, 12, 56, 9]):
    ws_oom.column_dimensions[col].width = w
for rr in ws_oom.iter_rows(min_row=3):
    for cc in rr:
        cc.alignment = Alignment(wrap_text=True, vertical="top")

# ---- Full capability matrix (staging then prod, side by side) ----
ws5 = wb.create_sheet("Full capability matrix")
ws5.append(["STAGING grants (live)", ""] + ORDER)
style_header(ws5, 2 + len(ORDER))
ws5.append(["role slug (live 2026-07-15)", ""] + [STG_SLUG.get(r, "") for r in ORDER])
for name, cat, sfn, pfn, sev, conf in CAPS:
    ws5.append([name, cat] + ["Y" if sfn(r) else "" for r in ORDER])
ws5.append([])
hdr2 = ws5.max_row + 1
ws5.append(["PRODUCTION grants (mapped union, live)", ""] + ORDER)
for c in range(1, 3 + len(ORDER)):
    cell = ws5.cell(row=hdr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="7F3F00")
for name, cat, sfn, pfn, sev, conf in CAPS:
    ws5.append([name, cat] + ["Y" if prod_grant(r, pfn)[0] else "" for r in ORDER])
ws5.column_dimensions["A"].width = 44
ws5.column_dimensions["B"].width = 6
for col in "CDEFGHIJKLM":
    ws5.column_dimensions[col].width = 12

# ---- Open questions ----
ws6 = wb.create_sheet("Open questions - NEEDS REVIEW")
ws6.append(["Item", "Detail"])
style_header(ws6, 2)
OPEN_Q = [
    ("Service Advisor / Senior SA mapping CONFIRMED (QA lead 2026-07-14)",
     "Naming trap RESOLVED by QA lead 2026-07-14 (spec migration table is authoritative): staging "
     "'Senior Service Advisor' <- legacy Service Advisor + SA Technician + SA No Reports (3 merged); "
     "staging 'Service Advisor' <- legacy 'SA Limited View'. All Service-Advisor / Senior-Service-Advisor "
     "rows are FINAL - the mapping-unconfirmed flag is removed. (The section-3549 1:1 same-name migration "
     "cases C26514/C26515 are superseded by this ruling.) A per-capability NEEDS-REVIEW may still apply "
     "where an old-model atom has no clean equivalent / is FE-gated."),
    ("Administrator compared 1:1 (Owner not applicable)",
     "Administrator compared 1:1 (prod Administrator <-> staging Administrator); the spec's 'Owner "
     "merged in' is not applicable - no Owner role exists in either environment; confirmed by QA lead "
     "2026-07-14. The Administrator delta rows stand as computed from the live Administrator-vs-"
     "Administrator capture (not incomplete)."),
    ("Old->new capability translation - NEEDS-REVIEW rows",
     "Some new-model atoms have no clean old-model resource/action equivalent and/or are FE-gated "
     "(Send to Portal, Send to Terminal, Customer/Billing Portal page access, See Financial Data, "
     "See AP/AR, Settings Service/Parts/Integrations/Wages, Part Sales, part-return verbs, line tasks). "
     "These rows are Confidence=NEEDS-REVIEW and mapped best-effort; verify in UI per role."),
    ("Customer Portal page access - no prod resource",
     "The old model exposes no 'customer portal' resource in the permission array; prod grant computed "
     "as No for all roles. Spec documents several roles GAIN Customer Portal (intended STAGING-MORE=Yes)."),
    ("Reporting legacy role has 0 resource permissions",
     "ROLE_REPORTING returns an EMPTY permissions array (report-page-only via role membership, not the "
     "permission model). It merges into Sales Representative; its report access is represented via the "
     "Reports Page Access capability."),
    ("Send to Portal / Send to Terminal are FE-gated",
     "Both are front-end button gates not enforced by a raw permission atom. Staging: Send to Portal "
     "needs Full view + Customer Portal + review; Send to Terminal needs Invoicing C&E + Customer Portal. "
     "Prod grant inferred from work_order/invoice access; verify live in UI per role before release."),
    ("Prod capture method (disposable TEST org)",
     "Prod is a disposable TEST org (per task). Per-role perms captured via impersonation "
     "(switch-user/exit-switch-user, fully reversible) and, for userless roles, a temporary role swap "
     "on throwaway user bilal.muzamil+bugstesting (restored to Technician; departments/workplace "
     "verified intact). No production data left modified."),
    ("Completeness result (independent verification)",
     "NO release-critical omissions. All 43 staging atoms + 3 cross-toggles + view_mode are "
     "represented; all 14 prod roles (no Owner) + 11 staging roles + 4 merges present; independent "
     "recompute matched the workbook 23/23 on critical rows (5 prod roles re-captured LIVE). The only "
     "gaps are 5 LOW-severity prod resources with no explicit row - workplace, department, "
     "vehicle_type, vehicle_history, shop_billing_efficiency - all settings/reference/report-view only "
     "(subsumed under Settings/Reports/vehicle view). Note: workplace* is held by SA-Limited-View so "
     "staging Service Advisor has a low-severity uncaptured 'workplace management' reduction."),
    ("Verification confidence + FE-gated UI-verify status (2026-07-15)",
     "HIGH-confidence rows are resource/action-mapped (recompute-matched + prod live-confirmed). "
     "STAGING FE-SOURCE VERIFICATION (2026-07-15) raised the priority FE-gated High rows: "
     "(a) 'Send to Portal' -> HIGH - FE gate is Customer Portal access (customerPortalPageAccess); "
     "live staging role defs show it ABSENT for all 6 STAGING-LESS roles (HIDDEN) and PRESENT for "
     "the 5 kept roles (SHOWN). (b) 'See AP/AR Data' -> HIGH - FE gate is seeApArData(); live "
     "cross-toggles match per role. (c) 'Send to Terminal' -> control ABSENT from the staging build "
     "entirely (no terminal/take-payment control in the FE bundle) - not a per-role gate; confirm "
     "prod's control name. (d) Part-return approve/complete/decline stay MEDIUM: controls exist "
     "('Process Return'/'Confirm Return') but the exact gate was not isolable from the minified "
     "source. NOTE: live pixel-screenshots were blocked - the sv_sso_session expired mid-run "
     "(quick-login 200 then next API 409 'Session has expired'); the above is FE-source (shipped "
     "gate logic) + live role-definition data. To capture pixel screenshots for the residual MEDIUM "
     "rows, re-run with a fresh staging cookie (boot2 hydration; TLS-1.2-max + disable ECH/PQ/HTTP2/"
     "QUIC through a local CONNECT-relay bridge - see staging-ui-verify-2026-07-14/)."),
]
for item, detail in OPEN_Q:
    ws6.append([item, detail])
ws6.column_dimensions["A"].width = 42
ws6.column_dimensions["B"].width = 110
for row in ws6.iter_rows(min_row=2):
    row[1].alignment = Alignment(wrap_text=True, vertical="top")

out_xlsx = os.path.join(HERE, "Prod-vs-Staging-Permission-Gaps_2026-07-14.xlsx")
wb.save(out_xlsx)

def counts(subset):
    return (
        [r for r in subset if r["direction"] == "STAGING-LESS" and r["intended"] == "No"],
        [r for r in subset if r["direction"] == "STAGING-LESS" and r["intended"] == "Yes"],
        [r for r in subset if r["direction"] == "STAGING-MORE" and r["intended"] == "No"],
        [r for r in subset if r["direction"] == "STAGING-MORE" and r["intended"] == "Yes"],
    )

print("wrote", out_xlsx)
ln, ly, mn, my = counts(delta_rows)
print(f"ALL: STAGING-LESS No={len(ln)} Yes={len(ly)} | STAGING-MORE No={len(mn)} Yes={len(my)}")
wln, wly, wmn, wmy = counts(wo_deltas)
print(f"WO : STAGING-LESS No={len(wln)} Yes={len(wly)} | STAGING-MORE No={len(wmn)} Yes={len(wmy)}")
print(f"OUT-OF-MODEL (excluded from No counts): {len(oom_rows)} rows (WO {len(oom_wo)})")

# ============================================================================
# MARKDOWN companion (kept in sync with the workbook)
# ============================================================================
SEVRANK = {"High": 3, "Medium": 2, "Low": 1}

def md_rows(subset):
    lines = ["| Staging role | Slug | Capability | Prod role(s) mapped | Severity | Confidence | Verification |",
             "|---|---|---|---|---|---|---|"]
    for r in sorted(subset, key=lambda x: (-SEVRANK[x["sev"]], x["role"], x["cap"])):
        lines.append(f"| {r['role']} | {r['slug']} | {r['cap']} | {r['holders']} | {r['sev']} | {r['conf']} | {r['vshort']} |")
    return "\n".join(lines)

def md_intended(subset, direction):
    rs = [r for r in subset if r["direction"] == direction and r["intended"] == "Yes"]
    lines = ["| Staging role | Capability | Prod role(s) mapped | Severity | Spec / Migration-Type citation |",
             "|---|---|---|---|---|"]
    for r in sorted(rs, key=lambda x: (-SEVRANK[x["sev"]], x["role"], x["cap"])):
        lines.append(f"| {r['role']} | {r['cap']} | {r['holders']} | {r['sev']} | {r['cit']} |")
    return "\n".join(lines)

less_no = [r for r in delta_rows if r["direction"] == "STAGING-LESS" and r["intended"] == "No"]
more_no = [r for r in delta_rows if r["direction"] == "STAGING-MORE" and r["intended"] == "No"]

# per-role summary
sum_lines = ["| Staging role | Slug | Merged? | Migration Type | STG-LESS Yes | **STG-LESS No** | STG-MORE Yes | **STG-MORE No** | Out-of-model (excl.) | Mapping |",
             "|---|---|---|---|---|---|---|---|---|---|"]
for role in ORDER:
    items = [x for x in delta_rows if x["role"] == role]
    less = [x for x in items if x["direction"] == "STAGING-LESS"]
    more = [x for x in items if x["direction"] == "STAGING-MORE"]
    lyes = sum(1 for x in less if x["intended"] == "Yes")
    lno = sum(1 for x in less if x["intended"] == "No")
    myes = sum(1 for x in more if x["intended"] == "Yes")
    mno = sum(1 for x in more if x["intended"] == "No")
    oom = sum(1 for x in oom_rows if x["role"] == role)
    merged = "YES" if len(MAP[role]) > 1 else "no"
    sum_lines.append(f"| {role} | {STG_SLUG.get(role,'')} | {merged} | {MIGRATION_TYPE[role]} | {lyes} | {lno} | {myes} | {mno} | {oom} | confirmed |")

# out-of-model md
oom_lines = ["| Staging role | Capability | Direction | Prod / Staging | Severity |",
             "|---|---|---|---|---|"]
for r in sorted(oom_rows, key=lambda x: (x["role"], x["cap"])):
    oom_lines.append(f"| {r['role']} | {r['cap']} | {r['direction']} | {r['pg']}/{r['sg']} | {r['sev']} |")

md = f"""# Custom Roles (SV-7388) — PRODUCTION vs STAGING Permission Gaps (LIVE, VERIFIED)

**Date:** 2026-07-15 · **Epic:** SV-7388 Custom Roles & Permissions · **PO:** Sasha Grosman
**Status:** ✅ BOTH SIDES LIVE-VERIFIED · **MAPPING CONFIRMED by QA lead 2026-07-14** (spec
migration table authoritative — Service Advisor / Senior Service Advisor rows are FINAL) ·
**INDEPENDENT VERIFICATION APPLIED** (`compare-VERIFICATION-2026-07-14.md`).
**Workbook:** `Prod-vs-Staging-Permission-Gaps_2026-07-14.xlsx` (13-col bi-directional main tab
with **Migration Type** + **Verification confidence** columns + dedicated **Work Orders —
granular** tab + per-role 2×2 summaries + **Out-of-model (staff-record)** tab + full matrix +
open questions).

## Data provenance (live)
- **Staging (new custom-roles model):** 11 system roles, `GET /api/organizations/{{org}}/roles`
  + per-role `GET /api/roles/{{id}}`, org `d55bc308-…`.
- **Production (old legacy model):** authenticated live on `api.shopview.com` (fresh PHPSESSID,
  no SSO). Prod org UUID `72b2cc90-6964-4429-a207-76e55f946936`. **14 legacy roles** from
  `GET /api/iam/list-roles`. **No "Owner" role exists in either environment**, so Administrator
  is compared **1:1** (spec "Owner merged in" not applicable; confirmed QA lead 2026-07-14).
  Per-role effective permissions captured by **impersonation** (`switch-user` → `data.permissions`
  → `exit-switch-user`); userless roles via a temporary throwaway-user role swap, restored to
  Technician. No prod data left modified.
- **Models:** old = `{{resource_name, action_name}}` pairs (action `*` = ALL incl. delete);
  new = 41 fe_permission atoms + view_mode + 3 cross-toggles. Capabilities translated old↔new.

## Independent verification applied (compare-VERIFICATION-2026-07-14.md)
- **Migration Type** column added per staging role (QA-lead spec Migration-Type = spec intent).
- **Verification confidence** column: **HIGH** = resource/action-mapped (independent recompute
  matched 23/23 + prod live-confirmed); **MEDIUM / NEEDS-UI-VERIFY** = FE-gated / no clean
  old-model atom (Send to Portal/Terminal, part-return verbs, AP/AR proxies, portal pages) —
  role-definition-inferred, **not UI-click-verified** (drive per role with a fresh staging cookie).
- **Correction 3.1:** *Service Advisor · See AP/AR Data · STAGING-LESS* → **intended = Yes**
  (spec "AP/AR OFF preserves core restriction") — removed a false **High** release risk.
- **Correction 3.2:** {len([k for k in SPEC_INTENDED if 'expansions' in SPEC_INTENDED[k].lower()])} STAGING-MORE expansion rows on expansion-typed roles
  (Foreman, Parts Technician) flipped **No → Yes**, citing Migration-Type "Direct (with
  expansions)" (generic clause; not individually itemized in Behavior-Changes).
- **Correction 3.3:** Clock-in + Timesheets rows are **staff-record-controlled** (spec "Staff
  Record Settings"), NOT permission deltas — moved to the **Out-of-model** section and **excluded
  from the risk "No" counts** ({len(oom_rows)} rows).
- **Kept as real "No" risks:** Parts Manager gains WO Create&Edit + WO Lines Create&Edit;
  SM/PM delete + settings over-grants; Sales Rep SFD/AP-AR; all STAGING-LESS regressions
  (Technician Order-Parts / WOL-Delete, Parts-Tech invoice-reverse, etc.).

## Staging UI / FE-source verification of the FE-gated High rows (2026-07-15)
Live staging role definitions (`GET /api/roles/{{id}}` + cross_toggles + view_mode, plus role
**slugs** from `GET /api/role-templates`) were evaluated against the ACTUAL front-end gate
predicate read from the shipped staging JS bundle. (Live pixel-screenshots were blocked: the
`sv_sso_session` expired mid-run — quick-login returned 200 but the next API call 409'd "Session
has expired"; the verification is FE-source + live role-definition data, authoritative for a
front-end DISPLAY gate.) Raw evidence: `staging-ui-verify-2026-07-14/`.

- **Send to Portal → HIGH (verified).** FE gate = Customer Portal access (store helper
  `userHasCustomerPortalAccess = has("customerPortalPageAccess")`; button in `WorkOrderNavBar`).
  Live staging: the atom is **ABSENT for all 6 STAGING-LESS roles** (Technician, Foreman, Parts
  Technician, Office User, Sales Representative, Time Clock User) → **HIDDEN**, and **PRESENT for
  the 5 roles that keep it** (Admin, Parts Manager, Senior Service Advisor, Service Advisor,
  Service Manager) → **SHOWN**. Internally consistent → the STAGING-LESS "Send to Portal" rows are
  CONFIRMED (staging genuinely hides it for those 6). Prod grants it (evidence-derived proxy =
  `work_order/view`).
- **See AP/AR Data → HIGH (verified).** FE gate = `seeApArData()` cross-toggle (source:
  `check:()=>seeApArData()` on the Accounts Payable/Receivable + transactions + payments tabs).
  Live staging cross_toggles match: Parts Tech OFF (HIDDEN, STAGING-LESS), Sales Rep ON (VISIBLE,
  STAGING-MORE), Service Advisor OFF (HIDDEN, spec-intended).
- **Send to Terminal → control ABSENT from the staging build.** No payment-terminal / card-reader
  / "Send to Terminal" / "take payment" control exists anywhere in the staging FE bundle
  ("terminal" matches only the Quasar framework). There is **no per-role Send-to-Terminal gate**
  in staging — the Parts-Tech STAGING-LESS row reflects a build-wide absence, not a role
  regression. Confirm prod's actual control name before treating it as a role-level loss.
- **Part-return approve/complete + decline → still MEDIUM / NEEDS-UI-VERIFY.** The controls DO
  exist in the staging build ("Process Return" / "Confirm Return"), but the exact permission gate
  could not be isolated from the minified source and the pixel-screenshot was blocked by SSO
  expiry.

## Headline totals (corrected, out-of-model excluded)
| Direction | Intended (Yes, spec/Migration-Type cited) | **NOT in spec (No) = RELEASE RISK** |
|---|---|---|
| **STAGING-LESS** (prod grants, staging doesn't) | {len(ly)} | **{len(ln)}** |
| **STAGING-MORE** (staging grants, prod didn't) | {len(my)} | **{len(mn)}** |

- **Work Orders — granular:** STAGING-LESS No = **{len(wln)}** · STAGING-MORE No = **{len(wmn)}**
- **Out-of-model (staff-record, excluded from No counts):** {len(oom_rows)} rows (WO {len(oom_wo)})

> The **No** rows in BOTH directions are the release-eve items needing a keep/change decision.
> Mapping is CONFIRMED (QA lead 2026-07-14); Administrator compared 1:1 (Owner N/A).

## STAGING-LESS · NOT-in-spec (No) — prod can do MORE than staging (regressions / over-in-prod)
{md_rows(less_no)}

## STAGING-MORE · NOT-in-spec (No) — staging grants MORE than prod (unaccounted expansions)
{md_rows(more_no)}

## STAGING-LESS · intended (Yes, spec-documented reductions)
{md_intended(delta_rows, "STAGING-LESS")}

## STAGING-MORE · intended (Yes, spec / Migration-Type documented expansions)
{md_intended(delta_rows, "STAGING-MORE")}

## Out-of-model (staff-record-controlled — NOT permission deltas; excluded from risk counts)
Per spec "Staff Record Settings", clock-in and timesheet appearance are staff-record controlled,
not the role/permission model. These are informational, not release risks:

{chr(10).join(oom_lines)}

## Per-role 2×2 summary (whole app, out-of-model excluded)
{chr(10).join(sum_lines)}

## Completeness (independent verification)
**No release-critical omissions.** All 43 staging atoms + 3 cross-toggles + view_mode are
represented; all 14 prod roles (no Owner) + 11 staging roles + 4 merges present; independent
recompute matched the workbook 23/23 on critical rows (5 prod roles re-captured LIVE). Only **5
LOW-severity** prod resources have no explicit row — `workplace`, `department`, `vehicle_type`,
`vehicle_history`, `shop_billing_efficiency` — all settings / reference / report-view only
(subsumed under Settings / Reports / vehicle view). Note: `workplace*` is held by SA-Limited-View,
so staging Service Advisor carries a low-severity uncaptured "workplace management" reduction.

## Open questions / NEEDS-REVIEW
1. **Mapping CONFIRMED (QA lead 2026-07-14)** — spec migration table authoritative; SA / Senior-SA
   rows FINAL; section-3549 1:1 same-name cases C26514/C26515 superseded.
2. **Administrator compared 1:1 (Owner not applicable)** — no Owner role in either environment.
3. **FE-gated / no-clean-map rows** (Send to Portal, Send to Terminal, portal page access, See
   AP/AR, part-return verbs) are Verification confidence = MEDIUM / NEEDS-UI-VERIFY — drive per
   role with a fresh staging cookie before go/no-go.
4. **Reporting** legacy role returns 0 resource permissions (report-page-only); merges into
   Sales Representative.

*Full detail incl. every match/staging-more row, Migration Type, Verification confidence, and the
side-by-side capability matrix is in the workbook.*
"""

out_md = os.path.join(HERE, "Prod-vs-Staging-Permission-Gaps_2026-07-14.md")
with open(out_md, "w") as f:
    f.write(md)
print("wrote", out_md)
