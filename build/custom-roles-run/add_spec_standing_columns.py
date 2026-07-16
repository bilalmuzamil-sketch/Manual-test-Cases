# add_spec_standing_columns.py — RUN-ONCE (2026-07-16). Appends the two annotation columns
# "Per Spec (v2)?" + "Per Standing Instructions?" to every role x capability x verdict tab of
# Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx. Sources: spec-conformance/spec-v2-permission-intent.md
# + spec-conformance/standing-permission-rules.md. NOT idempotent (re-running re-appends). The
# READ ME legend, Staging Live Grid note, and "Spec-Standing Conformance" summary tab were added by
# a companion one-shot step in the same session. Observed verdicts are never modified.
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

WB='build/custom-roles-run/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb=openpyxl.load_workbook(WB)

# ---- role normalization ----
def rk(s):
    if not s: return None
    l=str(s).lower()
    if 'time clock' in l: return 'timeclock'
    if 'sales representative' in l or 'sales rep' in l: return 'salesrep'
    if 'parts technician' in l or 'parts tech' in l: return 'pt'
    if 'parts manager' in l or 'parts mgr' in l: return 'pm'
    if 'senior service advisor' in l: return 'ssa'
    if 'service advisor' in l: return 'sa'
    if 'service manager' in l: return 'sm'
    if 'foreman' in l: return 'foreman'
    if 'office' in l: return 'office'
    if 'technician' in l: return 'tech'
    if l.strip() in ('admin','administrator') or 'administrator' in l or l.strip()=='admin': return 'admin'
    return None

RNAME={'admin':'Admin','sm':'Service Manager','ssa':'Senior Service Advisor','sa':'Service Advisor',
'foreman':'Foreman','tech':'Technician','pm':'Parts Manager','pt':'Parts Technician',
'office':'Office User','salesrep':'Sales Representative','timeclock':'Time Clock User'}

# ---- capability keys ----
def ck(s):
    if not s: return None
    l=str(s).lower()
    if 'send to terminal' in l: return 'terminal'
    if 'send to portal' in l: return 'portal'
    if 'approve/decline' in l: return 'approvedecline'
    if 'approve line' in l: return 'approve'
    if 'decline line' in l: return 'decline'
    if 'create/edit wo lines' in l or 'new line' in l: return 'newline'
    if 'review work orders' in l or 'reviewed' in l: return 'review'
    if 'see financial data' in l: return 'sfd'
    if 'invoicing/finance view' in l or 'finance tab' in l: return 'financeview'
    if 'take payment' in l: return 'takepayment'
    if 'see ap/ar' in l: return 'apar'
    if 'part return' in l: return 'partreturn'
    if l.startswith('finance') or 'new payment/reverse' in l: return 'financebundle'
    if 'wo-level history' in l or 'history tab' in l: return 'history'
    if 'order parts' in l: return 'orderparts'
    if 'wo notes' in l or 'notes tab' in l: return 'notes'
    if 'timesheets' in l: return 'timesheets'
    if 'wo delete' in l: return 'wodelete'
    if 'change customer' in l: return 'changecust'
    if 'change asset' in l: return 'changeasset'
    if 'receive parts' in l or 'accept a delivery' in l: return 'receive'
    if 'create work order' in l: return 'createwo'
    if 'create customer from new-wo' in l: return 'createcust'
    if 'create asset' in l: return 'createasset'
    return None

# ---- grant sets ----
ALL={'admin','sm','ssa','sa','foreman','tech','pm','pt','office','salesrep','timeclock'}
wo_ce={'admin','sm','ssa','sa','foreman','pm'}
wo_delete={'admin','sm','ssa'}
wol_ce={'admin','sm','ssa','sa','foreman','tech','pm'}
review={'admin','sm','ssa','sa','foreman','pm'}
sfd={'admin','sm','ssa','sa','foreman','pm','pt','office','salesrep'}
invoicing={'admin','sm','ssa','sa','foreman','pm','pt','office'}
order_parts={'admin','sm','ssa','sa','foreman','pm','pt'}
timesheets_v={'admin','sm','ssa','sa','foreman','pt','office','timeclock'}
apar={'admin','sm','ssa','pm','office','salesrep'}
vendor_ce={'admin','sm','ssa','sa','foreman','pm','pt'}
approve_s={'admin','sm','ssa','sa','foreman','pm'}
terminal_rg={'admin','sm','ssa','sa','pm'}
portal_grant={'admin','sm','ssa','sa','foreman','pm'}
portal_amb={'office','pt','salesrep'}

GRANT={'newline':wol_ce,'review':review,'sfd':sfd,'financeview':invoicing,'takepayment':invoicing,
'financebundle':invoicing,'history':wo_ce,'orderparts':order_parts,'notes':ALL,'timesheets':timesheets_v,
'wodelete':wo_delete,'changecust':wo_ce,'changeasset':wo_ce,'approve':approve_s,'approvedecline':approve_s,
'apar':apar,'partreturn':ALL,'receive':vendor_ce,'createwo':wo_ce,'createcust':wo_ce,'createasset':wo_ce,
'portal':portal_grant}

GATE={
'portal':"View Mode = Full View (spec §3/§4); Open Q6 also: 'anyone who can approve a WOL'",
'newline':"Work Order Lines -> Create & Edit (spec §1b/§13)",
'review':"WO sub-setting 'Review WOs' (spec §B WO Sub-Settings)",
'sfd':"See Financial Data toggle (spec §5a)",
'financeview':"Invoicing -> View; SFD gates Invoicing (spec §1i/§5a)",
'takepayment':"Invoicing -> Create & Edit (spec §1i, Open Q4)",
'financebundle':"Invoicing -> Create & Edit for New Payment (§1i); Reverse=WO Delete (28 Jun); Issue Credit=SPEC SILENT",
'terminal':"Invoicing -> Create & Edit + Customer Portal ON (spec §1i, 06 Jul) — build also adds an org-device gate",
'history':"Work Orders -> Create & Edit (spec §11, 07 Jul; WO-level audit log)",
'orderparts':"WO sub-setting 'Order Parts' (woOrderParts; needs WO View + SFD) (spec §1a/§7)",
'notes':"Work Orders -> View (create/edit any note) (spec §1a)",
'timesheets':"Timesheets -> View (spec §1j); clock-in/out + My Timesheets always on",
'wodelete':"Work Orders -> Delete (spec §1a)",
'changecust':"Work Orders -> Create & Edit (spec §1a/§14)",
'changeasset':"Work Orders -> Create & Edit (spec §1a/§14)",
'approve':"WO Lines -> Create & Edit + Full View (Open Q6)",
'decline':"SPEC SILENT (spec only says 'authorize lines')",
'approvedecline':"Approve = WOL Create&Edit + Full View; Decline = SPEC SILENT",
'apar':"Manage AP/AR toggle (tabs/fields, §5b) / AR-AP aging = Reports toggle (all-or-nothing, decoupled 3 Jul)",
'partreturn':"NO permission gate — return a WO part is available to everyone (29 Jun; §1a)",
'receive':"Vendor & Order Management -> Create & Edit (receive/return parts, spec §1g)",
'createwo':"Work Orders -> Create & Edit (spec §1a)",
'createcust':"Customer Management -> Create & Edit (create customer in New-WO flow, 01 Jun; behind WO Create&Edit)",
'createasset':"Customer Management -> Create & Edit (manage vehicles; behind WO Create&Edit)",
}

STAND={
'portal':"Consistent with standing rule (Rule 3.7 / §6): Send to Portal = Full View; enforcement model = FE display gate (view_mode). Standing rules cite bare Full View; build also requires WOL-approve (Open Q6).",
'newline':"Consistent with standing rule: add a new line = WO Lines Create&Edit (BE-enforced resource Edit).",
'review':"Consistent with enforcement model: 'Review WOs' is a WO sub-setting = FE display gate; no other standing rule addresses it.",
'sfd':"Consistent with standing rule: See Financial Data is the app-wide toggle gating Part Sales/Invoicing/Order Parts/Manage AP/AR (FE gate).",
'financeview':"Consistent with standing rule: Invoicing View gates the Finance tab; SFD gates Invoicing (enforcement: resource View).",
'takepayment':"Consistent with standing rule: payments = Invoicing Create&Edit; the Office-cannot-create-invoices hard-coded rule does NOT block Office from taking payment.",
'financebundle':"Consistent with standing rule: New Payment = Invoicing Create&Edit; Reverse Invoice = WO Delete (moved off Invoicing Delete, 28 Jun); no standing rule addresses Issue Credit.",
'terminal':"Consistent with standing rule (Rule 3.7): Send to Terminal = Invoicing C&E + Customer Portal ON; PLUS the build adds an org-device config gate. Enforcement: FE + org-config.",
'history':"Consistent with standing rule: History logs split WO-level (WO Create&Edit) vs line-level (WOL Create&Edit); WO-level = WO Create&Edit.",
'orderparts':"Consistent with standing rule: Order Parts requires See Financial Data and controls the WO Parts tab (FE sub-setting gate).",
'notes':"Consistent with standing rule: WO View = create/edit ANY note; WO Delete = delete any note (every role with WO View can see/add notes).",
'timesheets':"Consistent with standing rule: clock in/out + 'My Timesheets' always available regardless of Timesheets perm; the tab tracks Timesheets View (FE gate).",
'wodelete':"Consistent with standing rule: enforcement model — Delete is a FE display gate (BE enforces only View/Edit); WO Delete also gates Reverse Invoice.",
'changecust':"Consistent with standing rule: change customer/asset on an existing WO = WO Create&Edit (BE-enforced resource Edit).",
'changeasset':"Consistent with standing rule: change customer/asset on an existing WO = WO Create&Edit (BE-enforced resource Edit).",
'approve':"Consistent with standing rule: approve a line = WOL Create&Edit + Full View (FE gate); View Mode is UX not security, but approve is FE-gated by Full View.",
'decline':"No standing rule addresses Decline/Set Line Status as a discrete action (grouped with approve under WOL Create&Edit).",
'approvedecline':"Consistent with standing rule for Approve (WOL Create&Edit + Full View, FE gate); NO standing rule addresses Decline as a discrete action.",
'apar':"Consistent with standing rule: AP/AR tabs/fields = Manage AP/AR (requires SFD ON); AR/AP AGING reports FOLLOW the Reports permission (all-or-nothing), decoupled 3 Jul 2026 — NOT Manage AP/AR.",
'partreturn':"Consistent with standing rule: returning a part from a WO has NO permission gate — everyone (29 Jun 2026).",
'receive':"Consistent with enforcement model: receiving/returning parts to inventory/vendors = Vendor & Order Management Create&Edit (resource Edit). No more specific standing rule.",
'createwo':"Consistent with standing rule: create a WO = WO Create&Edit (BE-enforced resource Edit).",
'createcust':"Consistent with standing rule: Create/Edit customer also drives create-a-customer in the New-WO flow (Customer Management Create&Edit).",
'createasset':"Consistent with standing rule: manage vehicles = Customer Management Create&Edit.",
}

def direction_of(verdict):
    v=str(verdict or '').upper()
    if 'STAGING-LESS' in v: return 'LESS'
    if 'STAGING-MORE' in v: return 'MORE'
    if 'MATCH' in v: return 'MATCH'
    return None

def spec_verdict(cap,role,direction,staging_shown):
    rn=RNAME[role]
    if cap=='decline':
        return "Spec silent — Decline/Set Line Status not addressed (spec only says 'authorize lines'). No conformance inferred."
    if cap=='terminal':
        base=("Org-device config gate (NOT a role/spec delta): staging org has a card terminal, prod org has none. "
              "Spec role-gate = Invoicing C&E + Customer Portal ON.")
        if role in terminal_rg:
            return ("Per spec (role-gate met) + %s  %s meets the spec role-gate; the observed staging-vs-prod difference is ORG-CONFIG, not role/migration." % (base,rn))
        if role in {'foreman','office','pt'}:
            return ("DEVIATION (gating model) — spec requires Customer Portal ON, which %s has OFF, so the spec ROLE-gate would WITHHOLD Send to Terminal; the build instead gates on org-device presence and shows it. %s" % (rn,base))
        return ("Per spec (matches) — %s cannot open New Payment (no Invoicing C&E) so Send to Terminal is unreachable. %s" % (rn,base))
    if cap=='portal' and role in portal_amb:
        return ("Spec inconsistent/ambiguous — §3/§4 grant Send to Portal to ANY Full-View role (incl. %s), so a bare-Full-View reading would make this staging state NOT per spec; BUT Open Q6 conditions it on 'can approve a WOL', which %s lacks (no WOL Create&Edit) -> then per spec. Flagged, not resolved." % (rn,rn))
    note=""; eff=cap
    if cap=='approvedecline': eff='approve'; note=" (Decline component is SPEC SILENT.)"
    if cap=='financebundle': eff='financebundle'; note=" (Reverse=WO Delete; Issue Credit is SPEC SILENT.)"
    gset=GRANT[eff]; gate=GATE[cap]; granted=role in gset
    if direction=='LESS':
        if not granted: return "Per spec — expected reduction (spec does not grant to %s; gate: %s).%s"%(rn,gate,note)
        return "DEVIATION — spec GRANTS this to %s (gate: %s); staging loss is NOT per spec.%s"%(rn,gate,note)
    if direction=='MORE':
        if granted: return "Per spec — expected grant (%s holds the gate: %s).%s"%(rn,gate,note)
        return "DEVIATION — spec does NOT grant this to %s (gate: %s); possible over-grant.%s"%(rn,gate,note)
    # MATCH
    if granted and staging_shown: return "Per spec (matches) — spec grants to %s (gate: %s).%s"%(rn,gate,note)
    if (not granted) and (not staging_shown): return "Per spec (matches) — spec does not grant to %s (gate: %s).%s"%(rn,gate,note)
    if granted and not staging_shown: return "DEVIATION — spec grants to %s (gate: %s) but staging hides it (both envs hidden).%s"%(rn,gate,note)
    return "DEVIATION — staging grants to %s but spec does not (gate: %s) [both envs shown].%s"%(rn,gate,note)

def stand_verdict(cap,role):
    eff='approve' if cap=='approvedecline' else cap
    if cap=='approvedecline': return STAND['approvedecline']
    return STAND.get(cap,"No standing rule addresses this capability.")

def is_shown(cell):
    return str(cell or '').strip().lower().startswith('shown')

HDR_FONT=Font(bold=True)
WRAP=Alignment(wrap_text=True,vertical='top')
NOTEFILL=PatternFill('solid',fgColor='FFF2CC')

stats={'perspec':0,'deviation':0,'silent':0,'inconsistent':0,'orgconfig':0}
deviations=[]  # (tab,role,cap,direction,text)
conformance_rows=[]  # for md mirror of directional rows

def classify(text):
    t=text
    if t.startswith('DEVIATION'): return 'deviation'
    if t.startswith('Spec silent'): return 'silent'
    if t.startswith('Spec inconsistent') or t.startswith('Spec ambiguous'): return 'inconsistent'
    if t.startswith('Org-device'): return 'orgconfig'
    return 'perspec'

CAPLABEL={'portal':'Send to Portal','newline':'Create/Edit WO Lines','review':'Review WOs','sfd':'See Financial Data',
'financeview':'Finance view','takepayment':'Take Payment','financebundle':'Finance (Payment/Reverse/Credit)',
'terminal':'Send to Terminal','history':'WO-level History','orderparts':'Order Parts','notes':'WO Notes',
'timesheets':'Timesheets','wodelete':'WO Delete','changecust':'Change Customer','changeasset':'Change Asset',
'approve':'Approve line','decline':'Decline line','approvedecline':'Approve/Decline line','apar':'See AP/AR',
'partreturn':'Part Return','receive':'Receive parts','createwo':'Create WO','createcust':'Create Customer (New-WO)',
'createasset':'Create Asset (New-WO)'}

def annotate(tabname, header_row, role_col, cap_col, verdict_col, staging_col, prod_col, c1, c2, fixed_cap=None):
    ws=wb[tabname]
    ws.cell(header_row,c1,"Per Spec (v2)?").font=HDR_FONT
    ws.cell(header_row,c2,"Per Standing Instructions?").font=HDR_FONT
    ws.column_dimensions[openpyxl.utils.get_column_letter(c1)].width=60
    ws.column_dimensions[openpyxl.utils.get_column_letter(c2)].width=60
    n=0
    for r in range(header_row+1, ws.max_row+1):
        role=ws.cell(r,role_col).value
        cap=ws.cell(r,cap_col).value if cap_col else None
        rr=rk(role); cc=fixed_cap if fixed_cap else ck(cap)
        if not rr or not cc: continue
        verdict=ws.cell(r,verdict_col).value
        direction=direction_of(verdict)
        if direction is None: continue
        sc = ws.cell(r,staging_col).value if staging_col else None
        staging_shown=is_shown(sc)
        sv=spec_verdict(cc,rr,direction,staging_shown)
        st=stand_verdict(cc,rr)
        a=ws.cell(r,c1,sv); a.alignment=WRAP
        b=ws.cell(r,c2,st); b.alignment=WRAP
        cls=classify(sv); stats[cls]+=1
        if cls=='deviation':
            deviations.append((tabname,RNAME[rr],CAPLABEL.get(cc,cc),direction or '', sv))
        if direction in ('LESS','MORE'):
            conformance_rows.append((RNAME[rr],CAPLABEL.get(cc,cc),direction,cls,sv))
        n+=1
    # legend note row at bottom
    lr=ws.max_row+2
    ws.cell(lr,1,"LEGEND — 'Per Spec (v2)?': Per spec-expected/matches | DEVIATION (spec grants/withholds opposite of staging) | Spec silent (not addressed) | Spec inconsistent/ambiguous | Org-device config gate (out of role model). 'Per Standing Instructions?': Consistent / Conflicts / No rule — cites the specific standing rule. Existing observed verdicts UNCHANGED; columns are ADDED.").font=Font(italic=True,size=9)
    ws.cell(lr,1).fill=NOTEFILL
    return n

total=0
total+=annotate('Full Dual Matrix',1,1,3,6,5,4,8,9)
total+=annotate('Pass-11 LIVE (2026-07-16)',3,1,2,5,4,3,7,8)
total+=annotate('Pass-12 LIVE (2026-07-16)',4,1,2,5,4,3,7,8)
total+=annotate('Approve-Decline LIVE',3,1,None,4,3,2,6,7,fixed_cap='approvedecline')
total+=annotate('Send to Terminal LIVE',7,1,None,4,2,3,5,6,fixed_cap='terminal')
total+=annotate('Parts-Module Dual LIVE',2,1,3,6,5,4,12,13)
total+=annotate('New-WO Create Dual LIVE',2,1,3,6,5,4,12,13)

print("STATS",stats)
print("TOTAL annotated",total)
print("DEVIATIONS",len(deviations))
import json
json.dump({'stats':stats,'total':total,'deviations':deviations,'conformance':conformance_rows},open('/tmp/annot_out.json','w'))
wb.save(WB)
print("saved")
