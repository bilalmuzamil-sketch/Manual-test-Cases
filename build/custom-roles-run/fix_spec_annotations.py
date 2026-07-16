# fix_spec_annotations.py — RE-AUDIT CORRECTOR (2026-07-16).
# The "Per Spec (v2)?" annotations written by add_spec_standing_columns.py contained errors
# (user-reported, trust-critical). This one-shot script re-derives every annotation from the
# CANONICAL spec (build/custom-roles-spec-update/current-spec-2026-07-15.md — incl. the 7/14
# "Updated Office Role definition": Office Work Orders V -> —, Part Sales V -> —, Invoicing
# V -> V/E/D) and applies the corrections in place. Truth table + full derivation:
# build/custom-roles-run/spec-conformance/spec-truth-table.md
#
# Corrections applied (classes + citations):
#  1. Office User WO Notes (Full Dual Matrix)  : Per-spec  -> DEVIATION  (Office has NO WO access since 7/14)
#  2. Office User Part Return (Pass-11)        : Per-spec  -> DEVIATION (reachability)
#  3. Decline line (Full Dual Matrix, 11 rows) : Spec-silent -> Per-spec via §1b "authorize lines"
#     (WOL Create & Edit); Technician -> Spec inconsistent/ambiguous (§4 Tech View blocks only Approve)
#  4. Service Manager Finance/Reverse (Pass-11): Per-spec -> Spec inconsistent (matrix+28-Jun say SM CAN
#     reverse; migration Behavior-Changes table says "Loses Invoicing Delete (cannot reverse)")
#  5. New-WO Create Customer/Asset for Office/Parts Tech/Sales Rep: reason text corrected — those roles
#     DO hold Customer Management C&E; the control is absent because the New-WO dialog needs WO C&E
#  6. AP/AR aging rows (Pass-11): citation corrected to lead with the operative Reports gate (§2a)
#  7. Finance rows (Pass-11): per-component gates spelled out (New Payment=Invoicing C&E; Reverse=WO
#     Delete; Issue Credit=spec-silent component)
#  8. Combined Approve/Decline rows (Pass-12 + Approve-Decline LIVE): Decline no longer "SPEC SILENT"
#  9. Office SFD / Finance-view / Take-Payment / Timesheets rows: 14-Jul WO-entry-point caveat appended
# Summary tab, READ ME vocabulary + headline tallies rewritten. Observed verdicts NEVER modified.
import openpyxl, json
from openpyxl.styles import Font, Alignment

WB='build/custom-roles-run/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb=openpyxl.load_workbook(WB)
WRAP=Alignment(wrap_text=True,vertical='top')
diffs=[]

def setcell(tab,row,col,new,role,cap):
    ws=wb[tab]
    old=ws.cell(row,col).value
    if old!=new:
        c=ws.cell(row,col,new); c.alignment=WRAP
        diffs.append({'tab':tab,'row':row,'role':role,'cap':cap,'old':old,'new':new})

# ---------------- 1+2: Office WO Notes + Part Return -> DEVIATION ----------------
OFFICE_NOTES=("DEVIATION — the current spec (14 Jul 'Updated Office Role definition') gives Office User NO "
 "Work Orders access at all (Permission Matrix: Work Orders = '—'), so the Notes tab (gated by Work Orders -> "
 "View, spec §1a) should NOT be reachable for Office User; BOTH envs show it (both SHOWN — a MATCH row that "
 "disagrees with the spec). Staging does not implement the 14-Jul Office reduction. (Spec wrinkle also flagged: "
 "the matrix WO Lines row still shows 'V' for Office although §1b/matrix-note say WOL View is inherited from "
 "Work Orders View.) [Corrected 2026-07-16 — previously wrongly 'Per spec (matches)': the earlier extract "
 "carried the pre-7/14 Office column (Work Orders = V).]")
setcell('Full Dual Matrix',155,8,OFFICE_NOTES,'Office User','WO Notes (Notes tab)')

OFFICE_RETURN=("DEVIATION (reachability) — the Return control itself has NO permission gate (spec §1a: "
 "'Returning a part from a WOL does not require a permission'; 29 Jun: 'Everyone has access to Return a part "
 "from a WO'), BUT §1a adds 'In practice, the user will need WO view so they can see the point', and the "
 "current matrix (14 Jul Office update) gives Office User NO Work Orders access ('—'). So per the current spec "
 "Office User cannot reach a WO part to return; staging shows it (STAGING-MORE) = beyond the current spec. Same "
 "root cause as the Office WO-visibility deviation (see the Office User WO Notes row). [Corrected 2026-07-16 — "
 "previously wrongly 'Per spec — expected grant'.]")
setcell('Pass-11 LIVE (2026-07-16)',25,7,OFFICE_RETURN,'Office User','Part Return')

# ---------------- 3: Decline line rows (Full Dual Matrix col 8) ----------------
DECL_GATE=("Work Order Lines -> Create & Edit — spec §1b 'authorize lines' covers declining a pending line "
 "(corrected 2026-07-16: previously mislabeled 'spec silent')")
decl_rows={17:('Admin','granted_shown'),33:('Service Manager','granted_shown'),
 49:('Senior Service Advisor','granted_shown'),65:('Service Advisor','granted_shown'),
 81:('Foreman','granted_shown'),97:('Technician','tech'),113:('Parts Manager','more'),
 129:('Parts Technician','not_granted'),145:('Sales Representative','not_granted'),
 161:('Office User','not_granted'),177:('Time Clock User','not_granted')}
for r,(rn,kind) in decl_rows.items():
    if kind=='granted_shown':
        t="Per spec (matches) — spec grants to %s (gate: %s)."%(rn,DECL_GATE)
    elif kind=='more':
        t="Per spec — expected grant (%s holds the gate: %s)."%(rn,DECL_GATE)
    elif kind=='not_granted':
        t=("Per spec (matches) — spec does not grant to %s (gate: %s; %s has no WOL Create & Edit in the "
           "matrix)."%(rn,DECL_GATE,rn))
    else: # technician
        t=("Spec inconsistent/ambiguous — §1b grants Technician WO Lines Create & Edit, whose scope includes "
           "'authorize lines' (declining is part of authorizing), but the §4 Tech View restrictions block only "
           "Approve ('Cannot approve lines' / 'No approve action') and never mention Decline. Whether Tech View "
           "also withholds Decline is not resolvable from the spec; both envs hide it for Technician. Flagged, "
           "not inferred. [Corrected 2026-07-16 — previously blanket 'spec silent'.]")
    setcell('Full Dual Matrix',r,8,t,rn,'Decline line')

# ---------------- 8: combined Approve/Decline rows ----------------
OLD_FRAG="gate: Approve = WOL Create&Edit + Full View; Decline = SPEC SILENT). (Decline component is SPEC SILENT.)"
NEW_FRAG=("gate: Approve = WOL Create & Edit + Full View (§1b 'authorize lines' + §4 Tech-View approve-block / "
 "Open Q6); Decline = WOL Create & Edit (§1b 'authorize lines' — NOT spec-silent; corrected 2026-07-16)).")
TECH_NOTE=(" NOTE (Technician): §4 Tech View blocks only Approve and never mentions Decline while §1b grants "
 "Technician WOL Create & Edit — whether Tech View also withholds Decline is spec-ambiguous; both envs hide it "
 "(see the Full Dual Matrix Technician 'Decline line' row, classed Spec inconsistent/ambiguous).")
for tab,col,rows,techrow in (('Pass-12 LIVE (2026-07-16)',7,range(5,16),10),
                             ('Approve-Decline LIVE',6,range(4,15),9)):
    ws=wb[tab]
    for r in rows:
        cur=ws.cell(r,col).value
        if not cur or OLD_FRAG not in cur: continue
        rn=str(ws.cell(r,1).value)
        t=cur.replace(OLD_FRAG,NEW_FRAG)
        if r==techrow: t+=TECH_NOTE
        setcell(tab,r,col,t,rn,'Approve/Decline line')

# ---------------- 6: AP/AR aging rows (Pass-11 rows 4-14, col 7) ----------------
APAR_GATE=("AR/AP AGING reports follow the Reports toggle (§2a: 'a user with Reports ON sees all reports, "
 "including AR/AP aging, regardless of Manage AP/AR'; all-or-nothing, decoupled 3 Jul) — NOT Manage AP/AR. "
 "Manage AP/AR (§5b) separately gates the Customer/Vendor Unpaid-Invoices/Payments/Credits tabs + sensitive "
 "fields; in the matrix the SAME six roles hold both (Admin/Service Manager/Senior Service Advisor/Parts "
 "Manager/Office/Sales Rep). Citation corrected 2026-07-16 to lead with the operative Reports gate")
apar_rows={4:('Admin','match_shown'),5:('Office User','match_shown'),6:('Service Manager','more'),
 7:('Parts Manager','more'),8:('Sales Representative','more'),9:('Senior Service Advisor','dev'),
 10:('Service Advisor','match_hidden'),11:('Foreman','match_hidden'),12:('Technician','match_hidden'),
 13:('Parts Technician','match_hidden'),14:('Time Clock User','match_hidden')}
for r,(rn,kind) in apar_rows.items():
    if kind=='match_shown':
        t="Per spec (matches) — spec grants to %s: Reports = ON in the matrix (gate: %s)."%(rn,APAR_GATE)
    elif kind=='more':
        t="Per spec — expected grant (%s holds Reports = ON in the matrix; gate: %s)."%(rn,APAR_GATE)
    elif kind=='match_hidden':
        t="Per spec (matches) — spec does not grant to %s: Reports = OFF in the matrix (gate: %s)."%(rn,APAR_GATE)
    else:
        t=("DEVIATION — spec grants the AR/AP aging reports to Senior Service Advisor (Reports = ON in the "
           "matrix — §2a: aging reports follow Reports; Manage AP/AR is ALSO ON, §5b) but staging hides them "
           "(both envs hidden — a MATCH row that disagrees with the spec). Staging does not implement the spec "
           "grant for Senior Service Advisor. (Re-verified from the matrix 2026-07-16; citation corrected to "
           "lead with the operative Reports gate.)")
    setcell('Pass-11 LIVE (2026-07-16)',r,7,t,rn,'See AP/AR (aging reports)')

# ---------------- 4+7: Finance rows (Pass-11 rows 28-35, col 7) ----------------
TAIL="Issue Credit = SPEC SILENT (no issuance gate anywhere in the spec — only Credits-TAB visibility, §5b)."
fin={28:('Admin',"Per spec (matches) — New Payment = Invoicing -> Create & Edit (§1i, Open Q4), which Admin "
        "holds; Reverse Invoice = Work Orders -> Delete (28 Jun), which Admin holds (Reverse shown per spec). "+TAIL),
 29:('Service Manager',"Spec inconsistent — on the Reverse component (the New Payment grant itself is per "
        "spec): New Payment = Invoicing -> Create & Edit (§1i), which Service Manager holds (expected grant, "
        "STAGING-MORE). Reverse Invoice = Work Orders -> Delete (28 Jun) and the matrix gives Service Manager "
        "WO V/E/D (would GRANT Reverse), BUT the migration 'Behavior Changes' table says Service Manager "
        "'Loses Invoicing Delete (cannot reverse)'. Staging shows only New Payment + Issue Credit (no Reverse) "
        "— matching the migration table and contradicting the matrix + 28-Jun gate. Flagged, not resolved. "
        "[Corrected 2026-07-16 — previously 'Per spec — expected grant' with no Reverse flag.] "+TAIL),
 30:('Parts Manager',"Per spec — expected grant (Parts Manager holds Invoicing -> Create & Edit, §1i = New "
        "Payment); Reverse correctly absent (Parts Manager has no Work Orders -> Delete; Reverse = WO Delete, "
        "28 Jun). "+TAIL),
 31:('Parts Technician',"Per spec — expected grant (Parts Technician holds Invoicing -> Create & Edit, §1i = "
        "New Payment); Reverse correctly absent (no Work Orders -> Delete; Reverse = WO Delete, 28 Jun). "+TAIL),
 32:('Foreman',"Per spec — expected grant (Foreman holds Invoicing -> Create & Edit, §1i = New Payment); "
        "Reverse correctly absent (no Work Orders -> Delete; Reverse = WO Delete, 28 Jun). "+TAIL),
 33:('Office User',"Per spec — expected grant (Office User holds Invoicing -> Create & Edit — §1i; and the "
        "spec explicitly says Office users 'are expected to be able to make payments but not create invoices'); "
        "Reverse correctly absent (no Work Orders -> Delete). NOTE (14-Jul Office update): the matrix removed "
        "ALL Work Orders access for Office User, so per the current spec this Finance surface is reached via "
        "Customers/Part Sales entry points (§1i View), not the work-order screen where it was observed — the "
        "Office WO-visibility delta is logged as the Office WO Notes / Part Return DEVIATIONs. "+TAIL),
 34:('Senior Service Advisor',"Per spec (matches) — New Payment = Invoicing -> Create & Edit (§1i), which "
        "Senior Service Advisor holds; Reverse Invoice = Work Orders -> Delete (28 Jun), which Senior Service "
        "Advisor holds (Reverse shown per spec; migration table confirms 'Invoicing FULL' expansion). "+TAIL),
 35:('Service Advisor',"Per spec (matches) — New Payment = Invoicing -> Create & Edit (§1i), which Service "
        "Advisor holds; Reverse Invoice = Work Orders -> Delete (28 Jun), which Service Advisor does NOT hold "
        "(WO = V/E) — staging correctly hides Reverse (prod SA-Limited-View showing Reverse is the legacy "
        "model). "+TAIL)}
for r,(rn,t) in fin.items():
    setcell('Pass-11 LIVE (2026-07-16)',r,7,t,rn,'Finance (New Payment/Reverse/Issue Credit)')

# ---------------- 5: New-WO Create Customer/Asset for Office/PT/SalesRep ----------------
def newwo_fix(rn,asset,office):
    extra=" — indeed the 14-Jul update removed ALL Work Orders access for Office User" if office else ""
    what="Create Asset ('manage vehicles', §1d Edit)" if asset else "Create Customer in the New-WO flow (01 Jun change log; §1d Edit)"
    return ("Per spec (matches) — the control is correctly absent, but the REASON is corrected (2026-07-16): "
        "%s DOES hold Customer Management -> Create & Edit in the matrix (Customers row), so the earlier 'spec "
        "does not grant Customer Management C&E' reading was WRONG. The Add control is unreachable per spec "
        "because %s lacks Work Orders -> Create & Edit%s, so the New Work Order dialog itself never opens. "
        "Compound gate: Work Orders C&E (reach the New-WO dialog) + Customer Management C&E (%s)."%(rn,rn,extra,what))
for r,rn,asset in ((22,'Office User',False),(23,'Office User',True),(25,'Parts Technician',False),
                   (26,'Parts Technician',True),(28,'Sales Representative',False),(29,'Sales Representative',True)):
    setcell('New-WO Create Dual LIVE',r,12,newwo_fix(rn,asset,rn=='Office User'),rn,
            'Create Asset (New-WO)' if asset else 'Create Customer (New-WO)')

# ---------------- 9: Office WO-entry-point caveat (Full Dual rows 149,150,151,156) ----------------
CAVEAT=(" NOTE (14-Jul 'Updated Office Role definition'): the matrix now gives Office User NO Work Orders "
 "access ('—'), so per the current spec this surface would be reached via Customers/Part Sales entry points "
 "rather than the work-order screen where it was observed; the Office WO-visibility delta itself is logged as "
 "the Office WO Notes / Part Return DEVIATIONs (this row's own gate — which Office User holds — is unchanged).")
ws=wb['Full Dual Matrix']
for r,cap in ((149,'See Financial Data'),(150,'Invoicing/Finance view'),(151,'Take Payment'),(156,'Timesheets (tab)')):
    cur=ws.cell(r,8).value or ''
    if CAVEAT not in cur:
        setcell('Full Dual Matrix',r,8,cur+CAVEAT,'Office User',cap)

# ---------------- Rebuild the Spec-Standing Conformance tab (below row 4) ----------------
ws=wb['Spec-Standing Conformance']
for r in range(5,46):
    for c in range(1,6): ws.cell(r,c).value=None  # NB: ws.cell(r,c,None) would NOT clear
B=Font(bold=True)
def w(r,c,v,bold=False):
    cell=ws.cell(r,c); cell.value=v; cell.alignment=WRAP
    if bold: cell.font=B
w(5,1,"TALLY (297 annotated rows across 7 tabs) — RE-AUDITED + CORRECTED 2026-07-16 (every annotation re-derived from scratch against current-spec-2026-07-15.md; derivation + diff: spec-conformance/spec-truth-table.md):",True)
w(6,1,"Bucket",True); w(6,2,"Count",True); w(6,3,"Meaning",True)
w(7,1,"Per spec — expected / matches"); w(7,2,283); w(7,3,"STAGING grant/withhold agrees with the spec gate (intended). Includes the 10 Decline rows re-derived via §1b 'authorize lines' (previously mislabeled spec-silent).")
w(8,1,"DEVIATION"); w(8,2,9); w(8,3,"STAGING is opposite of what the spec prescribes (incl. gating-model divergence and MATCH rows where BOTH envs disagree with the spec).")
w(9,1,"Spec silent — not addressed"); w(9,2,0); w(9,3,"Row-level: none after the re-audit. Issue Credit remains spec-silent as a COMPONENT inside the Finance rows; remove-a-WO-part as a discrete atom has no annotated row.")
w(10,1,"Spec inconsistent / ambiguous"); w(10,2,5); w(10,3,"The spec contradicts itself for the role/capability — flagged, never resolved by inference.")
w(12,1,"KEY SIGNAL: the migration is LARGELY SPEC-ACCURATE — every STAGING-LESS loss and nearly every STAGING-MORE grant maps to an intended spec gate. Non-per-spec rows below.",True)
w(13,1,"Release-relevant DEVIATIONS (9 rows):",True)
w(14,1,"Tab",True); w(14,2,"Role",True); w(14,3,"Capability",True); w(14,4,"Direction",True); w(14,5,"Why (spec judgement)",True)
TERM=("DEVIATION (gating model) — spec requires Customer Portal ON, which %s has OFF, so the spec ROLE-gate would WITHHOLD Send to Terminal; the build instead gates on org-device presence and shows it. Org-config, not a role over-grant. Spec role-gate = Invoicing C&E + Customer Portal ON (§1i, 06 Jul).")
rows=[('Full Dual Matrix','Foreman','Send to Terminal','MORE',TERM%'Foreman'),
 ('Full Dual Matrix','Parts Technician','Send to Terminal','MORE',TERM%'Parts Technician'),
 ('Full Dual Matrix','Office User','Send to Terminal','MORE',TERM%'Office User'),
 ('Send to Terminal LIVE','Foreman','Send to Terminal','MORE',TERM%'Foreman'),
 ('Send to Terminal LIVE','Office User','Send to Terminal','MORE',TERM%'Office User'),
 ('Send to Terminal LIVE','Parts Technician','Send to Terminal','MORE',TERM%'Parts Technician'),
 ('Pass-11 LIVE (2026-07-16)','Senior Service Advisor','See AP/AR (aging reports)','MATCH (both hidden)',
  "DEVIATION — spec grants (Reports = ON in the matrix; §2a aging follows Reports; Manage AP/AR also ON) but staging hides them — staging does not implement the spec grant."),
 ('Full Dual Matrix','Office User','WO Notes (Notes tab)','MATCH (both SHOWN)',
  "DEVIATION — 14-Jul spec update removed ALL Work Orders access for Office User (WO = '—'), so the Notes tab should be unreachable; BOTH envs show it. [Corrected 2026-07-16 — was 'per spec'.]"),
 ('Pass-11 LIVE (2026-07-16)','Office User','Part Return','MORE',
  "DEVIATION (reachability) — Return itself is ungated (29 Jun; §1a) but §1a requires WO view in practice and the 14-Jul matrix gives Office User NO Work Orders access; staging shows it. [Corrected 2026-07-16 — was 'per spec'.]")]
for i,row in enumerate(rows):
    for c,v in enumerate(row,1): w(15+i,c,v)
w(25,1,"SPEC-INCONSISTENT / AMBIGUOUS rows (5 — flagged, not resolved, never inferred):",True)
w(26,1,"Tab",True); w(26,2,"Role",True); w(26,3,"Capability",True); w(26,4,"Direction",True); w(26,5,"Note",True)
inc=[('Full Dual Matrix','Parts Technician','Send to Portal','LESS',
  "§3/§4 bare Full-View would grant; Open Q6 'can approve a WOL' would withhold (no WOL Create&Edit). Flagged."),
 ('Full Dual Matrix','Office User','Send to Portal','LESS',
  "§3/§4 bare Full-View would grant; Open Q6 'can approve a WOL' would withhold (no WOL Create&Edit). Flagged."),
 ('Full Dual Matrix','Sales Representative','Send to Portal','MATCH',
  "Same §3/§4-vs-Open-Q6 ambiguity; both envs hidden."),
 ('Full Dual Matrix','Technician','Decline line','MATCH (both hidden)',
  "§1b grants Technician WOL Create & Edit ('authorize lines' covers declining) but §4 Tech View blocks only Approve and never mentions Decline — Tech-View treatment of Decline unresolvable. [Re-derived 2026-07-16 — was blanket 'spec silent'.]"),
 ('Pass-11 LIVE (2026-07-16)','Service Manager','Finance — Reverse Invoice component','MORE (bundle)',
  "Matrix (SM WO V/E/D) + 28-Jun rule (Reverse = WO Delete) say SM CAN reverse; migration Behavior-Changes table says SM 'Loses Invoicing Delete (cannot reverse)'; staging shows New Payment + Issue Credit only (no Reverse) — matches the migration table, contradicts the matrix. [Flagged 2026-07-16.]")]
for i,row in enumerate(inc):
    for c,v in enumerate(row,1): w(27+i,c,v)
w(33,1,"SPEC-SILENT components (conformance NOT inferred): Issue Credit (inside the 'Finance' rows — no issuance gate anywhere in the spec); remove-a-WO-part as a discrete atom (spec covers only: return = no gate, move = WOL C&E, remove-lines = WOL Delete). 'Decline line' is NO LONGER spec-silent — §1b 'authorize lines' addresses it (re-derived 2026-07-16).")
w(35,1,"RE-AUDIT ROOT CAUSE (2026-07-16): the earlier extract (spec-conformance/spec-v2-permission-intent.md §B) carried the PRE-7/14 Office column (Work Orders = V, Part Sales = V; WO-Lines slip '—'), so the generator treated Office as holding WO View -> Office WO Notes + Part Return were wrongly 'per spec'. The generator also blanket-labeled Decline spec-silent, reused the WO-C&E grant set for the New-WO Create-Customer/Asset reason text, missed the SM Reverse migration-table contradiction, and cited Manage AP/AR ahead of the operative Reports gate for the aging-reports surface. All corrected in place; truth table = spec-conformance/spec-truth-table.md.")

# ---------------- READ ME updates ----------------
ws=wb['READ ME - Coverage & Honesty']
for r in range(1,ws.max_row+1):
    v=ws.cell(r,1).value
    if not v: continue
    if "'Spec silent — not addressed'" in v:
        ws.cell(r,1,"  * 'Spec silent — not addressed'   : row-level NONE after the 2026-07-16 re-audit (Issue Credit stays spec-silent as a component inside the Finance rows; remove-a-WO-part atom has no annotated row). NOT inferred.")
    if "'Spec inconsistent/ambiguous'" in v and 'Send to Portal' in v:
        ws.cell(r,1,"  * 'Spec inconsistent/ambiguous'   : the spec contradicts itself for this role/capability (Send to Portal §3/§4 Full-View vs Open Q6 'can approve a WOL' for Office/Parts-Tech/Sales-Rep;")
    if v.strip().startswith('Core OK/Not-OK'):
        ws.cell(r,1,"                                      Technician Decline — §1b 'authorize lines' vs §4 approve-only Tech View; Service Manager Reverse — matrix+28-Jun vs migration-table 'cannot reverse'). Flagged, never resolved by inference.")
    if v.startswith('HEADLINE CONFORMANCE RESULT'):
        hl=r
NEW_HEAD=[
"HEADLINE CONFORMANCE RESULT (297 role x capability rows annotated across 7 tabs) — RE-AUDITED + CORRECTED 2026-07-16:",
"  Per-spec (expected/matches): 283   |   DEVIATION: 9   |   Spec silent: 0 (row-level)   |   Spec inconsistent/ambiguous: 5",
"  => The migration is LARGELY SPEC-ACCURATE: every STAGING-LESS loss and nearly every STAGING-MORE grant maps to an intended spec gate. The non-per-spec rows:",
"  (1) Send to Terminal for Foreman/Office User/Parts Technician = DEVIATION (gating model): build shows it via the org-device gate although the spec",
"      role-gate requires Customer Portal ON (those 3 roles have it OFF). Org-config, not a role over-grant. (6 rows across 2 tabs.)",
"  (2) See AP/AR aging for Senior Service Advisor = DEVIATION: spec grants (Reports ON; Manage AP/AR also ON) but BOTH envs hide it — staging does not implement the grant.",
"  (3) Office User WO Notes (both envs SHOWN) + Part Return (STAGING-MORE) = DEVIATION: the 14-Jul spec update removed ALL Work Orders access for Office User",
"      (Work Orders = '—'), so neither surface should be reachable for Office; staging (and prod) still show them. [Corrected 2026-07-16 — previously wrongly 'per spec'.]",
"  (4) Spec inconsistent/ambiguous (5): Send to Portal for Office User/Parts Tech (STAGING-LESS) + Sales Rep (MATCH) — Full-View vs Open Q6; Technician Decline —",
"      §1b 'authorize lines' vs §4 approve-only Tech View; Service Manager Reverse Invoice — matrix+28-Jun (can reverse) vs migration table ('cannot reverse').",
"",
"RE-AUDIT NOTE (2026-07-16): the 'Per Spec (v2)?' column was re-derived FROM SCRATCH against the canonical spec (current-spec-2026-07-15.md) after annotation",
"  errors were found. Root causes fixed: the spec extract carried the pre-7/14 Office column (WO=V) -> Office WO-Notes/Part-Return were wrongly 'per spec';",
"  Decline was wrongly blanket 'spec silent' (§1b 'authorize lines' addresses it); the New-WO Create-Customer/Asset rows cited a wrong gate reason (Office/",
"  Parts-Tech/Sales-Rep DO hold Customer Mgmt C&E — the control is absent for lack of WO C&E); AP/AR-aging citations now lead with the operative Reports gate;",
"  SM Reverse-Invoice flagged spec-inconsistent. Full role x gate truth table + diff: build/custom-roles-run/spec-conformance/spec-truth-table.md."]
for i,line in enumerate(NEW_HEAD):
    ws.cell(hl+i,1,line)

wb.save(WB)
json.dump(diffs,open('/tmp/annot_diffs.json','w'),indent=1)
print("cell corrections:",len(diffs))
for d in diffs: print("-",d['tab'],'r%d'%d['row'],d['role'],'|',d['cap'])
