#!/usr/bin/env python3
"""
Append a 'Parts-Module Dual LIVE' sheet to the LIVE-VERIFIED workbook.
Order Parts (New PO) + Receive, observed live per role on both envs (2026-07-15).
Staging: switch-user impersonation x7 + throwaway role-swap+switch-user x4 (throwaway restored).
Production: test-staff role-swap+self-login x14 (test staff restored to Office User).
Observed-only (Rules 10 & 12) — every row is DUAL LIVE-OBSERVED.
"""
import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = '/home/user/Manual-test-Cases/build/custom-roles-run'
STG = f'{BASE}/live-ui-2026-07-15/staging'
PROD = f'{BASE}/live-ui-2026-07-15/production'

def load(p):
    return json.load(open(p))

stg = {}
for d in os.listdir(STG):
    f = f'{STG}/{d}/parts-obs.json'
    if os.path.exists(f):
        o = load(f); stg[o['role']] = o
prod = {}
for d in os.listdir(PROD):
    f = f'{PROD}/{d}/prod-parts-obs.json'
    if os.path.exists(f):
        o = load(f); prod[o['role']] = o

# staging role -> (prod primary key, merge-component prod keys, note)
MAP = {
 'Admin': ('Administrator', [], ''),
 'Service_Manager': ('Service_Manager', [], ''),
 'Senior_Service_Advisor': ('Service_Advisor', ['SA_Technician', 'SA_No_Reports'], 'prod merge: Service Advisor (+ SA Technician, SA No Reports)'),
 'Parts_Manager': ('Parts_Manager', [], ''),
 'Service_Advisor': ('SA_Limited_View', [], 'confirmed mapping: staging Service Advisor <- prod SA Limited View'),
 'Foreman': ('Foreman', [], 'prod: Parts nav hidden but /parts/orders reachable directly'),
 'Office_User': ('Office_User', [], 'both envs: /parts/orders page viewable, but no New PO / no Receive'),
 'Parts_Technician': ('Parts_Technician', [], ''),
 'Sales_Representative': ('Sales_Representative', ['Reporting'], 'prod merge: Sales Representative (+ Reporting)'),
 'Technician': ('Technician', [], ''),
 'Time_Clock_User': ('Time_Clock_User', [], ''),
}
ORDER = ['Admin','Service_Manager','Senior_Service_Advisor','Parts_Manager','Service_Advisor',
         'Foreman','Office_User','Parts_Technician','Sales_Representative','Technician','Time_Clock_User']
NICE = lambda s: s.replace('_', ' ')

CAPS = [
 ('Order Parts (create Purchase Order — "New PO" on /parts/orders)', 'newPO',
  'Story: Order Parts controls the ability to create a purchase order. FE-gated.'),
 ('Receive parts / accept a delivery ("Receive" on /parts/orders)', 'receive',
  'Pick/Receive against an ordered PO. FE-gated; per-PO Receive button.'),
]

wb = load_workbook(f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx')
if 'Parts-Module Dual LIVE' in wb.sheetnames:
    del wb['Parts-Module Dual LIVE']
ws = wb.create_sheet('Parts-Module Dual LIVE')

H = Font(bold=True, color='FFFFFF'); HF = PatternFill('solid', fgColor='2F5496')
OK = PatternFill('solid', fgColor='C6EFCE'); YEL = PatternFill('solid', fgColor='FFF2CC')
ORG = PatternFill('solid', fgColor='FCE4D6'); NV = PatternFill('solid', fgColor='FFC7CE')
thin = Side(style='thin', color='BFBFBF'); BD = Border(thin, thin, thin, thin)
wrap = Alignment(wrap_text=True, vertical='top')

note = ws.cell(1, 1, 'PARTS-MODULE deep-flow — LIVE dual observation 2026-07-15. Both envs driven live to /parts/orders per role; '
       'controls (New PO = Order Parts; Receive = accept delivery) observed on-screen with full-page screenshots. '
       'Staging via switch-user impersonation (7) + throwaway role-swap+switch-user (4, throwaway restored). '
       'Production via test-staff role-swap+self-login (14 roles, test staff restored to Office User). '
       'Every row is DUAL LIVE-OBSERVED (Rules 10 & 12) — no inference.')
note.font = Font(bold=True, color='9C0006'); note.alignment = wrap
ws.merge_cells('A1:K1'); ws.row_dimensions[1].height = 60

hdr = ['Staging Role', 'Prod role compared', 'Capability', 'PROD observed', 'Staging observed',
       'Direction / verdict', 'Per-spec?', 'Confidence', 'Method', 'Prod screenshot', 'Staging screenshot / Notes']
ws.append(hdr)
for i, h in enumerate(hdr, 1):
    c = ws.cell(2, i); c.font = H; c.fill = HF; c.alignment = wrap; c.border = BD
for i, w in enumerate([20, 32, 42, 14, 15, 30, 10, 22, 30, 46, 50], 1):
    ws.column_dimensions[ws.cell(2, i).column_letter].width = w

def yn(v): return 'SHOWN' if v else 'hidden'

for role in ORDER:
    prole, merges, note_txt = MAP[role]
    s = stg[role]; p = prod[prole]
    pshot = f'{PROD}/{prole}/parts_orders.png'
    sshot = f'{STG}/{role}/parts_orders.png'
    for label, key, capnote in CAPS:
        sv = bool(s.get(key)); pv = bool(p.get(key))
        # merge consistency check
        merge_note = ''
        if merges:
            vals = {m: bool(prod[m].get(key)) for m in merges if m in prod}
            if all(v == pv for v in vals.values()):
                merge_note = f' (merge components {", ".join(NICE(m) for m in vals)} also {yn(pv)} — consistent)'
            else:
                merge_note = f' (merge components differ: ' + ', '.join(f'{NICE(m)}={yn(v)}' for m, v in vals.items()) + ')'
        if pv == sv:
            verdict = 'MATCH'
        elif pv and not sv:
            verdict = 'STAGING-LESS (prod grants more)'
        else:
            verdict = 'STAGING-MORE (staging grants more)'
        full_note = (note_txt + (' | ' if note_txt else '') + capnote + merge_note).strip(' |')
        ws.append([NICE(role), NICE(prole), label, yn(pv), yn(sv), verdict,
                   'yes' if verdict == 'MATCH' else 'see note', 'DUAL LIVE-OBSERVED',
                   'stg: switch-user/role-swap; prod: test-staff role-swap',
                   pshot if os.path.exists(pshot) else '', (sshot if os.path.exists(sshot) else '') + '  ' + full_note])
        r = ws.max_row
        for i in range(1, len(hdr) + 1):
            ws.cell(r, i).alignment = wrap; ws.cell(r, i).border = BD
        ws.cell(r, 4).fill = OK if pv else YEL
        ws.cell(r, 5).fill = OK if sv else YEL
        if verdict == 'MATCH':
            ws.cell(r, 6).fill = OK
        elif verdict.startswith('STAGING-LESS'):
            ws.cell(r, 6).fill = NV
        else:
            ws.cell(r, 6).fill = ORG

wb.save(f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx')
print('WROTE Parts-Module Dual LIVE sheet.')
# quick summary
match = 0; tot = 0
for role in ORDER:
    prole = MAP[role][0]
    for _, key, _ in CAPS:
        tot += 1
        if bool(stg[role].get(key)) == bool(prod[prole].get(key)):
            match += 1
print(f'Parts-module dual cells: {tot}, MATCH: {match}, non-match: {tot-match}')
