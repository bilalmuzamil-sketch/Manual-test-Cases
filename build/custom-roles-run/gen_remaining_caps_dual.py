#!/usr/bin/env python3
"""Rebuild 'Remaining-Caps Dual LIVE' + 'Prod Remaining-Caps (all 14)' tabs with REAL
dual verdicts now that prod was re-observed live (2026-07-15b). Observed-only (Rules 10 & 12).
Staging = remaining-caps-staging.json (7 roles live; 4 NOT VERIFIED staff/change 500 + Time Clock).
Prod    = prod-remaining-consolidated.json (14 roles live via test-staff role-swap)."""
import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
BASE='/home/user/Manual-test-Cases/build/custom-roles-run'
STGD=json.load(open('/tmp/custom-roles/remaining-caps-staging.json'))['roles']
PROD=json.load(open('/tmp/custom-roles/prod-remaining-consolidated.json'))
XLSX=f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'

# staging role -> primary prod role (+ merge note)
MAP={
 'Admin':('Administrator',''),
 'Service Manager':('Service Manager',''),
 'Senior Service Advisor':('Service Advisor','merge: + Service Advisor Technician + Service Advisor - No Reports'),
 'Parts Manager':('Parts Manager',''),
 'Service Advisor':('Service Advisor - Limited View',''),
 'Foreman':('Foreman',''),
 'Office User':('Office User',''),
 'Parts Technician':('Parts Technician',''),
 'Sales Representative':('Sales Representative','merge: + Reporting'),
 'Technician':('Technician',''),
 'Time Clock User':('Time Clock User',''),
}
ROLES=list(MAP.keys())
# (display, staging_key, prod_key)
CAPS=[('Part Return','PartReturn','partReturn'),
      ('Set Line Status','SetLineStatus','setLineStatus'),
      ('WO Delete (Delete Work Order)','WODelete','woDelete'),
      ('Invoicing create (New Payment)','InvoicingCreate_NewPayment','newPayment'),
      ('Invoice Reverse','InvoiceReverse','invoiceReverse'),
      ('Invoice Issue Credit','IssueCredit','issueCredit')]

def norm(v):
    if v is None: return 'NV'
    s=str(v).strip().lower()
    if s.startswith('shown') or s.startswith('partial'): return 'SHOWN'
    if s.startswith('hidden'): return 'HIDDEN'
    return 'NV'

def stg_val(role,key):
    r=STGD.get(role,{})
    if not r.get('rendered'):
        return f"NOT VERIFIED — {r.get('NOT_VERIFIED_reason','')}"
    return r.get(key,'NOT VERIFIED')

def verdict(pn,sn):
    if pn=='NV' or sn=='NV': return 'NOT VERIFIED'
    if pn==sn: return 'MATCH'
    if pn=='SHOWN' and sn=='HIDDEN': return 'STAGING-LESS (prod-more)'
    if pn=='HIDDEN' and sn=='SHOWN': return 'STAGING-MORE'
    return 'REVIEW'

wb=load_workbook(XLSX)
for t in ['Remaining-Caps Staging LIVE','Remaining-Caps Dual LIVE','Prod Remaining-Caps (all 14)']:
    if t in wb.sheetnames: del wb[t]

hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',fgColor='305496')
green=PatternFill('solid',fgColor='C6EFCE'); red=PatternFill('solid',fgColor='FFC7CE')
grey=PatternFill('solid',fgColor='D9D9D9'); amber=PatternFill('solid',fgColor='FFEB9C')

# ---- Tab 1: dual ----
ws=wb.create_sheet('Remaining-Caps Dual LIVE')
cols=['Staging Role','Prod Role(s) mapped','Capability','PROD Observed (live)','STAGING Observed (live)','Direction / Verdict','Method','Confidence','Evidence']
ws.append(cols)
for c in ws[1]: c.font=hdr; c.fill=hf; c.alignment=Alignment(wrap_text=True,vertical='top')
counts={'MATCH':0,'STAGING-LESS (prod-more)':0,'STAGING-MORE':0,'NOT VERIFIED':0,'REVIEW':0}
for role in ROLES:
    prole,note=MAP[role]
    pobs_all=PROD.get(prole,{})
    for disp,skey,pkey in CAPS:
        pv=pobs_all.get(pkey,'NOT VERIFIED')
        sv=stg_val(role,skey)
        pn,sn=norm(pv),norm(sv)
        vd=verdict(pn,sn); counts[vd]=counts.get(vd,0)+1
        prodcell=f"{pv}"+(f"  [{note}]" if note else '')
        smethod=STGD.get(role,{}).get('method','')
        conf = 'DUAL OBSERVED-LIVE' if vd not in('NOT VERIFIED',) else ('PROD live / STAGING NV' if pn!='NV' and sn=='NV' else ('STAGING live / PROD NV' if sn!='NV' and pn=='NV' else 'BOTH NV'))
        ev=f'live-ui-2026-07-15/production/{prole.replace(" ","_").replace("-","")}/  +  staging/{role.replace(" ","_")}/'
        ws.append([role,prole+(f" ({note})" if note else ''),disp,prodcell,sv,vd,f'prod: test-staff role-swap; stg: {smethod}',conf,ev])
        last=ws[ws.max_row]
        last[3].fill = green if pn=='SHOWN' else (red if pn=='HIDDEN' else grey)
        last[4].fill = green if sn=='SHOWN' else (red if sn=='HIDDEN' else grey)
        vc=last[5]
        vc.fill = amber if 'STAGING-LESS' in vd else (PatternFill('solid',fgColor='B4C7E7') if vd=='STAGING-MORE' else (grey if vd=='NOT VERIFIED' else green))
        for cell in last: cell.alignment=Alignment(wrap_text=True,vertical='top')
ws.insert_rows(1)
ws['A1']=('REAL DUAL verdicts for the remaining FE-gated caps — BOTH sides live-observed 2026-07-15b. '
 'Prod re-observed via renewable self-login + test-staff role-swap (restored to Office User). '
 'Part Return prod = NOT VERIFIED (control not surfacable via headless probe). '
 'Finance NOT VERIFIED where /api/work-orders/invoices/estimate HTTP 400 crashed the panel. '
 'Staging Service Manager/Foreman/Office User/Parts Technician = NOT VERIFIED (no role-holder + staff/change HTTP 500).')
ws['A1'].font=Font(bold=True,color='9C0006')
widths=[20,30,26,40,40,24,34,26,50]
for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w

# ---- Tab 2: prod all 14 ----
ws2=wb.create_sheet('Prod Remaining-Caps (all 14)')
pcols=['Prod Legacy Role','Part Return','Set Line Status','Core OK/Not-OK','WO Delete','New Payment','Invoice Reverse','Issue Credit','Finance method']
ws2.append(pcols)
for c in ws2[1]: c.font=hdr; c.fill=hf; c.alignment=Alignment(wrap_text=True,vertical='top')
for pr in ["Administrator","Service Manager","Service Advisor","Service Advisor - No Reports","Service Advisor Technician","Service Advisor - Limited View","Foreman","Technician","Parts Manager","Parts Technician","Sales Representative","Reporting","Office User","Time Clock User"]:
    o=PROD.get(pr,{})
    row=[pr,o.get('partReturn',''),o.get('setLineStatus',''),o.get('coreOkNotOk',''),o.get('woDelete',''),o.get('newPayment',''),o.get('invoiceReverse',''),o.get('issueCredit',''),o.get('finance_method','')]
    ws2.append(row)
    last=ws2[ws2.max_row]
    for idx in [2,3,4,5,6,7]:
        v=norm(last[idx].value)
        last[idx].fill = green if v=='SHOWN' else (red if v=='HIDDEN' else grey)
    for cell in last: cell.alignment=Alignment(wrap_text=True,vertical='top')
w2=[24,34,30,14,14,16,16,16,52]
for i,w in enumerate(w2,1): ws2.column_dimensions[chr(64+i)].width=w
wb.save(XLSX)
print('counts',counts)
print('tabs now:',wb.sheetnames)
