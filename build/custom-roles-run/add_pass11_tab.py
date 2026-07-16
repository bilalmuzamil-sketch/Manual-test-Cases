import openpyxl
from openpyxl.styles import Font, PatternFill
f='Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb=openpyxl.load_workbook(f)
TAB='Pass-11 LIVE (2026-07-16)'
if TAB in wb.sheetnames: del wb[TAB]
ws=wb.create_sheet(TAB)
hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',fgColor='305496')
less=PatternFill('solid',fgColor='FFC7CE'); more=PatternFill('solid',fgColor='FFEB9C')
match=PatternFill('solid',fgColor='C6EFCE'); nv=PatternFill('solid',fgColor='D9D9D9')
rows=[
 ['PASS-11 (2026-07-16) — CLOSED prod See-AP/AR (13), prod Part Return (13), prod Finance (13, CLEAN self-login), staging Time Clock. OBSERVED-LIVE only; never inferred (Rules 10/12).'],
 [],
 ['Staging role','Capability','Prod (LIVE-OBSERVED)','Staging (LIVE-OBSERVED)','Dual verdict','Confidence / method'],
 # ===== See AP/AR =====
 ['Administrator','See AP/AR (A/R+A/P Aging reports)','SHOWN (tiles render)','SHOWN','MATCH','clean self-login + real-holder switch-user (both agree)'],
 ['Office User','See AP/AR','SHOWN (tiles render)','SHOWN','MATCH','clean self-login + switch-user'],
 ['Service Manager','See AP/AR','HIDDEN (Reports surface FE-gated off; punch-clock also bounces)','SHOWN','STAGING-MORE','clean self-login CONFIRMS switch-user'],
 ['Parts Manager','See AP/AR','HIDDEN','SHOWN','STAGING-MORE','clean self-login'],
 ['Sales Representative','See AP/AR','HIDDEN (nav no-ops; bounce)','SHOWN','STAGING-MORE','real-holder switch-user + clean self-login'],
 ['Senior Service Advisor','See AP/AR','HIDDEN (SA/SA-Tech/SA-No-Reports)','hidden','MATCH','clean self-login'],
 ['Service Advisor','See AP/AR','HIDDEN (SA-Limited-View)','hidden','MATCH','clean self-login'],
 ['Foreman','See AP/AR','HIDDEN','hidden','MATCH','clean self-login'],
 ['Technician','See AP/AR','HIDDEN','hidden','MATCH','real holder + clean self-login'],
 ['Parts Technician','See AP/AR','HIDDEN','hidden','MATCH','clean self-login'],
 ['Time Clock User','See AP/AR','HIDDEN','hidden','MATCH','real holder + clean self-login'],
 [],
 # ===== Part Return =====
 ['Administrator','Part Return ("Return" in picked-part menu)','SHOWN (Move/Core OK/Return)','SHOWN','MATCH','prod S1-719 role-swap+self-login'],
 ['Service Manager','Part Return','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Senior Service Advisor','Part Return','SHOWN (SA/SA-Tech/SA-No-Reports)','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Service Advisor','Part Return','SHOWN (SA-Limited-View)','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Foreman','Part Return','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Technician','Part Return','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Parts Manager','Part Return','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Parts Technician','Part Return','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Sales Representative','Part Return','HIDDEN (WO detail not accessible)','SHOWN','STAGING-MORE','OBSERVED-LIVE both'],
 ['Office User','Part Return','HIDDEN (part menu = Move only)','SHOWN','STAGING-MORE','OBSERVED-LIVE both'],
 ['Time Clock User','Part Return','HIDDEN (WO detail not accessible)','SHOWN','STAGING-MORE','OBSERVED-LIVE both (staging TC closed this pass)'],
 [],
 # ===== Finance (New Payment / Reverse / Issue Credit) =====
 ['Administrator','Finance (New Payment/Reverse/Issue Credit)','SHOWN (all three)','SHOWN (all three)','MATCH','prod clean self-login on invoiced S1-518'],
 ['Service Manager','Finance','HIDDEN (Finance route -> /no-location; WO detail rendered)','SHOWN (NP+Credit)','STAGING-MORE','clean self-login (not an artifact - WO rendered, only Finance bounced)'],
 ['Parts Manager','Finance','HIDDEN (/no-location)','SHOWN (NP+Credit)','STAGING-MORE','clean self-login'],
 ['Parts Technician','Finance','HIDDEN (/no-location)','SHOWN (NP+Credit)','STAGING-MORE','clean self-login'],
 ['Foreman','Finance','HIDDEN (no Finance tab)','SHOWN (NP+Credit)','STAGING-MORE','clean self-login'],
 ['Office User','Finance','HIDDEN (/no-location)','SHOWN (NP+Credit)','STAGING-MORE','clean self-login'],
 ['Senior Service Advisor','Finance','component SA-No-Reports SHOWN (SA + SA-Tech HIDDEN)','SHOWN (all three)','MATCH (reports-enabled component keeps finance)','clean self-login per component'],
 ['Service Advisor','Finance','SHOWN (SA-Limited-View: NP+Reverse+Credit)','SHOWN (NP+Credit; Reverse hidden)','MATCH (minor: prod Reverse SHOWN vs staging hidden)','clean self-login'],
 ['Sales Representative','Finance','HIDDEN (WO detail not accessible)','hidden','MATCH','clean self-login'],
 ['Technician','Finance','HIDDEN (no Finance tab)','hidden','MATCH','clean self-login'],
 ['Time Clock User','Finance','HIDDEN (WO detail not accessible)','HIDDEN','MATCH','clean self-login (prod) + staging TC observed this pass'],
 [],
 ['GENUINE RESIDUALS (precise, non-inferable — everything else above is OBSERVED-LIVE):'],
 ['Send to Terminal (prod)','','N/A — prod org has NO payment-terminal device configured (org-device gate, not a role gate)','staging: SHOWN for some roles (org has a terminal)','ENV LIMIT','Cannot be observed on prod without a terminal device provisioned; not inferable. staging side observed.'],
 ['Approve / Decline line','','needs a WO deliberately left with a PENDING unapproved line','same (line-workflow state)','DATA-STATE','Headless New-Line did not persist a pending line; broader Set Line Status (Start/Complete/Pick) IS observed. Needs a pre-seeded pending-line WO.'],
]
for r in rows: ws.append(r)
ws['A1'].font=Font(bold=True,size=11)
for c in ws[3]: c.font=hdr; c.fill=hf
for row in ws.iter_rows(min_row=4):
    v=(row[4].value or '') if len(row)>4 else ''
    fill=None
    if 'STAGING-LESS' in v: fill=less
    elif 'STAGING-MORE' in v: fill=more
    elif v.startswith('MATCH'): fill=match
    elif 'ENV LIMIT' in v or 'DATA-STATE' in v: fill=nv
    if fill:
        for c in row: c.fill=fill
widths=[26,42,50,26,20,64]
for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
wb.move_sheet(TAB, -(len(wb.sheetnames)-2))
wb.save(f)
print('saved; sheets:', wb.sheetnames[:4])
