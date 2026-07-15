#!/usr/bin/env python3
"""Append 'Remaining-Caps Staging LIVE' tab to the LIVE-VERIFIED workbook + write md addendum.
Observed-only (Rules 10 & 12). Prod session expired this run -> every prod cell NOT VERIFIED."""
import json, os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
BASE='/home/user/Manual-test-Cases/build/custom-roles-run'
D=json.load(open('/tmp/custom-roles/remaining-caps-staging.json'))
XLSX=f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
STG=f'{BASE}/live-ui-2026-07-15/staging'
PROD_MAP={'Admin':'Administrator','Service Manager':'Service Manager','Senior Service Advisor':'Service Advisor (merge)',
 'Parts Manager':'Parts Manager','Service Advisor':'SA Limited View','Foreman':'Foreman','Office User':'Office User',
 'Parts Technician':'Parts Technician','Sales Representative':'Sales Representative (merge)','Technician':'Technician','Time Clock User':'Time Clock User'}
CAPS=[('Part Return','PartReturn'),('Set Line Status','SetLineStatus'),('WO Delete (Delete Work Order)','WODelete'),
 ('Invoicing create (New Payment)','InvoicingCreate_NewPayment'),('Invoice Reverse','InvoiceReverse'),('Invoice Issue Credit','IssueCredit')]
ROLES=['Admin','Service Manager','Senior Service Advisor','Parts Manager','Service Advisor','Foreman','Office User','Parts Technician','Sales Representative','Technician','Time Clock User']
wb=load_workbook(XLSX)
title='Remaining-Caps Staging LIVE'
if title in wb.sheetnames: del wb[title]
ws=wb.create_sheet(title)
hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',fgColor='305496')
green=PatternFill('solid',fgColor='C6EFCE'); red=PatternFill('solid',fgColor='FFC7CE'); grey=PatternFill('solid',fgColor='D9D9D9')
cols=['Staging Role','Prod Role (compare)','Capability','STAGING Observed (live)','PROD Observed','Verdict','Method','Confidence','Staging Evidence']
ws.append(cols)
for c in ws[1]: c.font=hdr; c.fill=hf; c.alignment=Alignment(wrap_text=True,vertical='top')
def shot(role):
    p=f'{STG}/{role.replace(" ","_")}/wo_lines.png'
    return f'live-ui-2026-07-15/staging/{role.replace(" ","_")}/wo_lines.png' if os.path.exists(p) else '(no render)'
rows=0
for role in ROLES:
    r=D['roles'].get(role,{})
    rendered=r.get('rendered')
    method=r.get('method','')
    for capname,key in CAPS:
        stg_obs = r.get(key) if rendered else f"NOT VERIFIED — {r.get('NOT_VERIFIED_reason','')}"
        prod_obs='NOT VERIFIED — prod session expired 2026-07-15'
        verdict='NOT VERIFIED (prod side)' if rendered else 'NOT VERIFIED (both sides)'
        conf='STAGING OBSERVED-LIVE / PROD NOT-VERIFIED' if rendered else 'NOT-VERIFIED'
        ev=shot(role) if rendered else '(not rendered)'
        ws.append([role,PROD_MAP[role],capname,stg_obs,prod_obs,verdict,method,conf,ev])
        rows+=1
        last=ws[ws.max_row]
        # colour staging cell
        sc=last[3]
        if rendered:
            sc.fill=green if str(stg_obs).startswith('SHOWN') else (red if str(stg_obs).startswith('hidden') else grey)
        else:
            sc.fill=grey
        last[4].fill=grey
        for cell in last: cell.alignment=Alignment(wrap_text=True,vertical='top')
# caps not observable at all
ws.append([])
ws.append(['CAPS NOT OBSERVABLE ON STAGING THIS RUN (both sides NOT VERIFIED)'])
ws[ws.max_row][0].font=Font(bold=True)
for cap,reason in D['caps_not_observable_staging'].items():
    ws.append(['(all roles)','(all)',cap,'NOT VERIFIED','NOT VERIFIED — prod session expired',reason,'','NOT-VERIFIED',''])
    for cell in ws[ws.max_row]: cell.alignment=Alignment(wrap_text=True,vertical='top'); cell.fill=grey
widths=[20,22,26,34,34,20,26,42,40]
for i,w in enumerate(widths,1): ws.column_dimensions[chr(64+i)].width=w
# banner note on the existing superseded gaps sheet + a note row at top of this sheet
ws.insert_rows(1)
ws['A1']='STAGING-side LIVE observations of the coordinator remaining caps (existing data, no seeding). PROD EXPIRED this run -> prod cells NOT VERIFIED. Rebuild dual once fresh prod cookies supplied.'
ws['A1'].font=Font(bold=True,color='9C0006')
wb.save(XLSX)
print('workbook tab written; data rows',rows)
