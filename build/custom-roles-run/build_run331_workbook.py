#!/usr/bin/env python3
"""
Build the Custom Roles RUN 331 per-status results workbook (+ CSV mirror).

Reproducible generator. Documentation only — reads local artifacts, writes:
  - CustomRoles_Run331_Results.xlsx  (7 tabs: Summary + 4 status tabs +
    Cases Updated (Phase 1) + Flagged for Decision)
  - CustomRoles_Run331_Results.csv   (combined status-tab mirror)

Sources of truth (all local, read-only):
  - run331-results-log.md       -> per-case status + plain Expected/Actual + notes
  - run331-tests.json           -> case_id -> Title / Section (section_path)
  - ../custom-roles-spec-update/updated-spec-diff.md -> 3 UPDATED + 16 FLAGGED

Dedup rule: a case may appear twice in the log (C26553 was posted Blocked then
re-posted Retest). Last occurrence in the log wins.
"""

import csv
import json
import os
import re

HERE = os.path.dirname(os.path.abspath(__file__))
LOG = os.path.join(HERE, "run331-results-log.md")
TESTS = os.path.join(HERE, "run331-tests.json")
XLSX = os.path.join(HERE, "CustomRoles_Run331_Results.xlsx")
CSV_OUT = os.path.join(HERE, "CustomRoles_Run331_Results.csv")

RUN_ID = 331
RUN_NAME = "Nightly Test Run - Jul 9, 2026"

# ---- The 3 master cases UPDATED in TestRail this phase (Phase 1) -------------
UPDATED = {
    "C2528": {
        "old": ("Customer Portal option hidden for many non-admin roles, "
                "including Service Manager, Service Advisor and Parts Manager."),
        "new": ("Per spec change #7 those roles now HAVE Customer Portal. "
                "Removed SM/SA/PM from the hidden-role list (kept Foreman, "
                "Technician, Parts Technician, Office, Time Clock User) and "
                "rewrote Expected to name the roles that DO get Customer Portal "
                "(Admin, SM, Senior SA, SA, Parts Manager)."),
        "spec": "#7 Customer Portal now ON for SA/SSA/SM/PM (+Admin)",
    },
    "C26424": {
        "old": ("Expected named the pre-10-Jun label 'View and Manage AP/AR "
                "Data' (and carried a pasted inline-style span)."),
        "new": ("Renamed to the current setting label 'Manage Accounts Payable "
                "and Receivable' and stripped the pasted style. Behaviour "
                "(enabling Invoicing Delete with AP/AR OFF prompts to enable "
                "AP/AR) unchanged and matches spec."),
        "spec": "#6 AP/AR setting renamed 'Manage Accounts Payable and Receivable'",
    },
    "C26475": {
        "old": ("Title/steps/expected said turning See Financial Data OFF "
                "'auto-clears / auto-unchecks' the Part-Sales + Invoicing "
                "checkboxes silently."),
        "new": ("Retitled + rewrote to the PROMPT model (spec #5): turning SFD "
                "OFF shows a prompt listing dependents to disable (Invoicing "
                "CRUD, Part Sales CRUD, Order Parts, Manage AP/AR); confirm "
                "disables, cancel keeps SFD ON."),
        "spec": "#5 SFD OFF prompts to disable dependents (not silent auto-clear)",
    },
}

# ---- The 16 cases FLAGGED for a product decision (not auto-updated) ----------
FLAGGED = {
    "C2480": ("#12 App Settings now covers Roles & Permissions and Office has App "
              "Settings ON, so Office may now see Roles & Permissions — contradicts "
              "the case's 'hidden' expectation. May be further-gated (Admin pages)."),
    "C2497": ("#3 Reverse Invoice now needs Work Orders Delete; Service Advisor has "
              "WO View+Edit but no Delete, so SA should NOT see Reverse. Bundled-roles "
              "case; carving reverse-invoice per role is ambiguous."),
    "C2500": ("#14 Timesheets have no Delete permission, only View + Create & Edit. "
              "Expected still says 'delete/modify' timesheet entries — 'delete' may mean "
              "per-entry removal vs the Delete atom. Ambiguous."),
    "C2561": ("§1a WO-detail edit fields need WO Create & Edit, but the case lists "
              "Technician (WO View only) as able to edit Lead Technician. Possible extra "
              "tech-specific rule — needs a product call."),
    "C2565": ("§1a Change Customer/Contact needs WO Create & Edit, but the case lists "
              "Office (WO View only). 07-Jul change removed the 'customer-setting controls "
              "WO tab' requirement — ambiguous."),
    "C2567": ("§1a Change Asset needs WO Create & Edit, but the case lists Office (WO "
              "View only). Same ambiguity as Change Customer — needs a product call."),
    "C26340": ("Already a 'decision needed from product' case: modal short labels "
               "(Admin/Parts Tech/Time Clock) vs Roles-list long labels "
               "(Administrator/Parts Technician/Time Clock User). Judgment call."),
    "C26419": ("Title/body mismatch: title says Catalog & Inventory enables Return to "
               "Inventory but the steps/expected are all about Vendor & Order Management. "
               "Spec attributes return-to-inventory to BOTH — needs cleanup + decision."),
    "C26488": ("#1 'View History Logs' was relabelled + repurposed to control only "
               "inventory Part History; WO audit log now needs WO Create & Edit and line "
               "story needs WOL View. A correct rewrite changes the case's whole subject."),
    "C26496": ("#3 Internal contradiction: steps say SM 'cannot reverse invoices' but "
               "Expected says SM CAN reverse. Reverse is now under WO Delete (which SM "
               "has). Spec self-contradicts — do not force."),
    "C26553": ("#2 Backend-API case gates AR Aging on ROLE_ACCOUNT_RECEIVABLE_REPORT::VIEW; "
               "new spec routes aging through all-or-nothing Reports. 403-for-tech still "
               "holds but the permission name/model may be stale."),
    "C27873": ("#8 Ties modify/delete of another user's CUSTOMER note to Work Orders "
               "Delete; spec now governs customer notes via Customer Management (View = "
               "edit any note; Delete = delete others'). Real but subtle discrepancy."),
    "C27418": ("#1 Combo case grants 'view WO history logs' from the History Logs "
               "permission; under the repurpose that permission no longer grants WO "
               "history. Adversarial/combo section — flag, don't rewrite."),
    "C27468": ("#1 Same as C27418 — 'Reviewer + history' combo grants WO history from "
               "the repurposed History Logs permission. Flag, don't rewrite."),
    "C27487": ("#1 Same as C27418 — 'Invoice view + history' combo grants WO history "
               "from the repurposed History Logs permission. Flag, don't rewrite."),
    "C27494": ("#1 Same as C27418 — 'WO + pick + history' combo grants WO history from "
               "the repurposed History Logs permission. Flag, don't rewrite."),
}

STATUS_ORDER = ["Passed", "Failed", "Retest", "Blocked", "Untested"]


def load_tests():
    with open(TESTS) as f:
        data = json.load(f)
    by_id = {}
    for t in data["tests"]:
        cid = "C%d" % t["case_id"]
        sec = t.get("section_path", "")
        # Trim the leading "Test Cases / " for readability where present.
        by_id[cid] = {"title": t.get("title", ""), "section": sec}
    return by_id


def parse_log():
    """Return dict case_id -> {status, expected, actual} (last occurrence wins)."""
    rows = {}
    row_re = re.compile(r"^\|\s*(C\d+)\s*\|\s*([^|]+?)\s*\|\s*(.*?)\s*\|\s*(.*?)\s*\|\s*$")
    with open(LOG) as f:
        for line in f:
            m = row_re.match(line.rstrip("\n"))
            if not m:
                continue
            cid, status, expected, actual = m.groups()
            status = status.strip()
            if status not in STATUS_ORDER:
                continue
            rows[cid] = {
                "status": status,
                "expected": expected.strip(),
                "actual": actual.strip(),
            }
    return rows


def notes_for(cid, status, actual):
    parts = []
    if cid in UPDATED:
        parts.append("UPDATED in TestRail (master case) — %s" % UPDATED[cid]["spec"])
    if cid in FLAGGED:
        parts.append("FLAGGED for decision: " + FLAGGED[cid])
    if status in ("Retest", "Blocked") and cid not in FLAGGED:
        # The reason/resume note lives in the Actual column; echo it here too.
        parts.append(actual)
    return "  ||  ".join(p for p in parts if p)


def main():
    tests = load_tests()
    log = parse_log()

    # Sanity: every logged case should map to a test in the run.
    missing = [c for c in log if c not in tests]
    extra = [c for c in tests if c not in log]

    # Assemble records
    records = {}
    for cid, r in log.items():
        meta = tests.get(cid, {"title": "", "section": ""})
        records[cid] = {
            "case_id": cid,
            "title": meta["title"],
            "section": meta["section"],
            "status": r["status"],
            "expected": r["expected"],
            "actual": r["actual"],
            "notes": notes_for(cid, r["status"], r["actual"]),
        }

    # Counts
    counts = {s: 0 for s in STATUS_ORDER}
    for r in records.values():
        counts[r["status"]] += 1
    total = len(records)

    print("=== RUN 331 computed tally (from log) ===")
    print("Total distinct cases resulted:", total)
    for s in STATUS_ORDER:
        print("  %-9s %d" % (s, counts[s]))
    if missing:
        print("WARN: logged cases not in run map:", missing)
    if extra:
        print("WARN: run cases with NO result in log:", extra)

    build_xlsx(records, counts, total, extra)
    build_csv(records)
    print("Wrote:", XLSX)
    print("Wrote:", CSV_OUT)


COLS = ["Case ID", "Title", "Section/Area", "Status",
        "Expected (plain)", "Actual (plain)", "Notes/Reason"]


def build_csv(records):
    order = {s: i for i, s in enumerate(STATUS_ORDER)}
    rows = sorted(records.values(),
                  key=lambda r: (order.get(r["status"], 99), r["case_id"]))
    with open(CSV_OUT, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(COLS)
        for r in rows:
            w.writerow([r["case_id"], r["title"], r["section"], r["status"],
                        r["expected"], r["actual"], r["notes"]])


def build_xlsx(records, counts, total, extra):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="305496")
    title_font = Font(bold=True, size=14)
    sub_font = Font(italic=True, size=10, color="555555")
    wrap = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    status_fill = {
        "Passed": "E2EFDA", "Failed": "FCE4E4", "Retest": "FFF2CC",
        "Blocked": "DDEBF7", "Untested": "F2F2F2",
    }

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
        ws.freeze_panes = "A%d" % (row + 1)

    def status_sheet(name, status):
        ws = wb.create_sheet(name)
        ws.append(COLS)
        style_header(ws, len(COLS))
        order_rows = sorted(
            [r for r in records.values() if r["status"] == status],
            key=lambda r: r["case_id"])
        for r in order_rows:
            ws.append([r["case_id"], r["title"], r["section"], r["status"],
                       r["expected"], r["actual"], r["notes"]])
        widths = [12, 46, 40, 11, 52, 60, 60]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        fill = PatternFill("solid", fgColor=status_fill.get(status, "FFFFFF"))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                max_col=len(COLS)):
            for cell in row:
                cell.alignment = wrap
                cell.border = border
            row[3].fill = fill
        return ws, len(order_rows)

    # ---- Summary sheet (first) ----
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Custom Roles & Permissions — TestRail RUN %d Results" % RUN_ID
    ws["A1"].font = title_font
    ws["A2"] = '"%s"  (project 1 / suite 1 "Master")' % RUN_NAME
    ws["A2"].font = sub_font
    ws["A3"] = ("Results posted LIVE to TestRail via add_result_for_case, each "
                "with a plain-language Expected/Actual comment. This workbook is a "
                "read-only export of the run331-results-log.md source of truth.")
    ws["A3"].font = sub_font
    ws["A3"].alignment = wrap

    r = 5
    ws.cell(row=r, column=1, value="Result").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Count").font = Font(bold=True)
    ws.cell(row=r, column=3, value="Share").font = Font(bold=True)
    for c in range(1, 4):
        ws.cell(row=r, column=c).fill = hdr_fill
        ws.cell(row=r, column=c).font = hdr_font
        ws.cell(row=r, column=c).border = border
    r += 1
    for s in STATUS_ORDER:
        ws.cell(row=r, column=1, value=s).border = border
        ws.cell(row=r, column=2, value=counts[s]).border = border
        pct = (counts[s] / total * 100) if total else 0
        ws.cell(row=r, column=3, value="%.1f%%" % pct).border = border
        ws.cell(row=r, column=1).fill = PatternFill(
            "solid", fgColor=status_fill.get(s, "FFFFFF"))
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total).font = Font(bold=True)
    ws.cell(row=r, column=3, value="100.0%").font = Font(bold=True)
    for c in range(1, 4):
        ws.cell(row=r, column=c).border = border
    total_row = r

    r += 2
    ws.cell(row=r, column=1, value="Status legend").font = Font(bold=True)
    r += 1
    legend = [
        ("Passed", "Live behaviour matches the (updated-spec) expected result."),
        ("Failed", "Real deviation — the spec change is not yet live on staging."),
        ("Retest", "Not fully driven this session (needs data seeding, multi-step "
                   "UI, endpoint discovery, a healthy payment endpoint, or a "
                   "product/spec decision) — resume note per case."),
        ("Blocked", "Cannot be exercised in this environment (external payment "
                    "processor / customer portal, or the IBS Multi-Tenancy feature "
                    "is not enabled on this org)."),
        ("Untested", "No result posted (none — every case in the run was resulted)."),
    ]
    for s, desc in legend:
        ws.cell(row=r, column=1, value=s).font = Font(bold=True)
        ws.cell(row=r, column=1).fill = PatternFill(
            "solid", fgColor=status_fill.get(s, "FFFFFF"))
        ws.cell(row=r, column=1).border = border
        ws.cell(row=r, column=2, value=desc).alignment = wrap
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value="TestRail case maintenance (Phase 1)").font = Font(bold=True)
    r += 1
    notes = [
        "3 MASTER cases were UPDATED in TestRail from the new spec: "
        "C2528, C26424, C26475 (see the 'Cases Updated (Phase 1)' tab).",
        "16 cases were FLAGGED for a product decision (NOT auto-updated) — they "
        "carry a spec conflict/ambiguity that needs a product call "
        "(see the 'Flagged for Decision' tab).",
        "The remaining ~141 cases were already consistent with the updated spec "
        "or are generic functional suites not driven by the changed rules.",
        "The 4 Failed cases are the real spec-not-yet-live deviations: C26387 & "
        "C26388 (Add Customer/Add Asset still shown with Customer Mgmt Edit OFF), "
        "C26475 (SFD OFF does not prompt/clear dependents), C26482 (aging reports "
        "still gated by Manage AP/AR instead of following Reports).",
    ]
    for n in notes:
        ws.cell(row=r, column=1, value="• " + n).alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    if extra:
        r += 1
        ws.cell(row=r, column=1,
                value="Discrepancy note: run cases with no result in log: "
                      + ", ".join(extra)).font = Font(bold=True, color="C00000")
        r += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    for col in ("D", "E", "F"):
        ws.column_dimensions[col].width = 14
    ws.freeze_panes = "A6"

    # ---- Status tabs ----
    status_sheet("Passed", "Passed")
    status_sheet("Failed", "Failed")
    status_sheet("Retest", "Retest")
    status_sheet("Blocked", "Blocked")

    # ---- Cases Updated (Phase 1) ----
    ws = wb.create_sheet("Cases Updated (Phase 1)")
    upd_cols = ["Case ID", "Title", "Spec change", "Old (summary)", "New (summary)"]
    ws.append(upd_cols)
    style_header(ws, len(upd_cols))
    for cid in ["C2528", "C26424", "C26475"]:
        u = UPDATED[cid]
        title = records.get(cid, {}).get("title", "")
        ws.append([cid, title, u["spec"], u["old"], u["new"]])
    for i, w in enumerate([12, 46, 42, 58, 66], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(upd_cols)):
        for cell in row:
            cell.alignment = wrap
            cell.border = border

    # ---- Flagged for Decision ----
    ws = wb.create_sheet("Flagged for Decision")
    fl_cols = ["Case ID", "Title", "Section/Area", "Result", "Reason (needs product call)"]
    ws.append(fl_cols)
    style_header(ws, len(fl_cols))
    for cid in sorted(FLAGGED, key=lambda c: int(c[1:])):
        meta = records.get(cid, {})
        ws.append([cid, meta.get("title", ""), meta.get("section", ""),
                   meta.get("status", ""), FLAGGED[cid]])
    for i, w in enumerate([12, 46, 40, 11, 78], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(fl_cols)):
        for cell in row:
            cell.alignment = wrap
            cell.border = border

    wb.save(XLSX)


if __name__ == "__main__":
    main()
