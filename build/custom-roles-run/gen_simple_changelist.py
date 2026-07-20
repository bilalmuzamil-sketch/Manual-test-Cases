#!/usr/bin/env python3
"""Simple change-list deliverable for Bilal & Vlad: the 29 cases that need a change or a
decision (the 240 OK cases are omitted). One plain row per case + driving ticket + whether
that ticket is Done. Second tab highlights cases waiting on a NOT-DONE ticket."""
import os
BASE = os.path.dirname(os.path.abspath(__file__))
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"
JIRA = "https://shopview.atlassian.net/browse/{}"
DATE = "2026-07-20"
AREA = {3529:'Create Role',3532:'Permission Summary',3534:'Work Orders',3535:'Work Order Lines',
 3538:'Parts Dept',3539:'Invoicing & Payments',3543:'View Mode',3545:'Cross-Cutting (AP/AR)',
 3546:'View Part History',3548:'Per-Role sets',4091:'API — Time Clock'}

# case_id -> (section, change (plain), tickets, ticket_status, action)
# ticket_status: DONE / NOT DONE (state) / OBSOLETE / none
M = {
26339:(3529,"Build does NOT hard-block a duplicate role NAME (only a soft 'similar role' warning on identical name+permissions). Case says a duplicate name is blocked.","—","no ticket","DECISION"),
26355:(3532,"Permission Summary lists a 3rd cross-cutting toggle 'View History Logs' — the build has only TWO (See Financial Data + View and Manage AP/AR Data). Drop the History toggle.","SV-8202","NOT DONE (Ready to Fix)","Apply update"),
26359:(3532,"Same as C26355 — remove the non-existent 'View History Logs' cross-cutting toggle from the summary list.","SV-8202","NOT DONE (Ready to Fix)","Apply update"),
26387:(3534,"In the New Work Order flow the 'Add Customer' button IS shown and works without Customers Create&Edit. Case wrongly expects it hidden.","SV-8002","DONE","Apply update"),
26388:(3534,"Same for 'Add Asset' in the New Work Order flow — shown & works without Customers Create&Edit. Case wrongly expects it hidden.","SV-8002","DONE","Apply update"),
26418:(3538,"Vendor Delete + reverse: the gate for REVERSING a vendor transaction is unsettled (old tickets went back and forth). Confirm the exact gate; vendor/PO delete part is correct.","SV-7871 / SV-7912","OBSOLETE","Apply update"),
26419:(3538,"Spec is self-contradictory on which permission gates 'Return to Inventory' (Catalog&Inventory Edit vs Vendor&Order Mgmt Edit). Needs a decision.","—","no ticket","DECISION"),
26424:(3539,"Case expects a prompt to enable AP/AR when ticking Invoicing Delete — that prompt is now BUILT, so the old 'Deviation' is resolved. Re-run should pass.","SV-8170","DONE","Apply update"),
26427:(3539,"'Send to Terminal' needs Invoicing Create&Edit AND Customer Portal ON. Case only requires Invoicing Create&Edit (would false-pass a Portal-OFF role).","SV-8087","DONE","Apply update"),
26459:(3543,"Does Tech View hide labor rates when See Financial Data is ON? Spec says hidden; SV-8107 (closed 'not a bug') says visible. Conflict — needs a decision.","SV-8107","OBSOLETE (no fix)","DECISION"),
26464:(3543,"Same labor-rate-in-Tech-View conflict as C26459 (only the labor-rate sentence). Needs the same decision.","SV-8107","OBSOLETE (no fix)","DECISION"),
26488:(3546,"'View History Logs' was renamed 'View Part History' and now controls ONLY the inventory Part History icon. WO/line audit log follows WO Create&Edit; line story follows WO Lines View. Rewrite the case.","SV-7989 / SV-8202","NOT DONE (Ready to Fix)","Apply update"),
26489:(3546,"Same rename/repurpose as C26488 (the OFF case).","SV-7989 / SV-8202","NOT DONE (Ready to Fix)","Apply update"),
26495:(3548,"Admin role summary lists 'View History Logs' — drop it (no such editor toggle). Permission set otherwise correct. AP/AR label stays as build.","SV-8202","NOT DONE (Ready to Fix)","Apply update"),
26496:(3548,"Service Manager DOES have Work Orders Delete (LIVE-confirmed). Case says 'no Delete'. Fix to V/E/D.","SV-8297","DONE","Apply update"),
26497:(3548,"Senior Service Advisor has Reports OFF in the build but spec says Reports ON. Editable role on a shared org — confirm on clean defaults / may need a bug.","—","no ticket","DECISION"),
26498:(3548,"Service Advisor Invoicing is View/Create&Edit in the build (no Delete) but spec/case say full. Confirm on clean defaults / may need a bug.","—","no ticket","DECISION"),
26500:(3548,"Technician shows WO Create&Edit, Customers Create&Edit and See Financial Data ON in the build vs spec View/View/off. Technician role is known-drifted on this shared org — re-derive on clean defaults.","—","no ticket (env drift)","DECISION"),
26502:(3548,"Parts Technician summary references 'View History Logs' — relabel to 'View Part History'. Permission set correct.","SV-8202","NOT DONE (Ready to Fix)","Apply update"),
26503:(3548,"Office User: Invoicing changed to View/Create&Edit/Delete (LIVE-confirmed). NOTE: the 7/14 spec that REMOVES Office Work Orders + Part Sales is NOT deployed (build still grants both); how Office makes payments without creating invoices is still open.","SV-7993","NOT DONE (Open)","Apply update"),
26504:(3548,"Sales Representative is NOT 'Reports only' — it has Work Orders View, WO Lines View, Customers View+Create&Edit, Part Sales View (LIVE-confirmed). Rewrite the case.","SV-8061","DONE","Apply update"),
27736:(3545,"AP/AR toggle label is correct as-is ('View and Manage AP/AR Data'), but the case also says it sits next to a 'View History Logs' toggle — remove that (only 2 cross-cutting toggles exist).","SV-8202","NOT DONE (Ready to Fix)","Apply update"),
27870:(3535,"Marking a core OK/Not-OK is gated by Work Orders VIEW (everyone), and a line's story history by WO Lines VIEW — not 'controlled by WO Lines Create&Edit' as the case says.","SV-8130 / SV-7989","DONE","Apply update"),
27873:(3534,"Customer notes follow the CUSTOMER permission, not Work Orders Delete: View lets you edit anyone's note (delete own); Delete lets you delete others'. Rewrite the stubbed case.","SV-8003","DONE","Apply update"),
29435:(3534,"Can a Pick/Order-Parts role (WO Lines View only) edit a part's quantity? SV-8136 says it needs WO Lines Create&Edit; SV-8055 suggests a Parts-tab exception. Conflict — needs a decision.","SV-8136 / SV-8055","DONE (conflicting)","DECISION"),
29457:(4091,"Time Clock user reading org Settings via API returns 200, not 403. PO confirmed backend leaks are EXPECTED (not enforced by design). Flip the 'BUG/403' expectation to the accepted 200.","SV-7958","DONE","Apply update"),
29458:(4091,"Time Clock user reading Taxes via API returns 200, not 403 — accepted per PO. Flip expectation.","SV-7958","DONE","Apply update"),
29459:(4091,"Time Clock user creating a Customer via API returns 201, not 403 — PO accepted this write-leak. Flip expectation (note: it IS a create-leak the PO knowingly accepted).","SV-7958","DONE","Apply update"),
29460:(4091,"Time Clock user Work Order create is not gated with 403 — accepted per PO. Flip expectation.","SV-7958","DONE","Apply update"),
}

order = sorted(M)
def verdict(cid): return 'DECISION' if M[cid][4]=='DECISION' else 'UPDATE'

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = openpyxl.Workbook()
hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
warnfill = PatternFill('solid', fgColor='FCE4D6')  # light orange for NOT-DONE ticket rows
okfill = PatternFill('solid', fgColor='E2EFDA')
thin = Border(*[Side(style='thin', color='D9D9D9')]*4)
HDR = ['Case ID','TestRail link','Area','What needs to change','Driving ticket','Ticket status','Action']

def sheet(ws, ids):
    ws.append(HDR)
    for c in ws[1]: c.fill=hf; c.font=hfont; c.alignment=wrap
    for cid in ids:
        sec,change,tix,status,action = M[cid]
        tlink = " ".join(JIRA.format(t.strip()) for t in tix.replace('/',' ').split() if t.startswith('SV-')) or tix
        ws.append([f"C{cid}", LINK.format(cid), AREA.get(sec,str(sec)), change, tix, status, action])
        r = ws[ws.max_row]
        fill = warnfill if status.startswith('NOT DONE') else None
        for c in r:
            c.alignment=wrap; c.border=thin
            if fill: c.fill=fill
    for i,w in enumerate([9,40,20,72,20,22,14],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes='A2'

# Tab 1 — all 29
ws=wb.active; ws.title='Change list (29)'
ws.append(['Custom Roles spec-recheck — cases that need a change or a decision (29 of 269).'])
ws['A1'].font=Font(bold=True,size=13)
ws.append(['The other 240 cases are fine as-is. Orange rows = waiting on a ticket that is NOT yet done. Nothing pushed to TestRail yet.'])
ws.append([])
start=ws.max_row+1
sheet(ws, order)
# move header styling: header is at row `start`
# (sheet() appended header at current position)

# Tab 2 — waiting on NOT-DONE tickets
nd = [c for c in order if M[c][3].startswith('NOT DONE')]
ws2=wb.create_sheet('Waiting on open tickets')
ws2.append(['Cases whose change depends on a ticket that is NOT yet done'])
ws2['A1'].font=Font(bold=True,size=13)
ws2.append(['These should NOT be finalised until the ticket ships; keep the case tracking the target and re-verify after the fix.'])
ws2.append([])
sheet(ws2, nd)

wb.save(os.path.join(BASE, f'CustomRoles_SpecRecheck_ChangeList_{DATE}.xlsx'))

# MD (simple)
with open(os.path.join(BASE, f'CustomRoles_SpecRecheck_ChangeList_{DATE}.md'),'w') as fh:
    fh.write(f"# Custom Roles spec-recheck — change list ({DATE})\n\n")
    fh.write("29 of 269 cases need a change or a decision; the other 240 are fine as-is. Nothing pushed to TestRail yet.\n\n")
    fh.write("**Legend:** Action = *Apply update* (wording/expected fix) or *Decision* (needs you/Vlad/PO to choose). ")
    fh.write("Ticket status shows whether the driving Jira ticket is Done.\n\n")
    fh.write("| Case | Area | What needs to change | Ticket | Ticket status | Action |\n|---|---|---|---|---|---|\n")
    for cid in order:
        sec,change,tix,status,action=M[cid]
        fh.write(f"| [C{cid}]({LINK.format(cid)}) | {AREA.get(sec,sec)} | {change} | {tix} | {status} | {action} |\n")
    nd=[c for c in order if M[c][3].startswith('NOT DONE')]
    fh.write(f"\n## Highlight — cases waiting on a ticket that is NOT yet done ({len(nd)})\n\n")
    fh.write("| Case | Ticket | Ticket status | Why it's blocked |\n|---|---|---|---|\n")
    for cid in nd:
        sec,change,tix,status,action=M[cid]
        fh.write(f"| [C{cid}]({LINK.format(cid)}) | {tix} | {status} | {change} |\n")
    fh.write("\n**Not-done tickets driving these:** SV-8202 (Ready to Fix — legacy 'View History Logs' still in build, must become Parts-scoped 'View Part History'); SV-7993 (Open — Office invoices vs payments rule). Also watch SV-8324 (Code Review — return permission dependencies vs spec).\n")

print(f"Wrote ChangeList: {len(order)} rows; {len(nd)} waiting on NOT-DONE tickets")
