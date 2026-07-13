#!/usr/bin/env python3
"""
Generator for the Time Clock Role — Backend API Enforcement evidence deliverable.

Reads the live-probe results (api-timeclock-2026-07-13/results.json) and the
authored TestRail case ids (embedded map below), emits:
  - CustomRoles_TimeClock-API-Enforcement_2026-07-13.xlsx  (Detail + Summary tabs)
  - CustomRoles_TimeClock-API-Enforcement_2026-07-13.md

One row per endpoint = Area | Endpoint | Method | Expected | Actual | Verdict |
TestRail C-ID + link. Summary counts correctly-blocked vs leaks (Standing Rule 8:
Case ID + clickable TestRail link).
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

BASE = os.path.dirname(os.path.abspath(__file__))
RESULTS = json.load(open(os.path.join(BASE, 'api-timeclock-2026-07-13', 'results.json')))
TR = "https://shopview.testrail.io/index.php?/cases/view/"

# endpoint (results row) -> TestRail case id that covers it
CASE_MAP = {
    "/api/work-orders?page=1": 29446,
    "/api/work-orders/view/{id}": 29446,
    "/api/calendar?date=..&end_date=..": 29447,
    "(timesheet-view API not identified; candidates 404)": None,
    "/api/inventory/parts?page=1": 29448,
    "/api/inventory/orders?page=1": 29449,
    "/api/customers?page=1": 29450,
    "/api/reporting/account-payable/unpaid-invoices-report": 29451,
    "/api/organizations/{org}/roles": 29452,
    "/api/staff?page=1": 29452,
    "/api/ibs/settings": 29453,
    "/api/departments": 29454,
    "/api/organizations/settings": 29457,
    "/api/taxes": 29458,
    "/api/organizations/settings/change": 29455,
    "/api/customers/delete": 29456,
    "/api/customers/create": 29459,
    "/api/work-orders/create": 29460,
}

rows = []
for r in RESULTS['results']:
    cid = CASE_MAP.get(r['endpoint'])
    rows.append({
        'area': r['area'], 'endpoint': r['endpoint'], 'method': r['method'],
        'expect': r['expect'], 'tech': r['tech'], 'verdict': r['verdict'],
        'note': r['note'], 'cid': cid,
        'link': (TR + str(cid)) if cid else '',
    })

# ---- counts ----
correctly_blocked = sum(1 for r in rows if r['verdict'] == 'Correctly blocked')
allowed_ok        = sum(1 for r in rows if r['verdict'] == 'Correct (allowed)')
leaks             = sum(1 for r in rows if r['verdict'] == 'LEAK-bug')
inconclusive      = sum(1 for r in rows if r['verdict'] == 'Inconclusive')

# ================= XLSX =================
wb = Workbook()
HEAD = PatternFill("solid", fgColor="1F4E78")
HF = Font(bold=True, color="FFFFFF")
RED = PatternFill("solid", fgColor="F8CBAD")
GRN = PatternFill("solid", fgColor="C6E0B4")
GRY = PatternFill("solid", fgColor="D9D9D9")

ws = wb.active
ws.title = "Enforcement Detail"
cols = ["Area", "Endpoint", "Method", "Expected Status", "Actual Status (Time Clock)",
        "Verdict", "TestRail C-ID", "TestRail Link", "Note"]
ws.append(cols)
for c in range(1, len(cols) + 1):
    ws.cell(1, c).fill = HEAD; ws.cell(1, c).font = HF
for r in rows:
    ws.append([r['area'], r['endpoint'], r['method'], r['expect'], r['tech'],
               r['verdict'], (f"C{r['cid']}" if r['cid'] else ""), r['link'], r['note']])
    row_i = ws.max_row
    fill = RED if r['verdict'] == 'LEAK-bug' else (GRN if r['verdict'].startswith('Correct') else GRY)
    ws.cell(row_i, 6).fill = fill
    if r['link']:
        ws.cell(row_i, 8).hyperlink = r['link']; ws.cell(row_i, 8).font = Font(color="0563C1", underline="single")
widths = [22, 46, 8, 15, 22, 18, 12, 52, 60]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w
ws.freeze_panes = "A2"

# Summary
sm = wb.create_sheet("Summary")
sm.append(["Time Clock Role — Backend API Permission Enforcement — 2026-07-13 (STAGING)"])
sm["A1"].font = Font(bold=True, size=13)
sm.append([])
sm.append(["Role under test", "Time Clock User (system role)"])
sm.append(["Role id used (staging)", RESULTS['role_id_used']])
sm.append(["Confirmed permissions", ", ".join(RESULTS['fe_permissions_confirmed']) + "; view_mode none; all cross-toggles OFF"])
sm.append(["Total checks", len(rows)])
sm.append([])
sm.append(["Verdict", "Count"])
hdr = sm.max_row
sm.append(["Allowed reads work (200 as expected)", allowed_ok])
sm.append(["Restricted calls correctly blocked (403)", correctly_blocked])
sm.append(["LEAKS — restricted call NOT properly blocked (bug)", leaks])
sm.append(["Inconclusive (endpoint not identified)", inconclusive])
for c in range(1, 3):
    sm.cell(hdr, c).fill = HEAD; sm.cell(hdr, c).font = HF
sm.cell(hdr + 3, 1).fill = RED
sm.append([])
sm.append(["Jira bug verdict", "PARTIALLY CONFIRMED — backend enforces resource View + Delete + Settings-write (403), but has 4 real enforcement gaps: Customer create (201, data created), Work Order create (never 403), Settings read (200), Taxes read (200)."])
sm.append(["Most serious leak", "POST /api/customers/create returns 201 and creates a real customer for a Time Clock user who cannot even VIEW customers (C29459)."])
for col, w in [("A", 46), ("B", 90)]:
    sm.column_dimensions[col].width = w

xlsx = os.path.join(BASE, "CustomRoles_TimeClock-API-Enforcement_2026-07-13.xlsx")
wb.save(xlsx)
print("wrote", xlsx)

# ================= MD =================
md = []
md.append("# Custom Roles — Time Clock Role: Backend API Permission Enforcement (Evidence)")
md.append("")
md.append("**Date:** 2026-07-13  **Env:** STAGING (`api.staging.shopview.com`)  "
          "**Role under test:** system **Time Clock User** "
          f"(id `{RESULTS['role_id_used']}`)")
md.append("")
md.append("**Confirmed permissions (GET /api/auth/me/fe-permissions):** "
          "`scheduleView, timesheetsView, workOrdersView`; view_mode **none**; all cross-toggles OFF — matches spec.")
md.append("")
md.append("**Linked Jira bug:** *BE — Time Clock Role Permissions Not Properly Enforced* "
          "(a Time Clock user can access restricted areas; API calls to restricted areas do not return proper 403).")
md.append("")
md.append("## Verdict")
md.append(f"- Allowed reads working (200 as expected): **{allowed_ok}**")
md.append(f"- Restricted calls correctly blocked (**403**): **{correctly_blocked}**")
md.append(f"- **LEAKS (restricted call NOT properly blocked = the bug): {leaks}**")
md.append(f"- Inconclusive (endpoint not identified): **{inconclusive}**")
md.append("")
md.append("**Jira bug = PARTIALLY CONFIRMED.** The backend correctly enforces resource-level "
          "*View* (403 on Parts/POs/Customers/Reports/Staff/Roles/Integrations/Departments) and "
          "*write* on Settings-change and Customer-delete (403 \"Access denied.\"). But there are "
          "**4 real enforcement gaps**: **Customer create (201 — a real customer is created)**, "
          "**Work Order create (never 403 — processed past the gate)**, **Settings read (200)** and "
          "**Taxes read (200)**. The create endpoints are missing the permission gate their view/delete counterparts have.")
md.append("")
md.append("## Endpoint-by-endpoint results")
md.append("")
md.append("| Area | Endpoint | Method | Expected | Actual (Time Clock) | Verdict | TestRail |")
md.append("|---|---|---|---|---|---|---|")
for r in rows:
    tag = "**LEAK-bug**" if r['verdict'] == 'LEAK-bug' else r['verdict']
    tr = f"[C{r['cid']}]({r['link']})" if r['cid'] else "—"
    md.append(f"| {r['area']} | `{r['endpoint']}` | {r['method']} | {r['expect']} | "
              f"{r['tech']} | {tag} | {tr} |")
md.append("")
md.append("## Notes")
for r in rows:
    if r['note']:
        md.append(f"- **{r['area']} ({r['method']} `{r['endpoint']}`):** {r['note']}")
md.append("")
md.append("## Provenance / cleanup")
md.append("- Live-probed as the Time Clock user; admin baselines confirm every tech-403 endpoint "
          "returns 200 for admin (so the 403s are genuine enforcement, not missing routes).")
md.append("- Two ZZAUTOTEST customers created by the create-leak probe were **deleted**; no work order persisted.")
md.append("- **Env drift:** the shared staging org's role ids were reseeded; the stale "
          "`a0359055` Time-Clock id no longer exists (invalid role_id → 500). Current Time Clock "
          "User = `be58f381-52fd-4958-9961-2d207bd1f09c`.")
md.append("- Raw evidence: `api-timeclock-2026-07-13/results.json`, `probe-results.tsv`.")
mdpath = os.path.join(BASE, "CustomRoles_TimeClock-API-Enforcement_2026-07-13.md")
open(mdpath, "w").write("\n".join(md) + "\n")
print("wrote", mdpath)
print(f"counts: allowed={allowed_ok} blocked={correctly_blocked} leaks={leaks} inconclusive={inconclusive}")
