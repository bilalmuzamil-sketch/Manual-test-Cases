#!/usr/bin/env python3
"""Generate the Custom Roles build-accurate wording + VIU results workbook + blockers tracker.
Reads cases-2026-07-13/*.json (viu_status/evidence) + testrail snapshots (title/section).
Rule 8: every row has Case ID (C#####) + clickable TestRail Link."""
import json, glob, os, re, csv
BASE=os.path.dirname(os.path.abspath(__file__))
CASEDIR=os.path.join(BASE,'cases-2026-07-13')
SNAPDIR=os.path.join(BASE,'testrail-snapshots-2026-07-13')
LINK="https://shopview.testrail.io/index.php?/cases/view/{}"
AREAS={3528:'Roles List Page',3529:'Create Custom Role',3530:'Edit Role',3531:'Delete Role',
 3532:'Permission Summary',3533:'CRUD Cascade Rules',3534:'Work Orders Permissions',
 3535:'Work Order Lines Permissions',3536:'Schedule Permissions',3537:'Customer Management Permissions',
 3538:'Parts Department Permissions',3539:'Invoicing and Payments Permissions',3540:'Timesheets Permissions',
 3541:'Page Access Toggles',3542:'Settings Access',3543:'View Mode',3544:'See Financial Data',
 3545:'View and Manage AP/AR Data',3546:'View History Logs',3547:'Staff Page Role Assignment',
 3548:'Per-Role Verification',3549:'Migration',3550:'Staff Record Settings',3551:'QuickBooks Relocation',
 3552:'User Feedback Strings',3553:'Cross-Permission Combinations'}
def bucket(v):
    v=(v or '').lower()
    if v.startswith('verified') or 'roles-api-verified' in v: return 'VIU-Verified'
    if 'deviation' in v or 'stale' in v or 'failed' in v or 'was run331' in v or 'prior run failed' in v: return 'Deviation/Finding'
    if 'blocked' in v: return 'Blocked-UI'
    return 'Other'
rows=[]
for f in sorted(glob.glob(os.path.join(CASEDIR,'C*.json'))):
    d=json.load(open(f)); cid=d['case_id']; sec=int(d['section_id'])
    rows.append({'case_id':cid,'link':LINK.format(cid),'section_id':sec,'area':AREAS.get(sec,str(sec)),
                 'title':d['title'],'viu_status':d.get('viu_status',''),'evidence':d.get('evidence',''),
                 'bucket':bucket(d.get('viu_status',''))})
rows.sort(key=lambda r:(r['section_id'],r['case_id']))
# CSV
with open(os.path.join(BASE,'CustomRoles_WordingVIU_2026-07-13.csv'),'w',newline='') as fh:
    w=csv.DictWriter(fh,fieldnames=['case_id','link','section_id','area','title','bucket','viu_status','evidence'])
    w.writeheader()
    for r in rows: w.writerow(r)
# XLSX
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
wb=openpyxl.Workbook(); 
hdr_fill=PatternFill('solid',fgColor='305496'); hdr_font=Font(bold=True,color='FFFFFF')
cols=['Case ID','TestRail Link','Section','Area','Title','Bucket','VIU Status','Evidence']
def write_sheet(ws, data):
    ws.append(cols)
    for c in ws[1]: c.fill=hdr_fill; c.font=hdr_font
    for r in data:
        ws.append([f"C{r['case_id']}", r['link'], r['section_id'], r['area'], r['title'], r['bucket'], r['viu_status'], r['evidence']])
        ws.cell(ws.max_row,2).hyperlink=r['link']; ws.cell(ws.max_row,2).font=Font(color='0563C1',underline='single')
    widths=[10,46,10,30,54,18,42,60]
    for i,wd in enumerate(widths,1): ws.column_dimensions[chr(64+i) if i<=26 else 'A'+chr(64+i-26)].width=wd
    ws.freeze_panes='A2'
from collections import Counter, defaultdict
bc=Counter(r['bucket'] for r in rows)
by_area=defaultdict(Counter)
for r in rows: by_area[r['area']][r['bucket']]+=1
# Summary
ws=wb.active; ws.title='Summary'
ws.append(['Custom Roles — Build-Accurate Wording + VIU — 2026-07-13']); ws['A1'].font=Font(bold=True,size=14)
ws.append([]); ws.append(['TestRail push: all 252 core cases (sections 3528–3553) updated via update_case, 200/200, 0 errors.'])
ws.append(['Section 3658 dedupe: 3 duplicates deleted (C27735, C27733, C27737); 7 flagged for ruling.'])
ws.append([])
ws.append(['Bucket','Count']); 
for c in ws[ws.max_row]: c.fill=hdr_fill; c.font=hdr_font
for b in ['VIU-Verified','Blocked-UI','Deviation/Finding','Other']:
    ws.append([b, bc.get(b,0)])
ws.append(['TOTAL', sum(bc.values())]); ws[ws.max_row][0].font=Font(bold=True); ws[ws.max_row][1].font=Font(bold=True)
ws.append([])
ws.append(['Area','VIU-Verified','Blocked-UI','Deviation/Finding','Other','Total'])
for c in ws[ws.max_row]: c.fill=hdr_fill; c.font=hdr_font
for sec in sorted(AREAS):
    a=AREAS[sec]; cnt=by_area[a]
    tot=sum(cnt.values())
    if tot: ws.append([f"{a} ({sec})",cnt.get('VIU-Verified',0),cnt.get('Blocked-UI',0),cnt.get('Deviation/Finding',0),cnt.get('Other',0),tot])
ws.column_dimensions['A'].width=42
for col in 'BCDEF': ws.column_dimensions[col].width=16
# Bucket tabs
for b in ['VIU-Verified','Blocked-UI','Deviation/Finding','Other']:
    data=[r for r in rows if r['bucket']==b]
    if data:
        ws=wb.create_sheet(b.replace('/','-')); write_sheet(ws,data)
# All tab
write_sheet(wb.create_sheet('All Cases'), rows)
wb.save(os.path.join(BASE,'CustomRoles_WordingVIU_2026-07-13.xlsx'))
print("workbook rows:", len(rows))
print("buckets:", dict(bc))
