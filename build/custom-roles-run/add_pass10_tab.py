import openpyxl
from openpyxl.styles import Font, PatternFill
f='Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb=openpyxl.load_workbook(f)
TAB='Pass-10 LIVE (2026-07-16)'
if TAB in wb.sheetnames: del wb[TAB]
ws=wb.create_sheet(TAB)
hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',fgColor='305496')
less=PatternFill('solid',fgColor='FFC7CE'); more=PatternFill('solid',fgColor='FFEB9C')
match=PatternFill('solid',fgColor='C6EFCE'); nv=PatternFill('solid',fgColor='D9D9D9')
rows=[
 ['PASS-10 (2026-07-16) — closing Pass-9 residuals. OBSERVED-LIVE only; NOT VERIFIED never inferred (Rules 10/12).'],
 [],
 ['Staging role','Capability','Prod (LIVE)','Staging (LIVE)','Dual verdict','Confidence / method'],
 # RESIDUAL 1 - prod finance (CLOSED)
 ['Service Manager','Finance payment controls (New Payment/Reverse/Credit)','NOT usable (Finance route -> /no-location)','SHOWN / SHOWN / SHOWN','STAGING-MORE','OBSERVED-LIVE both; role-deterministic bounce (SA-LV renders via /api/invoices/preview 200); invoice-view API=200'],
 ['Parts Manager','Finance payment controls','NOT usable (Finance route -> /no-location)','SHOWN / hidden / SHOWN','STAGING-MORE','OBSERVED-LIVE both; invoice-view API=200 (data readable)'],
 ['Parts Technician','Finance payment controls','NOT usable (Finance route -> /no-location)','SHOWN / hidden / SHOWN','STAGING-MORE','OBSERVED-LIVE both; invoice-view API=200'],
 ['Service Advisor (control)','Finance payment controls','SHOWN (renders via /api/invoices/preview 200)','SHOWN / hidden / SHOWN','Reverse STAGING-LESS','OBSERVED-LIVE (SA-LV self-login)'],
 ['Office User (control)','Finance data access','403 DENY (invoice-view)','SHOWN','STAGING-MORE','OBSERVED-LIVE (API 403 captured)'],
 # RESIDUAL 2 - staging core + return (CLOSED)
 ['Service Manager','Core OK/Not-OK','SHOWN (ok=2/notok=2, Pass-8 ref WO)','SHOWN','MATCH','OBSERVED-LIVE both (stg WO S9-25051 picked cored part AF26154)'],
 ['Foreman','Core OK/Not-OK','SHOWN (Pass-8 ref WO)','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Office User','Core OK/Not-OK','SHOWN (Pass-8 ref WO)','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Parts Technician','Core OK/Not-OK','SHOWN (Pass-8 ref WO)','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Service Manager','Part Return','NOT VERIFIED (no returnable part on prod ref WO)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE; prod data blocker'],
 ['Foreman','Part Return','NOT VERIFIED (prod data)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Office User','Part Return','NOT VERIFIED (prod data)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Parts Technician','Part Return','NOT VERIFIED (prod data)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 # RESIDUAL 3 - See AP/AR (staging all 11; prod Office only)
 ['Administrator','See AP/AR (A/R+A/P Aging reports)','NOT VERIFIED (route bounce)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Service Manager','See AP/AR','NOT VERIFIED (route bounce)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Parts Manager','See AP/AR','NOT VERIFIED (route bounce)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Sales Representative','See AP/AR','NOT VERIFIED (route bounce)','SHOWN','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Office User','See AP/AR','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both (only prod role whose reports route rendered)'],
 ['Senior Service Advisor','See AP/AR','NOT VERIFIED (route bounce)','hidden','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Service Advisor','See AP/AR','NOT VERIFIED (route bounce)','hidden','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Foreman','See AP/AR','NOT VERIFIED (route bounce)','hidden','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Technician','See AP/AR','NOT VERIFIED (route bounce)','hidden','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 ['Parts Technician','See AP/AR','NOT VERIFIED (route bounce)','hidden','NOT VERIFIED (prod)','stg OBSERVED-LIVE (flagged STAGING-LESS in compare; prod side NV)'],
 ['Time Clock User','See AP/AR','NOT VERIFIED (route bounce)','hidden','NOT VERIFIED (prod)','stg OBSERVED-LIVE'],
 # RESIDUAL 3 - build findings
 ['(all)','Bulk Receive','N/A (control absent)','N/A (control absent)','BUILD FINDING','No multi-select/bulk-receive in Custom Roles build; only single Receive (Simple-Flow feature)'],
 ['(all)','Fix Part#','N/A (control absent)','N/A (control absent)','BUILD FINDING','No distinct labeled Fix-Part# control found in build'],
 ['(all)','Assign Vendor','via New-PO/Order-Parts flow','via New-PO/Order-Parts flow','see Order Parts','Not a standalone WO control; part of Order Parts (already 22/22 dual MATCH)'],
 [],
 ['STILL NOT VERIFIED (precise, non-inferable):'],
 ['1) Prod Part Return per role','','','','','no prod WO has a returnable-state part; prod is read-only (needs dev/human-seeded returnable part)'],
 ['2) Prod See AP/AR for non-Office roles','','','','','prod reports route bounces test-staff self-login to /workorders (headless location artifact); AP/AR is FE-gated so API cannot substitute; needs real per-role prod login or attended headful session'],
]
for r in rows: ws.append(r)
ws['A1'].font=Font(bold=True,size=11)
for c in ws[3]: c.font=hdr; c.fill=hf
for row in ws.iter_rows(min_row=4):
    v=(row[4].value or '') if len(row)>4 else ''
    fill=None
    if 'STAGING-LESS' in v: fill=less
    elif 'STAGING-MORE' in v: fill=more
    elif v=='MATCH': fill=match
    elif 'BUILD FINDING' in v: fill=more
    elif 'NOT VERIFIED' in v: fill=nv
    if fill:
        for c in row:
            if c.value is not None or True: c.fill=fill
widths=[30,44,40,26,22,60]
for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
# move new tab right after READ ME (index 1)
wb.move_sheet(TAB, -(len(wb.sheetnames)-2))
wb.save(f)
print('saved; sheets now:', wb.sheetnames)
