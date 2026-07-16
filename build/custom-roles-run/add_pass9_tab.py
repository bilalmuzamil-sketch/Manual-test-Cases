import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
f='Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb=openpyxl.load_workbook(f)
if 'Pass-9 LIVE (2026-07-16)' in wb.sheetnames:
    del wb['Pass-9 LIVE (2026-07-16)']
ws=wb.create_sheet('Pass-9 LIVE (2026-07-16)')
hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',fgColor='305496')
less=PatternFill('solid',fgColor='FFC7CE'); more=PatternFill('solid',fgColor='FFEB9C')
match=PatternFill('solid',fgColor='C6EFCE'); nv=PatternFill('solid',fgColor='D9D9D9')
rows=[
 ['PASS-9 (2026-07-16) — self-service unblocks. OBSERVED-LIVE only; NOT VERIFIED never inferred (Rules 10/12).'],
 [],
 ['Staging role','Capability','Prod (LIVE)','Staging (LIVE)','Dual verdict','Confidence / method'],
 # TASK A - prod finance
 ['Service Advisor','New Payment','SHOWN','SHOWN','MATCH','OBSERVED-LIVE (prod SA-LV self-login; stg switch-user)'],
 ['Service Advisor','Invoice Reverse','SHOWN','hidden','STAGING-LESS (release risk)','OBSERVED-LIVE both'],
 ['Service Advisor','Issue Credit','SHOWN','SHOWN','MATCH','OBSERVED-LIVE both'],
 ['Foreman','New Payment / Finance','HIDDEN (no Finance tab)','SHOWN','STAGING-MORE','OBSERVED-LIVE both'],
 ['Foreman','Issue Credit','HIDDEN','SHOWN','STAGING-MORE','OBSERVED-LIVE both'],
 ['Office User','New Payment / Finance','403 DENY (Pass-8)','SHOWN','STAGING-MORE','OBSERVED-LIVE both'],
 ['Office User','Issue Credit','403 DENY (Pass-8)','SHOWN','STAGING-MORE','OBSERVED-LIVE both'],
 ['Service Manager','New Payment/Reverse/Credit','NOT VERIFIED','SHOWN/SHOWN/SHOWN','NOT VERIFIED (prod)','prod estimate-400 crash; invoice-view=200'],
 ['Parts Manager','New Payment/Reverse/Credit','NOT VERIFIED','SHOWN/hidden/SHOWN','NOT VERIFIED (prod)','prod estimate-400 crash'],
 ['Parts Technician','New Payment/Reverse/Credit','NOT VERIFIED','SHOWN/hidden/SHOWN','NOT VERIFIED (prod)','prod estimate-400 crash'],
 # TASK B - staging line caps
 ['Service Manager','Set Line Status','SHOWN','SHOWN (Approve/Decline)','MATCH','OBSERVED-LIVE (seeded WO S-25619)'],
 ['Service Manager','WO Delete','SHOWN','SHOWN','MATCH','OBSERVED-LIVE'],
 ['Foreman','Set Line Status','SHOWN','SHOWN (Approve/Decline)','MATCH','OBSERVED-LIVE'],
 ['Foreman','WO Delete','SHOWN','hidden','STAGING-LESS','OBSERVED-LIVE'],
 ['Office User','Set Line Status','no status buttons','hidden','MATCH','OBSERVED-LIVE'],
 ['Office User','WO Delete','SHOWN','hidden','STAGING-LESS','OBSERVED-LIVE'],
 ['Parts Technician','Set Line Status','Start/Complete only (Pass-8)','hidden','STAGING-LESS','OBSERVED-LIVE (stg); prod partial'],
 ['Parts Technician','WO Delete','SHOWN','hidden','STAGING-LESS','OBSERVED-LIVE'],
 # residual NV
 ['Service Manager/Parts Manager/Parts Technician','Prod finance (New Payment/Reverse/Credit)','NOT VERIFIED','—','NOT VERIFIED','prod estimate-400 -> /no-location on S1-518/543/517; needs dev fix or per-role prod login'],
 ['Service Mgr/Foreman/Office/Parts Tech','Core OK/Not-OK','—','NOT VERIFIED','NOT VERIFIED','needs picked cored part on a line (deep seed / attended session)'],
 ['Service Mgr/Foreman/Office/Parts Tech','Part Return','—','NOT VERIFIED','NOT VERIFIED','needs a picked part on a line'],
 ['(all)','Assign Vendor / Fix Part# / Bulk Receive / See AP-AR','NOT VERIFIED','NOT VERIFIED','NOT VERIFIED','PO-detail/deliveries/AP-AR in-app nav not driven this pass'],
]
for r in rows: ws.append(r)
ws['A1'].font=Font(bold=True,size=11)
for c in ws[3]: c.font=hdr; c.fill=hf
for row in ws.iter_rows(min_row=4):
    v=(row[4].value or '') if len(row)>4 else ''
    fill=None
    if 'STAGING-LESS' in v: fill=less
    elif 'STAGING-MORE' in v: fill=more
    elif v=='MATCH': fill=match
    elif 'NOT VERIFIED' in v: fill=nv
    if fill:
        for c in row: c.fill=fill
widths=[42,42,26,26,30,52]
for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
# move new tab right after READ ME
wb.move_sheet('Pass-9 LIVE (2026-07-16)', -(len(wb.sheetnames)-2))
wb.save(f)
print('saved; sheets now:', wb.sheetnames)
