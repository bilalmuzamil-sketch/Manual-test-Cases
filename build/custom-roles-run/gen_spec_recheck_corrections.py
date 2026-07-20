#!/usr/bin/env python3
"""Generate the Custom Roles spec-recheck PROPOSED-CORRECTIONS deliverable (Vlad's recheck).
Reads spec-recheck-2026-07-15/findings-G*.json + cases-2026-07-13/C*.json.
Output: xlsx + md + csv with tabs per verdict + Summary. Rule 8: C-ID + TestRail link on every row.
NO TestRail writes — this is a proposal for Bilal & Vlad to agree before pushing."""
import json, glob, os, re, csv
BASE = os.path.dirname(os.path.abspath(__file__))
FDIR = os.path.join(BASE, 'spec-recheck-2026-07-15')
CASEDIR = os.path.join(BASE, 'cases-2026-07-13')
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"
DATE = "2026-07-15"
AREAS = {3528:'Roles List Page',3529:'Create Custom Role',3530:'Edit Role',3531:'Delete Role',
 3532:'Permission Summary',3533:'CRUD Cascade Rules',3534:'Work Orders Permissions',
 3535:'Work Order Lines Permissions',3536:'Schedule Permissions',3537:'Customer Management Permissions',
 3538:'Parts Department Permissions',3539:'Invoicing and Payments Permissions',3540:'Timesheets Permissions',
 3541:'Page Access Toggles',3542:'Settings Access',3543:'View Mode',3544:'See Financial Data',
 3545:'Manage Accounts Payable and Receivable',3546:'View Part History',3547:'Staff Page Role Assignment',
 3548:'Per-Role Verification',3549:'Migration',3550:'Staff Record Settings',3551:'QuickBooks / Integrations',
 3552:'User Feedback Strings',3553:'Cross-Permission Combinations',4091:'API - Time Clock Enforcement'}

def strip(h):
    if not h: return ""
    h = re.sub(r'</li>', '\n', h); h = re.sub(r'</p>', '\n', h)
    h = re.sub(r'<[^>]+>', '', h)
    h = h.replace('&amp;','&').replace('&gt;','>').replace('&lt;','<').replace('&#39;',"'").replace('&quot;','"')
    return re.sub(r'\n{2,}', '\n', h).strip()

# load current case text
cases = {}
for f in glob.glob(os.path.join(CASEDIR, 'C*.json')):
    d = json.load(open(f)); cases[d['case_id']] = d

def current_text(cid, fields):
    d = cases.get(cid, {})
    parts = []
    fmap = {'title':'title','steps':'custom_steps','expected':'custom_expected','preconds':'custom_preconds'}
    for fld in (fields or []):
        key = fmap.get(fld)
        if key and d.get(key):
            parts.append(f"[{fld}] " + (d[key] if fld=='title' else strip(d[key])))
    return "\n\n".join(parts)

def proposed_text(pt):
    if pt is None: return ""
    if isinstance(pt, str): return pt
    if isinstance(pt, dict):
        fmap = {'title':'title','custom_steps':'steps','custom_expected':'expected','custom_preconds':'preconds'}
        out = []
        for k, v in pt.items():
            lbl = fmap.get(k, k)
            out.append(f"[{lbl}] " + (v if k=='title' else strip(v)))
        return "\n\n".join(out)
    return str(pt)

rows = []
for f in sorted(glob.glob(os.path.join(FDIR, 'findings-G*.json'))):
    grp = re.search(r'findings-(G\d)', f).group(1)
    for r in json.load(open(f)):
        cid = r['case_id']; sec = int(r['section_id'])
        cur = r.get('current_excerpt') or ""
        if not cur and r.get('fields_affected'):
            cur = current_text(cid, r['fields_affected'])
        rows.append({
            'group': grp, 'case_id': cid, 'link': LINK.format(cid), 'section_id': sec,
            'area': AREAS.get(sec, str(sec)), 'title': r.get('title',''),
            'verdict': r['verdict'], 'fields': ", ".join(r.get('fields_affected', []) or []),
            'current': cur, 'proposed': proposed_text(r.get('proposed_text')),
            'reason': r.get('reason',''), 'citation': r.get('citation',''),
            'confidence': r.get('confidence',''), 'note': r.get('note',''),
            'live': r.get('live_check',''), 'final': r.get('final_verdict', r['verdict'])})
rows.sort(key=lambda r:(r['section_id'], r['case_id']))
# route tabs by FINAL verdict (post 2026-07-20 live-build label check)
changed = [r for r in rows if r['final'] != 'OK']
updates = [r for r in rows if r['final'] == 'UPDATE']
openq = [r for r in rows if r['final'] == 'OPEN-QUESTION']
oks = [r for r in rows if r['final'] == 'OK']

# ---- CSV (all rows) ----
COLS = ['case_id','link','section_id','area','title','verdict','fields','current','proposed','reason','citation','confidence','note']
with open(os.path.join(BASE, f'CustomRoles_SpecRecheck_Proposed-Corrections_{DATE}.csv'),'w',newline='') as fh:
    w = csv.DictWriter(fh, fieldnames=COLS, extrasaction='ignore'); w.writeheader()
    for r in rows: w.writerow(r)

# ---- XLSX ----
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = openpyxl.Workbook()
hdr_fill = PatternFill('solid', fgColor='1F4E78'); hdr_font = Font(bold=True, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
thin = Border(*[Side(style='thin', color='D0D0D0')]*4)
HEADERS = ['Case ID','TestRail Link','Section','Area','Title','Final verdict','Spec-relative verdict','Field(s)',
           'Current (what is wrong)','Proposed correction','Reason','Spec / ticket citation','Live-build check (2026-07-20)','Confidence']

def fill_sheet(ws, data, include_note=True):
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font; c.alignment = wrap
    for r in data:
        prop = r['proposed']
        if include_note and r['note'] and not prop:
            prop = "(no wording change) " + r['note']
        ws.append([f"C{r['case_id']}", r['link'], r['section_id'], r['area'], r['title'],
                   r['final'], r['verdict'], r['fields'], r['current'], prop, r['reason'], r['citation'], r['live'], r['confidence']])
    widths = [10,42,9,24,38,14,15,15,46,55,40,36,44,11]
    for i, wd in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = wd
    for row in ws.iter_rows(min_row=2):
        for c in row: c.alignment = wrap; c.border = thin
    ws.freeze_panes = 'A2'

# Summary tab
ws = wb.active; ws.title = 'Summary'
ws.append(['Custom Roles (SV-7388) — Spec-Recheck Proposed Corrections'])
ws['A1'].font = Font(bold=True, size=14)
ws.append([f'Generated {DATE} for Vlad\'s spec-recheck. Ground truth: Confluence 565116952 (live export) + all DONE SV-7388 tickets (Sasha rulings, last-update-wins).'])
ws.append(['PROPOSAL ONLY — nothing pushed to TestRail. Counts below are the FINAL verdict AFTER the 2026-07-20 live-build label check.'])
ws.append(['Live check done 2026-07-20 (staging, as Admin): AP/AR toggle = "View and Manage AP/AR Data" (OLD label still in build) -> the 32 AP/AR rename proposals WITHDRAWN (cases are build-accurate). No "View History Logs"/"View Part History" toggle exists in the role editor. See LIVE-LABEL-CHECK-RESULT.md.'])
ws.append([])
ws.append(['Final verdict','Count','Meaning'])
for c in ws[6]: c.fill = hdr_fill; c.font = hdr_font
ws.append(['OK', len(oks), 'Case matches the current spec + ticket rulings + live build; no change. (Includes 32 AP/AR-label cases confirmed build-accurate on 2026-07-20.)'])
ws.append(['UPDATE', len(updates), 'Real drift from the current spec or a Sasha ruling; proposed correction given. Label-only renames were WITHDRAWN by the live check.'])
ws.append(['OPEN-QUESTION', len(openq), 'Spec is silent/self-contradictory or a PO decision is unresolved; needs Bilal/Vlad/PO agreement, not a silent rewrite.'])
ws.append(['TOTAL', len(rows), f'{len(cases)} local case bodies reconciled (core 3528-3553 + API 4091). Spec-relative was 220/44/5; live check moved 23 label rows OK.'])
ws.append([])
ws.append(['Headline drifts (highest priority):'])
for line in [
  'C26503 Office User — REDEFINED by the 7/14 spec update: Work Orders none, Part Sales none, Invoicing V/E/D (Create Invoice hard-blocked). Old case = WO/PS/Invoicing View.',
  'C26496 Service Manager — now HAS Work Orders Delete (SV-8297, reverses SV-8093). Old case = no WO Delete.',
  'C26504 Sales Representative — NOT Reports-only: WO View, WOL View, Customers V/E, Part Sales View (SV-8061). Old case = all areas off.',
  'C26488/C26489 View History Logs — repurposed to "View Part History" (inventory only); WO/line audit log now = WO C&E, story history = WOL View (spec 7/7).',
  'C29457-C29460 Time Clock API — the 4 "BUG" 403-guards assert a contract the PO DECLINED (SV-7958, 2026-07-14): backend leaves these open by design. Flip to accepted behaviour (likely Vlad nightly fails).',
  'C26387/C26388 New WO flow Add Customer/Add Asset — buttons ARE shown & work without Customers C&E (SV-8002). Old expected = hidden (inverted).',
  'LABEL LIVE-CHECK 2026-07-20: build still shows "View and Manage AP/AR Data" (spec 6/10 rename NOT deployed) -> 32 AP/AR rename proposals WITHDRAWN, cases kept as-is (build-accurate); build-vs-spec label gap flagged to dev. The role editor has NO "View History Logs"/"View Part History" toggle (cross-cutting card = 2 toggles) -> C26355/C26359/C27736/C26495/C26502/C26504 corrected to drop the non-existent History toggle; C26488/C26489 = inventory Part History behaviour.',
  'Open questions: C26339 name-uniqueness (spec vs build soft-warn); C26419 restock gate (Catalog vs Vendor Edit); C26459/C26464 Tech-View labor rate (spec sec4 vs SV-8107); C29435 qty edit (SV-8136 vs SV-8055).',
]:
    ws.append([line])
for i,wd in enumerate([16,10,90],1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=wd

for name, data in [('UPDATE (proposed edits)', updates), ('OPEN-QUESTION', openq), ('OK (no change)', oks)]:
    fill_sheet(wb.create_sheet(name), data)

wb.save(os.path.join(BASE, f'CustomRoles_SpecRecheck_Proposed-Corrections_{DATE}.xlsx'))

# ---- MD ----
def mdrow(r):
    cur = (r['current'] or '').replace('\n',' / ').replace('|','\\|')
    prop = (r['proposed'] or (('(no change) '+r['note']) if r['note'] else '')).replace('\n',' / ').replace('|','\\|')
    live = (r['live'] or '').replace('\n',' / ').replace('|','\\|')
    return f"| [C{r['case_id']}]({r['link']}) | {r['area']} | {r['final']} | {r['fields']} | {cur} | {prop} | {r['citation']} | {live} | {r['confidence']} |"
with open(os.path.join(BASE, f'CustomRoles_SpecRecheck_Proposed-Corrections_{DATE}.md'),'w') as fh:
    fh.write(f"# Custom Roles (SV-7388) — Spec-Recheck Proposed Corrections ({DATE})\n\n")
    fh.write("> Vlad's spec-recheck. Ground truth = live Confluence 565116952 + all DONE SV-7388 tickets (Sasha rulings, last-update-wins).\n")
    fh.write("> **PROPOSAL ONLY — nothing pushed to TestRail.** Verdicts below are FINAL, after the 2026-07-20 live-build label check.\n")
    fh.write("> **Live check (2026-07-20, staging as Admin):** AP/AR toggle still shows the OLD label \"View and Manage AP/AR Data\" (spec 6/10 rename not deployed) — 32 AP/AR rename proposals WITHDRAWN (cases are build-accurate). The role editor has NO \"View History Logs\"/\"View Part History\" toggle (cross-cutting card = 2 toggles only). See LIVE-LABEL-CHECK-RESULT.md.\n\n")
    fh.write(f"**Totals (final):** {len(rows)} cases reconciled — {len(oks)} OK, {len(updates)} UPDATE, {len(openq)} OPEN-QUESTION. (Spec-relative before live check: 220 OK / 44 UPDATE / 5 OPEN-QUESTION.)\n\n")
    fh.write("## UPDATE — proposed edits (real drift)\n\n")
    fh.write("| Case | Area | Final | Field(s) | Current (wrong) | Proposed correction | Citation | Live-build check | Conf |\n|---|---|---|---|---|---|---|---|---|\n")
    for r in updates: fh.write(mdrow(r)+"\n")
    fh.write("\n## OPEN-QUESTION — needs Bilal/Vlad/PO agreement\n\n")
    fh.write("| Case | Area | Final | Field(s) | Current | Proposed / question | Citation | Live-build check | Conf |\n|---|---|---|---|---|---|---|---|---|\n")
    for r in openq: fh.write(mdrow(r)+"\n")
    fh.write(f"\n## OK — no change ({len(oks)}) — includes 32 AP/AR-label cases confirmed build-accurate 2026-07-20\n\n")
    fh.write(", ".join(f"C{r['case_id']}" for r in oks) + "\n")

print(f"Reconciled {len(rows)} | OK {len(oks)} | UPDATE {len(updates)} | OPEN-QUESTION {len(openq)}")
print("Wrote xlsx/md/csv:", f'CustomRoles_SpecRecheck_Proposed-Corrections_{DATE}.*')
