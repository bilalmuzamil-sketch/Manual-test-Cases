#!/usr/bin/env python3
"""
Generate the LIVE-VERIFIED prod-vs-staging deliverable — DUAL-VERDICT edition (2026-07-15).
Both sides now have live observations for the roles impersonated. A row gets a REAL
prod-vs-staging verdict ONLY where BOTH sides were observed live; else NOT VERIFIED.
Observed-only (Rules 10 & 12). No inference.
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE='/home/user/Manual-test-Cases/build/custom-roles-run'
STG_DIR=f'{BASE}/live-ui-2026-07-15/staging'
PROD_DIR=f'{BASE}/live-ui-2026-07-15/production'
stg={r['role']:r for r in json.load(open('/tmp/custom-roles/consolidated-staging.json'))}
prod=json.load(open('/tmp/custom-roles/prod-consolidated.json'))

ROLE_ORDER=['Admin','Service Manager','Senior Service Advisor','Parts Manager','Service Advisor',
            'Foreman','Office User','Parts Technician','Sales Representative','Technician','Time Clock User']
PROD_MAP={
 'Admin':('Administrator',''),
 'Service Manager':('Service Manager',''),
 'Senior Service Advisor':('Service Advisor',' (merge: also SA Technician + SA No Reports — NOT VERIFIED)'),
 'Parts Manager':('Parts Manager',''),
 'Service Advisor':('SA Limited View',''),
 'Foreman':('Foreman',''),
 'Office User':('Office User',''),
 'Parts Technician':('Parts Technician',''),
 'Sales Representative':('Sales Representative',' (merge: also Reporting — NOT VERIFIED)'),
 'Technician':('Technician',''),
 'Time Clock User':('Time Clock User','')}
PROD_OBS_KEY={'Administrator':'Administrator','Service Advisor':'Service Advisor','Office User':'Office User',
              'Sales Representative':'Sales Representative','Technician':'Technician','Time Clock User':'Time Clock User'}

def pshot(prodrole,name):
    key=PROD_OBS_KEY.get(prodrole)
    if not key: return ''
    p=f'{PROD_DIR}/{key.replace(" ","_")}/{name}'
    return p if os.path.exists(p) else ''
def sshot(role,name):
    p=f'{STG_DIR}/{role.replace(" ","_")}/{name}'
    return p if os.path.exists(p) else ''

CAPS=[
 ('Send to Portal (WO action)','sendToPortal',True,'Spec Behavior-Changes: Technician "Loses Send to Portal".'),
 ('See Financial Data on WO (Rate/Margin columns)','rateMargin',True,'seeFinancialData cross-toggle.'),
 ('Create/Edit WO Lines ("New Line")','newLine',True,'workOrderLinesCreateAndEdit.'),
 ('Review Work Orders ("Reviewed")','reviewed',True,'woReviewWorkOrders.'),
 ('Finance tab on WO','financeTab',True,'seeFinancialData / invoicing view.'),
 ('WO line/bulk actions "⋮"','lineBulk',False,'Staging aria-label "Line bulk action"; prod old-build control name not confirmed.'),
 ('Send to Terminal (take payment)',None,False,'Prior "no control in build" was a source grep. Payment dialog not driven live on either env.'),
 ('Remove a WO part',None,False,'Line "⋮" submenu not opened per-role on either env.'),
 ('Delete Work Order',None,False,'Top "⋮" menu not opened systematically per-role on either env.'),
 ('Delete WO Line',None,False,'Line "⋮" submenu not opened per-role on either env.'),
 ('Order Parts (WO Parts tab)',None,False,'Parts-tab action not driven live on either env.'),
 ('Approve/Complete a part return',None,False,'Return flow not driven live on either env.'),
 ('Delete / Reverse an invoice',None,False,'Invoicing delete/reverse not driven live on either env.'),
 ('See AP/AR data',None,False,'AP/AR surface not navigated live on either env.'),
]

def prod_val(prodrole,stg_key):
    if prodrole not in prod: return None
    m={'sendToPortal':'portal','rateMargin':'rate','newLine':'newLine','reviewed':'reviewed','financeTab':'finance'}
    k=m.get(stg_key)
    if not k: return None
    return prod[prodrole].get(k)

wb=Workbook()
H=Font(bold=True,color='FFFFFF'); HF=PatternFill('solid',fgColor='2F5496')
WR=Font(bold=True,color='FFFFFF'); WF=PatternFill('solid',fgColor='C00000')
OK=PatternFill('solid',fgColor='C6EFCE'); NV=PatternFill('solid',fgColor='FFC7CE'); NVF=Font(color='9C0006')
YEL=PatternFill('solid',fgColor='FFF2CC'); ORG=PatternFill('solid',fgColor='FCE4D6')
thin=Side(style='thin',color='BFBFBF'); BD=Border(thin,thin,thin,thin)
wrap=Alignment(wrap_text=True,vertical='top')

ws=wb.active; ws.title='READ ME - Coverage & Honesty'
ws.column_dimensions['A'].width=122
def line(t,st=None):
    ws.append([t]); c=ws.cell(ws.max_row,1)
    if st=='banner': c.font=WR; c.fill=WF
    elif st=='h': c.font=Font(bold=True)
    elif st=='c': c.font=Font(name='Consolas')
for t,st in [
 ('LIVE-VERIFIED Prod-vs-Staging Permission Compare — DUAL VERDICT — 2026-07-15','banner'),
 ('',None),
 ('Observed-only rebuild (Standing Rules 10 & 12). A cell is a real result ONLY if the control was rendered on the real','h'),
 ('screen this run with a screenshot captured. Everything else = NOT VERIFIED with a reason. NOTHING inferred from role','n'),
 ('definitions, fe_permissions, atoms, or source code. The prior workbook (…Permission-Gaps…) is SUPERSEDED.','n'),
 ('',None),
 ('BOTH ENVIRONMENTS WERE OBSERVED LIVE THIS RUN:','h'),
 ('  • STAGING: all 11 system roles rendered via genuine impersonation (switch-user x7 + tech role-swap x4), WO-detail','n'),
 ('    controls observed + full-page screenshots.','n'),
 ('  • PRODUCTION: session came back ALIVE. 6 prod roles that had an ACTIVE user were observed live via switch-user on the','n'),
 ('    old-model SPA (Administrator, Office User, Sales Representative, Service Advisor, Technician, Time Clock User),','n'),
 ('    full-page screenshots, all exit-switch-user 200. 8 prod roles had NO active user and were NOT role-swapped (prod is','n'),
 ('    a real system + a dying session) -> those stay NOT VERIFIED.','n'),
 ('',None),
 ('REAL DUAL VERDICTS (both sides observed live) — Send to Portal:','h'),
 ('  Admin: prod SHOWN / staging SHOWN = MATCH','c'),
 ('  Technician: prod HIDDEN / staging HIDDEN = MATCH  (note: spec said Technician "loses Send to Portal", but prod','c'),
 ('     Technician never had it either -> NOT a real loss)','c'),
 ('  Office User: prod SHOWN / staging HIDDEN = STAGING-LESS  <-- REAL release risk (Office loses Send to Portal)','c'),
 ('  Sales Representative: prod HIDDEN / staging HIDDEN = MATCH  (vs the Sales Rep prod component; Reporting NOT VERIFIED)','c'),
 ('  Time Clock User: prod HIDDEN / staging HIDDEN = MATCH','c'),
 ('  Senior Service Advisor: prod Service Advisor HIDDEN / staging SHOWN = STAGING-MORE  (merge caveat: SA Technician +','c'),
 ('     SA No Reports components NOT VERIFIED)','c'),
 ('',None),
 ('STILL NOT VERIFIED (why): Service Manager / Foreman / Parts Manager / Parts Technician / SA Limited View(->staging','h'),
 ('Service Advisor) — no active prod user to impersonate, not role-swapped. Send to Terminal, Remove-a-WO-part, WO Delete,','n'),
 ('WO Lines Delete, Order Parts, part-return, Invoicing delete/reverse, See AP/AR — behind menus/dialogs not driven live','n'),
 ('on either env this run.','n'),
 ('',None),
 ('COVERAGE (capability x role x env cells):','h'),
]:
    line(t,st)
n_roles=len(ROLE_ORDER)
reliable=[c for c in CAPS if c[1] and c[2]]
stg_live=6*n_roles
prod_obs=sum(1 for r in ROLE_ORDER for c in reliable if prod_val(PROD_MAP[r][0],c[1]) is not None)
dual=prod_obs
total=len(CAPS)*n_roles*2
for t in [
 f'  Total cells (caps {len(CAPS)} x roles {n_roles} x envs 2) = {total}',
 f'  STAGING observed LIVE            = {stg_live}',
 f'  PRODUCTION observed LIVE         = {prod_obs}  (6 roles x 5 reliable caps)',
 f'  Cells with a REAL dual verdict   = {dual}  (both sides observed live)',
 f'  PRODUCTION NOT VERIFIED          = {len(CAPS)*n_roles - prod_obs}',
]:
    line(t,'c')
line('',None)
line('CLEANUP: all switch-user impersonations exited (200); staging tech restored to Technician; no prod role-swaps done; '
     'no throwaway data; no TestRail writes.','h')

ws=wb.create_sheet('Live Compare DUAL')
hdr=['Staging Role','Prod role compared','Capability','PROD observed','Staging observed','Direction / verdict',
     'Per-spec?','Confidence','Prod screenshot','Staging screenshot','Notes']
ws.append(hdr)
for i,h in enumerate(hdr,1):
    c=ws.cell(1,i); c.font=H; c.fill=HF; c.alignment=wrap; c.border=BD
for i,w in enumerate([20,34,38,16,16,26,10,26,46,46,44],1):
    ws.column_dimensions[ws.cell(1,i).column_letter].width=w
for role in ROLE_ORDER:
    prole,merge=PROD_MAP[role]
    d=stg[role]
    for label,key,rel,note in CAPS:
        s_state = ('SHOWN' if d.get(key) else 'hidden') if key else 'NOT VERIFIED'
        pv=prod_val(prole,key) if (key and rel) else None
        p_state = ('SHOWN' if pv else 'hidden') if pv is not None else 'NOT VERIFIED'
        if key and rel and pv is not None and d.get(key) is not None:
            sv=bool(d.get(key))
            if pv==sv: verdict='MATCH'
            elif pv and not sv: verdict='STAGING-LESS (prod grants more)'
            else: verdict='STAGING-MORE (staging grants more)'
            conf='DUAL LIVE-OBSERVED'
        else:
            miss=[]
            if p_state=='NOT VERIFIED': miss.append('prod')
            if s_state=='NOT VERIFIED': miss.append('staging')
            verdict='NOT VERIFIED'
            conf='NOT VERIFIED ('+(' & '.join(miss) if miss else 'not comparable')+' not observed)'
        psc=pshot(prole,'WO_ready_for_review.png') if pv is not None else ''
        ssc=sshot(role,'WO_ready_for_review.png') if key else ''
        ws.append([role,prole+merge,label,p_state,s_state,verdict,
                   ('' if verdict=='NOT VERIFIED' else ('yes' if 'MATCH' in verdict else 'see note')),
                   conf,psc,ssc,note])
        r=ws.max_row
        for i in range(1,len(hdr)+1): ws.cell(r,i).alignment=wrap; ws.cell(r,i).border=BD
        ws.cell(r,4).fill = (OK if pv else YEL) if pv is not None else NV
        if pv is None: ws.cell(r,4).font=NVF
        ws.cell(r,5).fill = (OK if d.get(key) else YEL) if key else NV
        if not key: ws.cell(r,5).font=NVF
        if verdict.startswith('STAGING-LESS'): ws.cell(r,6).fill=NV; ws.cell(r,6).font=NVF
        elif verdict.startswith('STAGING-MORE'): ws.cell(r,6).fill=ORG
        elif verdict=='MATCH': ws.cell(r,6).fill=OK
        else: ws.cell(r,6).fill=NV; ws.cell(r,6).font=NVF

ws=wb.create_sheet('Staging Live Grid')
gh=['Staging Role','Perms','View','Send to Portal','See Fin Data','New Line','Reviewed','Line ⋮','Finance tab','Screenshot']
ws.append(gh)
for i,h in enumerate(gh,1): c=ws.cell(1,i);c.font=H;c.fill=HF;c.alignment=wrap;c.border=BD
for i,w in enumerate([20,7,8,14,14,10,10,8,10,50],1): ws.column_dimensions[ws.cell(1,i).column_letter].width=w
yn=lambda v:'SHOWN' if v else 'hidden'
for role in ROLE_ORDER:
    d=stg[role]
    ws.append([role,d['perms'],str(d['view']),yn(d['sendToPortal']),yn(d['rateMargin']),yn(d['newLine']),yn(d['reviewed']),yn(d['lineBulk']),yn(d['financeTab']),sshot(role,'WO_ready_for_review.png')])
    r=ws.max_row
    for i in range(1,len(gh)+1): ws.cell(r,i).alignment=wrap;ws.cell(r,i).border=BD
    for ci,k in [(4,'sendToPortal'),(5,'rateMargin'),(6,'newLine'),(7,'reviewed'),(8,'lineBulk'),(9,'financeTab')]:
        ws.cell(r,ci).fill=OK if d[k] else YEL

ws=wb.create_sheet('Production Live Grid')
ph=['Prod Role','Maps to staging','Perms','Send to Portal','See Fin Data','New Line','Reviewed','Method','Screenshot']
ws.append(ph)
for i,h in enumerate(ph,1): c=ws.cell(1,i);c.font=H;c.fill=HF;c.alignment=wrap;c.border=BD
for i,w in enumerate([22,26,7,14,14,10,10,26,50],1): ws.column_dimensions[ws.cell(1,i).column_letter].width=w
S2P={'Administrator':'Admin','Service Advisor':'Senior Service Advisor (merge)','Office User':'Office User',
     'Sales Representative':'Sales Representative (merge)','Technician':'Technician','Time Clock User':'Time Clock User'}
for prole in ['Administrator','Service Advisor','Office User','Sales Representative','Technician','Time Clock User']:
    v=prod[prole]
    ws.append([prole,S2P[prole],v['perms'],yn(v['portal']),yn(v['rate']),yn(v['newLine']),yn(v['reviewed']),
               'switch-user impersonation (old-model SPA)',pshot(prole,'WO_ready_for_review.png')])
    r=ws.max_row
    for i in range(1,len(ph)+1): ws.cell(r,i).alignment=wrap;ws.cell(r,i).border=BD
    for ci,k in [(4,'portal'),(5,'rate'),(6,'newLine'),(7,'reviewed')]:
        ws.cell(r,ci).fill=OK if v[k] else YEL
for prole in ['Service Manager','Foreman','Parts Manager','Parts Technician','SA Limited View','SA Technician','SA No Reports','Reporting']:
    ws.append([prole,'-','-','NOT VERIFIED','NOT VERIFIED','NOT VERIFIED','NOT VERIFIED','no active prod user; not role-swapped',''])
    r=ws.max_row
    for i in range(1,len(ph)+1): ws.cell(r,i).alignment=wrap;ws.cell(r,i).border=BD;ws.cell(r,i).fill=NV

out=f'{BASE}/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb.save(out); print('WROTE',out)
