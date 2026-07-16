import openpyxl
from openpyxl.styles import Font, PatternFill
f='Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx'
wb=openpyxl.load_workbook(f)
TAB='Pass-12 LIVE (2026-07-16)'
if TAB in wb.sheetnames: del wb[TAB]
ws=wb.create_sheet(TAB)
hdr=Font(bold=True,color='FFFFFF'); hf=PatternFill('solid',fgColor='305496')
less=PatternFill('solid',fgColor='FFC7CE'); more=PatternFill('solid',fgColor='FFEB9C')
match=PatternFill('solid',fgColor='C6EFCE'); nv=PatternFill('solid',fgColor='D9D9D9')
rows=[
 ['PASS-12 (2026-07-16) — CLOSED the last two residuals: (1) Approve/Decline line BOTH envs, all roles; (2) Send-to-Terminal prod org-gate fully characterized. OBSERVED-LIVE only; never inferred (Rules 10/12). ZERO unexplained NOT-VERIFIED remain.'],
 [],
 ['--- APPROVE / DECLINE line (per-line "Approve" green + "Decline" red on a Needs-Approval line) ---'],
 ['Staging role','Capability','Prod (LIVE-OBSERVED)','Staging (LIVE-OBSERVED)','Dual verdict','Confidence / method'],
 ['Admin','Approve/Decline line','SHOWN','SHOWN','MATCH','prod: seeded pending line "SV 8180" on S1-723, role-swap+self-login. staging: real-holder switch-user on WO w/ pending line'],
 ['Service Manager','Approve/Decline line','SHOWN','SHOWN','MATCH','prod role-swap+self-login; staging role-swap+switch-user (2026-07-15 healthy, change 201)'],
 ['Senior Service Advisor','Approve/Decline line','SHOWN (prod SA/SA-Tech/SA-No-Reports/SA-Limited all SHOWN)','SHOWN','MATCH','prod role-swap+self-login; staging real-holder switch-user'],
 ['Service Advisor','Approve/Decline line','SHOWN','SHOWN','MATCH','prod + staging real-holder switch-user'],
 ['Foreman','Approve/Decline line','SHOWN','SHOWN','MATCH','prod role-swap+self-login; staging role-swap+switch-user (2026-07-15)'],
 ['Technician','Approve/Decline line','HIDDEN (line renders, no Approve/Decline)','HIDDEN (line renders, no Approve/Decline)','MATCH','prod role-swap+self-login; staging real-holder switch-user on S10-25071'],
 ['Parts Manager','Approve/Decline line','HIDDEN (line renders, no Approve/Decline)','SHOWN','STAGING-MORE','prod role-swap+self-login (HIDDEN) vs staging real-holder switch-user (SHOWN) — migration difference'],
 ['Parts Technician','Approve/Decline line','HIDDEN','HIDDEN','MATCH','prod role-swap+self-login; staging role-swap+switch-user (2026-07-15)'],
 ['Office User','Approve/Decline line','HIDDEN (line renders, no action buttons)','HIDDEN','MATCH','prod role-swap+self-login; staging role-swap+switch-user (2026-07-15)'],
 ['Sales Representative','Approve/Decline line','HIDDEN (WO detail not accessible — bounce)','HIDDEN (only Start shown)','MATCH','prod role-swap+self-login; staging real-holder switch-user'],
 ['Time Clock User','Approve/Decline line','HIDDEN (WO detail not accessible — bounce)','HIDDEN (WO viewable, no action buttons)','MATCH','prod role-swap+self-login; staging real-holder switch-user on S71-24170'],
 [],
 ['--- SEND TO TERMINAL (payment dialog, invoiced WO) — ORG-DEVICE GATE, fully characterized ---'],
 ['CONCLUSION','Send to Terminal','Send-to-Terminal is ORG-DEVICE gated, NOT role/migration gated.','','ORG-CONFIG','Definitive, evidence-backed.'],
 ['(all invoicing roles)','Send to Terminal','ABSENT for every prod role — prod org has NO terminal device AND no UI path to add one','SHOWN for invoicing roles (Admin/Parts Mgr/Senior SA...) — staging org HAS a terminal device','ORG-CONFIG (not a role/permission diff)','prod admin Settings full nav + Payment Methods page screenshots show NO Terminals/Card-Readers/Devices section; New Payment Method creates NAMED methods only; all terminal APIs (/api/terminals, /api/payment-terminals, /api/card-readers, /api/stripe/terminals) = 404; admin holds no terminal fe_permission. Provisioning a terminal requires external payment-processor/hardware registration (not a UI action). Migrating roles does NOT change Send-to-Terminal access.'],
 [],
 ['FINAL STATUS: All role-permission capability cells for the priority + broad set are OBSERVED-LIVE across both envs. The only non-role cell is Send-to-Terminal, now a FULLY-CHARACTERIZED org-device-config verdict (not an unexplained NOT-VERIFIED).'],
]
for r in rows: ws.append(r)
ws['A1'].font=Font(bold=True,size=11)
for c in ws[4]: c.font=hdr; c.fill=hf
for row in ws.iter_rows(min_row=5):
    v=(row[4].value or '') if len(row)>4 else ''
    fill=None
    if 'STAGING-LESS' in v: fill=less
    elif 'STAGING-MORE' in v: fill=more
    elif v.startswith('MATCH'): fill=match
    elif 'ORG-CONFIG' in v: fill=nv
    if fill:
        for c in row: c.fill=fill
widths=[26,26,52,44,26,90]
for i,w in enumerate(widths,1): ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
wb.move_sheet(TAB, -(len(wb.sheetnames)-2))
wb.save(f)
print('saved; sheets:', wb.sheetnames[:5])
