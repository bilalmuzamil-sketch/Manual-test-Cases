#!/usr/bin/env python3
"""
Build the SUPPORT-facing Custom Roles Settings/Permissions Quick-Reference.

Outputs (same directory as this script):
  - CustomRoles_Support_QuickReference.xlsx   (primary, one tab per section)
  - CustomRoles_Support_QuickReference.md     (readable mirror)

Everything here is grounded in the LIVE behaviour consolidated in
custom-roles-current-state.md section (a). Spec-pending behaviour is kept out of
the customer-facing tabs (only a small clearly-labelled "Coming soon" note).
Internal caveats (SV-8193, FE-gate nuance) live ONLY in the "Internal notes" tab.

No secrets, no test-case IDs, no TestRail/Jira IDs in the customer-facing tabs.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "CustomRoles_Support_QuickReference.xlsx")
MD = os.path.join(HERE, "CustomRoles_Support_QuickReference.md")

# ----------------------------------------------------------------------------
# CONTENT (single source of truth for both outputs)
# ----------------------------------------------------------------------------

COVER = [
    ("Custom Roles & Permissions — Support Quick-Reference", ""),
    ("", ""),
    ("What this is",
     "A plain-English aid for support staff to answer customer questions about "
     "ShopView roles, permissions and settings. Everything in the customer-facing "
     "tabs reflects how the product actually behaves today."),
    ("How to use it",
     "Use 'Permission catalog' to look up what a setting does; 'Role capability "
     "overview' for a quick can/can't by role; 'Customer FAQ' for ready answers; "
     "'Quick troubleshooting' when a customer says a role can't do something."),
    ("Important",
     "The 'Internal notes' tab is NOT for customers. Do not read it out or paste "
     "it to a customer."),
    ("Tabs",
     "1) How Custom Roles work  2) Permission catalog  3) Role capability "
     "overview  4) Customer FAQ  5) Quick troubleshooting  6) Internal notes "
     "(not for customers)"),
]

# --- Tab 1: How Custom Roles work -------------------------------------------
HOW_HEADERS = ["Topic", "In plain English"]
HOW_ROWS = [
    ["System roles vs custom roles",
     "ShopView ships with 11 built-in 'system' roles (Administrator, Office, Time "
     "Clock, Service Manager, Service Advisor, Senior Service Advisor, Foreman, "
     "Technician, Parts Manager, Parts Technician, Sales Representative). An admin "
     "can also create their own 'custom' roles with any mix of permissions."],
    ["Which roles can be changed",
     "No system role can be deleted. Office and Time Clock are locked (view-only). "
     "Every other system role, including Administrator, can be edited — but "
     "Administrator always keeps full access (its switches are shown but locked "
     "on)."],
    ["How an admin creates a role",
     "Settings > Roles & Permissions > 'Create custom role'. You can start from a "
     "template (pre-fills the switches) or from blank. Give it a name and turn on "
     "at least one permission, then Create. A role won't save without a name and "
     "at least one permission."],
    ["How an admin edits a role",
     "On the Roles list, click the pencil, change the switches, then Save. A "
     "'Confirm Permission Updates' box lists exactly what you added or removed. "
     "'Reset to Template' puts a template-based role back to its defaults."],
    ["Deleting a role",
     "A role can only be deleted when no users are assigned to it. If users are "
     "still on it, reassign them first — the Delete option stays hidden/disabled "
     "until then."],
    ["Assigning a role to a person",
     "Settings > Staff > open the person > Edit Staff Member > pick a Role. Roles "
     "are grouped as SYSTEM and CUSTOM, and an eye icon previews what a role can "
     "do."],
    ["Role change needs a re-login",
     "When you change someone's role, their current session ends right away and "
     "the new permissions take effect the next time they log in. That forced "
     "logout is normal, not an error."],
    ["Full view vs Tech view",
     "Every role has a work-order 'view mode'. FULL view shows the complete work "
     "order, including the per-line Approve/Decline buttons and the bulk-approve "
     "control. TECH view is a simpler, technician-focused screen that HIDES the "
     "Approve/Decline actions (and the bulk-approve control) — good for people who "
     "do the work but shouldn't approve. Note: seeing dollar amounts is a separate "
     "setting ('See Financial Data'), not the view mode."],
    ["The two-layer idea",
     "Some things need TWO things switched on: the permission itself AND a related "
     "toggle. The main example is money: a user only sees dollar amounts if 'See "
     "Financial Data' is on, on top of having access to the work order/part/"
     "invoice."],
    ["How switches cascade",
     "Turning on Delete for a resource automatically turns on Create & Edit and "
     "View for it. Turning View off clears everything under that resource. You "
     "can't turn on a sub-item while its parent View is off."],
]

# --- Tab 2: Permission catalog ----------------------------------------------
CAT_HEADERS = ["Permission / setting", "What it controls (plain English)",
               "View / Create & Edit / Delete meaning", "Depends on",
               "Where it appears in the app"]
CAT_ROWS = [
    ["Work Orders",
     "Access to work orders. Also lets a user create and edit ANY note on a work "
     "order (not just their own).",
     "View = open work orders and read notes. Create & Edit = create/change work "
     "orders and any note. Delete = delete a work order and delete ANY note.",
     "-",
     "Work Orders area; a work order's Notes tab."],
    ["Work Order Lines",
     "The individual job lines on a work order — marking lines OK/Not-OK, the line "
     "'story'/history, approving lines (in Full view).",
     "No separate View (it follows Work Orders: View). Create & Edit = add/change "
     "lines and line work. Delete = remove lines (a line can be deleted in any "
     "status except Complete).",
     "Work Orders: View (for visibility)",
     "Inside a work order > Lines tab."],
    ["Schedule",
     "The scheduling calendar (shows all users' appointments).",
     "View = see the calendar. Create & Edit = add/change appointments.",
     "-",
     "Schedule in the top nav."],
    ["Customer Management",
     "Customer records.",
     "View = see customers. Create & Edit = add/change customers. Delete = delete "
     "customers.",
     "-",
     "Customers in the top nav."],
    ["Parts Department",
     "The whole parts area — Catalog, Inventory, Vendors & Orders, and Part Sales "
     "(these are one combined card, not separate ones).",
     "View = browse parts. Create & Edit = add/change parts, vendors, inventory. "
     "Delete = remove them.",
     "See Financial Data (to see cost/price columns)",
     "Parts in the top nav (Catalog, Inventory, Vendors, Part Sales)."],
    ["Invoicing & Payments",
     "The Finance tab on a work order — invoices, payments, deposits, credits.",
     "View = see finance/invoices. Create & Edit = create invoices/payments. "
     "Delete = reverse/remove payments and transactions.",
     "See Financial Data (the Finance tab won't appear without it)",
     "Inside a work order > Finance tab; customer Payments tab."],
    ["Timesheets",
     "Timesheet activity.",
     "View = see timesheets. Create & Edit = edit them. (There is no Delete for "
     "timesheets.)",
     "-",
     "Reports > Timesheet Activities."],
    ["Reports",
     "The Reports area. This is one all-or-nothing switch (no separate view/edit).",
     "On = full access to the Reports area. Off = no Reports.",
     "-",
     "Reports in the top nav."],
    ["Settings",
     "Access to the admin/settings area, split into 6 sub-switches: App Settings, "
     "Service, Parts, Finance, Data Import, and View/Manage Wages.",
     "Each sub-switch on/off controls its part of Settings.",
     "-",
     "Settings sidebar (Roles, Staff, Locations, Labor Rates, Payment Methods, "
     "Wages, etc.)."],
    ["Order Parts (work-order sub-permission)",
     "Ordering parts on a work order. It also controls the work order's PARTS tab: "
     "with Order Parts on, the Parts tab appears; with it off, the Parts tab is "
     "hidden.",
     "Single on/off sub-permission under Work Orders.",
     "Goes together with See Financial Data (ordering shows prices/costs)",
     "Inside a work order (the Parts tab and the ordering actions)."],
    ["Pick Parts (work-order sub-permission)",
     "Picking in-stock parts for a work order. (Does not by itself show the Parts "
     "tab — the Parts tab is controlled by Order Parts.)",
     "Single on/off sub-permission under Work Orders.",
     "-",
     "Inside a work order (Pick action on in-stock parts)."],
    ["Review Work Orders (work-order sub-permission)",
     "The Review sign-off step on a work order. Until Review is done, Create "
     "Invoice stays disabled.",
     "Single on/off sub-permission under Work Orders.",
     "-",
     "Inside a work order (Review action)."],
    ["See Financial Data",
     "Whether the user sees dollar amounts on work orders, parts and invoices, and "
     "whether the Finance tab appears.",
     "Cross-cutting toggle (on/off).",
     "-",
     "Everywhere money shows: work orders, parts cost/price columns, invoices, the "
     "Finance tab."],
    ["Manage Accounts Payable & Receivable (AP/AR)",
     "The 7 sensitive customer fields (Credit Terms, Credit Limit, Default Labor "
     "Rate, Default Shop Supplies, Min & Max, Taxes, 'PO is required') and the "
     "customer AP/AR tabs. Today it also controls access to the AP/AR aging "
     "reports.",
     "Cross-cutting toggle (on/off).",
     "-",
     "Customer record (sensitive fields + AP/AR tabs); AP/AR aging reports."],
    ["View History Logs",
     "The work-order history. This is one combined history that shows BOTH "
     "work-order-level changes (e.g. the work order being created) AND line-level "
     "changes (e.g. a line being created/changed — the line 'story'). Work orders "
     "only — there is no history for Part Sales or Purchase Orders.",
     "Cross-cutting toggle (on/off).",
     "-",
     "Inside a work order (the History tab / Audit Log; line story)."],
]

# --- Tab 3: Role capability overview ----------------------------------------
ROLE_HEADERS = ["System role", "Typical use", "Sees money (financial data)?",
                "Work orders", "Parts", "Reports / AP-AR", "Notes"]
ROLE_ROWS = [
    ["Administrator", "Full admin", "Yes",
     "Full (create/edit/delete)", "Full", "Full",
     "Full access; cannot be deleted; always keeps full access."],
    ["Service Manager", "Runs the service dept", "Yes",
     "Full", "Yes", "Yes",
     "Broad access across work orders, customers, invoicing."],
    ["Senior Service Advisor", "Senior front-desk", "Yes",
     "Yes", "Yes", "Yes (Reports on)",
     "Wide access including Reports."],
    ["Service Advisor", "Front-desk advisor", "Yes",
     "Yes", "Yes", "Partial",
     "Day-to-day work-order and customer handling."],
    ["Foreman", "Shop lead", "Yes",
     "Yes", "Yes", "Partial",
     "Runs the floor; work orders and parts."],
    ["Technician", "Hands-on tech", "No (no See Financial Data by default)",
     "Works lines (Tech view; no Approve)", "Pick only (no ordering)", "No",
     "Does the work; no money visibility by default."],
    ["Parts Manager", "Runs parts dept", "Yes",
     "Supports work orders", "Full parts", "Partial",
     "Catalog, inventory, vendors, ordering."],
    ["Parts Technician", "Parts helper", "Depends on config",
     "Supports parts on work orders", "Parts (limited)", "No",
     "Receives/handles parts."],
    ["Office", "Back office (locked role)", "Config-dependent",
     "View-focused", "Limited", "Can open AP/AR + reports",
     "Locked/non-editable system role."],
    ["Sales Representative", "Reporting/sales", "Yes",
     "No work-order CRUD", "No", "Reports + AP/AR only",
     "Exactly Reports + See Financial Data + Manage AP/AR."],
    ["Time Clock", "Clock-in only (locked role)", "No",
     "Read-only minimal (clock UI)", "No", "No",
     "Locked/non-editable; minimal read access for the clock screen."],
]
ROLE_NOTE = ("High-level guide only. Exact capabilities depend on the shop's "
             "configuration and any custom roles. 'Partial' means it depends on "
             "the specific switches set for that shop.")

# --- Tab 4: Customer FAQ -----------------------------------------------------
FAQ_HEADERS = ["Customer question", "Answer support can use"]
FAQ_ROWS = [
    ["Why can a role order parts but the person still needs to see prices?",
     "Ordering parts and seeing financial data go hand in hand. Ordering involves "
     "prices and costs, so a user who orders parts also needs 'See Financial Data' "
     "turned on to see those amounts. Set both together."],
    ["Which permission controls the Parts area / parts work?",
     "The 'Parts Department' permission covers Catalog, Inventory, Vendors & "
     "Orders and Part Sales — they're one combined permission. To see cost and "
     "price columns there, the user also needs 'See Financial Data'."],
    ["What's the difference between Full view and Tech view?",
     "Full view is the complete work-order screen, including the per-line "
     "Approve/Decline buttons and the bulk-approve control. Tech view is a simpler, "
     "technician-focused screen that hides those approve actions — ideal for people "
     "who do the work but shouldn't approve. Seeing dollar amounts is a separate "
     "setting ('See Financial Data'), not the view mode."],
    ["Which permission shows or hides the Parts tab on a work order?",
     "The 'Order Parts' permission (under Work Orders). With Order Parts on, the "
     "work order's Parts tab appears; with it off, the Parts tab is hidden. (Pick "
     "Parts on its own does not show the tab.)"],
    ["How do I stop a role from seeing prices / dollar amounts?",
     "Turn off 'See Financial Data' for that role. With it off, dollar amounts are "
     "hidden on work orders, parts and invoices, and the Finance tab won't appear "
     "at all."],
    ["How do I control who sees the AP/AR aging reports?",
     "Right now, access to the AP/AR aging reports is controlled by the 'Manage "
     "Accounts Payable & Receivable' permission (in addition to Reports access). "
     "Turn that off for roles that shouldn't see aging reports."],
    ["How do I hide the sensitive customer fields (credit terms, limits, etc.)?",
     "Those 7 fields (Credit Terms, Credit Limit, Default Labor Rate, Default Shop "
     "Supplies, Min & Max, Taxes, 'PO is required') are controlled by 'Manage "
     "Accounts Payable & Receivable'. Turn it off to hide them; the basic customer "
     "fields stay visible."],
    ["Why can a user edit or delete a note they didn't create?",
     "That's how work-order notes work. Anyone with Work Orders: View (create/edit) "
     "can create and edit ANY note on a work order, and anyone with Work Orders: "
     "Delete can delete ANY note — notes aren't limited to their author."],
    ["Why can't this role delete a role I created?",
     "A role can only be deleted when no users are assigned to it. Reassign the "
     "users off the role first, then the Delete option becomes available."],
    ["Why did a user get logged out right after I changed their role?",
     "That's expected. Changing someone's role ends their current session "
     "immediately; the new permissions apply the next time they log in."],
    ["Why won't my new custom role save?",
     "A role needs a name AND at least one permission turned on. Also, the name "
     "must be unique — a duplicate name is rejected, and a duplicate set of "
     "permissions prompts a 'similar role already exists' confirmation."],
    ["Can the Administrator role be edited or removed?",
     "Administrator can be opened but always keeps full access (its switches are "
     "locked on), and no system role — including Administrator — can be deleted. "
     "Only Office and Time Clock are fully locked/view-only."],
    ["Where does 'View Permissions' live for a role?",
     "For Office and Time Clock there's an eye icon in the Actions column. For "
     "every other role (system or custom), View Permissions is inside the "
     "three-dot menu."],
    ["Does Part Sales or a Purchase Order have a history log?",
     "No. The history/audit log ('View History Logs') covers work orders only — "
     "both the work-order-level history and the line-level story. There is no "
     "history log for Part Sales or Purchase Orders."],
    ["What does 'Review Work Orders' do?",
     "It controls the Review sign-off step on a work order. Until the Review step "
     "is completed, the Create Invoice button stays disabled."],
    ["Turning on Delete also turned on Create & Edit and View — is that a bug?",
     "No, that's intended. For any resource, Delete requires Create & Edit and "
     "View, so switching Delete on turns those on too. Likewise, turning View off "
     "clears everything under that resource."],
]

# --- Tab 5: Quick troubleshooting -------------------------------------------
TS_HEADERS = ["Customer says...", "Check this permission / toggle"]
TS_ROWS = [
    ["\"They can't see any prices / dollar amounts.\"",
     "Turn on 'See Financial Data' for the role."],
    ["\"There's no Finance tab on the work order.\"",
     "Needs 'See Financial Data' AND Invoicing & Payments: View. Without See "
     "Financial Data the Finance tab never shows."],
    ["\"They can't create an invoice.\"",
     "Needs Invoicing & Payments: Create & Edit + See Financial Data. Also the "
     "work order's Review step must be completed and all parts received with a real "
     "part number."],
    ["\"There's no Parts tab on the work order.\"",
     "Turn on 'Order Parts' (under Work Orders) — it controls the Parts tab. Pick "
     "Parts alone does not show the tab."],
    ["\"They can't order parts on a work order.\"",
     "Turn on 'Order Parts' (under Work Orders) together with 'See Financial "
     "Data'."],
    ["\"They can't pick in-stock parts.\"",
     "Turn on 'Pick Parts' (under Work Orders)."],
    ["\"They can't approve work-order lines.\"",
     "Set the role to Full view (Tech view hides Approve) and give Work Order "
     "Lines: Create & Edit."],
    ["\"They can't see or edit the parts catalog / inventory / vendors.\"",
     "Give 'Parts Department' View or Create & Edit. For cost/price columns also "
     "turn on See Financial Data."],
    ["\"They can't see the sensitive customer fields (credit terms/limits).\"",
     "Turn on 'Manage Accounts Payable & Receivable'."],
    ["\"They can't open the AP/AR aging reports.\"",
     "Give Reports access AND 'Manage Accounts Payable & Receivable' (aging "
     "reports currently need both)."],
    ["\"They can't see the work-order history / audit log.\"",
     "Turn on 'View History Logs' (work orders only)."],
    ["\"They can't get into Settings / Staff / Roles.\"",
     "Turn on the relevant Settings sub-switch (App Settings covers Roles, Staff, "
     "Locations, Departments, Taxes)."],
    ["\"They can't delete a work order.\"",
     "Needs Work Orders: Delete. Note a work order must be moved to Uncomplete "
     "before it can be deleted."],
    ["\"They can't reverse an invoice.\"",
     "Reversing a work-order invoice needs Work Orders: Delete."],
    ["\"They can't reverse/delete a payment.\"",
     "Needs Invoicing & Payments: Delete."],
    ["\"The role won't delete.\"",
     "It has users assigned — reassign them to another role first, then delete."],
    ["\"The Customers menu / Parts menu disappeared for them.\"",
     "The parent permission (Customer Management: View / Parts Department) is off — "
     "turning a parent off hides the whole area."],
]

COMING_SOON = ("Coming soon (not live yet — do not promise these today): the AP/AR "
               "aging reports are planned to move to the Reports permission only; "
               "the role editor is planned to auto-link Order Parts with See "
               "Financial Data; and QuickBooks is planned to move under Finance "
               "settings. Until then, describe today's behaviour above.")

# --- Tab 6: Internal notes (NOT for customers) ------------------------------
INT_HEADERS = ["Internal caveat", "Detail"]
INT_ROWS = [
    ["NOT FOR CUSTOMERS",
     "This tab is for internal support/QA only. Do not read out or paste to "
     "customers."],
    ["Completed-inspection delete defect (SV-8193)",
     "A role with Work Order Lines View/Edit but NOT Delete (confirmed for "
     "Technician and Parts Manager) can actually DELETE a completed inspection — "
     "the bin is shown and the backend allows it (delete succeeds; inspection is "
     "gone). It is not just a wrongly-shown button. Fix is pending confirmation. "
     "Do not tell a customer this is prevented."],
    ["Front-end gate vs backend enforcement",
     "The backend only truly enforces resource-level View / Create & Edit. "
     "Granular Delete, the work-order sub-toggles, the cross-toggles and view mode "
     "are front-end display gates — the UI hides the control, but the underlying "
     "API often does not block it. So 'this role can't do X' based on a granular "
     "Delete or sub-toggle is a UI-level statement only."],
    ["Examples of the FE-only gate",
     "Editing/removing catalog parts succeeded via the API for view-only/no-delete "
     "roles; changing the service advisor succeeded for a Technician despite the "
     "button being hidden; missing sub-toggles return a validation error, not a "
     "hard 'access denied'."],
    ["Add Customer/Asset on New Work Order",
     "The Add Customer and Add Asset buttons on the New Work Order dialog appear "
     "and work even when Customer Management: Create & Edit is off (known FE "
     "gap)."],
    ["Reports Sales report + See Financial Data",
     "With See Financial Data off, the Reports > Sales report still shows financial "
     "figures. This appears to be by design (See Financial Data is scoped to work "
     "orders/parts/invoices; Reports is gated only by the Reports permission), but "
     "it is a common confusion point."],
    ["Spec-pending items (not live)",
     "AP/AR aging reports still require Manage AP/AR (the move to Reports-only "
     "isn't live); the role editor does not yet auto-link Order Parts with See "
     "Financial Data; QuickBooks is absent from Finance settings and the "
     "Integrations group is still present; various migration/rename UI items are "
     "not observable. Do not describe these to customers as working."],
    ["Verified in live VIU (2026-07-07)",
     "Confirmed on staging: (a) Order Parts controls the WO Parts tab (on shows / "
     "off hides; gated by Order Parts, not Pick Parts); (b) View History Logs is "
     "one feed carrying both WO-level and line-level events, work-orders-only "
     "(part-sales/PO history endpoints 404); (c) Full-vs-Tech view: Tech view hides "
     "per-line Approve/Decline and the bulk-approve icon and relabels hours 'Total "
     "Tech Hours' — money visibility is governed by See Financial Data, not view "
     "mode; (d) authoritative permission catalog captured (41 Admin atoms) in "
     "permission-catalog-source.json. There is NO separate 'Add Parts' permission "
     "atom."],
    ["Still not fully verified",
     "Core OK/Not-OK: the core UI surface is confirmed live (Parts grid 'Core' "
     "column; inventory core fields), but the OK/Not-OK action itself was not driven "
     "end-to-end (needs a manually-seeded received core part). Governed per spec by "
     "Work Order Lines: Create & Edit."],
]

# ----------------------------------------------------------------------------
# XLSX BUILD
# ----------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
INT_FILL = PatternFill("solid", fgColor="C00000")
NOTE_FONT = Font(italic=True, size=10, color="595959")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, ncols, row=1, fill=HDR_FILL):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = HDR_FONT
        cell.alignment = WRAP
        cell.border = BORDER


def add_table(ws, headers, rows, widths, start=1):
    for c, h in enumerate(headers, 1):
        ws.cell(row=start, column=c, value=h)
    style_header(ws, len(headers), row=start)
    r = start + 1
    for row in rows:
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
        r += 1
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return r


def build_xlsx():
    wb = Workbook()

    # Cover
    ws = wb.active
    ws.title = "Start Here"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 95
    r = 1
    for label, text in COVER:
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=text)
        a.alignment = WRAP
        b.alignment = WRAP
        if r == 1:
            a.font = TITLE_FONT
        else:
            a.font = Font(bold=True)
        r += 1

    # Tab 1
    ws = wb.create_sheet("1. How Custom Roles Work")
    add_table(ws, HOW_HEADERS, HOW_ROWS, [30, 105])

    # Tab 2
    ws = wb.create_sheet("2. Permission Catalog")
    add_table(ws, CAT_HEADERS, CAT_ROWS, [30, 55, 55, 32, 40])

    # Tab 3
    ws = wb.create_sheet("3. Role Capability Overview")
    end = add_table(ws, ROLE_HEADERS, ROLE_ROWS, [22, 22, 28, 28, 20, 22, 42])
    n = ws.cell(row=end + 1, column=1, value="Note: " + ROLE_NOTE)
    n.font = NOTE_FONT
    n.alignment = WRAP
    ws.merge_cells(start_row=end + 1, start_column=1, end_row=end + 1, end_column=7)

    # Tab 4
    ws = wb.create_sheet("4. Customer FAQ")
    end = add_table(ws, FAQ_HEADERS, FAQ_ROWS, [50, 90])
    n = ws.cell(row=end + 1, column=1, value=COMING_SOON)
    n.font = NOTE_FONT
    n.alignment = WRAP
    ws.merge_cells(start_row=end + 1, start_column=1, end_row=end + 1, end_column=2)

    # Tab 5
    ws = wb.create_sheet("5. Quick Troubleshooting")
    end = add_table(ws, TS_HEADERS, TS_ROWS, [48, 92])
    n = ws.cell(row=end + 1, column=1, value=COMING_SOON)
    n.font = NOTE_FONT
    n.alignment = WRAP
    ws.merge_cells(start_row=end + 1, start_column=1, end_row=end + 1, end_column=2)

    # Tab 6 (internal)
    ws = wb.create_sheet("6. Internal (NOT customer)")
    ws.cell(row=1, column=1,
            value="INTERNAL ONLY — NOT FOR CUSTOMERS")
    ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=12)
    ws.cell(row=1, column=1).fill = INT_FILL
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    add_table(ws, INT_HEADERS, INT_ROWS, [34, 106], start=2)
    style_header(ws, 2, row=2, fill=INT_FILL)

    # Freeze header rows + row heights
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    wb.save(XLSX)
    print("Wrote", XLSX)


# ----------------------------------------------------------------------------
# MARKDOWN BUILD
# ----------------------------------------------------------------------------
def md_table(headers, rows):
    out = ["| " + " | ".join(headers) + " |",
           "|" + "|".join(["---"] * len(headers)) + "|"]
    for row in rows:
        cells = [str(c).replace("\n", " ").replace("|", "\\|") for c in row]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def build_md():
    p = []
    p.append("# Custom Roles & Permissions — Support Quick-Reference\n")
    p.append("_Plain-English aid for support staff. The customer-facing sections "
             "reflect how ShopView behaves today. The final 'Internal notes' "
             "section is NOT for customers._\n")
    p.append("**Sections:** 1) How Custom Roles work · 2) Permission catalog · "
             "3) Role capability overview · 4) Customer FAQ · 5) Quick "
             "troubleshooting · 6) Internal notes (not for customers)\n")

    p.append("\n## 1. How Custom Roles Work\n")
    p.append(md_table(HOW_HEADERS, HOW_ROWS))

    p.append("\n## 2. Permission Catalog\n")
    p.append(md_table(CAT_HEADERS, CAT_ROWS))

    p.append("\n## 3. Role Capability Overview\n")
    p.append(md_table(ROLE_HEADERS, ROLE_ROWS))
    p.append("\n_Note: " + ROLE_NOTE + "_\n")

    p.append("\n## 4. Customer FAQ\n")
    p.append(md_table(FAQ_HEADERS, FAQ_ROWS))
    p.append("\n> " + COMING_SOON + "\n")

    p.append("\n## 5. Quick Troubleshooting\n")
    p.append(md_table(TS_HEADERS, TS_ROWS))
    p.append("\n> " + COMING_SOON + "\n")

    p.append("\n---\n")
    p.append("\n## 6. Internal Notes (NOT FOR CUSTOMERS)\n")
    p.append("> **Internal support/QA only. Do not read out or paste to "
             "customers.**\n")
    p.append(md_table(INT_HEADERS, INT_ROWS))

    with open(MD, "w") as f:
        f.write("\n".join(p) + "\n")
    print("Wrote", MD)


if __name__ == "__main__":
    build_xlsx()
    build_md()
