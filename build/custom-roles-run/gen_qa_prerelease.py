#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CustomRoles_QA-PreRelease-Checklist_2026-07-16.xlsx generator.

DETAILED QA-team pre-release analysis workbook for the ShopView Custom Roles
release, derived from the spec-compliance-annotated prod-vs-staging comparison.
Complements (does NOT replace) the executive file
CustomRoles_Release-Readiness_Prod-vs-Staging_EXEC_2026-07-16.xlsx.

GROUND TRUTH (every row traces to these):
  - Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx (Full Dual Matrix + Pass-11/
    Pass-12/Approve-Decline/Send to Terminal/Parts-Module/New-WO tabs +
    Spec-Standing Conformance tab)
  - spec-conformance/spec-truth-table.md (citation-backed spec truth table)
  - ../custom-roles-spec-update/current-spec-2026-07-15.md (canonical spec)
  - evidence: live-ui-2026-07-15/ and live-ui-2026-07-16/

Built-in ACCURACY GATE (runs after building, re-opens the written file):
  (a) every P1/P2 (and P3) row's role+capability+direction matches the source
      workbook cell(s) it cites;
  (b) every spec citation quote appears verbatim in the canonical spec file;
  (c) every TestRail C-id cited exists as a filename in cases-2026-07-13/;
  (d) Summary counts match the tab row counts;
  (e) zero forbidden placeholder strings anywhere.
Exits non-zero if any gate fails.
"""
import os
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
SRC_XLSX = os.path.join(BASE, "Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx")
SPEC_MD = os.path.join(BASE, "..", "custom-roles-spec-update", "current-spec-2026-07-15.md")
TRUTH_MD = os.path.join(BASE, "spec-conformance", "spec-truth-table.md")
CASES_DIR = os.path.join(BASE, "cases-2026-07-13")
OUT_XLSX = os.path.join(BASE, "CustomRoles_QA-PreRelease-Checklist_2026-07-16.xlsx")
TR_URL = "https://shopview.testrail.io/index.php?/cases/view/{}"
EV15 = "build/custom-roles-run/live-ui-2026-07-15"
EV16 = "build/custom-roles-run/live-ui-2026-07-16"

# --------------------------------------------------------------------------
# Source-workbook tab layouts (0-based column indexes) for the accuracy gate.
# --------------------------------------------------------------------------
TAB_CFG = {
    "Full Dual Matrix": {"role": 0, "cap": 2, "verdict": 5},
    "Pass-11 LIVE (2026-07-16)": {"role": 0, "cap": 1, "verdict": 4},
    "Pass-12 LIVE (2026-07-16)": {"role": 0, "cap": 1, "verdict": 4},
    "Approve-Decline LIVE": {"role": 0, "cap": None, "verdict": 3,
                             "fixed_cap": "Approve/Decline line"},
    "Send to Terminal LIVE": {"role": 0, "cap": None, "verdict": 3,
                              "fixed_cap": "Send to Terminal"},
    "Parts-Module Dual LIVE": {"role": 0, "cap": 2, "verdict": 5},
    "New-WO Create Dual LIVE": {"role": 0, "cap": 2, "verdict": 5},
}

# --------------------------------------------------------------------------
# Verbatim spec quotes (each MUST appear character-exact in current-spec-2026-07-15.md;
# gate (b) enforces this). Curly quotes / arrows are copied from the spec file.
# --------------------------------------------------------------------------
Q = {
    "stt": "To send to terminal use must have this and “Customer Portal: ON” enabled",
    "aging": "AR/AP aging reports are part of Reports — a user with Reports ON sees all reports, including AR/AP aging, regardless of Manage AP/AR",
    "return_nogate": "Returning a part from a WOL does not require a permission",
    "return_everyone": "Everyone has access to Return a part from a WO",
    "notes": "Users can also see the Notes tab, create notes, and edit any note",
    "office_update": "Updated Office Role definition",
    "sm_migration": "Loses Invoicing Delete (cannot reverse)",
    "reverse_gate": "For WO requires Work Order → Delete",
    "take_payment": "Create invoices, process payments directly from work orders and part sales, manage invoice fields and collect deposits",
    "office_pay": "Office users are expected to be able to make payments but not create invoices",
    "authorize": "authorize lines",
    "tech_no_approve": "Cannot approve work order lines (approve action hidden)",
    "portal_fullview": "Has access to “Send to Portal” button",
    "portal_q6": "Send to Portal button: can be anyone who can approve a WOL",
    "portal_tech": "The “Send to Portal” button is not visible and the user cannot take this action",
    "tech_lose_portal": "Technicians lose “send to portal”",
    "history": "Users can view the work order and work order line level audit logs",
    "change_cust": "edit customer details and change customer, change asset",
    "timesheets": "View timesheets from work orders. If OFF, the Timesheets top level nav item is hidden",
    "order_parts": "Enabling Order Parts requires See Financial Data setting",
    "review": "Without this setting, users will not see the Review option on work orders",
    "vendor_receive": "create and manage purchase orders, manage deliveries",
    "new_line": "Add new lines, edit line details, move parts between lines, authorize lines",
    "create_wo": "Create new work orders",
    "wo_delete": "Delete work orders, Reverse Invoices as long as validation criteria is met (e.g. no payments made)",
    "newwo_cust": "Create/Edit customer also affect the ability to create a customer in the New WO flow",
    "sfd": "Controls visibility of all financial data across the application",
}

SPEC_REF = "build/custom-roles-spec-update/current-spec-2026-07-15.md"
TRUTH_REF = "build/custom-roles-run/spec-conformance/spec-truth-table.md"

# helper to render the "Spec says" cell
def spec_says(quote_key, citation):
    return "“{}” ({})".format(Q[quote_key], citation)


def steps(*lines):
    return "\n".join("{}. {}".format(i + 1, s) for i, s in enumerate(lines))


LOGIN = ("On the staging build (app.staging.shopview.com), sign in as a user whose "
         "role is {role} (create a fresh ZZAUTOTEST staff member with that role and "
         "log in as them; do not reuse a drifted account)")

STT_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open Work Orders and open a work order whose status is Invoiced and that still has a balance due (seed a ZZAUTOTEST work order and invoice it if none exists)",
    "Open the \"Finance\" tab on the work order",
    "Click the \"New Payment\" button",
    "In the New Customer Payment dialog, look for the blue \"Send to Terminal\" button",
)

AGING_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open \"Reports\" from the main navigation",
    "Look for the A/R Aging and A/P Aging report tiles in the reports list",
)

FINANCE_TAB_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open Work Orders and open a work order whose status is Invoiced (seed and invoice a ZZAUTOTEST work order if needed)",
    "Look for the \"Finance\" tab on the work order detail page",
    "If the Finance tab is present, open it and note which buttons appear: \"New Payment\", \"Reverse\", \"Issue Credit\"",
)

WO_DELETE_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open Work Orders and open a ZZAUTOTEST work order in a deletable state (Estimate status, no payments)",
    "Click the ⋮ (three-dot) menu on the work order detail page",
    "Look for the \"Delete Work Order\" option",
)

PORTAL_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a work order that has a line eligible to send to the portal (an approved/estimate line)",
    "Look for the \"Send to Portal\" button on the work order",
)

NOTES_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open Work Orders from the main navigation (if the nav item is hidden, paste a work order detail URL directly into the browser)",
    "If a work order opens, open its \"Notes\" tab",
    "Try to add a note and to edit an existing note",
)

RETURN_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a work order that has a picked inventory part on a line (seed a ZZAUTOTEST work order and pick a part if needed)",
    "On the picked part, open the ⋮ (three-dot) part menu",
    "Look for the \"Return\" option",
)

HISTORY_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open any work order detail page",
    "Look for the \"History\" tab (the work-order-level audit log)",
)

CHANGE_CUST_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open an existing ZZAUTOTEST work order",
    "On the work order header, look for the edit (pencil) control next to the Customer",
    "If present, change the customer to another ZZAUTOTEST customer and save (then change it back)",
)

CHANGE_ASSET_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open an existing ZZAUTOTEST work order",
    "On the work order header, look for the edit (pencil) control next to the Asset",
    "If present, change the asset and save (then change it back)",
)

ORDER_PARTS_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open any work order detail page and look for the \"Parts\" tab",
    "Also open the Parts area (/parts/orders) and look for the \"New PO\" button",
)

TIMESHEETS_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Look for \"Timesheets\" in the main navigation and open it",
    "Separately confirm clock in / clock out and \"My Timesheets\" still work (they are always available)",
)

NEW_LINE_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a work order's lines page",
    "Look for the \"New Line\" (add line) button",
)

NEW_PAYMENT_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a ZZAUTOTEST work order that is Invoiced with a balance due",
    "Open the \"Finance\" tab",
    "Look for the \"New Payment\" button; click it and take a small test payment on the ZZAUTOTEST invoice",
)

APPROVE_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a work order that has a line whose Status = \"Needs Approval\" (seed one via a part request that requires authorization if none exists; delete it after)",
    "Look at the line's Action column for the green \"Approve\" and red \"Decline\" buttons",
)

NEWWO_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open the Work Orders list and look for the \"New\" button",
    "If present, click \"New\" and in the New Work Order dialog look for the \"Add\" button next to Customer and the \"Add\" button next to Asset (the Asset Add enables after a customer is chosen)",
)

REVIEWED_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a completed ZZAUTOTEST work order",
    "Look for the Review action (\"Reviewed\") on the work order",
)

SFD_STEPS_T = steps(
    LOGIN.format(role="{role}"),
    "Open a work order with labor and parts lines",
    "Look for money data: labor rate, part prices, total/labor/parts columns, margins",
)


# --------------------------------------------------------------------------
# Checklist items.
# Each: id, priority, category, roles, capability, observed, spec (spec_says cell),
#       why, steps, expected, cases (list[int] or None => coverage gap), evidence,
#       owner, sources = list of (tab, role_text, capability_text_or_None, verdict_prefix)
# --------------------------------------------------------------------------
ITEMS = []


def add(**kw):
    ITEMS.append(kw)


# ============================= P1 =========================================
add(
    id="QA-01", priority="P1", category="Deviation",
    roles="Foreman",
    capability="\"Send to Terminal\" button (card-terminal payment) shows for Foreman",
    observed="PROD: hidden (org has no terminal device). STAGING: SHOWN (org has a terminal; any role that can open New Payment sees it).",
    spec=spec_says("stt", "spec §1i Invoicing Edit; 7/6 change log; Foreman has Customer Portal = OFF in the Toggles table"),
    why="DEVIATION (gating model): the spec role-gate = Invoicing Create & Edit + Customer Portal ON. Foreman has Customer Portal OFF, so per spec the button should be withheld; the build instead gates only on the org having a terminal device and shows it.",
    steps=STT_STEPS_T.format(role="Foreman"),
    expected="Per spec the \"Send to Terminal\" button should NOT appear for Foreman (Customer Portal is OFF for that role), even though the org has a card terminal.",
    cases=[29434, 26427],
    evidence=EV15 + "/staging (SendToTerminal_dialog.png set — pixel-observed for the invoicing roles; NONE exists for this role: its staging cell is method-derived, see GAP-08) + " + EV16 + "/production/_terminal (Pass-12 prod org-gate captures)",
    owner="PO decision (accept org-device gating or enforce Customer Portal ON) + QA retest on release build",
    sources=[("Full Dual Matrix", "Foreman", "Send to Terminal", "STAGING-MORE")],
)
add(
    id="QA-02", priority="P1", category="Deviation",
    roles="Office User",
    capability="\"Send to Terminal\" button (card-terminal payment) shows for Office User",
    observed="PROD: hidden (org has no terminal device). STAGING: SHOWN (org has a terminal; any role that can open New Payment sees it).",
    spec=spec_says("stt", "spec §1i Invoicing Edit; 7/6 change log; Office User has Customer Portal = OFF in the Toggles table"),
    why="DEVIATION (gating model): spec role-gate = Invoicing Create & Edit + Customer Portal ON. Office User has Customer Portal OFF, so per spec the button should be withheld; the build shows it whenever the org has a terminal.",
    steps=STT_STEPS_T.format(role="Office User"),
    expected="Per spec the \"Send to Terminal\" button should NOT appear for Office User (Customer Portal is OFF for that role).",
    cases=[29434, 26427],
    evidence=EV15 + "/staging (SendToTerminal_dialog.png set — pixel-observed for the invoicing roles; NONE exists for this role: its staging cell is method-derived, see GAP-08) + " + EV16 + "/production/_terminal (Pass-12 prod org-gate captures)",
    owner="PO decision + QA retest on release build",
    sources=[("Full Dual Matrix", "Office User", "Send to Terminal", "STAGING-MORE")],
)
add(
    id="QA-03", priority="P1", category="Deviation",
    roles="Parts Technician",
    capability="\"Send to Terminal\" button (card-terminal payment) shows for Parts Technician",
    observed="PROD: hidden (org has no terminal device). STAGING: SHOWN (org has a terminal; any role that can open New Payment sees it).",
    spec=spec_says("stt", "spec §1i Invoicing Edit; 7/6 change log; Parts Technician has Customer Portal = OFF in the Toggles table"),
    why="DEVIATION (gating model): spec role-gate = Invoicing Create & Edit + Customer Portal ON. Parts Technician has Customer Portal OFF, so per spec the button should be withheld; the build shows it whenever the org has a terminal.",
    steps=STT_STEPS_T.format(role="Parts Technician"),
    expected="Per spec the \"Send to Terminal\" button should NOT appear for Parts Technician (Customer Portal is OFF for that role).",
    cases=[29434, 26427],
    evidence=EV15 + "/staging (SendToTerminal_dialog.png set — pixel-observed for the invoicing roles; NONE exists for this role: its staging cell is method-derived, see GAP-08) + " + EV16 + "/production/_terminal (Pass-12 prod org-gate captures)",
    owner="PO decision + QA retest on release build",
    sources=[("Full Dual Matrix", "Parts Technician", "Send to Terminal", "STAGING-MORE")],
)
add(
    id="QA-04", priority="P1", category="Deviation",
    roles="Senior Service Advisor",
    capability="A/R + A/P Aging reports (Reports area) missing for Senior Service Advisor",
    observed="PROD: hidden. STAGING: hidden. Both environments hide the aging report tiles, but the spec grants them.",
    spec=spec_says("aging", "spec §2a Reports note; Jul-3 change log (aging decoupled from Manage AP/AR); Senior Service Advisor has Reports = ON in the Toggles table"),
    why="DEVIATION: the spec grants the aging reports to Senior Service Advisor (Reports ON; Manage AP/AR also ON) but staging does not implement the grant — an unimplemented spec grant, not a migration loss.",
    steps=AGING_STEPS_T.format(role="Senior Service Advisor"),
    expected="Per spec the A/R Aging and A/P Aging report tiles SHOULD be listed for Senior Service Advisor (Reports is ON for that role; aging follows Reports, all-or-nothing).",
    cases=[26478, 26482],
    evidence=EV16 + "/production (Pass-11 AP/AR captures: apar-DUAL-verdicts.json + per-role clean_apar.png) + " + EV15 + "/staging (per-role apar_reports.png + apar-per-role.json)",
    owner="Dev fix + QA retest on release build",
    sources=[("Pass-11 LIVE (2026-07-16)", "Senior Service Advisor", "See AP/AR", "MATCH")],
)
add(
    id="QA-05", priority="P1", category="Deviation",
    roles="Office User",
    capability="Work order \"Notes\" tab reachable for Office User (over-grant vs the 7/14 spec)",
    observed="PROD: SHOWN. STAGING: SHOWN. Both environments let Office User reach a work order and its Notes tab.",
    spec=spec_says("office_update", "7/14 change log; the CRUD matrix now gives Office User Work Orders = — (no access); Notes ride on Work Orders View: " + "“" + Q["notes"] + "” (spec §1a View)"),
    why="DEVIATION: the 14-Jul spec update removed ALL Work Orders access for Office User, so the work order (and its Notes tab) should be unreachable; both envs still show it. Needs a PO confirm + dev change before customers see it.",
    steps=NOTES_STEPS_T.format(role="Office User"),
    expected="Per the 7/14 spec, Office User should NOT be able to open work orders at all, so the work order Notes tab should be unreachable for Office User.",
    cases=None,
    evidence=EV16 + "/ + " + EV15 + "/ (Office_User captures)",
    owner="PO decision (confirm the 7/14 Office lockdown) + Dev fix + QA retest",
    sources=[("Full Dual Matrix", "Office User", "WO Notes (Notes tab)", "MATCH")],
)
add(
    id="QA-06", priority="P1", category="Deviation",
    roles="Office User",
    capability="Part \"Return\" (picked-part menu on a work order) reachable for Office User",
    observed="PROD: hidden (part menu shows Move only). STAGING: SHOWN (Return appears in the picked-part menu).",
    spec=spec_says("return_nogate", "spec §1a; 6/29 change log “" + Q["return_everyone"] + "”; BUT the 7/14 matrix gives Office User Work Orders = —, so Office cannot reach a work order part"),
    why="DEVIATION (reachability): Return itself is ungated, but per the current spec Office User has no Work Orders access so the Return point should be unreachable; staging shows it (STAGING-MORE) — same root cause as QA-05.",
    steps=RETURN_STEPS_T.format(role="Office User"),
    expected="Per the 7/14 spec, Office User should not reach a work order part at all, so \"Return\" should be unreachable for Office User.",
    cases=None,
    evidence=EV16 + "/ (Pass-11 Part Return captures, e.g. production/_prod-partreturn json under live-ui-2026-07-15)",
    owner="PO decision + Dev fix + QA retest",
    sources=[("Pass-11 LIVE (2026-07-16)", "Office User", "Part Return", "STAGING-MORE")],
)
add(
    id="QA-07", priority="P1", category="Spec-contradiction",
    roles="Service Manager",
    capability="\"Reverse\" (reverse invoice) on the work order Finance tab — spec contradicts itself for Service Manager",
    observed="PROD: Finance tab hidden for the legacy role. STAGING: Finance tab SHOWN with \"New Payment\" + \"Issue Credit\" but NO \"Reverse\".",
    spec="Matrix gives Service Manager Work Orders V/E/D and the 28-Jun rule says Reverse follows WO Delete (“" + Q["reverse_gate"] + "”) → would GRANT Reverse; but the migration Behavior-Changes table says Service Manager “" + Q["sm_migration"] + "” → would WITHHOLD it. (28-Jun change log + migration table; flagged in " + TRUTH_REF + ")",
    why="The spec self-contradicts: matrix + 28-Jun gate say Service Manager CAN reverse; the migration table says cannot. Staging matches the migration table. Money-facing — must be resolved before release.",
    steps=FINANCE_TAB_STEPS_T.format(role="Service Manager"),
    expected="PO must rule which spec statement wins. If the matrix + 28-Jun rule win: \"Reverse\" SHOULD appear for Service Manager (dev change needed). If the migration table wins: current staging behavior (no Reverse) is correct.",
    cases=[26422, 27740],
    evidence=EV16 + "/ (Pass-11 Finance captures; staging Service_Manager)",
    owner="PO decision, then QA retest",
    sources=[("Pass-11 LIVE (2026-07-16)", "Service Manager", "Finance", "STAGING-MORE")],
)
add(
    id="QA-08", priority="P1", category="Spec-contradiction",
    roles="Technician",
    capability="Red \"Decline\" button on a \"Needs Approval\" line — spec ambiguous for Technician",
    observed="PROD: hidden (line renders, no Approve/Decline). STAGING: hidden (line renders, no Approve/Decline). Both envs hide it.",
    spec="Spec §1b Work Order Lines Edit includes “" + Q["authorize"] + "” (Technician HAS WO Lines Create & Edit) — would grant Decline; but §4 Tech View blocks only Approve (“" + Q["tech_no_approve"] + "”) and never mentions Decline — Tech-View treatment of Decline is unresolvable from the spec.",
    why="The spec grants Technician WOL Create & Edit ('authorize lines' covers declining) yet Tech View only blocks Approve. Whether Technician should see Decline is a genuine open product question; both envs currently hide it.",
    steps=APPROVE_STEPS_T.format(role="Technician"),
    expected="PO must state whether Tech View also hides \"Decline\". Current behavior in both envs: neither \"Approve\" nor \"Decline\" is shown to Technician.",
    cases=None,
    evidence=EV16 + "/staging/approve-decline-TECH-PT.json + production/approve-decline-ALL.json",
    owner="PO decision, then QA retest",
    sources=[("Full Dual Matrix", "Technician", "Decline line", "MATCH")],
)
add(
    id="QA-09", priority="P1", category="Spec-contradiction",
    roles="Parts Technician, Office User, Sales Representative",
    capability="\"Send to Portal\" button — spec wording contradiction (Full View vs Open Q6) for these 3 roles",
    observed="Parts Technician: PROD SHOWN / STAGING hidden. Office User: PROD SHOWN / STAGING hidden. Sales Representative: both hidden (MATCH).",
    spec="§4 Full View “" + Q["portal_fullview"] + "” + 6/10 change log (Full View required) would GRANT (all three are Full View); Open Q6 answer “" + Q["portal_q6"] + "” would WITHHOLD (none has WO Lines Create & Edit). Flagged, not resolved.",
    why="Customer-facing button with two contradictory spec gates. Staging currently follows the Open-Q6 (approve-a-WOL) reading — Parts Technician and Office User LOSE Send to Portal vs prod. PO must confirm the intended wording before release.",
    steps=PORTAL_STEPS_T.format(role="Parts Technician (repeat as Office User, then as Sales Representative)"),
    expected="PO must pick a reading. Current staging behavior = Open-Q6 reading: \"Send to Portal\" hidden for Parts Technician, Office User and Sales Representative.",
    cases=[26466, 26465],
    evidence=EV16 + "/ + " + EV15 + "/ (per-role portal captures)",
    owner="PO decision, then QA retest",
    sources=[
        ("Full Dual Matrix", "Parts Technician", "Send to Portal", "STAGING-LESS"),
        ("Full Dual Matrix", "Office User", "Send to Portal", "STAGING-LESS"),
        ("Full Dual Matrix", "Sales Representative (+Reporting)", "Send to Portal", "MATCH"),
    ],
)
add(
    id="QA-10", priority="P1", category="Intended-loss-to-confirm",
    roles="Service Advisor",
    capability="Service Advisor loses \"Reverse\" (reverse invoice) on the Finance tab — money-facing, confirm before release",
    observed="PROD (SA Limited View): Finance tab shows \"New Payment\" + \"Reverse\" + \"Issue Credit\". STAGING (Service Advisor): shows \"New Payment\" + \"Issue Credit\"; \"Reverse\" hidden.",
    spec=spec_says("reverse_gate", "28-Jun change log: Reverse Invoice moved to Work Orders → Delete; Service Advisor has Work Orders = V/E (no Delete) in the CRUD matrix"),
    why="Intended per spec (Reverse now follows WO Delete, which Service Advisor lacks) but it is a customer/money-facing REGRESSION vs prod — confirm the loss is intended and holds on the release build.",
    steps=FINANCE_TAB_STEPS_T.format(role="Service Advisor"),
    expected="\"New Payment\" and \"Issue Credit\" appear; \"Reverse\" does NOT appear for Service Advisor.",
    cases=[26422, 27740],
    evidence=EV16 + "/ (Pass-11 Finance captures; prod SA_Limited_View vs staging Service_Advisor)",
    owner="QA retest on release build (+ PO ack of the customer-facing loss)",
    sources=[("Pass-11 LIVE (2026-07-16)", "Service Advisor", "Finance", "MATCH")],
)
add(
    id="QA-11", priority="P1", category="Intended-grant-to-confirm",
    roles="Service Manager, Senior Service Advisor, Foreman, Parts Manager, Parts Technician, Office User",
    capability="\"New Payment\" (Take Payment) NEWLY GRANTED to 6 roles — money-facing, confirm each",
    observed="PROD: hidden for all 6 legacy counterparts. STAGING: SHOWN for all 6 (each can open the New Customer Payment dialog).",
    spec=spec_says("take_payment", "spec §1i Invoicing Create & Edit; Open Q4; all 6 roles hold Invoicing C&E in the CRUD matrix; Office explicitly: “" + Q["office_pay"] + "”"),
    why="Six roles that could NOT take payments in prod now can. Intended per spec, but money movement by new hands is release-critical — verify each role can take (and only take) a payment, and that Office User still cannot create invoices.",
    steps=NEW_PAYMENT_STEPS_T.format(role="each of: Service Manager, Senior Service Advisor, Foreman, Parts Manager, Parts Technician, Office User (one at a time)"),
    expected="\"New Payment\" appears and a test payment completes for each of the 6 roles. For Office User additionally confirm the Create Invoice button stays disabled (hard-coded Office rule).",
    cases=[26421, 26428, 26484],
    evidence=EV16 + "/ (Pass-11 Finance captures) + Full Dual Matrix Take Payment rows",
    owner="QA retest on release build",
    sources=[
        ("Full Dual Matrix", "Service Manager", "Take Payment (New Payment)", "STAGING-MORE"),
        ("Full Dual Matrix", "Senior Service Advisor (+SA Tech,+SA NoRep)", "Take Payment (New Payment)", "STAGING-MORE"),
        ("Full Dual Matrix", "Foreman", "Take Payment (New Payment)", "STAGING-MORE"),
        ("Full Dual Matrix", "Parts Manager", "Take Payment (New Payment)", "STAGING-MORE"),
        ("Full Dual Matrix", "Parts Technician", "Take Payment (New Payment)", "STAGING-MORE"),
        ("Full Dual Matrix", "Office User", "Take Payment (New Payment)", "STAGING-MORE"),
    ],
)
add(
    id="QA-12", priority="P1", category="Intended-loss-to-confirm",
    roles="Service Advisor, Foreman, Technician, Office User (lose) + Service Manager (gains)",
    capability="\"Delete Work Order\" now Admin / Service Manager / Senior Service Advisor ONLY — confirm the 4 losing roles + the SM gain",
    observed="STAGING-LESS (prod SHOWN → staging hidden): Service Advisor (prod SA Limited View), Foreman, Technician, Office User. STAGING-MORE: Service Manager (prod hidden → staging SHOWN). Admin + Senior Service Advisor: MATCH SHOWN. NOTE: the ground-truth matrix shows 4 losing roles (not 5).",
    spec=spec_says("wo_delete", "spec §1a Work Orders Delete; CRUD matrix Work Orders row: Delete only for Admin / Service Manager / Senior Service Advisor"),
    why="Destructive, customer-facing capability re-scoped. Intended per spec, but each losing role must be regression-confirmed (and the SM gain smoke-tested) on the release build.",
    steps=WO_DELETE_STEPS_T.format(role="each of: Service Advisor, Foreman, Technician, Office User, then Service Manager (one at a time)"),
    expected="\"Delete Work Order\" is ABSENT for Service Advisor, Foreman, Technician and Office User; PRESENT for Service Manager (and Admin / Senior Service Advisor).",
    cases=[26376, 26384],
    evidence=EV15 + "/ (per-role WO-detail captures; estimate-state WO)",
    owner="QA retest on release build",
    sources=[
        ("Full Dual Matrix", "Service Advisor", "WO Delete", "STAGING-LESS"),
        ("Full Dual Matrix", "Foreman", "WO Delete", "STAGING-LESS"),
        ("Full Dual Matrix", "Technician", "WO Delete", "STAGING-LESS"),
        ("Full Dual Matrix", "Office User", "WO Delete", "STAGING-LESS"),
        ("Full Dual Matrix", "Service Manager", "WO Delete", "STAGING-MORE"),
    ],
)

# ============================= P2 =========================================
P2_NOTE = "Intended per spec — regression-confirm on the release build."


def add_p2(idn, roles, capability, observed, spec, steps_, expected, cases, evidence, sources, category):
    add(id=idn, priority="P2", category=category, roles=roles, capability=capability,
        observed=observed, spec=spec, why=P2_NOTE, steps=steps_, expected=expected,
        cases=cases, evidence=evidence, owner="QA retest on release build", sources=sources)


# --- remaining STAGING-LESS losses ---
for i, role in enumerate(["Technician", "Parts Technician", "Office User"], start=13):
    add_p2("QA-%02d" % i, role,
           "Work order \"History\" tab (WO-level audit log) removed for " + role,
           "PROD: SHOWN. STAGING: hidden.",
           spec_says("history", "spec §1a Work Orders Edit + 7/7 change log; " + role + " lacks Work Orders Create & Edit in the CRUD matrix"),
           HISTORY_STEPS_T.format(role=role),
           "The \"History\" tab is ABSENT for " + role + ".",
           [26488, 26489, 27870],
           EV15 + "/ (per-role WO-detail captures)",
           [("Full Dual Matrix", role, "WO-level History (History tab)", "STAGING-LESS")],
           "Intended-loss-to-confirm")

for idn, role, cap, stp in [
    ("QA-16", "Technician", "Change Customer on WO", CHANGE_CUST_STEPS_T),
    ("QA-17", "Technician", "Change Asset on WO", CHANGE_ASSET_STEPS_T),
    ("QA-18", "Office User", "Change Customer on WO", CHANGE_CUST_STEPS_T),
    ("QA-19", "Office User", "Change Asset on WO", CHANGE_ASSET_STEPS_T),
]:
    what = "Change the Customer" if "Customer" in cap else "Change the Asset"
    add_p2(idn, role,
           what + " on an existing work order removed for " + role,
           "PROD: SHOWN. STAGING: hidden.",
           spec_says("change_cust", "spec §1a Work Orders Edit; " + role + " lacks Work Orders Create & Edit in the CRUD matrix"),
           stp.format(role=role),
           "The edit (pencil) control next to the " + ("Customer" if "Customer" in cap else "Asset") + " is ABSENT for " + role + ".",
           None,
           EV15 + "/ (per-role WO-detail captures)",
           [("Full Dual Matrix", role, cap, "STAGING-LESS")],
           "Intended-loss-to-confirm")

for idn, role in [("QA-20", "Technician"), ("QA-21", "Office User")]:
    add_p2(idn, role,
           "Work order \"Parts\" tab / Order Parts area removed for " + role,
           "PROD: SHOWN. STAGING: hidden.",
           spec_says("order_parts", "spec §1a Order Parts sub-setting (woOrderParts); Order Parts = — for " + role + " in the WO Sub-Settings table"),
           ORDER_PARTS_STEPS_T.format(role=role),
           "The work order \"Parts\" tab is ABSENT for " + role + " (no ordering surface).",
           [26381, 27868, 27869],
           EV15 + "/ (per-role parts_orders.png)",
           [("Full Dual Matrix", role, "Order Parts area (Parts tab)", "STAGING-LESS")],
           "Intended-loss-to-confirm")

for idn, role, extra in [
    ("QA-22", "Technician", "Explicit per spec: 28-Jun change log “" + Q["tech_lose_portal"] + "”."),
    ("QA-23", "Parts Technician", "Spec wording is contradictory for this role — see QA-09; the observed loss matches the Open-Q6 reading."),
    ("QA-24", "Office User", "Spec wording is contradictory for this role — see QA-09; the observed loss matches the Open-Q6 reading."),
]:
    add_p2(idn, role,
           "\"Send to Portal\" button removed for " + role,
           "PROD: SHOWN. STAGING: hidden. " + extra,
           spec_says("portal_q6", "Open Q6 answer; §4 View Mode; 6/10 change log; " + role + " has no WO Lines Create & Edit"),
           PORTAL_STEPS_T.format(role=role),
           "The \"Send to Portal\" button is ABSENT for " + role + " (per the Open-Q6 reading).",
           [26465, 26466],
           EV15 + "/ + " + EV16 + "/ (per-role portal captures)",
           [("Full Dual Matrix", role if role != "Technician" else "Technician", "Send to Portal", "STAGING-LESS")],
           "Intended-loss-to-confirm")

for idn, role in [("QA-25", "Technician"), ("QA-26", "Parts Manager")]:
    add_p2(idn, role,
           "\"Timesheets\" navigation item removed for " + role,
           "PROD: SHOWN. STAGING: hidden.",
           spec_says("timesheets", "spec §1j Timesheets View; Timesheets = — for " + role + " in the CRUD matrix; clock in/out + “My Timesheets” stay available"),
           TIMESHEETS_STEPS_T.format(role=role),
           "\"Timesheets\" is ABSENT from the main navigation for " + role + "; clock in/out and \"My Timesheets\" still work.",
           [26430, 26433, 27394],
           EV15 + "/ (per-role nav captures)",
           [("Full Dual Matrix", role, "Timesheets (tab)", "STAGING-LESS")],
           "Intended-loss-to-confirm")

add_p2("QA-27", "Technician",
       "Work order \"Finance\" tab removed for Technician",
       "PROD: SHOWN. STAGING: hidden.",
       spec_says("sfd", "spec §5a; plus §1i Invoicing View — Technician has Invoicing = — AND See Financial Data = — in the matrix/toggles"),
       steps(
           LOGIN.format(role="Technician"),
           "Open an Invoiced work order",
           "Look for the \"Finance\" tab",
       ),
       "The \"Finance\" tab is ABSENT for Technician.",
       [26420, 26426],
       EV15 + "/ (per-role WO-detail captures)",
       [("Full Dual Matrix", "Technician", "Invoicing/Finance view (Finance tab)", "STAGING-LESS")],
       "Intended-loss-to-confirm")

add_p2("QA-28", "Parts Technician",
       "\"New Line\" (add a work order line) removed for Parts Technician",
       "PROD: SHOWN. STAGING: hidden.",
       spec_says("new_line", "spec §1b WO Lines Edit; Parts Technician has WO Lines = V (no Create & Edit) in the CRUD matrix"),
       NEW_LINE_STEPS_T.format(role="Parts Technician"),
       "The \"New Line\" button is ABSENT for Parts Technician.",
       [26390, 26461, 27272],
       EV15 + "/ (per-role line captures)",
       [("Full Dual Matrix", "Parts Technician", "Create/Edit WO Lines (New Line)", "STAGING-LESS")],
       "Intended-loss-to-confirm")

# --- remaining STAGING-MORE grants ---
add_p2("QA-29", "Parts Manager",
       "Review action (\"Reviewed\") newly granted to Parts Manager",
       "PROD: hidden. STAGING: SHOWN.",
       spec_says("review", "spec §1a WO sub-setting Review Work Orders (woReviewWorkOrders); Review WOs = ON for Parts Manager in the WO Sub-Settings table"),
       REVIEWED_STEPS_T.format(role="Parts Manager"),
       "The Review action IS available to Parts Manager.",
       [26379],
       EV15 + "/ (per-role WO-detail captures)",
       [("Full Dual Matrix", "Parts Manager", "Review Work Orders (Reviewed)", "STAGING-MORE")],
       "Intended-grant-to-confirm")

for idn, cap, stp in [("QA-30", "Change Customer on WO", CHANGE_CUST_STEPS_T),
                      ("QA-31", "Change Asset on WO", CHANGE_ASSET_STEPS_T)]:
    what = "Customer" if "Customer" in cap else "Asset"
    add_p2(idn, "Parts Manager",
           "Change the " + what + " on an existing work order newly granted to Parts Manager",
           "PROD: hidden. STAGING: SHOWN.",
           spec_says("change_cust", "spec §1a Work Orders Edit; Parts Manager has Work Orders V/E in the CRUD matrix"),
           stp.format(role="Parts Manager"),
           "The edit (pencil) control next to the " + what + " IS available to Parts Manager.",
           None,
           EV15 + "/ (per-role WO-detail captures)",
           [("Full Dual Matrix", "Parts Manager", cap, "STAGING-MORE")],
           "Intended-grant-to-confirm")

for idn, cap in [("QA-32", "Approve line"), ("QA-33", "Decline line")]:
    add_p2(idn, "Parts Manager",
           "\"" + cap.split()[0] + "\" button on a \"Needs Approval\" line newly granted to Parts Manager",
           "PROD: hidden (line renders, no Approve/Decline). STAGING: SHOWN.",
           spec_says("authorize", "spec §1b WO Lines Edit (‘authorize lines’); Parts Manager has WO Lines V/E + Full View"),
           APPROVE_STEPS_T.format(role="Parts Manager"),
           "The green \"Approve\" and red \"Decline\" buttons ARE shown to Parts Manager on a Needs-Approval line.",
           [26390, 26458, 26466] if cap == "Approve line" else None,
           EV16 + "/staging/approve-decline-REALHOLDER-ALL.json + production/approve-decline-ALL.json",
           [("Full Dual Matrix", "Parts Manager", cap, "STAGING-MORE")],
           "Intended-grant-to-confirm")

add_p2("QA-34", "Parts Manager",
       "\"New\" work order button (create WO) newly granted to Parts Manager",
       "PROD: hidden (no New button). STAGING: SHOWN (dialog opens).",
       spec_says("create_wo", "spec §1a Work Orders Edit; Parts Manager has Work Orders V/E in the CRUD matrix"),
       NEWWO_STEPS_T.format(role="Parts Manager"),
       "The \"New\" button IS shown and the New Work Order dialog opens for Parts Manager.",
       [26375, 26383],
       EV15 + "/production/Parts_Manager/new_wo_nobutton.png + staging/Parts_Manager/new_wo_dialog.png",
       [("New-WO Create Dual LIVE", "Parts Manager", "Create Work Order (\"New\" button on Work Orders list)", "STAGING-MORE")],
       "Intended-grant-to-confirm")

add_p2("QA-35", "Parts Manager",
       "\"Add\" Customer inside the New Work Order dialog newly granted to Parts Manager",
       "PROD: hidden (dialog unreachable). STAGING: SHOWN.",
       spec_says("newwo_cust", "6/1 change log; §1d Customer Management Edit; Parts Manager holds Customers V/E/D + Work Orders V/E"),
       NEWWO_STEPS_T.format(role="Parts Manager"),
       "The \"Add\" button next to Customer IS shown in the New Work Order dialog for Parts Manager.",
       [26387],
       EV15 + "/staging/Parts_Manager/new_wo_dialog.png",
       [("New-WO Create Dual LIVE", "Parts Manager", "Create Customer from New-WO (\"Add\" next to Customer)", "STAGING-MORE")],
       "Intended-grant-to-confirm")

add_p2("QA-36", "Parts Manager",
       "\"Add\" Asset inside the New Work Order dialog newly granted to Parts Manager",
       "PROD: hidden (dialog unreachable). STAGING: SHOWN (enables after a customer is chosen).",
       spec_says("newwo_cust", "6/1 change log; §1d Customer Management Edit (manage vehicles); Parts Manager holds Customers V/E/D + Work Orders V/E"),
       NEWWO_STEPS_T.format(role="Parts Manager"),
       "The \"Add\" button next to Asset IS shown (enabled once a customer is chosen) for Parts Manager.",
       [26388],
       EV15 + "/staging/Parts_Manager/new_wo_dialog.png",
       [("New-WO Create Dual LIVE", "Parts Manager", "Create Asset control from New-WO (\"Add\" next to Asset)", "STAGING-MORE")],
       "Intended-grant-to-confirm")

add_p2("QA-37", "Sales Representative",
       "See Financial Data (rates / margins / money columns) newly granted to Sales Representative",
       "PROD: hidden. STAGING: SHOWN.",
       spec_says("sfd", "spec §5a; See Financial = ON for Sales Representative in the Cross-Cutting table"),
       SFD_STEPS_T.format(role="Sales Representative"),
       "Money data (rates, prices, totals, margins) IS visible to Sales Representative.",
       [26464, 26470],
       EV15 + "/ (per-role captures)",
       [("Full Dual Matrix", "Sales Representative (+Reporting)", "See Financial Data (Rate/Margin)", "STAGING-MORE")],
       "Intended-grant-to-confirm")

for idn, role_wb, role in [("QA-38", "Sales Representative (+Reporting)", "Sales Representative"),
                           ("QA-39", "Time Clock User", "Time Clock User")]:
    add_p2(idn, role,
           "Work order \"Notes\" tab newly reachable for " + role,
           "PROD: hidden. STAGING: SHOWN.",
           spec_says("notes", "spec §1a Work Orders View; " + role + " has Work Orders = V in the CRUD matrix"),
           NOTES_STEPS_T.format(role=role),
           "The \"Notes\" tab IS reachable (create + edit any note) for " + role + ".",
           None,
           EV15 + "/ (per-role WO-detail captures)",
           [("Full Dual Matrix", role_wb, "WO Notes (Notes tab)", "STAGING-MORE")],
           "Intended-grant-to-confirm")

add_p2("QA-40", "Time Clock User",
       "\"Timesheets\" navigation item newly granted to Time Clock User",
       "PROD: hidden. STAGING: SHOWN.",
       spec_says("timesheets", "spec §1j Timesheets View; Timesheets = V for Time Clock User in the CRUD matrix"),
       TIMESHEETS_STEPS_T.format(role="Time Clock User"),
       "\"Timesheets\" IS shown in the main navigation for Time Clock User.",
       [26430, 27394, 26432],
       EV15 + "/ (per-role nav captures)",
       [("Full Dual Matrix", "Time Clock User", "Timesheets (tab)", "STAGING-MORE")],
       "Intended-grant-to-confirm")

for idn, role in [("QA-41", "Service Manager"), ("QA-42", "Parts Manager"), ("QA-43", "Sales Representative")]:
    add_p2(idn, role,
           "A/R + A/P Aging reports newly visible to " + role,
           "PROD: hidden. STAGING: SHOWN (tiles render).",
           spec_says("aging", "spec §2a Reports note; Reports = ON for " + role + " in the Toggles table"),
           AGING_STEPS_T.format(role=role),
           "The A/R Aging and A/P Aging report tiles ARE listed for " + role + ".",
           [26478, 26482],
           EV16 + "/ (Pass-11 AP/AR captures)",
           [("Pass-11 LIVE (2026-07-16)", role, "See AP/AR", "STAGING-MORE")],
           "Intended-grant-to-confirm")

for idn, role in [("QA-44", "Sales Representative"), ("QA-45", "Time Clock User")]:
    add_p2(idn, role,
           "Part \"Return\" (picked-part menu) newly reachable for " + role,
           "PROD: hidden (WO detail not accessible). STAGING: SHOWN.",
           spec_says("return_nogate", "spec §1a + 6/29 change log; " + role + " has Work Orders = V so the Return point is reachable"),
           RETURN_STEPS_T.format(role=role),
           "The \"Return\" option IS shown in the picked-part menu for " + role + ".",
           None,
           EV16 + "/ (Pass-11 Part Return captures)",
           [("Pass-11 LIVE (2026-07-16)", role, "Part Return", "STAGING-MORE")],
           "Intended-grant-to-confirm")

# ============================= P3 =========================================
add(
    id="QA-46", priority="P3", category="Env-config",
    roles="Admin, Service Manager, Senior Service Advisor, Service Advisor, Parts Manager",
    capability="\"Send to Terminal\" appears on staging but not prod for the 5 roles that DO meet the spec role-gate — org-device config, not a role delta",
    observed="PROD: hidden for every role (the prod test org has NO card-terminal device and no UI path to add one; all terminal APIs 404). STAGING: SHOWN (the staging org has a terminal).",
    spec=spec_says("stt", "spec §1i Invoicing Edit; these 5 roles hold Invoicing C&E + Customer Portal ON, so the role-gate is met"),
    why="Fully-characterized ORG-DEVICE config gate (Pass-12), not a role/migration difference. Monitor only: confirm the release org's terminal-device configuration matches what customers should have.",
    steps=steps(
        "Confirm with the release/ops team whether the target org has a card terminal device registered with the payment processor",
        "If a terminal exists: as any of the 5 roles, open an Invoiced WO > \"Finance\" > \"New Payment\" and confirm \"Send to Terminal\" appears",
        "If no terminal exists: confirm the button is absent for every role (expected)",
    ),
    expected="Button presence follows the org's terminal device: shown for these 5 roles when a terminal is registered, absent for everyone when not.",
    cases=[29434, 26427],
    evidence=EV16 + "/ (Pass-12: prod Settings/Payment-Methods nav screenshots; terminal APIs 404 captures)",
    owner="QA monitor / ops config check",
    sources=[
        ("Send to Terminal LIVE", "Admin", None, "STAGING-MORE"),
        ("Send to Terminal LIVE", "Service Manager", None, "STAGING-MORE"),
        ("Send to Terminal LIVE", "Senior Service Advisor", None, "STAGING-MORE"),
        ("Send to Terminal LIVE", "Service Advisor", None, "STAGING-MORE"),
        ("Send to Terminal LIVE", "Parts Manager", None, "STAGING-MORE"),
    ],
)
add(
    id="QA-47", priority="P3", category="Spec-silent",
    roles="Service Manager, Parts Manager, Parts Technician, Foreman, Office User (observed holders)",
    capability="\"Issue Credit\" button (Finance tab) — the spec is SILENT on who may issue a credit",
    observed="STAGING: \"Issue Credit\" shows alongside \"New Payment\" for every role that reaches the Finance tab. PROD: Finance tab hidden for these legacy roles.",
    spec="The spec defines NO issuance gate for credits anywhere — §5b covers only the visibility of the Credits TAB on Customer/Vendor pages. (Spec-silent register: " + TRUTH_REF + " §5.)",
    why="Money-facing button with no spec gate. No action required for release, but the PO should define a gate; QA should monitor that it at least tracks Finance-tab access.",
    steps=steps(
        LOGIN.format(role="Service Manager (spot-check one or two of the listed roles)"),
        "Open an Invoiced ZZAUTOTEST work order > \"Finance\" tab",
        "Confirm \"Issue Credit\" appears next to \"New Payment\" and note which roles can see it",
    ),
    expected="No spec expectation exists (spec silent). Record observed behavior; ask the PO to define the intended gate.",
    cases=None,
    evidence=EV16 + "/ (Pass-11 Finance captures)",
    owner="PO decision (define a gate); QA monitor",
    sources=[("Pass-11 LIVE (2026-07-16)", "Service Manager", "Finance", "STAGING-MORE")],
)
add(
    id="QA-48", priority="P3", category="Intended-grant-to-confirm",
    roles="All roles except Technician (see QA-08) and Parts Manager (see QA-33)",
    capability="\"Decline\" button parity — Decline is gated together with Approve everywhere it was observed",
    observed="Both envs: Decline shows exactly where Approve shows, for every role (Admin/SM/SSA/SA/Foreman shown; PT/Office/Sales Rep/Time Clock hidden). Only deltas: Parts Manager (QA-33) and the Technician ambiguity (QA-08).",
    spec=spec_says("authorize", "spec §1b WO Lines Edit — ‘authorize lines’ covers declining a pending line"),
    why="Smoke-only if time permits: parity already observed live in both envs on 2026-07-16.",
    steps=steps(
        LOGIN.format(role="Admin (or any Approve-holding role)"),
        "Open a work order with a \"Needs Approval\" line",
        "Confirm the green \"Approve\" and red \"Decline\" buttons appear together",
    ),
    expected="\"Decline\" appears wherever \"Approve\" appears (same gate: WO Lines Create & Edit).",
    cases=None,
    evidence=EV16 + "/staging/approve-decline-REALHOLDER-ALL.json + production/approve-decline-ALL.json",
    owner="QA smoke (time permitting)",
    sources=[("Approve-Decline LIVE", "Admin", None, "MATCH")],
)
add(
    id="QA-49", priority="P3", category="Intended-grant-to-confirm",
    roles="All roles with Work Orders View",
    capability="Part \"Return\" parity (MATCH rows) — ungated return works the same in both envs for WO-View roles",
    observed="Both envs SHOWN for Admin, Service Manager, Senior Service Advisor, Service Advisor, Foreman, Technician, Parts Manager, Parts Technician (MATCH).",
    spec=spec_says("return_nogate", "spec §1a; 6/29 change log"),
    why="Smoke-only if time permits: parity already observed live (Pass-11).",
    steps=RETURN_STEPS_T.format(role="Technician (representative WO-View role)"),
    expected="\"Return\" appears in the picked-part menu for any role that can open the work order.",
    cases=None,
    evidence=EV16 + "/ (Pass-11 Part Return captures)",
    owner="QA smoke (time permitting)",
    sources=[("Pass-11 LIVE (2026-07-16)", "Technician", "Part Return", "MATCH")],
)
add(
    id="QA-50", priority="P3", category="Intended-grant-to-confirm",
    roles="Admin, Service Manager, Senior Service Advisor, Service Advisor, Foreman, Parts Manager, Parts Technician",
    capability="Parts module parity (MATCH rows): \"New PO\" + \"Receive\" on /parts/orders",
    observed="Both envs SHOWN for the 7 granted roles; both envs hidden for Office User, Sales Representative, Technician, Time Clock User (all MATCH).",
    spec=spec_says("order_parts", "spec §1a Order Parts sub-setting + §1g Vendor & Order Management Edit: “" + Q["vendor_receive"] + "”"),
    why="Smoke-only if time permits: full dual grid already observed live 2026-07-15.",
    steps=steps(
        LOGIN.format(role="Parts Manager (representative granted role)"),
        "Open the Parts area (/parts/orders)",
        "Confirm the \"New PO\" button and the per-PO \"Receive\" button appear",
    ),
    expected="\"New PO\" and \"Receive\" appear for the granted roles only.",
    cases=[27876, 26417, 26381],
    evidence=EV15 + "/ (per-role parts_orders.png, both envs)",
    owner="QA smoke (time permitting)",
    sources=[("Parts-Module Dual LIVE", "Admin", "Order Parts (create Purchase Order — \"New PO\" on /parts/orders)", "MATCH")],
)
add(
    id="QA-51", priority="P3", category="Intended-grant-to-confirm",
    roles="Admin, Service Manager, Senior Service Advisor, Service Advisor, Foreman",
    capability="New Work Order dialog parity (MATCH rows): \"New\" button + \"Add\" Customer / \"Add\" Asset",
    observed="Both envs SHOWN for the granted roles; both envs hidden for Office User, Parts Technician, Sales Representative, Technician, Time Clock User (all MATCH). Only delta = Parts Manager (QA-34/35/36).",
    spec=spec_says("create_wo", "spec §1a Work Orders Edit + 6/1 change log (Add Customer in the New-WO flow)"),
    why="Smoke-only if time permits: full dual grid already observed live 2026-07-15.",
    steps=NEWWO_STEPS_T.format(role="Service Advisor (representative granted role)"),
    expected="\"New\" opens the New Work Order dialog with \"Add\" next to Customer and Asset for the granted roles only.",
    cases=[26375, 26387, 26388],
    evidence=EV15 + "/ (per-role new_wo_dialog.png / new_wo_nobutton.png, both envs)",
    owner="QA smoke (time permitting)",
    sources=[("New-WO Create Dual LIVE", "Admin", "Create Work Order (\"New\" button on Work Orders list)", "MATCH")],
)
add(
    id="QA-52", priority="P3", category="Intended-grant-to-confirm",
    roles="Admin, Service Manager, Senior Service Advisor, Service Advisor, Foreman",
    capability="\"Approve\"/\"Decline\" line parity (MATCH rows) for the 5 approving roles",
    observed="Both envs SHOWN for Admin, Service Manager, Senior Service Advisor, Service Advisor, Foreman; both envs hidden for Technician, Parts Technician, Office User, Sales Representative, Time Clock User. Only delta = Parts Manager (QA-32/33).",
    spec=spec_says("authorize", "spec §1b WO Lines Edit + §4 Tech View approve-block"),
    why="Smoke-only if time permits: all 11 staging + 13 prod roles observed live (Pass-12, consolidated 2026-07-16; some staging role cells reuse 2026-07-15 captures — Parts Technician's is flagged for re-observe in GAP-09).",
    steps=APPROVE_STEPS_T.format(role="Foreman (representative approving role)"),
    expected="Green \"Approve\" + red \"Decline\" appear on a Needs-Approval line for the 5 approving roles only.",
    cases=[26390, 26458],
    evidence=EV16 + "/staging + /production (approve-decline JSON + PNG captures)",
    owner="QA smoke (time permitting)",
    sources=[("Pass-12 LIVE (2026-07-16)", "Foreman", "Approve/Decline line", "MATCH")],
)

# ========================= Coverage gaps (Tab 5) ===========================
GAPS = [
    dict(id="GAP-01", gtype="No existing TestRail case",
         roles="All roles", capability="Red \"Decline\" button on a \"Needs Approval\" line (role-based)",
         why="No case in the Custom Roles suite mentions Decline. It is a discrete money/authorization action gated with Approve (§1b ‘authorize lines’), and it carries an open Technician ambiguity (QA-08) and a Parts Manager migration delta (QA-33).",
         action="Author per-role cases: Decline shows/works for Admin/SM/SSA/SA/Foreman/Parts Manager; hidden for the rest; Technician pending PO ruling.",
         related="QA-08, QA-33, QA-48", evidence=EV16 + "/staging/approve-decline-REALHOLDER-ALL.json"),
    dict(id="GAP-02", gtype="No existing TestRail case",
         roles="All roles with Work Orders View", capability="Work order \"Notes\" tab (see tab, create note, edit any note)",
         why="Only a customer-notes case exists (C27873). No case covers the WO Notes tab riding on Work Orders View, which is also the surface of the Office User deviation (QA-05).",
         action="Author cases: WO Notes tab visible + create/edit-any-note for WO-View roles; unreachable for Office User (per the 7/14 spec).",
         related="QA-05, QA-38, QA-39", evidence=EV15 + "/ (per-role WO-detail captures)"),
    dict(id="GAP-03", gtype="No existing TestRail case",
         roles="All roles with Work Orders View", capability="Part \"Return\" from a work order (ungated picked-part menu action)",
         why="Existing return cases (C26419, C27871, C26417) cover inventory/vendor returns, a different surface. The ungated return-from-WO (29-Jun rule) has no case, and it carries the Office User deviation (QA-06).",
         action="Author cases: \"Return\" appears in the picked-part menu for every WO-View role; unreachable for Office User (per the 7/14 spec).",
         related="QA-06, QA-44, QA-45, QA-49", evidence=EV16 + "/ (Pass-11 Part Return captures)"),
    dict(id="GAP-04", gtype="No existing TestRail case",
         roles="Roles with / without Work Orders Create & Edit", capability="Change the Customer on an existing work order",
         why="C26375 covers status/lead-tech/service-advisor changes only; no case exercises changing the Customer on an existing WO, which is a live migration delta (lost by Technician/Office User, gained by Parts Manager).",
         action="Author cases: pencil control next to Customer present for WO-C&E roles (incl. Parts Manager), absent for Technician/Office User.",
         related="QA-16, QA-18, QA-30", evidence=EV15 + "/ (per-role WO-detail captures)"),
    dict(id="GAP-05", gtype="No existing TestRail case",
         roles="Roles with / without Work Orders Create & Edit", capability="Change the Asset on an existing work order",
         why="Same gap as GAP-04 for the Asset control (lost by Technician/Office User, gained by Parts Manager).",
         action="Author cases: pencil control next to Asset present for WO-C&E roles (incl. Parts Manager), absent for Technician/Office User.",
         related="QA-17, QA-19, QA-31", evidence=EV15 + "/ (per-role WO-detail captures)"),
    dict(id="GAP-06", gtype="No existing TestRail case",
         roles="Roles reaching the Finance tab", capability="\"Issue Credit\" button on the work order Finance tab",
         why="No case covers issuing a credit, and the spec is silent on its gate (QA-47). Money-facing surface with zero coverage.",
         action="Ask the PO to define the gate, then author per-role cases for \"Issue Credit\".",
         related="QA-47", evidence=EV16 + "/ (Pass-11 Finance captures)"),
    dict(id="GAP-07", gtype="No existing TestRail case",
         roles="Admin, Service Manager, Senior Service Advisor", capability="\"Reverse\" invoice via the Work Orders → Delete gate (positive case)",
         why="Only the negative case exists (C26422: Invoicing Delete does NOT reverse invoices). No positive case asserts that WO Delete holders see and can use \"Reverse\" on the Finance tab — and the SM answer is PO-pending (QA-07).",
         action="After the QA-07 PO ruling, author positive cases: \"Reverse\" present for Admin/SSA (and SM per ruling), absent for all non-WO-Delete roles.",
         related="QA-07, QA-10", evidence=EV16 + "/ (Pass-11 Finance captures)"),
    dict(id="GAP-08", gtype="Re-observe on release build",
         roles="Foreman, Office User, Parts Technician", capability="Staging \"Send to Terminal\" for these 3 roles — peer-observed method cell",
         why="The ground-truth workbook's \"Send to Terminal LIVE\" tab labels these 3 staging cells as method-derived: \"New Payment reachable (Take Payment observed SHOWN); org-device gate satisfied\" — the button itself was pixel-observed only for the invoicing-role set (SendToTerminal_dialog.png exists for Admin/Service Manager/Senior Service Advisor/Service Advisor/Parts Manager, not for these 3), and derived from the reachable New Payment dialog for these 3. Re-observe directly on the release build.",
         action="On the release build, as each of the 3 roles, open Finance > \"New Payment\" on an invoiced WO and pixel-confirm the \"Send to Terminal\" button state.",
         related="QA-01, QA-02, QA-03", evidence=EV15 + "/staging/*/SendToTerminal_dialog.png (pixel-observed roles only) + " + EV16 + "/production/_terminal (Pass-12 org-gate captures)"),
    dict(id="GAP-09", gtype="Re-observe on release build",
         roles="Parts Technician", capability="Staging \"Approve\"/\"Decline\" for Parts Technician — taken from the 2026-07-15 capture",
         why="The Approve-Decline LIVE tab sources the staging Parts Technician verdict from the 2026-07-15 role-swap capture (live-ui-2026-07-15/staging/Parts_Technician/refwo-caps.json) rather than the 2026-07-16 real-holder pass. Re-observe on the release build.",
         action="On the release build, as Parts Technician, open a WO with a \"Needs Approval\" line and confirm no \"Approve\"/\"Decline\" buttons appear.",
         related="QA-52", evidence=EV15 + "/staging/Parts_Technician/refwo-caps.json"),
]

# --------------------------------------------------------------------------
# Case-title index (for the TestRail columns)
# --------------------------------------------------------------------------
def load_case_titles():
    import json as _json
    titles = {}
    for fn in os.listdir(CASES_DIR):
        if fn.startswith("C") and fn.endswith(".json"):
            cid = int(fn[1:-5])
            try:
                with open(os.path.join(CASES_DIR, fn)) as f:
                    titles[cid] = _json.load(f).get("title", "")
            except Exception:
                titles[cid] = ""
    return titles


GAP_TEXT = "NO EXISTING CASE — coverage gap"


def cases_cell(cases, titles):
    if not cases:
        return GAP_TEXT
    parts = []
    for c in cases[:3]:
        parts.append("C{} — {}\n{}".format(c, titles.get(c, ""), TR_URL.format(c)))
    return "\n".join(parts)


# --------------------------------------------------------------------------
# Workbook build
# --------------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
P1_FILL = PatternFill("solid", fgColor="FFC7CE")
P2_FILL = PatternFill("solid", fgColor="FFEB9C")
P3_FILL = PatternFill("solid", fgColor="C6EFCE")
GAP_FILL = PatternFill("solid", fgColor="E4DFEC")
WRAP = Alignment(wrap_text=True, vertical="top")

ITEM_HEADERS = ["ID", "Category", "Role(s)", "Capability (plain language)",
                "What we observed (prod vs staging)", "Spec says (verbatim + citation)",
                "Why it needs checking", "Check steps (manual tester)",
                "Expected result per spec", "TestRail Case ID(s) + link(s)",
                "Evidence path", "Suggested owner", "Source row(s) in the ground-truth workbook"]
ITEM_WIDTHS = [8, 22, 26, 38, 42, 52, 42, 52, 42, 46, 40, 26, 46]


def src_cell(sources):
    out = []
    for (tab, role, cap, verdict) in sources:
        capd = cap if cap is not None else TAB_CFG[tab].get("fixed_cap", "")
        out.append("{} | {} | {} | {}".format(tab, role, capd, verdict))
    return "\n".join(out)


def write_item_tab(wb, title, items, fill, titles):
    ws = wb.create_sheet(title)
    ws.append(ITEM_HEADERS)
    for c in ws[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = WRAP
    for it in items:
        ws.append([it["id"], it["category"], it["roles"], it["capability"],
                   it["observed"], it["spec"], it["why"], it["steps"], it["expected"],
                   cases_cell(it["cases"], titles), it["evidence"], it["owner"],
                   src_cell(it["sources"])])
        r = ws.max_row
        for c in ws[r]:
            c.alignment = WRAP
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=1).font = Font(bold=True)
    for i, w in enumerate(ITEM_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"
    return ws


def build():
    titles = load_case_titles()
    p1 = [i for i in ITEMS if i["priority"] == "P1"]
    p2 = [i for i in ITEMS if i["priority"] == "P2"]
    p3 = [i for i in ITEMS if i["priority"] == "P3"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    cat_counts = {}
    for it in ITEMS:
        cat_counts[it["category"]] = cat_counts.get(it["category"], 0) + 1
    cat_counts["Coverage-gap"] = len(GAPS)

    mapped = sum(1 for it in ITEMS if it["cases"])
    unmapped = sum(1 for it in ITEMS if not it["cases"])

    rows = [
        ["CUSTOM ROLES — QA PRE-RELEASE CHECKLIST (2026-07-16)"],
        ["Detailed, actionable QA-team companion to the executive file "
         "CustomRoles_Release-Readiness_Prod-vs-Staging_EXEC_2026-07-16.xlsx. "
         "Every row traces to a LIVE-OBSERVED cell of the spec-annotated prod-vs-staging comparison."],
        [],
        ["HOW TO USE THIS FILE"],
        ["Work the tabs in order. P1 = must check before release (spec deviations, spec self-contradictions "
         "needing a PO ruling, and customer/money-facing permission changes) — do every row. "
         "P2 = intended-per-spec migration changes — regression-confirm each on the release build. "
         "P3 = monitor / smoke-only items (org-config gates, spec-silent components, parity already observed). "
         "Coverage Gaps = findings with no existing TestRail case (author these) plus two cells to re-observe "
         "directly on the release build. Each row carries exact build-accurate steps, the verbatim spec gate with "
         "citation, TestRail case links, the evidence path, and the exact source row(s) in the ground-truth workbook. "
         "Follow Standing Rule 12: mark a check done only after observing it live with evidence."],
        [],
        ["COUNTS BY PRIORITY"],
        ["P1 — Must check before release", len(p1)],
        ["P2 — Should check (intended per spec; regression-confirm)", len(p2)],
        ["P3 — Monitor / no action", len(p3)],
        ["Coverage Gaps (new cases to author + re-observe items)", len(GAPS)],
        ["TOTAL checklist rows", len(ITEMS) + len(GAPS)],
        [],
        ["COUNTS BY CATEGORY"],
    ]
    for cat in ["Deviation", "Spec-contradiction", "Spec-silent", "Intended-loss-to-confirm",
                "Intended-grant-to-confirm", "Coverage-gap", "Env-config"]:
        rows.append([cat, cat_counts.get(cat, 0)])
    rows += [
        [],
        ["TESTRAIL MAPPING (Standing Rule 8)"],
        ["Checklist rows mapped to existing TestRail case(s)", mapped],
        ["Checklist rows with NO existing case (see Coverage Gaps)", unmapped],
        [],
        ["GROUND-TRUTH LINEAGE (every row traces here)"],
        ["Source workbook", "build/custom-roles-run/Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx (Full Dual Matrix, 176 dual rows + Pass-11/Pass-12/Approve-Decline/Send to Terminal/Parts-Module/New-WO tabs + Spec-Standing Conformance)"],
        ["Spec truth table", TRUTH_REF],
        ["Canonical spec", SPEC_REF],
        ["Evidence", EV15 + "/ and " + EV16 + "/"],
        ["TestRail case source", "build/custom-roles-run/cases-2026-07-13/ (filename = C<id>)"],
        ["Generator", "build/custom-roles-run/gen_qa_prerelease.py (accuracy gate built in)"],
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    for label_row in (4, 7, 14, 22, 26):
        ws.cell(row=label_row, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 130
    for row in ws.iter_rows():
        for c in row:
            c.alignment = WRAP
    ws.freeze_panes = "A2"

    write_item_tab(wb, "P1 — Must check before release", p1, P1_FILL, titles)
    write_item_tab(wb, "P2 — Should check", p2, P2_FILL, titles)
    write_item_tab(wb, "P3 — Monitor - no action", p3, P3_FILL, titles)

    ws = wb.create_sheet("Coverage Gaps")
    headers = ["Gap ID", "Type", "Role(s)", "Capability / surface",
               "What is missing and why it matters", "Suggested action for QA",
               "Related checklist item(s)", "Evidence path"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
        c.alignment = WRAP
    for g in GAPS:
        ws.append([g["id"], g["gtype"], g["roles"], g["capability"], g["why"],
                   g["action"], g["related"], g["evidence"]])
        r = ws.max_row
        for c in ws[r]:
            c.alignment = WRAP
        ws.cell(row=r, column=1).fill = GAP_FILL
        ws.cell(row=r, column=1).font = Font(bold=True)
    for i, w in enumerate([9, 26, 30, 40, 60, 55, 22, 45], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "B2"

    wb.save(OUT_XLSX)
    print("wrote", OUT_XLSX)


# --------------------------------------------------------------------------
# ACCURACY GATE
# --------------------------------------------------------------------------
def load_src_rows():
    wb = openpyxl.load_workbook(SRC_XLSX, read_only=True)
    data = {}
    for tab, cfg in TAB_CFG.items():
        rows = []
        for row in wb[tab].iter_rows(values_only=True):
            rows.append([("" if v is None else str(v)) for v in row])
        data[tab] = rows
    return data


def find_src(data, tab, role, cap, verdict_prefix):
    cfg = TAB_CFG[tab]
    hits = []
    for row in data[tab]:
        if len(row) <= cfg["verdict"]:
            continue
        r_role = row[cfg["role"]].strip()
        if r_role != role:
            continue
        if cfg["cap"] is not None:
            r_cap = row[cfg["cap"]].strip()
            if r_cap != cap:
                continue
        v = row[cfg["verdict"]].strip()
        if not v:
            continue
        hits.append(v)
    ok = [v for v in hits if v.startswith(verdict_prefix)]
    return ok, hits


def verify():
    errors = []
    data = load_src_rows()
    spec_text = open(SPEC_MD, encoding="utf-8").read()

    # (a) source-cell match
    for it in ITEMS:
        for (tab, role, cap, verdict) in it["sources"]:
            capq = cap if cap is not None else TAB_CFG[tab].get("fixed_cap")
            ok, hits = find_src(data, tab, role, cap, verdict)
            if not ok:
                errors.append("(a) {}: no row in [{}] role={!r} cap={!r} verdict^={!r}; found verdicts={!r}"
                              .format(it["id"], tab, role, capq, verdict, hits[:3]))

    # (b) spec quotes verbatim
    for key, q in Q.items():
        if q not in spec_text:
            errors.append("(b) spec quote key {!r} NOT found verbatim in canonical spec: {!r}".format(key, q[:80]))

    # (c) cited C-ids exist
    for it in ITEMS:
        for c in (it["cases"] or []):
            if not os.path.exists(os.path.join(CASES_DIR, "C{}.json".format(c))):
                errors.append("(c) {}: C{} not in cases-2026-07-13/".format(it["id"], c))

    # re-open the built workbook for (d)+(e)
    wb = openpyxl.load_workbook(OUT_XLSX)
    tabrows = {}
    for name in ["P1 — Must check before release", "P2 — Should check",
                 "P3 — Monitor - no action", "Coverage Gaps"]:
        ws = wb[name]
        n = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r[0])
        tabrows[name] = n
    exp = {"P1 — Must check before release": sum(1 for i in ITEMS if i["priority"] == "P1"),
           "P2 — Should check": sum(1 for i in ITEMS if i["priority"] == "P2"),
           "P3 — Monitor - no action": sum(1 for i in ITEMS if i["priority"] == "P3"),
           "Coverage Gaps": len(GAPS)}
    for k, v in exp.items():
        if tabrows[k] != v:
            errors.append("(d) tab {!r} has {} rows, expected {}".format(k, tabrows[k], v))

    # (d) summary counts vs tab counts
    summary = {}
    for row in wb["Summary"].iter_rows(values_only=True):
        if row and row[0] is not None and len(row) > 1 and isinstance(row[1], (int, float)):
            summary[str(row[0])] = int(row[1])
    checks = {
        "P1 — Must check before release": exp["P1 — Must check before release"],
        "P2 — Should check (intended per spec; regression-confirm)": exp["P2 — Should check"],
        "P3 — Monitor / no action": exp["P3 — Monitor - no action"],
        "Coverage Gaps (new cases to author + re-observe items)": exp["Coverage Gaps"],
    }
    for k, v in checks.items():
        if summary.get(k) != v:
            errors.append("(d) Summary row {!r}={} but tab count={}".format(k, summary.get(k), v))
    # category counts
    cat_counts = {}
    for it in ITEMS:
        cat_counts[it["category"]] = cat_counts.get(it["category"], 0) + 1
    cat_counts["Coverage-gap"] = len(GAPS)
    for cat, v in cat_counts.items():
        if summary.get(cat) != v:
            errors.append("(d) Summary category {!r}={} expected {}".format(cat, summary.get(cat), v))

    # (e) forbidden strings
    forbidden = ["NOT VERIFIED", "TODO", "TBD", "PLACEHOLDER", "XXX", "FIXME", "LOREM"]
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    up = v.upper()
                    for f in forbidden:
                        if f in up:
                            errors.append("(e) forbidden string {!r} in tab {!r}: {!r}".format(f, ws.title, v[:80]))

    # every P1/P2/P3 row has a non-empty TestRail cell (either mapped or explicit gap text)
    for name in ["P1 — Must check before release", "P2 — Should check", "P3 — Monitor - no action"]:
        for r in wb[name].iter_rows(min_row=2, values_only=True):
            if r[0] and not r[9]:
                errors.append("(e) {} row {} has empty TestRail cell".format(name, r[0]))

    if errors:
        print("ACCURACY GATE FAILED ({} errors):".format(len(errors)))
        for e in errors:
            print("  -", e)
        return False
    print("ACCURACY GATE PASSED: (a) {} source refs, (b) {} spec quotes, (c) {} case citations, "
          "(d) counts match, (e) no placeholders.".format(
              sum(len(i["sources"]) for i in ITEMS), len(Q),
              sum(len(i["cases"] or []) for i in ITEMS)))
    return True


if __name__ == "__main__":
    build()
    ok = verify()
    sys.exit(0 if ok else 1)
