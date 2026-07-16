#!/usr/bin/env python3
"""
Generate the EXECUTIVE, boss-ready release-readiness deliverable for the Custom Roles
prod-vs-staging permission comparison.

SOURCE OF TRUTH (read live, never hard-coded where avoidable):
  Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx  (Full Dual Matrix + conformance tabs)

OUTPUT:
  CustomRoles_Release-Readiness_Prod-vs-Staging_EXEC_2026-07-16.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter, OrderedDict
import os

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, 'Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx')
OUT = os.path.join(HERE, 'CustomRoles_Release-Readiness_Prod-vs-Staging_EXEC_2026-07-16.xlsx')

# ---------------------------------------------------------------- read source
src = openpyxl.load_workbook(SRC, data_only=True)
fdm = src['Full Dual Matrix']
rows = [r for r in fdm.iter_rows(values_only=True)][1:]
data = [r for r in rows if r[0] is not None and r[5] is not None]


def norm_verdict(v):
    v = str(v)
    if v.startswith('MATCH'):
        return 'MATCH'
    if v.startswith('STAGING-MORE'):
        return 'STAGING-MORE'
    if v.startswith('STAGING-LESS'):
        return 'STAGING-LESS'
    raise ValueError('Unexpected verdict: ' + v)


def spec_class(v):
    v = str(v)
    if v.startswith('Per spec'):
        return 'per-spec'
    if v.startswith('DEVIATION'):
        return 'DEVIATION'
    if v.startswith('Spec inconsistent'):
        return 'inconsistent'
    if v.startswith('Spec silent'):
        return 'silent'
    raise ValueError('Unexpected spec class: ' + v[:80])


counts = Counter(norm_verdict(r[5]) for r in data)
N_TOTAL = len(data)
N_MATCH = counts['MATCH']
N_MORE = counts['STAGING-MORE']
N_LESS = counts['STAGING-LESS']
assert (N_TOTAL, N_MATCH, N_MORE, N_LESS) == (176, 130, 26, 20), (N_TOTAL, N_MATCH, N_MORE, N_LESS)

# conformance tally recomputed live from the 7 annotated tabs
CONF_TABS = ['Full Dual Matrix', 'Pass-11 LIVE (2026-07-16)', 'Pass-12 LIVE (2026-07-16)',
             'Approve-Decline LIVE', 'Send to Terminal LIVE', 'Parts-Module Dual LIVE',
             'New-WO Create Dual LIVE']
conf = Counter()
for name in CONF_TABS:
    ws = src[name]
    trows = list(ws.iter_rows(values_only=True))
    hdr_idx = col = None
    for i, r in enumerate(trows):
        for j, c in enumerate(r):
            if c == 'Per Spec (v2)?':
                hdr_idx, col = i, j
        if hdr_idx is not None:
            break
    for r in trows[hdr_idx + 1:]:
        if col < len(r) and r[col]:
            v = str(r[col])
            if v.startswith('LEGEND') or 'Per spec-expected/matches |' in v:
                continue
            conf[spec_class(v)] += 1
N_CONF = sum(conf.values())
assert N_CONF == 297 and conf['per-spec'] == 283 and conf['DEVIATION'] == 9 and conf['inconsistent'] == 5, dict(conf)

PCT_MATCH = round(100.0 * N_MATCH / N_TOTAL)          # 74
PCT_SPEC = round(100.0 * conf['per-spec'] / N_CONF)   # 95

# non-match rows for Tab 2 (strictly from the Full Dual Matrix)
nonmatch = [r for r in data if norm_verdict(r[5]) != 'MATCH']
assert len(nonmatch) == 46

ROLE_CLEAN = {
    'Senior Service Advisor (+SA Tech,+SA NoRep)': 'Senior Service Advisor',
    'Sales Representative (+Reporting)': 'Sales Representative',
}

PLAIN_CAP = {
    'Send to Portal': "Can send a work order to the customer's online portal",
    'Create/Edit WO Lines (New Line)': 'Can add or edit work order lines',
    'Review Work Orders (Reviewed)': 'Can mark a work order as reviewed',
    'See Financial Data (Rate/Margin)': 'Can see financial data (rates and margins)',
    'Invoicing/Finance view (Finance tab)': 'Can open the finance / invoicing area of a work order',
    'Take Payment (New Payment)': 'Can take a customer payment',
    'Send to Terminal': 'Can send a payment to the card terminal',
    'WO-level History (History tab)': "Can see a work order's change history",
    'Order Parts area (Parts tab)': 'Can order parts from a work order',
    'WO Notes (Notes tab)': 'Can see and add work order notes',
    'Timesheets (tab)': 'Can see the timesheets area',
    'WO Delete': 'Can delete a work order',
    'Change Customer on WO': 'Can change the customer on a work order',
    'Change Asset on WO': 'Can change the vehicle on a work order',
    'Approve line': 'Can approve a work order line that is waiting for authorization',
    'Decline line': 'Can decline a work order line that is waiting for authorization',
}


def perspec_label(cls, verdict):
    if cls == 'per-spec':
        return 'Yes — expected'
    if cls == 'DEVIATION':
        return 'DEVIATION'
    return 'Spec inconsistent'


def risk_for(cap, cls, verdict):
    if cap == 'Take Payment (New Payment)':
        return 'High'
    if cls == 'DEVIATION' or cls == 'inconsistent':
        return 'Medium'
    if cap in ('See Financial Data (Rate/Margin)', 'Approve line', 'Decline line'):
        return 'Medium'
    return 'Low'


def action_for(cap, cls, risk):
    if cls == 'DEVIATION':
        return ('Decide before release: align the build to the spec rule (payment terminal should '
                'follow the Customer Portal setting) or update the spec. Org-configuration item, '
                'not a data risk.')
    if cls == 'inconsistent':
        return 'Product Owner to resolve the contradictory spec wording, then re-check this role.'
    if risk == 'High':
        return 'Confirm with the Product Owner that this is intended (it matches the specification), then release.'
    if risk == 'Medium':
        return 'Confirm intended — it matches the specification.'
    return 'None — intended by the specification.'


# ---------------------------------------------------------------- styling
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=16, color='1F4E79')
SUB_FONT = Font(bold=True, size=12, color='1F4E79')
GREEN = PatternFill('solid', fgColor='C6EFCE')
GREEN_F = Font(color='006100', bold=True)
AMBER = PatternFill('solid', fgColor='FFEB9C')
AMBER_F = Font(color='9C6500', bold=True)
RED = PatternFill('solid', fgColor='FFC7CE')
RED_F = Font(color='9C0006', bold=True)
GREY = PatternFill('solid', fgColor='F2F2F2')
WRAP = Alignment(wrap_text=True, vertical='top')
THIN = Border(*[Side(style='thin', color='BFBFBF')] * 4)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = WRAP
        cell.border = THIN


def risk_style(cell, risk):
    if risk == 'High':
        cell.fill, cell.font = RED, RED_F
    elif risk == 'Medium':
        cell.fill, cell.font = AMBER, AMBER_F
    else:
        cell.fill, cell.font = GREEN, GREEN_F


wb = openpyxl.Workbook()

# ================================================================ TAB 1
ws = wb.active
ws.title = 'Executive Summary'
ws.sheet_view.showGridLines = False
widths = [3, 34, 96, 18]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

r = 2
ws.cell(row=r, column=2, value='Custom Roles & Permissions — Release Readiness').font = TITLE_FONT
r += 1
ws.cell(row=r, column=2, value='Production vs Staging (new roles model) — Executive Summary — 16 July 2026').font = Font(size=12, color='404040')
r += 2
ws.cell(row=r, column=2, value='Scope').font = SUB_FONT
r += 1
scope = ('All 14 production roles compared against all 11 new-model staging roles; every permission/function '
         'verified LIVE on-screen in both environments with captured evidence (screenshots + saved responses); '
         '{} role-capability comparisons; zero unverified items.').format(N_TOTAL)
ws.cell(row=r, column=2, value=scope).alignment = WRAP
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.row_dimensions[r].height = 44
r += 2

ws.cell(row=r, column=2, value='The verdict').font = SUB_FONT
r += 1
verdict = ('The new roles model is overwhelmingly faithful to both production behavior and the specification: '
           '{m} of {t} comparisons ({pm}%) match production exactly, and {pspec}% of all spec-conformance checks '
           '({ps} of {pt}) are per-spec. Every one of the {nm} differences was verified live, and {ok} of the {nm} '
           'are intended by the specification. What remains: 6 spec deviations (three of them one shared '
           'card-terminal configuration issue) and 3 spec self-contradictions that need decisions before release.'
           ).format(m=N_MATCH, t=N_TOTAL, pm=PCT_MATCH, pspec=PCT_SPEC, ps=conf['per-spec'], pt=N_CONF,
                    nm=len(nonmatch), ok=sum(1 for x in nonmatch if spec_class(x[7]) == 'per-spec'))
cell = ws.cell(row=r, column=2, value=verdict)
cell.alignment = WRAP
cell.font = Font(size=11, bold=True)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.row_dimensions[r].height = 74
r += 2

ws.cell(row=r, column=2, value='Key numbers').font = SUB_FONT
r += 1
keynums = [
    ('Role-capability comparisons (both environments, live)', '{}'.format(N_TOTAL)),
    ('   … match production exactly', '{}'.format(N_MATCH)),
    ('   … staging grants more than production', '{}'.format(N_MORE)),
    ('   … staging grants less than production', '{}'.format(N_LESS)),
    ('Spec-conformance checks (every verified row judged against the specification)', '{}'.format(N_CONF)),
    ('   … agree with the specification (per-spec)', '{}'.format(conf['per-spec'])),
    ('   … deviation rows (= 6 distinct findings)', '{}'.format(conf['DEVIATION'])),
    ('   … rows where the specification contradicts itself (= 3 distinct items)', '{}'.format(conf['inconsistent'])),
]
for label, val in keynums:
    ws.cell(row=r, column=2, value=label).alignment = WRAP
    c = ws.cell(row=r, column=3, value=int(val))
    c.font = Font(bold=True)
    if label.strip().startswith('…') or label.startswith('   '):
        ws.cell(row=r, column=2).font = Font(color='404040')
    else:
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=2).fill = GREY
        c.fill = GREY
    r += 1
r += 1

ws.cell(row=r, column=2, value='Decisions needed before release').font = SUB_FONT
r += 1
decisions = [
    ('RED', 'Service Advisor loses the ability to reverse an invoice',
     'In production this role can reverse an invoice; in the new model it cannot (the button is gone). This is '
     'exactly what the specification prescribes (invoice reversal now requires work-order delete rights, which '
     'this role no longer has) — but it is customer-billing-facing, so confirm it is intended before release.'),
    ('RED', 'Taking customer payments is newly granted to 6 roles',
     'Service Manager, Senior Service Advisor, Foreman, Parts Manager, Parts Technician and Office User can all '
     'take a customer payment in the new model; none of them can in production. This matches the specification — '
     'but it is money-handling, so confirm it is intended before release.'),
    ('RED', 'Senior Service Advisor: AR/AP aging reports promised but not delivered',
     'The specification grants this role the receivables/payables aging reports, but the build hides them '
     '(verified in both environments). This is an unimplemented spec grant — needs a dev fix or a spec change.'),
    ('AMBER', 'Office User still sees work-order notes and part returns',
     'The 14-July spec update removed ALL work-order access for Office User, yet the build still shows the notes '
     'tab (in both environments) and the part-return control (in staging). Both are small-surface over-grants '
     'versus the spec — decide: enforce the spec or amend it.'),
    ('AMBER', '"Send to Terminal" shows for three roles the spec would hide it from',
     'Foreman, Office User and Parts Technician see the card-terminal payment button on staging because the '
     'staging organization has a card terminal configured. The spec says these roles (Customer Portal switched '
     'off) should not see it. This is a difference in HOW the button is gated (organization device vs role '
     'setting), not a data risk — decide which gating rule should win.'),
    ('AMBER', 'The specification contradicts itself in three places — Product Owner to resolve',
     '(1) Service Manager invoice reversal: one spec table says the role can reverse, another says it cannot — '
     'the build follows the "cannot" table. (2) Technician declining a work-order line: one section grants it, '
     'another blocks only approving — unresolvable as written; both environments hide it. (3) "Send to Portal" '
     'for Office User, Parts Technician and Sales Representative: one section grants it to every full-view role, '
     'an answered open question restricts it to line-approvers — the build follows the stricter reading.'),
    ('GREEN', 'Card-terminal access is safe through the migration',
     'Verified: "Send to Terminal" depends on whether the organization has a card terminal device, not on the '
     'roles migration. Moving to the new roles model does NOT change anyone\'s terminal access.'),
    ('GREEN', 'Parts ordering and receiving match production exactly',
     'All 22 live parts-module checks (create purchase orders, receive deliveries) match production for every '
     'role — no change for the parts workflow.'),
    ('GREEN', 'Work-order deletion correctly tightened',
     'Only Admin, Service Manager and Senior Service Advisor can delete a work order in the new model — exactly '
     'as the specification prescribes (verified live for all 11 roles).'),
]
ws.cell(row=r, column=2, value='Risk').font = HDR_FONT
ws.cell(row=r, column=2).fill = HDR_FILL
ws.cell(row=r, column=3, value='Decision / status').font = HDR_FONT
ws.cell(row=r, column=3).fill = HDR_FILL
r += 1
for risk, head, body in decisions:
    c = ws.cell(row=r, column=2, value=risk)
    if risk == 'RED':
        c.fill, c.font = RED, RED_F
    elif risk == 'AMBER':
        c.fill, c.font = AMBER, AMBER_F
    else:
        c.fill, c.font = GREEN, GREEN_F
    c.alignment = Alignment(horizontal='center', vertical='top')
    c.border = THIN
    c2 = ws.cell(row=r, column=3, value='{} — {}'.format(head, body))
    c2.alignment = WRAP
    c2.border = THIN
    ws.row_dimensions[r].height = 78
    r += 1
r += 1

ws.cell(row=r, column=2, value='How much to trust this').font = SUB_FONT
r += 1
trust = ('Every result was observed live on-screen in both environments (screenshots and captured responses are '
         'archived); nothing was inferred from documents, role definitions or code. Spec conformance was derived '
         'from a verbatim, citation-backed truth table of the v2 specification.')
c = ws.cell(row=r, column=2, value=trust)
c.alignment = WRAP
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
ws.row_dimensions[r].height = 44
ws.freeze_panes = 'A2'

# ================================================================ TAB 2
ws2 = wb.create_sheet('Migration Deltas (by role)')
ws2.sheet_view.showGridLines = False
headers = ['Role', 'What changes for them', 'Direction', 'Per Spec?', 'Risk', 'Suggested action']
w2 = [26, 52, 12, 18, 10, 62]
for i, w in enumerate(w2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.cell(row=1, column=1, value='What each role gains or loses moving from production to the new roles model '
         '({} differences out of {} live comparisons: {} gains / {} losses)'.format(
             len(nonmatch), N_TOTAL, N_MORE, N_LESS)).font = Font(bold=True, size=12, color='1F4E79')
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
for j, h in enumerate(headers, 1):
    ws2.cell(row=3, column=j, value=h)
style_header(ws2, 3, len(headers))
ws2.freeze_panes = 'A4'

# group by role, preserving matrix role order
role_order = []
for rrow in nonmatch:
    role = ROLE_CLEAN.get(rrow[0], rrow[0])
    if role not in role_order:
        role_order.append(role)
rr = 4
for role in role_order:
    for rrow in nonmatch:
        if ROLE_CLEAN.get(rrow[0], rrow[0]) != role:
            continue
        cap = rrow[2]
        verdict = norm_verdict(rrow[5])
        cls = spec_class(rrow[7])
        direction = 'Gains' if verdict == 'STAGING-MORE' else 'Loses'
        risk = risk_for(cap, cls, verdict)
        what = PLAIN_CAP[cap]
        if cap == 'Send to Terminal':
            what += ' (only because the staging organization has a terminal device — the production organization has none)'
        vals = [role, what, direction, perspec_label(cls, verdict), risk,
                action_for(cap, cls, risk)]
        for j, v in enumerate(vals, 1):
            c = ws2.cell(row=rr, column=j, value=v)
            c.alignment = WRAP
            c.border = THIN
        # colors
        dcell = ws2.cell(row=rr, column=3)
        if direction == 'Gains':
            dcell.font = Font(color='1F4E79', bold=True)
        else:
            dcell.font = Font(color='9C0006', bold=True)
        pcell = ws2.cell(row=rr, column=4)
        if cls == 'per-spec':
            pcell.fill, pcell.font = GREEN, GREEN_F
        elif cls == 'DEVIATION':
            pcell.fill, pcell.font = RED, RED_F
        else:
            pcell.fill, pcell.font = AMBER, AMBER_F
        risk_style(ws2.cell(row=rr, column=5), risk)
        ws2.row_dimensions[rr].height = 42
        rr += 1
n_written = rr - 4
assert n_written == 46, n_written

# ================================================================ TAB 3
ws3 = wb.create_sheet('Spec Deviations & Open Qs')
ws3.sheet_view.showGridLines = False
h3 = ['#', 'Finding', 'Evidence (environment + what was observed)', 'Spec citation',
      'Recommended owner', 'Suggested next step']
w3 = [4, 40, 52, 40, 16, 42]
for i, w in enumerate(w3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.cell(row=1, column=1, value='Spec deviations (6 distinct findings, covering 9 verified rows), spec '
         'self-contradictions (3 items, 5 rows) and spec-silent components').font = Font(bold=True, size=12, color='1F4E79')
ws3.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
for j, h in enumerate(h3, 1):
    ws3.cell(row=3, column=j, value=h)
style_header(ws3, 3, len(h3))
ws3.freeze_panes = 'A4'

dev_rows = [
    ('D1', 'Foreman sees "Send to Terminal" though the spec role-rule would hide it',
     'Staging: button SHOWN for Foreman (staging org has a card terminal). Production: hidden for all roles '
     '(prod org has no terminal device; verified — no way to add one in the app, all terminal endpoints absent). '
     'Live-observed both environments, 15–16 July 2026.',
     'Spec §1i (06-Jul change): Send to Terminal requires Invoicing Create & Edit + "Customer Portal: ON"; '
     'Foreman has Customer Portal OFF.',
     'PO decision', 'Decide the gating rule: role-based (per spec) or organization-device-based (as built). '
     'Org-configuration item, not a data risk.'),
    ('D2', 'Parts Technician sees "Send to Terminal" though the spec role-rule would hide it',
     'Same observation as D1, for Parts Technician (Customer Portal OFF). Live-observed both environments.',
     'Spec §1i (06-Jul change): Invoicing Create & Edit + Customer Portal ON.',
     'PO decision', 'Same decision as D1.'),
    ('D3', 'Office User sees "Send to Terminal" though the spec role-rule would hide it',
     'Same observation as D1, for Office User (Customer Portal OFF). Live-observed both environments.',
     'Spec §1i (06-Jul change): Invoicing Create & Edit + Customer Portal ON.',
     'PO decision', 'Same decision as D1.'),
    ('D4', 'Senior Service Advisor: AR/AP aging reports granted by the spec but hidden by the build',
     'Both environments hide the receivables/payables aging reports for this role (staging live-observed; '
     'production live-observed for all three legacy roles that merge into it). The build does not implement '
     'the spec grant.',
     'Spec toggles matrix: Reports = ON for Senior Service Advisor; §2a: aging reports follow the Reports '
     'toggle (Manage AP/AR is also ON).',
     'Dev fix', 'Implement the grant (show the aging reports to this role) or amend the spec.'),
    ('D5', 'Office User can still open work-order notes though the 14-July spec removed all its work-order access',
     'BOTH environments show the notes tab to Office User (live-observed). The 14-July "Updated Office Role '
     'definition" sets Office\'s Work Orders access to none, so the notes tab should be unreachable.',
     'Spec permission matrix (14-Jul row): Office / Work Orders = "—"; notes are gated by Work Orders → View (§1a).',
     'PO decision', 'Decide: enforce the 14-July Office definition (hide the tab) or amend the spec. Small surface.'),
    ('D6', 'Office User can still use part return though the 14-July spec removed all its work-order access',
     'Staging shows the part-return control to Office User (production hides it — its part menu offers Move '
     'only). Live-observed both environments. The return control itself is ungated, but reaching it requires '
     'work-order view, which the 14-July matrix no longer gives Office.',
     'Spec §1a + 29-Jun note: returning a part needs no permission, but requires work-order view in practice; '
     '14-Jul matrix: Office / Work Orders = "—".',
     'PO decision', 'Same decision as D5 (same root cause: Office\'s residual work-order visibility).'),
]
inc_rows = [
    ('Q1', 'Spec self-contradiction — Service Manager invoice reversal',
     'Staging shows Service Manager New Payment + Issue Credit but NO invoice-reverse (live-observed). One spec '
     'table says the role CAN reverse; the migration table says it cannot. The build follows the migration table.',
     'Permission matrix (SM Work Orders = V/E/D) + 28-Jun rule (reverse = Work Orders Delete) vs migration '
     'behavior-changes table ("Loses Invoicing Delete (cannot reverse)").',
     'PO decision', 'Resolve which table wins; if the matrix wins, the build needs a change.'),
    ('Q2', 'Spec self-contradiction — Technician declining a work-order line',
     'Both environments hide Approve AND Decline from Technician (live-observed on a genuine needs-approval '
     'line). The spec both grants ("authorize lines") and, in its Tech-View section, blocks only Approve — the '
     'intended treatment of Decline is unresolvable as written.',
     '§1b Work Order Lines Create & Edit ("authorize lines") vs §4 Tech View (blocks Approve only, silent on Decline).',
     'PO decision', 'Clarify the spec; current build (hidden) is the conservative reading.'),
    ('Q3', 'Spec self-contradiction — "Send to Portal" for Office User, Parts Technician, Sales Representative',
     'Staging hides the button for all three (live-observed; for Office User and Parts Technician production '
     'shows it, so they lose it in the migration; for Sales Representative both environments hide it). One spec '
     'section grants the button to every full-view role; an answered open question restricts it to anyone who '
     'can approve a work-order line (which these roles cannot).',
     '§3/§4 "Full View has access to Send to Portal" vs Open Q6 answer: "can be anyone who can approve a WOL".',
     'PO decision', 'Resolve the wording; the build follows the stricter (approver-only) reading.'),
]
silent_rows = [
    ('S1', 'Spec-silent component — "Issue Credit"',
     'The Issue Credit button appears in the finance area for invoicing-capable roles (live-observed). The spec '
     'nowhere defines which permission should control issuing a credit (it covers only the credits-tab '
     'visibility), so conformance was deliberately NOT judged for this component.',
     'No issuance gate anywhere in the v2 spec; §5b covers tab visibility only.',
     'PO decision', 'Add an explicit gate for Issue Credit to the spec, then re-check.'),
]
rr = 4
for block, fill, font in ((dev_rows, RED, RED_F), (inc_rows, AMBER, AMBER_F), (silent_rows, GREY, Font(bold=True, color='404040'))):
    for row_vals in block:
        for j, v in enumerate(row_vals, 1):
            c = ws3.cell(row=rr, column=j, value=v)
            c.alignment = WRAP
            c.border = THIN
        idc = ws3.cell(row=rr, column=1)
        idc.fill, idc.font = fill, font
        ws3.row_dimensions[rr].height = 84
        rr += 1

# ================================================================ TAB 4
ws4 = wb.create_sheet('Coverage & Method')
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 3
ws4.column_dimensions['B'].width = 30
ws4.column_dimensions['C'].width = 105
ws4.cell(row=2, column=2, value='Coverage & Method').font = TITLE_FONT
method_rows = [
    ('Environments',
     'Production: the live ShopView production organization on the current (legacy) roles model. '
     'Staging: app.staging.shopview.com on the new custom-roles model.'),
    ('Roles covered',
     'All 14 production roles (Administrator; Service Manager; Service Advisor; SA Technician; SA No Reports; '
     'SA Limited View; Foreman; Technician; Parts Manager; Parts Technician; Sales Representative; Reporting; '
     'Office; Time Clock) compared against all 11 new-model staging roles, using the confirmed role-migration '
     'mapping (three SA variants merge into Senior Service Advisor; Sales Representative absorbs Reporting; '
     'legacy SA Limited View becomes the new Service Advisor).'),
    ('How each result was obtained',
     'Live login as each role, in each environment, and direct on-screen observation of each control/behavior '
     '(production: admin self-login + per-role switch-user or a role-swapped test staff member; staging: '
     'quick-login + real-holder switch-user for all 11 roles). Where a needed data state did not exist (e.g. a '
     'work-order line waiting for approval, an invoiced work order), it was seeded, observed, and cleaned up '
     'afterwards. No result was inferred from role definitions, permission data, documents or source code.'),
    ('Zero unverified items',
     'All {} role-capability comparisons carry a live-observed verdict. The single non-plain-observed item — '
     '"Send to Terminal" on production — is a fully characterized organization-configuration finding (the '
     'production organization has no card-terminal device and the app offers no way to add one), not an '
     'unverified gap.'.format(N_TOTAL)),
    ('Spec conformance',
     'Every one of the {} verified rows was additionally judged against the v2 specification using a verbatim, '
     'citation-backed truth table rebuilt from the canonical spec document; where the spec is silent or '
     'contradicts itself this is flagged, never resolved by guesswork.'.format(N_CONF)),
    ('Evidence archive',
     'build/custom-roles-run/live-ui-2026-07-15/ and build/custom-roles-run/live-ui-2026-07-16/ '
     '(full-page screenshots + captured per-role responses, both environments), plus '
     'build/custom-roles-run/compare-evidence-2026-07-14/.'),
    ('Full detail',
     'The complete row-by-row matrix behind this summary is in the detailed workbook: '
     'Prod-vs-Staging-LIVE-VERIFIED-2026-07-14.xlsx (tabs: Full Dual Matrix, Spec-Standing Conformance, '
     'per-capability live tabs, READ ME).'),
]
rr = 4
for label, text in method_rows:
    c = ws4.cell(row=rr, column=2, value=label)
    c.font = Font(bold=True, color='1F4E79')
    c.alignment = WRAP
    c2 = ws4.cell(row=rr, column=3, value=text)
    c2.alignment = WRAP
    c2.border = THIN
    c.border = THIN
    ws4.row_dimensions[rr].height = 72
    rr += 1
ws4.freeze_panes = 'A2'

wb.save(OUT)
print('WROTE', OUT)
print('Tab2 rows:', n_written, '| Gains:', N_MORE, '| Loses:', N_LESS)
print('Conformance:', dict(conf))
