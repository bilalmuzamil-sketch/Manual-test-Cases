#!/usr/bin/env python3
import json, glob, csv, os
from collections import defaultdict, OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = "build/custom-roles-run"
PREFIX = "Custom Roles - (Revised) > "
RUN_URL = "https://shopview.testrail.io/index.php?/runs/view/312"

def short_sec(path):
    return path[len(PREFIX):] if path.startswith(PREFIX) else path

# ---- Load run plan (in-scope) ----
plan = json.load(open(os.path.join(BASE, "run-plan.json")))
def is_excluded(p):
    return ("Regression Suite (Minja" in p) or ("Backend API and Security" in p)
inscope = OrderedDict()
for c in plan["cases"]:
    if not is_excluded(c["section_path"]):
        inscope[c["id"]] = c

# ---- Load results ----
results = {}
for f in sorted(glob.glob(os.path.join(BASE, "results", "*.jsonl"))):
    for line in open(f):
        line = line.strip()
        if not line:
            continue
        r = json.loads(line)
        results[r["case_id"]] = r

# ---- Merge: build per-case record for every in-scope case ----
def classify_type(r):
    reason = (r.get("reason") or "").lower()
    if "stale" in reason or "behavior ok" in reason or "behaviour ok" in reason:
        return "Stale case (UI/spec drift)"
    return "Real discrepancy"

records = []  # dicts with case_id, section(short), title, status, type, expected, actual, reason, logged
for cid, c in inscope.items():
    sec = short_sec(c["section_path"])
    if cid in results:
        r = results[cid]
        status = r["status"]
        typ = classify_type(r) if status == "Failed" else ""
        records.append({
            "case_id": cid, "section": sec, "title": r.get("title", c["title"]),
            "status": status, "type": typ,
            "expected": r.get("expected", ""), "actual": r.get("actual", ""),
            "reason": r.get("reason", ""), "logged": r.get("logged_to_testrail", False),
        })
    else:
        records.append({
            "case_id": cid, "section": sec, "title": c["title"],
            "status": "Not Run", "type": "",
            "expected": "", "actual": "", "reason": "", "logged": False,
        })

# ---- Tally ----
STATUS_ORDER = ["Passed", "Failed", "Retest", "Blocked", "Not Run"]
tally = OrderedDict((s, 0) for s in STATUS_ORDER)
for rec in records:
    tally[rec["status"]] += 1
grand_total = len(records)
executed = sum(tally[s] for s in ["Passed", "Failed", "Retest", "Blocked"])

# ---- Per-section table (in-scope order preserved) ----
sec_order = []
sec_stats = {}
for rec in records:
    s = rec["section"]
    if s not in sec_stats:
        sec_stats[s] = {"Total": 0, "Passed": 0, "Failed": 0, "Retest": 0, "Blocked": 0, "Not Run": 0}
        sec_order.append(s)
    sec_stats[s]["Total"] += 1
    sec_stats[s][rec["status"]] += 1

# ================= EXCEL =================
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFE699")
GREY = PatternFill("solid", fgColor="D9D9D9")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

def style_header(ws, ncols, row=1):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDER

def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_data_table(ws, headers, rows, widths, start_row, colorcol=None, colorfill=None, alt=True):
    """Generic table writer. colorcol: 1-based col to fill; colorfill: callable(rowdict/list)->fill or fixed fill."""
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, len(headers), row=start_row)
    r = start_row + 1
    for idx, row in enumerate(rows):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.alignment = WRAP_TOP
            c.border = BORDER
            if alt and idx % 2 == 1:
                c.fill = ALT_FILL
        if colorcol:
            fill = colorfill(row) if callable(colorfill) else colorfill
            if fill:
                cc = ws.cell(row=r, column=colorcol)
                cc.fill = fill
                cc.font = Font(bold=True)
        r += 1
    set_widths(ws, widths)
    return r

# ---------- Summary tab ----------
ws = wb.active
ws.title = "Summary"
ws.sheet_view.showGridLines = False
title = ws.cell(row=1, column=1, value="Custom Roles (Revised) - Test Run 312 Report")
title.font = Font(bold=True, size=16, color="1F3864")
ws.cell(row=2, column=1, value=RUN_URL).font = Font(color="0563C1", underline="single")
ws.cell(row=2, column=1).hyperlink = RUN_URL

# (a) status tally
r = 4
ws.cell(row=r, column=1, value="Status Tally").font = Font(bold=True, size=13)
r += 1
write_data_table(ws, ["Status", "Count"],
                 [[s, tally[s]] for s in STATUS_ORDER],
                 [16, 12], r, alt=True)
r += 1 + len(STATUS_ORDER)
gt = ws.cell(row=r, column=1, value="Grand Total"); gt.font = Font(bold=True); gt.border = BORDER
gc = ws.cell(row=r, column=2, value=grand_total); gc.font = Font(bold=True); gc.border = BORDER
r += 2

# (b) per-section table
ws.cell(row=r, column=1, value="Per-Section Breakdown").font = Font(bold=True, size=13)
r += 1
sec_rows = []
for s in sec_order:
    st = sec_stats[s]
    sec_rows.append([s, st["Total"], st["Passed"], st["Failed"], st["Retest"], st["Blocked"]])
next_r = write_data_table(ws, ["Section", "Total", "Passed", "Failed", "Retest", "Blocked"],
                          sec_rows, [46, 10, 10, 10, 10, 10], r, alt=True)
r = next_r + 1

# (c) narrative
ws.cell(row=r, column=1, value="Summary").font = Font(bold=True, size=13)
r += 1
notrun_note = "" if tally["Not Run"] == 0 else f" {tally['Not Run']} in-scope case was Not Run."
sentences = [
    f"{executed} of {grand_total} in-scope cases were executed ({tally['Passed']} Passed, "
    f"{tally['Failed']} Failed, {tally['Retest']} Retest, {tally['Blocked']} Blocked).{notrun_note}",
    f"Only the {tally['Passed']} Passed cases were logged to TestRail run 312 "
    f"(Failed/Retest/Blocked were reviewed locally and not written to the run).",
    "The 'Regression Suite (Minja's API file)' section (116 cases) and the 'Backend API and Security' "
    "section (38 cases) were excluded from this run per request and are not counted above.",
    f"Run URL: {RUN_URL}",
]
for s in sentences:
    c = ws.cell(row=r, column=1, value=s)
    c.alignment = WRAP_TOP
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 30
    r += 1
ws.freeze_panes = "A2"

# ---------- Passed tab ----------
ws = wb.create_sheet("Passed")
passed = [rec for rec in records if rec["status"] == "Passed"]
rows = [[rec["case_id"], rec["section"], rec["title"], rec["expected"], rec["actual"], "Yes"] for rec in passed]
write_data_table(ws, ["Case ID", "Section", "Title", "What Was Verified (expected)",
                      "What Happened (actual)", "Logged to TestRail"],
                 rows, [10, 30, 40, 50, 60, 12], 1, colorcol=1, colorfill=GREEN)
ws.freeze_panes = "A2"

# ---------- Failed tab ----------
ws = wb.create_sheet("Failed")
failed = [rec for rec in records if rec["status"] == "Failed"]
rows = [[rec["case_id"], rec["section"], rec["title"], rec["expected"], rec["actual"], rec["reason"], rec["type"]]
        for rec in failed]
def fail_color(row):
    return RED
write_data_table(ws, ["Case ID", "Section", "Title", "Expected", "What Actually Happened",
                      "Reason / Is-Stale-Case", "Type"],
                 rows, [10, 28, 38, 46, 52, 52, 26], 1, colorcol=1, colorfill=RED)
ws.freeze_panes = "A2"

# ---------- Retest tab ----------
ws = wb.create_sheet("Retest")
retest = [rec for rec in records if rec["status"] == "Retest"]
rows = [[rec["case_id"], rec["section"], rec["title"], rec["reason"], rec["actual"]] for rec in retest]
write_data_table(ws, ["Case ID", "Section", "Title", "Why Not Conclusive (reason)",
                      "What Was Partially Seen (actual)"],
                 rows, [10, 30, 40, 55, 60], 1, colorcol=1, colorfill=AMBER)
ws.freeze_panes = "A2"

# ---------- Blocked tab ----------
ws = wb.create_sheet("Blocked")
blocked = [rec for rec in records if rec["status"] == "Blocked"]
rows = [[rec["case_id"], rec["section"], rec["title"], rec["reason"], rec["actual"]] for rec in blocked]
write_data_table(ws, ["Case ID", "Section", "Title", "Why Blocked (reason)", "What's Needed To Run It"],
                 rows, [10, 30, 40, 58, 55], 1, colorcol=1, colorfill=GREY)
ws.freeze_panes = "A2"

# ---------- Not Run tab (only if any) ----------
notrun = [rec for rec in records if rec["status"] == "Not Run"]
if notrun:
    ws = wb.create_sheet("Not Run")
    rows = [[rec["case_id"], rec["section"], rec["title"]] for rec in notrun]
    write_data_table(ws, ["Case ID", "Section", "Title"], rows, [10, 36, 60], 1)
    ws.freeze_panes = "A2"

xlsx_path = os.path.join(BASE, "CustomRoles_Run312_Report.xlsx")
wb.save(xlsx_path)

# ================= CSV =================
csv_path = os.path.join(BASE, "CustomRoles_Run312_Report.csv")
with open(csv_path, "w", newline="") as f:
    w = csv.writer(f)
    w.writerow(["Case ID", "Section", "Title", "Status", "Type", "Expected", "Actual", "Reason", "Logged_to_TestRail"])
    for rec in records:
        typ = ""
        if rec["status"] == "Failed":
            typ = "stale" if rec["type"].startswith("Stale") else "real"
        w.writerow([rec["case_id"], rec["section"], rec["title"], rec["status"], typ,
                    rec["expected"], rec["actual"], rec["reason"],
                    "Yes" if rec["logged"] else "No"])

# ================= SUMMARY.md =================
md_path = os.path.join(BASE, "CustomRoles_Run312_SUMMARY.md")
real_fails = [rec for rec in failed if rec["type"].startswith("Real")]
stale_fails = [rec for rec in failed if rec["type"].startswith("Stale")]
with open(md_path, "w") as f:
    f.write("# Custom Roles (Revised) - Run 312 Summary\n\n")
    f.write(f"Run URL: {RUN_URL}\n\n")
    f.write(f"{executed} of {grand_total} in-scope cases executed. Only the {tally['Passed']} "
            f"Passed cases were logged to TestRail run 312.\n\n")
    f.write("## Status Tally\n\n")
    f.write("| Status | Count |\n|---|---|\n")
    for s in STATUS_ORDER:
        f.write(f"| {s} | {tally[s]} |\n")
    f.write(f"| **Grand Total** | **{grand_total}** |\n\n")
    f.write("## Per-Section Breakdown\n\n")
    f.write("| Section | Total | Passed | Failed | Retest | Blocked |\n|---|---|---|---|---|---|\n")
    for s in sec_order:
        st = sec_stats[s]
        f.write(f"| {s} | {st['Total']} | {st['Passed']} | {st['Failed']} | {st['Retest']} | {st['Blocked']} |\n")
    f.write("\n## Notable Failed Cases\n\n")
    f.write(f"### Real discrepancies ({len(real_fails)})\n\n")
    for rec in real_fails:
        f.write(f"- **{rec['case_id']}** ({rec['section']}): {rec['reason']}\n")
    f.write(f"\n### Stale cases (UI/spec drift) ({len(stale_fails)})\n\n")
    for rec in stale_fails:
        f.write(f"- **{rec['case_id']}** ({rec['section']}): {rec['reason']}\n")
    f.write(f"\n## Blocked ({len(blocked)}) - what's needed to run\n\n")
    for rec in blocked:
        f.write(f"- **{rec['case_id']}** ({rec['section']}): {rec['actual']}\n")
    if notrun:
        f.write(f"\n## Not Run ({len(notrun)})\n\n")
        for rec in notrun:
            f.write(f"- **{rec['case_id']}** ({rec['section']}): {rec['title']}\n")
    f.write("\n## Exclusions\n\n")
    f.write("The following sections were excluded from this run per request and are not counted above:\n\n")
    f.write("- Regression Suite (Minja's API file) - 116 cases\n")
    f.write("- Backend API and Security - 38 cases\n")

# ---- console report ----
print("XLSX:", xlsx_path)
print("CSV :", csv_path)
print("MD  :", md_path)
print("TALLY:", dict(tally), "grand_total", grand_total)
print("executed", executed)
print("logged_to_testrail true count:", sum(1 for rec in records if rec["logged"]))
print("real_fails", len(real_fails), [r["case_id"] for r in real_fails])
print("stale_fails", len(stale_fails), [r["case_id"] for r in stale_fails])
print("notrun", [(r["case_id"], r["section"], r["title"]) for r in notrun])
print("--- per section ---")
for s in sec_order:
    st = sec_stats[s]
    print(f"{s}: T{st['Total']} P{st['Passed']} F{st['Failed']} R{st['Retest']} B{st['Blocked']} N{st['Not Run']}")
