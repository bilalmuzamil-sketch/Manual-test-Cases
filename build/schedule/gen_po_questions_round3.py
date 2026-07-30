#!/usr/bin/env python3
"""Generate PO-Questions-Branko-Schedule-2026-07-31-Round-3.xlsx.

Mirrors the established Schedule/Filters PO-question workbook schema 1:1
(Standing Rule 16) — identical to gen_po_questions_techplan.py: 2 sheets —
"Questions for PO" (A1 title, blank row 2, header row 3, freeze A4, cols
# / Topic / What happens now / The question / Options / Your answer, widths
4/24/48/42/46/20) and "QA Internal Mapping" (A1 red QA-only banner, header
row 3, freeze A4, cols Q# / Affected internal case IDs (TestRail C-id) /
Source refs / What each answer resolves to, widths 4/34/52/60). The
"Your answer" column is left BLANK for the PO.

ROUND 3 supersedes the never-sent Round 2
(PO-Questions-Branko-Schedule-TechPlan_2026-07-30):
  * 3 questions WITHDRAWN — the live product write-up (Confluence v23) answers
    them outright, so re-asking would be the exact re-ask we must not make.
    Their tech-plan mismatches are DEV alignment items (Standing Rule 30).
  * 1 question ADDED (Q7) — three write-up sentences contradict rulings Branko
    has already given us, found by coverage-rederivation-2026-07-31 (F2/F3 +
    the section-9 VIN mismatch).
  * old Q2's spec-silent sub-ask survives as its own question (new Q2).

Reader-facing content is jargon-free (Rules 7/9): no case IDs, no C-ids, no
section numbers, no ticket keys, no spec/PRD/Story/Figma/TestRail/Jira/API/
HTTP/VIU terms. Verified by scan_jargon() below, which fails the build.
"""
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = "build/schedule/PO-Questions-Branko-Schedule-2026-07-31-Round-3.xlsx"

TITLE = "Schedule - Questions for Branko - 2026-07-31 (Round 3) - 7 questions"
QA_BANNER = (
    "QA-ONLY - do not send this sheet to the PO. ROUND 3, 2026-07-31. Supersedes the "
    "NEVER-SENT Round 2 (PO-Questions-Branko-Schedule-TechPlan_2026-07-30): 3 questions "
    "WITHDRAWN as answered by the live product write-up (Confluence version 23, 2026-07-30) "
    "with verbatim citations - see the WITHDRAWN rows - their build-plan mismatches re-routed "
    "to DEV (Standing Rule 30); 1 question ADDED (Q7, the three write-up sentences that "
    "contradict his own rulings, from coverage-rederivation-2026-07-31 flags F2/F3 + the "
    "section-9 VIN mismatch); old Q2's write-up-silent sub-ask (where closure days are set) "
    "survives as the new Q2. Branko's returned sheet for the sibling project came back BLANK "
    "- 0 answers - see build/filters/branko-sheet-check-2026-07-31/ANSWERS-INGEST-2026-07-31.md. "
    "TestRail C-ids from build/schedule/testrail-id-map.csv (Standing Rule 8, 165 rows): "
    "https://shopview.testrail.io/index.php?/cases/view/<id> ZERO case edits, ZERO TestRail writes."
)

# ---------------------------------------------------------------- reader-facing
QUESTIONS = [
    (
        "Shop closure days: does a multi-day job skip them?",
        "When a big job is spread across several days, the schedule plans one shift per day. "
        "Your written description says two opposite things about shop closure days (holidays, "
        "inventory days), and both sentences are still in it today. One says closure days are NOT "
        "skipped in the first release - a shift can land on a closure day, and only weekends are "
        "skipped (and only when no working hours are set for them). A later part says closure days "
        "stop the schedule from planning work on them. We have followed the first one, because it "
        "is the more recently edited of the two. The engineering build plan builds real closure-day "
        "skipping, so it follows the second.",
        "Which one is right for the first release - and please have the other sentence corrected so "
        "the description no longer says both.",
        "A) Closure days are NOT skipped in the first release (what we have followed); the other "
        "sentence should be corrected.\n"
        "B) Closure days ARE skipped after all; the first sentence should be corrected.\n"
        "C) Something else (please explain).",
    ),
    (
        "Where does someone set the shop's closure days?",
        "Your written description says closure days are set \"at the shop level\", but it never says "
        "on which screen. It does say where working hours live: a technician's own hours in the "
        "\"Edit Staff Member\" window, and the shop's hours in the \"Edit Location\" window. Closure "
        "days are the one piece with no home named anywhere, so we cannot write a test for setting "
        "them. The engineering build plan puts them on a brand-new \"Schedule Settings\" page in the "
        "Administration area, together with the shop's hours - but that is the plan's choice, not "
        "yours.",
        "Where should someone go to add or remove the shop's closure days?",
        "A) In the \"Edit Location\" window, alongside the shop's business hours.\n"
        "B) On a separate \"Schedule Settings\" page in the Administration area.\n"
        "C) Somewhere else (please say where).",
    ),
    (
        "Do meeting hours also count toward the \"OT\" tag and the hover breakdown?",
        "You confirmed that meeting hours DO use up a technician's time, so a 2-hour meeting makes "
        "the day's busy bar 2 hours fuller. Separately, the schedule shows a small \"OT\" tag on a "
        "day when one single technician goes over their own hours for that day, and a hover note "
        "that breaks the day down technician by technician. Your written description says the day's "
        "overall busy bar includes meeting hours, but it does not say whether meeting hours also "
        "feed that \"OT\" tag and that hover breakdown - it calls the overtime signal \"separate and "
        "independent\".",
        "Should meeting hours also count toward the \"OT\" tag and the per-technician hover "
        "breakdown, or only toward the day's overall busy bar?",
        "A) Yes - meeting hours count everywhere: the busy bar, the \"OT\" tag, and the hover "
        "breakdown.\n"
        "B) No - only the day's busy bar; the \"OT\" tag and the hover breakdown count job shifts "
        "only.\n"
        "C) Something else (please explain).",
    ),
    (
        "Can a meeting be created for a whole department instead of one technician?",
        "Today a meeting is placed on one named technician's row. Shops often hold a meeting (a "
        "safety briefing, a training session) for a whole department at once. Your written "
        "description does not mention this at all - a meeting only ever belongs to a single "
        "technician in it - so we do not know whether to test it. Engineering's working assumption "
        "is that a whole-department meeting would NOT use up each technician's time, but that is an "
        "engineering guess, not your decision.",
        "Should someone be able to create a meeting for a whole department in the first release - "
        "and if yes, does it use up each of those technicians' time?",
        "A) Yes - a whole-department meeting, and it uses up each of those technicians' time.\n"
        "B) Yes - a whole-department meeting, but it does not use up their time (it is just shown on "
        "their rows).\n"
        "C) No - one technician at a time only in the first release.\n"
        "D) Something else (please explain).",
    ),
    (
        "Can a meeting cover a whole day, and how should it show on the schedule?",
        "The meeting window has an \"all day\" switch, so a meeting can be set for a whole day with "
        "no start or end time. Now that meeting hours use up a technician's time, we do not know "
        "what a whole-day meeting should do - and your written description does not say. It also "
        "does not say where a whole-day meeting sits on the schedule, since it has no times to "
        "position it by. Engineering's working assumption is that a whole-day meeting is shown but "
        "uses up no time - again an engineering guess, not your decision.",
        "Should a whole-day meeting use up that technician's whole working day, or just be shown on "
        "their row without using up any time - and where should it appear?",
        "A) It uses up the whole working day, shown as a band across the top of that technician's "
        "row.\n"
        "B) It is only shown on the row and uses up no time.\n"
        "C) Something else (please explain).",
    ),
    (
        "If a user hides meetings from the view, do those hours stop counting?",
        "There is a display switch called \"Events\" that shows or hides meetings on the schedule. "
        "Meeting hours now count toward each day's busy bar. Your written description says only that "
        "this switch shows or hides the meeting blocks on the schedule - it does not say what should "
        "happen to the counting when someone switches meetings off. So the busy bars could either "
        "shrink or stay the same, and we cannot tell which is correct.",
        "When a user hides meetings from the view, should those hours also come OUT of the day's "
        "busy bars, or should the bars keep counting them and only the meetings disappear from the "
        "screen?",
        "A) The hours come out too - the busy bars recalculate as if there were no meetings.\n"
        "B) Only the meetings disappear from the screen - the busy bars keep counting their hours.\n"
        "C) Something else (please explain).",
    ),
    (
        "Three sentences in your description now disagree with decisions you already gave us "
        "(a request, not a choice)",
        "You have already settled three points for us, and we have followed your decisions. But the "
        "written description still carries the older wording in three places, so anyone reading it "
        "will think our tests are wrong. (1) You told us no money figures appear in the pop-up that "
        "opens when someone clicks a scheduled job - the description still says that pop-up shows "
        "the job lines \"with labor/total figures\". (2) You told us there is no right-click, only "
        "left-click - two sentences about who may edit the schedule still list a \"right-click "
        "context menu\" as one of the ways to create and edit. (3) You told us the vehicle "
        "identification number always shows in the hover note - the part describing the display "
        "switches still ties it to the switch being turned on. Separately, the work ticket for the "
        "job pop-up still lists a \"Reassign\" button, which you removed and told us was not "
        "wanted.",
        "Please have those three sentences corrected to match the decisions you already gave us, and "
        "the leftover \"Reassign\" wording removed from the work ticket. Nothing needs deciding - we "
        "are not changing any tests; we only want the description to stop contradicting you.",
        "A) Yes - I will correct all three (and the work ticket).\n"
        "B) One or more of them is actually right as written (please say which, so we can change our "
        "tests instead).",
    ),
]

# ------------------------------------------------------------------- QA mapping
QA_ROWS = [
    (
        "1",
        "SCH-EDGE-05 (C30089)\nSCH-SPREAD-07 (C29983)\nSCH-SPREAD-08 (C29984)\n"
        "SCH-SPREAD-11 (C38863)\nSCH-API-02 (C38873)",
        "Was Round-2 Q1 / originally NQ-1. GENUINELY OPEN, and it is a WRITE-UP "
        "SELF-CONTRADICTION - both sentences verified live in Confluence v23 this pass: "
        "section 4.5 (line 217) VERBATIM \"Shop closures and public holidays are not skipped in "
        "V1..\" versus section 12 (line 466) VERBATIM \"Shop closures (holidays, inventory days) "
        "are defined at the shop level and block the spread step from placing shifts on those "
        "days.\" Rule 32 latest-wins: the 4.5 sentence is the Confluence v22 edit, the section-12 "
        "sentence is untouched v18-era residue -> not-skipped, which is what SCH-EDGE-05 asserts. "
        "Tech plan D7 + Phase-7 E2E build real skipping. Reframed 2026-07-31 from a "
        "confirm-this-stands into a which-one-is-right, because the contradiction is his to "
        "resolve, not ours.",
        "A -> no case change; cases already correct; the section-12 sentence + the build plan are "
        "the corrections. B -> rewrite SCH-EDGE-05 (closure = skipped + struck through in preview) "
        "+ SCH-SPREAD-07 expected #3 + SCH-SPREAD-08 reason list + SCH-SPREAD-11 + SCH-API-02. "
        "Verify LIVE either way (Rule 12). Needs update_case authorization (Rule 6) + a run-357 "
        "re-check (Rule 34).",
    ),
    (
        "2",
        "no case exists yet (a closure-days CRUD case would be NEW, no C-ID yet)\n"
        "context: SCH-HRS-02 (C38847), SCH-HRS-03 (C38848), SCH-HRS-04 (C38849)",
        "NARROWED from Round-2 Q2 2026-07-31. The confirm-half of the old question is WITHDRAWN as "
        "answered by the write-up (see the WITHDRAWN W1 row). What survives is the genuinely "
        "WRITE-UP-SILENT sub-ask: section 12 says closures are \"defined at the shop level\" and "
        "NAMES NO SCREEN anywhere in the v23 body (full-text checked this pass: the only closure "
        "sentences are lines 217 and 466). Tech plan Phase 2 builds ScheduleSettings.vue in "
        "Administration + closures CRUD; the plan itself notes the design's Hours Settings file was "
        "an empty shell. Option A/B are the two real candidates, so the answer is directly "
        "authorable either way.",
        "Either answer -> author a closures-CRUD case on the named screen (NEW case, needs add_case "
        "authorization + a run-357 sync, Rule 34). B additionally means a new Administration page "
        "exists that no case covers at all. No existing case changes.",
    ),
    (
        "3",
        "SCH-CAP-03 (C30032)\nSCH-CAP-04 (C30033)",
        "Was Round-2 Q5 (internal id A1) - opened by Branko's own events->capacity answer. "
        "STILL GENUINELY OPEN, re-verified against live v23 this pass: section 4.12 line 307 puts "
        "events in the AGGREGATE verbatim \"aggregate technician-hours booked (shifts plus "
        "events)\", but line 305 calls overtime \"a separate per-technician signal, and the two are "
        "independent\", line 309 defines the OT tag per-technician (\"exceeds their own daily "
        "hours\") and line 310 the hover as \"assigned vs that tech's capacity\" - neither says "
        "whether events feed the per-technician side. Write-up silent (Rule 15). Could also be "
        "answered by dev.",
        "A -> assert event hours in the OT tag + hover breakdown on SCH-CAP-03/04. B -> assert "
        "shift-hours-only on both. Until answered, both cases assert only what IS pinned (neither "
        "outcome). Needs update_case authorization.",
    ),
    (
        "4",
        "SCH-CAP-01 (C30030)\nSCH-CAP-02 (C30031)\nSCH-CAP-03 (C30032)\nSCH-CAP-04 (C30033)\n"
        "SCH-EVT-08 (C30615)",
        "Was Round-2 Q6 (internal id A2). STILL GENUINELY OPEN, re-verified this pass: the Event "
        "entity in section 8 (line 393) is per-technician only - \"eid, name, rowKey (tech), date, "
        "startHour, endHour, allDay, color / Assigned to Technician\" - and departments appear in "
        "v23 ONLY as grid grouping (line 144), the Filter and Display toggles (lines 364/418) and "
        "grid-row derivation (section 14.4). No department-level event anywhere. The tech plan's "
        "working default is \"department-assigned events do NOT count toward capacity\" - an "
        "engineering default, not a product ruling.",
        "A -> author a department-event capacity case + extend SCH-CAP-01..04. B -> author a "
        "department-event display-only case. C -> no case; confirm at VIU that the UI offers no "
        "department option. New cases need add_case authorization + a run-357 sync.",
    ),
    (
        "5",
        "SCH-EVT-08 (C30615)\nSCH-EVT-03 (C30018)",
        "Was Round-2 Q7 (internal id A3). STILL GENUINELY OPEN, re-verified this pass: section 4.10 "
        "line 278 gives the Event modal an \"all-day toggle\" and section 8 line 393 gives Event an "
        "allDay field, but NOTHING in v23 says what an unbounded all-day event does to capacity or "
        "where it renders. Write-up silent. Tech plan working default = \"visual only\".",
        "A -> author an all-day-event capacity case (full working day consumed) + a render-position "
        "expectation. B -> assert display-only, zero capacity, on SCH-EVT-08. Needs "
        "update_case/add_case authorization.",
    ),
    (
        "6",
        "SCH-VIEW-05 (C30046)",
        "Was Round-2 Q8 (internal id A7, found in the Rule-28 sweep). STILL GENUINELY OPEN, "
        "re-verified this pass: section 9 line 428, the View-options table row, reads in full "
        "\"Events | On | Shows non-WO event blocks on the grid.\" - silent on the capacity "
        "consequence, even though section 4.12 now counts event time. Write-up silent.",
        "A -> SCH-VIEW-05 also asserts the capacity bars recalculate. B -> SCH-VIEW-05 asserts the "
        "bars are unchanged. Today it deliberately asserts only \"event blocks disappear from the "
        "grid while shifts remain\" - neither outcome. Needs update_case authorization.",
    ),
    (
        "7",
        "NO case changes - document hygiene only.\nCases already correct and STAY as they are: "
        "SCH-MODAL-04 (C30011), SCH-API-03 (C38874), SCH-PERM-02 (C30075), SCH-PERM-04 (C30077), "
        "SCH-MODAL-08 (C30015), SCH-REAS-01 (C30052)",
        "NEW 2026-07-31 - from coverage-rederivation-2026-07-31 flags F2/F3 plus the section-9 VIN "
        "mismatch already in the register. All three verified VERBATIM in the live v23 body this "
        "pass. (a) section 4.9 line 265: \"Scope summary and the scheduled line(s) with labor/total "
        "figures.\" - contradicts his 2026-07-22 Q3 no-money ruling, which SCH-MODAL-04 and "
        "SCH-API-03 follow. (b) section 14.1 lines 495 + 497: \"drag handles, drop targets, "
        "right-click context menu, resize handles..\" and \"including via right-click context menu "
        "and day-view click-to-create\" - contradicts his 2026-07-31 \"there is no right click, "
        "only left click\"; sections 4.10/7 were rewritten to left-click in v22 but 14.1 was "
        "missed; SCH-PERM-02/04 already follow left-click. (c) section 9 line 420: the VIN toggle "
        "\"Shows the VIN number as an additional line on shift blocks (day and week views) and in "
        "hover tooltips\" - but section 4.13 line 317 lists VIN in the shift tooltip "
        "unconditionally, and his Q6 answer made the toggle block-only. ALSO: Jira SV-8695 still "
        "lists a modal Reassign action that v23 deleted and he denied (\"B - No button\") - the "
        "TICKET is the stale artefact; we do not edit Jira (Rule 38 spirit).",
        "A -> ZERO case changes; the write-up and the ticket are corrected and the future-review "
        "cost disappears. B -> he names which sentence is actually right, and THAT case is "
        "re-derived whole against the corrected write-up (Rule 41) under a fresh authorization. "
        "Rule 33: his existing rulings already win, so nothing is hedged meanwhile.",
    ),
    # --- WITHDRAWN: answered by the live write-up (Confluence v23) ---
    (
        "W1 - WITHDRAWN\nwas Q2 (core)",
        "SCH-HRS-02 (C38847)\nSCH-HRS-03 (C38848)\nSCH-HRS-04 (C38849)",
        "WITHDRAWN 2026-07-31 - ANSWERED BY THE LIVE WRITE-UP. Section 4.2 line 181 VERBATIM: "
        "\"Hours settings (tech and business hours). Working hours are defined in two places: a "
        "technician's custom schedule in Edit Staff Member, and the shop's business hours in Edit "
        "Location.\" Line 183 VERBATIM names both toggles: \"Set custom hours for this technician\" "
        "/ \"Set business hours for this shop\". Asking him to re-confirm what his own document "
        "states twice over is the exact re-ask we must not make. The tech plan's ScheduleSettings "
        "page is the artefact that must change, and per Rule 30 engineering never overrules product "
        "-> a DEV item, not a PO question.",
        "Effective answer = option A (Edit Staff Member + Edit Location). SCH-HRS-02/03/04 stand "
        "as written - NO case change, NO TestRail write. The build-plan mismatch is logged for DEV. "
        "The write-up-silent sub-ask (where closure days are set) survives as reader-facing Q2. "
        "NEVER SENT to Branko in this form.",
    ),
    (
        "W2 - WITHDRAWN\nwas Q3",
        "SCH-HRS-05 (C38850)\nSCH-HRS-06 (C38851)",
        "WITHDRAWN 2026-07-31 - ANSWERED BY THE LIVE WRITE-UP. Section 4.2 line 184 VERBATIM: "
        "\"Per-day editor. One row per day (Mon-Sun): day name, with From -> To ranges on the "
        "right. Each day starts with a single range; 'Add hours' appends more to support split "
        "shifts, each removable.\" Line 185 VERBATIM: \"Overlap validation. If a day's ranges "
        "overlap, the offending range is flagged in red with an inline message ('These hours "
        "overlap. Adjust the times so they don't conflict.') and Save is disabled until it is "
        "resolved.\" Both halves of the old question are answered outright. Tech plan section 3 "
        "staff_working_hours = one start_minute/end_minute per weekday, unique (staff, workplace, "
        "day) - no split ranges -> the PLAN changes (Rule 30), not the product.",
        "Effective answer = option A (split days supported, with the overlap warning). "
        "SCH-HRS-05/06 stand as written - NO case change, NO TestRail write. Logged for DEV. "
        "NEVER SENT to Branko in this form.",
    ),
    (
        "W3 - WITHDRAWN\nwas Q4",
        "SCH-CONF-01 (C30023)\nSCH-CONF-05 (C30027)",
        "WITHDRAWN 2026-07-31 - ANSWERED BY THE LIVE WRITE-UP. This REVERSES the Round-2 sheet's "
        "own note that it was \"not settled by v23\"; the reversal is cited, not asserted. Section "
        "4.11 line 287 VERBATIM: \"The system continuously scans for scheduling issues and surfaces "
        "them in a toolbar pill:\" - and the FIRST ROW of the table that follows is VERBATIM "
        "\"Double-booked | Two different work orders overlap on the same technician at the same "
        "time.\" Line 298 VERBATIM: \"Conflicts appear as a warning icon on the affected block and "
        "are listed in a dropdown from the toolbar.\" Double-booked is therefore a listed conflict "
        "type, and listed conflict types populate the toolbar pill + dropdown = the problem counter "
        "and list. Tech plan D4 (\"FE soft warning, not a hard conflict\", BE detector covers "
        "outside-window/closure/non-working only) is what must change -> DEV (Rule 30).",
        "Effective answer = option A (double-bookings DO count in the counter and list). "
        "SCH-CONF-01 + SCH-CONF-05 already assert this - NO case change, NO TestRail write. Logged "
        "for DEV. If the QA lead disagrees with this withdrawal, the question is preserved here "
        "verbatim and can be restored unchanged. NEVER SENT to Branko in this form.",
    ),
    # --- ANSWERED EARLIER: never ask again ---
    (
        "ANSWERED\nearlier",
        "SCH-EVT-08 (C30615)\nSCH-CAP-01 (C30030)\nSCH-CAP-02 (C30031)\nSCH-CAP-03 (C30032)\n"
        "SCH-CAP-04 (C30033)\nSCH-CONF-01 (C30023)",
        "D1 - \"Do calendar events use up a technician's time?\" ANSWERED 2026-07-31, option A. "
        "Branko VERBATIM: \"A) 4.12 PRD is explicit: 'Event time is included in the utilization "
        "total alongside shifts, so meetings and training consume capacity.' A 2-hour meeting "
        "consumes 2 hours of capacity. Note the split in 4.11: events count toward capacity but are "
        "not conflict-checked. The design and the written plan already agree; this only needs "
        "confirming, not deciding.\" Corroborated word-for-word by Confluence section 4.12. Source: "
        "branko-answers-2026-07-31/answers-ingested.md Q1.",
        "NOT on the reader-facing sheet - DO NOT ASK AGAIN. HOLD LIFTED; reverses his earlier "
        "\"No\". Its write-up-silent follow-ons are reader-facing Q3/Q4/Q5/Q6 above.",
    ),
    (
        "ANSWERED\nearlier",
        "SCH-MODAL-08 (C30015)\nSCH-REAS-01 (C30052)\n(SCH-REAS-02 retired/deleted 2026-07-22)",
        "D4 - \"Should the shift pop-up have a 'Reassign' button?\" ANSWERED 2026-07-31, option B. "
        "Branko VERBATIM: \"B - No button\". Corroborated by Confluence v23 (2026-07-30) deleting "
        "\"and Reassign to another technician\" from section 4.9. Source: "
        "branko-answers-2026-07-31/answers-ingested.md Q2.",
        "NOT on the reader-facing sheet as a question - DO NOT ASK AGAIN. HOLD LIFTED; confirms our "
        "cases as written. The leftover Jira SV-8695 wording is folded into Q7's tidy request "
        "instead.",
    ),
    (
        "RE-ROUTED\nto dev",
        "SCH-PERM-09 (C30082) context; a new negative case only if the answer is \"yes, scoped\"",
        "NQ-5 - \"May a technician change other technicians' shifts?\" RE-ROUTED TO "
        "ENGINEERING/DEV 2026-07-31 and NOT re-added here. Branko on the sibling backend-scope "
        "question: \"I'm not sure if this question is for me Bilal.\" - and he is right, this is "
        "enforcement/scoping, not a product decision. Section 14.3 rules out own-only VIEWING and "
        "is SILENT on WRITING (re-verified against the live v23 body). Tech plan NFR-003 builds "
        "ManageShiftVoter own-data scoping (\"cross-tech own-data violation -> 403\").",
        "Logged as needing a DEV answer in tech-plan-2026-07-29/Questions-for-Branko-dev.md. On "
        "\"yes, scoped\" -> author the own-data write-negative (UI + backend halves). On \"no\" -> "
        "no case. A real behaviour still has NO case at all; we will not author against a guess.",
    ),
    (
        "DEV\nqueue",
        "no cases - build-plan alignment only",
        "The three items W1/W2/W3 withdrew from the PO sheet are DEV alignment items (Rule 30: "
        "engineering intent never overrules product truth, so the plan changes): (1) working hours "
        "live in Edit Staff Member + Edit Location, not a new Schedule Settings page; (2) a "
        "technician's day supports MULTIPLE working ranges, so one row per (staff, workplace, day) "
        "is the wrong data model; (3) double-booking IS a counted conflict, not a soft FE-only "
        "warning. Each is quoted verbatim against the live write-up in the W-rows above.",
        "Raise with engineering against epic SV-8685. No case changes - our cases already follow the "
        "write-up on all three. Recorded so the withdrawals are visibly re-routed, not dropped.",
    ),
    (
        "NOTE",
        "whole suite (165 active, all VIU-Pending)",
        "Branko's returned sheet for the sibling Filters project came back BLANK - all three Google "
        "export endpoints returned HTTP 200 with a valid workbook, it is provably OUR OWN sheet "
        "(104/104 content cells identical) and the answer column is empty on all 8 questions, with "
        "no cell comments. 0 answers ingested. Nothing was inferred from our own question text "
        "(Rule 12). Full forensics: "
        "build/filters/branko-sheet-check-2026-07-31/ANSWERS-INGEST-2026-07-31.md",
        "Round 3 therefore reflects NO new PO input - only our own re-derivation against the live "
        "write-up. Still no QA branch/environment (OQ-3): every case stays VIU-Pending, and "
        "write-up-pinned or design-pinned is NOT verified (Rule 12).",
    ),
]

# --------------------------------------------------------------- jargon scanner
BANNED = [
    (r"\bC\d{4,6}\b", "TestRail C-id"),
    (r"\bSCH-[A-Z]+-\d+\b", "internal case ID"),
    (r"\b(?:FD|SF|CR|FLT|SBC|SBR|PV|TU|WIP|IV)-[A-Z]+-\d+\b", "internal case ID"),
    (r"\bSV-\d+\b", "Jira ticket key"),
    (r"§", "section symbol"),
    (r"\bS\d+-R\d+\b", "spec anchor"),
    (r"\b\d+\.\d+\b(?!\s*(?:AM|PM))", "section number"),
    (r"\bPRD\b", "PRD"),
    (r"\bspec(?:s|ification)?\b", "spec"),
    (r"\bstor(?:y|ies)\b", "story"),
    (r"\bFigma\b", "Figma"),
    (r"\bTestRail\b", "TestRail"),
    (r"\bJira\b", "Jira"),
    (r"\bAPI\b", "API"),
    (r"\bHTTP\b", "HTTP"),
    (r"\bVIU\b", "VIU"),
    (r"\bendpoint\b", "endpoint"),
    (r"\b(?:20[01]|40[0-9]|500)\b(?=\s*(?:response|status|code))", "HTTP status code"),
    (r"\bRule \d+\b", "internal standing rule"),
    (r"\btest case\b", "test-case jargon"),
    (r"\bepic\b", "epic"),
    (r"\bConfluence\b", "Confluence"),
]


def scan_jargon(pairs):
    """pairs = [(where, text)]. Returns list of (where, term, match)."""
    hits = []
    for where, text in pairs:
        for pat, label in BANNED:
            flags = 0 if label == "section symbol" else re.IGNORECASE
            for m in re.finditer(pat, text, flags):
                hits.append((where, label, m.group(0)))
    return hits


def main():
    # ---- reader-facing jargon gate (title + all 4 content columns of every row)
    pairs = [("A1 title", TITLE)]
    for i, (topic, now, q, opts) in enumerate(QUESTIONS, 1):
        pairs += [
            (f"Q{i} Topic", topic),
            (f"Q{i} What happens now", now),
            (f"Q{i} The question", q),
            (f"Q{i} Options", opts),
        ]
    hits = scan_jargon(pairs)
    if hits:
        print("JARGON SCAN FAILED - reader-facing surfaces contain banned terms:")
        for where, label, tok in hits:
            print(f"  {where}: [{label}] {tok!r}")
        sys.exit(1)
    print(f"JARGON SCAN PASSED - {len(pairs)} reader-facing surfaces clean "
          f"({len(BANNED)} banned patterns each).")

    wb = Workbook()

    # ------------------------------------------------ Sheet 1: Questions for PO
    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    headers = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, (topic, now, q, opts) in enumerate(QUESTIONS, 1):
        r = 3 + i
        for c, v in enumerate([i, topic, now, q, opts, None], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # -------------------------------------------- Sheet 2: QA Internal Mapping
    qa = wb.create_sheet("QA Internal Mapping")
    qa["A1"] = QA_BANNER
    qa["A1"].font = Font(bold=True, color="C00000")
    qheaders = [
        "Q#",
        "Affected internal case IDs (TestRail C-id)",
        "Source refs",
        "What each answer resolves to",
    ]
    for c, h in enumerate(qheaders, 1):
        cell = qa.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, row in enumerate(QA_ROWS):
        r = 4 + i
        for c, v in enumerate(row, 1):
            cell = qa.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCD", [4, 34, 52, 60]):
        qa.column_dimensions[col].width = w
    qa.freeze_panes = "A4"

    wb.save(OUT)
    write_md()
    print(f"WROTE {OUT}")
    print(f"  'Questions for PO'     : {len(QUESTIONS)} questions "
          f"(rows 4-{3 + len(QUESTIONS)}), 'Your answer' column blank")
    print(f"  'QA Internal Mapping'  : {len(QA_ROWS)} rows "
          f"(7 reader-facing + 3 WITHDRAWN + 2 ANSWERED + 1 RE-ROUTED + 1 DEV + 1 NOTE)")
    print(f"  widths PO {[ws.column_dimensions[get_column_letter(i)].width for i in range(1, 7)]}"
          f"  QA {[qa.column_dimensions[get_column_letter(i)].width for i in range(1, 5)]}")


if __name__ == "__main__":
    main()
