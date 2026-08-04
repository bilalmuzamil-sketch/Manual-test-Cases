#!/usr/bin/env python3
"""Generate the Schedule v1 TestRail import file (CSV + XLSX) from the authored
per-case JSON in build/schedule/cases/.

CANONICAL FORMAT — PURE 1:1 MATCH to
testrail-import/fees-discounts-v1-testrail-import.csv,
testrail-import/simple-flow-v1-testrail-import.csv,
testrail-import/global-search-v2-testrail-import.csv and
testrail-import/filters-v1-testrail-import.csv:
  EIGHT named columns, IDENTICAL in name and order to the other project
  imports, followed by TWO trailing UNNAMED (blank) columns:
      Title, Section, Type, Priority, Preconditions, Steps, Expected Result,
      References, "", ""
  There are NO Schedule-specific columns. Standing-Rule-8 traceability
  (internal SCH- id <-> TestRail Case ID) lives in
  build/schedule/testrail-id-map.csv exactly as it does for the other
  projects; the import file itself is byte-format-identical to theirs.

CONTENT RULES enforced here (same as the other generators):
  1. VIU-word-free: internal `viu_status`, `notes`, `design_ref` are NOT emitted;
     any "(see SCH-...)" internal cross-refs are stripped.
  2. Feature-flag-free: no feature-flag phrasing survives (Schedule has none in
     the reader-facing case text; a sanity check confirms 0 occurrences).
  3. Sections = leaf area names. STANDING RULE 4: API-related cases route to an
     "API — <leaf>" section. Since 2026-07-29 the engineering tech plan supplies
     the backend contract (tech-plan-2026-07-29/TECH-PLAN-DELTAS.md §C), so the
     SCH-API-* cases (area "Schedule", api_related=true) emit into the
     "API — Schedule" section. (Before that the spec had no API contract and no
     API cases existed.)
  4. References = spec reference only (no internal SCH- ids, no VIU text).
  5/6. Preconditions/Steps/Expected kept as authored (numbered, line-broken),
     cleaned per rule 1.

Also writes build/schedule/testrail-id-map.csv (Standing Rule 8: all internal
ids, blank TestRail Case-ID column until a permitted push assigns C-ids).

Outputs (canonical location + naming, matching the other projects):
  testrail-import/schedule-v1-testrail-import.csv
  testrail-import/schedule-v1-testrail-import.xlsx
  build/schedule/testrail-id-map.csv
"""
import csv, json, glob, os, re
from collections import Counter

HERE = os.path.dirname(os.path.abspath(__file__))          # build/schedule/
BASE = os.path.dirname(HERE)                               # build/
ROOT = os.path.dirname(BASE)                               # repo root
CASES_DIR = os.path.join(HERE, "cases")
OUT_CSV = os.path.join(ROOT, "testrail-import", "schedule-v1-testrail-import.csv")
OUT_XLSX = os.path.join(ROOT, "testrail-import", "schedule-v1-testrail-import.xlsx")
OUT_IDMAP = os.path.join(HERE, "testrail-id-map.csv")

# Byte-identical (name + order) to the fees-discounts / simple-flow /
# global-search / filters imports: 8 named columns + 2 trailing UNNAMED
# (blank) columns.
HEADER = [
    "Title", "Section", "Type", "Priority",
    "Preconditions", "Steps", "Expected Result", "References",
    "", "",
]

# Deterministic, tidy section order (spec/user-journey order; API section last —
# fed by the tech-plan backend contract since 2026-07-29, see module docstring).
SECTION_ORDER = [
    "Navigation and Layout",
    "Sidebar - Mini Calendar",
    "Sidebar - Work Order List and Search",
    "Sidebar - Work Order Filters",
    "Sidebar - Line Drill-Down",
    "Drag-and-Drop Scheduling",
    "Scope Picker",
    "Shift Start Times and Unassigned Shifts",
    "Multi-Day Spread Scheduling",
    "Linked Series and Banners",
    "Shift Block Anatomy",
    "Overlap and Lane Stacking",
    "Day View Timeline",
    "Shift Detail Modal",
    "Events",
    "Conflict Detection",
    "Capacity Bars",
    "Hover Tooltips",
    "Grid Toolbar",
    "Filter and Display and View Options",
    "Reassignment and Context Menu",
    "Deletion, Series Scopes and Undo",
    "Keyboard Interactions",
    "Color System",
    "Working Hours Settings",
    "Week Export and Printing",
    "Permissions",
    "Edge Cases and Responsiveness",
    "Cross-Module and Rewrite Regression",
    "API — Schedule",
]


def clean(s):
    """Strip internal authoring markers; keep functional content (mirrors the
    other project generators)."""
    if not s:
        return s
    s = re.sub(r"\s*\(see (?:SCH|FLT|GS|SF|FD)-[A-Z0-9-]+(?:'s setup)?\)", "", s)
    s = re.sub(r"\s*\(from (?:SCH|FLT|GS|SF|FD)-[A-Z0-9-]+'s setup\)", "", s)
    s = re.sub(r"\s*\(per (?:SCH|FLT|GS|SF|FD)-[A-Z0-9-]+\)", "", s)
    # Strip remaining bare internal-id mentions defensively.
    s = re.sub(r"[,;]?\s*SCH-[A-Z]+-\d+(\.\.\d+)?", "", s)
    s = re.sub(r"feature[ -]flags?", "Schedule feature", s, flags=re.I)
    return s


def joinlines(lst):
    """Join a case field into the import's text form.

    The field may be a LIST of lines (how the cases were originally authored) or a
    single STRING already carrying its own newlines (how the 2026-08-04 re-sync from
    live TestRail now writes it).  A string MUST NOT be iterated: doing so joins it
    character by character and puts a newline between every letter, which is exactly
    how the 2026-08-04 import was corrupted across all 165 rows (and the Filters one
    across all 110).  So: split a string on its own newlines, and only then clean and
    rejoin.
    """
    if not lst:
        return ""
    if isinstance(lst, str):
        lst = lst.split("\n")
    return "\n".join(clean(x.rstrip()) for x in lst)


def load_cases():
    cases = []
    for f in sorted(glob.glob(os.path.join(CASES_DIR, "cases-*.json"))):
        cases += json.load(open(f))
    # Active cases only: retired cases (viu_status starts with 'Retired';
    # SCH-REAS-02/C30053 deleted from TestRail 2026-07-22 per user ruling) are
    # excluded from all deliverables + tally; body kept locally for the record.
    cases = [c for c in cases if not (c.get("viu_status") or "").startswith("Retired")]
    return cases


def section_for(c):
    """Leaf area name; API-related cases route to an 'API — <leaf>' section
    (STANDING RULE 4) — SCH-API-* (area 'Schedule') → 'API — Schedule'."""
    area = c["area"].strip()
    if c.get("api_related"):
        leaf = re.sub(r"^API\s*[—-]\s*", "", area).strip()
        return "API — " + leaf
    return area


def main():
    cases = load_cases()

    order = {s: i for i, s in enumerate(SECTION_ORDER)}
    cases.sort(key=lambda c: (order.get(section_for(c), 999), c["id"]))

    rows = []
    titles = []
    ids = []                 # parallel internal-id list (for dupe/empty checks only)
    idmap_rows = []
    api_sections = set()
    api_moved = 0
    for c in cases:
        title = clean(c["title"].strip())
        titles.append(title)
        ids.append(c["id"])
        section = section_for(c)
        if c.get("api_related"):
            api_sections.add(section)
            api_moved += 1
        # References = Rule-20 refs (Jira ticket + spec anchor); fall back to the
        # bare spec_ref for any case not yet backfilled.
        refs = clean((c.get("refs") or c.get("spec_ref") or "").strip())
        idmap_rows.append([c["id"], "", title, section, refs])
        rows.append([
            title,
            section,
            c.get("type", "Functional"),
            c["priority"].strip(),
            joinlines(c.get("preconditions")),
            joinlines(c.get("steps")),
            joinlines(c.get("expected")),
            refs,
            "",
            "",
        ])

    with open(OUT_CSV, "w", newline="") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(HEADER)
        w.writerows(rows)
    print("Wrote CSV:", OUT_CSV, "rows(data):", len(rows))

    with open(OUT_IDMAP, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["internal_id", "testrail_case_id", "title", "section", "refs"])
        w.writerows(idmap_rows)
    print("Wrote ID map:", OUT_IDMAP, "rows:", len(idmap_rows))

    # --- Sanity checks (must be zero / clean) ---
    dupes = [t for t, n in Counter(titles).items() if n > 1]
    print("Duplicate titles:", dupes if dupes else "NONE")
    dupids = [i for i, n in Counter(ids).items() if n > 1]
    print("Duplicate internal ids:", dupids if dupids else "NONE")
    blob = "\n".join("\t".join(str(x) for x in r) for r in rows).lower()
    print("VIU occurrences:", blob.count("viu"))
    print("'feature flag' occurrences:", blob.count("feature flag"))
    print("'flag on' occurrences:", blob.count("flag on"))
    print("'flag off' occurrences:", blob.count("flag off"))
    print("API sections created:", sorted(api_sections) if api_sections else "NONE (spec has no API contract)",
          "| API cases:", api_moved)
    empties = [iid for iid, r in zip(ids, rows)
               if not (r[4].strip() and r[5].strip() and r[6].strip())]
    print("Rows missing Preconditions/Steps/Expected:", empties if empties else "NONE")

    # --- xlsx review copy ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not available - skipped XLSX.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule V1"
    ws.append(HEADER)
    for r in rows:
        ws.append(r)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")

    widths = {"Title": 50, "Section": 40, "Type": 12, "Priority": 10,
              "Preconditions": 60, "Steps": 60, "Expected Result": 60,
              "References": 34}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    wrap = Alignment(wrap_text=True, vertical="top")
    for r in range(2, len(rows) + 2):
        for cidx in range(1, len(HEADER) + 1):
            ws.cell(row=r, column=cidx).alignment = wrap
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)
    print("Wrote XLSX:", OUT_XLSX)


if __name__ == "__main__":
    main()
