#!/usr/bin/env python3
"""Generate PO-Questions-Branko-Schedule-TechPlan_2026-07-30.xlsx (revised 2026-07-31).

Mirrors the established Schedule/Filters PO-question workbook schema 1:1
(Standing Rule 16): 2 sheets — "Questions for PO" (A1 title, blank row 2,
header row 3, freeze A4, cols # / Topic / What happens now / The question /
Options / Your answer, widths 4/24/48/42/46/20) and "QA Internal Mapping"
(A1 red QA-only banner, header row 3, freeze A4, cols Q# / Affected internal
case IDs (TestRail C-id) / Source refs / What each answer resolves to,
widths 4/34/52/60). The "Your answer" column is left BLANK for the PO.

Reader-facing content is jargon-free (Rules 7/9): no case IDs, no C-ids, no
section numbers, no ticket keys, no spec/PRD/Story/Figma/TestRail/Jira/API/
HTTP/VIU terms. Verified by scan_jargon() below, which fails the build.
"""
import re
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

OUT = "build/schedule/PO-Questions-Branko-Schedule-TechPlan_2026-07-30.xlsx"

TITLE = "Schedule - Questions for Branko - 2026-07-30 (revised 2026-07-31)"
QA_BANNER = (
    "QA-ONLY - do not send this sheet to the PO. Revised 2026-07-31 against the current "
    "product write-up (Confluence version 23, 2026-07-30). TestRail C-ids from "
    "build/schedule/testrail-id-map.csv (Standing Rule 8): "
    "https://shopview.testrail.io/index.php?/cases/view/<id>"
)

# ---------------------------------------------------------------- reader-facing
QUESTIONS = [
    (
        "Shop closure days: does a multi-day job skip them? (please confirm this still stands)",
        "When a big job is spread across several days, the schedule plans one shift per day. "
        "The current product write-up says shop closure days (holidays, inventory days) are NOT "
        "skipped in the first release - a shift can land on a closure day, and only weekends are "
        "skipped (and only when no working hours are set for them). Our tests are written that way. "
        "Two things still disagree with it: a later part of the same product write-up still says "
        "closure days stop the schedule from planning work on them, and the engineering build plan "
        "builds real closure-day skipping.",
        "Please confirm the current write-up stands - in the first release closure days are NOT "
        "skipped - and that the other sentence in the write-up and the build plan should be "
        "corrected.",
        "A) Confirmed - do not skip closure days in the first release; the other sentence and the "
        "build plan should change.\n"
        "B) No - closure days should be skipped after all.\n"
        "C) Something else (please explain).",
    ),
    (
        "Where do the shop's and technicians' working hours live? (please confirm this still stands)",
        "The current product write-up says working hours are set in two places: a technician's own "
        "hours inside the \"Edit Staff Member\" window (behind a switch called \"Set custom hours "
        "for this technician\"), and the shop's hours inside the \"Edit Location\" window (behind "
        "\"Set business hours for this shop\"). Our tests are written that way. The engineering "
        "build plan instead builds a separate \"Schedule Settings\" page in the Administration "
        "area, which would hold the shop's hours AND its closure days.",
        "Please confirm the current write-up stands - hours are set in Edit Staff Member and Edit "
        "Location - and tell us where the shop's closure days are set, since the write-up does not "
        "say.",
        "A) Confirmed - Edit Staff Member and Edit Location; the build plan should change. (Please "
        "say where closure days are set.)\n"
        "B) No - use a separate \"Schedule Settings\" page in Administration instead.\n"
        "C) Something else (please explain).",
    ),
    (
        "Can a technician have a split working day (two time ranges)? (please confirm this still stands)",
        "The current product write-up says each day starts with one working time range and an \"Add "
        "hours\" button adds more ranges, so a technician can have a split day (for example 8-12 "
        "and then 13-17), with a red warning and a blocked Save if two ranges overlap. Our tests "
        "are written that way. The engineering build plan stores only ONE working range per day for "
        "each technician - no second range at all.",
        "Please confirm the current write-up stands - a technician's day can have more than one "
        "working range in the first release - and that the build plan should change.",
        "A) Confirmed - more than one range per day, with the overlap warning; the build plan should "
        "change.\n"
        "B) No - one range per day only in the first release; \"Add hours\" is for later.\n"
        "C) Something else (please explain).",
    ),
    (
        "Does the problem counter include double-bookings?",
        "The schedule flags problems like a shift outside a technician's working hours, and a "
        "counter at the top shows how many problems there are. The product write-up also counts a "
        "technician being booked on two jobs at the same time (\"double-booked\") as one of these "
        "problems. The engineering build plan treats double-booking as a milder heads-up only - "
        "shown on the shift itself, but not counted in that counter at the top.",
        "When a technician is booked on two jobs at the same time, should that show up in the "
        "problem counter and list at the top, or only as a milder warning on the shift itself?",
        "A) Yes - double-bookings count in the problem counter and list at the top.\n"
        "B) No - double-bookings are only a milder warning on the shift, not in the counter.\n"
        "C) Something else (please explain).",
    ),
    (
        "Do meeting hours also count toward the \"OT\" tag and the hover breakdown?",
        "You confirmed that meeting hours DO use up a technician's time, so a 2-hour meeting makes "
        "the day's busy bar 2 hours fuller. Separately, the schedule shows a small \"OT\" tag on a "
        "day when one single technician goes over their own hours for that day, and a hover note "
        "that breaks the day down technician by technician. The product write-up says the day's busy "
        "bar includes meeting hours, but it does not say whether meeting hours also feed that \"OT\" "
        "tag and that hover breakdown - it calls the overtime signal \"separate and independent\".",
        "Should meeting hours also count toward the \"OT\" tag and the per-technician hover "
        "breakdown, or only toward the day's overall busy bar?",
        "A) Yes - meeting hours count everywhere: the busy bar, the \"OT\" tag, and the hover "
        "breakdown.\n"
        "B) No - only the day's busy bar; the \"OT\" tag and the hover breakdown count job shifts "
        "only.\n"
        "C) Something else (please explain).",
    ),
    (
        "Can a meeting be created for a whole department instead of one technician?",
        "Today a meeting is placed on one named technician's row. Shops often hold a meeting (a "
        "safety briefing, a training session) for a whole department at once. The product write-up "
        "does not mention this at all, so we do not know whether to test it. Engineering's working "
        "assumption is that a whole-department meeting would NOT use up each technician's time - but "
        "that is an engineering guess, not your decision.",
        "Should someone be able to create a meeting for a whole department in the first release - "
        "and if yes, does it use up each of those technicians' time?",
        "A) Yes - a whole-department meeting, and it uses up each of those technicians' time.\n"
        "B) Yes - a whole-department meeting, but it does not use up their time (it is just shown on "
        "their rows).\n"
        "C) No - one technician at a time only in the first release.\n"
        "D) Something else (please explain).",
    ),
    (
        "Can a meeting cover a whole day, and how should it show on the schedule?",
        "The meeting window has an \"all day\" switch, so a meeting can be set for a whole day with "
        "no start or end time. Now that meeting hours use up a technician's time, we do not know "
        "what a whole-day meeting should do - and the product write-up does not say. It also does "
        "not say where a whole-day meeting sits on the schedule, since it has no times to position "
        "it by. Engineering's working assumption is that a whole-day meeting is shown but uses up no "
        "time - again an engineering guess, not your decision.",
        "Should a whole-day meeting use up that technician's whole working day, or just be shown on "
        "their row without using up any time - and where should it appear?",
        "A) It uses up the whole working day, shown as a band across the top of that technician's "
        "row.\n"
        "B) It is only shown on the row and uses up no time.\n"
        "C) Something else (please explain).",
    ),
    (
        "If a user hides meetings from the view, do those hours stop counting?",
        "There is a display switch called \"Events\" that shows or hides meetings on the schedule. "
        "Meeting hours now count toward each day's busy bar. The product write-up does not say what "
        "should happen to that counting when someone switches meetings off - so the busy bars could "
        "either shrink or stay the same, and we cannot tell which is correct.",
        "When a user hides meetings from the view, should those hours also come OUT of the day's "
        "busy bars, or should the bars keep counting them and only the meetings disappear from the "
        "screen?",
        "A) The hours come out too - the busy bars recalculate as if there were no meetings.\n"
        "B) Only the meetings disappear from the screen - the busy bars keep counting their hours.\n"
        "C) Something else (please explain).",
    ),
]

# ------------------------------------------------------------------- QA mapping
QA_ROWS = [
    (
        "1",
        "SCH-EDGE-05 (C30089)\nSCH-SPREAD-07 (C29983)\nSCH-SPREAD-08 (C29984)\n"
        "SCH-SPREAD-11 (C38863)\nSCH-API-02 (C38873)",
        "Was NQ-1 (tech-plan-2026-07-29/Questions-for-Branko-dev.md). REFRAMED to a confirmation "
        "2026-07-31: spec 4.5 (Confluence v22, still standing in v23) = 'Shop closures and public "
        "holidays are not skipped in V1..' -> the spec now SIDES WITH our cases. Two "
        "counter-artefacts remain: spec 12 Edge cases still says closures 'block the spread step' "
        "(spec-internal contradiction X1, flagged in requirements.md), and tech plan D7 + Phase-7 "
        "E2E build real skipping.",
        "A (confirm) -> cases stand as written; raise the 12 sentence + the build plan as "
        "corrections. B -> rewrite SCH-EDGE-05 (closure = skipped + struck through in preview) + "
        "SCH-SPREAD-07 expected #3 + SCH-SPREAD-08 reason list. Verify LIVE either way (Rule 12).",
    ),
    (
        "2",
        "SCH-HRS-02 (C38847)\nSCH-HRS-03 (C38848)\nSCH-HRS-04 (C38849)",
        "Was NQ-3. REFRAMED to a confirmation 2026-07-31: spec 4.2 'Hours settings' block "
        "(Confluence v19, verbatim: 'a technician's custom schedule in Edit Staff Member, and the "
        "shop's business hours in Edit Location') -> the spec now SIDES WITH our cases. Tech plan "
        "Phase 2 builds ScheduleSettings.vue in Administration + closures CRUD instead; the plan "
        "itself notes the design's Hours Settings file was an empty shell. Closure-day location is "
        "spec-silent - hence the sub-ask.",
        "A (confirm) -> cases stand; the build plan should change; author a closures-CRUD case once "
        "he says where closures live. B -> re-home SCH-HRS-02/03/04 to the Schedule Settings page. "
        "(SCH-HRS-01/C38846 was merged into SCH-HRS-02 + deleted in the 2026-07-31 consolidation.)",
    ),
    (
        "3",
        "SCH-HRS-05 (C38850)\nSCH-HRS-06 (C38851)",
        "Was NQ-4. REFRAMED to a confirmation 2026-07-31: spec 4.2 verbatim \"'Add hours' appends "
        "more to support split shifts, each removable\" + the overlap-validation paragraph "
        "(Confluence v19) -> the spec now SIDES WITH our cases. Tech plan 3 staff_working_hours = "
        "one start_minute/end_minute per weekday, unique (staff, workplace, day) - no split ranges.",
        "A (confirm) -> cases stand; the build plan's data model must change. B -> retire/park "
        "SCH-HRS-05/06 (pending authorization, Rule 6). (SCH-HRS-07/C38852 was merged into "
        "SCH-HRS-06 + deleted in the 2026-07-31 consolidation.)",
    ),
    (
        "4",
        "SCH-CONF-01 (C30023)\nSCH-CONF-05 (C30027)",
        "Was NQ-2 - UNCHANGED, still a genuine open choice. Spec 4.11 conflict-type table lists "
        "'Double-booked' (unchanged in v23); tech plan D4 = double-booking is an FE soft warning, "
        "'not a hard conflict per the locked definition', BE detector covers "
        "outside-window/closure/non-working only. Not settled by v23.",
        "A -> cases stand. B -> rewrite SCH-CONF-01 expected #3/#4 (icon yes, pill no) and adjust "
        "SCH-CONF-05's count basis.",
    ),
    (
        "5",
        "SCH-CAP-03 (C30032)\nSCH-CAP-04 (C30033)",
        "NEW 2026-07-31 (internal id A1) - opened by Branko's own events->capacity answer. Spec "
        "4.12 (v19) includes event time in the aggregate total but calls overtime 'a separate "
        "per-technician signal, and the two are independent' and never says whether event hours feed "
        "the OT test or the per-tech hover breakdown. Spec silent (Rule 15). Could also be answered "
        "by dev.",
        "A -> assert event hours in the OT tag + hover breakdown on SCH-CAP-03/04. B -> assert "
        "shift-hours-only on both. Until answered, both cases assert only what IS pinned (neither "
        "outcome).",
    ),
    (
        "6",
        "SCH-CAP-01 (C30030)\nSCH-CAP-02 (C30031)\nSCH-CAP-03 (C30032)\nSCH-CAP-04 (C30033)\n"
        "SCH-EVT-08 (C30615)",
        "NEW 2026-07-31 (internal id A2). Department-level events are not in the spec at all (4.10 "
        "Events + 8 Event entity are per-technician: rowKey (tech)). The tech plan's working default "
        "is 'department-assigned events do NOT count toward capacity' - an engineering default, not "
        "a product ruling.",
        "A -> author a department-event capacity case + extend SCH-CAP-01..04. B -> author a "
        "department-event display-only case. C -> no case; confirm the UI offers no department "
        "option at VIU.",
    ),
    (
        "7",
        "SCH-EVT-08 (C30615)\nSCH-EVT-03 (C30018)",
        "NEW 2026-07-31 (internal id A3). Spec 4.10 gives the Event modal an 'all-day toggle' and 8 "
        "gives Event an allDay field, but nothing says what an unbounded all-day event does to "
        "capacity or where it renders. Spec silent. Tech plan working default = 'visual only'.",
        "A -> author an all-day-event capacity case (full working day consumed) + a render-position "
        "expectation. B -> assert display-only, zero capacity, on SCH-EVT-08.",
    ),
    (
        "8",
        "SCH-VIEW-05 (C30046)",
        "NEW 2026-07-31 (internal id A7, found in the Rule-28 sweep). Only became a question once "
        "events started consuming capacity (v19). Spec 9 View Options 'Events' toggle says only "
        "'Shows non-WO event blocks on the grid' - silent on the capacity consequence. Spec silent.",
        "A -> SCH-VIEW-05 also asserts the capacity bars recalculate. B -> SCH-VIEW-05 asserts the "
        "bars are unchanged. Today it deliberately asserts only 'event blocks disappear from the "
        "grid while shifts remain' - neither outcome.",
    ),
    # --- ANSWERED / withdrawn ---
    (
        "ANSWERED\nwas Q6",
        "SCH-EVT-08 (C30615)\nSCH-CAP-01 (C30030)\nSCH-CAP-02 (C30031)\nSCH-CAP-03 (C30032)\n"
        "SCH-CAP-04 (C30033)\nSCH-CONF-01 (C30023)",
        "D1 - 'Do calendar events use up a technician's time?' ANSWERED 2026-07-31, option A. "
        "Branko VERBATIM: \"A) 4.12 PRD is explicit: 'Event time is included in the utilization "
        "total alongside shifts, so meetings and training consume capacity.' A 2-hour meeting "
        "consumes 2 hours of capacity. Note the split in 4.11: events count toward capacity but are "
        "not conflict-checked. The design and the written plan already agree; this only needs "
        "confirming, not deciding.\" Corroborated word-for-word by Confluence v19 4.12. Source: "
        "branko-answers-2026-07-31/answers-ingested.md Q1.",
        "REMOVED from the reader-facing sheet - DO NOT ASK AGAIN. HOLD LIFTS; reverses his earlier "
        "'No'. Follow-on spec-silent items became reader-facing Q5/Q6/Q7/Q8 above.",
    ),
    (
        "ANSWERED\nwas Q7",
        "SCH-MODAL-08 (C30015)\nSCH-REAS-01 (C30052)\n(SCH-REAS-02 retired/deleted 2026-07-22)",
        "D4 - \"Should the shift pop-up have a 'Reassign' button?\" ANSWERED 2026-07-31, option B. "
        "Branko VERBATIM: \"B - No button\". Corroborated by Confluence v23 (2026-07-30) deleting "
        "'and Reassign to another technician' from 4.9. Source: "
        "branko-answers-2026-07-31/answers-ingested.md Q2.",
        "REMOVED from the reader-facing sheet - DO NOT ASK AGAIN. HOLD LIFTS; confirms our cases as "
        "written. Jira SV-8695 still lists a modal Reassign action -> SV-8695 is now the stale "
        "artefact; tell Branko/dev.",
    ),
    (
        "RE-ROUTED\nwas Q5",
        "SCH-PERM-09 (C30082) context; a new negative case would be authored only on answer 'yes'",
        "NQ-5 - 'May a technician change other technicians' shifts?' RE-ROUTED TO ENGINEERING/DEV "
        "2026-07-31. Branko on the sibling backend-scope question: \"I'm not sure if this question "
        "is for me Bilal.\" - and he is right, this is enforcement/scoping, not a product decision. "
        "Spec 14.3 rules out own-only VIEWING and is SILENT on WRITING (re-verified against the live "
        "v23 body). Tech plan NFR-003/4 builds ManageShiftVoter own-data scoping ('cross-tech "
        "own-data violation -> 403').",
        "REMOVED from the reader-facing PO sheet - it is not a product question. Logged as needing a "
        "DEV answer in tech-plan-2026-07-29/Questions-for-Branko-dev.md. On 'yes, scoped' -> author "
        "the own-data write-negative (UI + backend halves). On 'no' -> no case; confirm no 403 "
        "surprises at VIU.",
    ),
    (
        "NOTE",
        "SCH-EXP-01 (C38853)",
        "Week Export DESCOPED by Branko 2026-07-31, VERBATIM: 'No. There is nothing about this in "
        "the PRD, not in the future requirements.' Independently corroborated by a full-text scan of "
        "Confluence v23 (no export/print item in 6 Grid toolbar, 9 View options, or 15 Future "
        "considerations). NOT a question - a pending retire decision.",
        "RETIRE CANDIDATE - HELD, AWAITING EXPLICIT USER AUTHORIZATION (Rule 6). Nothing deleted. "
        "(SCH-EXP-02/C38854 was already merged away + deleted in the 2026-07-31 consolidation.) "
        "Known id-map gap: C38853 is absent from testrail-id-map.csv because its local body is "
        "flagged Retired while the TestRail case still exists - quote C38853 from the execution "
        "log/manifest until the map is reconciled.",
    ),
]

# --------------------------------------------------------------- jargon scanner
BANNED = [
    (r"\bC\d{4,6}\b", "TestRail C-id"),
    (r"\bSCH-[A-Z]+-\d+\b", "internal case ID"),
    (r"\b(?:FD|SF|CR|FLT|SBC|SBR|PV|TU|WIP|IV)-[A-Z]+-\d+\b", "internal case ID"),
    (r"\bSV-\d+\b", "Jira ticket key"),
    (r"§", "section symbol"),
    (r"\bS\d+-R\d+\b", "spec anchor"),
    (r"\b\d+\.\d+\b(?!\s*(?:AM|PM))", "section number"),
    (r"\bPRD\b", "PRD"),
    (r"\bspec(?:s|ification)?\b", "spec"),
    (r"\bstor(?:y|ies)\b", "story"),
    (r"\bFigma\b", "Figma"),
    (r"\bTestRail\b", "TestRail"),
    (r"\bJira\b", "Jira"),
    (r"\bAPI\b", "API"),
    (r"\bHTTP\b", "HTTP"),
    (r"\bVIU\b", "VIU"),
    (r"\bendpoint\b", "endpoint"),
    (r"\b(?:20[01]|40[0-9]|500)\b(?=\s*(?:response|status|code))", "HTTP status code"),
    (r"\bRule \d+\b", "internal standing rule"),
    (r"\btest case\b", "test-case jargon"),
    (r"\bepic\b", "epic"),
    (r"\bConfluence\b", "Confluence"),
]


def scan_jargon(pairs):
    """pairs = [(where, text)]. Returns list of (where, term, match)."""
    hits = []
    for where, text in pairs:
        for pat, label in BANNED:
            for m in re.finditer(pat, text, re.IGNORECASE if label != "section symbol" else 0):
                hits.append((where, label, m.group(0)))
    return hits


def main():
    # ---- reader-facing jargon gate (title + all 4 content columns of all 8 rows)
    pairs = [("A1 title", TITLE)]
    for i, (topic, now, q, opts) in enumerate(QUESTIONS, 1):
        pairs += [
            (f"Q{i} Topic", topic),
            (f"Q{i} What happens now", now),
            (f"Q{i} The question", q),
            (f"Q{i} Options", opts),
        ]
    hits = scan_jargon(pairs)
    if hits:
        print("JARGON SCAN FAILED - reader-facing surfaces contain banned terms:")
        for where, label, tok in hits:
            print(f"  {where}: [{label}] {tok!r}")
        sys.exit(1)
    print(f"JARGON SCAN PASSED - {len(pairs)} reader-facing surfaces clean "
          f"({len(BANNED)} banned patterns each).")

    wb = Workbook()

    # ------------------------------------------------ Sheet 1: Questions for PO
    ws = wb.active
    ws.title = "Questions for PO"
    ws["A1"] = TITLE
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    headers = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, (topic, now, q, opts) in enumerate(QUESTIONS, 1):
        r = 3 + i
        for c, v in enumerate([i, topic, now, q, opts, None], 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDEF", [4, 24, 48, 42, 46, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # -------------------------------------------- Sheet 2: QA Internal Mapping
    qa = wb.create_sheet("QA Internal Mapping")
    qa["A1"] = QA_BANNER
    qa["A1"].font = Font(bold=True, color="C00000")
    qheaders = [
        "Q#",
        "Affected internal case IDs (TestRail C-id)",
        "Source refs",
        "What each answer resolves to",
    ]
    for c, h in enumerate(qheaders, 1):
        cell = qa.cell(row=3, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for i, row in enumerate(QA_ROWS):
        r = 4 + i
        for c, v in enumerate(row, 1):
            cell = qa.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCD", [4, 34, 52, 60]):
        qa.column_dimensions[col].width = w
    qa.freeze_panes = "A4"

    wb.save(OUT)
    print(f"WROTE {OUT}")
    print(f"  'Questions for PO'     : {len(QUESTIONS)} questions "
          f"(rows 4-{3 + len(QUESTIONS)}), 'Your answer' column blank")
    print(f"  'QA Internal Mapping'  : {len(QA_ROWS)} rows "
          f"(8 reader-facing + 2 ANSWERED + 1 RE-ROUTED + 1 NOTE)")
    print(f"  widths PO {[ws.column_dimensions[get_column_letter(i)].width for i in range(1, 7)]}"
          f"  QA {[qa.column_dimensions[get_column_letter(i)].width for i in range(1, 5)]}")


if __name__ == "__main__":
    main()
