#!/usr/bin/env python3
"""Regenerate the Schedule Defects-for-Testers workbook (2026-08-20):
- merge the two C29945 Priority-filter rows into one DO-NOT-FILE reference row
- split into "Defects to file" (actionable) + "Reference - do not file" tabs
- regenerate .xlsx (real newlines, wrapped) and .md twin (<br>).
Source rows read from the existing single-tab workbook so verbatim text is preserved.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SRC = "Schedule_Defects-for-Testers_2026-08-20.xlsx"
OUT_XLSX = "Schedule_Defects-for-Testers_2026-08-20.xlsx"
OUT_MD = "Schedule_Defects-for-Testers_2026-08-20.md"

COLS = ["Report/Area", "Title", "Description", "Steps to Reproduce",
        "Expected behavior", "Source", "TestRail Case ID(s)",
        "TestRail Link(s)", "What needs to be done"]
WIDTHS = [24.0, 40.0, 52.0, 52.0, 44.0, 40.0, 20.0, 40.0, 52.0]

# --- read the existing 12 data rows (Defects tab, rows 2..13) verbatim ---
src = openpyxl.load_workbook(SRC)
sd = src["Defects"]
rows = []
for r in range(2, sd.max_row + 1):
    rows.append([sd.cell(row=r, column=c).value for c in range(1, 10)])
assert len(rows) == 12, f"expected 12 source rows, got {len(rows)}"

# index by the leading C-id / marker to be robust
# row 0 (#1) = C29945 DEFERRED ; row 5 (#6) = C29945/C29942 Branko ruling
r_deferred = rows[0]
r_branko = rows[5]
assert "C29945" in (r_deferred[6] or "") and "DEFERRED" in (r_deferred[8] or ""), "row1 mismatch"
assert "C29945" in (r_branko[6] or "") and "DO NOT FILE" in (r_branko[8] or ""), "row6 mismatch"

# --- build the merged C29945 row (latest authoritative word = Branko -> DO NOT FILE) ---
merged = [
    "Sidebar - Work Order Filters",
    "Priority filter group is not in the sidebar Filters panel",
    ("The sidebar Filters panel shows only Assignment and Status — there is no "
     "Priority (High / Medium / Low) group. Branko (the Product Owner) ruled on "
     "19 Aug 2026, verbatim \"Proceed without it, I'll remove that part from the "
     "PRD\", so the Priority filter is intentionally removed. The build is correct. "
     "NOT a defect — do not file."),
    ("1. Open Schedule.\n2. Click the sidebar Filters button.\n3. Look at the groups offered."),
    ("The Filters panel offers only two groups: Assignment and Status. Priority is "
     "removed per Branko's 19 Aug 2026 ruling; the PRD is to be updated."),
    ("Branko (PO) ruling 19 Aug 2026; Schedule specification v30 §5.1 (to be updated); "
     "story SV-8687; epic SV-8685."),
    "C29945, C29942",
    ("https://shopview.testrail.io/index.php?/cases/view/29945  "
     "https://shopview.testrail.io/index.php?/cases/view/29942"),
    ("Category: DO NOT FILE. Branko ruled the Priority filter out on 19 Aug 2026 — "
     "not a defect, do not file. History: this finding was earlier logged as DEFERRED "
     "(feature not found — pending Branko's confirmation on whether Priority filtering "
     "was in V1). Branko's 19 Aug 2026 ruling (\"Proceed without it, I'll remove that "
     "part from the PRD\") is the latest authoritative word (Standing Rule 32) and "
     "supersedes the deferral. The PRD (spec v30 §5.1) is to be updated to remove Priority."),
]

# --- split ---
# Defects to file = the create-new + reopen actionable rows (#2,#3,#4,#5)
defects = [rows[1], rows[2], rows[3], rows[4]]
# Reference - do not file = merged C29945 + #7,#8,#9,#10,#11,#12
reference = [merged, rows[6], rows[7], rows[8], rows[9], rows[10], rows[11]]

assert len(defects) == 4, len(defects)
assert len(reference) == 7, len(reference)
assert len(defects) + len(reference) == 11

# --- write xlsx ---
HDR_FILL = PatternFill("solid", fgColor="001F4E78")
HDR_FONT = Font(bold=True, color="00FFFFFF", size=11)
HDR_ALIGN = Alignment(wrap_text=True, vertical="top")
CELL_ALIGN = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()

# Summary tab
ws = wb.active
ws.title = "Summary"
ws.column_dimensions["A"].width = 20.0
ws.column_dimensions["B"].width = 95.0
summary = [
    ("SUMMARY", None),
    (None, None),
    ("Project", "Schedule"),
    ("Owning QA", "Ayesha"),
    ("Date", "2026-08-20"),
    ("Rows (pre-merge)", 12),
    ("Rows (post-merge, distinct)", 11),
    ("Split", "4 defects to file / 7 reference rows (do not file)"),
    ("Defects to file", "4 — the genuinely actionable Story Defects Ayesha should raise (create-new + reopen)."),
    ("Reference - do not file", "7 — informational only: not-a-defect, feature-not-built, ruled-out, and automation-held rows."),
    ("Merge note", "The two Priority-filter rows for case C29945 were merged into ONE reference row; this reduced the total from 12 to 11 distinct rows. Branko ruled the Priority filter out of scope on 19 Aug 2026, so the merged row is DO NOT FILE (its note preserves the earlier 'deferred' history)."),
    ("Purpose", "Schedule defects for the manual tester (Ayesha) to review. The 'Defects to file' tab lists what to raise as Jira Story Defects; the 'Reference - do not file' tab is informational only."),
    ("Note", "Expected behavior comes from the documented source (spec / epic / PO answer), never the build (Standing Rule 57). Jira ticket creation is on hold — this sheet is for the manual QA to read and then create the Story Defects."),
]
ws.cell(row=1, column=1, value="SUMMARY").font = Font(bold=True, size=14)
for i, (a, b) in enumerate(summary[2:], start=3):
    ca = ws.cell(row=i, column=1, value=a)
    ca.font = Font(bold=True)
    ca.alignment = Alignment(vertical="top")
    cb = ws.cell(row=i, column=2, value=b)
    cb.alignment = Alignment(wrap_text=True, vertical="top")

def write_tab(title, data):
    w = wb.create_sheet(title)
    for j, name in enumerate(COLS, start=1):
        c = w.cell(row=1, column=j, value=name)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = HDR_ALIGN
        w.column_dimensions[get_column_letter(j)].width = WIDTHS[j - 1]
    for i, row in enumerate(data, start=2):
        for j, val in enumerate(row, start=1):
            c = w.cell(row=i, column=j, value=val)
            c.alignment = CELL_ALIGN
    w.freeze_panes = "A2"

write_tab("Defects to file", defects)
write_tab("Reference - do not file", reference)
wb.save(OUT_XLSX)

# --- write md twin (<br> for newlines, double-space between links kept) ---
def md_cell(v):
    if v is None:
        return ""
    return str(v).replace("\n", "<br>").replace("|", "\\|").strip()

def md_table(data):
    lines = []
    lines.append("| # | " + " | ".join(COLS) + " |")
    lines.append("|---|" + "|".join(["---"] * len(COLS)) + "|")
    for i, row in enumerate(data, start=1):
        lines.append("| " + str(i) + " | " + " | ".join(md_cell(v) for v in row) + " |")
    return "\n".join(lines)

md = []
md.append("# Schedule — Defects for Manual Testers (2026-08-20)")
md.append("")
md.append("**Owning QA:** Ayesha. **Build:** v3.8. "
          "**Rows:** 12 pre-merge → 11 post-merge distinct "
          "(4 defects to file / 7 reference rows).")
md.append("")
md.append("Split into two tabs so actionable vs non-actionable is unambiguous. "
          "The **Defects to file** tab lists the genuinely actionable Story Defects "
          "for Ayesha to raise (create-new + reopen). The **Reference — do not file** "
          "tab is informational only (not-a-defect, feature-not-built, ruled-out, and "
          "automation-held rows), each keeping its explanation in the last column.")
md.append("")
md.append("Expected behavior comes from the documented **source** (spec / epic / PO answer), "
          "never the build (Standing Rule 57). Jira ticket creation is on hold — this sheet "
          "is for the manual QA to read and then create the Story Defects themselves.")
md.append("")
md.append("## Defects to file")
md.append("")
md.append("Genuinely actionable defects Ayesha should raise as Jira Story Defects.")
md.append("")
md.append(md_table(defects))
md.append("")
md.append("## Reference — do not file")
md.append("")
md.append("Informational only — not to be filed. Each row's last column explains why.")
md.append("")
md.append(md_table(reference))
md.append("")
md.append("## Summary")
md.append("")
md.append("- Project: Schedule")
md.append("- Owning QA: Ayesha")
md.append("- Date: 2026-08-20")
md.append("- Rows (pre-merge): 12")
md.append("- Rows (post-merge, distinct): 11")
md.append("- Split: 4 defects to file / 7 reference rows (do not file)")
md.append("- Merge note: the two Priority-filter rows for case C29945 were merged into ONE "
          "reference row (DO NOT FILE — Branko ruled it out on 19 Aug 2026; the merged row's "
          "note preserves the earlier 'deferred' history). This reduced the total from 12 to "
          "11 distinct rows.")
md.append("- Reference rows are informational only.")
md.append("")

with open(OUT_MD, "w") as f:
    f.write("\n".join(md))

print("defects to file:", len(defects))
print("reference:", len(reference))
print("total distinct:", len(defects) + len(reference), "(pre-merge 12 -> post-merge 11)")
