#!/usr/bin/env python3
"""Build SimpleFlow_Settings_QuickReference.xlsx — a glance-able settings matrix
for manual QA of the Simple Flow project. Grounded in requirements.md §3-4/§9,
viu-findings.md (confirmed toggles + baseline), and the QA Execution Guide."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = os.path.join(os.path.dirname(__file__),
                   "SimpleFlow_Settings_QuickReference.xlsx")

# ---- styling helpers -------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="257CFF")   # ShopView primary blue
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1A1A1A")
SUB_FONT = Font(italic=True, size=10, color="555555")
BOLD = Font(bold=True)
ON_FILL = PatternFill("solid", fgColor="D6EAD6")    # green-ish
OFF_FILL = PatternFill("solid", fgColor="F3D9D9")   # red-ish
NEU_FILL = PatternFill("solid", fgColor="FFF3CD")   # amber-ish
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center",
                                   horizontal="center")
        cell.border = BORDER


def val_fill(v):
    u = str(v).strip().upper()
    if u == "ON":
        return ON_FILL
    if u == "OFF":
        return OFF_FILL
    return NEU_FILL


# ---- Tab 1: Flow Matrix ----------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Flow Matrix"

ws["A1"] = "Simple Flow — Completion-Flow Settings Matrix"
ws["A1"].font = TITLE_FONT
ws["A2"] = ("Set the seven Work Order toggles as shown, Save, then run that "
            "variant's cases. Reset to the Baseline row between variants "
            "(settings are ORG-WIDE). Route: Administration > Settings > "
            "Work Orders tab > Save. Confirmed live 2026-07-06.")
ws["A2"].font = SUB_FONT
ws["A2"].alignment = WRAP
ws.merge_cells("A2:J2")
ws.row_dimensions[2].height = 42

toggles = [
    "Auto-approve Lines",
    "Require Vendor Invoice Number",
    "Require Review",
    "Require Tech Story",
    "Require Mileage",
    "Require Engine Hours",
    "Auto-pick Inventory",
]
headers = ["Flow variant"] + toggles + ["What you should see / cases"]
hdr_row = 4
for i, h in enumerate(headers, start=1):
    ws.cell(row=hdr_row, column=i, value=h)
style_header(ws, hdr_row, len(headers))

rows = [
    ("BASELINE (reset point)", "ON", "OFF", "OFF", "ON", "ON", "OFF", "ON",
     "Known-good env baseline (VIU 2026-07-06). Return to this between variants "
     "and at session end."),
    ("A. No-PO / skip completion", "ON", "OFF", "OFF", "OFF", "ON", "OFF", "ON",
     "Spec intent: Create POs OFF => no PO. NOT configurable on this env (no "
     "Create-POs toggle). Approximate with a no-parts / labor-only WO -> one "
     "confirm to Success. Cases: SF-COMP-02/03/04, SF-QB-02."),
    ("B. PO + Optional vendor invoice", "ON", "OFF", "OFF", "OFF", "ON", "OFF",
     "ON",
     "Part-bearing WO -> wizard shows Cancel / Complete Without Receiving / "
     "Receive Parts + 'N parts waiting to receive'. Cases: SF-COMP-11..17, "
     "SF-CORE-03/04/05/06."),
    ("C. PO + Required vendor invoice", "ON", "ON", "OFF", "OFF", "ON", "OFF",
     "ON",
     "Complete Work Order CTA DISABLED until all parts received (invoice # "
     "captured); NO 'Complete Without Receiving'. Cases: SF-COMP-18/19/20, "
     "SF-CORE-07, SF-VAL-05."),
    ("D. Require-Review ON (sign-off gate)", "ON", "OFF", "ON", "OFF", "ON",
     "OFF", "ON",
     "CTA relabels 'Complete & Send to Review'; WO -> Review (amber) -> Mark "
     "Reviewed (captures VIN) -> sign-off. Require Vendor Invoice is orthogonal "
     "(can be ON too). Cases: SF-REV-*, SF-PERM-04/07/08, SF-VAL-07."),
    ("Add-on: Tech-story gate", "-", "-", "-", "ON", "-", "-", "-",
     "Turn Require Tech Story ON on top of any variant -> tech-story modal opens "
     "before completion. Cases: SF-TECH-*."),
    ("Add-on: Auto-pick OFF", "-", "-", "-", "-", "-", "-", "OFF",
     "Turn Auto-pick Inventory OFF -> you must pick parts inside the completion "
     "modal. Case: SF-COMP-08."),
]

r = hdr_row + 1
for row in rows:
    for i, v in enumerate(row, start=1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.border = BORDER
        cell.alignment = WRAP
        if 2 <= i <= 8:            # toggle columns
            cell.alignment = CENTER
            if str(v).strip() != "-":
                cell.fill = val_fill(v)
                cell.font = BOLD
        if i == 1:
            cell.font = BOLD
    ws.row_dimensions[r].height = 58
    r += 1

ws.column_dimensions["A"].width = 30
for col in "BCDEFGH":
    ws.column_dimensions[col].width = 12
ws.column_dimensions["I"].width = 60
ws.freeze_panes = "B5"

# legend
r += 1
ws.cell(row=r, column=1, value="Legend:").font = BOLD
ws.cell(row=r, column=2, value="ON").fill = ON_FILL
ws.cell(row=r, column=2).alignment = CENTER
ws.cell(row=r, column=3, value="OFF").fill = OFF_FILL
ws.cell(row=r, column=3).alignment = CENTER
ws.cell(row=r, column=4, value="- = leave as baseline / orthogonal")


# ---- Tab 2: Toggle Reference ----------------------------------------------
ws2 = wb.create_sheet("Toggle Reference")
ws2["A1"] = "Work Order Settings — Toggle Reference"
ws2["A1"].font = TITLE_FONT
ws2["A2"] = ("The 7 toggles confirmed present on Administration > Settings > "
             "Work Orders tab (VIU 2026-07-06), in display order.")
ws2["A2"].font = SUB_FONT
ws2.merge_cells("A2:D2")

h = ["#", "Toggle", "Baseline", "What it does"]
for i, x in enumerate(h, start=1):
    ws2.cell(row=4, column=i, value=x)
style_header(ws2, 4, len(h))

tref = [
    (1, "Auto-approve Lines", "ON",
     "ON: each line approved the moment it is added. OFF: new lines land in "
     "'Needs Approval' with Approve/Decline. All lines must be approved to "
     "complete or Send to Review."),
    (2, "Require Vendor Invoice Number", "OFF",
     "ON: parts must be received and an invoice # captured before completing "
     "(Required flow C). OFF: complete now, receive later (Optional flow B)."),
    (3, "Require Review Before Completion", "OFF",
     "ON: completing sends the WO to a review / sign-off gate (Story 16) before "
     "it can be invoiced (flow D)."),
    (4, "Require Tech Story", "ON",
     "ON: every line needs a tech story before completion; the tech-story gate "
     "modal (Story 17) opens first."),
    (5, "Require Mileage", "ON",
     "ON: mileage is a required field at completion."),
    (6, "Require Engine Hours", "OFF",
     "ON: engine hours required at completion."),
    (7, "Automatically Pick Inventory Parts", "ON",
     "ON: in-stock parts auto-picked. OFF: pick parts inside the completion "
     "modal (SF-COMP-08)."),
]
r = 5
for row in tref:
    for i, v in enumerate(row, start=1):
        cell = ws2.cell(row=r, column=i, value=v)
        cell.border = BORDER
        cell.alignment = WRAP
        if i == 3 and str(v).strip() != "-":
            cell.fill = val_fill(v)
            cell.alignment = CENTER
            cell.font = BOLD
        if i == 2:
            cell.font = BOLD
    ws2.row_dimensions[r].height = 48
    r += 1

# deviation note rows
r += 1
ws2.cell(row=r, column=1,
         value="KNOWN DEVIATIONS (do not re-raise; see VIU findings):").font = BOLD
ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
r += 1
devs = [
    "No 'Create Purchase Orders' toggle / no createPurchaseOrders field -> POs "
    "always-on; pure No-PO/skip config not settable (SF-SET-03, SF-COMP-06, "
    "SF-QB-02).",
    "Save Settings button always enabled (no dirty-state gating) (SF-SET-13).",
    "Mark Reviewed dialog has no optional note field input_review_note "
    "(SF-REV-10).",
    "Review sign-off jumps straight to Complete; no distinct 'Reviewed' state "
    "observed (SF-REV-08, SF-REV-11).",
]
for d in devs:
    ws2.cell(row=r, column=1, value="- " + d).alignment = WRAP
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws2.row_dimensions[r].height = 30
    r += 1

ws2.column_dimensions["A"].width = 6
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 70


# ---- Tab 3: Where & How ----------------------------------------------------
ws3 = wb.create_sheet("Where & How")
ws3["A1"] = "How to set the Work Order settings"
ws3["A1"].font = TITLE_FONT
notes = [
    "",
    "Environment: https://sv7301.qa.shopview.com  (API host, reference only: "
    "https://sv7301api.qa.shopview.com)",
    "Log in with your OWN Admin QA credential. NO credentials in this file.",
    "",
    "Route to set toggles (confirmed live):",
    "  1. Sign in as Admin.",
    "  2. Go to /administration/settings.",
    "  3. Click the 'Work Orders' tab.",
    "  4. Toggle the settings per the Flow Matrix tab.",
    "  5. Click 'Save Settings'.",
    "  6. Reload the page to confirm the change persisted.",
    "",
    "Settings are ORG-WIDE and apply to FUTURE completions only (never "
    "retroactive). Reset to the Baseline row between flow variants and at the "
    "end of your session so you do not disrupt other testers/automation sharing "
    "this org.",
    "",
    "Backend check (optional, needs API access): GET "
    "/api/organizations/settings reads the saved settings object; save is POST "
    "/api/organizations/settings/change. Used by SF-SET-12, SF-PERM-06, SF-QB-*.",
    "",
    "See SimpleFlow_QA_Execution_Guide.md for full preconditions, accounts/"
    "roles, test-data recipes, per-area map, and cleanup.",
]
r = 2
for n in notes:
    c = ws3.cell(row=r, column=1, value=n)
    c.alignment = WRAP
    if n and not n.startswith(" ") and n.endswith(":"):
        c.font = BOLD
    r += 1
ws3.column_dimensions["A"].width = 100

wb.save(OUT)
print("wrote", OUT)
