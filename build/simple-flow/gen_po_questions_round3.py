#!/usr/bin/env python3
"""Generate the ROUND-3 PO question sheet for Milos (Simple Flow PO).

Outputs (regenerable — run from the repo root):
  - build/simple-flow/PO-Questions-Round3.xlsx
  - build/simple-flow/PO-Questions-Round3.md

Sheets / sections:
  - "Questions for PO"    : reader-facing, VERY SIMPLE layman language ONLY.
                            NO case IDs, SF-/BUG codes, API/HTTP terms, or jargon.
                            Columns: # | Topic | What happens now | The question |
                            Options | Your answer (blank).
  - "QA Internal Mapping" : QA-only. Per question, the gated MILOS-ANSWER case IDs
                            with TestRail Case ID C##### + clickable link (standing
                            rule 8; source testrail-id-map.csv), spec refs, and what
                            each answer option resolves to. Includes a ledger of all
                            15 gated MILOS-ANSWER cases and the SF-QB-09 dev-confirm
                            item (flagged for developers, NOT Milos).

Rounds 1 & 2 were answered by Milos (see milos-answers-mapping.md and
milos-round2-mapping.md). These 6 are the remaining genuine PRODUCT DECISIONS still
open after Round 2 + the RE-VIU BATCH 7/8 pass (PROJECT-STATE §5.F). Bugs/defects
are NOT included here (standing rule 7) — they go in SimpleFlow_Bug-Drafts.*.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/simple-flow/PO-Questions-Round3.xlsx"
MD_OUT = "build/simple-flow/PO-Questions-Round3.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no codes, no tech terms)
# ---------------------------------------------------------------------------
questions = [
    {
        "topic": "Should a work order have to be reviewed before it can be billed?",
        "now": ("When the \"needs a review before it is finished\" option is turned "
                "on, the person doing the review signs it off and the work order is "
                "finished right away. Today there is nothing that stops someone from "
                "creating the customer's bill (invoice) before that review sign-off "
                "has happened. (This is the same idea we asked about last time - it "
                "was not clear then, so here it is in plainer words.)"),
        "q": ("When review is required, should the app stop anyone from billing the "
              "customer until the work order has actually been reviewed and signed "
              "off?"),
        "opts": ("A) Yes - block billing until the review is done. No invoice can be "
                 "created until someone has reviewed and signed off the work order.\n"
                 "B) No - billing can happen any time; the review is just an extra "
                 "step and does not hold up billing."),
    },
    {
        "topic": "For a brand-new company, should \"require a review\" start ON or OFF?",
        "now": ("There is a setting that makes every work order go through a review "
                "before it is finished. When a brand-new company (organization) "
                "starts using the app, we need to know what that setting should be "
                "set to out of the box, before anyone changes it."),
        "q": ("When a new company first starts using the app, should \"require a "
              "review before finishing\" already be turned ON, or start OFF so they "
              "can turn it on if they want it?"),
        "opts": ("A) Start ON - new companies get review-required by default; they can "
                 "turn it off if they don't want it.\n"
                 "B) Start OFF - new companies get no review by default; they can turn "
                 "it on if they want it."),
    },
    {
        "topic": "What should the \"are you sure you want to close this?\" pop-up say and do?",
        "now": ("When someone tries to close a window that they were filling in, a "
                "small pop-up can appear to check they meant to. The exact wording "
                "and the two buttons on that pop-up have not been designed yet, so we "
                "need to know how it should behave: which button just closes the "
                "little pop-up and keeps their work, and which one backs all the way "
                "out to the previous screen."),
        "q": ("How should the confirmation pop-up work - what should the two buttons "
              "say, and what should each one do?"),
        "opts": ("A) One button = \"stay here\" (closes only the little pop-up and "
                 "keeps everything they typed); the other = \"leave\" (backs out to "
                 "the previous screen). Nothing they typed is ever thrown away just by "
                 "using this pop-up.\n"
                 "B) Something different - please describe the wording and what each "
                 "button should do."),
    },
    {
        "topic": "Where should parts that are missing a supplier sit in the receiving list?",
        "now": ("On the screens where parts are received, parts that don't yet have a "
                "supplier (vendor) chosen are grouped together on their own, because "
                "someone has to pick a supplier for them before they can be received. "
                "On one receiving screen this \"missing supplier\" group already sits "
                "at the top; on the newer receiving screen it currently sits at the "
                "bottom. We want them to be consistent."),
        "q": ("Should the \"missing supplier\" group always sit at the TOP of the "
              "list (so people deal with it first) on every receiving screen?"),
        "opts": ("A) Yes - always put the \"missing supplier\" group at the top on "
                 "every receiving screen, so it's the first thing people see and act "
                 "on.\n"
                 "B) No - leave it at the bottom on the newer screen (only the older "
                 "screen leads with it)."),
    },
    {
        "topic": "Can a work order be finished with a $0.00 selling price on a part?",
        "now": ("When a part is added, it has a selling price (what the customer is "
                "charged for it). The newer design shows a work order being finished "
                "even when a part's selling price is still $0.00 - it shows a small "
                "note that says \"$0.00 sell price, no action needed to continue\" "
                "and lets the person finish. An earlier write-up said a selling price "
                "must be filled in before saving. We need to know which one is right."),
        "q": ("Should a work order be allowed to be finished when a part still has a "
              "$0.00 selling price, or must every part have a real selling price "
              "first?"),
        "opts": ("A) Allow it - finishing with a $0.00 selling price is fine; just "
                 "show the little note and let the person continue (as the newer "
                 "design shows).\n"
                 "B) Require a price - a part must have a real selling price before "
                 "the work order can be finished (or before it can be saved)."),
    },
    {
        "topic": "Should adding a part with no supplier need the \"can see money figures\" permission?",
        "now": ("Some staff have permission to see money figures (costs and selling "
                "prices) and some do not. A person can add a part that has no supplier "
                "yet. We used to assume this required the \"can see money figures\" "
                "permission because a selling price had to be typed in - but you've "
                "since told us a selling price is NOT required here, so that reason no "
                "longer holds. We need to know whether that permission should still be "
                "required to add a part with no supplier."),
        "q": ("Should adding a part that has no supplier yet still require the \"can "
              "see money figures\" permission, or should anyone who can edit the work "
              "order be able to add one?"),
        "opts": ("A) Yes, still require it - only people who can see money figures can "
                 "add a part with no supplier.\n"
                 "B) No - anyone who can edit the work order can add a part with no "
                 "supplier (money figures aren't involved anymore)."),
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (rule 8: TestRail C##### + clickable link per case).
# Each entry: (q_no, [(sf_id, testrail_id), ...], refs, resolves_to)
# ---------------------------------------------------------------------------
internal_map = [
    (1,
     [("SF-REV-08", 29393), ("SF-REV-11", 29396), ("SF-REV-10", 29395)],
     "requirements.md Story 16 R5/R8 (distinct Reviewed state; invoicing blocked "
     "until reviewed). Re-ask of Round-1 Q8 (\"not sure what this means?\", "
     "milos-answers-mapping.md). SF-REV-10 = related review-dialog case.",
     "A -> SF-REV-08 expected keeps a distinct Reviewed holding state and SF-REV-11 "
     "expected keeps \"invoicing blocked until reviewed\" (both currently gated on "
     "this ruling). B -> rewrite both: sign-off completes directly and invoicing is "
     "NOT gated on review. NOTE: SF-REV-10 (review-note) already RESOLVED in Round-2 "
     "Q1 (note descoped, VIN-only) — listed for completeness, not re-asked."),
    (2,
     [("SF-REV-15", 29400)],
     "requirements.md Story 16 R Open (Require-Review default). Round-1 Q1 answer "
     "was \"ON for all orgs\"; this re-confirms the NEW-ORG out-of-box default "
     "specifically (and whether the live default matches).",
     "A -> SF-REV-15 expected = default ON for new orgs (if live default != ON that "
     "becomes a separate bug to verify). B -> SF-REV-15 expected = default OFF for "
     "new orgs."),
    (3,
     [("SF-UX-04", 29404)],
     "requirements.md Story 15 R4 (close-confirmation modal). Round-1 Q10 gave the "
     "Close/Cancel behavior but the design is still \"to be added\"; this confirms "
     "final wording + button behavior.",
     "A -> SF-UX-04 expected = Close closes only the modal and keeps entered data "
     "(stays on the WO); Cancel closes the modal and returns to the previous screen; "
     "nothing discarded. B -> capture Milos's alternate wording/behavior and rewrite "
     "SF-UX-04 accordingly."),
    (4,
     [("SF-RCV-05", 29373), ("SF-RCV-07", 29375)],
     "requirements.md Story 12 R1/R3 (vendor-missing group ordering). Round-1 Q11 "
     "recommended top/leads; RE-VIU BATCH 7 OBS-2: the Bulk Receive page renders the "
     "vendor-missing group LAST — should it also lead? Wording-only.",
     "A -> SF-RCV-05 expected #3 changed from \"at the bottom\" to \"leads (top)\" "
     "on every receive screen (incl. Bulk Receive); SF-RCV-07 already says \"leads "
     "(top)\" — confirmed, no change. B -> SF-RCV-05 stays \"at the bottom\" on the "
     "newer (Bulk Receive) screen; only the legacy Accept-Delivery screen leads."),
    (5,
     [("SF-VPART-01", 29331), ("SF-VPART-02", 29332), ("SF-VAL-09", 29423),
      ("SF-QB-06", 29431)],
     "PROJECT-STATE §5.F.2: design screenshot \"$0.00 sell price, no action needed "
     "to continue\" vs spec S5-R1 \"sell mandatory at save\". (Milos R2 Q4 already "
     "ruled sell NOT enforced on the vendorless part-request form; THIS is the "
     "remaining completion/receive-surface $0-sell tension.) Not one of the 15 gated "
     "MILOS cases — these are the genuinely-affected cases.",
     "A -> $0 sell allowed at completion; SF-VAL-09 / SF-QB-06 / SF-VPART-01/02 "
     "expecteds align to \"$0 sell permitted, note shown\" (consistent with R2 Q4). "
     "B -> sell mandatory before finish/save; add a completion-time sell-required "
     "gate to the affected expecteds."),
    (6,
     [("SF-VPART-02", 29332), ("SF-PERM-09", 29413)],
     "PROJECT-STATE §5.F.3 / milos-round2-mapping.md Q4 follow-up: the See-Financial-"
     "Data gate on vendorless part-add was premised on a mandatory sell price (now "
     "overturned by R2 Q4). Whether a permission gate still applies is open. Not one "
     "of the 15 gated MILOS cases — these are the affected cases.",
     "A -> keep the See-Financial-Data negative on SF-PERM-09 and SF-VPART-02 "
     "(permission still gates vendorless part-add). B -> drop the See-Financial-Data "
     "gate from both — any WO-edit role can add a vendorless part."),
]

# The full 15 gated MILOS-ANSWER cases (blockers tracker). The 8 below are gated
# cases already RESOLVED in Round 1/2 (or moved to a bug) and are NOT re-asked in
# Round 3 — listed for the complete ledger. (The other 7 appear under Q1-Q4 above:
# SF-REV-08, SF-REV-11, SF-REV-10, SF-REV-15, SF-UX-04, SF-RCV-05, SF-RCV-07.)
resolved_ledger = [
    ("SF-SET-03", 29277, "Story 1 R2 (Create POs toggle)",
     "RESOLVED Round-1 Q5 — Create-POs toggle descoped (POs always on). Not re-asked."),
    ("SF-SET-08", 29282, "Story 1 / §4 first-use defaults",
     "MOVED TO A BUG — wrong first-use defaults (see SimpleFlow_Bug-Drafts, bug T5). "
     "Not a PO question."),
    ("SF-SET-13", 29287, "Story 1 (Save dirty-state)",
     "RESOLVED Round-1 Q6 — Save-always-enabled accepted (nice-to-have). Not re-asked."),
    ("SF-COMP-06", 29295, "Story 2 (Create POs OFF completion)",
     "RETIRED Round-1 Q5 — Create-POs-OFF scenario no longer exists. Not re-asked."),
    ("SF-COMP-07", 29296, "Story 2 / §5 invariant 1 (inventory decrement)",
     "CONFIRMED Round-2 Q3 — in-stock parts decrement + Part History on completion. "
     "VIU-Pending on a live decrement drive; not a PO question."),
    ("SF-TECH-08", 29330, "Story 17 vs S15-R2 (tech-story placement)",
     "RESOLVED Round-2 Q2 — Story 17 authoritative (inline + gate modal; complete one "
     "or many at once). Not re-asked."),
    ("SF-QB-01", 29426, "§5 invariant 1 (inventory decrement / Part History)",
     "CONFIRMED Round-2 Q3 — same as SF-COMP-07. VIU-Pending on a live drive; not a "
     "PO question."),
    ("SF-QB-02", 29427, "§4/§5 (Create POs OFF QuickBooks integrity)",
     "RETIRED Round-1 Q5 — Create-POs-OFF scenario no longer exists. Not re-asked."),
]

# Dev-confirm item — QA-internal tab ONLY, flagged for developers (NOT Milos).
# SF-QB-09 is the one Open-Question case not yet imported to TestRail (no C-ID).
dev_item = (
    "SF-QB-09", None,
    "requirements.md §5 (shared order/status logic must not affect Part Sales).",
    "FOR DEVELOPERS, NOT MILOS. Open dev-confirm: verify the Part Sales flow is "
    "unaffected by the shared order/status logic Simple Flow introduces. Product "
    "decision not required — a code/behavior confirmation from the dev team. "
    "(Not yet imported to TestRail — no Case ID; the sole Open-Question case.)")

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
wb = Workbook()

ws = wb.active
ws.title = "Questions for PO"
ws.column_dimensions["A"].width = 5
for col, w in zip("BCDEF", [40, 56, 46, 60, 30]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Simple Mode - Round 3: A Few Quick Questions for You"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
intro = ("Hi Milos! Thanks for your answers to the earlier rounds - they are all "
         "actioned. While finishing our checks we found a few more spots where we "
         "need your call on how it should work. No wrong answers - for each one, "
         "pick an option (or write your own) in the \"Your answer\" box.")
ws["A2"] = intro
ws["A2"].alignment = WRAP
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 45

headers = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, item in enumerate(questions, start=1):
    row = HDR_ROW + i
    vals = [i, item["topic"], item["now"], item["q"], item["opts"], ""]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 165

# --- QA Internal Mapping sheet ---
wi = wb.create_sheet("QA Internal Mapping")
wi["A1"] = ("INTERNAL - for QA only. Do NOT share this tab (or any IDs/codes on it) "
            "with the PO.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:F1")

ihead = ["Q#", "Case (internal ID)", "TestRail Case ID", "TestRail link",
         "Refs", "What each answer option resolves to"]
iwid = [5, 16, 15, 46, 50, 66]
IH_ROW = 3
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for qno, cases, refs, resolves in internal_map:
    first = r
    for sf_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        rowvals = [qno, sf_id, f"C{tr_id}", url, refs, resolves]
        for c, v in enumerate(rowvals, start=1):
            cell = wi.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_CENTER if c in (1, 2, 3) else WRAP
            cell.border = BORDER
        link_cell = wi.cell(row=r, column=4)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT
        wi.row_dimensions[r].height = 130
        r += 1
    last = r - 1
    if last > first:
        for col in (1, 5, 6):
            wi.merge_cells(start_row=first, start_column=col,
                           end_row=last, end_column=col)

# --- Ledger: the remaining gated MILOS-ANSWER cases (resolved / not re-asked) ---
r += 1
ledger_title = wi.cell(row=r, column=1, value=(
    "Gated MILOS-ANSWER cases already resolved in Round 1/2 (or moved to a bug) - "
    "NOT re-asked in Round 3. Together with the 7 cases under Q1-Q4 above "
    "(SF-REV-08, SF-REV-11, SF-REV-10, SF-REV-15, SF-UX-04, SF-RCV-05, SF-RCV-07) "
    "these complete the 15 gated MILOS-ANSWER cases from the blockers tracker."))
ledger_title.font = Font(bold=True, color="1F4E79")
ledger_title.alignment = WRAP
wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
wi.row_dimensions[r].height = 45
r += 1
for c, h in enumerate(["", "Case (internal ID)", "TestRail Case ID", "TestRail link",
                       "Refs", "Status / resolution"], start=1):
    cell = wi.cell(row=r, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
r += 1
for sf_id, tr_id, refs, status in resolved_ledger:
    url = TR_LINK.format(tr_id)
    rowvals = ["", sf_id, f"C{tr_id}", url, refs, status]
    for c, v in enumerate(rowvals, start=1):
        cell = wi.cell(row=r, column=c, value=v)
        cell.alignment = WRAP_CENTER if c in (2, 3) else WRAP
        cell.border = BORDER
    lc = wi.cell(row=r, column=4)
    lc.hyperlink = url
    lc.font = LINK_FONT
    wi.row_dimensions[r].height = 60
    r += 1

# --- Dev-confirm item (developers, NOT Milos) ---
r += 1
dev_title = wi.cell(row=r, column=1, value=(
    "Dev-confirm item - FOR DEVELOPERS, NOT MILOS (no product decision needed):"))
dev_title.font = Font(bold=True, color="C00000")
wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
r += 1
sf_id, tr_id, refs, status = dev_item
tr_disp = f"C{tr_id}" if tr_id else "(not in TestRail)"
url = TR_LINK.format(tr_id) if tr_id else ""
rowvals = ["DEV", sf_id, tr_disp, url or "n/a", refs, status]
for c, v in enumerate(rowvals, start=1):
    cell = wi.cell(row=r, column=c, value=v)
    cell.alignment = WRAP_CENTER if c in (1, 2, 3) else WRAP
    cell.border = BORDER
if url:
    dc = wi.cell(row=r, column=4)
    dc.hyperlink = url
    dc.font = LINK_FONT
wi.row_dimensions[r].height = 60
r += 1

note_row = r + 1
wi.cell(row=note_row, column=1, value=(
    "Notes: Round-3 questions raised after Milos Round-2 answers + RE-VIU BATCH 7/8 "
    "(PROJECT-STATE.md §5.F). TestRail IDs sourced from testrail-id-map.csv "
    "(standing rule 8). Rounds 1 & 2 were answered by Milos - see "
    "milos-answers-mapping.md and milos-round2-mapping.md. Bugs/defects stay OUT of "
    "the PO-facing tab (standing rule 7) and are delivered separately in "
    "SimpleFlow_Bug-Drafts.xlsx; these 6 items are genuine product decisions.")
    ).alignment = WRAP
wi.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
wi.row_dimensions[note_row].height = 75

wb.save(XLSX_OUT)

# ---------------------------------------------------------------------------
# Markdown mirror
# ---------------------------------------------------------------------------
md = []
md.append("# Simple Mode — Round 3: A Few Quick Questions for You")
md.append("")
md.append("Hi Milos! Thanks for your answers to the earlier rounds — they are all")
md.append("actioned. While finishing our checks we found a few more spots where we")
md.append("need your call on how it should work.")
md.append("")
md.append("There are **no wrong answers**. For each item, pick an option (or write")
md.append("your own) on the **\"Your answer\"** line. It should take just a few")
md.append("minutes.")
md.append("")
md.append("---")
for i, item in enumerate(questions, start=1):
    md.append("")
    md.append(f"## {i}. {item['topic']}")
    md.append("")
    md.append("**What happens now**")
    md.append(item["now"])
    md.append("")
    md.append("**The question**")
    md.append(item["q"])
    md.append("")
    md.append("**Options**")
    for line in item["opts"].split("\n"):
        md.append(f"- {line}")
    md.append("")
    md.append("**Your answer:** ______________________________________________")
    md.append("")
    md.append("---")
md.append("")
md.append("## Thank you!")
md.append("")
md.append("That's everything for this round. Your answers will help us finish this")
md.append("feature the way you want it. Feel free to add any notes alongside your")
md.append("choices.")
md.append("")
md.append("---")
md.append("---")
md.append("")
md.append("## Internal — QA-only mapping (NOT for the PO)")
md.append("")
md.append("This section links each plain-English question above to its gated")
md.append("MILOS-ANSWER cases, TestRail cases, refs and what each answer resolves to.")
md.append("**Do not include this section (or any IDs/codes in it) in the PO-facing")
md.append("copy or the \"Questions for PO\" tab.**")
md.append("")
for qno, cases, refs, resolves in internal_map:
    md.append(f"### Q{qno}")
    md.append("")
    md.append("- **TestRail cases:**")
    for sf_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        md.append(f"  - {sf_id} — [C{tr_id}]({url})")
    md.append(f"- **Refs:** {refs}")
    md.append(f"- **Resolves to:** {resolves}")
    md.append("")
md.append("### Gated MILOS-ANSWER cases already resolved in Round 1/2 (or moved to a bug) — NOT re-asked")
md.append("")
md.append("Together with the 7 cases under Q1–Q4 above (SF-REV-08, SF-REV-11,")
md.append("SF-REV-10, SF-REV-15, SF-UX-04, SF-RCV-05, SF-RCV-07) these complete the")
md.append("15 gated MILOS-ANSWER cases from the blockers tracker.")
md.append("")
for sf_id, tr_id, refs, status in resolved_ledger:
    url = TR_LINK.format(tr_id)
    md.append(f"- {sf_id} — [C{tr_id}]({url}) — *{refs}* — {status}")
md.append("")
md.append("### Dev-confirm item — FOR DEVELOPERS, NOT MILOS")
md.append("")
sf_id, tr_id, refs, status = dev_item
if tr_id:
    md.append(f"- {sf_id} — [C{tr_id}]({TR_LINK.format(tr_id)}) — *{refs}* — {status}")
else:
    md.append(f"- {sf_id} — (not yet in TestRail) — *{refs}* — {status}")
md.append("")
md.append("**Notes:** Round-3 questions raised after Milos Round-2 answers + RE-VIU")
md.append("BATCH 7/8 (`PROJECT-STATE.md` §5.F). TestRail IDs sourced from")
md.append("`testrail-id-map.csv` (standing rule 8). Rounds 1 & 2 were answered by")
md.append("Milos — see `milos-answers-mapping.md` and `milos-round2-mapping.md`.")
md.append("Bugs/defects stay OUT of the PO-facing content (standing rule 7) and are")
md.append("delivered separately in `SimpleFlow_Bug-Drafts.xlsx`; these 6 items are")
md.append("genuine product decisions.")
md.append("")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md))

print(f"Wrote {XLSX_OUT} and {MD_OUT} with {len(questions)} round-3 PO questions.")
