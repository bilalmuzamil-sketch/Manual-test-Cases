#!/usr/bin/env python3
"""Assemble Simple Flow V1 manual test cases into an Excel workbook + CSV.

Reads the three group JSON case files (group A -> B -> C order) and emits:
  - SimpleFlow_V1_TestCases.xlsx  (Summary / Test Cases / VIU pending / Open Questions)
  - SimpleFlow_V1_TestCases.csv   (mirror of the Test Cases sheet)
"""
import csv
import json
import os
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
CASES_DIR = os.path.join(BASE, "cases")
XLSX = os.path.join(BASE, "SimpleFlow_V1_TestCases.xlsx")
CSV_PATH = os.path.join(BASE, "SimpleFlow_V1_TestCases.csv")

GROUP_FILES = [
    "group-A-settings-completion.json",
    "group-B-receiving-vendor.json",
    "group-C-review-permissions-validation-edge.json",
]

# ---------------------------------------------------------------- load + merge
cases = []
for gf in GROUP_FILES:
    with open(os.path.join(CASES_DIR, gf), encoding="utf-8") as fh:
        cases.extend(json.load(fh))

# ---- TestRail Case ID + link map (Standing Rule 8) ------------------------
TR_MAP = {}
_map_path = os.path.join(BASE, "testrail-id-map.csv")
if os.path.exists(_map_path):
    with open(_map_path, encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            if row.get("sf_id"):
                TR_MAP[row["sf_id"].strip()] = (row.get("ID") or "").strip()

def tr_id(sf_id):
    return TR_MAP.get(sf_id, "")

def tr_link(sf_id):
    tid = TR_MAP.get(sf_id, "")
    return ("https://shopview.testrail.io/index.php?/cases/view/" + tid) if tid else ""

# ---------------------------------------------------------------- area buckets
# (needle-in-area, bucket label). First match wins; order matters.
BUCKET_RULES = [
    ("Work Order Settings", "1 - Settings"),
    ("Completion — No-PO", "2 - Completion: No-PO/Skip"),
    ("Completion — PO + Optional", "3 - Completion: PO+Optional"),
    ("Completion — PO + Required", "4 - Completion: PO+Required"),
    ("Completion — Line approval", "Line-approval gate"),
    ("Completion — Idempotency", "Line-approval gate"),
    ("Core parts — Pre-Resolve", "18 - Core Pre-Resolve (SV-8353)"),
    ("Core Pre-Resolve", "18 - Core Pre-Resolve (SV-8353)"),
    ("Core parts", "Core parts (3/4/8/16)"),
    ("Vendorless / No-PN Part", "5 - Vendorless/No-PN part"),
    ("Vendor Missing", "6 - Vendor Missing flag"),
    ("PO Multi-Select", "7 - PO Multi-Select"),
    ("PO Bulk Receive", "8 - Bulk Receive page"),
    ("Apply Invoice", "9 - Apply Invoice"),
    ("Inline Part-Number Fix", "10 - Part-Number Fix"),
    ("Receive Button", "11 - Receive button on POs"),
    ("Accept Delivery", "12 - Accept Delivery"),
    ("Assign Vendor", "13 - Assign Vendor/Merge"),
    ("Waiting on Parts", "14 - Waiting on Parts"),
    ("UX Refinements", "15 - UX Refinements"),
    ("Review ON", "16 - Review ON"),
    ("Tech Story Flow", "17 - Tech Story Flow"),
    ("Permissions", "Permissions"),
    ("Validation", "Validation / Edge"),
    ("QuickBooks", "QuickBooks / Inventory"),
]

AREA_ORDER = [
    "1 - Settings",
    "2 - Completion: No-PO/Skip",
    "3 - Completion: PO+Optional",
    "4 - Completion: PO+Required",
    "Line-approval gate",
    "Core parts (3/4/8/16)",
    "18 - Core Pre-Resolve (SV-8353)",
    "5 - Vendorless/No-PN part",
    "6 - Vendor Missing flag",
    "7 - PO Multi-Select",
    "8 - Bulk Receive page",
    "9 - Apply Invoice",
    "10 - Part-Number Fix",
    "11 - Receive button on POs",
    "12 - Accept Delivery",
    "13 - Assign Vendor/Merge",
    "14 - Waiting on Parts",
    "15 - UX Refinements",
    "16 - Review ON",
    "17 - Tech Story Flow",
    "Permissions",
    "Validation / Edge",
    "QuickBooks / Inventory",
]


def bucket(area):
    for needle, label in BUCKET_RULES:
        if needle in (area or ""):
            return label
    return "Other"


for c in cases:
    c["_bucket"] = bucket(c["area"])

# ---------------------------------------------------------------- styles
DARK_BLUE = "1F4E78"
BRAND_BLUE = "257CFF"  # ShopView primary
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")
BANNER_FILL = PatternFill("solid", fgColor="C00000")
BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=14, color=DARK_BLUE)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

PRIORITY_FILL = {
    "Critical": PatternFill("solid", fgColor="F4B6B6"),
    "High": PatternFill("solid", fgColor="FFE08A"),
    "Medium": PatternFill("solid", fgColor="DCE6F5"),
    "Low": PatternFill("solid", fgColor="EDEDED"),
}
VIU_FILL = {
    "VIU-Verified": PatternFill("solid", fgColor="C6EFCE"),   # green
    "VIU-Pending": PatternFill("solid", fgColor="FFEB9C"),    # amber
    "Open-Question": PatternFill("solid", fgColor="F4B6B6"),  # red
}


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def n_join(items):
    return "\n".join(str(x) for x in (items or []))


wb = Workbook()

# ================================================================ Summary
ws = wb.active
ws.title = "Summary"

area_counts = Counter(c["_bucket"] for c in cases)
prio_counts = Counter(c["priority"] for c in cases)
viu_counts = Counter(c["viu_status"] for c in cases)

intro = (
    "Simple Flow V1 (Simple Mode) - Manual Test Cases\n\n"
    f"This workbook contains {len(cases)} manual test cases covering the ShopView "
    "\"Simple Mode - Streamlined Work Order Completion & Receiving\" feature (Epic "
    "SV-7301): the Work Order settings page; the four settings-driven completion "
    "flows (No-PO/skip, PO+Optional invoice, PO+Required invoice, and Review-ON); "
    "the tech-story and line-approval gates; core-part resolution; vendorless / "
    "Vendor-Missing part handling; PO multi-select, Bulk Receive, apply-invoice, "
    "part-number fix, the WO-PO Receive button, Accept Delivery and assign-vendor / "
    "merge; the Waiting-on-Parts column; UX refinements; permissions; validation / "
    "edge cases; and QuickBooks / inventory integrity.\n\n"
    "Authored from the COMPLETE spec (build/simple-flow/requirements.md, 17 stories) "
    "plus the design mockups / dev handoffs (design-notes.md) and the live VIU "
    "findings (viu-findings.md).\n\n"
    "VIU (Verify-in-UI on QA env sv7301) is PARTIAL because the feature is still "
    "under development. Each case carries a VIU Status: VIU-Verified (tested live and "
    "passed - screenshot cited in Notes), VIU-Pending (not yet verifiable - either a "
    "not-built story or a cookie/role-dependent check), or Open-Question (the expected "
    "result depends on an unresolved decision or a VIU deviation - see the Open "
    "Questions tab).\n\n"
    "NOT-BUILT stories (all cases VIU-Pending until dev completes): Story 7 (PO "
    "multi-select), Story 8 (Bulk Receive page), Story 9 (apply-invoice), Story 14 "
    "(Waiting-on-Parts column).\n\n"
    "Review the Open Questions tab BEFORE execution: it consolidates every spec/design "
    "conflict, unresolved decision, and the 4 VIU deviations/bugs found live."
)
ws["A1"] = intro
ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
ws["A1"].font = Font(size=11)
ws.merge_cells("A1:F1")
ws.row_dimensions[1].height = 360

# --- area table
r = 3
ws.cell(row=r, column=1, value="Cases by Area / Story").font = TITLE_FONT
r += 1
ws.cell(row=r, column=1, value="Area / Story").font = HEADER_FONT
ws.cell(row=r, column=2, value="# Cases").font = HEADER_FONT
for col in (1, 2):
    ws.cell(row=r, column=col).fill = HEADER_FILL
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
r += 1
for name in AREA_ORDER:
    ws.cell(row=r, column=1, value=name).border = BORDER
    ws.cell(row=r, column=2, value=area_counts.get(name, 0)).border = BORDER
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    r += 1
ws.cell(row=r, column=1, value="GRAND TOTAL").font = BOLD
ws.cell(row=r, column=2, value=sum(area_counts.values())).font = BOLD
for col in (1, 2):
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).fill = ALT_FILL
ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
r += 2

# --- priority table
ws.cell(row=r, column=1, value="Cases by Priority").font = TITLE_FONT
r += 1
ws.cell(row=r, column=1, value="Priority").font = HEADER_FONT
ws.cell(row=r, column=2, value="#").font = HEADER_FONT
for col in (1, 2):
    ws.cell(row=r, column=col).fill = HEADER_FILL
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
r += 1
for p in ["Critical", "High", "Medium", "Low"]:
    if p in prio_counts:
        ws.cell(row=r, column=1, value=p).border = BORDER
        ws.cell(row=r, column=1).fill = PRIORITY_FILL.get(p)
        ws.cell(row=r, column=2, value=prio_counts[p]).border = BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        r += 1
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=2, value=sum(prio_counts.values())).font = BOLD
for col in (1, 2):
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).fill = ALT_FILL
ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
r += 2

# --- VIU status table
ws.cell(row=r, column=1, value="Cases by VIU Status").font = TITLE_FONT
r += 1
ws.cell(row=r, column=1, value="VIU Status").font = HEADER_FONT
ws.cell(row=r, column=2, value="#").font = HEADER_FONT
for col in (1, 2):
    ws.cell(row=r, column=col).fill = HEADER_FILL
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
r += 1
for s in ["VIU-Verified", "VIU-Pending", "Open-Question"]:
    ws.cell(row=r, column=1, value=s).border = BORDER
    ws.cell(row=r, column=1).fill = VIU_FILL.get(s)
    ws.cell(row=r, column=2, value=viu_counts.get(s, 0)).border = BORDER
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    r += 1
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=2, value=sum(viu_counts.values())).font = BOLD
for col in (1, 2):
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).fill = ALT_FILL
ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
r += 2

# --- legend / notes
legend = (
    "Legend & notes:\n"
    "- VIU-Verified = tested live on QA env sv7301 and passed (screenshot cited in the case Notes).\n"
    "- VIU-Pending = still needs live verification: either a NOT-BUILT story (7 Bulk multi-select, "
    "8 Bulk Receive page, 9 apply-invoice, 14 Waiting-on-Parts) or a cookie/role-dependent check "
    "(non-admin role-gating negatives; full parts-receive + core round-trips; line-approval final block). "
    "See the 'VIU pending' tab.\n"
    "- Open-Question = the EXPECTED result depends on an unresolved decision or a live build deviation. "
    "See the 'Open Questions' tab (14 items: 3 RESOLVED by SV-8183, 11 still open pending Milos).\n"
    "- Permissions are now DEFINED by SV-8183: Simple Flow reuses existing Custom Roles atoms "
    "(no new atom) plus the NET-NEW reviewer!=completer rule. See requirements §9 for the "
    "action->atom map and per-role matrix; role-gating negatives still need live verification."
)
ws.cell(row=r, column=1, value=legend).font = Font(italic=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 160

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 12
for col in "CDEF":
    ws.column_dimensions[col].width = 16
ws.freeze_panes = "A2"

# ================================================================ Test Cases
tc = wb.create_sheet("Test Cases")
COLS = [
    ("ID", 14), ("Area", 28), ("Story Ref", 20), ("Title", 50),
    ("Priority", 12), ("Type", 12), ("Permissions", 32),
    ("Preconditions", 46), ("Steps", 52), ("Expected Result", 52),
    ("Design Ref", 28), ("VIU Status", 16), ("Notes", 48),
    ("TestRail Case ID", 16), ("TestRail Link", 48),
]
headers = [c[0] for c in COLS]
tc.append(headers)
style_header(tc, 1, len(headers))

csv_rows = [headers]
for i, c in enumerate(cases):
    row = [
        c["id"], c["area"], c["story_ref"], c["title"], c["priority"], c["type"],
        c["permissions_required"], n_join(c["preconditions"]), n_join(c["steps"]),
        n_join(c["expected"]), c["design_ref"], c["viu_status"], c.get("notes", ""),
        tr_id(c["id"]), tr_link(c["id"]),
    ]
    tc.append(row)
    csv_rows.append(row)
    excel_row = i + 2
    alt = (i % 2 == 1)
    for col in range(1, len(headers) + 1):
        cell = tc.cell(row=excel_row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL
    pfill = PRIORITY_FILL.get(c["priority"])
    if pfill:
        pc = tc.cell(row=excel_row, column=5)
        pc.fill = pfill
        pc.alignment = WRAP_TOP_CENTER
    vfill = VIU_FILL.get(c["viu_status"])
    if vfill:
        vc = tc.cell(row=excel_row, column=12)
        vc.fill = vfill
        vc.alignment = WRAP_TOP_CENTER
    tc.cell(row=excel_row, column=1).font = BOLD

for idx, (_, width) in enumerate(COLS, start=1):
    tc.column_dimensions[get_column_letter(idx)].width = width
tc.freeze_panes = "B2"

# ================================================================ VIU pending
viu = wb.create_sheet("VIU pending")
banner = (
    "VIU PENDING - Simple Flow is under development; this tab lists every case still "
    "needing live verification (NOT-BUILT stories + cookie/role-dependent checks). "
    "Fill in results as the build completes on QA env sv7301."
)
viu.append([banner])
viu.merge_cells("A1:G1")
viu["A1"].fill = BANNER_FILL
viu["A1"].font = BANNER_FONT
viu["A1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
viu.row_dimensions[1].height = 46

VIU_COLS = [
    ("ID", 14), ("Area", 28), ("Title", 52), ("VIU Status", 16),
    ("Why pending / what's needed to unblock", 52),
    ("VIU Result (Pass/Fail/Blocked)", 22), ("Observed Behavior + Screenshot", 40),
]
viu.append([c[0] for c in VIU_COLS])
style_header(viu, 2, len(VIU_COLS))

NOT_BUILT_BUCKETS = {
    "7 - PO Multi-Select", "8 - Bulk Receive page", "9 - Apply Invoice",
    "14 - Waiting on Parts",
}


def why_pending(c):
    b = c["_bucket"]
    notes = c.get("notes", "")
    if b in NOT_BUILT_BUCKETS:
        return "NOT BUILT - needs dev to complete the feature, then live verification."
    if "no non-admin" in notes or "non-admin" in notes or "tech key = 403" in notes \
            or c["type"] == "Permissions":
        return ("Role/cookie-dependent - needs a non-admin (or manager/foreman/office) "
                "credential; only the admin session works on sv7301 (tech key = 403). "
                "Expected also depends on the permissions matrix (TBD).")
    if "core" in c["area"].lower():
        return "Needs a cored part + a live receive/resolve round-trip walk."
    if "QuickBooks" in b:
        return "Needs QuickBooks / inventory inspection after a live receive/invoice."
    return "Not driven in the VIU run - needs a targeted live walk on sv7301."


viu_pending = [c for c in cases if c["viu_status"] == "VIU-Pending"]
for i, c in enumerate(viu_pending):
    viu.append([c["id"], c["area"], c["title"], c["viu_status"], why_pending(c), "", ""])
    excel_row = i + 3
    alt = (i % 2 == 1)
    for col in range(1, len(VIU_COLS) + 1):
        cell = viu.cell(row=excel_row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL
    viu.cell(row=excel_row, column=1).font = BOLD

for idx, (_, width) in enumerate(VIU_COLS, start=1):
    viu.column_dimensions[get_column_letter(idx)].width = width
viu.freeze_panes = "A3"

# ================================================================ Open Questions
oq = wb.create_sheet("Open Questions")
OQ_COLS = [
    ("#", 5), ("Topic", 26), ("Question / Issue", 58),
    ("Why it matters", 42), ("Affected cases", 30),
    ("Status / Resolution", 56),
]
oq.append([c[0] for c in OQ_COLS])
style_header(oq, 1, len(OQ_COLS))

RESOLVED_FILL = PatternFill("solid", fgColor="C6EFCE")  # green
OPEN_FILL = PatternFill("solid", fgColor="FFE08A")      # amber

open_questions = [
    ("No permissions / role matrix",
     "The spec supplied NO consolidated permissions/role matrix (§8 flagged it as "
     "unresolved). Which roles may complete WOs, bulk receive, edit settings, and "
     "sign off review was undefined.",
     "Every role-gating case's expected result depended on this.",
     "SF-PERM-01..10, SF-SET-11, SF-RCV-03, SF-REV-09",
     "RESOLVED by SV-8183. Simple Flow adds NO new permission atom - every action maps "
     "to an existing Custom Roles atom (see requirements §9 action->atom map + per-role "
     "matrix): Settings = App Settings (defaults Admin/Service Manager/Office); "
     "Complete WO = Work Orders C&E + WO Lines C&E/Full View; Bulk Receive = Vendor & "
     "Order Mgmt C&E + See Financial Data; Receive-on-WO = Order Parts (FE); Mark "
     "Reviewed = Review Work Orders + reviewer != completer. Source: SV-8183."),
    ("Spec V2.3 vs design V1.4 drift",
     "The product .doc header says V2.3 but both design handoffs cite 'Simple Mode "
     "V1.4'. It is unclear which is authoritative where they differ.",
     "Design-vs-spec differences (defaults, tech-story placement, review states) can't "
     "be adjudicated until the authoritative version is confirmed.",
     "Global - especially SF-SET-08, SF-TECH-08, SF-REV-08",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Settings first-use defaults conflict",
     "Spec §4/S1 defaults: Auto-approve OFF, Vendor invoice REQUIRED. Design "
     "HANDOFF.md defaults: Auto-approve ON, Vendor invoice Optional. The live org "
     "baseline shows autoApproveLines:true, requireVendorInvoiceNumber:false.",
     "The correct first-use defaults must be confirmed before pass/fail on the "
     "settings-default case and before framing the completion-flow matrix.",
     "SF-SET-08",
     "OPEN - pending Milos. Explicitly NOT addressed by SV-8183."),
    ("'Create Purchase Orders' toggle absent (build deviation #1)",
     "Spec S1-R2 requires a 'Create purchase orders' toggle (default ON) so POs can "
     "be turned OFF. It does NOT exist in the UI and there is no createPurchaseOrders "
     "field in the settings API - POs appear always-on.",
     "The pure Story-2 'No-PO / skip' configuration cannot be set up, so the Create-POs-"
     "OFF completion and its QB integrity cases are untestable as specified. May be an "
     "intentional descope - confirm.",
     "SF-SET-03, SF-COMP-06, SF-QB-02",
     "OPEN - pending Milos. Explicitly NOT addressed by SV-8183."),
    ("Tech-story placement (Story 17 vs S15-R2)",
     "S15-R2 (older) says the tech story stays on the line, not in a modal; Story 17 "
     "(SV-7876) supersedes with inline + gate-modal. Design handoff flags this "
     "[Confirm].",
     "Determines whether the tech-story gate-modal cases are authoritative. The live "
     "build showed the gate-modal working, so Story 17 appears current - confirm.",
     "SF-TECH-08 (and all SF-TECH-*)",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Cost at completion ($0-cost margins)",
     "§8 asks whether cost may be entered at completion to avoid $0-cost margins "
     "flowing to QuickBooks.",
     "Affects the expected QB margin figures at completion/invoice.",
     "SF-QB-06",
     "RESOLVED by SV-8183 (Decision 4). A vendorless / no-PN part add requires See "
     "Financial Data, so the mandatory sell price is captured at add time and $0-cost "
     "margins are avoided. Source: SV-8183."),
    ("Auto-receive of in-stock inventory on simple completion",
     "§8 asks whether in-stock inventory parts should auto-receive/decrement on simple "
     "completion. The spec also flags the 'bare status setter' skip path as an "
     "integrity risk (it can emit no events).",
     "Determines whether inventory decrement + Part History are expected on the skip "
     "path - a data-integrity invariant.",
     "SF-COMP-07, SF-QB-01",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Does the backend enforce Simple-Flow settings?",
     "§8 asks whether the backend should enforce the Simple-Flow settings, or whether "
     "they are front-end gates only.",
     "Sets the expected result for API-level negative tests (403/422 vs 200/201).",
     "SF-PERM-06",
     "RESOLVED by SV-8183. The backend DOES enforce the Simple-Flow settings AND the "
     "permission atoms (not FE-only), subject to the WO-atom collapse caveat: any role "
     "with WO Create & Edit can receive onto a WO, and the receive endpoint accepts the "
     "OR of ROLE_DELIVERY_CREATE_AND_EDIT / ROLE_WORK_ORDER_PART_CREATE / "
     "ROLE_WORK_ORDER_CREATE_AND_EDIT. Source: SV-8183."),
    ("Save Settings always enabled (build deviation #2)",
     "The Save Settings button is clickable with no pending changes (no dirty-state "
     "gating).",
     "Minor UX deviation from the expected 'disabled until changed' behavior; confirm "
     "intended.",
     "SF-SET-13",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Mark Reviewed missing optional note (build deviation #3)",
     "Story 16 R7/R10 specify an optional review note field (input_review_note). The "
     "live Mark Reviewed dialog exposes ONLY the VIN field.",
     "The optional-note case cannot pass as specified; confirm whether the field was "
     "descoped or is not yet built.",
     "SF-REV-10",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Review sign-off jumps to Complete (build deviation #4)",
     "Story 16 R5/R8 describe Review -> Reviewed (green, sign-off complete) -> a "
     "separate final Complete Work Order -> Complete. Live, Confirm Review went "
     "straight to Complete with no distinct Reviewed holding state observed.",
     "Determines whether a distinct Reviewed state + separate final Complete are "
     "expected (possible admin auto-progression). Affects the review state-machine "
     "cases and the invoicing-block-until-reviewed check.",
     "SF-REV-08, SF-REV-11",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Close-confirm modal design pending",
     "S15-R4 close-vs-cancel confirmation is explicitly 'Figma still to be added' - "
     "no design surface exists to test yet.",
     "The close-confirm case's expected behavior can't be finalized until the design "
     "ships.",
     "SF-UX-04",
     "OPEN - pending Milos. Not addressed by SV-8183."),
    ("Require-review default (new vs existing orgs)",
     "§8 leaves the Require-review default unresolved (on for bigger/existing shops?) "
     "with existing orgs preserved via backfill.",
     "Sets the expected default state for the review-default case.",
     "SF-REV-15",
     "OPEN - pending Milos. SV-8183 defined WHO can sign off review, but NOT the "
     "require-review default cohort."),
    ("Accept Delivery vendor-missing group ordering",
     "S12-R1 says the vendor-missing group sits at the BOTTOM; S12-R3 says the "
     "vendor-missing group LEADS (top). The spec contradicts itself.",
     "Determines the expected ordering on the Accept Delivery / Bulk Receive screens.",
     "SF-RCV-05, SF-RCV-07",
     "OPEN - pending Milos. Not addressed by SV-8183."),
]

for i, (topic, q, why, ids, status) in enumerate(open_questions, start=1):
    oq.append([i, topic, q, why, ids, status])
    excel_row = i + 1
    alt = (i % 2 == 1)
    for col in range(1, len(OQ_COLS) + 1):
        cell = oq.cell(row=excel_row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL
    oq.cell(row=excel_row, column=1).alignment = WRAP_TOP_CENTER
    scell = oq.cell(row=excel_row, column=6)
    scell.fill = RESOLVED_FILL if status.startswith("RESOLVED") else OPEN_FILL

resolved_n = sum(1 for x in open_questions if x[4].startswith("RESOLVED"))
open_n = len(open_questions) - resolved_n

for idx, (_, width) in enumerate(OQ_COLS, start=1):
    oq.column_dimensions[get_column_letter(idx)].width = width
oq.freeze_panes = "A2"

# ================================================================ save
wb.save(XLSX)

with open(CSV_PATH, "w", newline="", encoding="utf-8") as fh:
    writer = csv.writer(fh)
    for row in csv_rows:
        writer.writerow(row)

# ---------------------------------------------------------------- report
print("Total cases on Test Cases tab:", len(cases))
print("Open questions:", len(open_questions), "(RESOLVED:", resolved_n, "OPEN:", open_n, ")")
print("VIU-pending rows:", len(viu_pending))
print("Areas:")
for name in AREA_ORDER:
    print(f"  {area_counts.get(name, 0):3d}  {name}")
print("Priorities:", {p: prio_counts[p] for p in ["Critical", "High", "Medium", "Low"] if p in prio_counts})
print("VIU status:", dict(viu_counts))
print("Saved:", XLSX)
print("Saved:", CSV_PATH)
