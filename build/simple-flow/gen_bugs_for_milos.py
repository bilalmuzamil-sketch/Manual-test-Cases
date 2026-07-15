#!/usr/bin/env python3
"""Generate the MOST-SIMPLIFIED "possible problems" sheet for Milos (Simple Flow PO).

Purpose: give Milos (completely non-technical) a plain, everyday-scenario view of
the open bug drafts so HE can decide, per item, whether each is a real problem to
fix or is actually how it should work (expected). This is the PO-DECISION view of
the same items that live in the QA/dev-facing SimpleFlow_Bug-Drafts.xlsx — it does
NOT overwrite that file. (BUG-5 / TICKET 1 "reviewer can sign off own WO" was
DROPPED 2026-07-10 as expected behavior per Milos's ruling and is no longer listed.)

Outputs (regenerable — run from the repo root):
  - build/simple-flow/SimpleFlow_Bugs-for-Milos-Confirm.xlsx
  - build/simple-flow/SimpleFlow_Bugs-for-Milos-Confirm.md

Sheets:
  - "For Milos to confirm" : reader-facing, ZERO jargon (no API/HTTP/permission/
                             field/enum terms, no case IDs, no BUG codes). Columns:
                             # | Picture this | What happens today | Our question
                             for you | Your options | Your answer (blank).
  - "QA Internal Mapping"  : QA-only (NOT for Milos). Per item: internal BUG-code +
                             Jira TICKET#, affected case IDs with TestRail C##### +
                             clickable links (standing rule 8; source
                             testrail-id-map.csv), spec/story refs, current status,
                             and what each answer triggers on our side.

Source of the items: build/simple-flow/jira-bug-drafts.md (active drafts). The
reader-facing wording follows the friendly scenario tone of
gen_po_questions_round3.py Tab 1 (standing rule 7 — plain layman language).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/simple-flow/SimpleFlow_Bugs-for-Milos-Confirm.xlsx"
MD_OUT = "build/simple-flow/SimpleFlow_Bugs-for-Milos-Confirm.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STD_QUESTION = "Is this a problem we should fix, or is this how it should work?"
STD_OPTS_AB = ("A) This is a problem - please fix it.\n"
               "B) This is fine - it's how it should work.")

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, codes, or tech terms)
# ---------------------------------------------------------------------------
items = [
    # NOTE: BUG-5 / TICKET 1 (reviewer can sign off their own work order) was DROPPED
    # 2026-07-10 — settled as EXPECTED behavior by Milos's ruling (reviewer != completer
    # is not a v1 requirement). Removed from this PO-confirm sheet; see jira-bug-drafts.md
    # "Dropped — expected behavior per PO (Milos), 2026-07-10".
    {   # BUG-6 + BUG-7 / TICKET 2 (Medium)
        "picture": ("The screen correctly hides the \"finish\" and \"approve\" buttons "
                    "from people who aren't supposed to have them. But the system "
                    "behind the screen doesn't fully block those actions - so a "
                    "very technical person could still find a way around the screen "
                    "to do them."),
        "today": ("The buttons are hidden from the right people on screen, but the "
                  "block isn't fully enforced behind the scenes."),
        "question": STD_QUESTION,
        "opts": (STD_OPTS_AB + "\n"
                 "C) Fine for now, but fix it later."),
    },
    {   # BUG-8 / TICKET 3 (Medium)
        "picture": ("When finishing a repair job, the screen asks for details like the "
                    "mileage, the vehicle's ID number, or the engine hours. But the "
                    "system behind the screen doesn't truly require them - so a very "
                    "technical person could skip those details by going around the "
                    "screen."),
        "today": ("The screen asks for those details, but they aren't truly required "
                  "behind the scenes and could be skipped."),
        "question": STD_QUESTION,
        "opts": (STD_OPTS_AB + "\n"
                 "C) Fine for now, but fix it later."),
    },
    {   # BUG-11 / TICKET 4 (Low)
        "picture": ("There are two ways to receive parts that have arrived from a "
                    "supplier. The newer way works fine. The older way shows an error "
                    "message when a parts person tries to use it for parts that came "
                    "from a repair job."),
        "today": ("The newer receiving screen works. The older receiving screen shows "
                  "an error for these parts, so people should use the newer one."),
        "question": STD_QUESTION,
        "opts": (STD_OPTS_AB + "\n"
                 "C) Fine to just retire the old screen and keep the new one."),
    },
    {   # GAP-B / TICKET 5 (Medium)
        "picture": ("When a brand-new shop opens the app for the very first time, some "
                    "of the starting switches are set the wrong way by default - for "
                    "example, jobs auto-approve on their own, and a supplier's bill "
                    "isn't required when it should be."),
        "today": ("A brand-new shop starts with those switches set the opposite way "
                  "from what was agreed (auto-approve on, supplier's bill not "
                  "required)."),
        "question": STD_QUESTION,
        "opts": STD_OPTS_AB,
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (rule 8: TestRail C##### + clickable link per case).
# Each entry: bug_code, jira_ticket, priority, [(sf_id, testrail_id), ...],
#             refs, status, triggers
# ---------------------------------------------------------------------------
internal_map = [
    # BUG-5 / TICKET 1 REMOVED 2026-07-10: settled as EXPECTED by Milos (reviewer !=
    # completer is not a v1 requirement; a completer may review their own WO). No longer a
    # PO decision. SF-PERM-04/07 + SF-REV-09 re-adjudicated VIU-Verified; SF-PERM-08
    # obsolete. See jira-bug-drafts.md "Dropped — expected behavior per PO (Milos)".
    ("BUG-6 + BUG-7", "TICKET 2", "Medium",
     [("SF-PERM-06", 29410), ("SF-PERM-02", 29406), ("SF-PERM-07", 29411),
      ("SF-REV-09", 29394)],
     "SV-8183 backend-enforcement claim vs SV-7864 atom-collapse. Milos Round-2 Q5 "
     "already ruled UI gating = v1 pass; this tracks the backend gap.",
     "NOT filed. UI gating passes; backend allows the action (201) via API for a role "
     "without the permission (e.g. Technician). Recorded 'UI pass / API fail'.",
     "A (problem) -> file TICKET 2 (Medium) to add backend enforcement; keep cases "
     "'UI pass / API fail'. B (fine) -> accept UI-only gating as intended "
     "(atom-collapse per SV-7864); mark cases pass on UI, drop ticket. C -> file but "
     "defer (backlog)."),
    ("BUG-8", "TICKET 3", "Medium",
     [("SF-VAL-01", 29415), ("SF-VAL-02", 29416), ("SF-VAL-03", 29417),
      ("SF-COMP-05", 29294), ("SF-COMP-16", 29305), ("SF-REV-03", 29388)],
     "SV-8183 backend-enforcement claim; requirements.md §4 required-vehicle-field "
     "gates (mileage / VIN / engine hours). Related SV-7864 atom-collapse.",
     "NOT filed. Wizard blocks completion until fields are filled "
     "(viu-evidence/VIU2-02-mileage-gate.png) but backend completes without them "
     "(simple-complete returned 201 with mileage empty).",
     "A (problem) -> file TICKET 3 (Medium) to enforce required fields backend-side; "
     "cases stay Deviation until fixed. B (fine) -> UI-only enforcement accepted; "
     "mark SF-VAL-01/02/03 + related expected/pass on UI. C -> file but defer."),
    ("BUG-11", "TICKET 4", "Low",
     [("SF-COMP-13", 29302), ("SF-COMP-19", 29308), ("SF-VAL-05", 29419),
      ("SF-VAL-06", 29420), ("SF-PNFIX-02", 29364), ("SF-PNFIX-03", 29365),
      ("SF-PNFIX-04", 29366), ("SF-PNFIX-05", 29367), ("SF-PNFIX-06", 29368),
      ("SF-RCV-08", 29376), ("SF-VPART-07", 29337), ("SF-REV-04", 29389),
      ("SF-REV-14", 29399), ("SF-CORE-03", 29315), ("SF-CORE-04", 29316),
      ("SF-CORE-05", 29317), ("SF-CORE-07", 29319)],
     "SV-7301 Story 10 (receive creates/links part) / Story 8 (Bulk Receive = the "
     "working path). Downgraded to Low 2026-07-09: the 500 is confined to the LEGACY "
     "single-PO Accept-Delivery path; Bulk Receive works (200).",
     "NOT filed. Legacy Accept-Delivery receive of a WO-PO returns HTTP 500; new Bulk "
     "Receive receives the same WO PO fine. Evidence R7-01/R7-04/R7-06 in "
     "viu-evidence/. Affected cases now largely testable via Bulk Receive.",
     "A (problem) -> file TICKET 4 (Low) to fix the legacy Accept-Delivery 500. "
     "B (fine) -> accept legacy path as-is (use Bulk Receive); mark cases pass via "
     "Bulk Receive path. C -> retire the legacy single-PO Accept-Delivery surface "
     "(product/scope decision) and standardize on Bulk Receive."),
    ("GAP-B", "TICKET 5", "Medium",
     [("SF-SET-08", 29282)],
     "SV-7301 §4 / Story 1 first-use defaults (confirmed Milos Q3: Auto-approve Lines "
     "OFF, Create Purchase Orders ON, Vendor Invoice REQUIRED).",
     "NOT filed. First-use build ships Auto-approve Lines ON and Vendor Invoice "
     "Optional (autoApproveLines:true, requireVendorInvoiceNumber:false) - opposite "
     "of the confirmed defaults.",
     "A (problem) -> file TICKET 5 (Medium) to correct the first-use defaults; "
     "SF-SET-08 stays Deviation until fixed. B (fine) -> Milos re-confirms the shipped "
     "defaults are acceptable; update SF-SET-08 expected to match live and pass."),
]

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
wb = Workbook()

ws = wb.active
ws.title = "For Milos to confirm"
ws.column_dimensions["A"].width = 5
for col, w in zip("BCDEF", [54, 42, 40, 52, 28]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Simple Mode - A Few Things for You to Confirm"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
intro = ("Thanks so much, Milos! Below are a few things we spotted. For each one, "
         "just tell us if it's a problem we should fix, or if it's actually fine and "
         "how it should work - pick one option per row. There are no wrong answers, "
         "we just need your call.")
ws["A2"] = intro
ws["A2"].alignment = WRAP
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 48

headers = ["#", "Picture this", "What happens today", "Our question for you",
           "Your options", "Your answer"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, item in enumerate(items, start=1):
    row = HDR_ROW + i
    vals = [i, item["picture"], item["today"], item["question"], item["opts"], ""]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 150

# --- QA Internal Mapping sheet ---
wi = wb.create_sheet("QA Internal Mapping")
wi["A1"] = ("INTERNAL - for QA only. Do NOT share this tab (or any IDs/codes on it) "
            "with the PO / Milos.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:H1")

ihead = ["#", "Bug code", "Jira ticket", "Priority", "Affected cases (C-ID + link)",
         "Spec / story refs", "Current status", "What each answer triggers on our side"]
iwid = [4, 14, 11, 9, 40, 40, 42, 60]
IH_ROW = 3
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for i, (bug, ticket, prio, cases, refs, status, triggers) in enumerate(internal_map, start=1):
    # Build the affected-cases cell as text; put clickable links via a comment-free
    # approach — one hyperlink per cell isn't possible for many, so we list them and
    # add the first as the cell hyperlink is not ideal. Instead we render each case on
    # its own line "SF-ID  C#####" and set the cell hyperlink to the TestRail case
    # search is not possible; keep them as plain text with the links visible.
    cases_text = "\n".join(f"{sf}  C{tr}  {TR_LINK.format(tr)}" for sf, tr in cases)
    rowvals = [i, bug, ticket, prio, cases_text, refs, status, triggers]
    for c, v in enumerate(rowvals, start=1):
        cell = wi.cell(row=r, column=c, value=v)
        cell.alignment = WRAP_CENTER if c in (1, 2, 3, 4) else WRAP
        cell.border = BORDER
    # Make the affected-cases cell hyperlink to the first case (best-effort clickable);
    # the full clickable list is preserved in the .md mirror.
    if cases:
        link_cell = wi.cell(row=r, column=5)
        link_cell.hyperlink = TR_LINK.format(cases[0][1])
        link_cell.font = LINK_FONT
    wi.row_dimensions[r].height = max(120, 22 * len(cases))
    r += 1

note_row = r + 1
wi.cell(row=note_row, column=1, value=(
    "Notes: These items are the reconciled Simple Flow bug drafts (jira-bug-drafts.md, "
    "2026-07-09 post-Milos-Round-2, updated 2026-07-10). BUG-5/TICKET 1 (reviewer can "
    "sign off own WO) was DROPPED 2026-07-10 as expected behavior per Milos and is no "
    "longer listed. This sheet is the PO-DECISION view so Milos can "
    "confirm expected-vs-bug; it does NOT replace the QA/dev-facing "
    "SimpleFlow_Bug-Drafts.xlsx. TestRail Case IDs sourced from testrail-id-map.csv "
    "(standing rule 8); every affected case's clickable link is preserved in the .md "
    "mirror. Bugs stay OUT of the PO-facing tab (standing rule 7) - it only asks the "
    "product decision. None of the 5 are filed in Jira yet (Atlassian MCP was "
    "unavailable at run time).")
    ).alignment = WRAP
wi.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=8)
wi.row_dimensions[note_row].height = 90

wb.save(XLSX_OUT)

# ---------------------------------------------------------------------------
# Markdown mirror
# ---------------------------------------------------------------------------
md = []
md.append("# Simple Mode — A Few Things for You to Confirm")
md.append("")
md.append("Thanks so much, Milos! Below are a few things we spotted. **For each one,")
md.append("just tell us if it's a problem we should fix, or if it's actually fine and")
md.append("how it should work — pick one option per row.** There are no wrong answers,")
md.append("we just need your call.")
md.append("")
md.append("---")
for i, item in enumerate(items, start=1):
    md.append("")
    md.append(f"## {i}.")
    md.append("")
    md.append("**Picture this**")
    md.append(item["picture"])
    md.append("")
    md.append("**What happens today**")
    md.append(item["today"])
    md.append("")
    md.append("**Our question for you**")
    md.append(item["question"])
    md.append("")
    md.append("**Your options**")
    for line in item["opts"].split("\n"):
        md.append(f"- {line}")
    md.append("")
    md.append("**Your answer:** ______________________________________________")
    md.append("")
    md.append("---")
md.append("")
md.append("## Thank you!")
md.append("")
md.append("That's everything. Your answers tell us which of these to fix and which are")
md.append("fine as they are. Feel free to add any notes next to your choices.")
md.append("")
md.append("---")
md.append("---")
md.append("")
md.append("## Internal — QA-only mapping (NOT for the PO / Milos)")
md.append("")
md.append("Links each plain-English item above to its internal bug code + Jira ticket,")
md.append("affected cases (with TestRail links), refs, current status and what each")
md.append("answer triggers. **Do not include this section (or any IDs/codes in it) in")
md.append("the PO-facing copy or the \"For Milos to confirm\" tab.**")
md.append("")
for i, (bug, ticket, prio, cases, refs, status, triggers) in enumerate(internal_map, start=1):
    md.append(f"### Item {i} — {bug} ({ticket}, {prio})")
    md.append("")
    md.append("- **Affected cases:**")
    for sf, tr in cases:
        md.append(f"  - {sf} — [C{tr}]({TR_LINK.format(tr)})")
    md.append(f"- **Refs:** {refs}")
    md.append(f"- **Current status:** {status}")
    md.append(f"- **What each answer triggers:** {triggers}")
    md.append("")
md.append("**Notes:** These items are the reconciled Simple Flow bug drafts (BUG-5/")
md.append("TICKET 1 reviewer-self-review dropped 2026-07-10 as expected per Milos)")
md.append("(`jira-bug-drafts.md`, 2026-07-09 post-Milos-Round-2). This is the")
md.append("PO-DECISION view so Milos can confirm expected-vs-bug; it does NOT replace")
md.append("the QA/dev-facing `SimpleFlow_Bug-Drafts.xlsx`. TestRail IDs sourced from")
md.append("`testrail-id-map.csv` (standing rule 8). Bugs stay OUT of the PO-facing tab")
md.append("(standing rule 7). None of the 5 are filed in Jira yet (Atlassian MCP was")
md.append("unavailable at run time).")
md.append("")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md))

print(f"Wrote {XLSX_OUT} and {MD_OUT} with {len(items)} PO-confirm items.")
