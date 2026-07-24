#!/usr/bin/env python3
"""Generate SimpleFlow_SV-8183_vs-QA-Issues_Analysis_2026-07-24.xlsx
Tabs: Executive Summary, Per-Issue Analysis, Where Our Report Was Wrong,
      Root Cause, Corrective Actions, Was QA Right?
Rules 7/8/16/19: layman prose; C-id + TestRail link on every case ref; established workbook style.
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = "/home/user/Manual-test-Cases/build/simple-flow/sv8183/SimpleFlow_SV-8183_vs-QA-Issues_Analysis_2026-07-24.xlsx"

HDR = PatternFill("solid", fgColor="1F4E78")
HDRF = Font(bold=True, color="FFFFFF", size=11)
TITLEF = Font(bold=True, size=14, color="1F4E78")
SUBF = Font(italic=True, size=10, color="555555")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="top", horizontal="center")
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BAD = PatternFill("solid", fgColor="F8CBAD")
GOOD = PatternFill("solid", fgColor="C6E0B4")
WARN = PatternFill("solid", fgColor="FFE699")

def tl(cid):
    return f"https://shopview.testrail.io/index.php?/cases/view/{cid}"

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR; cell.font = HDRF; cell.alignment = WRAP; cell.border = BORDER

def put_table(ws, start_row, headers, rows, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for i, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = WRAP; cell.border = BORDER
        r += 1
    return r

wb = openpyxl.Workbook()

# ---------------- Tab 1: Executive Summary ----------------
ws = wb.active
ws.title = "Executive Summary"
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 120
ws.cell(row=1, column=2, value="Simple Flow — SV-8183 Permission Test vs. QA-found Issues: Reconciliation & Honest Analysis").font = TITLEF
ws.cell(row=2, column=2, value="Feature: Simple Flow (Epic SV-7301) — permission/role controls (Story SV-8183) · Environment: app.staging.shopview.com · PO: Milos · Date: 2026-07-24 · No TestRail writes").font = SUBF
ws.cell(row=3, column=2, value="Prior report under review: SimpleFlow_SV-8183_Permission-Test-Report_2026-07-23.md (concluded 11/11 PASS). Live re-verify on clean template roles (drift ruled out, Rule 26).").font = SUBF

lines = [
 ("CORRECTION 2026-07-24 — our 2026-07-23 report over-claimed.", True),
 ("On 2026-07-23 we published a permission report concluding \"11 of 11 test cases passed, zero mismatches, controls behave exactly as specified — PASS.\" QA (Ayesha) then raised three issues (SV-8515, SV-8516, SV-8541) against the same feature.", False),
 ("We re-checked all three LIVE on staging against clean (template-reset) roles, capturing the real screen and the real system response for each step. All three are real coverage gaps. Our report over-claimed: the 11 cases were correct for what they tested, but the SUITE did not cover the action paths these issues live on — and we reported \"all pass\" as if the whole feature were clean.", False),
 ("", False),
 ("SV-8515 — REAL defect (developer accepting the fix, \"Ready to Fix\").", True),
 ("A view-only parts user is correctly NOT shown the per-order Receive button, but can still tick several orders, click \"Receive Selected,\" and the full editable receiving screen opens — which should never appear for a view-only user. GOOD NEWS: when they actually try to receive, the system REFUSES it (Access denied / HTTP 403) — nothing is received or changed. So it is a misleading, dead-end screen that shouldn't be reachable, NOT a view-only user really receiving. QA's \"receive same as Admin / bypasses the permission model\" is right that there's a defect but OVERSTATED — the back end blocks the actual receive.", False),
 ("", False),
 ("SV-8516 — REAL, front-end already fixed, one back-end residue to flag.", True),
 ("A Time Clock user used to edit/cancel/return parts and change the vendor. The front-end fix HOLDS — that user now only sees \"Return.\" BUT the system will still accept a part EDIT sent directly to it (API returns 200 and the change persists). Per Standing Rule 24 a front-end block the API can still get past is FLAGGED, not a re-opened bug, unless the PO decides the back end must enforce it.", False),
 ("", False),
 ("SV-8541 — REAL, but pre-existing and awaiting a product decision.", True),
 ("A user WITHOUT \"Work Order Line: Create & Edit\" can still RESOLVE CORES — the system accepts it (returns success / HTTP 201) even for the lowest-privilege Time Clock role. This is NOT something Simple Flow broke; it behaves the SAME on Production, and the spec itself (§9.4) says these fine-grained work-order gates are front-end conveniences the back end does not separately enforce. Correctly OPEN with the developer (Sasha) as a clarification.", False),
 ("", False),
 ("CORRECTED VERDICT.", True),
 ("The 11 cases still pass for what they tested (the documented per-role matrix cells). But the report's FEATURE-WIDE \"PASS / everything behaves as specified\" claim is CORRECTED: there are three action-path coverage gaps — two real defects (one front-end-exposure defect the developer is already fixing; one front-end-fixed-but-back-end-open flag) and one pre-existing/spec-interpretation item open with the developer. QA caught three genuine gaps our pass missed.", False),
]
r = 5
for text, bold in lines:
    c = ws.cell(row=r, column=2, value=text)
    c.alignment = WRAP
    c.font = Font(bold=bold, size=11, color=("C00000" if bold else "000000"))
    r += 1
ws.freeze_panes = "A5"

# ---------------- Tab 2: Per-Issue Analysis ----------------
ws2 = wb.create_sheet("Per-Issue Analysis")
headers2 = ["Issue (link)", "What QA reported", "Role / env QA used",
            "Our prior finding (case + C-id + link)",
            "Live ground truth this run (UI + API status + evidence)",
            "Verdict", "Spec wording it deviates from (verbatim, Rule 25)",
            "Was QA right?", "Our miss (path not driven)"]
rows2 = [
 ["SV-8515\nhttps://shopview.atlassian.net/browse/SV-8515\n(Story Defect · Ready to Fix · Dusan)",
  "Office / Vendor & Order Mgmt VIEW-only user has no per-PO Receive button but can multi-select → \"Receive Selected\" → edit invoice/vendor and bulk receive \"same as Admin.\"",
  "Office user (real prereq = Vendor & Order Mgmt: View only). Staging vs Production.",
  f"SF-PERM-03 / C29407 ({tl(29407)}) — \"which roles can Bulk Receive\"; we said \"Office view-only per spec — PASS.\"\nAdjacent: SF-PERM-05 / C29409 ({tl(29409)}) — per-PO button hidden (still valid).",
  "UI: per-PO Receive HIDDEN (matches). BUT after multi-select the \"Receive Selected\" button appears and opens the full editable /bulk-receive \"Receive Vendor Parts\" screen (33 editable inputs — invoice#, date, cost $, tax), no front-end gate (sv8515-office-after-select.png, sv8515-bulk-receive-screen.png).\nAPI: read receive-view = 200 (why the screen loads); actual POST /api/inventory/orders/accept = HTTP 403 {\"error\":\"Access denied.\"} — receive does NOT complete (sv8515-recv5-net.json). Control: empty body = 400 (validation) for Admin vs 403 for Office = real back-end gate. Clean role (drift ruled out).",
  "REAL front-end-exposure defect (dev-accepted, Ready to Fix). NOT a data bypass — back end blocks the actual receive.",
  "§9.1: \"Bulk Receive page (accountant, PO-list driven) | 7/8/9 | Vendor & Order Mgmt: Create & Edit (route gate hasPartsPermissions) + See Financial Data...\" — front-end must require Create & Edit to reach the screen.\n§9.2 Office = \"No (4)\", footnote 4: \"Office has Vendor & Order Mgmt: View only → can open Bulk Receive but cannot receive.\" → back end matches spec; FRONT-END deviates by exposing the Receive-Selected entry point.",
  "YES — real front-end defect — but OVERSTATED the bypass (receive is blocked at the back end; no escalation / no data mutation).",
  "We never drove the multi-select \"Receive Selected\" alternate entry point for a View-only user. SF-PERM-03/C29407 stopped at route-level nav; SF-PERM-05/C29409 stopped at the per-PO button."],
 ["SV-8516\nhttps://shopview.atlassian.net/browse/SV-8516\n(Story Defect · Done · Staging_Verified · Dusan)",
  "Time Clock user could edit / cancel / return parts and change vendor (should have no access; Production blocks it).",
  "Time Clock user. Staging vs Production.",
  f"No dedicated part-action negative existed.\nNearest: SF-PERM-09 / C29413 ({tl(29413)}) — financial part-add gate; SF-PERM-10 / C29414 ({tl(29414)}) — per-role COMPLETION matrix only.",
  "UI (the fix): Time Clock part-row ⋮ menu now shows ONLY \"Return\" — Edit / Cancel / Change-Vendor removed (sv8516-tc-menus.json, sv8516-tc-wo-lines.png) → Done/Staging_Verified fix HOLDS at the front end.\nAPI (the residue): Time Clock (confirmed 3/3 perms live) can STILL edit a part via POST /api/work-orders/part/change-request = HTTP 200, change PERSISTED (re-GET showed new description) — same as Admin. Edit of an already-received part is blocked by STATE (\"can't be modified once received\"), not permission. Clean role (drift ruled out).",
  "Original over-grant was REAL, now front-end-FIXED; residual back-end-accepts-edit = Rule-24 API-possible FLAG — not a re-opened bug unless PO requires back-end enforcement.",
  "§9.2: Time Clock = \"No\" across every column (no access) — a Time Clock editing a part contradicts this.\n§9.1 (Sasha's mapping): part-request management (make/edit/cancel) → \"WO Lines: Create & Edit\"; Time Clock lacks it → back end should block.\n§9.4 counterweight: \"FE distinctions ... are conveniences, not BE-enforceable boundaries\" → residual back-end-possibility is spec-anticipated (flag).",
  "YES — real; front-end FIXED, back-end still OPEN (Rule-24 flag).",
  "We had NO per-role part edit / cancel / return negative anywhere in the suite."],
 ["SV-8541\nhttps://shopview.atlassian.net/browse/SV-8541\n(Story Defect / Clarification · Open · Sasha)",
  "A user WITHOUT \"WO Line: Create & Edit\" can return a received special-order part and resolve cores (OK/Not OK). Same on Staging AND Production.",
  "User lacking WOL C&E (tested Office and Time Clock). Staging + Production.",
  f"SF-REV-14 / C29399 ({tl(29399)}) — \"cores decided before receiving\" — touches core resolution but NO per-role permission-negative.",
  "API: POST /api/work-orders/{id}/pre-resolve-cores {cores:[{partRequestId,isCoreOk:true}]} = HTTP 201 {\"resolvedCount\":1} for OFFICE (WO 3996683a) AND TIME CLOCK (WO 1b6f0ae6) — endpoint applies ZERO permission check (LOG-SV-8541). Part-action endpoints returned STATE 400 (not permission 403) for Office & Time Clock.\nUI: low-priv roles also SEE the controls (Office shows Return/Core/Edit; Time Clock ⋮ shows Return). Clean template roles (drift ruled out, before==after).",
  "REAL & reproduces on clean roles. Time Clock holds NONE of the §9.4 collapsing atoms yet still succeeds → genuine missing back-end check — BUT pre-existing (matches Production), spec-anticipated per §9.4. Correctly OPEN for Sasha (Rule 24/25 clarification).",
  "§9.1: \"Resolve inventory / special-order cores (Ok/Not OK) | 3/4/16 | WO Lines: Create & Edit.\" — spec requires WOL C&E.\n§9.2: Time Clock = \"No\" every column.\nCounterweight (Rule 25): §9.4: collapse is \"a deliberate, spec-sanctioned low-privilege trade-off (SV-7864). FE distinctions ... are conveniences, not BE-enforceable boundaries.\" The one part §9.4 does NOT cover: a role (Time Clock) holding NONE of the collapsing atoms still succeeding = the gap for Sasha.",
  "YES — real; PRE-EXISTING / spec-interpretation (flag), correctly a clarification.",
  "SF-REV-14/C29399 exercised cores FUNCTIONALLY but never as a per-role permission-negative."],
]
rr = put_table(ws2, 1, headers2, rows2, [26, 34, 20, 34, 52, 30, 46, 26, 34])
# color verdict column (6) per row
for i, fill in enumerate([BAD, WARN, WARN]):
    ws2.cell(row=2 + i, column=6).fill = fill
ws2.freeze_panes = "A2"

# ---------------- Tab 3: Where Our Report Was Wrong ----------------
ws3 = wb.create_sheet("Where Our Report Was Wrong")
headers3 = ["#", "Dimension we missed", "What we DID do", "What we SHOULD have done", "Issue it hid"]
rows3 = [
 ["1", "Alternate entry points not driven per role",
  "Verified route-level nav (/parts/orders allowed/denied) and the per-PO Receive button being hidden.",
  "Also drive the multi-select → \"Receive Selected\" path into the editable Bulk-Receive screen. A hidden button on one path ≠ action unreachable on every path.",
  "SV-8515"],
 ["2", "Back end not probed per granular action",
  "Probed the settings SAVE endpoint (403/200) and concluded \"back end enforces.\"",
  "Probe the specific part-action endpoints (part/change-request, pre-resolve-cores, status-action) as EACH no-access role — the only way to catch front-end-exposure gaps AND Rule-24 API-possible flags.",
  "SV-8516, SV-8541"],
 ["3", "No cases existed for these negatives",
  "Suite covered completion / bulk-receive routing / review only.",
  "Have per-role part edit/cancel/return negatives and per-role core-resolution negatives. Several report rows were \"PASS (composition-verified)\" — inferred from the role's permission list + an inherited gate, not driven live per action (Rule 12: matching permission list ≠ observed enforced action).",
  "SV-8516, SV-8541"],
 ["4", "The completeness FRAMING was mis-scoped",
  "Reported \"110 combinations, zero mismatches, controls behave exactly as specified — PASS\" as FEATURE-WIDE completeness.",
  "Scope the claim: \"the documented §9.2 matrix cells pass; action-path and back-end-per-action coverage is not yet exhaustive.\" A passing matrix cell ≠ a fully-enforced action.",
  "All three"],
]
put_table(ws3, 2, headers3, rows3, [4, 34, 44, 60, 18])
ws3.cell(row=1, column=1, value="THE THREE DIMENSIONS WE MISSED").font = Font(bold=True, size=12, color="C00000")
# Got-right note
gr = 8
ws3.cell(row=gr, column=1, value="WHAT WE GOT RIGHT (honest credit)").font = Font(bold=True, size=12, color="375623")
c = ws3.cell(row=gr + 1, column=1, value="On SV-8515, our report said Office is \"view-only (can open, cannot receive) exactly per spec\" and that the back end enforces (403). That conclusion is CORRECT — the live accept call returns 403 and nothing is received. What we missed was the front-end EXPOSURE of the editable screen along the way. So: right about back-end enforcement, wrong to assume the front-end path was therefore clean.")
c.alignment = WRAP
ws3.merge_cells(start_row=gr + 1, start_column=1, end_row=gr + 1, end_column=5)
ws3.row_dimensions[gr + 1].height = 60
ws3.cell(row=gr + 1, column=1).fill = GOOD

# ---------------- Tab 4: Root Cause ----------------
ws4 = wb.create_sheet("Root Cause")
headers4 = ["Cause", "Detail"]
rows4 = [
 ["Summary", "Our VIU applied the four documented layers (composition / back-end endpoint / front-end route / on-screen element) to the 11 cases we had, and did that well for those cells. The failure was in SCOPE and METHOD BREADTH, not in the individual observations."],
 ["(a) Action-path coverage not exhaustive", "We tested WHETHER a role can reach a feature (route/nav/button) but not EVERY way to perform the action — alternate entry points (multi-select / bulk / kebab menus) were not driven per role."],
 ["(b) Per-action back-end probing not exhaustive", "We probed one endpoint family (settings) for 403/200 and generalised \"back end enforces.\" We did not send the specific granular action (part edit, core resolve) to the back end AS each no-access role — the only way to catch front-end-exposure gaps and Rule-24 API-possible flags."],
 ["(c) Suite had no cases", "No per-role part edit/cancel/return negative and no per-role core-resolution negative existed — so nothing prompted the checks."],
 ["(d) Completeness claim mis-scoped", "\"All pass\" was reported as feature-wide. A passing matrix cell is not a fully-enforced action, and several rows were composition-inferred rather than action-driven."],
 ["In one line", "We proved the documented gates were configured correctly, but did not adversarially try to BREAK every action from every role by every path and against the back end — so the enforcement holes hid in the paths and endpoints we never exercised."],
]
put_table(ws4, 1, headers4, rows4, [40, 108])
ws4.freeze_panes = "A2"

# ---------------- Tab 5: Corrective Actions ----------------
ws5 = wb.create_sheet("Corrective Actions")
headers5 = ["#", "Proposed case (NOT authored / NOT pushed)", "Assert / flag", "Driving ticket + spec anchor (Rule 20)", "Status"]
rows5 = [
 ["1", "(NEW) Bulk-Receive \"Receive Selected\" multi-select negative for a Vendor & Order Mgmt View-only user. Also UPDATE SF-PERM-03 / C29407 (" + tl(29407) + ") to drive the multi-select path, not just route nav.",
  "Assert: per-PO Receive hidden (passes) AND the multi-select \"Receive Selected\" must NOT open the editable /bulk-receive screen for a View-only user (currently the front-end EXPOSES it — defect) AND the actual receive is blocked at the back end (accept → 403).",
  "SV-8515 / SV-8183 (§9.1 Bulk-Receive route gate hasPartsPermissions; §9.2 footnote 4).",
  "Needs user approval to author + eventual TestRail add (Rule 6)."],
 ["2", "(NEW) No-access role (Time Clock) part edit / cancel / return negative.",
  "Assert: UI HIDES Edit / Cancel / Change-Vendor for Time Clock (passes — the SV-8516 front-end fix) AND flag per Rule 24 that the action is still possible via the API (part/change-request → 200, persisted). Not a bug unless PO requires back-end enforcement.",
  "SV-8516 / SV-8183 (§9.2 Time Clock row; §9.1 part-request → WO Lines: Create & Edit; §9.4 atom-collapse).",
  "Needs user approval to author + eventual TestRail add (Rule 6)."],
 ["3", "(NEW) WO-Lines-C&E-gated core-resolution / received-part-return negative.",
  "Assert: a role WITHOUT WO Lines: Create & Edit (test Time Clock) — intended rule is it cannot resolve cores / return a received special part; capture that the back end CURRENTLY returns 201 with no gate (pre-resolve-cores). Pending Sasha's SV-8541 ruling on whether back-end enforcement is required.",
  "SV-8541 / SV-8183 (§9.1 resolve cores → WO Lines: Create & Edit; §9.2 Time Clock row; §9.4 atom-collapse caveat).",
  "Needs user approval to author + eventual TestRail add (Rule 6)."],
]
put_table(ws5, 1, headers5, rows5, [4, 50, 60, 40, 26])
ws5.freeze_panes = "A2"

# ---------------- Tab 6: Was QA Right? ----------------
ws6 = wb.create_sheet("Was QA Right")  # '?' is invalid in Excel sheet titles
headers6 = ["Issue", "Our nearest case (C-id + link)", "QA's call", "Reality (live-verified)", "Was QA right?"]
rows6 = [
 ["SV-8515", f"SF-PERM-03 / C29407 ({tl(29407)}) · SF-PERM-05 / C29409 ({tl(29409)})",
  "Real defect; \"view-only can receive same as Admin — bypasses the permission model.\"",
  "Real FRONT-END-EXPOSURE defect (dev accepting, Ready to Fix). Back end BLOCKS the actual receive (403) — no data bypass.",
  "YES — real defect, but OVERSTATED the bypass."],
 ["SV-8516", f"SF-PERM-09 / C29413 ({tl(29413)}) · SF-PERM-10 / C29414 ({tl(29414)})",
  "Real over-grant (Time Clock could edit/cancel/return/change-vendor).",
  "Was real; front-end now FIXED (⋮ = only Return). Back end still accepts a part edit (change-request → 200) = Rule-24 flag.",
  "YES — real, front-end-fixed, back-end still open."],
 ["SV-8541", f"SF-REV-14 / C29399 ({tl(29399)})",
  "Real; user without WO Lines: C&E can resolve cores / return received part (Staging + Production).",
  "Real; PRE-EXISTING (matches Production), spec-anticipated (§9.4). Time Clock exceeds the documented collapse. Correctly OPEN for Sasha.",
  "YES — real, pre-existing / flag, correctly a clarification."],
 ["NET", "—",
  "3 issues raised.",
  "3 real coverage gaps confirmed. Two real defects (one front-end-exposure dev-accepted; one front-end-fixed / back-end-open Rule-24 flag) + one pre-existing/spec-interpretation clarification open with the dev.",
  "QA caught 3 real gaps. OUR REPORT OVER-CLAIMED completeness."],
]
put_table(ws6, 1, headers6, rows6, [10, 46, 40, 52, 34])
for i in range(2, 5):
    ws6.cell(row=i, column=5).fill = GOOD
ws6.cell(row=5, column=5).fill = WARN
ws6.freeze_panes = "A2"

# reasonable row heights
for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=2):
        sheet.row_dimensions[row[0].row].height = None

wb.save(OUT)
print("wrote", OUT)
