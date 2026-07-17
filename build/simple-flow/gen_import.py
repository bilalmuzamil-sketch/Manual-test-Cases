#!/usr/bin/env python3
"""Emit Simple Flow (Simple Mode) cases in the user's exact TestRail import CSV
format (matched to testrail-import/sv5319-testrail-import-MATCHED.csv), plus xlsx.

CONTENT RULES enforced here (see task PART 2):
  1. NO VIU wording. Internal `viu_status`, `notes`, `design_ref` are NOT emitted.
  2. Simple Flow is SETTINGS-driven, not flag-driven: legitimate settings
     preconditions (e.g. "Auto-approve Lines is ON") are KEPT. There is no
     feature-flag phrase in the source; the sanity check confirms 0 occurrences.
  3. Sections = leaf area names only (already leaf names, no "Simple" prefix).
  4. References = Jira story id(s) (SV-####) plus the spec-rule reference. No
     internal SF- ids, no VIU text.
  5/6. Preconditions/Steps/Expected kept as authored; undefined-permissions-matrix
     expected results were rewritten to functionally-defensible wording in the
     source cases, so no TBD/pending/depends-on-matrix meta text remains.
"""
import csv, json, re, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # build/
ROOT = os.path.dirname(BASE)
CASES_DIR = os.path.join(BASE, "simple-flow", "cases")
REF = os.path.join(ROOT, "testrail-import", "sv5319-testrail-import-MATCHED.csv")
OUT_CSV = os.path.join(ROOT, "testrail-import", "simple-flow-v1-testrail-import.csv")
OUT_XLSX = os.path.join(ROOT, "testrail-import", "simple-flow-v1-testrail-import.xlsx")

FILES = [
    "group-A-settings-completion.json",
    "group-B-receiving-vendor.json",
    "group-C-review-permissions-validation-edge.json",
]

# Story number -> Jira id (from requirements.md "Story <-> Jira ID map").
SV = {1: "SV-7696", 2: "SV-7697", 3: "SV-7698", 4: "SV-7699", 5: "SV-7700",
      6: "SV-7701", 7: "SV-7702", 8: "SV-7703", 9: "SV-7704", 10: "SV-7705",
      11: "SV-7706", 12: "SV-7707", 13: "SV-7708", 14: "SV-7709", 15: "SV-7710",
      16: "SV-7870", 17: "SV-7876", 18: "SV-8353"}

with open(REF, newline="") as f:
    REF_HEADER = next(csv.reader(f))
print("REFERENCE HEADER:", REF_HEADER)

cases = []
for fn in FILES:
    with open(os.path.join(CASES_DIR, fn)) as f:
        cases += json.load(f)
print("Total cases loaded:", len(cases))


def clean(s):
    """Strip internal authoring markers; keep functional content."""
    if not s:
        return s
    # Drop internal case cross-references like " (see SF-SET-03)".
    s = re.sub(r"\s*\(see (?:SF|FD)-[A-Z0-9-]+\)", "", s)
    m = re.match(r"^(\s*\d+\.\s*)EXPECTED PER SPEC:\s*(.*)$", s, re.I | re.S)
    if m:
        rest = m.group(2)
        rest = rest[:1].upper() + rest[1:] if rest else rest
        s = m.group(1) + rest
    return s


def joinlines(lst):
    return "\n".join(clean(x.rstrip()) for x in lst)


def story_numbers(c):
    nums = set()
    a = c["area"]
    sr = c.get("story_ref", "") or ""
    # From area "(Story N)" / "(Stories N/M)"
    for m in re.findall(r"\(Stor(?:y|ies)\s+([0-9/ and]+)\)", a):
        for n in re.findall(r"\d+", m):
            nums.add(int(n))
    # From story_ref: S#-R#, S#-C#, "S# ..." and "Story N"/"Stories N/M"
    for n in re.findall(r"\bS(\d+)-", sr):
        nums.add(int(n))
    for n in re.findall(r"\bS(\d+)\b(?!-)", sr):
        nums.add(int(n))
    for m in re.findall(r"\bStor(?:y|ies)\s+([0-9/ and]+)", sr):
        for n in re.findall(r"\d+", m):
            nums.add(int(n))
    # Tech-story rule refs (TS-R#) -> Story 17
    if re.search(r"\bTS-?R?\d*", sr):
        nums.add(17)
    # Bare review rule refs (R#) -> Story 16. Only when the ref carries no S#-
    # story-rule token (avoids range tails like "S1-R1..R8") and after removing
    # TS-R# tech-story tokens (avoids the R in "TS-R1" being misread).
    tmp = re.sub(r"\b(?:TS|C)-R\d+", "", sr)  # TS-R# = Story 17; C-R# = Story 18 core rules
    if not re.search(r"\bS\d+-", tmp) and re.search(r"\bR\d+\b", tmp):
        nums.add(16)
    return sorted(n for n in nums if n in SV)


def build_refs(c):
    nums = story_numbers(c)
    svs = [SV[n] for n in nums]
    spec = (c.get("story_ref") or "").strip()
    if svs and spec:
        return "{} ({})".format(", ".join(svs), spec)
    if svs:
        return ", ".join(svs)
    return spec


def section_for(c):
    """Leaf area name; API-related cases are routed to an 'API — <area>' section
    (STANDING RULE 4: API content must live under an API-titled section)."""
    area = c["area"].strip()
    if c.get("api_related"):
        return "API — " + area
    return area


rows = []
titles = []
api_sections = set()
api_moved = 0
for c in cases:
    title = clean(c["title"].strip())
    titles.append(title)
    section = section_for(c)
    if c.get("api_related"):
        api_sections.add(section)
        api_moved += 1
    row = [
        title,
        section,
        "Functional",
        c["priority"].strip(),
        joinlines(c.get("preconditions", [])),
        joinlines(c.get("steps", [])),
        joinlines(c.get("expected", [])),
        build_refs(c),
        "",
        "",
    ]
    rows.append(row)

with open(OUT_CSV, "w", newline="") as f:
    w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    w.writerow(REF_HEADER)
    for r in rows:
        w.writerow(r)

print("Wrote CSV:", OUT_CSV, "rows(data):", len(rows))

from collections import Counter
dupes = [t for t, n in Counter(titles).items() if n > 1]
print("Duplicate titles:", dupes if dupes else "NONE")

blob = "\n".join("\t".join(r) for r in rows).lower()
print("VIU occurrences:", blob.count("viu"))
print("'feature flag' occurrences:", blob.count("feature flag"))
print("Sections with 'Simple' prefix:",
      [r[1] for r in rows if r[1].lower().startswith("simple")] or "NONE")
print("API sections created:", len(api_sections), sorted(api_sections))
print("API-flagged cases routed to API sections:", api_moved)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Simple Flow V1"
ws.append(REF_HEADER)
for r in rows:
    ws.append(r)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
for col in range(1, len(REF_HEADER) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center", horizontal="left")

widths = {"Title": 50, "Section": 40, "Type": 12, "Priority": 10,
          "Preconditions": 60, "Steps": 60, "Expected Result": 60,
          "References": 22}
for i, name in enumerate(REF_HEADER, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 10)

wrap = Alignment(wrap_text=True, vertical="top")
for r in range(2, len(rows) + 2):
    for cidx in range(1, len(REF_HEADER) + 1):
        ws.cell(row=r, column=cidx).alignment = wrap

ws.freeze_panes = "A2"
wb.save(OUT_XLSX)
print("Wrote XLSX:", OUT_XLSX)
