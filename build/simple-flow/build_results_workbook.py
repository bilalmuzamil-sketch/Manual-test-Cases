#!/usr/bin/env python3
"""Simple Flow — per-status Results workbook generator.

Builds a per-status Excel workbook (+ CSV mirror) that snapshots the CURRENT
Verify-in-UI (VIU) state of every authored Simple Flow case. Mirrors the style of
the Custom Roles run-331 and Fees & Discounts per-status workbooks.

Source of truth (all recomputed live — no hard-coded counts):
  - build/simple-flow/cases/*.json          (viu_status + area/title/priority/notes)
  - build/simple-flow/gen_blockers.py        (blocker category + sub-bucket + owner + what's-needed)
  - build/simple-flow/bugs-log.md            (bug register + affects-cases; encoded below)

Outputs:
  - build/simple-flow/SimpleFlow_Results.xlsx  (Summary + 5 tabs)
  - build/simple-flow/SimpleFlow_Results.csv   (flat mirror of all cases)

Tabs:
  1. Summary            — total, counts by VIU status + by blocker category (+ sub-buckets), legend.
  2. VIU-Verified       — all VIU-Verified cases.
  3. VIU-Pending        — all VIU-Pending cases + sub-bucket + who-unblocks + what's-needed.
  4. Open-Question      — all Open-Question cases + the question.
  5. Deviation / Bugs   — cases tied to a bug/deviation + which BUG-#, plus a bug register.
  6. TestRail Sync      — what case-content is pushed vs pending; note no SF VIU run exists.
"""
import csv
import json
import os
from collections import Counter, OrderedDict

import gen_blockers  # reuse the authoritative classifier + sub-bucket map

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(BASE, "SimpleFlow_Results.xlsx")
OUT_CSV = os.path.join(BASE, "SimpleFlow_Results.csv")

# --- Bug / deviation register (encoded from bugs-log.md + finding-reclassification.md) --
# Order preserved; "affects" = the cases each bug/deviation touches (per bugs-log affects lists).
BUG_REGISTER = [
    ("BUG-1", "Medium", "OPEN (build-lag / Open Q5)",
     "No 'Create Purchase Orders' toggle / no createPurchaseOrders field — POs always-on. "
     "V2.4 retains the No-PO path, so this is a spec-vs-build gap (build lags V2.4), not a descope.",
     ["SF-SET-03", "SF-COMP-06", "SF-QB-02"]),
    ("BUG-2", "Low", "OPEN (Open Q6)",
     "Save Settings button always enabled (no dirty-state gating).",
     ["SF-SET-13"]),
    ("BUG-3", "Low", "CLOSED — NOT A BUG (Milos Round-2, 2026-07-09)",
     "Mark-Reviewed dialog has no optional review note. Milos removed the note from the "
     "design; v1 is VIN-only. SF-REV-10 expected updated to VIN-only — live now matches.",
     ["SF-REV-10"]),
    ("BUG-4", "Low", "RECLASSIFIED → EXPECTED (shortcut principle, 2026-07-08)",
     "Review sign-off jumps Review → Complete with no distinct 'Reviewed' holding state. "
     "A skipped intermediate state reaching the same end state with no error/corruption = expected.",
     ["SF-REV-08", "SF-REV-11"]),
    ("BUG-5", "High", "OPEN (needs dev/PO ruling)",
     "reviewer != completer rule NOT enforced — a user can Mark-Reviewed (sign off) their "
     "own completed / sent-to-review WO. The one net-new Simple-Flow permission rule is missing.",
     ["SF-PERM-08", "SF-PERM-04", "SF-PERM-07", "SF-REV-09"]),
    ("BUG-6", "High", "OPEN — API gap (v1 ruling: UI gating = PASS; API fix ticket)",
     "WO-completion permission is FE-only at the backend: a Technician (no workOrdersCreateAndEdit) "
     "can complete a WO via simple-complete API (201). FE hides the button; BE does not enforce "
     "(SV-7864 atom-collapse). Ruling 2026-07-09: record 'UI pass / API fail'.",
     ["SF-PERM-06", "SF-PERM-02"]),
    ("BUG-7", "High", "OPEN — API gap (v1 ruling: UI gating = PASS; API fix ticket)",
     "Review sign-off permission (woReviewWorkOrders) is FE-only at the backend: a Technician "
     "can drive review→complete change-status (201). Same FE-vs-BE ruling as BUG-6.",
     ["SF-PERM-07", "SF-REV-09"]),
    ("BUG-8", "Medium", "OPEN",
     "Mileage / VIN / engine-hours completion gates are FE-only (the simple-complete endpoint "
     "does not enforce them; only the wizard does). UI-level blocks are real; BE non-enforcement "
     "is the deviation. (Tech-story + all-lines-approved gates ARE BE-enforced.)",
     ["SF-VAL-01", "SF-VAL-02", "SF-VAL-03", "SF-COMP-05", "SF-COMP-16", "SF-REV-03"]),
    ("BUG-9", "Low", "CLOSED — INTENDED (Milos Round-2, 2026-07-09)",
     "Vendorless 'New Part Request' requires a Category and does not enforce Sell Price. "
     "Milos ruled current behavior is expected for v1 (Category required; Sell optional). "
     "SF-VPART-01/02 expected updated accordingly.",
     ["SF-VPART-01", "SF-VPART-02"]),
    ("BUG-10", "Medium", "RECLASSIFIED → EXPECTED (shortcut principle, 2026-07-08)",
     "No distinct 'Resolve Cores' step in the completion wizard for an inventory core — "
     "resolution is a line-level Ok/Not-OK control; wizard goes Details→Success (no error/"
     "corruption). Receive-dependent special-order-core paths remain blocked by BUG-11.",
     ["SF-CORE-01", "SF-CORE-02", "SF-CORE-03", "SF-CORE-04", "SF-CORE-05",
      "SF-CORE-06", "SF-CORE-07", "SF-CORE-08", "SF-CORE-09", "SF-CORE-10"]),
    ("BUG-11", "High → Low", "OPEN — workaround exists (Bulk Receive path works)",
     "WO-originated PO receive returns HTTP 500 on the LEGACY Accept-Delivery path "
     "(POST /api/inventory/orders/accept). The new Bulk Receive pipeline "
     "(receive-requested-parts) works (200), so the round-trip is achievable there; the "
     "legacy single-PO Accept-Delivery 500 should still be fixed.",
     ["SF-COMP-13", "SF-COMP-19", "SF-VAL-05", "SF-VAL-06", "SF-PNFIX-02", "SF-PNFIX-03",
      "SF-PNFIX-04", "SF-PNFIX-05", "SF-PNFIX-06", "SF-RCV-08", "SF-VPART-07", "SF-REV-04",
      "SF-REV-14", "SF-CORE-03", "SF-CORE-04", "SF-CORE-05", "SF-CORE-06", "SF-CORE-07"]),
]

# case_id -> list of BUG-# it is tied to
CASE_TO_BUGS = OrderedDict()
for bug_id, _sev, _status, _desc, affects in BUG_REGISTER:
    for cid in affects:
        CASE_TO_BUGS.setdefault(cid, []).append(bug_id)


TESTRAIL_MAP_CSV = os.path.join(BASE, "testrail-id-map.csv")
TESTRAIL_VIEW_URL = "https://shopview.testrail.io/index.php?/cases/view/%s"


def load_testrail_map():
    """sf_id -> TestRail numeric case_id (str). Column ID holds the case id."""
    m = {}
    with open(TESTRAIL_MAP_CSV, newline="") as f:
        for row in csv.DictReader(f):
            case_id = (row.get("ID") or row.get("testrail_case_id") or "").strip()
            sf_id = (row.get("sf_id") or "").strip()
            if sf_id and case_id:
                m[sf_id] = case_id
    return m


def testrail_cells(sf_id, tr_map):
    """Return (tr_id_display, tr_link_url) for a case; blanks if unmapped."""
    cid = tr_map.get(sf_id)
    if not cid:
        return "", ""
    return "C%s" % cid, TESTRAIL_VIEW_URL % cid


def load_cases():
    cases = []
    for fn in gen_blockers.FILES:
        cases += json.load(open(os.path.join(gen_blockers.CASES_DIR, fn)))
    return cases


def blocker_for(c):
    """Return (category_display, subbucket, owner, needs)."""
    cls = gen_blockers.classify(c)
    cat = cls["category"]
    cat_disp = {
        "READY (VIU-Verified)": "READY (VIU-Verified)",
        "DEV NOT BUILT": "BLOCKED — DEV NOT BUILT",
        "VIU PENDING (QA)": "BLOCKED — VIU PENDING (QA)",
        "MILOS ANSWER": "BLOCKED — MILOS ANSWER",
        "BUG/RULING": "BLOCKED — BUG/RULING",
    }.get(cat, cat)
    if cat == "VIU PENDING (QA)":
        sb, sbnote = gen_blockers.SUBBUCKET.get(c["id"], ("reachable-now", gen_blockers._REACHABLE))
    else:
        sb, sbnote = "—", ""
    return cat_disp, sb, cls["owner"], cls["needs"], sbnote


def main():
    cases = load_cases()
    assert len(cases) == 187, "expected 187 cases, got %d" % len(cases)

    tr_map = load_testrail_map()
    tr_blanks = [c["id"] for c in cases if c["id"] not in tr_map]
    n_mapped = len(cases) - len(tr_blanks)

    # ---- recompute counts ----
    def viu_bucket(vs):
        # Fold the V2.4 Δ (2026-07-13) retest strings into one "Pending / Retest" bucket.
        return "Pending / Retest" if (vs or "").startswith("Pending / Retest") else (vs or "?")
    viu_counts = Counter(viu_bucket(c.get("viu_status", "")) for c in cases)
    cat_counts = Counter()
    sub_counts = Counter()
    per_case = []
    for c in cases:
        cat_disp, sb, owner, needs, sbnote = blocker_for(c)
        cat_counts[cat_disp] += 1
        if cat_disp == "BLOCKED — VIU PENDING (QA)":
            sub_counts[sb] += 1
        per_case.append({
            "id": c["id"], "area": c["area"], "title": c["title"].strip(),
            "priority": c.get("priority", ""), "viu": c.get("viu_status", ""),
            "cat": cat_disp, "sub": sb, "owner": owner, "needs": needs,
            "subnote": sbnote, "notes": (c.get("notes") or "").strip(),
            "bugs": CASE_TO_BUGS.get(c["id"], []),
        })

    n_dev = sum(1 for pc in per_case if pc["bugs"])

    # ---------------- Styling helpers ----------------
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL = PatternFill("solid", fgColor="305496")
    WRAP = Alignment(wrap_text=True, vertical="top")
    TITLE_FONT = Font(bold=True, size=14)
    THIN = Side(style="thin", color="D9D9D9")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    VIU_FILL = {
        "VIU-Verified": "C6EFCE",   # green
        "VIU-Pending": "FFF2CC",    # amber
        "Open-Question": "D9E1F2",  # blue
    }
    CAT_FILL = {
        "READY (VIU-Verified)": "C6EFCE",
        "BLOCKED — DEV NOT BUILT": "F4CCCC",
        "BLOCKED — VIU PENDING (QA)": "FFF2CC",
        "BLOCKED — MILOS ANSWER": "D9E1F2",
        "BLOCKED — BUG/RULING": "FCE4D6",
    }
    SUB_FILL = {"reachable-now": "C6EFCE", "needs-data": "FFF2CC", "needs-account": "F4CCCC"}

    STATUS_COLS = ["TestRail ID", "TestRail Link", "SF ID (Case ID)", "Area",
                   "Title", "Priority", "VIU Status", "Blocker/Reason", "Notes"]
    LINK_FONT = Font(color="0563C1", underline="single")

    def style_header(ws, ncols, row=1):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
            cell.border = BORDER

    def set_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # ================= 1. Summary =================
    ss = wb.active
    ss.title = "Summary"
    r = 1
    ss.cell(row=r, column=1, value="Simple Flow (Simple Mode / SV-7301) — VIU Results Snapshot")
    ss.cell(row=r, column=1).font = TITLE_FONT
    r += 1
    ss.cell(row=r, column=1, value="Generated from build/simple-flow/cases/*.json (all counts recomputed live). "
                                   "Date of snapshot: 2026-07-09 (BATCH 8).")
    r += 2
    ss.cell(row=r, column=1, value="Total authored cases")
    ss.cell(row=r, column=1).font = Font(bold=True)
    ss.cell(row=r, column=2, value=len(cases))
    ss.cell(row=r, column=2).font = Font(bold=True)
    r += 2

    # By VIU status
    ss.cell(row=r, column=1, value="Count by VIU status")
    ss.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    hdr = r
    for i, h in enumerate(["VIU status", "Count", "Meaning"], start=1):
        ss.cell(row=hdr, column=i, value=h)
    style_header(ss, 3, hdr)
    r += 1
    VIU_MEAN = {
        "VIU-Verified": "Behavior confirmed live in the UI on sv7301.",
        "VIU-Pending": "Not yet driven / needs seed data, an account, or a bug fix (see VIU-Pending tab).",
        "Open-Question": "Behavior depends on a Product Owner (Milos) decision (see Open-Question tab).",
        "Pending / Retest": "Expected behavior changed by the V2.4 deltas (2026-07-13); needs a live re-VIU (see Pending / Retest tab).",
    }
    for st in ["VIU-Verified", "VIU-Pending", "Open-Question", "Pending / Retest"]:
        ss.cell(row=r, column=1, value=st)
        ss.cell(row=r, column=2, value=viu_counts.get(st, 0))
        ss.cell(row=r, column=3, value=VIU_MEAN[st])
        fill = PatternFill("solid", fgColor=VIU_FILL.get(st, "FCE4D6"))
        for c in range(1, 4):
            ss.cell(row=r, column=c).fill = fill
        r += 1
    ss.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ss.cell(row=r, column=2, value=sum(viu_counts.values())).font = Font(bold=True)
    r += 1
    ss.cell(row=r, column=1, value="Cases tied to a bug / deviation")
    ss.cell(row=r, column=1).font = Font(bold=True)
    ss.cell(row=r, column=2, value=n_dev).font = Font(bold=True)
    ss.cell(row=r, column=3, value="Cross-cut metric (not a VIU status) — see the Deviation / Bugs tab.")
    r += 1
    ss.cell(row=r, column=1, value="Current headline tally (2026-07-13): VIU-Verified %d / VIU-Pending %d / "
                                   "Open-Question %d / Pending-Retest %d = %d. Pending-Retest = the V2.4 Δ1-Δ4 "
                                   "cases (expected changed 2026-07-13) awaiting a live re-VIU. DEV-NOT-BUILT is 0."
                                   % (viu_counts.get("VIU-Verified", 0), viu_counts.get("VIU-Pending", 0),
                                      viu_counts.get("Open-Question", 0), viu_counts.get("Pending / Retest", 0),
                                      sum(viu_counts.values())))
    ss.cell(row=r, column=1).font = Font(italic=True)
    r += 2

    # TestRail mapping coverage
    ss.cell(row=r, column=1, value="TestRail ID coverage")
    ss.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    ss.cell(row=r, column=1, value="Cases mapped to a TestRail ID")
    ss.cell(row=r, column=2, value=n_mapped)
    r += 1
    ss.cell(row=r, column=1, value="Cases with NO TestRail ID (blank)")
    ss.cell(row=r, column=2, value=len(tr_blanks))
    if tr_blanks:
        ss.cell(row=r, column=3, value="Blank: " + ", ".join(tr_blanks) +
                " — add after the next TestRail sync (IDs not guessed).")
        ss.cell(row=r, column=2).fill = PatternFill("solid", fgColor="FCE4D6")
    else:
        ss.cell(row=r, column=3, value="All authored cases have a TestRail ID (C<id>) + view link.")
        ss.cell(row=r, column=2).fill = PatternFill("solid", fgColor="C6EFCE")
    r += 1
    ss.cell(row=r, column=1, value="Each status tab now carries: TestRail ID (C<id>) + a clickable "
                                   "TestRail Link (https://shopview.testrail.io/index.php?/cases/view/<id>) "
                                   "alongside the internal SF ID.")
    ss.cell(row=r, column=1).font = Font(italic=True)
    r += 2

    # By blocker category
    ss.cell(row=r, column=1, value="Count by blocker category")
    ss.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    hdr = r
    for i, h in enumerate(["Blocker category", "Count", "Owner"], start=1):
        ss.cell(row=hdr, column=i, value=h)
    style_header(ss, 3, hdr)
    r += 1
    CAT_ORDER = ["READY (VIU-Verified)", "BLOCKED — DEV NOT BUILT",
                 "BLOCKED — VIU PENDING (QA)", "BLOCKED — MILOS ANSWER",
                 "BLOCKED — BUG/RULING"]
    OWNER = {"READY (VIU-Verified)": "— (ready to upload)",
             "BLOCKED — DEV NOT BUILT": "Dev team",
             "BLOCKED — VIU PENDING (QA)": "QA",
             "BLOCKED — MILOS ANSWER": "Milos (PO)",
             "BLOCKED — BUG/RULING": "Dev / PO ruling"}
    for cat in CAT_ORDER:
        ss.cell(row=r, column=1, value=cat)
        ss.cell(row=r, column=2, value=cat_counts.get(cat, 0))
        ss.cell(row=r, column=3, value=OWNER[cat])
        fill = PatternFill("solid", fgColor=CAT_FILL[cat])
        for c in range(1, 4):
            ss.cell(row=r, column=c).fill = fill
        r += 1
    ss.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ss.cell(row=r, column=2, value=sum(cat_counts.values())).font = Font(bold=True)
    r += 2

    # VIU-PENDING (QA) sub-buckets
    ss.cell(row=r, column=1, value="BLOCKED — VIU PENDING (QA) — by sub-bucket")
    ss.cell(row=r, column=1).font = Font(bold=True, size=12)
    r += 1
    hdr = r
    for i, h in enumerate(["VIU sub-bucket", "Count", "Meaning"], start=1):
        ss.cell(row=hdr, column=i, value=h)
    style_header(ss, 3, hdr)
    r += 1
    SUB_MEAN = {
        "reachable-now": "admin+tech + normal data; just needs another VIU pass (no new inputs).",
        "needs-data": "needs a data state not seedable via the app (cores / WO-PO receive / QuickBooks / invoiced-WO).",
        "needs-account": "needs a role account we don't have (Office / Service Manager / Foreman ...).",
    }
    for sb in ["reachable-now", "needs-data", "needs-account"]:
        ss.cell(row=r, column=1, value=sb)
        ss.cell(row=r, column=2, value=sub_counts.get(sb, 0))
        ss.cell(row=r, column=3, value=SUB_MEAN[sb])
        fill = PatternFill("solid", fgColor=SUB_FILL[sb])
        for c in range(1, 3):
            ss.cell(row=r, column=c).fill = fill
        r += 1
    ss.cell(row=r, column=1, value="TOTAL VIU PENDING (QA)").font = Font(bold=True)
    ss.cell(row=r, column=2, value=sum(sub_counts.values())).font = Font(bold=True)
    r += 2

    # Legend
    ss.cell(row=r, column=1, value="Legend").font = Font(bold=True, size=12)
    r += 1
    legend = [
        ("VIU-Verified / READY", "Confirmed live in the UI; ready to upload to a run."),
        ("VIU-Pending", "Not yet verified — see the sub-bucket + what's-needed on the VIU-Pending tab."),
        ("Open-Question", "Verdict awaits a Product Owner (Milos) decision — see the Open-Question tab."),
        ("BLOCKED — BUG/RULING", "VIU-observed but the pass/fail verdict hangs on a dev/PO ruling (see Deviation / Bugs)."),
        ("Deviation / Bug", "A case whose live behavior differs from spec; tracked via a BUG-# in bugs-log.md."),
        ("DEV NOT BUILT", "Feature not yet built. Now 0 — Stories 7/8/9/14 are built & VIU-verified."),
    ]
    for k, v in legend:
        ss.cell(row=r, column=1, value=k).font = Font(bold=True)
        ss.cell(row=r, column=2, value=v)
        r += 1

    set_widths(ss, [34, 70, 40])
    for row in ss.iter_rows():
        for cell in row:
            if cell.alignment is None or cell.alignment.wrap_text is not True:
                cur = cell.alignment
                cell.alignment = Alignment(wrap_text=True, vertical="top",
                                           horizontal=cur.horizontal if cur else None)
    ss.freeze_panes = "A5"

    # ---------- shared status-tab builder ----------
    def status_tab(name, rows, extra_col=None):
        """rows: list of per_case dicts. extra_col: optional (header, fn) appended."""
        ws = wb.create_sheet(name)
        cols = list(STATUS_COLS)
        if extra_col:
            cols = STATUS_COLS[:-1] + [extra_col[0]] + [STATUS_COLS[-1]]
        ws.append(cols)
        style_header(ws, len(cols))
        link_col = cols.index("TestRail Link") + 1
        viu_col = cols.index("VIU Status") + 1
        for pc in rows:
            reason = pc["cat"]
            if pc["cat"] == "BLOCKED — VIU PENDING (QA)":
                reason = "%s · sub-bucket: %s · who: %s · needs: %s" % (
                    pc["cat"], pc["sub"], pc["owner"], pc["needs"])
            elif pc["cat"] == "BLOCKED — MILOS ANSWER":
                reason = "%s · %s" % (pc["cat"], pc["needs"])
            elif pc["cat"] == "BLOCKED — BUG/RULING":
                reason = "%s · %s" % (pc["cat"], pc["needs"])
            tr_id, tr_link = testrail_cells(pc["id"], tr_map)
            notes = pc["notes"]
            if not tr_id:
                notes = ("[No TestRail ID in map — add after next TestRail sync] " + notes).strip()
            row = [tr_id, tr_link, pc["id"], pc["area"], pc["title"], pc["priority"],
                   pc["viu"], reason]
            if extra_col:
                row.append(extra_col[1](pc))
            row.append(notes)
            ws.append(row)
        # style body
        ncols = len(cols)
        for ridx in range(2, len(rows) + 2):
            vcell = ws.cell(row=ridx, column=viu_col)  # VIU Status
            vfill = VIU_FILL.get(vcell.value)
            for cidx in range(1, ncols + 1):
                cell = ws.cell(row=ridx, column=cidx)
                cell.alignment = WRAP
                cell.border = BORDER
            if vfill:
                vcell.fill = PatternFill("solid", fgColor=vfill)
            lcell = ws.cell(row=ridx, column=link_col)
            if lcell.value:
                lcell.hyperlink = lcell.value
                lcell.font = LINK_FONT
        # widths
        base_w = {"TestRail ID": 12, "TestRail Link": 52, "SF ID (Case ID)": 16,
                  "Area": 34, "Title": 55, "Priority": 10, "VIU Status": 14,
                  "Blocker/Reason": 60, "Notes": 60}
        widths = [base_w.get(c, 30) for c in cols]
        set_widths(ws, widths)
        ws.freeze_panes = "A2"
        return ws

    # ================= 2. VIU-Verified =================
    verified = [pc for pc in per_case if pc["viu"] == "VIU-Verified"]
    status_tab("VIU-Verified", verified)

    # ================= 3. VIU-Pending =================
    pending = [pc for pc in per_case if pc["viu"] == "VIU-Pending"]
    status_tab("VIU-Pending", pending)

    # ================= 4. Open-Question =================
    openq = [pc for pc in per_case if pc["viu"] == "Open-Question"]
    status_tab("Open-Question", openq,
               extra_col=("Open Question / decision needed", lambda pc: pc["needs"]))

    # ========= 4b. Pending / Retest (V2.4 Δ 2026-07-13) =========
    retest = [pc for pc in per_case if (pc["viu"] or "").startswith("Pending / Retest")]
    status_tab("Pending _ Retest", retest)

    # ================= 5. Deviation / Bugs =================
    ws = wb.create_sheet("Deviation _ Bugs")
    ws.cell(row=1, column=1, value="Cases tied to a bug or deviation (from bugs-log.md)")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append([])
    hdr_row = 3
    dcols = ["TestRail ID", "TestRail Link", "SF ID (Case ID)", "Area", "Title",
             "Priority", "VIU Status", "BUG-#", "Notes"]
    for i, h in enumerate(dcols, start=1):
        ws.cell(row=hdr_row, column=i, value=h)
    style_header(ws, len(dcols), hdr_row)
    dev_rows = [pc for pc in per_case if pc["bugs"]]
    ridx = hdr_row + 1
    for pc in dev_rows:
        tr_id, tr_link = testrail_cells(pc["id"], tr_map)
        notes = pc["notes"]
        if not tr_id:
            notes = ("[No TestRail ID in map — add after next TestRail sync] " + notes).strip()
        ws.cell(row=ridx, column=1, value=tr_id)
        lcell = ws.cell(row=ridx, column=2, value=tr_link)
        if tr_link:
            lcell.hyperlink = tr_link
            lcell.font = LINK_FONT
        ws.cell(row=ridx, column=3, value=pc["id"])
        ws.cell(row=ridx, column=4, value=pc["area"])
        ws.cell(row=ridx, column=5, value=pc["title"])
        ws.cell(row=ridx, column=6, value=pc["priority"])
        ws.cell(row=ridx, column=7, value=pc["viu"])
        ws.cell(row=ridx, column=8, value=", ".join(pc["bugs"]))
        ws.cell(row=ridx, column=9, value=notes)
        vfill = VIU_FILL.get(pc["viu"])
        for cidx in range(1, len(dcols) + 1):
            cell = ws.cell(row=ridx, column=cidx)
            cell.alignment = WRAP
            cell.border = BORDER
        if vfill:
            ws.cell(row=ridx, column=7).fill = PatternFill("solid", fgColor=vfill)
        ridx += 1

    # --- Bug register block ---
    ridx += 1
    ws.cell(row=ridx, column=1, value="Bug / Deviation Register").font = Font(bold=True, size=13)
    ridx += 1
    reg_hdr = ridx
    rcols = ["BUG-#", "Severity", "Status", "Description", "Affected cases"]
    for i, h in enumerate(rcols, start=1):
        ws.cell(row=reg_hdr, column=i, value=h)
    style_header(ws, len(rcols), reg_hdr)
    ridx += 1
    STATUS_FILL = {"OPEN": "FCE4D6", "CLOSED": "C6EFCE", "EXPECTED": "C6EFCE"}
    for bug_id, sev, status, desc, affects in BUG_REGISTER:
        ws.cell(row=ridx, column=1, value=bug_id)
        ws.cell(row=ridx, column=2, value=sev)
        ws.cell(row=ridx, column=3, value=status)
        ws.cell(row=ridx, column=4, value=desc)
        ws.cell(row=ridx, column=5, value=", ".join(affects))
        # colour status cell
        fill = None
        if status.startswith("CLOSED") or "EXPECTED" in status:
            fill = "C6EFCE"
        elif status.startswith("OPEN"):
            fill = "FCE4D6"
        for cidx in range(1, len(rcols) + 1):
            cell = ws.cell(row=ridx, column=cidx)
            cell.alignment = WRAP
            cell.border = BORDER
        if fill:
            ws.cell(row=ridx, column=3).fill = PatternFill("solid", fgColor=fill)
        ridx += 1
    ridx += 1
    ws.cell(row=ridx, column=1,
            value="Note: BUG-5/6/7/8/11 are OPEN (BUG-6/7 carry a v1 'UI-pass / API-fail' ruling; "
                  "BUG-11 has a working Bulk-Receive path). BUG-3 and BUG-9 are CLOSED (Milos "
                  "Round-2: review-note descoped; Category-required/Sell-optional intended). "
                  "BUG-4 and BUG-10 are reclassified EXPECTED under the shortcut principle.")
    ws.cell(row=ridx, column=1).font = Font(italic=True)

    set_widths(ws, [12, 52, 16, 34, 55, 10, 14, 14, 60])
    ws.freeze_panes = "A4"

    # ================= 6. TestRail Sync Status =================
    ts = wb.create_sheet("TestRail Sync Status")
    ts.cell(row=1, column=1, value="Simple Flow — TestRail Sync Status").font = TITLE_FONT
    ts.append([])
    row = 3
    ts.cell(row=row, column=1, value="Environment").font = Font(bold=True)
    ts.cell(row=row, column=2, value="Project 1 · Suite 1 'Master' on https://shopview.testrail.io · "
                                     "parent section 4058 'Simple Flow'; leaf sections 4059–4090.")
    row += 2
    ts.cell(row=row, column=1, value="Item").font = Font(bold=True)
    ts.cell(row=row, column=2, value="State").font = Font(bold=True)
    ts.cell(row=row, column=3, value="Detail").font = Font(bold=True)
    style_header(ts, 3, row)
    row += 1
    sync_rows = [
        ("All 162 authored cases exist in TestRail", "PUSHED / VERIFIED",
         "1:1 title match (159 originally + SF-VMIS-07 + SF-RCV-10 adds). API-flagged cases moved to API-titled sections (4089/4090)."),
        ("SV-8183 permissions batch (case content)", "PUSHED / VERIFIED",
         "13 SV-8183 permission cases compared field-by-field — already identical (no-op) in the 2026-07-08 sync."),
        ("V2.4 batch (case content)", "PUSHED / VERIFIED",
         "18 in-place update_case + 2 add_case (SF-VMIS-07, SF-RCV-10) on 2026-07-08; all verified by re-fetch (testrail-push-v2.4-log.md)."),
        ("Milos Round-2 batch (case content)", "PUSHED / VERIFIED",
         "5 update_case on 2026-07-09 (SF-REV-10, SF-TECH-08, SF-VPART-01, SF-VPART-02, SF-PERM-06); all verified (milos-round2-mapping.md)."),
        ("SF-WOP-02 expected refinement (Bulk Receive)", "PENDING — HELD FOR APPROVAL",
         "Expected refined LOCALLY only (count-click opens Bulk Receive, not legacy Accept Delivery — OBS-5). Not yet pushed to TestRail; awaiting approval."),
        ("VIU pass/fail RESULTS in a TestRail run", "NOT IN TESTRAIL",
         "No Simple Flow TestRail RUN exists. All VIU pass/fail status lives ONLY in the case JSONs + this workbook + the blockers tracker. Nothing was logged to a run."),
        ("Cosmetic diffs (HTML numbering / ref comma-spacing)", "INTENTIONALLY LEFT",
         "19 cases HTML <ol> vs plain numbering; 13 cases ref comma-spacing. Visible text identical; not rewritten (flag if a normalize pass is wanted)."),
    ]
    for item, state, detail in sync_rows:
        ts.cell(row=row, column=1, value=item)
        ts.cell(row=row, column=2, value=state)
        ts.cell(row=row, column=3, value=detail)
        if "PUSHED" in state or "VERIFIED" in state:
            fill = "C6EFCE"
        elif "PENDING" in state or "HELD" in state:
            fill = "FFF2CC"
        elif "NOT IN" in state:
            fill = "FCE4D6"
        else:
            fill = "D9E1F2"
        for c in range(1, 4):
            cell = ts.cell(row=row, column=c)
            cell.alignment = WRAP
            cell.border = BORDER
        ts.cell(row=row, column=2).fill = PatternFill("solid", fgColor=fill)
        row += 1
    row += 1
    ts.cell(row=row, column=1,
            value="Standing rule: NEVER write to TestRail (cases/runs/results) without explicit user "
                  "permission. This workbook is documentation only — it does not touch TestRail.")
    ts.cell(row=row, column=1).font = Font(italic=True)
    set_widths(ts, [44, 26, 80])
    ts.freeze_panes = "A6"

    wb.save(OUT_XLSX)
    print("Wrote", OUT_XLSX)

    # ---------------- CSV mirror (flat, all cases) ----------------
    with open(OUT_CSV, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["TestRail ID", "TestRail Link", "SF ID (Case ID)", "Area",
                    "Title", "Priority", "VIU Status",
                    "Blocker Category", "VIU Sub-bucket", "Who Unblocks",
                    "What's Needed", "BUG-#", "Notes"])
        for pc in per_case:
            tr_id, tr_link = testrail_cells(pc["id"], tr_map)
            notes = pc["notes"]
            if not tr_id:
                notes = ("[No TestRail ID in map — add after next TestRail sync] " + notes).strip()
            w.writerow([tr_id, tr_link, pc["id"], pc["area"], pc["title"], pc["priority"], pc["viu"],
                        pc["cat"], pc["sub"] if pc["sub"] != "—" else "",
                        pc["owner"] if pc["owner"] != "—" else "",
                        pc["needs"] if pc["needs"] != "None — VIU-verified; uploadable now." else "",
                        ", ".join(pc["bugs"]), notes])
    print("Wrote", OUT_CSV)

    # ---------------- console recap ----------------
    print("\n=== RECOMPUTED COUNTS ===")
    print("Total:", len(cases))
    print("VIU status:", dict(viu_counts))
    print("Blocker category:", dict(cat_counts))
    print("VIU-PENDING sub-buckets:", dict(sub_counts))
    print("Cases tied to a bug/deviation:", n_dev)
    print("TestRail IDs mapped: %d / %d (blank: %d)" % (n_mapped, len(cases), len(tr_blanks)))
    if tr_blanks:
        print("  BLANK (no TestRail ID):", ", ".join(tr_blanks))
    print("  spot-check SF-VMIS-07 ->", tr_map.get("SF-VMIS-07"),
          "| SF-RCV-10 ->", tr_map.get("SF-RCV-10"))
    assert sum(viu_counts.values()) == 187
    assert sum(cat_counts.values()) == 187


if __name__ == "__main__":
    main()
