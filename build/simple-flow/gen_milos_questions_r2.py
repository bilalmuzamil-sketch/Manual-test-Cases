#!/usr/bin/env python3
"""Generate the 'Open Questions for Milos — Round 2' deliverable (md + xlsx).

TRIMMED 2026-07-08 under the Simple Flow shortcut principle: the two questions now
classified EXPECTED (the "distinct Reviewed state" question and the "Resolve-Cores
step" question) are REMOVED — a flow-skip that reaches the same end state with no
error/corruption is EXPECTED, not a PO question. See finding-reclassification.md.
Remaining questions renumbered 1..5.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
MD = os.path.join(BASE, "OpenQuestions-for-Milos-Round2.md")
XLSX = os.path.join(BASE, "OpenQuestions-for-Milos-Round2.xlsx")

FEATURE = ("Simple Mode / Streamlined Work Order Completion & Receiving "
           "(Epic SV-7301)")
INTRO = (
    "Thank you for answering our first round of questions. These are QA's "
    "remaining open items after Round 1 - the few you left unanswered or that came "
    "back ambiguous, plus a few new findings that were not on the first sheet. "
    "(Two earlier items - the 'distinct Reviewed state' and the 'Resolve-Cores "
    "wizard step' - have been withdrawn: they are now treated as expected Simple "
    "Mode shortcuts, since each reaches the same end state faster with no error or "
    "data loss.) Each question is written to stand on its own (with a concrete "
    "example where it helps), so no access to our test repository is needed to "
    "answer. Please add your answer in the last column for each row.")

COLUMNS = ["#", "Original Q ref", "Topic", "Question", "Why it matters",
           "Affected cases", "Answer (for Milos)"]

# (Original Q ref, Topic, Question, Why it matters, Affected cases)
Q = [
    ("Round-1 Q7",
     "Mark-Reviewed dialog - review note field",
     "The Mark Reviewed dialog captures only the vehicle identification number "
     "(VIN); there is no optional 'review note' field. We are treating this as an "
     "intended Simple Mode simplification (so we will set the expectation to 'no "
     "note field') - can you confirm that is correct for v1, or do you want the "
     "optional note field added back? Example: today a reviewer signs off by "
     "entering the VIN and confirming - there is nowhere to type a free-text note "
     "about the review.",
     "Just a quick confirmation so we lock in the right expectation. We assume 'no "
     "note field' (intended simplification) unless you say otherwise.",
     "SF-REV-10"),
    ("Round-1 Q9",
     "Tech-story entry points (Story 17 vs older wording)",
     "Does Story 17 (tech story captured via the completion modal at completion "
     "plus inline on the line) fully supersede the older 'tech story lives on the "
     "line only' behavior (S15-R2)? Please confirm the authoritative tech-story "
     "entry points for v1. Example: our live testing showed the completion modal "
     "gate for the tech story working, so we believe Story 17 is current - we just "
     "need your explicit confirmation.",
     "Confirms whether the tech-story gate-modal cases are authoritative, or "
     "whether we should revert to line-only behavior.",
     "SF-TECH-*"),
    ("Round-1 Q2",
     "Inventory lifecycle on completion (data-integrity)",
     "With the No-PO / skip path removed and a purchase order always created, when "
     "a work order is completed do in-stock / inventory parts still decrement "
     "on-hand inventory and post to Part History at completion (i.e., the real part "
     "lifecycle still runs)? Example: a work order uses 2 units of a part already "
     "in stock; on completion we expect on-hand to drop by 2 and a Part History "
     "movement to be recorded.",
     "This is a data-integrity invariant. Your Round-1 answer confirmed the No-PO "
     "path is removed but did not explicitly confirm that normal completion still "
     "moves inventory and writes history. (Under our shortcut rule, a faster "
     "completion is fine - but only if inventory and Part History stay correct.)",
     "SF-COMP-07, SF-QB-01"),
    ("New (BUG-9)",
     "'New Part Request' required fields (Category / Sell Price)",
     "The vendorless 'New Part Request' sub-form currently requires a Category, but "
     "spec S5-R1 lists only description + quantity + sell price. Is requiring a "
     "Category intended for v1 (so we add 'Category required' to the expected "
     "result), or is it a bug? And, separately, should Sell Price be a required "
     "field? Example: today, entering only description + quantity + sell price does "
     "not save - the form blocks with 'Category is a required field'; meanwhile "
     "Sell Price is not enforced as required at all.",
     "Determines the correct required-field set for the part-request cases; the "
     "spec and the build currently disagree in both directions (Category is "
     "required but not in spec; Sell Price is in spec but not enforced). Note: this "
     "is an ADDED required field, not a skipped step, so it is not covered by the "
     "Simple Mode shortcut rule.",
     "SF-VPART-01, SF-VPART-02"),
    ("New (SF-PERM-06 / BUG-6 / BUG-7)",
     "Backend enforcement of completion / review-sign-off permissions",
     "We found that the backend does not enforce the completion / review-sign-off "
     "permission atoms - a user without the permission (e.g. a Technician) can "
     "complete a work order or sign off a review directly via the API; only the UI "
     "hides the buttons. Is this intended (front-end-only gating, consistent with "
     "the Work-Order Create-&-Edit atom-collapse in SV-7864), or should the backend "
     "enforce these permissions? Example: a Technician who cannot see the 'Complete "
     "Work Order' button in the UI can still call the completion endpoint and the "
     "work order completes successfully (HTTP 201) instead of being rejected (HTTP "
     "403).",
     "Decides whether the permission cases pass (FE-only gating is acceptable) or "
     "fail (a backend enforcement gap). This resolves the SV-8183 'backend enforces "
     "the atoms' vs SV-7864 'atom-collapse' contradiction. This is an enforcement "
     "question, not a flow-skip.",
     "SF-PERM-06"),
]

# ---------------------------------------------------------------- markdown
lines = []
lines.append("# Open Questions for Milos — Round 2 — Simple Mode (Streamlined "
             "Work Order Completion & Receiving)")
lines.append("")
lines.append(f"**Feature:** {FEATURE}  ")
lines.append("**Raised by:** ShopView QA  ")
lines.append("**For:** Milos Vasic (Product Owner / spec author)")
lines.append("")
lines.append(INTRO)
lines.append("")
lines.append("| # | Original Q ref | Topic | Question | Why it matters | "
             "Affected cases | Answer (for Milos) |")
lines.append("|---|---|---|---|---|---|---|")


def md_cell(s):
    return s.replace("|", "\\|").replace("\n", "<br>")


for i, (ref, topic, q, why, aff) in enumerate(Q, start=1):
    lines.append("| {} | {} | {} | {} | {} | {} | |".format(
        i, md_cell(ref), md_cell(topic), md_cell(q), md_cell(why), md_cell(aff)))
lines.append("")
with open(MD, "w", encoding="utf-8") as fh:
    fh.write("\n".join(lines))

# ---------------------------------------------------------------- xlsx
DARK = "1F4E78"
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor=DARK)
ALT = PatternFill("solid", fgColor="F2F5FA")
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="top", horizontal="center")

wb = Workbook()
ws = wb.active
ws.title = "Open Questions R2"

ws.append(COLUMNS)  # row 1 header
for c in range(1, len(COLUMNS) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = BORDER

for i, (ref, topic, q, why, aff) in enumerate(Q, start=1):
    ws.append([i, ref, topic, q, why, aff, ""])
    r = 1 + i
    alt = (i % 2 == 1)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = WRAP
        cell.border = BORDER
        if alt:
            cell.fill = ALT
    ws.cell(row=r, column=1).alignment = WRAPC

widths = [5, 22, 30, 70, 45, 22, 30]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"

# About sheet
about = wb.create_sheet("About")
about_lines = [
    "Open Questions for Milos — Round 2",
    None,
    f"Feature: {FEATURE}",
    "Raised by: ShopView QA",
    "For: Milos Vasic (Product Owner / spec author)",
    None,
    "These are QA's remaining open items after Round 1 — the items left unanswered or",
    "ambiguous, plus new findings that were not on the first sheet. Two earlier items",
    "(the 'distinct Reviewed state' and the 'Resolve-Cores wizard step') have been",
    "withdrawn: they are now treated as expected Simple Mode shortcuts (each reaches",
    "the same end state faster with no error or data loss). Each question is",
    "self-contained (with a concrete example where it helps); no access to the QA test",
    "repository is needed to answer. Please add your answer in the last column of the",
    "'Open Questions R2' tab for each row.",
]
for txt in about_lines:
    about.append([txt])
about["A1"].font = Font(bold=True, size=14, color=DARK)
about.column_dimensions["A"].width = 90
for r in range(1, about.max_row + 1):
    about.cell(row=r, column=1).alignment = Alignment(wrap_text=False, vertical="top")

wb.save(XLSX)

print("Questions:", len(Q))
print("Saved:", MD)
print("Saved:", XLSX)
