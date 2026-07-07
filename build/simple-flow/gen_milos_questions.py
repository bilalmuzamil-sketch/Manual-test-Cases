#!/usr/bin/env python3
"""Generate the standalone 'Open Questions for Milos' deliverable (md + xlsx).

Milos Vasic is the Product Owner / spec author for Simple Mode (Epic SV-7301).
These are QA's open questions on the spec that REMAIN after SV-8183 (which
defined the Simple-Flow permissions model). Questions SV-8183 already answered
are intentionally excluded.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
MD = os.path.join(BASE, "OpenQuestions-for-Milos.md")
XLSX = os.path.join(BASE, "OpenQuestions-for-Milos.xlsx")

FEATURE = ("Simple Mode / Streamlined Work Order Completion & Receiving "
           "(Epic SV-7301)")
INTRO = (
    "These are the QA team's open questions on the Simple Mode spec. They are the "
    "items that are still unresolved AFTER story SV-8183 (which defined the Simple "
    "Flow permissions model). The permissions/role questions SV-8183 already "
    "answered are intentionally left out. Please add your answer in the last "
    "column for each row - the questions are written to stand on their own, so no "
    "access to our test repository is needed to answer them.")

COLUMNS = ["#", "Area / Story", "Question", "Context / why it matters",
           "Options (if any)", "Affected cases", "Answer (for Milos)"]

# (Area/Story, Question, Context, Options, Affected cases)
Q = [
    ("Story 1 & 16 - Settings / Review (SV-7696, SV-7870)",
     "Which roles or org types have 'Require Review Before Completion' turned ON "
     "by default? Should brand-new orgs default it ON or OFF, and do existing orgs "
     "keep today's behaviour via a backfill?",
     "Determines the default review behaviour for new and existing shops, and sets "
     "the expected default state for the settings and review test cases. The spec "
     "flags the require-review default as unresolved.",
     "(a) OFF for all by default; (b) ON for bigger/existing shops, OFF for new "
     "orgs; (c) some other cohort rule.",
     "SF-REV-15, SF-SET-14"),
    ("Story 2 - No-PO completion (SV-7697) / Data integrity",
     "When a work order is completed via the simple / skip path, should in-stock "
     "inventory parts auto-receive and decrement inventory (writing Part History), "
     "or not? What is the intended behaviour?",
     "The skip path's bare status setter can emit no events and bypass inventory "
     "movement / Part History / catalog creation. Whether inventory should still "
     "decrement is a data-integrity invariant.",
     "(a) Auto-decrement inventory + write Part History on completion; (b) no "
     "inventory movement on the skip path; (c) route the completion through the "
     "real receive lifecycle.",
     "SF-COMP-07, SF-QB-01"),
    ("Story 1 - Work Order Settings (SV-7696)",
     "The spec says the first-use defaults are Auto-approve OFF and Vendor Invoice "
     "REQUIRED, but the design mockups show Auto-approve ON and Vendor Invoice "
     "Optional (and omit the Require-Review toggle). Which is authoritative for "
     "first-use defaults?",
     "The correct defaults must be confirmed before we can pass/fail the settings-"
     "default case and frame the completion-flow matrix. The live org baseline "
     "currently shows Auto-approve ON and Vendor Invoice Optional.",
     "(a) Spec defaults (Auto-approve OFF / invoice REQUIRED); (b) Design defaults "
     "(Auto-approve ON / invoice Optional); (c) other.",
     "SF-SET-08"),
    ("Global - spec vs design versioning",
     "The product spec is labelled V2.3 but both design handoffs cite 'Simple Mode "
     "V1.4'. Which version governs where the two differ?",
     "Design-vs-spec differences (defaults, tech-story placement, review states) "
     "cannot be adjudicated until the authoritative version is confirmed.",
     "(a) Spec V2.3 governs; (b) Design V1.4 governs; (c) case-by-case.",
     "Global - especially SF-SET-08, SF-TECH-08, SF-REV-08"),
    ("Story 1 & 2 - Settings / Completion (SV-7696, SV-7697)",
     "Spec S1-R2 requires a 'Create purchase orders' toggle (default ON) so POs "
     "can be switched OFF, but it does not exist in the current build and POs "
     "appear always-on (no createPurchaseOrders field in the settings API). Is "
     "this intended (descoped) or a bug?",
     "Without the toggle the 'No-PO / skip' configuration and its QuickBooks-"
     "integrity cases cannot be set up as specified.",
     "(a) Intended descope - POs are always on; (b) Bug - the toggle should be "
     "added.",
     "SF-SET-03, SF-COMP-06, SF-QB-02"),
    ("Story 1 - Work Order Settings (SV-7696)",
     "The Save Settings button is clickable even when there are no unsaved changes "
     "(no dirty-state gating). Is this intended or a bug?",
     "Minor UX deviation from the expected 'disabled until a change is made' "
     "behaviour.",
     "(a) Intended; (b) Bug - Save should be disabled until something changes.",
     "SF-SET-13"),
    ("Story 16 - Review ON (SV-7870)",
     "Story 16 (R7/R10) specifies an optional review-note field (input_review_note) "
     "in the Mark Reviewed dialog, but the live dialog shows only the VIN field. Is "
     "the note field intended or a bug?",
     "The optional-note case cannot pass as specified until this is confirmed.",
     "(a) Descoped - no note field; (b) Bug - the note field should be added.",
     "SF-REV-10"),
    ("Story 16 - Review ON (SV-7870)",
     "Story 16 (R5/R8) describes Review -> Reviewed (green, 'sign-off complete') -> "
     "a separate final Complete Work Order -> Complete. Live, Confirm Review went "
     "straight to Complete with no distinct 'Reviewed' holding state. Is this "
     "intended or a bug?",
     "Determines whether a distinct Reviewed state plus a separate final Complete "
     "step are expected (this may be admin auto-progression). Affects the review "
     "state-machine cases and the invoicing-blocked-until-reviewed check.",
     "(a) Intended (single-step, e.g. for admins); (b) Bug - a distinct Reviewed "
     "state is expected before final Complete.",
     "SF-REV-08, SF-REV-11"),
    ("Story 15 vs Story 17 - Tech story (SV-7710, SV-7876)",
     "S15-R2 (older wording) says the tech story stays on the line, not in a modal; "
     "Story 17 supersedes it with an inline entry PLUS a gate-modal at completion. "
     "Which is authoritative?",
     "Determines whether the tech-story gate-modal cases are authoritative. The "
     "live build showed the gate-modal working, so Story 17 appears current - "
     "please confirm.",
     "(a) Story 17 (inline + gate-modal) governs; (b) S15-R2 (line-only) governs.",
     "SF-TECH-08 (and all SF-TECH-*)"),
    ("Story 15 - UX Refinements (SV-7710)",
     "The S15-R4 close-vs-cancel confirmation modal is marked 'Figma still to be "
     "added'. What is the intended behaviour, and when will the design ship?",
     "The close-confirm case's expected behaviour cannot be finalised until the "
     "design exists.",
     "(none yet - needs the design)",
     "SF-UX-04"),
    ("Story 12 - Accept Delivery (SV-7707)",
     "S12-R1 says the vendor-missing group sits at the BOTTOM of the Accept "
     "Delivery screen; S12-R3 says the vendor-missing group LEADS (top). The spec "
     "contradicts itself - which is correct?",
     "Determines the expected ordering of the vendor-missing group on the Accept "
     "Delivery and Bulk Receive screens.",
     "(a) Vendor-missing group at the bottom; (b) Vendor-missing group leads (top).",
     "SF-RCV-05, SF-RCV-07"),
]

# ---------------------------------------------------------------- markdown
lines = []
lines.append("# Open Questions for Milos - Simple Mode (Streamlined Work Order "
             "Completion & Receiving)")
lines.append("")
lines.append(f"**Feature:** {FEATURE}  ")
lines.append("**Raised by:** ShopView QA  ")
lines.append("**For:** Milos Vasic (Product Owner / spec author)")
lines.append("")
lines.append(INTRO)
lines.append("")
lines.append("| # | Area / Story | Question | Context / why it matters | "
             "Options (if any) | Affected cases | Answer (for Milos) |")
lines.append("|---|---|---|---|---|---|---|")


def md_cell(s):
    return s.replace("|", "\\|").replace("\n", "<br>")


for i, (area, q, ctx, opts, aff) in enumerate(Q, start=1):
    lines.append("| {} | {} | {} | {} | {} | {} | |".format(
        i, md_cell(area), md_cell(q), md_cell(ctx), md_cell(opts), md_cell(aff)))
lines.append("")
with open(MD, "w", encoding="utf-8") as fh:
    fh.write("\n".join(lines))

# ---------------------------------------------------------------- xlsx
DARK = "1F4E78"
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill("solid", fgColor=DARK)
ALT = PatternFill("solid", fgColor="F2F5FA")
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="top", horizontal="center")

wb = Workbook()
ws = wb.active
ws.title = "Open Questions for Milos"

ws["A1"] = ("Open Questions for Milos - " + FEATURE)
ws["A1"].font = Font(bold=True, size=14, color=DARK)
ws.merge_cells("A1:G1")
ws["A2"] = INTRO
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws["A2"].font = Font(size=11)
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 92

ws.append([])  # row 3 spacer
ws.append(COLUMNS)  # row 4 header
hdr_row = 4
for c in range(1, len(COLUMNS) + 1):
    cell = ws.cell(row=hdr_row, column=c)
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = BORDER

for i, (area, q, ctx, opts, aff) in enumerate(Q, start=1):
    ws.append([i, area, q, ctx, opts, aff, ""])
    r = hdr_row + i
    alt = (i % 2 == 1)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=r, column=c)
        cell.alignment = WRAP
        cell.border = BORDER
        if alt:
            cell.fill = ALT
    ws.cell(row=r, column=1).alignment = WRAPC

widths = [4, 30, 58, 48, 34, 26, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"
wb.save(XLSX)

print("Questions:", len(Q))
print("Saved:", MD)
print("Saved:", XLSX)
