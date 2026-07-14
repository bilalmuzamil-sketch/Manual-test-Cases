#!/usr/bin/env python3
"""Simple Flow — Blockers Tracker generator.

Classifies EVERY authored case (build/simple-flow/cases/*.json) into a delivery
state and, if blocked, WHAT it is blocked on and WHO unblocks it. Emits:
  - build/simple-flow/SimpleFlow_Blockers_Tracker.xlsx (Tracker + Summary tabs)
  - build/simple-flow/SimpleFlow_Blockers_Tracker.md

State model (derived from viu_status + viu-findings.md + OpenQuestions-for-Milos):
  READY                       — VIU-Verified, expected finalized, uploadable now.
  BLOCKED — DEV NOT BUILT     — Stories 7/8/9/14 + SF-PERM-03 (+ deps). Owner: Dev.
  BLOCKED — VIU PENDING (QA)  — couldn't drive yet / needs cookies+seed/accounts. Owner: QA.
  BLOCKED — MILOS ANSWER      — expected depends on the 11 Open Questions. Owner: Milos (PO).
  BLOCKED — BUG/RULING        — new VIU findings need a dev/PO ruling. Owner: Dev/PO.
"""
import csv, json, os, re

BASE = os.path.dirname(os.path.abspath(__file__))          # build/simple-flow
CASES_DIR = os.path.join(BASE, "cases")
OUT_XLSX = os.path.join(BASE, "SimpleFlow_Blockers_Tracker.xlsx")
OUT_MD = os.path.join(BASE, "SimpleFlow_Blockers_Tracker.md")

FILES = [
    "group-A-settings-completion.json",
    "group-B-receiving-vendor.json",
    "group-C-review-permissions-validation-edge.json",
]

# --- Classification inputs ---------------------------------------------------

# Cases whose PASS/FAIL verdict still hangs on a product/dev ruling.
# NOTE 2026-07-10: user ruled on the FE-vs-API enforcement gap ("If the front end is
# blocking it and just not blocked from the API then mark them as passed"). This resolves
# the cases whose ONLY gap is "UI blocks it, API does not" (BUG-6/BUG-7 = fix ticket T2):
#   SF-PERM-02 (WO-completion FE gate) and SF-PERM-06 (settings/WO-action FE gate) were
#   REMOVED from BUG_RULING; they now fall through to READY (viu_status already VIU-Verified)
#   with a PASS note applied in the case JSON.
# UPDATE 2026-07-10 (PO ruling, Milos, relayed by QA lead): the reviewer != completer hard
# rule is DESCOPED from v1 — a completer reviewing their own WO is EXPECTED, not a defect
# (origin: SV-8183 Decision-3/NET-NEW; Story 16/SV-7870 only ever needed a different ROLE).
# The four cases that were held on BUG-5 (SF-PERM-04, SF-PERM-07, SF-PERM-08, SF-REV-09) had
# the identity assertion removed and were re-adjudicated to VIU-Verified (permission-gating
# retained & verified). They now fall through to READY. SF-PERM-08 (the dedicated same-user
# case) is marked OBSOLETE / covered-by SF-PERM-04+07 in the case JSON (kept VIU-Verified so
# it exits BUG/RULING; flagged for the QA lead to retire in TestRail). BUG-5 / TICKET 1 dropped
# as expected. NOTE: these 4 cases' TestRail push is PENDING QA-lead authorization.
# BUG_RULING is now empty.
BUG_RULING = {}

# Open-Question / Milos-owned cases -> the specific Open Question number(s).
MILOS = {
    "SF-SET-03":  ("Q5", "'Create purchase orders' toggle absent / POs always-on — descope or bug?"),
    "SF-SET-08":  ("Q3, Q4", "First-use defaults: spec (Auto-approve OFF / invoice REQUIRED) vs design (ON / Optional)."),
    "SF-SET-13":  ("Q6", "Save Settings always enabled (no dirty-state gating) — intended or bug?"),
    "SF-COMP-06": ("Q5", "Create-POs-OFF => no PO config not possible (toggle absent)."),
    # SF-COMP-07 removed from MILOS 2026-07-10: Q2/R2-Q3 answered (decrement confirmed) AND
    #   live-verified (on-hand 6->5 on pick, persisted through simple-complete). Now VIU-Verified -> READY.
    "SF-TECH-08": ("Q9, Q4", "Tech-story placement: Story 17 (inline + gate-modal) vs S15-R2 (line-only)."),
    "SF-REV-08":  ("Q8, Q4", "Distinct 'Reviewed' state before final Complete — expected or single-step?"),
    "SF-REV-10":  ("Q7", "Optional review-note field absent — descope or bug?"),
    "SF-REV-11":  ("Q8", "Invoicing-blocked-until-reviewed depends on the Reviewed-state ruling (Q8)."),
    "SF-REV-15":  ("Q1", "Require-review default cohort rule + new-org preset."),
    "SF-UX-04":   ("Q10", "Close-vs-cancel confirm modal — design 'still to be added'."),
    # SF-QB-01 removed from MILOS 2026-07-10: Q2/R2-Q3 answered; decrement half now live-proven,
    #   remaining Part-History LOG surface is an env/QA blocker (see SUBBUCKET) -> VIU PENDING (QA).
    "SF-QB-02":   ("Q5", "QuickBooks integrity when Create-POs is OFF (toggle absent)."),
    "SF-RCV-05":  ("Q11", "Vendor-missing group ordering on Accept Delivery — spec contradicts itself."),
    "SF-RCV-07":  ("Q11", "Vendor-missing group ordering (S12-R1 bottom vs S12-R3 top)."),
}

# Dev-not-built: by area prefix -> (story label). Plus explicit case overrides.
# NOTE 2026-07-09: Stories 7/8/9/14 (SF-POSEL/SF-BULK/SF-INV/SF-WOP) were CONFIRMED
# BUILT & VIU-verified live on sv7301 (re-VIU pass) — they are no longer dev-not-built.
# Their per-case state is now driven by viu_status (VIU-Verified => READY; the few still
# unverified => VIU PENDING (QA)). DEV NOT BUILT is now empty.
DEV_STORY_BY_PREFIX = {}
DEV_CASE_OVERRIDE = {}

# --- VIU sub-bucket classification (only meaningful for VIU PENDING (QA) rows) --
# Buckets:
#   reachable-now = verifiable with admin+tech + normal data (just needs another
#                   VIU pass; no new inputs).
#   needs-data    = needs a data state that couldn't be seeded via the app.
#   needs-account = needs a role account we don't have (named).
_REACHABLE = "admin+tech + normal WO data; needs another VIU pass (no new inputs)."
SUBBUCKET = {
    # ---- reachable-now (19) ----
    "SF-SET-10":   ("reachable-now", _REACHABLE),
    "SF-COMP-08":  ("reachable-now", "FRESH VIU 2026-07-10: with autoPickInventoryParts=OFF, an added inventory part stays 'in_stock' and simple-complete is BE-BLOCKED (400 'All inventory and found parts must be picked...'); after line-level Pick it completes (201) — expected #2/#3 PROVEN. Remaining: surface the completion-WIZARD Pick step UI ('Pick all from default bins'/'Review individually') for expected #1 (drive the completion modal, not line-level pick)."),
    "SF-QB-01":    ("needs-data", "FRESH VIU 2026-07-10: decrement half PROVEN (P550848 on-hand 6->5 on pick, persisted through simple-complete 201). Part-History LOG surface BLOCKED-ENV: GET /api/inventory/parts/history -> 500; /parts/inventory/{id} detail page crashes ('page is totaled'); other history endpoints 404/405 (see bugs-log OBS-6). QuickBooks-integrity leg needs QB access."),
    "SF-COMP-10":  ("reachable-now", _REACHABLE),
    "SF-COMP-15":  ("reachable-now", "drive optional-flow Cancel and re-open; check no duplicate POs."),
    "SF-COMP-20":  ("reachable-now", "required-invoice flow (requireVendorInvoiceNumber=ON) + part-bearing WO; Cancel = no change."),
    "SF-COMP-23":  ("reachable-now", "re-run completion after a prior attempt; check no duplicate POs."),
    "SF-VPART-04": ("reachable-now", "edit an existing vendorless part inline (part_number/sell/qty row fields exist)."),
    "SF-VPART-06": ("reachable-now", "add a PN + vendor to a vendorless part and confirm it transitions out of vendorless."),
    "SF-VMIS-04":  ("needs-data", "BATCH 6: vendor-missing WO PO seeded (S2-15774), but NO assign-vendor dropdown / PN-edit on any reachable surface (PO list = Receive only; single /accept-delivery = Vendor Missing badge + invoice/qty/Receive only; /parts/deliveries = received invoices). Appears to require the Bulk Receive page (Story 8, NOT built)."),
    "SF-VMIS-05":  ("needs-data", "BATCH 6: no assign-vendor UI on reachable PO/receive surfaces to clear the flag (see SF-VMIS-04); part-level vendor+PN add DOES clear vendorless via WO Parts editor (part/change-request, SF-VPART-06). PO-surface flag-clear appears tied to Bulk Receive (Story 8, NOT built)."),
    "SF-PNFIX-01": ("needs-data", "BATCH 6: no inline Missing-part-number Edit on the vendor-missing PO Accept Delivery (empty PN cell, no edit control). Inline PN-fix appears tied to the Bulk Receive page (Story 8, NOT built)."),
    "SF-RCV-01":   ("reachable-now", "seed a WO-originated PO; confirm Receive action on PO list + detail."),
    "SF-VEND-01":  ("needs-data", "BATCH 6: vendor-missing WO PO seeded, but no PO-level vendor dropdown on any reachable surface (PO list/accept-delivery/deliveries). The vendor-missing-group vendor dropdown appears to live on the Bulk Receive page (Story 8/13 UI, NOT built)."),
    "SF-REV-03":   ("reachable-now", "enable Require Review; Details step collects mileage + engine hours (VIN captured later by reviewer)."),
    "SF-REV-07":   ("reachable-now", "enable Require Review; Send to Review locks lines to Complete + auto-picks inventory."),
    "SF-REV-12":   ("reachable-now", "enable Require Review; confirm a 'Ready for Review' list filter/column."),
    "SF-REV-13":   ("reachable-now", "enable Require Review; all lines must be approved to Send to Review (approve-line error)."),
    "SF-VAL-07":   ("reachable-now", "review flow; Confirm Review disabled until VIN entered in Mark Reviewed dialog."),
    "SF-VAL-08":   ("reachable-now", "re-complete after cancelling; check no duplicate POs."),
    # ---- needs-data (39) ----
    "SF-COMP-13":  ("needs-data", "BATCH 6: a deliverable ORDERED WO PO IS now achievable (Source=Vendor + real vendor + free-text PN). The shared Accept Delivery page opens for it via the PO Receive action, BUT the completion WIZARD Receive Parts still routes back to the WO lines (PO not placed until completion), and receiving the WO PO 500s (BUG-11)."),
    "SF-COMP-19":  ("needs-data", "BATCH 6: deliverable WO PO achieved, but the receive round-trip is blocked by BUG-11 (POST /api/inventory/orders/accept returns 500 for a WO PO)."),
    "SF-CORE-01":  ("needs-data", "an inventory/catalog part flagged is_core (genuine core) + a received cored line; the manual sub-form only sets core_charge (is_core stays false) and resolve-cores needs receiving."),
    "SF-CORE-02":  ("needs-data", "a genuine core part to prove the Resolve-Cores step appears, plus a no-core WO to prove it is skipped (is_core not seedable via canned/sub-form)."),
    "SF-CORE-03":  ("needs-data", "a special-order core part + optional-invoice completion (needs is_core part)."),
    "SF-CORE-04":  ("needs-data", "an unresolved special-order core at the invoice gate ('Cores pending' flag) — needs is_core part + receiving."),
    "SF-CORE-05":  ("needs-data", "resolve cores at the Create-Invoice gate -> receive the cored line (needs is_core part + receiving)."),
    "SF-CORE-06":  ("needs-data", "cancel the invoice-gate core resolution (needs is_core part + invoice gate)."),
    "SF-CORE-07":  ("needs-data", "special-order cores in the required-invoice receive round-trip (needs is_core part + receiving)."),
    "SF-CORE-08":  ("needs-data", "an unresolved special-order core that exists only as a PartRequest (needs is_core part)."),
    "SF-CORE-09":  ("needs-data", "a part-sale WO vs service WO with cores (needs is_core part + receiving)."),
    "SF-CORE-10":  ("needs-data", "the Resolve-Cores step live '+$ to invoice' total (needs is_core part + receiving)."),
    "SF-VPART-07": ("needs-data", "BATCH 6: deliverable WO PO achievable, but receiving it is blocked by BUG-11 (500), so the cannot-receive-until-PN+vendor proof cannot complete."),
    "SF-VMIS-03":  ("needs-data", "QuickBooks sync inspection (vendor-missing PO excluded from QB)."),
    "SF-VMIS-06":  ("needs-data", "reports data (Vendor Missing POs flagged 'needs vendor')."),
    "SF-PNFIX-02": ("needs-data", "receiving + inventory/catalog inspection (new PN creates inventory part + stock + Part History on receive)."),
    "SF-PNFIX-03": ("needs-data", "receiving + inventory inspection (existing PN links to item, updates stock/cost/history)."),
    "SF-PNFIX-04": ("needs-data", "a WO in invoiced/paid state (sell-field locking)."),
    "SF-PNFIX-05": ("needs-data", "an invoiced/paid WO + a receive attempt (cannot receive without PN)."),
    "SF-PNFIX-06": ("needs-data", "receiving + catalog/inventory back-end inspection (real creation/linking, not a stored string)."),
    "SF-RCV-02":   ("needs-data", "an ordered/deliverable PO so the Receive action opens Accept Delivery (optional-flow Receive Parts routed back to the WO in VIU)."),
    "SF-RCV-06":   ("needs-data", "BATCH 6: on a deliverable WO PO the Accept Delivery gates are present (vendor set, invoice # field), but Receive 500s (BUG-11); on a vendor-missing WO PO there is no assign-vendor / PN-edit control. Happy path was proven on an INVENTORY PO in BATCH 5; the WO-PO negative sub-gates need BUG-11 fixed."),
    "SF-RCV-08":   ("needs-data", "QuickBooks inspection (per-vendor vendor bill + separate AP entry) — and WO-PO receive additionally blocked by BUG-11 (500)."),
    "SF-RCV-09":   ("needs-data", "a receive with received qty > ordered qty (received-more-than-ordered warning)."),
    "SF-REV-04":   ("needs-data", "BATCH 6: deliverable WO PO achievable in the review flow, but receiving it is blocked by BUG-11 (500)."),
    "SF-REV-14":   ("needs-data", "cores + receiving in the review flow — blocked by BUG-10 (no wizard resolve step) + BUG-11 (WO-PO receive 500) + special-order cores not seedable."),
    "SF-VEND-02":  ("needs-data", "a PO where the assigned vendor already exists on the PO (Add-to-vendor merge vs keep-separate prompt)."),
    "SF-VEND-03":  ("needs-data", "two POs for the same WO with the same vendor (merge-POs prompt)."),
    "SF-VEND-04":  ("needs-data", "a vendor-missing PO + receive-enable + QB-flag-clear check after auto-assign."),
    "SF-VEND-05":  ("needs-data", "invoiced/paid WO + multi-PO merge guardrail state (match-by-ID, merge scoped to same WO, receive blocked when invoiced/paid)."),
    "SF-VAL-02":   ("needs-data", "an asset/vehicle with NO VIN so the non-review wizard prompts for VIN. BATCH 6: VIN prefills from the asset; the asset-creation flow (VIN likely required) was not reliably drivable in-harness. Needs a VIN-less asset seeded."),
    "SF-VAL-05":   ("needs-data", "BATCH 6: blocked by BUG-11 (WO-PO receive 500) — cannot complete a required-invoice receive to test the no-invoice-number negative."),
    "SF-VAL-06":   ("needs-data", "BATCH 6: blocked by BUG-11 (WO-PO receive 500) and no assign-vendor UI on the reachable receive surface."),
    "SF-QB-03":    ("needs-data", "QuickBooks / inventory back-end inspection (receive -> Delivery -> Vendor Bill -> QBO) — likely needs dev/QB access."),
    "SF-QB-04":    ("needs-data", "QuickBooks / inventory inspection (vendorless/no-PN part = zero inventory interaction) — likely needs dev/QB access."),
    "SF-QB-05":    ("needs-data", "QuickBooks inspection (Vendor-Missing POs excluded from QB until vendor + PN) — likely needs dev/QB access."),
    "SF-QB-06":    ("needs-data", "QuickBooks inspection (cost-at-completion to avoid $0-cost margins) — likely needs dev/QB access."),
    "SF-QB-07":    ("needs-data", "QuickBooks inspection (Journal Entry / Inventory sync fires on invoice creation) — likely needs dev/QB access."),
    "SF-QB-08":    ("needs-data", "Inventory Part History inspection for any part that becomes inventory-tracked — likely needs dev/QB access."),
    # ---- needs-account (1) ----
    "SF-PERM-10":  ("needs-account", "role accounts Office / Service Manager / Foreman (ideally also Senior SA / Parts Manager / Sales Rep / Time Clock) to run the per-role completion matrix (only admin+ and Technician- available)."),
    # ---- Stories 7/8/9/14 now BUILT — the few still-unverified re-VIU cases ----
    "SF-WOP-02":   ("reachable-now", "RE-VIU 2026-07-09: 'Waiting On Parts' column + unreceived count BUILT & verified (SF-WOP-01); only the click-count->Accept-Delivery navigation is undriven (harness column-persistence flakiness + no non-zero cell surfaced). Just needs a stable WO row with a non-zero count."),
    "SF-BULK-10":  ("needs-data", "RE-VIU 2026-07-09: Bulk Receive page BUILT; this case needs a genuine cored part received on the bulk page to expose Ok/Not-OK resolution (is_core part not seedable via canned/sub-form)."),
    "SF-VAL-09":   ("needs-data", "RE-VIU 2026-07-09: Bulk Receive field-locking BUILT (qty/cost/sell editable verified); the sell-lock-after-invoiced/paid clause needs an invoiced/paid WO to drive."),
    "SF-VAL-10":   ("reachable-now", "RE-VIU 2026-07-09: Apply-invoice BUILT & verified (SF-INV-01/02/03); the reused-invoice-number (uniqueness relaxed) clause just needs a targeted drive applying the same invoice # to multiple POs."),
}

# Tailored "what's needed" for the special QA-pending cases.
QA_OVERRIDE = {
    "SF-PERM-09": "QA VIU: a role account WITHOUT 'See Financial Data' to prove the "
                  "vendorless part-add gate (tech confirmed no seeFinancialData; sub-form not reached).",
    "SF-PERM-10": "QA VIU: additional role accounts (Office / Service Manager / Foreman) "
                  "to run the per-role completion matrix (only Technician-negative confirmed).",
    "SF-VPART-02": "QA VIU: drive the vendorless / no-PN add sub-form validation "
                   "(not reached in the last session's budget).",
}

# Area-family default "what's needed" for generic VIU-Pending cases.
def qa_default(cid, area):
    if cid.startswith("SF-CORE"):
        return ("QA VIU: seed a WO with core-charge parts, drive the completion "
                "Resolve-Cores modal + invoice-gate round-trip.")
    if cid.startswith("SF-VPART"):
        return "QA VIU: seed + drive the vendorless / no-PN part add flow."
    if cid.startswith("SF-VEND"):
        return "QA VIU: seed vendor-missing POs, drive Assign-Vendor + merge."
    if cid.startswith("SF-VMIS"):
        return "QA VIU: seed a WO PO with a vendor-missing part, drive the flag flow."
    if cid.startswith("SF-PNFIX"):
        return "QA VIU: drive the inline part-number fix flow with seeded PO lines."
    if cid.startswith("SF-RCV"):
        return "QA VIU: seed deliverable POs, drive the Receive / Accept-Delivery flow."
    if cid.startswith("SF-CORE"):
        return "QA VIU: seed core parts and drive the cores flow."
    if cid.startswith("SF-QB"):
        return ("QA VIU: complete a WO end-to-end and inspect the QuickBooks / inventory "
                "side-effects (Journal Entry, Part History, inventory decrement).")
    if cid.startswith("SF-COMP"):
        return "QA VIU: seed the specific completion configuration and drive the wizard to Success."
    if cid.startswith("SF-REV"):
        return "QA VIU: enable Require Review and drive the review/sign-off round-trip."
    if cid.startswith("SF-TECH"):
        return "QA VIU: drive the multi-line tech-story gate flow (with the specific test-id)."
    if cid.startswith("SF-VAL"):
        return "QA VIU: drive the specific validation/edge scenario with seeded data."
    if cid.startswith("SF-SET"):
        return "QA VIU: toggle the setting and confirm the downstream behaviour."
    return "QA VIU: fresh sv7301 cookies + seeded data to drive this scenario."


def classify(c):
    """Return dict: state, category, owner, needs, related."""
    cid = c["id"]
    area = c["area"]
    vs = c.get("viu_status", "")
    story = c.get("story_ref") or ""
    prefix = "-".join(cid.split("-")[:2])  # e.g. SF-POSEL

    # 1) Bug/ruling takes precedence (even over VIU-Verified).
    if cid in BUG_RULING:
        return dict(state="BLOCKED", category="BUG/RULING", owner="Dev / PO ruling",
                    needs=BUG_RULING[cid], related=story + " | see viu-findings BUGS #5/#6/#7")

    # 2) Dev-not-built.
    dev = DEV_CASE_OVERRIDE.get(cid) or DEV_STORY_BY_PREFIX.get(prefix)
    if dev:
        return dict(state="BLOCKED", category="DEV NOT BUILT", owner="Dev team",
                    needs="Dev deploys {}; then QA re-runs VIU.".format(dev),
                    related=dev + " | " + story)

    # 3) Milos-owned (Open-Question status or explicit Q dependency).
    if cid in MILOS:
        q, desc = MILOS[cid]
        return dict(state="BLOCKED", category="MILOS ANSWER", owner="Milos (Product Owner)",
                    needs="Milos answers Open Question {} — {}".format(q, desc),
                    related=q + " | " + story)

    # 4) VIU-Verified -> READY.
    if vs == "VIU-Verified":
        return dict(state="READY", category="READY (VIU-Verified)", owner="—",
                    needs="None — VIU-verified; uploadable now.", related=story)

    # 5) Everything else that is VIU-Pending -> QA pending.
    needs = QA_OVERRIDE.get(cid) or qa_default(cid, area)
    return dict(state="BLOCKED", category="VIU PENDING (QA)", owner="QA (needs cookies+seed data)",
                needs=needs, related=story)


def load_cases():
    cases = []
    for fn in FILES:
        cases += json.load(open(os.path.join(CASES_DIR, fn)))
    return cases


def section_for(c):
    """TestRail section for this case; API-related cases route to 'API — <area>'
    (STANDING RULE 4). Kept in sync with gen_import.py."""
    area = c["area"].strip()
    if c.get("api_related"):
        return "API — " + area
    return area


def main():
    cases = load_cases()
    # TestRail Case-ID map (Standing Rule 8: every case-listing deliverable carries
    # the C##### + a clickable TestRail link). Appended as the last two columns so
    # existing r[0..10] index references stay valid.
    import csv as _csv, os as _os
    _idmap = {}
    _mp = _os.path.join(_os.path.dirname(__file__), "testrail-id-map.csv")
    if _os.path.exists(_mp):
        for _r in _csv.DictReader(open(_mp)):
            _idmap[_r["sf_id"]] = _r["ID"]

    def _trlink(cid):
        tid = _idmap.get(cid, "")
        return (("C" + tid) if tid else "",
                ("https://shopview.testrail.io/index.php?/cases/view/" + tid) if tid else "")

    rows = []
    for c in cases:
        cls = classify(c)
        if cls["category"] == "VIU PENDING (QA)":
            sb, sbnote = SUBBUCKET.get(c["id"], ("reachable-now", _REACHABLE))
        else:
            sb, sbnote = "—", ""
        _trid, _link = _trlink(c["id"])
        rows.append([
            c["id"], section_for(c), c["title"].strip(), c.get("viu_status", ""),
            cls["state"], cls["category"], cls["owner"], cls["needs"], cls["related"],
            sb, sbnote, _trid, _link,
        ])

    from collections import Counter, OrderedDict
    cat_counts = Counter(r[5] for r in rows)
    state_counts = Counter(r[4] for r in rows)
    # VIU sub-bucket counts (only over VIU PENDING (QA) rows).
    sub_counts = Counter(r[9] for r in rows if r[5] == "VIU PENDING (QA)")

    CAT_ORDER = ["READY (VIU-Verified)", "BLOCKED — DEV NOT BUILT",
                 "VIU PENDING (QA)", "MILOS ANSWER", "BUG/RULING"]
    # normalise category display names
    def catkey(name):
        m = {"READY (VIU-Verified)": "READY (VIU-Verified)",
             "DEV NOT BUILT": "BLOCKED — DEV NOT BUILT",
             "VIU PENDING (QA)": "BLOCKED — VIU PENDING (QA)",
             "MILOS ANSWER": "BLOCKED — MILOS ANSWER",
             "BUG/RULING": "BLOCKED — BUG/RULING"}
        return m.get(name, name)

    # Rebuild counts with display names.
    disp_counts = Counter(catkey(r[5]) for r in rows)

    HEADER = ["Case ID", "Area", "Title", "Current VIU status", "State",
              "Blocker category", "Who unblocks", "What's needed to unblock",
              "Related story/question", "VIU sub-bucket",
              "VIU sub-bucket detail (QA-pending only)", "TestRail ID", "TestRail Link"]

    # ---- What to send next (batches) ----
    n_milos = disp_counts["BLOCKED — MILOS ANSWER"]
    n_bug = disp_counts["BLOCKED — BUG/RULING"]
    n_qa = disp_counts["BLOCKED — VIU PENDING (QA)"]
    # dev sub-batches
    dev_rows = [r for r in rows if r[5] == "DEV NOT BUILT"]
    STORY_CANON = {7: "Story 7 — PO multi-select (SV-7702)",
                   8: "Story 8 — PO Bulk Receive page (SV-7703)",
                   9: "Story 9 — Apply invoice to selected POs (SV-7704)",
                   14: "Story 14 — Waiting-on-Parts column (SV-7709)"}
    dev_by_story = Counter()
    for r in dev_rows:
        lbl = r[8].split(" | ")[0]
        m = re.search(r"Story (\d+)", lbl)
        canon = STORY_CANON.get(int(m.group(1)), lbl) if m else lbl
        dev_by_story[canon] += 1

    send_next = []
    send_next.append(("Milos's answers to the 11 Open Questions",
                      "unblocks {} cases (all the MILOS-ANSWER rows). Send the filled-in "
                      "OpenQuestions-for-Milos sheet.".format(n_milos)))
    for lbl, n in sorted(dev_by_story.items(), key=lambda x: -x[1]):
        send_next.append(("Dev deploys " + lbl,
                          "unblocks {} case(s); then I re-run VIU and send an update file.".format(n)))
    send_next.append(("Fresh QA cookies for sv7301 (admin + tech) + seeded test data",
                      "unblocks the bulk of the {} VIU-PENDING (QA) cases (cores, receiving, "
                      "vendor, validation round-trips).".format(n_qa)))
    send_next.append(("A 2nd/3rd role account (Office, Service Manager, Foreman) — some WITHOUT "
                      "'See Financial Data'",
                      "unblocks SF-PERM-09 and SF-PERM-10 (per-role completion + vendorless-add gate)."))
    send_next.append(("A dev/PO ruling on FE-only BE enforcement + the missing reviewer!=completer "
                      "rule (resolve SV-8183 'BE enforces' vs SV-7864 atom-collapse)",
                      "finalizes the {} BUG/RULING cases.".format(n_bug)))

    # ---------------- XLSX ----------------
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Blockers Tracker"
    ws.append(HEADER)

    STATE_FILL = {
        "READY": "C6EFCE",
    }
    CAT_FILL = {
        "READY (VIU-Verified)": "C6EFCE",
        "DEV NOT BUILT": "F4CCCC",
        "VIU PENDING (QA)": "FFF2CC",
        "MILOS ANSWER": "D9E1F2",
        "BUG/RULING": "FCE4D6",
    }
    for r in rows:
        ws.append(r)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")

    widths = {"Case ID": 13, "Area": 32, "Title": 55, "Current VIU status": 15,
              "State": 11, "Blocker category": 20, "Who unblocks": 26,
              "What's needed to unblock": 60, "Related story/question": 34,
              "VIU sub-bucket": 15,
              "VIU sub-bucket detail (QA-pending only)": 65}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    SUB_FILL = {"reachable-now": "C6EFCE", "needs-data": "FFF2CC",
                "needs-account": "F4CCCC"}
    wrap = Alignment(wrap_text=True, vertical="top")
    for ridx in range(2, len(rows) + 2):
        cat = ws.cell(row=ridx, column=6).value
        fill = CAT_FILL.get(cat)
        for cidx in range(1, len(HEADER) + 1):
            cell = ws.cell(row=ridx, column=cidx)
            cell.alignment = wrap
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
        # Override the VIU sub-bucket cell (col 10) with its own colour.
        sbval = ws.cell(row=ridx, column=10).value
        if sbval in SUB_FILL:
            ws.cell(row=ridx, column=10).fill = PatternFill("solid", fgColor=SUB_FILL[sbval])
    ws.freeze_panes = "A2"

    # ---- Summary tab ----
    ss = wb.create_sheet("Summary")
    ss.append(["Simple Flow — Blockers Tracker · Summary"])
    ss["A1"].font = Font(bold=True, size=14)
    ss.append([])
    ss.append(["Total authored cases", len(rows)])
    ss.append([])
    ss.append(["Blocker category", "Count", "Owner"])
    OWNER = {"READY (VIU-Verified)": "— (ready to upload)",
             "BLOCKED — DEV NOT BUILT": "Dev team",
             "BLOCKED — VIU PENDING (QA)": "QA",
             "BLOCKED — MILOS ANSWER": "Milos (PO)",
             "BLOCKED — BUG/RULING": "Dev / PO ruling"}
    order = ["READY (VIU-Verified)", "BLOCKED — DEV NOT BUILT",
             "BLOCKED — VIU PENDING (QA)", "BLOCKED — MILOS ANSWER",
             "BLOCKED — BUG/RULING"]
    hdr_row = ss.max_row
    for cidx in range(1, 4):
        ss.cell(row=hdr_row, column=cidx).font = Font(bold=True)
        ss.cell(row=hdr_row, column=cidx).fill = hdr_fill
        ss.cell(row=hdr_row, column=cidx).font = hdr_font
    for cat in order:
        ss.append([cat, disp_counts.get(cat, 0), OWNER[cat]])
        rr = ss.max_row
        f = CAT_FILL.get(cat.replace("BLOCKED — ", "").replace("READY (VIU-Verified)", "READY (VIU-Verified)"))
        # map display -> fill key
        fk = {"READY (VIU-Verified)": "READY (VIU-Verified)",
              "BLOCKED — DEV NOT BUILT": "DEV NOT BUILT",
              "BLOCKED — VIU PENDING (QA)": "VIU PENDING (QA)",
              "BLOCKED — MILOS ANSWER": "MILOS ANSWER",
              "BLOCKED — BUG/RULING": "BUG/RULING"}[cat]
        fill = CAT_FILL[fk]
        for cidx in range(1, 4):
            ss.cell(row=rr, column=cidx).fill = PatternFill("solid", fgColor=fill)
    ss.append(["TOTAL", len(rows), ""])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)

    # Dev sub-breakdown
    ss.append([])
    ss.append(["DEV NOT BUILT — by story", "Count"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)
    for lbl, n in sorted(dev_by_story.items(), key=lambda x: -x[1]):
        ss.append([lbl, n])

    # VIU PENDING (QA) sub-bucket breakdown
    ss.append([])
    ss.append(["VIU PENDING (QA) — by sub-bucket", "Count", "Meaning"])
    hr = ss.max_row
    for cidx in range(1, 4):
        ss.cell(row=hr, column=cidx).font = Font(bold=True)
    SUB_MEAN = {
        "reachable-now": "admin+tech + normal data; just needs another VIU pass (no new inputs).",
        "needs-data": "needs a data state not seedable via the app (see per-case detail).",
        "needs-account": "needs a role account we don't have (see per-case detail).",
    }
    SUB_ORDER = ["reachable-now", "needs-data", "needs-account"]
    for sb in SUB_ORDER:
        ss.append([sb, sub_counts.get(sb, 0), SUB_MEAN[sb]])
        rr = ss.max_row
        f = {"reachable-now": "C6EFCE", "needs-data": "FFF2CC",
             "needs-account": "F4CCCC"}[sb]
        for cidx in range(1, 3):
            ss.cell(row=rr, column=cidx).fill = PatternFill("solid", fgColor=f)
    ss.append(["TOTAL VIU PENDING (QA)", sum(sub_counts.values()), ""])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)

    ss.append([])
    ss.append(["WHAT TO SEND ME NEXT (to unblock each batch)"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True, size=12)
    for what, effect in send_next:
        ss.append(["• " + what, effect])
        ss.cell(row=ss.max_row, column=1).font = Font(bold=True)

    ss.column_dimensions["A"].width = 60
    ss.column_dimensions["B"].width = 70
    ss.column_dimensions["C"].width = 22
    for row in ss.iter_rows():
        for cell in row:
            if cell.alignment.wrap_text is not True:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT_XLSX)
    print("Wrote", OUT_XLSX)

    # ---------------- Markdown ----------------
    def md_esc(s):
        return (s or "").replace("|", "\\|").replace("\n", " ")

    lines = []
    lines.append("# Simple Flow — Blockers Tracker")
    lines.append("")
    lines.append("> Source of truth for what every authored Simple Flow case is waiting on "
                 "and who unblocks it. Regenerate with `python3 build/simple-flow/gen_blockers.py`.")
    lines.append("> Companion upload file: `testrail-import/simple-flow-v1-testrail-import.csv` "
                 "(all 159 cases). Update loop: `build/simple-flow/gen_update.py` "
                 "(+ `UPDATE-LOOP-README.md`).")
    lines.append("")
    lines.append("**Total authored cases: {}**".format(len(rows)))
    lines.append("")
    lines.append("## Summary — counts per category")
    lines.append("")
    lines.append("| Blocker category | Count | Owner |")
    lines.append("|---|---:|---|")
    for cat in order:
        lines.append("| {} | {} | {} |".format(cat, disp_counts.get(cat, 0), OWNER[cat]))
    lines.append("| **TOTAL** | **{}** | |".format(len(rows)))
    lines.append("")
    lines.append("### DEV NOT BUILT — by story")
    lines.append("")
    lines.append("| Story | Count |")
    lines.append("|---|---:|")
    for lbl, n in sorted(dev_by_story.items(), key=lambda x: -x[1]):
        lines.append("| {} | {} |".format(lbl, n))
    lines.append("")
    lines.append("### VIU PENDING (QA) — by sub-bucket")
    lines.append("")
    lines.append("| VIU sub-bucket | Count | Meaning |")
    lines.append("|---|---:|---|")
    _submean = {
        "reachable-now": "admin+tech + normal data; just needs another VIU pass (no new inputs).",
        "needs-data": "needs a data state not seedable via the app (see per-case detail).",
        "needs-account": "needs a role account we don't have (see per-case detail).",
    }
    for sb in ["reachable-now", "needs-data", "needs-account"]:
        lines.append("| {} | {} | {} |".format(sb, sub_counts.get(sb, 0), _submean[sb]))
    lines.append("| **TOTAL VIU PENDING (QA)** | **{}** | |".format(sum(sub_counts.values())))
    lines.append("")
    lines.append("## WHAT TO SEND ME NEXT (to unblock each batch)")
    lines.append("")
    for what, effect in send_next:
        lines.append("- **{}** → {}".format(what, effect))
    lines.append("")
    lines.append("## Full per-case tracker")
    lines.append("")
    lines.append("| Case ID | Area | Title | VIU status | State | Blocker category | "
                 "Who unblocks | What's needed | Related | VIU sub-bucket | Sub-bucket detail | "
                 "TestRail ID | TestRail Link |")
    lines.append("|---|---|---|---|---|---|---|---|---|---|---|---|---|")
    for r in rows:
        cat_disp = {"READY (VIU-Verified)": "READY (VIU-Verified)",
                    "DEV NOT BUILT": "BLOCKED — DEV NOT BUILT",
                    "VIU PENDING (QA)": "BLOCKED — VIU PENDING (QA)",
                    "MILOS ANSWER": "BLOCKED — MILOS ANSWER",
                    "BUG/RULING": "BLOCKED — BUG/RULING"}[r[5]]
        _lnk = ("[{}]({})".format(r[11], r[12]) if r[11] else "")
        lines.append("| {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} |".format(
            r[0], md_esc(r[1]), md_esc(r[2]), md_esc(r[3]), r[4], cat_disp,
            md_esc(r[6]), md_esc(r[7]), md_esc(r[8]), r[9], md_esc(r[10]), r[11], _lnk))
    lines.append("")
    open(OUT_MD, "w").write("\n".join(lines))
    print("Wrote", OUT_MD)

    print("\nState counts:", dict(state_counts))
    print("Category counts:", dict(disp_counts))
    assert sum(disp_counts.values()) == len(rows) == 170


if __name__ == "__main__":
    main()
