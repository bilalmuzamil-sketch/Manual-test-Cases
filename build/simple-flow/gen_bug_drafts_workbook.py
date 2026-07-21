#!/usr/bin/env python3
"""Generate the Simple Flow bug-drafts workbook in VERY SIMPLE layman language.

Outputs (regenerable — run from the repo root):
  - build/simple-flow/SimpleFlow_Bug-Drafts.xlsx
  - build/simple-flow/SimpleFlow_Bug-Drafts.md

Sheets:
  - "Bug Drafts"          : reader-facing, VERY SIMPLE layman language ONLY.
                            NO internal case IDs, BUG-codes, or API/HTTP jargon
                            (e.g. "enforced only in the UI, bypassable via API" ->
                            "the screen blocks it, but the system behind the screen
                            does not - someone technical could still do it").
                            Columns: # | Title | What happens now | What should
                            happen | How to see it | Severity.
  - "QA Internal Mapping" : QA-only. Per bug: internal BUG-code(s), affected case
                            IDs with TestRail Case ID C##### + clickable link
                            (standing rule 8; source testrail-id-map.csv), spec/story
                            refs, and current status.

Source of truth: build/simple-flow/jira-bug-drafts.md (4 active tickets TICKET 2-5,
post-Milos-Round-2, updated 2026-07-10 after BUG-5/TICKET 1 was dropped as expected
behavior) PLUS bug 5, a Round-3 deviation (Milos 2026-07-16 decision; live-observed
2026-07-16, evidence in viu-round3-2026-07-16/) — bug 5 was DROPPED / WON'T FILE on
2026-07-20 (user decision: vendor-missing-group position is cosmetic only, no
functional impact; kept in the list for the record, flipped to dropped status).
These are DEFECTS (dev tickets), kept OUT of any PO-facing deliverable (standing
rule 7).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/simple-flow/SimpleFlow_Bug-Drafts.xlsx"
MD_OUT = "build/simple-flow/SimpleFlow_Bug-Drafts.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="7C2128")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="7C2128")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no bug codes, no tech terms)
# ---------------------------------------------------------------------------
bugs = [
    # DROPPED 2026-07-10 — "A person can approve (sign off) their own work order review"
    # (BUG-5 / TICKET 1) was settled as EXPECTED behavior by Milos's ruling: reviewer !=
    # completer is not a v1 requirement, so a completer may review their own WO. Removed
    # from the active bug list. See jira-bug-drafts.md "Dropped — expected behavior per
    # PO (Milos), 2026-07-10".
    {
        "title": ("Permission to finish or review a work order is only enforced on "
                  "the screen, not behind the scenes"),
        "now": ("Some staff are not supposed to be able to finish a work order or "
                "sign off a review. The screen correctly hides those buttons from "
                "them - so through normal use they can't do it. But the check is only "
                "on the screen: the system behind the screen does NOT block the "
                "action, so someone technical could still finish a work order or sign "
                "off a review even without the permission. (It has been agreed the "
                "on-screen block is good enough for the first version; this is logged "
                "so the deeper block can be added later.)"),
        "should": ("The block should also happen behind the screen, not just on it - "
                   "so a person without the permission is refused even if they try to "
                   "go around the normal screen."),
        "steps": ("1. Sign in as a role that is NOT allowed to finish work orders "
                  "(for example a Technician).\n"
                  "2. Confirm the Finish / Mark-Reviewed buttons are hidden on the "
                  "screen - good.\n"
                  "3. Have someone technical trigger the same \"finish\" or \"sign "
                  "off\" action directly (not through the normal screen).\n"
                  "4. Notice the action still goes through - it should have been "
                  "refused."),
        "severity": "Medium",
    },
    {
        "title": ("Required fields at finish time (mileage, VIN, engine hours) are "
                  "only enforced on the screen"),
        "now": ("When settings say fields like mileage, VIN or engine hours are "
                "required to finish a work order, the finish screen correctly stops "
                "you until you fill them in. But the system behind the screen does "
                "NOT check for them - so a work order can still be finished with "
                "those required fields left empty if the action is triggered outside "
                "the normal screen."),
        "should": ("The system behind the screen should also refuse to finish a work "
                   "order when a required field (mileage / VIN / engine hours) is "
                   "missing - the same way the on-screen finish step already does."),
        "steps": ("1. In Work Order settings, turn on \"require mileage\" (and/or "
                  "VIN, engine hours).\n"
                  "2. Confirm the finish screen blocks you until those fields are "
                  "filled - good.\n"
                  "3. Have someone technical trigger the \"finish\" action directly, "
                  "leaving those fields empty.\n"
                  "4. Notice the work order finishes anyway with the fields blank - "
                  "it should have been refused."),
        "severity": "Medium",
    },
    {
        "title": ("Receiving a work-order part on the OLDER receiving screen fails "
                  "with an error (the newer bulk screen works)"),
        "now": ("When a part was ordered from a work order, trying to receive it on "
                "the older single-order \"Accept Delivery\" screen fails - the app "
                "shows a generic error and nothing is received. The SAME part "
                "receives fine on the newer \"bulk receive\" screen, and ordinary "
                "(non-work-order) parts receive fine on the older screen too. So it "
                "only breaks for work-order parts on that one older screen - and a "
                "working alternative already exists, which is why this is low "
                "urgency."),
        "should": ("Receiving a work-order part on the older Accept Delivery screen "
                   "should succeed and record the delivery, exactly like it already "
                   "does on the newer bulk receive screen and for ordinary parts."),
        "steps": ("1. Create a work order and add a part from a supplier, typing the "
                  "part number in by hand.\n"
                  "2. Finish the work order so the part becomes an order waiting to "
                  "be received.\n"
                  "3. Open the older \"Accept Delivery\" screen for that order.\n"
                  "4. Enter an invoice number and a received quantity, then click "
                  "Receive.\n"
                  "5. Notice it fails with an error - the newer bulk receive screen "
                  "would have accepted it."),
        "severity": "Low",
    },
    {
        "title": "A brand-new company starts with the wrong default settings",
        "now": ("When a company first starts using Simple Mode, two settings come "
                "out of the box set the wrong way: \"auto-approve lines\" is turned "
                "ON (it should be OFF) and the supplier invoice number is set to "
                "optional (it should be required). So a new company gets the wrong "
                "behaviour until someone notices and changes it by hand."),
        "should": ("Out of the box, a new company should have \"auto-approve lines\" "
                   "turned OFF and the supplier invoice number set to REQUIRED, "
                   "matching the agreed defaults."),
        "steps": ("1. On a brand-new company (or first use of Simple Mode), open Work "
                  "Order settings.\n"
                  "2. Look at \"auto-approve lines\" and the \"supplier invoice\" "
                  "setting.\n"
                  "3. Notice auto-approve is ON and the invoice is optional - they "
                  "should be OFF and required."),
        "severity": "Medium",
    },
    {
        "title": ("On the Receive screen, the \"Vendor Missing\" group of parts shows "
                  "at the top instead of the bottom"),
        "now": ("When you open a work order and click the \"Receive\" button, you land "
                "on the \"Purchase Order Details\" screen, where the parts are grouped "
                "by supplier. Parts that don't have a supplier assigned yet are put in "
                "a \"Vendor Missing\" group. On this Receive screen that \"Vendor "
                "Missing\" group appears at the TOP of the list, above all the "
                "supplier groups. It has been agreed it should appear at the BOTTOM "
                "here. (On the separate \"Bulk Receive\" / \"Receive Vendor Parts\" "
                "page the \"Vendor Missing\" group correctly shows at the top - that "
                "page is fine and should not change.)"),
        "should": ("On the Receive (\"Purchase Order Details\") screen reached from a "
                   "work order's \"Receive\" button, the \"Vendor Missing\" group "
                   "should appear at the BOTTOM of the list, below the supplier "
                   "groups. The \"Bulk Receive\" / \"Receive Vendor Parts\" page "
                   "should keep showing it at the top, unchanged."),
        "steps": ("1. Open a work order that has some parts with a supplier and at "
                  "least one part with no supplier assigned yet.\n"
                  "2. Click the \"Receive\" button on the work order to open the "
                  "\"Purchase Order Details\" screen.\n"
                  "3. Look at where the \"Vendor Missing\" group sits in the list.\n"
                  "4. Notice it is at the TOP, above the supplier groups - it should "
                  "be at the BOTTOM here.\n"
                  "5. For comparison, open the \"Bulk Receive\" / \"Receive Vendor "
                  "Parts\" page and confirm the \"Vendor Missing\" group is at the "
                  "top there - that one is correct."),
        "severity": ("DROPPED — WON'T FILE (cosmetic only, no functional impact; "
                     "user decision 2026-07-20)"),
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (rule 8: TestRail C##### + clickable link per case).
# Each entry: (bug_no, internal_codes, [(sf_id, testrail_id), ...], refs, status)
# ---------------------------------------------------------------------------
internal_map = [
    # BUG-5 / TICKET 1 REMOVED 2026-07-10: dropped as EXPECTED behavior by Milos
    # (reviewer != completer not a v1 requirement). SF-PERM-04/07 + SF-REV-09 now
    # VIU-Verified; SF-PERM-08 obsolete. Remaining bugs renumbered 1-4.
    (1,
     "BUG-6 + BUG-7. Jira draft: TICKET 2. Milos R2 Q5: UI gating = v1 PASS; this is "
     "the OPEN fix ticket for the behind-the-screen (API) gap.",
     [("SF-PERM-06", 29410), ("SF-PERM-02", 29406), ("SF-PERM-07", 29411),
      ("SF-REV-09", 29394)],
     "SV-8183 backend-enforcement claim vs SV-7864 atom-collapse "
     "(workOrderLinesCreateAndEdit collapses to ROLE_WORK_ORDER::VIEW+CREATE_AND_"
     "EDIT). Tech simple-complete -> 201; tech change-status->complete -> 201; by "
     "contrast tech settings/change -> 403 (settings atom IS enforced).",
     "OPEN — CONFIRMED bug (API gap kept open per Milos R2 Q5). Medium. SF-PERM-06 "
     "= API section (4090). Results recorded \"UI pass / API fail\"."),
    (2,
     "BUG-8. Jira draft: TICKET 3.",
     [("SF-VAL-01", 29415), ("SF-VAL-02", 29416), ("SF-VAL-03", 29417),
      ("SF-COMP-05", 29294), ("SF-COMP-16", 29305), ("SF-REV-03", 29388)],
     "SV-8183 backend-enforcement claim / SV-7864 atom-collapse. Evidence "
     "viu-evidence/VIU2-02-mileage-gate.png (wizard blocks) vs simple-complete {} "
     "-> 201 with mileage empty.",
     "OPEN — CONFIRMED bug, expected NOT rewritten. Medium. Required-field gates "
     "(mileage/VIN/engine hours) are UI-only; backend-checked blockers (tech story, "
     "line approval) ARE enforced."),
    (3,
     "BUG-11. Jira draft: TICKET 4. DOWNGRADED 2026-07-09 (RE-VIU BATCH 7): confined "
     "to the LEGACY single-PO Accept-Delivery path; the new Bulk Receive pipeline "
     "works (receive-requested-parts -> 200).",
     [("SF-COMP-13", 29302), ("SF-COMP-19", 29308), ("SF-VAL-05", 29419),
      ("SF-VAL-06", 29420), ("SF-PNFIX-02", 29364), ("SF-PNFIX-03", 29365),
      ("SF-PNFIX-04", 29366), ("SF-PNFIX-05", 29367), ("SF-PNFIX-06", 29368),
      ("SF-RCV-08", 29376), ("SF-VPART-07", 29337), ("SF-REV-04", 29389),
      ("SF-REV-14", 29399), ("SF-CORE-03", 29315), ("SF-CORE-04", 29316),
      ("SF-CORE-05", 29317), ("SF-CORE-07", 29319)],
     "SV-7301 / Story 10 (receive creates/links catalog+inventory part) / Story 8 "
     "(Bulk Receive = the working path). Legacy POST /api/inventory/orders/accept -> "
     "500 for WO POs (free-text/non-catalog part; manufacturer_id null). Evidence "
     "viu-evidence/R7-01-wo-po-accept-delivery.png, R7-06-received-full.png.",
     "OPEN — Low (downgraded). Affected cases now largely testable via the Bulk "
     "Receive path; this ticket only blocks the legacy single-PO Accept-Delivery "
     "surface."),
    (4,
     "GAP-B. Jira draft: TICKET 5.",
     [("SF-SET-08", 29282)],
     "§4 / S1 first-use defaults (confirmed Milos Q3): Auto-approve OFF, Create POs "
     "ON, Vendor Invoice REQUIRED. Live GET /api/organizations/settings shows "
     "autoApproveLines:true, requireVendorInvoiceNumber:false.",
     "OPEN — CONFIRMED bug, SF-SET-08 expected stays (authoritative spec default). "
     "Medium. Wrong first-use org defaults."),
    (5,
     "Round-3 deviation (Milos 2026-07-16 decision). No prior BUG-code; new dev "
     "ticket draft.",
     [("SF-RCV-05", 29373), ("SF-RCV-07", 29375)],
     "SV-7301 / Story 12 (Accept Delivery). Milos 2026-07-16: Vendor Missing group "
     "should sit at BOTTOM on the WO Receive (Purchase Order Details, grouped-by-vendor) "
     "surface but at TOP on the Bulk Receive (Receive Vendor Parts) page. Live-observed "
     "2026-07-16: Vendor Missing renders at TOP on the WO Receive screen (wrong); TOP on "
     "Bulk Receive (correct). Evidence viu-round3-2026-07-16/ORDER-RECV-S15878-full.png, "
     "ORDER-RECV-S15878-Aeboro-miss.png (WO Receive) vs BULK-groups-full.png (Bulk "
     "Receive); observations.json.",
     "DROPPED — WON'T FILE (cosmetic only, no functional impact; user decision "
     "2026-07-20). The Vendor Missing group still appears and functions; only its "
     "position differs (TOP vs Milos's ruled BOTTOM on the Accept-Delivery / Purchase "
     "Order Details screen) — purely visual, no functional/data/workflow impact, so "
     "not filed as a bug. SF-RCV-05 + SF-RCV-07 KEEP Deviation status (the build "
     "genuinely deviates from the ruling) but are annotated ACCEPTED COSMETIC / "
     "won't-file, not an open actionable bug. Bulk Receive surface is correct — no "
     "change there."),
]

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
wb = Workbook()

ws = wb.active
ws.title = "Bug Drafts"
ws.column_dimensions["A"].width = 5
for col, w in zip("BCDEF", [40, 60, 48, 58, 12]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Simple Mode - Bug Drafts (plain-English)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
intro = ("Five issues found while testing Simple Mode, written in plain English. "
         "Each row explains what happens now, what should happen instead, and simple "
         "steps to see it. (These are defects for the dev team - not questions for "
         "the product owner.)")
ws["A2"] = intro
ws["A2"].alignment = WRAP
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 45

headers = ["#", "Title", "What happens now", "What should happen",
           "How to see it (simple steps)", "Severity"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, b in enumerate(bugs, start=1):
    row = HDR_ROW + i
    vals = [i, b["title"], b["now"], b["should"], b["steps"], b["severity"]]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c in (1, 6) else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 185

# --- QA Internal Mapping sheet ---
wi = wb.create_sheet("QA Internal Mapping")
wi["A1"] = ("INTERNAL - for QA / dev only. The reader-facing \"Bug Drafts\" tab "
            "carries no IDs or codes.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:F1")

ihead = ["Bug #", "Internal code / Jira draft", "Case (internal ID)",
         "TestRail Case ID", "TestRail link", "Refs & current status"]
iwid = [7, 40, 16, 15, 46, 66]
IH_ROW = 3
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for bno, codes, cases, refs, status in internal_map:
    first = r
    refstat = f"Refs: {refs}\n\nStatus: {status}"
    for sf_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        rowvals = [bno, codes, sf_id, f"C{tr_id}", url, refstat]
        for c, v in enumerate(rowvals, start=1):
            cell = wi.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_CENTER if c in (1, 3, 4) else WRAP
            cell.border = BORDER
        link_cell = wi.cell(row=r, column=5)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT
        wi.row_dimensions[r].height = 40
        r += 1
    last = r - 1
    if last > first:
        for col in (1, 2, 6):
            wi.merge_cells(start_row=first, start_column=col,
                           end_row=last, end_column=col)

note_row = r + 1
wi.cell(row=note_row, column=1, value=(
    "Notes: Source of truth = jira-bug-drafts.md (4 active tickets TICKET 2-5, "
    "post-Milos-Round-2, updated 2026-07-10) PLUS bug 5, a Round-3 deviation "
    "(Milos 2026-07-16 decision; live-observed 2026-07-16). TestRail IDs sourced from "
    "testrail-id-map.csv (standing rule 8). These are DEFECTS for the dev team (Jira "
    "TICKET 2-5 + the Round-3 deviation under epic SV-7301, Product Area Work Orders) - "
    "NOT filed yet (no Atlassian MCP here; file from the chat app). Kept OUT of any "
    "PO-facing deliverable (standing rule 7). DROPPED / WON'T FILE: bug 5 "
    "(vendor-missing-group position — cosmetic only, no functional impact; user "
    "decision 2026-07-20; SF-RCV-05/07 stay Deviation, annotated accepted-cosmetic) "
    "and the earlier BUG-5/TICKET 1 (reviewer != completer descoped v1, Milos "
    "2026-07-10). CLOSED / not filed: BUG-3 "
    "(review-note descoped, Milos R2 Q1), BUG-9/GAP-A (vendorless category-req/"
    "sell-optional intended, Milos R2 Q4), BUG-1/2/4/10.")
    ).alignment = WRAP
wi.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
wi.row_dimensions[note_row].height = 90

wb.save(XLSX_OUT)

# ---------------------------------------------------------------------------
# Markdown mirror
# ---------------------------------------------------------------------------
md = []
md.append("# Simple Mode — Bug Drafts (plain-English)")
md.append("")
md.append("Five issues found while testing Simple Mode, written in plain English.")
md.append("Each entry explains what happens now, what should happen instead, and")
md.append("simple steps to see it. **These are defects for the dev team — not")
md.append("questions for the product owner.**")
md.append("")
md.append("---")
for i, b in enumerate(bugs, start=1):
    md.append("")
    md.append(f"## {i}. {b['title']}  _(Severity: {b['severity']})_")
    md.append("")
    md.append("**What happens now**")
    md.append(b["now"])
    md.append("")
    md.append("**What should happen**")
    md.append(b["should"])
    md.append("")
    md.append("**How to see it (simple steps)**")
    for line in b["steps"].split("\n"):
        md.append(line)
    md.append("")
    md.append("---")
md.append("")
md.append("---")
md.append("")
md.append("## Internal — QA/dev-only mapping (NOT for the PO)")
md.append("")
md.append("Links each plain-English bug above to its internal code, Jira draft,")
md.append("affected TestRail cases, refs and current status.")
md.append("")
for bno, codes, cases, refs, status in internal_map:
    md.append(f"### Bug {bno}")
    md.append("")
    md.append(f"- **Internal code / Jira draft:** {codes}")
    md.append("- **TestRail cases:**")
    for sf_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        md.append(f"  - {sf_id} — [C{tr_id}]({url})")
    md.append(f"- **Refs:** {refs}")
    md.append(f"- **Current status:** {status}")
    md.append("")
md.append("**Notes:** Source of truth = `jira-bug-drafts.md` (4 active tickets")
md.append("TICKET 2–5, post-Milos-Round-2, updated 2026-07-10) plus bug 5, a Round-3")
md.append("deviation (Milos 2026-07-16 decision; live-observed 2026-07-16). TestRail IDs")
md.append("sourced from `testrail-id-map.csv` (standing rule 8). These are DEFECTS for the")
md.append("dev team (Jira TICKET 2–5 + the Round-3 deviation under epic SV-7301, Product")
md.append("Area Work Orders) — NOT filed yet (no Atlassian MCP here; file from the chat")
md.append("app). Kept OUT of any")
md.append("PO-facing deliverable (standing rule 7). DROPPED / WON'T FILE: bug 5")
md.append("(vendor-missing-group position — cosmetic only, no functional impact; user")
md.append("decision 2026-07-20; SF-RCV-05/07 stay Deviation, annotated accepted-cosmetic)")
md.append("and the earlier BUG-5/TICKET 1 (reviewer != completer descoped v1, Milos")
md.append("2026-07-10). CLOSED / not filed:")
md.append("BUG-3 (review-note descoped, Milos R2 Q1), BUG-9/GAP-A (vendorless")
md.append("category-req/sell-optional intended, Milos R2 Q4), BUG-1/2/4/10.")
md.append("")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md))

print(f"Wrote {XLSX_OUT} and {MD_OUT} with {len(bugs)} bug drafts.")
