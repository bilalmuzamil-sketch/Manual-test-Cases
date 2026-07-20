#!/usr/bin/env python3
"""Generate the GROUNDED spec-V2.6 PO question sheet for Milos (Simple Flow PO).

Outputs (regenerable — run from the repo root):
  - build/simple-flow/PO-Questions-Milos-SpecV26_2026-07-17.xlsx
  - build/simple-flow/PO-Questions-Milos-SpecV26_2026-07-17.md

Mirrors the established SimpleFlow_PO-Decisions-for-Milos_2026-07-14 format 1:1
(gen_po_decisions_milos.py): a layman reader tab ("The situation / What the
written spec currently says / What the app actually does today / Why it needs
your decision / The options / Your decision") + a QA-internal evidence & mapping
tab. Every question SENT is grounded in the spec `_4` (V2.6, 2026-07-17) upload:
either the spec contradicts itself or leftover text survives a removal, so only
a product decision from Milos can settle it.

The three questions (all from the spec-`_4` apply pass, D5 flags in
spec-v4-2026-07-17/spec-diff-v4-2026-07-17.md):

  Q1 — S8-R7 leftover sentence (Δ14): new "$0-only cost editing" rule vs the
       surviving old tail "after it locks, only cost remains editable".
       Case: SF-BULK-06 (C29355).
  Q2 — Vendors-Expenses exclusion surface (Δ12 / S6-R6 rewrite): confirm
       "Vendors Expenses" is the exact report where vendor-missing spend is
       excluded. Case: SF-VMIS-06 (C29343).
  Q3 — S10-R2 residue (pre-existing, Δ7): the struck first-class-part rule vs
       the surviving Story-10 AC bullets + technical guardrails.
       Cases: SF-PNFIX-02/03/06 + SF-QB-08 (rescoped 2026-07-14).

Rules honoured: no case IDs / codes / HTTP / jargon in the reader-facing tab
(standing rule 7); TestRail Case ID + clickable link on the QA-internal tab
(standing rule 8); no bugs put in front of the PO (rule 7 -> bugs to dev).
NOTE: SF-CORE-05/06/09 were retired (user ruling 2026-07-17, executed
2026-07-20) and are NOT referenced here.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/simple-flow/PO-Questions-Milos-SpecV26_2026-07-17.xlsx"
MD_OUT = "build/simple-flow/PO-Questions-Milos-SpecV26_2026-07-17.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# READER-FACING content (layman ONLY — no IDs, no codes, no tech terms).
# Each question is grounded: spec_says quotes/paraphrases the actual clause,
# app_does states the observed build behavior.
# ---------------------------------------------------------------------------
INTRO = (
    "Hi Milos - thank you for the newest write-up. While folding it into our "
    "testing we ran into three small points where the text either disagrees with "
    "itself or leaves us unsure exactly what to check. These are not guesses: for "
    "each one we put the written spec and the app's real behaviour side by side. "
    "Please just pick one option per row (or add a note). Thank you!"
)

questions = [
    {
        # Q1 — SF-BULK-06 / S8-R7 tail vs the new $0-only rule (Δ14)
        "situation": (
            "When a delivery of parts is received, each line shows the cost of the "
            "item. The write-up sets rules for when that cost can still be changed "
            "on the receiving screen."),
        "spec_says": (
            "The write-up says the cost of an item can only be changed when it is "
            "0.00. But an older sentence in the same section still says that after "
            "a purchase order is locked \"only the cost stays changeable\". These "
            "two can't both be true."),
        "app_does": (
            "When we last checked the receiving screen (before this newest write-up "
            "arrived), the cost field could still be edited even when it was not "
            "0.00. We have not yet re-checked it against the new rule - we need to "
            "know which sentence wins first."),
        "why": (
            "We have to check the build against one rule or the other. If the new "
            "\"only when it is 0.00\" rule wins, the older sentence is leftover "
            "text; if the older sentence wins, the new rule can't be right as "
            "written. Only you can settle which one the app should follow."),
        "opts": (
            "A) The cost can only be changed while it is 0.00 - the older sentence "
            "is leftover text and will be cleaned up.\n"
            "B) The cost stays changeable after the purchase order locks - please "
            "explain how the two sentences fit together."),
    },
    {
        # Q2 — SF-VMIS-06 / Δ12 S6-R6 rewrite: Vendors-Expenses surface confirm
        "situation": (
            "Sometimes a part is bought before a supplier has been assigned to it. "
            "Your update explains what happens to that spending in the books while "
            "the supplier is still missing."),
        "spec_says": (
            "Your update says purchases from parts with no supplier assigned are "
            "left out of the QuickBooks supplier-bill export AND out of the "
            "\"Vendors Expenses\" report until a supplier is set. We want to make "
            "sure we check the right report."),
        "app_does": (
            "We can create such a no-supplier purchase in the app and we are ready "
            "to look for the missing spending - we just want to confirm the exact "
            "report name and place before we check it."),
        "why": (
            "If we look at the wrong report we could wrongly pass or fail this "
            "check. A one-word confirmation of the report name makes sure our test "
            "looks in the right place."),
        "opts": (
            "A) Yes - \"Vendors Expenses\" is the exact report where the missing "
            "spending should be visible.\n"
            "B) It's a different place - please name the report or screen we "
            "should check."),
    },
    {
        # Q3 — S10-R2 residue: struck rule vs surviving AC bullets + guardrails
        "situation": (
            "On the receiving screen a part can be given its proper part number on "
            "the spot. An earlier version of the write-up said that doing this also "
            "creates a brand-new part type in the catalogue."),
        "spec_says": (
            "An earlier update removed the rule about creating a brand-new part "
            "type from the receive screen, but two other places in the write-up "
            "still describe it. So one part of the text says the feature is out, "
            "while other parts still describe it as in."),
        "app_does": (
            "Today the app saves the typed part number and the part can be "
            "received normally. We stopped checking for a brand-new catalogue part "
            "type when the rule was removed."),
        "why": (
            "We need to know whether the leftover text is just clean-up that "
            "hasn't happened yet, or whether the feature is actually still meant "
            "to be in - that changes what we test."),
        "opts": (
            "A) The feature is out - the leftover text will be cleaned up.\n"
            "B) It is still in - please clarify what should happen when a part "
            "number is entered on the receiving screen."),
    },
]

# ---------------------------------------------------------------------------
# QA-INTERNAL mapping (rule 8: TestRail C##### + clickable link per case).
# Each entry: (q_no, title, [(sf_id, tr_id), ...], spec_clause, build_evidence,
#              ayesha_run325, resolves_to)
# ---------------------------------------------------------------------------
kept_map = [
    (1,
     "S8-R7 leftover sentence — $0-only cost rule vs \"only cost stays editable after lock\"",
     [("SF-BULK-06", 29355)],
     "SPEC SELF-CONTRADICTION introduced by spec `_4` (V2.6) Δ14 (spec-v4-2026-07-17/"
     "spec-diff-v4-2026-07-17.md §A Δ14 + §D5 flag 2): the NEW S8-R7 = \"Cost is "
     "editable (if the cost is 0, if cost is not 0 cost filed should not be "
     "editable)\" while the SURVIVING old tail of the same S8-R7 paragraph still "
     "reads \"After it locks, only cost remains editable.\" Both sentences are in "
     "the `_4` upload today. S10-R3/S12-R5 align with the new $0-only rule.",
     "OBSERVED build behavior (old build, 2026-07-13 labels): Bulk Receive allowed "
     "editing qty AND a non-zero cost (SF-BULK-06 was VIU-Verified against that old "
     "rule). The $0-only rule has NOT been observed yet — SF-BULK-06 was reworded to "
     "Δ14 on 2026-07-17 and flipped to VIU-Pending (re-VIU needed; if the build "
     "still allows editing a non-zero cost it becomes a deviation until dev ships). "
     "Case notes in group-B-receiving-vendor.json (SF-BULK-06).",
     "SF-BULK-06 C29355: in run 325 (Ayesha Khan) — see run325-status-map-2026-07-14"
     ".md (snapshot predates the 2026-07-17 Δ14 reword).",
     "A ($0-only wins) -> SF-BULK-06 stays as reworded (Δ14 $0-only cost + partial-"
     "receive qty note); the S8-R7 tail is confirmed leftover text; re-VIU against "
     "the $0-only rule (non-zero-cost editable => dev deviation). "
     "B (cost stays editable after lock) -> reword SF-BULK-06 expected back to "
     "cost-editable-after-lock per Milos's explanation and reconcile S8-R7 with "
     "S10-R3/S12-R5 (likely a fresh spec correction needed); re-VIU accordingly."),
    (2,
     "Vendors-Expenses exclusion surface confirm (Δ12 / S6-R6 rewrite)",
     [("SF-VMIS-06", 29343)],
     "Spec `_4` (V2.6) Δ12 (spec-v4-2026-07-17/spec-diff-v4-2026-07-17.md §A Δ12): "
     "S6-R6 REWRITTEN to match code — no dedicated PO report / no 'needs vendor' "
     "marker; a vendor-missing PO's spend \"does not appear in the QuickBooks "
     "Vendor Bill export (inner-join on vendor) and is not counted in the Vendors "
     "Expenses report\" until a vendor is assigned (change-log 2026-07-16, "
     "\"verified in VendorBillExportQueryHandler / VendorsExpensesQueryHandler\"). "
     "Question = confirm the REPORT SURFACE name (\"Vendors Expenses\") before the "
     "in-app re-VIU, so the exclusion is checked on the right report.",
     "OBSERVED build behavior: the old 2026-07-14 finding — the Reports area has NO "
     "'needs vendor' report — is now the spec itself (Deviation RESOLVED by "
     "rescope). The report leg is seedable (costed vendor-missing PO per Rule 14); "
     "the QB Vendor Bill export leg stays Blocked-Env (needs a QuickBooks-connected "
     "company + a human in QB). Case notes in group-B-receiving-vendor.json "
     "(SF-VMIS-06).",
     "SF-VMIS-06 C29343: in run 325 (Ayesha Khan) — see run325-status-map-2026-07-14"
     ".md (snapshot predates the 2026-07-17 Δ12 rescope).",
     "A (Vendors Expenses confirmed) -> re-VIU the report leg on that report (seed "
     "a costed vendor-missing PO, observe the exclusion, then assign a vendor and "
     "observe inclusion); QB leg stays Blocked-Env. "
     "B (different surface) -> reword SF-VMIS-06 expected to the named surface and "
     "re-VIU there; flag the S6-R6 report name for a spec correction."),
    (3,
     "Story-10 Δ7 residue — struck S10-R2 vs surviving AC bullets + technical guardrails",
     [("SF-PNFIX-02", 29364), ("SF-PNFIX-03", 29365), ("SF-PNFIX-06", 29368),
      ("SF-QB-08", 29433)],
     "PRE-EXISTING DOC SELF-CONTRADICTION (spec `_3` Δ7, still verbatim in `_4` — "
     "spec-v4-2026-07-17/spec-diff-v4-2026-07-17.md §D5 flag 3 + L58-59, L315, "
     "L416): S10-R2 (\"When a part number is added, the part becomes a first-class "
     "part\") is STRUCK THROUGH, but the surviving Story-10 AC bullets + technical "
     "guardrails + the §10 permission-matrix row still describe first-class-part "
     "creation. Last-update-wins applied the strike (QA-lead ruling 2026-07-14), "
     "but the leftover text has never been confirmed as clean-up by the PO.",
     "OBSERVED build behavior (VIU 2026-07-14, seeded vendor-missing PO S-15849): "
     "entering a NEW part number persists and the part becomes receivable "
     "(receive-requested-parts 200); the downstream first-class inventory/catalog/"
     "Part-History creation was DROPPED from the expected results per the Δ7 "
     "rescope (SF-PNFIX-02/03/06 + SF-QB-08 rescoped -> Verified on the remaining "
     "assertions). Case notes in group-B-receiving-vendor.json / group-C (SF-QB-08).",
     "SF-PNFIX-02 C29364 / SF-PNFIX-03 C29365 / SF-PNFIX-06 C29368 / SF-QB-08 "
     "C29433: in run 325 (Ayesha Khan) — see run325-status-map-2026-07-14.md "
     "(snapshot predates the rescope).",
     "A (feature out, leftover text) -> no case changes (the Δ7 rescope already "
     "matches); note the pending spec clean-up of the Story-10 AC bullets + "
     "guardrails + §10 matrix row. "
     "B (feature still in) -> REVERSE the Δ7 rescope: restore the first-class-part "
     "assertions to SF-PNFIX-02/03/06 + SF-QB-08, flip them off Verified pending "
     "re-VIU, and get the S10-R2 strike corrected in the spec."),
]

# ---------------------------------------------------------------------------
# No dropped section this round: these are the only 2 NEW Milos flags raised by
# the spec-`_4` apply (D5 flags 2 + the Δ12 surface confirm) plus the one
# PRE-EXISTING unconfirmed residue (D5 flag 3). D5 flag 1 (S12-R1 bottom vs
# S12-R3 leads) was already asked (Round-3 Q1 split ruling — answered, being
# folded); D5 flag 4 (design-vs-spec core-flow branch) is a DESIGN/dev item,
# not a PO A/B decision (rule 7).
# ---------------------------------------------------------------------------
not_sent_note = (
    "Not re-asked / not sent: D5 flag 1 (vendor-missing group \"bottom\" vs "
    "\"leads\") — already answered by Milos Round-3 Q1 (split ruling: Bulk Receive "
    "top / Receive bottom; the residual Receive-screen placement is dev deviation "
    "TICKET work, not a re-ask). D5 flag 4 (design files still showing the old "
    "resolve-after-receive core flow vs Story 18's resolve-first) — a design-"
    "revision/dev item, not a PO product decision (rule 7). The 5 unanswered "
    "earlier Milos questions remain on their existing sheets "
    "(PO-Questions-Round3 / PO-Decisions 2026-07-14) — not duplicated here."
)

# ===========================================================================
# XLSX
# ===========================================================================
wb = Workbook()

# --- Reader tab ---
ws = wb.active
ws.title = "Questions for you"
ws.column_dimensions["A"].width = 4
for col, w in zip("BCDEFG", [40, 46, 40, 40, 44, 26]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Simple Mode - Three Quick Questions on the Newest Write-Up"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
ws["A2"] = INTRO
ws["A2"].alignment = WRAP
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 90

headers = ["#", "The situation", "What the written spec currently says",
           "What the app actually does today", "Why it needs your decision",
           "The options", "Your decision"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, q in enumerate(questions, start=1):
    row = HDR_ROW + i
    vals = [i, q["situation"], q["spec_says"], q["app_does"], q["why"],
            q["opts"], ""]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 200

# --- QA Internal tab ---
wi = wb.create_sheet("Evidence & mapping")
wi["A1"] = ("INTERNAL - FOR THE QA LEAD ONLY. Do NOT share this tab (or any IDs / "
            "codes / clause numbers / links on it) with the PO.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:G1")

ihead = ["Q#", "Case (internal ID)", "TestRail Case ID", "TestRail link",
         "Exact spec clause (the citation)",
         "Build evidence (observed + pointer)",
         "Run-325 (Ayesha) status + remark",
         "What each answer option resolves to"]
iwid = [4, 15, 14, 42, 48, 50, 32, 58]
IH_ROW = 3
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for qno, title, cases, clause, evidence, ayesha, resolves in kept_map:
    first = r
    for sf_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        rowvals = [qno, sf_id, f"C{tr_id}", url, clause, evidence, ayesha, resolves]
        for c, v in enumerate(rowvals, start=1):
            cell = wi.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_CENTER if c in (1, 2, 3) else WRAP
            cell.border = BORDER
        lc = wi.cell(row=r, column=4)
        lc.hyperlink = url
        lc.font = LINK_FONT
        wi.row_dimensions[r].height = 150
        r += 1
    last = r - 1
    if last > first:
        for col in (1, 5, 6, 7, 8):
            wi.merge_cells(start_row=first, start_column=col,
                           end_row=last, end_column=col)

# --- Notes ---
r += 1
wi.cell(row=r, column=1, value=(
    "Notes: These are the 2 NEW Milos flags raised by the spec `_4` (V2.6) apply "
    "(D5 flag 2 = S8-R7 tail vs Δ14; the Δ12 Vendors-Expenses surface confirm) "
    "plus the 1 PRE-EXISTING unconfirmed residue (D5 flag 3 = S10-R2 strike vs "
    "the surviving Story-10 AC bullets + technical guardrails). " + not_sent_note +
    " TestRail IDs sourced from testrail-id-map.csv (standing rule 8); bugs stay "
    "off the PO sheet (standing rule 7). Spec citations: spec-v4-2026-07-17/"
    "spec-diff-v4-2026-07-17.md (§A Δ12/Δ14, §D5); requirements.md (V2.6).")
    ).alignment = WRAP
wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
wi.row_dimensions[r].height = 110

wb.save(XLSX_OUT)

# ===========================================================================
# Markdown mirror
# ===========================================================================
md = []
md.append("# Simple Mode - Three Quick Questions on the Newest Write-Up")
md.append("")
md.append(INTRO)
md.append("")
md.append("---")
for i, q in enumerate(questions, start=1):
    md.append("")
    md.append(f"## {i}.")
    md.append("")
    md.append("**The situation**")
    md.append(q["situation"])
    md.append("")
    md.append("**What the written spec currently says**")
    md.append(q["spec_says"])
    md.append("")
    md.append("**What the app actually does today**")
    md.append(q["app_does"])
    md.append("")
    md.append("**Why it needs your decision**")
    md.append(q["why"])
    md.append("")
    md.append("**The options**")
    for line in q["opts"].split("\n"):
        md.append(f"- {line}")
    md.append("")
    md.append("**Your decision:** ______________________________________________")
    md.append("")
    md.append("---")
md.append("")
md.append("Thank you! Just pick one option per row, or add a note. These three are "
          "the only points in the newest write-up where the text either disagrees "
          "with itself or leaves us unsure what to check - everything else we were "
          "able to confirm ourselves.")
md.append("")
md.append("---")
md.append("---")
md.append("")
md.append("## Internal - QA lead only (NOT for the PO)")
md.append("")
md.append("**Do not share this section (or any IDs / codes / clause numbers / links) "
          "with the PO.**")
md.append("")
md.append("### Kept questions - evidence & mapping")
md.append("")
for qno, title, cases, clause, evidence, ayesha, resolves in kept_map:
    md.append(f"#### Q{qno} - {title}")
    md.append("")
    md.append("- **TestRail cases:**")
    for sf_id, tr_id in cases:
        md.append(f"  - {sf_id} - [C{tr_id}]({TR_LINK.format(tr_id)})")
    md.append(f"- **Exact spec clause:** {clause}")
    md.append(f"- **Build evidence:** {evidence}")
    md.append(f"- **Run-325 (Ayesha):** {ayesha}")
    md.append(f"- **Resolves to:** {resolves}")
    md.append("")
md.append("### Not sent + why")
md.append("")
md.append(not_sent_note)
md.append("")
md.append("**Notes:** These are the 2 NEW Milos flags raised by the spec `_4` "
          "(V2.6) apply (D5 flag 2 + the Δ12 surface confirm) plus the 1 "
          "PRE-EXISTING unconfirmed residue (D5 flag 3). TestRail IDs from "
          "`testrail-id-map.csv` (rule 8); bugs stay off the PO sheet (rule 7). "
          "Spec citations: `spec-v4-2026-07-17/spec-diff-v4-2026-07-17.md` "
          "(§A Δ12/Δ14, §D5); `requirements.md` (V2.6). SF-CORE-05/06/09 were "
          "retired (user ruling 2026-07-17, executed 2026-07-20) and are not "
          "referenced here.")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md) + "\n")

print("Wrote:", XLSX_OUT)
print("Wrote:", MD_OUT)
