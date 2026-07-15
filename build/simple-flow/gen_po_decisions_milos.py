#!/usr/bin/env python3
"""Generate the GROUNDED PO decision sheet for Milos (Simple Flow PO).

Outputs (regenerable — run from the repo root):
  - build/simple-flow/SimpleFlow_PO-Decisions-for-Milos_2026-07-14.xlsx
  - build/simple-flow/SimpleFlow_PO-Decisions-for-Milos_2026-07-14.md

Why this deliverable exists
---------------------------
The earlier question sheets were dismissed as "assumption-based." This sheet is
the opposite: every question that is SENT is demonstrably grounded in FACT — it
shows (a) the exact clause in the CURRENT written spec that creates the
gap/contradiction, and (b) what the build actually does today (observed in
hands-on testing, with an evidence pointer). Anything that turned out to be an
assumption, or is already answered by the current spec, or that we can settle
ourselves, is DROPPED — with the reason recorded on the QA-internal tab.

Re-validation of the 8 previously "awaiting-Milos" cases against the CURRENT spec
(the 2026-07-14 `_3` upload = de-facto V2.5) + our build evidence:

  KEPT (3 cases -> 2 questions):
    - SF-RCV-05 + SF-RCV-07 : spec self-contradiction on vendor-missing group order
                              (S12-R1 "at the bottom" vs S12-R3 "leads"); build = TOP.
    - SF-REV-15             : S1-R4 "default per cohort (see §8)" + §8 lists the
                              new-org Require-Review preset as an OPEN QUESTION;
                              build has the toggle but no defined new-shop preset.

  DROPPED (5 cases):
    - SF-REV-11 : S1-R4 already makes review-before-invoicing a per-shop toggle;
                  §8 does NOT list it open; both legs built + VIU-observed ->
                  self-resolvable to VIU-Verified, not a product decision.
    - SF-UX-04  : S15-R4 already fully specifies Close/Cancel behavior; our case
                  wording already matches it; only the VISUAL close-confirm is
                  "design pending" (a design/build task, not a PO A/B decision).
    - SF-SET-08 : resolved by last-update-wins (spec first-use defaults
                  authoritative); the live-default mismatch is a DEV build gap
                  (GAP-B / bug draft T5), not a PO decision (rule 7).
    - SF-COMP-06: Milos already ruled Round-1 Q5 (POs always-on); residual
                  spec-vs-build is dev build-lag BUG-1, not a re-ask.
    - SF-QB-02  : same Create-POs-OFF scenario retired Round-1 Q5 AND QuickBooks
                  is not connected on sv7301 -> doubly non-actionable as a PO Q.

Rules honoured: no case IDs / codes / HTTP / jargon in the reader-facing tab
(standing rule 7); TestRail Case ID + clickable link on the QA-internal tab
(standing rule 8); no bugs put in front of the PO (rule 7 -> bugs to dev).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/simple-flow/SimpleFlow_PO-Decisions-for-Milos_2026-07-14.xlsx"
MD_OUT = "build/simple-flow/SimpleFlow_PO-Decisions-for-Milos_2026-07-14.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# READER-FACING content (layman ONLY — no IDs, no codes, no tech terms).
# Each question is grounded: spec_says quotes/paraphrases the actual clause,
# app_does states the observed build behavior.
# ---------------------------------------------------------------------------
INTRO = (
    "Hi Milos - thank you for the time on this. Below are two concrete points we "
    "ran into while testing the current build by hand against the latest written "
    "spec. These are not guesses: for each one we put the written spec and the "
    "app's real behaviour side by side. In each case the spec either disagrees "
    "with itself or leaves the point open, so only a product decision from you can "
    "settle it. Please just pick one option per row (or add a note). Thank you!"
)

questions = [
    {
        # SF-RCV-05 + SF-RCV-07
        "situation": (
            "When someone receives a delivery of parts, some of the parts don't yet "
            "have a supplier assigned. The app gathers all of those \"no supplier "
            "yet\" parts into their own group. The question is simply where that "
            "group should sit in the list."),
        "spec_says": (
            "The written spec gives TWO different answers in two different places. "
            "One place says the \"no supplier yet\" group should sit at the BOTTOM "
            "of the list. Another place says the very same group should LEAD - sit "
            "at the TOP. Both sentences are in the spec today, so the spec disagrees "
            "with itself."),
        "app_does": (
            "On the newer bulk-receiving screen, the \"no supplier yet\" group "
            "currently appears at the TOP of the list (it leads)."),
        "why": (
            "Because the spec says both \"top\" and \"bottom\" in different places, "
            "we genuinely cannot tell which one is correct. Only you can settle "
            "which the app should follow."),
        "opts": (
            "A) At the top (it leads) - this matches what the app does today.\n"
            "B) At the bottom of the list.\n"
            "C) Mixed in with all the other parts."),
    },
    {
        # SF-REV-15
        "situation": (
            "A brand-new shop opens the app for the very first time, having changed "
            "no settings yet. The app has an on/off setting for \"a job must be "
            "reviewed before it can be finished and billed.\" The question is only "
            "about brand-new shops - existing shops keep whatever they use today."),
        "spec_says": (
            "The spec leaves this open on purpose. It says the starting value should "
            "be decided \"per type of shop,\" and it lists this exact point under its "
            "own \"Open Questions\" as still undecided - including specifically what a "
            "brand-new shop should start with."),
        "app_does": (
            "The on/off setting exists on the settings screen and works, and existing "
            "shops keep their current behaviour. But because the spec never fixed a "
            "starting value for a brand-new shop, there is no defined starting value "
            "for a first-time shop for us to check against."),
        "why": (
            "The spec itself lists this as an open question and never set the starting "
            "value for a new shop, so there is nothing for us to test against. Only "
            "you can decide what a brand-new shop should start with."),
        "opts": (
            "A) Start turned ON for every new shop.\n"
            "B) Start turned OFF for every new shop.\n"
            "C) Start ON for larger/established shops and OFF for small new ones."),
    },
]

# ---------------------------------------------------------------------------
# QA-INTERNAL mapping (rule 8: TestRail C##### + clickable link per case).
# Each entry: (q_no, title, [(sf_id, tr_id), ...], spec_clause, build_evidence,
#              ayesha_run325, resolves_to)
# ---------------------------------------------------------------------------
kept_map = [
    (1,
     "Vendor-missing group order on the receive list",
     [("SF-RCV-05", 29373), ("SF-RCV-07", 29375)],
     "SPEC SELF-CONTRADICTION in the current spec (spec-source-2026-07-14.md): "
     "S12-R1 (L686) = \"...vendor-missing in their own group AT THE BOTTOM\" vs "
     "S12-R3 (L690) = \"...vendor-missing group LEADS\" (top). Both are present and "
     "unchanged in the _3 (V2.5) upload (spec-diff-2026-07-14.md §C Q4: "
     "\"top-vs-bottom ambiguity unchanged\").",
     "OBSERVED build behavior (VIU 2026-07-14, PROJECT-STATE §0-ZZ; seeded "
     "vendor-missing PO S-15845): on the Bulk Receive surface the \"Vendor Missing\" "
     "group renders as the FIRST / TOP group. Evidence: "
     "screenshots/grind-2026-07-14/OBSERVE-bulkreceive-groups.png (+ VMGATE-01..03). "
     "Case notes in group-B-receiving-vendor.json (SF-RCV-05/07).",
     "Both Untested in run 325 (Ayesha Khan) - no remark (SF-RCV-05 C29373, "
     "SF-RCV-07 C29375; run325-status-map-2026-07-14.md).",
     "A (TOP/leads) -> SF-RCV-05 expected #3 changed from \"at the bottom\" to "
     "\"leads (top)\" on every receive screen incl. Bulk Receive; SF-RCV-07 already "
     "says \"leads (top)\" - no change; BOTH flip VIU-observed-awaiting-Milos -> "
     "VIU-Verified (build already matches). "
     "B (BOTTOM) -> SF-RCV-05 stays \"at the bottom\" on the newer screen; the build "
     "(top) then DEVIATES from the chosen rule -> log a build deviation for dev. "
     "C (mixed) -> rewrite both expecteds; build deviates -> dev deviation."),
    (2,
     "New-shop starting default for \"require a review before completion\"",
     [("SF-REV-15", 29400)],
     "S1-R4 (L190, spec-source-2026-07-14.md) = \"Require review before completion. "
     "... Default PER COHORT (see §8).\" §8 Open Questions (L~1690) lists it "
     "unresolved: \"Require-review default - on for bigger/existing shops? + new-org "
     "preset (existing orgs keep today's behaviour via backfill).\" Unchanged in the "
     "_3 (V2.5) upload (spec-diff-2026-07-14.md §C Q2: default still \"per cohort "
     "(see §8)\"; §8 open).",
     "OBSERVED build behavior: the Require Review Before Completion toggle IS present "
     "and works on the Work Orders settings tab (screenshots/wording-2026-07-13/"
     "SET-workorders-tab.png; SF-SET-14 verified). Existing orgs keep their behavior. "
     "The brand-new-org PRESET cannot be observed on the long-lived shared sv7301 org "
     "(its value reflects prior test toggling, not first-use) -> no defined new-shop "
     "default exists to test. Case note SF-REV-15 (group-C...json).",
     "Untested in run 325 (Ayesha Khan) - no remark (SF-REV-15 C29400; "
     "run325-status-map-2026-07-14.md).",
     "A (ON for all new) -> SF-REV-15 expected = default ON for new orgs (if the "
     "live new-org default != ON that becomes a separate dev bug to verify). "
     "B (OFF for all new) -> expected = default OFF for new orgs. "
     "C (ON big / OFF small) -> expected = cohort-based new-org preset per the §8 "
     "wording; author the cohort split. In all cases existing orgs stay backfilled "
     "to today's behavior."),
]

# ---------------------------------------------------------------------------
# DROPPED (not sent) — the 5 of the 8 that are NOT genuine PO product decisions.
# (sf_id, tr_id, spec_status, build_status, drop_reason, self_resolution)
# ---------------------------------------------------------------------------
dropped = [
    ("SF-REV-11", 29396,
     "S1-R4 makes \"review before completion/invoicing\" a per-shop On/Off SETTING; "
     "§8 does NOT list \"make review mandatory\" as an open question. New Δ5/S16-R12 "
     "(auto-complete-on-last-line) confirms: Require Review OFF -> auto-Complete "
     "(invoice-ready); ON -> Ready for Review, invoicing blocked until signed off.",
     "Both legs BUILT + VIU-observed: direct sign-off completes the WO (Review->"
     "Complete, no separate final Complete); with Require Review ON invoicing is "
     "blocked until reviewed (SF-AUTO-05 verified the auto-complete trigger).",
     "NOT a product decision - the spec already answers it (it is a per-shop toggle, "
     "not an open item), and the behavior is already built and matches the spec. The "
     "original Round-1 Q8 phrasing was one Milos found confusing.",
     "SELF-RESOLVABLE: flip SF-REV-11 from VIU-observed-awaiting-Milos -> VIU-Verified "
     "(behavior matches S1-R4 + S16-R8 + Δ5/R12). No PO input needed."),
    ("SF-UX-04", 29404,
     "S15-R4 (L794) now FULLY specifies the behavior: \"Close = closes the modal "
     "only, no discard, stays on the WO (prominent/red); Cancel = closes the modal "
     "+ returns to the previous screen (text link, far left).\" Only a note remains: "
     "\"Design pending for the close-confirm specifically\" (the VISUAL only).",
     "The confirmation modal itself is not yet finished in the build (behavior not "
     "yet exercisable); our SF-UX-04 wording already matches S15-R4 exactly.",
     "NOT a product decision - the spec already defines the Close/Cancel behavior; "
     "what is outstanding is the VISUAL design + the build, i.e. a design/dev "
     "completion task, not a PO A/B choice.",
     "SELF-RESOLVABLE: case wording is already spec-accurate; keep it Blocked/pending "
     "until the modal is built, then VIU it. No PO input needed."),
    ("SF-SET-08", 29282,
     "Spec first-use defaults (§4: Auto-approve OFF / Create POs ON / Vendor Invoice "
     "REQUIRED) vs the design (ON / Optional). Already RECONCILED by last-update-wins "
     "-> spec is authoritative (contradiction-resolution.md: \"spec first-use "
     "defaults authoritative ... live defaults are a build gap (GAP-B)\").",
     "Brand-new-org first-use defaults are non-observable on the long-lived sv7301 "
     "org; the settings model exposes no createPurchaseOrders field.",
     "NOT a PO question - the decision is already made (spec wins). The remaining "
     "gap between spec defaults and the live defaults is a DEV build gap (GAP-B / bug "
     "draft T5); rule 7 keeps bugs off the PO sheet -> route to dev.",
     "Route to DEV as GAP-B / bug draft T5 (wrong live first-use defaults). Not sent "
     "to Milos."),
    ("SF-COMP-06", 29295,
     "V2.4/_3 spec documents a Create Purchase Orders toggle (S1-R2 \"Off -> no POs "
     "(default On)\"; §4 \"Create POs OFF => no PO\"). Milos ALREADY ruled Round-1 Q5 "
     "(\"we will Always have a PO\" = PO-OFF descoped); last-update-wins kept the "
     "V2.4 documentation, so the residual is a spec-vs-build lag.",
     "The Create Purchase Orders toggle is ABSENT from the settings tab and there is "
     "no createPurchaseOrders field; POs are always created for vendor parts "
     "(SF-SET-03 Deviation).",
     "NOT a re-ask - Milos already answered (Round-1 Q5, POs always-on). The residual "
     "spec-vs-build mismatch is tracked as build-lag BUG-1 (contradiction-resolution "
     "C2) = a dev item, not a PO decision.",
     "Route to DEV as build-lag BUG-1 (toggle not present). Not sent to Milos."),
    ("SF-QB-02", 29427,
     "Same Create-POs-OFF scenario as SF-COMP-06; the QuickBooks-integrity leg needs "
     "QB connected.",
     "Create Purchase Orders toggle absent (POs always-on) AND QuickBooks is not "
     "connected on sv7301 (no QB admin/API), so the scenario is doubly "
     "non-configurable/non-observable here.",
     "NOT a PO question - the Create-POs-OFF scenario was retired by Milos Round-1 Q5, "
     "and QB integrity is a Blocked-Env condition (needs a QB-connected company + a "
     "human in QB), not a product decision.",
     "Keep Blocked-Env (QB-not-connected) + covered by Milos Round-1 Q5 ruling. Not "
     "sent to Milos."),
]

# ===========================================================================
# XLSX
# ===========================================================================
wb = Workbook()

# --- Reader tab ---
ws = wb.active
ws.title = "Decisions we need from you"
ws.column_dimensions["A"].width = 4
for col, w in zip("BCDEFG", [40, 46, 40, 40, 44, 26]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Simple Mode - Decisions We Need From You"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")
ws["A2"] = INTRO
ws["A2"].alignment = WRAP
ws.merge_cells("A2:G2")
ws.row_dimensions[2].height = 90

headers = ["#", "The situation", "What the written spec currently says",
           "What the app actually does today", "Why it needs your decision",
           "The options", "Your decision"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, q in enumerate(questions, start=1):
    row = HDR_ROW + i
    vals = [i, q["situation"], q["spec_says"], q["app_does"], q["why"],
            q["opts"], ""]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 200

# --- QA Internal tab ---
wi = wb.create_sheet("Evidence & mapping")
wi["A1"] = ("INTERNAL - FOR THE QA LEAD ONLY. Do NOT share this tab (or any IDs / "
            "codes / clause numbers / links on it) with the PO.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:G1")

ihead = ["Q#", "Case (internal ID)", "TestRail Case ID", "TestRail link",
         "Exact spec clause (the citation)",
         "Build evidence (what we saw + pointer)",
         "Run-325 (Ayesha) status + remark",
         "What each answer option resolves to (on our side)"]
iwid = [4, 15, 14, 44, 50, 52, 34, 60]
IH_ROW = 3
# note: 8 columns now (added evidence + run325 + resolves); rebuild header
ihead = ["Q#", "Case (internal ID)", "TestRail Case ID", "TestRail link",
         "Exact spec clause (the citation)",
         "Build evidence (observed + pointer)",
         "Run-325 (Ayesha) status + remark",
         "What each answer option resolves to"]
iwid = [4, 15, 14, 42, 48, 50, 32, 58]
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for qno, title, cases, clause, evidence, ayesha, resolves in kept_map:
    first = r
    for sf_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        rowvals = [qno, sf_id, f"C{tr_id}", url, clause, evidence, ayesha, resolves]
        for c, v in enumerate(rowvals, start=1):
            cell = wi.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_CENTER if c in (1, 2, 3) else WRAP
            cell.border = BORDER
        lc = wi.cell(row=r, column=4)
        lc.hyperlink = url
        lc.font = LINK_FONT
        wi.row_dimensions[r].height = 150
        r += 1
    last = r - 1
    if last > first:
        for col in (1, 5, 6, 7, 8):
            wi.merge_cells(start_row=first, start_column=col,
                           end_row=last, end_column=col)

# --- Dropped section ---
r += 1
dtitle = wi.cell(row=r, column=1, value=(
    "DROPPED (NOT sent to the PO) - of the 8 previously \"awaiting-Milos\" cases, "
    "these 5 are NOT genuine product decisions once re-validated against the current "
    "spec + build (assumption / already-answered / self-resolvable / dev-bug). Kept "
    "here for the rationale."))
dtitle.font = Font(bold=True, color="1F4E79")
dtitle.alignment = WRAP
wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
wi.row_dimensions[r].height = 45
r += 1
dhead = ["", "Case (internal ID)", "TestRail Case ID", "TestRail link",
         "Spec position", "Build status", "Why dropped", "Self-resolution / routing"]
for c, h in enumerate(dhead, start=1):
    cell = wi.cell(row=r, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
r += 1
for sf_id, tr_id, spec_status, build_status, reason, selfres in dropped:
    url = TR_LINK.format(tr_id)
    rowvals = ["", sf_id, f"C{tr_id}", url, spec_status, build_status, reason, selfres]
    for c, v in enumerate(rowvals, start=1):
        cell = wi.cell(row=r, column=c, value=v)
        cell.alignment = WRAP_CENTER if c in (2, 3) else WRAP
        cell.border = BORDER
    lc = wi.cell(row=r, column=4)
    lc.hyperlink = url
    lc.font = LINK_FONT
    wi.row_dimensions[r].height = 130
    r += 1

# --- Notes ---
r += 1
wi.cell(row=r, column=1, value=(
    "Notes: Re-validated the 8 previously awaiting-Milos cases against the CURRENT "
    "spec (2026-07-14 _3 upload = de-facto V2.5) + our live build evidence. KEPT 3 "
    "cases -> 2 questions (both grounded in a spec self-contradiction or a spec-"
    "declared open question + observed build behavior). DROPPED 5 (already answered "
    "by the spec / a prior Milos ruling, self-resolvable, or a dev bug). TestRail IDs "
    "sourced from testrail-id-map.csv (standing rule 8); bugs stay off the PO sheet "
    "(standing rule 7). Spec citations: spec-source-2026-07-14.md; deltas: "
    "spec-diff-2026-07-14.md; reconciliation history: contradiction-resolution.md.")
    ).alignment = WRAP
wi.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
wi.row_dimensions[r].height = 90

wb.save(XLSX_OUT)

# ===========================================================================
# Markdown mirror
# ===========================================================================
md = []
md.append("# Simple Mode - Decisions We Need From You")
md.append("")
md.append(INTRO)
md.append("")
md.append("---")
for i, q in enumerate(questions, start=1):
    md.append("")
    md.append(f"## {i}.")
    md.append("")
    md.append("**The situation**")
    md.append(q["situation"])
    md.append("")
    md.append("**What the written spec currently says**")
    md.append(q["spec_says"])
    md.append("")
    md.append("**What the app actually does today**")
    md.append(q["app_does"])
    md.append("")
    md.append("**Why it needs your decision**")
    md.append(q["why"])
    md.append("")
    md.append("**The options**")
    for line in q["opts"].split("\n"):
        md.append(f"- {line}")
    md.append("")
    md.append("**Your decision:** ______________________________________________")
    md.append("")
    md.append("---")
md.append("")
md.append("Thank you! Just pick one option per row, or add a note. These two are the "
          "only points where the written spec either disagrees with itself or leaves "
          "the answer open - everything else we were able to confirm ourselves.")
md.append("")
md.append("---")
md.append("---")
md.append("")
md.append("## Internal - QA lead only (NOT for the PO)")
md.append("")
md.append("**Do not share this section (or any IDs / codes / clause numbers / links) "
          "with the PO.**")
md.append("")
md.append("### Kept questions - evidence & mapping")
md.append("")
for qno, title, cases, clause, evidence, ayesha, resolves in kept_map:
    md.append(f"#### Q{qno} - {title}")
    md.append("")
    md.append("- **TestRail cases:**")
    for sf_id, tr_id in cases:
        md.append(f"  - {sf_id} - [C{tr_id}]({TR_LINK.format(tr_id)})")
    md.append(f"- **Exact spec clause:** {clause}")
    md.append(f"- **Build evidence:** {evidence}")
    md.append(f"- **Run-325 (Ayesha):** {ayesha}")
    md.append(f"- **Resolves to:** {resolves}")
    md.append("")
md.append("### Dropped (not sent) + why")
md.append("")
md.append("Of the 8 previously \"awaiting-Milos\" cases, these 5 are NOT genuine PO "
          "product decisions once re-validated against the current spec + build:")
md.append("")
for sf_id, tr_id, spec_status, build_status, reason, selfres in dropped:
    md.append(f"- **{sf_id}** - [C{tr_id}]({TR_LINK.format(tr_id)})")
    md.append(f"  - *Spec position:* {spec_status}")
    md.append(f"  - *Build status:* {build_status}")
    md.append(f"  - *Why dropped:* {reason}")
    md.append(f"  - *Self-resolution / routing:* {selfres}")
    md.append("")
md.append("**Notes:** Re-validated the 8 previously awaiting-Milos cases against the "
          "CURRENT spec (2026-07-14 `_3` upload = de-facto V2.5) + live build "
          "evidence. KEPT 3 cases -> 2 questions (both grounded in a spec "
          "self-contradiction or a spec-declared open question PLUS observed build "
          "behavior). DROPPED 5. TestRail IDs from `testrail-id-map.csv` (rule 8); "
          "bugs stay off the PO sheet (rule 7). Citations: "
          "`spec-source-2026-07-14.md`, `spec-diff-2026-07-14.md`, "
          "`contradiction-resolution.md`.")
md.append("")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md))

print(f"Wrote {XLSX_OUT} and {MD_OUT}: {len(questions)} kept PO decisions, "
      f"{len(dropped)} dropped.")
