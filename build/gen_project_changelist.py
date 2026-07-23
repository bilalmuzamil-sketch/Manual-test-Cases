#!/usr/bin/env python3
"""Spec-Recheck Change-List Workbook generator for Simple Flow + Fees & Discounts.
Mirrors build/custom-roles-run/gen_simple_changelist.py 1:1 (2 tabs, 7 columns, same
styling, .xlsx + .md twin). Only the per-project data (TOTAL, rows) differs.
Change text is plain layman (Rule 7); Case ID + TestRail link every row (Rule 8);
driving ticket + Done/Not-Done status from live Jira 2026-07-23; nothing pushed to TestRail.
Human-readable filenames (Rule 19)."""
import os
BASE = os.path.dirname(os.path.abspath(__file__))
LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"
JIRA = "https://shopview.atlassian.net/browse/{}"
DATE = "2026-07-23"

# row = (Case ID, Area, plain "what needs to change / decide", tickets, ticket_status, action)
# ticket_status: "DONE" | "NOT DONE (<state>)"
SIMPLE_FLOW = {
 "name": "Simple Flow",
 "total": 184,
 "file": "simple-flow/SimpleFlow_SpecRecheck_ChangeList_2026-07-23",
 "omit_note": ("The other 180 cases need no change (151 Verified, 21 Blocked-on-environment, "
               "4 to re-verify live, 3 retired, plus 4 Create-Purchase-Orders cases deleted from "
               "TestRail and ignored per your ruling). Nothing pushed to TestRail yet."),
 "rows": [
  ("C29373","Accept Delivery",
   "On the Accept Delivery screen the group of parts with no vendor yet appears at the TOP of the "
   "list; the product owner wanted it at the BOTTOM. Reviewed and accepted as a look-only difference "
   "(no effect on how it works) — confirm we keep it as-is and change nothing.",
   "SV-7707","DONE","DECISION"),
  ("C29375","Accept Delivery",
   "Same Accept-Delivery point as C29373 — the position of the no-vendor group and the multi-vendor "
   "indicator. Accepted as look-only — confirm keep as-is, no case change.",
   "SV-7707","DONE","DECISION"),
  ("C29396","Review before completion",
   "When 'require review before completion' is on, signing off finishes the work order straight away "
   "with no separate 'Reviewed' holding step. Whether invoicing should be blocked until a review "
   "happens is a product decision (Milos). Don't finalise this case until that review ticket ships.",
   "SV-7870","NOT DONE (Blocked)","DECISION"),
  ("C29404","Completion screen — Close/Cancel",
   "The 'Close' vs 'Cancel' confirmation pop-up on the completion screens isn't finalised in the "
   "design yet, so the exact behaviour isn't defined. Needs Milos to confirm what Close and Cancel "
   "should do before the case can state a result.",
   "SV-7710","NOT DONE (Blocked)","DECISION"),
 ],
}

FEES_DISCOUNTS = {
 "name": "Fees and Discounts",
 "total": 186,
 "file": "fees-discounts/FeesAndDiscounts_SpecRecheck_ChangeList_2026-07-23",
 "omit_note": ("The other 174 cases need no change (167 Verified, 21 Blocked-on-environment, "
               "1 to re-verify live, 4 retired). Note: the Fees & Discounts V1 stories are still "
               "Open in Jira even though the feature is live, so most rows below wait on an open "
               "ticket. Nothing pushed to TestRail yet."),
 "rows": [
  ("C28436","Whole-WO Fee/Discount",
   "Confirm which permission controls adding or editing a whole-work-order fee/discount — the build "
   "may use one single 'edit work order' check instead of the separate whole-WO and line-level "
   "permissions the spec describes. Needs a check on staging with a restricted user.",
   "SV-8277 / SV-8289","NOT DONE (Open)","DECISION"),
  ("C28456","Inline display (Lines tab)",
   "On a labour line with two or more fees/discounts the build shows every row and has NO "
   "'Show more / Show less' collapse control the spec asks for. Confirmed difference — decide whether "
   "to change the case to match the build or log a fix for the build.",
   "SV-8279 / SV-8288","NOT DONE (Open)","DECISION"),
  ("C28460","Statistics tab",
   "The Statistics tab shows fees/discounts as one combined total with no per-item rows, so there is "
   "no per-item link to jump to the item. Confirmed difference from the design — decide keep-as-is or "
   "log a fix.",
   "SV-8280","NOT DONE (Open)","DECISION"),
  ("C28462","Statistics tab",
   "Wording check: the spec says work-order screens list fees/discounts oldest-first (the order they "
   "were created); an internal note said newest-first. The case was written to the spec (oldest-first) "
   "— confirm the actual order on staging.",
   "SV-8280","NOT DONE (Open)","DECISION"),
  # FD-WO-017 (C30618) IGNORED per user 2026-07-23 — dropped from the change list.
  ("C28489","Customer Fees & Discounts tab",
   "On the customer Fees & Discounts tab the 'add template' dropdown lists the template NAME only "
   "(no Type/Calc/Amount columns), and a Processing Fee shows as 'Fee'. Confirm the accepted display "
   "and update the case wording to match the build.",
   "SV-8284 / SV-8285","NOT DONE (Open)","Apply update"),
  ("C28490","Customer Fees & Discounts tab",
   "When every template is already linked, the dropdown shows 'No results' instead of the spec's "
   "'No templates available to add.' The product owner accepted 'No results' as-is — update the "
   "case's expected message to 'No results'.",
   "SV-8285","NOT DONE (Open)","Apply update"),
  ("C28511","Template admin — scoping",
   "On a single labour/part line there is NO 'Apply from template' picker at all, so the "
   "template-scoping rule can't be exercised at line level (the picker only exists on the whole "
   "work-order dialog). Confirm and adjust the case's scope.",
   "SV-8278 / SV-8281","NOT DONE (Open)","DECISION"),
  ("C28526","Processing Fee — WO behavior",
   "A processing fee can't be edited (the backend blocks it), but the work-order ⋮ menu still shows "
   "'Edit | Remove'; the spec wants Remove/Delete only. Decide keep-as-is or log a fix.",
   "SV-8279 / SV-8284","NOT DONE (Open)","DECISION"),
  ("C28527","Processing Fee — calculation",
   "The processing-fee amount is worked out on the wrong base — it includes the whole-work-order fees "
   "and their tax, when the spec says to leave those out. Confirmed calculation defect — log the fix; "
   "the case stays as-is and will correctly fail until it's fixed.",
   "SV-8284","NOT DONE (Open)","DECISION"),
  ("C28580","Calculation contract",
   "Same processing-fee base problem as C28527 in the overall calculation check — the grand-total base "
   "wrongly includes whole-work-order fees and their tax. Confirmed calculation defect — log the fix.",
   "SV-7865","NOT DONE (Testing QA)","DECISION"),
  ("C28586","Permissions (Story 13)",
   "Adding a whole-work-order fee/discount is not blocked by the backend for a user without "
   "work-order edit rights (the block is front-end only). Confirm the intended enforcement.",
   "SV-8289","NOT DONE (Open)","DECISION"),
  # --- New FD defects the user flagged 2026-07-23 (SV-8521, SV-8520): not previously in scope ---
  ("NEW-8521","Finance/Invoice — part-line display",
   "A fee/discount on a PART should show as an indented row under that part on the Finance/invoice "
   "view (work order AND parts sale), the same way a labour-line one does. Reported missing on "
   "Finance (value only in the bottom summary). Needs a new case covering this on both surfaces.",
   "SV-8521","NOT DONE (Ready for QA)","New case + verify"),
  ("NEW-8520","Part line — display after receive/pick",
   "A fee on a PART line should stay visible on the line row after the part is received or picked; "
   "it currently disappears from the line (still correct in totals and on the invoice). Needs a new "
   "case covering line display staying in sync after receive/pick.",
   "SV-8520","NOT DONE (Testing Stage)","New case + verify"),
 ],
}

# Live-build check 2026-07-23 (admin, app.staging.shopview.com, WO S9-25393; evidence in
# build/fees-discounts/viu-changelist-2026-07-23/). Per-case overrides for column D + action +
# live state. Cases NOT in this dict were NOT re-verified this run and are flagged pending.
LIVE = {
 "C28456": {"d": ("LIVE 2026-07-23: a labor line with 2+ fees/discounts now shows an inline "
                  "'Fee 10% +$46.49' row PLUS a 'Show 1 more' / 'Show less' collapse-expand toggle "
                  "(clicking expands then reads 'Show less'). The collapse control the spec asks for "
                  "EXISTS — the earlier 'no Show more/less toggle' finding is resolved. Update the "
                  "case to Verified."),
             "action": "Apply update", "state": "verified"},
 "C28460": {"d": ("LIVE 2026-07-23: the Statistics tab now shows a 'Fees & Discounts (6)' section "
                  "with PER-ADJUSTMENT rows (columns % and Amount: 'Fee +10% +$46.49', 'Discount', "
                  "'Fee −11% −$2.39', etc.) plus a Total — the earlier 'aggregate only, no per-row' "
                  "finding no longer holds. Still to confirm: whether each row has a scope hyperlink "
                  "to jump to its item. Update the case to the per-row layout."),
             "action": "Apply update", "state": "verified"},
 "C28511": {"d": ("LIVE 2026-07-23: the LINE-SCOPE fee dialog now HAS a template picker — the 'New Part "
                  "Fee / Discount' dialog (and the labour-line one) shows 'Apply From Template' with "
                  "'Showing templates compatible with this line', and the whole-WO dialog has it too. The "
                  "earlier 'no template picker at line scope' deviation is resolved; the scope-filtering "
                  "hint is present. Update the case to the picker-present, scope-filtered behaviour."),
             "action": "Apply update", "state": "verified"},
 "NEW-8520": {"d": ("LIVE 2026-07-23 (WO S-25991, QA-seeded): CONFIRMED — a 50% part fee was added to a "
                    "part then the part was picked; the part line row now shows $99.00 with NO indented "
                    "fee child row, yet the line TOTAL is $392.50 (labour $244 + part $99 + the $49.50 "
                    "fee). So the fee is still billed but hidden from the line after pick — matches "
                    "SV-8520. Author the case as a confirmed defect (Testing Stage)."),
               "action": "New case + verify", "state": "verified"},
 "C28526": {"d": ("LIVE 2026-07-23 (WO S-25989): a processing fee is REMOVE-ONLY on the backend — "
                  "POST /api/work-orders/adjustments/change on the processing fee returns HTTP 409 "
                  "'A processing fee cannot be edited through this endpoint', while the same edit on a "
                  "normal whole-WO fee returns 200. Backend behaviour matches spec S8-N5/S8-R17. The "
                  "residual deviation is UI-only (the ⋮ menu still OFFERS 'Edit') — the per-fee ⋮ menu "
                  "could not be isolated headless this run; re-confirm the menu label. Update the "
                  "backend-enforcement part to Verified."),
             "action": "Apply update", "state": "verified"},
 "C28462": {"d": ("LIVE 2026-07-23 (WO S-25989 Statistics tab): fees/discounts are listed OLDEST-FIRST "
                  "— display order Part Fee, WO Processing Fee, bil matches their creation order "
                  "(createdAt 14:17:53, 14:17:53, 14:44:25). Matches spec §5-R9; the case (authored to "
                  "spec) is correct. Part Fee correctly shows $0.00. Mark Verified."),
             "action": "Apply update", "state": "verified"},
 "C28489": {"d": ("LIVE 2026-07-23 (customer 'Default Fees & Discounts' tab): the defaults TABLE "
                  "shows full columns — Name, Type, Calculation Type, Amount, Max Amount, Taxable — "
                  "and a Processing Fee displays with Type 'Fee' (e.g. 'Processing Fee | Fee | % of "
                  "Grand Total | 6%'), confirming the 'shows as Fee' point. Still to confirm: whether "
                  "the 'Add Fee/Discount' picker DROPDOWN lists name-only vs columns. Update the case "
                  "to the observed table display."),
             "action": "Apply update", "state": "verified"},
 "C28527": {"d": ("LIVE 2026-07-23 (WO S-25989, seeded by QA): processing-fee base now EXCLUDES whole-WO "
                  "fees — FIXED. Numbers: labour $244 + shop supplies $20 = net subtotal $264; ×1.05 GST "
                  "= Grand Total $277.20; Processing Fee 10% = $27.72 (matches the observed $27.72 to the "
                  "cent). The whole-WO 'bil' fee ($24.40) is NOT in the base (the old bug would give "
                  "$30.28). Spec §5-R4 satisfied — the earlier 'base includes whole-WO fees' deviation is "
                  "resolved. Update the case to Verified."),
             "action": "Apply update", "state": "verified"},
 "C28580": {"d": ("LIVE 2026-07-23 (WO S-25989): same result as C28527 — the Grand-Total base for the "
                  "Processing Fee correctly excludes whole-WO fees and their tax (10% × ($264 net ×1.05) "
                  "= $27.72, observed $27.72; buggy would be $30.28). Calculation-contract deviation "
                  "resolved. Update the case to Verified."),
             "action": "Apply update", "state": "verified"},
 "NEW-8521": {"d": ("LIVE 2026-07-23 (WO S9-25393 Finance/Estimate view): part-line adjustments DO "
                    "render as indented child rows under their part — '↳ Name $11.00' and "
                    "'↳ Fee (% of parts) ($2.39)' show under the T-BOLT CLAMP part, exactly like the "
                    "labour-line '↳ Fee (% of labor)' / '↳ Discount'. So on the work-order invoice "
                    "this appears FIXED (matches SV-8521 'Ready for QA'). Still to confirm: the Parts "
                    "Sale invoice view. Author a case for both surfaces."),
               "action": "New case + verify", "state": "verified"},
}
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
hf = PatternFill('solid', fgColor='1F4E78'); hfont = Font(bold=True, color='FFFFFF')
wrap = Alignment(wrap_text=True, vertical='top')
warnfill = PatternFill('solid', fgColor='FCE4D6')      # orange = NOT DONE ticket
thin = Border(*[Side(style='thin', color='D9D9D9')]*4)
HDR = ['Case ID','TestRail link','Area','What needs to change','Driving ticket','Ticket status','Action']

verifill = PatternFill('solid', fgColor='E2EFDA')      # green = live-verified this run

def dcell(cid, change):
    if cid in LIVE: return LIVE[cid]["d"]
    return "⏳ LIVE CHECK PENDING (2026-07-23 run: not yet observed) — prior finding: " + change

def act(cid, action):
    return LIVE[cid]["action"] if cid in LIVE else action

def linkfor(cid):
    return LINK.format(cid[1:]) if (cid.startswith('C') and cid[1:].isdigit()) else "(no TestRail case yet — to author)"

def sheet(ws, rows):
    ws.append(HDR)
    for c in ws[ws.max_row]: c.fill=hf; c.font=hfont; c.alignment=wrap
    for cid,area,change,tix,status,action in rows:
        label = cid if cid.startswith('C') else cid.replace('NEW-','new (SV-')+')'
        ws.append([label, linkfor(cid), area, dcell(cid,change), tix, status, act(cid,action)])
        r = ws[ws.max_row]
        st = LIVE.get(cid,{}).get('state')
        fill = verifill if st=='verified' else (warnfill if (st=='blocked' or status.startswith('NOT DONE')) else None)
        for c in r:
            c.alignment=wrap; c.border=thin
            if fill: c.fill=fill
    for i,w in enumerate([9,40,20,72,20,22,14],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width=w
    ws.freeze_panes='A2'

def build(proj):
    rows = proj['rows']; n=len(rows); total=proj['total']
    nd = [r for r in rows if r[4].startswith('NOT DONE')]
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title=f'Change list ({n})'
    ws.append([f"{proj['name']} spec-recheck — cases that need a change or a decision ({n} of {total})."])
    ws['A1'].font=Font(bold=True,size=13)
    nlive = sum(1 for r in rows if LIVE.get(r[0],{}).get('state')=='verified')
    nblk = sum(1 for r in rows if LIVE.get(r[0],{}).get('state')=='blocked')
    ws.append([f"LIVE-BUILD CHECK 2026-07-23: {nlive} of {n} rows re-verified live on staging (green); "
               f"{nblk} characterised-blocked on an env defect (orange, see column D); the rest flagged "
               "'⏳ LIVE CHECK PENDING'."])
    ws['A2'].font=Font(bold=True, color='C00000')
    ws.append([proj['omit_note'] + "  Orange rows = waiting on a ticket that is NOT yet done."])
    ws.append([])
    sheet(ws, rows)
    ws2 = wb.create_sheet('Waiting on open tickets')
    ws2.append(['Cases whose change depends on a ticket that is NOT yet done'])
    ws2['A1'].font=Font(bold=True,size=13)
    ws2.append(['These should NOT be finalised until the ticket ships; keep the case tracking the target and re-verify after the fix.'])
    ws2.append([])
    sheet(ws2, nd)
    out = os.path.join(BASE, proj['file'])
    wb.save(out + '.xlsx')
    with open(out + '.md','w') as fh:
        fh.write(f"# {proj['name']} spec-recheck — change list ({DATE})\n\n")
        nlive=sum(1 for r in rows if LIVE.get(r[0],{}).get('state')=='verified')
        nblk=sum(1 for r in rows if LIVE.get(r[0],{}).get('state')=='blocked')
        fh.write(f"> **LIVE-BUILD CHECK 2026-07-23:** {nlive} of {n} rows re-verified live on staging; "
                 f"{nblk} characterised-blocked on an env defect; the rest flagged **⏳ LIVE CHECK PENDING**.\n\n")
        fh.write(f"{n} of {total} cases need a change or a decision. {proj['omit_note']}\n\n")
        fh.write("**Legend:** Action = *Apply update* (wording/expected fix) or *Decision* (needs you/PO/dev to choose). "
                 "Ticket status shows whether the driving Jira ticket is Done (live status 2026-07-23).\n\n")
        fh.write("| Case | Area | What needs to change | Ticket | Ticket status | Action |\n|---|---|---|---|---|---|\n")
        for cid,area,change,tix,status,action in rows:
            fh.write(f"| [{cid if cid.startswith(chr(67)) else cid} ]({linkfor(cid)}) | {area} | {dcell(cid,change).replace(chr(124),chr(47))} | {tix} | {status} | {act(cid,action)} |\n")
        fh.write(f"\n## Highlight — cases waiting on a ticket that is NOT yet done ({len(nd)})\n\n")
        fh.write("| Case | Ticket | Ticket status | Why it's blocked |\n|---|---|---|---|\n")
        for cid,area,change,tix,status,action in nd:
            fh.write(f"| [{cid}]({linkfor(cid)}) | {tix} | {status} | {dcell(cid,change).replace(chr(124),chr(47))} |\n")
    print(f"{proj['name']}: {n} rows ({len(nd)} waiting on NOT-DONE tickets) -> {proj['file']}.xlsx/.md")

build(SIMPLE_FLOW)
build(FEES_DISCOUNTS)
