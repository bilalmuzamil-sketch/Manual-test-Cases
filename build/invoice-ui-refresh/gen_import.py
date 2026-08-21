#!/usr/bin/env python3
"""Generate the TestRail import for Invoice UI Refresh (epic SV-8218, spec v38).

Mirrors build/report-suite/gen_import.py 1:1 (Rule 16): same HEADER, same clean()/
joinlines() (incl. the shredding guard), CRLF rows. References column carries the
case's `refs` (ticket + spec anchor, Rule 8/20). Sections = case `area`. All cases
are source-verified-only (Rule 85); their AUTOMATION marker is exported verbatim.

Regenerating BLANKS the id-map testrail_case_id column and there is no refs column in
the id-map by default here (kept simple, greenfield); re-merge C-IDs from live after a
push (core 3.6).
"""
import csv, glob, json, os, re, sys

ROOT = "build/invoice-ui-refresh"
CASES = sorted(glob.glob(os.path.join(ROOT, "cases", "cases-*.json")))
OUT_CSV = os.path.join(ROOT, "testrail-import", "Invoice-UI-Refresh_testrail-import.csv")
OUT_XLSX = os.path.join(ROOT, "testrail-import", "Invoice-UI-Refresh_testrail-import.xlsx")
OUT_IDMAP = os.path.join(ROOT, "testrail-id-map.csv")

HEADER = ["Title", "Section", "Type", "Priority",
          "Preconditions", "Steps", "Expected Result", "References", "", ""]

IDPAT = r"INV-[A-Z]+-\d+"

def clean(s):
    if not s:
        return s
    s = re.sub(r"\s*\((?:see|per|from)\s+" + IDPAT +
               r"(?:(?:,|\s+and)\s+" + IDPAT + r")*\)", "", s)
    s = re.sub(r"\s*\(" + IDPAT + r"\)", "", s)
    s = re.sub(r"[,;]?\s*(?:see\s+)?" + IDPAT, "", s)
    return s

def joinlines(v):
    if v is None:
        return ""
    if isinstance(v, str):      # shredding guard (core 3.7): split a string first
        v = v.split("\n")
    return "\n".join(clean(str(x).rstrip()) for x in v)

def load():
    out = []
    for f in CASES:
        for c in json.load(open(f)):
            st = str(c.get("viu_status", "")).lower()
            if st.startswith("retired"):
                continue
            out.append(c)
    return out

def main():
    cases = load()
    rows, idmap, titles_by_section = [], [], []
    for c in cases:
        title = clean(c["title"].strip())
        section = c["area"].strip()
        titles_by_section.append((section, title))
        idmap.append([c["id"], "", title, section, c.get("refs", "").strip()])
        rows.append([
            title, section,
            c.get("type", "Functional"), c["priority"].strip(),
            joinlines(c.get("preconditions")),
            joinlines(c.get("steps")),
            joinlines(c.get("expected")),
            (c.get("refs") or "").strip(),
            "", "",
        ])
    # CSV
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
        w.writerow(HEADER); w.writerows(rows)
    print("Wrote CSV:", OUT_CSV, "data rows:", len(rows))
    # XLSX
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        wb = Workbook(); ws = wb.active; ws.title = "Invoice UI Refresh"
        ws.append(HEADER)
        for cix in range(1, len(HEADER)+1):
            ws.cell(row=1, column=cix).font = Font(bold=True)
        for r in rows:
            ws.append(r)
        for r in range(2, len(rows)+2):
            for cix in range(1, len(HEADER)+1):
                ws.cell(row=r, column=cix).alignment = Alignment(wrap_text=True, vertical="top")
        widths=[42,26,12,10,40,40,60,26,3,3]
        from openpyxl.utils import get_column_letter
        for i,wd in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = wd
        wb.save(OUT_XLSX); print("Wrote XLSX:", OUT_XLSX)
    except Exception as e:
        print("XLSX skipped:", e)
    # id-map
    with open(OUT_IDMAP, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["internal_id", "testrail_case_id", "title", "section", "refs"])
        w.writerows(idmap)
    print("Wrote id-map:", OUT_IDMAP, "rows:", len(idmap))
    # --- sanity / shredding guard / leak checks ---
    from collections import Counter
    blob = "\n".join("\t".join(str(x) for x in r) for r in rows)
    dupes = [t for t,n in Counter(titles_by_section).items() if n>1]
    print("Duplicate titles within a section:", dupes or "NONE")
    print("Internal-id leaks in exported cells:", re.findall(IDPAT, blob) or "NONE")
    print("VIU-word occurrences:", blob.lower().count("viu"))
    # shredding guard: no field should have a newline between single characters
    shred = [r[0] for r in rows if re.search(r"(?:\S\n){6,}", r[4]+r[5]+r[6])]
    print("Shredding suspects:", shred or "NONE")
    # every row has exactly one AUTOMATION marker + one provenance line
    bad = [r[0] for r in rows if r[6].count("AUTOMATION:")!=1 or r[6].count("This is the expected behaviour")!=1]
    print("Rows w/o exactly one marker+provenance:", bad or "NONE")
    print("Total exported:", len(rows))

if __name__ == "__main__":
    main()
