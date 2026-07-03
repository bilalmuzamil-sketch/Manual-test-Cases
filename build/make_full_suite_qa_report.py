#!/usr/bin/env python3
"""Build a non-technical, manual-QA-readable Excel workbook for the FULL
Custom Roles test-suite run (batches A-E) with the blocked-recovery re-run
applied as status overrides.

Digital Inspections (DI-prefixed) cases are EXCLUDED per request, leaving
176 of the original 297 cases.

Output: CustomRoles_TestSuite_QA_Report.xlsx (repo root)
Tabs: Summary | Passed Tests | Failed Tests | Blocked Tests | Not Applicable
"""
import csv
import json
import os
import re

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # repo root
RESULTS = os.path.join(BASE, "build", "test-run-results")
OUT = os.path.join(BASE, "CustomRoles_TestSuite_QA_Report.xlsx")

BATCHES = ["batchA", "batchB", "batchC", "batchD", "batchE"]

# ---- styling constants (same look-and-feel as the SV5319 QA report) ----
HEADER_FILL = PatternFill("solid", fgColor="1F3864")       # dark blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
BODY_FONT = Font(size=11)
BODY_ALIGN = Alignment(vertical="top", wrap_text=True)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")          # very light blue-grey
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")        # light green
RED_FILL = PatternFill("solid", fgColor="FFC7CE")          # light red
AMBER_FILL = PatternFill("solid", fgColor="FFE699")        # amber / yellow
GREY_FILL = PatternFill("solid", fgColor="D9D9D9")         # grey
GRAND_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

GROUPS = [
    ("SP", "Role Permissions (individual permission toggles)"),
    ("TE", "Built-in Role Templates"),
    ("CB", "Combined Scenarios"),
]
GROUP_ORDER = {p: i for i, (p, _) in enumerate(GROUPS)}
GROUP_NAME = dict(GROUPS)


# ---------------------------------------------------------------- loading
def load_cases():
    cases = []
    for b in BATCHES:
        with open(os.path.join(RESULTS, b + ".json")) as fh:
            cases.extend(json.load(fh))
    return cases


def load_recovery():
    with open(os.path.join(RESULTS, "blocked-recovery.json")) as fh:
        return json.load(fh)


def apply_overrides(cases, recovery):
    """Apply blocked-recovery results as overrides.

    Per the earlier decision, Digital Inspections (DI) cases were ignored by
    the recovery pass, so they keep their original status/actual/notes.
    Returns (unmatched_recovery_ids).
    """
    by_id = {c["test_id"]: c for c in cases}
    unmatched = []
    for r in recovery:
        tid = r["test_id"]
        if tid not in by_id:
            unmatched.append(tid)
            continue
        if tid.startswith("DI"):
            continue  # keep original status for Digital Inspections
        c = by_id[tid]
        c["status"] = r["new_status"]
        if r.get("actual"):
            c["actual"] = r["actual"]
        if r.get("reason"):
            c["recovery_reason"] = r["reason"]
        c["recovered"] = True
    return unmatched


def harvest_spec_refs():
    """Read the existing by-status workbook and harvest the
    'Spec / Requirement Reference' (column L) text keyed by Test ID."""
    path = os.path.join(BASE, "custom-roles-test-run-by-status.xlsx")
    spec = {}
    wb = openpyxl.load_workbook(path, read_only=True)
    for ws in wb.worksheets:
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            tid_i = header.index("Test ID")
            spec_i = header.index("Spec / Requirement Reference")
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            tid = row[tid_i]
            if tid and row[spec_i]:
                spec[str(tid).strip()] = str(row[spec_i]).strip()
    wb.close()
    return spec


def harvest_ready_steps():
    """Steps to Reproduce (Manual) from the two ready-made QA CSVs, by Test ID."""
    steps = {}
    for fn in ("custom-roles-failed-tests-QA-report.csv",
               "custom-roles-blocked-recovery-QA-report.csv"):
        path = os.path.join(BASE, fn)
        if not os.path.exists(path):
            continue
        with open(path, newline="") as fh:
            for row in csv.DictReader(fh):
                tid = (row.get("Test ID") or "").strip()
                st = (row.get("Steps to Reproduce (Manual)") or "").strip()
                if tid and st and tid not in steps:
                    steps[tid] = st
    return steps


# ---------------------------------------------------------------- plain-English helpers
def split_camel(name):
    """workOrdersCreateAndEdit -> 'Work Orders Create And Edit'."""
    name = name.strip()
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", name)
    return s[:1].upper() + s[1:] if s else s


def humanize_config(cfg):
    """Turn the technical role/permission setup into plain words."""
    cfg = (cfg or "").strip()
    if not cfg:
        return ""
    # perms=[a, b, c]
    m = re.fullmatch(r"perms=\[(.*?)\]", cfg)
    if m:
        perms = [split_camel(p) for p in m.group(1).split(",") if p.strip()]
        if not perms:
            return "Custom role with no permissions turned on."
        return ("Custom role with only these permissions turned on: "
                + ", ".join(perms) + ".")
    # BILAL AUTOMATION = "X" profile (cloned from live system role, ...): long perm dump
    m = re.match(r'BILAL AUTOMATION = "(.+?)" profile \(cloned from live system role.*?\):', cfg)
    if m:
        return (f"Test role set up as an exact copy of the built-in \"{m.group(1)}\" role, "
                "assigned to a test technician account (fresh login used so the "
                "permissions load).")
    # BILAL AUTOMATION = X template; +[added]; -[removed]; ...
    m = re.match(r"BILAL AUTOMATION = (\w[\w ]*?) template\s*(.*)", cfg)
    if m:
        template = split_camel(m.group(1))
        rest = m.group(2)
        added = re.findall(r"\+\[(.*?)\]", rest)
        removed = re.findall(r"-\[(.*?)\]", rest)
        parts = [f"Test role based on the built-in \"{template}\" role template"]
        if added:
            parts.append("with these permissions turned ON in addition: "
                         + ", ".join(a.strip() for a in added))
        if removed:
            parts.append("and these turned OFF: "
                         + ", ".join(r.strip() for r in removed))
        return "; ".join(parts) + "."
    # BILAL AUTOMATION combo role -> A, B, C | extras
    m = re.match(r"BILAL AUTOMATION combo role -> (.+)", cfg)
    if m:
        return ("Custom role combining these permissions: "
                + m.group(1).split("|")[0].strip() + ".")
    # already plain English (e.g. "Start from blank/no-access role. ...")
    return cfg


def why_blocked(c):
    """Plain-language reason a case is still blocked."""
    reason = (c.get("recovery_reason") or "").strip()
    if not reason:
        notes = (c.get("notes") or "").strip()
        m = re.match(r"BLOCKED reason:\s*(.+)", notes, re.S)
        reason = m.group(1).strip() if m else notes
    r = reason.lower()
    if "di cases excluded" in r or (c["test_id"].startswith("DI") and "destructive" in r):
        # DI destructive reasons — keep original wording context
        if "fill of a real inspection" in r:
            return ("Running this test would have filled in a real inspection to 100% "
                    "on the shared test site, changing data other testers rely on, "
                    "so it was not executed.")
        if "pdf report" in r or "status change" in r:
            return ("Running this test would have generated a real report / changed the "
                    "status of live records on the shared test site, so it was not executed.")
        if "destructive mutation" in r:
            return ("Running this test would have permanently changed shared test data "
                    "(and the spec does not define the exact limit being checked), "
                    "so it was not executed.")
        return ("Running this test would have permanently changed shared test data, "
                "so it was not executed.")
    if "front-end-only display gate" in r or "rendered" in r or "not observable via api" in r \
            or "role-editor" in r:
        return ("This is a screen-display check: whether a button/field shows or hides "
                "on the actual page. Our automated checks work behind the scenes and "
                "cannot see the rendered screen, so a person needs to look at it.")
    if "could not provision" in r:
        m = re.search(r"could not provision a?\s*(throwaway [^;]+?) via api", r)
        thing = ("a " + m.group(1)) if m else "a throwaway (disposable) record"
        return (f"The test needs {thing} that can be safely deleted or reversed, and "
                "we could not create one automatically. Existing real records were "
                "deliberately left untouched.")
    return reason or "No reason recorded."


def unblock_needed(c):
    """Derive what is needed to unblock, from the blocked reason."""
    reason = (c.get("recovery_reason") or c.get("notes") or "").lower()
    tid = c["test_id"]
    if tid.startswith("DI"):
        return ("Needs a safe, disposable Digital Inspections work order (or agreement "
                "that shared staging data may be changed) so the action can be "
                "performed without harming other testers' data.")
    if "could not provision" in reason:
        m = re.search(r"could not provision a?\s*(throwaway [^;]+?) via api", reason)
        thing = ("a " + m.group(1)) if m else "a throwaway test record"
        return f"Needs {thing} staged as test data so the delete/reverse can be safely tried."
    if ("front-end-only" in reason or "rendered" in reason
            or "not observable via api" in reason or "role-editor" in reason):
        return ("Needs a quick manual visual check on the actual screen by a tester "
                "(log in with the role and look at the page).")
    return "Needs a manual re-run in an environment where this action is safe to perform."


def why_na(c):
    reason = (c.get("recovery_reason") or "").strip()
    r = reason.lower()
    m = re.search(r"\((.*?)\)", reason)
    detail = m.group(1) if m else ""
    if "not present in this build" in r or "not surfaced in this build" in r \
            or "not standalone in this build" in r:
        base = ("The feature/permission this test checks does not exist in the "
                "current build, so there is nothing to test.")
        return base + (f" ({detail}.)" if detail else "")
    return reason or "Marked Not Applicable during the blocked-case re-run."


def numbered_steps_from_case(c):
    """Fallback: build a clean numbered layman step list from config+expected."""
    steps = [
        "Log in as an Administrator and go to Settings > Roles & Permissions.",
    ]
    cfg = humanize_config(c.get("config"))
    if cfg:
        steps.append(f"Set up the role like this: {cfg}")
    steps.append("Assign the role to a test user, then log in as that user "
                 "(fresh login so the permissions load).")
    exp = (c.get("expected") or "").strip()
    if exp:
        steps.append(f"Try the action being tested and compare with what should happen: {exp}")
    return "\n".join(f"{i}. {s}" for i, s in enumerate(steps, 1))


# ---------------------------------------------------------------- workbook helpers
def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_body(ws, first_data_row, last_row, ncols, alt=True):
    for r in range(first_data_row, last_row + 1):
        shade = alt and ((r - first_data_row) % 2 == 1)
        for col in range(1, ncols + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = BORDER
            if shade:
                cell.fill = ALT_FILL


def est_row_height(values, widths):
    max_lines = 1
    for val, w in zip(values, widths):
        s = str(val or "")
        chars_per_line = max(int(w * 1.05), 1)
        lines = 0
        for seg in s.split("\n"):
            lines += max(1, -(-len(seg) // chars_per_line))
        max_lines = max(max_lines, lines)
    return min(max(max_lines * 15, 22), 409)


def mark_id_cells(ws, fill, first=2):
    for r in range(first, ws.max_row + 1):
        cell = ws.cell(row=r, column=1)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(size=11, bold=True)


def sort_key(c):
    tid = c["test_id"]
    prefix = tid.split("-")[0]
    m = re.search(r"(\d+)$", tid)
    num = int(m.group(1)) if m else 0
    return (GROUP_ORDER.get(prefix, 99), tid.rsplit("-", 1)[0], num)


def expected_with_spec(c, spec):
    exp = (c.get("expected") or "").strip()
    ref = spec.get(c["test_id"], "")
    if ref:
        return f"{exp}\n\nSpec reference: {ref}"
    return exp


def fill_tab(ws, headers, widths, rows):
    ws.append(headers)
    style_header(ws, len(headers))
    set_widths(ws, widths)
    for row in rows:
        ws.append(row)
        ws.row_dimensions[ws.max_row].height = est_row_height(row, widths)
    apply_body(ws, 2, ws.max_row, len(headers))


# ---------------------------------------------------------------- main
def main():
    cases = load_cases()
    recovery = load_recovery()
    unmatched = apply_overrides(cases, recovery)
    # Digital Inspections (DI) cases are excluded from this report per request.
    cases = [c for c in cases if not c["test_id"].upper().startswith("DI")]
    spec = harvest_spec_refs()
    ready_steps = harvest_ready_steps()

    cases.sort(key=sort_key)
    passed = [c for c in cases if c["status"] == "PASS"]
    failed = [c for c in cases if c["status"] == "FAIL"]
    blocked = [c for c in cases if c["status"] == "BLOCKED"]
    na = [c for c in cases if c["status"] in ("NA", "N-A", "N/A")]

    wb = Workbook()

    # ---------------- TAB 1: Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    headers = ["Test Group", "Total", "Passed", "Failed", "Blocked", "N/A"]
    ws.append(headers)
    style_header(ws, len(headers))
    totals = [0] * 5
    for prefix, gname in GROUPS:
        gc = [c for c in cases if c["test_id"].startswith(prefix)]
        p = sum(1 for c in gc if c["status"] == "PASS")
        f = sum(1 for c in gc if c["status"] == "FAIL")
        b = sum(1 for c in gc if c["status"] == "BLOCKED")
        n = sum(1 for c in gc if c["status"] in ("NA", "N-A", "N/A"))
        ws.append([gname, len(gc), p, f, b, n])
        for i, v in enumerate([len(gc), p, f, b, n]):
            totals[i] += v
    apply_body(ws, 2, ws.max_row, len(headers))
    grand_row = ws.max_row + 1
    ws.append(["GRAND TOTAL"] + totals)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=grand_row, column=col)
        cell.font = GRAND_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.append([])
    sentence_row = ws.max_row + 1
    sentence = (
        f"We ran the Custom Roles test suite: {totals[0]} tests in total "
        "(Digital Inspections cases excluded per request). "
        f"{totals[1]} passed, {totals[2]} failed, {totals[3]} could not be completed "
        f"(blocked), and {totals[4]} turned out not to apply to this build. "
        "The 'Failed Tests' tab has full step-by-step instructions to reproduce each "
        "problem; the 'Blocked Tests' tab explains why each blocked test could not "
        "be run and what is needed to run it. Blocked tests were re-tried on "
        "disposable test data where possible before being counted here."
    )
    cell = ws.cell(row=sentence_row, column=1, value=sentence)
    cell.font = Font(size=11, italic=True)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=sentence_row, start_column=1,
                   end_row=sentence_row, end_column=len(headers))
    ws.row_dimensions[sentence_row].height = 90
    set_widths(ws, [48, 10, 10, 10, 10, 10])

    # ---------------- TAB 2: Passed Tests ----------------
    ws = wb.create_sheet("Passed Tests")
    headers = ["Test ID", "What We Tested", "Role / Setup Used",
               "What Should Happen", "What Actually Happened", "Notes"]
    widths = [15, 42, 48, 60, 60, 42]
    rows = [[c["test_id"], (c.get("title") or "").strip(),
             humanize_config(c.get("config")),
             expected_with_spec(c, spec),
             (c.get("actual") or "").strip(),
             (c.get("notes") or "").strip()] for c in passed]
    fill_tab(ws, headers, widths, rows)
    mark_id_cells(ws, GREEN_FILL)

    # ---------------- TAB 3: Failed Tests ----------------
    ws = wb.create_sheet("Failed Tests")
    headers = ["Test ID", "What We Tested", "Role / Setup Used",
               "Steps To Reproduce", "What Should Happen (per spec)",
               "What Actually Happened (the bug)", "Notes"]
    widths = [15, 38, 42, 70, 60, 62, 38]
    rows = []
    for c in failed:
        steps = ready_steps.get(c["test_id"]) or numbered_steps_from_case(c)
        rows.append([c["test_id"], (c.get("title") or "").strip(),
                     humanize_config(c.get("config")),
                     steps,
                     expected_with_spec(c, spec),
                     (c.get("actual") or "").strip(),
                     (c.get("notes") or "").strip()])
    fill_tab(ws, headers, widths, rows)
    mark_id_cells(ws, RED_FILL)

    # ---------------- TAB 4: Blocked Tests ----------------
    ws = wb.create_sheet("Blocked Tests")
    headers = ["Test ID", "What We Tested", "Role / Setup Used",
               "Why It Was Blocked", "What's Needed To Unblock", "Notes"]
    widths = [15, 42, 46, 62, 55, 42]
    rows = [[c["test_id"], (c.get("title") or "").strip(),
             humanize_config(c.get("config")),
             why_blocked(c),
             unblock_needed(c),
             (c.get("notes") or "").strip()] for c in blocked]
    fill_tab(ws, headers, widths, rows)
    mark_id_cells(ws, AMBER_FILL)

    # ---------------- TAB 5: Not Applicable ----------------
    ws = wb.create_sheet("Not Applicable")
    headers = ["Test ID", "What We Tested", "Why It Doesn't Apply", "Notes"]
    widths = [15, 48, 65, 48]
    rows = [[c["test_id"], (c.get("title") or "").strip(),
             why_na(c),
             (c.get("notes") or "").strip()] for c in na]
    fill_tab(ws, headers, widths, rows)
    mark_id_cells(ws, GREY_FILL)

    wb.save(OUT)
    print("Saved:", OUT)
    print("Tabs:", wb.sheetnames)
    print(f"Passed: {len(passed)}  Failed: {len(failed)}  "
          f"Blocked: {len(blocked)}  NA: {len(na)}  Total: {len(cases)}")
    if unmatched:
        print("WARNING - recovery test ids with no matching batch case:", unmatched)
    else:
        print("All blocked-recovery test ids matched batch cases.")


if __name__ == "__main__":
    main()
