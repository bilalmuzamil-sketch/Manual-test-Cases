#!/usr/bin/env python3
"""The Defects-for-Testers workbook (skill 04 section 6.1) for both suites being handed over.

Layout per the skill: a TAB PER VERDICT STATUS plus a Summary tab. One row per case that did NOT
pass, with the C-id and its TestRail link (Rule 8), what the DOCUMENT requires with its anchor and
version, what the BUILD actually does as observed with the evidence reference, the verdict, the
Rule-91 freshness badge, and - never omitted - a plain "What needs to be done" a non-technical QA can
act on. A bare DEVIATION or Blocked with no next step is what leaves a finding sitting for a week.

Numbers and rows are GENERATED from each suite's verdicts.py, never transcribed (G7).
"""
import sys, json, os, collections
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

ROOT = '/home/user/Manual-test-Cases'
BUILD = 'v26.35.6-598cc8a'
BADGE = f'GREEN - checked 1 Sep 2026 on build {BUILD} (0 days old)'
link = lambda i: f'https://shopview.testrail.io/index.php?/cases/view/{i}'

def load(path, cases_json):
    sys.path.insert(0, path)
    import importlib, verdicts
    importlib.reload(verdicts)
    cases = {c['id']: c for c in json.load(open(cases_json))}
    v = dict(verdicts.V)
    sys.path.remove(path)
    del sys.modules['verdicts']
    return v, cases

SUITES = [
 ('Inline Add and Edit Parts',
  f'{ROOT}/build/inline-add-edit-parts/build-verify-2026-09-01/verdicts',
  '/tmp/inl6597/cases6597.json',
  'Inline Add and Edit Parts on Work Order Lines specification version 16'),
 ('Printer Friendly Work Orders',
  f'{ROOT}/build/printer-friendly-wo/build-verify-2026-09-01/verdicts',
  '/tmp/pf6617/cases6617.json',
  'Printer Friendly Work Orders specification version 8'),
]

# the plain next step, per case. Written for a non-technical reader, and never left blank.
NEXT = {
 45068: 'RUN IT AND EXPECT IT TO FAIL. Type a part into the inline row without saving, then click the '
        'pencil on another part on the same line. The "Discard this part?" question should appear '
        'first; it does not - the Edit Part Request window opens straight away and your typed part is '
        'left behind it. Mark the case FAILED and add nothing else.',
 44993: 'RUN THE PART YOU CAN. The case now names three statuses - Complete, Invoiced, Paid - and only '
        'Paid exists in the data here. Open a Paid work order and check the Add Part button is not '
        'there; that part is confirmed working. Then mark the case Blocked with the note "only the '
        'Paid status could be checked". (Two statuses this case used to name, Declined and Imported, '
        'are not statuses this product has at all.)',
 44994: 'Same as the case above, for the pencil (Edit) control instead of the Add Part button. Note: '
        'this case had been narrowed by hand to three statuses and my write pass reverted it; that is '
        'repaired, and the case you open now says three.',
 44996: 'RUN IT AND EXPECT IT TO FAIL. Pick a line with NO parts on it, click Approve, then mark it '
        'Complete from the same row (going straight to Complete is refused, and a line that has parts '
        'cannot be completed at all). With the badge reading "Complete", look at that line\'s Parts '
        'section: the "+ Add Part" button is still there, and it should not be. Mark the case FAILED.',
 45034: 'THIS ONE REALLY DOES NEED A SECOND PERSON. Ask a colleague to change or delete the same part '
        'while your edit row is open, then press Save. If you cannot arrange that, leave it Untested '
        'and tell the QA lead - do not guess. We tried it from a second connection rather than a '
        'second person and could not get the row open at the right moment, so nothing is known about '
        'this behaviour either way.',
 45060: 'RUN IT AND EXPECT IT TO FAIL. Click Add Part, type "F40010212" in the Part number box and '
        'click the suggestion marked "Catalog" (Slack Adjuster - it is stocked nowhere and has no '
        'price on record). The Cost and Sell price boxes should open EMPTY and stop you saving until '
        'you fill them; instead they open showing "0.00" and the part saves at 0.00. Mark the case '
        'FAILED.',
 45239: 'RUNS AND PASSES - kept here only so you know how to reach the state. Click Add Part, type '
        '"F40010212" and pick the suggestion marked "Catalog": it sits in no bin, so you get no bin '
        'chip and no "Pulled from" line, which is correct.',
 45220: 'NOT YOURS TO RUN OR CHANGE - this case belongs to Vladimir Tomovic and it has no steps '
        'written in it. Leave it alone entirely.',
 45088: 'RUN THE THREE YOU CAN. The Print option is confirmed present on Estimate, Approved and Paid '
        'work orders. Note that the product itself only has seven work order statuses - Estimate, '
        'Approved, In progress, Review, Complete, Invoiced, Paid - and only those three exist in the '
        'data here, so mark the case Blocked with the note "only three of the statuses exist here".',
 45107: 'DO NOT RUN IT. It cannot be done: on a work order with no lines the Print option is greyed '
        'out, so you can never see the printout it describes. The written description contradicts '
        'itself here and the product owner has to settle it. Leave it Untested.',
 45116: 'DO NOT RUN IT - same reason as the case above. You cannot print a work order that has no '
        'lines, so there is no summary to look at.',
 45090: 'RUNS AND PASSES. To set it up: Settings > Roles & Permissions > pencil on a role > switch its '
        'work-order viewing OFF, and switch the work-order line editing and part-picking permissions '
        'off in the same role (viewing alone will not stay off - the others depend on it). A user on '
        'that role is then bounced off the work order entirely, "Work Orders" disappears from the top '
        'menu, and there is no More menu and no print option. Put the permissions back afterwards.',
 45097: 'DO NOT RUN IT - it cannot be done. The app will not create a work order without a customer: '
        'leaving Customer empty on the New Work Order window answers "Customer is a required field" '
        'and nothing is saved. So the printout this case describes can never exist. It is waiting on a '
        'product-owner ruling. Leave it Untested.',
 45098: 'DO NOT RUN IT - it cannot be done, same as the customer one. Choosing a customer and pressing '
        'Save with Add Asset empty answers "Asset is a required field". Waiting on a product-owner '
        'ruling. Leave it Untested.',
 45104: 'DO NOT RUN IT - there is no "Cancelled" status for a work order line in this product. A line '
        'offers only Authorization required, Declined, Authorized and Complete, and those are the only '
        'four the system accepts. Waiting on a product-owner ruling. Leave it Untested.',
 45111: 'RUNS AND PASSES. To set it up: open the work order\'s Lines tab, click a line, paste about a '
        'full paragraph (500+ characters) into its Tech Story box and save. The whole story then '
        'prints, with no cut-off and no "Show more".',
 45123: 'RUN IT NORMALLY - the behaviour is correct. One thing to expect: the history row is called '
        '"Work order printed history", not "Work Order Printed" as the case says. That difference is '
        'already reported, so do NOT raise it again; pass the case on the behaviour.',
}
FLAVOUR = {'FAIL': 'DEVIATION', 'PARTIAL': 'PARTLY BLOCKED', 'UNREACHABLE': 'CANNOT BE RUN',
           'NOTVER': 'BLOCKED - DATA MISSING', None: 'NOT CHECKED YET', 'FOREIGN': 'NOT OURS'}
HEAD = ['C-ID', 'TestRail link', 'Suite', 'Case title', 'What the document requires',
        'What the build actually does (observed)', 'Verdict', 'Freshness', 'What needs to be done']

rows_by_status = collections.defaultdict(list)
summary = []
for suite, vdir, cjson, specref in SUITES:
    V, cases = load(vdir, cjson)
    c = collections.Counter(v[0] for v in V.values())
    summary.append((suite, len(V), c))
    for cid, (verdict, ev, note) in sorted(V.items()):
        if verdict == 'PASS':
            continue
        flav = FLAVOUR.get(verdict, str(verdict))
        rows_by_status[flav].append([
            f'C{cid}', link(cid), suite, cases[cid]['title'],
            f'{specref} - see the case\'s own provenance line for the exact section',
            note or 'not exercised by this pass',
            flav, BADGE,
            NEXT.get(cid, 'ASK THE QA LEAD before running this one - no instruction has been written '
                          'for it, which is itself something to report.'),
        ])
    # C45123 is a PASS but carries an instruction, so it rides along on its own tab
    if suite.startswith('Printer') and 45123 in V:
        rows_by_status['PASS - BUT READ THE NOTE'].append([
            'C45123', link(45123), suite, cases[45123]['title'],
            f'{specref}, section S6-R1', V[45123][2], 'PASS with a wording difference', BADGE, NEXT[45123]])

wb = openpyxl.Workbook(); wb.remove(wb.active)
ws = wb.create_sheet('Summary')
ws['A1'] = 'WHAT DID NOT PASS - both suites, checked 1 September 2026'
ws['A1'].font = Font(bold=True, size=13)
ws['A2'] = (f'Build checked: {BUILD} on https://sv9315.qa.shopview.com. Everything not listed on the '
            'other tabs PASSED. A case on this list is not necessarily broken - most of them just need '
            'something that does not exist on the test system yet, and each row says exactly what to do.')
ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
ws.append([]); ws.append(['Suite', 'Cases', 'Passed', 'Needs attention', 'Breakdown'])
for c_ in range(1, 6): ws.cell(row=4, column=c_).font = Font(bold=True)
for suite, n, c in summary:
    ws.append([suite, n, c.get('PASS', 0), n - c.get('PASS', 0),
               ', '.join(f'{FLAVOUR.get(k, k)}: {v}' for k, v in c.items() if k != 'PASS')])
ws.append([])
ws.append(['Tab', 'Rows', 'What it means'])
ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
ws.cell(row=ws.max_row, column=2).font = Font(bold=True)
ws.cell(row=ws.max_row, column=3).font = Font(bold=True)
MEANING = {
 'DEVIATION': 'The build genuinely does the wrong thing. Run it, expect it to fail, mark it FAILED.',
 'PARTLY BLOCKED': 'Part of the case can be run today and part cannot. Run the part you can.',
 'CANNOT BE RUN': 'The situation the case describes cannot be created at all. Leave it Untested.',
 'BLOCKED - DATA MISSING': 'The case is fine but the test system has nothing to run it against yet.',
 'NOT CHECKED YET': 'We did not get to it. Leave it Untested unless the row says otherwise.',
 'NOT OURS': 'Somebody else owns this case. Do not run or change it.',
 'PASS - BUT READ THE NOTE': 'It passes, but something on screen differs from the wording in the case.',
}
for k, v in rows_by_status.items():
    ws.append([k, len(v), MEANING.get(k, '')])
for col, w in zip('ABCDE', (36, 10, 10, 18, 62)): ws.column_dimensions[col].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')

for status, rows in rows_by_status.items():
    t = wb.create_sheet(status[:31])
    t.append(HEAD)
    for i in range(1, len(HEAD) + 1):
        t.cell(row=1, column=i).font = Font(bold=True)
        t.cell(row=1, column=i).fill = PatternFill('solid', fgColor='DDDDDD')
    for r in rows: t.append(r)
    for col, w in zip('ABCDEFGHI', (10, 46, 26, 52, 44, 62, 22, 40, 78)):
        t.column_dimensions[col].width = w
    for row in t.iter_rows(min_row=2):
        for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')
    t.freeze_panes = 'A2'

out = f'{ROOT}/build/handoff-2026-09-01/Inline-Add-and-Edit-Parts_and_Printer-Friendly-Work-Orders_Defects-for-Testers_2026-09-01.xlsx'
wb.save(out)
print('written', out)
for k, v in rows_by_status.items(): print(f'  tab {k}: {len(v)} rows')
missing = [r[0] for rows in rows_by_status.values() for r in rows if 'ASK THE QA LEAD before running' in r[8]]
print('rows with no written next step:', missing or 'none')
