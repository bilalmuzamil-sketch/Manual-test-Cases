#!/usr/bin/env python3
"""Reformat the ShopView Custom Roles test suite into the team's STANDARD column format.

Standard columns (exact order):
  Title | Section | Type | Priority | Preconditions | Steps | Expected Result | References

Outputs:
  - custom-roles-test-cases.xlsx  (tabs: Index, Single Permission, Template Edit,
        Combination - Representative, Combination - Random, Digital Inspections)
  - custom-roles-test-cases.csv   (one combined sheet)
  - build/di-perrole.csv          (clean tracked copy of the uploaded DI CSV)
"""
import csv
import json
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BUILD = os.path.join(REPO, "build")
UPLOAD = "/root/.claude/uploads/dd1d42ba-2c47-5229-9b17-b8f94e3eb99a/3f8dd06f-SV7388_DigitalInspections_PerRole_TestCases_1.csv"
DI_COPY = os.path.join(BUILD, "di-perrole.csv")

HEADERS = ["Title", "Section", "Type", "Priority", "Preconditions",
           "Steps", "Expected Result", "References"]

VIU_FINDINGS = os.path.join(BUILD, "viu-findings.json")
VIU_HEADERS = ["ID", "Topic", "Spec Expectation", "Actual (Staging)",
               "Verdict", "Evidence", "Notes"]

SP_MAP = {
    "WO": "Work Orders", "WOL": "Work Order Lines", "SCH": "Schedule",
    "CUST": "Customer Management", "PS": "Part Sales",
    "CAT": "Catalog & Inventory", "VEND": "Vendor & Order Management",
    "INV": "Invoicing & Payments", "TS": "Timesheets",
    "WOSUB": "WO Sub-Settings", "RPT": "Reports", "CPORT": "Customer Portal",
    "BPORT": "Billing Portal", "PDEPT": "Parts Department", "SET": "Settings",
    "VM": "View Mode", "FIN": "See Financial Data", "APAR": "Manage AP/AR",
    "HIST": "View History Logs",
}
TE_MAP = {
    "ADMIN": "Administrator", "SM": "Service Manager",
    "SSA": "Senior Service Advisor", "SADV": "Service Advisor",
    "FORE": "Foreman", "TECH": "Technician", "PM": "Parts Manager",
    "PT": "Parts Technician", "OFFICE": "Office",
    "SALES": "Sales Representative", "TIMECLK": "Time Clock",
}


def build_row(case):
    """Transform one JSON case object into a standard 8-field row."""
    tid = case["test_id"]
    title = f'{tid} — {case["title"]}'

    code = tid.split("-")[1]
    if tid.startswith("SP-"):
        category = "Single Permission"
        leaf = SP_MAP.get(code, code)
    elif tid.startswith("TE-"):
        category = "Template Edit"
        leaf = TE_MAP.get(code, code)
    elif tid.startswith("CB-REP"):
        category = "Combination (Representative)"
        leaf = case["permission"]
    elif tid.startswith("CB-RND"):
        category = "Combination (Random)"
        leaf = case["permission"]
    else:
        raise ValueError(f"Unknown test_id prefix: {tid}")
    section = f"Custom Roles and Permissions > {category} > {leaf}"

    row_type = "UI" if case["type"] == "UI" else "Functional"
    priority = case["priority"]

    pre_parts = []
    if case.get("preconditions"):
        pre_parts.append(case["preconditions"])
    if case.get("role_setup"):
        pre_parts.append("Role setup: " + case["role_setup"])
    if case.get("test_data"):
        pre_parts.append("Test data: " + case["test_data"])
    pre_parts.append(
        "Permission under test: " + case["permission"]
        + " | Original type: " + case["type"]
        + " | Dependency mode: " + case["dependency_mode"]
    )
    preconditions = "\n".join(pre_parts)

    step_lines = []
    exp_lines = []
    for i, s in enumerate(case["steps"], start=1):
        n = s.get("n", i)
        step_lines.append(f"{n}. {s['action']}")
        exp_lines.append(f"Step {n}: {s['expected']}")
    steps = "\n".join(step_lines)

    expected_result = "\n".join(exp_lines) + "\n\n" \
        + "Overall: " + case["expected_final"] + "\n" \
        + "VIU status: " + case["source_viu"]

    references = case["jira"]

    return [title, section, row_type, priority, preconditions,
            steps, expected_result, references]


def load_json(name):
    with open(os.path.join(BUILD, name)) as f:
        return json.load(f)


def copy_di_csv():
    """Read the uploaded DI CSV and write a clean tracked copy; return data rows."""
    with open(UPLOAD, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    header = rows[0][:8]
    data = []
    for r in rows[1:]:
        if not any(cell.strip() for cell in r):
            continue  # skip fully blank lines
        data.append(r[:8])
    # write clean copy with 8 columns
    with open(DI_COPY, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        w.writerows(data)
    return data


# ---- Column formatting spec ----
COL_WIDTHS = {
    "Title": 40, "Section": 35, "Type": 12, "Priority": 12,
    "Preconditions": 45, "Steps": 60, "Expected Result": 60, "References": 16,
}
WRAP_COLS = {"Preconditions", "Steps", "Expected Result"}
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(bold=True)
TOP_LEFT = Alignment(vertical="top", horizontal="left", wrap_text=False)
TOP_WRAP = Alignment(vertical="top", horizontal="left", wrap_text=True)


def style_sheet(ws, nrows):
    for ci, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[h]
    for r in range(2, nrows + 2):
        for ci, h in enumerate(HEADERS, start=1):
            cell = ws.cell(row=r, column=ci)
            cell.alignment = TOP_WRAP if h in WRAP_COLS else TOP_LEFT
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{max(nrows + 1, 1)}"


def write_data_sheet(wb, title, rows):
    ws = wb.create_sheet(title)
    ws.append(HEADERS)
    for r in rows:
        ws.append(r)
    style_sheet(ws, len(rows))
    return ws


def load_viu_findings():
    with open(VIU_FINDINGS) as f:
        return json.load(f)


def viu_tallies(findings):
    disc = sum(1 for x in findings if "DISCREPANCY" in x["verdict"].upper())
    match = len(findings) - disc
    return len(findings), match, disc


def build_index_sheet(wb, counts, viu_stats, verified_counts):
    total_viu, match_viu, disc_viu = viu_stats
    ver_match, ver_disc = verified_counts
    ws = wb.create_sheet("Index")
    ws["A1"] = "ShopView Custom Roles & Permissions — Test Suite Index"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Epic: SV-7388 (Custom Roles and Permissions)"
    ws["A4"] = (
        f"Verification status: VIU (Verify-in-UI) pass 3 complete — per-role verified "
        f"as the restricted Tech user (tech@shopview.com) on staging, 2026-07-01. "
        f"{total_viu} findings logged "
        f"({match_viu} MATCH / {disc_viu} DISCREPANCY) — see the 'VIU Findings Log' "
        f"tab. Pass 2 (live admin session) confirmed exact Custom Roles labels, the "
        f"silent CRUD cascade, the 'Enable See Financial Data?' popup, and the "
        f"kebab-menu variants. Pass 3 logged in AS Tech and verified all 11 per-role "
        f"gate configs behave PER SPEC: CRUD gating (View-only read-only/no New; "
        f"+Create&Edit shows New; +Delete per-row delete), page-access nav gating "
        f"(Reports/Customers/combination exact), Settings routes under '/administration', "
        f"and view-mode/parts/financial effects carried in the permission wrapper "
        f"(effects inside the WO detail / invoicing). Discrepancies/limits logged: "
        f"'/settings' 404s (admin lives at '/administration'), a zero-permission role is "
        f"not creatable (min 1 permission enforced), and permission changes require the "
        f"user to re-login + settle. Earlier build-vs-spec discrepancies still stand "
        f"(doubled '/api/api/sso/check' path, no unsaved-changes guard on Create Role "
        f"X-close, silent-cascade vs confirmation-popup dependency-UX inconsistency). "
        f"{ver_match + ver_disc} test cases now VERIFIED "
        f"({ver_match} MATCH / {ver_disc} DISCREPANCY); remaining cases are "
        f"UNVERIFIED — VIU pending.")
    ws["A5"] = ("Standard column format (test-case tabs): Title | Section | Type | "
                "Priority | Preconditions | Steps | Expected Result | References")

    ws["A7"] = "Tab"
    ws["B7"] = "Row Count"
    for c in ("A7", "B7"):
        ws[c].font = HEADER_FONT
        ws[c].fill = HEADER_FILL
    r = 8
    total = 0
    for tab, n in counts:
        ws.cell(row=r, column=1, value=tab)
        ws.cell(row=r, column=2, value=n)
        total += n
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total).font = Font(bold=True)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 14
    for row in ws.iter_rows():
        for cell in row:
            if cell.alignment.wrap_text is not True:
                cell.alignment = Alignment(vertical="top", horizontal="left")
    for a in ("A4",):
        ws[a].alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    return ws


def build_viu_sheet(wb, findings):
    ws = wb.create_sheet("VIU Findings Log")
    ws.append(VIU_HEADERS)
    for fnd in findings:
        ws.append([
            fnd["id"], fnd["topic"], fnd["spec_expectation"],
            fnd["actual_staging"], fnd["verdict"], fnd["evidence"], fnd["notes"],
        ])
    widths = {"ID": 10, "Topic": 24, "Spec Expectation": 40,
              "Actual (Staging)": 50, "Verdict": 26, "Evidence": 24, "Notes": 40}
    wrap = {"Spec Expectation", "Actual (Staging)", "Verdict", "Notes", "Evidence"}
    for ci, h in enumerate(VIU_HEADERS, start=1):
        c = ws.cell(row=1, column=ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = widths[h]
    for rr in range(2, len(findings) + 2):
        for ci, h in enumerate(VIU_HEADERS, start=1):
            cell = ws.cell(row=rr, column=ci)
            cell.alignment = TOP_WRAP if h in wrap else TOP_LEFT
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(VIU_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(findings) + 1}"
    return ws


def main():
    sp_crud = load_json("sp-crud.json")
    sp_noncrud = load_json("sp-noncrud.json")
    te = load_json("te.json")
    combo = load_json("combo.json")

    assert len(sp_crud) == 52, len(sp_crud)
    assert len(sp_noncrud) == 65, len(sp_noncrud)
    assert len(te) == 41, len(te)
    assert len(combo) == 18, len(combo)

    cb_rep = [c for c in combo if c["test_id"].startswith("CB-REP")]
    cb_rnd = [c for c in combo if c["test_id"].startswith("CB-RND")]

    single_rows = [build_row(c) for c in sp_crud] + [build_row(c) for c in sp_noncrud]
    te_rows = [build_row(c) for c in te]
    rep_rows = [build_row(c) for c in cb_rep]
    rnd_rows = [build_row(c) for c in cb_rnd]

    json_total = len(single_rows) + len(te_rows) + len(rep_rows) + len(rnd_rows)

    di_rows = copy_di_csv()
    di_count = len(di_rows)

    # ---- VIU findings + verified-case tallies ----
    findings = load_viu_findings()
    viu_stats = viu_tallies(findings)
    all_cases = sp_crud + sp_noncrud + te + combo
    ver_match = sum(1 for c in all_cases
                    if "VERIFIED (staging" in c["source_viu"]
                    and "MATCH" in c["source_viu"].split("VERIFIED", 1)[1])
    ver_disc = sum(1 for c in all_cases
                   if "VERIFIED (staging" in c["source_viu"]
                   and "DISCREPANCY" in c["source_viu"].split("VERIFIED", 1)[1])

    # ---- XLSX ----
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    counts = [
        ("Single Permission", len(single_rows)),
        ("Template Edit", len(te_rows)),
        ("Combination - Representative", len(rep_rows)),
        ("Combination - Random", len(rnd_rows)),
        ("Digital Inspections", di_count),
    ]
    build_index_sheet(wb, counts, viu_stats, (ver_match, ver_disc))
    build_viu_sheet(wb, findings)
    write_data_sheet(wb, "Single Permission", single_rows)
    write_data_sheet(wb, "Template Edit", te_rows)
    write_data_sheet(wb, "Combination - Representative", rep_rows)
    write_data_sheet(wb, "Combination - Random", rnd_rows)
    write_data_sheet(wb, "Digital Inspections", di_rows)
    xlsx_path = os.path.join(REPO, "custom-roles-test-cases.xlsx")
    wb.save(xlsx_path)

    # ---- Combined CSV ----
    csv_path = os.path.join(REPO, "custom-roles-test-cases.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in single_rows + te_rows + rep_rows + rnd_rows:
            w.writerow(r)
        for r in di_rows:
            w.writerow(r)

    # =========== VERIFY ===========
    print("=== VERIFY ===")
    print(f"JSON-sourced total = {json_total} (expected 176: 52+65+41+18)")
    assert json_total == 176, json_total
    print(f"DI rows parsed from CSV = {di_count}")

    wb2 = load_workbook(xlsx_path)
    print("\n-- Sheet order --")
    expected_order = ["Index", "VIU Findings Log", "Single Permission",
                      "Template Edit", "Combination - Representative",
                      "Combination - Random", "Digital Inspections"]
    print("  " + " | ".join(wb2.sheetnames))
    assert wb2.sheetnames == expected_order, wb2.sheetnames
    print("  Sheet order OK.")

    viu_ws = wb2["VIU Findings Log"]
    viu_data_rows = viu_ws.max_row - 1
    print(f"\n-- VIU Findings Log rows = {viu_data_rows} (expected 31) --")
    assert viu_data_rows == 31, viu_data_rows

    print("\n-- XLSX per-sheet data-row counts --")
    for name in ["Single Permission", "Template Edit",
                 "Combination - Representative", "Combination - Random",
                 "Digital Inspections"]:
        ws = wb2[name]
        n = ws.max_row - 1  # minus header
        print(f"  {name}: {n}")
    assert wb2["Single Permission"].max_row - 1 == 117
    assert wb2["Template Edit"].max_row - 1 == 41
    assert wb2["Combination - Representative"].max_row - 1 == 10
    assert wb2["Combination - Random"].max_row - 1 == 8
    assert wb2["Digital Inspections"].max_row - 1 == di_count
    print("  All sheet-count assertions passed.")

    with open(csv_path, newline="", encoding="utf-8") as f:
        rr = list(csv.reader(f))
    data_rows = rr[1:]
    print(f"\nCombined CSV data rows = {len(data_rows)} "
          f"(expected {176 + di_count})")
    assert len(data_rows) == 176 + di_count
    bad = [(i, len(r)) for i, r in enumerate(rr) if len(r) != 8]
    print(f"Rows with != 8 columns (including header): {len(bad)}")
    assert not bad, bad[:5]
    print("Every CSV row has exactly 8 columns.")

    print("\n-- SPOT CHECK: Single Permission row (SP-WO-001) --")
    sp_row = single_rows[0]
    for h, v in zip(HEADERS, sp_row):
        print(f"[{h}]\n{v}\n")

    print("-- SPOT CHECK: Template Edit row (TE-ADMIN-001) --")
    te_row = te_rows[0]
    for h, v in zip(HEADERS, te_row):
        print(f"[{h}]\n{v}\n")

    verified_total = sum(1 for c in all_cases
                         if c["source_viu"].startswith("VERIFIED")
                         or "VERIFIED (staging" in c["source_viu"])
    print("-- VIU VERIFICATION SUMMARY --")
    print(f"Case rows with source_viu containing VERIFIED = {verified_total} "
          f"({ver_match} MATCH / {ver_disc} DISCREPANCY); "
          f"remaining UNVERIFIED = {len(all_cases) - verified_total}")


if __name__ == "__main__":
    main()
