#!/usr/bin/env python3
"""Emit Fees & Discounts V1 cases in the user's exact TestRail import CSV format
(matched to testrail-import/sv5319-testrail-import-MATCHED.csv), plus a review xlsx.

CONTENT RULES enforced here (see task PART 2):
  1. NO VIU wording anywhere. The internal `viu_status`, `notes`, and `design_ref`
     fields are NOT emitted. No "(Ref: FD-... — VIU pending...)" traceability clause.
  2. NO feature-flag preconditions. The literal phrase "feature flag" is reworded to
     "Fees & Discounts feature" so 0 occurrences of "feature flag" remain, while the
     genuine flag-gating test cases stay coherent.
  3. Sections = leaf area names only (already leaf names in the source JSON),
     EXCEPT api_related cases which route to "API — <leaf area>" (STANDING RULE 4:
     any case with API endpoints/methods/status-codes/backend request-response
     checks must live under an API-titled section).
  4. References = spec/story reference only (no internal FD- ids, no VIU text).
  5/6. Preconditions/Steps/Expected kept as authored, cleaned per rules above; no
     TBD/pending meta text (the source expected results are clean).
"""
import csv, json, re, os

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # build/
ROOT = os.path.dirname(BASE)
CASES_DIR = os.path.join(BASE, "fees-discounts", "cases")
REF = os.path.join(ROOT, "testrail-import", "sv5319-testrail-import-MATCHED.csv")
OUT_CSV = os.path.join(ROOT, "testrail-import", "fees-discounts-v1-testrail-import.csv")
OUT_XLSX = os.path.join(ROOT, "testrail-import", "fees-discounts-v1-testrail-import.xlsx")

FILES = [
    "group-A-wo-parts.json",
    "group-B-customer-admin-finance.json",
    "group-C-calc-permissions-validation.json",
]

# --- Read reference header exactly ---
with open(REF, newline="") as f:
    REF_HEADER = next(csv.reader(f))
print("REFERENCE HEADER:", REF_HEADER)

cases = []
for fn in FILES:
    with open(os.path.join(CASES_DIR, fn)) as f:
        cases += json.load(f)
print("Total cases loaded (incl. retired):", len(cases))
# Active cases only: retired cases (viu_status 'Retired — …'; FD-CUST-016 retired
# 2026-07-20, duplicate of FD-VAL-007/C28605, ex-C28500 deleted from TestRail) are
# kept in the JSON for the record but excluded from every deliverable.
cases = [c for c in cases if not (c.get("viu_status") or "").startswith("Retired") and not c.get("dev_authored")]
print("Active cases (retired excluded):", len(cases))


def clean(s):
    """Strip internal/VIU markers and rewrite the banned 'feature flag' phrase."""
    if not s:
        return s
    # Drop internal case cross-references like " (see FD-WO-003)".
    s = re.sub(r"\s*\(see (?:SF|FD)-[A-Z0-9-]+\)", "", s)
    # Drop the admin Feature-Flags nav parenthetical (avoids the banned phrase).
    s = s.replace(" (Administration → Feature Flags)", "")
    s = s.replace("(Administration → Feature Flags)", "")
    # Rewrite flag phrases -> the feature name (keeps gating cases coherent).
    # Handle product-name-prefixed forms first to avoid doubling the name.
    s = s.replace("Fees & Discounts feature flag", "Fees & Discounts feature")
    s = s.replace("FeesAndDiscounts feature flag", "Fees & Discounts feature")
    s = s.replace("FeesAndDiscounts flag", "Fees & Discounts feature")
    s = re.sub(r"feature[ -]flags?", "Fees & Discounts feature", s, flags=re.I)
    # Strip any leftover internal "EXPECTED PER SPEC:" authoring prefix.
    m = re.match(r"^(\s*\d+\.\s*)EXPECTED PER SPEC:\s*(.*)$", s, re.I | re.S)
    if m:
        rest = m.group(2)
        rest = rest[:1].upper() + rest[1:] if rest else rest
        s = m.group(1) + rest
    return s


def joinlines(lst):
    return "\n".join(clean(x.rstrip()) for x in lst)


def build_refs(c):
    parts = []
    sr = (c.get("story_ref") or "").strip()
    if sr:
        parts.append(sr)
    for jira in sorted(set(re.findall(r"SV-\d+", str(c.get("story_ref", ""))))):
        if jira not in " ".join(parts):
            parts.append(jira)
    return clean(" ".join(parts).strip())


def build_preconditions(c):
    # Authored preconditions only. No VIU clause, no notes, cleaned per rules.
    return joinlines(c.get("preconditions", []))


def section_for(c):
    """Leaf area name; API-related cases are routed to an 'API — <area>' section
    (STANDING RULE 4: API content must live under an API-titled section).
    Kept in sync with build/simple-flow/gen_import.py."""
    area = c["area"].strip()
    if c.get("api_related"):
        return "API — " + area
    return area


rows = []
titles = []
api_sections = set()
api_moved = 0
for c in cases:
    title = clean(c["title"].strip())
    titles.append(title)
    section = section_for(c)
    if c.get("api_related"):
        api_sections.add(section)
        api_moved += 1
    row = [
        title,
        section,
        "Functional",
        c["priority"].strip(),
        build_preconditions(c),
        joinlines(c.get("steps", [])),
        joinlines(c.get("expected", [])),
        build_refs(c),
        "",
        "",
    ]
    rows.append(row)

# --- Write CSV matching reference: CRLF row endings, LF inside cells, minimal quoting ---
with open(OUT_CSV, "w", newline="") as f:
    w = csv.writer(f, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    w.writerow(REF_HEADER)
    for r in rows:
        w.writerow(r)

print("Wrote CSV:", OUT_CSV, "rows(data):", len(rows))

# --- Duplicate title check ---
from collections import Counter
dupes = [t for t, n in Counter(titles).items() if n > 1]
print("Duplicate titles:", dupes if dupes else "NONE")

# --- Sanity checks (must be zero) ---
blob = "\n".join("\t".join(r) for r in rows).lower()
print("VIU occurrences:", blob.count("viu"))
print("'feature flag' occurrences:", blob.count("feature flag"))
print("API sections created:", len(api_sections), sorted(api_sections))
print("API-flagged cases routed to API sections:", api_moved)

# --- xlsx review copy ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Fees & Discounts V1"
ws.append(REF_HEADER)
for r in rows:
    ws.append(r)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="305496")
for col in range(1, len(REF_HEADER) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(vertical="center", horizontal="left")

widths = {"Title": 50, "Section": 40, "Type": 12, "Priority": 10,
          "Preconditions": 60, "Steps": 60, "Expected Result": 60,
          "References": 18}
for i, name in enumerate(REF_HEADER, start=1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 10)

wrap = Alignment(wrap_text=True, vertical="top")
for r in range(2, len(rows) + 2):
    for cidx in range(1, len(REF_HEADER) + 1):
        ws.cell(row=r, column=cidx).alignment = wrap

ws.freeze_panes = "A2"
wb.save(OUT_XLSX)
print("Wrote XLSX:", OUT_XLSX)
