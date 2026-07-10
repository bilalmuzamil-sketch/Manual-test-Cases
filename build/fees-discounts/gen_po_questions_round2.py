#!/usr/bin/env python3
"""Generate the ROUND-2 PO question sheet for Chris Ward (Fees & Discounts PO).

Outputs (regenerable — run from the repo root):
  - build/fees-discounts/PO-Questions-Round2.xlsx
  - build/fees-discounts/PO-Questions-Round2.md

Sheets:
  - "Questions for PO"    : reader-facing, VERY SIMPLE layman language ONLY.
                            NO case IDs, FDBUG codes, API/HTTP terms, or jargon.
                            Columns: # | Topic | What happens now | The question |
                            Options | Your answer (blank).
  - "QA Internal Mapping" : QA-only. One row per (question, TestRail case) with
                            internal bug/case refs, TestRail Case ID C##### +
                            clickable link (standing rule 8; source
                            testrail-id-map.csv), spec refs, and status notes.

Round-1 reference: gen_po_questions_simple.py / PO-Questions-SIMPLE.md (all 6
round-1 questions were answered by Chris Ward 2026-07-09). These 4 questions are
NEW product decisions surfaced by the fresh full VIU pass of 2026-07-10.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/fees-discounts/PO-Questions-Round2.xlsx"
MD_OUT = "build/fees-discounts/PO-Questions-Round2.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no bug codes, no tech terms)
# ---------------------------------------------------------------------------
questions = [
    {
        "topic": "A discount bigger than the bill saves with no warning",
        "now": ("If someone applies a discount that is larger than the whole work "
                "order's total, it just saves — there is no warning or \"are you "
                "sure?\" step. The bill's total becomes $0.00 and the extra amount "
                "is quietly kept as a credit on the customer's account. The credited "
                "amount itself is exact — nothing is lost — but the user is never "
                "told any of this is happening. The original write-up expected a "
                "warning first."),
        "q": ("Should the app warn the user and ask them to confirm before saving a "
              "discount bigger than the bill, or is saving silently fine as long as "
              "the extra amount is kept as an exact customer credit?"),
        "opts": ("A) Add the warning - the user must see what will happen (total "
                 "becomes $0.00, the extra becomes a customer credit) and confirm "
                 "before it saves.\n"
                 "B) Silent is fine - no warning needed, as long as the credited "
                 "amount is exact (which it is today)."),
    },
    {
        "topic": "Typing 0 as a fee's maximum removes the limit instead of applying it",
        "now": ("When creating a percentage fee or discount, there is a \"maximum "
                "amount\" box so the charge can never go above a chosen ceiling. If "
                "someone types 0 into that box today, the 0 is saved but then "
                "ignored - the fee is applied with NO maximum at all, as if the box "
                "had been left empty."),
        "q": ("What should typing 0 into the maximum box mean: no limit (as it works "
              "today), a limit of zero (so nothing is charged), or should the app "
              "simply not accept 0 in that box?"),
        "opts": ("A) 0 means \"no limit\" - keep it working the way it does today.\n"
                 "B) 0 means \"cap at zero\" - the fee/discount amount becomes $0.00 "
                 "(charge nothing).\n"
                 "C) Don't allow it - the app should refuse 0 in the maximum box and "
                 "ask for a real amount (or an empty box)."),
    },
    {
        "topic": "Very small percentages are quietly rounded up",
        "now": ("The smallest percentage the app is meant to work with is 0.01%. If "
                "someone types in something even smaller - for example 0.005% - the "
                "app accepts it and quietly changes it to 0.01% without saying "
                "anything. The user thinks they saved 0.005% but the app is actually "
                "using 0.01%."),
        "q": ("Is quietly rounding tiny percentages up to 0.01% acceptable, or should "
              "the app keep exactly what was typed, or refuse values that small?"),
        "opts": ("A) Rounding is fine - quietly using 0.01% for anything smaller is "
                 "acceptable.\n"
                 "B) Keep the exact value - the app should save and use exactly what "
                 "was typed (for example 0.005%).\n"
                 "C) Refuse it - the app should reject anything smaller than 0.01% "
                 "with a clear message, so the user knowingly picks a valid value."),
    },
    {
        "topic": "A processing fee's \"minimum amount\" is quietly thrown away",
        "now": ("A processing fee (the fee that covers card-processing costs) is not "
                "supposed to have a minimum amount. But if someone does type a "
                "minimum amount on one, the app doesn't complain - it just saves the "
                "fee WITHOUT the minimum and never tells the user the number they "
                "typed was thrown away."),
        "q": ("Should processing fees support a minimum amount, or - if they "
              "shouldn't - should the app make that clear instead of silently "
              "dropping the number?"),
        "opts": ("A) Support it - a processing fee should be able to have a minimum "
                 "amount, and the app should honor it.\n"
                 "B) Don't support it - but make that clear: remove/disable the box "
                 "for processing fees (or show a message) so nothing a user types is "
                 "ever silently thrown away."),
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (rule 8: TestRail C##### + clickable link per case;
# IDs sourced from build/fees-discounts/testrail-id-map.csv)
# ---------------------------------------------------------------------------
# Each entry: (q_no, internal_refs, [(fd_id, testrail_id), ...], spec_refs, status)
internal_map = [
    (1,
     "FDBUG-15 (over-discount saves silently, no warn/confirm). Primary case "
     "FD-QB-014; companion over-discount thread FD-QB-012 (floor worked example, "
     "Verified) + FD-QB-015 (excess -> customer credit; in-app half VERIFIED "
     "2026-07-10: credit of exactly -117.24; QB goodwill-memo half Blocked-Env). "
     "Customer-document (FD-DOC) surfaces show the floored $0.00 totals per "
     "S6-R10 - only the warning step is missing.",
     [("FD-QB-014", 28557), ("FD-QB-012", 28555), ("FD-QB-015", 28558)],
     "requirements.md §7.1: S6-R12 (mandatory warn/confirm before saving when "
     "discounts exceed the net subtotal); context S6-R10 (subtotal floors at "
     "$0.00) + S6-R11/R13 (excess recorded as customer credit / QB tax-exempt "
     "goodwill credit memo).",
     "FD-QB-014 = VIU-Deviation (FDBUG-15 CONFIRMED AGAIN 2026-07-10: over-"
     "discount saves 201 with no warning payload; batch-6 UI shots show no "
     "warn/confirm dialog). Currently bucketed case-update pending this PO "
     "ruling. A=defect ticket + keep spec expected; B=case-update FD-QB-014 to "
     "silent-carry expected."),
    (2,
     "FDBUG-9 (maxCap 0 accepted as \"no cap\"). Cases FD-CALC-008 (0 must force "
     "$0.00), FD-VAL-006 (0/empty edge behavior), FD-TMPL-011 (template dialog "
     "stores 0). Jira draft exists: jira-bug-drafts.md TICKET 4 (not filed).",
     [("FD-CALC-008", 28575), ("FD-VAL-006", 28604), ("FD-TMPL-011", 28512)],
     "Spec contradiction the PO answer settles: §5-R6 (Max $0 forces resolve to "
     "$0.00) vs S7-R12e/R14 (0 treated as empty / never sent; design-notes §6 "
     "\"Max cap min=0\"). Live build matches NEITHER reading for 0 (0 = "
     "unlimited).",
     "FD-CALC-008 / FD-VAL-006 / FD-TMPL-011 = VIU-Deviation (FDBUG-9 CONFIRMED "
     "AGAIN 2026-07-10: maxCap 0 stored, 10% resolved 34.15 = uncapped). "
     "A=case-update all 3 to \"0 = no cap\" + drop TICKET 4; B=file TICKET 4 as "
     "drafted (§5-R6); C=new validation requirement + case updates."),
    (3,
     "FDBUG-10 (below-minimum percent silently rounded up, not rejected). Case "
     "FD-CALC-006. Jira draft exists: jira-bug-drafts.md TICKET 5 (not filed).",
     [("FD-CALC-006", 28573)],
     "requirements.md §7: §5-R1 (minimums - Flat $0.01 / Percentage 0.01%; "
     "below-minimum input is rejected, expected HTTP 400).",
     "FD-CALC-006 = VIU-Deviation (FDBUG-10 CONFIRMED AGAIN 2026-07-10: pct "
     "0.005 accepted 201 and coerced to 0.01, resolved 0.02; flat 0.005 stored "
     "as 0.01). A=case-update to expect coercion + drop TICKET 5; B=dev change "
     "(store exact, likely new precision spec); C=file TICKET 5 as drafted."),
    (4,
     "FD-PROC-014 (Processing Fee minimum-amount rejection). No FDBUG - the §8 "
     "no-minimum invariant holds; the deviation is silent-ignore vs explicit "
     "reject. Related Story-8 context: builder UI absent (TICKET 11, PO round-1 "
     "Q3=B in-scope).",
     [("FD-PROC-014", 28532)],
     "requirements.md §9.2: S8-N6 (system rejects a Processing Fee carrying a "
     "minimum amount) + §5-R6 Min Amount data-model note (min supported for "
     "fee/discount kinds, not processing fees).",
     "FD-PROC-014 = VIU-Verified with a standing wording note (fresh pass "
     "2026-07-10: pfee minimum silently STRIPPED on create - 201, no min field "
     "persisted). A=spec/data-model change + new cases for pfee minimums; "
     "B=case-update FD-PROC-014 to expect explicit reject/absent field vs "
     "today's silent strip (minor dev tweak or accepted-behavior wording)."),
]

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
wb = Workbook()

ws = wb.active
ws.title = "Questions for PO"
ws.column_dimensions["A"].width = 5
for col, w in zip("BCDEF", [34, 52, 42, 56, 30]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Fees & Discounts - Round 2: Four Quick Questions for You"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
intro = ("Hi Chris! Thanks for your answers to the first set - they are all "
         "actioned. While finishing our checks we found four more spots where we "
         "need your call on how it should work. No wrong answers - for each one, "
         "pick an option (or write your own) in the \"Your answer\" box.")
ws["A2"] = intro
ws["A2"].alignment = WRAP
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 45

headers = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, item in enumerate(questions, start=1):
    row = HDR_ROW + i
    vals = [i, item["topic"], item["now"], item["q"], item["opts"], ""]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 150

# --- QA Internal Mapping sheet ---
wi = wb.create_sheet("QA Internal Mapping")
wi["A1"] = ("INTERNAL - for QA only. Do NOT share this tab (or any IDs/codes on it) "
            "with the PO.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:G1")

ihead = ["Q#", "Internal bug/case refs", "Case (internal ID)", "TestRail Case ID",
         "TestRail link", "Spec refs", "Current status notes"]
iwid = [5, 48, 16, 14, 44, 44, 52]
IH_ROW = 3
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for qno, refs, cases, spec, status in internal_map:
    first = r
    for fd_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        rowvals = [qno, refs, fd_id, f"C{tr_id}", url, spec, status]
        for c, v in enumerate(rowvals, start=1):
            cell = wi.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_CENTER if c in (1, 3, 4) else WRAP
            cell.border = BORDER
        link_cell = wi.cell(row=r, column=5)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT
        wi.row_dimensions[r].height = 120
        r += 1
    last = r - 1
    if last > first:  # merge the per-question columns across its case rows
        for col in (1, 2, 6, 7):
            wi.merge_cells(start_row=first, start_column=col,
                           end_row=last, end_column=col)

note_row = r + 1
wi.cell(row=note_row, column=1, value=(
    "Notes: Round-2 questions raised after the FRESH FULL VIU PASS 2026-07-10 "
    "(FeesDiscounts_FreshVIU_2026-07-10.xlsx). TestRail IDs sourced from "
    "testrail-id-map.csv (standing rule 8). Round-1 (6 questions) was answered by "
    "Chris Ward 2026-07-09 - see PO-Questions-SIMPLE.md / spec-v1-reconciliation.md. "
    "Related unfiled Jira drafts: jira-bug-drafts.md TICKET 4 (maxCap 0) and "
    "TICKET 5 (tiny-percent rounding) - hold both until these rulings land. "
    "Bugs/defects stay OUT of the PO-facing tab (standing rule 7); these 4 items "
    "are genuine product decisions (which behavior is intended), not defect "
    "reports.")).alignment = WRAP
wi.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
wi.row_dimensions[note_row].height = 75

wb.save(XLSX_OUT)

# ---------------------------------------------------------------------------
# Markdown mirror
# ---------------------------------------------------------------------------
md = []
md.append("# Fees & Discounts — Round 2: Four Quick Questions for You")
md.append("")
md.append("Hi Chris! Thanks for your answers to the first set — they are all")
md.append("actioned. While finishing our checks we found **four** more spots where")
md.append("we need your call on how it should work.")
md.append("")
md.append("There are **no wrong answers**. For each item, pick an option (or write")
md.append("your own) on the **\"Your answer\"** line. It should take just a few")
md.append("minutes.")
md.append("")
md.append("---")
for i, item in enumerate(questions, start=1):
    md.append("")
    md.append(f"## {i}. {item['topic']}")
    md.append("")
    md.append("**What happens now**")
    md.append(item["now"])
    md.append("")
    md.append("**The question**")
    md.append(item["q"])
    md.append("")
    md.append("**Options**")
    for line in item["opts"].split("\n"):
        md.append(f"- {line}")
    md.append("")
    md.append("**Your answer:** ______________________________________________")
    md.append("")
    md.append("---")
md.append("")
md.append("## Thank you!")
md.append("")
md.append("That's everything for this round. Your answers will help us finish this")
md.append("feature the way you want it. Feel free to add any notes alongside your")
md.append("choices.")
md.append("")
md.append("---")
md.append("---")
md.append("")
md.append("## Internal — QA-only mapping (NOT for the PO)")
md.append("")
md.append("This section links each plain-English question above to its internal")
md.append("bug/case refs, TestRail cases, spec refs and current status, so the")
md.append("answers can be actioned. **Do not include this section (or any IDs/codes")
md.append("in it) in the PO-facing copy or the \"Questions for PO\" tab.**")
md.append("")
for qno, refs, cases, spec, status in internal_map:
    md.append(f"### Q{qno}")
    md.append("")
    md.append(f"- **Internal refs:** {refs}")
    md.append("- **TestRail cases:**")
    for fd_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        md.append(f"  - {fd_id} — [C{tr_id}]({url})")
    md.append(f"- **Spec refs:** {spec}")
    md.append(f"- **Current status:** {status}")
    md.append("")
md.append("**Notes:** Round-2 questions raised after the FRESH FULL VIU PASS")
md.append("2026-07-10 (`FeesDiscounts_FreshVIU_2026-07-10.xlsx`). TestRail IDs")
md.append("sourced from `testrail-id-map.csv` (standing rule 8). Round-1 (6")
md.append("questions) was answered by Chris Ward 2026-07-09 — see")
md.append("`PO-Questions-SIMPLE.md` / `spec-v1-reconciliation.md`. Related unfiled")
md.append("Jira drafts: `jira-bug-drafts.md` TICKET 4 (maxCap 0) and TICKET 5")
md.append("(tiny-percent rounding) — hold both until these rulings land.")
md.append("Bugs/defects stay OUT of the PO-facing content (standing rule 7); these")
md.append("4 items are genuine product decisions (which behavior is intended), not")
md.append("defect reports.")
md.append("")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md))

print(f"Wrote {XLSX_OUT} and {MD_OUT} with {len(questions)} round-2 PO questions.")
