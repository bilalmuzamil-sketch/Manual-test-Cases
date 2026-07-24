#!/usr/bin/env python3
"""Fees & Discounts V1 — Blockers Tracker generator.

Classifies EVERY authored case (build/fees-discounts/cases/*.json) into a delivery
state and, if blocked, WHAT it is blocked on and WHO unblocks it. Emits:
  - build/fees-discounts/FeesDiscounts_Blockers_Tracker.xlsx (Tracker + Summary tabs)
  - build/fees-discounts/FeesDiscounts_Blockers_Tracker.md

Mirrors build/simple-flow/gen_blockers.py so the two projects read the same way.

State model (derived from each case's viu_status, viu-qb-findings.md FDBUG register
and bugs-log.md):
  READY                        — VIU-Verified, uploadable now.
  BLOCKED — DEVIATION          — built but deviates from spec (code-bug / PO-question /
                                 copy-drift). Owner: Dev or PO ruling.
  BLOCKED — DEV NOT BUILT      — surface absent (Story 8 Processing-Fee builder UI,
                                 Story 11 Part Sales). Owner: Dev.
  BLOCKED — ENV                — QuickBooks (Story 6) / flag-off / shared-env limits.
                                 Owner: QA env / dev.
  BLOCKED — NEEDS-ACCOUNT      — Story-13 per-role negatives needing role logins / a
                                 self-service staff role-switch. Owner: QA.
  BLOCKED — VIU PENDING (QA)   — built surface, not yet driven; needs fresh cookies /
                                 seeded data / another pass. Owner: QA.
"""
import json, os, re, csv
from collections import Counter
from whats_needed import whats_needed

BASE = os.path.dirname(os.path.abspath(__file__))          # build/fees-discounts
CASES_DIR = os.path.join(BASE, "cases")

# TestRail case-id map (standing rule 8: every case-listing deliverable carries the
# TestRail Case ID + clickable link). Source: testrail-id-map.csv.
TR_URL = "https://shopview.testrail.io/index.php?/cases/view/{}"
IDMAP = {}
with open(os.path.join(BASE, "testrail-id-map.csv"), encoding="utf-8") as _f:
    for _r in csv.DictReader(_f):
        IDMAP[_r["fd_id"]] = _r["ID"]


def tr_id(fd):
    return IDMAP.get(fd, "")


def tr_link(fd):
    v = IDMAP.get(fd, "")
    return TR_URL.format(v) if v and v.isdigit() else ""

# Batch-2 VIU wrote verbose annotations into some viu_status fields (e.g.
# "VIU-Verified qb 2026-07-08 — …" or "VIU-Pending — DEVIATION (flagged …)").
# Normalize any such string back to its base enum for classification.
ENUMS = ["VIU-Verified", "VIU-Deviation", "VIU-Blocked-NotBuilt", "VIU-Blocked-Env",
         "VIU-Pending"]


def norm_status(s):
    s = (s or "").strip()
    for e in ENUMS:
        if s.startswith(e):
            return e
    # V1_2 (2026-07-13): history-gating re-VIU cases + the new §5-R15 case carry a
    # "Pending ..." status; collapse to VIU-Pending for classification (see RETEST_V12).
    if s.startswith("Pending"):
        return "VIU-Pending"
    return s
OUT_XLSX = os.path.join(BASE, "FeesDiscounts_Blockers_Tracker.xlsx")
OUT_MD = os.path.join(BASE, "FeesDiscounts_Blockers_Tracker.md")

# Freshness stamp — bump when the deliverables are regenerated from the live source.
DATA_AS_OF = "2026-07-24"

FILES = [
    "group-A-wo-parts.json",
    "group-B-customer-admin-finance.json",
    "group-C-calc-permissions-validation.json",
]

# --- Classification inputs ---------------------------------------------------

# Story-13 per-role negatives still blocked after batch-2. The tech quick-login user
# is NOT in the org staff table on qb (staff/{id}/view → 404) and quick-login only
# supports admin/tech, so the OTHER 9 roles cannot be logged in / role-switched on
# this env — these need a real restricted-role account. (Batch-2 verified the rest of
# the Story-13 matrix by DERIVING per-role capability from roles-matrix.json, which is
# why FD-PERM-001/003/005/006/009, FD-HIST-005/006, FD-LABOR-007, FD-REMOVE-002 all
# flipped to VIU-Verified.)
NEEDS_ACCOUNT = {
    "FD-PERM-004": "Story 13 — Part-Sale adjustment permission (needs a non-Tech role login).",
    "FD-PERM-008": "Story 13 — customer-default view/change gate (Manage AP/AR + Cust Mgmt); needs the specific role.",
    "FD-PERM-010": "Story 13 — BOTH gates required (flag + permission) proven per-role; needs a restricted account.",
    "FD-CUST-015": "Story 13 — customer F&D tab per-role enforcement; needs a non-Tech role login.",
}

# The remaining VIU-Blocked-Env cases = true environment blockers (QuickBooks not
# present on qb, org-flag toggle skipped on shared env, shared library can't be
# emptied). Sub-note keyed by id-prefix / explicit.
ENV_FLAGOFF = {
    "FD-FLAG-001": "flag-off window (org-level FeesAndDiscounts toggle skipped on shared env).",
    "FD-FLAG-002": "flag-off exception (history log visible with flag off).",
    "FD-FLAG-003": "flag ON + permission-required interaction (flag-side complement).",
    "FD-HIST-004": "flag-off history visibility (org flag toggle skipped on shared env).",
    "FD-TMPL-012": "shared env: cannot empty the location's template library (real templates in use).",
}

# Deviations (VIU-Deviation status) that hang on a PRODUCT-OWNER ruling.
DEVIATION_PO = {
    "FD-STATS-001": "Stats F&D section is aggregate (Fees/Discounts/Net), not the spec's "
                    "per-adjustment %+Amount rows (BUG-FD-2/FDBUG-6). PO: is aggregate the intended V1?",
    "FD-PERM-002": "Whole-WO adjustment add/edit/remove is FE-only at the BE (tech got 201) "
                   "(BUG-FD-3). PO/dev: should whole-WO writes be BE-enforced for V1?",
    "FD-WO-013": "Whole-WO starting-places hidden without WO Create&Edit is FE-only "
                 "(BUG-FD-3). Same FE-vs-BE ruling as FD-PERM-002.",
}

# Deviations that are confirmed CODE BUGS (FDBUG register) — need a dev fix.
DEVIATION_BUG = {
    "FD-DOC-011": "FDBUG-1 (MAJOR) — WO/estimate Subtotal/Total exclude adjustments while GST taxes them.",
    "FD-PROC-009": "FDBUG-2 — processing-fee Grand-Total base wrongly includes whole-WO fees + their tax.",
    "FD-CALC-013": "FDBUG-2 — processing-fee base includes whole-WO adjustments.",
    "FD-HIST-001": "FDBUG-3 — auto-applied adjustments write NO history-log entry.",
    "FD-CALC-006": "FDBUG-10 — percent below minimum silently coerced up instead of rejected.",
    "FD-CALC-008": "FDBUG-9 — maxCap 0 accepted and treated as NO cap (should force $0.00).",
    "FD-VAL-006": "FDBUG-9 — maxCap 0 not treated as empty/never-sent.",
}

# VIU-Pending cases that batch-2 explicitly flagged "VIU-Pending — DEVIATION" for a PO
# ruling (surfaced as a sub-bucket so they aren't lost in the generic QA-pending pile).
PENDING_PO_FLAG = {
    "FD-WO-005": "BUG-FD-4 — Add-dialog 'Add' button not disabled on an empty form (inline-error vs disabled-button).",
    "FD-VAL-001": "BUG-FD-4 — Add button enabled on empty form; build validates on submit.",
    "FD-INLINE-003": "BUG-FD-5 — no 'Show N more' collapse when a line has ≥2 adjustments.",
    "FD-STATS-002": "BUG-FD-2 — Stats aggregate layout blocks per-adjustment row verification.",
    "FD-STATS-004": "BUG-FD-2 — Stats aggregate blocks creation-order row verification.",
    "FD-CUST-005": "NOTE-FD-5 — customer-default picker is a single-select dropdown, not a multi-select checkbox list.",
}

# V1_2 (2026-07-13): cases whose EXPECTED changed with the V1_2 spec (history-log
# gating flip per S13-R10 + the new §5-R15 dialog case) — each needs a live re-VIU.
RETEST_V12 = {
    "FD-PERM-009", "FD-HIST-006", "FD-HIST-001", "FD-HIST-002", "FD-HIST-003",
    "FD-HIST-004", "FD-HIST-005", "FD-HIST-007", "FD-HIST-008", "FD-FLAG-002",
    "FD-WO-016",
}

# VIU-Verified cases that still carry a PO confirmation ask (double-add did NOT
# reproduce on the batch-2 build — backend dedupes; PO to confirm the S9 fix shipped
# and re-scope EXPECTED to "single adjustment"). Kept READY, surfaced in "what's needed".
PO_CONFIRM_VERIFIED = {
    "FD-CUST-016": "Double-add (BUG-FD-1) did NOT reproduce on batch-2 (backend dedupes) — PO to confirm the S9 fix shipped; re-scope to single adjustment.",
    "FD-VAL-007": "Double-add (BUG-FD-1) not reproduced — PO to confirm fixed.",
}


def group_label(fn):
    return {"group-A-wo-parts.json": "A (WO/Parts)",
            "group-B-customer-admin-finance.json": "B (Customer/Admin/Finance)",
            "group-C-calc-permissions-validation.json": "C (Calc/Perms/Validation)"}[fn]


def load_cases():
    cases = []
    for fn in FILES:
        for c in json.load(open(os.path.join(CASES_DIR, fn))):
            c["_group"] = group_label(fn)
            cases.append(c)
    # Active cases only: retired cases (viu_status 'Retired — …') are kept in the
    # JSON for the record but excluded from every deliverable.
    # Retired 2026-07-20 (user ruling): FD-CUST-016 (duplicate of FD-VAL-007/C28605,
    # ex-C28500 deleted from TestRail). Active suite = 184 (185 authored - 1 retired).
    return [c for c in cases if not (c.get("viu_status") or "").startswith("Retired") and not c.get("dev_authored")]


def section_for(c):
    """TestRail section for this case; API-related cases route to 'API — <area>'
    (STANDING RULE 4). Kept in sync with gen_import.py."""
    area = c["area"].strip()
    if c.get("api_related"):
        return "API — " + area
    return area


def classify(c):
    cid = c["id"]
    vs = norm_status(c.get("viu_status", ""))
    story = c.get("story_ref") or ""

    if vs == "VIU-Verified":
        needs = "None — VIU-verified; uploadable now."
        sub = "verified"
        if cid in PO_CONFIRM_VERIFIED:
            needs = "Verified, but " + PO_CONFIRM_VERIFIED[cid]
            sub = "verified (PO-confirm)"
        return dict(state="READY", category="READY (VIU-Verified)", owner="—",
                    needs=needs, sub=sub, related=story)

    if vs == "VIU-Deviation":
        if cid in DEVIATION_BUG:
            return dict(state="BLOCKED", category="DEVIATION", owner="Dev team",
                        needs="Dev fix — " + DEVIATION_BUG[cid], sub="code-bug", related=story)
        if cid in DEVIATION_PO:
            return dict(state="BLOCKED", category="DEVIATION", owner="Product Owner / Dev ruling",
                        needs="PO ruling — " + DEVIATION_PO[cid], sub="PO-question", related=story)
        return dict(state="BLOCKED", category="DEVIATION", owner="QA (case-text update)",
                    needs="Build deviates from spec wording/UX — update case text once the build "
                          "is confirmed intended (label/copy/UX drift; see viu-qb-findings.md).",
                    sub="case-update", related=story)

    if vs == "VIU-Blocked-NotBuilt":
        if cid.startswith("FD-PROC"):
            dev = "Story 8 — Processing-Fee builder UI (SV-73xx)"
        elif cid.startswith("FD-PCOL") or cid == "FD-PERM-004":
            dev = "Story 11 — Part Sales fees/discounts (SV-73xx)"
        else:
            dev = "Story not built"
        return dict(state="BLOCKED", category="DEV NOT BUILT", owner="Dev team",
                    needs="Dev deploys {}; then QA re-runs VIU.".format(dev),
                    sub=dev, related=story)

    if vs == "VIU-Blocked-Env":
        if cid in NEEDS_ACCOUNT:
            return dict(state="BLOCKED", category="NEEDS-ACCOUNT", owner="QA (restricted-role session)",
                        needs="Restricted-role login (tech quick-login flaky on qb) or a "
                              "self-service staff role-switch: " + NEEDS_ACCOUNT[cid],
                        sub="needs-account", related=story)
        if cid.startswith("FD-QB"):
            return dict(state="BLOCKED", category="ENV", owner="Dev / QA env",
                        needs="QuickBooks integration not present on qb env (Story 6). Needs a "
                              "QB-connected env or dev/QB-side inspection.",
                        sub="quickbooks", related=story)
        note = ENV_FLAGOFF.get(cid, "environment limitation on the shared qb env.")
        return dict(state="BLOCKED", category="ENV", owner="QA env",
                    needs="Environment window — " + note, sub="flag-off/env", related=story)

    if vs == "VIU-Pending":
        if cid in RETEST_V12:
            return dict(state="BLOCKED", category="VIU PENDING (QA)", owner="QA (V1_2 re-VIU)",
                        needs="V1_2 spec change — EXPECTED updated (S13-R10 history gating / §5-R15 "
                              "note); needs a live re-VIU (re-derive the qb roles matrix first; "
                              "tech quick-login flaky).",
                        sub="v1_2-retest", related=story)
        if cid in PENDING_PO_FLAG:
            return dict(state="BLOCKED", category="VIU PENDING (QA)",
                        owner="QA + PO ruling",
                        needs="Batch-2 flagged as a deviation for PO confirmation, not rewritten — "
                              + PENDING_PO_FLAG[cid],
                        sub="pending (PO-flagged deviation)", related=story)
        return dict(state="BLOCKED", category="VIU PENDING (QA)", owner="QA (needs cookies+seed data)",
                    needs="Fresh qb cookies + seeded/throwaway data; drive this built surface "
                          "(parts UI flows, invoice-time walk, misc retests — see viu-qb-findings.md batch-2 backlog).",
                    sub="reachable/needs-data", related=story)

    return dict(state="BLOCKED", category="UNKNOWN", owner="?", needs="?", sub="?", related=story)


# Display order + colours.
CAT_DISPLAY = {
    "READY (VIU-Verified)": "READY (VIU-Verified)",
    "DEVIATION": "BLOCKED — DEVIATION",
    "DEV NOT BUILT": "BLOCKED — DEV NOT BUILT",
    "ENV": "BLOCKED — ENV",
    "NEEDS-ACCOUNT": "BLOCKED — NEEDS-ACCOUNT",
    "VIU PENDING (QA)": "BLOCKED — VIU PENDING (QA)",
}
ORDER = ["READY (VIU-Verified)", "BLOCKED — DEVIATION", "BLOCKED — DEV NOT BUILT",
         "BLOCKED — ENV", "BLOCKED — NEEDS-ACCOUNT", "BLOCKED — VIU PENDING (QA)"]
OWNER = {
    "READY (VIU-Verified)": "— (ready to upload)",
    "BLOCKED — DEVIATION": "Dev fix / PO ruling / QA case-update",
    "BLOCKED — DEV NOT BUILT": "Dev team (Stories 8 & 11)",
    "BLOCKED — ENV": "Dev / QA env (QuickBooks, flag-off)",
    "BLOCKED — NEEDS-ACCOUNT": "QA (restricted-role session)",
    "BLOCKED — VIU PENDING (QA)": "QA (fresh cookies + seed data)",
}
CAT_FILL = {
    "READY (VIU-Verified)": "C6EFCE",
    "BLOCKED — DEVIATION": "FCE4D6",
    "BLOCKED — DEV NOT BUILT": "F4CCCC",
    "BLOCKED — ENV": "D9E1F2",
    "BLOCKED — NEEDS-ACCOUNT": "E2D9F2",
    "BLOCKED — VIU PENDING (QA)": "FFF2CC",
}


def main():
    cases = load_cases()
    rows = []
    for c in cases:
        cls = classify(c)
        rows.append([
            c["id"], c["_group"], section_for(c), c["title"].strip(),
            norm_status(c.get("viu_status", "")),
            whats_needed(c["id"], c.get("viu_status", "")),
            cls["state"], CAT_DISPLAY[cls["category"]],
            cls["owner"], cls["needs"], cls["related"], cls["sub"],
            tr_id(c["id"]) or "pending-create", tr_link(c["id"]),
        ])

    # Column indices after inserting "What needs to be done (plain)" at index 5:
    # 0 id, 1 group, 2 area, 3 title, 4 status, 5 plain-what-to-do, 6 state,
    # 7 category, 8 owner, 9 needs, 10 related, 11 sub, 12 TR id, 13 TR link.
    disp_counts = Counter(r[7] for r in rows)
    group_counts = Counter(r[1] for r in rows)
    # Deviation sub-buckets.
    dev_sub = Counter(r[11] for r in rows if r[7] == "BLOCKED — DEVIATION")
    notbuilt_sub = Counter(r[11] for r in rows if r[7] == "BLOCKED — DEV NOT BUILT")
    env_sub = Counter(r[11] for r in rows if r[7] == "BLOCKED — ENV")

    # ---- What to send next ----
    n_pend = disp_counts["BLOCKED — VIU PENDING (QA)"]
    n_acct = disp_counts["BLOCKED — NEEDS-ACCOUNT"]
    n_notbuilt = disp_counts["BLOCKED — DEV NOT BUILT"]
    n_env_qb = env_sub.get("quickbooks", 0)
    n_env_flag = sum(v for k, v in env_sub.items() if k != "quickbooks")
    n_dev_bug = dev_sub.get("code-bug", 0)
    n_dev_po = dev_sub.get("PO-question", 0)
    n_dev_cu = dev_sub.get("case-update", 0)
    n_proc = notbuilt_sub.get("Story 8 — Processing-Fee builder UI (SV-73xx)", 0)
    n_pcol = notbuilt_sub.get("Story 11 — Part Sales fees/discounts (SV-73xx)", 0)

    send_next = [
        ("Fresh qb QA cookies (admin; tech quick-login is FLAKY — retest each run)",
         "unblocks the {} VIU-PENDING cases (parts UI flows, invoice-time walk, retests) AND — via "
         "the self-service staff role-switch — the {} NEEDS-ACCOUNT Story-13 per-role negatives.".format(n_pend, n_acct)),
        ("Dev deploys Story 8 (Processing-Fee builder UI) + Story 11 (Part Sales fees/discounts)",
         "unblocks the {} DEV-NOT-BUILT cases (Story 8 = {}, Story 11 = {}); then QA re-runs VIU.".format(n_notbuilt, n_proc, n_pcol)),
        ("A QuickBooks-connected env (or dev/QB-side inspection) for Story 6",
         "unblocks the {} QuickBooks ENV cases (mapping guard, sync, negative-total credit memo).".format(n_env_qb)),
        ("A flag-off maintenance window on a non-shared env",
         "unblocks the {} flag-off/shared-env cases (FD-FLAG-001/002/003, FD-HIST-004, FD-TMPL-012).".format(n_env_flag)),
        ("PO/dev rulings on the deviations + double-add + NOTE-FD-4",
         "finalizes the {} PO-question deviations (Stats layout, whole-WO FE-vs-BE enforcement), the "
         "double-add confirmation (FD-VAL-007; the duplicate FD-CUST-016/C28500 was retired 2026-07-20), and NOTE-FD-4 (BE accepts processing_fee). "
         "Dev fixes finalize the {} code-bug deviations; QA updates the {} copy/UX-drift deviations once "
         "the build is confirmed intended.".format(n_dev_po, n_dev_bug, n_dev_cu)),
        ("Restricted-role accounts (or confirmation the self-service staff role-switch is usable on qb)",
         "unblocks the {} NEEDS-ACCOUNT Story-13 per-role negatives (restore Tech afterward).".format(n_acct)),
    ]

    HEADER = ["Case ID", "Group", "Area (TestRail section)", "Title", "Current VIU status",
              "What needs to be done (plain)",
              "State", "Blocker category", "Who unblocks", "What's needed to unblock",
              "Related story/req", "Sub-bucket", "TestRail Case ID", "TestRail Link"]

    # ---------------- XLSX ----------------
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Blockers Tracker"
    ws.append(HEADER)
    for r in rows:
        ws.append(r)

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="305496")
    for col in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(vertical="center", horizontal="left")

    widths = {"Case ID": 14, "Group": 12, "Area (TestRail section)": 34, "Title": 55,
              "Current VIU status": 16, "What needs to be done (plain)": 60,
              "State": 10, "Blocker category": 24,
              "Who unblocks": 28, "What's needed to unblock": 62,
              "Related story/req": 30, "Sub-bucket": 18,
              "TestRail Case ID": 16, "TestRail Link": 52}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    wrap = Alignment(wrap_text=True, vertical="top")
    for ridx in range(2, len(rows) + 2):
        cat = ws.cell(row=ridx, column=8).value  # "Blocker category" is now column 8
        fill = CAT_FILL.get(cat)
        for cidx in range(1, len(HEADER) + 1):
            cell = ws.cell(row=ridx, column=cidx)
            cell.alignment = wrap
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
    ws.freeze_panes = "A2"

    # ---- Summary tab ----
    ss = wb.create_sheet("Summary")
    ss.append(["Fees & Discounts V1 — Blockers Tracker · Summary"])
    ss["A1"].font = Font(bold=True, size=14)
    ss.append(["Data as of:", DATA_AS_OF])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)
    ss.append([])
    ss.append(["Total authored cases", len(rows)])
    ss.append([])
    ss.append(["Blocker category", "Count", "Owner"])
    hr = ss.max_row
    for cidx in range(1, 4):
        ss.cell(row=hr, column=cidx).font = hdr_font
        ss.cell(row=hr, column=cidx).fill = hdr_fill
    for cat in ORDER:
        ss.append([cat, disp_counts.get(cat, 0), OWNER[cat]])
        rr = ss.max_row
        for cidx in range(1, 4):
            ss.cell(row=rr, column=cidx).fill = PatternFill("solid", fgColor=CAT_FILL[cat])
    ss.append(["TOTAL", len(rows), ""])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)

    ss.append([])
    ss.append(["By authoring group", "Count"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)
    for g in ["A (WO/Parts)", "B (Customer/Admin/Finance)", "C (Calc/Perms/Validation)"]:
        ss.append([g, group_counts.get(g, 0)])

    ss.append([])
    ss.append(["DEVIATION — by sub-bucket", "Count"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)
    for k in ["code-bug", "PO-question", "case-update"]:
        ss.append([k, dev_sub.get(k, 0)])

    ss.append([])
    ss.append(["DEV NOT BUILT — by story", "Count"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)
    for k, v in notbuilt_sub.most_common():
        ss.append([k, v])

    ss.append([])
    ss.append(["ENV — by sub-bucket", "Count"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True)
    ss.cell(row=ss.max_row, column=2).font = Font(bold=True)
    for k, v in env_sub.most_common():
        ss.append([k, v])

    ss.append([])
    ss.append(["WHAT TO SEND ME NEXT (to unblock each batch)"])
    ss.cell(row=ss.max_row, column=1).font = Font(bold=True, size=12)
    for what, effect in send_next:
        ss.append(["• " + what, effect])
        ss.cell(row=ss.max_row, column=1).font = Font(bold=True)

    ss.column_dimensions["A"].width = 62
    ss.column_dimensions["B"].width = 72
    ss.column_dimensions["C"].width = 30
    for row in ss.iter_rows():
        for cell in row:
            if cell.alignment.wrap_text is not True:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT_XLSX)
    print("Wrote", OUT_XLSX)

    # ---------------- Markdown ----------------
    def md(s):
        return (s or "").replace("|", "\\|").replace("\n", " ")

    L = []
    L.append("# Fees & Discounts V1 — Blockers Tracker")
    L.append("")
    L.append("> Source of truth for what every authored F&D case is waiting on and who "
             "unblocks it. Regenerate with `python3 build/fees-discounts/gen_blockers.py`.")
    L.append("> Canonical resume snapshot: `build/fees-discounts/PROJECT-STATE.md`. "
             "Interim upload file: `testrail-import/fees-discounts-v1-testrail-import.csv` (all {}).".format(len(rows)))
    L.append("")
    L.append("**Data as of: {}**".format(DATA_AS_OF))
    L.append("")
    L.append("**Total authored cases: {}**".format(len(rows)))
    L.append("")
    L.append("## Summary — counts per category")
    L.append("")
    L.append("| Blocker category | Count | Owner |")
    L.append("|---|---:|---|")
    for cat in ORDER:
        L.append("| {} | {} | {} |".format(cat, disp_counts.get(cat, 0), OWNER[cat]))
    L.append("| **TOTAL** | **{}** | |".format(len(rows)))
    L.append("")
    L.append("### By authoring group")
    L.append("")
    L.append("| Group | Count |")
    L.append("|---|---:|")
    for g in ["A (WO/Parts)", "B (Customer/Admin/Finance)", "C (Calc/Perms/Validation)"]:
        L.append("| {} | {} |".format(g, group_counts.get(g, 0)))
    L.append("")
    L.append("### DEVIATION — by sub-bucket")
    L.append("")
    L.append("| Sub-bucket | Count | Meaning |")
    L.append("|---|---:|---|")
    L.append("| code-bug | {} | confirmed FDBUG code defect — needs a dev fix |".format(dev_sub.get("code-bug", 0)))
    L.append("| PO-question | {} | expected hangs on a Product-Owner ruling |".format(dev_sub.get("PO-question", 0)))
    L.append("| case-update | {} | label/copy/UX drift — update case text once build confirmed intended |".format(dev_sub.get("case-update", 0)))
    L.append("")
    L.append("### DEV NOT BUILT — by story")
    L.append("")
    L.append("| Story | Count |")
    L.append("|---|---:|")
    for k, v in notbuilt_sub.most_common():
        L.append("| {} | {} |".format(k, v))
    L.append("")
    L.append("### ENV — by sub-bucket")
    L.append("")
    L.append("| Sub-bucket | Count |")
    L.append("|---|---:|")
    for k, v in env_sub.most_common():
        L.append("| {} | {} |".format(k, v))
    L.append("")
    L.append("## WHAT TO SEND ME NEXT (to unblock each batch)")
    L.append("")
    for what, effect in send_next:
        L.append("- **{}** → {}".format(what, effect))
    L.append("")
    L.append("## Full per-case tracker")
    L.append("")
    L.append("| Case ID | Group | Area | Title | VIU status | What needs to be done (plain) | "
             "State | Blocker category | "
             "Who unblocks | What's needed | Related | Sub-bucket | TestRail Case ID | TestRail Link |")
    L.append("|---|---|---|---|---|---|---|---|---|---|---|---|---|---|")
    for r in rows:
        trlink = "[C{0}]({1})".format(r[12], r[13]) if r[13] else r[12]
        L.append("| {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} | {} |".format(
            r[0], md(r[1]), md(r[2]), md(r[3]), md(r[4]), md(r[5]), r[6], r[7],
            md(r[8]), md(r[9]), md(r[10]), md(r[11]), r[12], trlink))
    L.append("")
    open(OUT_MD, "w").write("\n".join(L))
    print("Wrote", OUT_MD)

    print("\nCategory counts:", dict(disp_counts))
    print("Deviation sub:", dict(dev_sub))
    print("NotBuilt sub:", dict(notbuilt_sub))
    print("Env sub:", dict(env_sub))
    assert sum(disp_counts.values()) == len(rows) == 199, "count mismatch"  # 202 active - 3 retired 2026-07-22 (FD-LABOR-003/PCOL-003/PCOL-007 SV-8479/8480 consolidation)


if __name__ == "__main__":
    main()
