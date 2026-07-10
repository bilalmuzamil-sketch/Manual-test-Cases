#!/usr/bin/env python3
"""Generate FeesDiscounts_FreshVIU_2026-07-10.xlsx (+ .csv) — the fresh full-VIU results
workbook for the manual tester.

Tabs: Summary · one tab per status (Verified / Deviation / Blocked-NotBuilt /
Blocked-Env / Pending) · "Known Deviations & Blocked Areas" (plain layman English).
Every case row carries the TestRail Case ID (C#####) + clickable link
(standing rule 8; source: testrail-id-map.csv).

Regenerate: python3 build/fees-discounts/gen_fresh_viu_workbook.py
"""
import json, csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
FRESH = '2026-07-10'
OUT_X = os.path.join(HERE, f'FeesDiscounts_FreshVIU_{FRESH}.xlsx')
OUT_C = os.path.join(HERE, f'FeesDiscounts_FreshVIU_{FRESH}.csv')
TR_URL = 'https://shopview.testrail.io/index.php?/cases/view/{}'

# ---- load id map ----
idmap = {}
with open(os.path.join(HERE, 'testrail-id-map.csv'), encoding='utf-8') as f:
    for row in csv.DictReader(f):
        idmap[row['fd_id']] = row['ID']

# ---- load cases ----
cases = []
for fn in ['group-A-wo-parts.json', 'group-B-customer-admin-finance.json',
           'group-C-calc-permissions-validation.json']:
    d = json.load(open(os.path.join(HERE, 'cases', fn), encoding='utf-8'))
    cases += d['cases'] if isinstance(d, dict) else d

def fresh_note(c):
    n = c.get('notes', '')
    tag = 'FRESH VIU 2026-07-10: '
    i = n.rfind(tag)
    return n[i + len(tag):] if i >= 0 else n[-300:]

STATUSES = [
    ('VIU-Verified', 'Verified (fresh)'),
    ('VIU-Deviation', 'Deviations'),
    ('VIU-Blocked-NotBuilt', 'Blocked - Not Built'),
    ('VIU-Blocked-Env', 'Blocked - Environment'),
    ('VIU-Pending', 'Pending'),
]
by_status = {s: [c for c in cases if c['viu_status'] == s] for s, _ in STATUSES}

# ---- Known Deviations & Blocked Areas (plain layman English, one row per issue) ----
DEVIATIONS_PLAIN = [
    # (Area, What the tester will SEE, What the SPEC says should happen, Cases, Severity/Owner)
    ("Over-discount saves silently",
     "If you add a discount bigger than the work-order total, it saves immediately. The subtotal just becomes $0.00 and the extra amount is kept as a customer credit — you are never warned.",
     "A warning must appear first, telling you the total will floor at $0.00, that tax may still be owed, and how much extra will be carried as credit. You must confirm before it saves. (FDBUG-15)",
     "FD-QB-014", "Bug — dev fix"),
    ("Processing Fee cannot be created in the admin screen",
     "The template builder only offers 'Fee' and 'Discount' as types. There is no way to create a Processing Fee from the screen (re-checked today).",
     "'Processing Fee' should be a third type in the template builder. The PO confirmed this belongs in this release. (FDBUG-8)",
     "FD-PROC-001..004", "Missing feature — dev"),
    ("Part Sales has no fees & discounts at all",
     "The Part Sales page has no 'Fees & Discounts' column, no Add button, and no breakdown popup (re-checked today on a fresh part sale).",
     "Story 11: part sales should support fees/discounts per part with a column + popup.",
     "FD-PCOL-001..007, FD-PERM-004", "Missing feature — dev"),
    ("Automatic fees/discounts leave no history trail",
     "When a fee or discount is added automatically (from a location template or customer default), NOTHING is written to the work-order history. Only manual add/edit/remove are logged.",
     "Every adjustment that lands on a work order should be logged. (FDBUG-3)",
     "FD-HIST-001, FD-HIST-007", "Bug — dev fix"),
    ("Processing Fee is calculated on the wrong base",
     "A % Processing Fee includes other whole-work-order fees/discounts (and their tax) in its base. Example today: with only a $20 fee on the WO, a 3% processing fee charged $0.63 instead of $0.00.",
     "The Processing Fee base must EXCLUDE all whole-work-order fees/discounts. (FDBUG-2)",
     "FD-PROC-009, FD-CALC-013", "Bug — dev fix"),
    ("'Max $0' does not force the amount to zero",
     "Setting Max Amount to 0 on a percentage adjustment is accepted but behaves like NO cap at all (today: 20% resolved to $34.15 with Max 0).",
     "Max $0 should force the resolved amount to $0.00. (FDBUG-9)",
     "FD-CALC-008, FD-VAL-006, FD-TMPL-011", "Bug — dev fix"),
    ("Too-small percentages are silently rounded up",
     "Typing 0.005% is accepted and silently becomes 0.01% instead of being rejected.",
     "Values below the minimum (0.01%) should be rejected with an error. (FDBUG-10)",
     "FD-CALC-006", "Bug — dev fix"),
    ("Add buttons are never greyed out",
     "In the add fee/discount dialog (and the template builder), the Add/Create button is clickable even when the form is empty; errors only appear after you click.",
     "The button should stay disabled until Name, Type, Calculation type and Amount are all filled in. PO confirmed: grey it out. (BUG-FD-4)",
     "FD-WO-005, FD-VAL-001", "Bug — dev fix"),
    ("No 'Show more' on stacked line fees",
     "A line/part with 2+ fees shows every row at once — there is no 'Show N more' collapse (re-checked today).",
     "Only the first row should show, with a 'Show N more' toggle. PO confirmed this was in the design. (BUG-FD-5)",
     "FD-INLINE-003", "Bug — dev fix"),
    ("Stats tab shows only totals, not a table",
     "The Stats tab shows one combined 'Fees & Discounts (N) $X' figure instead of a table listing each fee/discount with its % and amount (re-checked today).",
     "Each adjustment should be its own row with Value/% and Amount columns. PO confirmed: per-row was the design. (BUG-FD-2 / FDBUG-6)",
     "FD-STATS-001, FD-STATS-002, FD-STATS-004", "Bug — dev fix"),
    ("Part fee/discount dialog labels are wrong",
     "On a part, the add dialog works but: the subtitle doesn't name the line/part number the spec way, the calculation shows a raw 'Pct_parts' value, and the part-percentage option is mislabeled '% of Labor Total' (it still calculates on the part total).",
     "Labels should read 'Line {N} Part — {name} ({number})' and '% of Parts Total'. (FDBUG-14)",
     "FD-PART-001", "Bug — dev fix (labels only)"),
    ("Empty name accepted by the back door",
     "NEW TODAY: the raw API accepts a fee/discount with an EMPTY name (the screen still blocks it, so normal users won't hit this).",
     "The server should reject empty names like it used to. (FDBUG-16 — API-only regression)",
     "FD-WO-008, FD-VAL-003 (note)", "Bug — dev fix (low)"),
    ("Creating work-order lines is broken on this test environment",
     "STILL BROKEN TODAY: adding any line to any work order fails with a server error (500) — from the screen or the API, plain or canned. You cannot build a fresh invoiceable work order.",
     "Line creation should work. This blocks: fresh part flows (requested→received), a Processing Fee on a money-bearing new WO, and clean QuickBooks export tests.",
     "FD-PART-005 (+ blocks re-tests)", "ENV bug — dev/env"),
    ("QuickBooks export of re-invoiced work orders fails",
     "Re-invoicing a work order that was invoiced before makes the QuickBooks export fail with a duplicate-document-number error (seen again today; entries land in Unexported Items).",
     "Export should succeed (or handle the number collision).",
     "FD-QB-001..003, FD-QB-015 (QB half)", "ENV/dev bug"),
    ("QuickBooks line-level checks can't be verified from ShopView",
     "What each invoice line looks like INSIDE QuickBooks (item, description, tax code, class, penny-capped discount lines, goodwill credit memo) cannot be read from ShopView — there is no read API. The item-unmapping switch also errors (500) so the 'blocked add' messages can't be triggered.",
     "Needs someone to eyeball the synced invoice inside the QuickBooks UI on a cleanly connected org.",
     "FD-QB-001..011, FD-QB-013, FD-QB-016, FD-CALC-017 (QB half), FD-QB-015 (memo half)", "Blocked — needs QB UI access"),
    ("Feature-flag OFF tests not run",
     "Turning the Fees & Discounts feature OFF for the whole org was deliberately NOT done today because other testers (including you) share this environment.",
     "Flag-off behavior (everything hidden, history preserved) still needs a private window.",
     "FD-FLAG-001..003, FD-HIST-004, FD-TMPL-012", "Blocked — env window"),
    ("Whole-WO permission only enforced by the screen",
     "Hiding 'add fee/discount' for users without Work-Order edit rights is only done by the UI; the server accepts the write anyway (per earlier testing; today the shared Technician role was changed so this could not be re-tested).",
     "Dev decision needed: should the server also enforce it? (BUG-FD-3)",
     "FD-PERM-002, FD-WO-013", "Dev decision"),
    ("Good news — two old bugs look FIXED",
     "1) The old 'totals exclude adjustments' money bug (FDBUG-1) did NOT reproduce again today — document Subtotal/Tax/Total all include adjustments and match the app exactly (3rd clean pass). 2) Customers created via the API now DO inherit auto-apply defaults (FDBUG-12 fixed).",
     "Keep an eye out during manual testing but treat both as fixed on this build.",
     "FD-DOC-011, FD-CUST-014", "FYI — verify in passing"),
]

HEAD = ['FD ID', 'TestRail Case ID', 'TestRail Link', 'Area', 'Title', 'Priority',
        'Fresh Status (2026-07-10)', 'Fresh Evidence / Note']

def row_for(c):
    tid = idmap.get(c['id'], '')
    return [c['id'], f'C{tid}' if tid else '(unmapped)',
            TR_URL.format(tid) if tid else '', c.get('area', ''), c['title'],
            c.get('priority', ''), c['viu_status'], fresh_note(c)]

wb = Workbook()
bold = Font(bold=True); wrap = Alignment(wrap_text=True, vertical='top')
fill = PatternFill('solid', fgColor='DDDDDD')

# Summary tab
ws = wb.active; ws.title = 'Summary'
ws.append([f'Fees & Discounts V1 — FRESH full VIU pass, {FRESH} (all 182 cases re-adjudicated)'])
ws['A1'].font = Font(bold=True, size=13)
ws.append([])
ws.append(['Status', 'Count', 'Meaning'])
for cell in ws[3]: cell.font = bold; cell.fill = fill
MEAN = {
    'VIU-Verified': 'Exercised on the live build today (or evidence re-validated) and matches the spec',
    'VIU-Deviation': 'Built but differs from the spec — see the Known Deviations tab',
    'VIU-Blocked-NotBuilt': 'The screen/feature does not exist yet (Story 8 builder UI, Story 11 Part Sales)',
    'VIU-Blocked-Env': 'Cannot be tested on this environment (QuickBooks internals, flag-off window, broken unmap)',
    'VIU-Pending': 'Not yet verifiable (blocked by the line-create environment bug)',
}
for s, label in STATUSES:
    ws.append([label, len(by_status[s]), MEAN[s]])
ws.append(['TOTAL', len(cases), ''])
ws.append([])
ws.append(['Environment', 'app qb.qa.shopview.com · API sv7387api.qa.shopview.com · FeesAndDiscounts flag ON · QuickBooks connected & both items mapped'])
ws.append(['Env health today', 'Awake & healthy. KNOWN ENV BUGS: work-order line creation 500s everywhere; QB export of re-invoiced WOs fails on duplicate numbers; QB item-unmap (PUT settings) 500s.'])
ws.append(['Cleanup', 'Batch-6 leftovers cleaned (S-15895 restored to baseline 182.76/9.14/191.90; failed exports marked done; 3 of 4 leftover WOs deleted — S-15947 is Complete and undeletable).'])
for col, w in enumerate([28, 10, 120], 1): ws.column_dimensions[get_column_letter(col)].width = w
for r in ws.iter_rows():
    for c in r: c.alignment = wrap

# Per-status tabs
for s, label in STATUSES:
    t = wb.create_sheet(label[:31])
    t.append(HEAD)
    for cell in t[1]: cell.font = bold; cell.fill = fill
    for c in sorted(by_status[s], key=lambda x: x['id']):
        t.append(row_for(c))
        link = t.cell(row=t.max_row, column=3)
        if link.value: link.hyperlink = link.value; link.font = Font(color='0563C1', underline='single')
    widths = [14, 14, 44, 30, 60, 10, 22, 90]
    for col, w in enumerate(widths, 1): t.column_dimensions[get_column_letter(col)].width = w
    for r in t.iter_rows(min_row=2):
        for c in r: c.alignment = wrap

# Known Deviations & Blocked Areas tab
t = wb.create_sheet('Known Deviations & Blocked')
t.append(['#', 'Area / Issue', 'What you will SEE on the build', 'What the SPEC says should happen', 'Related cases', 'Type / Owner'])
for cell in t[1]: cell.font = bold; cell.fill = fill
for i, (area, see, spec, cs, owner) in enumerate(DEVIATIONS_PLAIN, 1):
    t.append([i, area, see, spec, cs, owner])
for col, w in enumerate([4, 34, 62, 62, 34, 22], 1): t.column_dimensions[get_column_letter(col)].width = w
for r in t.iter_rows(min_row=2):
    for c in r: c.alignment = wrap

wb.save(OUT_X)

# CSV (flat, all cases)
with open(OUT_C, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(HEAD)
    for s, _ in STATUSES:
        for c in sorted(by_status[s], key=lambda x: x['id']):
            w.writerow(row_for(c))

print('written', OUT_X)
print('written', OUT_C)
print({s: len(v) for s, v in by_status.items()})
