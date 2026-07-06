#!/usr/bin/env python3
"""Assemble Fees & Discounts V1 manual test cases into an Excel workbook + CSV.

Data-processing only. Reads the three group JSON case files, merges (keeping IDs,
group A -> B -> C order), and emits:
  - FeesDiscounts_V1_TestCases.xlsx  (Summary / Test Cases / VIU (pending) / Open Questions)
  - FeesDiscounts_V1_TestCases.csv   (the Test Cases sheet)
"""
import csv
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
CASES_DIR = os.path.join(BASE, "cases")
XLSX = os.path.join(BASE, "FeesDiscounts_V1_TestCases.xlsx")
CSV_PATH = os.path.join(BASE, "FeesDiscounts_V1_TestCases.csv")

GROUP_FILES = [
    "group-A-wo-parts.json",
    "group-B-customer-admin-finance.json",
    "group-C-calc-permissions-validation.json",
]

# ---------------------------------------------------------------- load + merge
cases = []
for gf in GROUP_FILES:
    with open(os.path.join(CASES_DIR, gf), encoding="utf-8") as fh:
        cases.extend(json.load(fh))  # group A then B then C, original order within

# ---------------------------------------------------------------- area buckets
def bucket(area):
    a = area or ""
    # order matters: check "Customer document" before generic "Customer"
    rules = [
        ("Work Order — Whole-WO", "Whole-WO"),
        ("Work Order — Labor-line", "Labor line"),
        ("Part-line", "Part line"),
        ("Parts page", "Parts column"),
        ("Work Order — Inline display", "Inline display"),
        ("Work Order — Statistics", "Stats"),
        ("Work Order — Financial Info", "Financial Info"),
        ("Work Order — Sidebar", "Financial Info"),
        ("Work Order — Edit", "Edit"),
        ("Work Order — Remove", "Remove"),
        ("Work Order — Multiple adjustments", "Stacking"),
        ("Customer document", "Finance/Invoice"),
        ("Customer", "Customer defaults"),
        ("Template admin", "Template admin"),
        ("Processing Fee", "Processing Fee"),
        ("QuickBooks", "QuickBooks"),
        ("History log", "History"),
        ("Calculation contract", "Calculation"),
        ("Permissions", "Permissions"),
        ("Feature-flag", "Feature-flag"),
        ("Validation", "Validation"),
    ]
    for needle, name in rules:
        if a.startswith(needle) or (needle == "Part-line" and needle in a):
            return name
    return "Other"

AREA_ORDER = [
    "Whole-WO", "Labor line", "Part line", "Parts column", "Inline display",
    "Stats", "Financial Info", "Edit", "Remove", "Stacking",
    "Customer defaults", "Template admin", "Processing Fee", "Finance/Invoice",
    "QuickBooks", "History", "Calculation", "Permissions", "Feature-flag",
    "Validation",
]

for c in cases:
    c["_bucket"] = bucket(c["area"])

# ---------------------------------------------------------------- styles
DARK_BLUE = "1F4E78"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
ALT_FILL = PatternFill("solid", fgColor="F2F5FA")
BANNER_FILL = PatternFill("solid", fgColor="C00000")
BANNER_FONT = Font(bold=True, color="FFFFFF", size=12)
TITLE_FONT = Font(bold=True, size=14, color=DARK_BLUE)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_TOP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")

PRIORITY_FILL = {
    "Critical": PatternFill("solid", fgColor="F4B6B6"),  # red-ish
    "High": PatternFill("solid", fgColor="FFE08A"),      # amber
    "Medium": PatternFill("solid", fgColor="DCE6F5"),    # light blue
    "Low": PatternFill("solid", fgColor="EDEDED"),       # light grey
}


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def n_join(items):
    return "\n".join(str(x) for x in (items or []))


wb = Workbook()

# ================================================================ Summary
ws = wb.active
ws.title = "Summary"

intro = (
    "Fees & Discounts V1 - Manual Test Cases\n\n"
    "This workbook contains 182 manual test cases covering the ShopView "
    "\"Fees & Discounts\" V1 feature: whole-work-order, labor-line and part-line "
    "fees/discounts; inline and statistics display; the Financial Info and sidebar "
    "cards; the Parts page column and breakdown modal; edit/remove/stacking flows; "
    "customer-level defaults; template administration; the Processing Fee; "
    "estimate/invoice (finance) rendering; QuickBooks sync; the history log; the "
    "calculation contract; permissions (Story 13); feature-flag gating; and "
    "validation/edge cases.\n\n"
    "The cases were authored from the complete written spec plus the design "
    "mockups (see build/fees-discounts/requirements.md and design-notes.md).\n\n"
    "VIU (Verify-in-UI on staging) is PENDING: the feature is not yet deployed on "
    "staging. The feature flag is ON, but the backend and UI have not shipped "
    "(expected in roughly two weeks). Because of this, EVERY case is marked "
    "\"VIU pending\" - the cases are written to the spec and must be re-verified "
    "against the live app once it is deployed. See the \"VIU (pending)\" tab.\n\n"
    "Where the design mockups and the written spec disagree, or where the current "
    "build is known to differ, the case Notes flag it and the item is consolidated "
    "on the \"Open Questions\" tab - review that tab before execution."
)
ws["A1"] = intro
ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
ws["A1"].font = Font(size=11)
ws.merge_cells("A1:E1")
ws.row_dimensions[1].height = 300

# area table
r = 3
ws.cell(row=r, column=1, value="Cases by Area").font = TITLE_FONT
r += 1
ws.cell(row=r, column=1, value="Area").font = HEADER_FONT
ws.cell(row=r, column=2, value="# Cases").font = HEADER_FONT
for col in (1, 2):
    ws.cell(row=r, column=col).fill = HEADER_FILL
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
area_counts = {}
for c in cases:
    area_counts[c["_bucket"]] = area_counts.get(c["_bucket"], 0) + 1
r += 1
area_start = r
for name in AREA_ORDER:
    ws.cell(row=r, column=1, value=name).border = BORDER
    ws.cell(row=r, column=2, value=area_counts.get(name, 0)).border = BORDER
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
    r += 1
ws.cell(row=r, column=1, value="GRAND TOTAL").font = BOLD
ws.cell(row=r, column=2, value=sum(area_counts.values())).font = BOLD
for col in (1, 2):
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).fill = ALT_FILL
ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
r += 2

# priority table
ws.cell(row=r, column=1, value="Cases by Priority").font = TITLE_FONT
r += 1
ws.cell(row=r, column=1, value="Priority").font = HEADER_FONT
ws.cell(row=r, column=2, value="#").font = HEADER_FONT
for col in (1, 2):
    ws.cell(row=r, column=col).fill = HEADER_FILL
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).alignment = Alignment(horizontal="center")
prio_counts = {}
for c in cases:
    prio_counts[c["priority"]] = prio_counts.get(c["priority"], 0) + 1
r += 1
for p in ["Critical", "High", "Medium", "Low"]:
    if p in prio_counts:
        ws.cell(row=r, column=1, value=p).border = BORDER
        ws.cell(row=r, column=1).fill = PRIORITY_FILL.get(p)
        ws.cell(row=r, column=2, value=prio_counts[p]).border = BORDER
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        r += 1
ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=2, value=sum(prio_counts.values())).font = BOLD
for col in (1, 2):
    ws.cell(row=r, column=col).border = BORDER
    ws.cell(row=r, column=col).fill = ALT_FILL
ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
r += 2

note = (
    "Note: See the \"Open Questions\" tab for every design vs spec conflict and "
    "current-build caveat that must be resolved, and the \"VIU (pending)\" tab to "
    "record on-staging verification once the feature is deployed."
)
ws.cell(row=r, column=1, value=note).font = Font(italic=True)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 45

ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 12
for col in "CDE":
    ws.column_dimensions[col].width = 16
ws.freeze_panes = "A2"

# ================================================================ Test Cases
tc = wb.create_sheet("Test Cases")
COLS = [
    ("Test ID", 14), ("Area", 26), ("Story Ref", 16), ("Title", 50),
    ("Priority", 12), ("Type", 12), ("Permissions Required", 34),
    ("Preconditions", 52), ("Steps To Reproduce", 56), ("Expected Result", 56),
    ("Design Ref", 30), ("VIU Status", 20), ("Notes", 46),
]
headers = [c[0] for c in COLS]
tc.append(headers)
style_header(tc, 1, len(headers))

csv_rows = [headers]
for i, c in enumerate(cases):
    row = [
        c["id"], c["area"], c["story_ref"], c["title"], c["priority"], c["type"],
        c["permissions_required"], n_join(c["preconditions"]), n_join(c["steps"]),
        n_join(c["expected"]), c["design_ref"], c["viu_status"], c.get("notes", ""),
    ]
    tc.append(row)
    csv_rows.append(row)
    excel_row = i + 2
    alt = (i % 2 == 1)
    for col in range(1, len(headers) + 1):
        cell = tc.cell(row=excel_row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL
    # priority colour (col 5)
    pfill = PRIORITY_FILL.get(c["priority"])
    if pfill:
        pc = tc.cell(row=excel_row, column=5)
        pc.fill = pfill
        pc.alignment = WRAP_TOP_CENTER
    tc.cell(row=excel_row, column=1).font = BOLD

for idx, (_, width) in enumerate(COLS, start=1):
    tc.column_dimensions[get_column_letter(idx)].width = width
tc.freeze_panes = "B2"  # freeze header row + Test ID column

# ================================================================ VIU (pending)
viu = wb.create_sheet("VIU (pending)")
banner = (
    "VIU PENDING - Fees & Discounts not yet deployed on staging. Fill this tab "
    "once the feature is live (see build/fees-discounts/RESUME-STRATEGY.md)."
)
viu.append([banner])
viu.merge_cells("A1:H1")
viu["A1"].fill = BANNER_FILL
viu["A1"].font = BANNER_FONT
viu["A1"].alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
viu.row_dimensions[1].height = 42

VIU_COLS = [
    ("Test ID", 14), ("Area", 26), ("Title", 60), ("Permissions Required", 34),
    ("VIU Result (Pass/Fail/Blocked)", 24),
    ("Observed Behavior on Staging", 46), ("Screenshot", 22), ("Notes", 40),
]
viu.append([c[0] for c in VIU_COLS])
style_header(viu, 2, len(VIU_COLS))

for i, c in enumerate(cases):
    viu.append([c["id"], c["area"], c["title"], c["permissions_required"],
                "", "", "", ""])
    excel_row = i + 3
    alt = (i % 2 == 1)
    for col in range(1, len(VIU_COLS) + 1):
        cell = viu.cell(row=excel_row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL
    viu.cell(row=excel_row, column=1).font = BOLD

for idx, (_, width) in enumerate(VIU_COLS, start=1):
    viu.column_dimensions[get_column_letter(idx)].width = width
viu.freeze_panes = "A3"

# ================================================================ Open Questions
oq = wb.create_sheet("Open Questions")
OQ_COLS = [
    ("#", 5), ("Topic/Area", 24), ("The Question or Conflict", 70),
    ("Spec Ref", 22), ("How To Resolve", 22), ("Related Case IDs", 34),
]
oq.append([c[0] for c in OQ_COLS])
style_header(oq, 1, len(OQ_COLS))

open_questions = [
    ("Calculation base selector",
     "The design Add/Edit modal has NO calculation-base selector and computes a "
     "percentage against a single flat WO total. Spec requires the base to depend "
     "on scope/method (% of Labor Total / Parts Total / Subtotal for whole-WO). "
     "Confirm staging offers the correct base options and resolves against them.",
     "S5-R4 / §5-R10", "VIU / Dev",
     "FD-WO-003, FD-WO-011, FD-LABOR-002, FD-CALC-001, FD-CALC-012, FD-VAL-001"),
    ("Taxable field in dialog",
     "The design Add/Edit modal collects NO Taxable field. Spec requires a Taxable "
     "Yes/No dropdown (default Yes) and a 'Tax is recalculated on save.' preview "
     "footer. Confirm the live dialog has a Taxable control.",
     "S2-R26 / S2-R35", "VIU / Dev",
     "FD-WO-004, FD-CALC-011, FD-VAL-001"),
    ("Max Amount (cap) scope",
     "The mock labels the field 'Max cap', shows it unconditionally and even caps "
     "Flat amounts. Spec restricts 'Max Amount (Optional)' to percentage methods "
     "only (never Flat / Processing Fee). Confirm staging hides it for Flat and "
     "treats an entered 0 as no maximum.",
     "S2-R24 / §5-R6", "VIU / Dev",
     "FD-WO-006, FD-TMPL-011, FD-CALC-007, FD-VAL-006"),
    ("Preview fee/discount colour inverted",
     "The design preview renders discounts green and fees amber - the OPPOSITE of "
     "the spec (fee green, discount red). Sidebar/inline amounts should be plain "
     "grey outside the preview. Confirm the actual colours on staging.",
     "S2-R33 / S12-R1", "VIU / Product Owner",
     "FD-WO-015, FD-FIN-004, FD-INLINE-001"),
    ("Adjustment ordering",
     "Group-A task wording says 'descending build order', but the spec is explicit "
     "that WO screens list adjustments in creation order OLDEST first. Confirm the "
     "actual order on staging.",
     "§5-R9", "VIU / Product Owner",
     "FD-STATS-004"),
    ("Remove vs Delete wording",
     "The design uses 'Delete'/'Edit' on the inline line/part menu and WO card but "
     "'Remove' on the customer tab and parts breakdown. Spec confirm dialog is "
     "titled 'Remove Fee / Discount'. Regardless of label, removal uses the "
     "Create & Edit permission (not Delete). Capture the actual label per surface.",
     "S13-R7 / §10", "VIU / Product Owner",
     "FD-REMOVE-001, FD-REMOVE-002, FD-INLINE-005, FD-PCOL-006, FD-PERM-006"),
    ("Add-button label fixed vs dynamic",
     "The design uses a fixed 'Add' button label. Spec says the WO dialog flips to "
     "'Add Fee' / 'Add Discount' by type, while the template create dialog keeps a "
     "fixed 'Add Fee / Discount'. Confirm the button label behavior on staging.",
     "S2-R27 / S7-R17", "VIU",
     "FD-WO-002, FD-WO-007, FD-TMPL-003"),
    ("Sidebar card Add control",
     "The v1 mockup sidebar card includes an 'Add Fee / Discount' button, which "
     "contradicts the spec (no Add control on the card). Confirm whether staging "
     "shows an in-card Add button.",
     "S3-R10", "VIU / Product Owner",
     "FD-FIN-004, FD-FIN-005"),
    ("'History mode' undefined",
     "The spec does not define what puts a user into 'history mode' (which hides "
     "the add/edit controls). This precondition cannot be made concrete until the "
     "trigger is confirmed on staging.",
     "requirements §14 item 3", "VIU / Product Owner",
     "FD-PERM-011, FD-WO-001"),
    ("Auto-apply + customer-default double-add",
     "A fee reaching a WO via BOTH auto-apply and a customer default may be added "
     "twice. This is a KNOWN DEFECT, not a spec requirement - log actual behavior "
     "in the VIU tab rather than failing the case against spec. (Manual re-apply of "
     "a template to one WO is separately allowed.)",
     "requirements §14 item 4 / S9", "Dev / Product Owner",
     "FD-STACK-003, FD-CUST-016, FD-VAL-007"),
    ("Whole-WO Flat Amount base",
     "The spec base table lists explicit bases only for the three percentage "
     "methods; Whole-WO Flat Amount base = 'the set amount' by inference (no "
     "explicit Whole-WO Flat base row). Confirm none is expected.",
     "requirements §14 item 2 / §5-R4", "Dev / Product Owner",
     "FD-WO-002, FD-CALC-003"),
    ("Processing Fee legal-disclosure text",
     "The Processing Fee legal-disclosure literal text is NOT in the exported spec. "
     "Capture the exact wording from the live dialog before asserting it. Spec says "
     "'toggle' but the control is a Yes/No dropdown.",
     "requirements §14 item 5", "VIU / Product Owner",
     "FD-PROC-004"),
    ("Processing Fee absent from mocks",
     "The Processing Fee UI is not modeled in the HTML mockups (template builder, "
     "methods, WO remove-only behavior). Verify the whole Processing Fee flow live.",
     "S8", "VIU",
     "FD-PROC-001, FD-CALC-013, FD-TMPL-014"),
    ("Admin page breadth vs Settings -> Finance",
     "Current build may show the admin Fees & Discounts page to any user with a "
     "location; spec (S13-R8) tightens visibility to Settings -> Finance. Record "
     "which gating staging actually enforces.",
     "requirements §10.4 / S7-R7b / S13-R8", "VIU / Dev",
     "FD-TMPL-016, FD-PERM-007"),
    ("Possible unsplit WO-edit permission",
     "The current build may use one WO-edit permission check where the spec splits "
     "whole-WO (Create & Edit) from line-level (Work Order Lines: Create & Edit). "
     "Record which permission staging actually enforces.",
     "requirements §10.4 / S13-R3 / S13-R4", "VIU / Dev",
     "FD-WO-013, FD-PERM-002, FD-PERM-003"),
    ("History 'processing_fee' label",
     "Current build may show raw 'processing_fee' on the history 'Applied to:' line "
     "instead of 'Full invoice' for a Processing Fee. Verify the labels on staging.",
     "requirements §14 item 6", "VIU / Dev",
     "FD-HIST-007, FD-PERM-009"),
    ("Estimate/invoice & admin layouts have no mockup",
     "There is no HTML mockup for the customer estimate/invoice (finance) view, the "
     "admin Fees & Discounts page, or the Edit dialog / locked-field behavior. "
     "Confirm all these layouts live.",
     "S3 / S7 / §10", "VIU",
     "FD-FIN-001 (Group B), FD-TMPL-001, FD-EDIT-001"),
    ("Stats total placeholder mismatch",
     "The design mock's Statistics total placeholder (-$1,198.02) disagrees with "
     "its Financial Info total (-$2,055.25) for the same '(5)' set - a mock "
     "placeholder inconsistency. Confirm the correct signed sum on staging.",
     "design §3", "VIU",
     "FD-STATS-003"),
    ("Inline validation error text not modeled",
     "The mock only disables the Add button and shows NO inline error copy for "
     "empty Name/Amount, over-100% discount, or negative input. Spec expects inline "
     "error text. Capture the actual rejection UX on staging.",
     "S2-N1 / S2-N2 / §10", "VIU / Dev",
     "FD-WO-005, FD-WO-008, FD-WO-010, FD-VAL-002, FD-VAL-004, FD-VAL-005, FD-CALC-005"),
    ("Statistics column header label",
     "Spec Statistics tab uses a '%' column; the design mock uses a 'Value' column. "
     "Confirm the actual header on staging.",
     "S4-R2", "VIU",
     "FD-STATS-001"),
    ("Processing Fee Edit control on WO",
     "On a WO a Processing Fee is remove-only (no Edit), but the current build "
     "reportedly still shows an Edit control that fails. Record actual behavior.",
     "S8-R17 / requirements §14 item 6", "VIU / Dev",
     "FD-PROC-008, FD-CALC-013"),
    ("Manage AP/AR still gates customer F&D defaults",
     "Per project notes, Manage AP/AR no longer gates aging reports, but it IS "
     "still the gate (with Customer Management) for customer Fees & Discounts "
     "defaults. Verify staging enforces both.",
     "S13-R9", "VIU / Dev",
     "FD-PERM-008"),
]

for i, (topic, q, ref, resolve, ids) in enumerate(open_questions, start=1):
    oq.append([i, topic, q, ref, resolve, ids])
    excel_row = i + 1
    alt = (i % 2 == 1)
    for col in range(1, len(OQ_COLS) + 1):
        cell = oq.cell(row=excel_row, column=col)
        cell.alignment = WRAP_TOP
        cell.border = BORDER
        if alt:
            cell.fill = ALT_FILL
    oq.cell(row=excel_row, column=1).alignment = WRAP_TOP_CENTER

for idx, (_, width) in enumerate(OQ_COLS, start=1):
    oq.column_dimensions[get_column_letter(idx)].width = width
oq.freeze_panes = "A2"

# ================================================================ save
wb.save(XLSX)

with open(CSV_PATH, "w", newline="", encoding="utf-8") as fh:
    writer = csv.writer(fh)
    for row in csv_rows:
        writer.writerow(row)

# ---------------------------------------------------------------- report
print("Total cases on Test Cases tab:", len(cases))
print("Open questions:", len(open_questions))
print("Areas:")
for name in AREA_ORDER:
    print(f"  {area_counts.get(name,0):3d}  {name}")
print("Priorities:", {p: prio_counts[p] for p in ["Critical","High","Medium","Low"] if p in prio_counts})
print("Saved:", XLSX)
print("Saved:", CSV_PATH)
