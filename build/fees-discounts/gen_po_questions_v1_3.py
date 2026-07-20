#!/usr/bin/env python3
"""Generate the V1_3 follow-up PO question sheet for Chris Ward (Fees & Discounts PO).

Outputs (regenerable — run from the repo root):
  - build/fees-discounts/PO-Questions-Chris-V1_3_2026-07-17.xlsx
  - build/fees-discounts/PO-Questions-Chris-V1_3_2026-07-17.md

Sheets (format mirrors gen_po_questions_round2.py 1:1):
  - "Questions for PO"    : reader-facing, VERY SIMPLE layman language ONLY.
                            NO case IDs, FDBUG codes, API/HTTP terms, or jargon.
                            Columns: # | Topic | What happens now | The question |
                            Options | Your answer (blank).
  - "QA Internal Mapping" : QA-only. One row per (question, TestRail case) with
                            internal bug/case refs, TestRail Case ID C##### +
                            clickable link (standing rule 8; source
                            testrail-id-map.csv), spec refs, and status notes.

Source of the two questions: the V1_3 spec ambiguities flagged (do-not-resolve)
in build/fees-discounts/spec-v3-2026-07-17/spec-diff-v3-2026-07-17.md §H
(a = note placement inconsistency; b = SFD-gate observability at the WO dialog).
Round-2 reference: gen_po_questions_round2.py / PO-Questions-Round2.md (all 4
round-2 questions were answered by Chris Ward 2026-07-14).
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

XLSX_OUT = "build/fees-discounts/PO-Questions-Chris-V1_3_2026-07-17.xlsx"
MD_OUT = "build/fees-discounts/PO-Questions-Chris-V1_3_2026-07-17.md"

TR_LINK = "https://shopview.testrail.io/index.php?/cases/view/{}"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ---------------------------------------------------------------------------
# Reader-facing content (layman ONLY — no IDs, no bug codes, no tech terms)
# ---------------------------------------------------------------------------
questions = [
    {
        "topic": "Where exactly should the tax-area note appear?",
        "now": ("Your update says the tax-area note (\"Fees may vary based on the "
                "customer's tax jurisdiction\") is shown only to people who can see "
                "money amounts. But the write-up points to different places for the "
                "note itself: one part says it appears in the window where fee "
                "templates are created and edited, another part says it appears in "
                "the Processing-Fee window, and a third phrase says it sits below "
                "\"every\" place where Taxable can be chosen. These don't fully "
                "agree."),
        "q": ("Exactly where should this tax-area note appear?"),
        "opts": ("A) Only the Processing-Fee window.\n"
                 "B) Every window that has a Taxable choice - including the "
                 "ordinary window where fee templates are created and edited.\n"
                 "C) Somewhere else - please describe."),
        "answer": "",
        "action": "",
    },
    {
        "topic": "Who would ever notice the note being hidden?",
        "now": ("People who cannot see money amounts also cannot open the fee "
                "window on a work order at all - so for the work-order window, "
                "hiding the note from them changes nothing anyone can see. The "
                "only place we can genuinely check the new \"only people who can "
                "see money amounts\" rule is the admin window where fee templates "
                "are created and edited."),
        "q": ("Is that expected?"),
        "opts": ("A) Yes - the rule mainly matters in the admin window.\n"
                 "B) No - there is another place where a person who cannot see "
                 "money amounts would still see fees; please describe it."),
        "answer": "",
        "action": "",
    },
]

# ---------------------------------------------------------------------------
# QA-internal mapping (rule 8: TestRail C##### + clickable link per case;
# IDs sourced from build/fees-discounts/testrail-id-map.csv)
# ---------------------------------------------------------------------------
# Each entry: (q_no, internal_refs, [(fd_id, testrail_id), ...], spec_refs, status)
internal_map = [
    (1,
     "V1_3 §5-R15 note-placement inconsistency (spec-diff-v3-2026-07-17.md §H-a; "
     "spec inconsistent, flagged - do not resolve unilaterally). Affected cases: "
     "FD-WO-016 (WO Add/Edit dialog note, §5-R15 SFD-gate expected applied "
     "2026-07-17) + FD-PROC-004 (Processing-Fee dialog folded §5-R15 note check). "
     "Answer decides whether the plain template dialog (S7-R12f Taxable control) "
     "also needs the note -> possible new case / FD-PROC-004-style fold.",
     [("FD-WO-016", 29441), ("FD-PROC-004", 28522)],
     "requirements.md §17 (V1_3): §5-R15 body names the WO Add/Edit dialog "
     "(S2-R26) + the Processing Fee dialog (S8-R11); the 2026-07-14 change-log "
     "entry says \"the work-order Add / Edit dialog and the template dialog\"; "
     "§5-R15 opens with \"Below every Taxable control\"; Story 7's own template "
     "dialog spec (S7-R12f) carries no §5-R15 reference. Three readings conflict.",
     "FD-WO-016 = VIU-Deviation (note absent in the build for an SFD user); "
     "FD-PROC-004 = Blocked-NotBuilt (Story-8 builder UI absent). A = FD-PROC-004 "
     "unchanged + no template-dialog case; B = add a template-dialog note case "
     "(area Templates - admin) + keep both existing; C = per description. No "
     "TestRail write until answered + authorized."),
    (2,
     "V1_3 SFD-gate testability caveat (spec-diff-v3-2026-07-17.md §H-b). "
     "Stories 1/2 prerequisites already require See Financial Data to open the "
     "WO Add/Edit dialog, so the no-SFD negative on the §5-R15 note is only "
     "independently observable at the admin template/Processing-Fee dialog "
     "(Story 7 prerequisites = administration access only, no SFD). Determines "
     "where the SFD-gate negative case (fold in FD-WO-016 vs standalone) must "
     "observe, and whether a WO-side negative is even meaningful.",
     [("FD-WO-016", 29441), ("FD-PROC-004", 28522)],
     "requirements.md §17 (V1_3): §5-R15 gate (\"visible only to users with See "
     "Financial Data\") + change-log rationale (restricted roles in admin "
     "shouldn't flag a missing note); Stories 1/2 prerequisites (SFD required to "
     "open the WO fee dialog) vs Story 7 prerequisites (admin access only).",
     "FD-WO-016 carries the folded SFD-negative expected (applied 2026-07-17, "
     "C29441 update 200); status VIU-Deviation until the note ships. A = admin "
     "dialog is the negative's observation point (current folding stands); B = "
     "author additional coverage at the described surface. No TestRail write "
     "until answered + authorized."),
]

# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------
wb = Workbook()

ws = wb.active
ws.title = "Questions for PO"
ws.column_dimensions["A"].width = 5
for col, w in zip("BCDEF", [34, 52, 42, 56, 30]):
    ws.column_dimensions[col].width = w

ws["A1"] = "Fees & Discounts - Two Quick Questions About Your Latest Update"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:F1")
intro = ("Hi Chris! We have worked your latest write-up (the tax-area note and "
         "the switch to \"audit log\" wording) into our checks. Two spots in the "
         "write-up point in different directions, so we need your call. No wrong "
         "answers - for each one, pick an option (or write your own) in the "
         "\"Your answer\" box.")
ws["A2"] = intro
ws["A2"].alignment = WRAP
ws.merge_cells("A2:F2")
ws.row_dimensions[2].height = 45

headers = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
HDR_ROW = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=HDR_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
ws.freeze_panes = "A5"

for i, item in enumerate(questions, start=1):
    row = HDR_ROW + i
    answer_cell = ("ANSWER: " + item["answer"] + "\n\nRESULTING ACTION: "
                   + item["action"]) if item.get("answer") else ""
    vals = [i, item["topic"], item["now"], item["q"], item["opts"], answer_cell]
    for c, v in enumerate(vals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    ws.row_dimensions[row].height = 150

# --- QA Internal Mapping sheet ---
wi = wb.create_sheet("QA Internal Mapping")
wi["A1"] = ("INTERNAL - for QA only. Do NOT share this tab (or any IDs/codes on it) "
            "with the PO.")
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:G1")

ihead = ["Q#", "Internal bug/case refs", "Case (internal ID)", "TestRail Case ID",
         "TestRail link", "Spec refs", "Current status notes"]
iwid = [5, 48, 16, 14, 44, 44, 52]
IH_ROW = 3
for c, (h, w) in enumerate(zip(ihead, iwid), start=1):
    cell = wi.cell(row=IH_ROW, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
wi.freeze_panes = "A4"

r = IH_ROW + 1
for qno, refs, cases, spec, status in internal_map:
    first = r
    for fd_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        rowvals = [qno, refs, fd_id, f"C{tr_id}", url, spec, status]
        for c, v in enumerate(rowvals, start=1):
            cell = wi.cell(row=r, column=c, value=v)
            cell.alignment = WRAP_CENTER if c in (1, 3, 4) else WRAP
            cell.border = BORDER
        link_cell = wi.cell(row=r, column=5)
        link_cell.hyperlink = url
        link_cell.font = LINK_FONT
        wi.row_dimensions[r].height = 120
        r += 1
    last = r - 1
    if last > first:  # merge the per-question columns across its case rows
        for col in (1, 2, 6, 7):
            wi.merge_cells(start_row=first, start_column=col,
                           end_row=last, end_column=col)

note_row = r + 1
wi.cell(row=note_row, column=1, value=(
    "Notes: V1_3 follow-up questions sourced from the spec-diff FLAGS section "
    "(build/fees-discounts/spec-v3-2026-07-17/spec-diff-v3-2026-07-17.md §H a/b; "
    "V1_3 applied to cases + TestRail 2026-07-17, see testrail-update-log.md in "
    "the same folder). TestRail IDs sourced from testrail-id-map.csv (standing "
    "rule 8). Round-1 (6 questions) answered by Chris Ward 2026-07-09; Round-2 "
    "(4 questions) answered by Chris Ward 2026-07-14 - see PO-Questions-SIMPLE.md "
    "/ PO-Questions-Round2.md. These 2 items are genuine product decisions (the "
    "spec points in different directions), not defect reports - bugs/defects stay "
    "OUT of the PO-facing tab (standing rule 7).")).alignment = WRAP
wi.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=7)
wi.row_dimensions[note_row].height = 75

wb.save(XLSX_OUT)

# ---------------------------------------------------------------------------
# Markdown mirror
# ---------------------------------------------------------------------------
md = []
md.append("# Fees & Discounts — Two Quick Questions About Your Latest Update")
md.append("")
md.append("Hi Chris! We have worked your latest write-up (the tax-area note and")
md.append("the switch to \"audit log\" wording) into our checks. Two spots in the")
md.append("write-up point in different directions, so we need your call.")
md.append("")
md.append("There are **no wrong answers**. For each item, pick an option (or write")
md.append("your own) on the **\"Your answer\"** line. It should take just a couple")
md.append("of minutes.")
md.append("")
md.append("---")
for i, item in enumerate(questions, start=1):
    md.append("")
    md.append(f"## {i}. {item['topic']}")
    md.append("")
    md.append("**What happens now**")
    md.append(item["now"])
    md.append("")
    md.append("**The question**")
    md.append(item["q"])
    md.append("")
    md.append("**Options**")
    for line in item["opts"].split("\n"):
        md.append(f"- {line}")
    md.append("")
    if item.get("answer"):
        md.append(f"**Your answer:** {item['answer']}")
        md.append("")
        md.append(f"**Resulting action:** {item['action']}")
    else:
        md.append("**Your answer:** ______________________________________________")
    md.append("")
    md.append("---")
md.append("")
md.append("## Thank you!")
md.append("")
md.append("That's everything for this round. Your answers will help us finish this")
md.append("feature the way you want it. Feel free to add any notes alongside your")
md.append("choices.")
md.append("")
md.append("---")
md.append("---")
md.append("")
md.append("## Internal — QA-only mapping (NOT for the PO)")
md.append("")
md.append("This section links each plain-English question above to its internal")
md.append("bug/case refs, TestRail cases, spec refs and current status, so the")
md.append("answers can be actioned. **Do not include this section (or any IDs/codes")
md.append("in it) in the PO-facing copy or the \"Questions for PO\" tab.**")
md.append("")
for qno, refs, cases, spec, status in internal_map:
    md.append(f"### Q{qno}")
    md.append("")
    md.append(f"- **Internal refs:** {refs}")
    md.append("- **TestRail cases:**")
    for fd_id, tr_id in cases:
        url = TR_LINK.format(tr_id)
        md.append(f"  - {fd_id} — [C{tr_id}]({url})")
    md.append(f"- **Spec refs:** {spec}")
    md.append(f"- **Current status:** {status}")
    md.append("")
md.append("**Notes:** V1_3 follow-up questions sourced from the spec-diff FLAGS")
md.append("section (`spec-v3-2026-07-17/spec-diff-v3-2026-07-17.md` §H a/b; V1_3")
md.append("applied to cases + TestRail 2026-07-17 — see `testrail-update-log.md`")
md.append("in the same folder). TestRail IDs sourced from `testrail-id-map.csv`")
md.append("(standing rule 8). Round-1 (6 questions) answered by Chris Ward")
md.append("2026-07-09; Round-2 (4 questions) answered by Chris Ward 2026-07-14 —")
md.append("see `PO-Questions-SIMPLE.md` / `PO-Questions-Round2.md`. These 2 items")
md.append("are genuine product decisions (the spec points in different")
md.append("directions), not defect reports — bugs/defects stay OUT of the")
md.append("PO-facing content (standing rule 7).")
md.append("")

with open(MD_OUT, "w") as f:
    f.write("\n".join(md))

print(f"Wrote {XLSX_OUT} and {MD_OUT} with {len(questions)} V1_3 PO questions.")
