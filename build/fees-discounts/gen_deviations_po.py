#!/usr/bin/env python3
"""Generate Deviations-and-Questions-for-PO.xlsx from the curated rows.

One 'Deviations & Questions' sheet with the columns the user asked for:
# | Case/Bug ref | Type | What we see | Spec/expected | Proposed resolution | Answer
Plus a Summary sheet. Mirrors the Simple Flow OpenQuestions workbook style.
Run: python3 build/fees-discounts/gen_deviations_po.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = os.path.join(os.path.dirname(__file__), "Deviations-and-Questions-for-PO.xlsx")

# Type values: PO-question / case-update / bug-confirm
ROWS = [
    # PO questions
    ("FD-STATS-001 (BUG-FD-2/FDBUG-6; +FD-STATS-002/004)", "PO-question",
     "Statistics tab F&D area is an aggregate ('Fees (3) $227.90 / Discounts (0) $0.00 / Net $227.90'); no per-adjustment rows.",
     "S4-R2/R3: 'Fees & Discounts (N)' section listing each adjustment with a %/Value column and an Amount column, signed.",
     "PO ruling: aggregate layout intended for V1, or is the per-adjustment table still to be built? (Settles FD-STATS-002/004 too.)"),
    ("FD-PERM-002 (BUG-FD-3)", "PO-question",
     "Technician without workOrdersCreateAndEdit can add a whole-WO adjustment: POST /work-orders/adjustments/add scope=whole_wo -> 201 (batch-2 reconfirmed). FE-only gate; BE does not enforce. (Templates + customer-defaults ARE BE-enforced -> 403.)",
     "S13-R3: whole-WO add/edit/remove requires Work Orders: Create and Edit.",
     "PO/dev: should whole-WO adjustment writes be backend-enforced for V1, or is FE-only gating acceptable?"),
    ("FD-WO-013 (BUG-FD-3)", "PO-question",
     "Same root cause as FD-PERM-002 - whole-WO starting places hidden in FE only; BE does not enforce. Per-role UI hiding not exercisable on qb (admin/tech logins only).",
     "S1-N2/S13-N2: starting places hidden without Work Orders: Create & Edit.",
     "Same ruling as FD-PERM-002. Also confirm split whole-WO vs line permission model (S13-R3/R4) vs the build's single WO-edit check (req. 10.4)."),
    ("FD-CUST-016 / FD-VAL-007 (BUG-FD-1 double-add)", "PO-question",
     "Old double-add defect did NOT reproduce. Controlled repro (template that is BOTH auto-apply AND a customer default, fresh WO create) -> exactly ONE adjustment per template (backend dedupes).",
     "Sec.14 item 4 / S9 known gap: intended result is exactly ONE adjustment.",
     "PO: confirm the S9 dedupe fix has shipped so we treat double-add as fixed and keep these cases as 'single adjustment'."),
    ("NOTE-FD-4 (Story 8)", "PO-question",
     "Template builder Type dropdown offers only Fee/Discount (no Processing Fee), but the BE accepts POST /adjustment-templates {kind:'processing_fee', calculationType:'pct_grand_total'} -> 201 and it auto-applies/resolves correctly.",
     "Story 8: Processing Fee is a first-class template type with its own builder UI.",
     "PO/dev: is Story 8 partially built on purpose (BE ready, UI pending)? Should the BE keep accepting kind:processing_fee before the builder ships, or reject it until the UI lands?"),
    ("FD-WO-005 / FD-VAL-001 (BUG-FD-4)", "PO-question",
     "Add Fee/Discount dialog confirm button is enabled with Name+Amount empty; required-field validation fires on submit (inline 'Amount must be greater than 0'), not by disabling the button.",
     "Design Sec.6 validateForm() / S2-N1/N2: confirm button disabled until form is valid.",
     "PO: is inline-error-on-submit acceptable for V1, or should the button be disabled-until-valid as designed?"),
    ("FD-INLINE-003 (BUG-FD-5)", "PO-question",
     "With two line-level adjustments on one line, both inline rows render; no 'Show N more'/'Show less' collapse toggle observed.",
     "S3-R15/R16 / S12-R6: show first row + a 'Show N more' toggle when >=2.",
     "PO: is the collapse toggle planned for V1, or is always-expanded acceptable?"),
    ("FD-CUST-005 (NOTE-FD-5/FDBUG-7; ties FD-CUST-003/004/006)", "PO-question",
     "Customer-default 'Add Fee/Discount' control is a single-select dropdown + Save (no caption; generic 'No results' empty; direct trash-icon remove). Backend accepts a templateIds array, so multi-add works under the hood - only the UI multi-select is missing.",
     "S9-R18..R24: caption + checkbox multi-select list with 'Add'; empty state 'No templates available to add.'; 3-dot 'Remove'; removal toast.",
     "PO: is the single-select dropdown the intended V1 picker, or should the checkbox multi-select + exact copy be built? (Settles FD-CUST-003/004/006/007 case-updates.)"),
    # Case-update deviations
    ("FD-WO-001", "case-update",
     "Dialog opens at whole-WO scope (correct); title reads 'Add new fee/discount' with WO number as subtitle; menu item 'Add Fee/Discount'.",
     "Title 'New Fee / Discount'; WO-toolbar label 'Add Work Order Fee / Discount'.",
     "Update expected: title 'Add new fee/discount' (WO number subtitle); menu/toolbar label 'Add Fee/Discount'."),
    ("FD-LABOR-001", "case-update",
     "Dialog scope-locked to the labor line (correct); subtitle 'Applying to: {line name}' - omits 'Line {N} Labor -' prefix.",
     "Subtitle 'Applying to: Line {N} Labor - {name}'.",
     "Update expected: subtitle 'Applying to: {line name}' (no 'Line N Labor -' prefix). [Same prefix-omission class as part-line FDBUG-14 - if prefix still wanted, stays a build gap.]"),
    ("FD-FIN-004", "case-update",
     "Card lists whole-WO adjustments correctly; title 'WO Fees & Discounts'; hover menu 'Edit | Remove'.",
     "Card title 'Work Order Fee / Discount'; hover menu 'Edit / Delete'.",
     "Update expected: card title 'WO Fees & Discounts'; hover menu 'Edit | Remove'."),
    ("FD-REMOVE-001", "case-update",
     "Confirm title 'Remove Fee / Discount' (correct); message 'Are you sure you want to remove this fee?'; toast 'Fee removed'; remove -> 204.",
     "Message 'Remove \"{name}\" from this work order?'; toast 'Discount removed'.",
     "Update expected: message 'Are you sure you want to remove this fee?'; toast 'Fee removed'/'Discount removed' (confirm if it varies by kind)."),
    ("FD-CUST-003", "case-update",
     "Add works end-to-end; toast 'Fee / discount added.' (correct). Picker is a single dropdown + Cancel/Save, no caption.",
     "Picker = checkbox list of rows (Name/Type/Calc/Amount); toast 'Fee / discount added.'",
     "Update expected: picker is a single-select dropdown + Save; toast 'Fee / discount added.' [Contingent on the FD-CUST-005 picker ruling.]"),
    ("FD-CUST-004", "case-update",
     "UI cannot multi-select (single dropdown); backend accepts a templateIds array (multi-link 201) - only UI + plural toast missing.",
     "Multi-select 3 templates in one action; toast '3 fees / discounts added.'",
     "Keep as build gap (if multi-select wanted) OR rewrite to single-add. [Blocked on FD-CUST-005 ruling.]"),
    ("FD-CUST-006", "case-update",
     "With every template linked, the dropdown shows the generic 'No results'.",
     "Empty picker message 'No templates available to add.'",
     "Update expected: empty state 'No results' (or dev fixes copy - contingent on FD-CUST-005)."),
    ("FD-CUST-007", "case-update",
     "Remove = direct trash icon per row; no confirm (correct, S9-R24); NO removal toast observed; row/count update.",
     "Row 3-dot menu 'Remove'; no confirm; toast 'Fee / discount removed.'",
     "Update expected: remove via direct trash icon, no confirm, row+count update; confirm whether a 'Fee / discount removed' toast is expected."),
    ("FD-TMPL-001", "case-update",
     "Page at Administration -> Finance -> 'Fees & Discounts' (below Payment Methods), route /administration/adjustment-templates. Matches S13-R8 target.",
     "Page at Administration -> Service -> Fees & Discounts, below Canned Lines.",
     "Update expected: page at Administration -> Finance -> 'Fees & Discounts'."),
    ("FD-TMPL-003", "case-update",
     "Create works. Button 'Create'; toast 'Template created'; amount label '$ Default Amount'; Taxable & auto-apply are toggles; extra 'Description (Optional)' (255) field; auto-apply label 'Auto-apply to new work orders'. Title 'New Fee / Discount' correct.",
     "Button 'Add Fee / Discount'; toast 'Fee added'; amount label 'Amount'; Taxable/auto-apply Yes/No dropdown + checkbox; no Description field; auto-apply '...at this location'.",
     "Update expected to reflect actual labels/controls (button 'Create', toast 'Template created', '$ Default Amount', toggles, Description field, 'Auto-apply to new work orders')."),
    ("FD-TMPL-004", "case-update",
     "Discount template created fine; toast is the generic 'Template created'.",
     "Toast 'Discount added'.",
     "Update expected: toast 'Template created' (generic, not per-type)."),
    ("FD-TMPL-006", "case-update",
     "Row-click opens 'Edit Fee / Discount' + 'Save' (correct); toast 'Template updated'; Type + Calc locked in template edit.",
     "Toast 'Fee updated'; Type/Calc editable in template edit.",
     "Update expected: toast 'Template updated'; Type/Calc locked in edit."),
    ("FD-TMPL-008", "case-update",
     "Warning present + backed by delete-precondition API; reads 'This template is set as a default for 1 customer. Deleting it will remove it from them.'",
     "Warning '...set as a default for [N] customer(s). Their defaults will be removed too.'",
     "Update expected: 'This template is set as a default for {N} customer(s). Deleting it will remove it from them.'"),
    ("FD-TMPL-010", "case-update",
     "Line-scope Add dialog has NO template picker at all (control absent); whole-WO dialog picker lists every template.",
     "Line-scope picker filters by method + hint 'Showing templates compatible with this line.'; excludes Processing Fee.",
     "Update expected: line-scope Add dialog has no template picker (templates via whole-WO dialog). [Overlaps FDBUG-13 - if filtered line-scope picker still planned, stays a build gap.]"),
    ("FD-TMPL-011", "case-update",
     "Max Amount shows for % only (correct). But maxCap:0 is stored and resolves with NO cap (10% of $324.60 -> $32.46 despite maxCap 0).",
     "Max Amount for % only (verified) AND Max Amount 0 treated as empty (no cap -> $0.00).",
     "Keep expected #1/#2. Do NOT rewrite 0-handling - it is a code bug tracked as FDBUG-9; fix in build."),
    ("FD-PROC-008", "case-update",
     "Removal works; BE correctly 409s an edit ('A processing fee cannot be edited through this endpoint.'), but the WO card menu still shows 'Edit | Remove'. Matches Sec.14 current-build gap.",
     "A Processing Fee offers 'Delete' only (no Edit).",
     "Update expected: menu shows 'Edit | Remove' but Edit fails with 409; removal works - pending S8-R17 cleanup (remove the dead Edit control)."),
    ("FD-HIST-002", "case-update",
     "Detail shows Name, Amount (set rate), Applied-to 'Full invoice' (all correct) - but there is no 'Type:' line.",
     "History detail shows Name, Type, Amount (set rate), Applied-to.",
     "Keep Name/Amount/Applied-to. Missing 'Type:' line is a code bug tracked as FDBUG-11; do not drop from expected unless PO descopes S10-R6b."),
    # bug-confirm
    ("FD-DOC-011 (FDBUG-1 inconsistency)", "bug-confirm",
     "Batch 1/2: FDBUG-1 reproduced (WO total_cost + Financial Info + estimate Subtotal/Total EXCLUDED net adjustments while GST INCLUDED their tax). Batch 4: did NOT reproduce on estimate docs (totals reconciled on 3 WOs).",
     "S5-R5: Adjustments block sits BEFORE Subtotal (i.e. included in Subtotal & Total).",
     "Dev: was a totals/tax fix shipped 07-08->07-09 (known/partial)? Which surfaces/scenarios? QA to run a controlled re-check to pin the trigger (suspect discount-heavy/excess-credit/specific surface). FD-DOC-011 expected left unchanged. Filed as High bug draft."),
]

def main():
    wb = Workbook()

    # ---- main sheet ----
    ws = wb.active
    ws.title = "Deviations & Questions"
    headers = ["#", "Case/Bug ref", "Type", "What we see", "Spec/expected",
               "Proposed resolution", "Answer (for PO/dev)"]
    ws.append(headers)

    hfill = PatternFill("solid", fgColor="1F4E78")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    type_fill = {
        "PO-question": PatternFill("solid", fgColor="FCE4D6"),
        "case-update": PatternFill("solid", fgColor="E2EFDA"),
        "bug-confirm": PatternFill("solid", fgColor="FFF2CC"),
    }
    for c in ws[1]:
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for i, (ref, typ, see, spec, prop) in enumerate(ROWS, start=1):
        ws.append([i, ref, typ, see, spec, prop, ""])
        r = ws.max_row
        for c in ws[r]:
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = border
        ws.cell(r, 1).alignment = Alignment(horizontal="center", vertical="top")
        ws.cell(r, 3).fill = type_fill.get(typ, PatternFill())
        ws.cell(r, 3).font = Font(bold=True)

    widths = [4, 34, 14, 52, 40, 52, 30]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{ws.max_row}"

    # ---- summary sheet ----
    s = wb.create_sheet("Summary")
    po = sum(1 for r in ROWS if r[1] == "PO-question")
    cu = sum(1 for r in ROWS if r[1] == "case-update")
    bc = sum(1 for r in ROWS if r[1] == "bug-confirm")
    s.append(["Fees & Discounts V1 — Deviations & Open Questions for PO / Dev"])
    s["A1"].font = Font(bold=True, size=13)
    s.append([])
    s.append(["Type", "Count", "Notes"])
    for c in s[3]:
        c.fill = hfill; c.font = hfont; c.border = border
    rows = [
        ("PO-question", po, "Behaviour differs from spec; needs a product ruling. (FD-PERM-002 & FD-WO-013 share one ruling; FD-CUST-005 governs FD-CUST case-updates.)"),
        ("case-update", cu, "App is acceptable; our expected wording is stale spec text. 13 safe; 4 overlap FDBUG code bugs (FD-TMPL-011->FDBUG-9, FD-HIST-002->FDBUG-11, FD-TMPL-010->FDBUG-13, FD-CUST-003/004/006 blocked on the picker ruling)."),
        ("bug-confirm", bc, "FDBUG-1 reproduced in batch 1/2, not batch 4 - needs a controlled re-check + a known/partial-fix answer. Also filed as a High bug draft."),
        ("TOTAL", po + cu + bc, "All rows in the 'Deviations & Questions' sheet."),
    ]
    for name, cnt, note in rows:
        s.append([name, cnt, note])
        for c in s[s.max_row]:
            c.border = border; c.alignment = Alignment(vertical="top", wrap_text=True)
        if name == "TOTAL":
            for c in s[s.max_row]:
                c.font = Font(bold=True)
    s.append([])
    s.append(["Confirmed code bugs are written up as ready-to-file tickets in jira-bug-drafts.md."])
    s.append(["Env: qb.qa.shopview.com / sv7387api.qa.shopview.com (SV-7387), FeesAndDiscounts flag ON."])
    s.append(["Nothing filed; nothing written to TestRail; case JSONs unchanged."])
    s.column_dimensions["A"].width = 16
    s.column_dimensions["B"].width = 8
    s.column_dimensions["C"].width = 100

    wb.save(OUT)
    print("wrote", OUT, "| rows:", len(ROWS), "| PO", po, "case-update", cu, "bug-confirm", bc)

if __name__ == "__main__":
    main()
