#!/usr/bin/env python3
"""Generate PO-Questions-SIMPLE.xlsx — plain-English Fees & Discounts questions for the PO.

Sheets:
  - About        : friendly intro
  - PO Questions  : # | Topic | What happens now | The question | Options | Your answer
  - Internal      : QA-only mapping (Q# -> case IDs / refs); NOT for the PO
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT = "build/fees-discounts/PO-Questions-SIMPLE.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

questions = [
    {
        "topic": "Stats page shows a combined total, not a line-by-line list",
        "now": ("On the Statistics page, fees and discounts are shown as a single "
                "rolled-up total (for example, \"Fees (3): $227.90\"). They are not "
                "listed out one at a time."),
        "q": ("Is the combined total what you want for this release, or should each "
              "fee and discount be listed separately with its own row?"),
        "opts": ("A) Keep it as it is - one combined total is fine for now.\n"
                 "B) Change it so each fee and discount is listed on its own row."),
    },
    {
        "topic": "A customer's default fee now adds only once - please confirm",
        "now": ("Previously there was a worry that a customer's default fee could get "
                "added twice to a new work order. In our testing it now adds only once, "
                "which looks fixed."),
        "q": "Can you confirm that adding it only once is the intended behavior?",
        "opts": ("A) Yes - adding it only once is correct. Treat this as settled.\n"
                 "B) No - it should behave differently (please describe)."),
    },
    {
        "topic": "\"Processing Fee\" isn't visible in the app yet, but is partly ready",
        "now": ("The \"Processing Fee\" option isn't available to pick in the app yet. "
                "However, the underlying system will already partly accept it."),
        "q": ("Is Processing Fee meant to be part of this release, or is it planned for "
              "a later one?"),
        "opts": ("A) It's coming in a later release - leave the visible option out for now.\n"
                 "B) It should be part of this release - the visible option needs to be added."),
    },
    {
        "topic": "The \"Add\" button on the fee form is clickable before the form is filled in",
        "now": ("When adding a fee, the \"Add\" button can be clicked even before the "
                "form is complete. If you click it too early, it shows an error message "
                "instead of staying greyed out until everything is filled in correctly."),
        "q": ("Is showing an error on click acceptable for this release, or should the "
              "\"Add\" button stay greyed out until the form is valid?"),
        "opts": ("A) Keep it as it is - showing an error on click is fine for now.\n"
                 "B) Change it so the button is greyed out until the form is filled in correctly."),
    },
    {
        "topic": "When a line has several fees/discounts, they all show at once (no \"show more\")",
        "now": ("When a single line has more than one fee or discount, all of them are "
                "shown at the same time. There's no \"show more / show less\" option to "
                "collapse the list."),
        "q": ("Is showing them all at once fine for this release, or should there be a "
              "\"show more\" collapse when there are several?"),
        "opts": ("A) Keep it as it is - showing them all at once is fine.\n"
                 "B) Change it so extra ones collapse under a \"show more\" option."),
    },
    {
        "topic": "On customer-defaults you add templates one at a time from a dropdown",
        "now": ("On the customer-defaults screen, you pick fee/discount templates from "
                "a dropdown one at a time. There isn't a checklist where you can tick "
                "several and add them all together."),
        "q": ("Is picking them one at a time acceptable for this release, or should "
              "there be a checklist to add several at once?"),
        "opts": ("A) Keep it as it is - adding one at a time is fine.\n"
                 "B) Change it so you can tick several and add them all at once."),
    },
]

internal_map = [
    ("1", "Part 1 #1", "FD-STATS-001 (BUG-FD-2 / FDBUG-6); also settles FD-STATS-002, FD-STATS-004"),
    ("2", "Part 1 #4", "FD-CUST-016 / FD-VAL-007 (BUG-FD-1 double-add; did not reproduce - confirm S9 dedupe shipped)"),
    ("3", "Part 1 #5", "NOTE-FD-4 (Story 8 Processing Fee - backend accepts it, builder UI missing)"),
    ("4", "Part 1 #6", "FD-WO-005 / FD-VAL-001 (BUG-FD-4 - confirm button enabled; validation on submit)"),
    ("5", "Part 1 #7", "FD-INLINE-003 (BUG-FD-5 - no \"Show N more\" collapse on line adjustments)"),
    ("6", "Part 1 #8", "FD-CUST-005 (NOTE-FD-5 / FDBUG-7); ruling also settles FD-CUST-003/004/006/007"),
]

wb = Workbook()

# --- About sheet ---
ws = wb.active
ws.title = "About"
ws.column_dimensions["A"].width = 100
ws["A1"] = "Fees & Discounts - A Few Quick Questions for You"
ws["A1"].font = TITLE_FONT
about_lines = [
    "",
    "This is about the new Fees & Discounts feature.",
    "",
    ("While checking it over, we found a handful of spots where the app behaves a little "
     "differently than the original write-up described. None of these are problems on "
     "their own - we just want your quick call on each one so we know which way you'd "
     "like it to work."),
    "",
    ("There are no wrong answers. On the \"PO Questions\" tab, each row has a short "
     "description of what happens now, the question, and simple A/B options. Please pick "
     "an option (or write your own) in the \"Your answer\" column."),
    "",
    "It should take just a few minutes. Thank you!",
]
r = 2
for line in about_lines:
    ws.cell(row=r, column=1, value=line).alignment = WRAP
    ws.row_dimensions[r].height = 30 if len(line) > 60 else 15
    r += 1

# --- PO Questions sheet ---
wq = wb.create_sheet("PO Questions")
headers = ["#", "Topic", "What happens now", "The question", "Options", "Your answer"]
widths = [5, 34, 46, 40, 46, 26]
for c, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = wq.cell(row=1, column=1 + c - 1, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wq.column_dimensions[chr(64 + c)].width = w
wq.freeze_panes = "A2"

for i, item in enumerate(questions, start=1):
    row = i + 1
    vals = [i, item["topic"], item["now"], item["q"], item["opts"], ""]
    for c, v in enumerate(vals, start=1):
        cell = wq.cell(row=row, column=c, value=v)
        cell.alignment = WRAP_CENTER if c == 1 else WRAP
        cell.border = BORDER
    wq.row_dimensions[row].height = 110

# --- Internal sheet ---
wi = wb.create_sheet("Internal (QA only)")
wi["A1"] = "INTERNAL - for QA only. Do NOT share with the PO."
wi["A1"].font = Font(bold=True, color="C00000", size=12)
wi.merge_cells("A1:C1")
ih = ["Q#", "Source thread", "Cases / refs covered"]
iw = [6, 22, 90]
for c, (h, w) in enumerate(ih, start=1) if False else []:
    pass
for c, (h, w) in enumerate(zip(ih, iw), start=1):
    cell = wi.cell(row=3, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = WRAP_CENTER
    cell.border = BORDER
    wi.column_dimensions[chr(64 + c)].width = w
for i, (qn, src, refs) in enumerate(internal_map, start=4):
    for c, v in enumerate([qn, src, refs], start=1):
        cell = wi.cell(row=i, column=c, value=v)
        cell.alignment = WRAP
        cell.border = BORDER
    wi.row_dimensions[i].height = 30
note_row = 4 + len(internal_map) + 1
wi.cell(row=note_row, column=1,
        value=("Note: Two threads were REMOVED from the PO document as dev bugs (not "
               "product decisions), already covered in jira-bug-drafts.md: the whole-WO "
               "permission \"hidden but not backend-enforced\" item (BUG-FD-3), and the "
               "customer total/estimate leaving out the fee/discount amount (FDBUG-1). "
               "Pure code-bug tickets and the Part 2 case-update wording proposals are "
               "also excluded. Source: "
               "build/fees-discounts/Deviations-and-Questions-for-PO.md")).alignment = WRAP
wi.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
wi.row_dimensions[note_row].height = 45

wb.save(OUT)
print(f"Wrote {OUT} with {len(questions)} PO questions.")
