#!/usr/bin/env python3
"""Generate ANTICIPATED-QUESTIONS-AND-ANSWERS.md + .xlsx from register_data.py.

FORMAT (Standing Rule 16 - mirrors the established workbook conventions used by
build/report-suite/gen_po_questions_2026-07-31.py and build/filters/gen_po_questions.py):
  title row in A1 (bold), blank row, header row on row 3 with fill 1F4E79 / white bold /
  wrap, freeze at A4, explicit column widths, wrap+top alignment on every data cell.

Tabs:
  1 "Summary"            - what this is, the counts per category / risk / status
  2 "TOP 10"             - ranked by likelihood x risk, paste-ready answers
  3 "Report Suite"       - the project registers, one tab each
  4 "Schedule"
  5 "Filters"
  6 "Cross-project"
  7 "Conceded"           - the honest concession list
  8 "All rows"           - every row in one flat sheet for filtering

Standing Rule 8: every named case carries its TestRail Case ID and a clickable link -
columns "Test case(s)" (readable) and "TestRail link(s)" (URLs).
Re-run:  python3 gen_preemptive_register.py
"""
import os
import sys
from collections import Counter, OrderedDict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

from register_data import CATEGORIES, CONCEDED, COUNTS, ROWS, TOP10, TR  # noqa: E402

BASE = "ANTICIPATED-QUESTIONS-AND-ANSWERS"
DATE = "2026-07-31"
PROJECT_ORDER = ["Report Suite", "Schedule", "Filters", "Cross-project"]
RISK_ORDER = ["HIGH", "MEDIUM", "LOW"]
STATUS_ORDER = ["SETTLED", "AWAITING ANSWER", "AWAITING LIVE BUILD", "SCHEDULED", "ACCEPTED"]

TITLE = ("Anticipated questions and ready answers - a defensibility register for the QA lead "
         f"- {DATE}")


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def numbered():
    """ROWS with a stable 1-based number, in list order."""
    out = []
    for i, r in enumerate(ROWS, 1):
        d = dict(r)
        d["n"] = i
        out.append(d)
    return out


def cases_text(cases):
    """Readable case string: internal ID = C-id (link) for every case (Rule 8)."""
    if isinstance(cases, str):
        return cases
    return " · ".join(f"{iid} = C{cid} ({TR}{cid})" for iid, cid in cases)


def cases_md(cases):
    if isinstance(cases, str):
        return cases
    return " · ".join(f"**{iid} = [C{cid}]({TR}{cid})**" for iid, cid in cases)


def cases_plain(cases):
    """Just the IDs, no URLs - for the workbook's readable column."""
    if isinstance(cases, str):
        return cases
    return " · ".join(f"{iid} = C{cid}" for iid, cid in cases)


def cases_links(cases):
    if isinstance(cases, str):
        return ""
    return "\n".join(f"{TR}{cid}" for _, cid in cases)


def resolve_top10(rows):
    """Match TOP10 keys onto rows; fail loudly rather than silently mis-rank."""
    out = []
    for proj, cat, prefix in TOP10:
        hit = [r for r in rows
               if r["project"] == proj and r["category"] == cat
               and r["say"].startswith(prefix)]
        if len(hit) != 1:
            raise SystemExit(
                f"TOP10 key did not resolve to exactly one row: {proj}/{cat}/{prefix!r} "
                f"-> {len(hit)} matches")
        out.append(hit[0])
    return out


# --------------------------------------------------------------------------- #
# markdown
# --------------------------------------------------------------------------- #
def write_md(rows, top):
    cat_counts = Counter(r["category"] for r in rows)
    risk_counts = Counter(r["risk"] for r in rows)
    status_counts = Counter(r["status"] for r in rows)
    proj_counts = Counter(r["project"] for r in rows)

    L = []
    a = L.append

    a(f"# Anticipated questions and ready answers — {DATE}")
    a("")
    a("**What this is.** A defensibility register. Every row is something a QA, an engineer or "
      "a product owner could challenge about our test cases, paired with the one-paragraph "
      "answer you can paste straight into a channel and the exact evidence behind it. The "
      "failure mode this exists to prevent is you being blindsided.")
    a("")
    a("**Scope:** the three active projects — **Report Suite** (PO Chris Ward, epic SV-8582), "
      "**Schedule** (PO Branko, epic SV-8685), **Filters** (PO Branko, **no epic exists**) — "
      "plus cross-project items.")
    a("")
    a("**Who this anticipates:** Ahtesham Amjad and Mudassir Qamar (manual QA), Vladimir "
      "Tomovic (automation), Stefan Mitrovic (engineering manager), Chris Ward and Branko "
      "Cicovic (product owners), and anyone reading a run cold in a public channel.")
    a("")
    a("**How to use it:** find the row, paste the answer, and only open the evidence if you are "
      "pushed. Where the honest answer is *\"you are right\"*, the row says so — those are "
      "collected in **[What is genuinely conceded](#what-is-genuinely-conceded)** at the end.")
    a("")

    # ---- source currency -------------------------------------------------
    a("---")
    a("")
    a("## SOURCE-CURRENCY (Standing Rule 31)")
    a("")
    a("Every claim in this register was re-read from its source this pass. This register is "
      "**analysis of our own artefacts** — it makes no new product claim of its own.")
    a("")
    a("| Source | Identifier | Version / last-updated | Checked | Verdict |")
    a("|---|---|---|---|---|")
    a("| Report Suite specs ×6 | Confluence 577634305 / 585629698 / 620888066 / 641400833 / "
      "703660034 / 720142338 | SBC **v12** · SBR **v15** · PV **v4** · TU **v5** · WIP **v6** · "
      "IV **v3**, all updated 2026-07-29 | 2026-07-31 (captures in "
      "`build/report-suite/spec-current-2026-07-31/`) | **CURRENT** |")
    a("| Schedule spec | Confluence **713031682** | Confluence version **23**, "
      "2026-07-30T10:40:32Z (the in-body \"Version\" field still reads 1.0 — do not trust it) | "
      "2026-07-31 (`build/schedule/spec-current-2026-07-31/`) | **CURRENT** |")
    a("| Filters spec | Confluence **572030978** | Confluence version **12** = spec **v1.6**, "
      "2026-07-28 | 2026-07-31 (`build/filters/spec-current-2026-07-31/`) | **CURRENT** |")
    a("| Epics | **SV-8582** (97 children) · **SV-8685** (15 children) · Filters = **none, "
      "proven** (all 170 SV epics enumerated) | Tier-1 currency check 2026-07-31 | 2026-07-31 "
      "(`build/epic-recheck-2026-07-31/`) | **CURRENT** |")
    a("| Designs | Report Suite **none exist** · Schedule = the Claude prototype "
      "(authoritative, Branko Q0) · Filters Figma `DR4gEODShYgJqkozs3mF5q` | Filters **73 of "
      "85 boards rendered — 12 pending** | 2026-07-31 | **PARTIAL (Filters)** — Rule-35 queue "
      "OPEN, `build/filters/design-2026-07-31/PENDING-FIGMA-FETCH.md` |")
    a("| Engineering tech plans | all three projects | reconciled 2026-07-29 / 2026-07-30 | "
      "2026-07-31 | **CURRENT** |")
    a("| PO answers | Chris Ward 2026-07-28 · 2026-07-29 · 2026-07-31 · Branko 2026-07-17 · "
      "2026-07-20 · 2026-07-31 | newest = 2026-07-31 | 2026-07-31 | **CURRENT** |")
    a("| Live build | Report Suite · Schedule · Filters | **no environment exists for any of "
      "the three** | 2026-07-31 | **MISSING — nothing in this register is live-verified** "
      "(Rules 12/22) |")
    a("")
    a("**TestRail:** read-only. **Zero writes were made producing this register** — no case, "
      "section, run or result was created, updated, deleted or moved.")
    a("")

    # ---- counts ----------------------------------------------------------
    a("---")
    a("")
    a("## The register at a glance")
    a("")
    a(f"**{len(rows)} rows.**")
    a("")
    a("| Category | What it means | Rows |")
    a("|---|---|---|")
    for k in sorted(CATEGORIES):
        a(f"| **{k}** | {CATEGORIES[k]} | **{cat_counts.get(k, 0)}** |")
    a("")
    a("| Project | Rows | Our active cases | Live in TestRail | Tester run |")
    a("|---|---|---|---|---|")
    for p in PROJECT_ORDER:
        c = COUNTS.get(p)
        if c:
            a(f"| **{p}** | {proj_counts.get(p, 0)} | **{c['ours']}** | {c['live']} | "
              f"run **{c['run']}** = {c['run_tests']} tests, {c['run_results']} results |")
        else:
            a(f"| **{p}** | {proj_counts.get(p, 0)} | — | — | — |")
    a("")
    a("| Risk if raised publicly | Rows | What it means |")
    a("|---|---|---|")
    a(f"| **HIGH** | **{risk_counts.get('HIGH', 0)}** | We would have to concede something. "
      "Know these cold. |")
    a(f"| **MEDIUM** | **{risk_counts.get('MEDIUM', 0)}** | Defensible, but it needs a short "
      "explanation, not a one-liner. |")
    a(f"| **LOW** | **{risk_counts.get('LOW', 0)}** | Quote the source and it is over. |")
    a("")
    a("| Status | Rows |")
    a("|---|---|")
    for s in STATUS_ORDER:
        a(f"| **{s}** | {status_counts.get(s, 0)} |")
    a("")

    # ---- top 10 ----------------------------------------------------------
    a("---")
    a("")
    a("## TOP 10 MOST LIKELY TO BE RAISED")
    a("")
    a("Ranked by **likelihood × risk**. Each answer is written to be pasted as-is.")
    a("")
    for i, r in enumerate(top, 1):
        a(f"### {i}. {r['say']}")
        a("")
        a(f"*{r['project']} · category {r['category']} · **{r['status']}** · risk "
          f"**{r['risk']}** · register row {r['n']}*")
        a("")
        a(f"> {r['answer']}")
        a("")
        a(f"**Evidence:** {r['evidence']}")
        a("")
        a(f"**Test case(s):** {cases_md(r['cases'])}")
        a("")
        a(f"**Who can close it:** {r['closer']}")
        a("")

    # ---- per project -----------------------------------------------------
    a("---")
    a("")
    a("## The full register, by project")
    a("")
    for p in PROJECT_ORDER:
        prows = [r for r in rows if r["project"] == p]
        if not p or not prows:
            continue
        c = COUNTS.get(p)
        a(f"## {p}")
        a("")
        if c:
            a(f"*{c['ours']} active cases of ours ({c['retired_note']}) · live in TestRail "
              f"{c['live']} · tester run {c['run']} = {c['run_tests']} tests with "
              f"{c['run_results']} recorded results · **every case VIU-Pending, no environment "
              f"exists**.*")
            a("")
        by_cat = OrderedDict()
        for r in prows:
            by_cat.setdefault(r["category"], []).append(r)
        for cat in sorted(by_cat):
            a(f"### {p} — category {cat}: {CATEGORIES[cat]}")
            a("")
            for r in by_cat[cat]:
                a(f"#### {r['n']}. “{r['say']}”")
                a("")
                a(f"**Our answer.** {r['answer']}")
                a("")
                a(f"**The evidence.** {r['evidence']}")
                a("")
                a(f"**Test case(s).** {cases_md(r['cases'])}")
                a("")
                a(f"**Who can close it.** {r['closer']}  ·  **Status.** {r['status']}  ·  "
                  f"**Risk if raised publicly.** {r['risk']}")
                a("")
        a("---")
        a("")

    # ---- conceded --------------------------------------------------------
    a("## What is genuinely conceded")
    a("")
    a("If any of these is raised, the honest answer starts with *\"you are right\"*. They are "
      "listed openly because being blindsided is worse than the concession.")
    a("")
    for i, (what, detail, need) in enumerate(CONCEDED, 1):
        a(f"### {i}. {what}")
        a("")
        a(detail)
        a("")
        a(f"**What it takes to close it:** {need}")
        a("")

    # ---- outstanding -----------------------------------------------------
    a("---")
    a("")
    a("## OUTSTANDING — what I need from you")
    a("")
    a("Swept across all six Standing-Rule-36 categories. The full cross-project list is "
      "`build/OUTSTANDING-ITEMS-REGISTER.md`; this is what **this pass** needs.")
    a("")
    a("**1. Missing sources.** The **Filters Jira epic** — or your confirmation that the work "
      "genuinely is not ticketed. Without it, none of the 110 Filters cases can ever cite a "
      "ticket, and half of our own traceability rule stays unsatisfiable. *Since 2026-07-17; "
      "proven absent 2026-07-31.* Also: **12 of 85 Filters design boards** still have no "
      "rendered picture (auto-retrying, no authorization needed) — but on a fresh container the "
      "**Figma token** has to be re-supplied by you.")
    a("")
    a("**2. Unanswered questions.** **Chris Ward** — the 8 spec-text corrections headed by the "
      "**WIP asset identifier** (he believes he already made that edit; he has not), plus 3 "
      "questions from his own answers (does \"normal reports access\" collapse the other five "
      "reports' permissions · does \"Rep is active?\" become \"Representative\" · the exact "
      "renamed assignments file name). **Branko** — the **8-question Schedule sheet** and the "
      "**8-question Filters sheet** are both written and **not yet sent**; his Filters Q1 (the "
      "numbered Parts/Reports write-up) came back **blank**. **Engineering** — Schedule NQ-5, "
      "may a technician change another technician's shifts (Branko declined it, correctly).")
    a("")
    a("**3. Missing go-aheads.** A **dev change ticket against SV-8582** for the Sales By "
      "Customer permission gap (draft written) — without it a tester reads three deliberate "
      "failures as our defect. Plus three small authorized touches: the Schedule "
      "**`SCH-HRS-04`** stray-bracket fix, deleting the now-empty Schedule section **5406**, "
      "and the 3 optional Filters under-merge findings.")
    a("")
    a("**4. Access and credentials.** **The QA environment for all three projects**, plus the "
      "feature-flag state and fresh cookies. **749 cases are written, traced, audited and in "
      "the testers' runs, and not one has been checked against a running build.** This is the "
      "single biggest item on this list.")
    a("")
    a("**5. Deferred or held decisions.** **Run 278** (Custom Permissions — 9 cases missing, "
      "**3,521 graded results**; Custom Roles is an active recurring project so your "
      "completed-projects ruling does not cover it). **Simple Flow V2 (SV-8683)** — Open with 7 "
      "children against a project we recorded as completed; this needs a yes/no, not analysis. "
      "The **19 Filters dropdown merges**, correctly held until someone can look at a build.")
    a("")
    a("**6. What another team owes.** **Chris Ward** — the SBR v15 self-contradiction on the CSV "
      "headers, and the six still-open spec-watch items (two of which now contradict rulings he "
      "gave us afterwards). **Branko** — the Filters PRD's \"hidden\" Status-chip prose in six "
      "places, still unaligned with his own answer eight versions later, plus three Schedule "
      "spec sentences that contradict his own rulings. **Dev** — Jira **SV-8695** still lists a "
      "modal Reassign action the spec deleted. **Fabian / you** — the sell-price bug ticket key "
      "for three Simple Flow cases still carrying a placeholder.")
    a("")
    a("---")
    a("")
    a(f"*Generated by `gen_preemptive_register.py` from `register_data.py`. {len(rows)} rows. "
      "Read-only pass — zero TestRail writes.*")

    path = os.path.join(HERE, BASE + ".md")
    with open(path, "w") as fh:
        fh.write("\n".join(L) + "\n")
    return path


# --------------------------------------------------------------------------- #
# workbook
# --------------------------------------------------------------------------- #
HDR_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

COLS = ["#", "Project", "Category", "What they could say",
        "Our answer (plain, 1-3 sentences)", "The evidence", "Test case(s)",
        "TestRail link(s)", "Who can close it", "Status", "Risk if raised publicly"]
WIDTHS = [5, 15, 34, 52, 62, 78, 46, 44, 34, 22, 14]


def _sheet(wb, name, title, first=False):
    ws = wb.active if first else wb.create_sheet(name)
    if first:
        ws.title = name
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    return ws


def _row_table(ws, rows, start=3):
    for j, c in enumerate(COLS, 1):
        cell = ws.cell(row=start, column=j, value=c)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
    r = start + 1
    for d in rows:
        vals = [d["n"], d["project"], f"{d['category']} - {CATEGORIES[d['category']]}",
                d["say"], d["answer"], d["evidence"], cases_plain(d["cases"]),
                cases_links(d["cases"]), d["closer"], d["status"], d["risk"]]
        for j, v in enumerate(vals, 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    for col, w in zip("ABCDEFGHIJK", WIDTHS):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = f"A{start + 1}"
    return r


def write_xlsx(rows, top):
    wb = openpyxl.Workbook()
    cat_counts = Counter(r["category"] for r in rows)
    risk_counts = Counter(r["risk"] for r in rows)
    status_counts = Counter(r["status"] for r in rows)
    proj_counts = Counter(r["project"] for r in rows)

    # --- Summary ---
    ws = _sheet(wb, "Summary", TITLE, first=True)
    r = 3
    blocks = [
        ("What this is",
         ["A defensibility register. Every row = something someone could challenge about our",
          "test cases + the paste-ready answer + the evidence behind it.",
          "Scope: Report Suite (PO Chris Ward, epic SV-8582), Schedule (PO Branko, epic",
          "SV-8685), Filters (PO Branko, NO epic exists), plus cross-project items.",
          "READ-ONLY pass: ZERO TestRail writes were made producing this register.",
          "NOTHING in this register is live-verified - no QA environment exists for any of the",
          "three active projects (Standing Rules 12 and 22)."]),
    ]
    for head, lines in blocks:
        ws.cell(row=r, column=1, value=head).font = Font(bold=True)
        r += 1
        for ln in lines:
            ws.cell(row=r, column=1, value=ln).alignment = WRAP
            r += 1
        r += 1

    def table(head, cols, data, widths):
        nonlocal r
        ws.cell(row=r, column=1, value=head).font = Font(bold=True)
        r += 1
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row=r, column=j, value=c)
            cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
        r += 1
        for tup in data:
            for j, v in enumerate(tup, 1):
                ws.cell(row=r, column=j, value=v).alignment = WRAP
            r += 1
        r += 1
        for col, w in zip("ABCDE", widths):
            if ws.column_dimensions[col].width < w:
                ws.column_dimensions[col].width = w

    for col, w in zip("ABCDE", [56, 20, 20, 20, 20]):
        ws.column_dimensions[col].width = w

    table("Rows per category", ["Category", "What it means", "Rows"],
          [(k, CATEGORIES[k], cat_counts.get(k, 0)) for k in sorted(CATEGORIES)],
          [56, 56, 10])
    table("Rows per project (with the case counts, verified 2026-07-31)",
          ["Project", "Rows", "Our active cases", "Live in TestRail", "Tester run"],
          [(p, proj_counts.get(p, 0),
            COUNTS[p]["ours"] if p in COUNTS else "-",
            COUNTS[p]["live"] if p in COUNTS else "-",
            (f"run {COUNTS[p]['run']} = {COUNTS[p]['run_tests']} tests, "
             f"{COUNTS[p]['run_results']} results") if p in COUNTS else "-")
           for p in PROJECT_ORDER],
          [56, 10, 20, 20, 40])
    table("Rows per risk level", ["Risk", "Rows", "What it means"],
          [("HIGH", risk_counts.get("HIGH", 0), "We would have to concede something."),
           ("MEDIUM", risk_counts.get("MEDIUM", 0), "Defensible, needs a short explanation."),
           ("LOW", risk_counts.get("LOW", 0), "Quote the source and it is over.")],
          [56, 10, 56])
    table("Rows per status", ["Status", "Rows"],
          [(s, status_counts.get(s, 0)) for s in STATUS_ORDER], [56, 10])
    table("Tabs in this workbook", ["Tab", "What is on it"],
          [("TOP 10", "the ten most likely challenges, ranked by likelihood x risk"),
           ("Report Suite / Schedule / Filters / Cross-project", "the per-project registers"),
           ("Conceded", "what we would have to concede if challenged"),
           ("All rows", "every row in one flat sheet, for filtering")],
          [56, 78])

    # --- TOP 10 ---
    ws = _sheet(wb, "TOP 10",
                "TOP 10 MOST LIKELY TO BE RAISED - ranked by likelihood x risk. "
                "Column E is paste-ready.")
    _row_table(ws, top)

    # --- per project ---
    for p in PROJECT_ORDER:
        prows = sorted([r for r in rows if r["project"] == p], key=lambda d: (d["category"], d["n"]))
        if not prows:
            continue
        c = COUNTS.get(p)
        sub = (f"{c['ours']} active cases of ours ({c['retired_note']}); live {c['live']}; run "
               f"{c['run']} = {c['run_tests']} tests / {c['run_results']} results; every case "
               f"VIU-Pending - no environment exists.") if c else \
              "Items that do not belong to a single project."
        ws = _sheet(wb, p[:31], f"{p} - anticipated questions and ready answers. {sub}")
        _row_table(ws, prows)

    # --- Conceded ---
    ws = _sheet(wb, "Conceded",
                "WHAT IS GENUINELY CONCEDED - if raised, the honest answer starts with "
                "\"you are right\". Listed openly on purpose.")
    cols = ["#", "What we would concede", "The detail", "What it takes to close it"]
    for j, cc in enumerate(cols, 1):
        cell = ws.cell(row=3, column=j, value=cc)
        cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, WRAP
    r = 4
    for i, (what, detail, need) in enumerate(CONCEDED, 1):
        for j, v in enumerate([i, what, detail, need], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP
        r += 1
    for col, w in zip("ABCD", [5, 62, 90, 62]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # --- All rows ---
    ws = _sheet(wb, "All rows",
                "Every row in one sheet. TestRail Case IDs in column G, clickable links in "
                "column H (Standing Rule 8).")
    _row_table(ws, rows)

    path = os.path.join(HERE, BASE + ".xlsx")
    wb.save(path)
    return path


if __name__ == "__main__":
    rows = numbered()
    top = resolve_top10(rows)
    md = write_md(rows, top)
    xl = write_xlsx(rows, top)
    cat_counts = Counter(r["category"] for r in rows)
    risk_counts = Counter(r["risk"] for r in rows)
    print("wrote", os.path.basename(md), "and", os.path.basename(xl))
    print(f"  rows: {len(rows)} | top10: {len(top)} | conceded: {len(CONCEDED)}")
    print("  per category:", {k: cat_counts.get(k, 0) for k in sorted(CATEGORIES)})
    print("  per risk:", {k: risk_counts.get(k, 0) for k in RISK_ORDER})
    print("  per project:", dict(Counter(r["project"] for r in rows)))
